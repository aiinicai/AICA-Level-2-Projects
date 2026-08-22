"""
excel_handler.py — Excel template generation, upload validation, import, and export
"""

import io
import re
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

from utils import validate_gstin, parse_date, format_date

# ─────────────────────────────────────────────
# Column mapping: Excel header → DB field name
# ─────────────────────────────────────────────

COLUMN_MAP = {
    "Client Name":                    "client_name",
    "GSTIN":                          "gstin",
    "Notice/Reference Number":        "notice_number",
    "Notice Issue Date":              "notice_issue_date",
    "Notice Issued Under Section":    "notice_section",
    "Act Type (CGST/SGST/IGST)":      "act_type",
    "Issuing Officer":                "issuing_officer",
    "Officer Designation":            "officer_designation",
    "Due Date":                       "due_date",
    "Notice Type/Subject":            "notice_type",
    "Client Data Collection Status":  "client_data_status",
    "Data Requested":                 "data_requested",
    "Date Data Requested":            "date_data_requested",
    "Date Data Received":             "date_data_received",
    "Assigned Team Member":           "assigned_team_member",
    "Response Filing Date":           "response_filing_date",
    "Response Status":                "response_status",
    "Remarks":                        "remarks",
}

REQUIRED_COLUMNS = list(COLUMN_MAP.keys())

REQUIRED_FIELDS = [
    "Client Name",
    "GSTIN",
    "Notice/Reference Number",
    "Due Date",
]

DATE_COLUMNS = [
    "Notice Issue Date",
    "Due Date",
    "Date Data Requested",
    "Date Data Received",
    "Response Filing Date",
]

VALID_ACT_TYPES = {"CGST", "SGST", "IGST", "CGST/SGST", "CGST/IGST", "SGST/IGST", ""}


# ─────────────────────────────────────────────
# Template Generation
# ─────────────────────────────────────────────

def generate_template() -> bytes:
    """Return a styled openpyxl workbook as bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GST Notice Import"

    # ── Style definitions ──────────────────────────────────────────────
    header_fill   = PatternFill("solid", fgColor="1F3864")
    req_fill      = PatternFill("solid", fgColor="C00000")   # required fields highlight
    sample_fill   = PatternFill("solid", fgColor="EBF3FB")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    req_font      = Font(bold=True, color="FFD700", size=10)
    sample_font   = Font(size=10)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align    = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin"),   right=Side(style="thin"),
        top=Side(style="thin"),    bottom=Side(style="thin"),
    )

    # ── Header row ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    for col_idx, header in enumerate(REQUIRED_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        is_required = header in REQUIRED_FIELDS
        cell.fill   = req_fill if is_required else header_fill
        cell.font   = req_font if is_required else header_font
        cell.alignment = center_align
        cell.border    = thin_border

    # ── Sample row ────────────────────────────────────────────────────
    sample_data = [
        "ABC Pvt Ltd",           # Client Name
        "27AABCU9603R1ZX",       # GSTIN
        "GSTN/2024/001",         # Notice/Reference Number
        "01-04-2024",            # Notice Issue Date
        "Section 73",            # Notice Issued Under Section
        "CGST",                  # Act Type
        "ITO GST",               # Issuing Officer
        "Inspector",             # Officer Designation
        "31-08-2024",            # Due Date
        "Show Cause Notice",     # Notice Type/Subject
        "Pending",               # Client Data Collection Status
        "Bank statements, invoices",  # Data Requested
        "05-04-2024",            # Date Data Requested
        "10-04-2024",            # Date Data Received
        "Rahul Sharma",          # Assigned Team Member
        "",                      # Response Filing Date
        "Pending",               # Response Status
        "Follow up required",    # Remarks
    ]

    ws.row_dimensions[2].height = 25
    for col_idx, value in enumerate(sample_data, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill      = sample_fill
        cell.font      = sample_font
        cell.alignment = left_align
        cell.border    = thin_border

    # ── Column widths ────────────────────────────────────────────────
    col_widths = [
        20, 18, 22, 16, 25, 10, 18, 18, 14, 22,
        24, 25, 18, 18, 20, 18, 16, 22,
    ]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Instructions sheet ───────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    instructions = [
        ("GST Notice Tracker — Excel Import Instructions", True),
        ("", False),
        ("RED column headers = Required fields (cannot be blank)", False),
        ("BLUE column headers = Optional fields", False),
        ("", False),
        ("GSTIN Format: 2-digit state code + 5 alpha + 4 digits + 1 alpha + 1 alpha/digit + Z + 1 alphanumeric", False),
        ("  Example: 27AABCU9603R1ZX", False),
        ("", False),
        ("Date Format: DD-MM-YYYY (e.g., 31-08-2024)  or  DD/MM/YYYY", False),
        ("", False),
        ("Act Type must be one of: CGST, SGST, IGST (or a combination like CGST/SGST)", False),
        ("", False),
        ("Client Data Collection Status options:", False),
        ("  Pending | Awaiting Client Data | Partially Received | Data Received | Not Applicable", False),
        ("", False),
        ("Response Status options:", False),
        ("  Pending | In Progress | Filed | Completed | Not Applicable", False),
        ("", False),
        ("Deduplication Key: GSTIN + Notice/Reference Number", False),
        ("  Uploading the same GSTIN+Notice No. again will UPDATE the record (in Update mode)", False),
        ("  or SKIP it (in Add New mode).", False),
    ]
    wi.column_dimensions["A"].width = 90
    for row_idx, (text, bold) in enumerate(instructions, start=1):
        cell = wi.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=11 if bold else 10)
        cell.alignment = left_align

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
# Excel Validation
# ─────────────────────────────────────────────

def validate_excel(df_raw: pd.DataFrame, existing_keys: set) -> dict:
    """
    Validate uploaded DataFrame.
    Returns:
    {
        valid_records:   list[dict],   # ready to insert into DB
        invalid_rows:    list[dict],   # {"row": N, "errors": [...], "data": {...}}
        duplicate_rows:  list[dict],   # rows already in DB
        summary: {
            total, valid, invalid, duplicates
        }
    }
    """
    # 1. Column check ─────────────────────────────────────────────────
    missing_cols = [c for c in REQUIRED_COLUMNS if c not in df_raw.columns]
    if missing_cols:
        return {
            "column_error": f"Missing columns: {', '.join(missing_cols)}",
            "valid_records": [],
            "invalid_rows": [],
            "duplicate_rows": [],
            "summary": {"total": 0, "valid": 0, "invalid": 0, "duplicates": 0},
        }

    valid_records  = []
    invalid_rows   = []
    duplicate_rows = []

    for i, row in df_raw.iterrows():
        excel_row = i + 2   # account for 0-index + header row
        errors = []

        # ── Required fields ───────────────────────────────────────────
        for req_col in REQUIRED_FIELDS:
            val = str(row.get(req_col, "")).strip()
            if not val or val.lower() in ("nan", "none"):
                errors.append(f"'{req_col}' is required")

        # ── GSTIN validation ──────────────────────────────────────────
        raw_gstin = str(row.get("GSTIN", "")).strip()
        if raw_gstin and raw_gstin.lower() not in ("nan", "none"):
            if not validate_gstin(raw_gstin):
                errors.append(f"Invalid GSTIN '{raw_gstin}'")

        # ── Date validation ───────────────────────────────────────────
        parsed_dates = {}
        for date_col in DATE_COLUMNS:
            raw_val = row.get(date_col, None)
            # Treat empty / NaN as blank (allowed for optional date fields)
            if raw_val is None or (isinstance(raw_val, float) and pd.isna(raw_val)):
                parsed_dates[date_col] = None
                continue
            if isinstance(raw_val, str) and raw_val.strip() in ("", "nan", "none", "None"):
                parsed_dates[date_col] = None
                continue
            dt = parse_date(raw_val)
            if dt is None:
                errors.append(f"Invalid date in '{date_col}': {raw_val}")
            else:
                parsed_dates[date_col] = dt

        # ── Act Type ──────────────────────────────────────────────────
        act_val = str(row.get("Act Type (CGST/SGST/IGST)", "")).strip().upper()
        if act_val and act_val.lower() not in ("nan", "none") and act_val not in VALID_ACT_TYPES:
            errors.append(f"Act Type must be CGST, SGST, or IGST (got '{act_val}')")

        # ── Build record dict ─────────────────────────────────────────
        record = {}
        for excel_col, db_field in COLUMN_MAP.items():
            raw = row.get(excel_col, "")
            if excel_col in DATE_COLUMNS:
                dt = parsed_dates.get(excel_col)
                record[db_field] = format_date(dt) if dt else ""
            elif excel_col == "GSTIN":
                record[db_field] = str(raw).strip().upper() if str(raw).strip().lower() not in ("nan", "none", "") else ""
            else:
                val = str(raw).strip() if raw is not None else ""
                record[db_field] = "" if val.lower() in ("nan", "none") else val

        row_info = {"row": excel_row, "data": record}

        if errors:
            invalid_rows.append({**row_info, "errors": errors})
            continue

        # ── Duplicate check ───────────────────────────────────────────
        key = (record["gstin"].upper(), record["notice_number"].upper())
        if key in existing_keys:
            duplicate_rows.append({**row_info, "errors": ["Duplicate: already exists in database"]})
        else:
            valid_records.append(record)

    return {
        "valid_records":  valid_records,
        "invalid_rows":   invalid_rows,
        "duplicate_rows": duplicate_rows,
        "summary": {
            "total":      len(df_raw),
            "valid":      len(valid_records),
            "invalid":    len(invalid_rows),
            "duplicates": len(duplicate_rows),
        },
        "column_error": None,
    }


def validate_excel_for_update(df_raw: pd.DataFrame, existing_keys: set) -> dict:
    """
    Validate uploaded DataFrame for Upsert mode (Update existing & Add new).
    - Records matching existing_keys will be updated in DB while preserving existing data for empty fields.
    - New records will be added.
    - Duplicates within the uploaded Excel file itself are flagged and skipped.
    - Complete / Filed status counts are aggregated for user inspection.
    """
    result = validate_excel(df_raw, set())   # skip DB duplicate check inside validate_excel

    seen_keys = set()
    valid_unique = []
    file_duplicates = []

    for rec in result["valid_records"]:
        key = (rec["gstin"].upper(), rec["notice_number"].upper())
        if key in seen_keys:
            file_duplicates.append({
                "row": None,
                "data": rec,
                "errors": ["Duplicate record within uploaded Excel file"],
            })
        else:
            seen_keys.add(key)
            valid_unique.append(rec)

    result["duplicate_rows"].extend(file_duplicates)

    update_records = []
    insert_records = []
    completed_count = 0
    filed_count = 0
    pending_count = 0

    for rec in valid_unique:
        key = (rec["gstin"].upper(), rec["notice_number"].upper())
        if key in existing_keys:
            update_records.append(rec)
        else:
            insert_records.append(rec)

        r_status = (rec.get("response_status") or "").strip().lower()
        if r_status == "completed":
            completed_count += 1
        elif r_status == "filed":
            filed_count += 1
        else:
            pending_count += 1

    result["valid_records"] = valid_unique
    result["update_records"] = update_records
    result["insert_records"] = insert_records
    result["summary"]["valid"] = len(valid_unique)
    result["summary"]["duplicates"] = len(result["duplicate_rows"])
    result["summary"]["will_update"] = len(update_records)
    result["summary"]["will_insert"] = len(insert_records)
    result["summary"]["completed_count"] = completed_count
    result["summary"]["filed_count"] = filed_count
    result["summary"]["pending_count"] = pending_count
    return result


def validate_excel_for_replace(df_raw: pd.DataFrame) -> dict:
    """
    Validate uploaded DataFrame for 'Replace Existing Data' mode.
    Existing DB records are ignored since existing data will be wiped.
    Duplicates within the uploaded Excel file itself are flagged.
    """
    result = validate_excel(df_raw, set())   # skip DB duplicate check
    seen_keys = set()
    valid_unique = []
    file_duplicates = []

    for rec in result["valid_records"]:
        key = (rec["gstin"].upper(), rec["notice_number"].upper())
        if key in seen_keys:
            file_duplicates.append({
                "row": None,
                "data": rec,
                "errors": ["Duplicate record within uploaded Excel file"],
            })
        else:
            seen_keys.add(key)
            valid_unique.append(rec)

    result["valid_records"] = valid_unique
    result["duplicate_rows"].extend(file_duplicates)
    result["summary"]["valid"] = len(valid_unique)
    result["summary"]["duplicates"] = len(result["duplicate_rows"])
    return result


# ─────────────────────────────────────────────
# Export
# ─────────────────────────────────────────────


def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Export the notice register DataFrame to a styled xlsx file.
    The sheet uses the same column names as the import template so the file
    can be uploaded back for updates.
    """
    # Map DB columns → Excel display columns
    export_col_map = {v: k for k, v in COLUMN_MAP.items()}

    # Add computed columns
    from utils import calc_days_remaining, calc_urgency

    out_rows = []
    for _, row in df.iterrows():
        due = row.get("due_date", "")
        days = calc_days_remaining(due)
        urgency = calc_urgency(due)

        out_row = {}
        for db_col, excel_col in export_col_map.items():
            out_row[excel_col] = row.get(db_col, "")

        out_row["Days Remaining"] = days if days is not None else ""
        out_row["Urgency"]        = urgency
        out_rows.append(out_row)

    if not out_rows:
        out_df = pd.DataFrame(columns=list(export_col_map.values()) + ["Days Remaining", "Urgency"])
    else:
        out_df = pd.DataFrame(out_rows)

    # Build styled workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notice Register Export"

    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    green_fill   = PatternFill("solid", fgColor="D5F5E3")
    amber_fill   = PatternFill("solid", fgColor="FDEBD0")
    red_fill     = PatternFill("solid", fgColor="FADBD8")
    over_fill    = PatternFill("solid", fgColor="E8DAEF")
    thin_border  = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    headers = list(out_df.columns)

    # Header row
    ws.row_dimensions[1].height = 36
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border

    # Data rows
    urgency_fill_map = {
        "GREEN":   green_fill,
        "AMBER":   amber_fill,
        "RED":     red_fill,
        "OVERDUE": over_fill,
    }
    urgency_col = headers.index("Urgency") + 1 if "Urgency" in headers else None

    for row_idx, row_data in enumerate(out_rows, start=2):
        urgency_val = row_data.get("Urgency", "")
        row_fill = urgency_fill_map.get(urgency_val)
        ws.row_dimensions[row_idx].height = 20
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(h, ""))
            cell.alignment = center_align if col_idx == urgency_col else left_align
            cell.border    = thin_border
            if row_fill:
                cell.fill = row_fill

    # Auto column widths
    for col_idx, h in enumerate(headers, 1):
        max_len = max(
            len(str(h)),
            *(len(str(r.get(h, ""))) for r in out_rows) if out_rows else [0]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
