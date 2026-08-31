"""Excel export. Build Prompt v2 §11.3.

Export only — §1.1 excludes Excel *import* from v1, deliberately: a
spreadsheet is not a controlled input to a statutory document.
"""

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.engagement import Engagement
from app.models.enums import CommentStatus, EngagementStatus
from app.models.masters import Client, ClientProfile, Firm
from app.models.workflow import ReviewComment

HEADER_FILL = PatternFill("solid", fgColor="1A3260")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)


def _sheet(workbook: Workbook, title: str, headers: tuple[str, ...]) -> Worksheet:
    sheet = workbook.create_sheet(title)
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(column)].width = max(14, len(header) + 4)
    sheet.freeze_panes = "A2"
    return sheet


def _write(sheet: Worksheet, rows: list[tuple[object, ...]]) -> None:
    for index, row in enumerate(rows, start=2):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=index, column=column, value=value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def export_workbook(session: Session, target: Path) -> Path:
    """Client list, engagement status, pending reviews and open findings."""
    firm = session.scalar(select(Firm))
    generated = datetime.now(UTC).strftime("%d-%b-%Y %H:%M UTC")

    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover["A1"] = firm.firm_name if firm else "AuditCraft"
    cover["A1"].font = Font(name="Arial", size=14, bold=True, color="1A3260")
    cover["A2"] = f"FRN {firm.frn}" if firm else ""
    cover["A3"] = f"Generated {generated}"
    cover["A3"].font = Font(name="Arial", size=9, italic=True)
    cover.column_dimensions["A"].width = 60

    # --- clients ---------------------------------------------------------
    clients = _sheet(
        workbook,
        "Clients",
        ("Code", "Name", "CIN", "PAN", "Type", "Framework", "Incorporated"),
    )
    rows = session.execute(
        select(Client, ClientProfile)
        .join(ClientProfile, ClientProfile.client_id == Client.client_id)
        .where(ClientProfile.is_current.is_(True))
        .order_by(ClientProfile.company_name)
    ).all()
    _write(
        clients,
        [
            (
                client.client_code,
                profile.company_name,
                client.cin,
                client.pan,
                profile.company_type.value,
                profile.framework.value,
                client.date_of_incorp,
            )
            for client, profile in rows
        ],
    )

    # --- engagements -----------------------------------------------------
    engagements = _sheet(
        workbook,
        "Engagements",
        ("Client", "FY", "Status", "Opinion", "Going concern", "Report date", "Locked"),
    )
    engagement_rows = session.execute(
        select(Engagement, ClientProfile.company_name)
        .join(Client, Client.client_id == Engagement.client_id)
        .join(
            ClientProfile,
            (ClientProfile.client_id == Client.client_id) & ClientProfile.is_current.is_(True),
        )
        .order_by(ClientProfile.company_name, Engagement.fy_end.desc())
    ).all()
    _write(
        engagements,
        [
            (
                name,
                engagement.fy_code,
                engagement.status.value.replace("_", " "),
                engagement.opinion_type.value if engagement.opinion_type else "",
                engagement.going_concern.value,
                engagement.report_date,
                "yes" if engagement.is_locked else "",
            )
            for engagement, name in engagement_rows
        ],
    )

    # --- pending reviews -------------------------------------------------
    pending = _sheet(workbook, "Pending review", ("Client", "FY", "Status"))
    _write(
        pending,
        [
            (name, engagement.fy_code, engagement.status.value.replace("_", " "))
            for engagement, name in engagement_rows
            # Decision 29 removed the two reviewer states; "pending" is now
            # prepared and not yet approved.
            if engagement.status is EngagementStatus.PREPARED
        ],
    )

    # --- open comments ---------------------------------------------------
    findings = _sheet(
        workbook, "Open comments", ("Engagement", "Field", "Raised by", "Status", "Comment")
    )
    comments = session.scalars(
        select(ReviewComment)
        .where(ReviewComment.status != CommentStatus.RESOLVED)
        .order_by(ReviewComment.comment_id)
    ).all()
    _write(
        findings,
        [
            (
                comment.engagement_id,
                comment.field_key or comment.document or "",
                comment.raised_by,
                comment.status.value,
                comment.body,
            )
            for comment in comments
        ],
    )

    # --- year on year ----------------------------------------------------
    yoy = _sheet(workbook, "Year on year", ("Client", "FY", "Opinion", "Turnover"))
    _write(
        yoy,
        [
            (
                name,
                engagement.fy_code,
                engagement.opinion_type.value if engagement.opinion_type else "",
                None,
            )
            for engagement, name in engagement_rows
        ],
    )

    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(target))
    return target
