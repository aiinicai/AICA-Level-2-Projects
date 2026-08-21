"""Phase 10 — Excel export, backup and retention. §11.3, §11.4."""

from __future__ import annotations

import zipfile
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.masters import Firm
from app.services.excel import export_workbook


class TestExcelExport:
    @pytest.fixture
    def workbook_path(self, db: Session, tmp_path: Path) -> Path:
        return export_workbook(db, tmp_path / "auditcraft.xlsx")

    def test_it_writes_a_readable_workbook(self, workbook_path: Path) -> None:
        assert workbook_path.exists()
        assert load_workbook(workbook_path) is not None

    def test_the_required_sheets_exist(self, workbook_path: Path) -> None:
        # §11.3 names all five.
        sheets = set(load_workbook(workbook_path).sheetnames)
        assert {
            "Clients",
            "Engagements",
            "Pending review",
            "Open comments",
            "Year on year",
        } <= sheets

    def test_the_cover_carries_the_firm_and_a_timestamp(
        self, workbook_path: Path, db: Session
    ) -> None:
        # Asserted against the Firm record rather than a literal. The seeded
        # placeholder used to be "Your Firm Name, Chartered Accountants" and
        # this test matched on the suffix — which broke when the suffix was
        # removed, because every signature block adds "Chartered Accountants"
        # on its own line and a firm name carrying it printed it twice.
        firm = db.scalar(select(Firm))
        assert firm is not None
        cover = load_workbook(workbook_path)["Cover"]
        assert firm.firm_name in str(cover["A1"].value)
        assert "Generated" in str(cover["A3"].value)

    def test_clients_are_listed(self, workbook_path: Path) -> None:
        sheet = load_workbook(workbook_path)["Clients"]
        values = [sheet.cell(row=2, column=c).value for c in range(1, 8)]
        assert values[0] == "ABC001"
        assert values[1] == "ABC Private Limited"

    def test_engagements_are_listed_with_status(self, workbook_path: Path) -> None:
        sheet = load_workbook(workbook_path)["Engagements"]
        codes = {sheet.cell(row=r, column=2).value for r in range(2, 5)}
        assert {"2025-26", "2024-25"} <= codes

    def test_headers_are_frozen(self, workbook_path: Path) -> None:
        assert load_workbook(workbook_path)["Clients"].freeze_panes == "A2"


class TestBackup:
    @pytest.fixture(autouse=True)
    def _isolated_data_dir(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
        """Point the backup at a scratch data directory.

        Backing up the developer's real database during a test run would be
        harmless but slow and surprising.
        """
        from app.config import Settings, get_settings

        data = tmp_path / "data"
        (data / "clients" / "ABC001").mkdir(parents=True)
        (data / "clients" / "ABC001" / "doc.docx").write_bytes(b"docx")

        source_db = data / "auditcraft.db"
        import sqlite3

        with sqlite3.connect(str(source_db)) as connection:
            connection.execute("CREATE TABLE t (a INTEGER)")
            connection.execute("INSERT INTO t VALUES (1)")

        settings = Settings(
            data_dir=data,
            document_dir=data / "documents",
            database_url=f"sqlite:///{source_db}",
            content_dir=Path("content"),
        )
        monkeypatch.setattr("app.config.get_settings", lambda: settings)
        monkeypatch.setattr("scripts.backup.get_settings", lambda: settings)
        get_settings.cache_clear()
        yield settings
        get_settings.cache_clear()

    def test_a_backup_contains_the_database_and_documents(self) -> None:
        from scripts.backup import create_backup

        archive = create_backup()
        assert archive.exists()
        with zipfile.ZipFile(archive) as zipped:
            names = zipped.namelist()
        assert any(name.startswith("database/") for name in names)
        assert any(name.startswith("documents/") for name in names)
        assert any(name.startswith("content/") for name in names)

    def test_the_copied_database_is_usable(self, tmp_path: Path) -> None:
        """A plain file copy under WAL can capture a mid-write database.

        `sqlite3.backup` is used instead; this proves the archived file
        actually opens and holds the data.
        """
        import sqlite3

        from scripts.backup import create_backup

        archive = create_backup()
        with zipfile.ZipFile(archive) as zipped:
            name = next(n for n in zipped.namelist() if n.startswith("database/"))
            zipped.extract(name, tmp_path / "restored")

        restored = tmp_path / "restored" / name
        with sqlite3.connect(str(restored)) as connection:
            assert connection.execute("SELECT a FROM t").fetchone() == (1,)

    def test_retention_prunes_only_old_archives(self, _isolated_data_dir) -> None:
        from scripts.backup import STAMP, create_backup, list_backups, prune

        fresh = create_backup()
        stale_stamp = (datetime.now(UTC) - timedelta(days=45)).strftime(STAMP)
        stale = fresh.parent / f"auditcraft-{stale_stamp}.zip"
        stale.write_bytes(b"old")

        removed = prune()
        assert stale in removed
        assert stale not in list_backups()
        assert fresh.exists()

    def test_an_unrecognised_filename_is_left_alone(self, _isolated_data_dir) -> None:
        from scripts.backup import create_backup, prune

        archive = create_backup()
        stray = archive.parent / "auditcraft-not-a-timestamp.zip"
        stray.write_bytes(b"x")
        prune()
        assert stray.exists()
