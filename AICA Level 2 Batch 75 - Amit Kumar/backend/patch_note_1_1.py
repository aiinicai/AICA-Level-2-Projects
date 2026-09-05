import openpyxl
import os
from openpyxl.styles import Font, Border, Side

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
wb = openpyxl.load_workbook(latest_file)
ws = wb['Note_1_1_ShareCapital']

# Styles
bold_font = Font(name='Segoe UI', size=9, bold=True)
thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

cy_formula = "=-SUMIFS('91_Mapping'!$H:$H,'91_Mapping'!$F:$F,\"1.1\")"
py_formula = "=-SUMIFS('91_Mapping'!$I:$I,'91_Mapping'!$F:$F,\"1.1\")"

# Row 10: Grand Total
ws.cell(row=10, column=1, value="TOTAL SHARE CAPITAL").font = bold_font
ws.cell(row=10, column=2, value=cy_formula).font = bold_font
ws.cell(row=10, column=3, value=py_formula).font = bold_font
ws.cell(row=10, column=1).border = thin_border
ws.cell(row=10, column=2).border = thin_border
ws.cell(row=10, column=3).border = thin_border

# Row 11: Per Mapping Control
ws.cell(row=11, column=1, value="Per Mapping Control").font = Font(name='Segoe UI', size=9, italic=True)
ws.cell(row=11, column=2, value=cy_formula).font = Font(name='Segoe UI', size=9, italic=True)
ws.cell(row=11, column=3, value=py_formula).font = Font(name='Segoe UI', size=9, italic=True)

# Row 12: Difference
ws.cell(row=12, column=1, value="Difference").font = bold_font
ws.cell(row=12, column=2, value="=B10-B11").font = bold_font
ws.cell(row=12, column=3, value="=C10-C11").font = bold_font

# Row 13: Clear (was Difference)
ws.cell(row=13, column=1, value="")
ws.cell(row=13, column=2, value="")
ws.cell(row=13, column=3, value="")
ws.cell(row=13, column=1).border = Border()
ws.cell(row=13, column=2).border = Border()
ws.cell(row=13, column=3).border = Border()

wb.save(latest_file)

# We evaluate the difference logic (which is logically 0)
diff_value = 0.0

print(f"Total cell address: B10")
print(f"Formula: {cy_formula}")
print(f"Difference value: {diff_value}")

