import openpyxl
import os

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
wb = openpyxl.load_workbook(latest_file)
ws_bs = wb['02_Balance_Sheet']

# Find Share Capital row
target_row = None
for r in range(1, ws_bs.max_row + 1):
    val_b = str(ws_bs.cell(row=r, column=2).value).strip()
    val_a = str(ws_bs.cell(row=r, column=1).value).strip()
    if 'Share Capital' in val_b or 'Share capital' in val_b or 'Share Capital' in val_a:
        # Avoid matching the main header or totals if there are any
        if 'TOTAL' not in val_b.upper() and 'TOTAL' not in val_a.upper():
            target_row = r
            break

if target_row:
    # Assuming C is CY and D is PY based on standard Schedule III
    # Let's verify by checking headers or just assume C/D
    # Let's write the formulas
    cy_formula = "='Note_1_1_ShareCapital'!B10"
    py_formula = "='Note_1_1_ShareCapital'!C10"
    
    ws_bs.cell(row=target_row, column=3).value = cy_formula
    ws_bs.cell(row=target_row, column=4).value = py_formula
    
    wb.save(latest_file)
    
    print("Source sheet\nNote_1_1_ShareCapital")
    print("Source cell\nB10")
    print("Destination sheet\n02_Balance_Sheet")
    print(f"Destination cell\nC{target_row}")
    print(f"Formula\n{cy_formula}")
else:
    print("Share Capital row not found in Balance Sheet.")
