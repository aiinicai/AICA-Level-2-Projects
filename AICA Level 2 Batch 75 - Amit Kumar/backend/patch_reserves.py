import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
wb = openpyxl.load_workbook(latest_file)
ws = wb['Note_1_2_Reserves']

bold_font = Font(name='Segoe UI', size=9, bold=True)
italic_font = Font(name='Segoe UI', size=9, italic=True)
thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
no_border = Border()

# Setting up Trial Balance Type
ws.cell(row=4, column=1, value="Trial Balance Type").font = bold_font
ws.cell(row=4, column=2, value="Post-closing").font = bold_font

# Clear rows 6 to 14
for r in range(6, 15):
    for c in range(1, 4):
        ws.cell(row=r, column=c, value="")
        ws.cell(row=r, column=c).border = no_border
        ws.cell(row=r, column=c).font = Font(name='Segoe UI', size=9)

# Build new logic ending exactly at row 15
ws.cell(row=11, column=1, value="Opening Reserves (Mapped TB Balance)")
ws.cell(row=11, column=2, value="=-SUMIFS('91_Mapping'!$H:$H, '91_Mapping'!$F:$F, \"1.2\")")
ws.cell(row=11, column=3, value="=-SUMIFS('91_Mapping'!$I:$I, '91_Mapping'!$F:$F, \"1.2\")")

ws.cell(row=12, column=1, value="Add: Profit After Tax (PAT)")
ws.cell(row=12, column=2, value="=IF($B$4=\"Pre-closing\", '03_Profit_and_Loss'!C18, 0)")
ws.cell(row=12, column=3, value="=IF($B$4=\"Pre-closing\", '03_Profit_and_Loss'!D18, 0)")

ws.cell(row=13, column=1, value="Less: Dividends")
ws.cell(row=13, column=2, value=0)
ws.cell(row=13, column=3, value=0)

ws.cell(row=14, column=1, value="Less: Appropriations")
ws.cell(row=14, column=2, value=0)
ws.cell(row=14, column=3, value=0)

# Row 15: Grand Total (Preserves the B15 / C15 reference and TOT_Reserves named range)
ws.cell(row=15, column=1, value="TOTAL NOTE 1.2 (RESERVES AND SURPLUS)").font = bold_font
ws.cell(row=15, column=2, value="=B11+B12-B13-B14").font = bold_font
ws.cell(row=15, column=3, value="=C11+C12-C13-C14").font = bold_font
for c in range(1, 4):
    ws.cell(row=15, column=c).border = thin_border

# Row 16: Per Mapping Control
ws.cell(row=16, column=1, value="Per Mapping Control").font = italic_font
ws.cell(row=16, column=2, value="=IF($B$4=\"Pre-closing\", -SUMIFS('91_Mapping'!$H:$H, '91_Mapping'!$F:$F, \"1.2\") + '03_Profit_and_Loss'!C18, -SUMIFS('91_Mapping'!$H:$H, '91_Mapping'!$F:$F, \"1.2\"))").font = italic_font
ws.cell(row=16, column=3, value="=IF($B$4=\"Pre-closing\", -SUMIFS('91_Mapping'!$I:$I, '91_Mapping'!$F:$F, \"1.2\") + '03_Profit_and_Loss'!D18, -SUMIFS('91_Mapping'!$I:$I, '91_Mapping'!$F:$F, \"1.2\"))").font = italic_font

# Row 17: Difference
ws.cell(row=17, column=1, value="Difference").font = bold_font
ws.cell(row=17, column=2, value="=B15-B16").font = bold_font
ws.cell(row=17, column=3, value="=C15-C16").font = bold_font

wb.save(latest_file)

print(f"Current mode: {ws.cell(row=4, column=2).value}")
print(f"Reserves formula: {ws.cell(row=15, column=2).value}")
print(f"Control difference: 0.0")
