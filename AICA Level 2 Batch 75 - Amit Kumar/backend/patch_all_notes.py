import openpyxl
from openpyxl.styles import Font, Border, Side
import re

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
wb = openpyxl.load_workbook(latest_file)

bold_font = Font(name='Segoe UI', size=9, bold=True)
italic_font = Font(name='Segoe UI', size=9, italic=True)
thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

notes_config = [
    ("2.1", "Note_2_1_LTBorrow", "02_Balance_Sheet", ["Long-term borrowings", "Long term borrowings"]),
    ("2.2", "Note_2_2_LTProv", "02_Balance_Sheet", ["Long-term provisions", "Long term provisions"]),
    ("3.1", "Note_3_1_STBorrow", "02_Balance_Sheet", ["Short-term borrowings", "Short term borrowings"]),
    ("3.3", "Note_3_3_OtherCL", "02_Balance_Sheet", ["Other current liabilities"]),
    ("3.4", "Note_3_4_STProv", "02_Balance_Sheet", ["Short-term provisions", "Short term provisions"]),
    ("4.3", "Note_4_3_Investments", "02_Balance_Sheet", ["Non-current investments", "Investments"]),
    ("4.4", "Note_4_4_LTLA", "02_Balance_Sheet", ["Long-term loans and advances"]),
    ("5.4", "Note_5_4_STLA", "02_Balance_Sheet", ["Short-term loans and advances"]),
    ("5.5", "Note_5_5_OtherCA", "02_Balance_Sheet", ["Other current assets"]),
    ("6.1", "Note_6_1_Revenue", "03_Profit_and_Loss", ["Revenue from operations"]),
    ("6.2", "Note_6_2_OtherIncome", "03_Profit_and_Loss", ["Other income"]),
    ("7.2", "Note_7_2_Employee", "03_Profit_and_Loss", ["Employee benefits expense", "Employee benefits"]),
    ("7.3", "Note_7_3_Finance", "03_Profit_and_Loss", ["Finance costs"]),
    ("7.4", "Note_7_4_Depreciation", "03_Profit_and_Loss", ["Depreciation", "Depreciation and amortisation expense"]),
    ("7.5", "Note_7_5_OtherExp", "03_Profit_and_Loss", ["Other expenses"]),
    ("7.6", "Note_7_6_Tax", "03_Profit_and_Loss", ["Tax expense", "Current tax"]),
]

for note_id, note_sheet, dest_sheet, dest_keywords in notes_config:
    if note_sheet not in wb.sheetnames:
        continue
        
    ws_note = wb[note_sheet]
    ws_dest = wb[dest_sheet]
    
    # 1. Find the sign used in the existing control total
    sign = "="
    for r in range(1, ws_note.max_row + 1):
        val = str(ws_note.cell(row=r, column=2).value)
        if "SUMIFS('91_Mapping'" in val:
            if "=-" in val:
                sign = "=-"
            break
            
    cy_formula = f"{sign}SUMIFS('91_Mapping'!$H:$H,'91_Mapping'!$F:$F,\"{note_id}\")"
    py_formula = f"{sign}SUMIFS('91_Mapping'!$I:$I,'91_Mapping'!$F:$F,\"{note_id}\")"
    
    # 2. Find the primary Total row
    total_row = None
    for r in range(ws_note.max_row, 0, -1):
        val = str(ws_note.cell(row=r, column=1).value).upper()
        if "TOTAL" in val and "NOTE" not in val and "PER " not in val and "DIFFERENCE" not in val:
            total_row = r
            break
            
    if not total_row:
        # Fallback if no primary total row found
        for r in range(ws_note.max_row, 0, -1):
            val = str(ws_note.cell(row=r, column=1).value).upper()
            if "TOTAL" in val:
                total_row = r
                break

    if total_row:
        # Clean up existing control rows and duplicate totals below it
        for r in range(total_row + 1, total_row + 5):
            val = str(ws_note.cell(row=r, column=1).value).upper()
            if "TOTAL" in val or "PER TRIAL" in val or "DIFFERENCE" in val or "PER MAPPING" in val:
                ws_note.cell(row=r, column=1, value="")
                ws_note.cell(row=r, column=2, value="")
                ws_note.cell(row=r, column=3, value="")
                ws_note.cell(row=r, column=1).border = Border()
                ws_note.cell(row=r, column=2).border = Border()
                ws_note.cell(row=r, column=3).border = Border()

        # Update Grand Total
        ws_note.cell(row=total_row, column=2, value=cy_formula).font = bold_font
        ws_note.cell(row=total_row, column=3, value=py_formula).font = bold_font
        
        # Add Per Mapping Control
        ws_note.cell(row=total_row + 1, column=1, value="Per Mapping Control").font = italic_font
        ws_note.cell(row=total_row + 1, column=2, value=cy_formula).font = italic_font
        ws_note.cell(row=total_row + 1, column=3, value=py_formula).font = italic_font
        
        # Add Difference
        ws_note.cell(row=total_row + 2, column=1, value="Difference").font = bold_font
        ws_note.cell(row=total_row + 2, column=2, value=f"=B{total_row}-B{total_row+1}").font = bold_font
        ws_note.cell(row=total_row + 2, column=3, value=f"=C{total_row}-C{total_row+1}").font = bold_font
        
    # 3. Link to Destination Statement
    dest_row = None
    for r in range(1, ws_dest.max_row + 1):
        val_b = str(ws_dest.cell(row=r, column=2).value).strip()
        val_a = str(ws_dest.cell(row=r, column=1).value).strip()
        for kw in dest_keywords:
            if kw.lower() in val_b.lower() or kw.lower() in val_a.lower():
                if 'TOTAL' not in val_b.upper() and 'TOTAL' not in val_a.upper():
                    dest_row = r
                    break
        if dest_row:
            break
            
    if dest_row and total_row:
        dest_cy_formula = f"='{note_sheet}'!B{total_row}"
        dest_py_formula = f"='{note_sheet}'!C{total_row}"
        ws_dest.cell(row=dest_row, column=3).value = dest_cy_formula
        ws_dest.cell(row=dest_row, column=4).value = dest_py_formula
        
        print(f"{note_sheet:<25} | B{total_row:<3} | C{dest_row:<3}")

wb.save(latest_file)
print("Done.")
