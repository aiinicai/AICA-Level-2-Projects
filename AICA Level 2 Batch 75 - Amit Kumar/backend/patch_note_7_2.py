import openpyxl
from openpyxl.styles import Font, Border, Side

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
wb = openpyxl.load_workbook(latest_file)

bold_font = Font(name='Segoe UI', size=9, bold=True)
italic_font = Font(name='Segoe UI', size=9, italic=True)

note_id = "7.2"
note_sheet = "Note_7_2_Employee"
dest_sheet = "03_Profit_and_Loss"

ws_note = wb[note_sheet]
ws_dest = wb[dest_sheet]

sign = "="
for r in range(1, ws_note.max_row + 1):
    val = str(ws_note.cell(row=r, column=2).value)
    if "SUMIFS('91_Mapping'" in val:
        if "=-" in val:
            sign = "=-"
        break
        
cy_formula = f"{sign}SUMIFS('91_Mapping'!$H:$H,'91_Mapping'!$F:$F,\"{note_id}\")"
py_formula = f"{sign}SUMIFS('91_Mapping'!$I:$I,'91_Mapping'!$F:$F,\"{note_id}\")"

total_row = None
for r in range(ws_note.max_row, 0, -1):
    val = str(ws_note.cell(row=r, column=1).value).upper()
    if "TOTAL" in val and "NOTE" not in val and "PER " not in val and "DIFFERENCE" not in val:
        total_row = r
        break

if total_row:
    # Clean up
    for r in range(total_row + 1, total_row + 5):
        val = str(ws_note.cell(row=r, column=1).value).upper()
        if "TOTAL" in val or "PER TRIAL" in val or "DIFFERENCE" in val or "PER MAPPING" in val:
            ws_note.cell(row=r, column=1, value="")
            ws_note.cell(row=r, column=2, value="")
            ws_note.cell(row=r, column=3, value="")
            ws_note.cell(row=r, column=1).border = Border()
            ws_note.cell(row=r, column=2).border = Border()
            ws_note.cell(row=r, column=3).border = Border()

    # Update Grand Total
    ws_note.cell(row=total_row, column=2, value=cy_formula).font = bold_font
    ws_note.cell(row=total_row, column=3, value=py_formula).font = bold_font
    
    ws_note.cell(row=total_row + 1, column=1, value="Per Mapping Control").font = italic_font
    ws_note.cell(row=total_row + 1, column=2, value=cy_formula).font = italic_font
    ws_note.cell(row=total_row + 1, column=3, value=py_formula).font = italic_font
    
    ws_note.cell(row=total_row + 2, column=1, value="Difference").font = bold_font
    ws_note.cell(row=total_row + 2, column=2, value=f"=B{total_row}-B{total_row+1}").font = bold_font
    ws_note.cell(row=total_row + 2, column=3, value=f"=C{total_row}-C{total_row+1}").font = bold_font
    
    # Dest is row 11 in P&L
    dest_row = 11
    ws_dest.cell(row=dest_row, column=3).value = f"='{note_sheet}'!B{total_row}"
    ws_dest.cell(row=dest_row, column=4).value = f"='{note_sheet}'!C{total_row}"
    
    print(f"{note_sheet:<25} | B{total_row:<3} | C{dest_row:<3}")

wb.save(latest_file)
print("Done.")
