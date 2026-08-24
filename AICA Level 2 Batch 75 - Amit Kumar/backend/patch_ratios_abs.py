import os
import openpyxl

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'

# 1. Write the formulas using openpyxl
wb = openpyxl.load_workbook(latest_file)
ws = wb['07_Financial_Ratios']

# Note: Using ABS to ensure ratios display as positive standard figures
# Current Ratio (Row 5)
f_cr_cy = "=ABS('02_Balance_Sheet'!C36/'02_Balance_Sheet'!C20)"
ws.cell(row=5, column=4, value=f_cr_cy)

# Debt Equity Ratio (Row 6)
f_de_cy = "=ABS(('02_Balance_Sheet'!C12+'02_Balance_Sheet'!C16)/('02_Balance_Sheet'!C8+'02_Balance_Sheet'!C9))"
ws.cell(row=6, column=4, value=f_de_cy)

# Net Profit Ratio (Row 7)
f_npr_cy = "='03_Profit_and_Loss'!C18 / ABS('03_Profit_and_Loss'!C6)"
ws.cell(row=7, column=4, value=f_npr_cy)

# ROE (Row 9)
f_roe_cy = "='03_Profit_and_Loss'!C18 / ABS('02_Balance_Sheet'!C8+'02_Balance_Sheet'!C9)"
ws.cell(row=9, column=4, value=f_roe_cy)

# Trade Receivable Days (Row 10)
f_trd_cy = "=ABS('02_Balance_Sheet'!C32 / '03_Profit_and_Loss'!C6) * 365"
ws.cell(row=10, column=4, value=f_trd_cy)

# Trade Payable Days (Row 11)
f_tpd_cy = "=ABS('02_Balance_Sheet'!C17 / '03_Profit_and_Loss'!C10) * 365"
ws.cell(row=11, column=4, value=f_tpd_cy)

# Inventory Days (Row 12)
f_invd_cy = "=ABS('02_Balance_Sheet'!C31 / '03_Profit_and_Loss'!C10) * 365"
ws.cell(row=12, column=4, value=f_invd_cy)

wb.save(latest_file)

# 2. Evaluate with win32com and print
try:
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    full_path = os.path.abspath(latest_file)
    wb_com = excel.Workbooks.Open(full_path)
    ws_com = wb_com.Sheets('07_Financial_Ratios')
    
    ratios_to_print = [
        ("Current Ratio", f_cr_cy, ws_com.Range("D5").Value),
        ("Debt Equity", f_de_cy, ws_com.Range("D6").Value),
        ("ROE", f_roe_cy, ws_com.Range("D9").Value),
        ("Trade Receivable Days", f_trd_cy, ws_com.Range("D10").Value),
        ("Trade Payable Days", f_tpd_cy, ws_com.Range("D11").Value),
        ("Inventory Days", f_invd_cy, ws_com.Range("D12").Value),
        ("Net Profit Ratio", f_npr_cy, ws_com.Range("D7").Value),
    ]
    
    wb_com.Close(False)
    excel.Quit()
    
    # Print formatted output
    print(f"{'Ratio':<25} | {'Formula':<75} | {'Value'}")
    print("-" * 120)
    for name, f, val in ratios_to_print:
        # Format the value if it's a number
        if isinstance(val, (int, float)):
            v_str = f"{val:.2f}"
            if name in ["ROE", "Net Profit Ratio"]:
                v_str = f"{val*100:.2f}%"
        else:
            v_str = str(val)
        print(f"{name:<25} | {f:<75} | {v_str}")
        
except Exception as e:
    print(f"win32com evaluation failed: {e}")
