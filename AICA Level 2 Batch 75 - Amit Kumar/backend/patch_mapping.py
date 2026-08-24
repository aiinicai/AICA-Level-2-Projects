import openpyxl
import glob
import os
import sys

# Find the latest Financial Statements xlsx
latest_file = r"c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx"
print(f'Modifying {os.path.basename(latest_file)}...')

wb = openpyxl.load_workbook(latest_file)
print("Available sheets:", wb.sheetnames)
if '91_Mapping' not in wb.sheetnames:
    print('91_Mapping not found in workbook.')
    sys.exit(0)

ws = wb['91_Mapping']
print(f'{"Row":<5} | {"Ledger":<40} | {"CY Formula":<60} | {"PY Formula":<60}')
print('-' * 170)

for row in range(2, ws.max_row + 1):
    ledger = ws.cell(row=row, column=1).value
    if ledger:
        cy_formula = f"=SUMIFS('90_Trial_Balance'!$D:$D,'90_Trial_Balance'!$A:$A,$A{row})"
        py_formula = f"=SUMIFS('90_Trial_Balance'!$E:$E,'90_Trial_Balance'!$A:$A,$A{row})"
        
        ws.cell(row=row, column=8).value = cy_formula
        ws.cell(row=row, column=9).value = py_formula
        
        print(f'{row:<5} | {str(ledger)[:38]:<40} | {cy_formula:<60} | {py_formula:<60}')

wb.save(latest_file)
print('\nDone.')
