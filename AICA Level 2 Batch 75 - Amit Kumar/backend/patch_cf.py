import openpyxl

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
wb = openpyxl.load_workbook(latest_file)
ws_cf = wb['04_Cash_Flow_Statement']
ws_adj = wb['29_Cash_Flow_Adjustments']

# 1. Zero out auto-plugs in 29_Cash_Flow_Adjustments
for r in [2, 3, 4, 5, 6]:
    ws_adj.cell(row=r, column=3).value = 0
    ws_adj.cell(row=r, column=4).value = 0

# 2. Update 04_Cash_Flow_Statement to link directly from BS and P&L
# Depreciation
ws_cf.cell(row=9, column=2).value = "='03_Profit_and_Loss'!C13"
ws_cf.cell(row=9, column=3).value = "='03_Profit_and_Loss'!D13"

# Finance Cost
ws_cf.cell(row=10, column=2).value = "='03_Profit_and_Loss'!C12"
ws_cf.cell(row=10, column=3).value = "='03_Profit_and_Loss'!D12"

# Working Capital
ws_cf.cell(row=14, column=2).value = "=-('02_Balance_Sheet'!C31 - '02_Balance_Sheet'!D31)"
ws_cf.cell(row=15, column=2).value = "=-('02_Balance_Sheet'!C32 - '02_Balance_Sheet'!D32)"
ws_cf.cell(row=16, column=2).value = "='02_Balance_Sheet'!C17 - '02_Balance_Sheet'!D17"
ws_cf.cell(row=17, column=2).value = "='02_Balance_Sheet'!C18 - '02_Balance_Sheet'!D18"
ws_cf.cell(row=18, column=2).value = "='02_Balance_Sheet'!C19 - '02_Balance_Sheet'!D19"

# Closing Cash per BS
ws_cf.cell(row=39, column=2).value = "='02_Balance_Sheet'!C33"
ws_cf.cell(row=39, column=3).value = "='02_Balance_Sheet'!D33"

wb.save(latest_file)

# Evaluate with win32com to get values
try:
    import win32com.client
    import os
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    full_path = os.path.abspath(latest_file)
    wb_com = excel.Workbooks.Open(full_path)
    ws_com = wb_com.Sheets('04_Cash_Flow_Statement')
    
    closing_bs = ws_com.Range("B39").Value
    computed_closing = ws_com.Range("B38").Value
    diff = ws_com.Range("B40").Value
    
    wb_com.Close(False)
    excel.Quit()
    
    print(f"Closing cash per BS: {closing_bs}")
    print(f"Computed closing cash: {computed_closing}")
    print(f"Difference: {diff}")
    
except Exception as e:
    print(f"win32com evaluation failed: {e}")
    print("Closing cash per BS: cell B39")
    print("Computed closing cash: cell B38")
    print("Difference: cell B40")
