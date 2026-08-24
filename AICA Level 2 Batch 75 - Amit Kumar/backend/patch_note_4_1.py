import openpyxl
from openpyxl.styles import Font, Border, Side

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
wb = openpyxl.load_workbook(latest_file)
ws_note = wb['Note_4_1_PPE']
ws_bs = wb['02_Balance_Sheet']

bold_font = Font(name='Segoe UI', size=9, bold=True)
italic_font = Font(name='Segoe UI', size=9, italic=True)

# Row 13: Supporting Schedule Total
ws_note.cell(row=13, column=1, value="Supporting Schedule Total").font = bold_font
ws_note.cell(row=13, column=2, value="=J12").font = bold_font
ws_note.cell(row=13, column=3, value="=K12").font = bold_font

# Row 14: Per Trial Balance Total
ws_note.cell(row=14, column=1, value="Per Trial Balance Total").font = bold_font
ws_note.cell(row=14, column=2, value="=SUMIFS('91_Mapping'!$H:$H, '91_Mapping'!$F:$F, \"4.1\")").font = bold_font
ws_note.cell(row=14, column=3, value="=SUMIFS('91_Mapping'!$I:$I, '91_Mapping'!$F:$F, \"4.1\")").font = bold_font

# Row 15: Difference
ws_note.cell(row=15, column=1, value="Difference").font = bold_font
ws_note.cell(row=15, column=2, value="=B13-B14").font = bold_font
ws_note.cell(row=15, column=3, value="=C13-C14").font = bold_font

# Row 16: Status
ws_note.cell(row=16, column=1, value="Status").font = bold_font
ws_note.cell(row=16, column=2, value="=IF(ABS(B15)>0.01, \"REVIEW REQUIRED\", \"TALLIED\")").font = bold_font
ws_note.cell(row=16, column=3, value="=IF(ABS(C15)>0.01, \"REVIEW REQUIRED\", \"TALLIED\")").font = bold_font

# Update Balance Sheet to use TB control total (Row 14) instead of Supporting Schedule
dest_row = None
for r in range(1, ws_bs.max_row + 1):
    val_b = str(ws_bs.cell(row=r, column=2).value).strip()
    if 'Property, plant and equipment' in val_b or 'Property, Plant and Equipment' in val_b:
        dest_row = r
        break

if dest_row:
    ws_bs.cell(row=dest_row, column=3).value = "='Note_4_1_PPE'!B14"
    ws_bs.cell(row=dest_row, column=4).value = "='Note_4_1_PPE'!C14"

wb.save(latest_file)

print("FAR total: B13")
print("TB total: B14")
print("Difference: B15")
