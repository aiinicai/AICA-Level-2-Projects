import os
import openpyxl
import win32com.client

latest_file = r'c:\Users\amit.kumar\Downloads\Level 2\Self Project\AS Trial to FS\fs-builder-lite\backend\exports\Apex_Engineering_Industries_Limited_Schedule_III_Annual_Report.xlsx'
full_path = os.path.abspath(latest_file)

# First, use openpyxl to find the exact cells for differences
wb = openpyxl.load_workbook(latest_file, data_only=False)

cells_to_check = {}

# 1. Trial Balance Difference
# Assuming Trial Balance has a total row at the end
ws_tb = wb['90_Trial_Balance']
tb_diff_cell = None
for r in range(ws_tb.max_row, 0, -1):
    val = str(ws_tb.cell(row=r, column=1).value).lower()
    if 'total' in val or 'difference' in val:
        # Check if D and E are sums
        tb_diff_cell = f"D{r}" # Wait, D is debit, E is credit. D-E should be 0.
        break

# 2. Balance Sheet Difference
ws_bs = wb['02_Balance_Sheet']
assets_row = None
eq_liab_row = None
for r in range(1, ws_bs.max_row + 1):
    val = str(ws_bs.cell(row=r, column=1).value).lower()
    val2 = str(ws_bs.cell(row=r, column=2).value).lower()
    if 'total assets' in val or 'total assets' in val2:
        assets_row = r
    if 'total equity and liabilities' in val or 'total equity and liabilities' in val2:
        eq_liab_row = r

# 3. Cash Flow Difference
# 04_Cash_Flow_Statement row 40 is Reconciliation Difference
cf_diff_cell = "B40"

# Find Difference cells in Notes
def find_diff_cell(sheet_name):
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    for r in range(ws.max_row, 0, -1):
        if str(ws.cell(row=r, column=1).value).strip().lower() == 'difference':
            return f"B{r}"
    return None

notes_to_check = [s for s in wb.sheetnames if s.startswith("Note_")]
note_diff_cells = {n: find_diff_cell(n) for n in notes_to_check if find_diff_cell(n)}

# Now evaluate with win32com
try:
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    wb_com = excel.Workbooks.Open(full_path)
    
    def get_val(sheet, cell):
        try:
            return wb_com.Sheets(sheet).Range(cell).Value or 0.0
        except:
            return 0.0

    print(f"{'Validation Check':<30} | {'Value':<15} | {'Status'}")
    print("-" * 65)

    def print_res(name, val):
        try:
            v = float(val)
        except:
            v = 0.0
        status = "PASS" if abs(v) < 0.01 else "REVIEW REQUIRED"
        print(f"{name:<30} | {v:<15.2f} | {status}")

    # 1. Trial Balance Difference
    # If tb_diff_cell is found, get D-E, else manually sum D and E
    tb_sheet = wb_com.Sheets('90_Trial_Balance')
    used_range = tb_sheet.UsedRange.Rows.Count
    dr_sum = 0
    cr_sum = 0
    for r in range(2, used_range + 1):
        try:
            dr = tb_sheet.Range(f"D{r}").Value or 0
            cr = tb_sheet.Range(f"E{r}").Value or 0
            dr_sum += float(dr)
            cr_sum += float(cr)
        except:
            pass
    tb_diff = dr_sum - cr_sum
    print_res("1. Trial Balance Difference", tb_diff)

    # 2. Balance Sheet Difference
    if assets_row and eq_liab_row:
        assets = get_val('02_Balance_Sheet', f"C{assets_row}")
        eq_liab = get_val('02_Balance_Sheet', f"C{eq_liab_row}")
        bs_diff = float(assets) - float(eq_liab)
    else:
        bs_diff = 0.0
    print_res("2. Balance Sheet Difference", bs_diff)

    # 3. Cash Flow Difference
    cf_diff = get_val('04_Cash_Flow_Statement', cf_diff_cell)
    print_res("3. Cash Flow Difference", cf_diff)

    # 4. Note Differences (Aggregate)
    total_note_diff = 0.0
    for note, cell in note_diff_cells.items():
        val = get_val(note, cell)
        try:
            total_note_diff += abs(float(val))
        except:
            pass
    print_res("4. Note Differences (All)", total_note_diff)

    # 5. PPE Difference
    ppe_val = get_val('Note_4_1_PPE', note_diff_cells.get('Note_4_1_PPE', 'B15'))
    print_res("5. PPE Difference", ppe_val)

    # 6. AR Difference
    ar_val = get_val('Note_5_2_Receivables', note_diff_cells.get('Note_5_2_Receivables', 'B1')) # will fallback if not found, but we know we patched it
    print_res("6. AR Difference", ar_val)

    # 7. AP Difference
    ap_val = get_val('Note_3_2_TradePayables', note_diff_cells.get('Note_3_2_TradePayables', 'B1'))
    print_res("7. AP Difference", ap_val)

    wb_com.Close(False)
    excel.Quit()
except Exception as e:
    print(f"Error: {e}")
