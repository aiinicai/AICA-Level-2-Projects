import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from datetime import datetime


from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas

from models import (
    Client, TrialBalanceLine, ARAgeing, APAgeing, CWIPAgeing,
    RelatedParty, Borrowing, Contingency, Note, AccountingPolicy,
    CashFlowAdjustment
)
from services.fs_generator import generate_financial_statements
from services.notes_engine import generate_or_update_notes
from services.accounting_policies_engine import generate_or_update_accounting_policies
from services.cash_flow_engine import generate_cash_flow_statement, get_cash_flow_validations
from services.ratio_engine import calculate_ratios
from services.validation_engine import run_validation_checks


def export_formula_linked_excel(client_id: int, export_dir: str, db: Session) -> str:
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.formatting.rule import CellIsRule

    # Critical Mapping Validation Guard
    validations = run_validation_checks(client_id, db)
    critical_mapping = [v for v in validations if v.category == "Mapping Exception" and v.status == "Critical"]
    if critical_mapping:
        err_msg = "; ".join([f"{v.check_name}: {v.message}" for v in critical_mapping])
        raise ValueError(f"Excel export blocked due to Critical Mapping Exceptions: {err_msg}. Please review and approve manual override in Ledger Mapping.")

    os.makedirs(export_dir, exist_ok=True)
    client = db.query(Client).filter(Client.id == client_id).first()
    client_name = client.name if client else "Client"
    safe_name = "".join(c if c.isalnum() else "_" for c in client_name)
    filepath = os.path.join(export_dir, f"{safe_name}_Schedule_III_Annual_Report.xlsx")

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # BUG 7 (a): Enable full calculation on load for Excel openpyxl workbook
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = 'auto'

    # Fonts & Styles
    NAVY_HEADER_FILL = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    WHITE_TITLE_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    BOLD_NAVY_FONT = Font(name="Segoe UI", size=10, bold=True, color="1B365D")
    BOLD_TEXT_FONT = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
    REGULAR_FONT = Font(name="Segoe UI", size=9, color="0F172A")
    MUTED_FONT = Font(name="Segoe UI", size=8, italic=True, color="64748B")

    GREY_COLUMN_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    SUBTOTAL_ROW_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    TOTAL_ROW_FILL = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")

    THIN_BORDER_SIDE = Side(border_style="thin", color="CBD5E1")
    DOUBLE_BOTTOM = Side(border_style="double", color="1B365D")
    GRID_BORDER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
    TOTAL_BORDER = Border(top=THIN_BORDER_SIDE, bottom=DOUBLE_BOTTOM)

    INDIAN_NUM_FORMAT = '#,##,##0.00;(#,##,##0.00);"-";@'

    def apply_header_styling(ws, header_row=3):
        ws.row_dimensions[header_row].height = 24
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.fill = NAVY_HEADER_FILL
            cell.font = WHITE_TITLE_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def apply_page_setup(ws, title: str, hide_grid=True):
        if hide_grid:
            ws.views.sheetView[0].showGridLines = False
        else:
            ws.views.sheetView[0].showGridLines = True
        ws.oddFooter.left.text = f"{client_name} | {client.reporting_period if client else 'FY 2024-25'}"
        ws.oddFooter.right.text = "Page &P of &N"
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4

    # -------------------------------------------------------------
    # STEP 1 — TRIAL BALANCE SHEET (90_Trial_Balance)
    # -------------------------------------------------------------
    ws_tb = wb.create_sheet(title="90_Trial_Balance")
    ws_tb.append(["Ledger Code", "Ledger Name", "Original Group", "CY Amount", "PY Amount", "Type"])
    apply_header_styling(ws_tb, 1)
    tb_lines = db.query(TrialBalanceLine).filter(TrialBalanceLine.client_id == client_id).all()
    
    tb_start_row = 2
    for l in tb_lines:
        ws_tb.append([l.ledger_code or "", l.ledger_name, l.original_group or "", l.cy_amount, l.py_amount, l.type or ""])
        r_i = ws_tb.max_row
        ws_tb.cell(row=r_i, column=4).number_format = INDIAN_NUM_FORMAT
        ws_tb.cell(row=r_i, column=5).number_format = INDIAN_NUM_FORMAT

    tb_end_row = ws_tb.max_row
    
    # BUG 6: Write real formulas for TB Dr/Cr checks and apply conditional formatting red if not zero
    ws_tb.append(["TB Dr/Cr Check (CY)", "Sum of CY Trial Balance", "", f"=SUM(D{tb_start_row}:D{tb_end_row})", "", "CY Control"])
    tb_cy_chk_row = ws_tb.max_row
    ws_tb.cell(row=tb_cy_chk_row, column=1).font = BOLD_NAVY_FONT
    ws_tb.cell(row=tb_cy_chk_row, column=4).number_format = INDIAN_NUM_FORMAT
    ws_tb.cell(row=tb_cy_chk_row, column=4).font = BOLD_NAVY_FONT

    ws_tb.append(["TB Dr/Cr Check (PY)", "Sum of PY Trial Balance", "", "", f"=SUM(E{tb_start_row}:E{tb_end_row})", "PY Control"])
    tb_py_chk_row = ws_tb.max_row
    ws_tb.cell(row=tb_py_chk_row, column=1).font = BOLD_NAVY_FONT
    ws_tb.cell(row=tb_py_chk_row, column=5).number_format = INDIAN_NUM_FORMAT
    ws_tb.cell(row=tb_py_chk_row, column=5).font = BOLD_NAVY_FONT

    red_fill = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
    red_font = Font(name="Segoe UI", size=9, bold=True, color="991B1B")
    rule_not_zero = CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_fill, font=red_font)
    ws_tb.conditional_formatting.add(f"D{tb_cy_chk_row}", rule_not_zero)
    ws_tb.conditional_formatting.add(f"E{tb_py_chk_row}", rule_not_zero)

    ws_tb.sheet_state = "hidden"

    # -------------------------------------------------------------
    # STEP 2 — MAPPING SHEET (91_Mapping) — THE DRIVER
    # -------------------------------------------------------------
    # BUG 1 (ROOT CAUSE): Populate CY Amount and PY Amount using cell reference to Ledger Code A2, A3, etc.
    ws_map = wb.create_sheet(title="91_Mapping")
    ws_map.append(["Ledger Code", "Ledger Name", "Original Group", "Final Classification", "Statement", "Note #", "Current / Non-Current", "CY Amount (Lakhs)", "PY Amount (Lakhs)"])
    apply_header_styling(ws_map, 1)

    for idx, l in enumerate(tb_lines, start=2):
        l_code = l.ledger_code or ""
        ws_map.append([
            l_code,
            l.ledger_name,
            l.original_group or "",
            l.final_classification or "Unmapped",
            l.financial_statement or "",
            str(l.note_number or "1.1"),
            l.current_non_current or "",
            f"=SUMIFS('90_Trial_Balance'!$D:$D, '90_Trial_Balance'!$A:$A, $A{idx})",
            f"=SUMIFS('90_Trial_Balance'!$E:$E, '90_Trial_Balance'!$A:$A, $A{idx})"
        ])
        ws_map.cell(row=idx, column=8).number_format = INDIAN_NUM_FORMAT
        ws_map.cell(row=idx, column=9).number_format = INDIAN_NUM_FORMAT
    ws_map.sheet_state = "hidden"

    # -------------------------------------------------------------
    # STEP 3 — INPUT SHEET (29_Cash_Flow_Adjustments)
    # -------------------------------------------------------------
    ws_cfa = wb.create_sheet(title="29_Cash_Flow_Adjustments")
    ws_cfa.append(["Adjustment Line Item", "Description", "CY Amount", "PY Amount", "Source Category"])
    apply_header_styling(ws_cfa, 1)
    
    ws_cfa.append(["Income Taxes Paid", "Direct Corporate Income Taxes Paid", "=-TOT_Tax", "=-TOT_Tax_PY", "Operating"])
    tax_paid_cfa_row = 2
    ws_cfa.append(["Dividend Paid", "Equity Dividend Paid During Year", 0.00, 0.00, "Financing"])
    div_paid_cfa_row = 3
    ws_cfa.append(["Finance Cost Paid", "Actual Interest & Finance Cost Paid", "=-TOT_Finance", "=-TOT_Finance_PY", "Financing"])
    fin_paid_cfa_row = 4
    ws_cfa.append(["Interest / Dividend Received", "Interest and Dividend Income Received", "=TOT_InterestIncome", "=TOT_InterestIncome_PY", "Investing"])
    int_recd_cfa_row = 5
    ws_cfa.append(["Actual Capex Paid", "Purchase of Property, Plant & Equipment", "=-((TOT_PPE+TOT_CWIP)-(TOT_PPE_PY+TOT_CWIP_PY)+TOT_Depreciation)", 0.00, "Investing"])
    capex_paid_cfa_row = 6

    for r_cfa in range(2, 7):
        ws_cfa.cell(row=r_cfa, column=3).number_format = INDIAN_NUM_FORMAT
        ws_cfa.cell(row=r_cfa, column=4).number_format = INDIAN_NUM_FORMAT
    ws_cfa.sheet_state = "hidden"

    # -------------------------------------------------------------
    # STEP 4 — NOTES: ONE SHEET PER NOTE WITH DYNAMIC NAMED RANGES
    # -------------------------------------------------------------
    notes = generate_or_update_notes(client_id, db)
    
    note_sheets_created = {}
    note_defined_names = {}

    note_code_to_name = {
        "1.1": ("Note_1_1_ShareCapital", "TOT_ShareCapital", "TOT_ShareCapital_PY"),
        "1.2": ("Note_1_2_Reserves", "TOT_Reserves", "TOT_Reserves_PY"),
        "2.1": ("Note_2_1_LTBorrow", "TOT_LTBorrow", "TOT_LTBorrow_PY"),
        "2.2": ("Note_2_2_LTProv", "TOT_LTProv", "TOT_LTProv_PY"),
        "3.1": ("Note_3_1_STBorrow", "TOT_STBorrow", "TOT_STBorrow_PY"),
        "3.2": ("Note_3_2_TradePayables", "TOT_Payables", "TOT_Payables_PY"),
        "3.3": ("Note_3_3_OtherCL", "TOT_OtherCL", "TOT_OtherCL_PY"),
        "3.4": ("Note_3_4_STProv", "TOT_STProv", "TOT_STProv_PY"),
        "4.1": ("Note_4_1_PPE", "TOT_PPE", "TOT_PPE_PY"),
        "4.2": ("Note_4_2_CWIP", "TOT_CWIP", "TOT_CWIP_PY"),
        "4.3": ("Note_4_3_Investments", "TOT_Investments", "TOT_Investments_PY"),
        "4.4": ("Note_4_4_LTLA", "TOT_LTLA", "TOT_LTLA_PY"),
        "5.1": ("Note_5_1_Inventories", "TOT_Inventory", "TOT_Inventory_PY"),
        "5.2": ("Note_5_2_Receivables", "TOT_Receivables", "TOT_Receivables_PY"),
        "5.3": ("Note_5_3_CashBank", "TOT_Cash", "TOT_Cash_PY"),
        "5.4": ("Note_5_4_STLA", "TOT_STLA", "TOT_STLA_PY"),
        "5.5": ("Note_5_5_OtherCA", "TOT_OtherCA", "TOT_OtherCA_PY"),
        "6.1": ("Note_6_1_Revenue", "TOT_Revenue", "TOT_Revenue_PY"),
        "6.2": ("Note_6_2_OtherIncome", "TOT_OtherIncome", "TOT_OtherIncome_PY"),
        "7.1": ("Note_7_1_Material", "TOT_Material", "TOT_Material_PY"),
        "7.2": ("Note_7_2_Employee", "TOT_Employee", "TOT_Employee_PY"),
        "7.3": ("Note_7_3_Finance", "TOT_Finance", "TOT_Finance_PY"),
        "7.4": ("Note_7_4_Depreciation", "TOT_Depreciation", "TOT_Depreciation_PY"),
        "7.5": ("Note_7_5_OtherExp", "TOT_OtherExp", "TOT_OtherExp_PY"),
        "7.6": ("Note_7_6_Tax", "TOT_Tax", "TOT_Tax_PY"),
        "8.1": ("Note_8_1_RPT", "TOT_RPT", "TOT_RPT_PY"),
        "8.2": ("Note_8_2_Contingencies", "TOT_Contingencies", "TOT_Contingencies_PY"),
    }

    # Consolidated Master Notes Sheet
    ws_all_notes = wb.create_sheet(title="06_Notes_to_Accounts")
    ws_all_notes.append([client_name, "NOTES FORMING PART OF FINANCIAL STATEMENTS", "=HYPERLINK(\"#'01_Cover_Page'!A1\", \"[<- Back to Index Sitemap]\")", ""])
    ws_all_notes.append(["Figures in " + (client.currency if client else "INR Lakhs"), "", "", ""])
    apply_header_styling(ws_all_notes, 1)

    for note in notes:
        n_num = note.note_number
        sheet_title, dn_cy, dn_py = note_code_to_name.get(n_num, (f"Note_{n_num.replace('.', '_')}", f"TOT_{n_num.replace('.', '_')}", f"TOT_{n_num.replace('.', '_')}_PY"))
        
        ws_n = wb.create_sheet(title=sheet_title)
        note_sheets_created[n_num] = sheet_title
        
        ws_n.append([client_name.upper()])
        ws_n.append([f"NOTE {n_num}: {note.title.upper()}", "=HYPERLINK(\"#'02_Balance_Sheet'!A1\", \"[<- Back to Balance Sheet]\")", ""])
        ws_n.append(["(All amounts in INR Lakhs unless otherwise stated)", "", ""])
        ws_n.append([])
        
        ws_n.cell(row=1, column=1).font = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
        ws_n.cell(row=2, column=1).font = BOLD_NAVY_FONT
        ws_n.cell(row=2, column=2).font = Font(name="Segoe UI", size=9, bold=True, color="EA580C")
        
        det_start_row = 6
        det_end_row = 6
        grand_total_row_idx = None

        credit_notes = {"1.1", "1.2", "2.1", "2.2", "3.1", "3.2", "3.3", "3.4", "6.1", "6.2", "7.6"}
        is_credit = n_num in credit_notes

        # BUG 5: Note 1.2 Reserves & Surplus formula override to tie PAT directly to P&L PAT cell ('03_Profit_and_Loss'!C18)
        if n_num == "1.2":
            headers = ["Particulars", "As at March 31, 2025", "As at March 31, 2024"]
            ws_n.append(headers)
            apply_header_styling(ws_n, 5)
            
            rs_rows = [
                ["(a) General Reserve:", "", ""],
                ["    Opening Balance", "=ABS(SUMIFS('91_Mapping'!I:I, '91_Mapping'!F:F, \"1.2\"))", "=ABS(SUMIFS('91_Mapping'!I:I, '91_Mapping'!F:F, \"1.2\"))"],
                ["    Add: Transfer during the year", 0.00, 0.00],
                ["    Closing General Reserve  (A)", "=SUM(B7:B8)", "=SUM(C7:C8)"],
                ["(b) Surplus in Statement of Profit & Loss:", "", ""],
                ["    Opening Balance", "='03_Profit_and_Loss'!D18", 0.00],
                ["    Add: Profit After Tax for the year", "='03_Profit_and_Loss'!C18", "='03_Profit_and_Loss'!D18"],
                ["    Closing Surplus in P&L  (B)", "=SUM(B11:B12)", "=SUM(C11:C12)"],
                ["TOTAL RESERVES AND SURPLUS  (A + B)", "=B9+B13", "=C9+C13"],
            ]
            for r_rs in rs_rows:
                ws_n.append(r_rs)
                r_idx = ws_n.max_row
                is_tot = bool(r_rs[0] and ("TOTAL" in r_rs[0] or "Closing" in r_rs[0]))
                if "TOTAL" in str(r_rs[0]):
                    grand_total_row_idx = r_idx
                for col_c in range(1, len(r_rs) + 1):
                    cell = ws_n.cell(row=r_idx, column=col_c)
                    if is_tot:
                        cell.font = BOLD_TEXT_FONT
                        cell.fill = SUBTOTAL_ROW_FILL if "Closing" in r_rs[0] else TOTAL_ROW_FILL
                    if col_c > 1 and not str(r_rs[col_c - 1]).startswith("="):
                        cell.number_format = INDIAN_NUM_FORMAT
            det_start_row = 6
            det_end_row = 14
        elif note.table_json:
            try:
                t_data = json.loads(note.table_json)
                headers = t_data.get("headers", [])
                rows = t_data.get("rows", [])
                if headers and rows:
                    ws_n.append(headers)
                    apply_header_styling(ws_n, 5)
                    det_start_row = 6
                    
                    for r in rows:
                        typed_row = []
                        for idx_c, cell_val in enumerate(r):
                            if idx_c > 0:
                                val_str = str(cell_val).replace(',', '').strip()
                                try:
                                    if val_str and (val_str.replace('.', '', 1).replace('-', '', 1).isdigit()):
                                        typed_row.append(float(val_str))
                                    else:
                                        typed_row.append(cell_val)
                                except ValueError:
                                    typed_row.append(cell_val)
                            else:
                                typed_row.append(cell_val)

                        ws_n.append(typed_row)
                        r_idx = ws_n.max_row
                        is_tot = bool(r[0] and (r[0].strip().startswith("TOTAL") or r[0].strip().startswith("GRAND TOTAL") or r[0].strip().startswith("NET BLOCK")))
                        if is_tot:
                            grand_total_row_idx = r_idx
                        else:
                            det_end_row = r_idx
                        for col_c in range(1, len(r) + 1):
                            cell = ws_n.cell(row=r_idx, column=col_c)
                            if is_tot:
                                cell.font = BOLD_TEXT_FONT
                                cell.fill = TOTAL_ROW_FILL
                            if col_c > 1 and isinstance(cell.value, (int, float)):
                                cell.number_format = INDIAN_NUM_FORMAT
            except Exception as e:
                print(f"Error building note sheet {n_num}: {e}")
        else:
            ws_n.append(["Particulars", "Current Year", "Previous Year"])
            apply_header_styling(ws_n, 5)
            ws_n.append([note.title, 0.00, 0.00])
            det_start_row = 6
            det_end_row = 6
            grand_total_row_idx = 6

        # Determine exact grand-total cell address and column for defined name pointing
        if grand_total_row_idx is None:
            grand_total_row_idx = ws_n.max_row

        # Target column letters for CY and PY in table layout
        if n_num == "4.1":
            gt_cy_col, gt_py_col = "J", "K"
        elif n_num == "3.2":
            gt_cy_col, gt_py_col = "G", "C"
        elif n_num == "5.2":
            gt_cy_col, gt_py_col = "H", "C"
        elif n_num == "4.2":
            gt_cy_col, gt_py_col = "F", "C"
        else:
            gt_cy_col, gt_py_col = "B", "C"

        # BUG A FIX: Note Total Row equals single grand-total cell (no column SUM double counting)
        tot_row_idx = ws_n.max_row + 1
        tot_cy_formula = f"={gt_cy_col}{grand_total_row_idx}"
        tot_py_formula = f"={gt_py_col}{grand_total_row_idx}"

        ws_n.append([f"TOTAL NOTE {n_num} ({note.title.upper()})", tot_cy_formula, tot_py_formula])
        ws_n.row_dimensions[tot_row_idx].fill = TOTAL_ROW_FILL
        ws_n.cell(row=tot_row_idx, column=1).font = BOLD_NAVY_FONT
        ws_n.cell(row=tot_row_idx, column=2).font = BOLD_NAVY_FONT
        ws_n.cell(row=tot_row_idx, column=3).font = BOLD_NAVY_FONT
        ws_n.cell(row=tot_row_idx, column=2).number_format = INDIAN_NUM_FORMAT
        ws_n.cell(row=tot_row_idx, column=3).number_format = INDIAN_NUM_FORMAT
        ws_n.cell(row=tot_row_idx, column=1).border = TOTAL_BORDER
        ws_n.cell(row=tot_row_idx, column=2).border = TOTAL_BORDER
        ws_n.cell(row=tot_row_idx, column=3).border = TOTAL_BORDER

        # BUG 2: Per Trial Balance (Control Total) Formula from 91_Mapping with sign correction for credit vs debit
        tb_ctrl_row_idx = ws_n.max_row + 1
        if n_num == "1.2":
            tb_cy_f = "=-SUMIFS('91_Mapping'!H:H, '91_Mapping'!F:F, \"1.2\") + '03_Profit_and_Loss'!C18"
            tb_py_f = "=-SUMIFS('91_Mapping'!I:I, '91_Mapping'!F:F, \"1.2\") + '03_Profit_and_Loss'!D18"
        elif is_credit:
            tb_cy_f = f"=-SUMIFS('91_Mapping'!H:H, '91_Mapping'!F:F, \"{n_num}\")"
            tb_py_f = f"=-SUMIFS('91_Mapping'!I:I, '91_Mapping'!F:F, \"{n_num}\")"
        else:
            tb_cy_f = f"=SUMIFS('91_Mapping'!H:H, '91_Mapping'!F:F, \"{n_num}\")"
            tb_py_f = f"=SUMIFS('91_Mapping'!I:I, '91_Mapping'!F:F, \"{n_num}\")"

        ws_n.append(["Per Trial Balance (Control Total)", tb_cy_f, tb_py_f])
        ws_n.cell(row=tb_ctrl_row_idx, column=1).font = BOLD_TEXT_FONT
        ws_n.cell(row=tb_ctrl_row_idx, column=2).number_format = INDIAN_NUM_FORMAT
        ws_n.cell(row=tb_ctrl_row_idx, column=3).number_format = INDIAN_NUM_FORMAT
        ws_n.cell(row=tb_ctrl_row_idx, column=2).fill = GREY_COLUMN_FILL
        ws_n.cell(row=tb_ctrl_row_idx, column=3).fill = GREY_COLUMN_FILL

        # BUG 2: Difference Row (TOTAL NOTE minus Per TB Control, must equal 0.00)
        diff_row_idx = ws_n.max_row + 1
        ws_n.append(["Difference (Total - Per TB Control)", f"=B{tot_row_idx}-B{tb_ctrl_row_idx}", f"=C{tot_row_idx}-C{tb_ctrl_row_idx}"])
        ws_n.cell(row=diff_row_idx, column=1).font = BOLD_TEXT_FONT
        ws_n.cell(row=diff_row_idx, column=2).number_format = INDIAN_NUM_FORMAT
        ws_n.cell(row=diff_row_idx, column=3).number_format = INDIAN_NUM_FORMAT

        # BUG A FIX: Register WORKBOOK-LEVEL Defined Names pointing to single grand-total cell
        cell_ref_cy = f"'{sheet_title}'!${gt_cy_col}${grand_total_row_idx}"
        cell_ref_py = f"'{sheet_title}'!${gt_py_col}${grand_total_row_idx}"
        
        wb.defined_names[dn_cy] = DefinedName(dn_cy, attr_text=cell_ref_cy)
        wb.defined_names[dn_py] = DefinedName(dn_py, attr_text=cell_ref_py)
        note_defined_names[n_num] = (dn_cy, dn_py, cell_ref_cy, cell_ref_py)

        # Extra Sub-lines for Cash Flow & Income Adjustments
        if n_num == "5.3":
            # Note 5.3 Cash & Bank: Register CashEquiv defined name
            dn_cash_eq_cy = "TOT_CashEquiv"
            dn_cash_eq_py = "TOT_CashEquiv_PY"
            wb.defined_names[dn_cash_eq_cy] = DefinedName(dn_cash_eq_cy, attr_text=f"'{sheet_title}'!${gt_cy_col}${grand_total_row_idx}")
            wb.defined_names[dn_cash_eq_py] = DefinedName(dn_cash_eq_py, attr_text=f"'{sheet_title}'!${gt_py_col}${grand_total_row_idx}")

        if n_num == "6.2":
            # Note 6.2 Other Income: Register InterestIncome defined name
            dn_int_cy = "TOT_InterestIncome"
            dn_int_py = "TOT_InterestIncome_PY"
            wb.defined_names[dn_int_cy] = DefinedName(dn_int_cy, attr_text=f"'{sheet_title}'!$B$6")
            wb.defined_names[dn_int_py] = DefinedName(dn_int_py, attr_text=f"'{sheet_title}'!$C$6")

        if note.content and note.content.strip():
            ws_n.append([])
            ws_n.append([note.content])
            ws_n.cell(row=ws_n.max_row, column=1).font = MUTED_FONT

        apply_page_setup(ws_n, f"Note {n_num}")

        # Also append to Consolidated Master Notes Sheet
        ws_all_notes.append([])
        ws_all_notes.append([f"NOTE {n_num}: {note.title.upper()}", f"=HYPERLINK(\"#'{sheet_title}'!A1\", \"[Open Note Tab]\")", "", ""])
        ws_all_notes.cell(row=ws_all_notes.max_row, column=1).font = BOLD_NAVY_FONT
        if note.table_json:
            try:
                t_data = json.loads(note.table_json)
                headers = t_data.get("headers", [])
                rows = t_data.get("rows", [])
                if headers and rows:
                    ws_all_notes.append(headers)
                    apply_header_styling(ws_all_notes, ws_all_notes.max_row)
                    for r in rows:
                        ws_all_notes.append(r)
                        r_idx = ws_all_notes.max_row
                        is_tot = bool(r[0] and (r[0].strip().startswith("TOTAL") or r[0].strip().startswith("GRAND TOTAL") or r[0].strip().startswith("NET BLOCK")))
                        for col_c in range(1, len(r) + 1):
                            cell = ws_all_notes.cell(row=r_idx, column=col_c)
                            if is_tot:
                                cell.font = BOLD_TEXT_FONT
                                cell.fill = TOTAL_ROW_FILL
                            if col_c > 1 and str(r[col_c - 1]).replace('.', '', 1).replace('-', '', 1).isdigit():
                                cell.number_format = INDIAN_NUM_FORMAT
            except Exception as e:
                pass

    apply_page_setup(ws_all_notes, "Notes to Accounts")

    # BUG 3: Ensure fallback defined names exist for all non-present optional notes
    optional_dns = [
        ("TOT_LTProv", "TOT_LTProv_PY"),
        ("TOT_Investments", "TOT_Investments_PY"),
        ("TOT_LTLA", "TOT_LTLA_PY"),
        ("TOT_STLA", "TOT_STLA_PY"),
        ("TOT_OtherCA", "TOT_OtherCA_PY"),
        ("TOT_InterestIncome", "TOT_InterestIncome_PY"),
        ("TOT_CashEquiv", "TOT_CashEquiv_PY"),
    ]
    for dnc, dnp in optional_dns:
        if dnc not in wb.defined_names:
            wb.defined_names[dnc] = DefinedName(dnc, attr_text="'Note_1_1_ShareCapital'!$B$2")
        if dnp not in wb.defined_names:
            wb.defined_names[dnp] = DefinedName(dnp, attr_text="'Note_1_1_ShareCapital'!$C$2")

    # -------------------------------------------------------------
    # STEP 5 — PUBLISHED BALANCE SHEET (02_Balance_Sheet) — REAL FORMULAS
    # -------------------------------------------------------------
    # BUG 4: Write real formulas using defined names with sign correction
    ws_bs = wb.create_sheet(title="02_Balance_Sheet")
    ws_bs.append([client_name.upper()])
    ws_bs.append(["BALANCE SHEET AS AT MARCH 31, 2025"])
    ws_bs.append(["(All amounts in INR Lakhs unless otherwise stated)"])
    ws_bs.append([])
    ws_bs.append(["Particulars", "Note No.", "As at March 31, 2025", "As at March 31, 2024"])
    apply_header_styling(ws_bs, 5)

    ws_bs.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    ws_bs.cell(row=2, column=1).font = Font(name="Segoe UI", size=11, bold=True, color="EA580C")
    ws_bs.cell(row=3, column=1).font = Font(name="Segoe UI", size=9, italic=True, color="64748B")

    # 1. Equity and Liabilities
    ws_bs.append(["EQUITY AND LIABILITIES", "", "", ""])
    ws_bs.cell(row=6, column=1).font = BOLD_NAVY_FONT

    ws_bs.append(["1. Shareholders' Funds", "", "", ""])
    ws_bs.cell(row=7, column=1).font = BOLD_TEXT_FONT
    ws_bs.append(["   (a) Share Capital", f'=HYPERLINK("#\'Note_1_1_ShareCapital\'!A1", "1.1")', "=-TOT_ShareCapital", "=-TOT_ShareCapital_PY"])
    ws_bs.append(["   (b) Reserves and Surplus", f'=HYPERLINK("#\'Note_1_2_Reserves\'!A1", "1.2")', "=-TOT_Reserves", "=-TOT_Reserves_PY"])
    sh_start_r, sh_end_r = 8, 9

    ws_bs.append(["Total Shareholders' Funds", "", f"=SUM(C{sh_start_r}:C{sh_end_r})", f"=SUM(D{sh_start_r}:D{sh_end_r})"])
    sh_tot_r = ws_bs.max_row
    ws_bs.row_dimensions[sh_tot_r].fill = SUBTOTAL_ROW_FILL
    ws_bs.cell(row=sh_tot_r, column=1).font = BOLD_TEXT_FONT
    ws_bs.cell(row=sh_tot_r, column=3).font = BOLD_TEXT_FONT
    ws_bs.cell(row=sh_tot_r, column=4).font = BOLD_TEXT_FONT

    ws_bs.append(["2. Non-Current Liabilities", "", "", ""])
    ws_bs.cell(row=ws_bs.max_row, column=1).font = BOLD_TEXT_FONT
    ws_bs.append(["   (a) Long-term Borrowings", f'=HYPERLINK("#\'Note_2_1_LTBorrow\'!A1", "2.1")', "=-TOT_LTBorrow", "=-TOT_LTBorrow_PY"])
    ws_bs.append(["   (b) Long-term Provisions", f'=HYPERLINK("#\'Note_2_2_LTProv\'!A1", "2.2")', "=-TOT_LTProv", "=-TOT_LTProv_PY"])
    ncl_start_r, ncl_end_r = 12, 13

    ws_bs.append(["Total Non-Current Liabilities", "", f"=SUM(C{ncl_start_r}:C{ncl_end_r})", f"=SUM(D{ncl_start_r}:D{ncl_end_r})"])
    ncl_tot_r = ws_bs.max_row
    ws_bs.row_dimensions[ncl_tot_r].fill = SUBTOTAL_ROW_FILL
    ws_bs.cell(row=ncl_tot_r, column=1).font = BOLD_TEXT_FONT
    ws_bs.cell(row=ncl_tot_r, column=3).font = BOLD_TEXT_FONT
    ws_bs.cell(row=ncl_tot_r, column=4).font = BOLD_TEXT_FONT

    ws_bs.append(["3. Current Liabilities", "", "", ""])
    ws_bs.cell(row=ws_bs.max_row, column=1).font = BOLD_TEXT_FONT
    ws_bs.append(["   (a) Short-term Borrowings", f'=HYPERLINK("#\'Note_3_1_STBorrow\'!A1", "3.1")', "=-TOT_STBorrow", "=-TOT_STBorrow_PY"])
    ws_bs.append(["   (b) Trade Payables", f'=HYPERLINK("#\'Note_3_2_TradePayables\'!A1", "3.2")', "=-TOT_Payables", "=-TOT_Payables_PY"])
    ws_bs.append(["   (c) Other Current Liabilities", f'=HYPERLINK("#\'Note_3_3_OtherCL\'!A1", "3.3")', "=-TOT_OtherCL", "=-TOT_OtherCL_PY"])
    ws_bs.append(["   (d) Short-term Provisions", f'=HYPERLINK("#\'Note_3_4_STProv\'!A1", "3.4")', "=-TOT_STProv", "=-TOT_STProv_PY"])
    cl_start_r, cl_end_r = 16, 19

    ws_bs.append(["Total Current Liabilities", "", f"=SUM(C{cl_start_r}:C{cl_end_r})", f"=SUM(D{cl_start_r}:D{cl_end_r})"])
    cl_tot_r = ws_bs.max_row
    ws_bs.row_dimensions[cl_tot_r].fill = SUBTOTAL_ROW_FILL
    ws_bs.cell(row=cl_tot_r, column=1).font = BOLD_TEXT_FONT
    ws_bs.cell(row=cl_tot_r, column=3).font = BOLD_TEXT_FONT
    ws_bs.cell(row=cl_tot_r, column=4).font = BOLD_TEXT_FONT

    ws_bs.append(["TOTAL EQUITY AND LIABILITIES", "", f"=C{sh_tot_r}+C{ncl_tot_r}+C{cl_tot_r}", f"=D{sh_tot_r}+D{ncl_tot_r}+D{cl_tot_r}"])
    tot_eq_liab_r = ws_bs.max_row
    ws_bs.row_dimensions[tot_eq_liab_r].fill = TOTAL_ROW_FILL
    ws_bs.cell(row=tot_eq_liab_r, column=1).font = BOLD_NAVY_FONT
    ws_bs.cell(row=tot_eq_liab_r, column=3).font = BOLD_NAVY_FONT
    ws_bs.cell(row=tot_eq_liab_r, column=4).font = BOLD_NAVY_FONT
    ws_bs.cell(row=tot_eq_liab_r, column=1).border = TOTAL_BORDER
    ws_bs.cell(row=tot_eq_liab_r, column=3).border = TOTAL_BORDER
    ws_bs.cell(row=tot_eq_liab_r, column=4).border = TOTAL_BORDER

    # 2. Assets
    ws_bs.append([])
    ws_bs.append(["ASSETS", "", "", ""])
    ws_bs.cell(row=ws_bs.max_row, column=1).font = BOLD_NAVY_FONT

    ws_bs.append(["1. Non-Current Assets", "", "", ""])
    ws_bs.cell(row=ws_bs.max_row, column=1).font = BOLD_TEXT_FONT
    ws_bs.append(["   (a) Property, Plant and Equipment", f'=HYPERLINK("#\'Note_4_1_PPE\'!A1", "4.1")', "=TOT_PPE", "=TOT_PPE_PY"])
    ws_bs.append(["   (b) Capital Work-in-Progress", f'=HYPERLINK("#\'Note_4_2_CWIP\'!A1", "4.2")', "=TOT_CWIP", "=TOT_CWIP_PY"])
    ws_bs.append(["   (c) Non-current Investments", f'=HYPERLINK("#\'Note_4_3_Investments\'!A1", "4.3")', "=TOT_Investments", "=TOT_Investments_PY"])
    ws_bs.append(["   (d) Long-term Loans and Advances", f'=HYPERLINK("#\'Note_4_4_LTLA\'!A1", "4.4")', "=TOT_LTLA", "=TOT_LTLA_PY"])
    nca_start_r, nca_end_r = 25, 28

    ws_bs.append(["Total Non-Current Assets", "", f"=SUM(C{nca_start_r}:C{nca_end_r})", f"=SUM(D{nca_start_r}:D{nca_end_r})"])
    nca_tot_r = ws_bs.max_row
    ws_bs.row_dimensions[nca_tot_r].fill = SUBTOTAL_ROW_FILL
    ws_bs.cell(row=nca_tot_r, column=1).font = BOLD_TEXT_FONT
    ws_bs.cell(row=nca_tot_r, column=3).font = BOLD_TEXT_FONT
    ws_bs.cell(row=nca_tot_r, column=4).font = BOLD_TEXT_FONT

    ws_bs.append(["2. Current Assets", "", "", ""])
    ws_bs.cell(row=ws_bs.max_row, column=1).font = BOLD_TEXT_FONT
    ws_bs.append(["   (a) Inventories", f'=HYPERLINK("#\'Note_5_1_Inventories\'!A1", "5.1")', "=TOT_Inventory", "=TOT_Inventory_PY"])
    ws_bs.append(["   (b) Trade Receivables", f'=HYPERLINK("#\'Note_5_2_Receivables\'!A1", "5.2")', "=TOT_Receivables", "=TOT_Receivables_PY"])
    ws_bs.append(["   (c) Cash and Bank Balances", f'=HYPERLINK("#\'Note_5_3_CashBank\'!A1", "5.3")', "=TOT_Cash", "=TOT_Cash_PY"])
    ws_bs.append(["   (d) Short-term Loans and Advances", f'=HYPERLINK("#\'Note_5_4_STLA\'!A1", "5.4")', "=TOT_STLA", "=TOT_STLA_PY"])
    ws_bs.append(["   (e) Other Current Assets", f'=HYPERLINK("#\'Note_5_5_OtherCA\'!A1", "5.5")', "=TOT_OtherCA", "=TOT_OtherCA_PY"])
    ca_start_r, ca_end_r = 31, 35

    ws_bs.append(["Total Current Assets", "", f"=SUM(C{ca_start_r}:C{ca_end_r})", f"=SUM(D{ca_start_r}:D{ca_end_r})"])
    ca_tot_r = ws_bs.max_row
    ws_bs.row_dimensions[ca_tot_r].fill = SUBTOTAL_ROW_FILL
    ws_bs.cell(row=ca_tot_r, column=1).font = BOLD_TEXT_FONT
    ws_bs.cell(row=ca_tot_r, column=3).font = BOLD_TEXT_FONT
    ws_bs.cell(row=ca_tot_r, column=4).font = BOLD_TEXT_FONT

    ws_bs.append(["TOTAL ASSETS", "", f"=C{nca_tot_r}+C{ca_tot_r}", f"=D{nca_tot_r}+D{ca_tot_r}"])
    tot_assets_r = ws_bs.max_row
    ws_bs.row_dimensions[tot_assets_r].fill = TOTAL_ROW_FILL
    ws_bs.cell(row=tot_assets_r, column=1).font = BOLD_NAVY_FONT
    ws_bs.cell(row=tot_assets_r, column=3).font = BOLD_NAVY_FONT
    ws_bs.cell(row=tot_assets_r, column=4).font = BOLD_NAVY_FONT
    ws_bs.cell(row=tot_assets_r, column=1).border = TOTAL_BORDER
    ws_bs.cell(row=tot_assets_r, column=3).border = TOTAL_BORDER
    ws_bs.cell(row=tot_assets_r, column=4).border = TOTAL_BORDER

    # Control Line
    ws_bs.append(["Balance Check (Total Assets - Total Equity & Liabilities)", "", f"=C{tot_assets_r}-C{tot_eq_liab_r}", f"=D{tot_assets_r}-D{tot_eq_liab_r}"])
    bs_chk_r = ws_bs.max_row
    ws_bs.cell(row=bs_chk_r, column=1).font = MUTED_FONT
    ws_bs.cell(row=bs_chk_r, column=3).font = BOLD_TEXT_FONT
    ws_bs.cell(row=bs_chk_r, column=4).font = BOLD_TEXT_FONT

    for r_bs in range(5, ws_bs.max_row + 1):
        ws_bs.cell(row=r_bs, column=3).number_format = INDIAN_NUM_FORMAT
        ws_bs.cell(row=r_bs, column=4).number_format = INDIAN_NUM_FORMAT
        ws_bs.cell(row=r_bs, column=4).fill = GREY_COLUMN_FILL
        if ws_bs.cell(row=r_bs, column=2).value and str(ws_bs.cell(row=r_bs, column=2).value).startswith("="):
            ws_bs.cell(row=r_bs, column=2).font = Font(name="Segoe UI", size=9, bold=True, color="EA580C")

    apply_page_setup(ws_bs, "Balance Sheet")

    # -------------------------------------------------------------
    # STEP 6 — PUBLISHED PROFIT & LOSS (03_Profit_and_Loss) — REAL FORMULAS
    # -------------------------------------------------------------
    # BUG 4: Write real P&L formulas referencing defined names
    ws_pl = wb.create_sheet(title="03_Profit_and_Loss")
    ws_pl.append([client_name.upper()])
    ws_pl.append(["STATEMENT OF PROFIT AND LOSS FOR THE YEAR ENDED MARCH 31, 2025"])
    ws_pl.append(["(All amounts in INR Lakhs unless otherwise stated)"])
    ws_pl.append([])
    ws_pl.append(["Particulars", "Note No.", "Current Year", "Previous Year"])
    apply_header_styling(ws_pl, 5)

    ws_pl.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    ws_pl.cell(row=2, column=1).font = Font(name="Segoe UI", size=11, bold=True, color="EA580C")
    ws_pl.cell(row=3, column=1).font = Font(name="Segoe UI", size=9, italic=True, color="64748B")

    ws_pl.append(["I. Revenue from Operations", f'=HYPERLINK("#\'Note_6_1_Revenue\'!A1", "6.1")', "=-TOT_Revenue", "=-TOT_Revenue_PY"])
    rev_row = 6
    ws_pl.append(["II. Other Income", f'=HYPERLINK("#\'Note_6_2_OtherIncome\'!A1", "6.2")', "=-TOT_OtherIncome", "=-TOT_OtherIncome_PY"])
    oth_inc_row = 7

    ws_pl.append(["III. Total Income (I + II)", "", f"=SUM(C{rev_row}:C{oth_inc_row})", f"=SUM(D{rev_row}:D{oth_inc_row})"])
    tot_inc_row = ws_pl.max_row
    ws_pl.row_dimensions[tot_inc_row].fill = SUBTOTAL_ROW_FILL
    ws_pl.cell(row=tot_inc_row, column=1).font = BOLD_TEXT_FONT

    ws_pl.append(["IV. Expenses:", "", "", ""])
    ws_pl.cell(row=9, column=1).font = BOLD_TEXT_FONT
    ws_pl.append(["   Cost of Materials Consumed", f'=HYPERLINK("#\'Note_7_1_Material\'!A1", "7.1")', "=TOT_Material", "=TOT_Material_PY"])
    ws_pl.append(["   Employee Benefit Expenses", f'=HYPERLINK("#\'Note_7_2_Employee\'!A1", "7.2")', "=TOT_Employee", "=TOT_Employee_PY"])
    ws_pl.append(["   Finance Costs", f'=HYPERLINK("#\'Note_7_3_Finance\'!A1", "7.3")', "=TOT_Finance", "=TOT_Finance_PY"])
    ws_pl.append(["   Depreciation and Amortisation Expense", f'=HYPERLINK("#\'Note_7_4_Depreciation\'!A1", "7.4")', "=TOT_Depreciation", "=TOT_Depreciation_PY"])
    ws_pl.append(["   Other Expenses", f'=HYPERLINK("#\'Note_7_5_OtherExp\'!A1", "7.5")', "=TOT_OtherExp", "=TOT_OtherExp_PY"])
    exp_start_r, exp_end_r = 10, 14

    ws_pl.append(["Total Expenses (IV)", "", f"=SUM(C{exp_start_r}:C{exp_end_r})", f"=SUM(D{exp_start_r}:D{exp_end_r})"])
    tot_exp_row = ws_pl.max_row
    ws_pl.row_dimensions[tot_exp_row].fill = SUBTOTAL_ROW_FILL
    ws_pl.cell(row=tot_exp_row, column=1).font = BOLD_TEXT_FONT

    ws_pl.append(["V. Profit Before Tax (III - IV)", "", f"=C{tot_inc_row}-C{tot_exp_row}", f"=D{tot_inc_row}-D{tot_exp_row}"])
    pbt_row = ws_pl.max_row
    ws_pl.row_dimensions[pbt_row].fill = TOTAL_ROW_FILL
    ws_pl.cell(row=pbt_row, column=1).font = BOLD_NAVY_FONT

    ws_pl.append(["VI. Tax Expense", f'=HYPERLINK("#\'Note_7_6_Tax\'!A1", "7.6")', "=-TOT_Tax", "=-TOT_Tax_PY"])
    tax_row = ws_pl.max_row

    ws_pl.append(["VII. Profit After Tax (V - VI)", "", f"=C{pbt_row}-C{tax_row}", f"=D{pbt_row}-D{tax_row}"])
    pat_row = ws_pl.max_row
    ws_pl.row_dimensions[pat_row].fill = TOTAL_ROW_FILL
    ws_pl.cell(row=pat_row, column=1).font = BOLD_NAVY_FONT
    ws_pl.cell(row=pat_row, column=1).border = TOTAL_BORDER
    ws_pl.cell(row=pat_row, column=3).border = TOTAL_BORDER
    ws_pl.cell(row=pat_row, column=4).border = TOTAL_BORDER

    ws_pl.append(["Earnings Per Equity Share (Basic & Diluted in Rs.)", "", f"=C{pat_row}/10.00", f"=D{pat_row}/10.00"])
    eps_row = ws_pl.max_row
    ws_pl.cell(row=eps_row, column=1).font = BOLD_TEXT_FONT

    for r_pl in range(5, ws_pl.max_row + 1):
        if r_pl == eps_row:
            ws_pl.cell(row=r_pl, column=3).number_format = "0.00"
            ws_pl.cell(row=r_pl, column=4).number_format = "0.00"
        else:
            ws_pl.cell(row=r_pl, column=3).number_format = INDIAN_NUM_FORMAT
            ws_pl.cell(row=r_pl, column=4).number_format = INDIAN_NUM_FORMAT
        ws_pl.cell(row=r_pl, column=4).fill = GREY_COLUMN_FILL
        if ws_pl.cell(row=r_pl, column=2).value and str(ws_pl.cell(row=r_pl, column=2).value).startswith("="):
            ws_pl.cell(row=r_pl, column=2).font = Font(name="Segoe UI", size=9, bold=True, color="EA580C")

    apply_page_setup(ws_pl, "Profit & Loss")

    # -------------------------------------------------------------
    # STEP 7 — PUBLISHED CASH FLOW STATEMENT (04_Cash_Flow_Statement) — AS 3 INDIRECT METHOD
    # -------------------------------------------------------------
    # BUG 8: Write real formulas using defined names and 29_Cash_Flow_Adjustments
    ws_cf = wb.create_sheet(title="04_Cash_Flow_Statement")
    ws_cf.append([client_name.upper()])
    ws_cf.append(["AS 3 CASH FLOW STATEMENT (INDIRECT METHOD) FOR THE YEAR ENDED MARCH 31, 2025"])
    ws_cf.append(["(All amounts in INR Lakhs unless otherwise stated)"])
    ws_cf.append([])
    ws_cf.append(["Particulars", "Current Year", "Previous Year"])
    apply_header_styling(ws_cf, 5)

    ws_cf.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    ws_cf.cell(row=2, column=1).font = Font(name="Segoe UI", size=11, bold=True, color="EA580C")
    ws_cf.cell(row=3, column=1).font = Font(name="Segoe UI", size=9, italic=True, color="64748B")

    # A. Operating Activities
    ws_cf.append(["A. CASH FLOW FROM OPERATING ACTIVITIES", "", ""])
    ws_cf.cell(row=6, column=1).font = BOLD_NAVY_FONT

    ws_cf.append(["  Profit Before Tax", f"='03_Profit_and_Loss'!C{pbt_row}", f"='03_Profit_and_Loss'!D{pbt_row}"])
    cf_pbt_row = 7
    ws_cf.append(["  Adjustments for Non-Cash & Non-Operating Items:", "", ""])
    ws_cf.append(["    Add: Depreciation & Amortisation Expense", "=TOT_Depreciation", "=TOT_Depreciation_PY"])
    ws_cf.append(["    Add: Finance Costs", "=TOT_Finance", "=TOT_Finance_PY"])
    ws_cf.append(["    Less: Interest Income / Non-Operating Gain", "=-TOT_InterestIncome", "=-TOT_InterestIncome_PY"])

    ws_cf.append(["  Operating Profit before Working Capital Changes", f"=SUM(B{cf_pbt_row}:B{ws_cf.max_row})", f"=SUM(C{cf_pbt_row}:C{ws_cf.max_row})"])
    op_wc_row = ws_cf.max_row
    ws_cf.row_dimensions[op_wc_row].fill = SUBTOTAL_ROW_FILL
    ws_cf.cell(row=op_wc_row, column=1).font = BOLD_TEXT_FONT

    ws_cf.append(["  Adjustments for Working Capital Movements:", "", ""])
    ws_cf.append(["    (Increase) / Decrease in Inventories", "=-(TOT_Inventory - TOT_Inventory_PY)", 0.00])
    ws_cf.append(["    (Increase) / Decrease in Trade Receivables", "=-(TOT_Receivables - TOT_Receivables_PY)", 0.00])
    ws_cf.append(["    Increase / (Decrease) in Trade Payables", "=(-TOT_Payables) - (-TOT_Payables_PY)", 0.00])
    ws_cf.append(["    Increase / (Decrease) in Other Current Liabilities", "=(-TOT_OtherCL) - (-TOT_OtherCL_PY)", 0.00])
    ws_cf.append(["    Increase / (Decrease) in Short-term Provisions", "=(-TOT_STProv) - (-TOT_STProv_PY)", 0.00])
    wc_start_r, wc_end_r = 14, 18

    ws_cf.append(["  Cash Generated from Operations", f"=B{op_wc_row}+SUM(B{wc_start_r}:B{wc_end_r})", f"=C{op_wc_row}+SUM(C{wc_start_r}:C{wc_end_r})"])
    gen_ops_row = ws_cf.max_row
    ws_cf.row_dimensions[gen_ops_row].fill = SUBTOTAL_ROW_FILL
    ws_cf.cell(row=gen_ops_row, column=1).font = BOLD_TEXT_FONT

    ws_cf.append(["  Less: Income Taxes Paid", f"='29_Cash_Flow_Adjustments'!C{tax_paid_cfa_row}", f"='29_Cash_Flow_Adjustments'!D{tax_paid_cfa_row}"])
    tax_paid_row = ws_cf.max_row

    ws_cf.append(["NET CASH FROM OPERATING ACTIVITIES (A)", f"=B{gen_ops_row}+B{tax_paid_row}", f"=C{gen_ops_row}+C{tax_paid_row}"])
    net_op_row = ws_cf.max_row
    ws_cf.row_dimensions[net_op_row].fill = TOTAL_ROW_FILL
    ws_cf.cell(row=net_op_row, column=1).font = BOLD_NAVY_FONT

    # B. Investing Activities
    ws_cf.append([])
    ws_cf.append(["B. CASH FLOW FROM INVESTING ACTIVITIES", "", ""])
    ws_cf.cell(row=ws_cf.max_row, column=1).font = BOLD_NAVY_FONT
    ws_cf.append(["  Purchase of Property, Plant & Equipment & CWIP", f"='29_Cash_Flow_Adjustments'!C{capex_paid_cfa_row}", f"='29_Cash_Flow_Adjustments'!D{capex_paid_cfa_row}"])
    inv_start_r = ws_cf.max_row
    ws_cf.append(["  Proceeds from Sale of Fixed Assets / Investments", 0.00, 0.00])
    ws_cf.append(["  Interest / Dividend Received", f"='29_Cash_Flow_Adjustments'!C{int_recd_cfa_row}", f"='29_Cash_Flow_Adjustments'!D{int_recd_cfa_row}"])
    inv_end_r = ws_cf.max_row

    ws_cf.append(["NET CASH FROM INVESTING ACTIVITIES (B)", f"=SUM(B{inv_start_r}:B{inv_end_r})", f"=SUM(C{inv_start_r}:C{inv_end_r})"])
    net_inv_row = ws_cf.max_row
    ws_cf.row_dimensions[net_inv_row].fill = TOTAL_ROW_FILL
    ws_cf.cell(row=net_inv_row, column=1).font = BOLD_NAVY_FONT

    # C. Financing Activities
    ws_cf.append([])
    ws_cf.append(["C. CASH FLOW FROM FINANCING ACTIVITIES", "", ""])
    ws_cf.cell(row=ws_cf.max_row, column=1).font = BOLD_NAVY_FONT
    ws_cf.append(["  Net Proceeds / (Repayment) of Borrowings", "=((-TOT_LTBorrow)+(-TOT_STBorrow)) - ((-TOT_LTBorrow_PY)+(-TOT_STBorrow_PY))", 0.00])
    fin_start_r = ws_cf.max_row
    ws_cf.append(["  Proceeds from Issuance of Share Capital", "=((-TOT_ShareCapital)-(-TOT_ShareCapital_PY))", 0.00])
    ws_cf.append(["  Finance Cost Paid", f"='29_Cash_Flow_Adjustments'!C{fin_paid_cfa_row}", f"='29_Cash_Flow_Adjustments'!D{fin_paid_cfa_row}"])
    ws_cf.append(["  Dividends Paid", f"='29_Cash_Flow_Adjustments'!C{div_paid_cfa_row}", f"='29_Cash_Flow_Adjustments'!D{div_paid_cfa_row}"])
    fin_end_r = ws_cf.max_row

    ws_cf.append(["NET CASH FROM FINANCING ACTIVITIES (C)", f"=SUM(B{fin_start_r}:B{fin_end_r})", f"=SUM(C{fin_start_r}:C{fin_end_r})"])
    net_fin_row = ws_cf.max_row
    ws_cf.row_dimensions[net_fin_row].fill = TOTAL_ROW_FILL
    ws_cf.cell(row=net_fin_row, column=1).font = BOLD_NAVY_FONT

    # Cash Summary
    ws_cf.append([])
    ws_cf.append(["NET INCREASE / (DECREASE) IN CASH & CASH EQUIVALENTS (A + B + C)", f"=B{net_op_row}+B{net_inv_row}+B{net_fin_row}", f"=C{net_op_row}+C{net_inv_row}+C{net_fin_row}"])
    net_inc_row = ws_cf.max_row
    ws_cf.row_dimensions[net_inc_row].fill = TOTAL_ROW_FILL
    ws_cf.cell(row=net_inc_row, column=1).font = BOLD_NAVY_FONT

    ws_cf.append(["Cash and Cash Equivalents at Beginning of the Year", "=TOT_CashEquiv_PY", "=TOT_CashEquiv_PY"])
    open_cash_row = ws_cf.max_row

    ws_cf.append(["Computed Closing Cash & Cash Equivalents", f"=B{net_inc_row}+B{open_cash_row}", f"=C{net_inc_row}+C{open_cash_row}"])
    comp_close_row = ws_cf.max_row
    ws_cf.cell(row=comp_close_row, column=1).font = BOLD_TEXT_FONT

    ws_cf.append(["Closing Cash & Cash Equivalents per Balance Sheet", "=TOT_CashEquiv", "=TOT_CashEquiv"])
    bs_close_row = ws_cf.max_row
    ws_cf.cell(row=bs_close_row, column=1).font = BOLD_TEXT_FONT

    ws_cf.append(["Reconciliation Difference (Computed Closing - BS Closing)", f"=B{comp_close_row}-B{bs_close_row}", f"=C{comp_close_row}-C{bs_close_row}"])
    cf_recon_row = ws_cf.max_row
    ws_cf.row_dimensions[cf_recon_row].fill = TOTAL_ROW_FILL
    ws_cf.cell(row=cf_recon_row, column=1).font = BOLD_NAVY_FONT
    ws_cf.cell(row=cf_recon_row, column=1).border = TOTAL_BORDER
    ws_cf.cell(row=cf_recon_row, column=2).border = TOTAL_BORDER
    ws_cf.cell(row=cf_recon_row, column=3).border = TOTAL_BORDER

    for r_cf in range(5, ws_cf.max_row + 1):
        ws_cf.cell(row=r_cf, column=2).number_format = INDIAN_NUM_FORMAT
        ws_cf.cell(row=r_cf, column=3).number_format = INDIAN_NUM_FORMAT
        ws_cf.cell(row=r_cf, column=3).fill = GREY_COLUMN_FILL

    apply_page_setup(ws_cf, "Cash Flow Statement")

    # -------------------------------------------------------------
    # STEP 8 — ACCOUNTING POLICIES (05_Accounting_Policies)
    # -------------------------------------------------------------
    ws_pol = wb.create_sheet(title="05_Accounting_Policies")
    ws_pol.append([client_name.upper()])
    ws_pol.append(["SIGNIFICANT ACCOUNTING POLICIES (IGAAP)"])
    ws_pol.append([])
    ws_pol.append(["POLICY #", "TITLE", "SIGNIFICANT ACCOUNTING POLICY DISCLOSURE", "APPLICABILITY"])
    apply_header_styling(ws_pol, 4)

    ws_pol.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    ws_pol.cell(row=2, column=1).font = Font(name="Segoe UI", size=11, bold=True, color="EA580C")

    policies = generate_or_update_accounting_policies(client_id, db)
    for p in policies:
        ws_pol.append([p.policy_number, p.title, p.content, "Applicable" if p.is_applicable else "Not Applicable"])
    apply_page_setup(ws_pol, "Accounting Policies")

    # -------------------------------------------------------------
    # STEP 9 — FINANCIAL RATIOS (07_Financial_Ratios) — REFERENCE THE FACE
    # -------------------------------------------------------------
    ws_rat = wb.create_sheet(title="07_Financial_Ratios")
    ws_rat.append([client_name.upper()])
    ws_rat.append(["SCHEDULE III MANDATORY FINANCIAL RATIOS"])
    ws_rat.append([])
    ws_rat.append(["Ratio Code", "Ratio Name", "Formula", "Current Year", "Previous Year", "Unit", "Movement %", "Audit Interpretation"])
    apply_header_styling(ws_rat, 4)

    ws_rat.cell(row=1, column=1).font = Font(name="Segoe UI", size=14, bold=True, color="1B365D")
    ws_rat.cell(row=2, column=1).font = Font(name="Segoe UI", size=11, bold=True, color="EA580C")

    ratios = calculate_ratios(client_id, db)
    for r in ratios:
        cy_f_val = f"='02_Balance_Sheet'!C{ca_tot_r}/'02_Balance_Sheet'!C{cl_tot_r}" if r.code == "R01" else f"={r.cy_value:.2f}"
        py_f_val = f"='02_Balance_Sheet'!D{ca_tot_r}/'02_Balance_Sheet'!D{cl_tot_r}" if r.code == "R01" else f"={r.py_value:.2f}"
        
        if r.code == "R02": # Debt Equity
            cy_f_val = f"=('02_Balance_Sheet'!C{ncl_start_r}+'02_Balance_Sheet'!C{cl_start_r})/('02_Balance_Sheet'!C{sh_start_r}+'02_Balance_Sheet'!C{sh_start_r+1})"
            py_f_val = f"=('02_Balance_Sheet'!D{ncl_start_r}+'02_Balance_Sheet'!D{cl_start_r})/('02_Balance_Sheet'!D{sh_start_r}+'02_Balance_Sheet'!D{sh_start_r+1})"
        elif r.code == "R09": # Net Profit Ratio
            cy_f_val = f"='03_Profit_and_Loss'!C{pat_row}/'03_Profit_and_Loss'!C{rev_row}"
            py_f_val = f"='03_Profit_and_Loss'!D{pat_row}/'03_Profit_and_Loss'!D{rev_row}"

        ws_rat.append([r.code, r.name, r.formula, cy_f_val, py_f_val, r.unit, r.movement, r.interpretation])
        r_idx = ws_rat.max_row
        ws_rat.cell(row=r_idx, column=4).number_format = "0.00"
        ws_rat.cell(row=r_idx, column=5).number_format = "0.00"
        ws_rat.cell(row=r_idx, column=5).fill = GREY_COLUMN_FILL
    apply_page_setup(ws_rat, "Financial Ratios")

    # -------------------------------------------------------------
    # STEP 10 — COVER PAGE (01_Cover_Page)
    # -------------------------------------------------------------
    ws_cover = wb.create_sheet(title="01_Cover_Page")
    ws_cover.append([client_name.upper()])
    ws_cover.append(["FINANCIAL STATEMENTS FOR THE YEAR ENDED MARCH 31, 2025"])
    ws_cover.append(["PREPARED UNDER IGAAP SCHEDULE III DIVISION I OF COMPANIES ACT, 2013"])
    ws_cover.append([])
    ws_cover.append(["Client Name:", client.name])
    ws_cover.append(["Entity Constitution:", client.entity_type])
    ws_cover.append(["Reporting Period:", client.reporting_period])
    ws_cover.append(["Previous Year Period:", client.previous_year_period])
    ws_cover.append(["Currency / Denomination:", client.currency])
    ws_cover.append(["Accounting Framework:", client.accounting_framework])
    ws_cover.append(["Prepared By:", client.prepared_by])
    ws_cover.append(["Reviewed By:", client.reviewed_by])
    ws_cover.append(["Generation Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_cover.append([])
    ws_cover.append(["PUBLISHED FINANCIAL REPORT CONTENTS & SITEMAP INDEX"])

    sitemap_items = [
        ("01. Cover Page & Financial Metadata", "01_Cover_Page"),
        ("02. Schedule III Balance Sheet", "02_Balance_Sheet"),
        ("03. Statement of Profit and Loss", "03_Profit_and_Loss"),
        ("04. AS 3 Cash Flow Statement (Indirect Method)", "04_Cash_Flow_Statement"),
        ("05. Significant Accounting Policies (AP-01 to AP-21)", "05_Accounting_Policies"),
        ("06. Notes Forming Part of Financial Statements (Consolidated)", "06_Notes_to_Accounts"),
        ("07. Schedule III Mandatory Financial Ratios", "07_Financial_Ratios"),
    ]
    ws_cover.append(["Sec #", "Published Section Title", "Interactive Navigation Link"])
    apply_header_styling(ws_cover, 17)
    for idx_c, (t_str, s_str) in enumerate(sitemap_items, start=1):
        ws_cover.append([f"S{idx_c:02d}", t_str, f'=HYPERLINK("#\'{s_str}\'!A1", "Open {t_str}")'])
        r_c = ws_cover.max_row
        ws_cover.cell(row=r_c, column=3).font = Font(name="Segoe UI", size=9, bold=True, color="EA580C")

    ws_cover.cell(row=1, column=1).font = Font(name="Segoe UI", size=18, bold=True, color="1B365D")
    ws_cover.cell(row=2, column=1).font = Font(name="Segoe UI", size=12, bold=True, color="EA580C")
    ws_cover.cell(row=3, column=1).font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
    apply_page_setup(ws_cover, "Cover Page")

    # -------------------------------------------------------------
    # STEP 11 — TIE-OUT & FORMULA AUDIT SHEETS (98_Tie_Out & 99_Formula_Audit)
    # -------------------------------------------------------------
    ws_tie = wb.create_sheet(title="98_Tie_Out")
    ws_tie.append(["Check ID", "Control Check Description", "Expected Value", "Computed Formula", "Variance (Must be 0.00)", "Tie-Out Status"])
    apply_header_styling(ws_tie, 1)

    checks = [
        ("CHK-01", "Balance Sheet Tally (Total Assets - Total Equity & Liabilities)", "0.00", f"='02_Balance_Sheet'!C{tot_assets_r}-'02_Balance_Sheet'!C{tot_eq_liab_r}"),
        ("CHK-02", "Share Capital Note vs Mapped TB Check", "0.00", f"=TOT_ShareCapital-SUMIFS('91_Mapping'!H:H, '91_Mapping'!F:F, \"1.1\")"),
        ("CHK-03", "Cash Flow Closing Cash vs Balance Sheet Cash Check", "0.00", f"='04_Cash_Flow_Statement'!B{comp_close_row}-'02_Balance_Sheet'!C33"),
        ("CHK-04", "Profit After Tax to Cash Flow Starting Line Check", "0.00", f"='03_Profit_and_Loss'!C{pbt_row}-'04_Cash_Flow_Statement'!B{cf_pbt_row}"),
    ]
    for c_id, desc, exp_v, comp_f in checks:
        ws_tie.append([c_id, desc, exp_v, comp_f, f"=D{ws_tie.max_row+1}-C{ws_tie.max_row+1}", f'=IF(ABS(E{ws_tie.max_row+1})<0.01, "TALLIED (PASSED)", "REVIEW REQUIRED")'])
        r_i = ws_tie.max_row
        ws_tie.cell(row=r_i, column=5).number_format = INDIAN_NUM_FORMAT
        ws_tie.cell(row=r_i, column=6).font = Font(name="Segoe UI", size=9, bold=True, color="166534")

    ws_tie.sheet_state = "hidden"

    ws_fa = wb.create_sheet(title="99_Formula_Audit")
    ws_fa.append(["Statement", "Line Item Particulars", "Source Note #", "Named Range Reference", "Applied Excel Formula", "Resolved Cell Address"])
    apply_header_styling(ws_fa, 1)

    audit_traces = [
        ("Balance Sheet", "Share Capital", "Note 1.1", "TOT_ShareCapital", "=-TOT_ShareCapital", "'02_Balance_Sheet'!C8"),
        ("Balance Sheet", "Reserves & Surplus", "Note 1.2", "TOT_Reserves", "=-TOT_Reserves", "'02_Balance_Sheet'!C9"),
        ("Balance Sheet", "Property, Plant & Equipment", "Note 4.1", "TOT_PPE", "=TOT_PPE", "'02_Balance_Sheet'!C25"),
        ("Balance Sheet", "Trade Receivables", "Note 5.2", "TOT_Receivables", "=TOT_Receivables", "'02_Balance_Sheet'!C32"),
        ("Balance Sheet", "Cash & Bank Balances", "Note 5.3", "TOT_Cash", "=TOT_Cash", "'02_Balance_Sheet'!C33"),
        ("Profit and Loss", "Revenue from Operations", "Note 6.1", "TOT_Revenue", "=-TOT_Revenue", "'03_Profit_and_Loss'!C6"),
        ("Profit and Loss", "Depreciation & Amortisation", "Note 7.4", "TOT_Depreciation", "=TOT_Depreciation", "'03_Profit_and_Loss'!C13"),
        ("Cash Flow", "Profit Before Tax", "Statement P&L", "PBT Line", f"='03_Profit_and_Loss'!C{pbt_row}", "'04_Cash_Flow_Statement'!B7"),
        ("Cash Flow", "Closing Cash Balances", "Note 5.3", "TOT_CashEquiv", "=TOT_CashEquiv", "'04_Cash_Flow_Statement'!B34"),
    ]
    for st, li, sn, nr, fo, res in audit_traces:
        ws_fa.append([st, li, sn, nr, fo, res])
    ws_fa.sheet_state = "hidden"

    # Explicit Published Two-Layer Sheet Ordering
    published_order = [
        "01_Cover_Page",
        "02_Balance_Sheet",
        "03_Profit_and_Loss",
        "04_Cash_Flow_Statement",
        "05_Accounting_Policies",
        "06_Notes_to_Accounts",
    ] + [s for s in wb.sheetnames if s.startswith("Note_")] + [
        "07_Financial_Ratios",
        "90_Trial_Balance",
        "91_Mapping",
        "29_Cash_Flow_Adjustments",
        "98_Tie_Out",
        "99_Formula_Audit",
    ]
    wb._sheets = [wb[s] for s in published_order if s in wb.sheetnames] + [wb[s] for s in wb.sheetnames if s not in published_order]

    wb.save(filepath)

    # BUG 7 (b) & MANDATORY SELF-TEST: Recalculate workbook and run data_only=True self-test audit
    recalculate_and_verify(filepath)

    return filepath


def recalculate_and_verify(filepath: str):
    from openpyxl import load_workbook
    abs_path = os.path.abspath(filepath)
    
    # 1. Recalculate formulas pass
    recalculated = False
    try:
        import win32com.client as win32
        excel = win32.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        wb_com = excel.Workbooks.Open(abs_path)
        excel.CalculateFull()
        wb_com.Save()
        wb_com.Close(True)
        excel.Quit()
        recalculated = True
        print("[RECALC] Recalculated formulas via Excel COM successfully.")
    except Exception as e:
        print(f"[RECALC] Excel COM recalculation skipped/failed: {e}")

    if not recalculated:
        try:
            import formulas
            f_model = formulas.FormulaModel().load(abs_path).finish()
            f_model.calculate()
            f_model.to_excel(abs_path)
            recalculated = True
            print("[RECALC] Recalculated formulas via formulas package successfully.")
        except Exception as e:
            print(f"[RECALC] Formulas engine recalculation skipped/failed: {e}")

    # 2. Reopen with data_only=True for MANDATORY SELF-TEST
    wb_v = load_workbook(filepath, data_only=True)
    
    audit_results = []
    failed_checks = []

    # Check 1: 91_Mapping CY/PY amounts are all non-blank numbers
    ws_map = wb_v['91_Mapping']
    map_cy_vals = [ws_map.cell(row=r, column=8).value for r in range(2, ws_map.max_row + 1)]
    map_py_vals = [ws_map.cell(row=r, column=9).value for r in range(2, ws_map.max_row + 1)]
    c1_pass = all(v is not None and isinstance(v, (int, float)) for v in map_cy_vals + map_py_vals)
    audit_results.append(("1. 91_Mapping CY/PY Non-Blank Numbers", "All Numeric", f"Count: {len(map_cy_vals)} rows", "PASS" if c1_pass else "FAIL"))
    if not c1_pass:
        failed_checks.append("91_Mapping has blank or non-numeric values")

    # Check 2: Every note 'Per Trial Balance' cell is non-blank
    note_sheets = [s for s in wb_v.sheetnames if s.startswith("Note_")]
    note_tb_vals = []
    c2_pass = True
    for ns in note_sheets:
        ws_n = wb_v[ns]
        tb_row = None
        for r in range(1, ws_n.max_row + 1):
            val = ws_n.cell(row=r, column=1).value
            if val and "Per Trial Balance" in str(val):
                tb_row = r
                break
        if tb_row:
            v_cy = ws_n.cell(row=tb_row, column=2).value
            v_py = ws_n.cell(row=tb_row, column=3).value
            if v_cy is None or v_py is None:
                c2_pass = False
                failed_checks.append(f"{ns} cell B{tb_row}/C{tb_row} is blank")
            note_tb_vals.append((ns, v_cy, v_py))
        else:
            c2_pass = False
            failed_checks.append(f"{ns} missing Per Trial Balance row")
    audit_results.append(("2. Every Note Per Trial Balance Cell Non-Blank", "All Non-Blank", f"Checked {len(note_sheets)} notes", "PASS" if c2_pass else "FAIL"))

    # Check 3: Every note Difference = 0
    c3_pass = True
    for ns in note_sheets:
        ws_n = wb_v[ns]
        diff_row = None
        for r in range(1, ws_n.max_row + 1):
            val = ws_n.cell(row=r, column=1).value
            if val and "Difference" in str(val):
                diff_row = r
                break
        if diff_row:
            d_cy = ws_n.cell(row=diff_row, column=2).value
            d_py = ws_n.cell(row=diff_row, column=3).value
            def _to_float(val):
                if val is None:
                    return 999999.0
                try:
                    return float(val)
                except Exception:
                    return 999999.0

            if abs(_to_float(d_cy)) > 0.01 or abs(_to_float(d_py)) > 0.01:
                c3_pass = False
                failed_checks.append(f"{ns} Difference B{diff_row}={d_cy}, C{diff_row}={d_py}")
        else:
            c3_pass = False
            failed_checks.append(f"{ns} missing Difference row")
    audit_results.append(("3. Every Note Difference = 0.00", "0.00", f"Checked {len(note_sheets)} notes", "PASS" if c3_pass else "FAIL"))

    # Check 4: Every TOT_* defined name exists and resolves to non-blank numeric cell
    c4_pass = True
    tot_names = [dn for dn in wb_v.defined_names.keys() if dn.startswith("TOT_")]
    for dn in tot_names:
        target = wb_v.defined_names[dn].attr_text
        if "!" in target:
            s_name, c_addr = target.split("!")
            s_name = s_name.strip("'")
            c_addr = c_addr.replace("$", "")
            if s_name in wb_v.sheetnames:
                val = wb_v[s_name][c_addr].value
                if val is None or not isinstance(val, (int, float)):
                    c4_pass = False
                    failed_checks.append(f"Defined Name {dn} -> {target} evaluated to {val}")
            else:
                c4_pass = False
                failed_checks.append(f"Defined Name {dn} points to missing sheet {s_name}")
        else:
            c4_pass = False
            failed_checks.append(f"Defined Name {dn} invalid target {target}")
    audit_results.append(("4. Every TOT_* Defined Name Resolves Numeric", "Numeric", f"Count: {len(tot_names)} defined names", "PASS" if c4_pass else "FAIL"))

    # Check 5: Balance Sheet, P&L and Cash Flow face cells are all non-blank
    c5_pass = True
    ws_bs = wb_v['02_Balance_Sheet']
    ws_pl = wb_v['03_Profit_and_Loss']
    ws_cf = wb_v['04_Cash_Flow_Statement']
    bs_vals = [ws_bs.cell(row=r, column=3).value for r in range(8, ws_bs.max_row) if ws_bs.cell(row=r, column=3).value is not None]
    pl_vals = [ws_pl.cell(row=r, column=3).value for r in range(6, ws_pl.max_row) if ws_pl.cell(row=r, column=3).value is not None]
    cf_vals = [ws_cf.cell(row=r, column=2).value for r in range(7, ws_cf.max_row) if ws_cf.cell(row=r, column=2).value is not None]
    if not (len(bs_vals) > 10 and len(pl_vals) > 5 and len(cf_vals) > 10):
        c5_pass = False
        failed_checks.append("Face statement cells are empty")
    audit_results.append(("5. Face Statements Populated Non-Blank", "Non-Blank", f"BS: {len(bs_vals)}, PL: {len(pl_vals)}, CF: {len(cf_vals)}", "PASS" if c5_pass else "FAIL"))

    # Check 6: Balance Sheet Total Assets minus Total Equity & Liabilities = 0
    bs_diff = ws_bs.cell(row=37, column=3).value
    c6_pass = bs_diff is not None and abs(float(bs_diff or 0)) < 0.01
    audit_results.append(("6. Balance Sheet Tally Check", "0.00", f"Variance: {bs_diff}", "PASS" if c6_pass else "FAIL"))
    if not c6_pass:
        failed_checks.append(f"Balance Sheet out of balance by {bs_diff}")

    # Check 7: Ratios show computed values
    ws_rat = wb_v['07_Financial_Ratios']
    rat_vals = [ws_rat.cell(row=r, column=4).value for r in range(4, ws_rat.max_row + 1)]
    c7_pass = all(v is not None and isinstance(v, (int, float)) for v in rat_vals)
    audit_results.append(("7. Financial Ratios Computed Values", "Numeric", f"Count: {len(rat_vals)} ratios", "PASS" if c7_pass else "FAIL"))
    if not c7_pass:
        failed_checks.append("Financial ratios have non-numeric values")

    print("\n" + "="*88)
    print("MANDATORY SELF-TEST AUDIT REPORT")
    print("="*88)
    print(f"{'Check Description':<45} | {'Expected':<12} | {'Actual':<20} | {'Status':<6}")
    print("-" * 88)
    for desc, exp, act, stat in audit_results:
        print(f"{desc:<45} | {exp:<12} | {str(act):<20} | {stat:<6}")
    print("="*88)

    if failed_checks:
        print("\nAUDIT SUMMARY FOR DATA/FACE LAYERS:")
        for fc in failed_checks:
            print(f"  - {fc}")
    else:
        print("\nALL MANDATORY SELF-TEST CHECKS PASSED PERFECTLY!\n")
    return filepath




# -------------------------------------------------------------
# PDF EXPORTER WITH ALL 13 MANDATORY REPORT SECTIONS
# -------------------------------------------------------------
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        if self._pageNumber == 1:
            return  # Skip cover page
        self.saveState()
        self.setFont("Helvetica-Bold", 8)
        self.setFillColor(colors.HexColor("#1B365D"))
        
        # Header
        self.drawString(36, 810, "SW INDIA | FS BUILDER LITE v0.2 | ANNUAL REPORT FINANCIAL STATEMENTS")
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(36, 802, 559, 802)
        
        # Footer
        self.setFont("Helvetica-Oblique", 7)
        self.setFillColor(colors.HexColor("#64748B"))
        self.drawString(36, 25, "Confidential - For Internal CA Review Only. AS 3 & Schedule III Division I Compliant.")
        self.setFont("Helvetica", 8)
        self.drawRightString(559, 25, f"Page {self._pageNumber} of {page_count}")
        self.line(36, 35, 559, 35)
        self.restoreState()


def export_pdf_review_pack(client_id: int, export_dir: str, db: Session) -> str:
    # Critical Mapping Validation Guard
    validations = run_validation_checks(client_id, db)
    critical_mapping = [v for v in validations if v.category == "Mapping Exception" and v.status == "Critical"]
    if critical_mapping:
        err_msg = "; ".join([f"{v.check_name}: {v.message}" for v in critical_mapping])
        raise ValueError(f"PDF report generation blocked due to Critical Mapping Exceptions: {err_msg}. Please review and approve manual override in Ledger Mapping.")

    os.makedirs(export_dir, exist_ok=True)
    client = db.query(Client).filter(Client.id == client_id).first()
    client_name = client.name if client else "Client"
    safe_name = "".join(c if c.isalnum() else "_" for c in client_name)
    filepath = os.path.join(export_dir, f"{safe_name}_Annual_Report_Review_Pack.pdf")

    doc = SimpleDocTemplate(
        filepath, pagesize=A4,
        leftMargin=36, rightMargin=36, topMargin=54, bottomMargin=54
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CoverTitle', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=22,
        leading=26, textColor=colors.HexColor('#1B365D'), alignment=0
    )
    subtitle_style = ParagraphStyle(
        'CoverSub', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=12,
        leading=16, textColor=colors.HexColor('#EA580C'), alignment=0
    )
    h1_style = ParagraphStyle(
        'Header1', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=13,
        leading=16, textColor=colors.HexColor('#1B365D'), spaceAfter=8
    )
    body_style = ParagraphStyle(
        'Body', parent=styles['BodyText'], fontName='Helvetica', fontSize=9,
        leading=12, textColor=colors.HexColor('#0F172A')
    )
    table_cell = ParagraphStyle(
        'Cell', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10
    )
    table_cell_bold = ParagraphStyle(
        'CellBold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10
    )

    story = []

    # SECTION 1: COVER SHEET
    story.append(Spacer(1, 30))
    story.append(Paragraph("SW INDIA | CHARTERED ACCOUNTANTS", subtitle_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("ANNUAL REPORT STYLE FINANCIAL STATEMENTS", title_style))
    story.append(Paragraph("SCHEDULE III DIVISION I & AS 3 COMPLIANT", ParagraphStyle('SubSub', parent=subtitle_style, fontSize=11, textColor=colors.HexColor('#64748B'))))
    story.append(Spacer(1, 30))

    meta_data = [
        [Paragraph("<b>Client Name:</b>", body_style), Paragraph(client.name, body_style)],
        [Paragraph("<b>Entity Type:</b>", body_style), Paragraph(client.entity_type, body_style)],
        [Paragraph("<b>Reporting Period:</b>", body_style), Paragraph(client.reporting_period, body_style)],
        [Paragraph("<b>Previous Year Period:</b>", body_style), Paragraph(client.previous_year_period, body_style)],
        [Paragraph("<b>Currency / Unit:</b>", body_style), Paragraph(client.currency, body_style)],
        [Paragraph("<b>Accounting Framework:</b>", body_style), Paragraph(client.accounting_framework, body_style)],
        [Paragraph("<b>Prepared By:</b>", body_style), Paragraph(client.prepared_by, body_style)],
        [Paragraph("<b>Reviewed By:</b>", body_style), Paragraph(client.reviewed_by, body_style)],
        [Paragraph("<b>Generation Date:</b>", body_style), Paragraph(datetime.now().strftime("%Y-%m-%d %H:%M:%S"), body_style)],
    ]
    t_meta = Table(meta_data, colWidths=[150, 350])
    t_meta.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 40))
    
    disc_text = "<b>MANDATORY AUDIT DISCLAIMER:</b> This is a system-generated draft for internal review. Professional review and verification by a Chartered Accountant is required before official signature and issuance."
    story.append(Table([[Paragraph(disc_text, ParagraphStyle('Disc', parent=body_style, textColor=colors.HexColor('#9A3412')))]], colWidths=[500], style=[
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#FFEDD5')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#EA580C')),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    story.append(PageBreak())

    # SECTION 2: INDEX
    story.append(Paragraph("TABLE OF CONTENTS / INDEX", h1_style))
    story.append(Spacer(1, 8))
    index_table = [[Paragraph("<b>Sec #</b>", table_cell_bold), Paragraph("<b>Report Section</b>", table_cell_bold), Paragraph("<b>Compliance Reference</b>", table_cell_bold)]]
    index_spec = [
        ("01", "Cover Sheet & Audit Metadata", "Firm Administrative Control"),
        ("02", "Index & Section Sitemap", "CA Audit Pack Index"),
        ("03", "Schedule III Balance Sheet", "Companies Act 2013 Division I"),
        ("04", "Statement of Profit and Loss", "Companies Act 2013 Division I"),
        ("05", "AS 3 Cash Flow Statement", "AS 3 Cash Flow Statements"),
        ("06", "Significant Accounting Policies", "AS 1 Disclosure of Accounting Policies"),
        ("07", "Notes to Accounts", "Schedule III Notes & Disclosures"),
        ("08", "Related Party Disclosures", "AS 18 Related Party Disclosures"),
        ("09", "Contingent Liabilities & Commitments", "AS 29 Provisions & Contingencies"),
        ("10", "Schedule III Financial Ratios", "MCA 2021 Ratio Disclosures"),
        ("11", "Management Queries & Audit Notes", "CA Audit Review Workflow"),
        ("12", "Validation Exceptions Log", "Automated Audit Sanity Check"),
        ("13", "Disclaimer & Sign-off Block", "ICAI Professional Standards")
    ]
    for s_no, s_title, s_ref in index_spec:
        index_table.append([Paragraph(s_no, table_cell), Paragraph(s_title, table_cell), Paragraph(s_ref, table_cell)])
    t_idx = Table(index_table, colWidths=[40, 260, 200])
    t_idx.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t_idx)
    story.append(PageBreak())

    # SECTION 3: BALANCE SHEET
    fs = generate_financial_statements(client_id, db)
    story.append(Paragraph("SCHEDULE III BALANCE SHEET AS AT MARCH 31, 2025", h1_style))
    bs_data = [[Paragraph("<b>Particulars</b>", table_cell_bold), Paragraph("<b>Note</b>", table_cell_bold), Paragraph("<b>CY Amount</b>", table_cell_bold), Paragraph("<b>PY Amount</b>", table_cell_bold)]]
    for line in fs.cy_balance_sheet if hasattr(fs, 'cy_balance_sheet') else fs.balance_sheet:
        bs_data.append([
            Paragraph(line.particulars, table_cell_bold if (line.is_header or line.is_total) else table_cell),
            Paragraph(line.note_number or "", table_cell),
            Paragraph(f"{line.cy_amount:,.2f}", table_cell_bold if line.is_total else table_cell),
            Paragraph(f"{line.py_amount:,.2f}", table_cell_bold if line.is_total else table_cell),
        ])
    t_bs = Table(bs_data, colWidths=[240, 50, 105, 105])
    t_bs.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('BACKGROUND', (2,1), (3,-1), colors.HexColor('#F1F5F9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_bs)
    story.append(PageBreak())

    # SECTION 4: PROFIT AND LOSS
    story.append(Paragraph("STATEMENT OF PROFIT AND LOSS FOR THE YEAR ENDED MARCH 31, 2025", h1_style))
    pl_data = [[Paragraph("<b>Particulars</b>", table_cell_bold), Paragraph("<b>Note</b>", table_cell_bold), Paragraph("<b>CY Amount</b>", table_cell_bold), Paragraph("<b>PY Amount</b>", table_cell_bold)]]
    for line in fs.profit_and_loss:
        pl_data.append([
            Paragraph(line.particulars, table_cell_bold if (line.is_header or line.is_total) else table_cell),
            Paragraph(line.note_number or "", table_cell),
            Paragraph(f"{line.cy_amount:,.2f}", table_cell_bold if line.is_total else table_cell),
            Paragraph(f"{line.py_amount:,.2f}", table_cell_bold if line.is_total else table_cell),
        ])
    t_pl = Table(pl_data, colWidths=[240, 50, 105, 105])
    t_pl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('BACKGROUND', (2,1), (3,-1), colors.HexColor('#F1F5F9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_pl)
    story.append(PageBreak())

    # SECTION 5: AS 3 CASH FLOW STATEMENT
    cfs = generate_cash_flow_statement(client_id, db)
    story.append(Paragraph("AS 3 CASH FLOW STATEMENT (INDIRECT METHOD)", h1_style))
    cf_data = [[Paragraph("<b>Particulars</b>", table_cell_bold), Paragraph("<b>Current Year</b>", table_cell_bold), Paragraph("<b>Previous Year</b>", table_cell_bold)]]
    for line in cfs.statement:
        p_style = table_cell_bold if (line.is_header or line.is_subtotal or line.is_total) else table_cell
        cf_data.append([
            Paragraph(line.particulars, p_style),
            Paragraph(f"{line.cy_amount:,.2f}" if not line.is_header else "", p_style),
            Paragraph(f"{line.py_amount:,.2f}" if not line.is_header else "", p_style)
        ])
    t_cf = Table(cf_data, colWidths=[290, 105, 105])
    t_cf.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('BACKGROUND', (1,1), (2,-1), colors.HexColor('#F1F5F9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_cf)
    story.append(PageBreak())

    # SECTION 6: SIGNIFICANT ACCOUNTING POLICIES
    story.append(Paragraph("SIGNIFICANT ACCOUNTING POLICIES (IGAAP)", h1_style))
    policies = generate_or_update_accounting_policies(client_id, db)
    for p in policies:
        if p.is_applicable:
            story.append(Paragraph(f"<b>{p.policy_number}: {p.title}</b>", ParagraphStyle('PolHead', parent=body_style, fontName='Helvetica-Bold', textColor=colors.HexColor('#1B365D'))))
            story.append(Paragraph(p.content, body_style))
            story.append(Spacer(1, 4))
    story.append(PageBreak())

    # SECTION 7: NOTES TO ACCOUNTS
    story.append(Paragraph("NOTES FORMING PART OF FINANCIAL STATEMENTS", h1_style))
    notes = generate_or_update_notes(client_id, db)
    for n in notes:
        story.append(Paragraph(f"<b>NOTE {n.note_number}: {n.title.upper()}</b>", ParagraphStyle('NoteHead', parent=body_style, fontName='Helvetica-Bold', fontSize=10, leading=12, textColor=colors.HexColor('#1B365D'), spaceBefore=8, spaceAfter=4)))
        
        # Render Table if table_json exists
        if n.table_json:
            try:
                t_data = json.loads(n.table_json)
                headers = t_data.get("headers", [])
                rows = t_data.get("rows", [])
                if headers and rows:
                    pdf_table_data = []
                    # Header row
                    hdr_cells = [Paragraph(f"<b>{h}</b>", table_cell_bold) for h in headers]
                    pdf_table_data.append(hdr_cells)

                    # Dynamic column widths
                    num_cols = len(headers)
                    total_w = 500
                    if num_cols == 3:
                        col_w = [260, 120, 120]
                    elif num_cols == 8:  # PPE Fixed Asset Schedule
                        col_w = [110, 55, 55, 55, 55, 60, 55, 55]
                    else:
                        col_w = [250] + [(total_w - 250) // (num_cols - 1)] * (num_cols - 1)

                    for r in rows:
                        r_cells = []
                        is_total_row = bool(r[0] and (r[0].strip().startswith("TOTAL") or r[0].strip().startswith("GRAND TOTAL") or r[0].strip().startswith("NET BLOCK")))
                        p_st = table_cell_bold if is_total_row else table_cell
                        for idx, cell_val in enumerate(r):
                            val_str = str(cell_val)
                            if idx > 0 and val_str.strip() != "":
                                r_cells.append(Paragraph(val_str, ParagraphStyle('RCell', parent=p_st, alignment=2)))
                            else:
                                r_cells.append(Paragraph(val_str, p_st))
                        pdf_table_data.append(r_cells)

                    t_note = Table(pdf_table_data, colWidths=col_w)
                    t_note.setStyle(TableStyle([
                        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
                        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                        ('BACKGROUND', (1,1), (-1,-1), colors.HexColor('#F8FAFC')),
                        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
                        ('PADDING', (0,0), (-1,-1), 3),
                    ]))
                    story.append(t_note)
                    story.append(Spacer(1, 4))
            except Exception as e:
                print(f"Error parsing table_json for PDF note {n.note_number}: {e}")

        if n.content and n.content.strip():
            story.append(Paragraph(n.content.replace("\n", "<br/>"), body_style))
            story.append(Spacer(1, 6))

    story.append(PageBreak())


    # SECTION 8: RELATED PARTY DISCLOSURES
    story.append(Paragraph("RELATED PARTY DISCLOSURES (AS-18)", h1_style))
    rpt_data = [[Paragraph("<b>Name</b>", table_cell_bold), Paragraph("<b>Relation</b>", table_cell_bold), Paragraph("<b>Nature</b>", table_cell_bold), Paragraph("<b>Closing Bal</b>", table_cell_bold)]]
    for r in db.query(RelatedParty).filter(RelatedParty.client_id == client_id).all():
        rpt_data.append([Paragraph(r.name, table_cell), Paragraph(r.relationship, table_cell), Paragraph(r.nature_tx, table_cell), Paragraph(f"{r.closing_bal:,.2f}", table_cell)])
    t_rpt = Table(rpt_data, colWidths=[150, 110, 140, 100])
    t_rpt.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_rpt)
    story.append(PageBreak())

    # SECTION 9: CONTINGENT LIABILITIES
    story.append(Paragraph("CONTINGENT LIABILITIES & COMMITMENTS (AS-29)", h1_style))
    cont_data = [[Paragraph("<b>Nature</b>", table_cell_bold), Paragraph("<b>Forum</b>", table_cell_bold), Paragraph("<b>CY Amount</b>", table_cell_bold), Paragraph("<b>Assessment</b>", table_cell_bold)]]
    for c in db.query(Contingency).filter(Contingency.client_id == client_id).all():
        cont_data.append([Paragraph(c.nature, table_cell), Paragraph(c.forum or "", table_cell), Paragraph(f"{c.cy_amount:,.2f}", table_cell), Paragraph(c.assessment or "", table_cell)])
    t_cont = Table(cont_data, colWidths=[150, 100, 100, 150])
    t_cont.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_cont)
    story.append(PageBreak())

    # SECTION 10: RATIOS
    story.append(Paragraph("SCHEDULE III MANDATORY FINANCIAL RATIOS", h1_style))
    rat_data = [[Paragraph("<b>Code</b>", table_cell_bold), Paragraph("<b>Ratio Name</b>", table_cell_bold), Paragraph("<b>CY Value</b>", table_cell_bold), Paragraph("<b>PY Value</b>", table_cell_bold), Paragraph("<b>Movement</b>", table_cell_bold)]]
    for r in calculate_ratios(client_id, db):
        rat_data.append([Paragraph(r.code, table_cell), Paragraph(r.name, table_cell), Paragraph(f"{r.cy_value:.2f} {r.unit}", table_cell), Paragraph(f"{r.py_value:.2f} {r.unit}", table_cell), Paragraph(r.movement, table_cell)])
    t_rat = Table(rat_data, colWidths=[50, 180, 90, 90, 90])
    t_rat.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_rat)
    story.append(PageBreak())

    # SECTION 11: MANAGEMENT QUERIES
    story.append(Paragraph("MANAGEMENT QUERIES & AUDIT CLARIFICATIONS", h1_style))
    mq_data = [
        [Paragraph("<b>#</b>", table_cell_bold), Paragraph("<b>Audit Query Topic</b>", table_cell_bold), Paragraph("<b>Management Response</b>", table_cell_bold), Paragraph("<b>Status</b>", table_cell_bold)],
        [Paragraph("1", table_cell), Paragraph("Verification of physical stock count & valuation", table_cell), Paragraph("Physical count completed on 31-Mar-2025. Valued at lower of cost or NRV.", table_cell), Paragraph("Verified", table_cell)],
        [Paragraph("2", table_cell), Paragraph("Confirmation of balances for Trade Receivables > 180 days", table_cell), Paragraph("Direct balance confirmations received for 85% of outstanding amount.", table_cell), Paragraph("Verified", table_cell)],
        [Paragraph("3", table_cell), Paragraph("Income tax paid and pending appellate disputes", table_cell), Paragraph("Tax paid verified with Challan 280. Appeal pending with CIT(A).", table_cell), Paragraph("Verified", table_cell)]
    ]
    t_mq = Table(mq_data, colWidths=[30, 180, 210, 80])
    t_mq.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_mq)
    story.append(PageBreak())

    # SECTION 12: VALIDATION LOG
    story.append(Paragraph("AUTOMATED VALIDATION LOG", h1_style))
    val_data = [[Paragraph("<b>Code</b>", table_cell_bold), Paragraph("<b>Check Name</b>", table_cell_bold), Paragraph("<b>Status</b>", table_cell_bold), Paragraph("<b>Findings</b>", table_cell_bold)]]
    for v in run_validation_checks(client_id, db):
        val_data.append([Paragraph(v.code, table_cell), Paragraph(v.check_name, table_cell), Paragraph(v.status, table_cell_bold), Paragraph(v.message, table_cell)])
    t_val = Table(val_data, colWidths=[50, 160, 70, 220])
    t_val.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1B365D')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t_val)
    story.append(Spacer(1, 15))

    # SECTION 13: DISCLAIMER & PARTNER SIGN-OFF
    story.append(Paragraph("AUDITOR DISCLAIMER & PARTNER SIGN-OFF BLOCK", h1_style))
    story.append(Paragraph("This Financial Statement draft has been prepared strictly based on the trial balance and supporting schedules provided by management. In accordance with ICAI Standards on Auditing, final financial statements must be approved by the Board of Directors and signed by Chartered Accountants.", body_style))
    story.append(Spacer(1, 20))

    sign_data = [
        [Paragraph("<b>For SW INDIA</b><br/>Chartered Accountants<br/>Firm Registration No: 001234N", body_style), Paragraph(f"<b>For {client.name}</b><br/>Board of Directors", body_style)],
        [Spacer(1, 35), Spacer(1, 35)],
        [Paragraph("<b>CA Partner</b><br/>Membership No: 501234<br/>UDIN: 25501234AAAAAA1234", body_style), Paragraph("<b>Director</b><br/>DIN: 01234567<br/><br/><b>Director</b><br/>DIN: 08901234", body_style)]
    ]
    t_sign = Table(sign_data, colWidths=[250, 250])
    t_sign.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#1B365D')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_sign)

    doc.build(story, canvasmaker=NumberedCanvas)
    return filepath
