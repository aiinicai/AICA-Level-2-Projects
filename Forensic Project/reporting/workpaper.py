"""
Multi-sheet Excel Working Paper Export for Red Flag Engine.
Generates an audit-ready 9-sheet Excel workbook:
1. Summary
2. Exceptions_TB
3. Exceptions_LG
4. Exceptions_FS
5. Models
6. Derived_Statements
7. Hypothesis_Register
8. Coverage
9. Rule_Contributions
10. Suppressed
11. Custody_Log
"""
import io
import json
import datetime
from typing import Dict, List, Any, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from reporting.hypotheses import build_hypothesis_register

EXC_COLS = ["rule_id", "rule_name", "flag", "fy_span", "subject", "observed_value",
            "threshold_value", "occurrences", "materiality_factor", "flag_score",
            "ml_outlier_score", "branch", "scheme", "source", "detail"]


def _slice(df, cols):
    """Return df restricted to cols, tolerating any that are absent."""
    if df is None or df.empty:
        return pd.DataFrame(columns=cols)
    return df[[c for c in cols if c in df.columns]]


def generate_excel_workpaper(
    scoring_result: Dict[str, Any],
    derived: pd.DataFrame,
    coverage_result: Dict[str, Any],
    custody_entry: Dict[str, Any],
    client_name: str = "Client Engagement"
) -> bytes:
    """
    Generate complete 9-sheet audit working paper workbook as bytes.
    """
    output = io.BytesIO()
    scored_df = scoring_result.get("scored_exceptions", pd.DataFrame())
    
    # Style definitions
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F497D")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        wb = writer.book
        
        # 1. Sheet: Summary
        st_ = scoring_result.get("stats", {})
        summary_rows = [
            {"Engagement Metric": "Client Name", "Value": client_name},
            {"Engagement Metric": "Audit Firm", "Value": custody_entry.get("parameters", {}).get("firm_name", "")},
            {"Engagement Metric": "Lead Operator", "Value": custody_entry.get("operator", "Forensic Auditor")},
            {"Engagement Metric": "Run ID", "Value": custody_entry.get("run_id", "")},
            {"Engagement Metric": "Report Generated (UTC)", "Value": datetime.datetime.now(datetime.timezone.utc).isoformat()},
            {"Engagement Metric": "Predication Note", "Value": custody_entry.get("predication_note", "")},
            {"Engagement Metric": "Performance Materiality", "Value": custody_entry.get("parameters", {}).get("performance_materiality", "")},
            {"Engagement Metric": "", "Value": ""},
            {"Engagement Metric": "ENTITY RISK SCORE (0-100)", "Value": scoring_result.get("entity_score", 0.0)},
            {"Engagement Metric": "Risk Classification Bucket", "Value": scoring_result.get("bucket", "GREEN")},
            {"Engagement Metric": "Recommended Audit Action", "Value": scoring_result.get("bucket_action", "")},
            {"Engagement Metric": "Bucket Thresholds", "Value": "RED >= 40 | YELLOW 18-40 | GREEN < 18"},
            {"Engagement Metric": "Score Numerator (weighted contributions)", "Value": scoring_result.get("raw_weighted_sum", 0.0)},
            {"Engagement Metric": "Score Denominator (executed rule battery)", "Value": scoring_result.get("score_denominator", 0.0)},
            {"Engagement Metric": "Score before Governance Overlay", "Value": scoring_result.get("entity_score_pre_governance", 0.0)},
            {"Engagement Metric": "Governance Assessment Status", "Value": scoring_result.get("governance_status", "not assessed")},
            {"Engagement Metric": "Governance Overlay Factor", "Value": scoring_result.get("governance_factor", 1.0)},
            {"Engagement Metric": "Green Flag Score (0-100, never netted)", "Value": scoring_result.get("green_score", 0.0)},
            {"Engagement Metric": "", "Value": ""},
            {"Engagement Metric": "Raw Exception Instances", "Value": st_.get("raw_instances", len(scored_df))},
            {"Engagement Metric": "After De-duplication (rule x subject)", "Value": st_.get("after_dedup", "")},
            {"Engagement Metric": "Audit Leads Retained", "Value": st_.get("retained", len(scored_df))},
            {"Engagement Metric": "Suppressed as Systemic", "Value": st_.get("suppressed", 0)},
            {"Engagement Metric": "Max Subjects Reported per Rule", "Value": st_.get("max_instances_per_rule", "")},
            {"Engagement Metric": "Distinct Rules Fired", "Value": st_.get("distinct_rules_fired", "")},
            {"Engagement Metric": "Distinct Subjects Flagged", "Value": st_.get("distinct_subjects", "")},
            {"Engagement Metric": "Available Forensic Methods", "Value": f"{coverage_result.get('available_count', 0)} / {coverage_result.get('total_implemented', 0)}"},
            {"Engagement Metric": "", "Value": ""},
            {"Engagement Metric": "PROFESSIONAL STANDARD NOTICE", "Value": "Indicators are not evidence. This working paper establishes predication under ICAI FAIS 130 / SA 240; it does not conclude that an offence has occurred."},
        ]
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        
        # 2. Sheet: Exceptions_TB
        tb_df = scored_df[scored_df["module"] == "TB"] if not scored_df.empty and "module" in scored_df.columns else pd.DataFrame()
        if tb_df.empty:
            tb_df = pd.DataFrame(columns=EXC_COLS)
        else:
            tb_df = _slice(tb_df, EXC_COLS)
        tb_df.to_excel(writer, sheet_name="Exceptions_TB", index=False)
        
        # 3. Sheet: Exceptions_LG
        lg_df = scored_df[scored_df["module"] == "LG"] if not scored_df.empty and "module" in scored_df.columns else pd.DataFrame()
        if lg_df.empty:
            lg_df = pd.DataFrame(columns=EXC_COLS)
        else:
            lg_df = _slice(lg_df, EXC_COLS)
        lg_df.to_excel(writer, sheet_name="Exceptions_LG", index=False)
        
        # 4. Sheet: Exceptions_FS
        fs_df = scored_df[scored_df["module"] == "FS"] if not scored_df.empty and "module" in scored_df.columns else pd.DataFrame()
        if fs_df.empty:
            fs_df = pd.DataFrame(columns=EXC_COLS)
        else:
            fs_df = _slice(fs_df, EXC_COLS)
        fs_df.to_excel(writer, sheet_name="Exceptions_FS", index=False)
        
        # 5. Sheet: Models
        ms_df = scored_df[scored_df["module"] == "MS"] if not scored_df.empty and "module" in scored_df.columns else pd.DataFrame()
        if ms_df.empty:
            ms_df = pd.DataFrame(columns=EXC_COLS)
        else:
            ms_df = _slice(ms_df, EXC_COLS)
        ms_df.to_excel(writer, sheet_name="Models", index=False)
        
        # 6. Sheet: Derived_Statements
        derived.to_excel(writer, sheet_name="Derived_Statements", index=False)
        
        # 7. Sheet: Hypothesis_Register
        hyp_df = build_hypothesis_register(scored_df)
        hyp_df.to_excel(writer, sheet_name="Hypothesis_Register", index=False)
        
        # 8. Sheet: Coverage
        cov_rows = []
        for m in coverage_result.get("implemented", []):
            cov_rows.append({
                "Method ID": m.get("id"),
                "Method Name": m.get("name"),
                "Status": m.get("status"),
                "Execution Feasibility": m.get("reasons"),
                "Associated Rules": m.get("rules")
            })
        for m in coverage_result.get("not_implemented", []):
            cov_rows.append({
                "Method ID": m.get("id"),
                "Method Name": m.get("name"),
                "Status": m.get("status"),
                "Execution Feasibility": f"Blocked by: {m.get('blocked_by')}",
                "Associated Rules": f"Unlocked by: {m.get('unlocked_by')}"
            })
        pd.DataFrame(cov_rows).to_excel(writer, sheet_name="Coverage", index=False)
        
        # 8b. Sheet: Rule_Contributions — audit trail behind the headline score
        contrib = scoring_result.get("rule_contributions", pd.DataFrame())
        if contrib is None or contrib.empty:
            contrib = pd.DataFrame(columns=["rule_id", "rule_name", "module", "flag", "weight",
                                            "confidence", "subjects_hit", "materiality_factor",
                                            "pervasiveness", "contribution", "max_possible"])
        contrib.to_excel(writer, sheet_name="Rule_Contributions", index=False)

        # 8c. Sheet: Suppressed — systemic hits below the per-rule reporting cap
        supp = scoring_result.get("suppressed_exceptions", pd.DataFrame())
        _slice(supp, EXC_COLS).to_excel(writer, sheet_name="Suppressed", index=False)

        # 9. Sheet: Custody_Log
        custody_rows = [
            {"Custody Field": "Run ID", "Details": custody_entry.get("run_id")},
            {"Custody Field": "Timestamp UTC", "Details": custody_entry.get("timestamp_utc")},
            {"Custody Field": "Operator", "Details": custody_entry.get("operator")},
            {"Custody Field": "Predication Note", "Details": custody_entry.get("predication_note")},
            {"Custody Field": "Source Files Uploaded", "Details": json.dumps(custody_entry.get("files", []), indent=2)},
            {"Custody Field": "Rule Versions (SHA-256)", "Details": json.dumps(custody_entry.get("rule_versions", {}), indent=2)},
            {"Custody Field": "Parameters", "Details": json.dumps(custody_entry.get("parameters", {}), indent=2)},
            {"Custody Field": "Confirmations", "Details": json.dumps(custody_entry.get("confirmations", []), indent=2)},
        ]
        pd.DataFrame(custody_rows).to_excel(writer, sheet_name="Custody_Log", index=False)

    # Format styling with openpyxl
    output.seek(0)
    wb = openpyxl.load_workbook(output)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        ws.views.sheetView[0].showGridLines = True
        
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = regular_font
                cell.border = thin_border
                
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 60)
            
    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()
