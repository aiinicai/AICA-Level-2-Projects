"""
Script to generate the standardized Trial Balance Excel Template for Red Flag Engine.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def make_template(output_path: str = "templates/trial_balance_template.xlsx"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb = openpyxl.Workbook()
    
    # ------------------ Sheet 1: Instructions ------------------
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info.views.sheetView[0].showGridLines = True
    
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    accent_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    ws_info.merge_cells("B2:H2")
    ws_info["B2"] = "Red Flag Engine — Standard Trial Balance Template"
    ws_info["B2"].font = title_font
    
    instructions = [
        ("Overview", "This workbook serves as the standardized input format for the Red Flag Forensic Accounting Engine."),
        ("Required Years", "At least 3 financial years must be supplied (e.g. FY22, FY23, FY24) for comparative forensic modeling."),
        ("Format Options", "You may supply data in Long format (Sheet: 'Trial_Balance_Long') or Multi-sheet format (one sheet per year named 'FY22', 'FY23', 'FY24')."),
        ("Schedule III Groups", "Every ledger must have an assigned Schedule III Primary Group (e.g., 'Sundry Debtors', 'Bank Accounts', 'Direct Expenses')."),
        ("Opening Balances", "Opening debit/credit balances are mandatory for all years to enable Indirect Cash Flow Derivation (cfo_indirect)."),
        ("Arithmetic Integrity", "Sum of Debits must equal Sum of Credits for each financial year. Unbalanced trial balances will be flagged under TB-01."),
        ("File Formats", "Save as .xlsx or .xls. The engine operates completely locally and never modifies original source files.")
    ]
    
    row_idx = 4
    for title, desc in instructions:
        ws_info.cell(row=row_idx, column=2, value=title).font = bold_font
        ws_info.cell(row=row_idx, column=2).fill = accent_fill
        ws_info.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=8)
        ws_info.cell(row=row_idx, column=3, value=desc).font = regular_font
        row_idx += 2
        
    ws_info.column_dimensions["B"].width = 24
    for c in ["C", "D", "E", "F", "G", "H"]:
        ws_info.column_dimensions[c].width = 15

    # ------------------ Sheet 2: Long Format Template ------------------
    ws_data = wb.create_sheet(title="Trial_Balance_Long")
    ws_data.views.sheetView[0].showGridLines = True
    
    headers = [
        "Financial Year", "Ledger Name", "Group", "Sub Group",
        "Opening Dr (Rs.)", "Opening Cr (Rs.)",
        "Turnover Dr (Rs.)", "Turnover Cr (Rs.)",
        "Closing Dr (Rs.)", "Closing Cr (Rs.)"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    sample_rows = [
        ("FY22", "HDFC Bank A/c 5020001", "Bank Accounts", "Current Accounts", 500000.0, 0.0, 12000000.0, 11800000.0, 700000.0, 0.0),
        ("FY22", "Sundry Debtors — Acme Corp", "Sundry Debtors", "Domestic Trade", 850000.0, 0.0, 4500000.0, 4200000.0, 1150000.0, 0.0),
        ("FY22", "Sundry Creditors — Metal Supply Ltd", "Sundry Creditors", "Raw Materials", 0.0, 600000.0, 3800000.0, 4100000.0, 0.0, 900000.0),
        ("FY22", "Domestic Sales 18%", "Revenue from Operations", "Domestic Sales", 0.0, 0.0, 0.0, 4500000.0, 0.0, 4500000.0),
        ("FY22", "Raw Material Purchases", "Cost of Materials Consumed", "Direct Materials", 0.0, 0.0, 3550000.0, 0.0, 3550000.0, 0.0),
        ("FY23", "HDFC Bank A/c 5020001", "Bank Accounts", "Current Accounts", 700000.0, 0.0, 14000000.0, 13900000.0, 800000.0, 0.0),
        ("FY24", "HDFC Bank A/c 5020001", "Bank Accounts", "Current Accounts", 800000.0, 0.0, 16000000.0, 15850000.0, 950000.0, 0.0),
    ]
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    for r_idx, row in enumerate(sample_rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
            cell.font = regular_font
            cell.border = thin_border
            if c_idx >= 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws_data.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 4, 16)
        
    wb.save(output_path)
    print(f"Created template at: {output_path}")

if __name__ == "__main__":
    make_template()
