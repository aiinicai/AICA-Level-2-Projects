"""
Excel Writer — writes the final RACM rows to a formatted .xlsx file,
plus a .csv backup, using the locked 17-column schema.
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from racm_schema import RACM_COLUMNS


def write_racm_to_excel(rows, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "RACM"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, col_name in enumerate(RACM_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(RACM_COLUMNS, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_name == "Design Deficiency (Yes/No)" and str(value).strip().lower() == "yes":
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    widths = [8, 20, 20, 10, 35, 35, 35, 22, 14, 16, 14, 12, 35, 14, 35, 22, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Excel file saved: {output_path}")


def write_racm_to_csv(rows, output_path):
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=RACM_COLUMNS)
        writer.writeheader()
        for row in rows:
            writer.writerow({col: row.get(col, "") for col in RACM_COLUMNS})
    print(f"CSV file saved: {output_path}")