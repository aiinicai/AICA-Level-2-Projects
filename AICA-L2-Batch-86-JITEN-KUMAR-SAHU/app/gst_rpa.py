"""
=============================================================
  ClientLedger India — GST RPA Server  v2.0
  Local server on http://localhost:8765

  REQUIREMENTS:
    python -m pip install flask playwright
    python -m playwright install chromium

  RUN:
    python gst_rpa.py

  Then open ClientLedger India -> GST Filing Status -> RPA Mode
=============================================================
"""

import sys, os

# ── Force UTF-8 console output, with a crash-proof fallback ────────
# Windows consoles default to the 'cp1252' codepage, which cannot
# encode the checkmarks/symbols (✓ ✗ ⏸ ✅ ❌ 💾 ⚠ etc.) used throughout
# this file's ~20 print()/logging call sites — without a fix, any one
# of them can crash whatever thread calls it ("UnicodeEncodeError:
# 'charmap' codec can't encode character..."), which is exactly what
# takes down background workers like the GSTIN Directory enrichment.
#
# Two layers, so this can never happen again even in an environment
# where the first layer doesn't fully apply (e.g. a frozen/windowed
# build's console is not a real reconfigurable TextIOWrapper):
#   1) Try to reconfigure the real stream to UTF-8 — this is what
#      makes the symbols actually display correctly on modern
#      terminals (Windows Terminal, VS Code, etc.).
#   2) Regardless of whether (1) succeeded, wrap the stream so any
#      character that still can't be encoded is silently replaced
#      instead of raising — so a decorative symbol can never crash
#      a worker thread again.
class _SafeStream:
    def __init__(self, stream):
        self._stream = stream
        enc = getattr(stream, "encoding", None) or "utf-8"
        self._encoding = enc

    def write(self, s):
        try:
            return self._stream.write(s)
        except UnicodeEncodeError:
            safe = s.encode(self._encoding, errors="replace").decode(self._encoding, errors="replace")
            return self._stream.write(safe)
        except Exception:
            return 0

    def flush(self):
        try:
            self._stream.flush()
        except Exception:
            pass

    def __getattr__(self, name):
        return getattr(self._stream, name)


for _stream_name in ("stdout", "stderr"):
    _stream = getattr(sys, _stream_name, None)
    if _stream is None:
        continue
    if hasattr(_stream, "reconfigure"):
        try:
            _stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    try:
        setattr(sys, _stream_name, _SafeStream(_stream))
    except Exception:
        pass

# ── Startup diagnostics ───────────────────────────────────────
print("=" * 56)
print("  ClientLedger India — GST RPA Server v2.0")
print("=" * 56)

errors = []
try:
    import flask
    print("  ✓ Flask        OK")
except ImportError:
    errors.append("Flask not found. Run:  python -m pip install flask")
    print("  ✗ Flask        MISSING")

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    PLAYWRIGHT_AVAILABLE = True
    print("  ✓ Playwright   OK")
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    errors.append("Playwright not found. Run:  python -m pip install playwright  "
                  "then  python -m playwright install chromium")
    print("  ✗ Playwright   MISSING")

if errors:
    print()
    for e in errors:
        print(f"  → {e}")
    print()
    input("  Press Enter to exit...")
    sys.exit(1)

import base64, threading, time, json, logging, io
from datetime import datetime
from flask import Flask, jsonify, request, Response, send_file

import config as app_config
import dbstore

# ── Resolve the base data folder (runs the first-run folder-picker
#    wizard the very first time the app is launched) and derive every
#    working path from it. Everything below that used to be hardcoded
#    relative to this script now reads from PATHS instead. ──────────
PATHS = app_config.Paths.load()
dbstore.init(PATHS.db_file)

# Log to BOTH console and a persistent file — console alone is useless
# once packaged into a windowed .exe (nothing displays it), and without
# a FileHandler here, log.info()/log.error() calls never reach disk at
# all, regardless of how many diagnostic lines get added elsewhere.
_log_file_path = PATHS.log_file
os.makedirs(os.path.dirname(_log_file_path), exist_ok=True)
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s %(levelname)s %(message)s',
                    handlers=[
                        logging.StreamHandler(sys.stdout),
                        logging.FileHandler(_log_file_path, encoding="utf-8"),
                    ])
log = logging.getLogger('gst_rpa')
log.info(f"Data folder: {PATHS.base}")
log.info(f"Database:    {PATHS.db_file}")

# Global safety net: log ANY uncaught exception from ANY background
# thread (the app spawns many — GSTR1/2A/2B/3B, TDS, GSTIN Directory
# enrichment, combined download, etc.). Python's default behavior on
# an uncaught thread exception is to print a traceback to sys.stderr —
# but in a frozen/windowed build (console=False) there is no real
# stderr for that to go to, so it vanishes completely, leaving no
# trace anywhere that anything went wrong. This turns that silent
# failure into a logged, diagnosable one for every thread, not just
# the two worker functions that already got dedicated try/except
# wrappers for this same reason.
def _log_uncaught_thread_exception(args):
    import traceback as _tb
    tb_text = "".join(_tb.format_exception(args.exc_type, args.exc_value, args.exc_traceback))
    log.error(f"[UNCAUGHT] Thread '{args.thread.name}' died with an unhandled "
              f"exception:\n{tb_text}")

threading.excepthook = _log_uncaught_thread_exception

# Build fingerprint — print this to console AND write it to the log
# file at every startup, so it's trivial to confirm whether a given
# build actually contains a given round of fixes, instead of guessing.
# Bump the date/tag here whenever a meaningful fix goes in.
BUILD_FINGERPRINT = "2026-08-31-d (fix: DevTools no longer auto-opens on every launch — now opt-in via CLIENTLEDGER_DEBUG=1 env var, since the origin-mismatch bug it helped diagnose is now fixed and confirmed)"
print(f"  Build: {BUILD_FINGERPRINT}")
log.info(f"Build fingerprint: {BUILD_FINGERPRINT}")

# ── Flask app ─────────────────────────────────────────────────
app = Flask(__name__)

def _cors(response):
    response.headers['Access-Control-Allow-Origin']  = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    response.headers['Cache-Control']                = 'no-store'
    return response

@app.after_request
def after_request(r): return _cors(r)

@app.before_request
def handle_options():
    if request.method == 'OPTIONS':
        return _cors(Response(status=204))

# ── Excel export: save to disk (in the relevant GSTR folder) AND stream
#    to the browser, so the file is always traceable on disk regardless
#    of whatever the browser's own download settings happen to be. ────
def save_excel_and_respond(buf, save_dir, fname, mimetype=None, extra_header=True, log_fn=None):
    """
    buf: a BytesIO already containing the .xlsx bytes (as returned by
         the various *_json_to_excel() builders below).
    save_dir: absolute folder to save the permanent copy into — pass
         the same <GSTRx dir>/<gstin>/<fy> folder the raw downloads
         for that client/period already live in, so everything for a
         given GSTIN+FY sits together.
    log_fn: the module's own activity-log function (g2a_log, g2b_log,
         g1_log, g3b_log, tds_log) so "Saved to: ..." shows up right
         in the on-screen activity panel the user is already watching
         — not just in the console/log file, which they'd never see.
    """
    mimetype = mimetype or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    os.makedirs(save_dir, exist_ok=True)
    saved_path = os.path.join(save_dir, fname)
    buf.seek(0)
    data = buf.read()
    with open(saved_path, "wb") as f:
        f.write(data)
    log.info(f"Excel export saved: {saved_path}")
    if log_fn:
        try:
            log_fn(f"📁 Excel saved to: {saved_path}", "info")
        except Exception:
            pass

    resp = send_file(
        io.BytesIO(data),
        mimetype=mimetype,
        as_attachment=True,
        download_name=fname,
    )
    if extra_header:
        # Front-end reads this to show "Saved to: <path>" — browsers'
        # own download-location settings are otherwise the only place
        # this info would appear (or not appear at all).
        resp.headers["X-Saved-Path"] = saved_path
        resp.headers["Access-Control-Expose-Headers"] = "X-Saved-Path"
    return resp

# ── Shared state ──────────────────────────────────────────────
lock  = threading.Lock()
state = {
    "status":         "idle",
    "current_client": None,
    "captcha_image":  None,
    "captcha_answer": None,
    "otp_answer":     None,
    "results":        [],
    "log":            [],
    "error":          None,
    "progress":       0,
    "start_time":     None,
    "user_continue":  False,
}

# ── Constants ─────────────────────────────────────────────────
GST_LOGIN_URL  = "https://services.gst.gov.in/services/login"
GST_PORTAL_URL = "https://services.gst.gov.in"

SELECTORS = {
    "username": ["#user_name", "input[name='user_name']", "input[placeholder*='sername']"],
    "password": ["#user_pass",  "input[name='user_pass']",  "input[type='password']"],
    "captcha_input": [
        "input[name='captcha']",
        "input[placeholder*='Characters']",
        "input[placeholder*='characters']",
        "input[placeholder*='shown below']",
        "input[placeholder*='aptcha']",
        "input[id*='captcha']",
        "input[id*='Captcha']",
        "input[name*='captchaAnswer']",
        "#captchaAnswer",
    ],
    "captcha_image": ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                      "#captchaImage", ".captchaImg"],
    "login_btn": ["#loginBtn", "button[type='submit']", "input[type='submit']",
                  "button:has-text('LOGIN')", "button:has-text('Login')"],
    "otp_input":  ["input[id*='otp']", "input[name*='otp']", "input[placeholder*='OTP']"],
    "otp_submit": ["button:has-text('SUBMIT')", "button:has-text('Verify')",
                   "button:has-text('Submit')"],
}

# ── Activity log ──────────────────────────────────────────────
LOG_FILE  = PATHS.log_file
_full_log = []

def push_log(msg, level="info"):
    ts   = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    with lock:
        state["log"].append(line)
        if len(state["log"]) > 300:
            state["log"] = state["log"][-300:]
        _full_log.append(line)
    # Route through the shared logging.FileHandler (opened once, kept
    # open for the process lifetime) instead of a fresh
    # open(LOG_FILE, "a") on every single call. This function is the
    # most heavily used logger in the app — every RPA flow calls it
    # constantly — so this was the highest-impact place this pattern
    # could cause trouble. Confirmed elsewhere (_comb_log, under real
    # antivirus interference) that repeatedly opening/closing the same
    # file in quick succession can hang indefinitely; that risk now
    # only exists in this one already-open handle, not on every call.
    (log.error if level == "error" else log.info)(msg)

# ── Playwright helpers ────────────────────────────────────────
def wait_for_user(field, timeout_sec=180):
    """Block until ClientLedger UI posts the requested field value."""
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        with lock:
            val = state.get(field)
            if val:
                state[field] = None
                return val
        time.sleep(0.4)
    return None

def find_el(page, selectors, timeout=5000):
    for sel in selectors:
        try:
            el = page.wait_for_selector(sel, timeout=timeout)
            if el: return el
        except Exception:
            continue
    return None

def type_into(page, selectors, text, label="field"):
    """
    Type text into an Angular input using press_sequentially() which
    fires real key events that Angular's ngModel responds to.
    """
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=5000)
            loc.scroll_into_view_if_needed()
            loc.click()
            time.sleep(0.2)
            loc.select_text()
            time.sleep(0.1)
            loc.press("Delete")
            time.sleep(0.1)
            loc.press_sequentially(text, delay=60)
            time.sleep(0.2)
            val = loc.input_value()
            push_log(f"    {label} → selector={sel!r} → value='{val}'")
            if val.strip():
                return True
        except Exception as e:
            push_log(f"    type_into skip {sel!r}: {e}")
    return False

def clk(page, selectors, label, timeout=8000):
    """Click first visible element from selector list."""
    for sel in selectors:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=timeout)
            loc.scroll_into_view_if_needed()
            loc.click()
            push_log(f"  ✓ {label}")
            return True
        except Exception:
            continue
    push_log(f"  ✗ Not found: {label}", "error")
    return False

def capture_captcha_img(page):
    """
    Capture captcha exactly as rendered. Element screenshot first (most reliable),
    then canvas, then full viewport fallback.
    """
    # Method 1: element screenshot of captcha img
    try:
        for sel in (SELECTORS.get("captcha_image", []) +
                    ["#imgCaptcha", "img[id*='aptcha' i]", "form img", "img"]):
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    push_log(f"  ✓ Captcha captured ({sel})")
                    return "data:image/png;base64," + base64.b64encode(png).decode()
            except Exception:
                continue
    except Exception:
        pass

    # Method 2: canvas toDataURL
    try:
        data_url = page.evaluate("""() => {
            let best = null;
            for (const cv of document.querySelectorAll('canvas')) {
                if (cv.width > 30 && cv.height > 10) {
                    if (!best || cv.width * cv.height > best.width * best.height)
                        best = cv;
                }
            }
            return best ? best.toDataURL('image/png') : null;
        }""")
        if data_url and data_url.startswith('data:image'):
            push_log("  ✓ Captcha captured from canvas")
            return data_url
    except Exception:
        pass

    # Method 3: full viewport fallback
    try:
        png = page.screenshot(full_page=False)
        push_log("  ✓ Captcha via full screenshot (fallback)")
        return "data:image/png;base64," + base64.b64encode(png).decode()
    except Exception as e:
        push_log(f"  ✗ Captcha capture failed: {e}", "error")
        return None

def fill_captcha_field(page, text):
    """Fill the captcha text input, with JS-discovery fallback."""
    if type_into(page, SELECTORS["captcha_input"], text, "captcha"):
        return True
    try:
        sel = page.evaluate("""() => {
            for (const img of document.querySelectorAll('img')) {
                if (!(img.src||'').toLowerCase().includes('captcha')) continue;
                let node = img.parentElement;
                for (let i = 0; i < 6 && node; i++) {
                    for (const inp of node.querySelectorAll('input')) {
                        if (inp.readOnly || inp.disabled || inp.type === 'hidden') continue;
                        if (inp.id)   return '#' + inp.id;
                        if (inp.name) return 'input[name="' + inp.name + '"]';
                    }
                    node = node.parentElement;
                }
            }
            return null;
        }""")
        if sel:
            return type_into(page, [sel], text, "captcha-js")
    except Exception:
        pass
    return False

def check_login_success(page):
    """True = logged in, False = failed, None = uncertain."""
    try:
        url = page.url.lower()
        push_log(f"    Current URL: {page.url[:80]}")
        if any(x in url for x in ["fowelcome", "/auth/", "dashboard", "returns", "profile"]):
            return True
        for sel in ["#mnuSideBar", ".userMenu", "#lnkLogout",
                    "a:has-text('Logout')", "a:has-text('Dashboard')"]:
            try:
                if page.query_selector(sel): return True
            except Exception:
                pass
        try:
            body_up = page.locator("body").inner_text(timeout=3000).upper()
            for err in ["INVALID CAPTCHA", "WRONG CAPTCHA", "INCORRECT CAPTCHA",
                        "PLEASE ENTER VALID", "USER LOCKED", "INVALID CREDENTIALS"]:
                if err in body_up:
                    push_log(f"    Login error: {err}")
                    return False
        except Exception:
            pass
        if "services/login" in url:
            return False
    except Exception as ex:
        push_log(f"    check_login_success error: {ex}")
    return None

# ── Core extraction ───────────────────────────────────────────
def nav_to_search_page(page):
    """
    Navigate to Search Taxpayer → Search by GSTIN/UIN tab.
    Called ONCE per session — before the client loop.
    Goes directly to services.gst.gov.in/services/auth/searchtp
    (authenticated URL — avoids www.gst.gov.in redirect delays).
    """
    push_log("  Nav: goto Search Taxpayer (authenticated)...")
    page.goto("https://services.gst.gov.in/services/auth/searchtp",
              wait_until="domcontentloaded", timeout=25000)
    time.sleep(1.5)

    # If redirected to login, session is gone
    if "login" in (page.url or "").lower():
        raise RuntimeError(f"Session expired — redirected to login: {page.url}")

    push_log(f"  Nav: on {page.url[:60]}")

    # Click Search by GSTIN/UIN tab to activate input
    page.evaluate("""() => {
        for (const a of document.querySelectorAll('a')) {
            const t = (a.textContent||'').trim().toLowerCase();
            if (t.includes('gstin') || t.includes('uin')) { a.click(); return; }
        }
    }""")
    time.sleep(1)

    # Wait for GSTIN input to be visible
    try:
        page.wait_for_selector(
            "input[id*='gstin' i], input[name*='gstin' i], input[type='text']",
            state="visible", timeout=8000)
        push_log("  ✓ Search Taxpayer page ready")
    except Exception:
        push_log("  ⚠ GSTIN input not found — proceeding anyway")


def extract_one_gstin(page, gstin):
    """
    Steps 4–7: Enter GSTIN → SEARCH → Show Filing Table → capture API.
    Stays on same searchtp page — no re-login needed between clients.
    """
    result = {
        "gstin":          gstin,
        "gstr1_status":   "unknown", "gstr1_date":  None, "gstr1_period":  None,
        "gstr3b_status":  "unknown", "gstr3b_date": None, "gstr3b_period": None,
        "trade_name":     None,
        "fetched_at":     datetime.now().isoformat(),
        "source":         "GST Portal RPA",
    }

    try:
        # Step 4: Enter GSTIN
        push_log(f"  Step 4: entering GSTIN {gstin}")
        if not type_into(page, [
            "input[id*='gstin' i]", "input[name*='gstin' i]",
            "input[placeholder*='GSTIN' i]", "input[type='text']",
        ], gstin, "GSTIN"):
            result["error"] = "GSTIN field not found"
            return result
        time.sleep(0.5)

        # Step 5: SEARCH (id='lotsearch' — GSTIN lookup button)
        push_log("  Step 5: SEARCH")
        page.evaluate("""() => {
            const b = document.getElementById('lotsearch');
            if (b) { b.click(); return; }
            for (const el of document.querySelectorAll('button'))
                if (el.textContent.trim().toUpperCase() === 'SEARCH') { el.click(); return; }
        }""")
        time.sleep(3)

        # Step 6: Show Filing Table
        push_log("  Step 6: Show Filing Table")
        if not clk(page, [
            "#filingTable",
            "button:has-text('Show Filing')",
            "button:has-text('SHOW FILING')",
            "button:has-text('Filing Table')",
        ], "SHOW FILING TABLE", timeout=8000):
            result["error"] = "SHOW FILING TABLE not found"
            return result
        time.sleep(1)

        # Step 7: Select correct FY in portal dropdown → SEARCH → capture API response
        #
        # Portal: after "Show Filing Table" a FY <select> appears (screenshot confirms it
        # defaults to "2026-2027"). We MUST select the correct FY before clicking SEARCH.
        #
        # In April (transition month):
        #   Primary FY   = 2025-2026  (March 2026 filings due by end of April 2026)
        #   Secondary FY = 2026-2027  (fallback)
        # Other months:
        #   Primary FY   = current FY (e.g. 2026-2027 from May onwards)
        #   Secondary FY = previous FY
        #
        # Portal dropdown format: "2025-2026" (full YYYY-YYYY with hyphen)
        now          = datetime.now()
        cur_fy_start = now.year if now.month >= 4 else now.year - 1
        prev_fy_start= cur_fy_start - 1
        cur_fy  = f"{cur_fy_start}-{cur_fy_start+1}"    # e.g. "2026-2027"
        prev_fy = f"{prev_fy_start}-{prev_fy_start+1}"  # e.g. "2025-2026"

        # April is a transition month: previous FY filings (March) still being completed
        if now.month == 4:
            primary_fy   = prev_fy   # "2025-2026" ← check this first in April
            secondary_fy = cur_fy    # "2026-2027" ← fallback
        else:
            primary_fy   = cur_fy
            secondary_fy = prev_fy

        push_log(f"  Step 7: selecting FY '{primary_fy}' in dropdown → SEARCH...")

        # ── Wait up to 10s for the FY dropdown to appear after "Show Filing Table" ──
        fy_selected = False
        deadline    = time.time() + 10
        while time.time() < deadline:
            try:
                sel_info = page.evaluate("""() => {
                    return Array.from(document.querySelectorAll('select')).map((s, i) => ({
                        idx:  i,
                        opts: Array.from(s.options).map(o => o.text.trim())
                    }));
                }""")
                for info in sel_info:
                    # FY dropdown has options like "2026-2027" — length 9, hyphen at pos 4
                    fy_opts = [o for o in info.get("opts", [])
                               if len(o) == 9 and o[4] == '-' and o[:4].isdigit()]
                    if fy_opts:
                        push_log(f"  FY dropdown found at select[{info['idx']}]: {fy_opts[:4]}")
                        # Try primary FY first, fall back to secondary
                        for fy_try in [primary_fy, secondary_fy]:
                            if fy_try in fy_opts:
                                try:
                                    loc = page.locator("select").nth(info["idx"])
                                    loc.select_option(label=fy_try)
                                    time.sleep(0.4)
                                    actual = loc.evaluate(
                                        "el => el.options[el.selectedIndex]?.text?.trim() || ''")
                                    if actual == fy_try:
                                        push_log(f"  ✓ FY selected: '{fy_try}'")
                                        primary_fy = fy_try   # use this for record filtering
                                        fy_selected = True
                                        break
                                except Exception as _se:
                                    push_log(f"  ⚠ FY select error: {_se}")
                        break
            except Exception:
                pass
            if fy_selected:
                break
            time.sleep(0.4)

        if not fy_selected:
            push_log(f"  ⚠ FY dropdown not found or selection failed — using default")

        time.sleep(0.3)

        # ── Intercept API then click SEARCH (must enter context BEFORE click) ──
        try:
            with page.expect_response(
                lambda r: "taxpayerReturnDetails" in r.url,
                timeout=120000
            ) as resp_info:
                clicked = page.evaluate("""() => {
                    for (const b of document.querySelectorAll('button')) {
                        const t = b.textContent.trim().toUpperCase();
                        if ((t === 'SEARCH' || t === 'GO') && b.id !== 'lotsearch')
                            { b.click(); return b.textContent.trim(); }
                    }
                    return null;
                }""")
                if clicked:
                    push_log(f"  ✓ Clicked '{clicked}' to fetch filing data")
                else:
                    push_log("  ⚠ No SEARCH/GO button found")

            body = resp_info.value.text()
            data = json.loads(body)
            push_log(f"  ✓ Response: {len(body)} bytes")

            raw = data.get("filingStatus") or data.get("EFiledlist") or data.get("returns") or []
            all_records = []
            for item in raw:
                if isinstance(item, list):  all_records.extend(item)
                elif isinstance(item, dict): all_records.append(item)

            # Filter: primary FY first, fall back to secondary, then all
            records = [r for r in all_records
                       if isinstance(r, dict) and str(r.get("fy","")).strip() == primary_fy]
            push_log(f"  {len(records)} records for primary FY '{primary_fy}'")
            if not records:
                records = [r for r in all_records
                           if isinstance(r, dict) and str(r.get("fy","")).strip() == secondary_fy]
                push_log(f"  {len(records)} records for secondary FY '{secondary_fy}'")
            if not records:
                records = all_records
                push_log(f"  Using all {len(records)} records (no FY match)")

            for ret in records:
                if not isinstance(ret, dict): continue
                rtype = (str(ret.get("rtntype") or ret.get("retType") or "")
                         .upper().replace("-","").replace(" ",""))
                sraw  = str(ret.get("status") or ret.get("sts") or "").strip()
                sv    = ("filed"     if sraw.lower() == "filed" else
                         "not_filed" if "not" in sraw.lower() else
                         "filed"     if sraw.upper() in ("FLD","F","Y") else "unknown")
                dof   = str(ret.get("dof") or ret.get("dateOfFiling") or "")
                taxp  = str(ret.get("taxp") or ret.get("taxPeriod") or "")
                fy_r  = str(ret.get("fy") or "")

                if rtype in ("GSTR1","R1","GSTR-1") and result["gstr1_status"] == "unknown":
                    result["gstr1_status"] = sv
                    result["gstr1_date"]   = dof
                    # Store plain MMYYYY (parseable by parsePeriodStr format-1)
                    result["gstr1_period"] = taxp if taxp else fy_r
                    push_log(f"  ✓ GSTR-1:  {sv} | {dof} | {taxp}")

                elif rtype in ("GSTR3B","R3B","3B","GSTR-3B") and result["gstr3b_status"] == "unknown":
                    result["gstr3b_status"] = sv
                    result["gstr3b_date"]   = dof
                    result["gstr3b_period"] = taxp if taxp else fy_r
                    push_log(f"  ✓ GSTR-3B: {sv} | {dof} | {taxp}")

        except Exception as ex:
            push_log(f"  ⚠ expect_response error: {ex}", "error")

        # Trade name (best-effort)
        for sel in ["#lgnm", "#tradeName", ".legal-name", ".legalName", "#legalName"]:
            try:
                t = page.locator(sel).first.inner_text(timeout=2000).strip()
                if 2 < len(t) < 100:
                    result["trade_name"] = t
                    break
            except Exception:
                pass

    except Exception as e:
        push_log(f"  ✗ extract_one_gstin error: {e}", "error")

    return result


# ── RPA worker ────────────────────────────────────────────────
def rpa_worker(clients, use_profile, username="", password=""):
    """
    Thin wrapper around _rpa_worker_impl that guarantees any exception —
    including one raised before browser launch even starts — is logged
    and reflected in `state`, instead of silently killing the thread
    with nothing recorded anywhere (see the matching, more detailed
    comment on combined_download_worker for why this matters in a
    frozen/windowed build).
    """
    try:
        _rpa_worker_impl(clients, use_profile, username, password)
    except Exception as fatal:
        import traceback as _tb
        try:
            push_log(f"✗ FATAL (uncaught): {fatal}", "error")
        except Exception:
            pass
        try:
            with lock:
                state["status"] = "error"
                state["error"] = str(fatal)
        except Exception:
            pass
        log.error(f"[RPA] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _rpa_worker_impl(clients, use_profile, username="", password=""):
    """
    LOGIN ONCE (with provided credentials) → Navigate to Search Taxpayer ONCE →
    Loop: enter GSTIN → search → extract filing table → next client →
    LOGOUT ONCE.
    No re-login, no re-captcha, no logout between clients.
    Credentials are passed from the UI (single login for all GSTINs),
    not from individual client records.
    """
    total = len(clients)

    # NOTE: previously this used open(LOG_FILE, "w", ...) here, which
    # ERASED the entire log file — including the startup diagnostics
    # ("Data folder:", "Build fingerprint:") and anything logged by
    # OTHER features (like the GSTIN Directory enrichment) — every
    # single time this particular RPA flow started. That's why a log
    # file could look "not updating": whichever feature ran LAST wiped
    # out everything anyone had seen from earlier features. Switched to
    # append mode so history survives across different features/runs;
    # the "=== RPA Run ... ===" line below still clearly marks where
    # each new run starts within that history.
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            f.write(f"=== RPA Run {datetime.now().isoformat()} ===\n")
    except Exception:
        pass

    # sync_playwright().start() launches Playwright's own Node.js driver
    # subprocess — if it throws (not just hangs), that was previously
    # uncaught here, silently killing this thread with nothing logged
    # anywhere in a frozen/windowed build.
    try:
        pw = sync_playwright().start()
    except Exception as e:
        push_log(f"✗ Playwright driver failed to start: {e}", "error")
        with lock:
            state["status"] = "error"
            state["error"] = f"Playwright driver failed to start: {e}"
        return
    try:
        profile_dir = os.path.join(PATHS.profiles_dir, "gst_profile")
        os.makedirs(profile_dir, exist_ok=True)

        # Remove stale SingletonLock from previous crash
        lock_file = os.path.join(profile_dir, "SingletonLock")
        if os.path.exists(lock_file):
            try: os.remove(lock_file)
            except Exception: pass

        push_log("Launching browser...")
        try:
            if use_profile:
                context = pw.chromium.launch_persistent_context(
                    user_data_dir=profile_dir,
                    headless=False, slow_mo=80,
                    args=["--disable-blink-features=AutomationControlled",
                          "--no-sandbox", "--start-maximized"],
                    no_viewport=True,
                )
                page = context.pages[0] if context.pages else context.new_page()
            else:
                browser = pw.chromium.launch(
                    headless=False, slow_mo=80,
                    args=["--disable-blink-features=AutomationControlled", "--no-sandbox"])
                context = browser.new_context(viewport={"width": 1280, "height": 800})
                page = context.new_page()
        except Exception as e:
            push_log(f"Browser launch failed: {e}", "error")
            with lock:
                state["status"] = "error"
                state["error"]  = f"Browser launch failed: {e}"
            return

        page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        results = []

        if not username or not password:
            push_log("✗ No credentials provided — cannot start", "error")
            with lock: state["status"] = "error"; state["error"] = "No credentials. Enter GST Username & Password in the RPA panel."
            return

        # ── Open login page ───────────────────────────────────────
        push_log("Opening GST login page...")
        try:
            page.goto(GST_LOGIN_URL, wait_until="domcontentloaded", timeout=20000)
            time.sleep(2)
        except Exception as e:
            push_log(f"✗ Cannot open login page: {e}", "error")
            with lock: state["status"] = "error"; state["error"] = str(e)
            return

        # ── Fill credentials ──────────────────────────────────────
        if not type_into(page, SELECTORS["username"], username, "username"):
            push_log("✗ Username field not found", "error"); return
        push_log("✓ Username filled")

        if not type_into(page, SELECTORS["password"], password, "password"):
            push_log("✗ Password field not found", "error"); return
        push_log("✓ Password filled")

        # ── Captcha ───────────────────────────────────────────────
        push_log("Capturing captcha...")
        cap_img = capture_captcha_img(page)
        with lock:
            state["status"]        = "waiting_captcha"
            state["captcha_image"] = cap_img
        push_log("⏸ Waiting for captcha answer...")
        answer = wait_for_user("captcha_answer", 180)
        if not answer:
            push_log("✗ Captcha timeout", "error"); return
        with lock:
            state["status"]        = "running"
            state["captcha_image"] = None

        # ── Submit login ──────────────────────────────────────────
        if not fill_captcha_field(page, answer.strip()):
            push_log("✗ Captcha field not found", "error"); return
        push_log(f"✓ Captcha filled: {answer.strip()}")
        time.sleep(0.4)
        login_clicked = False
        for sel in SELECTORS["login_btn"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.click(); login_clicked = True; break
            except Exception: continue
        if not login_clicked:
            page.keyboard.press("Enter")
        push_log("✓ Login submitted")
        time.sleep(3)

        # ── OTP (if triggered) ────────────────────────────────────
        otp_el = None
        for sel in SELECTORS["otp_input"]:
            try:
                page.wait_for_selector(sel, timeout=3000)
                otp_el = page.query_selector(sel)
                if otp_el: break
            except Exception: continue
        if otp_el:
            push_log("📱 OTP required — waiting for user...")
            with lock: state["status"] = "waiting_otp"
            otp = wait_for_user("otp_answer", 300)
            if not otp:
                push_log("✗ OTP timeout", "error"); return
            with lock: state["status"] = "running"
            try:
                otp_el.click(); time.sleep(0.2)
                page.keyboard.press("Control+a")
                page.keyboard.type(otp.strip(), delay=60)
                sub = find_el(page, SELECTORS["otp_submit"])
                if sub: sub.click()
                else:   page.keyboard.press("Enter")
                push_log("✓ OTP submitted")
                time.sleep(2.5)
            except Exception as e:
                push_log(f"✗ OTP error: {e}", "error"); return
        else:
            push_log("✓ No OTP needed")

        # ── Verify login ──────────────────────────────────────────
        time.sleep(2)
        login_ok = check_login_success(page)
        push_log(f"Login check: {login_ok}")

        if login_ok is False:
            push_log("✗ Login failed — refreshing captcha for retry")
            for sel in ["button:has-text('Refresh')", "a:has-text('Refresh')",
                        "[title*='efresh']", "#refreshCaptcha"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=2000)
                    loc.click(); push_log("✓ Captcha refreshed"); break
                except Exception: continue
            time.sleep(2)
            cap_img2 = capture_captcha_img(page)
            with lock:
                state["status"]        = "waiting_captcha"
                state["captcha_image"] = cap_img2
            push_log("⏸ Waiting for new captcha answer...")
            answer2 = wait_for_user("captcha_answer", 180)
            if not answer2:
                push_log("✗ Captcha retry timeout", "error"); return
            with lock:
                state["status"]        = "running"
                state["captcha_image"] = None
            fill_captcha_field(page, answer2.strip())
            btn2 = find_el(page, SELECTORS["login_btn"])
            if btn2: btn2.click()
            else:    page.keyboard.press("Enter")
            time.sleep(3)
            push_log("✓ Retry submitted")
            login_ok = check_login_success(page)
            if login_ok is False:
                push_log("✗ Login still failed — aborting", "error")
                with lock: state["status"] = "error"; state["error"] = "Login failed"
                return

        push_log("✓ Logged in!")

        # Dismiss Principal Place of Business popup (safe — never crashes)
        try:
            g2a_dismiss_popup(page)
        except Exception as e:
            push_log(f"  ⚠ Popup dismiss error (ignored): {e}")

        # ── Navigate to Search Taxpayer — ONCE ───────────────────
        try:
            nav_to_search_page(page)
        except Exception as e:
            push_log(f"  ✗ nav_to_search_page failed: {e}", "error")
            with lock: state["status"] = "error"; state["error"] = str(e)
            context.close(); return

        # ═══ MAIN LOOP — one GSTIN at a time, stay on same page ══
        for idx, client in enumerate(clients):
            gstin = client.get("gstin", "")
            name  = client.get("name",  "Unknown")

            with lock:
                state.update({
                    "current_client": {"name": name, "gstin": gstin,
                                       "index": idx+1, "total": total},
                    "progress":       int((idx / total) * 100),
                    "status":         "running",
                })

            push_log(f"━━━ [{idx+1}/{total}] {name} ({gstin}) ━━━")
            result         = extract_one_gstin(page, gstin)
            result["name"] = name
            results.append(result)
            push_log(f"  ✓ GSTR-1={result['gstr1_status']}  GSTR-3B={result['gstr3b_status']}")

            # ── Auto-save to gst_cache.json immediately (survives browser close) ──
            try:
                with _gst_cache_lock:
                    disk   = _load_gst_cache_file()
                    dcache = disk.get("cache", {}) if isinstance(disk, dict) else {}
                    cd     = dict(dcache.get(result["gstin"], {}))
                    for fld in ("gstr1_status","gstr1_date","gstr1_period",
                                "gstr3b_status","gstr3b_date","gstr3b_period",
                                "trade_name","fetched_at"):
                        if result.get(fld) and result[fld] != "unknown":
                            cd[fld] = result[fld]
                    cd["rpa_updated"] = datetime.now().isoformat()
                    dcache[result["gstin"]] = cd
                    _save_gst_cache_file({"cache": dcache,
                                          "last_fetch": datetime.now().isoformat()})
            except Exception as _ae:
                log.warning(f"  ⚠ gst_cache auto-save: {_ae}")

            # Navigate back to search form for next GSTIN
            if idx < total - 1:
                push_log("  ↩ Back to search form...")
                # Try clicking a back/new-search link (faster — no reload)
                page.evaluate("""() => {
                    for (const a of document.querySelectorAll('a')) {
                        const t = (a.textContent||'').trim().toLowerCase();
                        if (t.includes('search another') || t.includes('new search')
                            || t.includes('back to search')) { a.click(); return; }
                    }
                }""")
                time.sleep(1)
                # If not on searchtp, reload it
                if "searchtp" not in page.url:
                    page.goto("https://services.gst.gov.in/services/auth/searchtp",
                              wait_until="domcontentloaded", timeout=15000)
                    time.sleep(1)
                # Re-click GSTIN/UIN tab to ensure input is visible
                page.evaluate("""() => {
                    for (const a of document.querySelectorAll('a')) {
                        const t = (a.textContent||'').trim().toLowerCase();
                        if (t.includes('gstin') || t.includes('uin')) { a.click(); return; }
                    }
                }""")
                try:
                    page.wait_for_selector(
                        "input[id*='gstin' i], input[name*='gstin' i], input[type='text']",
                        state="visible", timeout=10000)
                    push_log("  ✓ Search form ready")
                except Exception:
                    push_log("  ⚠ Search form wait timed out — proceeding anyway")
                time.sleep(0.5)

        # ═══ All done — logout once ═══════════════════════════════
        try:
            page.goto(f"{GST_PORTAL_URL}/services/logout", timeout=8000)
            time.sleep(1.5)
            push_log(f"✓ Logged out — {total} GSTIN(s) processed")
        except Exception:
            pass
        try:
            context.close()
        except Exception:
            pass
    finally:
        try: pw.stop()
        except Exception: pass

    with lock:
        state["status"]   = "done"
        state["results"]  = results
        state["progress"] = 100
    push_log(f"✅ RPA complete — {len(results)} clients processed")


# ── API routes ────────────────────────────────────────────────

# ══════════════════════════════════════════════════════════════════════════════
#  CLIENT DATA PERSISTENCE — server-side JSON (survives browser cache clears)
# ══════════════════════════════════════════════════════════════════════════════
_CLIENTS_FILE = PATHS.clients_backup_file
_clients_lock = threading.Lock()

def _load_clients_file():
    try:
        if os.path.exists(_CLIENTS_FILE):
            with open(_CLIENTS_FILE, encoding="utf-8") as f:
                data = json.load(f)
            return data.get("clients", []) if isinstance(data, dict) else data
    except Exception as e:
        log.error(f"clients.json load error: {e}")
    return []

def _save_clients_file(clients):
    try:
        payload = {"version":1,"saved_at":datetime.now().isoformat(),
                   "count":len(clients),"clients":clients}
        tmp = _CLIENTS_FILE + ".tmp"
        with open(tmp,"w",encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp, _CLIENTS_FILE)
        return True
    except Exception as e:
        log.error(f"clients.json save error: {e}"); return False

@app.route("/clients/load")
def clients_load():
    with _clients_lock: clients = _load_clients_file()
    return jsonify({"ok":True,"clients":clients,"count":len(clients),"file":_CLIENTS_FILE})

@app.route("/clients/save", methods=["POST"])
def clients_save():
    data = request.get_json(force=True) or {}
    clients = data.get("clients", [])
    if not isinstance(clients, list):
        return jsonify({"error":"clients must be an array"}), 400
    with _clients_lock: ok = _save_clients_file(clients)
    if ok: return jsonify({"ok":True,"count":len(clients),"file":_CLIENTS_FILE})
    return jsonify({"error":"Failed to write clients.json"}), 500

@app.route("/clients/export")
def clients_export():
    with _clients_lock: clients = _load_clients_file()
    payload = {"version":1,"exported_at":datetime.now().isoformat(),
               "count":len(clients),"clients":clients}
    from io import BytesIO
    buf = BytesIO(json.dumps(payload,ensure_ascii=False,indent=2).encode())
    buf.seek(0)
    fname = f"ClientLedger_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return send_file(buf, mimetype="application/json", as_attachment=True, download_name=fname)

@app.route("/clients/import", methods=["POST"])
def clients_import_route():
    data=request.get_json(force=True) or {}
    incoming=data.get("clients",[]); merge=bool(data.get("merge",False))
    if not isinstance(incoming,list): return jsonify({"error":"clients must be an array"}),400
    with _clients_lock:
        if merge:
            existing=_load_clients_file(); existing_ids={c.get("id") for c in existing}
            added=[c for c in incoming if c.get("id") not in existing_ids]
            final=existing+added
        else: final=incoming; added=incoming
        ok=_save_clients_file(final)
    if ok: return jsonify({"ok":True,"count":len(final),"added":len(added)})
    return jsonify({"error":"Failed to write clients.json"}),500




# ══════════════════════════════════════════════════════════════════════════════
#  GST FILING STATUS CACHE — server-side persistence (gst_cache.json)
#  Saves to gst_cache.json next to gst_rpa.py.
#  Survives browser cache clears / system formats.
# ══════════════════════════════════════════════════════════════════════════════
_GST_CACHE_FILE = PATHS.gst_cache_file
_gst_cache_lock = threading.Lock()

def _load_gst_cache_file():
    try:
        if os.path.exists(_GST_CACHE_FILE):
            with open(_GST_CACHE_FILE, encoding="utf-8") as f:
                return json.load(f)
    except Exception as e:
        log.error(f"gst_cache.json load error: {e}")
    return {}

def _save_gst_cache_file(payload):
    try:
        payload["saved_at"] = datetime.now().isoformat()
        tmp = _GST_CACHE_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        os.replace(tmp, _GST_CACHE_FILE)
        return True
    except Exception as e:
        log.error(f"gst_cache.json save error: {e}"); return False

@app.route("/gst_cache/load")
def gst_cache_load():
    """Return saved GST filing cache. Called by browser on startup."""
    with _gst_cache_lock:
        data = _load_gst_cache_file()
    return jsonify({
        "ok":         True,
        "cache":      data.get("cache",      {}),
        "last_fetch": data.get("last_fetch", None),
        "saved_at":   data.get("saved_at",   None),
        "count":      len(data.get("cache",  {})),
        "file":       _GST_CACHE_FILE,
    })

@app.route("/gst_cache/save", methods=["POST"])
def gst_cache_save():
    """Save GST filing cache from browser to disk.
    Body: { cache: {GSTIN: {...}}, last_fetch: "timestamp" }
    """
    data       = request.get_json(force=True) or {}
    cache      = data.get("cache",      {})
    last_fetch = data.get("last_fetch", None)
    if not isinstance(cache, dict):
        return jsonify({"error": "cache must be an object"}), 400
    with _gst_cache_lock:
        ok = _save_gst_cache_file({"cache": cache, "last_fetch": last_fetch})
    if ok:
        return jsonify({"ok": True, "count": len(cache), "file": _GST_CACHE_FILE})
    return jsonify({"error": "Failed to write gst_cache.json"}), 500





@app.route("/health")
def health():
    return jsonify({"status": "ok", "playwright": PLAYWRIGHT_AVAILABLE,
                    "server": "ClientLedger GST RPA", "version": "2.0",
                    "build": BUILD_FINGERPRINT,
                    "time": datetime.now().isoformat()})

@app.route("/state")
def get_state():
    with lock:
        return jsonify({
            "status":         state["status"],
            "current_client": state["current_client"],
            "has_captcha":    state["captcha_image"] is not None,
            "progress":       state["progress"],
            "results_count":  len(state["results"]),
            "log":            state["log"][-40:],
            "error":          state["error"],
        })

@app.route("/captcha_image")
def get_captcha_image():
    with lock:
        img = state.get("captcha_image")
    if not img:
        return jsonify({"error": "No captcha available"}), 404
    return jsonify({"image": img})

@app.route("/submit_captcha", methods=["POST"])
def submit_captcha():
    data   = request.get_json(force=True) or {}
    answer = (data.get("captcha") or "").strip()
    if not answer:
        return jsonify({"error": "captcha required"}), 400
    with lock:
        state["captcha_answer"] = answer
        state["captcha_image"]  = None
    return jsonify({"ok": True})

@app.route("/submit_otp", methods=["POST"])
def submit_otp():
    data = request.get_json(force=True) or {}
    otp  = (data.get("otp") or "").strip()
    if not otp:
        return jsonify({"error": "otp required"}), 400
    with lock:
        state["otp_answer"] = otp
    return jsonify({"ok": True})

@app.route("/results")
def get_results():
    with lock:
        return jsonify({"status": state["status"], "results": state["results"]})

@app.route("/start", methods=["POST"])
def start_rpa():
    data     = request.get_json(force=True) or {}
    clients  = data.get("clients", [])
    if not clients:
        return jsonify({"error": "clients array required"}), 400
    # Accept top-level credentials (preferred — single login for all GSTINs)
    # Fall back to first client's credentials for backward compatibility
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not username:
        username = (clients[0].get("username") or "").strip()
    if not password:
        password = (clients[0].get("password") or "").strip()

    # Atomic check-and-claim under a single lock acquisition — see the
    # matching, more detailed comment on gstin_dir_enrich_start()/
    # combined_start() for why this must not be a separate check then
    # a separate later claim (the gap between them let concurrent
    # requests both pass the check and both spawn a worker).
    with lock:
        if state["status"] in ("running", "waiting_captcha", "waiting_otp"):
            return jsonify({"error": "RPA already running"}), 409
        state.update({"status": "running", "results": [], "log": [],
                      "error": None, "progress": 0,
                      "start_time": datetime.now().isoformat()})

    threading.Thread(target=rpa_worker,
                     args=(clients, data.get("persistent_profile", True),
                           username, password),
                     daemon=True).start()
    return jsonify({"ok": True, "clients": len(clients)})

@app.route("/stop", methods=["POST"])
def stop_rpa():
    with lock:
        state["status"] = "idle"
        state["error"]  = "Stopped by user"
    return jsonify({"ok": True})

@app.route("/reset", methods=["POST"])
def reset():
    with lock:
        state.update({"status": "idle", "current_client": None,
                      "captcha_image": None, "captcha_answer": None,
                      "otp_answer": None, "results": [], "log": [],
                      "error": None, "progress": 0})
    return jsonify({"ok": True})

@app.route("/download_log")
def download_log():
    with lock:
        content = "\n".join(_full_log)
    return Response(content, mimetype="text/plain",
        headers={"Content-Disposition": "attachment; filename=gst_rpa_activity.log"})

@app.route("/continue_scrape", methods=["POST"])
def continue_scrape():
    with lock:
        state["user_continue"] = True
    return jsonify({"ok": True})


# ══════════════════════════════════════════════════════════════════════════════
# GSTIN NAMES MODULE — Per-return-type persistent JSON store
# ─────────────────────────────────────────────────────────────────────────────
# Folder layout:
#   gstin_names/
#     2a/  name_lookup.json  ← supplier GSTINs from G2A + trdnm from G2B scan
#     2b/  name_lookup.json  ← supplier GSTINs from G2B + trdnm from G2B JSON
#     1/   name_lookup.json  ← buyer GSTINs from G1 b2b
#     enriched/ name_lookup.json ← portal enrichment: legal_name + trade_name
#
# Written by:
#   gstin_dir_scan  → writes 2a/, 2b/, 1/ files
#   enrichment worker → writes enriched/ file (live, after each portal lookup)
# Read by:
#   _gnames_load_for(return_type) → merges per-return + enriched (enriched wins)
# No IndexedDB / browser involvement — pure server-side file persistence.
# ══════════════════════════════════════════════════════════════════════════════
_APP_DIR         = getattr(sys, "_MEIPASS", None) or \
                    os.path.dirname(os.path.abspath(__file__))
GSTIN_NAMES_DIR  = PATHS.gstin_names_dir
_GNAMES_ENRICHED = os.path.join(GSTIN_NAMES_DIR, "enriched", "name_lookup.json")
_GNAMES_2A       = os.path.join(GSTIN_NAMES_DIR, "2a",       "name_lookup.json")
_GNAMES_2B       = os.path.join(GSTIN_NAMES_DIR, "2b",       "name_lookup.json")
_GNAMES_1        = os.path.join(GSTIN_NAMES_DIR, "1",        "name_lookup.json")
_gnames_lock     = threading.Lock()

_GNAMES_FILES = {"2a": _GNAMES_2A, "2b": _GNAMES_2B, "1": _GNAMES_1,
                 "enriched": _GNAMES_ENRICHED}


def _gnames_load_file(fpath):
    """Load one name_lookup JSON file. Returns {} if missing/corrupt."""
    if not fpath:
        return {}
    try:
        with open(fpath, encoding="utf-8") as fh:
            data = json.load(fh)
        return data if isinstance(data, dict) else {}
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _gnames_load_for(return_type):
    """
    Load merged name lookup for a specific return type.
    Priority: enriched (portal, legal+trade) > per-return (JSON scan, trade only) > ""
    Returns {GSTIN_UPPER: {trade_name, legal_name}}.
    """
    per_return = _gnames_load_file(_GNAMES_FILES.get(return_type, ""))
    enriched   = _gnames_load_file(_GNAMES_ENRICHED)
    merged     = {}
    for k in set(per_return) | set(enriched):
        pr = per_return.get(k, {})
        en = enriched.get(k,   {})
        merged[k] = {
            "trade_name": (en.get("trade_name") or pr.get("trade_name") or "").strip(),
            "legal_name": (en.get("legal_name") or pr.get("legal_name") or "").strip(),
        }
    return merged


def _gnames_load():
    """Backward-compat alias — returns the enriched store only."""
    return _gnames_load_file(_GNAMES_ENRICHED)


def _gnames_write_file(fpath, store):
    """Atomic write of a name_lookup dict to fpath."""
    os.makedirs(os.path.dirname(fpath), exist_ok=True)
    tmp = fpath + ".tmp"
    with open(tmp, "w", encoding="utf-8") as fh:
        json.dump(store, fh, ensure_ascii=False, indent=2)
    os.replace(tmp, fpath)


def _gnames_write_return_batch(return_type, entries):
    """
    Write/merge a batch of {gstin, trade_name} into the per-return file.
    entries = {gstin_upper: trade_name_str}
    Non-empty values win; existing entries preserved.
    """
    fpath = _GNAMES_FILES.get(return_type)
    if not fpath:
        return
    with _gnames_lock:
        store = _gnames_load_file(fpath)
        for gstin, trdnm in entries.items():
            key = (gstin or "").strip().upper()
            if not key or len(key) != 15:
                continue
            existing_trdnm = store.get(key, {}).get("trade_name", "")
            store[key] = {
                "trade_name": trdnm or existing_trdnm or "",
                "legal_name": store.get(key, {}).get("legal_name", ""),
            }
        _gnames_write_file(fpath, store)
    return len(store)


def _gnames_upsert(supplier_gstin, record):
    """
    Thread-safe upsert into the enriched store.
    Called by enrichment worker after each portal lookup.
    record = {trade_name, legal_name, status, reg_date,
              constitution, taxpayer_type, turnover_slab, enrich_ts}
    Non-empty values overwrite; empty never blanks existing data.
    """
    key = (supplier_gstin or "").strip().upper()
    if not key or len(key) != 15:
        return
    os.makedirs(os.path.dirname(_GNAMES_ENRICHED), exist_ok=True)
    with _gnames_lock:
        store = _gnames_load_file(_GNAMES_ENRICHED)
        ex    = store.get(key, {})
        store[key] = {
            "trade_name":    record.get("trade_name",    "") or ex.get("trade_name",    ""),
            "legal_name":    record.get("legal_name",    "") or ex.get("legal_name",    ""),
            "status":        record.get("status",        "") or ex.get("status",        ""),
            "reg_date":      record.get("reg_date",      "") or ex.get("reg_date",      ""),
            "constitution":  record.get("constitution",  "") or ex.get("constitution",  ""),
            "taxpayer_type": record.get("taxpayer_type", "") or ex.get("taxpayer_type", ""),
            "turnover_slab": record.get("turnover_slab", "") or ex.get("turnover_slab", ""),
            "enrich_ts":     record.get("enrich_ts",     "") or ex.get("enrich_ts",     ""),
        }
        _gnames_write_file(_GNAMES_ENRICHED, store)


def _gnames_save_batch(records):
    """
    Upsert a list of {gstin, trade_name, legal_name, ...} into the enriched store.
    Returns total count of entries after save.
    """
    os.makedirs(os.path.dirname(_GNAMES_ENRICHED), exist_ok=True)
    with _gnames_lock:
        store = _gnames_load_file(_GNAMES_ENRICHED)
        for rec in records:
            key = (rec.get("gstin") or "").strip().upper()
            if not key or len(key) != 15:
                continue
            ex = store.get(key, {})
            store[key] = {
                "trade_name":    rec.get("trade_name",    "") or ex.get("trade_name",    ""),
                "legal_name":    rec.get("legal_name",    "") or ex.get("legal_name",    ""),
                "status":        rec.get("status",        "") or ex.get("status",        ""),
                "reg_date":      rec.get("reg_date",      "") or ex.get("reg_date",      ""),
                "constitution":  rec.get("constitution",  "") or ex.get("constitution",  ""),
                "taxpayer_type": rec.get("taxpayer_type", "") or ex.get("taxpayer_type", ""),
                "turnover_slab": rec.get("turnover_slab", "") or ex.get("turnover_slab", ""),
                "enrich_ts":     rec.get("enrich_ts",     "") or ex.get("enrich_ts",     ""),
            }
        _gnames_write_file(_GNAMES_ENRICHED, store)
        return len(store)


# ── Flask routes for GSTIN Names ──────────────────────────────────────────────

# ── GSTR-2A module globals ─────────────────────────────────────
g2a_lock  = threading.Lock()
G2A_DOWNLOAD_DIR = PATHS.gstr2a_dir
g2a_state = {
    "status":         "idle",
    "log":            [],
    "error":          None,
    "progress":       0,
    "current_month":  None,
    "total_months":   0,
    "done_months":    0,
    "files":          [],
    "gstin":          None,
    "fy":             None,
    "specific_month": None,
    "captcha_image":  None,
    "captcha_answer": None,
    "otp_answer":     None,
}


def g2a_log(msg, level="info"):
    ts   = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    with g2a_lock:
        g2a_state["log"].append(line)
        if len(g2a_state["log"]) > 400:
            g2a_state["log"] = g2a_state["log"][-400:]
    (log.error if level == "error" else log.info)(msg)


def g2a_set(updates, log_msg=None):
    with g2a_lock:
        g2a_state.update(updates)
    if log_msg:
        g2a_log(log_msg)


def g2a_wait_field(field, timeout_sec=180):
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        time.sleep(0.4)
        with g2a_lock:
            val = g2a_state.get(field)
            if val:
                g2a_state[field] = None
                return val
    return None


# ── Month list builder ─────────────────────────────────────────
def g2a_months_for_fy(fy):
    """Return all elapsed months for the given FY (April → March).
    FY string format: '2025-26'.  Only months up to today are returned."""
    fy_year = int(fy.split("-")[0])
    now     = datetime.now()
    order   = [
        (4,"April"),(5,"May"),(6,"June"),(7,"July"),
        (8,"August"),(9,"September"),(10,"October"),
        (11,"November"),(12,"December"),
        (1,"January"),(2,"February"),(3,"March"),
    ]
    months = []
    for mon_num, mon_name in order:
        yr = fy_year if mon_num >= 4 else fy_year + 1
        if yr > now.year or (yr == now.year and mon_num > now.month):
            continue
        months.append({
            "num":     mon_num,
            "year":    yr,
            "name":    mon_name,
            "abbr":    mon_name[:3].upper(),
            "display": f"{mon_name} {yr}",
            "period":  f"{mon_num:02d}{yr}",   # MMYYYY — confirmed portal format
        })
    return months


# ══════════════════════════════════════════════════════════════════════
#  FY CONFIG SYSTEM  (Requirement 3)
#  ─────────────────────────────────────────────────────────────────────
#  Stored in  fy_config.json  next to gst_rpa.py.
#  Schema:
#    {
#      "extra_fys":  ["2026-27", "2027-28"],   ← user-added future FYs
#      "stale_days": 10                         ← re-download if older than N days
#    }
#
#  Auto-range  2017-18 → (current FY + 1)  is always available without any
#  config, so the Settings panel only needs to handle FYs beyond that window.
# ══════════════════════════════════════════════════════════════════════
_FY_CONFIG_FILE  = PATHS.fy_config_file
_fy_config_lock  = threading.Lock()
_DEFAULT_STALE_DAYS = 10


def g2a_load_fy_config():
    """Load fy_config.json; return safe defaults on any error."""
    defaults = {"extra_fys": [], "stale_days": _DEFAULT_STALE_DAYS}
    try:
        with _fy_config_lock:
            if os.path.isfile(_FY_CONFIG_FILE):
                with open(_FY_CONFIG_FILE, encoding="utf-8") as f:
                    saved = json.load(f)
                defaults["extra_fys"]  = [str(x) for x in saved.get("extra_fys", [])]
                defaults["stale_days"] = max(1, int(saved.get("stale_days",
                                                              _DEFAULT_STALE_DAYS)))
    except Exception as _e:
        log.warning(f"fy_config load error (using defaults): {_e}")
    return defaults


def g2a_save_fy_config(cfg):
    """Write fy_config.json atomically.  Returns True on success."""
    try:
        with _fy_config_lock:
            tmp = _FY_CONFIG_FILE + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2)
            os.replace(tmp, _FY_CONFIG_FILE)
        return True
    except Exception as _e:
        log.error(f"fy_config save error: {_e}")
        return False


def g2a_available_fys():
    """
    Return sorted list of all selectable FY strings.

    Auto-range: 2017-18 → (current FY + 1) so users never have to add
    recent or upcoming FYs manually.  The +1 means next year appears in
    the dropdown from Day 1 of the new FY (April 1), useful for firms
    that need to prepare in advance.

    Extra FYs beyond that window (e.g. 2028-29 added today) are stored in
    fy_config.json and merged here without duplicates.
    """
    now          = datetime.now()
    # Current FY start year: if month >= April use this year, else last year
    cur_fy_start = now.year if now.month >= 4 else now.year - 1
    auto_start   = 2017
    auto_end     = cur_fy_start + 1      # one FY ahead of current

    auto_fys = []
    for y in range(auto_start, auto_end + 1):
        short = str(y + 1)[-2:]          # e.g. 2025 → "26"
        auto_fys.append(f"{y}-{short}")

    cfg       = g2a_load_fy_config()
    extra_fys = cfg.get("extra_fys", [])

    # Merge, deduplicate, sort ascending by start year
    combined = list(dict.fromkeys(auto_fys + extra_fys))
    combined.sort(key=lambda x: int(x.split("-")[0]))
    return combined


def g2a_stale_days():
    """Return the configured stale-threshold (default 10 days)."""
    return g2a_load_fy_config().get("stale_days", _DEFAULT_STALE_DAYS)


# ══════════════════════════════════════════════════════════════════════
#  FILE-AGE CACHE CHECK  (Requirement 1)
#  ─────────────────────────────────────────────────────────────────────
#  Called at the START of g2a_worker, before any browser is opened.
#  Splits the requested month list into two buckets:
#    fresh_results  — file exists and is < stale_days old  → return as-is
#    needed_months  — file missing OR older than stale_days → must download
#
#  The 'force' flag (passed from /g2a/start) bypasses the check entirely,
#  treating every month as needed regardless of file age.
# ══════════════════════════════════════════════════════════════════════
def g2a_check_cached_months(gstin_dir, months, force=False):
    """
    Args:
        gstin_dir : folder that holds GSTR2A_*.json files for this client/FY
        months    : list of month-dicts from g2a_months_for_fy()
        force     : if True, skip cache — every month is marked 'needed'

    Returns:
        (fresh_results, needed_months)
        fresh_results : list of result-dicts (same shape as download results)
        needed_months : list of month-dicts that must be downloaded
    """
    stale_days    = g2a_stale_days()
    cutoff_epoch  = time.time() - stale_days * 86400
    fresh_results = []
    needed_months = []

    for month in months:
        abbr      = month["display"].replace(" ", "_")
        json_path = os.path.join(gstin_dir, f"GSTR2A_{abbr}.json")

        if os.path.isfile(json_path) and not force:
            mtime   = os.path.getmtime(json_path)
            age_d   = (time.time() - mtime) / 86400
            size_kb = max(1, os.path.getsize(json_path) // 1024)

            if mtime >= cutoff_epoch:
                # ── Fresh: skip download ───────────────────────────────────
                g2a_log(f"  ⏭  {month['display']} — cached "
                        f"({age_d:.1f}d old, {size_kb} KB) "
                        f"< {stale_days}-day threshold → skipping")
                fresh_results.append({
                    "month":    month["display"],
                    "filename": f"GSTR2A_{abbr}.json",
                    "size_kb":  size_kb,
                    "period":   month["period"],
                    "cached":   True,
                    "age_days": round(age_d, 1),
                })
            else:
                # ── Stale: re-download ─────────────────────────────────────
                g2a_log(f"  🔄  {month['display']} — stale "
                        f"({age_d:.1f}d old ≥ {stale_days}d) → re-downloading")
                needed_months.append(month)
        else:
            if force and os.path.isfile(json_path):
                g2a_log(f"  🔃  {month['display']} — force re-download requested")
            needed_months.append(month)

    g2a_log(f"  Cache summary: {len(fresh_results)} fresh | "
            f"{len(needed_months)} to download "
            f"(threshold: {stale_days} days{', FORCED' if force else ''})")
    return fresh_results, needed_months


# ══════════════════════════════════════════════════════════════════════
#  WHAT THE CURRENT GST PORTAL ACTUALLY DOES (from inspecting behavior)
#  ─────────────────────────────────────────────────────────────────────
#  "View" mode:  Uses deprecated section API  ?action=B2B ...
#                Returns status:0 even when data exists. ABANDONED.
#
#  "Download" button → "Generate JSON file to download":
#    GET /returns/auth/gstr/auth/api/offline/download/generate
#        ?flag=0&rtn_prd=MMYYYY&rtn_typ=GSTR2A    → trigger, ack
#        ?flag=1&rtn_prd=MMYYYY&rtn_typ=GSTR2A    → poll ready (status=9)
#        ?flag=2&rtn_prd=MMYYYY&rtn_typ=GSTR2A    → stream ZIP/JSON
#    Server-side generation takes ~3-20 min.  This IS the correct API.
#
#  WHY ClearTax IS FASTER:
#    ClearTax is a licensed GSP.  They call GSTN's dedicated server API at
#    api.gst.gov.in with an RSA-encrypted app_key.  That API responds
#    instantly.  Not reproducible without a GSP licence.
#
#  OUR REAL OPTIMISATION  (no GSP licence needed):
#    Submit ALL months' flag=0 (trigger generate) SIMULTANEOUSLY.
#    All 11 generation jobs run concurrently on GSTN's servers.
#    Total wait: ~20 min flat for the whole year, not per month.
#    Then intercept + capture actual network calls so we learn real URLs.
# ══════════════════════════════════════════════════════════════════════

# ── Offline-download API constants ─────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────
#  Network interceptor — log API calls the portal makes
# ─────────────────────────────────────────────────────────────────────
_captured_api_calls = []
_captured_lock = threading.Lock()

def g2a_install_interceptors(page):
    """Log all XHR calls to GST domains for debugging."""
    def on_response(response):
        url = response.url
        if any(d in url for d in ["return.gst.gov.in", "services.gst.gov.in"]):
            with _captured_lock:
                _captured_api_calls.append((response.request.method, url, response.status))
                if len(_captured_api_calls) > 200:
                    _captured_api_calls.pop(0)
    try:
        page.on("response", on_response)
    except Exception:
        pass

def g2a_log_discovered_apis():
    with _captured_lock:
        calls = list(_captured_api_calls)
    if not calls:
        return
    g2a_log("📡 API calls intercepted from portal:")
    for method, url, status in calls[-15:]:
        g2a_log(f"  [{status}] {method} {url}")


# ─────────────────────────────────────────────────────────────────────
#  Quarter helper
# ─────────────────────────────────────────────────────────────────────
def g2a_quarter_for_month(month_num):
    if month_num in (4,5,6):      return "Quarter 1 (Apr - Jun)", 1
    elif month_num in (7,8,9):    return "Quarter 2 (Jul - Sep)", 2
    elif month_num in (10,11,12): return "Quarter 3 (Oct - Dec)", 3
    else:                         return "Quarter 4 (Jan - Mar)", 4


# ─────────────────────────────────────────────────────────────────────
#  Popup dismissal
# ─────────────────────────────────────────────────────────────────────
def g2a_dismiss_popup(page, context=None):
    """
    Dismiss the 'Principal Place of Business metadata' dialog.
    Button exact text: NO-REMIND ME LATER

    Retries for up to 12 seconds across all open pages.
    """
    # All pages to check — dialog may be on any tab
    def all_pages():
        pages = [page]
        if context:
            for p in context.pages:
                if p not in pages:
                    pages.append(p)
        return pages

    deadline = time.time() + 12
    attempt  = 0
    while time.time() < deadline:
        attempt += 1
        for p in all_pages():
            # JS: find button with exact text "NO-REMIND ME LATER"
            result = p.evaluate("""() => {
                const targets = [
                    'NO-REMIND ME LATER',
                    'No-Remind Me Later',
                    'NO REMIND ME LATER',
                ];
                for (const btn of document.querySelectorAll('button, a')) {
                    const t = (btn.textContent || '').trim();
                    if (targets.includes(t) || t.toUpperCase() === 'NO-REMIND ME LATER') {
                        btn.click();
                        return t;
                    }
                }
                return null;
            }""")
            if result:
                g2a_log(f"  ✓ Dismissed popup: '{result}' (attempt {attempt})")
                time.sleep(0.3)
                return True
        time.sleep(0.5)

    g2a_log(f"  ℹ Popup not found after {attempt} attempts (may not have appeared)")
    return False


# ─────────────────────────────────────────────────────────────────────
#  Session activation — reach return.gst.gov.in after login
# ─────────────────────────────────────────────────────────────────────
def g2a_activate_session(page, context):
    """
    Navigate from services.gst.gov.in/fowelcome → return.gst.gov.in/dashboard.

    Portal DOM structure (confirmed from screenshots):
    ┌─ Top navbar ────────────────────────────────────────────────┐
    │ Dashboard │ Services ▼ │ GST Law │ Downloads │ ...          │
    └─────────────────────────────────────────────────────────────┘
         ↓  Click "Services"
    ┌─ Services mega-dropdown (Bootstrap) ────────────────────────┐
    │  [Registration] [Ledgers] [Returns] [Payments] [User Svc]   │  ← nav-tabs
    │                                                             │
    │  (tab content area — changes on tab click)                  │
    └─────────────────────────────────────────────────────────────┘
         ↓  Click "Returns" tab  (Bootstrap tab, needs click not hover)
    ┌─ Returns tab content ───────────────────────────────────────┐
    │  Returns Dashboard │ View Filed Returns                     │
    │  Track Return Status │ Transition Forms                     │
    │  ITC Forms │ Annual Return  ...                             │
    └─────────────────────────────────────────────────────────────┘
         ↓  Click "Returns Dashboard"  (native Playwright click)
    → return.gst.gov.in/returns/auth/dashboard

    IMPORTANT:
    - "Returns" is a Bootstrap nav-tab → CLICK (not just hover) activates it
    - "Returns Dashboard" must use native loc.click(), NOT dispatchEvent()
      (Angular/React portals ignore synthetic MouseEvents)
    """
    DASHBOARD_URL = "https://return.gst.gov.in/returns/auth/dashboard"

    g2a_log(f"  -> Starting at: {page.url[:80]}")
    g2a_dismiss_popup(page, context)

    # ── Helpers ──────────────────────────────────────────────────────────────
    def real_url(p):
        try:
            return p.evaluate("location.href") or p.url
        except Exception:
            return p.url

    def arrived(p):
        u = real_url(p)
        return ("return.gst.gov.in" in u
                and "accessdenied" not in u
                and "login" not in u)

    def sync_page(p, label=""):
        try:
            p.wait_for_load_state("domcontentloaded", timeout=8000)
        except Exception:
            pass
        u = real_url(p)
        if label:
            g2a_log(f"  ✓ {label}: {u[:70]}")
        return u

    def get_arrived_page():
        for p in context.pages:
            if arrived(p):
                return p
        return None

    # Already there?
    if arrived(page):
        sync_page(page, "Already on Returns portal")
        return True, page

    # ── Step 1: Click "Services" in top navbar ───────────────────────────────
    g2a_log("  -> Step 1: Click 'Services'...")
    for sel in [
        "ul.nav.navbar-nav li.dropdown > a:has-text('Services')",
        "li.dropdown > a:has-text('Services')",
        "nav a:has-text('Services')",
        "a:text-is('Services')",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=4000)
            loc.click()
            g2a_log(f"    ✓ 'Services' clicked")
            time.sleep(0.7)
            break
        except Exception:
            continue

    # ── Step 2: HOVER "Returns" tab — reveals sub-menu without navigating ──────
    # HOVER only (clicking "Returns" tab navigates to another page).
    # Use exact text to avoid matching "Returns Dashboard", "View Filed Returns" etc.
    g2a_log("  -> Step 2: Hover 'Returns' tab...")
    returns_ok = False
    for sel in [
        "ul.nav-tabs a:text-is('Returns')",
        ".nav-tabs li a:text-is('Returns')",
        ".service-sub-cat a:text-is('Returns')",
        "ul.sub-menu a:text-is('Returns')",
        "a:text-is('Returns')",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            loc.hover()
            g2a_log(f"    ✓ 'Returns' hovered — sub-menu should be visible")
            returns_ok = True
            time.sleep(0.8)
            break
        except Exception:
            continue

    if not returns_ok:
        g2a_log("    ⚠ 'Returns' hover failed — sub-menu may not be visible")

    # ── Step 3: CLICK "Returns Dashboard" with native Playwright click ───────
    # Native loc.click() fires real browser events — Angular handles it correctly.
    # dispatchEvent(MouseEvent) is intentionally NOT used here.
    g2a_log("  -> Step 3: Click 'Returns Dashboard'...")
    dash_ok = False
    for sel in [
        "a:text-is('Returns Dashboard')",
        "a:has-text('Returns Dashboard')",
        ".dropdown-menu a:has-text('Returns Dashboard')",
        "ul.list-unstyled a:has-text('Returns Dashboard')",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=4000)
            loc.click()
            g2a_log(f"    ✓ 'Returns Dashboard' clicked")
            dash_ok = True
            break
        except Exception:
            continue

    if not dash_ok:
        g2a_log("    ⚠ Click failed — extracting href and using page.goto()...")
        try:
            href = page.evaluate("""() => {
                for (const a of document.querySelectorAll('a')) {
                    const t = (a.textContent || '').trim();
                    if (t === 'Returns Dashboard') {
                        return a.href || a.getAttribute('href');
                    }
                }
                return null;
            }""")
            if href and "return.gst.gov.in" in href:
                g2a_log(f"    goto: {href[:70]}")
                page.goto(href, wait_until="domcontentloaded", timeout=20000)
                time.sleep(1.5)
        except Exception as e:
            g2a_log(f"    href fallback error: {e}")

    # ── Wait for navigation to return.gst.gov.in ─────────────────────────────
    g2a_log("  -> Waiting for return.gst.gov.in...")
    for i in range(15):
        time.sleep(1)
        if arrived(page):
            sync_page(page, f"Arrived after {i+1}s")
            return True, page
        active = get_arrived_page()
        if active:
            sync_page(active, f"Arrived (other tab) after {i+1}s")
            return True, active
        u = real_url(page)
        g2a_log(f"    [{i+1}s] {u[:65]}")

    # ── SSO quicklinks intermediary ───────────────────────────────────────────
    g2a_log("  -> Checking for SSO quicklinks page...")
    for p in context.pages:
        u = real_url(p)
        if "quicklinks" in u:
            g2a_log(f"    SSO page found: {u[:70]}")
            try:
                # Find and click the return.gst.gov.in link
                lnk = p.locator("a[href*='return.gst.gov.in']").first
                lnk.wait_for(state="visible", timeout=3000)
                lnk.click()
                time.sleep(2)
                if arrived(p):
                    sync_page(p, "SSO → Returns portal")
                    return True, p
            except Exception:
                pass
            # Direct navigation from SSO page
            try:
                p.goto(DASHBOARD_URL, wait_until="domcontentloaded", timeout=15000)
                time.sleep(1.5)
                if arrived(p):
                    sync_page(p, "SSO → direct URL")
                    return True, p
            except Exception:
                pass

    # ── 3-minute manual fallback ──────────────────────────────────────────────
    live_urls = [real_url(p)[:55] for p in context.pages]
    g2a_log(f"  ⚠ Auto navigation failed. Pages: {live_urls}")
    g2a_log("  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    g2a_log("  📋 ACTION NEEDED — in the browser window:")
    g2a_log("     Services  →  Returns  →  Returns Dashboard")
    g2a_log("  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    deadline = time.time() + 180
    while time.time() < deadline:
        time.sleep(2)
        if arrived(page):
            sync_page(page, "Manual nav")
            return True, page
        active = get_arrived_page()
        if active:
            sync_page(active, "Manual nav (other tab)")
            return True, active
    g2a_log("  ✗ Timed out.")
    return False, page


# ─────────────────────────────────────────────────────────────────────
#  Browser login
# ─────────────────────────────────────────────────────────────────────
GST_LOGIN = "https://services.gst.gov.in/services/login"

def g2a_do_browser_login(page, username, password):
    """Handle GST portal login: fill credentials, wait for captcha, submit."""
    g2a_log("Opening GST login page...")
    try:
        page.goto(GST_LOGIN, wait_until="domcontentloaded", timeout=25000)
        try: page.wait_for_selector(
                "input#username, input[name='username'], input[placeholder*='username' i]",
                state="visible", timeout=8000)
        except Exception: pass
    except Exception as e:
        g2a_log(f"  ✗ Could not open login page: {e}", "error")
        g2a_set({"status":"error","error":str(e)}); return False

    # Username
    for sel in ["input#username","input[name='username']","input[placeholder*='username' i]"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=4000)
            page.locator(sel).first.fill(username)
            g2a_log("  ✓ Username filled"); break
        except Exception: continue

    # Password
    for sel in ["input#user_pass","input[name='user_pass']",
                "input[type='password']","input[placeholder*='password' i]"]:
        try:
            page.locator(sel).first.fill(password)
            g2a_log("  ✓ Password filled"); break
        except Exception: continue

    # Proceed immediately to captcha capture

    # Capture captcha image — try element screenshot first (most reliable)
    cap_img = None
    try:
        import base64 as _b64
        # Method 1: direct element screenshot of the captcha <img>
        for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                    "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                    ".captchaImage img", "img.captcha",
                    # GST portal specific: img below password field
                    "form img", "img"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
                    g2a_log(f"  ✓ Captcha captured via element screenshot ({sel})")
                    break
            except Exception:
                continue

        # Method 2: canvas toDataURL
        if not cap_img:
            data_url = page.evaluate("""() => {
                for (const cv of document.querySelectorAll('canvas')) {
                    if (cv.width > 30 && cv.height > 10) {
                        try { return cv.toDataURL('image/png'); } catch(e) {}
                    }
                }
                return null;
            }""")
            if data_url and data_url.startswith("data:image"):
                cap_img = data_url
                g2a_log("  ✓ Captcha captured via canvas")

        # Method 3: full page screenshot fallback
        if not cap_img:
            png = page.screenshot(full_page=False)
            cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
            g2a_log("  ✓ Captcha captured via full screenshot (fallback)")

    except Exception as e:
        g2a_log(f"  ⚠ Captcha capture error: {e}")

    g2a_set({"status":"waiting_captcha", "captcha_image": cap_img})
    g2a_log("  ⏸ Waiting for captcha...")

    # Wait for user to submit captcha answer
    answer = g2a_wait_field("captcha_answer", timeout_sec=300)
    if not answer:
        g2a_log("  ✗ Captcha timeout", "error")
        g2a_set({"status":"error","error":"Captcha timeout"}); return False

    # Hide captcha box immediately — answer received, processing login
    g2a_set({"status":"running", "captcha_image": None})

    # Fill captcha
    for sel in ["input[placeholder*='Characters' i]","input#captcha","input[name='captcha']"]:
        try:
            page.locator(sel).first.fill(str(answer))
            g2a_log(f"  ✓ Captcha field: {sel}")
            g2a_log(f"  ✓ Captcha filled: {repr(answer)}")
            g2a_log(f"  ✓ Captcha: {answer}")
            break
        except Exception: continue

    # Submit login
    for sel in ["button[type='submit']","input[type='submit']","button:has-text('LOGIN')"]:
        try:
            page.locator(sel).first.click()
            g2a_log("  ✓ Login submitted")
            break
        except Exception: continue
    try: page.wait_for_load_state("domcontentloaded", timeout=4000)
    except Exception: time.sleep(0.8)

    # Handle OTP if needed
    otp_needed = False
    for _ in range(8):
        time.sleep(1)
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.wait_for(state="visible", timeout=1500)
            otp_needed = True; break
        except Exception:
            if check_login_success(page) is not False:
                break

    if otp_needed:
        g2a_set({"status":"waiting_otp"})
        g2a_log("  ⏸ Waiting for OTP...")
        otp = g2a_wait_field("otp_answer", timeout_sec=180)
        if not otp:
            g2a_log("  ✗ OTP timeout", "error")
            g2a_set({"status":"error","error":"OTP timeout"}); return False
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.click(); time.sleep(0.2)
            otp_el.fill(str(otp))
            page.locator("button[type='submit'],input[type='submit']").first.click()
            g2a_log("  ✓ OTP submitted"); time.sleep(1)
        except Exception as e:
            g2a_log(f"  ✗ OTP error: {e}", "error")
            return False

    time.sleep(1)

    # ── Captcha retry loop (up to 3 total attempts) ──────────
    MAX_CAPTCHA_ATTEMPTS = 3
    for _attempt in range(MAX_CAPTCHA_ATTEMPTS):
        result = check_login_success(page)
        if result is not False:
            # Logged in (True) or uncertain (None → proceed optimistically)
            break

        # Login failed — portal has already auto-refreshed the captcha image
        remaining = MAX_CAPTCHA_ATTEMPTS - _attempt - 1
        if remaining == 0:
            g2a_log(f"  ✗ Login failed after {MAX_CAPTCHA_ATTEMPTS} attempts", "error")
            g2a_set({"status":"error","error":f"Login failed after {MAX_CAPTCHA_ATTEMPTS} captcha attempts"})
            return False

        g2a_log(f"  ✗ Login failed — re-capturing fresh captcha (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")

        # Small wait for portal to settle after failure + captcha refresh
        time.sleep(2)

        # Re-capture the new captcha the portal already rendered
        new_cap = None
        try:
            import base64 as _b64
            for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                        "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                        ".captchaImage img", "img.captcha", "form img", "img"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=2000)
                    png = loc.screenshot()
                    if png and len(png) > 500:
                        new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                        g2a_log(f"  ✓ Fresh captcha re-captured ({sel})")
                        break
                except Exception:
                    continue
            if not new_cap:
                png = page.screenshot(full_page=False)
                new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                g2a_log("  ✓ Fresh captcha via screenshot (fallback)")
        except Exception as ce:
            g2a_log(f"  ⚠ Re-capture error: {ce}")

        # Signal UI to show new captcha — force=True by clearing then setting
        # so the JS !alreadyShown guard triggers and fetches the new image
        g2a_set({"status":"running",  "captcha_image": None})
        time.sleep(0.1)
        g2a_set({"status":"waiting_captcha", "captcha_image": new_cap})
        g2a_log(f"  ⏸ Waiting for captcha answer (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")

        new_answer = g2a_wait_field("captcha_answer", timeout_sec=300)
        if not new_answer:
            g2a_log("  ✗ Captcha retry timeout", "error")
            g2a_set({"status":"error","error":"Captcha retry timeout"})
            return False

        g2a_set({"status":"running", "captcha_image": None})

        # Clear old captcha field, fill new answer, re-submit
        for sel in ["input[placeholder*='Characters' i]","input#captcha","input[name='captcha']"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.fill("")
                loc.fill(str(new_answer))
                g2a_log(f"  ✓ Re-filled captcha: {repr(new_answer)}")
                break
            except Exception:
                continue

        for sel in ["button[type='submit']","input[type='submit']","button:has-text('LOGIN')"]:
            try:
                page.locator(sel).first.click()
                g2a_log("  ✓ Login re-submitted")
                break
            except Exception:
                continue

        time.sleep(2)
        # Handle OTP in case it appears on retry (unlikely but safe)
        for _ in range(5):
            time.sleep(1)
            try:
                otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
                otp_el.wait_for(state="visible", timeout=1000)
                # OTP appeared on retry — handle it
                g2a_set({"status":"waiting_otp"})
                g2a_log("  ⏸ OTP required (retry)...")
                otp_r = g2a_wait_field("otp_answer", timeout_sec=180)
                if not otp_r:
                    g2a_log("  ✗ OTP timeout (retry)", "error")
                    g2a_set({"status":"error","error":"OTP timeout on retry"})
                    return False
                otp_el.click(); time.sleep(0.2)
                otp_el.fill(str(otp_r))
                page.locator("button[type='submit'],input[type='submit']").first.click()
                g2a_log("  ✓ OTP re-submitted"); time.sleep(1)
                break
            except Exception:
                if check_login_success(page) is not False:
                    break
        # Loop back to check_login_success at top of for loop

    g2a_log("  ✅ Logged in successfully")
    g2a_dismiss_popup(page)
    return True


# ─────────────────────────────────────────────────────────────────────
#  Download file from URL
# ─────────────────────────────────────────────────────────────────────
def g2a_fetch_download_link(page, url, gstin_dir, month_display,
                            expected_period=""):
    """
    Download the JSON/ZIP from 'url' and save it as GSTR2A_<Month_Year>.json.

    PERIOD CORRECTNESS GUARD (Requirement 2)
    ──────────────────────────────────────────
    The portal has a known bug: when the RPA fails to select a particular
    month, it sometimes falls back to showing the *most recently generated*
    file's download link.  Clicking that link downloads the wrong month's
    data, which previously got saved under the wrong filename.

    Fix: after downloading the raw bytes, parse the JSON and read the 'fp'
    field (MMYYYY, e.g. '042025' for April 2025).  If it does NOT match
    expected_period the file is REJECTED — not saved — and the function
    returns None so the multi-pass sweep retries that month later.

    Args:
        expected_period : 'MMYYYY' string matching month['period'].
                          Empty string → skip verification (safe fallback).
    """
    abbr      = month_display.replace(" ", "_")
    json_path = os.path.join(gstin_dir, f"GSTR2A_{abbr}.json")
    try:
        resp = page.request.get(url, headers={
            "Accept":  "*/*",
            "Referer": "https://return.gst.gov.in/returns/auth/gstr/offlinedownload",
        }, timeout=120000)
        raw = resp.body()
        ct  = resp.headers.get("content-type", "").lower()
        g2a_log(f"    [{resp.status}] {ct[:40]}  {len(raw)} bytes")
        if resp.status != 200 or not raw:
            return None

        # Unzip if needed
        import zipfile, io as _io
        if raw[:2] == b"PK" or "zip" in ct:
            with zipfile.ZipFile(_io.BytesIO(raw)) as zf:
                names  = zf.namelist()
                target = next((n for n in names if n.endswith(".json")), names[0])
                data   = zf.read(target)
        else:
            data = raw

        # ── Period correctness check ──────────────────────────────────────
        if expected_period:
            try:
                parsed    = json.loads(data.decode("utf-8", errors="replace"))
                actual_fp = (parsed.get("fp") or "").strip()
                if actual_fp and actual_fp != expected_period:
                    # WRONG MONTH — reject entirely, do not write to disk
                    g2a_log(
                        f"    ❌ PERIOD MISMATCH — wanted fp='{expected_period}' "
                        f"but file contains fp='{actual_fp}' "
                        f"(portal returned wrong month's data)", "error")
                    g2a_log(f"       Rejecting file for {month_display}. "
                            f"Will retry in next pass.")
                    # Save a .mismatch debug copy so user can inspect if needed
                    debug = json_path + f".mismatch_{actual_fp}"
                    with open(debug, "wb") as f: f.write(data)
                    g2a_log(f"       Debug copy → {os.path.basename(debug)}")
                    return None          # ← triggers retry in multi-pass loop
                elif actual_fp:
                    g2a_log(f"    ✓ Period verified: fp='{actual_fp}' "
                            f"matches expected '{expected_period}'")
                else:
                    g2a_log(f"    ⚠ No fp field found in JSON — "
                            f"accepting without period check")
            except Exception as _pe:
                g2a_log(f"    ⚠ Period check error ({_pe}) — accepting anyway")

        # ── Save verified data ────────────────────────────────────────────
        with open(json_path, "wb") as f: f.write(data)
        size_kb = max(1, len(data) // 1024)
        g2a_log(f"    ✅ {month_display} — {size_kb} KB → {os.path.basename(json_path)}")
        return {"month": month_display, "filename": f"GSTR2A_{abbr}.json",
                "size_kb": size_kb}

    except Exception as e:
        g2a_log(f"    ✗ Download error: {e}", "error")
        return None


def g2a_navigate_offline_download(page, fy):
    """
    Navigate Returns Dashboard → select FY/Quarter/Period → SEARCH
    → click GSTR2A DOWNLOAD → reach offlinedownload page.

    Also captures the real download URL from "Click here to download JSON - File 1"
    for use in parallel downloads.
    Returns True if offlinedownload page reached.
    """
    now = datetime.now()
    month_names = {
        1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
        7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"
    }
    fy_start  = int(fy.split("-")[0])
    fy_months = list(range(4,13)) + list(range(1,4))
    fy_years  = [fy_start]*9 + [fy_start+1]*3
    # Most recent completed month in FY
    use_idx = -1
    for idx, (m, y) in enumerate(zip(fy_months, fy_years)):
        if y < now.year or (y == now.year and m < now.month):
            use_idx = idx
    use_month = fy_months[use_idx] if use_idx >= 0 else 4
    use_year  = fy_years[use_idx]  if use_idx >= 0 else fy_start
    quarter_label, _ = g2a_quarter_for_month(use_month)
    month_name = month_names[use_month]
    g2a_log(f"  -> Selecting: FY={fy}  Q={quarter_label}  Period={month_name}")

    def js_select(nth, label):
        result = page.evaluate(f"""() => {{
            const sel = document.querySelectorAll('select')[{nth}];
            if (!sel) return {{ok:false, opts:[]}};
            const opts = Array.from(sel.options).map(o => o.text.trim());
            for (const opt of sel.options) {{
                if (opt.text.trim() === '{label}') {{
                    sel.value = opt.value;
                    sel.dispatchEvent(new Event('change', {{bubbles:true}}));
                    return {{ok:true, selected:opt.text.trim(), opts:opts}};
                }}
            }}
            return {{ok:false, opts:opts}};
        }}""")
        g2a_log(f"    select[{nth}] '{label}': ok={result.get('ok')} "
                f"opts={result.get('opts',[])} selected='{result.get('selected','')}'")
        return result.get('ok', False)

    def wait_for_option(nth, label, timeout=5):
        deadline = time.time() + timeout
        while time.time() < deadline:
            opts = page.evaluate(f"""() => {{
                const sel = document.querySelectorAll('select')[{nth}];
                return sel ? Array.from(sel.options).map(o=>o.text.trim()) : [];
            }}""")
            if label in opts:
                return True
            time.sleep(0.3)
        return False

    # FY
    js_select(0, fy); time.sleep(0.5)
    # Quarter — wait for options to reload after FY
    wait_for_option(1, quarter_label, timeout=4)
    js_select(1, quarter_label)
    # Period — wait for month options to reload after Quarter
    g2a_log(f"    Waiting for '{month_name}' in Period dropdown...")
    if wait_for_option(2, month_name, timeout=6):
        js_select(2, month_name)
    else:
        g2a_log("    ⚠ Period did not reload — trying anyway")
        js_select(2, month_name)

    # SEARCH
    for s in ["button:has-text('SEARCH')","button:has-text('Search')"]:
        try:
            page.locator(s).first.wait_for(state="visible", timeout=4000)
            page.locator(s).first.click()
            g2a_log(f"    Clicked SEARCH ({s})")
            break
        except Exception: continue
    time.sleep(2)

    # Dump tiles for diagnosis
    btns = page.evaluate("""() =>
        Array.from(document.querySelectorAll('button,a.btn,.btn'))
            .map(e => (e.textContent||'').trim().slice(0,40))
            .filter(t => t.length > 0)
    """)
    g2a_log(f"    Buttons after search: {btns[:20]}")

    # Click DOWNLOAD inside GSTR2A tile
    clicked = page.evaluate("""() => {
        // Find element with text GSTR2A or GSTR-2A
        const all = Array.from(document.querySelectorAll('*'));
        for (const el of all) {
            const t = (el.childNodes.length === 1
                ? el.textContent : el.getAttribute('data-title') || '')
                .trim();
            if (t === 'GSTR2A' || t === 'GSTR-2A') {
                let container = el;
                for (let i = 0; i < 6; i++) {
                    container = container.parentElement;
                    if (!container) break;
                    for (const btn of container.querySelectorAll('button,a')) {
                        const bt = (btn.textContent||'').trim().toUpperCase();
                        if (bt === 'DOWNLOAD') {
                            btn.click();
                            return 'DOWNLOAD clicked in ' + (container.className||'?').slice(0,30);
                        }
                    }
                }
            }
        }
        return null;
    }""")
    g2a_log(f"    GSTR2A DOWNLOAD click: {clicked}")

    try:
        page.wait_for_url("*offlinedownload*", timeout=10000)
    except Exception:
        pass
    time.sleep(1.5)
    g2a_log(f"    -> Now on: {page.url[:70]}")

    on_page = "offlinedownload" in page.url
    if on_page:
        # Capture download link URL pattern for parallel downloads
        link = page.evaluate("""() => {
            for (const a of document.querySelectorAll('a')) {
                const t = (a.textContent||'').toLowerCase();
                if (t.includes('click here') || t.includes('json - file') || t.includes('json-file'))
                    return a.href;
            }
            return null;
        }""")
        if link:
            g2a_log(f"    📡 Existing download link: {link[:100]}")
            with g2a_lock:
                g2a_state["_download_link_pattern"] = link

    g2a_log_discovered_apis()
    return on_page


def g2a_download_all_months(page, months, gstin, gstin_dir, fy):
    """
    ══════════════════════════════════════════════════════════════════════
    MULTI-PASS SWEEP STRATEGY
    ══════════════════════════════════════════════════════════════════════
    The GST portal takes 5–10 minutes to generate a JSON file after you
    click GENERATE.  The old approach waited inline per month (sequential):
       12 months × 10 min = up to 2 hours.

    New approach — parallel generation via sweep passes:

    PASS 1 — "Trigger sweep":
      Visit each month quickly.  For every month:
        • Navigate dashboard → select FY/Quarter/Period → SEARCH
        • Click GSTR2A DOWNLOAD → offlinedownload page
        • If link already exists  → download immediately  ✅
        • If no link              → click GENERATE (kicks off server-side
                                    generation) then wait only 5 seconds
                                    → move on to next month  ⏩
      Result: all months whose JSON wasn't pre-generated are now queuing
              on the server simultaneously.

    PASS 2, 3, 4 … — "Collection sweeps":
      Re-visit every still-pending month (same navigation flow).
      Wait 5 seconds for the link.
        • Link appeared → download  ✅
        • Still not ready → leave for next pass  ⏩
      Continue until all months collected or max 25-minute wall-clock limit.

    Net effect: 12 months in ~8–12 minutes instead of potentially 2 hours.
    ══════════════════════════════════════════════════════════════════════
    """
    LINK_WAIT_SEC  = 5      # seconds to wait per visit before moving on
    MAX_WALL_MIN   = 25     # overall timeout in minutes
    dashboard_url  = "https://return.gst.gov.in/returns/auth/dashboard"

    month_names = {
        1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
        7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"
    }

    files_done   = []
    pending      = list(months)   # months still needing a file
    wall_start   = time.time()
    pass_num     = 0

    # ── Shared helpers ─────────────────────────────────────────────────────────
    def pw_select(nth, label, timeout=8):
        """Select nth <select> using Playwright native API — fires real browser
        events that Angular responds to (unlike synthetic dispatchEvent)."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                loc = page.locator("select").nth(nth)
                loc.wait_for(state="visible", timeout=2000)
                loc.select_option(label=label)
                time.sleep(0.4)
                cur = loc.evaluate(
                    "el => el.options[el.selectedIndex] "
                    "? el.options[el.selectedIndex].text.trim() : ''")
                if cur == label:
                    return True
                loc.select_option(label=label)   # retry once if Angular reset it
                time.sleep(0.5)
                return True
            except Exception:
                time.sleep(0.4)
        return False

    def wait_for_option(nth, label, timeout=8):
        """Poll until label appears as an option in nth <select>."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                opts = page.locator("select").nth(nth).evaluate(
                    "el => Array.from(el.options).map(o => o.text.trim())")
                if label in opts:
                    return True
            except Exception:
                pass
            time.sleep(0.4)
        return False

    def get_json_link():
        """Return href of the 'Click here to download JSON' link, or None."""
        return page.evaluate("""() => {
            for (const a of document.querySelectorAll('a')) {
                const t = (a.textContent||'').toLowerCase();
                if (t.includes('click here') || t.includes('json - file')
                        || t.includes('json-file'))
                    return a.href;
            }
            return null;
        }""")

    def navigate_to_month(month):
        """
        Dashboard → FY/Qtr/Period dropdowns → SEARCH → GSTR2A DOWNLOAD tile
        → offlinedownload page.
        Uses page.goto() (not window.location.href) so Playwright stays
        synchronised and page.evaluate() never hits a mid-navigation context.
        Uses pw_select() (Playwright native) so Angular responds to FY change.
        """
        mon_num       = month["num"]
        mon_name      = month_names[mon_num]
        quarter_label, _ = g2a_quarter_for_month(mon_num)

        # ── Navigate to dashboard ─────────────────────────────────────────
        # Use window.location.href (preserves GST session/cookies).
        # page.goto() breaks the session — portal redirects to access denied.
        # If the context is already destroyed, fall back to page.goto().
        # Then wait_for_load_state() re-syncs Playwright's frame reference.
        try:
            page.evaluate(f"window.location.href = '{dashboard_url}'")
        except Exception:
            page.goto(dashboard_url, wait_until="domcontentloaded", timeout=15000)
        try:
            page.wait_for_load_state("domcontentloaded", timeout=12000)
        except Exception:
            pass
        time.sleep(1.5)

        # ── FY dropdown (index 0) via Playwright native ───────────────────
        g2a_log(f"    Selecting FY: {fy}")
        pw_select(0, fy)
        time.sleep(0.5)

        # ── Detect layout: poll for ≥2 selects to load after FY change ────
        #   3 selects → FY + Quarter + Month  (normal)
        #   2 selects → FY + Month only       (no-quarter portal variant)
        n_selects = 0
        for _w in range(15):           # up to 4.5 s
            try:
                n_selects = page.evaluate(
                    "() => document.querySelectorAll('select').length")
            except Exception:
                pass
            if n_selects >= 2:
                break
            time.sleep(0.3)
        g2a_log(f"    {n_selects} select(s) detected after FY change")

        if n_selects >= 3:
            # ── Normal layout: FY → Quarter → Month ──────────────────────
            g2a_log(f"    Selecting Quarter: {quarter_label}")
            wait_for_option(1, quarter_label, timeout=6)
            pw_select(1, quarter_label)
            g2a_log(f"    Selecting Month: {mon_name}")
            wait_for_option(2, mon_name, timeout=8)
            pw_select(2, mon_name)
        else:
            # ── 2-dropdown layout: FY → Month (no quarter) ───────────────
            g2a_log(f"    Selecting Month: {mon_name} (no quarter dropdown)")
            wait_for_option(1, mon_name, timeout=8)
            pw_select(1, mon_name)

        # ── SEARCH ────────────────────────────────────────────────────────
        for s in ["button:has-text('SEARCH')", "button:has-text('Search')"]:
            try:
                page.locator(s).first.wait_for(state="visible", timeout=4000)
                page.locator(s).first.click()
                g2a_log(f"    ✓ SEARCH clicked")
                break
            except Exception:
                continue
        time.sleep(2)

        # ── Click DOWNLOAD on GSTR2A tile ─────────────────────────────────
        clicked = page.evaluate("""() => {
            for (const el of document.querySelectorAll('*')) {
                const t = (el.childNodes.length === 1
                    ? el.textContent : el.getAttribute('data-title')||'').trim();
                if (t === 'GSTR2A' || t === 'GSTR-2A') {
                    let c = el;
                    for (let i=0; i<6; i++) {
                        c = c.parentElement;
                        if (!c) break;
                        for (const btn of c.querySelectorAll('button,a')) {
                            if ((btn.textContent||'').trim().toUpperCase() === 'DOWNLOAD') {
                                btn.click(); return 'ok';
                            }
                        }
                    }
                }
            }
            return null;
        }""")
        if not clicked:
            g2a_log(f"    ⚠ GSTR2A DOWNLOAD button not found for {month['display']}")
            return False

        try:
            page.wait_for_url("*offlinedownload*", timeout=10000)
        except Exception:
            pass
        time.sleep(1)

        if "offlinedownload" not in page.url:
            g2a_log(f"    ⚠ Did not reach offlinedownload — on {page.url[:60]}")
            return False

        return True

    # ══════════════════════════════════════════════════════════════════════
    # SWEEP LOOP
    # ══════════════════════════════════════════════════════════════════════
    while pending and (time.time() - wall_start) < (MAX_WALL_MIN * 60):
        pass_num  += 1
        still_pending = []

        if pass_num == 1:
            g2a_log(f"\n🔁 PASS 1 — Trigger sweep: visiting all {len(pending)} months, "
                    f"waiting {LINK_WAIT_SEC}s per month for link")
        else:
            elapsed = int((time.time() - wall_start) / 60)
            g2a_log(f"\n🔁 PASS {pass_num} — Collection sweep: "
                    f"{len(pending)} month(s) still pending  [{elapsed}m elapsed]")

        for i, month in enumerate(pending):
            mon_display = month["display"]
            period      = month["period"]
            total_done  = len(months) - len(pending) + i  # approx progress
            g2a_set({
                "status":        "downloading",
                "current_month": f"{mon_display} (pass {pass_num})",
                "progress":      int((total_done / len(months)) * 100),
            })
            g2a_log(f"  [{i+1}/{len(pending)}] {mon_display}  (pass {pass_num})")

            # Navigate to the offlinedownload page for this month
            on_page = navigate_to_month(month)
            if not on_page:
                g2a_log(f"    ⏩ Navigation failed — will retry next pass")
                still_pending.append(month)
                continue

            # ── Check for pre-existing link ───────────────────────────────────
            link_href = get_json_link()
            if link_href:
                g2a_log(f"    ✅ Link already present — downloading immediately")
            else:
                # ── Click GENERATE to trigger server-side generation ──────────
                for s in ["button:has-text('GENERATE JSON FILE TO DOWNLOAD')",
                          "button:has-text('GENERATE JSON')",
                          "button:has-text('GENERATE')"]:
                    try:
                        page.locator(s).first.wait_for(state="visible", timeout=4000)
                        page.locator(s).first.click()
                        g2a_log(f"    ▶ Clicked GENERATE — waiting {LINK_WAIT_SEC}s "
                                f"for link to appear...")
                        break
                    except Exception:
                        continue

                # ── Wait only LINK_WAIT_SEC seconds then move on ──────────────
                deadline = time.time() + LINK_WAIT_SEC
                while time.time() < deadline:
                    time.sleep(0.5)
                    link_href = get_json_link()
                    if link_href:
                        g2a_log(f"    ⚡ Link appeared within {LINK_WAIT_SEC}s!")
                        break

            # ── Download if link available ────────────────────────────────────
            if link_href:
                result = g2a_fetch_download_link(page, link_href, gstin_dir,
                                                 mon_display,
                                                 expected_period=period)
                if result:
                    result["period"] = period
                    files_done.append(result)
                    g2a_set({"files": files_done.copy(),
                             "done_months": len(files_done)})
                    g2a_log(f"    ✅ {mon_display} downloaded  "
                            f"({len(files_done)}/{len(months)} total)")
                    # DO NOT add to still_pending — done
                    continue
                else:
                    g2a_log(f"    ⚠ Download returned None (404 / error / period mismatch)"
                            f" — re-clicking GENERATE to request a fresh file")
                    # Re-trigger generation so the next pass finds a fresh link
                    for s in ["button:has-text('GENERATE JSON FILE TO DOWNLOAD')",
                              "button:has-text('GENERATE JSON')",
                              "button:has-text('GENERATE')"]:
                        try:
                            page.locator(s).first.wait_for(state="visible", timeout=3000)
                            page.locator(s).first.click()
                            g2a_log(f"    ▶ GENERATE re-clicked — will collect next pass")
                            break
                        except Exception: continue
            else:
                g2a_log(f"    ⏩ Not ready yet — server still generating. "
                        f"Will revisit in next pass.")

            still_pending.append(month)

        pending = still_pending

        if pending and pass_num == 1:
            g2a_log(f"\n  ✅ Pass 1 complete — triggered generation for "
                    f"{len(pending)} month(s).")
            g2a_log(f"  ⏳ Portal is generating files in background "
                    f"(takes 5–10 min per month).")
            g2a_log(f"  🔁 Starting collection passes now...")
        elif pending:
            g2a_log(f"  {len(pending)} month(s) still pending — starting next pass...")

    # ── Final report ──────────────────────────────────────────────────────────
    elapsed_min = (time.time() - wall_start) / 60
    if pending:
        g2a_log(f"\n⚠ Timed out after {elapsed_min:.1f} min — "
                f"{len(pending)} month(s) not downloaded: "
                f"{[m['display'] for m in pending]}", "error")
    else:
        g2a_log(f"\n✅ All months collected in {pass_num} pass(es), "
                f"{elapsed_min:.1f} min total")

    return files_done


def g2a_worker(gstin, fy, username, password, specific_month="", force=False):
    """
    Thin wrapper guaranteeing any exception is logged and reflected in
    g2a_state instead of silently killing the thread — same reasoning
    as the matching wrapper on combined_download_worker.
    """
    try:
        _g2a_worker_impl(gstin, fy, username, password, specific_month, force)
    except Exception as fatal:
        import traceback as _tb
        try:
            g2a_log(f"✗ FATAL (uncaught): {fatal}", "error")
        except Exception:
            pass
        try:
            g2a_set({"status": "error", "error": f"Uncaught: {fatal}"})
        except Exception:
            pass
        log.error(f"[G2A] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _g2a_worker_impl(gstin, fy, username, password, specific_month="", force=False):
    os.makedirs(G2A_DOWNLOAD_DIR, exist_ok=True)
    gstin_dir = os.path.join(G2A_DOWNLOAD_DIR, gstin, fy.replace("-","_"))
    os.makedirs(gstin_dir, exist_ok=True)

    all_months = g2a_months_for_fy(fy)
    if not all_months:
        g2a_set({"status":"error","error":f"No months available for FY {fy}"}); return

    if specific_month:
        sm = specific_month.strip()
        filtered = [m for m in all_months if (
            sm.lower() == m["display"].lower() or sm == m["period"] or
            sm.lower() == m["name"].lower()    or sm.upper() == m["abbr"])]
        if not filtered:
            g2a_set({"status":"error",
                     "error":f"Month '{specific_month}' not found in FY {fy}. "
                             f"Available: {[m['display'] for m in all_months]}"}); return
        months = filtered
    else:
        months = all_months

    # Cache check — skip months whose file is fresh enough
    fresh_results, months_needed = g2a_check_cached_months(gstin_dir, months, force=force)
    g2a_set({"total_months":len(months),"done_months":len(fresh_results),"files":fresh_results.copy()})

    if not months_needed:
        g2a_log(f"GSTR-2A | All {len(months)} month(s) fresh (< {g2a_stale_days()} days). "
                f"Tick Force Re-download to refresh.")
        g2a_set({"status":"done","progress":100,"current_month":None})
        time.sleep(8)
        with g2a_lock:
            if g2a_state.get("status") == "done": g2a_state["status"] = "idle"
        return

    g2a_log(f"GSTR-2A | GSTIN: {gstin} | FY: {fy} | "
            f"{len(months_needed)} to download, {len(fresh_results)} from cache")

    profile_dir = os.path.join(PATHS.profiles_dir, "gst_profile_g2a")
    os.makedirs(profile_dir, exist_ok=True)
    lock_file = os.path.join(profile_dir, "SingletonLock")
    if os.path.exists(lock_file):
        try: os.remove(lock_file); g2a_log("  ✓ Removed stale SingletonLock")
        except Exception: pass

    # sync_playwright().start() launches Playwright's own Node.js driver
    # subprocess — if it throws (not just hangs), that was previously
    # uncaught here, silently killing this thread with nothing logged.
    try:
        pw = sync_playwright().start()
    except Exception as e:
        g2a_log(f"✗ Playwright driver failed to start: {e}", "error")
        g2a_set({"status": "error", "error": f"Playwright driver failed to start: {e}"})
        return
    try:
        g2a_log("🌐 Launching browser...")
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=False, slow_mo=40,
                args=["--disable-blink-features=AutomationControlled",
                      "--no-sandbox","--start-maximized"],
                no_viewport=True, accept_downloads=True,
            )
        except Exception as e:
            g2a_log(f"  ✗ Browser launch error: {e}", "error")
            g2a_set({"status":"error","error":str(e)}); return

        page = context.pages[0] if context.pages else context.new_page()
        page.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        g2a_install_interceptors(page)

        if not g2a_do_browser_login(page, username, password):
            context.close(); return

        g2a_log("🔑 Activating session on Returns portal...")
        ok, page = g2a_activate_session(page, context)
        if not ok:
            g2a_set({"status":"error","error":"Could not reach Returns portal"})
            context.close(); return
        g2a_log(f"  Using page: {page.url[:70]}")
        g2a_install_interceptors(page)

        g2a_log(f"📥 Starting GSTR-2A multi-pass download sweep...")
        new_files = g2a_download_all_months(page, months_needed, gstin, gstin_dir, fy)

        all_files = fresh_results + new_files
        g2a_set({"status":"done","progress":100,"current_month":None,
                 "files":all_files,"done_months":len(all_files)})
        g2a_log(f"\n✅ Complete — {len(new_files)} downloaded, "
                f"{len(fresh_results)} from cache, {len(all_files)} total")
        g2a_log(f"   Folder: {gstin_dir}")

        try: context.close(); g2a_log("🌐 Browser closed")
        except Exception as _ce: g2a_log(f"  ⚠ Browser close: {_ce}")
    except Exception as _e:
        g2a_set({"status":"error","error":str(_e)})
        g2a_log(f"✗ Fatal worker error: {_e}", "error")
        import traceback; g2a_log(traceback.format_exc(), "error")
    finally:
        try: pw.stop()
        except Exception: pass

    time.sleep(8)
    with g2a_lock:
        if g2a_state.get("status") == "done": g2a_state["status"] = "idle"
    g2a_log("🔁 GSTR-2A RPA reset to idle")



# ── GSTR-2A → Excel Converter ────────────────────────────────
def g2a_json_to_excel(gstin, fy, name_lookup=None):
    """
    Read all downloaded GSTR-2A JSON files for gstin+fy,
    combine all months into a single Excel workbook.

    Sheets:
      B2B Purchases     — all b2b + b2ba invoice line items
      Credit-Debit Notes— all cdnr + cdnra note line items
      Import of Goods   — impg (Bill of Entry)
      ISD Credits       — isd + isda
      Summary           — month-wise totals
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    import glob, re

    fy_dir     = fy.replace("-", "_")
    folder     = os.path.join(G2A_DOWNLOAD_DIR, gstin, fy_dir)
    json_files = sorted(glob.glob(os.path.join(folder, "*.json")))
    if not json_files:
        return None, f"No JSON files found in {folder}"

    # ── Styles ────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")
    HDR2_FILL = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")
    TOT_FILL  = PatternFill("solid", fgColor="D6E4F0")
    RISK_RED  = PatternFill("solid", fgColor="FFD7D7")   # cfs=N  (ITC blocked)
    RISK_AMB  = PatternFill("solid", fgColor="FFF3CD")   # cfs3b=N only (reversal risk)
    AMEND_FILL= PatternFill("solid", fgColor="FFE0B2")   # Amended row (b2ba/cdna)
    SUPERSED  = PatternFill("solid", fgColor="FF6600")   # Superseded original — DELETE
    CDN_CR    = PatternFill("solid", fgColor="FFE8E8")   # Credit Note row
    CDN_DR    = PatternFill("solid", fgColor="E8F5E9")   # Debit Note row
    IMP_FILL  = PatternFill("solid", fgColor="EDF2FB")   # Import BOE row
    TCS_FILL  = PatternFill("solid", fgColor="F3E5F5")   # TCS row
    TDS_FILL  = PatternFill("solid", fgColor="E8EAF6")   # TDS row (indigo tint)
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    HDR2_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    BODY_FONT = Font(name="Arial", size=9)
    BOLD_FONT = Font(name="Arial", bold=True, size=9)
    TOT_FONT  = Font(name="Arial", bold=True, size=9, color="1F3864")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    thin      = Side(style="thin", color="B0C4DE")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM_FMT   = '#,##0.00'
    INT_FMT   = '#,##0'

    def style_header_row(ws, row, col_count, fill=HDR_FILL, font=HDR_FONT):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill   = fill
            cell.font   = font
            cell.alignment = CENTER
            cell.border = BORDER

    def style_body_row(ws, row, col_count, alt=False):
        fill = ALT_FILL if alt else None
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            if fill: cell.fill = fill
            cell.font   = BODY_FONT
            cell.border = BORDER
            # right-align numeric columns (detect by value)
            if isinstance(cell.value, (int, float)):
                cell.alignment = RIGHT
                cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    def style_total_row(ws, row, col_count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill   = TOT_FILL
            cell.font   = TOT_FONT
            cell.border = BORDER
            if isinstance(cell.value, (int, float)):
                cell.alignment = RIGHT
                cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def month_from_filename(fname):
        """Extract month display from filename like GSTR2A_April_2024.json"""
        base = os.path.basename(fname).replace("GSTR2A_", "").replace(".json", "")
        return base.replace("_", " ")

    def fp_to_month_display(fp):
        """Convert fp like '042024' → 'April 2024'"""
        months = {
            "01":"January","02":"February","03":"March","04":"April",
            "05":"May","06":"June","07":"July","08":"August",
            "09":"September","10":"October","11":"November","12":"December"
        }
        if fp and len(fp) >= 6:
            mm = fp[:2]; yyyy = fp[2:6]
            return f"{months.get(mm, mm)} {yyyy}"
        return fp or "Unknown"

    # ── Load all months ────────────────────────────────────────
    months_data = []
    for fpath in json_files:
        try:
            with open(fpath, encoding="utf-8") as f:
                data = json.load(f)
            month_display = month_from_filename(fpath)
            if not month_display or month_display == "Unknown":
                month_display = fp_to_month_display(data.get("fp",""))
            months_data.append((month_display, data))
        except Exception as e:
            g2a_log(f"  ⚠ Skipping {fpath}: {e}")

    if not months_data:
        return None, "Could not parse any JSON files"

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ════════════════════════════════════════════════════════════
    # STATE CODE LOOKUP
    # ════════════════════════════════════════════════════════════
    STATE_CODES = {
        "01":"Jammu & Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
        "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"Uttar Pradesh",
        "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
        "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
        "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh",
        "24":"Gujarat","25":"Daman & Diu","26":"Dadra & Nagar Haveli","27":"Maharashtra",
        "28":"Andhra Pradesh (Old)","29":"Karnataka","30":"Goa","31":"Lakshadweep",
        "32":"Kerala","33":"Tamil Nadu","34":"Puducherry","35":"Andaman & Nicobar",
        "36":"Telangana","37":"Andhra Pradesh","38":"Ladakh","97":"Other Territory",
        "99":"Centre Jurisdiction",
    }
    def pos_display(code):
        s = str(code).zfill(2)
        return f"{s}-{STATE_CODES.get(s, code)}" if s in STATE_CODES else str(code)

    INV_TYPES = {
        "R":"Regular","SEWP":"SEZ w/ Payment","SEWOP":"SEZ w/o Payment",
        "DE":"Deemed Export","CBW":"Customs Bonded Warehouse",
    }

    # ════════════════════════════════════════════════════════════
    # SHEET 1 — B2B PURCHASES
    # Col layout:
    # A=Month  B=SupplierGSTIN  C=InvNo  D=InvDate  E=InvValue
    # F=InvType  G=POS  H=RevCharge  I=TaxRate  J=TaxableVal
    # K=IGST  L=CGST  M=SGST  N=Cess  O=TotalTax
    # P=SupplierFiledDate  Q=SupplierPeriod
    # ════════════════════════════════════════════════════════════
    # ── Load names: merge per-return (2a/) + enriched/ store ─────────────────
    # Priority: gstin_names/enriched/ (portal, legal+trade) >
    #           gstin_names/2a/ (scan-built, trade from G2B) >
    #           G2B trdnm scan (last resort) > ""
    _disk_names = _gnames_load_for("2a")   # {GSTIN_UPPER: {trade_name, legal_name}}
    if name_lookup is None:
        name_lookup = {}
    # Merge disk names (win) over any POST-body names
    _all_keys = set(_disk_names) | set(name_lookup)
    _merged   = {}
    for _k in _all_keys:
        _fl = _disk_names.get(_k,  {})
        _nl = name_lookup.get(_k, {})
        _merged[_k] = {
            "trade_name": _fl.get("trade_name","") or _nl.get("trade_name","") or "",
            "legal_name": _fl.get("legal_name", "") or _nl.get("legal_name", "") or "",
        }
    name_lookup = _merged
    g2a_log(f"  names: {sum(1 for v in name_lookup.values() if v['trade_name'])} trade, "
            f"{sum(1 for v in name_lookup.values() if v['legal_name'])} legal "
            f"(from gstin_names/2a/ + enriched/)")

    ws_b2b = wb.create_sheet("B2B Purchases")
    # ── Column map (26 cols) ────────────────────────────────────
    # Zone 1 – Identity (A–H):
    #   A=Month  B=SupplierGSTIN  C=TradeName  D=LegalName
    #   E=InvNo  F=InvDate  G=InvType  H=SrcType  I=IRN  J=IRNGenDate
    # Zone 2 – Financial (K–T):
    #   K=InvValue  L=POS  M=RevCharge  N=TaxRate%
    #   O=TaxableValue  P=IGST  Q=CGST  R=SGST  S=Cess  T=TotalTax
    # Zone 3 – Supplier compliance + metadata (U–X):
    #   U=GSTR1Status  V=GSTR3BStatus  W=SupplierFiledDate  X=SupplierPeriod
    # Zone 4 – Amended original ref (Y–Z)
    #   Y=OrigInvoiceNo  Z=OrigInvoiceDate
    # ────────────────────────────────────────────────────────────
    B2B_COLS = [
        "Month","Supplier GSTIN","Trade Name","Legal Name",
        "Invoice No","Invoice Date","Inv Type","Source Type","IRN","IRN Gen Date",
        "Invoice Value","Place of Supply","Rev. Charge","Tax Rate %",
        "Taxable Value","IGST","CGST","SGST","Cess","Total Tax",
        "GSTR-1 Status","GSTR-3B Status",
        "Supplier Filed Date","Supplier Period",
        "Orig Invoice No","Orig Invoice Date"
    ]
    B2B_WIDTHS = [
        14, 20, 28, 28,           # A–D  (Trade/Legal Name wide)
        24, 13, 14, 12, 30, 13,   # E–J
        14, 22, 10, 10,           # K–N
        14, 12, 12, 12, 10, 12,   # O–T
        14, 15,                   # U–V  compliance
        18, 14,                   # W–X  metadata
        22, 13                    # Y–Z  orig inv (b2ba)
    ]

    ws_b2b.merge_cells(f"A1:{get_column_letter(len(B2B_COLS))}1")
    t = ws_b2b["A1"]
    t.value     = f"GSTR-2A | B2B Purchase Register | GSTIN: {gstin} | FY: {fy}"
    t.font      = Font(name="Arial", bold=True, size=11, color="1F3864")
    t.alignment = CENTER
    t.fill      = PatternFill("solid", fgColor="D6E4F0")
    ws_b2b.row_dimensions[1].height = 22

    # Sub-headers: shade the four zones differently
    ZONE_FILLS = {
        range(1,11):  PatternFill("solid", fgColor="1F3864"),  # identity  — dark blue (A-J)
        range(11,21): PatternFill("solid", fgColor="1F5C8B"),  # financial — mid blue  (K-T)
        range(21,25): PatternFill("solid", fgColor="2E75B6"),  # compliance+metadata    (U-X)
        range(25,27): PatternFill("solid", fgColor="375623"),  # amended orig ref—green (Y-Z)
    }
    for c, col in enumerate(B2B_COLS, 1):
        cell = ws_b2b.cell(row=2, column=c, value=col)
        cell.font      = HDR_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        for rng, fill in ZONE_FILLS.items():
            if c in rng:
                cell.fill = fill
                break
    ws_b2b.row_dimensions[2].height = 32

    SRCTYPES = {
        "E-Invoice":"E-Invoice","EINV":"E-Invoice",
        "ECOM":"E-Commerce","ISD":"ISD","IMPG":"Import",
    }

    b2b_rows = []

    # ── Build trdnm_map from GSTR-2B JSON files — LAST RESORT fallback ────────
    # Used only for suppliers not yet in gstin_names/name_lookup.json.
    # name_lookup (file_lookup merged above) is always checked first.
    # ROOT CAUSE: GSTR-2A offline download JSON does NOT contain 'trdnm'.
    #   b2b supplier entry has only: {ctin, cfs, cfs3b, fldtr1, flprdr1, inv[]}
    # GSTR-2B JSON DOES contain 'trdnm' in data.docdata.b2b[].
    # We scan ALL G2B files for this GSTIN (across all FYs) to build the map.
    # Priority: gstin_names/name_lookup.json (enrichment) > G2B trdnm > ""
    trdnm_map = {}   # {ctin_upper: trade_name_string}

    def _build_trdnm_from_g2b(base_dir, owner_gstin):
        owner_dir = os.path.join(base_dir, owner_gstin)
        if not os.path.isdir(owner_dir):
            return
        for fy_dir in os.listdir(owner_dir):
            fd = os.path.join(owner_dir, fy_dir)
            if not os.path.isdir(fd):
                continue
            for fname in os.listdir(fd):
                if not fname.endswith(".json"):
                    continue
                try:
                    with open(os.path.join(fd, fname), encoding="utf-8") as fh:
                        jd = json.load(fh)
                    inner   = jd.get("data", jd)
                    docdata = inner.get("docdata", {}) if isinstance(inner, dict) else {}
                    for sec_key in ("b2b", "cdnr"):
                        for sup in docdata.get(sec_key, []):
                            _ctin  = (sup.get("ctin",  "") or "").strip().upper()
                            _trdnm = (sup.get("trdnm", "") or "").strip()
                            if _ctin and _trdnm and _ctin not in trdnm_map:
                                trdnm_map[_ctin] = _trdnm
                except Exception:
                    pass

    _build_trdnm_from_g2b(G2B_DOWNLOAD_DIR, gstin)
    g2a_log(f"  trdnm_map: {len(trdnm_map)} supplier names loaded from G2B files")

    for month_display, data in months_data:
        for section in ["b2b", "b2ba"]:
            for supplier in data.get(section, []):
                ctin    = supplier.get("ctin", "")
                fldtr1  = supplier.get("fldtr1", "")    # "09-May-25"
                flprdr1 = supplier.get("flprdr1", "")   # "Apr-25"
                # ── ITC compliance status (supplier-level) ─────────────
                # cfs   : "Y"=GSTR-1 filed, "N"=not filed → ITC blocked Sec 16(2)(aa)
                # cfs3b : "Y"=GSTR-3B filed/tax paid, "N"=not paid → reversal risk Sec 16(2)(c)
                _cfs   = supplier.get("cfs",   "")
                _cfs3b = supplier.get("cfs3b", "")
                cfs_disp   = {"Y":"Filed ✓","N":"NOT FILED ✗"}.get(_cfs,   _cfs   or "—")
                cfs3b_disp = {"Y":"Filed ✓","N":"NOT FILED ✗"}.get(_cfs3b, _cfs3b or "—")
                is_b2ba = (section == "b2ba")
                for inv in supplier.get("inv", []):
                    inum       = inv.get("inum", "")
                    idt        = inv.get("idt",  "")
                    val        = inv.get("val",  0) or 0
                    inv_typ    = INV_TYPES.get(inv.get("inv_typ","R"), inv.get("inv_typ","R"))
                    srctyp     = SRCTYPES.get(inv.get("srctyp",""), inv.get("srctyp",""))
                    irn        = inv.get("irn", "")
                    irngendate = inv.get("irngendate", "")
                    pos        = pos_display(inv.get("pos", ""))
                    rchrg      = "Yes" if inv.get("rchrg","N") == "Y" else "No"
                    # ── Name lookup ───────────────────────────────────────────
                    # Priority: GSTIN Directory (enriched) → trdnm from JSON → ""
                    ninfo      = name_lookup.get(ctin.strip().upper(), {})
                    trd_nm     = ninfo.get("trade_name", "") or trdnm_map.get(ctin.strip().upper(), "") or ""
                    leg_nm     = ninfo.get("legal_name",  "")
                    for itm in inv.get("itms", [{}]):
                        # Handle both nested itm_det and flat item formats
                        d     = itm.get("itm_det", itm)
                        rt    = d.get("rt",    0) or 0
                        txval = d.get("txval", 0) or 0
                        iamt  = d.get("iamt",  0) or 0
                        camt  = d.get("camt",  0) or 0
                        samt  = d.get("samt",  0) or 0
                        csamt = d.get("csamt", 0) or 0
                        ttax  = iamt + camt + samt + csamt
                        oinum = inv.get("oinum","") if is_b2ba else ""
                        oidt  = inv.get("oidt", "") if is_b2ba else ""
                        b2b_rows.append([
                            # Zone 1 – Identity (indices 0-9)
                            month_display, ctin, trd_nm, leg_nm,
                            inum, idt, inv_typ, srctyp, irn, irngendate,
                            # Zone 2 – Financial (indices 10-19)
                            val, pos, rchrg, rt,
                            txval, iamt, camt, samt, csamt, ttax,
                            # Zone 3 – Compliance + Metadata (indices 20-25)
                            cfs_disp, cfs3b_disp, fldtr1, flprdr1,
                            oinum, oidt,
                            # Raw flags (not written to sheet — indices 26-28)
                            _cfs, _cfs3b, is_b2ba
                        ])

    # Track ITC risk totals for Summary sheet
    itc_risk_blocked   = {"inv":0,"txval":0,"tax":0}   # cfs=N (GSTR-1 not filed)
    itc_risk_reversal  = {"inv":0,"txval":0,"tax":0}   # cfs3b=N but cfs=Y

    # Build superseded invoice set (oinum from b2ba)
    superseded_inv_keys = set()
    for row in b2b_rows:
        if len(row) > 28 and row[28]:  # is_b2ba flag & oinum present
            oinum_val = row[24]        # orig invoice no (index unchanged)
            if oinum_val:
                superseded_inv_keys.add((row[1], oinum_val))  # (ctin, oinum)

    for i, row in enumerate(b2b_rows):
        r = i + 3
        # row[26]=_cfs raw, row[27]=_cfs3b raw, row[28]=is_b2ba — NOT written
        _raw_cfs   = row[26] if len(row) > 26 else ""
        _raw_cfs3b = row[27] if len(row) > 27 else ""
        _is_b2ba   = row[28] if len(row) > 28 else False
        for c, val in enumerate(row[:26], 1):   # write only first 26 cols (data only)
            ws_b2b.cell(row=r, column=c, value=val)

        # Determine compliance fill
        # Check if this is a superseded original (amended by b2ba)
        is_superseded = (row[1], row[4]) in superseded_inv_keys  # (ctin, inum)

        if is_superseded:
            row_fill = SUPERSED
        elif _is_b2ba:
            row_fill = AMEND_FILL
        elif _raw_cfs == "N":
            row_fill = RISK_RED
            itc_risk_blocked["inv"]   += 1
            itc_risk_blocked["txval"] += row[14] or 0
            itc_risk_blocked["tax"]   += row[19] or 0
        elif _raw_cfs3b == "N":
            row_fill = RISK_AMB
            itc_risk_reversal["inv"]   += 1
            itc_risk_reversal["txval"] += row[14] or 0
            itc_risk_reversal["tax"]   += row[19] or 0
        else:
            row_fill = ALT_FILL if (i % 2 == 1) else None

        for c in range(1, len(B2B_COLS) + 1):
            cell = ws_b2b.cell(row=r, column=c)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if row_fill:
                cell.fill = row_fill
            if isinstance(cell.value, (int, float)):
                cell.alignment = RIGHT
                cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    # Totals row
    # New column layout (with Trade Name + Legal Name inserted at C/D):
    # K=InvValue(11) O=TaxableVal(15) P=IGST(16) Q=CGST(17) R=SGST(18) S=Cess(19) T=TotalTax(20)
    tr = len(b2b_rows) + 3
    ws_b2b.cell(row=tr, column=1, value="TOTAL")
    ws_b2b.cell(row=tr, column=2, value=f"{len(b2b_rows)} line items")
    if b2b_rows:
        ds = 3; de = tr - 1
        for col_idx, col_letter in [(11,"K"),(15,"O"),(16,"P"),(17,"Q"),(18,"R"),(19,"S"),(20,"T")]:
            ws_b2b.cell(row=tr, column=col_idx,
                value=(f"=SUM({col_letter}{ds}:{col_letter}{de})" if de >= ds else 0))
    style_total_row(ws_b2b, tr, len(B2B_COLS))
    ws_b2b.row_dimensions[tr].height = 18
    set_col_widths(ws_b2b, B2B_WIDTHS)
    ws_b2b.freeze_panes = "E3"   # freeze A-D (Month, Supplier GSTIN, Trade Name, Legal Name)

    # ════════════════════════════════════════════════════════════
    # SHEET 2 — CREDIT / DEBIT NOTES
    # JSON key is 'cdn' (NOT 'cdnr') in GSTR-2A downloads
    # Amended notes use 'cdna' (NOT 'cdnra')
    #
    # Col map (26 cols):
    # A=Month  B=SupGSTIN  C=NoteType  D=NoteNo  E=NoteDate
    # F=NoteValue(signed)  G=InvType  H=SrcType  I=POS  J=RevCharge
    # K=TaxRate  L=TaxableVal(signed)  M=IGST(signed)  N=CGST(signed)
    # O=SGST(signed)  P=Cess(signed)  Q=TotalTax(signed)
    # R=IRN  S=IRNGenDate  T=Declared
    # U=GSTR1Status  V=GSTR3BStatus  W=FiledDate  X=Period
    # Y=OrigNoteNo  Z=OrigNoteDate
    #
    # Sign rule: Credit Note → negate val/txval/iamt/camt/samt/csamt/ttax
    #            Debit Note  → keep positive
    # ════════════════════════════════════════════════════════════
    # ══ E-COMMERCE (ECO) SHEET — Table 9: Inward Supplies via E-Commerce Operators
    # dtcancel: ECO registration cancellation date (supplier-group level)
    # No irn/srctyp/irngendate — ECO invoices not on e-invoice network
    ws_eco = wb.create_sheet("E-Commerce (ECO)")
    ECO_COLS = [
        "Month","ECO GSTIN","Trade Name","Legal Name",
        "Invoice No","Invoice Date","Inv Type",
        "Invoice Value","Place of Supply","Rev. Charge","Tax Rate %",
        "Taxable Value","IGST","CGST","SGST","Cess","Total Tax",
        "GSTR-1 Status","GSTR-3B Status","Filed Date","Period",
        "ECO Cancellation Date",
    ]
    ECO_WIDTHS=[14,22,28,28, 22,13,14, 14,22,10,10, 14,12,12,12,10,12, 14,15,18,14, 22]
    ws_eco.merge_cells(f"A1:{get_column_letter(len(ECO_COLS))}1")
    t_eco=ws_eco["A1"]
    t_eco.value=(f"GSTR-2A | E-Commerce (ECO) — Table 9 | GSTIN: {gstin} | FY: {fy}  |  "
                 "⚠ PINK rows = ECO registration CANCELLED — verify ITC eligibility")
    t_eco.font=Font(name="Arial",bold=True,size=11,color="7B1D00")
    t_eco.alignment=CENTER; t_eco.fill=PatternFill("solid",fgColor="FDECEA")
    ws_eco.row_dimensions[1].height=24
    ECO_ZONE={
        range(1,8):  PatternFill("solid",fgColor="4A235A"),
        range(8,18): PatternFill("solid",fgColor="6C3483"),
        range(18,22):PatternFill("solid",fgColor="884EA0"),
        range(22,23):PatternFill("solid",fgColor="C0392B"),
    }
    for c,col in enumerate(ECO_COLS,1):
        cell=ws_eco.cell(row=2,column=c,value=col)
        cell.font=HDR_FONT; cell.alignment=CENTER; cell.border=BORDER
        for rng,fill in ECO_ZONE.items():
            if c in rng: cell.fill=fill; break
    ws_eco.row_dimensions[2].height=32

    eco_rows=[]
    ECO_NORM=PatternFill("solid",fgColor="F5EEF8")
    ECO_ALT =PatternFill("solid",fgColor="EBE2F5")
    ECO_CXLD_FILL=PatternFill("solid",fgColor="FADBD8")

    for month_display,data in months_data:
        for supplier in data.get("eco",[]):
            ctin     = supplier.get("ctin","")
            _cfs     = supplier.get("cfs","")
            _cfs3b   = supplier.get("cfs3b","")
            fldtr1   = supplier.get("fldtr1","")
            flprdr1  = supplier.get("flprdr1","")
            dtcancel = supplier.get("dtcancel","")
            cfs_d    = {"Y":"Filed ✓","N":"NOT FILED ✗"}.get(_cfs,  _cfs  or "—")
            cfs3b_d  = {"Y":"Filed ✓","N":"NOT FILED ✗"}.get(_cfs3b,_cfs3b or "—")
            ninfo    = name_lookup.get(ctin.strip().upper(),{})
            trd_nm   = ninfo.get("trade_name","") or trdnm_map.get(ctin.strip().upper(),"") or ""
            leg_nm   = ninfo.get("legal_name","")
            for inv in supplier.get("inv",[]):
                inum  = inv.get("inum","")
                idt   = inv.get("idt","")
                val   = inv.get("val",0) or 0
                inv_t = {"R":"Regular","DE":"Deemed Export","SEWP":"SEZ w/ Pay",
                         "SEWOP":"SEZ wo/ Pay"}.get(inv.get("inv_typ","R"),inv.get("inv_typ","R"))
                pos   = pos_display(inv.get("pos",""))
                rchrg = "Yes" if inv.get("rchrg","N")=="Y" else "No"
                for itm in inv.get("itms",[{}]):
                    d     = itm.get("itm_det",itm)
                    rt    = d.get("rt",0) or 0
                    txval = d.get("txval",0) or 0
                    iamt  = d.get("iamt",0) or 0
                    camt  = d.get("camt",0) or 0
                    samt  = d.get("samt",0) or 0
                    csamt = d.get("csamt",0) or 0
                    ttax  = iamt+camt+samt+csamt
                    eco_rows.append([
                        month_display,ctin,trd_nm,leg_nm,
                        inum,idt,inv_t,
                        val,pos,rchrg,rt,
                        txval,iamt,camt,samt,csamt,ttax,
                        cfs_d,cfs3b_d,fldtr1,flprdr1,dtcancel,
                        _cfs,_cfs3b,bool(dtcancel),  # raw flags indices 22,23,24
                    ])

    WRITTEN_ECO=len(ECO_COLS)
    for i,row in enumerate(eco_rows):
        r=i+3
        is_cxld=bool(row[24]) if len(row)>24 else False
        fill=ECO_CXLD_FILL if is_cxld else (ECO_NORM if i%2==0 else ECO_ALT)
        for c,val in enumerate(row[:WRITTEN_ECO],1):
            cell=ws_eco.cell(row=r,column=c,value=val)
            cell.font=BODY_FONT; cell.border=BORDER; cell.fill=fill
            if isinstance(val,(int,float)):
                cell.alignment=RIGHT; cell.number_format=NUM_FMT
            else:
                cell.alignment=LEFT
    tr_eco=len(eco_rows)+3
    ws_eco.cell(row=tr_eco,column=1,value="TOTAL")
    ws_eco.cell(row=tr_eco,column=2,value=f"{len(eco_rows)} line items")
    if eco_rows:
        ds=3;de=tr_eco-1
        for ci,cl in [(8,"H"),(12,"L"),(13,"M"),(14,"N"),(15,"O"),(16,"P"),(17,"Q")]:
            ws_eco.cell(row=tr_eco,column=ci,
                value=(f"=SUM({cl}{ds}:{cl}{de})" if de>=ds else 0))
    style_total_row(ws_eco,tr_eco,WRITTEN_ECO)
    set_col_widths(ws_eco,ECO_WIDTHS)
    ws_eco.freeze_panes="E3"

    ws_cdn = wb.create_sheet("Credit-Debit Notes")
    CDN_COLS = [
        "Month","Supplier GSTIN","Trade Name","Legal Name",
        "Note Type","Note No","Note Date",
        "Note Value","Inv Type","Source Type","Place of Supply","Rev. Charge",
        "Tax Rate %","Taxable Value","IGST","CGST","SGST","Cess","Total Tax",
        "IRN","IRN Gen Date","Declared",
        "GSTR-1 Status","GSTR-3B Status","Supplier Filed Date","Supplier Period",
        "Orig Note No","Orig Note Date"
    ]
    CDN_WIDTHS = [
        14, 20, 28, 28,              # A–D  Month, SupGSTIN, TradeName, LegalName
        12, 22, 13,                  # E–G  NoteType, NoteNo, NoteDate
        13, 14, 12, 22, 10,          # H–L  NoteVal, InvType, SrcType, POS, RevCharge
        10, 14, 12, 12, 12, 10, 12,  # M–S  TaxRate, TaxVal, IGST, CGST, SGST, Cess, TotalTax
        30, 13, 10,                  # T–V  IRN, IRNGenDate, Declared
        14, 15, 18, 14,              # W–Z  GSTR1Sts, GSTR3BSts, FiledDate, Period
        22, 13                       # AA–AB OrigNoteNo, OrigNoteDate
    ]

    ws_cdn.merge_cells(f"A1:{get_column_letter(len(CDN_COLS))}1")
    t2 = ws_cdn["A1"]
    t2.value     = f"GSTR-2A | Credit / Debit Note Register | GSTIN: {gstin} | FY: {fy}"
    t2.font      = Font(name="Arial", bold=True, size=11, color="1F3864")
    t2.alignment = CENTER
    t2.fill      = PatternFill("solid", fgColor="D6E4F0")
    ws_cdn.row_dimensions[1].height = 22

    # Zone fills for header (all ranges shifted +2 for Trade Name + Legal Name)
    CDN_ZONE_FILLS = {
        range(1,8):   PatternFill("solid", fgColor="1F3864"),  # identity  A-G
        range(8,20):  PatternFill("solid", fgColor="1F5C8B"),  # financial H-S
        range(20,23): PatternFill("solid", fgColor="2E75B6"),  # e-invoice T-V
        range(23,29): PatternFill("solid", fgColor="375623"),  # compliance+amended W-AB
    }
    for c, col in enumerate(CDN_COLS, 1):
        cell = ws_cdn.cell(row=2, column=c, value=col)
        cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
        for rng, fill in CDN_ZONE_FILLS.items():
            if c in rng: cell.fill = fill; break
    ws_cdn.row_dimensions[2].height = 32

    NOTE_TYPES = {"C":"Credit Note","D":"Debit Note","R":"Refund Voucher"}

    cdn_rows = []
    for month_display, data in months_data:
        # 'cdn'  = current period CDN   (GSTR-2A uses 'cdn', NOT 'cdnr')
        # 'cdna' = amended CDN          (GSTR-2A uses 'cdna', NOT 'cdnra')
        for section in ["cdn", "cdna"]:
            is_amended = (section == "cdna")
            for supplier in data.get(section, []):
                ctin    = supplier.get("ctin", "")
                fldtr1  = supplier.get("fldtr1", "")
                flprdr1 = supplier.get("flprdr1", "")
                _cfs    = supplier.get("cfs",   "")
                _cfs3b  = supplier.get("cfs3b", "")
                cfs_disp   = {"Y":"Filed ✓","N":"NOT FILED ✗"}.get(_cfs,   _cfs   or "—")
                cfs3b_disp = {"Y":"Filed ✓","N":"NOT FILED ✗"}.get(_cfs3b, _cfs3b or "—")
                for nt in supplier.get("nt", []):
                    _ntty      = nt.get("ntty", "C")
                    ntty_disp  = NOTE_TYPES.get(_ntty, _ntty)
                    is_credit  = (_ntty == "C")          # Credit Note → negate
                    sign       = -1 if is_credit else 1
                    nt_num     = nt.get("nt_num", "")
                    nt_dt      = nt.get("nt_dt",  "")
                    raw_val    = nt.get("val", 0) or 0
                    val        = raw_val * sign
                    inv_typ    = INV_TYPES.get(nt.get("inv_typ","R"), nt.get("inv_typ","R"))
                    srctyp     = SRCTYPES.get(nt.get("srctyp",""), nt.get("srctyp",""))
                    irn        = nt.get("irn", "")
                    irngendate = nt.get("irngendate", "")
                    d_flag     = nt.get("d_flag", "")
                    declared   = {"Y":"Yes ✓","N":"No"}.get(d_flag, d_flag or "—")
                    pos        = pos_display(nt.get("pos", ""))
                    rchrg      = "Yes" if nt.get("rchrg","N") == "Y" else "No"
                    # Amended note original reference
                    orig_nt_num = nt.get("ont_num", "") if is_amended else ""
                    orig_nt_dt  = nt.get("ont_dt",  "") if is_amended else ""
                    # ── Name lookup (same as B2B — fallback to JSON trdnm) ───
                    ninfo_cdn  = name_lookup.get(ctin.strip().upper(), {})
                    trd_nm_cdn = ninfo_cdn.get("trade_name", "") or trdnm_map.get(ctin.strip().upper(), "") or ""
                    leg_nm_cdn = ninfo_cdn.get("legal_name",  "")
                    for itm in nt.get("itms", []):
                        d     = itm.get("itm_det", {})
                        rt    = d.get("rt",    0) or 0
                        txval = (d.get("txval", 0) or 0) * sign
                        iamt  = (d.get("iamt",  0) or 0) * sign
                        camt  = (d.get("camt",  0) or 0) * sign
                        samt  = (d.get("samt",  0) or 0) * sign
                        csamt = (d.get("csamt", 0) or 0) * sign
                        ttax  = iamt + camt + samt + csamt
                        cdn_rows.append([
                            # indices 0-6: identity
                            month_display, ctin, trd_nm_cdn, leg_nm_cdn,
                            ntty_disp, nt_num, nt_dt,
                            # indices 7-18: financial (signed)
                            val, inv_typ, srctyp, pos, rchrg,
                            rt, txval, iamt, camt, samt, csamt, ttax,
                            # indices 19-21: e-invoice
                            irn, irngendate, declared,
                            # indices 22-25: compliance
                            cfs_disp, cfs3b_disp, fldtr1, flprdr1,
                            # indices 26-27: amended original ref
                            orig_nt_num, orig_nt_dt,
                            # indices 28-31: raw flags (not written to sheet)
                            _cfs, _cfs3b, _ntty, is_amended
                        ])

    for i, row in enumerate(cdn_rows):
        r = i + 3
        _raw_cfs   = row[28] if len(row) > 28 else ""
        _raw_cfs3b = row[29] if len(row) > 29 else ""
        _raw_ntty  = row[30] if len(row) > 30 else "C"
        _is_amend  = row[31] if len(row) > 31 else False
        for c, val in enumerate(row[:28], 1):
            ws_cdn.cell(row=r, column=c, value=val)
        # Row fill: amended first, then compliance, then note type
        if _is_amend:
            row_fill = AMEND_FILL
        elif _raw_cfs == "N":
            row_fill = RISK_RED
        elif _raw_cfs3b == "N":
            row_fill = RISK_AMB
        elif _raw_ntty == "C":
            row_fill = CDN_CR
        else:
            row_fill = CDN_DR if _raw_ntty == "D" else (ALT_FILL if i%2==1 else None)
        for c in range(1, len(CDN_COLS) + 1):
            cell = ws_cdn.cell(row=r, column=c)
            cell.font = BODY_FONT; cell.border = BORDER
            if row_fill: cell.fill = row_fill
            if isinstance(cell.value, (int, float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    # CDN totals
    # Numeric: H=NoteVal(8) N=TaxableVal(14) O=IGST(15) P=CGST(16) Q=SGST(17) R=Cess(18) S=TotalTax(19)
    cn_count = sum(1 for r in cdn_rows if len(r)>30 and r[30]=="C")
    dn_count = sum(1 for r in cdn_rows if len(r)>30 and r[30]=="D")
    tr2 = len(cdn_rows) + 3
    ws_cdn.cell(row=tr2, column=1, value="TOTAL")
    ws_cdn.cell(row=tr2, column=2, value=f"{cn_count} Credit Notes | {dn_count} Debit Notes")
    if cdn_rows:
        ds = 3; de = tr2 - 1
        for col_idx, col_letter in [(8,"H"),(14,"N"),(15,"O"),(16,"P"),(17,"Q"),(18,"R"),(19,"S")]:
            ws_cdn.cell(row=tr2, column=col_idx,
                value=(f"=SUM({col_letter}{ds}:{col_letter}{de})" if de >= ds else 0))
    style_total_row(ws_cdn, tr2, len(CDN_COLS))
    ws_cdn.row_dimensions[tr2].height = 18
    set_col_widths(ws_cdn, CDN_WIDTHS)
    ws_cdn.freeze_panes = "E3"

    # ════════════════════════════════════════════════════════════
    # SHEET 3 — IMPORT OF GOODS (Bill of Entry)
    # ════════════════════════════════════════════════════════════
    ws_imp = wb.create_sheet("Import of Goods")
    IMP_COLS = [
        "Month","Bill of Entry No","BOE Date","BOE Value",
        "Port Code","Tax Rate %","Taxable Value","IGST","Cess","Total Tax"
    ]
    IMP_WIDTHS = [14, 18, 13, 14, 12, 10, 14, 14, 10, 12]

    ws_imp.merge_cells(f"A1:{get_column_letter(len(IMP_COLS))}1")
    t3 = ws_imp["A1"]
    t3.value     = f"GSTR-2A | Import of Goods (Bill of Entry) | GSTIN: {gstin} | FY: {fy}"
    t3.font      = Font(name="Arial", bold=True, size=11, color="1F3864")
    t3.alignment = CENTER
    t3.fill      = PatternFill("solid", fgColor="D6E4F0")
    ws_imp.row_dimensions[1].height = 22

    for c, col in enumerate(IMP_COLS, 1):
        ws_imp.cell(row=2, column=c, value=col)
    style_header_row(ws_imp, 2, len(IMP_COLS))
    ws_imp.row_dimensions[2].height = 28

    imp_rows = []
    for month_display, data in months_data:
        for imp in data.get("impg", []):
            boe_num  = imp.get("boe_num","")
            boe_dt   = imp.get("boe_dt","")
            boe_val  = imp.get("boe_val", 0) or 0
            port     = imp.get("port_code","")
            for itm in imp.get("itms", []):
                d = itm.get("itm_det", {})
                rt    = d.get("rt", 0) or 0
                txval = d.get("txval", 0) or 0
                iamt  = d.get("iamt", 0) or 0
                csamt = d.get("csamt", 0) or 0
                ttax  = iamt + csamt
                imp_rows.append([
                    month_display, boe_num, boe_dt, boe_val,
                    port, rt, txval, iamt, csamt, ttax
                ])

    for i, row in enumerate(imp_rows):
        r = i + 3
        for c, val in enumerate(row, 1):
            ws_imp.cell(row=r, column=c, value=val)
        style_body_row(ws_imp, r, len(IMP_COLS), alt=(i % 2 == 1))

    tr3 = len(imp_rows) + 3
    ws_imp.cell(row=tr3, column=1, value="TOTAL")
    if imp_rows:
        ds = 3; de = tr3 - 1
        for c, col_letter in [(4,"D"),(7,"G"),(8,"H"),(9,"I"),(10,"J")]:
            ws_imp.cell(row=tr3, column=c,
                value=(f"=SUM({col_letter}{ds}:{col_letter}{de})" if de >= ds else 0))
    style_total_row(ws_imp, tr3, len(IMP_COLS))
    set_col_widths(ws_imp, IMP_WIDTHS)
    ws_imp.freeze_panes = "A3"

    # ════════════════════════════════════════════════════════════
    # SHEET 4 — TCS (Tax Collected at Source by E-Commerce Operators)
    # ════════════════════════════════════════════════════════════
    ws_tcs = wb.create_sheet("TCS")
    TCS_COLS = [
        "Month","E-Commerce Operator Name","Operator GSTIN",
        "Gross Supply (₹)","Net Supply (₹)","Net Returns (₹)",
        "IGST","CGST","SGST","Total Tax","Flag","Period"
    ]
    TCS_WIDTHS = [14, 30, 20, 16, 16, 16, 12, 12, 12, 12, 8, 14]

    ws_tcs.merge_cells(f"A1:{get_column_letter(len(TCS_COLS))}1")
    t_tcs = ws_tcs["A1"]
    t_tcs.value     = f"GSTR-2A | TCS (Tax Collected at Source) | GSTIN: {gstin} | FY: {fy}"
    t_tcs.font      = Font(name="Arial", bold=True, size=11, color="1F3864")
    t_tcs.alignment = CENTER
    t_tcs.fill      = PatternFill("solid", fgColor="EDE7F6")
    ws_tcs.row_dimensions[1].height = 22
    for c, col in enumerate(TCS_COLS, 1):
        ws_tcs.cell(row=2, column=c, value=col)
    style_header_row(ws_tcs, 2, len(TCS_COLS))
    ws_tcs.row_dimensions[2].height = 28

    tcs_rows = []
    TCS_FLAGS = {"A": "Added", "D": "Deleted", "U": "Unchanged"}
    for month_display, data in months_data:
        for entry in data.get("tcs", []):
            tcs_rows.append([
                month_display,
                entry.get("sup_name", ""),
                entry.get("sup_gstin", ""),
                entry.get("grossSup", 0) or 0,
                entry.get("supR", 0) or 0,
                entry.get("retsupR", 0) or 0,
                entry.get("iamt", 0) or 0,
                entry.get("camt", 0) or 0,
                entry.get("samt", 0) or 0,
                (entry.get("iamt",0) or 0)+(entry.get("camt",0) or 0)+(entry.get("samt",0) or 0),
                TCS_FLAGS.get(entry.get("flag",""), entry.get("flag","")),
                entry.get("month", ""),
            ])
    for i, row in enumerate(tcs_rows):
        r = i + 3
        for c, val in enumerate(row, 1):
            ws_tcs.cell(row=r, column=c, value=val)
        for c in range(1, len(TCS_COLS)+1):
            cell = ws_tcs.cell(row=r, column=c)
            cell.font = BODY_FONT; cell.border = BORDER; cell.fill = TCS_FILL
            if isinstance(cell.value,(int,float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT
    tr_tcs = len(tcs_rows) + 3
    ws_tcs.cell(row=tr_tcs, column=1, value="TOTAL")
    if tcs_rows:
        ds=3; de=tr_tcs-1
        for col_idx,col_letter in [(4,"D"),(7,"G"),(8,"H"),(9,"I"),(10,"J")]:
            ws_tcs.cell(row=tr_tcs,column=col_idx,value=(f"=SUM({col_letter}{ds}:{col_letter}{de})" if de >= ds else 0))
    style_total_row(ws_tcs, tr_tcs, len(TCS_COLS))
    set_col_widths(ws_tcs, TCS_WIDTHS)
    ws_tcs.freeze_panes = "A3"

    # ════════════════════════════════════════════════════════════
    # SHEET 5 — TDS (Tax Deducted at Source — Section 51 CGST Act)
    # Deductors: Govt depts, PSUs, local authorities (GSTR-7 filers)
    # TDS credit goes to Electronic Cash Ledger (not ITC credit ledger)
    # Key: 'tds' in GSTR-2A JSON
    # Fields: tds_gstin, bflag, iamt, camt, samt, csamt,
    #         tot_amt (contract base), amtf (exempt amount), month
    # ════════════════════════════════════════════════════════════
    ws_tds = wb.create_sheet("TDS")
    # Column layout (confirmed against actual GSTR-2A JSON structure):
    #   gstin_deductor = party who deducted TDS (Deductor)
    #   deductor_name  = name of the deductor
    #   gstin_ded      = deductee GSTIN (our GSTIN / recipient)
    #   amt_ded        = amount on which TDS was deducted
    #   iamt/camt/samt = IGST/CGST/SGST deducted (1% each for intra-state)
    #   month          = return period (MMYYYY)
    TDS_COLS = [
        "Month",
        "Deductor Name",
        "Deductor GSTIN",
        "Deductee GSTIN",
        "TDS Base Amount (₹)",
        "IGST Deducted (₹)",
        "CGST Deducted (₹)",
        "SGST/UTGST Deducted (₹)",
        "Total TDS (₹)",
        "Period",
    ]
    # Col index map (1-based):
    # A=1 Month  B=2 Deductor Name  C=3 Deductor GSTIN  D=4 Deductee GSTIN
    # E=5 TDS Base  F=6 IGST  G=7 CGST  H=8 SGST  I=9 Total TDS  J=10 Period
    TDS_WIDTHS = [14, 36, 24, 24, 22, 18, 18, 22, 18, 14]

    ws_tds.merge_cells(f"A1:{get_column_letter(len(TDS_COLS))}1")
    t_tds = ws_tds["A1"]
    t_tds.value     = (f"GSTR-2A | TDS — Tax Deducted at Source (Sec 51 CGST Act) | "
                       f"GSTIN: {gstin} | FY: {fy}")
    t_tds.font      = Font(name="Arial", bold=True, size=11, color="1A237E")
    t_tds.alignment = CENTER
    t_tds.fill      = PatternFill("solid", fgColor="E8EAF6")
    ws_tds.row_dimensions[1].height = 22

    # Sub-header row 2 — colour zones
    TDS_ZONE = {
        range(1, 5):  PatternFill("solid", fgColor="283593"),  # identity cols
        range(5, 9):  PatternFill("solid", fgColor="0277BD"),  # tax amounts
        range(9, 10): PatternFill("solid", fgColor="01579B"),  # total TDS
        range(10,11): PatternFill("solid", fgColor="283593"),  # period
    }
    for c, col in enumerate(TDS_COLS, 1):
        cell = ws_tds.cell(row=2, column=c, value=col)
        cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
        for rng, fill in TDS_ZONE.items():
            if c in rng: cell.fill = fill; break
    ws_tds.row_dimensions[2].height = 32

    # Note row
    ws_tds.merge_cells(f"A3:{get_column_letter(len(TDS_COLS))}3")
    note_cell = ws_tds["A3"]
    note_cell.value = ("ℹ  TDS u/s 51 CGST Act: credited to Electronic Cash Ledger (not ITC ledger). "
                       "Rate: 1% CGST + 1% SGST (intra-state) or 2% IGST (inter-state) on taxable value. "
                       "Use TDS credit to offset GST cash payment liability.")
    note_cell.font      = Font(name="Arial", italic=True, size=9, color="1A237E")
    note_cell.fill      = PatternFill("solid", fgColor="EEF2FF")
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note_cell.border    = BORDER
    ws_tds.row_dimensions[3].height = 30

    tds_rows = []
    for month_display, data in months_data:
        for entry in data.get("tds", []):
            # Correct field names from actual GSTR-2A JSON
            deductor_name = entry.get("deductor_name", "") or ""
            gstin_deductor = entry.get("gstin_deductor", "") or ""   # who deducted
            gstin_ded      = entry.get("gstin_ded", "") or ""        # deductee (our GSTIN)
            amt_ded        = entry.get("amt_ded",   0) or 0          # TDS base amount
            iamt           = entry.get("iamt",      0) or 0          # IGST deducted
            camt           = entry.get("camt",      0) or 0          # CGST deducted
            samt           = entry.get("samt",      0) or 0          # SGST deducted
            total_tds      = iamt + camt + samt
            period         = entry.get("month",     "") or ""
            tds_rows.append([
                month_display,
                deductor_name,
                gstin_deductor,
                gstin_ded,
                amt_ded,
                iamt, camt, samt,
                total_tds,
                period,
            ])

    DATA_START_TDS = 4
    for i, row in enumerate(tds_rows):
        r = i + DATA_START_TDS
        for c, val in enumerate(row, 1):
            cell = ws_tds.cell(row=r, column=c, value=val)
            cell.font   = BODY_FONT
            cell.border = BORDER
            cell.fill   = TDS_FILL if i % 2 == 0 else ALT_FILL
            if isinstance(val, (int, float)):
                cell.alignment    = RIGHT
                cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    # Totals row
    # Numeric cols: E=5 TDS Base, F=6 IGST, G=7 CGST, H=8 SGST, I=9 Total
    tr_tds = len(tds_rows) + DATA_START_TDS
    ws_tds.cell(row=tr_tds, column=1, value="TOTAL")
    ws_tds.cell(row=tr_tds, column=2, value=f"{len(tds_rows)} deductor(s)")
    if tds_rows:
        ds = DATA_START_TDS; de = tr_tds - 1
        for col_idx, col_letter in [(5,"E"),(6,"F"),(7,"G"),(8,"H"),(9,"I")]:
            ws_tds.cell(row=tr_tds, column=col_idx,
                value=(f"=SUM({col_letter}{ds}:{col_letter}{de})" if de >= ds else 0))
    style_total_row(ws_tds, tr_tds, len(TDS_COLS))
    ws_tds.row_dimensions[tr_tds].height = 20
    set_col_widths(ws_tds, TDS_WIDTHS)
    ws_tds.freeze_panes = "C4"   # freeze Month+DeductorName, data from row 4

    # ════════════════════════════════════════════════════════════
    # SHEET 6 — CONSOLIDATED VIEW  (was SHEET 5)
    # All transactions in one sheet for reconciliation
    # Credit Notes shown with NEGATIVE financial values
    # Superseded originals (amended by b2ba/cdna) shown in ORANGE
    # ════════════════════════════════════════════════════════════
    ws_con = wb.create_sheet("Consolidated")

    CONSOL_COLS = [
        "Doc Type","Month","Supplier GSTIN","Trade Name","Legal Name",
        "Document No","Document Date",
        "Inv/Note Type","Source Type","IRN","IRN Gen Date",
        "Document Value","Place of Supply","Rev. Charge","Tax Rate %",
        "Taxable Value","IGST","CGST","SGST","Cess","Total Tax",
        "GSTR-1 Status","GSTR-3B Status","Filed Date","Period",
        "Declared","Orig Doc No","Orig Doc Date","Remarks"
    ]
    CONSOL_WIDTHS = [
        16, 14, 20, 28, 28,         # A–E  Doc Type,Month,GSTIN,TradeName,LegalName
        24, 13,                     # F–G  Doc No, Doc Date
        14, 12, 28, 13,             # H–K  Inv Type,Source,IRN,IRN Date
        14, 22, 10, 10,             # L–O  Doc Val,POS,RevChg,Rate
        14, 12, 12, 12, 10, 12,     # P–U  Tax cols
        14, 15, 18, 14,             # V–Y  Compliance
        10, 22, 13, 32              # Z–AC Declared,Orig,Remarks
    ]

    ws_con.merge_cells(f"A1:{get_column_letter(len(CONSOL_COLS))}1")
    t_con = ws_con["A1"]
    t_con.value     = (f"GSTR-2A | Consolidated Register | GSTIN: {gstin} | FY: {fy}  |  "
                       "⚠ ORANGE rows = Superseded by Amendment — DELETE before reconciliation")
    t_con.font      = Font(name="Arial", bold=True, size=11, color="7B3F00")
    t_con.alignment = CENTER
    t_con.fill      = PatternFill("solid", fgColor="FFF3CD")
    ws_con.row_dimensions[1].height = 24

    # Zone header for consolidated
    CON_ZONE = {
        range(1, 6):  PatternFill("solid", fgColor="1F3864"),   # Doc Type,Month,GSTIN,Trade,Legal
        range(6, 12): PatternFill("solid", fgColor="1F5C8B"),   # DocNo,DocDate,InvType,Src,IRN,IRNDate
        range(12, 22):PatternFill("solid", fgColor="1F5C8B"),   # Value,POS,RevChg,Rate,Tax cols
        range(22, 26):PatternFill("solid", fgColor="2E75B6"),   # Compliance
        range(26, 30):PatternFill("solid", fgColor="375623"),   # Declared,Orig,Remarks
    }
    for c, col in enumerate(CONSOL_COLS, 1):
        cell = ws_con.cell(row=2, column=c, value=col)
        cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
        for rng, fill in CON_ZONE.items():
            if c in rng: cell.fill = fill; break
    ws_con.row_dimensions[2].height = 32

    # Build superseded CDN note set
    superseded_cdn_keys = set()
    for row in cdn_rows:
        if len(row) > 31 and row[31]:       # is_amended flag (idx 31 after +2 shift)
            orig_nt = row[26]               # orig_nt_num    (idx 26 after +2 shift)
            if orig_nt:
                superseded_cdn_keys.add((row[1], orig_nt))   # (ctin, orig_nt_num)

    con_rows_data = []   # (row_values, fill)

    # ── B2B rows ─────────────────────────────────────────────
    for row in b2b_rows:
        _is_b2ba   = row[28] if len(row)>28 else False
        _raw_cfs   = row[26] if len(row)>26 else ""
        _raw_cfs3b = row[27] if len(row)>27 else ""
        is_superseded = (row[1], row[4]) in superseded_inv_keys
        if is_superseded:
            remarks = "⚠ SUPERSEDED — DELETE (amended by b2ba)"
            fill    = SUPERSED
        elif _is_b2ba:
            remarks = "✏ AMENDED INVOICE (replaces original)"
            fill    = AMEND_FILL
        elif _raw_cfs == "N":
            remarks = "❌ ITC BLOCKED — GSTR-1 not filed"
            fill    = RISK_RED
        elif _raw_cfs3b == "N":
            remarks = "⚠ REVERSAL RISK — GSTR-3B not filed"
            fill    = RISK_AMB
        else:
            fill    = None
            remarks = ""
        con_rows_data.append(([
            "Invoice",           # Doc Type
            row[0],              # Month
            row[1],              # Supplier GSTIN
            row[2],              # Trade Name       ← NOW INCLUDED
            row[3],              # Legal Name       ← NOW INCLUDED
            row[4],              # Invoice No       (idx shifted +2)
            row[5],              # Invoice Date     (idx shifted +2)
            row[6],              # Inv Type         (idx shifted +2)
            row[7],              # Source Type      (idx shifted +2)
            row[8],              # IRN              (idx shifted +2)
            row[9],              # IRN Gen Date     (idx shifted +2)
            row[10],             # Invoice Value    (idx shifted +2)
            row[11],             # POS              (idx shifted +2)
            row[12],             # Rev Charge       (idx shifted +2)
            row[13],             # Tax Rate         (idx shifted +2)
            row[14],             # Taxable Value    (idx shifted +2)
            row[15],             # IGST             (idx shifted +2)
            row[16],             # CGST             (idx shifted +2)
            row[17],             # SGST             (idx shifted +2)
            row[18],             # Cess             (idx shifted +2)
            row[19],             # Total Tax        (idx shifted +2)
            row[20],             # GSTR-1 Status    (idx shifted +2)
            row[21],             # GSTR-3B Status   (idx shifted +2)
            row[22],             # Filed Date       (idx shifted +2)
            row[23],             # Period           (idx shifted +2)
            "",                  # Declared (N/A for invoices)
            row[24],             # Orig Inv No (b2ba only) (idx unchanged)
            row[25],             # Orig Inv Date (b2ba only) (idx unchanged)
            remarks,
        ], fill))

    # ── CDN rows ─────────────────────────────────────────────
    for row in cdn_rows:
        _raw_cfs   = row[28] if len(row)>28 else ""
        _raw_cfs3b = row[29] if len(row)>29 else ""
        _raw_ntty  = row[30] if len(row)>30 else "C"
        _is_amend  = row[31] if len(row)>31 else False
        is_superseded_cdn = (row[1], row[5]) in superseded_cdn_keys  # (ctin, nt_num idx 5)
        doc_type = {"C":"Credit Note","D":"Debit Note","R":"Refund Voucher"}.get(_raw_ntty, _raw_ntty)
        if is_superseded_cdn:
            remarks = "⚠ SUPERSEDED — DELETE (amended by cdna)"
            fill    = SUPERSED
        elif _is_amend:
            remarks = "✏ AMENDED NOTE (replaces original)"
            fill    = AMEND_FILL
        elif _raw_cfs == "N":
            remarks = "❌ ITC BLOCKED — GSTR-1 not filed"
            fill    = RISK_RED
        elif _raw_cfs3b == "N":
            remarks = "⚠ REVERSAL RISK — GSTR-3B not filed"
            fill    = RISK_AMB
        elif _raw_ntty == "C":
            fill    = CDN_CR
            remarks = "Credit Note — values are NEGATIVE"
        else:
            fill    = CDN_DR
            remarks = "Debit Note — values are POSITIVE"
        con_rows_data.append(([
            doc_type,            # Doc Type
            row[0],              # Month
            row[1],              # Supplier GSTIN
            row[2],              # Trade Name        ← NOW INCLUDED
            row[3],              # Legal Name        ← NOW INCLUDED
            row[5],              # Note No          (idx 5 after +2 shift)
            row[6],              # Note Date         (idx 6)
            row[8],              # Inv Type          (idx 8)
            row[9],              # Source Type       (idx 9)
            row[19],             # IRN               (idx 19)
            row[20],             # IRN Gen Date      (idx 20)
            row[7],              # Note Value (signed)(idx 7)
            row[10],             # POS               (idx 10)
            row[11],             # Rev Charge        (idx 11)
            row[12],             # Tax Rate          (idx 12)
            row[13],             # Taxable Value (signed)(idx 13)
            row[14],             # IGST (signed)     (idx 14)
            row[15],             # CGST (signed)     (idx 15)
            row[16],             # SGST (signed)     (idx 16)
            row[17],             # Cess (signed)     (idx 17)
            row[18],             # Total Tax (signed)(idx 18)
            row[22],             # GSTR-1 Status     (idx 22)
            row[23],             # GSTR-3B Status    (idx 23)
            row[24],             # Filed Date        (idx 24)
            row[25],             # Period            (idx 25)
            row[21],             # Declared (d_flag) (idx 21)
            row[26],             # Orig Note No (cdna only)(idx 26)
            row[27],             # Orig Note Date (cdna only)(idx 27)
            remarks,
        ], fill))

    # ── E-Commerce (ECO) rows ─────────────────────────────────────────────────
    ECO_CXLD = PatternFill("solid", fgColor="FADBD8")
    for row in eco_rows:
        if len(row) < 22: continue
        is_cancelled = bool(row[24]) if len(row) > 24 else False
        fill_eco = ECO_CXLD if is_cancelled else None
        cancel_dt = row[21] if len(row) > 21 else ""
        remarks = (f"⚠ ECO CANCELLED {cancel_dt} — verify ITC eligibility"
                   if is_cancelled else "")
        con_rows_data.append(([
            "Invoice (ECO)",    # Doc Type
            row[0],             # Month
            row[1],             # ECO GSTIN
            row[2],             # Trade Name
            row[3],             # Legal Name
            row[4],             # Invoice No
            row[5],             # Invoice Date
            row[6],             # Inv Type
            "",                 # Source Type  (N/A for ECO)
            "",                 # IRN          (N/A for ECO)
            "",                 # IRN Gen Date (N/A for ECO)
            row[7],             # Invoice Value
            row[8],             # Place of Supply
            row[9],             # Rev. Charge
            row[10],            # Tax Rate %
            row[11],            # Taxable Value
            row[12],            # IGST
            row[13],            # CGST
            row[14],            # SGST
            row[15],            # Cess
            row[16],            # Total Tax
            row[17],            # GSTR-1 Status
            row[18],            # GSTR-3B Status
            row[19],            # Filed Date
            row[20],            # Period
            "",                 # Declared
            "",                 # Orig Doc No
            cancel_dt,          # ECO Cancellation Date → Orig Doc Date col
            remarks,
        ], fill_eco))

    # ── Import of Goods rows ──────────────────────────────────
    for row in imp_rows:
        con_rows_data.append(([
            "Import-BOE",        # Doc Type
            row[0],              # Month
            "",                  # Supplier GSTIN (N/A for BOE)
            "",                  # Trade Name     (N/A for BOE)
            "",                  # Legal Name     (N/A for BOE)
            row[1],              # BOE No
            row[2],              # BOE Date
            "Import",            # Inv Type
            "",                  # Source Type
            "",                  # IRN
            "",                  # IRN Gen Date
            row[3],              # BOE Value
            "",                  # POS
            "",                  # Rev Charge
            row[5],              # Tax Rate
            row[6],              # Taxable Value
            row[7],              # IGST
            0,                   # CGST
            0,                   # SGST
            row[8],              # Cess
            row[9],              # Total Tax
            "", "", "", "",      # compliance cols (N/A)
            "", "", "",          # declared, orig, remarks
            "",
        ], IMP_FILL))

    # Write consolidated rows
    for i, (vals, fill) in enumerate(con_rows_data):
        r = i + 3
        for c, val in enumerate(vals, 1):
            ws_con.cell(row=r, column=c, value=val)
        for c in range(1, len(CONSOL_COLS)+1):
            cell = ws_con.cell(row=r, column=c)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if fill: cell.fill = fill
            if isinstance(cell.value,(int,float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    # Consolidated totals
    con_tr = len(con_rows_data) + 3
    ws_con.cell(row=con_tr, column=1, value="TOTAL")
    ws_con.cell(row=con_tr, column=2, value=f"{len(con_rows_data)} documents")
    if con_rows_data:
        ds=3; de=con_tr-1
        # Col indices after +2 shift for Trade Name + Legal Name:
        # Doc Value=12(L), Txval=16(P), IGST=17(Q), CGST=18(R), SGST=19(S), Cess=20(T), Total=21(U)
        for col_idx,col_letter in [(12,"L"),(16,"P"),(17,"Q"),(18,"R"),(19,"S"),(20,"T"),(21,"U")]:
            ws_con.cell(row=con_tr,column=col_idx,
                value=(f"=SUM({col_letter}{ds}:{col_letter}{de})" if de >= ds else 0))
    style_total_row(ws_con, con_tr, len(CONSOL_COLS))
    ws_con.row_dimensions[con_tr].height = 18

    # Colour legend below totals
    leg_r = con_tr + 2
    legend_items = [
        ("Invoice (B2B/B2BA)", "FFFFFF"),
        ("Credit Note (negative values)", "FFE8E8"),
        ("Debit Note (positive values)", "E8F5E9"),
        ("Import BOE", "EDF2FB"),
        ("Amended document (b2ba/cdna)", "FFE0B2"),
        ("⚠ SUPERSEDED — DELETE this row", "FF6600"),
        ("ITC Blocked (GSTR-1 not filed)", "FFD7D7"),
        ("Reversal Risk (GSTR-3B not filed)", "FFF3CD"),
    ]
    ws_con.merge_cells(f"A{leg_r}:B{leg_r}")
    ws_con.cell(row=leg_r,column=1,value="COLOUR LEGEND").font = Font(name="Arial",bold=True,size=9)
    for idx,(label,color) in enumerate(legend_items):
        lr = leg_r + idx + 1
        cell = ws_con.cell(row=lr, column=1, value=label)
        cell.fill   = PatternFill("solid", fgColor=color)
        cell.font   = Font(name="Arial", size=9)
        cell.border = BORDER
        cell.alignment = LEFT

    set_col_widths(ws_con, CONSOL_WIDTHS)
    ws_con.freeze_panes = "D3"

    # ════════════════════════════════════════════════════════════
    # SHEET 7 — SUMMARY (month-wise)
    # ════════════════════════════════════════════════════════════
    # ════════════════════════════════════════════════════════════════════════
    # Build summary data (month → metrics) — same as before
    # ════════════════════════════════════════════════════════════════════════
    from collections import defaultdict

    summary = defaultdict(lambda: {
        "b2b_inv":0,"b2b_txval":0.0,"b2b_igst":0.0,"b2b_cgst":0.0,
        "b2b_sgst":0.0,"b2b_cess":0.0,
        "eco_inv":0,"eco_txval":0.0,"eco_igst":0.0,"eco_cgst":0.0,
        "eco_sgst":0.0,"eco_cess":0.0,
        "cdn_nt":0,"cdn_txval":0.0,"cdn_igst":0.0,"cdn_cgst":0.0,
        "cdn_sgst":0.0,"cdn_cess":0.0,
        "imp_cnt":0,"imp_val":0.0,"imp_igst":0.0,"imp_cess":0.0,
    })

    for row in b2b_rows:
        m = row[0]
        summary[m]["b2b_inv"]   += 1
        summary[m]["b2b_txval"] += row[14] or 0
        summary[m]["b2b_igst"]  += row[15] or 0
        summary[m]["b2b_cgst"]  += row[16] or 0
        summary[m]["b2b_sgst"]  += row[17] or 0
        summary[m]["b2b_cess"]  += row[18] or 0

    for row in eco_rows:
        if len(row) > 15:
            m = row[0]
            summary[m]["eco_inv"]   += 1
            summary[m]["eco_txval"] += row[11] or 0
            summary[m]["eco_igst"]  += row[12] or 0
            summary[m]["eco_cgst"]  += row[13] or 0
            summary[m]["eco_sgst"]  += row[14] or 0
            summary[m]["eco_cess"]  += row[15] or 0

    for row in cdn_rows:
        m = row[0]
        summary[m]["cdn_nt"]    += 1
        summary[m]["cdn_txval"] += row[13] or 0
        summary[m]["cdn_igst"]  += row[14] or 0
        summary[m]["cdn_cgst"]  += row[15] or 0
        summary[m]["cdn_sgst"]  += row[16] or 0
        summary[m]["cdn_cess"]  += row[17] or 0

    for row in imp_rows:
        m = row[0]
        summary[m]["imp_cnt"]  += 1
        summary[m]["imp_val"]  += row[3] or 0
        summary[m]["imp_igst"] += row[7] or 0
        summary[m]["imp_cess"] += row[8] or 0

    # ── Sort months Apr → Mar ────────────────────────────────────────────────
    _G2A_MONTH_NAMES = ["April","May","June","July","August","September",
                        "October","November","December","January","February","March"]
    _G2A_MON_ABBR    = {n: n[:3] for n in _G2A_MONTH_NAMES}
    _FY_START = int(fy.split("-")[0])

    def _g2a_month_key(mdisp):
        parts = mdisp.split()
        try: return _G2A_MONTH_NAMES.index(parts[0])
        except (ValueError, IndexError): return 99

    sorted_months = sorted(summary.keys(), key=_g2a_month_key)

    # ════════════════════════════════════════════════════════════════════════
    # SUMMARY SHEET — Column-per-month layout (Apr → Mar → Total)
    # Row A = Particular, Cols B onwards = months, last col = Total
    # ════════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("Summary", 0)

    # ── Month column headers (abbreviated, e.g. "Apr '25") ───────────────────
    def _mon_col_label(mdisp):
        parts = mdisp.split()
        if len(parts) == 2:
            return f"{parts[0][:3]} '{parts[1][-2:]}"
        return mdisp[:8]

    mon_labels = [_mon_col_label(m) for m in sorted_months]
    n_months   = len(mon_labels)
    # Cols: A=Particular (col 1), B..N=months, last=Total
    PART_COL   = 1
    MON_START  = 2
    MON_END    = MON_START + n_months - 1
    TOT_COL    = MON_END + 1
    NCOLS      = TOT_COL

    # ── Row 1: Title ─────────────────────────────────────────────────────────
    ws_sum.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    tc = ws_sum["A1"]
    tc.value = f"GSTR-2A | Month-wise Summary | GSTIN: {gstin} | FY: {fy}"
    tc.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    tc.alignment = CENTER
    tc.fill      = HDR_FILL
    ws_sum.row_dimensions[1].height = 26

    # ── Row 2: Column headers ─────────────────────────────────────────────────
    ws_sum.cell(row=2, column=PART_COL, value="Particulars")
    for ci, lbl in enumerate(mon_labels, MON_START):
        ws_sum.cell(row=2, column=ci, value=lbl)
    ws_sum.cell(row=2, column=TOT_COL, value="Total")
    style_header_row(ws_sum, 2, NCOLS,
                     fill=HDR_FILL, font=Font(name="Arial", bold=True, color="FFFFFF", size=9))
    ws_sum.row_dimensions[2].height = 30

    # ── Helper: write one data row ────────────────────────────────────────────
    _sum_row = 3

    def _sum_write(label, extractor, alt=False, bold=False,
                   sect_fill=None, num_fmt=None, is_count=False):
        """Write a metric row: label | Apr | May | ... | Total"""
        nonlocal _sum_row
        r = _sum_row
        lc = ws_sum.cell(row=r, column=PART_COL, value=label)
        lc.font      = Font(name="Arial", bold=bold, size=9)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = BORDER
        if sect_fill:
            lc.fill = sect_fill
        elif alt:
            lc.fill = ALT_FILL

        total = 0.0
        for ci, m in enumerate(sorted_months, MON_START):
            v = extractor(m)
            cell = ws_sum.cell(row=r, column=ci, value=v if v else 0)
            cell.font   = Font(name="Arial", size=9)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="right", vertical="center")
            if sect_fill:
                cell.fill = sect_fill
            elif alt:
                cell.fill = ALT_FILL
            if is_count:
                cell.number_format = '#,##0'
            else:
                cell.number_format = '#,##0.00'
                if v: total += v

        # Total column — use SUM formula for numeric, sum for count
        tc = ws_sum.cell(row=r, column=TOT_COL)
        if n_months > 0:
            st_ltr = get_column_letter(MON_START)
            en_ltr = get_column_letter(MON_END)
            tc.value = f"=SUM({st_ltr}{r}:{en_ltr}{r})"
        else:
            tc.value = 0
        tc.font   = Font(name="Arial", bold=True, size=9)
        tc.border = BORDER
        tc.fill   = TOT_FILL
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.number_format = '#,##0' if is_count else '#,##0.00'
        ws_sum.row_dimensions[r].height = 16
        _sum_row += 1

    def _sum_section(label, fill_color):
        """Write a section header row."""
        nonlocal _sum_row
        r = _sum_row
        ws_sum.merge_cells(f"A{r}:{get_column_letter(NCOLS)}{r}")
        c = ws_sum.cell(row=r, column=1, value=label)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = BORDER
        ws_sum.row_dimensions[r].height = 18
        _sum_row += 1

    # ── SECTION 1: B2B PURCHASES ─────────────────────────────────────────────
    _sum_section("B2B Purchases (Inward Supplies from Registered Suppliers)", "1F3864")
    _sum_write("  Invoices (Count)", lambda m: summary[m]["b2b_inv"],   is_count=True)
    _sum_write("  Taxable Value",    lambda m: summary[m]["b2b_txval"])
    _sum_write("  IGST",             lambda m: summary[m]["b2b_igst"])
    _sum_write("  CGST",             lambda m: summary[m]["b2b_cgst"])
    _sum_write("  SGST",             lambda m: summary[m]["b2b_sgst"])
    _sum_write("  Cess",             lambda m: summary[m]["b2b_cess"])
    _sum_write("  Total Tax",
               lambda m: (summary[m]["b2b_igst"]+summary[m]["b2b_cgst"]
                          +summary[m]["b2b_sgst"]+summary[m]["b2b_cess"]),
               bold=True)

    # ── SECTION 1b: E-COMMERCE (ECO) PURCHASES ──────────────────────────────
    _sum_section("E-Commerce (ECO) — Inward Supplies via E-Commerce Operators","4A235A")
    _sum_write("  Line Items (Count)",lambda m: summary[m]["eco_inv"],  is_count=True)
    _sum_write("  Taxable Value",     lambda m: summary[m]["eco_txval"])
    _sum_write("  IGST",              lambda m: summary[m]["eco_igst"])
    _sum_write("  CGST",              lambda m: summary[m]["eco_cgst"])
    _sum_write("  SGST",              lambda m: summary[m]["eco_sgst"])
    _sum_write("  Cess",              lambda m: summary[m]["eco_cess"])
    _sum_write("  Total Tax",
               lambda m:(summary[m]["eco_igst"]+summary[m]["eco_cgst"]
                         +summary[m]["eco_sgst"]+summary[m]["eco_cess"]),bold=True)

    # ── SECTION 2: CREDIT/DEBIT NOTES ────────────────────────────────────────
    _sum_section("Credit / Debit Notes (Inward from Registered Suppliers)", "2E75B6")
    _sum_write("  Notes (Count)",    lambda m: summary[m]["cdn_nt"],    is_count=True)
    _sum_write("  Taxable Value",    lambda m: summary[m]["cdn_txval"])
    _sum_write("  IGST",             lambda m: summary[m]["cdn_igst"])
    _sum_write("  CGST",             lambda m: summary[m]["cdn_cgst"])
    _sum_write("  SGST",             lambda m: summary[m]["cdn_sgst"])
    _sum_write("  Cess",             lambda m: summary[m]["cdn_cess"])
    _sum_write("  Total Tax",
               lambda m: (summary[m]["cdn_igst"]+summary[m]["cdn_cgst"]
                          +summary[m]["cdn_sgst"]+summary[m]["cdn_cess"]),
               bold=True)

    # ── SECTION 3: IMPORT OF GOODS (BOE) ─────────────────────────────────────
    _sum_section("Import of Goods — Bill of Entry (IMPG)", "375623")
    _sum_write("  BOE Count",        lambda m: summary[m]["imp_cnt"],   is_count=True)
    _sum_write("  BOE Value",        lambda m: summary[m]["imp_val"])
    _sum_write("  IGST",             lambda m: summary[m]["imp_igst"])
    _sum_write("  Cess",             lambda m: summary[m]["imp_cess"])
    _sum_write("  Total Tax",
               lambda m: summary[m]["imp_igst"]+summary[m]["imp_cess"],
               bold=True)

    # ── Column widths ─────────────────────────────────────────────────────────
    ws_sum.column_dimensions["A"].width = 44
    for ci in range(MON_START, TOT_COL + 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 13
    ws_sum.freeze_panes = f"B3"

    # ── ITC Risk Section (below summary, same as before) ─────────────────────
    risk_start = _sum_row + 1

    ws_sum.merge_cells(f"A{risk_start}:F{risk_start}")
    rh = ws_sum.cell(row=risk_start, column=1)
    rh.value     = "⚠  ITC COMPLIANCE RISK SUMMARY"
    rh.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    rh.fill      = PatternFill("solid", fgColor="C00000")
    rh.alignment = CENTER
    ws_sum.row_dimensions[risk_start].height = 22

    risk_hdr = risk_start + 1
    for c, lbl in enumerate(["Risk Category","Description","Invoices","Taxable Value","Total Tax","Action Required"],1):
        cell = ws_sum.cell(row=risk_hdr, column=c, value=lbl)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor="9B0000")
        cell.alignment = CENTER
        cell.border    = BORDER
    ws_sum.row_dimensions[risk_hdr].height = 22
    ws_sum.column_dimensions["A"].width = 44
    ws_sum.column_dimensions["B"].width = 45
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 16
    ws_sum.column_dimensions["E"].width = 14
    ws_sum.column_dimensions["F"].width = 40

    r1 = risk_start + 2
    risk_data = [
        ("ITC BLOCKED",
         "Supplier GSTR-1 NOT filed — ITC unavailable u/s 16(2)(aa)",
         itc_risk_blocked["inv"], itc_risk_blocked["txval"], itc_risk_blocked["tax"],
         "Do NOT claim ITC; wait for supplier to file GSTR-1",
         RISK_RED, Font(name="Arial", bold=True, size=9, color="7B0000")),
        ("REVERSAL RISK",
         "Supplier GSTR-3B not filed — tax unpaid; reversal risk u/s 16(2)(c)",
         itc_risk_reversal["inv"], itc_risk_reversal["txval"], itc_risk_reversal["tax"],
         "Verify tax payment before claiming ITC; monitor GSTR-2B",
         RISK_AMB, Font(name="Arial", bold=True, size=9, color="7B5200")),
        ("SAFE — Both Filed",
         "Supplier filed both GSTR-1 and GSTR-3B — ITC eligible",
         len(b2b_rows) - itc_risk_blocked["inv"] - itc_risk_reversal["inv"],
         None, None,
         "ITC claimable — verify GSTR-2B auto-population",
         PatternFill("solid", fgColor="E2EFDA"), Font(name="Arial", bold=True, size=9, color="375623")),
    ]
    for _idx, (cat, desc, inv_cnt, txval, tax, action, fill, font) in enumerate(risk_data):
        r = r1 + _idx
        vals = [cat, desc, inv_cnt, txval, tax, action]
        for c, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=r, column=c, value=v)
            cell.fill   = fill
            cell.font   = font
            cell.border = BORDER
            if isinstance(v, (int, float)) and v is not None:
                cell.alignment   = RIGHT
                cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT
        ws_sum.row_dimensions[r].height = 18

    leg_r = r1 + len(risk_data) + 1
    ws_sum.merge_cells(f"A{leg_r}:F{leg_r}")
    leg = ws_sum.cell(row=leg_r, column=1,
        value="LEGEND:  PINK rows in B2B sheet = ITC Blocked (GSTR-1 not filed)  |  "
              "AMBER rows = Reversal Risk (GSTR-3B not filed)  |  "
              "Refer columns U (GSTR-1 Status) and V (GSTR-3B Status) in B2B sheet")
    leg.font      = Font(name="Arial", italic=True, size=8, color="595959")
    leg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    leg.fill      = PatternFill("solid", fgColor="F2F2F2")
    ws_sum.row_dimensions[leg_r].height = 30

    # ── Save ──────────────────────────────────────────────────
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


# ── GSTR-2A Flask Routes ───────────────────────────────────────
@app.route("/g2a/start", methods=["POST"])
def g2a_start():
    data     = request.get_json(force=True) or {}
    gstin    = (data.get("gstin")    or "").strip().upper()
    fy       = (data.get("fy")       or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()

    if not all([gstin, fy, username, password]):
        return jsonify({"error":"gstin, fy, username and password required"}), 400

    specific_month = (data.get("specific_month") or "").strip()
    # force=true bypasses the 10-day cache check and re-downloads everything
    force = bool(data.get("force", False))

    # Atomic check-and-claim under one lock acquisition — see the
    # detailed comment on gstin_dir_enrich_start()/combined_start() for
    # why a separate check-then-later-claim leaves a race window.
    with g2a_lock:
        if g2a_state["status"] in ("running","waiting_captcha","waiting_otp","downloading"):
            return jsonify({"error":"Already running — reset first"}), 409
        g2a_state.update({
            "status":"running","log":[],"error":None,
            "captcha_image":None,"captcha_answer":None,"otp_answer":None,
            "progress":0,"current_month":None,"total_months":0,
            "done_months":0,"files":[],"gstin":gstin,"fy":fy,
            "specific_month": specific_month or None,
        })

    threading.Thread(target=g2a_worker,
                     args=(gstin,fy,username,password,specific_month,force),
                     daemon=True).start()
    return jsonify({"ok":True, "specific_month": specific_month or None,
                    "force": force})


@app.route("/g2a/state")
def g2a_get_state():
    with g2a_lock:
        return jsonify({
            "status":         g2a_state["status"],
            "log":            g2a_state["log"][-40:],
            "error":          g2a_state["error"],
            "progress":       g2a_state["progress"],
            "current_month":  g2a_state["current_month"],
            "total_months":   g2a_state["total_months"],
            "done_months":    g2a_state["done_months"],
            "files":          g2a_state["files"],
            "gstin":          g2a_state["gstin"],
            "fy":             g2a_state["fy"],
            "specific_month": g2a_state.get("specific_month"),
            "has_captcha":    g2a_state["captcha_image"] is not None,
        })


@app.route("/g2a/captcha_image")
def g2a_captcha_image():
    with g2a_lock:
        img = g2a_state.get("captcha_image")
    if not img:
        return jsonify({"error":"no captcha available"}), 404
    return jsonify({"image": img})


@app.route("/g2a/submit_captcha", methods=["POST"])
def g2a_submit_captcha():
    data   = request.get_json(force=True) or {}
    answer = (data.get("captcha") or "").strip()
    if not answer:
        return jsonify({"error":"captcha required"}), 400
    with g2a_lock:
        g2a_state["captcha_answer"] = answer
        g2a_state["captcha_image"]  = None
    return jsonify({"ok":True})


@app.route("/g2a/submit_otp", methods=["POST"])
def g2a_submit_otp():
    data = request.get_json(force=True) or {}
    otp  = (data.get("otp") or "").strip()
    if not otp:
        return jsonify({"error":"otp required"}), 400
    with g2a_lock:
        g2a_state["otp_answer"] = otp
    return jsonify({"ok":True})


@app.route("/g2a/export_excel", methods=["GET","POST"])
def g2a_export_excel():
    """Convert all downloaded GSTR-2A JSON files → single Excel, return for download."""
    try:
        if request.method == "POST":
            body  = request.get_json(force=True) or {}
            gstin = (body.get("gstin","") or "").strip().upper()
            fy    = (body.get("fy","")    or "").strip()
            # name_lookup: [{gstin, trade_name, legal_name}, ...]
            raw_lookup = body.get("name_lookup", []) or []
        else:
            gstin      = request.args.get("gstin","").strip().upper()
            fy         = request.args.get("fy","").strip()
            raw_lookup = []

        if not gstin or not fy:
            return jsonify({"error":"gstin and fy are required"}), 400

        # Build {gstin→{trade_name, legal_name}} dict
        name_lookup = {}
        for entry in raw_lookup:
            g = (entry.get("gstin") or "").strip().upper()
            if g:
                name_lookup[g] = {
                    "trade_name": (entry.get("trade_name") or "").strip(),
                    "legal_name":  (entry.get("legal_name") or "").strip(),
                }

        try:
            import openpyxl  # noqa
        except ImportError:
            return jsonify({"error":
                "openpyxl is not installed. Run: pip install openpyxl"}), 500

        buf, err = g2a_json_to_excel(gstin, fy, name_lookup=name_lookup)
        if err:
            return jsonify({"error": err}), 404

        fname = f"GSTR2A_{gstin}_{fy.replace('-','_')}.xlsx"
        save_dir = os.path.join(G2A_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
        return save_excel_and_respond(buf, save_dir, fname, log_fn=g2a_log)
    except Exception as e:
        import traceback
        g2a_log(f"export_excel error: {e} | {traceback.format_exc()}", "error")
        return jsonify({"error": str(e)}), 500


@app.route("/g2a/check_excel_deps")
def g2a_check_excel_deps():
    try:
        import openpyxl
        return jsonify({"ok": True, "version": openpyxl.__version__})
    except ImportError:
        return jsonify({"ok": False,
                        "error": "openpyxl not installed",
                        "fix": "pip install openpyxl"}), 200


@app.route("/g2a/reset", methods=["POST"])
def g2a_reset():
    g2a_set({
        "status":"idle","log":[],"error":None,
        "captcha_image":None,"captcha_answer":None,"otp_answer":None,
        "progress":0,"current_month":None,"total_months":0,
        "done_months":0,"files":[],"gstin":None,"fy":None,
        "specific_month":None,
    })
    return jsonify({"ok":True})


@app.route("/g2a/files")
def g2a_list_files():
    """List all downloaded GSTR-2A files with age / freshness metadata."""
    stale = g2a_stale_days()
    files = []
    if os.path.isdir(G2A_DOWNLOAD_DIR):
        for gstin_d in sorted(os.listdir(G2A_DOWNLOAD_DIR)):
            gp = os.path.join(G2A_DOWNLOAD_DIR, gstin_d)
            if not os.path.isdir(gp): continue
            for fy_dir in sorted(os.listdir(gp), reverse=True):
                fp = os.path.join(gp, fy_dir)
                if not os.path.isdir(fp): continue
                for fname in sorted(os.listdir(fp)):
                    if not fname.endswith(".json"): continue
                    fpath   = os.path.join(fp, fname)
                    mtime   = os.path.getmtime(fpath)
                    age_d   = (time.time() - mtime) / 86400
                    size_kb = max(1, os.path.getsize(fpath) // 1024)
                    files.append({
                        "gstin":    gstin_d,
                        "fy":       fy_dir.replace("_","-"),
                        "filename": fname,
                        "size_kb":  size_kb,
                        "age_days": round(age_d, 1),
                        "fresh":    age_d < stale,
                        "url":      f"/g2a/file/{gstin_d}/{fy_dir}/{fname}",
                    })
    return jsonify({"files": files, "stale_days": stale})


@app.route("/g2a/file/<gstin>/<fy_dir>/<fname>")
def g2a_serve_file(gstin, fy_dir, fname):
    fpath = os.path.join(G2A_DOWNLOAD_DIR, gstin, fy_dir, fname)
    if not os.path.isfile(fpath):
        return jsonify({"error":"File not found"}), 404
    return send_file(fpath, as_attachment=True, download_name=fname)


# ── File status per client/FY — used by UI before starting download ──────────
@app.route("/g2a/file_status")
def g2a_file_status():
    """
    Return per-month cache status for a given GSTIN + FY.
    The UI calls this when the user selects a client/FY to show which months
    are already cached (green), stale (amber), or missing (red) before they
    click Start Download.

    Query params: gstin, fy
    """
    gstin = request.args.get("gstin","").strip().upper()
    fy    = request.args.get("fy","").strip()
    if not gstin or not fy:
        return jsonify({"error":"gstin and fy required"}), 400

    stale     = g2a_stale_days()
    gstin_dir = os.path.join(G2A_DOWNLOAD_DIR, gstin, fy.replace("-","_"))
    months    = g2a_months_for_fy(fy)
    result    = []

    for month in months:
        abbr  = month["display"].replace(" ","_")
        fpath = os.path.join(gstin_dir, f"GSTR2A_{abbr}.json")
        if os.path.isfile(fpath):
            mtime   = os.path.getmtime(fpath)
            age_d   = (time.time() - mtime) / 86400
            size_kb = max(1, os.path.getsize(fpath) // 1024)
            fresh   = age_d < stale
            result.append({
                "month":   month["display"],
                "period":  month["period"],
                "status":  "fresh" if fresh else "stale",
                "age_days": round(age_d, 1),
                "size_kb": size_kb,
                "url":     f"/g2a/file/{gstin}/{fy.replace('-','_')}/GSTR2A_{abbr}.json",
            })
        else:
            result.append({
                "month":   month["display"],
                "period":  month["period"],
                "status":  "missing",
                "age_days": None,
                "size_kb": 0,
                "url":     None,
            })

    total   = len(result)
    fresh_n = sum(1 for r in result if r["status"] == "fresh")
    stale_n = sum(1 for r in result if r["status"] == "stale")
    miss_n  = sum(1 for r in result if r["status"] == "missing")
    return jsonify({
        "gstin":      gstin,
        "fy":         fy,
        "stale_days": stale,
        "months":     result,
        "summary":    {"total": total, "fresh": fresh_n,
                       "stale": stale_n, "missing": miss_n},
    })


# ── FY Config routes (Requirement 3) ─────────────────────────────────────────
@app.route("/g2a/fy_config", methods=["GET"])
def g2a_get_fy_config():
    """Return current FY config and the full computed list of available FYs."""
    cfg = g2a_load_fy_config()
    return jsonify({
        "extra_fys":  cfg.get("extra_fys", []),
        "stale_days": cfg.get("stale_days", _DEFAULT_STALE_DAYS),
        "all_fys":    g2a_available_fys(),       # auto-range + extras combined
    })


@app.route("/g2a/fy_config", methods=["POST"])
def g2a_update_fy_config():
    """
    Modify FY config from Settings.  All body fields are optional:

      { "add_fy":    "2027-28" }     ← add a future FY to the dropdown
      { "remove_fy": "2027-28" }     ← remove a manually-added FY
      { "stale_days": 7 }            ← change the re-download threshold (days)

    Multiple fields can be combined in one call.
    Returns updated config + fresh all_fys list so the UI dropdown can
    refresh immediately.
    """
    data = request.get_json(force=True) or {}
    cfg  = g2a_load_fy_config()
    changed = False

    # ── Stale-days ──────────────────────────────────────────────────────────
    if "stale_days" in data:
        sd = int(data["stale_days"])
        if sd < 1 or sd > 365:
            return jsonify({"error":"stale_days must be between 1 and 365"}), 400
        cfg["stale_days"] = sd
        changed = True

    # ── Add FY ──────────────────────────────────────────────────────────────
    if "add_fy" in data:
        raw = str(data["add_fy"]).strip()
        # Accept "YYYY-YY" or "YYYY-YYYY" — normalise to "YYYY-YY"
        parts = raw.split("-")
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].isdigit():
            return jsonify({"error":
                f"Invalid FY format '{raw}'. Use YYYY-YY, e.g. '2027-28'"}), 400
        y1     = int(parts[0])
        y2_exp = str(y1 + 1)[-2:]       # expected short suffix
        if parts[1] not in (y2_exp, str(y1 + 1)):
            return jsonify({"error":
                f"Year mismatch in '{raw}': expected {y1}-{y2_exp}"}), 400
        fy_norm = f"{y1}-{y2_exp}"
        extras  = cfg.get("extra_fys", [])
        if fy_norm not in extras:
            extras.append(fy_norm)
            extras.sort(key=lambda x: int(x.split("-")[0]))
            cfg["extra_fys"] = extras
            changed = True

    # ── Remove FY ────────────────────────────────────────────────────────────
    if "remove_fy" in data:
        fy_del = str(data["remove_fy"]).strip()
        before = len(cfg.get("extra_fys",[]))
        cfg["extra_fys"] = [f for f in cfg.get("extra_fys",[]) if f != fy_del]
        if len(cfg["extra_fys"]) != before:
            changed = True

    if not changed:
        return jsonify({"ok": True, "message": "No changes made",
                        "extra_fys":  cfg["extra_fys"],
                        "stale_days": cfg["stale_days"],
                        "all_fys":    g2a_available_fys()})

    if g2a_save_fy_config(cfg):
        return jsonify({
            "ok":        True,
            "extra_fys": cfg["extra_fys"],
            "stale_days":cfg["stale_days"],
            "all_fys":   g2a_available_fys(),
        })
    return jsonify({"error":"Could not write fy_config.json"}), 500


# ── Serve the HTML UI directly ───────────────────────────────


# ══════════════════════════════════════════════════════════════════════════════
#  GSTR-2B MODULE  (Static ITC Register — u/s 16(2)(aa) CGST Act)
# ══════════════════════════════════════════════════════════════════════════════



# ╔══════════════════════════════════════════════════════════════════════════════╗
# ║                          GSTR-2B MODULE                                    ║
# ║  Static ITC Register — generated on 14th of following month               ║
# ║  Statutory basis for ITC u/s 16(2)(aa) CGST Act                           ║
# ║                                                                              ║
# ║  KEY SCHEMA DIFFERENCES FROM GSTR-2A:                                       ║
# ║  • JSON wrapped in  {"data":{...}}   (handle both flat + wrapped)           ║
# ║  • Period key  'rtnprd'  (not 'fp')                                         ║
# ║  • genDt field  — generation date (frozen, statutory snapshot)              ║
# ║  • Credit/Debit Notes: 'cdnr'/'cdnra'  (not 'cdn'/'cdna' like 2A)         ║
# ║  • New: itcavl "Y"/"N"  — ITC eligibility per invoice                      ║
# ║  • New: diffprcnt 1 / 0.65  — differential % (eligible portion)            ║
# ║  • New: reason  — why ITC is blocked                                        ║
# ║  • New sections: 'eco'/'ecoa'  — E-Commerce operator purchases             ║
# ║  • ISD: flat structure  (no nested itms)                                    ║
# ║  • Monthly availability: after 14th of following month                      ║
# ╚══════════════════════════════════════════════════════════════════════════════╝

# ── GSTR-2B module globals ─────────────────────────────────────────────────────
g2b_lock         = threading.Lock()
G2B_DOWNLOAD_DIR = PATHS.gstr2b_dir
g2b_state = {
    "status":         "idle",
    "log":            [],
    "error":          None,
    "progress":       0,
    "current_month":  None,
    "total_months":   0,
    "done_months":    0,
    "files":          [],
    "gstin":          None,
    "fy":             None,
    "specific_month": None,
    "captcha_image":  None,
    "captcha_answer": None,
    "otp_answer":     None,
}


def g2b_log(msg, level="info"):
    ts   = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    with g2b_lock:
        g2b_state["log"].append(line)
        if len(g2b_state["log"]) > 400:
            g2b_state["log"] = g2b_state["log"][-400:]
    (log.error if level == "error" else log.info)(f"G2B: {msg}")


def g2b_set(updates, log_msg=None):
    with g2b_lock:
        g2b_state.update(updates)
    if log_msg:
        g2b_log(log_msg)


def g2b_wait_field(field, timeout_sec=300):
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        time.sleep(0.4)
        with g2b_lock:
            val = g2b_state.get(field)
            if val:
                g2b_state[field] = None
                return val
    return None


# ── Month availability — GSTR-2B available only after 14th of following month ──
def g2b_months_for_fy(fy):
    """
    GSTR-2B for month M is published on 14th of M+1.
    Example: February 2026 -> available from 14-March-2026.
    Only returns months where the publication date has already passed.
    Months still pending are logged so the user knows why they are skipped.
    """
    fy_year = int(fy.split("-")[0])
    now     = datetime.now()
    order   = [
        (4,"April"),(5,"May"),(6,"June"),(7,"July"),
        (8,"August"),(9,"September"),(10,"October"),
        (11,"November"),(12,"December"),
        (1,"January"),(2,"February"),(3,"March"),
    ]
    months  = []
    skipped = []
    for mon_num, mon_name in order:
        yr      = fy_year if mon_num >= 4 else fy_year + 1
        gen_yr  = yr + 1  if mon_num == 12 else yr
        gen_mon = 1       if mon_num == 12 else mon_num + 1
        # Not yet in the publication month
        if gen_yr > now.year or (gen_yr == now.year and gen_mon > now.month):
            skipped.append(f"{mon_name} {yr} (publishes 14-{_g2b_mon_abbr(gen_mon)}-{gen_yr})")
            continue
        # In publication month but before 14th
        if gen_yr == now.year and gen_mon == now.month and now.day < 14:
            skipped.append(
                f"{mon_name} {yr} (publishes on 14-{_g2b_mon_abbr(gen_mon)}-{gen_yr}, "
                f"today is {now.day}-{_g2b_mon_abbr(now.month)}-{now.year})")
            continue
        months.append({
            "num":    mon_num,
            "year":   yr,
            "name":   mon_name,
            "abbr":   mon_name[:3].upper(),
            "display":f"{mon_name} {yr}",
            "period": f"{mon_num:02d}{yr}",
        })
    if skipped:
        g2b_log(f"  Skipping {len(skipped)} not-yet-published month(s):")
        for s in skipped:
            g2b_log(f"    x {s}")
    return months


def _g2b_mon_abbr(m):
    return ["","Jan","Feb","Mar","Apr","May","Jun",
            "Jul","Aug","Sep","Oct","Nov","Dec"][m]


def g2b_unwrap(raw_json):
    """
    GSTR-2B JSON may be wrapped: {"data": {...}} or flat {...}.
    Returns the inner data dict.
    """
    if isinstance(raw_json, dict):
        if "data" in raw_json and isinstance(raw_json["data"], dict):
            return raw_json["data"]
        return raw_json
    return {}


# ── Login — same portal, parallel state ───────────────────────────────────────
def g2b_do_browser_login(page, username, password):
    """Login to GST portal — mirrors g2a_do_browser_login but uses g2b_state."""
    g2b_log("Opening GST login page...")
    try:
        page.goto(GST_LOGIN, wait_until="domcontentloaded", timeout=25000)
        try: page.wait_for_selector(
                "input#username, input[name='username'], input[placeholder*='username' i]",
                state="visible", timeout=8000)
        except Exception: pass
    except Exception as e:
        g2b_log(f"  ✗ Could not open login page: {e}", "error")
        g2b_set({"status":"error","error":str(e)}); return False

    for sel in ["input#username","input[name='username']","input[placeholder*='username' i]"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=4000)
            page.locator(sel).first.fill(username); g2b_log("  ✓ Username filled"); break
        except Exception: continue
    for sel in ["input#user_pass","input[name='user_pass']",
                "input[type='password']","input[placeholder*='password' i]"]:
        try: page.locator(sel).first.fill(password); g2b_log("  ✓ Password filled"); break
        except Exception: continue
    # Proceed immediately to captcha capture

    # Capture captcha
    cap_img = None
    try:
        import base64 as _b64
        for sel in ["#imgCaptcha","img[id*='aptcha' i]","img[src*='captcha' i]",
                    "img[src*='kaptcha' i]","img[alt*='captcha' i]",
                    ".captchaImage img","img.captcha","form img","img"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
                    g2b_log(f"  ✓ Captcha captured ({sel})"); break
            except Exception: continue
        if not cap_img:
            data_url = page.evaluate("""() => {
                for (const cv of document.querySelectorAll('canvas')) {
                    if (cv.width > 30 && cv.height > 10) {
                        try { return cv.toDataURL('image/png'); } catch(e) {}
                    }
                }
                return null;
            }""")
            if data_url and data_url.startswith("data:image"):
                cap_img = data_url; g2b_log("  ✓ Captcha via canvas")
        if not cap_img:
            png = page.screenshot(full_page=False)
            cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
            g2b_log("  ✓ Captcha via screenshot fallback")
    except Exception as e:
        g2b_log(f"  ⚠ Captcha capture: {e}")

    g2b_set({"status":"waiting_captcha","captcha_image":cap_img})
    g2b_log("  ⏸ Waiting for captcha...")

    answer = g2b_wait_field("captcha_answer", timeout_sec=300)
    if not answer:
        g2b_log("  ✗ Captcha timeout","error")
        g2b_set({"status":"error","error":"Captcha timeout"}); return False

    g2b_set({"status":"running","captcha_image":None})

    for sel in ["input[placeholder*='Characters' i]","input#captcha","input[name='captcha']"]:
        try: page.locator(sel).first.fill(str(answer)); break
        except Exception: continue
    for sel in ["button[type='submit']","input[type='submit']","button:has-text('LOGIN')"]:
        try: page.locator(sel).first.click(); g2b_log("  ✓ Login submitted"); break
        except Exception: continue
    time.sleep(1.5)

    # OTP handling
    otp_needed = False
    for _ in range(8):
        time.sleep(1)
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.wait_for(state="visible", timeout=1500)
            otp_needed = True; break
        except Exception:
            if check_login_success(page) is not False: break
    if otp_needed:
        g2b_set({"status":"waiting_otp"}); g2b_log("  ⏸ Waiting for OTP...")
        otp = g2b_wait_field("otp_answer", timeout_sec=180)
        if not otp:
            g2b_log("  ✗ OTP timeout","error")
            g2b_set({"status":"error","error":"OTP timeout"}); return False
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.click(); time.sleep(0.2); otp_el.fill(str(otp))
            page.locator("button[type='submit'],input[type='submit']").first.click()
            g2b_log("  ✓ OTP submitted"); time.sleep(1)
        except Exception as e:
            g2b_log(f"  ✗ OTP error: {e}","error"); return False

    time.sleep(1)

    # Captcha retry loop (up to 3 attempts)
    MAX_CAPTCHA_ATTEMPTS = 3
    for _attempt in range(MAX_CAPTCHA_ATTEMPTS):
        result = check_login_success(page)
        if result is not False:
            break
        remaining = MAX_CAPTCHA_ATTEMPTS - _attempt - 1
        if remaining == 0:
            g2b_log(f"  ✗ Login failed after {MAX_CAPTCHA_ATTEMPTS} captcha attempts","error")
            g2b_set({"status":"error","error":f"Login failed after {MAX_CAPTCHA_ATTEMPTS} captcha attempts"})
            return False
        g2b_log(f"  ✗ Login failed — re-capturing captcha (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")
        time.sleep(2)
        new_cap = None
        try:
            import base64 as _b64
            for sel in ["#imgCaptcha","img[id*='aptcha' i]","img[src*='captcha' i]",
                        "img[src*='kaptcha' i]","form img","img"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=2000)
                    png = loc.screenshot()
                    if png and len(png) > 500:
                        new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                        g2b_log(f"  ✓ Fresh captcha re-captured ({sel})"); break
                except Exception: continue
            if not new_cap:
                png = page.screenshot(full_page=False)
                new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
        except Exception as ce:
            g2b_log(f"  ⚠ Re-capture error: {ce}")
        g2b_set({"status":"running","captcha_image":None})
        time.sleep(0.1)
        g2b_set({"status":"waiting_captcha","captcha_image":new_cap})
        g2b_log(f"  ⏸ Waiting for captcha (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")
        new_answer = g2b_wait_field("captcha_answer", timeout_sec=300)
        if not new_answer:
            g2b_log("  ✗ Captcha retry timeout","error")
            g2b_set({"status":"error","error":"Captcha retry timeout"}); return False
        g2b_set({"status":"running","captcha_image":None})
        for sel in ["input[placeholder*='Characters' i]","input#captcha","input[name='captcha']"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.fill(""); loc.fill(str(new_answer)); break
            except Exception: continue
        for sel in ["button[type='submit']","input[type='submit']","button:has-text('LOGIN')"]:
            try: page.locator(sel).first.click(); g2b_log("  ✓ Login re-submitted"); break
            except Exception: continue
        time.sleep(2)
        for _ in range(5):
            time.sleep(1)
            try:
                otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
                otp_el.wait_for(state="visible", timeout=1000)
                g2b_set({"status":"waiting_otp"})
                g2b_log("  ⏸ OTP required (retry)...")
                otp_r = g2b_wait_field("otp_answer", timeout_sec=180)
                if not otp_r:
                    g2b_log("  ✗ OTP timeout (retry)","error")
                    g2b_set({"status":"error","error":"OTP timeout on retry"}); return False
                otp_el.click(); time.sleep(0.2); otp_el.fill(str(otp_r))
                page.locator("button[type='submit'],input[type='submit']").first.click()
                g2b_log("  ✓ OTP re-submitted"); time.sleep(1); break
            except Exception:
                if check_login_success(page) is not False: break

    g2b_log("  ✅ Logged in successfully")
    g2a_dismiss_popup(page)   # same portal — reuse popup dismissal
    return True


# ── Navigate Returns Dashboard to GSTR-2B offline download ────────────────────
def g2b_navigate_offline_download(page, fy):
    """
    Navigate Returns Dashboard → select FY / Quarter / Month → SEARCH
    → click DOWNLOAD on GSTR-2B tile → reach gstr2b/offlinedownload.
    Returns True if GSTR-2B page reached.
    """
    now = datetime.now()
    month_names = {
        1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
        7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"
    }
    fy_start  = int(fy.split("-")[0])
    fy_months = list(range(4,13)) + list(range(1,4))
    fy_years  = [fy_start]*9 + [fy_start+1]*3
    use_idx   = -1
    for idx, (m, y) in enumerate(zip(fy_months, fy_years)):
        if y < now.year or (y == now.year and m < now.month):
            use_idx = idx
    use_month     = fy_months[use_idx] if use_idx >= 0 else 4
    use_year      = fy_years[use_idx]  if use_idx >= 0 else fy_start
    quarter_label, _ = g2a_quarter_for_month(use_month)
    month_name        = month_names[use_month]
    g2b_log(f"  -> Selecting: FY={fy}  Q={quarter_label}  Period={month_name}")

    def js_select(nth, label):
        result = page.evaluate(f"""() => {{
            const sel = document.querySelectorAll('select')[{nth}];
            if (!sel) return {{ok:false, opts:[]}};
            const opts = Array.from(sel.options).map(o => o.text.trim());
            for (const opt of sel.options) {{
                if (opt.text.trim() === '{label}') {{
                    sel.value = opt.value;
                    sel.dispatchEvent(new Event('change', {{bubbles:true}}));
                    return {{ok:true, selected:opt.text.trim()}};
                }}
            }}
            return {{ok:false, opts:opts}};
        }}""")
        g2b_log(f"    select[{nth}] '{label}': ok={result.get('ok')}")
        return result.get('ok', False)

    def wait_for_option(nth, label, timeout=5):
        deadline = time.time() + timeout
        while time.time() < deadline:
            opts = page.evaluate(f"""() => {{
                const sel = document.querySelectorAll('select')[{nth}];
                return sel ? Array.from(sel.options).map(o=>o.text.trim()) : [];
            }}""")
            if label in opts: return True
            time.sleep(0.3)
        return False

    js_select(0, fy); time.sleep(0.5)
    wait_for_option(1, quarter_label, timeout=4)
    js_select(1, quarter_label)
    g2b_log(f"    Waiting for '{month_name}' in Period dropdown...")
    if wait_for_option(2, month_name, timeout=6):
        js_select(2, month_name)
    else:
        g2b_log("    ⚠ Period did not reload — trying anyway")
        js_select(2, month_name)

    for s in ["button:has-text('SEARCH')","button:has-text('Search')"]:
        try:
            page.locator(s).first.wait_for(state="visible", timeout=4000)
            page.locator(s).first.click()
            g2b_log(f"    Clicked SEARCH"); break
        except Exception: continue
    time.sleep(2)

    # Click DOWNLOAD / VIEW inside GSTR-2B tile
    clicked = page.evaluate("""() => {
        const all = Array.from(document.querySelectorAll('*'));
        for (const el of all) {
            const t = (el.childNodes.length === 1
                ? el.textContent : el.getAttribute('data-title') || '').trim();
            if (t === 'GSTR2B' || t === 'GSTR-2B' || t === 'GSTR 2B'
                    || t.toUpperCase() === 'GSTR2B') {
                let container = el;
                for (let i = 0; i < 6; i++) {
                    container = container.parentElement;
                    if (!container) break;
                    for (const btn of container.querySelectorAll('button,a')) {
                        const bt = (btn.textContent||'').trim().toUpperCase();
                        if (bt === 'DOWNLOAD' || bt === 'VIEW/DOWNLOAD'
                                || bt === 'VIEW' || bt === 'PROCEED') {
                            btn.click();
                            return 'CLICKED: ' + bt + ' in ' + (container.className||'?').slice(0,30);
                        }
                    }
                }
            }
        }
        return null;
    }""")
    g2b_log(f"    GSTR-2B tile click: {clicked}")

    try:
        page.wait_for_url("*gstr2b*", timeout=10000)
    except Exception:
        pass
    time.sleep(1.5)
    g2b_log(f"    -> Now on: {page.url[:70]}")
    return "gstr2b" in page.url.lower() or "offlinedownload" in page.url


# ── Download one month's JSON from the offlinedownload page ───────────────────
def g2b_fetch_download_link(page, url, gstin_dir, month_display):
    """Download GSTR-2B JSON/ZIP from URL, save as GSTR2B_<Month>_<Year>.json."""
    import zipfile, io as _io
    import base64 as _b64
    abbr      = month_display.replace(" ", "_")
    json_path = os.path.join(gstin_dir, f"GSTR2B_{abbr}.json")
    try:
        resp = page.request.get(url, headers={
            "Accept":  "*/*",
            "Referer": "https://return.gst.gov.in/returns/auth/gstr2b/offlinedownload",
        }, timeout=120000)
        raw = resp.body()
        ct  = resp.headers.get("content-type","").lower()
        g2b_log(f"    [{resp.status}] {ct[:40]}  {len(raw)} bytes")
        if resp.status != 200 or not raw:
            return None
        if raw[:2] == b"PK" or "zip" in ct:
            with zipfile.ZipFile(_io.BytesIO(raw)) as zf:
                names  = zf.namelist()
                target = next((n for n in names if n.endswith(".json")), names[0])
                data   = zf.read(target)
            with open(json_path, "wb") as f: f.write(data)
            size_kb = max(1, len(data)//1024)
        else:
            with open(json_path, "wb") as f: f.write(raw)
            size_kb = max(1, len(raw)//1024)
        g2b_log(f"    ✅ {month_display} — {size_kb} KB saved")
        return {"month":month_display, "filename":f"GSTR2B_{abbr}.json", "size_kb":size_kb}
    except Exception as e:
        g2b_log(f"    ✗ Download error: {e}", "error")
        return None


# ── Download all months via offline download UI ────────────────────────────────
def g2b_download_all_months(page, months, gstin, gstin_dir, fy):
    """
    Per-month GSTR-2B downloader.

    KEY FIXES (v3):
      1. FY selection — uses Playwright-native select_option() (same as g2a)
         instead of custom js_sel(); tries both "YYYY-YY" and "YYYY-YYYY"
         formats; logs available options if neither matches.
      2. Download timeout — 5 seconds max for expect_download().
         If it times out → skip to next month (whole-year run) or break
         (specific-month run).  No more 90-second hangs.
      3. Period verification — after every successful download, read the JSON
         fp/rtnprd field and confirm it matches the expected MMYYYY period.
         If mismatch → rename and log so the user can see the actual period.
      4. Dropdown layout detection — same logic as g2a:
           3 selects → FY + Quarter + Month
           2 selects → FY + Month only
    """
    QUARTER_END   = {3, 6, 9, 12}
    DOWNLOAD_TIMEOUT_MS = 5_000   # Issue 2: 5 seconds max per download attempt

    month_names = {
        1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
        7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"
    }
    dashboard_url = "https://return.gst.gov.in/returns/auth/dashboard"
    files_done    = []

    # ── Inner helpers (same pattern as g2a) ──────────────────────────────────
    def pw_sel(nth, label, timeout=8):
        """Playwright-native select — fires real browser events Angular needs."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                loc = page.locator("select").nth(nth)
                loc.wait_for(state="visible", timeout=2000)
                loc.select_option(label=label)
                time.sleep(0.3)
                cur = loc.evaluate(
                    "el => el.options[el.selectedIndex] "
                    "? el.options[el.selectedIndex].text.trim() : ''")
                if cur == label:
                    return True
                loc.select_option(label=label)   # retry once
                time.sleep(0.4)
                return True
            except Exception:
                time.sleep(0.3)
        return False

    def wait_opt(nth, label, timeout=8):
        """Poll until label appears as an option in nth <select>."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                opts = page.locator("select").nth(nth).evaluate(
                    "el => Array.from(el.options).map(o => o.text.trim())")
                if label in opts:
                    return True
            except Exception:
                pass
            time.sleep(0.3)
        return False

    def get_select_opts(nth):
        try:
            return page.locator("select").nth(nth).evaluate(
                "el => Array.from(el.options).map(o => o.text.trim())")
        except Exception:
            return []

    def select_fy():
        """
        Issue 1 fix — try FY in both portal formats.
        Portal may show "2024-25" (short) or "2024-2025" (full).
        Logs available options if neither matches.
        """
        fy_long = fy                             # e.g. "2024-25"
        parts   = fy.split("-")
        # Build long form: "2024-25" → "2024-2025"
        if len(parts) == 2 and len(parts[1]) == 2:
            fy_full = f"{parts[0]}-20{parts[1]}"  # "2024-2025"
        else:
            fy_full = fy_long

        for fmt in [fy_long, fy_full]:
            if wait_opt(0, fmt, timeout=5):
                ok = pw_sel(0, fmt)
                if ok:
                    g2b_log(f"    ✓ FY selected: '{fmt}'")
                    return True
            g2b_log(f"    ⚠ FY '{fmt}' not in dropdown")

        # Log what's actually available
        avail = get_select_opts(0)
        g2b_log(f"    ✗ Could not select FY. Available options: {avail}")
        return False

    def verify_period(json_path, expected_period, mon_display):
        """
        Issue 3 — read saved JSON and confirm fp/rtnprd matches expected period.
        expected_period = MMYYYY  e.g. "042024"
        """
        import json as _json
        try:
            with open(json_path, encoding="utf-8") as f:
                raw = _json.load(f)
            data = raw.get("data", raw) if isinstance(raw, dict) and "data" in raw else raw
            fp   = str(data.get("fp","") or data.get("rtnprd","")).strip()
            if not fp:
                g2b_log(f"    ℹ Period field not found in JSON — assuming correct")
                return True
            if fp == expected_period:
                g2b_log(f"    ✓ Period verified: {fp} == {expected_period}")
                return True
            # Mismatch — rename file to reflect actual period
            import calendar as _cal
            try:
                actual_mon = int(fp[:2]); actual_yr = int(fp[2:])
                actual_name = f"{_cal.month_name[actual_mon]}_{actual_yr}"
            except Exception:
                actual_name = fp
            new_path = os.path.join(
                os.path.dirname(json_path),
                f"GSTR2B_{actual_name}.json")
            os.rename(json_path, new_path)
            g2b_log(f"    ⚠ Period mismatch: got {fp}, expected {expected_period}. "
                    f"Renamed to GSTR2B_{actual_name}.json")
            return False
        except Exception as e:
            g2b_log(f"    ⚠ Period verification error: {e}")
            return True   # assume OK if can't verify

    # ── Main month loop ───────────────────────────────────────────────────────
    for i, month in enumerate(months):
        mon_display   = month["display"]
        mon_num       = month["num"]
        period        = month["period"]        # MMYYYY e.g. "042024"
        quarter_label, _ = g2a_quarter_for_month(mon_num)
        mon_name      = month_names[mon_num]

        g2b_set({
            "status":        "downloading",
            "current_month": mon_display,
            "progress":      int((i / len(months)) * 100),
        })
        g2b_log(f"\n[{i+1}/{len(months)}] {mon_display}  (period={period})")

        # Skip if already downloaded and period-verified
        existing = [f for f in os.listdir(gstin_dir)
                    if mon_display.replace(" ", "_") in f and f.endswith(".json")]
        if existing:
            g2b_log(f"  ✓ Already downloaded — skipping")
            continue

        # ── Navigate to dashboard ─────────────────────────────────────────────
        try:
            page.evaluate(f"window.location.href = '{dashboard_url}'")
        except Exception:
            page.goto(dashboard_url, wait_until="domcontentloaded", timeout=15000)
        try:
            page.wait_for_load_state("domcontentloaded", timeout=12000)
        except Exception:
            pass
        time.sleep(1.5)

        # ── Issue 1 Fix: FY selection ─────────────────────────────────────────
        if not select_fy():
            g2b_log(f"  ✗ FY selection failed for {mon_display} — skipping")
            continue
        time.sleep(0.5)

        # ── Detect layout: 3 selects (FY+Qtr+Month) or 2 selects (FY+Month) ──
        n_selects = 0
        for _w in range(15):
            try:
                n_selects = page.evaluate(
                    "() => document.querySelectorAll('select').length")
            except Exception:
                pass
            if n_selects >= 2:
                break
            time.sleep(0.3)
        g2b_log(f"    {n_selects} select(s) after FY change")

        if n_selects >= 3:
            # Normal layout: FY → Quarter → Month
            g2b_log(f"    Selecting Quarter: {quarter_label}")
            wait_opt(1, quarter_label, timeout=5)
            pw_sel(1, quarter_label)
            g2b_log(f"    Selecting Month:   {mon_name}")
            wait_opt(2, mon_name, timeout=6)
            pw_sel(2, mon_name)
        else:
            # 2-dropdown layout: FY → Month only
            g2b_log(f"    Selecting Month: {mon_name} (no quarter dropdown)")
            wait_opt(1, mon_name, timeout=6)
            pw_sel(1, mon_name)

        # ── SEARCH ────────────────────────────────────────────────────────────
        search_clicked = False
        for s in ["button:has-text('SEARCH')", "button:has-text('Search')"]:
            try:
                page.locator(s).first.wait_for(state="visible", timeout=4000)
                page.locator(s).first.click()
                g2b_log(f"    ✓ SEARCH clicked")
                search_clicked = True
                break
            except Exception:
                continue
        if not search_clicked:
            g2b_log(f"  ✗ SEARCH button not found — skipping {mon_display}")
            continue
        time.sleep(2.5)

        # ── Click DOWNLOAD on GSTR-2B tile ────────────────────────────────────
        clicked = page.evaluate("""() => {
            for (const el of document.querySelectorAll('*')) {
                const txt = (el.textContent || '').trim().toUpperCase()
                              .split(' ').filter(s => s).join(' ');
                if ((txt === 'GSTR2B' || txt === 'GSTR-2B' || txt === 'GSTR 2B')
                        && el.children.length === 0) {
                    let container = el.parentElement;
                    for (let i = 0; i < 10; i++) {
                        if (!container) break;
                        for (const btn of container.querySelectorAll('button,a')) {
                            if ((btn.textContent || '').trim().toUpperCase() === 'DOWNLOAD') {
                                btn.click();
                                return 'DOWNLOAD';
                            }
                        }
                        container = container.parentElement;
                    }
                }
            }
            return null;
        }""")

        if not clicked:
            clicked = page.evaluate("""() => {
                for (const btn of document.querySelectorAll('button,a')) {
                    if ((btn.textContent || '').trim().toUpperCase() !== 'DOWNLOAD') continue;
                    let p = btn.parentElement;
                    for (let i = 0; i < 8; i++) {
                        if (!p) break;
                        const pt = (p.textContent || '').toUpperCase();
                        if (pt.includes('GSTR2B') || pt.includes('GSTR-2B')
                                || pt.includes('ITC STATEMENT')) {
                            btn.click();
                            return 'DOWNLOAD (fallback)';
                        }
                        p = p.parentElement;
                    }
                }
                return null;
            }""")

        if not clicked:
            if mon_num not in QUARTER_END:
                g2b_log(f"  ⏭ {mon_display} — IFF month (QRMP for this quarter): "
                        f"no GSTR-2B tile")
            else:
                g2b_log(f"  ⚠ GSTR-2B DOWNLOAD button not found for {mon_display}")
            continue

        g2b_log(f"  ✓ Clicked {clicked}")

        # ── Wait for GSTR-2B download page ────────────────────────────────────
        try:
            page.wait_for_url("*gstr2b*dwld*", timeout=14000)
        except Exception:
            pass
        time.sleep(1)

        cur_url = page.url.lower()
        g2b_log(f"  -> {page.url[:80]}")

        if "accessdenied" in cur_url or "login" in cur_url:
            g2b_log(f"  ⚠ Access denied — skipping {mon_display}")
            continue

        if "gstr2bdwld" not in cur_url and "gstr2bqdwld" not in cur_url:
            g2b_log(f"  ⚠ Not on GSTR-2B download page — skipping {mon_display}")
            continue

        if "gstr2bqdwld" in cur_url:
            g2b_log(f"  ℹ Quarterly GSTR-2B (QRMP)")
        else:
            g2b_log(f"  ℹ Monthly GSTR-2B")

        # ── Issue 2 Fix: GENERATE with 5-second download timeout ─────────────
        abbr       = mon_display.replace(" ", "_")
        json_path  = os.path.join(gstin_dir, f"GSTR2B_{abbr}.json")
        downloaded = False

        for s in [
            "button:has-text('GENERATE JSON FILE TO DOWNLOAD')",
            "button:has-text('GENERATE JSON FILE')",
            "button:has-text('GENERATE JSON')",
            "button:has-text('GENERATE')",
        ]:
            try:
                page.locator(s).first.wait_for(state="visible", timeout=8000)
                g2b_log(f"    -> Clicking GENERATE…")
                try:
                    with page.expect_download(timeout=DOWNLOAD_TIMEOUT_MS) as dl_info:
                        page.locator(s).first.click()
                    dl = dl_info.value

                    import zipfile, io as _io
                    raw = open(dl.path(), "rb").read()
                    if raw[:2] == b"PK":   # zip file
                        with zipfile.ZipFile(_io.BytesIO(raw)) as zf:
                            names  = zf.namelist()
                            target = next((n for n in names if n.endswith(".json")), names[0])
                            data   = zf.read(target)
                        with open(json_path, "wb") as f:
                            f.write(data)
                        size_kb = max(1, len(data) // 1024)
                    else:
                        with open(json_path, "wb") as f:
                            f.write(raw)
                        size_kb = max(1, len(raw) // 1024)

                    g2b_log(f"    ✅ {mon_display} — {size_kb} KB saved")

                    # Issue 3: Verify the JSON contains the correct period
                    verify_period(json_path, period, mon_display)

                    result = {
                        "month":    mon_display,
                        "filename": f"GSTR2B_{abbr}.json",
                        "size_kb":  size_kb,
                        "period":   period,
                    }
                    files_done.append(result)
                    g2b_set({
                        "files":       files_done.copy(),
                        "done_months": len(files_done),
                        "progress":    int(((i + 1) / len(months)) * 100),
                    })
                    downloaded = True
                    break

                except Exception as dl_err:
                    err_str = str(dl_err).lower()
                    # Issue 2: timeout or error → skip, don't wait
                    if "timeout" in err_str or "download" in err_str:
                        g2b_log(f"    ⚠ Download not triggered within "
                                f"{DOWNLOAD_TIMEOUT_MS//1000}s — "
                                f"portal may show error or file not ready. "
                                f"Skipping {mon_display}.")
                    else:
                        g2b_log(f"    ⚠ Download error: {str(dl_err)[:100]}")
                    break   # Don't try other GENERATE selectors — move to next month

            except Exception as e:
                g2b_log(f"    ⚠ GENERATE button error: {str(e)[:80]}")
                continue

        if not downloaded:
            g2b_log(f"  ✗ {mon_display} — not downloaded, continuing with next month")

    return files_done




# ── Main worker ────────────────────────────────────────────────────────────────
def g2b_worker(gstin, fy, username, password, specific_month=""):
    """
    Thin wrapper guaranteeing any exception is logged and reflected in
    g2b_state instead of silently killing the thread.
    """
    try:
        _g2b_worker_impl(gstin, fy, username, password, specific_month)
    except Exception as fatal:
        import traceback as _tb
        try:
            g2b_log(f"✗ FATAL (uncaught): {fatal}", "error")
        except Exception:
            pass
        try:
            g2b_set({"status": "error", "error": f"Uncaught: {fatal}"})
        except Exception:
            pass
        log.error(f"[G2B] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _g2b_worker_impl(gstin, fy, username, password, specific_month=""):
    os.makedirs(G2B_DOWNLOAD_DIR, exist_ok=True)
    gstin_dir = os.path.join(G2B_DOWNLOAD_DIR, gstin, fy.replace("-","_"))
    os.makedirs(gstin_dir, exist_ok=True)

    all_months = g2b_months_for_fy(fy)
    if not all_months:
        g2b_set({"status":"error",
                 "error":f"No GSTR-2B months available for FY {fy}. "
                         f"GSTR-2B is available only after the 14th of the following month."}); return

    if specific_month:
        sm = specific_month.strip()
        filtered = [m for m in all_months if (
            sm.lower() == m["display"].lower() or sm == m["period"] or
            sm.lower() == m["name"].lower()    or sm.upper() == m["abbr"]
        )]
        if not filtered:
            g2b_set({"status":"error",
                     "error":f"Month '{specific_month}' not found or not yet generated. "
                             f"Available: {[m['display'] for m in all_months]}"}); return
        months = filtered
        g2b_log(f"?? Specific month: {months[0]['display']}")
    else:
        months = all_months

    g2b_set({"total_months":len(months),"done_months":0,"files":[]})
    g2b_log(f"GSTR-2B | GSTIN: {gstin} | FY: {fy} | {len(months)} month(s)")
    g2b_log(f"Note: GSTR-2B is a static frozen register — statutory for ITC u/s 16(2)(aa)")

    profile_dir = os.path.join(PATHS.profiles_dir, "gst_profile_g2b")
    os.makedirs(profile_dir, exist_ok=True)
    lock_file = os.path.join(profile_dir, "SingletonLock")
    if os.path.exists(lock_file):
        try: os.remove(lock_file); g2b_log("  ✓ Removed stale SingletonLock")
        except Exception: pass

    # sync_playwright().start() launches Playwright's own Node.js driver
    # subprocess — if it throws (not just hangs), that was previously
    # uncaught here, silently killing this thread with nothing logged.
    try:
        pw = sync_playwright().start()
    except Exception as e:
        g2b_log(f"✗ Playwright driver failed to start: {e}", "error")
        g2b_set({"status": "error", "error": f"Playwright driver failed to start: {e}"})
        return
    try:
        g2b_log("?? Launching browser...")
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=False, slow_mo=40,
                args=["--disable-blink-features=AutomationControlled","--no-sandbox","--start-maximized"],
                no_viewport=True, accept_downloads=True,
            )
        except Exception as e:
            g2b_log(f"  ✗ Browser launch error: {e}","error")
            g2b_set({"status":"error","error":str(e)}); return

        page = context.pages[0] if context.pages else context.new_page()
        page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        g2a_install_interceptors(page)

        if not g2b_do_browser_login(page, username, password):
            context.close(); return

        g2b_log("?? Activating session on Returns portal...")
        ok, page = g2a_activate_session(page, context)   # same portal — reuse
        if not ok:
            g2b_set({"status":"error","error":"Could not reach Returns portal"})
            context.close(); return
        g2b_log(f"  Using page: {page.url[:70]}")
        g2a_install_interceptors(page)

        g2b_log("📥 Starting GSTR-2B month-by-month download...")
        g2b_log(f"   Months: {', '.join(m['display'] for m in months)}")
        files_done = g2b_download_all_months(page, months, gstin, gstin_dir, fy)

        g2b_set({"status":"done","progress":100,"current_month":None,"files":files_done})
        g2b_log(f"\
✅ Complete — {len(files_done)}/{len(months)} months downloaded")
        g2b_log(f"   Saved to: {gstin_dir}")

        try: context.close(); g2b_log("?? Browser closed")
        except Exception as ce: g2b_log(f"  ⚠ Browser close: {ce}")
    except Exception as _e:
        g2b_set({"status":"error","error":str(_e)})
        g2b_log(f"✗ Fatal worker error: {_e}", "error")
        import traceback; g2b_log(traceback.format_exc(), "error")
    finally:
        try: pw.stop()
        except Exception: pass

    time.sleep(8)
    with g2b_lock:
        if g2b_state.get("status") == "done":
            g2b_state["status"] = "idle"
    g2b_log("?? RPA reset to idle — ready for next download")


# ══════════════════════════════════════════════════════════════════════════════
#  GSTR-2B → EXCEL CONVERTER
#  9 Sheets:
#    1. Summary              — month-wise totals + ITC Eligibility Analysis
#    2. B2B Purchases        — b2b+b2ba with itcavl/reason/diffprcnt
#    3. Credit-Debit Notes   — cdnr+cdnra (KEY: 2B uses 'cdnr' not 'cdn')
#    4. E-Commerce (ECO)     — eco+ecoa  (NEW — GSTR-2B only)
#    5. Import of Goods      — impg+impgsez
#    6. ISD Credits          — isd+isda  (flat structure)
#    7. TCS Credits          — tcs (Tax Collected at Source by ECO)
#    8. TDS Credits          — tds (Tax Deducted at Source u/s 51)
#    9. Consolidated         — all docs, ITC eligibility colour-coded
# ══════════════════════════════════════════════════════════════════════════════
def g2b_json_to_excel(gstin, fy):
    """
    Convert all GSTR-2B JSON files for a GSTIN/FY to a multi-sheet Excel workbook.

    2025-26 Portal JSON Schema (docdata flat format):
      raw → raw["data"] → data keys: gstin, rtnprd, gendt, version, docdata, itcsumm, cpsumm
      data["docdata"] keys: b2b, cdnr, ecom   (sections in downloaded JSON)

      ECO supplier entry: {ctin, trdnm, supfildt, supprd, inv:[...]}
      ECO invoice:        {inum, dt, val, txval, igst, cgst, sgst, cess,
                           rev, itcavl, rsn, typ, pos, imsStatus}
                           NOTE: Same flat structure as B2B. No irn/srctyp.

      B2B supplier entry: {ctin, trdnm, supfildt, supprd, inv:[...]}
      B2B invoice:        {inum, dt, val, txval, igst, cgst, sgst, cess,
                           rev, itcavl, rsn, typ, pos, imsStatus}
                           NOTE: NO itms / itm_det nesting — taxes are FLAT

      CDNR supplier entry: {ctin, trdnm, supfildt, supprd, nt:[...]}
      CDNR note:           {ntnum, dt, val, txval, igst, cgst, sgst, cess,
                            rev, itcavl, rsn, typ, pos, suptyp, imsStatus}
                            typ: "C"=Credit / "D"=Debit
                            suptyp: "R"=Regular / "SEWP"/"SEWOP"/etc.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    import glob, io as _io

    fy_dir     = fy.replace("-","_")
    folder     = os.path.join(G2B_DOWNLOAD_DIR, gstin, fy_dir)
    json_files = sorted(glob.glob(os.path.join(folder, "*.json")))
    if not json_files:
        return None, f"No GSTR-2B JSON files found in {folder}"

    # ── Styles ────────────────────────────────────────────────────────────────
    HDR_FILL   = PatternFill("solid", fgColor="1A237E")   # deep indigo
    ALT_FILL   = PatternFill("solid", fgColor="EEF2FF")
    TOT_FILL   = PatternFill("solid", fgColor="C5CAE9")
    ITC_Y_FILL = PatternFill("solid", fgColor="E8F5E9")   # ITC Available
    ITC_N_FILL = PatternFill("solid", fgColor="FFEBEE")   # ITC Blocked
    ITC_R_FILL = PatternFill("solid", fgColor="FFF8E1")   # Reverse charge
    AMEND_FILL = PatternFill("solid", fgColor="FFF3E0")
    CDN_CR     = PatternFill("solid", fgColor="FCE4EC")
    CDN_DR     = PatternFill("solid", fgColor="F1F8E9")
    IMS_P_FILL = PatternFill("solid", fgColor="FFF9C4")   # IMS Pending (yellow)
    IMS_R_FILL = PatternFill("solid", fgColor="FFEEDD")   # IMS Rejected

    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    BODY_FONT = Font(name="Arial", size=9)
    TOT_FONT  = Font(name="Arial", bold=True, size=9, color="1A237E")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    thin      = Side(style="thin", color="9FA8DA")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM_FMT   = '#,##0.00'

    def make_title(ws, ncols, text, fg="1A237E"):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value     = text
        c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=fg)
        c.alignment = CENTER
        ws.row_dimensions[1].height = 26

    def make_hdr(ws, labels, row=2, fills=None):
        for ci, lbl in enumerate(labels, 1):
            cell = ws.cell(row=row, column=ci, value=lbl)
            cell.font      = HDR_FONT
            cell.alignment = CENTER
            cell.border    = BORDER
            cell.fill      = (fills[ci-1] if fills and ci-1 < len(fills) else HDR_FILL)
        ws.row_dimensions[row].height = 32

    def wrow(ws, r, vals, fill=None):
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.font   = BODY_FONT
            cell.border = BORDER
            if fill: cell.fill = fill
            if isinstance(v, (int, float)):
                cell.alignment    = RIGHT
                cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    def trow(ws, r, ncols, sum_cols, ds=3):
        """Write totals row with Python-computed sums (no =SUM formulas)."""
        sums = {}
        for ci in sum_cols:
            total = 0.0
            for dr in range(ds, r):
                v = ws.cell(row=dr, column=ci).value
                if isinstance(v, (int, float)):
                    total += v
            sums[ci] = total
        for ci in range(1, ncols+1):
            cell = ws.cell(row=r, column=ci)
            cell.fill   = TOT_FILL
            cell.font   = TOT_FONT
            cell.border = BORDER
            if ci in sums:
                cell.value         = sums[ci]
                cell.alignment     = RIGHT
                cell.number_format = NUM_FMT
            elif isinstance(cell.value, (int, float)):
                cell.alignment     = RIGHT
                cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ── Reference data ────────────────────────────────────────────────────────
    STATE_CODES = {
        "01":"Jammu & Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
        "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"UP",
        "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
        "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
        "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh",
        "24":"Gujarat","27":"Maharashtra","29":"Karnataka","30":"Goa","32":"Kerala",
        "33":"Tamil Nadu","34":"Puducherry","36":"Telangana","37":"Andhra Pradesh",
        "38":"Ladakh","97":"Other Territory","99":"Centre",
    }
    def pos_disp(code):
        s = str(code).zfill(2)
        return f"{s}-{STATE_CODES.get(s, code)}"

    INV_TYPES = {"R":"Regular","SEWP":"SEZ w/ Payment","SEWOP":"SEZ w/o Payment",
                 "DE":"Deemed Export","CBW":"Customs Bonded Warehouse"}
    NOTE_TYPES = {"C":"Credit Note","D":"Debit Note","R":"Refund Voucher"}
    IMS_STATUS = {"A":"Accepted ✓","P":"Pending ⏳","R":"Rejected ✗","":"—"}

    MONTH_ORDER = ["April","May","June","July","August","September",
                   "October","November","December","January","February","March"]
    def mon_key(m):
        parts = m.split()
        try: return MONTH_ORDER.index(parts[0]) if parts else 99
        except ValueError: return 99

    def fp_to_display(fp):
        MN={"01":"January","02":"February","03":"March","04":"April","05":"May",
            "06":"June","07":"July","08":"August","09":"September","10":"October",
            "11":"November","12":"December"}
        if fp and len(fp)>=6:
            return f"{MN.get(fp[:2],fp[:2])} {fp[2:6]}"
        return fp or "Unknown"

    # ── Load and unwrap all JSON files ────────────────────────────────────────
    months_data = []   # list of (mdisp, docdata_dict, meta_dict)
    for fpath in json_files:
        try:
            with open(fpath, encoding="utf-8") as f:
                raw = json.load(f)
            data = g2b_unwrap(raw)          # handles flat / data / data.docdata wrapping

            # In new schema data has 'docdata' subkey; in old it's flat
            docdata = data.get("docdata", data)   # prefer docdata, fallback flat
            meta = {
                "gstin":    data.get("gstin",""),
                "rtnprd":   data.get("rtnprd",""),
                "gendt":    data.get("gendt",""),
                "itcsumm":  data.get("itcsumm",{}),
                "version": data.get("version",""),
            }
            period  = meta["rtnprd"] or data.get("fp","")
            fname   = os.path.basename(fpath)
            mdisp   = (fname.replace("GSTR2B_","").replace(".json","")
                           .replace("_"," ").strip())
            if not mdisp or mdisp == "Unknown":
                mdisp = fp_to_display(period)
            months_data.append((mdisp, docdata, meta))
        except Exception as e:
            g2b_log(f"  ⚠ Skipping {fpath}: {e}")

    if not months_data:
        return None, "No valid GSTR-2B JSON data found"

    months_data.sort(key=lambda x: mon_key(x[0]))

    # ── Load enriched legal names for supplementing G2B trdnm ────────────────
    # G2B JSON already has trdnm (trade name). The enriched store adds legal_name.
    _g2b_name_lookup = _gnames_load_for("2b")   # {GSTIN: {trade_name, legal_name}}

    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — B2B PURCHASES
    # New flat schema: no itms/itm_det nesting; taxes directly on invoice
    # New fields: trdnm (trade name), supfildt (supplier file date),
    #             supprd (supplier period), imsStatus (IMS action)
    # ══════════════════════════════════════════════════════════════════════════
    ws_b2b = wb.create_sheet("B2B Purchases")
    B2B_H = [
        "Month","Gen Date",
        # Supplier info
        "Supplier GSTIN","Supplier Trade Name","Supplier Legal Name","Supplier Filed Date","Supplier Period",
        # Invoice identity
        "Invoice No","Invoice Date","Invoice Value","Place of Supply",
        # Invoice type / compliance
        "Inv Type","Rev. Charge","IMS Status",
        # Tax (flat — no itms)
        "Taxable Value","IGST","CGST","SGST","Cess","Total Tax",
        # ITC
        "ITC Available","Reason (if Blocked)",
        # E-Invoice details
        "Source Type","IRN","IRN Gen Date",
    ]
    B2B_W = [14,12, 20,28,28,16,12, 22,12,14,22, 14,10,14, 14,12,12,12,10,12, 14,30, 14,64,14]

    # Zone-colour header fills
    Z_ID  = PatternFill("solid", fgColor="1A237E")
    Z_SUP = PatternFill("solid", fgColor="283593")
    Z_FIN = PatternFill("solid", fgColor="1565C0")
    Z_ITC = PatternFill("solid", fgColor="1B5E20")
    B2B_FILLS = ([Z_ID]*2 + [Z_SUP]*4 + [Z_ID]*3 + [Z_FIN]*2 +
                 [Z_ID]*1 + [Z_FIN]*1 + [Z_FIN]*6 + [Z_ITC]*2)

    make_title(ws_b2b, len(B2B_H),
        f"GSTR-2B | B2B Purchase Register | GSTIN: {gstin} | FY: {fy}  "
        f"|  Green=ITC Available  ·  Red=ITC Blocked  ·  Amber=Rev Charge  "
        f"·  Yellow=IMS Pending  ·  Orange=IMS Rejected", "1A237E")
    make_hdr(ws_b2b, B2B_H, fills=B2B_FILLS)
    ws_b2b.freeze_panes = "G3"

    b2b_rows = []   # raw row data for Summary and Consolidated sheets
    for mdisp, docdata, meta in months_data:
        gen_dt = meta.get("gendt","")
        for sup in docdata.get("b2b",[]):
            ctin      = sup.get("ctin","")
            trdnm     = sup.get("trdnm","")
            supfildt  = sup.get("supfildt","")
            supprd    = sup.get("supprd","")
            for inv in sup.get("inv",[]):
                inum       = inv.get("inum","")
                dt         = inv.get("dt","")
                val        = inv.get("val",0) or 0
                pos        = pos_disp(inv.get("pos",""))
                ityp       = INV_TYPES.get(inv.get("typ","R"), inv.get("typ","R"))
                rc         = "Yes" if inv.get("rev","N") == "Y" else "No"
                itcavl     = inv.get("itcavl","")
                rsn        = inv.get("rsn","")
                ims        = IMS_STATUS.get(inv.get("imsStatus",""), inv.get("imsStatus",""))
                txval      = inv.get("txval",0) or 0
                igst       = inv.get("igst",0)  or 0
                cgst       = inv.get("cgst",0)  or 0
                sgst       = inv.get("sgst",0)  or 0
                cess       = inv.get("cess",0)  or 0
                ttax       = igst+cgst+sgst+cess
                # E-Invoice details — only present on e-invoiced documents
                irn        = inv.get("irn","") or ""
                irngendate = inv.get("irngendate","") or ""
                srctyp     = inv.get("srctyp","") or ""
                itc_d    = ("Available ✓" if itcavl=="Y" else
                            f"Blocked ✗{(' — '+rsn) if rsn else ''}" if itcavl=="N"
                            else itcavl or "—")
                # Row fill based on ITC and IMS status
                ims_raw  = inv.get("imsStatus","")
                if   itcavl == "N":       fill = ITC_N_FILL
                elif inv.get("rev")=="Y": fill = ITC_R_FILL
                elif ims_raw == "P":      fill = IMS_P_FILL
                elif ims_raw == "R":      fill = IMS_R_FILL
                elif itcavl == "Y":       fill = ITC_Y_FILL
                else:                     fill = ALT_FILL
                # Supplement trdnm from JSON with legal name from enriched store
                _g2b_nl   = _g2b_name_lookup.get(ctin.strip().upper(), {})
                _g2b_lgnm = _g2b_nl.get("legal_name", "")
                # If enriched has a better trade name (from portal), use it; else keep JSON trdnm
                _g2b_trdnm = _g2b_nl.get("trade_name") or trdnm
                row_vals = [
                    mdisp, gen_dt,
                    ctin, _g2b_trdnm, _g2b_lgnm, supfildt, supprd,
                    inum, dt, val, pos,
                    ityp, rc, ims,
                    txval, igst, cgst, sgst, cess, ttax,
                    itc_d, rsn,
                    srctyp, irn, irngendate,
                    # hidden raw flags (index 24+) for summary computation
                    itcavl, inv.get("rev","N"),
                ]
                b2b_rows.append((row_vals, fill))

    WRITTEN_B2B = len(B2B_H)
    for i, (rv, fill) in enumerate(b2b_rows):
        wrow(ws_b2b, i+3, rv[:WRITTEN_B2B], fill)
    tr_b2b = len(b2b_rows)+3
    ws_b2b.cell(row=tr_b2b, column=1, value="TOTAL")
    ws_b2b.cell(row=tr_b2b, column=2, value=f"{len(b2b_rows)} invoices")
    trow(ws_b2b, tr_b2b, WRITTEN_B2B, [10,15,16,17,18,19,20])
    set_widths(ws_b2b, B2B_W)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — CREDIT / DEBIT NOTES  (cdnr)
    # New fields: ntnum (was nt_num), dt (was nt_dt), typ C/D,
    #             suptyp (supply type of original invoice), imsStatus
    # ══════════════════════════════════════════════════════════════════════════
    # ══ E-COMMERCE (ECO) SHEET — GSTR-2B: Inward Supplies via E-Commerce Operators
    # Structure identical to B2B (flat taxes, same invoice fields)
    # No irn/srctyp — ECO invoices not on e-invoice network
    ws_eco = wb.create_sheet("E-Commerce (ECO)")
    ECO_H = [
        "Month","Gen Date",
        "ECO GSTIN","Supplier Trade Name","Supplier Legal Name",
        "Supplier Filed Date","Supplier Period",
        "Invoice No","Invoice Date","Invoice Value","Place of Supply",
        "Inv Type","Rev. Charge","IMS Status",
        "Taxable Value","IGST","CGST","SGST","Cess","Total Tax",
        "ITC Available","Reason (if Blocked)",
    ]
    ECO_W = [14,12, 22,28,28,16,12, 22,12,14,22, 14,10,14, 14,12,12,12,10,12, 14,30]
    make_title(ws_eco, len(ECO_H),
        f"GSTR-2B | E-Commerce (ECO) Purchases | GSTIN: {gstin} | FY: {fy}  "
        f"|  Inward supplies received through E-Commerce Operators (Table 3.1)", "4A235A")
    ECO_HDR_FILLS = {
        range(1, 3):  PatternFill("solid", fgColor="4A235A"),  # month,gendt
        range(3, 8):  PatternFill("solid", fgColor="6C3483"),  # supplier
        range(8, 15): PatternFill("solid", fgColor="7D3C98"),  # invoice details
        range(15, 21):PatternFill("solid", fgColor="884EA0"),  # tax
        range(21, 23):PatternFill("solid", fgColor="A569BD"),  # ITC
    }
    for c, col in enumerate(ECO_H, 1):
        cell = ws_eco.cell(row=2, column=c, value=col)
        cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
        for rng, fill in ECO_HDR_FILLS.items():
            if c in rng: cell.fill = fill; break
    ws_eco.row_dimensions[2].height = 32
    ws_eco.freeze_panes = "C3"

    eco_rows = []   # kept for consolidated sheet
    WRITTEN_ECO = len(ECO_H)
    ECO_FILL  = PatternFill("solid", fgColor="F5EEF8")
    ALT_ECO   = PatternFill("solid", fgColor="EBE2F5")

    for mdisp, docdata, meta in months_data:
        gen_dt = meta.get("gendt","")
        for sup in docdata.get("ecom",[]):
            ctin     = sup.get("ctin","")
            trdnm    = sup.get("trdnm","")
            supfildt = sup.get("supfildt","")
            supprd   = sup.get("supprd","")
            # Legal name from enrichment lookup (same as B2B in 2B: _g2b_name_lookup)
            _eco_nl    = _g2b_name_lookup.get(ctin.strip().upper(), {})
            _eco_lgnm  = _eco_nl.get("legal_name","")
            _eco_trdnm = _eco_nl.get("trade_name") or trdnm
            for inv in sup.get("inv",[]):
                inum    = inv.get("inum","")
                dt      = inv.get("dt","")
                val     = inv.get("val",0) or 0
                pos     = pos_disp(inv.get("pos",""))
                ityp    = INV_TYPES.get(inv.get("typ","R"), inv.get("typ","R"))
                rc      = "Yes" if inv.get("rev","N") == "Y" else "No"
                itcavl  = inv.get("itcavl","")
                rsn     = inv.get("rsn","")
                ims     = IMS_STATUS.get(inv.get("imsStatus",""), inv.get("imsStatus",""))
                txval   = inv.get("txval",0) or 0
                igst    = inv.get("igst",0)  or 0
                cgst    = inv.get("cgst",0)  or 0
                sgst    = inv.get("sgst",0)  or 0
                cess    = inv.get("cess",0)  or 0
                ttax    = igst+cgst+sgst+cess
                itc_d   = ("Available ✓" if itcavl=="Y" else
                            f"Blocked ✗{(' — '+rsn) if rsn else ''}" if itcavl=="N"
                            else itcavl or "—")
                row_vals = [
                    mdisp, gen_dt,
                    ctin, _eco_trdnm, _eco_lgnm, supfildt, supprd,
                    inum, dt, val, pos,
                    ityp, rc, ims,
                    txval, igst, cgst, sgst, cess, ttax,
                    itc_d, rsn,
                    itcavl, inv.get("rev","N"),   # hidden flags (index 22+)
                ]
                eco_rows.append((row_vals, ECO_FILL))

    for i, (rv, fill) in enumerate(eco_rows):
        fill = ECO_FILL if i % 2 == 0 else ALT_ECO
        wrow(ws_eco, i+3, rv[:WRITTEN_ECO], fill)
    tr_eco = len(eco_rows)+3
    ws_eco.cell(row=tr_eco, column=1, value="TOTAL")
    ws_eco.cell(row=tr_eco, column=2, value=f"{len(eco_rows)} invoices")
    if eco_rows:
        ds=3; de=tr_eco-1
        # J=10 InvVal, O=15 Txval, P=16 IGST, Q=17 CGST, R=18 SGST, S=19 Cess, T=20 Total
        for ci, cl in [(10,"J"),(15,"O"),(16,"P"),(17,"Q"),(18,"R"),(19,"S"),(20,"T")]:
            ws_eco.cell(row=tr_eco, column=ci,
                value=(f"=SUM({cl}{ds}:{cl}{de})" if de>=ds else 0))
    trow(ws_eco, tr_eco, WRITTEN_ECO, [10,15,16,17,18,19,20])
    set_widths(ws_eco, ECO_W)

    ws_cdn = wb.create_sheet("Credit-Debit Notes")
    CDN_H = [
        "Month","Gen Date",
        "Supplier GSTIN","Supplier Trade Name","Supplier Legal Name","Supplier Filed Date","Supplier Period",
        "Note Type","Note No","Note Date","Note Value","Place of Supply",
        "Supply Type","Rev. Charge","IMS Status",
        "Taxable Value","IGST","CGST","SGST","Cess","Total Tax",
        "ITC Available","Reason (if Blocked)",
    ]
    CDN_W = [14,12, 20,28,28,16,12, 12,22,12,14,22, 14,10,14, 14,12,12,12,10,12, 14,30]
    CDN_FILLS = ([Z_ID]*2 + [Z_SUP]*4 + [Z_ID]*6 + [Z_FIN]*2 +
                 [Z_ID]*1 + [Z_FIN]*1 + [Z_FIN]*6 + [Z_ITC]*2)
    make_title(ws_cdn, len(CDN_H),
        f"GSTR-2B | Credit / Debit Note Register | GSTIN: {gstin} | FY: {fy}  "
        f"|  Credit Note values are NEGATIVE  ·  Blocked ITC shown in Red", "1A237E")
    make_hdr(ws_cdn, CDN_H, fills=CDN_FILLS)
    ws_cdn.freeze_panes = "G3"

    cdn_rows = []
    for mdisp, docdata, meta in months_data:
        gen_dt = meta.get("gendt","")
        for sup in docdata.get("cdnr",[]):
            ctin      = sup.get("ctin","")
            trdnm     = sup.get("trdnm","")
            supfildt  = sup.get("supfildt","")
            supprd    = sup.get("supprd","")
            for nt in sup.get("nt",[]):
                _typ     = nt.get("typ","C")       # "C" or "D"
                sign     = -1 if _typ == "C" else 1
                ntty_d   = NOTE_TYPES.get(_typ, _typ)
                nt_num   = nt.get("ntnum","")
                nt_dt    = nt.get("dt","")
                val      = (nt.get("val",0) or 0) * sign
                pos      = pos_disp(nt.get("pos",""))
                suptyp   = INV_TYPES.get(nt.get("suptyp","R"), nt.get("suptyp","R"))
                rc       = "Yes" if nt.get("rev","N") == "Y" else "No"
                itcavl   = nt.get("itcavl","")
                rsn      = nt.get("rsn","")
                ims      = IMS_STATUS.get(nt.get("imsStatus",""), nt.get("imsStatus",""))
                txval    = (nt.get("txval",0) or 0) * sign
                igst     = (nt.get("igst",0)  or 0) * sign
                cgst     = (nt.get("cgst",0)  or 0) * sign
                sgst     = (nt.get("sgst",0)  or 0) * sign
                cess     = (nt.get("cess",0)  or 0) * sign
                ttax      = igst+cgst+sgst+cess
                nt_irn    = nt.get("irn","") or ""
                nt_irngdt = nt.get("irngendate","") or ""
                nt_srctyp = nt.get("srctyp","") or ""
                itc_d    = ("Available ✓" if itcavl=="Y" else
                            f"Blocked ✗{(' — '+rsn) if rsn else ''}" if itcavl=="N"
                            else itcavl or "—")
                ims_raw  = nt.get("imsStatus","")
                if   itcavl == "N": fill = ITC_N_FILL
                elif ims_raw == "P":fill = IMS_P_FILL
                elif ims_raw == "R":fill = IMS_R_FILL
                elif _typ == "C":   fill = CDN_CR
                else:               fill = CDN_DR
                # Add legal name from enriched store
                _g2b_cdn_nl   = _g2b_name_lookup.get(ctin.strip().upper(), {})
                _g2b_cdn_lgnm = _g2b_cdn_nl.get("legal_name", "")
                _g2b_cdn_trdnm = _g2b_cdn_nl.get("trade_name") or trdnm
                row_vals = [
                    mdisp, gen_dt,
                    ctin, _g2b_cdn_trdnm, _g2b_cdn_lgnm, supfildt, supprd,
                    ntty_d, nt_num, nt_dt, val, pos,
                    suptyp, rc, ims,
                    txval, igst, cgst, sgst, cess, ttax,
                    itc_d, rsn,
                    nt_srctyp, nt_irn, nt_irngdt,
                    # hidden flags
                    itcavl, _typ,
                ]
                cdn_rows.append((row_vals, fill))

    WRITTEN_CDN = len(CDN_H)

    for i, (rv, fill) in enumerate(cdn_rows):
            wrow(ws_cdn, i+3, rv[:WRITTEN_CDN], fill)
    tr_cdn = len(cdn_rows)+3
    ws_cdn.cell(row=tr_cdn, column=1, value="TOTAL")
    ws_cdn.cell(row=tr_cdn, column=2,
        value=f"{sum(1 for r,_ in cdn_rows if len(r)>24 and r[24]=='C')} Credit  |  "
              f"{sum(1 for r,_ in cdn_rows if len(r)>24 and r[24]=='D')} Debit")
    trow(ws_cdn, tr_cdn, WRITTEN_CDN, [11,16,17,18,19,20,21])
    set_widths(ws_cdn, CDN_W)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — ITC SUMMARY from itcsumm (portal computed)
    # ══════════════════════════════════════════════════════════════════════════
    # ══════════════════════════════════════════════════════════════════════════
    # CONSOLIDATED SHEET — All sections merged for reconciliation
    # Combines: B2B + ECO + Credit-Debit Notes in one sheet
    # ══════════════════════════════════════════════════════════════════════════
    ws_con = wb.create_sheet("Consolidated")

    CON_H = [
        "Doc Type", "Month", "Gen Date",
        "Supplier GSTIN", "Trade Name", "Legal Name",
        "Supplier Filed Date", "Supplier Period",
        "Document No", "Document Date",
        "Document Value", "Place of Supply",
        "Inv/Note Type", "Source Type", "Rev. Charge", "IMS Status",
        "Taxable Value", "IGST", "CGST", "SGST", "Cess", "Total Tax",
        "ITC Available", "Reason (if Blocked)",
        "IRN", "IRN Gen Date", "Remarks",
    ]
    CON_W = [
        16, 14, 12,                    # A–C  Doc Type, Month, Gen Date
        22, 28, 28, 16, 12,            # D–I  Supplier info
        24, 13,                        # J–K  Doc No, Date
        14, 22, 14, 14, 10, 14,        # L–Q  Value/Type/Compliance
        14, 12, 12, 12, 10, 12,        # R–W  Tax cols
        14, 30,                        # X–Y  ITC
        64, 13, 32,                    # Z–AB IRN, IRN Date, Remarks
    ]

    ws_con.merge_cells(f"A1:{get_column_letter(len(CON_H))}1")
    t_con        = ws_con["A1"]
    t_con.value  = (f"GSTR-2B | Consolidated Register | GSTIN: {gstin} | FY: {fy}  |  "
                    "PINK = ITC Blocked  |  AMBER = ITC Reversed risk")
    t_con.font      = Font(name="Arial", bold=True, size=11, color="1B3A6B")
    t_con.alignment = CENTER
    t_con.fill      = PatternFill("solid", fgColor="DDEEFF")
    ws_con.row_dimensions[1].height = 24

    CON_ZONE = {
        range(1, 4):  PatternFill("solid", fgColor="1F3864"),   # identity cols
        range(4, 9):  PatternFill("solid", fgColor="1F5C8B"),   # supplier
        range(9, 17): PatternFill("solid", fgColor="1F5C8B"),   # doc + type
        range(17, 23):PatternFill("solid", fgColor="2E75B6"),   # tax cols
        range(23, 25):PatternFill("solid", fgColor="375623"),   # ITC
        range(25, 28):PatternFill("solid", fgColor="4A4A4A"),   # IRN + Remarks
    }
    for c, col in enumerate(CON_H, 1):
        cell = ws_con.cell(row=2, column=c, value=col)
        cell.font = HDR_FONT; cell.alignment = CENTER; cell.border = BORDER
        for rng, fill in CON_ZONE.items():
            if c in rng: cell.fill = fill; break
    ws_con.row_dimensions[2].height = 32

    # Fills for different conditions
    CON_NORM  = None                                     # default — no fill
    CON_BLKD  = PatternFill("solid", fgColor="FFCCCC")  # ITC blocked (itcavl=N)
    CON_ECO   = PatternFill("solid", fgColor="F5EEF8")  # ECO — light purple
    CON_CDN_C = PatternFill("solid", fgColor="FFF0F0")  # Credit Note — light red
    CON_CDN_D = PatternFill("solid", fgColor="F0FFF0")  # Debit Note  — light green

    con_data = []   # list of (row_values, fill)

    # ── B2B rows ──────────────────────────────────────────────────────────────
    # b2b row_vals: [0]=mdisp,[1]=gendt,[2]=ctin,[3]=trdnm,[4]=lgnm,
    #               [5]=supfildt,[6]=supprd,[7]=inum,[8]=dt,[9]=val,[10]=pos,
    #               [11]=ityp,[12]=rc,[13]=ims,[14]=txval,[15-18]=igst/cgst/sgst/cess,
    #               [19]=ttax,[20]=itc_d,[21]=rsn,[22]=srctyp,[23]=irn,[24]=irngendate
    #               [25]=itcavl_raw,[26]=rev_raw
    for rv, _fill in b2b_rows:
        itcavl_raw = rv[25] if len(rv) > 25 else ""
        fill = CON_BLKD if itcavl_raw == "N" else CON_NORM
        con_data.append(([
            "Invoice",  rv[0], rv[1],           # Doc Type, Month, Gen Date
            rv[2], rv[3], rv[4],                # GSTIN, Trade Name, Legal Name
            rv[5], rv[6],                       # Filed Date, Period
            rv[7], rv[8],                       # Doc No, Date
            rv[9],  rv[10],                     # Value, POS
            rv[11], rv[22], rv[12], rv[13],     # Inv Type, Source, Rev, IMS
            rv[14], rv[15], rv[16], rv[17], rv[18], rv[19],  # Tax cols
            rv[20], rv[21],                     # ITC, Reason
            rv[23], rv[24],                     # IRN, IRN Gen Date
            "",                                 # Remarks
        ], fill))

    # ── ECO rows ──────────────────────────────────────────────────────────────
    # eco row_vals: [0]=mdisp,[1]=gendt,[2]=ctin,[3]=trdnm,[4]=lgnm,
    #               [5]=supfildt,[6]=supprd,[7]=inum,[8]=dt,[9]=val,[10]=pos,
    #               [11]=ityp,[12]=rc,[13]=ims,[14]=txval,[15-18]=igst/cgst/sgst/cess,
    #               [19]=ttax,[20]=itc_d,[21]=rsn
    #               [22]=itcavl_raw,[23]=rev_raw
    for rv, _fill in eco_rows:
        itcavl_raw = rv[22] if len(rv) > 22 else ""
        fill = CON_BLKD if itcavl_raw == "N" else CON_ECO
        con_data.append(([
            "Invoice (ECO)", rv[0], rv[1],      # Doc Type, Month, Gen Date
            rv[2], rv[3], rv[4],                # GSTIN, Trade Name, Legal Name
            rv[5], rv[6],                       # Filed Date, Period
            rv[7], rv[8],                       # Doc No, Date
            rv[9],  rv[10],                     # Value, POS
            rv[11], "E-Commerce", rv[12], rv[13],  # Inv Type, Source(ECO), Rev, IMS
            rv[14], rv[15], rv[16], rv[17], rv[18], rv[19],  # Tax cols
            rv[20], rv[21],                     # ITC, Reason
            "", "",                             # IRN, IRN Gen Date (N/A for ECO)
            "",                                 # Remarks
        ], fill))

    # ── CDN rows ──────────────────────────────────────────────────────────────
    # cdn row_vals: [0]=mdisp,[1]=gendt,[2]=ctin,[3]=trdnm,[4]=lgnm,
    #               [5]=supfildt,[6]=supprd,[7]=ntty_d,[8]=nt_num,[9]=ntdt,
    #               [10]=val(signed),[11]=pos,[12]=suptyp,[13]=rc,[14]=ims,
    #               [15]=txval(signed),[16-19]=igst/cgst/sgst/cess(signed),
    #               [20]=ttax(signed),[21]=itc_d,[22]=rsn,
    #               [23]=srctyp,[24]=irn,[25]=irngdt
    #               [26]=itcavl_raw,[27]=_typ
    for rv, _fill in cdn_rows:
        itcavl_raw = rv[26] if len(rv) > 26 else ""
        note_typ   = rv[27] if len(rv) > 27 else "C"
        fill = CON_BLKD if itcavl_raw == "N" else (CON_CDN_C if note_typ=="C" else CON_CDN_D)
        con_data.append(([
            "Credit Note" if note_typ == "C" else "Debit Note",
            rv[0], rv[1],                       # Month, Gen Date
            rv[2], rv[3], rv[4],                # GSTIN, Trade Name, Legal Name
            rv[5], rv[6],                       # Filed Date, Period
            rv[8], rv[9],                       # Doc No (nt_num), Date (ntdt)
            rv[10], rv[11],                     # Value(signed), POS
            rv[12], rv[23], rv[13], rv[14],     # Suptyp, Source, Rev, IMS
            rv[15], rv[16], rv[17], rv[18], rv[19], rv[20],  # Tax (signed)
            rv[21], rv[22],                     # ITC, Reason
            rv[24], rv[25],                     # IRN, IRN Gen Date
            "",                                 # Remarks
        ], fill))

    # ── Write consolidated rows ───────────────────────────────────────────────
    WRITTEN_CON = len(CON_H)
    for i, (vals, fill) in enumerate(con_data):
        r = i + 3
        for c, val in enumerate(vals, 1):
            cell = ws_con.cell(row=r, column=c, value=val)
            cell.font = BODY_FONT; cell.border = BORDER
            if fill: cell.fill = fill
            if isinstance(val, (int, float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    # ── Totals row ────────────────────────────────────────────────────────────
    con_tr = len(con_data) + 3
    ws_con.cell(row=con_tr, column=1, value="TOTAL")
    ws_con.cell(row=con_tr, column=2, value=f"{len(con_data)} documents")
    if con_data:
        ds = 3; de = con_tr - 1
        # Numeric cols: K=11 DocVal, R=17 Txval, S=18 IGST, T=19 CGST,
        #               U=20 SGST, V=21 Cess, W=22 Total Tax
        for ci, cl in [(11,"K"),(17,"R"),(18,"S"),(19,"T"),(20,"U"),(21,"V"),(22,"W")]:
            ws_con.cell(row=con_tr, column=ci,
                value=(f"=SUM({cl}{ds}:{cl}{de})" if de >= ds else 0))
    trow(ws_con, con_tr, WRITTEN_CON, [])
    ws_con.row_dimensions[con_tr].height = 20
    set_widths(ws_con, CON_W)
    ws_con.freeze_panes = "D3"

    ws_itc = wb.create_sheet("ITC Summary")
    ITC_H = ["Month","Gen Date","Category","Section",
             "Taxable Value","IGST","CGST","SGST","Cess","Total Tax"]
    make_title(ws_itc, len(ITC_H),
        f"GSTR-2B | ITC Summary (Portal-Computed) | GSTIN: {gstin} | FY: {fy}  "
        f"|  Source: itcsumm block — statutory basis u/s 16(2)(aa)", "1B5E20")
    make_hdr(ws_itc, ITC_H)
    ws_itc.freeze_panes = "A3"
    itc_r = 3
    for mdisp, docdata, meta in months_data:
        # itcsumm is on the parent data level; re-read the raw file for it
        gen_dt = meta.get("gendt","")
        # itcsumm is stored in meta from the raw file
        itcsumm = meta.get("itcsumm", {})
        itcavl  = itcsumm.get("itcavl",  {}).get("nonrevsup", {})
        itcunavl = itcsumm.get("itcunavl",{}).get("nonrevsup", {})
        if not itcsumm: continue

        def _itc_row(r, label, section_d):
            """Write one ITC summary row; returns next row index."""
            if not section_d: return r
            ws_itc.cell(row=r, column=1, value=mdisp)
            ws_itc.cell(row=r, column=2, value=gen_dt)
            ws_itc.cell(row=r, column=3, value=label)
            ws_itc.cell(row=r, column=4,
                value="Available" if "avl" in label.lower() else "Unavailable")
            txv   = section_d.get("txval",0) or 0
            _igst = section_d.get("igst",0) or 0
            _cgst = section_d.get("cgst",0) or 0
            _sgst = section_d.get("sgst",0) or 0
            _cess = section_d.get("cess",0) or 0
            for ci, v in enumerate([txv,_igst,_cgst,_sgst,_cess,_igst+_cgst+_sgst+_cess],5):
                cell = ws_itc.cell(row=r, column=ci, value=v)
                cell.number_format = "#,##0.00"; cell.alignment = RIGHT
            for c in range(1, len(ITC_H)+1):
                cell = ws_itc.cell(row=r, column=c)
                cell.font = BODY_FONT; cell.border = BORDER
                if c % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F0FFF0")
            return r + 1

        itc_r = _itc_row(itc_r, "B2B — ITC Available (Non-Rev Charge)",   itcavl.get("b2b",{}))
        itc_r = _itc_row(itc_r, "ECO — ITC Available (Non-Rev Charge)",   itcavl.get("ecom",{}))
        itc_r = _itc_row(itc_r, "B2B — ITC Unavailable (Non-Rev Charge)", itcunavl.get("b2b",{}))
    ws_itc.cell(row=itc_r, column=1,
        value="ℹ ITC Summary (itcsumm) is available in the raw JSON at data.itcsumm. "
              "The B2B Purchases sheet above already colour-codes each invoice by ITC eligibility.")
    ws_itc.cell(row=itc_r, column=1).font = Font(name="Arial", italic=True, size=9)
    set_widths(ws_itc, [14,12,16,14,16,14,14,14,10,14])

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — MONTH-WISE SUMMARY  (first sheet, inserted at position 0)
    # ══════════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("Summary", 0)
    SUM_H = [
        "Month","Gen Date",
        "B2B Invoices","B2B Taxable","B2B IGST","B2B CGST","B2B SGST","B2B Cess","B2B Total Tax",
        "CDN Notes","CDN Taxable","CDN IGST","CDN CGST","CDN SGST","CDN Cess","CDN Total Tax",
        "ITC Available (Tax)","ITC Blocked (Tax)","ITC Rev.Charge (Tax)",
        "IMS Accepted","IMS Pending","IMS Rejected",
    ]
    SUM_W = [16,12,
             12,16,14,14,14,10,14,
             10,16,14,14,14,10,14,
             16,16,16, 14,12,12]
    make_title(ws_sum, len(SUM_H),
        f"GSTR-2B | Month-wise Summary & ITC Analysis | GSTIN: {gstin} | FY: {fy}",
        fg="1A237E")
    SUM_FILLS = (
        [Z_ID]*2 + [Z_FIN]*7 + [Z_FIN]*7 + [Z_ITC]*3 +
        [PatternFill("solid",fgColor="1B5E20"),
         PatternFill("solid",fgColor="F57F17"),
         PatternFill("solid",fgColor="B71C1C")]
    )
    make_hdr(ws_sum, SUM_H, fills=SUM_FILLS)
    ws_sum.freeze_panes = "C3"

    summary = defaultdict(lambda: dict(
        gen_dt="",
        b2b_inv=0,b2b_txval=0.0,b2b_igst=0.0,b2b_cgst=0.0,b2b_sgst=0.0,b2b_cess=0.0,
        cdn_nt=0,cdn_txval=0.0,cdn_igst=0.0,cdn_cgst=0.0,cdn_sgst=0.0,cdn_cess=0.0,
        itc_avl=0.0,itc_blk=0.0,itc_rc=0.0,
        ims_a=0,ims_p=0,ims_r=0,
    ))

    for rv, _ in b2b_rows:
        m        = rv[0]
        itcavl   = rv[22] if len(rv)>22 else ""
        rc_raw   = rv[23] if len(rv)>23 else "N"
        ims_raw  = rv[13]   # "Accepted ✓" / "Pending ⏳" / "Rejected ✗"
        if not summary[m]["gen_dt"]: summary[m]["gen_dt"] = rv[1]
        summary[m]["b2b_inv"]   += 1
        summary[m]["b2b_txval"] += rv[14] or 0
        summary[m]["b2b_igst"]  += rv[15] or 0
        summary[m]["b2b_cgst"]  += rv[16] or 0
        summary[m]["b2b_sgst"]  += rv[17] or 0
        summary[m]["b2b_cess"]  += rv[18] or 0
        ttax = (rv[15] or 0)+(rv[16] or 0)+(rv[17] or 0)+(rv[18] or 0)
        if   itcavl=="Y" and rc_raw!="Y": summary[m]["itc_avl"] += ttax
        elif rc_raw=="Y":                  summary[m]["itc_rc"]  += ttax
        elif itcavl=="N":                  summary[m]["itc_blk"] += ttax
        if   "Accepted" in str(ims_raw):   summary[m]["ims_a"]  += 1
        elif "Pending"  in str(ims_raw):   summary[m]["ims_p"]  += 1
        elif "Rejected" in str(ims_raw):   summary[m]["ims_r"]  += 1

    for rv, _ in cdn_rows:
        m = rv[0]
        summary[m]["cdn_nt"]    += 1
        summary[m]["cdn_txval"] += rv[15] or 0
        summary[m]["cdn_igst"]  += rv[16] or 0
        summary[m]["cdn_cgst"]  += rv[17] or 0
        summary[m]["cdn_sgst"]  += rv[18] or 0
        summary[m]["cdn_cess"]  += rv[19] or 0

    sorted_months = sorted(summary.keys(), key=mon_key)
    for i, month in enumerate(sorted_months):
        s = summary[month]
        r = i+3
        b2b_ttax = s["b2b_igst"]+s["b2b_cgst"]+s["b2b_sgst"]+s["b2b_cess"]
        cdn_ttax = s["cdn_igst"]+s["cdn_cgst"]+s["cdn_sgst"]+s["cdn_cess"]
        wrow(ws_sum, r, [
            month, s["gen_dt"],
            s["b2b_inv"],s["b2b_txval"],s["b2b_igst"],s["b2b_cgst"],s["b2b_sgst"],s["b2b_cess"],b2b_ttax,
            s["cdn_nt"],s["cdn_txval"],s["cdn_igst"],s["cdn_cgst"],s["cdn_sgst"],s["cdn_cess"],cdn_ttax,
            s["itc_avl"],s["itc_blk"],s["itc_rc"],
            s["ims_a"],s["ims_p"],s["ims_r"],
        ], ALT_FILL if i%2==1 else None)
        # Colour ITC columns
        for ci, fi in [(17,ITC_Y_FILL),(18,ITC_N_FILL),(19,ITC_R_FILL)]:
            c = ws_sum.cell(row=r, column=ci)
            if isinstance(c.value,(int,float)) and c.value > 0: c.fill = fi

    gtr = len(sorted_months)+3
    ws_sum.cell(row=gtr, column=1, value="GRAND TOTAL")
    # Python-computed grand totals
    for ci in range(3, len(SUM_H)+1):
        total = sum(
            ws_sum.cell(row=dr, column=ci).value or 0
            for dr in range(3, gtr)
            if isinstance(ws_sum.cell(row=dr, column=ci).value, (int, float))
        )
        ws_sum.cell(row=gtr, column=ci, value=total)
    trow(ws_sum, gtr, len(SUM_H), list(range(3, len(SUM_H)+1)), ds=3)
    set_widths(ws_sum, SUM_W)

    # ── Colour Legend ─────────────────────────────────────────────────────────
    leg_r = gtr + 2
    ws_sum.merge_cells(f"A{leg_r}:E{leg_r}")
    ws_sum.cell(row=leg_r, column=1, value="COLOUR LEGEND").font = \
        Font(name="Arial", bold=True, size=9)
    for idx, (lbl, color) in enumerate([
        ("ITC Available — claim in GSTR-3B 4A(5)",           "E8F5E9"),
        ("ITC Blocked — DO NOT claim (supplier non-filer)",   "FFEBEE"),
        ("Reverse Charge — pay under RCM first (Sec 9(3))",   "FFF8E1"),
        ("IMS Status: Pending — take action on GST portal",   "FFF9C4"),
        ("IMS Status: Rejected by you",                       "FFEEDD"),
        ("Credit Note (values shown as NEGATIVE)",            "FCE4EC"),
        ("Debit Note (values positive)",                      "F1F8E9"),
    ]):
        lr = leg_r+idx+1
        c = ws_sum.cell(row=lr, column=1, value=lbl)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(name="Arial", size=9); c.border = BORDER; c.alignment = LEFT
        for col in range(2,6):
            d = ws_sum.cell(row=lr, column=col)
            d.fill = PatternFill("solid", fgColor=color); d.border = BORDER

    # ── ITC Analysis block ────────────────────────────────────────────────────
    itc_blk_r = leg_r + 10
    ws_sum.merge_cells(f"A{itc_blk_r}:{get_column_letter(len(SUM_H))}{itc_blk_r}")
    ih = ws_sum.cell(row=itc_blk_r, column=1,
        value="💡  GSTR-2B ITC ELIGIBILITY — Statutory Basis u/s 16(2)(aa) CGST Act")
    ih.font  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    ih.fill  = PatternFill("solid", fgColor="1B5E20")
    ih.alignment = CENTER
    ws_sum.row_dimensions[itc_blk_r].height = 28

    total_avl = sum(summary[m]["itc_avl"] for m in summary)
    total_blk = sum(summary[m]["itc_blk"] for m in summary)
    total_rc  = sum(summary[m]["itc_rc"]  for m in summary)
    hdr_r2 = itc_blk_r+1
    for ci, lbl in enumerate(["ITC Category","Total Tax (₹)","Invoices","Action"],1):
        cell = ws_sum.cell(row=hdr_r2, column=ci, value=lbl)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill = PatternFill("solid", fgColor="2E7D32")
        cell.alignment = CENTER; cell.border = BORDER

    # rv[22]=itcavl raw flag, rv[23]=rev raw flag (+1 shift after Legal Name at idx 4)
    itc_table = [
        ("✅ ITC AVAILABLE (100%) — itcavl='Y'", total_avl,
         sum(1 for rv,_ in b2b_rows if len(rv)>22 and rv[22]=="Y"),
         "Claim in GSTR-3B Table 4A(5)", ITC_Y_FILL),
        ("❌ ITC BLOCKED — itcavl='N' (supplier not filed)", total_blk,
         sum(1 for rv,_ in b2b_rows if len(rv)>22 and rv[22]=="N"),
         "DO NOT claim — await supplier compliance", ITC_N_FILL),
        ("⚠ REVERSE CHARGE — rchrg='Y'", total_rc,
         sum(1 for rv,_ in b2b_rows if len(rv)>23 and rv[23]=="Y"),
         "Pay GST under RCM first, then claim in 4A(3)", ITC_R_FILL),
    ]
    for idx2,(lbl,amt,cnt,action,fi) in enumerate(itc_table):
        tr2 = hdr_r2+idx2+1
        for ci2, val in enumerate([lbl,amt,cnt,action],1):
            cell = ws_sum.cell(row=tr2, column=ci2, value=val)
            cell.fill = fi; cell.border = BORDER
            cell.font = Font(name="Arial", size=9)
            if isinstance(val,(int,float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    legal_r = hdr_r2 + len(itc_table) + 2
    ws_sum.merge_cells(f"A{legal_r}:{get_column_letter(len(SUM_H))}{legal_r}")
    leg = ws_sum.cell(row=legal_r, column=1,
        value=("⚖ LEGAL NOTE: GSTR-2B is a STATIC monthly register (frozen on ~14th of following month). "
               "ITC u/s 16(2)(aa) claimable ONLY for invoices in GSTR-2B with itcavl='Y'. "
               "IMS (Invoice Management System): Accept/Reject invoices on GST portal before filing GSTR-3B. "
               "TDS/TCS credits flow to Electronic Cash Ledger — NOT ITC Ledger."))
    leg.font = Font(name="Arial", italic=True, size=8.5, color="1A237E")
    leg.fill = PatternFill("solid", fgColor="EEF2FF")
    leg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    leg.border = BORDER
    ws_sum.row_dimensions[legal_r].height = 36

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf, None




# ── GSTR-2B Flask Routes ────────────────────────────────────────────────────
@app.route("/g2b/start", methods=["POST"])
def g2b_start():
    data     = request.get_json(force=True) or {}
    gstin    = (data.get("gstin")    or "").strip().upper()
    fy       = (data.get("fy")       or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not all([gstin, fy, username, password]):
        return jsonify({"error":"gstin, fy, username and password required"}), 400
    specific_month = (data.get("specific_month") or "").strip()

    with g2b_lock:
        if g2b_state["status"] in ("running","waiting_captcha","waiting_otp","downloading"):
            return jsonify({"error":"Already running — reset first"}), 409
        g2b_state.update({
            "status":"running","log":[],"error":None,
            "captcha_image":None,"captcha_answer":None,"otp_answer":None,
            "progress":0,"current_month":None,"total_months":0,
            "done_months":0,"files":[],"gstin":gstin,"fy":fy,
            "specific_month":specific_month or None,
        })
    threading.Thread(target=g2b_worker,
                     args=(gstin,fy,username,password,specific_month), daemon=True).start()
    return jsonify({"ok":True,"specific_month":specific_month or None})


@app.route("/g2b/state")
def g2b_get_state():
    with g2b_lock:
        return jsonify({
            "status":         g2b_state["status"],
            "log":            g2b_state["log"][-40:],
            "error":          g2b_state["error"],
            "progress":       g2b_state["progress"],
            "current_month":  g2b_state["current_month"],
            "total_months":   g2b_state["total_months"],
            "done_months":    g2b_state["done_months"],
            "files":          g2b_state["files"],
            "gstin":          g2b_state["gstin"],
            "fy":             g2b_state["fy"],
            "specific_month": g2b_state.get("specific_month"),
            "has_captcha":    g2b_state["captcha_image"] is not None,
        })


@app.route("/g2b/captcha_image")
def g2b_captcha_image():
    with g2b_lock:
        img = g2b_state.get("captcha_image")
    if not img:
        return jsonify({"error":"no captcha available"}), 404
    return jsonify({"image": img})


@app.route("/g2b/submit_captcha", methods=["POST"])
def g2b_submit_captcha():
    data   = request.get_json(force=True) or {}
    answer = (data.get("captcha") or "").strip()
    if not answer:
        return jsonify({"error":"captcha required"}), 400
    with g2b_lock:
        g2b_state["captcha_answer"] = answer
        g2b_state["captcha_image"]  = None
    return jsonify({"ok":True})


@app.route("/g2b/submit_otp", methods=["POST"])
def g2b_submit_otp():
    data = request.get_json(force=True) or {}
    otp  = (data.get("otp") or "").strip()
    if not otp:
        return jsonify({"error":"otp required"}), 400
    with g2b_lock:
        g2b_state["otp_answer"] = otp
    return jsonify({"ok":True})


@app.route("/g2b/export_excel", methods=["GET","POST"])
def g2b_export_excel():
    try:
        if request.method == "POST":
            body  = request.get_json(force=True) or {}
            gstin = (body.get("gstin","") or "").strip().upper()
            fy    = (body.get("fy",   "") or "").strip()
        else:
            gstin = request.args.get("gstin","").strip().upper()
            fy    = request.args.get("fy",   "").strip()
        if not gstin or not fy:
            return jsonify({"error":"gstin and fy required"}), 400
        try: import openpyxl
        except ImportError:
            return jsonify({"error":"openpyxl not installed. Run: pip install openpyxl"}), 500
        buf, err = g2b_json_to_excel(gstin, fy)
        if err:
            return jsonify({"error":err}), 404
        fname = f"GSTR2B_{gstin}_{fy.replace('-','_')}.xlsx"
        save_dir = os.path.join(G2B_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
        return save_excel_and_respond(buf, save_dir, fname, log_fn=g2b_log)
    except Exception as e:
        import traceback
        g2b_log(f"export_excel error: {e} | {traceback.format_exc()}", "error")
        return jsonify({"error":str(e)}), 500


@app.route("/g2b/reset", methods=["POST"])
def g2b_reset():
    g2b_set({
        "status":"idle","log":[],"error":None,
        "captcha_image":None,"captcha_answer":None,"otp_answer":None,
        "progress":0,"current_month":None,"total_months":0,
        "done_months":0,"files":[],"gstin":None,"fy":None,
        "specific_month":None,
    })
    return jsonify({"ok":True})


@app.route("/g2b/files")
def g2b_list_files():
    files = []
    if os.path.isdir(G2B_DOWNLOAD_DIR):
        for gstin_dir in sorted(os.listdir(G2B_DOWNLOAD_DIR)):
            gp = os.path.join(G2B_DOWNLOAD_DIR, gstin_dir)
            if not os.path.isdir(gp): continue
            for fy_dir in sorted(os.listdir(gp), reverse=True):
                fp = os.path.join(gp, fy_dir)
                if not os.path.isdir(fp): continue
                for fname in sorted(os.listdir(fp)):
                    fpath = os.path.join(fp, fname)
                    files.append({
                        "gstin":    gstin_dir,
                        "fy":       fy_dir.replace("_","-"),
                        "filename": fname,
                        "size_kb":  max(1, os.path.getsize(fpath)//1024),
                        "url":      f"/g2b/file/{gstin_dir}/{fy_dir}/{fname}",
                    })
    return jsonify({"files":files})


@app.route("/g2b/file/<gstin>/<fy_dir>/<fname>")
def g2b_serve_file(gstin, fy_dir, fname):
    fpath = os.path.join(G2B_DOWNLOAD_DIR, gstin, fy_dir, fname)
    if not os.path.isfile(fpath):
        return jsonify({"error":"File not found"}), 404
    return send_file(fpath, as_attachment=True, download_name=fname)




# ══════════════════════════════════════════════════════════════════════════════
#  GSTR-1 MODULE   (Outward Supplies — Sales Return)
# ══════════════════════════════════════════════════════════════════════════════
#
#  STATUTORY BACKGROUND
#  ─────────────────────
#  GSTR-1 is the statement of outward (sales) supplies filed by every registered
#  supplier u/s 37 CGST Act 2017.
#  • Monthly filers  : due by 11th of the following month
#  • Quarterly (QRMP): due by 13th of month following the quarter end
#    (Apr-Jun → 13-Jul, Jul-Sep → 13-Oct, Oct-Dec → 13-Jan, Jan-Mar → 13-Apr)
#  • IFF (Invoice Furnishing Facility): first 2 months of a quarter for QRMP
#    taxpayers (due 13th of succeeding month)
#
#  HOW TO DOWNLOAD FILED GSTR-1 JSON FROM PORTAL
#  ───────────────────────────────────────────────
#  Login → Services → Returns → Returns Dashboard
#    → Select FY + Tax Period → SEARCH
#    → GSTR-1 tile → click "PREPARE OFFLINE"
#    → Navigates to GSTR-1 prepare offline page
#    → Click "Download" tab
#    → Click "GENERATE FILE" (server generates ZIP — may take up to 20 min)
#    → "Click Here" download link appears → click → ZIP contains JSON
#
#  URL PATTERNS (confirmed from portal):
#    Dashboard     : return.gst.gov.in/returns/auth/dashboard
#    GSTR-1 Offline: *gstr1*prepoffline* OR *gstr1*offline*
#
#  JSON SCHEMA (top-level keys, confirmed from GST docs + API references)
#  ──────────────────────────────────────────────────────────────────────
#  {
#    "gstin": "29AAAAA0000A1Z5",
#    "fp":    "032025",           ← MMYYYY (same as 2A; NOT 'rtnprd' like 2B)
#    "gt":    5000000.00,         ← Gross Turnover (FY total)
#    "cur_gt": 400000.00,         ← Current month turnover
#    "b2b":  [...],  ← B2B: registered buyers (Table 4A/4B/4C/6B/6C)
#    "b2cl": [...],  ← B2CL: inter-state, unregistered, invoice val > ₹2.5L (Table 5A)
#    "b2cs": [...],  ← B2CS: intra-state unregistered + small inter-state (Table 7, rate-wise)
#    "cdnr": [...],  ← Credit/Debit Notes — Registered (Table 9B)
#    "cdnur":[...],  ← Credit/Debit Notes — Unregistered (Table 9B)
#    "exp":  [...],  ← Exports (Table 6A)
#    "nil":  {...},  ← Nil/Exempt/Non-GST supplies (Table 8A/8B/8C/8D)
#    "hsn":  {...},  ← HSN Summary (Table 12) — mandatory from Jan 2025 (Phase 3)
#    "docs": {...},  ← Document Issued Summary (Table 13)
#    "at":   [...],  ← Advance Tax received (Table 11A)
#    "ata":  [...],  ← Advance Tax received — amendment (Table 11A1)
#    "txp":  [...],  ← Tax on advance payment (Table 11B)
#    "txpa": [...],  ← Tax on advance — amendment (Table 11B1)
#    "b2ba": [...],  ← B2B amendments (Table 9A)
#    "b2cla":[...],  ← B2CL amendments (Table 9A)
#    "b2csa":[...],  ← B2CS amendments (Table 10)
#    "cdnra":[...],  ← CDN Registered amendments (Table 9C)
#    "cdnura":[...], ← CDN Unregistered amendments (Table 9C)
#    "expa": [...]   ← Export amendments (Table 9A)
#  }
#
#  KEY B2B INNER STRUCTURE:
#  { "ctin":"buyer_gstin", "inv":[{
#      "inum":"INV001", "idt":"DD-MM-YYYY", "val":118000,
#      "pos":"27", "rchrg":"N", "inv_typ":"R",
#      "itms":[{"num":1,"itm_det":{"rt":18,"txval":100000,
#               "iamt":18000,"camt":0,"samt":0,"csamt":0}}]
#  }]}
#
#  MULTI-PASS SWEEP (mirrors g2a pattern)
#  ───────────────────────────────────────
#  Pass 1 — "Trigger": For each month, navigate to GSTR-1 offline page,
#            check for existing download link. If none, click GENERATE FILE.
#  Wait 5 min, then Pass 2+ — "Harvest": Revisit each undownloaded month,
#            check if link appeared, download immediately if ready.
#  Repeat until all done or 30-min deadline.
# ══════════════════════════════════════════════════════════════════════════════

# ── Module globals ─────────────────────────────────────────────────────────────
g1_lock         = threading.Lock()
G1_DOWNLOAD_DIR = PATHS.gstr1_dir
g1_state = {
    "status":         "idle",
    "log":            [],
    "error":          None,
    "progress":       0,
    "current_month":  None,
    "total_months":   0,
    "done_months":    0,
    "files":          [],
    "gstin":          None,
    "fy":             None,
    "specific_month": None,
    "captcha_image":  None,
    "captcha_answer": None,
    "otp_answer":     None,
}


def g1_log(msg, level="info"):
    ts   = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    with g1_lock:
        g1_state["log"].append(line)
        if len(g1_state["log"]) > 400:
            g1_state["log"] = g1_state["log"][-400:]
    (log.error if level == "error" else log.info)(f"G1: {msg}")


def g1_set(updates, log_msg=None):
    with g1_lock:
        g1_state.update(updates)
    if log_msg:
        g1_log(log_msg)


def g1_wait_field(field, timeout_sec=300):
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        time.sleep(0.4)
        with g1_lock:
            val = g1_state.get(field)
            if val:
                g1_state[field] = None
                return val
    return None


# ── Month availability for GSTR-1 ─────────────────────────────────────────────
def g1_months_for_fy(fy):
    """
    GSTR-1 is filed by the 11th of the following month (monthly filers).
    For QRMP quarterly filers, it's 13th of month after quarter end.
    We return all months up to and including those that are due — i.e.,
    any month where the 11th of the following month has already passed.
    This is looser than 2B (which waits until 14th) — we allow any elapsed month.
    """
    fy_year = int(fy.split("-")[0])
    now     = datetime.now()
    order   = [
        (4,"April"),(5,"May"),(6,"June"),(7,"July"),
        (8,"August"),(9,"September"),(10,"October"),
        (11,"November"),(12,"December"),
        (1,"January"),(2,"February"),(3,"March"),
    ]
    months = []
    for mon_num, mon_name in order:
        yr = fy_year if mon_num >= 4 else fy_year + 1
        # Due date: 11th of following month
        due_yr  = yr + 1 if mon_num == 12 else yr
        due_mon = 1       if mon_num == 12 else mon_num + 1
        # Skip if the due month hasn't started at all
        if due_yr > now.year or (due_yr == now.year and due_mon > now.month):
            continue
        months.append({
            "num":    mon_num,
            "year":   yr,
            "name":   mon_name,
            "abbr":   mon_name[:3].upper(),
            "display":f"{mon_name} {yr}",
            "period": f"{mon_num:02d}{yr}",
        })
    return months


# ── Login — same GST portal, mirrored login code using g1_state ────────────────
def g1_do_browser_login(page, username, password):
    """Login to GST portal — mirrors g2a_do_browser_login exactly."""
    g1_log("Opening GST login page...")
    try:
        page.goto(GST_LOGIN, wait_until="domcontentloaded", timeout=25000)
        try: page.wait_for_selector(
                "input#username, input[name='username'], input[placeholder*='username' i]",
                state="visible", timeout=8000)
        except Exception: pass
    except Exception as e:
        g1_log(f"  ✗ Could not open login page: {e}", "error")
        g1_set({"status":"error","error":str(e)}); return False

    # Username
    for sel in ["input#username","input[name='username']","input[placeholder*='username' i]"]:
        try:
            page.locator(sel).first.fill(username)
            g1_log("  ✓ Username filled"); break
        except Exception: continue

    # Password
    for sel in ["input#user_pass","input[name='user_pass']",
                "input[type='password']","input[placeholder*='password' i]"]:
        try:
            page.locator(sel).first.fill(password)
            g1_log("  ✓ Password filled"); break
        except Exception: continue

    time.sleep(0.3)

    # Capture captcha image — try element screenshot first (most reliable)
    cap_img = None
    try:
        import base64 as _b64
        # Method 1: direct element screenshot of the captcha <img>
        for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                    "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                    ".captchaImage img", "img.captcha",
                    # GST portal specific: img below password field
                    "form img", "img"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
                    g1_log(f"  ✓ Captcha captured via element screenshot ({sel})")
                    break
            except Exception:
                continue

        # Method 2: canvas toDataURL
        if not cap_img:
            data_url = page.evaluate("""() => {
                for (const cv of document.querySelectorAll('canvas')) {
                    if (cv.width > 30 && cv.height > 10) {
                        try { return cv.toDataURL('image/png'); } catch(e) {}
                    }
                }
                return null;
            }""")
            if data_url and data_url.startswith("data:image"):
                cap_img = data_url
                g1_log("  ✓ Captcha captured via canvas")

        # Method 3: full page screenshot fallback
        if not cap_img:
            png = page.screenshot(full_page=False)
            cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
            g1_log("  ✓ Captcha captured via full screenshot (fallback)")

    except Exception as e:
        g1_log(f"  ⚠ Captcha capture error: {e}")

    g1_set({"status":"waiting_captcha", "captcha_image": cap_img})
    g1_log("  ⏸ Waiting for captcha...")

    # Wait for user to submit captcha answer
    answer = g1_wait_field("captcha_answer", timeout_sec=300)
    if not answer:
        g1_log("  ✗ Captcha timeout", "error")
        g1_set({"status":"error","error":"Captcha timeout"}); return False

    # Hide captcha box immediately — answer received, processing login
    g1_set({"status":"running", "captcha_image": None})

    # Fill captcha
    for sel in ["input[placeholder*='Characters' i]","input#captcha","input[name='captcha']"]:
        try:
            page.locator(sel).first.fill(str(answer))
            g1_log(f"  ✓ Captcha field: {sel}")
            g1_log(f"  ✓ Captcha filled: {repr(answer)}")
            g1_log(f"  ✓ Captcha: {answer}")
            break
        except Exception: continue

    # Submit login
    for sel in ["button[type='submit']","input[type='submit']","button:has-text('LOGIN')"]:
        try:
            page.locator(sel).first.click()
            g1_log("  ✓ Login submitted")
            break
        except Exception: continue
    time.sleep(1.5)

    # Handle OTP if needed
    otp_needed = False
    for _ in range(8):
        time.sleep(1)
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.wait_for(state="visible", timeout=1500)
            otp_needed = True; break
        except Exception:
            if check_login_success(page) is not False:
                break

    if otp_needed:
        g1_set({"status":"waiting_otp"})
        g1_log("  ⏸ Waiting for OTP...")
        otp = g1_wait_field("otp_answer", timeout_sec=180)
        if not otp:
            g1_log("  ✗ OTP timeout", "error")
            g1_set({"status":"error","error":"OTP timeout"}); return False
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.click(); time.sleep(0.2)
            otp_el.fill(str(otp))
            page.locator("button[type='submit'],input[type='submit']").first.click()
            g1_log("  ✓ OTP submitted"); time.sleep(1)
        except Exception as e:
            g1_log(f"  ✗ OTP error: {e}", "error")
            return False

    time.sleep(1)

    # ── Captcha retry loop (up to 3 total attempts) ──────────
    MAX_CAPTCHA_ATTEMPTS = 3
    for _attempt in range(MAX_CAPTCHA_ATTEMPTS):
        result = check_login_success(page)
        if result is not False:
            # Logged in (True) or uncertain (None → proceed optimistically)
            break

        # Login failed — portal has already auto-refreshed the captcha image
        remaining = MAX_CAPTCHA_ATTEMPTS - _attempt - 1
        if remaining == 0:
            g1_log(f"  ✗ Login failed after {MAX_CAPTCHA_ATTEMPTS} attempts", "error")
            g1_set({"status":"error","error":f"Login failed after {MAX_CAPTCHA_ATTEMPTS} captcha attempts"})
            return False

        g1_log(f"  ✗ Login failed — re-capturing fresh captcha (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")

        # Small wait for portal to settle after failure + captcha refresh
        time.sleep(2)

        # Re-capture the new captcha the portal already rendered
        new_cap = None
        try:
            import base64 as _b64
            for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                        "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                        ".captchaImage img", "img.captcha", "form img", "img"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=2000)
                    png = loc.screenshot()
                    if png and len(png) > 500:
                        new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                        g1_log(f"  ✓ Fresh captcha re-captured ({sel})")
                        break
                except Exception:
                    continue
            if not new_cap:
                png = page.screenshot(full_page=False)
                new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                g1_log("  ✓ Fresh captcha via screenshot (fallback)")
        except Exception as ce:
            g1_log(f"  ⚠ Re-capture error: {ce}")

        # Signal UI to show new captcha — force=True by clearing then setting
        # so the JS !alreadyShown guard triggers and fetches the new image
        g1_set({"status":"running",  "captcha_image": None})
        time.sleep(0.1)
        g1_set({"status":"waiting_captcha", "captcha_image": new_cap})
        g1_log(f"  ⏸ Waiting for captcha answer (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")

        new_answer = g1_wait_field("captcha_answer", timeout_sec=300)
        if not new_answer:
            g1_log("  ✗ Captcha retry timeout", "error")
            g1_set({"status":"error","error":"Captcha retry timeout"})
            return False

        g1_set({"status":"running", "captcha_image": None})

        # Clear old captcha field, fill new answer, re-submit
        for sel in ["input[placeholder*='Characters' i]","input#captcha","input[name='captcha']"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.fill("")
                loc.fill(str(new_answer))
                g1_log(f"  ✓ Re-filled captcha: {repr(new_answer)}")
                break
            except Exception:
                continue

        for sel in ["button[type='submit']","input[type='submit']","button:has-text('LOGIN')"]:
            try:
                page.locator(sel).first.click()
                g1_log("  ✓ Login re-submitted")
                break
            except Exception:
                continue

        time.sleep(2)
        # Handle OTP in case it appears on retry (unlikely but safe)
        for _ in range(5):
            time.sleep(1)
            try:
                otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
                otp_el.wait_for(state="visible", timeout=1000)
                # OTP appeared on retry — handle it
                g1_set({"status":"waiting_otp"})
                g1_log("  ⏸ OTP required (retry)...")
                otp_r = g1_wait_field("otp_answer", timeout_sec=180)
                if not otp_r:
                    g1_log("  ✗ OTP timeout (retry)", "error")
                    g1_set({"status":"error","error":"OTP timeout on retry"})
                    return False
                otp_el.click(); time.sleep(0.2)
                otp_el.fill(str(otp_r))
                page.locator("button[type='submit'],input[type='submit']").first.click()
                g1_log("  ✓ OTP re-submitted"); time.sleep(1)
                break
            except Exception:
                if check_login_success(page) is not False:
                    break
        # Loop back to check_login_success at top of for loop

    g1_log("  ✅ Logged in successfully")
    g2a_dismiss_popup(page)
    return True


# ─────────────────────────────────────────────────────────────────────
#  Download file from URL
# ─────────────────────────────────────────────────────────────────────
def g1_navigate_to_prepare_offline(page, fy, mon_name, quarter_label):
    """
    Portal flow (confirmed from screenshots):
      1. Returns Dashboard → select FY / Quarter / Month → SEARCH
      2. Tile "Details of outward supplies / GSTR1 / Quarterly" shows DOWNLOAD button
         (IFF non-end-quarter months also show DOWNLOAD on the IFF tile)
      3. Click DOWNLOAD → lands on /returns/auth/gstr/offlinedownload
      4. Caller clicks "GENERATE JSON FILE TO DOWNLOAD"

    KEY: Uses Playwright native select_option() — NOT synthetic JS events.
    Angular ignores synthetic change events; only real browser interaction works.
    """
    dashboard_url = "https://return.gst.gov.in/returns/auth/dashboard"
    try:
        page.evaluate(f"window.location.href = '{dashboard_url}'")
    except Exception:
        page.goto(dashboard_url, wait_until="domcontentloaded", timeout=15000)
    try:
        page.wait_for_load_state("domcontentloaded", timeout=12000)
    except Exception:
        pass
    time.sleep(1.5)

    # ── Helper: select by visible label text using Playwright native API ──────
    # select_option(label=...) triggers real browser focus/change events that
    # Angular's form bindings actually respond to.
    def pw_select(nth, label, timeout=8):
        """Select <option> with matching text in the nth <select> on the page.
           Returns True on success, False on failure."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                sel_loc = page.locator("select").nth(nth)
                sel_loc.wait_for(state="visible", timeout=2000)
                sel_loc.select_option(label=label)
                time.sleep(0.4)   # let Angular process the change
                # Verify it actually changed
                current = sel_loc.evaluate(
                    "el => el.options[el.selectedIndex] ? el.options[el.selectedIndex].text.trim() : ''")
                if current == label:
                    return True
                # Sometimes Angular resets it — try once more
                sel_loc.select_option(label=label)
                time.sleep(0.5)
                return True
            except Exception as _e:
                time.sleep(0.4)
        return False

    def wait_for_option(nth, label, timeout=8):
        """Poll until <label> appears as an option in nth select."""
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                opts = page.locator("select").nth(nth).evaluate(
                    "el => Array.from(el.options).map(o => o.text.trim())")
                if label in opts:
                    return True
            except Exception:
                pass
            time.sleep(0.4)
        return False

    # ── Step 1: Select Financial Year ─────────────────────────────────────────
    g1_log(f"    Selecting FY: {fy}")
    if not pw_select(0, fy, timeout=8):
        g1_log(f"    ⚠ FY '{fy}' select failed — trying anyway")
    time.sleep(0.6)

    # ── Detect portal layout ───────────────────────────────────────────────────
    # Count selects after FY change (Angular may inject/remove the quarter dropdown)
    n_selects = 0
    for _w in range(15):   # up to 4.5 s for Angular to render dropdowns
        try:
            n_selects = page.evaluate("() => document.querySelectorAll('select').length")
        except Exception:
            pass
        if n_selects >= 2:
            break
        time.sleep(0.3)
    g1_log(f"    {n_selects} select(s) detected after FY change")

    if n_selects >= 3:
        # ── Normal layout: FY (0) → Quarter (1) → Month (2) ──────────────────
        g1_log(f"    Selecting Quarter: {quarter_label}")
        if not wait_for_option(1, quarter_label, timeout=6):
            g1_log(f"    ⚠ Quarter '{quarter_label}' option never appeared")
        if not pw_select(1, quarter_label, timeout=6):
            g1_log(f"    ⚠ Quarter select failed")
        time.sleep(0.6)   # wait for Angular to reload Month dropdown

        g1_log(f"    Selecting Month: {mon_name}")
        if not wait_for_option(2, mon_name, timeout=6):
            g1_log(f"    ⚠ Month '{mon_name}' option never appeared in dropdown")
        if not pw_select(2, mon_name, timeout=6):
            g1_log(f"    ⚠ Month select failed")
        time.sleep(0.4)

    else:
        # ── 2-dropdown layout: FY (0) → Month (1), no Quarter ─────────────────
        g1_log(f"    ℹ No Quarter dropdown ({n_selects} selects) — selecting Month at index 1")
        if not wait_for_option(1, mon_name, timeout=6):
            g1_log(f"    ⚠ Month '{mon_name}' never appeared in index-1")
        if not pw_select(1, mon_name, timeout=6):
            g1_log(f"    ⚠ Month select failed")
        time.sleep(0.4)

    # ── Step 2: Click SEARCH ───────────────────────────────────────────────────
    _searched = False
    for s in ["button:has-text('SEARCH')", "button:has-text('Search')",
              "input[value='SEARCH']", "input[value='Search']"]:
        try:
            _loc = page.locator(s).first
            _loc.wait_for(state="visible", timeout=4000)
            _loc.click()
            _searched = True
            g1_log(f"    ✓ SEARCH clicked")
            break
        except Exception:
            continue
    if not _searched:
        g1_log(f"    ⚠ SEARCH button not found")
    time.sleep(2.5)   # Angular re-renders tiles

    # ── Step 3: Wait for tiles to render ──────────────────────────────────────
    _tile_ready = False
    for _tw in range(25):    # up to ~7.5 s
        try:
            _found = page.evaluate("""() => {
                for (const btn of document.querySelectorAll('button,a')) {
                    if ((btn.textContent || '').trim().toUpperCase() === 'DOWNLOAD')
                        return true;
                }
                return false;
            }""")
            if _found:
                _tile_ready = True
                break
        except Exception:
            pass
        time.sleep(0.3)
    if not _tile_ready:
        g1_log(f"    ⚠ DOWNLOAD button not found after SEARCH — proceeding anyway")

    # ── Step 4: Click DOWNLOAD on GSTR-1 / IFF tile ONLY ────────────────────
    # Strategy: find every DOWNLOAD button on the page; for each one, walk UP
    # through its ancestors checking whether ANY ancestor:
    #   (a) contains a GSTR-1/IFF label  → this is the right tile → click it
    #   (b) also contains a label from another return type → too high, wrong tile
    # If GSTR-1/IFF tile is found on the page but has no DOWNLOAD button,
    # return a "NO DOWNLOAD" signal so the caller can skip (not retry) the month.
    clicked = page.evaluate("""() => {
        const G1_LABELS   = ['GSTR1','GSTR-1','Invoice Furnishing Facility'];
        const OTHER_TYPES = ['GSTR3B','GSTR-3B','GSTR2A','GSTR-2A',
                             'GSTR2B','GSTR-2B','GSTR4','GSTR-4',
                             'TDS','TCS','GSTR5','GSTR-5','GSTR9','GSTR-9'];

        function elText(el) {
            // Return trimmed text only for leaf-like elements (single text node)
            return (el.childNodes.length === 1
                ? (el.textContent||'') : (el.getAttribute('data-title')||'')).trim();
        }

        // ── Check if this tile's label is a GSTR-1/IFF label ──────────────
        function ancestorHasG1Label(btn) {
            let node = btn;
            for (let depth = 0; depth < 8; depth++) {
                node = node.parentElement;
                if (!node) return {found: false, hasOther: false};

                // Scan direct descendants (not recursive) at each level
                for (const child of node.querySelectorAll('*')) {
                    const t = elText(child);
                    if (G1_LABELS.includes(t))  return {found: true,  node, hasOther: false};
                    if (OTHER_TYPES.includes(t)) return {found: false, node, hasOther: true};
                }
            }
            return {found: false, hasOther: false};
        }

        // Try each DOWNLOAD button
        for (const btn of document.querySelectorAll('button, a')) {
            if ((btn.textContent||'').trim().toUpperCase() !== 'DOWNLOAD') continue;
            const {found} = ancestorHasG1Label(btn);
            if (found) {
                btn.click();
                return 'GSTR-1: DOWNLOAD clicked';
            }
        }

        // No DOWNLOAD button found — check if the tile exists at all
        for (const el of document.querySelectorAll('*')) {
            if (G1_LABELS.includes(elText(el))) {
                return 'GSTR-1: NO DOWNLOAD (not filed — skip)';
            }
        }

        return null;   // GSTR-1 tile not found on this dashboard
    }""")
    g1_log(f"    Dashboard tile result: {clicked}")

    # ── Handle tile result ──────────────────────────────────────────────────
    if clicked is None:
        g1_log(f"    ⚠ GSTR-1/IFF tile not found on dashboard")
        return False

    if isinstance(clicked, str) and "NO DOWNLOAD" in clicked:
        # Tile found but no DOWNLOAD button — GSTR-1 not filed for this period
        g1_log(f"    ℹ GSTR-1 not filed for this period — no download available")
        return "not_filed"

    # ── Step 5: Wait for offlinedownload URL ───────────────────────────────────
    try:
        page.wait_for_url("*offlinedownload*", timeout=12000)
    except Exception:
        pass
    time.sleep(1.5)

    current_url = page.evaluate("() => location.href").lower()
    g1_log(f"    -> URL: {current_url[:80]}")
    return "offlinedownload" in current_url


def g1_click_download_tab(page):
    """
    Previously clicked a 'Download' tab on an older portal layout.
    Current portal flow: clicking DOWNLOAD on the dashboard tile lands
    directly on /gstr/offlinedownload — there is no separate tab to click.
    This function is kept for safety but is now effectively a no-op.
    """
    return True   # already on offlinedownload page


def g1_get_download_link(page):
    """Return the href of the 'Click here to download JSON - File 1' link, or None.
    Matches only <a> tags (not buttons) — same approach as g2a get_json_link.
    Conditions deliberately kept tight to avoid false-matching the GENERATE button
    whose text 'GENERATE JSON FILE TO DOWNLOAD' would match 'json file'."""
    return page.evaluate("""() => {
        for (const a of document.querySelectorAll('a')) {
            const t = (a.textContent || '').toLowerCase().trim();
            if (t.includes('click here') || t.includes('json - file')
                    || t.includes('json-file')) {
                return a.href || null;
            }
        }
        return null;
    }""")


def g1_trigger_generate(page):
    """
    Click the GENERATE JSON FILE TO DOWNLOAD button on /gstr/offlinedownload.
    Exact button text confirmed from portal screenshot:
        "GENERATE JSON FILE TO DOWNLOAD"
    """
    for sel in [
        "button:has-text('GENERATE JSON FILE TO DOWNLOAD')",
        "button:has-text('GENERATE JSON FILE')",
        "button:has-text('GENERATE JSON')",
        "button:has-text('GENERATE FILE')",
        "button:has-text('Generate File')",
        "button:has-text('GENERATE')",
        "input[value*='GENERATE' i]",
    ]:
        try:
            el = page.locator(sel).first
            el.wait_for(state="visible", timeout=4000)
            el.click()
            g1_log(f"    ✓ Generate button clicked: {sel}")
            return True
        except Exception: continue
    g1_log("    ⚠ Generate button not found")
    return False


def g1_fetch_download(page, url, gstin_dir, month_display):
    """Download GSTR-1 JSON/ZIP from URL, save as GSTR1_<Month>_<Year>.json"""
    import zipfile, io as _io
    abbr      = month_display.replace(" ", "_")
    json_path = os.path.join(gstin_dir, f"GSTR1_{abbr}.json")
    try:
        resp = page.request.get(url, headers={
            "Accept":  "*/*",
            "Referer": "https://return.gst.gov.in/returns/auth/gstr1/",
        }, timeout=120000)
        raw = resp.body()
        ct  = resp.headers.get("content-type","").lower()
        g1_log(f"    [{resp.status}] {ct[:40]}  {len(raw)} bytes")
        if resp.status != 200 or not raw:
            return None
        if raw[:2] == b"PK" or "zip" in ct:
            with zipfile.ZipFile(_io.BytesIO(raw)) as zf:
                names  = zf.namelist()
                target = next((n for n in names if n.endswith(".json")), names[0])
                data   = zf.read(target)
            with open(json_path, "wb") as f: f.write(data)
            size_kb = max(1, len(data)//1024)
        else:
            with open(json_path, "wb") as f: f.write(raw)
            size_kb = max(1, len(raw)//1024)
        g1_log(f"    ✅ {month_display} — {size_kb} KB saved")
        return {"month":month_display,
                "filename":f"GSTR1_{abbr}.json", "size_kb":size_kb}
    except Exception as e:
        g1_log(f"    ✗ Download error: {e}", "error")
        return None


# ── Multi-pass sweep download ──────────────────────────────────────────────────
def g1_download_all_months(page, months, gstin, gstin_dir, fy,
                           inter_sweep_wait=300, max_sweeps=None):
    """
    MULTI-PASS SWEEP — mirrors g2a pattern exactly.
    inter_sweep_wait: seconds between sweeps (default 300). Pass 0 for combined mode.
    max_sweeps:       if set, stop after N sweeps and return pending months.
    Returns: (files_done, pending)
    """
    LINK_WAIT_SEC = 5   # seconds to wait per month for link to appear after GENERATE

    month_names = {
        1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
        7:"July",8:"August",9:"September",10:"October",
        11:"November",12:"December"
    }
    files_done  = []
    pending     = list(months)
    sweep       = 0
    deadline    = time.time() + (35 * 60)

    while pending and time.time() < deadline:
        sweep += 1
        g1_log(f"\n━━ Sweep {sweep} — {len(pending)} month(s) remaining ━━")
        still_pending = []

        for i, month in enumerate(pending):
            mon_num       = month["num"]
            mon_name      = month_names[mon_num]
            mon_display   = month["display"]
            period        = month["period"]
            quarter_label, _ = g2a_quarter_for_month(mon_num)

            g1_set({
                "status":        "downloading",
                "current_month": mon_display,
                "progress":      int((len(files_done) / len(months)) * 100),
            })
            g1_log(f"  [{sweep}.{i+1}] {mon_display}")

            on_page = g1_navigate_to_prepare_offline(
                page, fy, mon_name, quarter_label)

            if on_page == "not_filed":
                # GSTR-1/IFF tile found but no DOWNLOAD button → not filed → skip permanently
                g1_log(f"    ⏭ {mon_display}: GSTR-1 not filed — skipping (no download available)")
                files_done.append({"month": mon_display, "filename": None,
                                   "skipped": True, "reason": "not_filed"})
                continue   # do NOT add to still_pending

            if not on_page:
                g1_log(f"    ⚠ Could not reach GSTR-1 offlinedownload for {mon_display}")
                still_pending.append(month); continue

            # ── Check for pre-existing link ───────────────────────────────────
            link_href = g1_get_download_link(page)
            if link_href:
                g1_log(f"    ✅ Link already present — downloading immediately")
            else:
                # ── Click GENERATE JSON FILE TO DOWNLOAD ──────────────────────
                g1_trigger_generate(page)
                g1_log(f"    ▶ GENERATE clicked — waiting {LINK_WAIT_SEC}s for link...")

                # ── Wait LINK_WAIT_SEC seconds polling for link ────────────────
                _lw_deadline = time.time() + LINK_WAIT_SEC
                while time.time() < _lw_deadline:
                    time.sleep(0.5)
                    link_href = g1_get_download_link(page)
                    if link_href:
                        g1_log(f"    ⚡ Link appeared!")
                        break

            # ── Download if link available ────────────────────────────────────
            if link_href:
                result = g1_fetch_download(page, link_href, gstin_dir, mon_display)
                if result:
                    result["period"] = period
                    files_done.append(result)
                    g1_set({"files": files_done.copy(),
                            "done_months": len(files_done),
                            "progress": int((len(files_done)/len(months))*100)})
                    g1_log(f"    ✅ {mon_display} downloaded  "
                           f"({len(files_done)}/{len(months)} total)")
                    continue   # do NOT add to still_pending
                else:
                    g1_log(f"    ⚠ Download failed (404/error) — "
                           f"re-clicking GENERATE for fresh link")
                    g1_trigger_generate(page)
                    g1_log(f"    ▶ GENERATE re-clicked — will collect next sweep")
            else:
                g1_log(f"    ⏩ Link not ready — server still generating. "
                       f"Will revisit next sweep.")

            still_pending.append(month)

        pending = still_pending

        if pending and time.time() < deadline:
            if max_sweeps and sweep >= max_sweeps:
                break   # combined mode: caller runs other modules now
            if inter_sweep_wait > 0:
                g1_log(f"\n  ⏸ Waiting {inter_sweep_wait}s for portal to generate files...")
                for s in range(inter_sweep_wait):
                    if time.time() >= deadline: break
                    if s % 60 == 0 and s > 0:
                        g1_log(f"    ... {inter_sweep_wait-s}s remaining")
                    time.sleep(1)

    if pending:
        g1_log(f"  ⚠ {len(pending)} month(s) not downloaded (timeout or portal issue):",
               "error")
        for m in pending:
            g1_log(f"    • {m['display']}", "error")

    return files_done, pending


# ── Main worker ────────────────────────────────────────────────────────────────
def g1_worker(gstin, fy, username, password, specific_month=""):
    """
    Thin wrapper guaranteeing any exception is logged and reflected in
    g1_state instead of silently killing the thread.
    """
    try:
        _g1_worker_impl(gstin, fy, username, password, specific_month)
    except Exception as fatal:
        import traceback as _tb
        try:
            g1_log(f"✗ FATAL (uncaught): {fatal}", "error")
        except Exception:
            pass
        try:
            g1_set({"status": "error", "error": f"Uncaught: {fatal}"})
        except Exception:
            pass
        log.error(f"[G1] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _g1_worker_impl(gstin, fy, username, password, specific_month=""):
    os.makedirs(G1_DOWNLOAD_DIR, exist_ok=True)
    gstin_dir = os.path.join(G1_DOWNLOAD_DIR, gstin, fy.replace("-","_"))
    os.makedirs(gstin_dir, exist_ok=True)

    all_months = g1_months_for_fy(fy)
    if not all_months:
        g1_set({"status":"error",
                "error":f"No GSTR-1 months available for FY {fy}. "
                        f"GSTR-1 is due by 11th of the following month."}); return

    if specific_month:
        sm = specific_month.strip()
        filtered = [m for m in all_months if (
            sm.lower() == m["display"].lower() or sm == m["period"] or
            sm.lower() == m["name"].lower()    or sm.upper() == m["abbr"]
        )]
        if not filtered:
            g1_set({"status":"error",
                    "error":f"Month '{specific_month}' not found or not yet due. "
                            f"Available: {[m['display'] for m in all_months]}"}); return
        months = filtered
    else:
        months = all_months

    # ── Cache check — skip months already downloaded ─────────────────────────
    def _g1_cached_path(m):
        fname = f"GSTR1_{m['display'].replace(' ', '_')}.json"
        return os.path.join(gstin_dir, fname), fname

    cached_files = []
    months_needed = []
    for m in months:
        fpath, fname = _g1_cached_path(m)
        if os.path.exists(fpath) and os.path.getsize(fpath) > 10:
            cached_files.append({
                "month":    m["display"],
                "filename": fname,
                "size_kb":  max(1, os.path.getsize(fpath) // 1024),
            })
        else:
            months_needed.append(m)

    if cached_files:
        g1_log(f"  ✅ {len(cached_files)} month(s) already in cache (skipped):")
        for c in cached_files:
            g1_log(f"     {c['month']}  ({c['filename']},  {c['size_kb']} KB)")

    if not months_needed:
        # All months cached — no browser needed
        g1_set({"status": "done", "progress": 100, "current_month": None,
                "files": cached_files, "done_months": len(cached_files),
                "total_months": len(months)})
        g1_log(f"\n✅ All {len(cached_files)} month(s) already in cache — no download needed")
        g1_log(f"   Saved to: {gstin_dir}")
        time.sleep(5)
        with g1_lock:
            if g1_state.get("status") == "done":
                g1_state["status"] = "idle"
        g1_log("🔁 GSTR-1 RPA reset to idle")
        return

    g1_log(f"  📥 {len(months_needed)}/{len(months)} month(s) need downloading"
           f" ({len(cached_files)} cached)")
    months = months_needed   # only download what's not already on disk

    g1_set({"total_months": len(months), "done_months": 0,
            "files": cached_files})   # pre-populate with cached files
    g1_log(f"GSTR-1 | GSTIN: {gstin} | FY: {fy} | {len(months)} month(s)")
    g1_log(f"Note: GSTR-1 = Outward Supplies (Sales). "
           f"Download uses PREPARE OFFLINE → Download tab → GENERATE FILE.")

    profile_dir = os.path.join(PATHS.profiles_dir, "gst_profile_g1")
    os.makedirs(profile_dir, exist_ok=True)
    lock_file = os.path.join(profile_dir, "SingletonLock")
    if os.path.exists(lock_file):
        try: os.remove(lock_file); g1_log("  ✓ Removed stale SingletonLock")
        except Exception: pass

    # sync_playwright().start() launches Playwright's own Node.js driver
    # subprocess — if it throws (not just hangs), that was previously
    # uncaught here, silently killing this thread with nothing logged.
    try:
        pw = sync_playwright().start()
    except Exception as e:
        g1_log(f"✗ Playwright driver failed to start: {e}", "error")
        g1_set({"status": "error", "error": f"Playwright driver failed to start: {e}"})
        return
    try:
        g1_log("🚀 Launching browser...")
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=False, slow_mo=40,
                args=["--disable-blink-features=AutomationControlled",
                      "--no-sandbox","--start-maximized"],
                no_viewport=True, accept_downloads=True,
            )
        except Exception as e:
            g1_log(f"  ✗ Browser launch error: {e}", "error")
            g1_set({"status":"error","error":str(e)}); return

        page = context.pages[0] if context.pages else context.new_page()
        page.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        g2a_install_interceptors(page)

        if not g1_do_browser_login(page, username, password):
            context.close(); return

        g1_log("🔄 Activating session on Returns portal...")
        ok, page = g2a_activate_session(page, context)
        if not ok:
            g1_set({"status":"error","error":"Could not reach Returns portal"})
            context.close(); return
        g1_log(f"  Using: {page.url[:70]}")
        g2a_install_interceptors(page)

        g1_log("📥 Starting GSTR-1 multi-pass download sweep...")
        files_done, _ = g1_download_all_months(page, months, gstin, gstin_dir, fy)

        all_files = cached_files + files_done
        g1_set({"status":"done","progress":100,"current_month":None,
                "files": all_files, "done_months": len(all_files)})
        g1_log(f"\n✅ Complete — {len(files_done)} downloaded, "
               f"{len(cached_files)} from cache, {len(all_files)} total")
        g1_log(f"   Saved to: {gstin_dir}")

        try: context.close(); g1_log("🔒 Browser closed")
        except Exception as ce: g1_log(f"  ⚠ Browser close: {ce}")
    except Exception as _e:
        g1_set({"status":"error","error":str(_e)})
        g1_log(f"✗ Fatal worker error: {_e}", "error")
        import traceback; g1_log(traceback.format_exc(), "error")
    finally:
        try: pw.stop()
        except Exception: pass

    time.sleep(8)
    with g1_lock:
        if g1_state.get("status") == "done":
            g1_state["status"] = "idle"
    g1_log("🔁 GSTR-1 RPA reset to idle")


# ══════════════════════════════════════════════════════════════════════════════
#  GSTR-1 → EXCEL CONVERTER
#  Sheets:
#    1. Summary           — month-wise totals (Taxable, IGST, CGST, SGST, Cess)
#    2. B2B               — Registered buyer invoices (b2b + b2ba amendments)
#    3. B2CL              — Unregistered inter-state > ₹2.5L (b2cl + b2cla)
#    4. B2CS              — Unregistered intra-state + small value (b2cs + b2csa)
#    5. Credit-Debit Notes— CDNR (registered) + CDNUR (unregistered)
#    6. Exports           — exp (WPAY/WOPAY) + expa amendments
#    7. NIL / Exempt      — nil section (exempt, nil-rated, non-GST)
#    8. HSN Summary       — hsn.data (mandatory from Jan 2025, Phase 3)
#    9. Document Summary  — docs.doc_det (serialised document tracking)
#   10. Advance / TXP     — at, ata, txp, txpa sections
# ══════════════════════════════════════════════════════════════════════════════
def g1_json_to_excel(gstin, fy):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import glob

    fy_dir     = fy.replace("-","_")
    folder     = os.path.join(G1_DOWNLOAD_DIR, gstin, fy_dir)
    json_files = sorted(glob.glob(os.path.join(folder, "*.json")))
    if not json_files:
        return None, f"No GSTR-1 JSON files found in {folder}"

    # ── Colours (Green brand for GSTR-1 — distinct from 2A navy, 2B indigo) ──
    HDR_FILL   = PatternFill("solid", fgColor="1B5E20")   # dark green
    HDR2_FILL  = PatternFill("solid", fgColor="2E7D32")
    ALT_FILL   = PatternFill("solid", fgColor="F1F8E9")   # light green
    TOT_FILL   = PatternFill("solid", fgColor="C8E6C9")
    B2B_FILL   = PatternFill("solid", fgColor="E8F5E9")
    B2CL_FILL  = PatternFill("solid", fgColor="FFF3E0")
    B2CS_FILL  = PatternFill("solid", fgColor="E3F2FD")
    CDN_CR     = PatternFill("solid", fgColor="FCE4EC")   # credit note
    CDN_DR     = PatternFill("solid", fgColor="F3E5F5")   # debit note
    EXP_FILL   = PatternFill("solid", fgColor="E8EAF6")
    NIL_FILL   = PatternFill("solid", fgColor="FFF8E1")
    HSN_FILL   = PatternFill("solid", fgColor="F3E5F5")
    AMEND_FILL = PatternFill("solid", fgColor="FFF3E0")

    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    BODY_FONT = Font(name="Arial", size=9)
    TOT_FONT  = Font(name="Arial", bold=True, size=9, color="1B5E20")
    CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT      = Alignment(horizontal="left",   vertical="center")
    RIGHT     = Alignment(horizontal="right",  vertical="center")
    thin      = Side(style="thin", color="A5D6A7")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM_FMT   = '#,##0.00'

    def make_title(ws, col_count, text, fg="1B5E20"):
        ws.merge_cells(f"A1:{get_column_letter(col_count)}1")
        c = ws["A1"]
        c.value = text
        c.font  = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        c.fill  = PatternFill("solid", fgColor=fg)
        c.alignment = CENTER
        ws.row_dimensions[1].height = 26

    def make_hdr(ws, row, labels, fill=None):
        for c, lbl in enumerate(labels, 1):
            cell = ws.cell(row=row, column=c, value=lbl)
            cell.font = HDR_FONT; cell.alignment = CENTER
            cell.border = BORDER; cell.fill = fill or HDR_FILL
        ws.row_dimensions[row].height = 32

    def write_row(ws, r, vals, row_fill=None):
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY_FONT; cell.border = BORDER
            if row_fill: cell.fill = row_fill
            if isinstance(v, (int, float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT

    def tot_row(ws, r, ncols, num_col_idxs, ds=3):
        """
        Write a TOTAL row with SUM formulas.
        Guard: when r == ds (no data rows, total immediately after header),
        de = r-1 < ds which creates a reversed range =SUM(N3:N2).
        Excel auto-corrects reversed ranges to forward (N2:N3) making the
        formula cell PART of its own range → circular reference.
        Fix: write 0 instead of SUM when de < ds (no data rows).
        """
        for c in range(1, ncols+1):
            cell = ws.cell(row=r, column=c)
            cell.fill = TOT_FILL; cell.font = TOT_FONT
            cell.border = BORDER
            if isinstance(cell.value, (int, float)):
                cell.alignment = RIGHT; cell.number_format = NUM_FMT
            else:
                cell.alignment = LEFT
        de = r - 1
        for col_idx in num_col_idxs:
            col_ltr = get_column_letter(col_idx)
            cell = ws.cell(row=r, column=col_idx)
            if de >= ds:
                # Normal case: data rows exist between ds and de
                cell.value = f"=SUM({col_ltr}{ds}:{col_ltr}{de})"
            else:
                # No data rows — write 0 to avoid circular reference
                cell.value = 0
            cell.number_format = '#,##0.00'
            cell.alignment = RIGHT

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # State codes
    STATE_CODES = {
        "01":"J&K","02":"HP","03":"Punjab","04":"Chandigarh",
        "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan",
        "09":"UP","10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh",
        "13":"Nagaland","14":"Manipur","15":"Mizoram","16":"Tripura",
        "17":"Meghalaya","18":"Assam","19":"West Bengal","20":"Jharkhand",
        "21":"Odisha","22":"Chhattisgarh","23":"MP","24":"Gujarat",
        "27":"Maharashtra","29":"Karnataka","30":"Goa","32":"Kerala",
        "33":"Tamil Nadu","34":"Puducherry","36":"Telangana","37":"AP",
        "38":"Ladakh",
    }
    def pos_name(code):
        s = str(code).zfill(2)
        return f"{s}-{STATE_CODES.get(s,code)}"

    INV_TYPES  = {"R":"Regular","SEWP":"SEZ w/ Payment",
                  "SEWOP":"SEZ w/o Payment","DE":"Deemed Export",
                  "CBW":"Customs Bonded Warehouse"}
    NOTE_TYPES = {"C":"Credit Note","D":"Debit Note","R":"Refund Voucher"}
    EXP_TYPES  = {"WPAY":"With Payment","WOPAY":"Without Payment"}

    MONTH_ORDER = ["April","May","June","July","August","September",
                   "October","November","December","January","February","March"]

    def fp_to_display(fp):
        MN = {"01":"January","02":"February","03":"March","04":"April",
              "05":"May","06":"June","07":"July","08":"August",
              "09":"September","10":"October","11":"November","12":"December"}
        return f"{MN.get(fp[:2],fp[:2])} {fp[2:6]}" if fp and len(fp)>=6 else fp or "?"

    # ── Load all JSON files ────────────────────────────────────────────────────
    months_data = []
    for fpath in json_files:
        try:
            with open(fpath, encoding="utf-8") as f:
                raw = json.load(f)
            # GSTR-1 JSON may be flat or wrapped in {"data":{...}}
            if isinstance(raw, dict) and "data" in raw and isinstance(raw["data"], dict):
                data = raw["data"]
            else:
                data = raw
            period = data.get("fp","")
            fname  = os.path.basename(fpath)
            mdisp  = fname.replace("GSTR1_","").replace(".json","").replace("_"," ").strip()
            if not mdisp:
                mdisp = fp_to_display(period)
            months_data.append((mdisp, data))
        except Exception as e:
            log.warning(f"G1 Excel: error loading {fpath}: {e}")

    if not months_data:
        return None, "All JSON files failed to load"

    months_data.sort(key=lambda x: MONTH_ORDER.index(x[0].split()[0])
                     if x[0].split() and x[0].split()[0] in MONTH_ORDER else 99)

    # ── Load buyer names from per-return (1/) + enriched store ───────────────
    _g1_name_lookup = _gnames_load_for("1")   # {buyer_GSTIN: {trade_name, legal_name}}

    wb = Workbook()
    wb.remove(wb.active)

    # ═══════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ═══════════════════════════════════════════════════════
    # ── Build month-wise summary dict for the transposed Summary sheet ─────────
    _G1_MO_ORDER = ["April","May","June","July","August","September",
                    "October","November","December","January","February","March"]
    _FY_START_G1 = int(fy.split("-")[0])

    def _g1_mon_key(mdisp):
        parts = mdisp.split()
        try: return _G1_MO_ORDER.index(parts[0])
        except (ValueError, IndexError): return 99

    def _itm_tax(d):
        """
        Extract txval, igst, cgst, sgst, cess from an item dict.
        Handles BOTH formats:
          - Nested:  {"itm_det": {"rt":18, "txval":52618}}  → pass itm.get("itm_det",{})
          - Flat:    {"rt":18, "txval":52618}               → pass itm directly
        When called with an empty dict (old pattern on flat items), returns all zeros.
        """
        return (d.get("txval",0) or 0, d.get("iamt",0) or 0,
                d.get("camt",0) or 0,  d.get("samt",0) or 0, d.get("csamt",0) or 0)

    def _itm_data(itm):
        """
        Return the data dict from an item — handles both flat and itm_det-wrapped formats.
        Flat:   {"rt":18, "txval":52618}             → returns itm directly
        Nested: {"itm_det": {"rt":18, "txval":52618}} → returns itm["itm_det"]
        """
        return itm.get("itm_det", itm)

    from collections import defaultdict
    _g1_summary = defaultdict(lambda: {
        # B2B (registered buyers)
        "b2b_cnt":0,  "b2b_txval":0.0, "b2b_igst":0.0,"b2b_cgst":0.0,"b2b_sgst":0.0,"b2b_cess":0.0,
        # B2CL (inter-state unregistered > 2.5L)
        "b2cl_cnt":0, "b2cl_txval":0.0,"b2cl_igst":0.0,"b2cl_cess":0.0,
        # B2CS current (intra-state unregistered + small inter-state)
        "b2cs_cnt":0, "b2cs_txval":0.0,"b2cs_igst":0.0,"b2cs_cgst":0.0,"b2cs_sgst":0.0,"b2cs_cess":0.0,
        # B2CSA amended (separate totals so summary can split them)
        "b2csa_cnt":0,"b2csa_txval":0.0,"b2csa_cgst":0.0,"b2csa_sgst":0.0,"b2csa_cess":0.0,
        # CDN Registered
        "cdnr_cnt":0, "cdnr_txval":0.0,"cdnr_igst":0.0,"cdnr_cgst":0.0,"cdnr_sgst":0.0,"cdnr_cess":0.0,
        # CDN Unregistered
        "cdnur_cnt":0,"cdnur_txval":0.0,"cdnur_igst":0.0,"cdnur_cess":0.0,
        # Exports
        "exp_cnt":0,  "exp_txval":0.0, "exp_igst":0.0,"exp_cess":0.0,
        # Nil/Exempt/Non-GST
        "nil_exempt":0.0,"nil_nil":0.0,"nil_nongst":0.0,
        # Advances Received (Table 11A) — at + ata
        "at_cnt":0,  "at_adv":0.0,"at_igst":0.0,"at_cgst":0.0,"at_sgst":0.0,
        "ata_cnt":0, "ata_adv":0.0,"ata_igst":0.0,"ata_cgst":0.0,"ata_sgst":0.0,
        # Advance Tax Adjusted (Table 11B) — txpd + txpda
        "txpd_cnt":0, "txpd_adv":0.0,"txpd_igst":0.0,"txpd_cgst":0.0,"txpd_sgst":0.0,
        "txpda_cnt":0,"txpda_adv":0.0,"txpda_igst":0.0,"txpda_cgst":0.0,"txpda_sgst":0.0,
        # E-Commerce supply through operators (Table 14)
        "ecom_supply":0.0,"ecom_igst":0.0,"ecom_cgst":0.0,"ecom_sgst":0.0,"ecom_cess":0.0,
    })

    for mdisp, data in months_data:
        ms = _g1_summary[mdisp]
        # B2B + B2BA
        for sec in ["b2b","b2ba"]:
            for buyer in data.get(sec,[]):
                for inv in buyer.get("inv",[]):
                    ms["b2b_cnt"] += 1
                    for itm in inv.get("itms",[]):
                        tv,ig,cg,sg,cs = _itm_tax(_itm_data(itm))
                        ms["b2b_txval"]+=tv; ms["b2b_igst"]+=ig
                        ms["b2b_cgst"]+=cg;  ms["b2b_sgst"]+=sg; ms["b2b_cess"]+=cs
        # B2CL + B2CLA
        for sec in ["b2cl","b2cla"]:
            for grp in data.get(sec,[]):
                for inv in grp.get("inv",[]):
                    ms["b2cl_cnt"] += 1
                    for itm in inv.get("itms",[]):
                        tv,ig,_cg,_sg,cs = _itm_tax(_itm_data(itm))
                        ms["b2cl_txval"]+=tv; ms["b2cl_igst"]+=ig; ms["b2cl_cess"]+=cs
        # B2CS — flat records (current period)
        for r in data.get("b2cs",[]):
            ms["b2cs_cnt"]  += 1
            ms["b2cs_txval"] += r.get("txval",0) or 0
            ms["b2cs_igst"]  += r.get("iamt",0)  or 0
            ms["b2cs_cgst"]  += r.get("camt",0)  or 0
            ms["b2cs_sgst"]  += r.get("samt",0)  or 0
            ms["b2cs_cess"]  += r.get("csamt",0) or 0
        # B2CSA — amended records; each record has itms[] with per-rate rows
        # txval/camt/samt live at itm level NOT at the record level
        for r in data.get("b2csa",[]):
            for itm in r.get("itms",[]):
                ms["b2csa_cnt"]  += 1
                ms["b2csa_txval"] += itm.get("txval",0) or 0
                ms["b2csa_cgst"]  += itm.get("camt",0)  or 0
                ms["b2csa_sgst"]  += itm.get("samt",0)  or 0
                ms["b2csa_cess"]  += itm.get("csamt",0) or 0
        # CDN Registered (cdnr + cdnra)
        # Sign rule: Credit Note (ntty="C") reduces taxable supply → negative
        #            Debit Note  (ntty="D") increases taxable supply → positive
        for sec in ["cdnr","cdnra"]:
            for grp in data.get(sec,[]):
                for nt in grp.get("nt",[]):
                    ms["cdnr_cnt"] += 1
                    cdn_sign = -1 if nt.get("ntty","D") == "C" else 1
                    for itm in nt.get("itms",[]):
                        tv,ig,cg,sg,cs = _itm_tax(_itm_data(itm))
                        ms["cdnr_txval"] += tv * cdn_sign
                        ms["cdnr_igst"]  += ig * cdn_sign
                        ms["cdnr_cgst"]  += cg * cdn_sign
                        ms["cdnr_sgst"]  += sg * cdn_sign
                        ms["cdnr_cess"]  += cs * cdn_sign
        # CDN Unregistered (cdnur + cdnura)
        # Same sign rule: ntty="C" → Credit Note → negative
        for nt in data.get("cdnur",[]) + data.get("cdnura",[]):
            ms["cdnur_cnt"] += 1
            cdnur_sign = -1 if nt.get("ntty","D") == "C" else 1
            for itm in nt.get("itms",[]):
                tv,ig,_cg,_sg,cs = _itm_tax(_itm_data(itm))
                ms["cdnur_txval"] += tv * cdnur_sign
                ms["cdnur_igst"]  += ig * cdnur_sign
                ms["cdnur_cess"]  += cs * cdnur_sign
        # Exports (exp + expa)
        for sec in ["exp","expa"]:
            for grp in data.get(sec,[]):
                for inv in grp.get("inv",[]):
                    ms["exp_cnt"] += 1
                    for itm in inv.get("itms",[]):
                        tv,ig,_cg,_sg,cs = _itm_tax(_itm_data(itm))
                        ms["exp_txval"]+=tv; ms["exp_igst"]+=ig; ms["exp_cess"]+=cs
        # Nil/Exempt/Non-GST
        nil_raw = data.get("nil",{})
        inv_list = nil_raw.get("inv",[]) if isinstance(nil_raw,dict) else nil_raw
        for item in (inv_list if isinstance(inv_list,list) else []):
            if item:
                ms["nil_exempt"]  += item.get("expt_amt",0) or 0
                ms["nil_nil"]     += item.get("nil_amt",0)  or 0
                ms["nil_nongst"]  += item.get("ngsup_amt",0)or 0
        # Advances Received (at) — Table 11A, itms[] structure
        for entry in data.get("at",[]):
            for itm in entry.get("itms",[]):
                ms["at_cnt"]  += 1
                ms["at_adv"]  += itm.get("ad_amt",0) or 0
                ms["at_igst"] += itm.get("iamt",0)   or 0
                ms["at_cgst"] += itm.get("camt",0)   or 0
                ms["at_sgst"] += itm.get("samt",0)   or 0
        # Amendments to Advances (ata) — Table 11A, has omon
        for entry in data.get("ata",[]):
            for itm in entry.get("itms",[]):
                ms["ata_cnt"]  += 1
                ms["ata_adv"]  += itm.get("ad_amt",0) or 0
                ms["ata_igst"] += itm.get("iamt",0)   or 0
                ms["ata_cgst"] += itm.get("camt",0)   or 0
                ms["ata_sgst"] += itm.get("samt",0)   or 0
        # Advance Tax Adjusted (txpd) — Table 11B
        for entry in data.get("txpd",[]):
            for itm in entry.get("itms",[]):
                ms["txpd_cnt"]  += 1
                ms["txpd_adv"]  += itm.get("ad_amt",0) or 0
                ms["txpd_igst"] += itm.get("iamt",0)   or 0
                ms["txpd_cgst"] += itm.get("camt",0)   or 0
                ms["txpd_sgst"] += itm.get("samt",0)   or 0
        # E-Commerce Supply through Operators (supeco) — Table 14
        for r in (data.get("supeco") or {}).get("paytx",[]):
            ms["ecom_supply"] += r.get("suppval",0) or 0
            ms["ecom_cgst"]   += r.get("cgst",0)    or 0
            ms["ecom_sgst"]   += r.get("sgst",0)    or 0
            ms["ecom_igst"]   += r.get("igst",0)    or 0
            ms["ecom_cess"]   += r.get("cess",0)    or 0

        # Amendment to Advance Tax Adjusted (txpda) — Table 11B, has omon
        for entry in data.get("txpda",[]):
            for itm in entry.get("itms",[]):
                ms["txpda_cnt"]  += 1
                ms["txpda_adv"]  += itm.get("ad_amt",0) or 0
                ms["txpda_igst"] += itm.get("iamt",0)   or 0
                ms["txpda_cgst"] += itm.get("camt",0)   or 0
                ms["txpda_sgst"] += itm.get("samt",0)   or 0

    sorted_months_g1 = sorted(_g1_summary.keys(), key=_g1_mon_key)

    # ════════════════════════════════════════════════════════════════════════
    # 1-Summary SHEET — Column-per-month layout (Apr → Mar → Total)
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("1-Summary")

    def _g1_mon_col_lbl(mdisp):
        parts = mdisp.split()
        if len(parts) == 2:
            return f"{parts[0][:3]} '{parts[1][-2:]}"
        return mdisp[:8]

    _g1_mon_labels = [_g1_mon_col_lbl(m) for m in sorted_months_g1]
    _g1_n   = len(_g1_mon_labels)
    _g1_PC  = 1           # Particulars column
    _g1_MS  = 2           # Month columns start
    _g1_ME  = _g1_MS + _g1_n - 1
    _g1_TC  = _g1_ME + 1  # Total column
    _g1_NC  = _g1_TC      # total col count

    # Row 1: Title
    ws.merge_cells(f"A1:{get_column_letter(_g1_NC)}1")
    tc1 = ws["A1"]
    tc1.value = f"GSTR-1 | Month-wise Summary | GSTIN: {gstin} | FY: {fy} | Outward Supplies (Sales)"
    tc1.font      = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    tc1.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tc1.fill      = HDR_FILL
    ws.row_dimensions[1].height = 28

    # Row 2: Column headers
    ws.cell(row=2, column=_g1_PC, value="Particulars")
    for ci, lbl in enumerate(_g1_mon_labels, _g1_MS):
        ws.cell(row=2, column=ci, value=lbl)
    ws.cell(row=2, column=_g1_TC, value="Total")
    for ci in range(1, _g1_NC+1):
        cell = ws.cell(row=2, column=ci)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = Border(
            left=Side(style="thin",color="B0C4DE"),
            right=Side(style="thin",color="B0C4DE"),
            top=Side(style="thin",color="B0C4DE"),
            bottom=Side(style="thin",color="B0C4DE"))
    ws.row_dimensions[2].height = 30

    _g1_row = 3

    def _g1_sec(label, color="1B5E20"):
        nonlocal _g1_row
        ws.merge_cells(f"A{_g1_row}:{get_column_letter(_g1_NC)}{_g1_row}")
        c = ws.cell(row=_g1_row, column=1, value=label)
        c.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = Border(
            left=Side(style="thin",color="B0C4DE"),
            right=Side(style="thin",color="B0C4DE"),
            top=Side(style="thin",color="B0C4DE"),
            bottom=Side(style="thin",color="B0C4DE"))
        ws.row_dimensions[_g1_row].height = 18
        _g1_row += 1

    def _g1_row_write(label, extractor, is_count=False, bold=False):
        nonlocal _g1_row
        r = _g1_row
        lc = ws.cell(row=r, column=_g1_PC, value=label)
        lc.font      = Font(name="Arial", bold=bold, size=9)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border    = Border(
            left=Side(style="thin",color="B0C4DE"),
            right=Side(style="thin",color="B0C4DE"),
            top=Side(style="thin",color="B0C4DE"),
            bottom=Side(style="thin",color="B0C4DE"))
        for ci, m in enumerate(sorted_months_g1, _g1_MS):
            v = extractor(m)
            cell = ws.cell(row=r, column=ci, value=v if v else 0)
            cell.font   = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = Border(
                left=Side(style="thin",color="B0C4DE"),
                right=Side(style="thin",color="B0C4DE"),
                top=Side(style="thin",color="B0C4DE"),
                bottom=Side(style="thin",color="B0C4DE"))
            cell.number_format = '#,##0' if is_count else '#,##0.00'
        tc = ws.cell(row=r, column=_g1_TC)
        if _g1_n > 0:
            st = get_column_letter(_g1_MS); en = get_column_letter(_g1_ME)
            tc.value = f"=SUM({st}{r}:{en}{r})"
        else:
            tc.value = 0
        tc.font   = Font(name="Arial", bold=True, size=9)
        tc.fill   = TOT_FILL
        tc.alignment = Alignment(horizontal="right", vertical="center")
        tc.border = Border(
            left=Side(style="thin",color="B0C4DE"),
            right=Side(style="thin",color="B0C4DE"),
            top=Side(style="thin",color="B0C4DE"),
            bottom=Side(style="thin",color="B0C4DE"))
        tc.number_format = '#,##0' if is_count else '#,##0.00'
        ws.row_dimensions[r].height = 16
        _g1_row += 1

    s = _g1_summary  # alias

    # B2B
    _g1_sec("B2B — Registered Buyers (Table 4A/4B/4C/6B/6C)", "1B5E20")
    _g1_row_write("  Invoices (Count)",  lambda m: s[m]["b2b_cnt"],  is_count=True)
    _g1_row_write("  Taxable Value",     lambda m: s[m]["b2b_txval"])
    _g1_row_write("  IGST",              lambda m: s[m]["b2b_igst"])
    _g1_row_write("  CGST",              lambda m: s[m]["b2b_cgst"])
    _g1_row_write("  SGST",              lambda m: s[m]["b2b_sgst"])
    _g1_row_write("  Cess",              lambda m: s[m]["b2b_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["b2b_igst"]+s[m]["b2b_cgst"]+s[m]["b2b_sgst"]+s[m]["b2b_cess"],
                  bold=True)

    # B2CL
    _g1_sec("B2CL — Inter-State Unregistered >₹2.5L (Table 5A)", "2E7D32")
    _g1_row_write("  Invoices (Count)",  lambda m: s[m]["b2cl_cnt"],  is_count=True)
    _g1_row_write("  Taxable Value",     lambda m: s[m]["b2cl_txval"])
    _g1_row_write("  IGST",              lambda m: s[m]["b2cl_igst"])
    _g1_row_write("  Cess",              lambda m: s[m]["b2cl_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["b2cl_igst"]+s[m]["b2cl_cess"], bold=True)

    # B2CS
    _g1_sec("B2CS — Unregistered / Small Value, Current Period (Table 7)", "388E3C")
    _g1_row_write("  Records (Count)",       lambda m: s[m]["b2cs_cnt"],  is_count=True)
    _g1_row_write("  Taxable Value",         lambda m: s[m]["b2cs_txval"])
    _g1_row_write("  IGST",                  lambda m: s[m]["b2cs_igst"])
    _g1_row_write("  CGST",                  lambda m: s[m]["b2cs_cgst"])
    _g1_row_write("  SGST",                  lambda m: s[m]["b2cs_sgst"])
    _g1_row_write("  Cess",                  lambda m: s[m]["b2cs_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["b2cs_igst"]+s[m]["b2cs_cgst"]+s[m]["b2cs_sgst"]+s[m]["b2cs_cess"],
                  bold=True)
    _g1_sec("B2CSA — Amended B2CS (Table 7 Amendments)", "2E7D32")
    _g1_row_write("  Amended Records (Count)",lambda m: s[m]["b2csa_cnt"], is_count=True)
    _g1_row_write("  Taxable Value",          lambda m: s[m]["b2csa_txval"])
    _g1_row_write("  CGST",                   lambda m: s[m]["b2csa_cgst"])
    _g1_row_write("  SGST",                   lambda m: s[m]["b2csa_sgst"])
    _g1_row_write("  Cess",                   lambda m: s[m]["b2csa_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["b2csa_cgst"]+s[m]["b2csa_sgst"]+s[m]["b2csa_cess"],
                  bold=True)

    # CDN Registered
    _g1_sec("CDN Registered — Credit/Debit Notes (Table 9B)", "43A047")
    _g1_row_write("  Notes (Count)",     lambda m: s[m]["cdnr_cnt"],  is_count=True)
    _g1_row_write("  Taxable Value",     lambda m: s[m]["cdnr_txval"])
    _g1_row_write("  IGST",              lambda m: s[m]["cdnr_igst"])
    _g1_row_write("  CGST",              lambda m: s[m]["cdnr_cgst"])
    _g1_row_write("  SGST",              lambda m: s[m]["cdnr_sgst"])
    _g1_row_write("  Cess",              lambda m: s[m]["cdnr_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["cdnr_igst"]+s[m]["cdnr_cgst"]+s[m]["cdnr_sgst"]+s[m]["cdnr_cess"],
                  bold=True)

    # CDN Unregistered
    _g1_sec("CDN Unregistered — Credit/Debit Notes (Table 9B)", "558B2F")
    _g1_row_write("  Notes (Count)",     lambda m: s[m]["cdnur_cnt"], is_count=True)
    _g1_row_write("  Taxable Value",     lambda m: s[m]["cdnur_txval"])
    _g1_row_write("  IGST",              lambda m: s[m]["cdnur_igst"])
    _g1_row_write("  Cess",              lambda m: s[m]["cdnur_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["cdnur_igst"]+s[m]["cdnur_cess"], bold=True)

    # Exports
    _g1_sec("Exports (With / Without Payment of Tax) — Table 6A", "689F38")
    _g1_row_write("  Invoices (Count)",  lambda m: s[m]["exp_cnt"],   is_count=True)
    _g1_row_write("  Taxable Value",     lambda m: s[m]["exp_txval"])
    _g1_row_write("  IGST",              lambda m: s[m]["exp_igst"])
    _g1_row_write("  Cess",              lambda m: s[m]["exp_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["exp_igst"]+s[m]["exp_cess"], bold=True)

    # Nil / Exempt / Non-GST
    _g1_sec("Nil / Exempt / Non-GST Supplies — Table 8", "7CB342")
    _g1_row_write("  Exempt Amount",     lambda m: s[m]["nil_exempt"])
    _g1_row_write("  Nil-Rated Amount",  lambda m: s[m]["nil_nil"])
    _g1_row_write("  Non-GST Amount",    lambda m: s[m]["nil_nongst"])
    _g1_row_write("  Total",
                  lambda m: s[m]["nil_exempt"]+s[m]["nil_nil"]+s[m]["nil_nongst"], bold=True)

    # Advances (Table 11A — at + ata)
    _g1_sec("Advances Received — Table 11A (at) & Amendments (ata)", "0277BD")
    _g1_row_write("  Advances Received (Count)",    lambda m: s[m]["at_cnt"],   is_count=True)
    _g1_row_write("  Gross Advance Amount (at)",    lambda m: s[m]["at_adv"])
    _g1_row_write("  IGST on Advances",             lambda m: s[m]["at_igst"])
    _g1_row_write("  CGST on Advances",             lambda m: s[m]["at_cgst"])
    _g1_row_write("  SGST on Advances",             lambda m: s[m]["at_sgst"])
    _g1_row_write("  Total Tax on Advances",
                  lambda m: s[m]["at_igst"]+s[m]["at_cgst"]+s[m]["at_sgst"], bold=True)
    _g1_row_write("  Amended Advances (Count)",     lambda m: s[m]["ata_cnt"],  is_count=True)
    _g1_row_write("  Gross Advance Amount (ata)",   lambda m: s[m]["ata_adv"])
    _g1_row_write("  IGST on Amended Advances",     lambda m: s[m]["ata_igst"])
    _g1_row_write("  CGST on Amended Advances",     lambda m: s[m]["ata_cgst"])
    _g1_row_write("  SGST on Amended Advances",     lambda m: s[m]["ata_sgst"])
    _g1_row_write("  Total Tax on Amended Advances",
                  lambda m: s[m]["ata_igst"]+s[m]["ata_cgst"]+s[m]["ata_sgst"], bold=True)

    # Advance Tax Adjusted (Table 11B — txpd + txpda)
    _g1_sec("Advance Tax Adjusted — Table 11B (txpd) & Amendments (txpda)", "01579B")
    _g1_row_write("  Tax Adjusted (Count)",         lambda m: s[m]["txpd_cnt"], is_count=True)
    _g1_row_write("  Advance Amount Adjusted",      lambda m: s[m]["txpd_adv"])
    _g1_row_write("  IGST Adjusted",                lambda m: s[m]["txpd_igst"])
    _g1_row_write("  CGST Adjusted",                lambda m: s[m]["txpd_cgst"])
    _g1_row_write("  SGST Adjusted",                lambda m: s[m]["txpd_sgst"])
    _g1_row_write("  Total Tax Adjusted",
                  lambda m: s[m]["txpd_igst"]+s[m]["txpd_cgst"]+s[m]["txpd_sgst"], bold=True)
    _g1_row_write("  Amended Tax Adjusted (Count)", lambda m: s[m]["txpda_cnt"],is_count=True)
    _g1_row_write("  Advance Amount (txpda)",       lambda m: s[m]["txpda_adv"])
    _g1_row_write("  IGST Adjusted (Amended)",      lambda m: s[m]["txpda_igst"])
    _g1_row_write("  CGST Adjusted (Amended)",      lambda m: s[m]["txpda_cgst"])
    _g1_row_write("  SGST Adjusted (Amended)",      lambda m: s[m]["txpda_sgst"])
    _g1_row_write("  Total Tax Adjusted (Amended)",
                  lambda m: s[m]["txpda_igst"]+s[m]["txpda_cgst"]+s[m]["txpda_sgst"], bold=True)

    # E-Commerce Supply (Table 14 — supeco)
    _g1_sec("E-Commerce Supply through Operators — Table 14 (supeco/supecoa)", "0277BD")
    _g1_row_write("  Supply Value",         lambda m: s[m]["ecom_supply"])
    _g1_row_write("  IGST",                 lambda m: s[m]["ecom_igst"])
    _g1_row_write("  CGST",                 lambda m: s[m]["ecom_cgst"])
    _g1_row_write("  SGST",                 lambda m: s[m]["ecom_sgst"])
    _g1_row_write("  Cess",                 lambda m: s[m]["ecom_cess"])
    _g1_row_write("  Total Tax",
                  lambda m: s[m]["ecom_igst"]+s[m]["ecom_cgst"]+s[m]["ecom_sgst"]+s[m]["ecom_cess"],
                  bold=True)

    # Grand Total row
    _g1_sec("Gross Taxable Turnover (All Sections)", "1F3864")
    _g1_row_write("  Total Taxable Value (All Sections)",
                  lambda m: (s[m]["b2b_txval"]+s[m]["b2cl_txval"]+s[m]["b2cs_txval"]
                             +s[m]["cdnr_txval"]+s[m]["cdnur_txval"]+s[m]["exp_txval"]
                             +s[m]["nil_exempt"]+s[m]["nil_nil"]+s[m]["nil_nongst"]),
                  bold=True)
    _g1_row_write("  Total Tax (B2B+B2CL+B2CS+CDN)",
                  lambda m: (s[m]["b2b_igst"]+s[m]["b2b_cgst"]+s[m]["b2b_sgst"]+s[m]["b2b_cess"]
                             +s[m]["b2cl_igst"]+s[m]["b2cl_cess"]
                             +s[m]["b2cs_igst"]+s[m]["b2cs_cgst"]+s[m]["b2cs_sgst"]+s[m]["b2cs_cess"]
                             +s[m]["cdnr_igst"]+s[m]["cdnr_cgst"]+s[m]["cdnr_sgst"]+s[m]["cdnr_cess"]
                             +s[m]["cdnur_igst"]+s[m]["cdnur_cess"]
                             +s[m]["exp_igst"]+s[m]["exp_cess"]),
                  bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 50
    for ci in range(_g1_MS, _g1_TC + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 13
    ws.freeze_panes = "B3"

    # ═══════════════════════════════════════════════════════
    # SHEET 2 — B2B (Registered Buyers)
    # ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("2-B2B")
    HDR = ["Period","Amend","Buyer GSTIN","Trade Name","Legal Name",
           "Invoice No","Invoice Date",
           "Invoice Value","POS","Reverse Charge","Inv Type","IFF %",
           "Rate %","Taxable Value","IGST","CGST","SGST","Cess"]
    make_title(ws, len(HDR),
               "GSTR-1 — B2B Invoices (Registered Buyers) | Table 4A/4B/4C/6B/6C", "1B5E20")
    make_hdr(ws, 2, HDR)
    row = 3
    for mdisp, data in months_data:
        for is_amend, sec_key in [(False,"b2b"),(True,"b2ba")]:
            for buyer in data.get(sec_key,[]):
                ctin       = buyer.get("ctin","")
                _g1_nl     = _g1_name_lookup.get(ctin.strip().upper(), {})
                _g1_trd    = _g1_nl.get("trade_name", "")
                _g1_leg    = _g1_nl.get("legal_name",  "")
                for inv in buyer.get("inv",[]):
                    inum   = inv.get("inum","")
                    idt    = inv.get("idt","")
                    val    = inv.get("val",0) or 0
                    pos    = pos_name(inv.get("pos",""))
                    rchrg  = inv.get("rchrg","N")
                    itype  = INV_TYPES.get(inv.get("inv_typ","R"),
                                           inv.get("inv_typ","R"))
                    items  = inv.get("itms",[])
                    if not items:
                        fill = AMEND_FILL if is_amend else B2B_FILL
                        write_row(ws, row, [mdisp, "A" if is_amend else "",
                                            ctin, _g1_trd, _g1_leg,
                                            inum, idt, val, pos,
                                            rchrg, itype,
                                            inv.get("diff_percent",""),
                                            "", 0, 0, 0, 0, 0],
                                  row_fill=fill); row += 1
                    else:
                        for itm_i, itm in enumerate(items):
                            d = _itm_data(itm)
                            rt  = d.get("rt",0)
                            txv = d.get("txval",0) or 0
                            igt = d.get("iamt",0)  or 0
                            cgt = d.get("camt",0)  or 0
                            sgt = d.get("samt",0)  or 0
                            csa = d.get("csamt",0) or 0
                            fill = AMEND_FILL if is_amend else (
                                B2B_FILL if itm_i == 0 else ALT_FILL)
                            write_row(ws, row, [
                                mdisp  if itm_i==0 else "",
                                "A"    if is_amend else "",
                                ctin   if itm_i==0 else "",
                                _g1_trd if itm_i==0 else "",
                                _g1_leg if itm_i==0 else "",
                                inum   if itm_i==0 else "",
                                idt    if itm_i==0 else "",
                                val    if itm_i==0 else "",
                                pos    if itm_i==0 else "",
                                rchrg  if itm_i==0 else "",
                                itype  if itm_i==0 else "",
                                inv.get("diff_percent","") if itm_i==0 else "",
                                rt, txv, igt, cgt, sgt, csa,
                            ], row_fill=fill)
                            row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    tot_row(ws, row, len(HDR), [14,15,16,17,18], ds=3)
    set_widths(ws, [14,6,20,28,28,18,12,14,18,10,16,6,8,14,12,12,12,10])
    ws.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════
    # SHEET 3 — B2CL (Unregistered, Inter-state, > ₹2.5L)
    # ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("3-B2CL")
    HDR = ["Period","Amend","POS","Invoice No","Invoice Date","Invoice Value",
           "Rate %","Taxable Value","IGST","Cess"]
    make_title(ws, len(HDR),
               "GSTR-1 — B2CL (Inter-state, Unregistered, Invoice > ₹2.5L) | Table 5A", "2E7D32")
    make_hdr(ws, 2, HDR)
    row = 3
    for mdisp, data in months_data:
        for is_amend, sec_key in [(False,"b2cl"),(True,"b2cla")]:
            for grp in data.get(sec_key,[]):
                pos = pos_name(grp.get("pos",""))
                for inv in grp.get("inv",[]):
                    inum = inv.get("inum",""); idt = inv.get("idt","")
                    val  = inv.get("val",0) or 0
                    items = inv.get("itms",[])
                    for itm_i, itm in enumerate(items or [{}]):
                        d = _itm_data(itm)
                        fill = AMEND_FILL if is_amend else B2CL_FILL
                        write_row(ws, row, [
                            mdisp if itm_i==0 else "",
                            "A" if is_amend else "",
                            pos  if itm_i==0 else "",
                            inum if itm_i==0 else "",
                            idt  if itm_i==0 else "",
                            val  if itm_i==0 else "",
                            d.get("rt",0), d.get("txval",0) or 0,
                            d.get("iamt",0) or 0, d.get("csamt",0) or 0,
                        ], row_fill=fill); row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    tot_row(ws, row, len(HDR), [8,9,10], ds=3)
    set_widths(ws, [14,6,18,18,12,14,8,14,12,10])
    ws.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════
    # SHEET 4 — B2CS + B2CSA (Unregistered, intra-state + small inter-state)
    # ═══════════════════════════════════════════════════════
    # b2cs:  flat record — rt/txval/iamt/camt/samt/csamt at record level
    # b2csa: nested — one record per pos+omon, itms[] has per-rate rows;
    #        txval/camt/samt/csamt are at itm level (NOT record level)
    #        omon = original month (MMYYYY format) that is being amended
    ws = wb.create_sheet("4-B2CS")
    HDR = ["Period","Section","Orig Month","Supply Type","Type","Rate %","POS",
           "Taxable Value","IGST","CGST","SGST","Cess","E-Commerce GSTIN"]
    make_title(ws, len(HDR),
               "GSTR-1 — B2CS & B2CSA (Intra-state + Small Inter-state Unregistered) | Table 7", "388E3C")
    make_hdr(ws, 2, HDR)
    row = 3
    SPLY_TY = {"INTER":"Inter-state","INTRA":"Intra-state"}  # defined at function scope

    def _omon_display(omon):
        """Convert MMYYYY to 'Mon YYYY', e.g. '032025' → 'Mar 2025'."""
        _MN = {"01":"Jan","02":"Feb","03":"Mar","04":"Apr","05":"May","06":"Jun",
               "07":"Jul","08":"Aug","09":"Sep","10":"Oct","11":"Nov","12":"Dec"}
        if omon and len(omon) >= 6:
            return f"{_MN.get(omon[:2], omon[:2])} {omon[2:6]}"
        return omon or ""

    for mdisp, data in months_data:
        # Current period B2CS — flat structure
        for r in data.get("b2cs", []):
            write_row(ws, row, [
                mdisp, "Current", "",
                SPLY_TY.get(r.get("sply_ty",""), r.get("sply_ty","")),
                r.get("typ",""),
                r.get("rt",0),
                pos_name(r.get("pos","")),
                r.get("txval",0) or 0,
                r.get("iamt",0)  or 0,
                r.get("camt",0)  or 0,
                r.get("samt",0)  or 0,
                r.get("csamt",0) or 0,
                r.get("etin",""),
            ], row_fill=B2CS_FILL); row += 1

        # Amended B2CS — nested itms[] structure, one row per rate per pos+omon
        for r in data.get("b2csa", []):
            orig_mon = _omon_display(r.get("omon",""))
            pos      = pos_name(r.get("pos",""))
            sply_ty  = SPLY_TY.get(r.get("sply_ty",""), r.get("sply_ty",""))
            typ      = r.get("typ","")
            for itm in r.get("itms", []):
                write_row(ws, row, [
                    mdisp, "Amendment", orig_mon,
                    sply_ty, typ,
                    itm.get("rt",0),
                    pos,
                    itm.get("txval",0)  or 0,
                    0,                          # IGST = 0 for intra-state B2CS
                    itm.get("camt",0)   or 0,
                    itm.get("samt",0)   or 0,
                    itm.get("csamt",0)  or 0,
                    r.get("etin",""),
                ], row_fill=AMEND_FILL); row += 1

    ws.cell(row=row, column=1, value="TOTAL")
    tot_row(ws, row, len(HDR), [8,9,10,11,12], ds=3)
    set_widths(ws, [14,10,11,12,6,8,18,14,12,12,12,10,20])
    ws.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════
    # SHEET 5 — Credit / Debit Notes (Registered + Unregistered)
    # ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("5-CDN")
    HDR = ["Period","Amend","Type","Counterparty GSTIN","Trade Name","Legal Name",
           "Note Type",
           "Note No","Note Date","Note Value","POS","Reverse Charge",
           "Inv Type","Rate %","Taxable Value","IGST","CGST","SGST","Cess"]
    make_title(ws, len(HDR),
               "GSTR-1 — Credit/Debit Notes (Registered + Unregistered) | Table 9B", "43A047")
    make_hdr(ws, 2, HDR)
    row = 3

    def write_cdn_section(data_dict, sec_key, party_type, is_amend):
        nonlocal row
        nt_key = "nt" if not is_amend else "nt"
        for grp in data_dict.get(sec_key,[]):
            ctin       = grp.get("ctin","") if party_type == "Registered" else "—"
            # Look up Trade/Legal name for this buyer (same lookup as B2B sheet)
            _cdn_nl    = _g1_name_lookup.get(ctin.strip().upper(), {})
            _cdn_trd   = _cdn_nl.get("trade_name", "")
            _cdn_leg   = _cdn_nl.get("legal_name",  "")
            for nt in grp.get("nt",[]) if not is_amend else grp.get("nt",[]):
                ntty   = NOTE_TYPES.get(nt.get("ntty",""),nt.get("ntty",""))
                nt_num = nt.get("nt_num",""); nt_dt = nt.get("nt_dt","")
                val    = nt.get("val",0) or 0
                pos    = pos_name(nt.get("pos","")) if party_type=="Registered" else ""
                rchrg  = nt.get("rchrg","N")
                itype  = INV_TYPES.get(nt.get("inv_typ","R"),nt.get("inv_typ","R"))
                items  = nt.get("itms",[])
                is_credit  = nt.get("ntty","D") == "C"
                det_sign   = -1 if is_credit else 1
                base_fill  = CDN_CR if is_credit else CDN_DR
                fill = AMEND_FILL if is_amend else base_fill
                for itm_i, itm in enumerate(items or [{}]):
                    d = _itm_data(itm)
                    write_row(ws, row, [
                        mdisp      if itm_i==0 else "",
                        "A"        if is_amend else "",
                        party_type,
                        ctin       if itm_i==0 else "",
                        _cdn_trd   if itm_i==0 else "",
                        _cdn_leg   if itm_i==0 else "",
                        ntty       if itm_i==0 else "",
                        nt_num     if itm_i==0 else "",
                        nt_dt      if itm_i==0 else "",
                        val        if itm_i==0 else "",
                        pos        if itm_i==0 else "",
                        rchrg      if itm_i==0 else "",
                        itype      if itm_i==0 else "",
                        d.get("rt",0),
                        (d.get("txval",0) or 0) * det_sign,
                        (d.get("iamt",0)  or 0) * det_sign,
                        (d.get("camt",0)  or 0) * det_sign,
                        (d.get("samt",0)  or 0) * det_sign,
                        (d.get("csamt",0) or 0) * det_sign,
                    ], row_fill=fill); row += 1

    for mdisp, data in months_data:
        write_cdn_section(data,"cdnr","Registered",False)
        write_cdn_section(data,"cdnra","Registered",True)
        write_cdn_section(data,"cdnur","Unregistered",False)
        write_cdn_section(data,"cdnura","Unregistered",True)

    ws.cell(row=row, column=1, value="TOTAL")
    tot_row(ws, row, len(HDR), [15,16,17,18,19], ds=3)   # shifted +2 for Trade+Legal
    set_widths(ws, [14,6,14,20,28,28,12,16,12,14,18,10,16,8,14,12,12,12,10])
    ws.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════
    # SHEET 6 — Exports
    # ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("6-Exports")
    HDR = ["Period","Amend","Export Type","Invoice No","Invoice Date",
           "Invoice Value","Port Code","Shipping Bill No","Shipping Bill Date",
           "Rate %","Taxable Value","IGST","Cess"]
    make_title(ws, len(HDR),
               "GSTR-1 — Exports (With/Without Payment) | Table 6A", "558B2F")
    make_hdr(ws, 2, HDR)
    row = 3
    for mdisp, data in months_data:
        for is_amend, sec_key in [(False,"exp"),(True,"expa")]:
            for grp in data.get(sec_key,[]):
                exp_typ = EXP_TYPES.get(grp.get("exp_typ",""),grp.get("exp_typ",""))
                for inv in grp.get("inv",[]):
                    inum   = inv.get("inum",""); idt = inv.get("idt","")
                    val    = inv.get("val",0) or 0
                    pcode  = inv.get("sbpcode","")
                    sbnum  = inv.get("sbnum",""); sbdt = inv.get("sbdt","")
                    items  = inv.get("itms",[])
                    fill   = AMEND_FILL if is_amend else EXP_FILL
                    for itm_i, itm in enumerate(items or [{}]):
                        d = _itm_data(itm)
                        write_row(ws, row, [
                            mdisp if itm_i==0 else "",
                            "A" if is_amend else "",
                            exp_typ if itm_i==0 else "",
                            inum if itm_i==0 else "",
                            idt  if itm_i==0 else "",
                            val  if itm_i==0 else "",
                            pcode if itm_i==0 else "",
                            sbnum if itm_i==0 else "",
                            sbdt  if itm_i==0 else "",
                            d.get("rt",0), d.get("txval",0) or 0,
                            d.get("iamt",0) or 0, d.get("csamt",0) or 0,
                        ], row_fill=fill); row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    tot_row(ws, row, len(HDR), [11,12,13], ds=3)
    set_widths(ws, [14,6,14,18,12,14,12,18,14,8,14,12,10])
    ws.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════
    # SHEET 7 — NIL / Exempt / Non-GST Supplies
    # ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("7-NIL-Exempt")
    HDR = ["Period","Supply Type","Exempt Amount","Nil-Rated Amount","Non-GST Amount","Total"]
    make_title(ws, len(HDR),
               "GSTR-1 — Nil, Exempt, Non-GST Supplies | Table 8A/8B/8C/8D", "689F38")
    make_hdr(ws, 2, HDR)
    row = 3
    NIL_SPLY = {
        "INTRB2B": "Inter-state B2B",
        "INTRAB2B":"Intra-state B2B",
        "INTRB2C": "Inter-state B2C",
        "INTRAB2C":"Intra-state B2C",
        "INTER":   "Inter-state",
        "INTRA":   "Intra-state",
    }
    for mdisp, data in months_data:
        nil_data = data.get("nil",{})
        inv_list = nil_data.get("inv",[]) if isinstance(nil_data,dict) else nil_data
        for item in (inv_list if isinstance(inv_list,list) else [inv_list]):
            if not item: continue
            raw_ty  = item.get("sply_ty","")
            sply_ty = NIL_SPLY.get(raw_ty, raw_ty)
            expt = item.get("expt_amt",0) or 0
            nil  = item.get("nil_amt",0)  or 0
            ngs  = item.get("ngsup_amt",0) or 0
            write_row(ws, row, [mdisp, sply_ty, expt, nil, ngs, expt+nil+ngs],
                      row_fill=NIL_FILL); row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    tot_row(ws, row, len(HDR), [3,4,5,6], ds=3)
    set_widths(ws, [14,14,18,18,18,16])

    # ═══════════════════════════════════════════════════════
    # SHEET 8 — HSN Summary (mandatory from Jan 2025 Phase 3)
    # Portal JSON schema (Jan 2026):
    #   hsn is a DICT: {flag, hsn_b2b:[...], hsn_b2c:[...], chksum}
    #   Each item: {num, hsn_sc, desc, user_desc, uqc, qty,
    #               rt, txval, iamt, camt, samt, csamt}
    # ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("8-HSN Summary")
    HDR = ["Period","Section","HSN/SAC","User Description","System Description",
           "UQC","Quantity","Rate %","Taxable Value","IGST","CGST","SGST","Cess"]
    make_title(ws, len(HDR),
               "GSTR-1 — HSN/SAC Summary | Table 12 (Mandatory Phase 3 from Jan 2025)"
               "  |  B2B = Registered Supplies  ·  B2C = Unregistered/Consumer Supplies", "7CB342")
    make_hdr(ws, 2, HDR)
    row = 3
    HSN_SECS = [("hsn_b2b","B2B"), ("hsn_b2c","B2C")]
    for mdisp, data in months_data:
        hsn_raw = data.get("hsn",{})
        # Handle both old format {data:[...]} and new format {hsn_b2b:[...], hsn_b2c:[...]}
        if isinstance(hsn_raw, dict):
            if "hsn_b2b" in hsn_raw or "hsn_b2c" in hsn_raw:
                # New portal format (Jan 2025+)
                item_sources = [(sec_lbl, hsn_raw.get(sec_key,[]))
                                for sec_key, sec_lbl in HSN_SECS]
            elif "data" in hsn_raw:
                # Old offline tool format
                item_sources = [("Combined", hsn_raw.get("data",[]))]
            else:
                item_sources = []
        elif isinstance(hsn_raw, list):
            item_sources = [("Combined", hsn_raw)]
        else:
            item_sources = []

        for sec_lbl, items in item_sources:
            for itm in items:
                fill = ALT_FILL if row % 2 == 0 else HSN_FILL
                write_row(ws, row, [
                    mdisp,
                    sec_lbl,
                    itm.get("hsn_sc",""),
                    itm.get("user_desc",""),
                    itm.get("desc",""),
                    itm.get("uqc","NA"),
                    itm.get("qty",0)   or 0,
                    itm.get("rt",0),
                    itm.get("txval",0) or 0,
                    itm.get("iamt",0)  or 0,
                    itm.get("camt",0)  or 0,
                    itm.get("samt",0)  or 0,
                    itm.get("csamt",0) or 0,
                ], row_fill=fill); row += 1
    ws.cell(row=row, column=1, value="TOTAL")
    tot_row(ws, row, len(HDR), [7,9,10,11,12,13], ds=3)
    set_widths(ws, [14,8,14,26,20,8,10,8,16,12,12,12,10])
    ws.freeze_panes = "C3"

    # ═══════════════════════════════════════════════════════
    # SHEET 9 — Document Summary
    # ═══════════════════════════════════════════════════════
    ws = wb.create_sheet("9-Doc Summary")
    HDR = ["Period","Doc Type","From Serial","To Serial","Total Issued",
           "Cancelled","Net Issued"]
    make_title(ws, len(HDR),
               "GSTR-1 — Document Issued Summary | Table 13", "8BC34A")
    make_hdr(ws, 2, HDR)
    row = 3
    DOC_TYPES = {
        "1":"Tax Invoice","2":"Supplementary Invoice","3":"Revised Invoice",
        "4":"Credit Note","5":"Debit Note","6":"Delivery Challan",
        "7":"Bill of Supply","8":"Receipt Voucher","9":"Payment Voucher",
        "10":"Refund Voucher",
    }
    for mdisp, data in months_data:
        # Portal JSON: "doc_issue" key holds {flag, doc_det:[...]}
        # Offline tool: "docs" key with same structure
        docs_raw = data.get("doc_issue", data.get("docs",{}))
        doc_det = docs_raw.get("doc_det",[]) if isinstance(docs_raw,dict) else []
        for doc_grp in doc_det:
            doc_type = DOC_TYPES.get(str(doc_grp.get("doc_num","")),
                                     f"Type {doc_grp.get('doc_num','')}")
            for d in doc_grp.get("docs",[]):
                frm  = d.get("from",""); to_ = d.get("to","")
                tot  = d.get("totnum",0) or 0
                canc = d.get("cancel",0) or 0
                net  = d.get("net_issue", tot - canc)
                write_row(ws, row, [mdisp, doc_type, frm, to_, tot, canc, net])
                row += 1
    set_widths(ws, [14,20,20,20,14,14,14])

    # ═══════════════════════════════════════════════════════
    # SHEET 10 — Advances (Tables 11A/11B) — at, ata, txpd, txpda
    # ═══════════════════════════════════════════════════════
    # Table 11A: at  = Advances Received (current period)
    #            ata = Amendments to Advances Received (has omon = original month)
    # Table 11B: txpd  = Advance Tax Adjusted (current period)
    #            txpda = Amendments to Advance Tax Adjusted (has omon)
    # All four use itms[] array for per-rate rows.
    # at/txpd have no omon; ata/txpda have omon (MMYYYY = original month being amended).
    has_at = any(d.get("at") or d.get("ata") or d.get("txpd") or d.get("txpda")
                 for _, d in months_data)
    if has_at:
        ws = wb.create_sheet("10-Advances")
        HDR = ["Period","Table","Section","Orig Month (Amended)",
               "Supply Type","Place of Supply",
               "Rate %","Gross Advance / Adj Amt","IGST","CGST","SGST","Cess"]
        make_title(ws, len(HDR),
                   "GSTR-1 — Advances Received & Adjusted | Tables 11A (at/ata) & 11B (txpd/txpda)",
                   "0277BD")
        make_hdr(ws, 2, HDR)
        row = 3

        ADV_FILL  = PatternFill("solid", fgColor="E3F2FD")   # at  — light blue
        ATA_FILL  = PatternFill("solid", fgColor="FFF3E0")   # ata — amber (amendment)
        TXPD_FILL = PatternFill("solid", fgColor="E8F5E9")   # txpd — light green
        TXPDA_FILL= PatternFill("solid", fgColor="FCE4EC")   # txpda — light pink (amendment)

        SEC_CFG = [
            # (json_key, table_label, section_label, has_omon, fill)
            ("at",    "11A", "Advance Received",             False, ADV_FILL),
            ("ata",   "11A", "Advance Received (Amended)",   True,  ATA_FILL),
            ("txpd",  "11B", "Advance Tax Adjusted",         False, TXPD_FILL),
            ("txpda", "11B", "Advance Tax Adjusted (Amended)",True, TXPDA_FILL),
        ]

        for mdisp, data in months_data:
            for sec_key, tbl_lbl, sec_lbl, has_omon, sec_fill in SEC_CFG:
                for entry in data.get(sec_key, []):
                    pos     = pos_name(entry.get("pos",""))
                    sply_ty = SPLY_TY.get(entry.get("sply_ty",""), entry.get("sply_ty",""))
                    omon    = _omon_display(entry.get("omon","")) if has_omon else ""
                    for itm in entry.get("itms", []):
                        # Handle both flat itm and nested itm_det (portal variant)
                        d = _itm_data(itm)
                        write_row(ws, row, [
                            mdisp, tbl_lbl, sec_lbl, omon,
                            sply_ty, pos,
                            d.get("rt",0),
                            d.get("ad_amt",0) or 0,
                            d.get("iamt",0)   or 0,
                            d.get("camt",0)   or 0,
                            d.get("samt",0)   or 0,
                            d.get("csamt",0)  or 0,
                        ], row_fill=sec_fill); row += 1

        ws.cell(row=row, column=1, value="TOTAL")
        tot_row(ws, row, len(HDR), [8,9,10,11,12], ds=3)
        set_widths(ws, [14,7,26,16,12,18,8,18,12,12,12,10])


    # ═══════════════════════════════════════════════════════
    # SHEET 11 — Supply through E-Commerce Operators
    # supeco  = Table 14: current period   (paytx = supplier pays tax)
    # supecoa = Table 14: amendments        (paytxa / clttxa)
    #   paytx  : etin, suppval, igst, cgst, sgst, cess
    #   clttxa : oetin, etin, suppval, igst, cgst, sgst, cess, omon
    #   paytxa : omon, oetin, etin, suppval, igst, cgst, sgst, cess
    # ═══════════════════════════════════════════════════════
    has_supeco = any(
        d.get("supeco") or d.get("supecoa") for _, d in months_data)
    if has_supeco:
        SECO_FILL  = PatternFill("solid", fgColor="E8EAF6")   # current — indigo tint
        SECOA_FILL = PatternFill("solid", fgColor="FFF3E0")   # amended — amber

        ws11 = wb.create_sheet("11-EComm-Supply-14")
        HDR11 = ["Period","Section","Orig Month","E-Comm GSTIN (ETIN)",
                 "Orig ETIN","Supply Value","IGST","CGST","SGST","Cess","Total Tax"]
        make_title(ws11, len(HDR11),
                   "GSTR-1 — Supply through E-Commerce Operators | Table 14 (supeco/supecoa)",
                   "3949AB")
        make_hdr(ws11, 2, HDR11)
        row11 = 3

        for mdisp, data in months_data:
            # paytx — current period, supplier pays tax
            for r in (data.get("supeco") or {}).get("paytx", []):
                write_row(ws11, row11, [
                    mdisp, "Current (Supplier pays tax)", "",
                    r.get("etin",""), "",
                    r.get("suppval",0) or 0,
                    r.get("igst",0)   or 0,
                    r.get("cgst",0)   or 0,
                    r.get("sgst",0)   or 0,
                    r.get("cess",0)   or 0,
                    (r.get("igst",0) or 0)+(r.get("cgst",0) or 0)+
                    (r.get("sgst",0) or 0)+(r.get("cess",0) or 0),
                ], row_fill=SECO_FILL); row11 += 1

            # supecoa paytxa — amended (supplier pays)
            for r in (data.get("supecoa") or {}).get("paytxa", []):
                write_row(ws11, row11, [
                    mdisp, "Amendment (Supplier pays)", _omon_display(r.get("omon","")),
                    r.get("etin",""), r.get("oetin",""),
                    r.get("suppval",0) or 0,
                    r.get("igst",0)   or 0,
                    r.get("cgst",0)   or 0,
                    r.get("sgst",0)   or 0,
                    r.get("cess",0)   or 0,
                    (r.get("igst",0) or 0)+(r.get("cgst",0) or 0)+
                    (r.get("sgst",0) or 0)+(r.get("cess",0) or 0),
                ], row_fill=SECOA_FILL); row11 += 1

            # supecoa clttxa — amended (e-comm collects tax)
            for r in (data.get("supecoa") or {}).get("clttxa", []):
                write_row(ws11, row11, [
                    mdisp, "Amendment (E-Comm collects)", _omon_display(r.get("omon","")),
                    r.get("etin",""), r.get("oetin",""),
                    r.get("suppval",0) or 0,
                    r.get("igst",0)   or 0,
                    r.get("cgst",0)   or 0,
                    r.get("sgst",0)   or 0,
                    r.get("cess",0)   or 0,
                    (r.get("igst",0) or 0)+(r.get("cgst",0) or 0)+
                    (r.get("sgst",0) or 0)+(r.get("cess",0) or 0),
                ], row_fill=SECOA_FILL); row11 += 1

        ws11.cell(row=row11, column=1, value="TOTAL")
        tot_row(ws11, row11, len(HDR11), [6,7,8,9,10,11], ds=3)
        set_widths(ws11, [14,28,12,22,20,16,12,12,12,10,12])

    # ═══════════════════════════════════════════════════════
    # SHEET 12 — E-Commerce Operator Sales  (Table 8A)
    # ecom  = current period  { b2b, b2c, urp2b, urp2c }
    # ecoma = amendments      { b2ba, b2ca, urp2ba, urp2ca }
    #
    # Sub-section key:
    #   b2b   — E-Comm B2B (seller GSTIN known, buyer registered)
    #            stin=seller, rtin=recipient, inv[] with itm_det
    #   b2c   — E-Comm B2C (flat per seller+pos+rate)
    #            stin, pos, rt, txval, iamt/camt/samt, csamt
    #   urp2b — E-Comm Unregistered Seller → Registered Buyer
    #            rtin=recipient, inv[] with itm_det
    #   urp2c — E-Comm Unregistered Seller → Consumer (flat)
    #            pos, txval, rt, iamt/camt/samt, csamt
    #   *a    — Amendment variants (same structure + oinum/oidt or omon)
    # ═══════════════════════════════════════════════════════
    has_ecom = any(d.get("ecom") or d.get("ecoma") for _, d in months_data)
    if has_ecom:
        ECOM_B2B_FILL  = PatternFill("solid", fgColor="E3F2FD")
        ECOM_B2C_FILL  = PatternFill("solid", fgColor="E8F5E9")
        ECOM_URP_FILL  = PatternFill("solid", fgColor="FFF8E1")
        ECOM_AMEND_FILL= PatternFill("solid", fgColor="FFF3E0")

        ws12 = wb.create_sheet("12-EComm-Sales-8A")
        HDR12 = ["Period","Section","Amend","Seller GSTIN","Recipient GSTIN",
                 "Invoice No","Invoice Date","Orig Inv No","Orig Inv Dt",
                 "Orig Month","Invoice Value","POS","Supply Type","Inv Type",
                 "Rate %","Taxable Value","IGST","CGST","SGST","Cess"]
        make_title(ws12, len(HDR12),
                   "GSTR-1 — E-Commerce Operator Sales | Table 8A (ecom / ecoma)",
                   "0288D1")
        make_hdr(ws12, 2, HDR12)
        row12 = 3

        def _ecom_write(ws, row, mdisp, section, is_amend,
                        stin, rtin, inum, idt, oinum, oidt, omon,
                        val, pos, sply_ty, inv_typ,
                        rt, txv, igt, cgt, sgt, csa, fill):
            write_row(ws, row, [
                mdisp, section, "A" if is_amend else "",
                stin, rtin, inum, idt, oinum, oidt,
                _omon_display(omon) if omon else "",
                val, pos_name(pos) if pos else "",
                SPLY_TY.get(sply_ty, sply_ty),
                INV_TYPES.get(inv_typ, inv_typ),
                rt, txv, igt, cgt, sgt, csa,
            ], row_fill=fill)

        for mdisp, data in months_data:
            # ── CURRENT PERIOD (ecom) ────────────────────────────────────
            ecom_d = data.get("ecom") or {}

            # b2b — E-Comm B2B (invoice-level)
            for grp in ecom_d.get("b2b", []):
                stin = grp.get("stin","")
                rtin = grp.get("rtin","")
                for inv in grp.get("inv", []):
                    for i, itm in enumerate(inv.get("itms", [])):
                        d = _itm_data(itm)
                        _ecom_write(ws12, row12, mdisp,
                            "E-Comm B2B", False,
                            stin, rtin,
                            inv.get("inum","") if i==0 else "",
                            inv.get("idt","")  if i==0 else "",
                            "","","",
                            inv.get("val",0) if i==0 else "",
                            inv.get("pos",""),
                            inv.get("sply_ty",""),
                            inv.get("inv_typ",""),
                            d.get("rt",0), d.get("txval",0) or 0,
                            d.get("iamt",0) or 0, d.get("camt",0) or 0,
                            d.get("samt",0) or 0, d.get("csamt",0) or 0,
                            ECOM_B2B_FILL); row12 += 1

            # b2c — E-Comm B2C flat (one row per entry)
            for r in ecom_d.get("b2c", []):
                _ecom_write(ws12, row12, mdisp,
                    "E-Comm B2C (flat)", False,
                    r.get("stin",""), "",
                    "","","","","",
                    "", r.get("pos",""), r.get("sply_ty",""), "",
                    r.get("rt",0),
                    r.get("txval",0) or 0,
                    r.get("iamt",0)  or 0,
                    r.get("camt",0)  or 0,
                    r.get("samt",0)  or 0,
                    r.get("csamt",0) or 0,
                    ECOM_B2C_FILL); row12 += 1

            # urp2b — Unregistered seller → registered buyer
            for grp in ecom_d.get("urp2b", []):
                rtin = grp.get("rtin","")
                for inv in grp.get("inv", []):
                    for i, itm in enumerate(inv.get("itms", [])):
                        d = _itm_data(itm)
                        _ecom_write(ws12, row12, mdisp,
                            "E-Comm Unreg Seller→Reg Buyer", False,
                            "", rtin,
                            inv.get("inum","") if i==0 else "",
                            inv.get("idt","")  if i==0 else "",
                            "","","",
                            inv.get("val",0) if i==0 else "",
                            inv.get("pos",""),
                            inv.get("sply_ty",""),
                            inv.get("inv_typ",""),
                            d.get("rt",0), d.get("txval",0) or 0,
                            d.get("iamt",0) or 0, d.get("camt",0) or 0,
                            d.get("samt",0) or 0, d.get("csamt",0) or 0,
                            ECOM_URP_FILL); row12 += 1

            # urp2c — Unregistered seller → consumer flat
            for r in ecom_d.get("urp2c", []):
                _ecom_write(ws12, row12, mdisp,
                    "E-Comm Unreg Seller→Consumer", False,
                    "", "",
                    "","","","","",
                    "", r.get("pos",""), r.get("sply_ty",""), "",
                    r.get("rt",0),
                    r.get("txval",0) or 0,
                    r.get("iamt",0)  or 0,
                    r.get("camt",0)  or 0,
                    r.get("samt",0)  or 0,
                    r.get("csamt",0) or 0,
                    ECOM_URP_FILL); row12 += 1

            # ── AMENDMENTS (ecoma) ───────────────────────────────────────
            ecoma_d = data.get("ecoma") or {}

            # b2ba — amended B2B
            for grp in ecoma_d.get("b2ba", []):
                stin = grp.get("stin","")
                rtin = grp.get("rtin","")
                for inv in grp.get("inv", []):
                    for i, itm in enumerate(inv.get("itms", [])):
                        d = _itm_data(itm)
                        _ecom_write(ws12, row12, mdisp,
                            "E-Comm B2B Amended", True,
                            stin, rtin,
                            inv.get("inum","")  if i==0 else "",
                            inv.get("idt","")   if i==0 else "",
                            inv.get("oinum","") if i==0 else "",
                            inv.get("oidt","")  if i==0 else "",
                            "",
                            inv.get("val",0) if i==0 else "",
                            inv.get("pos",""),
                            inv.get("sply_ty",""),
                            inv.get("inv_typ",""),
                            d.get("rt",0), d.get("txval",0) or 0,
                            d.get("iamt",0) or 0, d.get("camt",0) or 0,
                            d.get("samt",0) or 0, d.get("csamt",0) or 0,
                            ECOM_AMEND_FILL); row12 += 1

            # b2ca — amended B2C (has posItms[] nested)
            for grp in ecoma_d.get("b2ca", []):
                outer_pos = grp.get("pos","")
                for pi in grp.get("posItms", []):
                    stin = pi.get("stin","")
                    omon = pi.get("omon","")
                    for itm in pi.get("itms", []):
                        _ecom_write(ws12, row12, mdisp,
                            "E-Comm B2C Amended", True,
                            stin, "", "","","","",
                            omon,
                            "", outer_pos or pi.get("pos",""),
                            pi.get("sply_ty",""), "",
                            itm.get("rt",0),
                            itm.get("txval",0) or 0,
                            itm.get("iamt",0)  or 0,
                            itm.get("camt",0)  or 0,
                            itm.get("samt",0)  or 0,
                            itm.get("csamt",0) or 0,
                            ECOM_AMEND_FILL); row12 += 1

            # urp2ba — amended unregistered seller B2B
            for grp in ecoma_d.get("urp2ba", []):
                rtin = grp.get("rtin","")
                for inv in grp.get("inv", []):
                    for i, itm in enumerate(inv.get("itms", [])):
                        d = _itm_data(itm)
                        _ecom_write(ws12, row12, mdisp,
                            "E-Comm Unreg B2B Amended", True,
                            "", rtin,
                            inv.get("inum","")  if i==0 else "",
                            inv.get("idt","")   if i==0 else "",
                            inv.get("oinum","") if i==0 else "",
                            inv.get("oidt","")  if i==0 else "",
                            "",
                            inv.get("val",0) if i==0 else "",
                            inv.get("pos",""),
                            inv.get("sply_ty",""),
                            inv.get("inv_typ",""),
                            d.get("rt",0), d.get("txval",0) or 0,
                            d.get("iamt",0) or 0, d.get("camt",0) or 0,
                            d.get("samt",0) or 0, d.get("csamt",0) or 0,
                            ECOM_AMEND_FILL); row12 += 1

            # urp2ca — amended unregistered seller B2C flat
            for r in ecoma_d.get("urp2ca", []):
                _ecom_write(ws12, row12, mdisp,
                    "E-Comm Unreg B2C Amended", True,
                    "", "", "","","","",
                    r.get("omon",""),
                    "", r.get("pos",""),
                    r.get("sply_ty",""), "",
                    r.get("itms",[{}])[0].get("rt",0)    if r.get("itms") else 0,
                    r.get("itms",[{}])[0].get("txval",0) if r.get("itms") else 0,
                    r.get("itms",[{}])[0].get("iamt",0)  if r.get("itms") else 0,
                    r.get("itms",[{}])[0].get("camt",0)  if r.get("itms") else 0,
                    r.get("itms",[{}])[0].get("samt",0)  if r.get("itms") else 0,
                    r.get("itms",[{}])[0].get("csamt",0) if r.get("itms") else 0,
                    ECOM_AMEND_FILL); row12 += 1

        ws12.cell(row=row12, column=1, value="TOTAL")
        tot_row(ws12, row12, len(HDR12), [16,17,18,19,20], ds=3)
        set_widths(ws12, [14,28,6,22,22,14,12,14,12,11,14,18,12,14,8,14,12,12,12,10])

    # ── Save to BytesIO ────────────────────────────────────────────────────────
    import io as _io
    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf, None


# ── GSTR-1 Flask Routes ────────────────────────────────────────────────────────
@app.route("/g1/start", methods=["POST"])
def g1_start():
    data     = request.get_json(force=True) or {}
    gstin    = (data.get("gstin")    or "").strip().upper()
    fy       = (data.get("fy")       or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    if not all([gstin, fy, username, password]):
        return jsonify({"error":"gstin, fy, username and password required"}), 400
    specific_month = (data.get("specific_month") or "").strip()

    with g1_lock:
        if g1_state["status"] in ("running","waiting_captcha","waiting_otp","downloading"):
            return jsonify({"error":"Already running — reset first"}), 409
        g1_state.update({
            "status":"running","log":[],"error":None,
            "captcha_image":None,"captcha_answer":None,"otp_answer":None,
            "progress":0,"current_month":None,"total_months":0,
            "done_months":0,"files":[],"gstin":gstin,"fy":fy,
            "specific_month":specific_month or None,
        })
    threading.Thread(target=g1_worker,
                     args=(gstin,fy,username,password,specific_month),
                     daemon=True).start()
    return jsonify({"ok":True,"specific_month":specific_month or None})


@app.route("/g1/state")
def g1_get_state():
    with g1_lock:
        return jsonify({
            "status":         g1_state["status"],
            "log":            g1_state["log"][-40:],
            "error":          g1_state["error"],
            "progress":       g1_state["progress"],
            "current_month":  g1_state["current_month"],
            "total_months":   g1_state["total_months"],
            "done_months":    g1_state["done_months"],
            "files":          g1_state["files"],
            "gstin":          g1_state["gstin"],
            "fy":             g1_state["fy"],
            "specific_month": g1_state.get("specific_month"),
            "has_captcha":    g1_state["captcha_image"] is not None,
        })


@app.route("/g1/captcha_image")
def g1_captcha_image():
    with g1_lock:
        img = g1_state.get("captcha_image")
    if not img:
        return jsonify({"error":"no captcha available"}), 404
    return jsonify({"image": img})


@app.route("/g1/submit_captcha", methods=["POST"])
def g1_submit_captcha():
    data   = request.get_json(force=True) or {}
    answer = (data.get("captcha") or "").strip()
    if not answer:
        return jsonify({"error":"captcha required"}), 400
    with g1_lock:
        g1_state["captcha_answer"] = answer
        g1_state["captcha_image"]  = None
    return jsonify({"ok":True})


@app.route("/g1/submit_otp", methods=["POST"])
def g1_submit_otp():
    data = request.get_json(force=True) or {}
    otp  = (data.get("otp") or "").strip()
    if not otp:
        return jsonify({"error":"otp required"}), 400
    with g1_lock:
        g1_state["otp_answer"] = otp
    return jsonify({"ok":True})


@app.route("/g1/export_excel", methods=["GET","POST"])
def g1_export_excel():
    try:
        if request.method == "POST":
            body  = request.get_json(force=True) or {}
            gstin = (body.get("gstin","") or "").strip().upper()
            fy    = (body.get("fy",   "") or "").strip()
        else:
            gstin = request.args.get("gstin","").strip().upper()
            fy    = request.args.get("fy",   "").strip()
        if not gstin or not fy:
            return jsonify({"error":"gstin and fy required"}), 400
        try: import openpyxl
        except ImportError:
            return jsonify({"error":"openpyxl not installed. Run: pip install openpyxl"}), 500
        buf, err = g1_json_to_excel(gstin, fy)
        if err:
            return jsonify({"error":err}), 404
        fname = f"GSTR1_{gstin}_{fy.replace('-','_')}.xlsx"
        save_dir = os.path.join(G1_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
        return save_excel_and_respond(buf, save_dir, fname, log_fn=g1_log)
    except Exception as e:
        import traceback
        g1_log(f"export_excel error: {e} | {traceback.format_exc()}", "error")
        return jsonify({"error":str(e)}), 500


@app.route("/g1/check_excel_deps")
def g1_check_excel_deps():
    try:
        import openpyxl
        return jsonify({"ok":True, "openpyxl":openpyxl.__version__})
    except ImportError:
        return jsonify({"ok":False,
                        "error":"openpyxl not installed",
                        "install":"pip install openpyxl"}), 200


@app.route("/g1/reset", methods=["POST"])
def g1_reset():
    g1_set({
        "status":"idle","log":[],"error":None,
        "captcha_image":None,"captcha_answer":None,"otp_answer":None,
        "progress":0,"current_month":None,"total_months":0,
        "done_months":0,"files":[],"gstin":None,"fy":None,
        "specific_month":None,
    })
    return jsonify({"ok":True})


@app.route("/g1/files")
def g1_list_files():
    files = []
    if os.path.isdir(G1_DOWNLOAD_DIR):
        for gstin_dir in sorted(os.listdir(G1_DOWNLOAD_DIR)):
            gp = os.path.join(G1_DOWNLOAD_DIR, gstin_dir)
            if not os.path.isdir(gp): continue
            for fy_dir in sorted(os.listdir(gp), reverse=True):
                fp = os.path.join(gp, fy_dir)
                if not os.path.isdir(fp): continue
                for fname in sorted(os.listdir(fp)):
                    fpath = os.path.join(fp, fname)
                    files.append({
                        "gstin":    gstin_dir,
                        "fy":       fy_dir.replace("_","-"),
                        "filename": fname,
                        "size_kb":  max(1, os.path.getsize(fpath)//1024),
                        "url":      f"/g1/file/{gstin_dir}/{fy_dir}/{fname}",
                    })
    return jsonify({"files":files})


@app.route("/g1/file/<gstin>/<fy_dir>/<fname>")
def g1_serve_file(gstin, fy_dir, fname):
    fpath = os.path.join(G1_DOWNLOAD_DIR, gstin, fy_dir, fname)
    if not os.path.isfile(fpath):
        return jsonify({"error":"File not found"}), 404
    return send_file(fpath, as_attachment=True, download_name=fname)




# ══════════════════════════════════════════════════════════════════════════════
#  GST TDS / TCS CREDIT RECEIVED — Download & Excel Export
# ══════════════════════════════════════════════════════════════════════════════
#
#  PORTAL FLOW (researched from GST portal manual):
#    Services → Returns → TDS and TCS Credit Received
#    URL: https://return.gst.gov.in/returns/auth/tdstcscredreceived
#
#    Page shows 4 tables (all auto-populated from GSTR-7/8 filed by deductors):
#      Table 1 — TDS Credit Received       (from GSTR-7 of deductors)
#      Table 2 — Amendment to TDS Credit   (amendments filed by deductors)
#      Table 3 — TCS Credit Received       (from GSTR-8 of e-comm operators)
#      Table 4 — Amendment to TCS Credit   (amendments filed by operators)
#
#    Each record has: deductor GSTIN / trade name, invoice no, invoice date,
#                     taxable value, IGST/CGST/SGST, action (Accept/Reject/Pending)
#
#    Download approach: GST portal shows these as paginated HTML tables.
#    The "DOWNLOAD CSV" button on each table lets you download CSV data.
#    We click this button for each table and parse the downloaded CSVs.
#    If no CSV button, we scrape the visible table rows via JS.
#
#  WHAT WE DO:
#    1. Login (shared session from g2b / g1 login)
#    2. Navigate to TDS/TCS Credit Received page for selected period
#    3. Download/scrape all 4 tables
#    4. Convert to Excel (multi-sheet with summary)
# ══════════════════════════════════════════════════════════════════════════════

import threading, json, os, time, glob, csv, io as _io
from flask import jsonify, request, send_file

TDS_DOWNLOAD_DIR = PATHS.tdstcs_dir

tds_lock  = threading.Lock()
tds_state = {
    "status":         "idle",      # idle / running / waiting_captcha / waiting_otp / downloading / done / error
    "log":            [],
    "progress":       0,
    "current_month":  "",
    "done_months":    0,
    "files":          [],
    "error":          "",
    "captcha_image":  None,
    "captcha_answer": None,
    "otp_answer":     None,
}

def tds_log(msg, level="info"):
    with tds_lock:
        tds_state["log"].append({"msg": msg, "level": level,
                                  "ts": time.strftime("%H:%M:%S")})
        if len(tds_state["log"]) > 500:
            tds_state["log"] = tds_state["log"][-500:]
    # Same fix as _comb_log/_gdir_log/push_log: route through the
    # already-open logging.FileHandler instead of only an in-memory
    # list + console print() (which never reached disk at all) or a
    # fresh open() per call (which can hang under antivirus scanning).
    (log.error if level == "error" else log.info)(f"[TDS] {msg}")

def tds_set(d):
    with tds_lock:
        tds_state.update(d)


def tds_wait_field(field, timeout_sec=300):
    """Poll tds_state[field] until set by a /tdstcs/submit_* route."""
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        time.sleep(0.4)
        with tds_lock:
            val = tds_state.get(field)
            if val:
                tds_state[field] = None
                return val
    return None


def _scrape_tds_table(page, table_selector_hint, table_label):
    """
    Scrape a TDS/TCS table from the GST portal page.
    Returns list of dicts with the scraped data.
    The portal uses Angular tables; we extract via JS evaluation.
    """
    rows = []
    try:
        # Extract all table data via JavaScript — handles pagination by
        # clicking 'Next' until no more pages, or reads all visible rows.
        result = page.evaluate(f"""() => {{
            // Find the table matching our hint
            const allTables = document.querySelectorAll('table');
            let targetTable = null;
            // Try to find table near a heading matching our label
            const headings = document.querySelectorAll('h4,h5,h6,.panel-heading,.card-header,.section-head,th');
            for (const h of headings) {{
                if ((h.textContent||'').toUpperCase().includes('{table_label.upper()[:20]}')) {{
                    // Find closest table
                    let p = h.parentElement;
                    for (let i=0;i<8;i++) {{
                        if (!p) break;
                        const t = p.querySelector('table');
                        if (t) {{ targetTable = t; break; }}
                        p = p.parentElement;
                    }}
                    if (targetTable) break;
                }}
            }}
            if (!targetTable && allTables.length > 0) {{
                // Fallback: use index hint from label
                const idx = {{"TDS":0,"TDS AMENDMENT":1,"TCS":2,"TCS AMENDMENT":3}};
                const n = '{table_label.split()[0].upper()}';
                const isAmend = '{table_label}'.toUpperCase().includes('AMENDMENT');
                // Pick first table that looks like data (has >1 row)
                for (const t of allTables) {{
                    if (t.rows.length > 1) {{ targetTable = t; break; }}
                }}
            }}
            if (!targetTable) return [];
            const rows = [];
            const headerRow = targetTable.querySelector('thead tr') ||
                              targetTable.rows[0];
            const headers = headerRow ? Array.from(headerRow.cells)
                                .map(c=>c.textContent.trim()) : [];
            const bodyRows = targetTable.querySelectorAll('tbody tr');
            for (const tr of bodyRows) {{
                const cells = Array.from(tr.cells).map(c=>c.textContent.trim());
                if (cells.length === 0 || (cells.length===1 && cells[0]==='')) continue;
                const obj = {{}};
                cells.forEach((v,i)=>{{ obj[headers[i]||('col'+i)] = v; }});
                rows.push(obj);
            }}
            return rows;
        }}""")
        rows = result or []
    except Exception as e:
        tds_log(f"    ⚠ Error scraping {table_label}: {e}")
    return rows


def _click_download_csv(page, table_idx):
    """Click DOWNLOAD CSV / DOWNLOAD button for a table section. Returns True if clicked."""
    try:
        clicked = page.evaluate(f"""() => {{
            // Find all buttons/links with 'download' or 'csv' text
            const btns = document.querySelectorAll('button,a');
            const candidates = [];
            for (const b of btns) {{
                const t = (b.textContent||'').trim().toUpperCase();
                if (t.includes('DOWNLOAD CSV') || t.includes('DOWNLOAD EXCEL') ||
                    t.includes('DOWNLOAD') || t === 'CSV') {{
                    candidates.push(b);
                }}
            }}
            if (candidates[{table_idx}]) {{
                candidates[{table_idx}].click();
                return true;
            }}
            return false;
        }}""")
        return clicked
    except Exception:
        return False


def tds_do_browser_login(page, username, password):
    """Login to GST portal — mirrors g2b_do_browser_login but uses tds_state."""
    import base64 as _b64
    tds_log("Opening GST login page...")
    try:
        page.goto(GST_LOGIN, wait_until="domcontentloaded", timeout=25000)
        try: page.wait_for_selector(
                "input#username, input[name='username'], input[placeholder*='username' i]",
                state="visible", timeout=8000)
        except Exception: pass
    except Exception as e:
        tds_log(f"  ✗ Could not open login page: {e}", "error")
        tds_set({"status": "error", "error": str(e)}); return False

    for sel in ["input#username", "input[name='username']", "input[placeholder*='username' i]"]:
        try: page.locator(sel).first.fill(username); tds_log("  ✓ Username filled"); break
        except Exception: continue
    for sel in ["input#user_pass", "input[name='user_pass']",
                "input[type='password']", "input[placeholder*='password' i]"]:
        try: page.locator(sel).first.fill(password); tds_log("  ✓ Password filled"); break
        except Exception: continue
    time.sleep(0.3)

    # Capture captcha image
    cap_img = None
    try:
        for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                    "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                    ".captchaImage img", "img.captcha", "form img", "img"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
                    tds_log(f"  ✓ Captcha captured ({sel})"); break
            except Exception: continue
        if not cap_img:
            data_url = page.evaluate("""() => {
                for (const cv of document.querySelectorAll('canvas')) {
                    if (cv.width > 30 && cv.height > 10) {
                        try { return cv.toDataURL('image/png'); } catch(e) {}
                    }
                }
                return null;
            }""")
            if data_url and data_url.startswith("data:image"):
                cap_img = data_url; tds_log("  ✓ Captcha via canvas")
        if not cap_img:
            png = page.screenshot(full_page=False)
            cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
            tds_log("  ✓ Captcha via page screenshot")
    except Exception as e:
        tds_log(f"  ⚠ Captcha capture error: {e}")

    tds_set({"status": "waiting_captcha", "captcha_image": cap_img})
    tds_log("  ⏸ Waiting for captcha answer...")

    answer = tds_wait_field("captcha_answer", timeout_sec=300)
    if not answer:
        tds_log("  ✗ Captcha timeout", "error")
        tds_set({"status": "error", "error": "Captcha timeout"}); return False

    tds_set({"status": "running", "captcha_image": None})

    for sel in ["input[placeholder*='Characters' i]", "input#captcha", "input[name='captcha']"]:
        try: page.locator(sel).first.fill(str(answer)); break
        except Exception: continue
    for sel in ["button[type='submit']", "input[type='submit']", "button:has-text('LOGIN')"]:
        try: page.locator(sel).first.click(); tds_log("  ✓ Login submitted"); break
        except Exception: continue
    time.sleep(1.5)

    # OTP handling
    otp_needed = False
    for _ in range(8):
        time.sleep(1)
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.wait_for(state="visible", timeout=1500)
            otp_needed = True; break
        except Exception:
            if check_login_success(page) is not False: break
    if otp_needed:
        tds_set({"status": "waiting_otp"}); tds_log("  ⏸ Waiting for OTP...")
        otp = tds_wait_field("otp_answer", timeout_sec=180)
        if not otp:
            tds_log("  ✗ OTP timeout", "error")
            tds_set({"status": "error", "error": "OTP timeout"}); return False
        try:
            otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.click(); time.sleep(0.2); otp_el.fill(str(otp))
            page.locator("button[type='submit'],input[type='submit']").first.click()
            tds_log("  ✓ OTP submitted"); time.sleep(1)
        except Exception as e:
            tds_log(f"  ✗ OTP submit error: {e}", "error"); return False

    time.sleep(1)

    # Captcha retry loop (up to 3 attempts)
    MAX_CAPTCHA_ATTEMPTS = 3
    for _attempt in range(MAX_CAPTCHA_ATTEMPTS):
        result = check_login_success(page)
        if result is not False:
            break
        remaining = MAX_CAPTCHA_ATTEMPTS - _attempt - 1
        if remaining == 0:
            tds_log(f"  ✗ Login failed after {MAX_CAPTCHA_ATTEMPTS} captcha attempts", "error")
            tds_set({"status": "error", "error": f"Login failed after {MAX_CAPTCHA_ATTEMPTS} captcha attempts"})
            return False
        tds_log(f"  ✗ Login failed — re-capturing captcha (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")
        time.sleep(2)
        new_cap = None
        try:
            for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                        "img[src*='kaptcha' i]", "form img", "img"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=2000)
                    png = loc.screenshot()
                    if png and len(png) > 500:
                        new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                        tds_log(f"  ✓ Fresh captcha ({sel})"); break
                except Exception: continue
            if not new_cap:
                png = page.screenshot(full_page=False)
                new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
        except Exception as ce:
            tds_log(f"  ⚠ Re-capture error: {ce}")
        tds_set({"status": "running", "captcha_image": None})
        time.sleep(0.1)
        tds_set({"status": "waiting_captcha", "captcha_image": new_cap})
        tds_log(f"  ⏸ Waiting for captcha (attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")
        new_answer = tds_wait_field("captcha_answer", timeout_sec=300)
        if not new_answer:
            tds_log("  ✗ Captcha retry timeout", "error")
            tds_set({"status": "error", "error": "Captcha retry timeout"}); return False
        tds_set({"status": "running", "captcha_image": None})
        for sel in ["input[placeholder*='Characters' i]", "input#captcha", "input[name='captcha']"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.fill(""); loc.fill(str(new_answer)); break
            except Exception: continue
        for sel in ["button[type='submit']", "input[type='submit']", "button:has-text('LOGIN')"]:
            try: page.locator(sel).first.click(); tds_log("  ✓ Login re-submitted"); break
            except Exception: continue
        time.sleep(2)
        for _ in range(5):
            time.sleep(1)
            try:
                otp_el = page.locator("input[placeholder*='OTP' i], input[id*='otp' i]").first
                otp_el.wait_for(state="visible", timeout=1000)
                tds_set({"status": "waiting_otp"})
                tds_log("  ⏸ OTP required (retry)...")
                otp_r = tds_wait_field("otp_answer", timeout_sec=180)
                if not otp_r:
                    tds_log("  ✗ OTP timeout (retry)", "error")
                    tds_set({"status": "error", "error": "OTP timeout on retry"}); return False
                otp_el.click(); time.sleep(0.2); otp_el.fill(str(otp_r))
                page.locator("button[type='submit'],input[type='submit']").first.click()
                tds_log("  ✓ OTP re-submitted"); time.sleep(1); break
            except Exception:
                if check_login_success(page) is not False: break

    tds_log("  ✅ Logged in successfully")
    g2a_dismiss_popup(page)
    return True


def tds_activate_session(page, context):
    """
    Navigate from services.gst.gov.in/fowelcome to the TDS/TCS Credit Received page.

    Correct portal flow (NO direct URLs — disabled on GST portal):
      1. Click "Services" in the top navbar
      2. HOVER "Returns" tab (click navigates away — hover only reveals sub-menu)
      3. Native Playwright CLICK on "TDS and TCS credit received"
         (dispatchEvent/JS click is NOT used — Angular ignores synthetic events)

    The result URL should contain "comptds" (returns2/auth/comptds).
    """
    TDS_TEXT_TARGETS = [
        "a:text-is('TDS and TCS credit received')",
        "a:has-text('TDS and TCS credit received')",
        "a:has-text('TDS & TCS credit received')",
        "a:has-text('TDS/TCS credit received')",
    ]

    tds_log(f"  -> Starting at: {page.url[:80]}")
    g2a_dismiss_popup(page, context)

    def real_url(p):
        try:    return p.evaluate("location.href") or p.url
        except: return p.url

    def on_comptds(p):
        return "comptds" in real_url(p).lower()

    def get_comptds_page():
        for p in context.pages:
            if on_comptds(p): return p
        return None

    # Already there?
    if on_comptds(page):
        tds_log("  -> Already on TDS/TCS page")
        return True, page

    # ── Step 1: Click "Services" in the top navbar ───────────────────────────
    tds_log("  -> Step 1: Click 'Services'...")
    for sel in [
        "ul.nav.navbar-nav li.dropdown > a:has-text('Services')",
        "li.dropdown > a:has-text('Services')",
        "nav a:has-text('Services')",
        "a:text-is('Services')",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=4000)
            loc.click()
            tds_log(f"    ✓ 'Services' clicked")
            time.sleep(0.7)
            break
        except Exception:
            continue

    # ── Step 2: HOVER "Returns" tab (hover only — clicking navigates away) ───
    tds_log("  -> Step 2: Hover 'Returns' tab...")
    for sel in [
        "ul.nav-tabs a:text-is('Returns')",
        ".nav-tabs li a:text-is('Returns')",
        ".service-sub-cat a:text-is('Returns')",
        "a:text-is('Returns')",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            loc.hover()
            tds_log(f"    ✓ 'Returns' hovered — sub-menu visible")
            time.sleep(0.8)
            break
        except Exception:
            continue

    # ── Step 3: Native click on "TDS and TCS credit received" ────────────────
    tds_log("  -> Step 3: Click 'TDS and TCS credit received'...")
    tds_clicked = False
    for sel in TDS_TEXT_TARGETS:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=4000)
            loc.click()
            tds_log(f"    ✓ 'TDS and TCS credit received' clicked")
            tds_clicked = True
            break
        except Exception:
            continue

    if not tds_clicked:
        tds_log("    ⚠ Link not found via selectors — trying href extraction...")
        try:
            href = page.evaluate("""() => {
                const targets = [
                    'tds and tcs credit received',
                    'tds & tcs credit received',
                    'tds/tcs credit received',
                ];
                for (const a of document.querySelectorAll('a')) {
                    const t = (a.textContent || '').trim().toLowerCase();
                    if (targets.some(k => t.includes(k))) {
                        return a.getAttribute('href') || a.href || null;
                    }
                }
                return null;
            }""")
            if href and href.startswith('http'):
                tds_log(f"    Found href: {href[:70]}")
                # Navigate using the href extracted from the anchor
                page.locator(f"a[href='{href}']").first.click()
                tds_log(f"    ✓ Clicked via href anchor")
            else:
                tds_log(f"    ✗ No valid href found on page")
        except Exception as e:
            tds_log(f"    href extraction error: {e}")

    # ── Wait for navigation to TDS/TCS page ─────────────────────────────────
    tds_log("  -> Waiting for TDS/TCS Credit Received page...")
    for i in range(15):
        time.sleep(1)
        if on_comptds(page):
            try: page.wait_for_load_state("domcontentloaded", timeout=8000)
            except: pass
            tds_log(f"  ✓ TDS/TCS page reached after {i+1}s: {real_url(page)[:70]}")
            return True, page
        arrived = get_comptds_page()
        if arrived:
            try: arrived.wait_for_load_state("domcontentloaded", timeout=8000)
            except: pass
            tds_log(f"  ✓ TDS/TCS page (other tab) after {i+1}s: {real_url(arrived)[:70]}")
            return True, arrived
        tds_log(f"    [{i+1}s] {real_url(page)[:60]}")

    # ── 3-minute manual fallback ─────────────────────────────────────────────
    tds_log("  ⚠ Auto navigation failed")
    tds_log("  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    tds_log("  📋 ACTION NEEDED — in the browser window:")
    tds_log("     Services → hover Returns → TDS and TCS credit received")
    tds_log("  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    deadline = time.time() + 180
    while time.time() < deadline:
        time.sleep(2)
        if on_comptds(page):
            tds_log(f"  ✓ Manual nav successful")
            return True, page
        arrived = get_comptds_page()
        if arrived:
            tds_log(f"  ✓ Manual nav (other tab)")
            return True, arrived
    tds_log("  ✗ Timed out")
    return False, page


def tds_download_worker(gstin, fy, username, password, months):
    """
    Thin wrapper guaranteeing any exception is logged and reflected in
    tds_state instead of silently killing the thread.
    """
    try:
        _tds_download_worker_impl(gstin, fy, username, password, months)
    except Exception as fatal:
        import traceback as _tb
        try:
            tds_log(f"✗ FATAL (uncaught): {fatal}", "error")
        except Exception:
            pass
        try:
            tds_set({"status": "error", "error": f"Uncaught: {fatal}"})
        except Exception:
            pass
        log.error(f"[TDS] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _tds_download_worker_impl(gstin, fy, username, password, months):
    """
    Download TDS & TCS Credit Received JSON for each requested month.

    PORTAL FLOW (confirmed from screenshots):
      URL: https://return.gst.gov.in/returns2/auth/comptds
      1. Login via GST portal (same captcha+OTP flow as GSTR-2B)
      2. Navigate to returns2/auth/comptds
      3. Select Financial Year + Return Filing Period (month) → SEARCH
      4a. If FILED:     click DOWNLOAD button
      4b. If NOT FILED: click PREPARE OFFLINE button
      → Both lead to: return.gst.gov.in/returns2/auth/gstr2xco/offlinedownload
      5. Click "DOWNLOAD TDS AND TCS CREDIT RECEIVED JSON FILE (JSON)"
      6. Intercept browser download (instant — no link appears)
         Quick 3-second check; if no download, mark pending.
      After all months → Sweep 2: revisit pending months, wait up to 5 min.
    """
    os.makedirs(TDS_DOWNLOAD_DIR, exist_ok=True)
    gstin_dir = os.path.join(TDS_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    os.makedirs(gstin_dir, exist_ok=True)

    tds_set({"status": "running", "log": [], "progress": 0, "files": [], "error": "",
             "done_months": 0, "current_month": "",
             "captcha_image": None, "captcha_answer": None, "otp_answer": None})
    tds_log(f"TDS/TCS download — GSTIN: {gstin}  FY: {fy}")
    tds_log(f"Months requested: {len(months)}")

    COMPTDS_URL  = "https://return.gst.gov.in/returns2/auth/comptds"
    OFFLINE_PATH = "gstr2xco/offlinedownload"

    # Persistent profile (keeps session cookies across runs)
    profile_dir = os.path.join(PATHS.profiles_dir, "gst_profile_tdstcs")
    os.makedirs(profile_dir, exist_ok=True)
    lock_file = os.path.join(profile_dir, "SingletonLock")
    if os.path.exists(lock_file):
        try: os.remove(lock_file); tds_log("  ✓ Removed stale SingletonLock")
        except Exception: pass

    # sync_playwright().start() launches Playwright's own Node.js driver
    # subprocess — if it throws (not just hangs), that was previously
    # uncaught here, silently killing this thread with nothing logged.
    try:
        pw = sync_playwright().start()
    except Exception as e:
        tds_log(f"✗ Playwright driver failed to start: {e}", "error")
        tds_set({"status": "error", "error": f"Playwright driver failed to start: {e}"})
        return
    try:
        tds_log("🌐 Launching browser...")
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=False, slow_mo=40,
                args=["--disable-blink-features=AutomationControlled",
                      "--no-sandbox", "--start-maximized"],
                no_viewport=True, accept_downloads=True,
            )
        except Exception as e:
            tds_log(f"  ✗ Browser launch error: {e}", "error")
            tds_set({"status": "error", "error": str(e)}); return

        page = context.pages[0] if context.pages else context.new_page()
        page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")

        # ── Login ─────────────────────────────────────────────────────
        if not tds_do_browser_login(page, username, password):
            context.close(); return

        # ── Activate session → navigate to TDS/TCS Credit Received page ─────
        # Uses a dedicated activation (NOT g2a_activate_session) because TDS needs
        # Services → hover Returns → click "TDS and TCS credit received"
        # NOT "Returns Dashboard".  Direct URLs are disabled on the GST portal.
        tds_log("🌐 Activating session — navigating to TDS/TCS Credit Received...")
        ok, page = tds_activate_session(page, context)
        if not ok:
            tds_set({"status": "error",
                      "error": "Could not reach TDS/TCS Credit Received page"})
            context.close(); return
        tds_log(f"  ✓ On: {page.url[:70]}")

        tds_set({"status": "downloading"})

        # ── Helpers ────────────────────────────────────────────────────
        def navigate_to_comptds(mon_name_str):
            """
            Navigate to TDS/TCS Credit Received page via the portal menu:
              Services → hover Returns → click 'TDS and TCS credit received'
            Then select FY + month and click SEARCH.
            Falls back to direct URL if menu navigation fails.
            """
            nonlocal page   # page may be reassigned if menu opens a new tab

            def real_url(p):
                try: return p.evaluate("location.href") or p.url
                except Exception: return p.url

            on_comptds = "comptds" in real_url(page).lower()

            if not on_comptds:
                tds_log(f"  -> Not on comptds ({real_url(page)[:55]}) — re-navigating...")

                # Re-navigate via Services menu click path only (no direct URLs).
                # Same 3-step sequence: Click Services → hover Returns → click TDS link.
                nav_ok = False

                # Step 1: Click Services
                for sel in [
                    "ul.nav.navbar-nav li.dropdown > a:has-text('Services')",
                    "li.dropdown > a:has-text('Services')",
                    "nav a:has-text('Services')",
                    "a:text-is('Services')",
                ]:
                    try:
                        loc = page.locator(sel).first
                        loc.wait_for(state="visible", timeout=4000)
                        loc.click()
                        tds_log("    ✓ Services clicked")
                        time.sleep(0.7); break
                    except Exception: continue

                # Step 2: Hover Returns (hover only — clicking navigates away)
                for sel in [
                    "ul.nav-tabs a:text-is('Returns')",
                    ".nav-tabs li a:text-is('Returns')",
                    "a:text-is('Returns')",
                ]:
                    try:
                        loc = page.locator(sel).first
                        loc.wait_for(state="visible", timeout=3000)
                        loc.hover()
                        tds_log("    ✓ Returns hovered")
                        time.sleep(0.8); break
                    except Exception: continue

                # Step 3: Native click on TDS/TCS link
                for sel in [
                    "a:text-is('TDS and TCS credit received')",
                    "a:has-text('TDS and TCS credit received')",
                    "a:has-text('TDS & TCS credit received')",
                    "a:has-text('TDS/TCS credit received')",
                ]:
                    try:
                        loc = page.locator(sel).first
                        loc.wait_for(state="visible", timeout=4000)
                        loc.click()
                        tds_log("    ✓ 'TDS and TCS credit received' clicked")
                        nav_ok = True; break
                    except Exception: continue

                if not nav_ok:
                    tds_log("    ⚠ Link not found — repeating via tds_activate_session")
                    ok2, page = tds_activate_session(page, context)
                    if not ok2:
                        tds_log("    ✗ Re-activation failed")
                        return False

                # Wait for comptds page
                for i in range(15):
                    time.sleep(1)
                    if "comptds" in real_url(page).lower(): break
                    for p in context.pages:
                        if "comptds" in real_url(p).lower():
                            page = p; break

                tds_log(f"  -> Now at: {real_url(page)[:70]}")
                try: page.wait_for_load_state("domcontentloaded", timeout=8000)
                except Exception: pass
                time.sleep(1.0)

            # Select Financial Year (select[0])
            page.evaluate(f"""() => {{
                const sel = document.querySelectorAll('select')[0];
                if (!sel) return;
                for (const opt of sel.options) {{
                    if (opt.text.trim() === '{fy}') {{
                        sel.value = opt.value;
                        sel.dispatchEvent(new Event('change', {{bubbles:true}}));
                        return;
                    }}
                }}
            }}""")
            time.sleep(0.5)

            # Select Return Filing Period / month (select[1])
            page.evaluate(f"""() => {{
                const sel = document.querySelectorAll('select')[1];
                if (!sel) return;
                for (const opt of sel.options) {{
                    if (opt.text.trim() === '{mon_name_str}') {{
                        sel.value = opt.value;
                        sel.dispatchEvent(new Event('change', {{bubbles:true}}));
                        return;
                    }}
                }}
            }}""")

            # Click SEARCH
            for s in ["button:has-text('SEARCH')", "button:has-text('Search')"]:
                try:
                    page.locator(s).first.wait_for(state="visible", timeout=5000)
                    page.locator(s).first.click(); break
                except Exception: continue
            time.sleep(2)

        def get_real_url():
            """Always use location.href — page.url can be stale after Angular routing."""
            try: return page.evaluate("location.href") or page.url
            except Exception: return page.url

        def click_to_offline_download():
            """
            Click DOWNLOAD (filed) or PREPARE OFFLINE (unfiled).
            If we land on the Upload tab, click the Download tab using
            Playwright locator (Angular routing needs a real browser click).
            Returns 'filed', 'not_filed', or None.
            """
            result = page.evaluate("""() => {
                const btns = Array.from(document.querySelectorAll('button,a'));
                for (const b of btns) {
                    if ((b.textContent||'').trim().toUpperCase() === 'DOWNLOAD') {
                        b.click(); return 'filed';
                    }
                }
                for (const b of btns) {
                    const t = (b.textContent||'').trim().toUpperCase();
                    if (t === 'PREPARE OFFLINE' || t.includes('PREPARE OFFLINE')) {
                        b.click(); return 'not_filed';
                    }
                }
                return null;
            }""")
            if not result:
                return None

            # Wait for navigation to the offline page (upload or download tab)
            try:
                page.wait_for_url("*gstr2xco*", timeout=10000)
            except Exception:
                pass
            # Also accept gstr/offlineupload (not-filed path on some accounts)
            for _ in range(5):
                cur = get_real_url()
                if "gstr2xco" in cur or "offlineupload" in cur or "offlinedownload" in cur:
                    break
                time.sleep(1)
            time.sleep(1)

            cur = get_real_url()
            tds_log(f"  -> After click, URL: {cur[30:80]}")

            # If on Download tab already — done
            if OFFLINE_PATH in cur:
                return result

            # On Upload tab — click Download tab using Playwright locator
            tds_log(f"  -> On Upload tab — clicking Download tab...")
            try:
                # Tab is an <a> or <li> with exact text "Download"
                tab_loc = page.locator("a:has-text('Download'), li:has-text('Download')")
                # Filter to the one that's NOT "offlinedownload" button (just the tab)
                for i in range(tab_loc.count()):
                    t = (tab_loc.nth(i).text_content() or "").strip().lower()
                    if t == "download":
                        tab_loc.nth(i).click()
                        tds_log(f"    -> Download tab clicked (locator)")
                        break
                else:
                    # Fallback: click first visible element with text "Download"
                    page.locator("text=Download").first.click()
                    tds_log(f"    -> Download tab clicked (text fallback)")
            except Exception as e:
                tds_log(f"    ⚠ Tab click error: {e} — trying direct URL")
                page.goto(
                    "https://return.gst.gov.in/returns2/auth/gstr2xco/offlinedownload",
                    wait_until="domcontentloaded", timeout=12000
                )
                time.sleep(1.5)
                return result

            # Wait for URL to become offlinedownload
            try:
                page.wait_for_url(f"*{OFFLINE_PATH}*", timeout=8000)
            except Exception:
                pass
            time.sleep(1)
            cur = get_real_url()
            tds_log(f"  -> After Download tab, URL: {cur[30:80]}")

            # Final fallback if still not on offlinedownload
            if OFFLINE_PATH not in cur:
                tds_log(f"  -> Tab click did not navigate — using direct URL")
                page.goto(
                    "https://return.gst.gov.in/returns2/auth/gstr2xco/offlinedownload",
                    wait_until="domcontentloaded", timeout=12000
                )
                time.sleep(1.5)

            return result

        def _save_download(dl_value, json_path, mon_display, period):
            """Save a Playwright Download object → json_path. Returns result dict or None."""
            import zipfile, io as _io
            try:
                raw = open(dl_value.path(), "rb").read()
                if raw[:2] == b"PK":
                    with zipfile.ZipFile(_io.BytesIO(raw)) as zf:
                        names  = zf.namelist()
                        target = next((n for n in names if n.endswith(".json")), names[0])
                        data   = zf.read(target)
                else:
                    data = raw
                with open(json_path, "wb") as jf: jf.write(data)
                size_kb = max(1, len(data) // 1024)
                tds_log(f"    ✅ {mon_display} — {size_kb} KB saved")
                return {"file": json_path, "period": period,
                        "month": mon_display, "size_kb": size_kb}
            except Exception as e:
                tds_log(f"    ⚠ Save error: {e}", "error")
                return None

        def click_json_and_intercept(mon_display, period):
            """
            Download TDS/TCS JSON from the offline download page.

            Flow (confirmed from portal screenshots):
              Step 1: Click 'DOWNLOAD TDS AND TCS CREDIT RECEIVED JSON FILE (JSON)' button
              Step 2: Wait up to 5s for 'Click here to download - File N' link to appear
              Step 3: Click that link → intercept browser download
              If link never appears → return None (caller marks month as pending for sweep 2)
            """
            safe_m    = mon_display.replace(" ", "_").replace("/", "_")
            json_path = os.path.join(gstin_dir, f"TDSTCS_{safe_m}.json")

            # ── Step 1: Click the JSON button ─────────────────────────────
            clicked = False
            for s in [
                "button:has-text('DOWNLOAD TDS AND TCS CREDIT RECEIVED JSON FILE')",
                "button:has-text('JSON FILE')",
                "button:has-text('JSON')",
            ]:
                try:
                    page.locator(s).first.wait_for(state="visible", timeout=5000)
                    page.locator(s).first.click()
                    tds_log(f"    -> JSON button clicked")
                    clicked = True
                    break
                except Exception as e:
                    tds_log(f"    ⚠ {s[:50]}: {e}"); continue

            if not clicked:
                tds_log(f"    ⚠ JSON button not found on page")
                return None

            # ── Step 2: Wait up to 5s for 'Click here to download' link ──
            def find_download_link():
                return page.evaluate("""() => {
                    for (const a of document.querySelectorAll('a')) {
                        const t = (a.textContent || '').trim().toLowerCase();
                        if (t.includes('click here to download') ||
                            (t.includes('click here') && t.includes('file'))) {
                            return true;
                        }
                    }
                    return false;
                }""")

            link_appeared = False
            for _ in range(5):
                time.sleep(1)
                if find_download_link():
                    link_appeared = True
                    break

            if not link_appeared:
                tds_log(f"    -> Download link did not appear — marking for sweep 2")
                return None

            # ── Step 3: Click the link and intercept the download ─────────
            tds_log(f"    -> 'Click here to download' link appeared — downloading...")
            dl_result = None
            for _dl_attempt in range(2):   # up to 2 attempts
                try:
                    with page.expect_download(timeout=30000) as dl_info:
                        page.evaluate("""() => {
                            for (const a of document.querySelectorAll('a')) {
                                const t = (a.textContent || '').trim().toLowerCase();
                                if (t.includes('click here to download') ||
                                    (t.includes('click here') && t.includes('file'))) {
                                    a.removeAttribute('target');
                                    a.click();
                                    return;
                                }
                            }
                        }""")
                    dl_result = _save_download(
                        dl_info.value, json_path, mon_display, period)
                    break   # success
                except Exception as e:
                    tds_log(f"    ⚠ Link-click attempt {_dl_attempt+1} failed: {e}")
                    if _dl_attempt == 0:
                        # Re-request a fresh download: click JSON button again
                        tds_log(f"    -> Clicking JSON button again to request fresh file...")
                        for s in [
                            "button:has-text('DOWNLOAD TDS AND TCS CREDIT RECEIVED JSON FILE')",
                            "button:has-text('JSON FILE')",
                            "button:has-text('JSON')",
                        ]:
                            try:
                                page.locator(s).first.wait_for(state="visible", timeout=4000)
                                page.locator(s).first.click()
                                tds_log(f"    -> JSON button re-clicked — waiting 5s for fresh link")
                                break
                            except Exception: continue
                        # Wait for the fresh link to appear
                        for _ in range(5):
                            time.sleep(1)
                            if find_download_link():
                                tds_log(f"    -> Fresh link appeared")
                                break
            return dl_result

        # ══════════════════════════════════════════════════════════════
        # SWEEP 1: Navigate to each month, place JSON download request.
        #          Download completes instantly on click.
        #          If it fails (e.g., file still generating), mark pending.
        # ══════════════════════════════════════════════════════════════
        files_done = []
        pending    = []

        for i, month in enumerate(months):
            mon_display = month["display"]
            period      = month["period"]
            mon_name    = mon_display.split()[0]   # "January" from "January 2026"

            tds_set({"current_month": mon_display,
                     "progress": int((i / len(months)) * 100)})
            tds_log(f"\n[Sweep 1 · {i+1}/{len(months)}] {mon_display}")

            # Skip if already downloaded
            safe_m = mon_display.replace(" ", "_").replace("/", "_")
            if os.path.isfile(os.path.join(gstin_dir, f"TDSTCS_{safe_m}.json")):
                tds_log(f"  ✓ Already downloaded — skipping")
                continue

            navigate_to_comptds(mon_name)

            status = click_to_offline_download()
            if not status:
                tds_log(f"  ⚠ No DOWNLOAD or PREPARE OFFLINE button found — "
                        f"no TDS/TCS records for {mon_display}")
                continue

            tds_log(f"  -> Status: {status} — on offline download page")

            if OFFLINE_PATH not in get_real_url():
                tds_log(f"  ⚠ Not on offline page: {get_real_url()[:60]}")
                pending.append(month); continue

            result = click_json_and_intercept(mon_display, period)
            if result:
                files_done.append(result)
                tds_set({"files": files_done.copy(), "done_months": len(files_done)})
            else:
                tds_log(f"  -> Download not immediate — marked for sweep 2")
                pending.append(month)

        # ══════════════════════════════════════════════════════════════
        # SWEEP 2: Revisit pending months, wait up to 5 min per month
        # ══════════════════════════════════════════════════════════════
        if pending:
            tds_log("=" * 50)
            tds_log(f"SWEEP 2: {len(pending)} pending month(s)...")

        for i, month in enumerate(pending):
            mon_display = month["display"]
            period      = month["period"]
            mon_name    = mon_display.split()[0]

            safe_m = mon_display.replace(" ", "_").replace("/", "_")
            if os.path.isfile(os.path.join(gstin_dir, f"TDSTCS_{safe_m}.json")):
                continue

            tds_set({"current_month": mon_display})
            tds_log(f"\n[Sweep 2 · {i+1}/{len(pending)}] {mon_display}")

            navigate_to_comptds(mon_name)
            status = click_to_offline_download()
            if not status or OFFLINE_PATH not in get_real_url():
                tds_log(f"  ⚠ Still no offline page for {mon_display}"); continue

            # Try up to 5 minutes with retries
            deadline = time.time() + 300
            result   = None
            attempt  = 0
            while time.time() < deadline and not result:
                attempt += 1
                result = click_json_and_intercept(mon_display, period)
                if result: break
                if attempt < 10:
                    tds_log(f"    [{attempt}] Retrying in 30s...")
                    time.sleep(30)
                    # Re-navigate to refresh page state
                    navigate_to_comptds(mon_name)
                    click_to_offline_download()

            if result:
                files_done.append(result)
                tds_set({"files": files_done.copy(), "done_months": len(files_done),
                         "progress": int(((len(months)-len(pending)+i+1)/len(months))*100)})
            else:
                tds_log(f"  ✗ Could not download {mon_display} after 5 min", "error")

        tds_set({"status": "done", "progress": 100})
        tds_log(f"\n✅ TDS/TCS download complete — {len(files_done)} month(s) downloaded")

        try: context.close(); tds_log("🌐 Browser closed")
        except Exception as ce: tds_log(f"  ⚠ Browser close: {ce}")

    except Exception as e:
        tds_set({"status": "error", "error": str(e)})
        tds_log(f"✗ Fatal: {e}", "error")
        import traceback; tds_log(traceback.format_exc(), "error")
    finally:
        try: pw.stop()
        except Exception: pass

    time.sleep(8)
    with tds_lock:
        if tds_state.get("status") == "done":
            tds_state["status"] = "idle"
    tds_log("🔄 TDS/TCS reset to idle")


def tds_json_to_excel(gstin, fy):
    """
    Convert all downloaded TDS/TCS JSON files for a GSTIN+FY to a structured Excel workbook.

    Actual JSON schema (confirmed from portal downloads):
    ─────────────────────────────────────────────────────
    Top-level keys:
        gstin       – Recipient's own GSTIN
        fp          – Financial period (e.g. "2025-26")
        leg_name    – Legal name of recipient
        trade_name  – Trade name of recipient
        fil_status  – "FIL" (filed) or absent
        arn         – Acknowledgement Reference Number
        arn_date    – Date of ARN
        date_gn     – Date this JSON was generated
        rtn_prd     – Return period in MMYYYY (e.g. "042025")
        tds         – [ TDS entry, ... ]
        tcs         – [ TCS entry, ... ]

    TDS entry keys:
        ctin        – Deductor GSTIN
        ctin_name   – Deductor name
        amt_ded     – Amount deducted (Gross Payment u/s 51)
        iamt        – IGST deducted
        camt        – CGST deducted
        samt        – SGST/UTGST deducted
        flag        – "A"=Accepted, "R"=Rejected
        month       – MMYYYY (same as rtn_prd)
        inum        – Invoice number (optional – invoice-level detail months)
        idt         – Invoice date (optional)
        ival        – Invoice value / taxable value (optional)

    TCS entry keys:
        ctin        – E-commerce Operator GSTIN
        ctin_name   – Operator name
        pos         – Place of Supply (2-digit state code)
        amt         – Net amount of taxable supply
        iamt        – IGST collected
        camt        – CGST collected
        samt        – SGST/UTGST collected
        supR        – Supply to registered persons
        retsupR     – Returns from registered persons
        supU        – Supply to unregistered persons
        retsupU     – Returns from unregistered persons
        flag        – "A"=Accepted, "R"=Rejected

    Sheets produced:
        1. Summary              – One row per month: filing info + tax totals
        2. TDS Credit Received  – All TDS entries (invoice-level where available)
        3. TCS Credit Received  – All TCS entries with supply breakdown
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fy_dir     = fy.replace("-", "_")
    folder     = os.path.join(TDS_DOWNLOAD_DIR, gstin, fy_dir)
    json_files = sorted(glob.glob(os.path.join(folder, "*.json")))
    if not json_files:
        return None, f"No TDS/TCS JSON files found in {folder}"

    # ── Styles ────────────────────────────────────────────────────────────────
    C_TDS_HDR  = "880E4F"   # deep magenta  — TDS headers
    C_TCS_HDR  = "1B5E20"   # deep green    — TCS headers
    C_SUM_HDR  = "1A237E"   # deep indigo   — Summary header
    C_TOT      = "F8BBD0"   # light pink    — totals row
    C_ACC      = "C8E6C9"   # accepted
    C_REJ      = "FFCDD2"   # rejected
    C_PND      = "FFF9C4"   # pending

    def _fill(c): return PatternFill("solid", fgColor=c)
    def _font(bold=False, size=9, color="000000", italic=False):
        return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT    = Alignment(horizontal="left",   vertical="center")
    RIGHT   = Alignment(horizontal="right",  vertical="center")

    def _border(color="BDBDBD"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    BDR     = _border("CE93D8")
    BDR_TCS = _border("81C784")
    BDR_SUM = _border("9FA8DA")
    NUM_FMT = '#,##0.00'

    def make_title(ws, ncols, text, hdr_color):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value     = text
        c.font      = _font(bold=True, size=11, color="FFFFFF")
        c.fill      = _fill(hdr_color)
        c.alignment = CENTER
        ws.row_dimensions[1].height = 28

    def make_hdr(ws, labels, hdr_color, bdr=None, row=2):
        if bdr is None: bdr = _border("FFFFFF")
        for ci, lbl in enumerate(labels, 1):
            c = ws.cell(row=row, column=ci, value=lbl)
            c.font      = _font(bold=True, size=9, color="FFFFFF")
            c.fill      = _fill(hdr_color)
            c.alignment = CENTER
            c.border    = bdr
        ws.row_dimensions[row].height = 34

    def wrow(ws, r, vals, row_fill=None, bdr=None):
        if bdr is None: bdr = BDR
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font   = _font(size=9)
            c.border = bdr
            if row_fill: c.fill = _fill(row_fill)
            c.alignment = RIGHT if isinstance(v, (int, float)) else LEFT
            if isinstance(v, (int, float)):
                c.number_format = NUM_FMT

    def total_row(ws, r, ncols, sum_cols, ds=3, bdr=None):
        if bdr is None: bdr = BDR
        for ci in range(1, ncols + 1):
            c = ws.cell(row=r, column=ci)
            c.fill   = _fill(C_TOT)
            c.font   = _font(bold=True, size=9)
            c.border = bdr
            if ci in sum_cols:
                val = sum(
                    ws.cell(row=dr, column=ci).value or 0
                    for dr in range(ds, r)
                    if isinstance(ws.cell(row=dr, column=ci).value, (int, float))
                )
                c.value     = val
                c.alignment = RIGHT
                c.number_format = NUM_FMT
            else:
                c.alignment = LEFT

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def sf(v):
        """Safe float."""
        try: return float(str(v).replace(",", "").strip())
        except: return 0.0

    # ── Month helpers ─────────────────────────────────────────────────────────
    MN = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]
    FY_ORDER = {m: (i if i >= 4 else i + 12) for i, m in enumerate(MN, 1)}

    def period_to_display(p):
        """'042025' → 'April 2025'"""
        try: return f"{MN[int(p[:2])-1]} {p[2:]}"
        except: return p

    def period_sort_key(p):
        try: return FY_ORDER.get(MN[int(p[:2])-1], 99)
        except: return 99

    def flag_display(f):
        s = str(f).upper()
        if s == "A": return "Accepted"
        if s == "R": return "Rejected"
        return "Pending"

    def flag_fill(f):
        s = str(f).upper()
        if s == "A": return C_ACC
        if s == "R": return C_REJ
        return C_PND

    STATE_CODES = {
        "01":"Jammu and Kashmir","02":"Himachal Pradesh","03":"Punjab",
        "04":"Chandigarh","05":"Uttarakhand","06":"Haryana","07":"Delhi",
        "08":"Rajasthan","09":"Uttar Pradesh","10":"Bihar","11":"Sikkim",
        "12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur","15":"Mizoram",
        "16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
        "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh",
        "24":"Gujarat","25":"Daman and Diu","26":"Dadra and Nagar Haveli",
        "27":"Maharashtra","28":"Andhra Pradesh","29":"Karnataka",
        "30":"Goa","31":"Lakshadweep","32":"Kerala","33":"Tamil Nadu",
        "34":"Puducherry","35":"Andaman and Nicobar Islands","36":"Telangana",
        "37":"Andhra Pradesh (new)","38":"Ladakh",
    }

    def pos_display(pos):
        code = str(pos).zfill(2)
        return f"{code} – {STATE_CODES.get(code, 'Unknown')}"

    # ── Load and sort all JSON files ──────────────────────────────────────────
    all_data = []
    for fpath in json_files:
        try:
            with open(fpath, encoding="utf-8") as f:
                d = json.load(f)
            d["_mon_display"] = period_to_display(d.get("rtn_prd", ""))
            d["_sort_key"]    = period_sort_key(d.get("rtn_prd", ""))
            all_data.append(d)
        except Exception:
            pass

    all_data.sort(key=lambda d: d["_sort_key"])

    if not all_data:
        return None, "No valid TDS/TCS JSON data found"

    wb = Workbook()
    wb.remove(wb.active)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — TDS CREDIT RECEIVED
    # ══════════════════════════════════════════════════════════════════════════
    ws_tds = wb.create_sheet("TDS Credit Received")

    TDS_H = [
        "Month", "Return Period", "Deductor GSTIN", "Deductor Name",
        "Invoice No.", "Invoice Date", "Invoice Value (₹)",
        "Amount Deducted (₹)", "IGST (₹)", "CGST (₹)", "SGST/UTGST (₹)",
        "Total TDS (₹)", "Status",
    ]
    TDS_SUM_COLS = {7, 8, 9, 10, 11, 12}

    make_title(ws_tds, len(TDS_H),
        f"TDS Credit Received  |  GSTIN: {gstin}  |  FY: {fy}  "
        f"|  u/s 51 CGST Act  |  Green=Accepted · Red=Rejected · Yellow=Pending",
        C_TDS_HDR)
    make_hdr(ws_tds, TDS_H, hdr_color=C_TDS_HDR)
    ws_tds.freeze_panes = "A3"

    tds_rows = []
    for d in all_data:
        mon = d["_mon_display"]
        rtn = d.get("rtn_prd", "")
        for e in d.get("tds", []):
            iamt = sf(e.get("iamt", 0))
            camt = sf(e.get("camt", 0))
            samt = sf(e.get("samt", 0))
            ival = e.get("ival")
            tds_rows.append((
                [
                    mon, rtn,
                    e.get("ctin", ""),
                    e.get("ctin_name", ""),
                    e.get("inum", ""),
                    e.get("idt", ""),
                    sf(ival) if ival not in (None, "", 0) else "",
                    sf(e.get("amt_ded", 0)),
                    iamt, camt, samt,
                    iamt + camt + samt,
                    flag_display(e.get("flag", "")),
                ],
                flag_fill(e.get("flag", ""))
            ))

    for i, (vals, rf) in enumerate(tds_rows):
        wrow(ws_tds, i + 3, vals, row_fill=rf)

    tr = len(tds_rows) + 3
    ws_tds.cell(row=tr, column=1).value = "GRAND TOTAL"
    ws_tds.cell(row=tr, column=1).font  = _font(bold=True)
    ws_tds.cell(row=tr, column=2).value = f"{len(tds_rows)} entries"
    ws_tds.cell(row=tr, column=2).font  = _font(bold=True)
    total_row(ws_tds, tr, len(TDS_H), TDS_SUM_COLS)

    set_col_widths(ws_tds,
        [16, 13, 22, 34, 13, 13, 18, 18, 14, 14, 14, 14, 12])

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — TCS CREDIT RECEIVED
    # ══════════════════════════════════════════════════════════════════════════
    ws_tcs = wb.create_sheet("TCS Credit Received")

    TCS_H = [
        "Month", "Return Period", "Operator GSTIN", "Operator Name",
        "Place of Supply",
        "Supply to Registered (₹)", "Returns from Registered (₹)",
        "Supply to Unregistered (₹)", "Returns from Unregistered (₹)",
        "Net Taxable Supply (₹)",
        "IGST (₹)", "CGST (₹)", "SGST/UTGST (₹)", "Total TCS (₹)",
        "Status",
    ]
    TCS_SUM_COLS = {6, 7, 8, 9, 10, 11, 12, 13, 14}

    make_title(ws_tcs, len(TCS_H),
        f"TCS Credit Received  |  GSTIN: {gstin}  |  FY: {fy}  "
        f"|  u/s 52 CGST Act  |  Green=Accepted · Red=Rejected · Yellow=Pending",
        C_TCS_HDR)
    make_hdr(ws_tcs, TCS_H, hdr_color=C_TCS_HDR)
    ws_tcs.freeze_panes = "A3"

    tcs_rows = []
    for d in all_data:
        mon = d["_mon_display"]
        rtn = d.get("rtn_prd", "")
        for e in d.get("tcs", []):
            iamt = sf(e.get("iamt", 0))
            camt = sf(e.get("camt", 0))
            samt = sf(e.get("samt", 0))
            tcs_rows.append((
                [
                    mon, rtn,
                    e.get("ctin", ""),
                    e.get("ctin_name", ""),
                    pos_display(e.get("pos", "")),
                    sf(e.get("supR",    0)),
                    sf(e.get("retsupR", 0)),
                    sf(e.get("supU",    0)),
                    sf(e.get("retsupU", 0)),
                    sf(e.get("amt",     0)),
                    iamt, camt, samt,
                    iamt + camt + samt,
                    flag_display(e.get("flag", "")),
                ],
                flag_fill(e.get("flag", ""))
            ))

    for i, (vals, rf) in enumerate(tcs_rows):
        wrow(ws_tcs, i + 3, vals, row_fill=rf, bdr=BDR_TCS)

    tr2 = len(tcs_rows) + 3
    ws_tcs.cell(row=tr2, column=1).value = "GRAND TOTAL"
    ws_tcs.cell(row=tr2, column=1).font  = _font(bold=True)
    ws_tcs.cell(row=tr2, column=2).value = f"{len(tcs_rows)} entries"
    ws_tcs.cell(row=tr2, column=2).font  = _font(bold=True)
    total_row(ws_tcs, tr2, len(TCS_H), TCS_SUM_COLS, bdr=BDR_TCS)

    set_col_widths(ws_tcs,
        [16, 13, 22, 34, 30, 20, 22, 22, 22, 20, 14, 14, 14, 14, 12])

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY (inserted at position 0)
    # ══════════════════════════════════════════════════════════════════════════
    ws_sum = wb.create_sheet("Summary", 0)

    SUM_H = [
        "Month", "Return Period", "Legal Name", "Trade Name",
        "Filing Status", "ARN", "ARN Date", "Generated On",
        "TDS Entries", "Amt Deducted (₹)",
        "TDS IGST (₹)", "TDS CGST (₹)", "TDS SGST (₹)", "Total TDS Tax (₹)",
        "TCS Entries", "Net Supply (₹)",
        "TCS IGST (₹)", "TCS CGST (₹)", "TCS SGST (₹)", "Total TCS Tax (₹)",
        "Grand Total Tax (₹)",
    ]
    SUM_SUM_COLS = {9,10,11,12,13,14, 15,16,17,18,19,20, 21}

    make_title(ws_sum, len(SUM_H),
        f"TDS / TCS Credit Received — Summary  |  GSTIN: {gstin}  |  FY: {fy}  "
        f"|  Services → Returns → TDS and TCS Credit Received",
        C_SUM_HDR)
    make_hdr(ws_sum, SUM_H, hdr_color=C_SUM_HDR, bdr=BDR_SUM)
    ws_sum.freeze_panes = "E3"

    for i, d in enumerate(all_data):
        tds_l = d.get("tds", [])
        tcs_l = d.get("tcs", [])

        tds_amt  = sum(sf(e.get("amt_ded", 0)) for e in tds_l)
        tds_igst = sum(sf(e.get("iamt",    0)) for e in tds_l)
        tds_cgst = sum(sf(e.get("camt",    0)) for e in tds_l)
        tds_sgst = sum(sf(e.get("samt",    0)) for e in tds_l)
        tds_tax  = tds_igst + tds_cgst + tds_sgst

        tcs_net  = sum(sf(e.get("amt",  0)) for e in tcs_l)
        tcs_igst = sum(sf(e.get("iamt", 0)) for e in tcs_l)
        tcs_cgst = sum(sf(e.get("camt", 0)) for e in tcs_l)
        tcs_sgst = sum(sf(e.get("samt", 0)) for e in tcs_l)
        tcs_tax  = tcs_igst + tcs_cgst + tcs_sgst

        fil = d.get("fil_status", "")
        fil_disp = "Filed ✓" if fil == "FIL" else ("Unfiled" if not fil else fil)

        wrow(ws_sum, i + 3, [
            d["_mon_display"],
            d.get("rtn_prd", ""),
            d.get("leg_name", ""),
            d.get("trade_name", ""),
            fil_disp,
            d.get("arn", ""),
            d.get("arn_date", ""),
            d.get("date_gn", ""),
            len(tds_l), tds_amt, tds_igst, tds_cgst, tds_sgst, tds_tax,
            len(tcs_l), tcs_net, tcs_igst, tcs_cgst, tcs_sgst, tcs_tax,
            tds_tax + tcs_tax,
        ], bdr=BDR_SUM)

    gtr = len(all_data) + 3
    ws_sum.cell(row=gtr, column=1).value = "GRAND TOTAL"
    ws_sum.cell(row=gtr, column=1).font  = _font(bold=True)
    total_row(ws_sum, gtr, len(SUM_H), SUM_SUM_COLS, bdr=BDR_SUM)

    # Legal note
    leg_r = gtr + 2
    ws_sum.merge_cells(f"A{leg_r}:{get_column_letter(len(SUM_H))}{leg_r}")
    leg = ws_sum.cell(row=leg_r, column=1,
        value=(
            "⚖  TDS/TCS LEGAL NOTE:  "
            "TDS (u/s 51 CGST Act) is deducted @ 1% CGST + 1% SGST (= 2% total) "
            "by Govt/PSU deductors on payments above ₹2.5 lakh.  "
            "TCS (u/s 52 CGST Act) is collected @ 0.5% CGST + 0.5% SGST (= 1% total) "
            "by e-commerce operators.  "
            "Both TDS and TCS are credited to Electronic CASH Ledger — NOT ITC Ledger.  "
            "Accept/Reject credits here BEFORE filing GSTR-3B to use them for tax payment.  "
            "No due date or late fee for filing GSTR-7X (TDS/TCS Credit Received)."
        ))
    leg.font      = _font(italic=True, size=8.5, color="1A237E")
    leg.fill      = _fill("E8EAF6")
    leg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    leg.border    = BDR_SUM
    ws_sum.row_dimensions[leg_r].height = 52

    set_col_widths(ws_sum, [
        16, 13, 26, 26, 11, 24, 13, 14,
        11, 18, 14, 14, 14, 16,
        11, 18, 14, 14, 14, 16,
        18,
    ])

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


# ── TDS/TCS Flask Routes ──────────────────────────────────────────────────────

@app.route("/tdstcs/start", methods=["POST"])
def tdstcs_start():
    data     = request.get_json(force=True) or {}
    gstin    = (data.get("gstin","")).strip().upper()
    fy       = (data.get("fy","")).strip()
    username = (data.get("username","")).strip()
    password = (data.get("password","")).strip()
    months   = data.get("months",[])
    if not gstin or not fy or not username or not password or not months:
        return jsonify({"error":"gstin / fy / username / password / months required"}),400

    with tds_lock:
        if tds_state["status"] in ("running","waiting_captcha","downloading"):
            return jsonify({"error":"Already running — reset first"}), 409
        tds_state.update({"status":"running","log":[],"progress":0,"files":[],"error":"",
                          "done_months":0,"current_month":""})

    threading.Thread(
        target=tds_download_worker,
        args=(gstin, fy, username, password, months),
        daemon=True,
    ).start()
    return jsonify({"ok":True})


@app.route("/tdstcs/state")
def tdstcs_state_route():
    with tds_lock:
        return jsonify(dict(tds_state))


@app.route("/tdstcs/export_excel", methods=["GET","POST"])
def tdstcs_export_excel():
    if request.method == "POST":
        body  = request.get_json(force=True) or {}
        gstin = (body.get("gstin","") or "").strip().upper()
        fy    = (body.get("fy",   "") or "").strip()
    else:
        gstin = request.args.get("gstin","").strip().upper()
        fy    = request.args.get("fy","").strip()
    if not gstin or not fy:
        return jsonify({"error":"gstin and fy required"}), 400
    buf, err = tds_json_to_excel(gstin, fy)
    if err:
        return jsonify({"error":err}), 404
    fname = f"TDS_TCS_{gstin}_{fy.replace('-','_')}.xlsx"
    save_dir = os.path.join(TDS_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    return save_excel_and_respond(buf, save_dir, fname, log_fn=tds_log)


@app.route("/tdstcs/reset", methods=["POST"])
def tdstcs_reset():
    tds_set({"status":"idle","log":[],"progress":0,"files":[],"error":"",
              "done_months":0,"current_month":"",
              "captcha_image":None,"captcha_answer":None,"otp_answer":None})
    return jsonify({"ok":True})


@app.route("/tdstcs/submit_captcha", methods=["POST"])
def tdstcs_submit_captcha():
    data = request.get_json(force=True) or {}
    ans  = (data.get("captcha") or "").strip()
    if not ans:
        return jsonify({"error": "captcha required"}), 400
    tds_set({"captcha_answer": ans})
    return jsonify({"ok": True})


@app.route("/tdstcs/submit_otp", methods=["POST"])
def tdstcs_submit_otp():
    data = request.get_json(force=True) or {}
    otp  = (data.get("otp") or "").strip()
    if not otp:
        return jsonify({"error": "otp required"}), 400
    tds_set({"otp_answer": otp})
    return jsonify({"ok": True})


@app.route("/tdstcs/files")
def tdstcs_files():
    result = []
    if os.path.isdir(TDS_DOWNLOAD_DIR):
        for gd in sorted(os.listdir(TDS_DOWNLOAD_DIR)):
            gp = os.path.join(TDS_DOWNLOAD_DIR, gd)
            if not os.path.isdir(gp): continue
            for fy_d in sorted(os.listdir(gp)):
                fp = os.path.join(gp, fy_d)
                if not os.path.isdir(fp): continue
                flist = [f for f in sorted(os.listdir(fp)) if f.endswith(".json")]
                if flist:
                    result.append({"gstin":gd,"fy":fy_d.replace("_","-"),
                                   "count":len(flist),"files":flist})
    return jsonify(result)








# ══════════════════════════════════════════════════════════════════════════════
# GSTR-3B WORKER — Full DOM scraping (no PDF, no external conversion tools)
# Supports both Monthly filers (every month) and QRMP filers (quarter-end only)
# Login + session activation mirrors g2a_worker exactly.
# Section navigation uses Angular client-side routing (never window.location.href)
# so period context is preserved across all 7 sub-sections.
# ══════════════════════════════════════════════════════════════════════════════

# ── State ─────────────────────────────────────────────────────────────────────
g3b_state = {
    "status":         "idle",
    "log":            [],
    "captcha_image":  None,
    "captcha_answer": None,
    "otp_answer":     None,
    "files":          [],
    "fresh_files":    [],
    "error":          None,
}
g3b_lock = threading.Lock()


def g3b_log(msg, level="info"):
    ts   = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    with g3b_lock:
        g3b_state["log"].append(line)
    # Same fix as _comb_log/_gdir_log/push_log/tds_log — route through
    # the already-open logging.FileHandler. Also added the missing
    # `level` parameter: this function was previously called with a
    # second argument in one place (the excel-export error handler),
    # which would have raised TypeError ("takes 1 positional argument
    # but 2 were given") the moment that error path was ever hit.
    (log.error if level == "error" else log.info)(f"[G3B] {msg}")


def g3b_set(updates, log_msg=None):
    with g3b_lock:
        g3b_state.update(updates)
    if log_msg:
        g3b_log(log_msg)


def g3b_wait_field(field, timeout_sec=180):
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        time.sleep(0.4)
        with g3b_lock:
            val = g3b_state.get(field)
            if val:
                g3b_state[field] = None
                return val
    return None


# ── Constants ──────────────────────────────────────────────────────────────────
G3B_DOWNLOAD_DIR = PATHS.gstr3b_dir
_G3B_DASHBOARD = "https://return.gst.gov.in/returns/auth/dashboard"
_G3B_QTR_END   = {3, 6, 9, 12}
_G3B_MON = {
    1: "January",  2: "February", 3: "March",    4: "April",
    5: "May",      6: "June",     7: "July",      8: "August",
    9: "September",10: "October", 11: "November", 12: "December",
}
_G3B_QTR = {
    4:  "Quarter 1 (Apr - Jun)", 5:  "Quarter 1 (Apr - Jun)", 6:  "Quarter 1 (Apr - Jun)",
    7:  "Quarter 2 (Jul - Sep)", 8:  "Quarter 2 (Jul - Sep)", 9:  "Quarter 2 (Jul - Sep)",
    10: "Quarter 3 (Oct - Dec)", 11: "Quarter 3 (Oct - Dec)", 12: "Quarter 3 (Oct - Dec)",
    1:  "Quarter 4 (Jan - Mar)", 2:  "Quarter 4 (Jan - Mar)", 3:  "Quarter 4 (Jan - Mar)",
}
# Angular route fragment for each section (used for client-side nav)
_G3B_SECTION_PATH = {
    "3_1":   "iosup",
    "3_1_1": "supplyCGSTAct",
    "3_2":   "interstatesupplies",
    "4":     "elgITC",
    "5":     "inwardSup",
    "5_1":   "interestLateFee",
    "6_1":   "payment",
}


# ── Popup dismiss (exact copy of g2a_dismiss_popup) ───────────────────────────
def g3b_dismiss_popup(page, context=None):
    """Dismiss 'Principal Place of Business metadata' dialog (NO-REMIND ME LATER)."""
    def all_pages():
        pages = [page]
        if context:
            for p in context.pages:
                if p not in pages:
                    pages.append(p)
        return pages

    deadline = time.time() + 12
    attempt  = 0
    while time.time() < deadline:
        attempt += 1
        for p in all_pages():
            result = p.evaluate("""() => {
                const targets = [
                    'NO-REMIND ME LATER',
                    'No-Remind Me Later',
                    'NO REMIND ME LATER',
                ];
                for (const btn of document.querySelectorAll('button, a')) {
                    const t = (btn.textContent || '').trim();
                    if (targets.includes(t) || t.toUpperCase() === 'NO-REMIND ME LATER') {
                        btn.click();
                        return t;
                    }
                }
                return null;
            }""")
            if result:
                g3b_log(f"  \u2713 Dismissed popup: '{result}' (attempt {attempt})")
                time.sleep(0.3)
                return True
        time.sleep(0.5)
    g3b_log(f"  \u2139 Popup not found after {attempt} attempts (may not have appeared)")
    return False


# ── Session activation (exact copy of g2a_activate_session) ───────────────────
def g3b_activate_session(page, context):
    """Navigate from fowelcome/SSO to return.gst.gov.in/returns/auth/dashboard."""
    g3b_log(f"  -> Post-login: {page.url[:80]}")
    g3b_dismiss_popup(page, context)

    def real_url(p):
        try:    return p.evaluate("location.href") or p.url
        except: return p.url

    def arrived(p):
        u = real_url(p)
        return ("return.gst.gov.in" in u
                and "accessdenied" not in u
                and "login" not in u)

    def sync_and_check(p):
        try:    p.wait_for_load_state("domcontentloaded", timeout=5000)
        except: pass
        u = real_url(p)
        g3b_log(f"    synced URL: {u[:70]}")
        return u

    def get_active_page():
        for p in context.pages:
            if arrived(p): return p
        return None

    # Step 1: Click Services
    g3b_log("  -> Step 1: Clicking Services...")
    for sel in ["nav a:has-text('Services')",
                "ul.nav > li > a:has-text('Services')",
                "li.dropdown > a:has-text('Services')"]:
        try:
            page.locator(sel).first.wait_for(state="visible", timeout=3000)
            page.locator(sel).first.click(); time.sleep(0.8)
            g3b_log(f"    Clicked Services ({sel})"); break
        except Exception: continue

    # Step 2: HOVER "Returns" tab — reveals sub-menu (clicking navigates away)
    g3b_log("  -> Step 2: Hover 'Returns' tab...")
    returns_ok = False
    for sel in [
        "ul.nav-tabs a:text-is('Returns')",
        ".nav-tabs li a:text-is('Returns')",
        ".service-sub-cat a:text-is('Returns')",
        "ul.sub-menu a:text-is('Returns')",
        "a:text-is('Returns')",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            loc.hover()
            g3b_log(f"    ✓ 'Returns' hovered — sub-menu visible")
            returns_ok = True
            time.sleep(0.8)
            break
        except Exception:
            continue

    if not returns_ok:
        g3b_log("    ⚠ 'Returns' hover failed")

    # Step 3: CLICK "Returns Dashboard" — native Playwright click, NOT dispatchEvent
    g3b_log("  -> Step 3: Click 'Returns Dashboard'...")
    dash_ok = False
    for sel in [
        "a:text-is('Returns Dashboard')",
        "a:has-text('Returns Dashboard')",
        ".dropdown-menu a:has-text('Returns Dashboard')",
        "ul.list-unstyled a:has-text('Returns Dashboard')",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=4000)
            loc.click()
            g3b_log(f"    ✓ 'Returns Dashboard' clicked")
            dash_ok = True
            break
        except Exception:
            continue

    if not dash_ok:
        g3b_log("    ⚠ Click failed — trying href fallback...")
        try:
            href = page.evaluate("""() => {
                for (const a of document.querySelectorAll('a')) {
                    if ((a.textContent||'').trim() === 'Returns Dashboard')
                        return a.href || a.getAttribute('href');
                }
                return null;
            }""")
            if href and "return.gst.gov.in" in href:
                g3b_log(f"    goto: {href[:70]}")
                page.goto(href, wait_until="domcontentloaded", timeout=20000)
                time.sleep(1.5)
        except Exception as e:
            g3b_log(f"    href fallback error: {e}")

    g3b_log("  -> Waiting for return.gst.gov.in...")
    for i in range(15):
        time.sleep(1)
        live = real_url(page)
        g3b_log(f"    [{i+1}s] {live[:70]}")
        if "return.gst.gov.in" in live and "accessdenied" not in live:
            sync_and_check(page)
            if arrived(page):
                g3b_log(f"  ✓ Session active: {page.url[:70]}")
                return True, page
        active = get_active_page()
        if active:
            sync_and_check(active)
            g3b_log(f"  ✓ Session active (other tab): {active.url[:70]}")
            return True, active

    # quicklinks SSO handler
    for p in context.pages:
        if "quicklinks" in real_url(p):
            g3b_log(f"  -> SSO quicklinks: {real_url(p)[:70]}")
            try:
                lnk = p.locator("a[href*='return.gst.gov.in']").first
                lnk.wait_for(state="visible", timeout=3000)
                lnk.click()
                time.sleep(2)
                if arrived(p):
                    sync_and_check(p)
                    g3b_log(f"  ✓ SSO → Returns portal")
                    return True, p
            except Exception:
                pass

    # Manual fallback
    live_urls = [real_url(p)[:50] for p in context.pages]
    g3b_log(f"  ⚠ Auto nav failed. Pages: {live_urls}")
    g3b_log("  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    g3b_log("  📋 ACTION NEEDED — in the browser window:")
    g3b_log("     Services  →  hover Returns  →  click Returns Dashboard")
    g3b_log("  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    deadline = time.time() + 180
    while time.time() < deadline:
        time.sleep(2)
        if arrived(page):
            sync_and_check(page); return True, page
        active = get_active_page()
        if active:
            sync_and_check(active); return True, active
    g3b_log("  ✗ Timed out.")
    return False, page


# ── Browser login (exact mirror of g2a_do_browser_login) ──────────────────────
def g3b_do_browser_login(page, username, password):
    """Handle GST portal login: fill credentials, wait for captcha, submit."""
    g3b_log("Opening GST login page...")
    try:
        page.goto(GST_LOGIN, wait_until="domcontentloaded", timeout=25000)
        try: page.wait_for_selector(
                "input#username, input[name='username'], input[placeholder*='username' i]",
                state="visible", timeout=8000)
        except Exception: pass
    except Exception as e:
        g3b_log(f"  \u2717 Could not open login page: {e}")
        g3b_set({"status": "error", "error": str(e)}); return False

    for sel in ["input#username", "input[name='username']",
                "input[placeholder*='username' i]"]:
        try: page.locator(sel).first.fill(username); g3b_log("  \u2713 Username filled"); break
        except Exception: continue

    for sel in ["input#user_pass", "input[name='user_pass']",
                "input[type='password']", "input[placeholder*='password' i]"]:
        try: page.locator(sel).first.fill(password); g3b_log("  \u2713 Password filled"); break
        except Exception: continue

    time.sleep(0.3)

    cap_img = None
    try:
        import base64 as _b64
        for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                    "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                    ".captchaImage img", "img.captcha", "form img", "img"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
                    g3b_log(f"  \u2713 Captcha captured via element screenshot ({sel})")
                    break
            except Exception: continue

        if not cap_img:
            data_url = page.evaluate("""() => {
                for (const cv of document.querySelectorAll('canvas')) {
                    if (cv.width > 30 && cv.height > 10) {
                        try { return cv.toDataURL('image/png'); } catch(e) {}
                    }
                }
                return null;
            }""")
            if data_url and data_url.startswith("data:image"):
                cap_img = data_url; g3b_log("  \u2713 Captcha captured via canvas")

        if not cap_img:
            png     = page.screenshot(full_page=False)
            cap_img = "data:image/png;base64," + _b64.b64encode(png).decode()
            g3b_log("  \u2713 Captcha captured via full screenshot (fallback)")
    except Exception as e:
        g3b_log(f"  \u26a0 Captcha capture error: {e}")

    g3b_set({"status": "waiting_captcha", "captcha_image": cap_img})
    g3b_log("  \u23f8 Waiting for captcha...")

    answer = g3b_wait_field("captcha_answer", timeout_sec=300)
    if not answer:
        g3b_log("  \u2717 Captcha timeout")
        g3b_set({"status": "error", "error": "Captcha timeout"}); return False

    g3b_set({"status": "running", "captcha_image": None})

    for sel in ["input[placeholder*='Characters' i]", "input#captcha",
                "input[name='captcha']"]:
        try:
            page.locator(sel).first.fill(str(answer))
            g3b_log(f"  \u2713 Captcha filled: {repr(answer)}"); break
        except Exception: continue

    for sel in ["button[type='submit']", "input[type='submit']",
                "button:has-text('LOGIN')"]:
        try: page.locator(sel).first.click(); g3b_log("  \u2713 Login submitted"); break
        except Exception: continue
    time.sleep(1.5)

    otp_needed = False
    for _ in range(8):
        time.sleep(1)
        try:
            otp_el = page.locator(
                "input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.wait_for(state="visible", timeout=1500)
            otp_needed = True; break
        except Exception:
            if check_login_success(page) is not False: break

    if otp_needed:
        g3b_set({"status": "waiting_otp"})
        g3b_log("  \u23f8 Waiting for OTP...")
        otp = g3b_wait_field("otp_answer", timeout_sec=180)
        if not otp:
            g3b_log("  \u2717 OTP timeout")
            g3b_set({"status": "error", "error": "OTP timeout"}); return False
        try:
            otp_el = page.locator(
                "input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.click(); time.sleep(0.2)
            otp_el.fill(str(otp))
            page.locator("button[type='submit'],input[type='submit']").first.click()
            g3b_log("  \u2713 OTP submitted"); time.sleep(1)
        except Exception as e:
            g3b_log(f"  \u2717 OTP error: {e}"); return False

    time.sleep(1)

    # Captcha retry loop — up to 3 total attempts
    MAX_CAPTCHA_ATTEMPTS = 3
    for _attempt in range(MAX_CAPTCHA_ATTEMPTS):
        result = check_login_success(page)
        if result is not False:
            break

        remaining = MAX_CAPTCHA_ATTEMPTS - _attempt - 1
        if remaining == 0:
            g3b_log(f"  \u2717 Login failed after {MAX_CAPTCHA_ATTEMPTS} attempts")
            g3b_set({"status": "error",
                     "error": f"Login failed after {MAX_CAPTCHA_ATTEMPTS} captcha attempts"})
            return False

        g3b_log(f"  \u2717 Login failed \u2014 re-capturing fresh captcha "
                f"(attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")
        time.sleep(2)
        new_cap = None
        try:
            import base64 as _b64
            for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                        "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                        ".captchaImage img", "img.captcha", "form img", "img"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=2000)
                    png = loc.screenshot()
                    if png and len(png) > 500:
                        new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                        g3b_log(f"  \u2713 Fresh captcha re-captured ({sel})")
                        break
                except Exception: continue
            if not new_cap:
                png     = page.screenshot(full_page=False)
                new_cap = "data:image/png;base64," + _b64.b64encode(png).decode()
                g3b_log("  \u2713 Fresh captcha via screenshot (fallback)")
        except Exception as ce:
            g3b_log(f"  \u26a0 Re-capture error: {ce}")

        g3b_set({"status": "running",          "captcha_image": None})
        time.sleep(0.1)
        g3b_set({"status": "waiting_captcha",  "captcha_image": new_cap})
        g3b_log(f"  \u23f8 Waiting for captcha answer "
                f"(attempt {_attempt+2}/{MAX_CAPTCHA_ATTEMPTS})...")

        new_answer = g3b_wait_field("captcha_answer", timeout_sec=300)
        if not new_answer:
            g3b_log("  \u2717 Captcha retry timeout")
            g3b_set({"status": "error", "error": "Captcha retry timeout"})
            return False

        g3b_set({"status": "running", "captcha_image": None})

        for sel in ["input[placeholder*='Characters' i]", "input#captcha",
                    "input[name='captcha']"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.fill(""); loc.fill(str(new_answer))
                g3b_log(f"  \u2713 Re-filled captcha: {repr(new_answer)}"); break
            except Exception: continue

        for sel in ["button[type='submit']", "input[type='submit']",
                    "button:has-text('LOGIN')"]:
            try:
                page.locator(sel).first.click()
                g3b_log("  \u2713 Login re-submitted"); break
            except Exception: continue

        time.sleep(2)
        for _ in range(5):
            time.sleep(1)
            try:
                otp_el = page.locator(
                    "input[placeholder*='OTP' i], input[id*='otp' i]").first
                otp_el.wait_for(state="visible", timeout=1000)
                g3b_set({"status": "waiting_otp"})
                g3b_log("  \u23f8 OTP required (retry)...")
                otp_r = g3b_wait_field("otp_answer", timeout_sec=180)
                if not otp_r:
                    g3b_log("  \u2717 OTP timeout (retry)")
                    g3b_set({"status": "error", "error": "OTP timeout on retry"})
                    return False
                otp_el.click(); time.sleep(0.2); otp_el.fill(str(otp_r))
                page.locator("button[type='submit'],input[type='submit']").first.click()
                g3b_log("  \u2713 OTP re-submitted"); time.sleep(1); break
            except Exception:
                if check_login_success(page) is not False: break

    g3b_log("  \u2705 Logged in successfully")
    g3b_dismiss_popup(page)
    return True


# Section headings — unique text that appears in the teal banner when that section loads.
# We poll for this text after Angular route change to confirm the DOM has updated.
_G3B_SECTION_HEADING = {
    "3_1":   "outward and reverse",        # "3.1 Tax on outward and reverse charge..."
    "3_1_1": "9(5)",                       # "3.1.1 Supplies notified under section 9(5)"
    "3_2":   "inter-state supplies",       # "3.2 Of the supplies shown in 3.1(a)..."
    "4":     "eligible itc",               # "4. Eligible ITC"
    "5":     "non-gst",                    # "5. Values of exempt, nil-rated, non-GST..."
    "5_1":   "interest",                   # "5.1 Interest and Late fee"
    "6_1":   "payment of tax",             # "6.1 Payment of tax"
}


def _g3b_wait_heading(page, heading_fragment, timeout=15):
    """
    Poll until the section heading text appears in page body.
    This is the ONLY reliable signal that Angular has rendered the section content.
    Replaces wait_for_url (which never works — Angular changes internal route,
    not the browser URL) and avoids 8-second wasted timeouts.
    """
    frag     = heading_fragment.lower()
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            found = page.evaluate(
                f"() => (document.body.innerText || '').toLowerCase().includes('{frag}')")
            if found:
                return True
        except Exception:
            pass
        time.sleep(0.4)
    return False


def _g3b_wait_angular_http(page, timeout=8):
    """
    Wait for Angular's $http pending requests to reach 0.
    Each section fires XHR calls to fetch auto-populated data (GSTR-1/2B).
    Without this, inputs are 0 because the server hasn't responded yet.
    """
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            pending = page.evaluate("""() => {
                try {
                    const inj   = angular.element(document.body).injector();
                    const $http = inj.get('$http');
                    return $http.pendingRequests ? $http.pendingRequests.length : 0;
                } catch(e) { return 0; }
            }""")
            if pending == 0:
                return True
        except Exception:
            return True    # if we can't check, assume done
        time.sleep(0.4)
    return False


def _g3b_click_section(page, section_key):
    """
    Navigate to a GSTR-3B sub-section using Angular's internal router.

    KEY FINDING: Angular changes its internal route state but does NOT always
    change the browser URL. wait_for_url() is therefore useless and wastes 8
    seconds every call. Removed entirely.

    Correct approach after navigation:
      1. Poll page body for section heading text  → confirms DOM updated
      2. Poll Angular $http.pendingRequests == 0  → confirms server data loaded
      3. Then read inputs

    Strategy priority:
      1. AngularJS $location.url() via injector  (fastest, no reload)
      2. Playwright locator click on sidebar <a> links
      3. JS .click() on href/ui-sref/ng-href elements
      4. page.goto() direct URL (when Angular injector unavailable after reload)
    """
    path    = _G3B_SECTION_PATH[section_key]
    heading = _G3B_SECTION_HEADING.get(section_key, "")

    def _post_nav_wait(label=""):
        """Common wait sequence after any navigation attempt."""
        if heading:
            found = _g3b_wait_heading(page, heading)
            if not found:
                g3b_log(f"      ⚠ heading '{heading}' not confirmed{' (' + label + ')' if label else ''}")
        _g3b_wait_angular_http(page)
        time.sleep(0.8)
        _g3b_wait_inputs(page)

    # ── Strategy 1: AngularJS $location service ──────────────────────────────
    result = "err:not_run"
    try:
        result = page.evaluate(f"""() => {{
            try {{
                const inj  = angular.element(document.body).injector();
                if (!inj) return 'no_angular';
                const $loc = inj.get('$location');
                const $rs  = inj.get('$rootScope');
                const cur  = $loc.url() || '';
                const parts  = cur.split('/');
                const gIdx   = parts.lastIndexOf('gstr3b');
                const base   = gIdx >= 0
                    ? parts.slice(0, gIdx + 1).join('/')
                    : '/returns/auth/gstr3b';
                const newUrl = base + '/{path}';
                $loc.url(newUrl);
                $rs.$apply();
                return 'ok:' + newUrl;
            }} catch(e) {{
                return 'err:' + String(e).slice(0, 80);
            }}
        }}""")
    except Exception:
        # Context destroyed = Angular triggered a full page reload (normal for 3.2)
        g3b_log(f"      Angular $location: context destroyed (reload in progress)")
        try: page.wait_for_load_state("domcontentloaded", timeout=10000)
        except Exception: pass
        time.sleep(1.0)
        _post_nav_wait("after-reload")
        g3b_log(f"      Recovered after reload: {page.url[-50:]}")
        return True

    g3b_log(f"      Angular $location: {result}")

    if result.startswith("ok:"):
        _post_nav_wait()
        return True

    # ── Strategy 2: Playwright locator click on sidebar nav ──────────────────
    for sel in [
        f"a[href*='{path}']",
        f"a[ui-sref*='{path}']",
        f"a[ng-href*='{path}']",
        f"li[ui-sref-active*='{path}'] a",
    ]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            loc.click()
            _post_nav_wait("strategy-2")
            return True
        except Exception:
            continue

    # ── Strategy 3: JS click on any element referencing the path ─────────────
    clicked = page.evaluate(f"""() => {{
        const t = '{path}'.toLowerCase();
        for (const el of document.querySelectorAll('a, li, span, div')) {{
            const h  = (el.getAttribute('href')    || '').toLowerCase();
            const sr = (el.getAttribute('ui-sref') || '').toLowerCase();
            const ng = (el.getAttribute('ng-href') || '').toLowerCase();
            if (h.includes(t) || sr.includes(t) || ng.includes(t)) {{
                el.click(); return true;
            }}
        }}
        return false;
    }}""")
    if clicked:
        _post_nav_wait("strategy-3")
        return True

    # ── Strategy 4: page.goto() direct URL ───────────────────────────────────
    g3b_log(f"      ⚠ Strategies 1-3 failed — using page.goto fallback")
    _G3B_BASE = "https://return.gst.gov.in/returns/auth/gstr3b"
    try:
        page.goto(f"{_G3B_BASE}/{path}", wait_until="domcontentloaded", timeout=15000)
        time.sleep(1.0)
        _post_nav_wait("goto")
        g3b_log(f"      ✓ page.goto succeeded: {page.url[-50:]}")
        return True
    except Exception as _ge:
        g3b_log(f"      ✗ page.goto failed: {_ge}")

    g3b_log(f"      ⚠ Cannot navigate to {section_key} ({path}) — reading page as-is")
    return False


def _g3b_wait_inputs(page, timeout=10):
    """Wait until Angular has rendered at least one table <input> on the page."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            n = page.evaluate(
                "() => document.querySelectorAll('table input').length")
            if n > 0:
                return True
        except Exception:
            pass
        time.sleep(0.4)
    return False


def _g3b_parse(s):
    if not s: return 0.0
    try:    return float(str(s).replace(",", "").replace("\u20b9", "").strip())
    except: return 0.0


def _g3b_table_vals(page):
    for _attempt in range(3):
        try:
            raw = page.evaluate("""() =>
                Array.from(document.querySelectorAll('table input'))
                     .map(el => (el.value || '').replace(/,/g, '').trim())
            """)
            return [_g3b_parse(x) for x in raw]
        except Exception as _e:
            if "context was destroyed" in str(_e).lower() or \
               "execution context" in str(_e).lower():
                # Page reloaded — wait and retry
                try:
                    page.wait_for_load_state("domcontentloaded", timeout=8000)
                except Exception:
                    pass
                time.sleep(2.0)
                _g3b_wait_inputs(page)
            else:
                raise
    return []


def _g(v, i): return v[i] if i < len(v) else 0.0


def _g3b_log_verify(vals, label):
    """Log a verification summary so the user can see real data was captured."""
    nonzero = [(i, v) for i, v in enumerate(vals) if v != 0.0]
    if nonzero:
        sample = ", ".join(f"[{i}]={v}" for i, v in nonzero[:4])
        g3b_log(f"      \u2714 {label}: {len(nonzero)}/{len(vals)} non-zero ({sample})")
    else:
        g3b_log(f"      \u26a0 {label}: all values are 0 (nil return or unfiled?)")


# ── Section scrapers ──────────────────────────────────────────────────────────

def g3b_scrape_3_1(page):
    _g3b_click_section(page, "3_1")
    # Dismiss IGST export popup that appears on section 3.1
    for sel in ["button:has-text('OK')", "button:has-text('Ok')",
                "button:has-text('CLOSE')", "button:has-text('Close')"]:
        try:
            btn = page.locator(sel).first
            btn.wait_for(state="visible", timeout=3000)
            btn.click(); time.sleep(0.4)
            g3b_log("      \u2713 Dismissed IGST popup"); break
        except Exception: continue
    v = _g3b_table_vals(page)
    _g3b_log_verify(v, "3.1")
    return {
        "a": {"taxable_value": _g(v,0),  "igst": _g(v,1),  "cgst": _g(v,2),  "sgst": _g(v,3),  "cess": _g(v,4)},
        "b": {"taxable_value": _g(v,5),  "igst": _g(v,6),  "cess": _g(v,7)},
        "c": {"taxable_value": _g(v,8)},
        "d": {"taxable_value": _g(v,9),  "igst": _g(v,10), "cgst": _g(v,11), "sgst": _g(v,12), "cess": _g(v,13)},
        "e": {"taxable_value": _g(v,14)},
    }


def g3b_scrape_3_1_1(page):
    _g3b_click_section(page, "3_1_1")
    v = _g3b_table_vals(page)
    _g3b_log_verify(v, "3.1.1")
    return {
        "i":  {"taxable_value": _g(v,0), "igst": _g(v,1), "cgst": _g(v,2), "sgst": _g(v,3), "cess": _g(v,4)},
        "ii": {"taxable_value": _g(v,5)},
    }


def g3b_scrape_3_2(page):
    """
    Section 3.2 — Inter-state supplies (state-wise).

    Portal structure (confirmed from screenshots):
      Three accordion panels on the page, each identified by heading text:
        • "Supplies made to Unregistered Persons"
        • "Supplies made to Composition Taxable Persons"
        • "Supplies made to UIN holders"

      Each panel has a "+" button (collapsed) or "−" (expanded) at the RIGHT edge
      of the heading row.  Clicking the heading row (or the +/−) toggles the panel.

      When expanded, the panel shows a table:
        col 0: checkbox  (ignore)
        col 1: <select>  Place of Supply (State/UT)   — text e.g. "34 - Puducherry"
        col 2: <input>   Total Taxable value (₹)
        col 3: <input>   Amount of Integrated Tax (₹)

      Only rows where the select is NOT "Select" carry real data.
    """
    _g3b_click_section(page, "3_2")

    # Wait for the page to fully load after navigation
    try:
        page.wait_for_load_state("domcontentloaded", timeout=8000)
    except Exception:
        pass
    time.sleep(1.5)

    SECTIONS = [
        # (output_key, substring_to_find_heading)
        ("unregistered", "Unregistered"),
        ("composition",  "Composition"),
        ("uin",          "UIN"),
    ]
    result = {}

    for key, label_fragment in SECTIONS:
        # ── Step 1: click the accordion header to expand it ───────────────────
        # The heading row is an element whose text contains label_fragment.
        # It has a sibling/child "+" sign.  We click the heading row itself.
        # Use Playwright locator so it works even after a partial reload.
        try:
            # Find heading rows that contain the label text AND a "+" sign
            # Strategy A: heading row contains "+" in its own text
            expanded = page.evaluate(f"""() => {{
                const fragment = '{label_fragment}'.toLowerCase();
                for (const el of document.querySelectorAll(
                        'tr, div, li, h4, h5, h6, span, a, p, td')) {{
                    const txt = (el.textContent || '').trim();
                    if (!txt.toLowerCase().includes(fragment)) continue;
                    // Check for "+" or "−" indicator nearby
                    const parent = el.closest('tr,div,li') || el.parentElement;
                    if (!parent) continue;
                    const parentTxt = (parent.textContent || '');
                    const hasMinus = parentTxt.includes('−') || parentTxt.includes('–')
                                  || parentTxt.includes('-') || parentTxt.includes('−');
                    const hasPlus  = parentTxt.includes('+');
                    if (hasMinus && !hasPlus) {{
                        return 'already_open';   // already expanded
                    }}
                    if (hasPlus || !parent.querySelector('table')) {{
                        el.click();
                        return 'clicked';
                    }}
                }}
                return 'not_found';
            }}""")
            g3b_log(f"      3.2 [{label_fragment}] accordion: {expanded}")
            if expanded not in ("already_open", "clicked"):
                # Fallback: use Playwright locator to find and click the heading
                for sel in [
                    f"text={label_fragment}",
                    f"*:has-text('{label_fragment}')",
                ]:
                    try:
                        loc = page.locator(sel).first
                        loc.wait_for(state="visible", timeout=3000)
                        loc.click()
                        g3b_log(f"      3.2 [{label_fragment}] clicked via Playwright ({sel})")
                        break
                    except Exception:
                        continue
        except Exception as _e:
            g3b_log(f"      3.2 [{label_fragment}] expand error: {_e}")
        time.sleep(1.2)

        # ── Step 2: read the rows from the now-expanded panel ─────────────────
        rows = []
        for _attempt in range(3):
            try:
                rows = page.evaluate(f"""() => {{
                    const fragment = '{label_fragment}'.toLowerCase();
                    // Find the container that holds both the heading and the table
                    let panel = null;
                    // Walk all elements; find deepest one containing fragment + table
                    for (const el of document.querySelectorAll(
                            'div, section, li, article, tbody, table')) {{
                        const txt = (el.textContent || '').toLowerCase();
                        if (!txt.includes(fragment)) continue;
                        if (!el.querySelector('table, tr')) continue;
                        // Prefer narrower containers
                        if (!panel || el.contains(panel)) continue;
                        panel = el;
                    }}
                    // Also try: find the heading element, then its next sibling container
                    if (!panel) {{
                        for (const el of document.querySelectorAll('*')) {{
                            const own = (el.textContent || '').trim().toLowerCase();
                            if (own.includes(fragment) && own.length < 120) {{
                                // Look at siblings and parent's children for a table
                                let sib = el.nextElementSibling;
                                while (sib) {{
                                    if (sib.querySelector && sib.querySelector('table')) {{
                                        panel = sib; break;
                                    }}
                                    sib = sib.nextElementSibling;
                                }}
                                if (!panel) {{
                                    const p = el.parentElement;
                                    if (p && p.nextElementSibling &&
                                        p.nextElementSibling.querySelector('table')) {{
                                        panel = p.nextElementSibling;
                                    }}
                                }}
                                if (panel) break;
                            }}
                        }}
                    }}
                    if (!panel) return [];

                    const out = [];
                    const rows = panel.querySelectorAll('tr');
                    for (const tr of rows) {{
                        const selects = tr.querySelectorAll('select');
                        const inputs  = tr.querySelectorAll('input[type="text"],input:not([type])');
                        if (selects.length === 0 || inputs.length < 2) continue;
                        const sel0 = selects[0];
                        const stateText = sel0.options[sel0.selectedIndex]
                                        ? sel0.options[sel0.selectedIndex].text.trim()
                                        : '';
                        if (!stateText || stateText.toLowerCase() === 'select') continue;
                        const taxable = (inputs[0].value || '').replace(/,/g, '').trim();
                        const igst    = (inputs[1].value || '').replace(/,/g, '').trim();
                        out.push({{ state: stateText, taxable, igst }});
                    }}
                    return out;
                }}""")
                break  # success
            except Exception as _e2:
                if "context was destroyed" in str(_e2).lower() or \
                   "execution context" in str(_e2).lower():
                    try:
                        page.wait_for_load_state("domcontentloaded", timeout=8000)
                    except Exception:
                        pass
                    time.sleep(2.0)
                else:
                    g3b_log(f"      3.2 [{label_fragment}] read error: {_e2}")
                    break

        result[key] = [
            {"state":         r["state"],
             "taxable_value": _g3b_parse(r["taxable"]),
             "igst":          _g3b_parse(r["igst"])}
            for r in rows
        ]
        if result[key]:
            sample = result[key][0]
            g3b_log(f"      ✔ 3.2 [{label_fragment}]: {len(result[key])} rows "
                    f"(e.g. {sample['state']} tv={sample['taxable_value']} igst={sample['igst']})")
        else:
            g3b_log(f"      ✔ 3.2 [{label_fragment}]: 0 rows (nil or empty)")

    total = sum(len(v) for v in result.values())
    g3b_log(f"      ✔ 3.2 total: {total} state rows across all sub-sections")
    return result



def g3b_scrape_4(page):
    """
    Section 4 — Eligible ITC.

    Table structure (from screenshot):
      Columns: Details | Integrated Tax | Central Tax | State/UT Tax | CESS
      Rows:
        A(1) Import of goods          — IGST + CESS only  (CGST/SGST blank)
        A(2) Import of services       — IGST + CESS only
        A(3) Inward RCM (other 1&2)   — all 4
        A(4) Inward from ISD          — all 4
        A(5) All other ITC            — all 4
        B(1) Rules 38,42,43 & 17(5)   — all 4
        B(2) Others                   — all 4
        C    Net ITC (A-B)            — all 4  (computed, read-only)
        D(1) ITC reclaimed            — all 4
        D(2) Ineligible ITC 16(4)/PoS — all 4

    We read each row by matching label text, then reading inputs in order.
    """
    _g3b_click_section(page, "4")

    def _read_row(label_fragment, n_inputs):
        """Find the <tr> containing label_fragment, return up to n_inputs values."""
        try:
            vals = page.evaluate(f"""() => {{
                const frag = '{label_fragment}'.toLowerCase();
                for (const tr of document.querySelectorAll('tr')) {{
                    if (!(tr.textContent || '').toLowerCase().includes(frag)) continue;
                    const inputs = tr.querySelectorAll('input');
                    if (inputs.length === 0) continue;
                    return Array.from(inputs)
                               .slice(0, {n_inputs})
                               .map(i => (i.value || '').replace(/,/g, '').trim());
                }}
                return [];
            }}""")
            return [_g3b_parse(x) for x in vals]
        except Exception:
            return []

    def _v(vals, i): return vals[i] if i < len(vals) else 0.0

    # A(1) Import of goods: IGST, CESS only
    a1 = _read_row("Import of goods",  2)
    # A(2) Import of services: IGST, CESS only
    a2 = _read_row("Import of services", 2)
    # A(3) Inward RCM (other than 1 & 2): IGST, CGST, SGST, CESS
    a3 = _read_row("reverse charge",   4)
    # A(4) Inward from ISD: IGST, CGST, SGST, CESS
    a4 = _read_row("from ISD",         4)
    # A(5) All other ITC: IGST, CGST, SGST, CESS
    a5 = _read_row("All other ITC",    4)
    # B(1) Rules 38,42,43: IGST, CGST, SGST, CESS
    b1 = _read_row("rules 38",         4)
    # B(2) Others: IGST, CGST, SGST, CESS
    b2 = _read_row("Others",           4)
    # C Net ITC: IGST, CGST, SGST, CESS
    c  = _read_row("Net ITC",          4)
    # D(1) ITC reclaimed: IGST, CGST, SGST, CESS
    d1 = _read_row("reclaimed",        4)
    # D(2) Ineligible ITC 16(4): IGST, CGST, SGST, CESS
    d2 = _read_row("16(4)",            4)

    result = {
        "A": {
            "1": {"igst": _v(a1,0), "cess": _v(a1,1)},
            "2": {"igst": _v(a2,0), "cess": _v(a2,1)},
            "3": {"igst": _v(a3,0), "cgst": _v(a3,1), "sgst": _v(a3,2), "cess": _v(a3,3)},
            "4": {"igst": _v(a4,0), "cgst": _v(a4,1), "sgst": _v(a4,2), "cess": _v(a4,3)},
            "5": {"igst": _v(a5,0), "cgst": _v(a5,1), "sgst": _v(a5,2), "cess": _v(a5,3)},
        },
        "B": {
            "1": {"igst": _v(b1,0), "cgst": _v(b1,1), "sgst": _v(b1,2), "cess": _v(b1,3)},
            "2": {"igst": _v(b2,0), "cgst": _v(b2,1), "sgst": _v(b2,2), "cess": _v(b2,3)},
        },
        "C":  {"igst": _v(c,0),  "cgst": _v(c,1),  "sgst": _v(c,2),  "cess": _v(c,3)},
        "D": {
            "1": {"igst": _v(d1,0), "cgst": _v(d1,1), "sgst": _v(d1,2), "cess": _v(d1,3)},
            "2": {"igst": _v(d2,0), "cgst": _v(d2,1), "sgst": _v(d2,2), "cess": _v(d2,3)},
        },
    }
    # Log non-zero values for verification
    nz = [(k1+k2, f"{vv}") for k1,sec in result.items()
          for k2,row in (sec.items() if isinstance(sec,dict) and not any(isinstance(v,float) for v in sec.values()) else {k1:sec}.items())
          for _k,vv in (row.items() if isinstance(row,dict) else {k2:row}.items())
          if isinstance(vv,float) and vv != 0.0]
    if nz:
        g3b_log(f"      ✔ 4: {len(nz)} non-zero values (e.g. {nz[0][0]}={nz[0][1]})")
    else:
        g3b_log(f"      ⚠ 4: all values are 0")
    return result



def g3b_scrape_5(page):
    _g3b_click_section(page, "5")
    v = _g3b_table_vals(page)
    _g3b_log_verify(v, "5")
    return {
        "composition_exempt_nil": {"inter_state": _g(v,0), "intra_state": _g(v,1)},
        "non_gst":                {"inter_state": _g(v,2), "intra_state": _g(v,3)},
    }


def g3b_scrape_5_1(page):
    _g3b_click_section(page, "5_1")
    v = _g3b_table_vals(page)
    _g3b_log_verify(v, "5.1")
    return {
        "interest": {"igst": _g(v,0), "cgst": _g(v,1), "sgst": _g(v,2), "cess": _g(v,3)},
        "late_fee": {"cgst": _g(v,4), "sgst": _g(v,5)},
    }


def g3b_scrape_6_1(page):
    """
    Section 6.1 — Payment of tax.

    TWO TABLES on the page (confirmed from screenshots):

    TABLE 1 — Cash Ledger Balance + Credit Ledger Balance (top):
      Identified by header text containing "Cash Ledger Balance"
      Rows: "Tax" | "Interest" | "Late Fees"
      Cash cols:   IGST | CGST | SGST | CESS | Total  (5 inputs)
      Credit cols: IGST | CGST | SGST | CESS | Total  (5 inputs)
      Tax row: 10 inputs (5 cash + 5 credit)
      Interest row: 5 inputs (cash only)
      Late Fees row: 3 inputs (CGST + SGST + Total, cash only)

    TABLE 2 — Tax Payment detail (bottom, scrollable):
      Identified by header text containing "Net Tax Payable"
      Row labels: "Integrated Tax (₹)" | "Central Tax (₹)" | "State/UT Tax (₹)" | "CESS (₹)"
      Columns per row (left→right, up to 14 inputs):
        0: net_payable_rcm    (col 6, "Reverse charge and supplies u/s 9(5)")
        1: net_payable_other  (col 7, "Other than reverse charge")
        2: itc_igst           (col 8, "Paid through ITC - IGST")
        3: itc_cgst           (col 9, "Paid through ITC - CGST")
        4: itc_sgst           (col 10, "Paid through ITC - SGST")
        5: itc_cess           (col 11, "Paid through ITC - CESS")
        6: cash_other         (col 12, "Other than reverse charge - Cash")
        7: cash_rcm           (col 13, "Reverse charge Tax to be paid in Cash")
        8: interest_payable   (col 14)
        9: interest_cash      (col 15)
       10: late_fee_payable   (col 16)
       11: late_fee_cash      (col 17)
       12: utilizable_cash    (col 18)
       13: additional_cash    (col 19)

    CRITICAL: _row("integrated tax") wrongly matches the Cash Ledger "Tax" row first.
    Fix: find each table by its unique header, then scope TR searches WITHIN that table.
    """
    _g3b_click_section(page, "6_1")
    time.sleep(2.0)

    def _v(lst, i): return lst[i] if i < len(lst) else 0.0

    # ── Read the entire page once — two separate table extractions ────────────
    raw = page.evaluate("""() => {
        function parseVal(s) {
            return parseFloat((s || '').replace(/,/g, '').trim()) || 0;
        }

        // ── TABLE 1: find container with "Cash Ledger Balance" ────────────────
        // Look for a TABLE element that contains that header text
        let t1 = null;
        for (const el of document.querySelectorAll('table')) {
            if ((el.textContent || '').includes('Cash Ledger Balance')) {
                t1 = el; break;
            }
        }
        // Fallback: any ancestor div/section containing it
        if (!t1) {
            for (const el of document.querySelectorAll('div, section')) {
                if ((el.textContent || '').includes('Cash Ledger Balance')
                    && el.querySelector('table')) {
                    t1 = el.querySelector('table'); break;
                }
            }
        }

        let table1 = {
            tax:      [],
            interest: [],
            late_fee: [],
        };
        if (t1) {
            for (const tr of t1.querySelectorAll('tr')) {
                const txt = (tr.textContent || '').trim().toLowerCase();
                const inputs = Array.from(tr.querySelectorAll('input'))
                                   .map(i => parseVal(i.value));
                if (inputs.length === 0) continue;
                if (txt.includes('interest')) { table1.interest = inputs; continue; }
                if (txt.includes('late fee') || txt.includes('late fees'))
                                              { table1.late_fee = inputs; continue; }
                // "Tax" row — must NOT include "interest", "late" etc.
                if (txt.includes('tax') && !txt.includes('interest')
                    && !txt.includes('late') && !txt.includes('integrated')
                    && !txt.includes('central') && !txt.includes('state')
                    && !txt.includes('cess') && inputs.length >= 3)
                                              { table1.tax = inputs; }
            }
        }

        // ── TABLE 2: find container with "Net Tax Payable" ────────────────────
        let t2 = null;
        for (const el of document.querySelectorAll('table')) {
            if ((el.textContent || '').includes('Net Tax Payable')) {
                t2 = el; break;
            }
        }
        if (!t2) {
            for (const el of document.querySelectorAll('div, section')) {
                if ((el.textContent || '').includes('Net Tax Payable')
                    && el.querySelector('table')) {
                    t2 = el.querySelector('table'); break;
                }
            }
        }

        let table2 = {
            igst: [],
            cgst: [],
            sgst: [],
            cess: [],
        };
        if (t2) {
            for (const tr of t2.querySelectorAll('tr')) {
                const txt = (tr.textContent || '').trim().toLowerCase();

                // The payment table cells are NOT <input> elements.
                // They are read-only display cells rendered as <td> or <span>/<div>
                // styled to look like input boxes.
                // Strategy: read all <td> cells in this row, parse numeric text.

                // Helper: extract numeric value from a cell
                function cellVal(td) {
                    // First try an <input> inside the cell
                    const inp = td.querySelector('input');
                    if (inp) return parseVal(inp.value);
                    // Then try text content of the cell itself
                    const t = (td.textContent || '').replace(/,/g, '').trim();
                    return parseVal(t);
                }

                const tds = Array.from(tr.querySelectorAll('td'));
                if (tds.length < 2) continue;

                // First TD is the row label — skip it; rest are values
                // Also skip rows where ALL cells are zero/empty (header/separator rows)
                const vals = tds.slice(1).map(cellVal);
                const hasData = vals.some(v => v !== 0);

                // Match by row label text (first TD or overall TR text)
                const rowLabel = (tds[0].textContent || '').trim().toLowerCase();
                const isIgst = rowLabel.includes('integrated tax') || txt.startsWith('integrated tax');
                const isCgst = rowLabel.includes('central tax')    || txt.startsWith('central tax');
                const isSgst = (rowLabel.includes('state/ut') || rowLabel.includes('state / ut'))
                             || txt.startsWith('state/ut');
                const isCess = (rowLabel === 'cess (₹)' || rowLabel === 'cess'
                             || rowLabel.startsWith('cess'))
                             && !rowLabel.includes('integrated')
                             && !rowLabel.includes('central');

                if (isIgst && vals.length > 0) { table2.igst = vals; continue; }
                if (isCgst && vals.length > 0) { table2.cgst = vals; continue; }
                if (isSgst && vals.length > 0) { table2.sgst = vals; continue; }
                if (isCess && vals.length > 0) { table2.cess = vals; continue; }
            }
        }

        return { table1, table2,
                 t1_found: !!t1, t2_found: !!t2 };
    }""")

    t1     = raw.get("table1", {})
    t2     = raw.get("table2", {})
    t1_ok  = raw.get("t1_found", False)
    t2_ok  = raw.get("t2_found", False)

    tax_r  = t1.get("tax",      [])
    int_r  = t1.get("interest", [])
    lf_r   = t1.get("late_fee", [])

    igst_r = t2.get("igst", [])
    cgst_r = t2.get("cgst", [])
    sgst_r = t2.get("sgst", [])
    cess_r = t2.get("cess", [])

    g3b_log(f"      6.1 table1={'found' if t1_ok else 'NOT FOUND'} "
            f"(tax={len(tax_r)} int={len(int_r)} lf={len(lf_r)} vals), "
            f"table2={'found' if t2_ok else 'NOT FOUND'} "
            f"(igst={len(igst_r)} cgst={len(cgst_r)} sgst={len(sgst_r)} "
            f"cess={len(cess_r)} vals)")

    cash_ledger = {
        "tax": {
            "igst":  _v(tax_r, 0), "cgst":  _v(tax_r, 1),
            "sgst":  _v(tax_r, 2), "cess":  _v(tax_r, 3),
            "total": _v(tax_r, 4),
        },
        "interest": {
            "igst":  _v(int_r, 0), "cgst":  _v(int_r, 1),
            "sgst":  _v(int_r, 2), "cess":  _v(int_r, 3),
            "total": _v(int_r, 4),
        },
        "late_fee": {
            "cgst":  _v(lf_r, 0), "sgst":  _v(lf_r, 1), "total": _v(lf_r, 2),
        },
    }
    credit_ledger = {
        "tax": {
            "igst":  _v(tax_r, 5), "cgst":  _v(tax_r, 6),
            "sgst":  _v(tax_r, 7), "cess":  _v(tax_r, 8),
            "total": _v(tax_r, 9),
        },
    }

    def _payment_row(r):
        return {
            "net_payable_rcm":   _v(r, 0),
            "net_payable_other": _v(r, 1),
            "itc_igst":          _v(r, 2),
            "itc_cgst":          _v(r, 3),
            "itc_sgst":          _v(r, 4),
            "itc_cess":          _v(r, 5),
            "cash_other":        _v(r, 6),
            "cash_rcm":          _v(r, 7),
            "interest_payable":  _v(r, 8),
            "interest_cash":     _v(r, 9),
            "late_fee_payable":  _v(r, 10),
            "late_fee_cash":     _v(r, 11),
            "utilizable_cash":   _v(r, 12),
            "additional_cash":   _v(r, 13),
        }

    result = {
        "cash_ledger":   cash_ledger,
        "credit_ledger": credit_ledger,
        "tax_payment": {
            "igst": _payment_row(igst_r),
            "cgst": _payment_row(cgst_r),
            "sgst": _payment_row(sgst_r),
            "cess": _payment_row(cess_r),
        },
    }

    # Verification log
    nz_cash = _v(tax_r, 4)
    nz_igst = _v(igst_r, 1)
    nz_cgst = _v(cgst_r, 1)
    nz_sgst = _v(sgst_r, 1)
    parts = [f"cash_ledger.tax.total={nz_cash}"]
    if nz_igst: parts.append(f"igst.net_payable_other={nz_igst}")
    if nz_cgst: parts.append(f"cgst.net_payable_other={nz_cgst}")
    if nz_sgst: parts.append(f"sgst.net_payable_other={nz_sgst}")
    # Also log first non-zero payment value from each row for debugging
    for label, row in [("igst", igst_r), ("cgst", cgst_r), ("sgst", sgst_r), ("cess", cess_r)]:
        nz = [(i, v) for i, v in enumerate(row) if v != 0]
        if nz:
            parts.append(f"{label}[{nz[0][0]}]={nz[0][1]}")
    any_nz = any([nz_cash, nz_igst, nz_cgst, nz_sgst,
                  any(v != 0 for v in igst_r + cgst_r + sgst_r + cess_r)])
    if any_nz:
        g3b_log(f"      ✔ 6.1: non-zero — {', '.join(parts)}")
    else:
        g3b_log(f"      ⚠ 6.1: all values are 0 (nil or not filed?)")
    return result






# ── Tile reader ───────────────────────────────────────────────────────────────
def _g3b_read_tile(page):
    """
    Read the GSTR-3B tile from the dashboard after SEARCH.

    What the portal actually shows (confirmed from screenshots):
      QRMP filer  → card header "Quarterly Return | GSTR-3B | Quarterly"
                    info banner "...quarterly frequency..."
                    button:  "VIEW GSTR3B"  (filed)
                             "PREPARE ONLINE" (not yet filed)

      Monthly filer → card header "Monthly Return | GSTR-3B"
                      info banner "...monthly frequency..."
                      button:  "VIEW GSTR3B"  (filed)
                               "PREPARE ONLINE" (not yet filed)

    Returns dict:
        button    : 'view' | 'prepare_online' | 'resume' | 'none' | 'not_found'
        frequency : 'quarterly' | 'monthly' | 'unknown'
    """
    return page.evaluate("""() => {
        // ── Filing frequency from info banner ─────────────────────────────────
        // The portal shows a blue info banner AFTER search with the exact phrases:
        //   QRMP:    "...quarterly frequency..."
        //   Monthly: "...monthly frequency..."
        // This is the ONLY reliable QRMP signal — IFF/PMT-06/QRMP text can appear
        // as global news tickers for ALL taxpayer types and must NOT be used.
        let frequency = 'unknown';
        const pageText = (document.body.innerText || '').toLowerCase();
        if (pageText.includes('quarterly frequency'))       frequency = 'quarterly';
        else if (pageText.includes('monthly frequency'))    frequency = 'monthly';

        // ── GSTR-3B tile button ───────────────────────────────────────────────
        // The tile is identified by a text node whose sole content is "GSTR-3B"
        // or "GSTR3B". Walk up to find the card, then read the button inside it.
        // Button text is "VIEW GSTR3B" (filed), "PREPARE ONLINE" (unfiled), "RESUME".
        for (const el of document.querySelectorAll('*')) {
            // Only match elements whose entire text content is the form code
            const own = (el.childNodes.length === 1 && el.childNodes[0].nodeType === 3
                ? el.textContent
                : (el.getAttribute('data-title') || '')).trim();
            if (own !== 'GSTR3B' && own !== 'GSTR-3B') continue;

            // Walk up toward the card container
            let card = el.parentElement;
            for (let i = 0; i < 14 && card && card !== document.body; i++) {
                for (const btn of card.querySelectorAll('button, a')) {
                    const bt = (btn.textContent || '').trim().toUpperCase()
                                                           .replace(/\\s+/g, ' ');
                    // "VIEW GSTR3B" / "VIEW GSTR 3B" / "VIEW GSTR-3B"
                    if (bt.startsWith('VIEW GSTR'))  return {button: 'view',          frequency};
                    if (bt === 'PREPARE ONLINE')     return {button: 'prepare_online', frequency};
                    if (bt === 'RESUME')             return {button: 'resume',         frequency};
                }
                card = card.parentElement;
            }
            // Found the GSTR-3B label but no button → tile exists without button
            return {button: 'none', frequency};
        }
        return {button: 'not_found', frequency};
    }""")


def _g3b_click_tile(page):
    """
    Click the VIEW GSTR3B (or PREPARE ONLINE / RESUME) button inside the
    GSTR-3B tile card.  Returns the button text that was clicked, or None.
    """
    return page.evaluate("""() => {
        for (const el of document.querySelectorAll('*')) {
            const own = (el.childNodes.length === 1 && el.childNodes[0].nodeType === 3
                ? el.textContent
                : (el.getAttribute('data-title') || '')).trim();
            if (own !== 'GSTR3B' && own !== 'GSTR-3B') continue;

            let card = el.parentElement;
            for (let i = 0; i < 14 && card && card !== document.body; i++) {
                for (const btn of card.querySelectorAll('button, a')) {
                    const bt = (btn.textContent || '').trim().toUpperCase()
                                                           .replace(/\\s+/g, ' ');
                    if (bt.startsWith('VIEW GSTR') ||
                        bt === 'PREPARE ONLINE'    ||
                        bt === 'RESUME') {
                        btn.click();
                        return bt;
                    }
                }
                card = card.parentElement;
            }
            return null;  // tile found but no button
        }
        return null;  // tile not found at all
    }""")


# ── Dashboard navigation: FY/Quarter/Month → GSTR-3B tile → VIEW ──────────────
def g3b_navigate_to_period(page, context, fy, mon_num, already_qrmp=False):
    """
    Go to Returns Dashboard, select the period, click VIEW GSTR3B.

    IMPORTANT: The portal opens the GSTR-3B Angular app in a NEW TAB
    (target=_blank on the VIEW GSTR3B button). After clicking we must scan
    context.pages for the gstr3b URL — NOT check page.url which stays on
    the dashboard forever.

    Returns (True,  result, frequency, form_page) on success
            (False, reason, frequency, None)       on failure
    """
    mon_name  = _G3B_MON[mon_num]
    qtr_label = _G3B_QTR[mon_num]

    try:
        page.evaluate(f"window.location.href = '{_G3B_DASHBOARD}'")
    except Exception:
        page.goto(_G3B_DASHBOARD, wait_until="domcontentloaded", timeout=15000)
    try:
        page.wait_for_load_state("domcontentloaded", timeout=12000)
    except Exception:
        pass
    time.sleep(1.5)

    def pw_select(nth, label, timeout=8):
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                loc = page.locator("select").nth(nth)
                loc.wait_for(state="visible", timeout=2000)
                loc.select_option(label=label)
                time.sleep(0.4)
                cur = loc.evaluate(
                    "el => el.options[el.selectedIndex]"
                    " ? el.options[el.selectedIndex].text.trim() : ''")
                if cur == label: return True
                loc.select_option(label=label); time.sleep(0.5); return True
            except Exception:
                time.sleep(0.4)
        return False

    def wait_for_opt(nth, label, timeout=8):
        deadline = time.time() + timeout
        while time.time() < deadline:
            try:
                opts = page.locator("select").nth(nth).evaluate(
                    "el => Array.from(el.options).map(o => o.text.trim())")
                if label in opts: return True
            except Exception:
                pass
            time.sleep(0.4)
        return False

    g3b_log(f"    Selecting FY: {fy}")
    pw_select(0, fy)
    time.sleep(0.5)

    n_sel = 0
    for _ in range(15):
        try:
            n_sel = page.evaluate(
                "() => document.querySelectorAll('select').length")
        except Exception:
            pass
        if n_sel >= 2: break
        time.sleep(0.3)
    g3b_log(f"    {n_sel} select(s) detected after FY change")

    if n_sel >= 3:
        g3b_log(f"    Selecting Quarter: {qtr_label}")
        wait_for_opt(1, qtr_label, timeout=6)
        pw_select(1, qtr_label)
        g3b_log(f"    Selecting Month: {mon_name}")
        wait_for_opt(2, mon_name, timeout=8)
        pw_select(2, mon_name)
    else:
        g3b_log(f"    Selecting Month: {mon_name}")
        wait_for_opt(1, mon_name, timeout=8)
        pw_select(1, mon_name)

    for s in ["button:has-text('SEARCH')", "button:has-text('Search')"]:
        try:
            page.locator(s).first.wait_for(state="visible", timeout=4000)
            page.locator(s).first.click()
            g3b_log("    \u2713 SEARCH clicked"); break
        except Exception: continue
    time.sleep(2.5)

    # ── Read tile ─────────────────────────────────────────────────────────────
    tile = _g3b_read_tile(page)
    result    = tile["button"]    # 'view'|'prepare_online'|'resume'|'none'|'not_found'
    frequency = tile["frequency"] # 'quarterly'|'monthly'|'unknown'
    g3b_log(f"    Dashboard result: {result}  |  frequency: {frequency}")

    if result in ("none", "not_found"):
        return False, "tile_not_found", frequency, None
    if result not in ("view", "prepare_online", "resume"):
        return False, result, frequency, None

    # ── Click the VIEW GSTR3B button ──────────────────────────────────────────
    # The portal sometimes opens gstr3b in the SAME tab (same-tab navigation via
    # Angular router), sometimes in a NEW TAB (target=_blank button).
    # Strategy:
    #   1. Use Playwright native locator click (not JS evaluate) so Playwright
    #      tracks the navigation properly.
    #   2. Wait for the SAME page to navigate to gstr3b (most common case).
    #   3. If same-tab doesn't happen, scan context.pages for a new gstr3b tab.
    clicked_text = None
    try:
        # Find and Playwright-click the button (not JS evaluate)
        from playwright.sync_api import TimeoutError as PWTimeout
        for sel in [
            "button:has-text('VIEW GSTR3B')",
            "button:has-text('VIEW GSTR 3B')",
            "button:has-text('VIEW GSTR-3B')",
            "a:has-text('VIEW GSTR3B')",
            "a:has-text('VIEW GSTR 3B')",
            "button:has-text('PREPARE ONLINE')",
            "button:has-text('RESUME')",
        ]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                clicked_text = loc.inner_text().strip()
                loc.click()
                g3b_log(f"    Clicked: {clicked_text!r} (Playwright click)")
                break
            except Exception:
                continue
    except Exception as ce:
        g3b_log(f"    ⚠ Playwright click error: {ce}")

    if not clicked_text:
        # JS fallback
        clicked_text = _g3b_click_tile(page)
        g3b_log(f"    Clicked: {clicked_text!r} (JS fallback)")

    # ── Wait for gstr3b URL — same tab first (longer timeout), then scan ────────
    form_page   = None
    is_new_tab  = False

    # Strategy 1: same-tab navigation — wait up to 20s (portal can be slow)
    try:
        page.wait_for_url("*gstr3b*", timeout=20000)
        if "gstr3b" in page.url.lower():
            form_page  = page
            is_new_tab = False
            g3b_log(f"    ✓ Same-tab navigation to gstr3b")
    except Exception:
        # Timed out — check if page is already there despite the exception
        try:
            if "gstr3b" in page.url.lower():
                form_page  = page
                is_new_tab = False
                g3b_log(f"    ✓ Same-tab (URL check after timeout)")
        except Exception:
            pass

    # Strategy 2: scan all context pages for gstr3b
    # CRITICAL: if page itself is the gstr3b page → same tab (is_new_tab=False)
    # Only set is_new_tab=True if it is a DIFFERENT page object
    if not form_page:
        for _ in range(12):
            time.sleep(1)
            for p in context.pages:
                try:
                    if "gstr3b" in p.url.lower():
                        form_page  = p
                        is_new_tab = (p is not page)
                        tab_type   = "New-tab" if is_new_tab else "Same-tab (found in scan)"
                        g3b_log(f"    ✓ {tab_type} navigation to gstr3b")
                        break
                except Exception:
                    continue
            if form_page:
                break

    if not form_page:
        g3b_log(f"    ❌ No gstr3b page found after click. Open tabs:")
        for p in context.pages:
            try: g3b_log(f"       {p.url[:80]}")
            except Exception: pass
        return False, "wrong_url", frequency, None

    # Sync
    try:
        form_page.wait_for_load_state("domcontentloaded", timeout=8000)
    except Exception:
        pass
    time.sleep(1.5)

    # Dismiss "System generated summary for GSTR-3B" popup (CLOSE button)
    for sel in ["button:has-text('CLOSE')", "button:has-text('Close')",
                "button:has-text('OK')",    "button:has-text('Ok')"]:
        try:
            btn = form_page.locator(sel).first
            btn.wait_for(state="visible", timeout=4000)
            btn.click()
            g3b_log(f"    ✓ Dismissed summary popup ({sel})")
            time.sleep(0.5)
            break
        except Exception:
            continue

    g3b_log(f"    ✓ Form loaded ({result}): {form_page.url[:70]}")
    # is_new_tab is returned so the caller knows whether to close the tab
    return True, result, frequency, (form_page, is_new_tab)


# ── Period scraper ─────────────────────────────────────────────────────────────
def g3b_scrape_period(page, fy, mon_num, gstin, gstin_dir, filing_type):
    """
    Scrape all 7 sections of the GSTR-3B form for the given period.
    IMPORTANT: page must already be inside the Angular GSTR-3B app (period loaded).
    We navigate between sections using Angular's own router — never window.location.href.
    """
    mon_name   = _G3B_MON[mon_num]
    fy_year    = int(fy[:4])
    year_label = str(fy_year) if mon_num >= 4 else str(fy_year + 1)
    fy_tag     = fy.replace("-", "_")
    fname      = f"GSTR3B_{mon_name}_{fy_tag}.json"
    outpath    = os.path.join(gstin_dir, fname)

    g3b_log(f"  \U0001f4cb {mon_name} {year_label} \u2014 scraping all sections")
    g3b_log(f"    Current URL: {page.url[:70]}")

    try:
        # DO NOT navigate away from this page — Angular period context is here.
        g3b_log(f"    \u2192 3.1  Outward + RCM")
        t31  = g3b_scrape_3_1(page)
        g3b_log(f"    \u2192 3.1.1 ECO / s9(5)")
        t311 = g3b_scrape_3_1_1(page)
        g3b_log(f"    \u2192 3.2  Inter-state supplies (state-wise)")
        t32  = g3b_scrape_3_2(page)
        g3b_log(f"    \u2192 4    Eligible ITC")
        t4   = g3b_scrape_4(page)
        g3b_log(f"    \u2192 5    Exempt / nil / non-GST inward")
        t5   = g3b_scrape_5(page)
        g3b_log(f"    \u2192 5.1  Interest & Late fee")
        t51  = g3b_scrape_5_1(page)
        g3b_log(f"    \u2192 6.1  Payment of tax")
        t61  = g3b_scrape_6_1(page)

        payload = {
            "gstin":        gstin,
            "fy":           fy,
            "period":       f"{mon_name} {year_label}",
            "period_num":   mon_num,
            "filing_type":  filing_type,
            "scraped_at":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_url":   page.url,
            "table_3_1":    t31,
            "table_3_1_1":  t311,
            "table_3_2":    t32,
            "table_4":      t4,
            "table_5":      t5,
            "table_5_1":    t51,
            "table_6_1":    t61,
        }
        os.makedirs(gstin_dir, exist_ok=True)
        with open(outpath, "w") as f:
            json.dump(payload, f, indent=2)

        size_kb = os.path.getsize(outpath) / 1024
        # Check for suspicious all-zero file
        all_vals = []
        for tbl in [t31, t311, t4, t5, t51]:
            for row in tbl.values():
                if isinstance(row, dict):
                    all_vals.extend(v for v in row.values() if isinstance(v, (int, float)))
        nonzero_count = sum(1 for v in all_vals if v != 0.0)
        if nonzero_count == 0:
            g3b_log(f"    \u26a0 WARNING: all values are zero in {fname} "
                    f"({size_kb:.1f} KB) \u2014 Angular nav may not have worked")
        else:
            g3b_log(f"    \u2705 Saved: {fname} ({size_kb:.1f} KB, "
                    f"{nonzero_count} non-zero values)")
        return outpath

    except Exception as e:
        import traceback as _tb
        g3b_log(f"    \u274c Error in {mon_name}: {e}")
        g3b_log(_tb.format_exc())
        return None


# ── Upfront QRMP probe ────────────────────────────────────────────────────────
def g3b_probe_qrmp(page, fy):
    """
    Determine filing frequency (QRMP vs Monthly) before the scraping loop.

    Uses _g3b_read_tile() which reads the info banner text for the exact phrases:
      "quarterly frequency" → QRMP
      "monthly frequency"   → Monthly
    This is the ONLY reliable signal — IFF/PMT-06/QRMP text appears as global
    news tickers for ALL taxpayer types and must NOT be used.

    Algorithm:
      Stage 1 — Navigate to a non-quarter-end month (May).
                If info banner says "quarterly frequency" → QRMP confirmed.
                If info banner says "monthly frequency"   → Monthly confirmed.
      Stage 2 — If banner is absent/unclear:
                tile button = 'view' on non-quarter month → Monthly filer
                              (they actually filed that month)
                tile button = 'prepare_online' → cross-check quarter-end month
                tile not found                 → QRMP (no 3B tile for non-quarter)
      Stage 3 — Cross-check with June (quarter-end):
                June tile = 'view' + May = 'prepare_online' → QRMP confirmed
    """
    g3b_log("  \U0001f50d Probing filing frequency (QRMP vs Monthly)...")

    fy_start = int(fy[:4])
    now      = datetime.now()

    def _nav_and_read(mon_num):
        """Navigate dashboard to mon_num, return _g3b_read_tile() dict."""
        mon_name  = _G3B_MON[mon_num]
        qtr_label = _G3B_QTR[mon_num]
        try:
            page.evaluate(f"window.location.href = '{_G3B_DASHBOARD}'")
        except Exception:
            page.goto(_G3B_DASHBOARD, wait_until="domcontentloaded", timeout=15000)
        try: page.wait_for_load_state("domcontentloaded", timeout=12000)
        except Exception: pass
        time.sleep(1.5)

        def _pw(nth, label, timeout=6):
            deadline = time.time() + timeout
            while time.time() < deadline:
                try:
                    loc = page.locator("select").nth(nth)
                    loc.wait_for(state="visible", timeout=2000)
                    loc.select_option(label=label); time.sleep(0.4)
                    cur = loc.evaluate(
                        "el => el.options[el.selectedIndex]"
                        " ? el.options[el.selectedIndex].text.trim() : ''")
                    if cur == label: return True
                    loc.select_option(label=label); time.sleep(0.5); return True
                except Exception: time.sleep(0.4)
            return False

        def _wait(nth, label, timeout=6):
            deadline = time.time() + timeout
            while time.time() < deadline:
                try:
                    opts = page.locator("select").nth(nth).evaluate(
                        "el => Array.from(el.options).map(o => o.text.trim())")
                    if label in opts: return True
                except Exception: pass
                time.sleep(0.4)
            return False

        _pw(0, fy); time.sleep(0.5)
        n_sel = 0
        for _ in range(15):
            try: n_sel = page.evaluate(
                "() => document.querySelectorAll('select').length")
            except Exception: pass
            if n_sel >= 2: break
            time.sleep(0.3)

        if n_sel >= 3:
            _wait(1, qtr_label); _pw(1, qtr_label)
            _wait(2, mon_name);  _pw(2, mon_name)
        else:
            _wait(1, mon_name);  _pw(1, mon_name)

        for s in ["button:has-text('SEARCH')", "button:has-text('Search')"]:
            try:
                page.locator(s).first.wait_for(state="visible", timeout=4000)
                page.locator(s).first.click(); break
            except Exception: continue
        time.sleep(2.5)

        return _g3b_read_tile(page)

    # ── Probe May (non-quarter-end) ───────────────────────────────────────────
    probe_non_qtr = 5   # May
    probe_qtr_end = 6   # June

    may_year = fy_start if probe_non_qtr >= 4 else fy_start + 1
    if datetime(may_year, probe_non_qtr, 1) > now:
        probe_non_qtr = 4   # FY just started — use April

    g3b_log(f"    Checking {_G3B_MON[probe_non_qtr]} (non-quarter month)...")
    tile = _nav_and_read(probe_non_qtr)
    btn  = tile.get("button",    "unknown")
    freq = tile.get("frequency", "unknown")
    g3b_log(f"    {_G3B_MON[probe_non_qtr]}: button={btn!r}  frequency={freq!r}")

    # Stage 1 — info banner is definitive
    if freq == "quarterly":
        g3b_log("  \U0001f4cc QRMP confirmed (info banner: 'quarterly frequency') "
                "\u2014 quarter-end months only (Jun/Sep/Dec/Mar)")
        return True
    if freq == "monthly":
        g3b_log("  \u2714 Monthly filer confirmed (info banner: 'monthly frequency')")
        return False

    # Stage 2 — no banner; fall back to tile button
    if btn == "not_found":
        g3b_log("  \U0001f4cc QRMP confirmed (no GSTR-3B tile for "
                f"{_G3B_MON[probe_non_qtr]})")
        return True
    if btn == "view":
        g3b_log(f"  \u2714 Monthly filer confirmed ({_G3B_MON[probe_non_qtr]} = VIEW)")
        return False

    # Stage 3 — prepare_online on non-quarter: cross-check quarter-end
    jun_year = fy_start if probe_qtr_end >= 4 else fy_start + 1
    if datetime(jun_year, probe_qtr_end, 1) <= now:
        g3b_log(f"    {_G3B_MON[probe_non_qtr]} = PREPARE ONLINE \u2014 "
                f"cross-checking {_G3B_MON[probe_qtr_end]}...")
        tile2 = _nav_and_read(probe_qtr_end)
        btn2  = tile2.get("button",    "unknown")
        freq2 = tile2.get("frequency", "unknown")
        g3b_log(f"    {_G3B_MON[probe_qtr_end]}: button={btn2!r}  frequency={freq2!r}")

        if freq2 == "quarterly":
            g3b_log("  \U0001f4cc QRMP confirmed on cross-check (quarterly banner)")
            return True
        if freq2 == "monthly":
            g3b_log("  \u2714 Monthly filer confirmed on cross-check (monthly banner)")
            return False
        if btn2 == "view":
            g3b_log(f"  \U0001f4cc QRMP confirmed: {_G3B_MON[probe_qtr_end]}=VIEW "
                    f"(quarterly filed) + {_G3B_MON[probe_non_qtr]}=PREPARE ONLINE")
            return True

    # Inconclusive — safer to treat as monthly
    g3b_log(f"  \u2714 Treating as Monthly (probe inconclusive — "
            f"{_G3B_MON[probe_non_qtr]}=btn:{btn}/freq:{freq})")
    return False


# ── Main worker ────────────────────────────────────────────────────────────────
def g3b_worker(gstin, fy, username, password, base_dir, force=False, specific_month=""):
    """
    Thin wrapper guaranteeing any exception is logged and reflected in
    g3b_state instead of silently killing the thread — belt-and-suspenders
    alongside the existing catch-all further down in _g3b_worker_impl.
    """
    try:
        _g3b_worker_impl(gstin, fy, username, password, base_dir, force, specific_month)
    except Exception as fatal:
        import traceback as _tb
        try:
            g3b_log(f"✗ FATAL (uncaught): {fatal}", "error")
        except Exception:
            pass
        try:
            g3b_set({"status": "error", "error": f"Uncaught: {fatal}"})
        except Exception:
            pass
        log.error(f"[G3B] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _g3b_worker_impl(gstin, fy, username, password, base_dir, force=False, specific_month=""):
    import traceback
    from playwright.sync_api import sync_playwright

    g3b_set({
        "status":         "running",
        "log":            [],
        "files":          [],
        "fresh_files":    [],
        "error":          None,
        "captcha_image":  None,
        "captcha_answer": None,
        "otp_answer":     None,
    })
    g3b_log(f"GSTR-3B Scraper | GSTIN: {gstin} | FY: {fy}")

    fy_start  = int(fy[:4])
    fy_tag    = fy.replace("-", "_")
    gstin_dir = os.path.join(base_dir, gstin, fy_tag)

    months_in_fy  = list(range(4, 13)) + [1, 2, 3]
    now           = datetime.now()
    months_needed = []
    fresh_files   = []

    for m in months_in_fy:
        year  = fy_start if m >= 4 else fy_start + 1
        if datetime(year, m, 1) > now:
            continue
        fname = f"GSTR3B_{_G3B_MON[m]}_{fy_tag}.json"
        fpath = os.path.join(gstin_dir, fname)
        if not force and os.path.exists(fpath):
            fresh_files.append(fpath)
            g3b_log(f"  \u2713 Cache: {fname}")
        else:
            months_needed.append(m)

    # ── Apply specific_month filter if requested ──────────────────────────────
    if specific_month:
        sm = specific_month.strip()
        before = list(months_needed)
        months_needed = [
            m for m in months_needed
            if sm.lower() == _G3B_MON[m].lower()
            or sm.lower().startswith(_G3B_MON[m].lower()[:3])
        ]
        if not months_needed:
            # Check if it was already cached (no need to scrape)
            all_possible = [m for m in ([4,5,6,7,8,9,10,11,12,1,2,3])
                            if sm.lower() == _G3B_MON[m].lower()
                            or sm.lower().startswith(_G3B_MON[m].lower()[:3])]
            if not all_possible:
                g3b_set({"status":"error","error":
                         f"Month '{specific_month}' not found in FY {fy}."}); return
            g3b_log(f"  \u2713 {specific_month} is already cached — "
                    f"use Force re-scrape to overwrite")
            g3b_set({"status":"done",
                     "files":[os.path.basename(f) for f in fresh_files],
                     "fresh_files":[os.path.basename(f) for f in fresh_files]}); return
        else:
            g3b_log(f"  [Month] Specific month: {_G3B_MON[months_needed[0]]}")

    if not months_needed:
        g3b_log("\u2705 All months already cached \u2014 nothing to scrape")
        g3b_set({"files":        [os.path.basename(f) for f in fresh_files],
                 "fresh_files":  [os.path.basename(f) for f in fresh_files],
                 "status":       "done"})
        return

    g3b_log(f"  {len(months_needed)} month(s) to scrape, "
            f"{len(fresh_files)} from cache")
    g3b_set({"fresh_files": [os.path.basename(f) for f in fresh_files]})

    # Profile dir (per-GSTIN, same pattern as g2a) — previously used
    # os.path.expanduser("~")/.gst_rpa_profiles directly, the only
    # worker in the app that stored its browser profile outside the
    # user-chosen data folder instead of PATHS.profiles_dir. Fixed for
    # consistency with every other worker (g2a, g2b, g1, tds, GDIR,
    # combined) — the whole point of the config/data-folder setup is
    # that everything the app writes lives in one place the user chose.
    profile_base = PATHS.profiles_dir
    os.makedirs(profile_base, exist_ok=True)
    profile_dir  = os.path.join(profile_base, f"g3b_{gstin}")
    lock_file    = os.path.join(profile_dir, "SingletonLock")
    if os.path.exists(lock_file):
        try: os.remove(lock_file); g3b_log("  \u2713 Removed stale SingletonLock")
        except Exception: pass

    # sync_playwright().start() launches Playwright's own Node.js driver
    # subprocess — if it throws (not just hangs), that was previously
    # uncaught here, silently killing this thread with nothing logged.
    try:
        pw = sync_playwright().start()
    except Exception as e:
        g3b_log(f"✗ Playwright driver failed to start: {e}", "error")
        g3b_set({"status": "error", "error": f"Playwright driver failed to start: {e}"})
        return
    try:
        g3b_log("\U0001f310 Launching browser...")
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=False, slow_mo=40,
                args=["--disable-blink-features=AutomationControlled",
                      "--no-sandbox", "--start-maximized"],
                no_viewport=True, accept_downloads=True,
            )
        except Exception as e:
            g3b_log(f"  \u2717 Browser launch error: {e}", )
            g3b_set({"status": "error", "error": str(e)}); return

        page = context.pages[0] if context.pages else context.new_page()
        page.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")

        # ── Login (mirrors g2a exactly) ────────────────────────────────────────
        if not g3b_do_browser_login(page, username, password):
            context.close(); return

        # ── Session activation (mirrors g2a exactly) ───────────────────────────
        g3b_log("\U0001f511 Activating session on Returns portal...")
        ok, page = g3b_activate_session(page, context)
        if not ok:
            g3b_set({"status": "error",
                     "error":  "Could not reach Returns portal"})
            context.close(); return
        g3b_log(f"  Using page: {page.url[:70]}")

        # ── No upfront QRMP probe — every month is tried independently ────────
        # Taxpayers can switch QRMP ↔ Monthly at each quarter end, so a single
        # probe of May cannot determine the frequency for Jul-Sep or Oct-Dec.
        g3b_log(f"\U0001f4cb Starting GSTR-3B scrape sweep (per-month frequency detection)...")

        new_files = []

        for mon_num in months_needed:
            mon_name = _G3B_MON[mon_num]
            g3b_log(f"\n[{mon_name}] Processing...")
            g3b_set({"status": "running"})

            # Every month is tried independently — no global QRMP assumption.
            ok, reason, freq, nav_result = g3b_navigate_to_period(
                page, context, fy, mon_num, already_qrmp=False)

            # Per-month frequency from the info banner (set by _g3b_read_tile)
            # freq = "quarterly" / "monthly" / "unknown" for THIS period only
            mon_is_qrmp = (freq == "quarterly")
            if freq == "quarterly":
                g3b_log(f"  ℹ {mon_name}: quarterly frequency (QRMP for this quarter)")
            elif freq == "monthly":
                g3b_log(f"  ℹ {mon_name}: monthly frequency")

            if not ok:
                if reason == "tile_not_found" and mon_num not in _G3B_QTR_END:
                    # No GSTR-3B tile for a non-quarter-end month = IFF month.
                    # QRMP for this quarter only — other quarters are independent.
                    g3b_log(f"  ⏭ {mon_name}: no GSTR-3B tile — IFF month "
                            f"(QRMP for this quarter), skipping this month only")
                elif reason == "qrmp_pmt06":
                    g3b_log(f"  ⏭ {mon_name}: PMT-06 / QRMP signal — "
                            f"IFF month, skipping this month only")
                elif reason == "wrong_url":
                    g3b_log(f"  ✗ {mon_name}: portal navigated to wrong form — "
                            f"skipping (check manually)")
                else:
                    g3b_log(f"  ⚠ {reason} for {mon_name} — skipping")
                continue

            # Only scrape VIEW status — return is filed and has real data.
            # prepare_online = unfiled (IFF month, or genuinely not filed yet).
            # resume         = draft saved but not submitted.
            if reason != "view":
                if reason == "prepare_online":
                    if mon_num not in _G3B_QTR_END:
                        g3b_log(f"  ⏭ {mon_name} = PREPARE ONLINE (non-quarter-end) — "
                                f"IFF month for this quarter, skipping this month only")
                    else:
                        g3b_log(f"  ⏭ {mon_name} = PREPARE ONLINE — "
                                f"return not yet filed, skipping")
                elif reason == "resume":
                    g3b_log(f"  ⏭ {mon_name} = RESUME (draft not submitted) — skipping")
                else:
                    g3b_log(f"  ⏭ {mon_name} = {reason} — skipping")
                continue

            form_page, is_new_tab = nav_result
            # Determine filing type for THIS specific period from the banner signal
            filing_type = "qrmp" if mon_is_qrmp else "monthly"
            fpath = g3b_scrape_period(
                form_page, fy, mon_num, gstin, gstin_dir, filing_type)
            if fpath:
                new_files.append(fpath)
            # Only close if it was a new tab — if same-tab, we need
            # this page to navigate back to dashboard for the next month
            if is_new_tab:
                try:
                    form_page.close()
                    g3b_log(f"    📄 Closed GSTR-3B tab for {mon_name}")
                except Exception:
                    pass
            else:
                # Same-tab: navigate back to dashboard for next iteration
                try:
                    page.evaluate(f"window.location.href = '{_G3B_DASHBOARD}'")
                    page.wait_for_load_state("domcontentloaded", timeout=8000)
                    time.sleep(1.0)
                except Exception:
                    pass

        all_files     = fresh_files + new_files
        all_basenames = [os.path.basename(f) for f in all_files]
        g3b_log(f"\n✅ Complete (per-quarter frequency detection) — "
                f"{len(new_files)} scraped, "
                f"{len(fresh_files)} from cache, "
                f"{len(all_files)} total")
        g3b_log(f"   Folder: {gstin_dir}")
        g3b_set({"status":      "done",
                 "files":       all_basenames,
                 "fresh_files": [os.path.basename(f) for f in fresh_files]})

        try:   context.close(); g3b_log("\U0001f310 Browser closed")
        except Exception as ce: g3b_log(f"  \u26a0 Browser close: {ce}")

    except Exception as _e:
        g3b_set({"status": "error", "error": str(_e)})
        g3b_log(f"\u2717 Fatal worker error: {_e}")
        g3b_log(traceback.format_exc())
    finally:
        try: pw.stop()
        except Exception: pass

    time.sleep(8)
    with g3b_lock:
        if g3b_state.get("status") == "done":
            g3b_state["status"] = "idle"
    g3b_log("\U0001f501 GSTR-3B RPA reset to idle")



# ── GSTR-3B JSON → Excel ───────────────────────────────────────────────────────
def g3b_json_to_excel(gstin, fy):
    """
    Read all GSTR-3B JSON files for gstin+fy, build a multi-month
    Excel workbook that exactly matches the standard GSTR-3B filed report
    template (2 sheets: GSTR 3B + Section 3.2).

    Returns (BytesIO, error_string).
    """
    import glob
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
    from openpyxl.utils import get_column_letter

    fy_dir     = fy.replace("-", "_")
    folder     = os.path.join(G3B_DOWNLOAD_DIR, gstin, fy_dir)
    json_files = sorted(glob.glob(os.path.join(folder, "GSTR3B_*.json")))
    if not json_files:
        return None, f"No GSTR-3B JSON files found in {folder}"

    # ── Load each month ────────────────────────────────────────────────────────
    MONTHS_ORDER = [4,5,6,7,8,9,10,11,12,1,2,3]
    MON_ABBR = {4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",
                10:"Oct",11:"Nov",12:"Dec",1:"Jan",2:"Feb",3:"Mar"}
    MON_NAME = _G3B_MON  # already defined globally

    fy_start = int(fy[:4])

    data = {}   # mon_num -> loaded JSON dict
    for fpath in json_files:
        try:
            with open(fpath) as f:
                d = json.load(f)
            mn = d.get("period_num")
            if mn:
                data[mn] = d
        except Exception:
            pass

    if not data:
        return None, "All JSON files are empty or unreadable"

    def g(d, *keys):
        """Safe nested get with float default."""
        cur = d
        for k in keys:
            if not isinstance(cur, dict):
                return 0.0
            cur = cur.get(k, 0.0)
        return float(cur) if cur else 0.0

    def mon_val(mon_num, *path):
        """Get value from data[mon_num] at path, or 0 if month not loaded."""
        d = data.get(mon_num, {})
        return g(d, *path) if d else 0.0

    # ── Style helpers ──────────────────────────────────────────────────────────
    TEAL      = "1F7A8C"
    TEAL_LT   = "D6EEF2"
    NAVY      = "1F3864"
    PINK      = "C0143C"
    GREY_H    = "4A4A4A"
    GREY_LT   = "F2F2F2"
    BLUE_LT   = "EBF3FB"
    GOLD_LT   = "FFF8E1"
    GREEN_LT  = "E8F5E9"
    PURPLE_LT = "F3E5F5"
    RED_LT    = "FFEBEE"
    WHITE     = "FFFFFF"

    def fill(hex_): return PatternFill("solid", fgColor=hex_)
    def font(bold=False, color="000000", size=9, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    def align(h="center", v="center", wrap=True):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    thin  = Side(style="thin",   color="B0C4DE")
    thick = Side(style="medium", color="4A4A4A")
    def border(left=thin, right=thin, top=thin, bottom=thin):
        return Border(left=left, right=right, top=top, bottom=bottom)

    NUM_FMT = '#,##0.00'
    INT_FMT = '#,##0'

    def style_cell(cell, bold=False, bg=None, fc="000000", size=9,
                   h_align="center", wrap=True, num_fmt=None, italic=False):
        cell.font      = font(bold=bold, color=fc, size=size, italic=italic)
        cell.alignment = align(h=h_align, wrap=wrap)
        if bg:    cell.fill = fill(bg)
        if num_fmt: cell.number_format = num_fmt

    # ── Build "GSTR 3B" sheet ─────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR 3B"

    # Column widths
    ws.column_dimensions["A"].width = 1.5
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 48
    for col_idx in range(4, 17):   # D-P
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("B1:P1")
    c = ws["B1"]
    c.value = "Multi Month GSTR-3B Filed Report"
    style_cell(c, bold=True, bg=NAVY, fc=WHITE, size=13, h_align="center")
    ws.row_dimensions[1].height = 22

    # ── Row 2: Client + Duration ──────────────────────────────────────────────
    fy_start_y = int(fy[:4])
    fy_end_y   = fy_start_y + 1
    # Try to find client name from any JSON
    client_name = ""
    for d in data.values():
        if d.get("gstin"):
            client_name = d.get("gstin","")
            break
    label2 = f"{gstin}"
    dur    = f"Duration : Apr {fy_start_y} to Mar {fy_end_y}"

    ws.merge_cells("B2:H2")
    c = ws["B2"]; c.value = label2
    style_cell(c, bold=True, bg=TEAL, fc=WHITE, size=11, h_align="left")

    ws.merge_cells("I2:P2")
    c = ws["I2"]; c.value = dur
    style_cell(c, bold=True, bg=TEAL, fc=WHITE, size=11, h_align="right")
    ws.row_dimensions[2].height = 18

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    headers = ["", "S.No.", "PARTICULARS"] +               [f"{MON_ABBR[m]} '{str(fy_start_y if m>=4 else fy_end_y)[2:]}" for m in MONTHS_ORDER] +               ["Total"]
    for ci, hv in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=hv)
        style_cell(c, bold=True, bg=NAVY, fc=WHITE, size=9, h_align="center")
    ws.row_dimensions[3].height = 30

    # ── Helper: write a data row ──────────────────────────────────────────────
    def data_row(row_num, sno, label, extractor, bg=None, bold=False, section_bg=None):
        """
        Write one data row.
        extractor: callable(mon_num) -> float value
        """
        ws.cell(row=row_num, column=2, value=sno)
        c3 = ws.cell(row=row_num, column=3, value=label)
        style_cell(c3, bold=bold, bg=bg or GREY_LT, fc=GREY_H, size=9, h_align="left")

        col_vals = []
        for col_i, m in enumerate(MONTHS_ORDER, 4):
            v = extractor(m)
            c = ws.cell(row=row_num, column=col_i, value=v if v else 0)
            style_cell(c, bold=bold, bg=bg or WHITE, num_fmt=NUM_FMT)
            col_vals.append(v)

        # Total column (P = col 16)
        d_letter = get_column_letter(4)
        o_letter = get_column_letter(15)
        tc = ws.cell(row=row_num, column=16)
        tc.value = f"=SUM({d_letter}{row_num}:{o_letter}{row_num})"
        style_cell(tc, bold=True, bg=TEAL_LT, num_fmt=NUM_FMT)
        ws.row_dimensions[row_num].height = 16

    def section_header(row_num, label, bg=NAVY, fc=WHITE, bold=True, span_cols=True):
        """Full-width section header row."""
        if span_cols:
            ws.merge_cells(f"B{row_num}:P{row_num}")
        c = ws.cell(row=row_num, column=2, value=label)
        style_cell(c, bold=bold, bg=bg, fc=fc, size=9, h_align="left")
        ws.row_dimensions[row_num].height = 18

    def sub_header(row_num, sno, label, bg=BLUE_LT):
        ws.cell(row=row_num, column=2, value=sno)
        ws.merge_cells(f"C{row_num}:P{row_num}")
        c = ws.cell(row=row_num, column=3, value=label)
        style_cell(c, bold=True, bg=bg, fc=NAVY, size=9, h_align="left")
        ws.row_dimensions[row_num].height = 30

    def formula_row(row_num, sno, label, formula_fn, bg=GOLD_LT, bold=True):
        """Row computed from other cells via formula."""
        ws.cell(row=row_num, column=2, value=sno)
        c3 = ws.cell(row=row_num, column=3, value=label)
        style_cell(c3, bold=bold, bg=bg, fc=GREY_H, size=9, h_align="left")
        for col_i, m in enumerate(MONTHS_ORDER, 4):
            ltr = get_column_letter(col_i)
            c = ws.cell(row=row_num, column=col_i, value=formula_fn(ltr, row_num))
            style_cell(c, bold=bold, bg=bg, num_fmt=NUM_FMT)
        tc = ws.cell(row=row_num, column=16)
        d_l = get_column_letter(4)
        o_l = get_column_letter(15)
        tc.value = f"=SUM({d_l}{row_num}:{o_l}{row_num})"
        style_cell(tc, bold=True, bg=TEAL_LT, num_fmt=NUM_FMT)
        ws.row_dimensions[row_num].height = 16

    # ══════════════════════════════════════════════════════════════════════════
    # ROW MAP — matches the demo template exactly
    # ══════════════════════════════════════════════════════════════════════════

    # ── Section 3.1 ──────────────────────────────────────────────────────────
    section_header(4, "3.1 Details of Outward Supplies and inward supplies liable to reverse charge (other than those covered in 3.1.1)")
    sub_header(5, "(a)", "Outward taxable supplies (other than zero rated, nil rated and exempted)", BLUE_LT)
    data_row(6,  "(i)",   "Total Taxable Value",   lambda m: mon_val(m,"table_3_1","a","taxable_value"))
    data_row(7,  "(ii)",  "Integrated Tax",        lambda m: mon_val(m,"table_3_1","a","igst"))
    data_row(8,  "(iii)", "Central Tax",           lambda m: mon_val(m,"table_3_1","a","cgst"))
    data_row(9,  "(iv)",  "State/UT Tax",          lambda m: mon_val(m,"table_3_1","a","sgst"))
    data_row(10, "(v)",   "Cess",                  lambda m: mon_val(m,"table_3_1","a","cess"))
    sub_header(11,"(b)", "Outward taxable supplies (zero rated)", BLUE_LT)
    data_row(12, "(i)",   "Total Taxable Value",   lambda m: mon_val(m,"table_3_1","b","taxable_value"))
    data_row(13, "(ii)",  "Integrated Tax",        lambda m: mon_val(m,"table_3_1","b","igst"))
    data_row(14, "(iii)", "Cess",                  lambda m: mon_val(m,"table_3_1","b","cess"))
    sub_header(15,"(c)", "Other Outward Taxable supplies (Nil rated, exempted)", BLUE_LT)
    data_row(16, "(i)",   "Total Taxable Value",   lambda m: mon_val(m,"table_3_1","c","taxable_value"))
    sub_header(17,"(d)", "Inward Supplies Liable to Reverse charge", BLUE_LT)
    data_row(18, "(i)",   "Total Taxable Value",   lambda m: mon_val(m,"table_3_1","d","taxable_value"))
    data_row(19, "(ii)",  "Integrated Tax",        lambda m: mon_val(m,"table_3_1","d","igst"))
    data_row(20, "(iii)", "Central Tax",           lambda m: mon_val(m,"table_3_1","d","cgst"))
    data_row(21, "(iv)",  "State/UT Tax",          lambda m: mon_val(m,"table_3_1","d","sgst"))
    data_row(22, "(v)",   "Cess",                  lambda m: mon_val(m,"table_3_1","d","cess"))
    sub_header(23,"(e)", "Non-GST Outward Supplies", BLUE_LT)
    data_row(24, "(i)",   "Total Taxable Value",   lambda m: mon_val(m,"table_3_1","e","taxable_value"))

    # ── Section 3.1.1 ────────────────────────────────────────────────────────
    section_header(25, "3.1.1 Details of supplies notified under sub-section (5) of section 9 (ECO)")
    sub_header(26, "(i)", "Taxable supplies on which ECO pays tax u/s 9(5) [by ECO]", PURPLE_LT)
    data_row(27, "",      "Total Taxable Value",   lambda m: mon_val(m,"table_3_1_1","i","taxable_value"), bg=PURPLE_LT)
    data_row(28, "",      "Integrated Tax",        lambda m: mon_val(m,"table_3_1_1","i","igst"),          bg=PURPLE_LT)
    data_row(29, "",      "Central Tax",           lambda m: mon_val(m,"table_3_1_1","i","cgst"),          bg=PURPLE_LT)
    data_row(30, "",      "State/UT Tax",          lambda m: mon_val(m,"table_3_1_1","i","sgst"),          bg=PURPLE_LT)
    data_row(31, "",      "Cess",                  lambda m: mon_val(m,"table_3_1_1","i","cess"),          bg=PURPLE_LT)
    sub_header(32, "(ii)","Taxable supplies made by registered person through ECO [by registered person]", PURPLE_LT)
    data_row(33, "",      "Total Taxable Value",   lambda m: mon_val(m,"table_3_1_1","ii","taxable_value"), bg=PURPLE_LT)

    # ── Section 4 ─────────────────────────────────────────────────────────────
    section_header(34, "4.  Eligible ITC")
    sub_header(35, "A", "ITC Available (Whether in full or part)", GREEN_LT)
    sub_header(36, "1", " Import of goods", GREEN_LT)
    data_row(37, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","A","1","igst"), bg=GREEN_LT)
    data_row(38, "(ii)",  "Cess",                  lambda m: mon_val(m,"table_4","A","1","cess"), bg=GREEN_LT)
    sub_header(39, "2", "Import of services", GREEN_LT)
    data_row(40, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","A","2","igst"), bg=GREEN_LT)
    data_row(41, "(ii)",  "Cess",                  lambda m: mon_val(m,"table_4","A","2","cess"), bg=GREEN_LT)
    sub_header(42, "3", "Inward supplies liable to reverse charge (other than 1 & 2 above)", GREEN_LT)
    data_row(43, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","A","3","igst"), bg=GREEN_LT)
    data_row(44, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","A","3","cgst"), bg=GREEN_LT)
    data_row(45, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","A","3","sgst"), bg=GREEN_LT)
    data_row(46, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","A","3","cess"), bg=GREEN_LT)
    sub_header(47, "4", "Inward supplies from ISD", GREEN_LT)
    data_row(48, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","A","4","igst"), bg=GREEN_LT)
    data_row(49, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","A","4","cgst"), bg=GREEN_LT)
    data_row(50, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","A","4","sgst"), bg=GREEN_LT)
    data_row(51, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","A","4","cess"), bg=GREEN_LT)
    sub_header(52, "5", "All other ITC", GREEN_LT)
    data_row(53, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","A","5","igst"), bg=GREEN_LT)
    data_row(54, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","A","5","cgst"), bg=GREEN_LT)
    data_row(55, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","A","5","sgst"), bg=GREEN_LT)
    data_row(56, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","A","5","cess"), bg=GREEN_LT)
    sub_header(57, "B", "ITC Reversed", RED_LT)
    sub_header(58, "1", "As per rules 38, 42 and 43 of CGST Rules and sub-section (5) of section 17", RED_LT)
    data_row(59, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","B","1","igst"), bg=RED_LT)
    data_row(60, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","B","1","cgst"), bg=RED_LT)
    data_row(61, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","B","1","sgst"), bg=RED_LT)
    data_row(62, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","B","1","cess"), bg=RED_LT)
    sub_header(63, "2", "Others", RED_LT)
    data_row(64, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","B","2","igst"), bg=RED_LT)
    data_row(65, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","B","2","cgst"), bg=RED_LT)
    data_row(66, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","B","2","sgst"), bg=RED_LT)
    data_row(67, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","B","2","cess"), bg=RED_LT)
    sub_header(68, "C", "Net ITC Available (A-B)", TEAL_LT)
    data_row(69, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","C","igst"), bg=TEAL_LT, bold=True)
    data_row(70, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","C","cgst"), bg=TEAL_LT, bold=True)
    data_row(71, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","C","sgst"), bg=TEAL_LT, bold=True)
    data_row(72, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","C","cess"), bg=TEAL_LT, bold=True)
    sub_header(73, "D", "Other Details", GREY_LT)
    sub_header(74, "1", "ITC reclaimed which was reversed under Table 4(B)(2) in earlier tax period", GREY_LT)
    data_row(75, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","D","1","igst"))
    data_row(76, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","D","1","cgst"))
    data_row(77, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","D","1","sgst"))
    data_row(78, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","D","1","cess"))
    sub_header(79, "2", "Ineligible ITC under section 16(4) and ITC restricted due to PoS provisions", GREY_LT)
    data_row(80, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_4","D","2","igst"))
    data_row(81, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_4","D","2","cgst"))
    data_row(82, "(iii)", "State/UT Tax",          lambda m: mon_val(m,"table_4","D","2","sgst"))
    data_row(83, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_4","D","2","cess"))

    # ── Section 5 ─────────────────────────────────────────────────────────────
    section_header(84, "5. Values of exempt, Nil-rated and non-GST inward supplies")
    sub_header(85, "1", "From a supplier under composition scheme, Exempt and Nil rated supply", BLUE_LT)
    data_row(86, "(i)",  "Inter-State supplies",  lambda m: mon_val(m,"table_5","composition_exempt_nil","inter_state"), bg=BLUE_LT)
    data_row(87, "(ii)", "Intra-state supplies",  lambda m: mon_val(m,"table_5","composition_exempt_nil","intra_state"), bg=BLUE_LT)
    sub_header(88, "2", "Non GST Supplies", BLUE_LT)
    data_row(89, "(i)",  "Inter-State supplies",  lambda m: mon_val(m,"table_5","non_gst","inter_state"), bg=BLUE_LT)
    data_row(90, "(ii)", "Intra-state supplies",  lambda m: mon_val(m,"table_5","non_gst","intra_state"), bg=BLUE_LT)

    # ── Section 5.1 ───────────────────────────────────────────────────────────
    section_header(91, "5.1 Interest & late fee payable")
    sub_header(92, "1", "Interest", GOLD_LT)
    data_row(93, "(i)",   "Integrated Tax",        lambda m: mon_val(m,"table_5_1","interest","igst"), bg=GOLD_LT)
    data_row(94, "(ii)",  "Central Tax",           lambda m: mon_val(m,"table_5_1","interest","cgst"), bg=GOLD_LT)
    data_row(95, "(iii)", "State Tax",             lambda m: mon_val(m,"table_5_1","interest","sgst"), bg=GOLD_LT)
    data_row(96, "(iv)",  "Cess",                  lambda m: mon_val(m,"table_5_1","interest","cess"), bg=GOLD_LT)
    sub_header(97, "2", "Late fee", GOLD_LT)
    data_row(98, "(i)",   "Central Tax",           lambda m: mon_val(m,"table_5_1","late_fee","cgst"), bg=GOLD_LT)
    data_row(99, "(ii)",  "State/UT Tax",          lambda m: mon_val(m,"table_5_1","late_fee","sgst"), bg=GOLD_LT)

    # ── Section 6.1 ───────────────────────────────────────────────────────────
    section_header(100, "6.1 Payment of Tax")
    sub_header(101, "6.1", "Tax Payable", NAVY)
    section_header(102, "6.1(A)  Other than Reverse Charge", bg="2E75B6")
    sub_header(103, "1", "IGST", BLUE_LT)
    data_row(104, "(i)",   "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","igst","net_payable_other"), bg=BLUE_LT)
    data_row(105, "(ii)",  "Interest payable",      lambda m: mon_val(m,"table_6_1","tax_payment","igst","interest_payable"),  bg=BLUE_LT)
    sub_header(106, "2", "CGST", BLUE_LT)
    data_row(107, "(i)",   "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","cgst","net_payable_other"), bg=BLUE_LT)
    data_row(108, "(ii)",  "Interest payable",      lambda m: mon_val(m,"table_6_1","tax_payment","cgst","interest_payable"),  bg=BLUE_LT)
    data_row(109, "(iii)", "Late fee payable",      lambda m: mon_val(m,"table_6_1","tax_payment","cgst","late_fee_payable"),  bg=BLUE_LT)
    sub_header(110, "3", "SGST", BLUE_LT)
    data_row(111, "(i)",   "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","sgst","net_payable_other"), bg=BLUE_LT)
    data_row(112, "(ii)",  "Interest payable",      lambda m: mon_val(m,"table_6_1","tax_payment","sgst","interest_payable"),  bg=BLUE_LT)
    data_row(113, "(iii)", "Late fee payable",      lambda m: mon_val(m,"table_6_1","tax_payment","sgst","late_fee_payable"),  bg=BLUE_LT)
    sub_header(114, "4", "Cess", BLUE_LT)
    data_row(115, "(i)",   "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","cess","net_payable_other"), bg=BLUE_LT)
    data_row(116, "(ii)",  "Interest payable",      lambda m: mon_val(m,"table_6_1","tax_payment","cess","interest_payable"),  bg=BLUE_LT)

    # Net Tax Payable (Other than RC) — computed
    section_header(117, "Net Tax Payable (Other than Reverse Charge)", bg=TEAL)
    for roff, (sno, lbl, rows_to_sum) in enumerate([
        ("IGST",  "IGST Net",  [104, 105]),
        ("CGST",  "CGST Net",  [107, 108, 109]),
        ("SGST",  "SGST Net",  [111, 112, 113]),
        ("Cess",  "Cess Net",  [115, 116]),
    ], 118):
        row_num = roff
        ws.cell(row=row_num, column=2, value=sno)
        c3 = ws.cell(row=row_num, column=3, value=lbl)
        style_cell(c3, bold=True, bg=TEAL_LT, fc=NAVY, size=9, h_align="left")
        for col_i in range(4, 16):
            ltr = get_column_letter(col_i)
            refs = "+".join(f"{ltr}{r}" for r in rows_to_sum)
            c = ws.cell(row=row_num, column=col_i, value=f"={refs}")
            style_cell(c, bold=True, bg=TEAL_LT, num_fmt=NUM_FMT)
        tc = ws.cell(row=row_num, column=16)
        tc.value = f"=SUM({get_column_letter(4)}{row_num}:{get_column_letter(15)}{row_num})"
        style_cell(tc, bold=True, bg=TEAL_LT, num_fmt=NUM_FMT)
        ws.row_dimensions[row_num].height = 16

    section_header(122, "6.1(B)  Reverse Charge", bg="2E75B6")
    sub_header(123, "1", "IGST", RED_LT)
    data_row(124, "(i)",  "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","igst","net_payable_rcm"), bg=RED_LT)
    sub_header(125, "2", "CGST", RED_LT)
    data_row(126, "(i)",  "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","cgst","net_payable_rcm"), bg=RED_LT)
    sub_header(127, "3", "SGST", RED_LT)
    data_row(128, "(i)",  "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","sgst","net_payable_rcm"), bg=RED_LT)
    sub_header(129, "4", "Cess", RED_LT)
    data_row(130, "(i)",  "Tax payable",           lambda m: mon_val(m,"table_6_1","tax_payment","cess","net_payable_rcm"), bg=RED_LT)

    # ── Tax paid through ITC ──────────────────────────────────────────────────
    section_header(131, "6.1(A)(i)  Tax Paid through ITC", bg="2E75B6")
    sub_header(132, "IGST", "IGST Payable → paid through ITC", GREEN_LT)
    data_row(133, "",  "IGST ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","igst","itc_igst"), bg=GREEN_LT)
    data_row(134, "",  "CGST ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","igst","itc_cgst"), bg=GREEN_LT)
    data_row(135, "",  "SGST ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","igst","itc_sgst"), bg=GREEN_LT)
    sub_header(136, "CGST", "CGST Payable → paid through ITC", GREEN_LT)
    data_row(137, "",  "IGST ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","cgst","itc_igst"), bg=GREEN_LT)
    data_row(138, "",  "CGST ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","cgst","itc_cgst"), bg=GREEN_LT)
    sub_header(139, "SGST", "SGST Payable → paid through ITC", GREEN_LT)
    data_row(140, "",  "IGST ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","sgst","itc_igst"), bg=GREEN_LT)
    data_row(141, "",  "SGST ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","sgst","itc_sgst"), bg=GREEN_LT)
    sub_header(142, "CESS", "CESS Payable → paid through ITC", GREEN_LT)
    data_row(143, "",  "CESS ITC utilized",       lambda m: mon_val(m,"table_6_1","tax_payment","cess","itc_cess"), bg=GREEN_LT)

    # ── Tax paid through Cash ─────────────────────────────────────────────────
    section_header(144, "6.1(A)(ii)  Tax Paid through Cash (Other than Reverse Charge)", bg="2E75B6")
    data_row(145, "IGST",  "Tax Paid (Cash)",      lambda m: mon_val(m,"table_6_1","tax_payment","igst","cash_other"),    bg=GOLD_LT, bold=True)
    data_row(146, "",      "Interest Paid",        lambda m: mon_val(m,"table_6_1","tax_payment","igst","interest_cash"), bg=GOLD_LT)
    data_row(147, "CGST",  "Tax Paid (Cash)",      lambda m: mon_val(m,"table_6_1","tax_payment","cgst","cash_other"),    bg=GOLD_LT, bold=True)
    data_row(148, "",      "Interest Paid",        lambda m: mon_val(m,"table_6_1","tax_payment","cgst","interest_cash"), bg=GOLD_LT)
    data_row(149, "",      "Late Fees Paid",       lambda m: mon_val(m,"table_6_1","tax_payment","cgst","late_fee_cash"), bg=GOLD_LT)
    data_row(150, "SGST",  "Tax Paid (Cash)",      lambda m: mon_val(m,"table_6_1","tax_payment","sgst","cash_other"),    bg=GOLD_LT, bold=True)
    data_row(151, "",      "Interest Paid",        lambda m: mon_val(m,"table_6_1","tax_payment","sgst","interest_cash"), bg=GOLD_LT)
    data_row(152, "",      "Late Fees Paid",       lambda m: mon_val(m,"table_6_1","tax_payment","sgst","late_fee_cash"), bg=GOLD_LT)
    data_row(153, "CESS",  "Tax Paid (Cash)",      lambda m: mon_val(m,"table_6_1","tax_payment","cess","cash_other"),    bg=GOLD_LT, bold=True)
    data_row(154, "",      "Interest Paid",        lambda m: mon_val(m,"table_6_1","tax_payment","cess","interest_cash"), bg=GOLD_LT)

    section_header(155, "6.1(B)(i)  Tax Paid through Cash (Reverse Charge)", bg="2E75B6")
    data_row(156, "IGST",  "Tax Paid (RC Cash)",   lambda m: mon_val(m,"table_6_1","tax_payment","igst","cash_rcm"),      bg=RED_LT, bold=True)
    data_row(157, "CGST",  "Tax Paid (RC Cash)",   lambda m: mon_val(m,"table_6_1","tax_payment","cgst","cash_rcm"),      bg=RED_LT, bold=True)
    data_row(158, "SGST",  "Tax Paid (RC Cash)",   lambda m: mon_val(m,"table_6_1","tax_payment","sgst","cash_rcm"),      bg=RED_LT, bold=True)
    data_row(159, "CESS",  "Tax Paid (RC Cash)",   lambda m: mon_val(m,"table_6_1","tax_payment","cess","cash_rcm"),      bg=RED_LT, bold=True)

    # ── Cash Ledger Balance ───────────────────────────────────────────────────
    section_header(160, "Cash Ledger Balance (as shown in 6.1 Payment screen)")
    data_row(161, "Tax",      "IGST Cash Bal.",    lambda m: mon_val(m,"table_6_1","cash_ledger","tax","igst"))
    data_row(162, "",         "CGST Cash Bal.",    lambda m: mon_val(m,"table_6_1","cash_ledger","tax","cgst"))
    data_row(163, "",         "SGST Cash Bal.",    lambda m: mon_val(m,"table_6_1","cash_ledger","tax","sgst"))
    data_row(164, "",         "CESS Cash Bal.",    lambda m: mon_val(m,"table_6_1","cash_ledger","tax","cess"))
    data_row(165, "",         "Total Cash Bal.",   lambda m: mon_val(m,"table_6_1","cash_ledger","tax","total"),   bold=True, bg=TEAL_LT)
    data_row(166, "Interest", "CGST Int. Cash",   lambda m: mon_val(m,"table_6_1","cash_ledger","interest","cgst"))
    data_row(167, "",         "SGST Int. Cash",   lambda m: mon_val(m,"table_6_1","cash_ledger","interest","sgst"))
    data_row(168, "Late Fee", "CGST LF Cash",     lambda m: mon_val(m,"table_6_1","cash_ledger","late_fee","cgst"))
    data_row(169, "",         "SGST LF Cash",     lambda m: mon_val(m,"table_6_1","cash_ledger","late_fee","sgst"))

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2: Section 3.2 — State-wise inter-state supplies
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Section 3.2")
    ws2.column_dimensions["A"].width = 1.5
    ws2.column_dimensions["B"].width = 18   # Section type
    ws2.column_dimensions["C"].width = 28   # State name
    for ci in range(4, 4 + 12*2 + 2):      # 12 months × 2 cols + 2 total
        ws2.column_dimensions[get_column_letter(ci)].width = 10

    # Row 1: Title
    ws2.merge_cells(f"B1:{get_column_letter(4+12*2)}1")
    c = ws2["B1"]
    c.value = "Multi Month GSTR-3B Filed Report — Section 3.2 Inter-State Supplies"
    style_cell(c, bold=True, bg=NAVY, fc=WHITE, size=12, h_align="center")
    ws2.row_dimensions[1].height = 22

    # Row 2: Client info
    ws2.merge_cells(f"B2:{get_column_letter(4+12*2)}2")
    c = ws2["B2"]; c.value = f"{gstin}   ·   {dur}"
    style_cell(c, bold=True, bg=TEAL, fc=WHITE, size=10, h_align="center")
    ws2.row_dimensions[2].height = 16

    # Row 3: Month headers (each month has 2 cols: TV, IGST)
    ws2.cell(row=3, column=2, value="Section Type")
    style_cell(ws2.cell(row=3, column=2), bold=True, bg=NAVY, fc=WHITE)
    ws2.cell(row=3, column=3, value="Place of Supply (State/UT)")
    style_cell(ws2.cell(row=3, column=3), bold=True, bg=NAVY, fc=WHITE)
    col3 = 4
    for m in MONTHS_ORDER:
        yr = fy_start_y if m >= 4 else fy_end_y
        mlabel = f"{MON_ABBR[m]} '{str(yr)[2:]}"
        ws2.merge_cells(start_row=3, start_column=col3, end_row=3, end_column=col3+1)
        c = ws2.cell(row=3, column=col3, value=mlabel)
        style_cell(c, bold=True, bg=NAVY, fc=WHITE)
        # Sub-headers row 4
        ws2.cell(row=4, column=col3,   value="Taxable Value").fill = fill("2E75B6")
        ws2.cell(row=4, column=col3+1, value="IGST").fill           = fill("2E75B6")
        for cc in [col3, col3+1]:
            style_cell(ws2.cell(row=4, column=cc), bold=True, bg="2E75B6", fc=WHITE, size=8)
        col3 += 2

    # Total header
    ws2.merge_cells(start_row=3, start_column=col3, end_row=3, end_column=col3+1)
    c = ws2.cell(row=3, column=col3, value="Total"); style_cell(c, bold=True, bg=TEAL, fc=WHITE)
    ws2.cell(row=4, column=col3,   value="Taxable Value"); style_cell(ws2.cell(row=4, column=col3),   bold=True, bg=TEAL, fc=WHITE, size=8)
    ws2.cell(row=4, column=col3+1, value="IGST");          style_cell(ws2.cell(row=4, column=col3+1), bold=True, bg=TEAL, fc=WHITE, size=8)
    ws2.row_dimensions[3].height = 22
    ws2.row_dimensions[4].height = 22

    # Collect all state rows across all months
    # Structure: {(section_type, state_code): {mon_num: (tv, igst)}}
    state_data = {}
    SEC_TYPES   = [("unregistered","Unregistered"), ("composition","Composition"), ("uin","UIN holders")]
    for mon_num, d in data.items():
        t32 = d.get("table_3_2", {})
        for sec_key, sec_label in SEC_TYPES:
            rows = t32.get(sec_key, [])
            for row in rows:
                sc = row.get("state","").strip()
                if not sc or sc.lower() == "select":
                    continue
                key = (sec_label, sc)
                if key not in state_data:
                    state_data[key] = {}
                state_data[key][mon_num] = (row.get("taxable_value",0), row.get("igst",0))

    # Sort: by section type then state
    order = {"Unregistered":0, "Composition":1, "UIN holders":2}
    sorted_states = sorted(state_data.keys(), key=lambda k: (order.get(k[0],9), k[1]))

    SECTION_COLORS = {"Unregistered": BLUE_LT, "Composition": GREEN_LT, "UIN holders": PURPLE_LT}
    current_row3  = 5
    current_sec   = None
    total_col     = 4 + 12*2

    for (sec_label, state_name) in sorted_states:
        # Print section divider
        if sec_label != current_sec:
            current_sec = sec_label
            ws2.merge_cells(f"B{current_row3}:{get_column_letter(total_col+1)}{current_row3}")
            c = ws2.cell(row=current_row3, column=2, value=f"3.2 — Supplies to {sec_label}")
            style_cell(c, bold=True, bg=NAVY, fc=WHITE, size=9, h_align="left")
            ws2.row_dimensions[current_row3].height = 16
            current_row3 += 1

        bg_c = SECTION_COLORS.get(sec_label, GREY_LT)
        ws2.cell(row=current_row3, column=2, value=sec_label)
        style_cell(ws2.cell(row=current_row3, column=2), bg=bg_c, h_align="center", size=8)
        ws2.cell(row=current_row3, column=3, value=state_name)
        style_cell(ws2.cell(row=current_row3, column=3), bg=bg_c, h_align="left")

        col_c = 4
        for m in MONTHS_ORDER:
            tv, igst = state_data[(sec_label, state_name)].get(m, (0, 0))
            c_tv = ws2.cell(row=current_row3, column=col_c,   value=tv   if tv   else 0)
            c_ig = ws2.cell(row=current_row3, column=col_c+1, value=igst if igst else 0)
            style_cell(c_tv, bg=bg_c, num_fmt=NUM_FMT)
            style_cell(c_ig, bg=bg_c, num_fmt=NUM_FMT)
            col_c += 2

        # Total columns
        tv_cols = "+".join(get_column_letter(4 + i*2) for i in range(12))
        ig_cols = "+".join(get_column_letter(5 + i*2) for i in range(12))
        c_ttv = ws2.cell(row=current_row3, column=total_col,   value=f"={tv_cols}")
        c_tig = ws2.cell(row=current_row3, column=total_col+1, value=f"={ig_cols}")
        style_cell(c_ttv, bold=True, bg=TEAL_LT, num_fmt=NUM_FMT)
        style_cell(c_tig, bold=True, bg=TEAL_LT, num_fmt=NUM_FMT)
        ws2.row_dimensions[current_row3].height = 15
        current_row3 += 1

    if not sorted_states:
        ws2.merge_cells(f"B5:{get_column_letter(total_col+1)}5")
        c = ws2.cell(row=5, column=2, value="No inter-state supply data found for this period")
        style_cell(c, bg=GOLD_LT, fc=GREY_H, size=9, h_align="center")

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes  = "D4"
    ws2.freeze_panes = "D5"

    # ── Save ──────────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, None


# ── Flask route ────────────────────────────────────────────────────────────────
@app.route("/g3b/export_excel", methods=["GET","POST"])
def g3b_export_excel():
    """Convert all scraped GSTR-3B JSON files → Excel workbook, return for download."""
    try:
        if request.method == "POST":
            body  = request.get_json(force=True) or {}
            gstin = (body.get("gstin","") or "").strip().upper()
            fy    = (body.get("fy",   "") or "").strip()
        else:
            gstin = request.args.get("gstin","").strip().upper()
            fy    = request.args.get("fy","").strip()
        if not gstin or not fy:
            return jsonify({"error":"gstin and fy are required"}), 400
        try:
            import openpyxl  # noqa
        except ImportError:
            return jsonify({"error":"openpyxl is not installed. Run: pip install openpyxl"}), 500
        buf, err = g3b_json_to_excel(gstin, fy)
        if err:
            return jsonify({"error": err}), 404
        fname = f"GSTR3B_{gstin}_{fy.replace('-','_')}.xlsx"
        save_dir = os.path.join(G3B_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
        return save_excel_and_respond(buf, save_dir, fname, log_fn=lambda m, lvl="info": g3b_log(m))
    except Exception as e:
        import traceback
        g3b_log(f"export_excel error: {e} | {traceback.format_exc()}", "error")
        return jsonify({"error": str(e)}), 500


# ── Flask routes ───────────────────────────────────────────────────────────────

@app.route("/g3b/start", methods=["POST"])
def g3b_start():
    data     = request.json or {}
    gstin    = (data.get("gstin")    or "").strip().upper()
    fy       = (data.get("fy")       or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    force    = bool(data.get("force", False))
    specific_month = (data.get("specific_month") or "").strip()
    if not all([gstin, fy, username, password]):
        return jsonify({"error": "gstin, fy, username, password required"}), 400

    # This route previously didn't even use a lock for the busy check —
    # a plain unsynchronized dict read, with no state claim before
    # spawning the thread at all (it relied entirely on g3b_worker
    # itself setting status="running" moments later). That's the least
    # protected version of the same race found and fixed everywhere
    # else in this file — fixed the same way here for consistency.
    busy = ("running", "login", "waiting_captcha", "waiting_otp")
    with g3b_lock:
        if g3b_state.get("status") in busy:
            return jsonify({"error": "already running"}), 409
        g3b_state["status"] = "running"

    threading.Thread(
        target=g3b_worker,
        args=(gstin, fy, username, password, G3B_DOWNLOAD_DIR, force, specific_month),
        daemon=True
    ).start()
    return jsonify({"ok": True, "specific_month": specific_month or None})


@app.route("/g3b/state")
def g3b_state_route():
    with g3b_lock:
        return jsonify({
            "status":        g3b_state["status"],
            "log":           g3b_state["log"][-100:],
            "captcha_image": g3b_state.get("captcha_image"),
            "files":         g3b_state.get("files",       []),
            "fresh_files":   g3b_state.get("fresh_files", []),
            "error":         g3b_state.get("error"),
        })


@app.route("/g3b/submit_captcha", methods=["POST"])
def g3b_submit_captcha():
    val = ((request.json or {}).get("captcha") or "").strip()
    with g3b_lock:
        g3b_state["captcha_answer"] = val
    return jsonify({"ok": True})


@app.route("/g3b/submit_otp", methods=["POST"])
def g3b_submit_otp():
    val = ((request.json or {}).get("otp") or "").strip()
    with g3b_lock:
        g3b_state["otp_answer"] = val
    return jsonify({"ok": True})


@app.route("/g3b/reset", methods=["POST"])
def g3b_reset():
    g3b_set({
        "status": "idle", "log": [], "captcha_image": None,
        "error":  None,   "captcha_answer": None, "otp_answer": None,
        "files":  [],     "fresh_files": [],
    })
    return jsonify({"ok": True})


@app.route("/g3b/files")
def g3b_files_route():
    gstin = (request.args.get("gstin") or "").upper()
    fy    = (request.args.get("fy")    or "")
    if not gstin or not fy:
        return jsonify({"files": []})
    d = os.path.join(G3B_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    if not os.path.isdir(d):
        return jsonify({"files": [], "dir": d})
    files = sorted([f for f in os.listdir(d)
                    if f.startswith("GSTR3B_") and f.endswith(".json")])
    out = []
    for fn in files:
        fp = os.path.join(d, fn)
        out.append({"filename": fn,
                    "gstin":    gstin,
                    "fy":       fy,
                    "size_kb":  round(os.path.getsize(fp)/1024, 1)})
    return jsonify({"files": out, "dir": d})


@app.route("/g3b/file/<gstin>/<fy_dir>/<fname>")
def g3b_file_route(gstin, fy_dir, fname):
    fpath = os.path.join(G3B_DOWNLOAD_DIR, gstin, fy_dir, fname)
    if not os.path.exists(fpath):
        return jsonify({"error": "not found"}), 404
    with open(fpath) as f:
        return jsonify(json.load(f))


# ══════════════════════════════════════════════════════════════════════
#  BROWSER CHANNEL — set to "chrome" to use installed Chrome instead of Chromium
#  Options: "chromium" (default), "chrome", "msedge"
# ══════════════════════════════════════════════════════════════════════
BROWSER_CHANNEL = "chromium"


def _get_browser_kwargs(headless=False, slow_mo=0, no_viewport=False, accept_downloads=False):
    """Return launch kwargs; adds channel= when BROWSER_CHANNEL != 'chromium'."""
    kw = {"headless": headless}
    if slow_mo:
        kw["slow_mo"] = slow_mo
    if no_viewport:
        kw["no_viewport"] = True
    if accept_downloads:
        kw["accept_downloads"] = True
    if BROWSER_CHANNEL != "chromium":
        kw["channel"] = BROWSER_CHANNEL
    else:
        kw["args"] = [
            "--start-maximized",
            "--no-sandbox",
            "--disable-blink-features=AutomationControlled",
            "--no-first-run",
            "--no-default-browser-check",
        ]
    return kw


# ══════════════════════════════════════════════════════════════════════
#  PORTAL RPA LOGIN — GST · TRACES · Income Tax (TAN/PAN)
# ──────────────────────────────────────────────────────────────────────
#  Architecture:
#    • Uses Playwright's launch_persistent_context (headless=False) for
#      DIRECT browser control — no subprocess+CDP reconnect instability.
#    • Playwright + browser objects stored in _portal_sessions dict so
#      they are NOT garbage-collected → browser stays alive after thread
#      exits until the user closes it manually or gst_rpa.py exits.
#    • Fills via press_sequentially() (char-by-char keyboard events) so
#      AngularJS/Angular/JSF form-change detectors all fire correctly.
# ══════════════════════════════════════════════════════════════════════
import uuid as _uuid
import tempfile as _tempfile

_VALID_PORTALS = {"gst", "traces", "itax_tan", "itax_pan"}

_PORTAL_URLS = {
    "gst":      "https://services.gst.gov.in/services/login",
    "traces":   "https://traces.tdscpc.gov.in/auth/login/loginScreen",
    "itax_tan": "https://eportal.incometax.gov.in/iec/foservices/#/login",
    "itax_pan": "https://eportal.incometax.gov.in/iec/foservices/#/login",
}

_PORTAL_LABELS = {
    "gst":      "GST Portal",
    "traces":   "TRACES",
    "itax_tan": "Income Tax (TAN Login)",
    "itax_pan": "Income Tax (PAN Login)",
}

# ── Per-portal login architecture ─────────────────────────────────────────────
#
# GST  (services.gst.gov.in/services/login)
#   Tech: AngularJS SPA
#   Fields: Username (#user_name) + Password (#user_pass) + Captcha (user fills)
#   Confirmed selectors from working SELECTORS constant in this file (line ~97):
#     username → #user_name  /  input[name='user_name']
#     password → #user_pass  /  input[name='user_pass']
#
# TRACES  (traces.tdscpc.gov.in/auth/login/loginScreen)
#   Tech: Angular SPA (complete redesign, live 2026)
#   Tabs: Deductor | Taxpayer | PAO  → click "Deductor" first
#   Fields (Deductor tab):
#     1. User ID (placeholder "Enter TAN") — TAN IS the User ID now
#     2. "Do you want to login as a sub-user?" Yes/No radio → select No
#     3. Password
#     4. Verification Code (captcha — user fills manually)
#   No separate TAN Number field anymore.
#   Angular Material inputs: use get_by_placeholder() + .fill() for reliable fill.
#
# Income Tax ePortal  (eportal.incometax.gov.in/iec/foservices/#/login)
#   Tech: Angular 12+ SPA with Angular Material (mat-form-field)
#   Login page has TWO TABS:
#     Tab 1 "Individual"  → for PAN / Aadhaar login (default active tab)
#     Tab 2 "Others"      → for TAN, ERI, ITDREIN, etc.
#   Fields (same in both tabs):
#     User ID  → input inside mat-form-field (formControlName="userId")
#     Password → input[type=password] (formControlName="password")
#   For PAN login: page loads with Individual tab active → fill directly
#   For TAN login: must click "Others" tab first, then fill
#   Angular Material inputs need: click → clear → press_sequentially (char-by-char)
#   This fires Angular's ControlValueAccessor which updates the reactive form.
# ─────────────────────────────────────────────────────────────────────────────

_PORTAL_CFG = {
    "gst": {
        "tab_labels": None,          # no tab to click
        "tan_sels":   [],            # no separate TAN field
        "uid_sels": [
            "#user_name",
            "input[name='user_name']",
            "input[id='user_name']",
            "input#username",
            "input[name='username']",
            "input[placeholder*='sername' i]",
            "input[type='text']",
        ],
        "pwd_sels": [
            "#user_pass",
            "input[name='user_pass']",
            "input[id='user_pass']",
            "input[type='password']",
        ],
        "banner": "User ID & Password filled ✓  —  solve captcha then click LOGIN",
    },
    "traces": {
        # New Angular portal (2026) — click Deductor tab first
        "tab_labels": ["Deductor"],
        # TAN IS the User ID now — no separate TAN Number field
        "tan_sels": [],
        # Angular Material inputs — get_by_placeholder() used directly in login flow
        "uid_sels": [
            "input[placeholder='Enter TAN']",
            "input[placeholder*='Enter TAN' i]",
            "input[placeholder*='TAN' i]",
            "input[formcontrolname='userId']",
            "input[formcontrolname='username']",
            "mat-form-field:first-of-type input",
            "input[type='text']:first-of-type",
        ],
        "pwd_sels": [
            "input[type='password']",
            "input[placeholder='Enter Password']",
            "input[placeholder*='Password' i]",
            "input[formcontrolname='password']",
        ],
        "sub_user_no_sels": [
            "input[type='radio'][value='N']",
            "input[type='radio'][value='No']",
            "label:has-text('No') input[type='radio']",
            "mat-radio-button:has-text('No') input",
            "mat-radio-button:has-text('No')",
        ],
        "banner": ("TAN (User ID) + Password filled ✓  |  "
                   "Sub-user = No  |  Solve the Verification Code then Login"),
    },
    "itax_tan": {
        "tab_labels": None,    # same field as PAN — no tab switch needed
        "tan_sels":   [],           # TAN is the User ID on IT portal
        "uid_sels": [
            "input[formcontrolname='userId']",
            "input[id='userId']",
            "input[name='userId']",
            "input[placeholder*='user id' i]",
            "input[placeholder*='TAN' i]",
            ".mat-form-field:first-of-type input",
            "input[type='text']",
        ],
        "pwd_sels": [
            "input[formcontrolname='password']",
            "input[id='password']",
            "input[type='password']",
        ],
        "banner": "TAN login — Step 1: User ID → Continue → Step 2: Checkbox + Password → Continue",
    },
    "itax_pan": {
        "tab_labels": None,         # Individual tab is default — no click needed
        "tan_sels":   [],
        "uid_sels": [
            "input[formcontrolname='userId']",
            "input[id='userId']",
            "input[name='userId']",
            "input[placeholder*='user id' i]",
            "input[placeholder*='PAN' i]",
            ".mat-form-field:first-of-type input",
            "input[type='text']",
        ],
        "pwd_sels": [
            "input[formcontrolname='password']",
            "input[id='password']",
            "input[type='password']",
        ],
        "banner": "PAN login — Step 1: User ID → Continue → Step 2: Checkbox + Password → Continue",
    },
}

# Keeps (playwright_instance, browser_context) alive so browser doesn't close
_portal_sessions: dict = {}
_portal_sessions_lock = threading.Lock()


def _portal_click_tab(page, tab_labels):
    """Click the first visible tab whose text matches any label."""
    for label in (tab_labels or []):
        for sel in [
            f"button:has-text('{label}')",
            f"[role='tab']:has-text('{label}')",
            f"a:has-text('{label}')",
            f"li:has-text('{label}')",
            f"span:has-text('{label}')",
            f"div[role='tab']:has-text('{label}')",
        ]:
            try:
                el = page.locator(sel).first
                el.wait_for(state="visible", timeout=4000)
                el.click()
                log.info(f"  [RPA] ✓ Tab clicked: '{label}'")
                time.sleep(1.5)   # wait for Angular form to render after tab switch
                try:
                    page.wait_for_selector(
                        "input[type='text'], input[type='password']",
                        state="visible", timeout=5000)
                except Exception:
                    pass
                return True
            except Exception:
                continue
    log.warning(f"  [RPA] Tab not found for: {tab_labels}")
    return False


def _portal_fill_field(page, selectors, value, label):
    """
    Fill an input reliably for Angular/AngularJS/JSF portals.

    Strategy:
      1. For each selector, try to locate the element.
      2. Click it (focus), triple-click (select all existing text), then
         press_sequentially() — this fires real keyboard events that all
         framework event listeners (Angular zone, AngularJS $watch, JSF) pick up.
      3. Also dispatch input/change/blur via JS as belt-and-suspenders.

    Returns True on first successful fill.
    """
    js_val = value.replace("\\", "\\\\").replace("'", "\\'")
    for sel in selectors:
        try:
            el = page.locator(sel).first
            el.wait_for(state="visible", timeout=2000)   # field confirmed visible by caller
            el.scroll_into_view_if_needed()

            # Focus + select any existing text
            el.click()
            time.sleep(0.05)
            el.select_text()

            # Type character-by-character at 15ms — triggers all Angular listeners,
            # fast enough that normal portals don't miss events
            el.press_sequentially(value, delay=15)

            # Belt-and-suspenders: fire JS events for frameworks that need them
            try:
                page.evaluate(f"""
                    () => {{
                        const q = (s) => document.querySelector(s);
                        const el = q('{sel.replace("'", "\\'")}');
                        if (!el) return;
                        ['input','change','blur'].forEach(ev =>
                            el.dispatchEvent(new Event(ev, {{bubbles:true}}))
                        );
                    }}
                """)
            except Exception:
                pass

            # Verify value actually got in
            actual = el.input_value()
            if actual == value:
                log.info(f"  [RPA] ✓ {label} filled via '{sel}'")
                return True
            elif actual:
                # Something typed but value mismatch — still better than nothing
                log.info(f"  [RPA] ✓ {label} partially filled via '{sel}' "
                         f"(got {len(actual)} chars)")
                return True
            else:
                log.warning(f"  [RPA]   '{sel}' — typed but field reads empty")

        except Exception as ex:
            log.info(f"  [RPA]   selector '{sel}' — {ex}")
            continue

    log.error(f"  [RPA] ✗ Could not fill {label} — all selectors exhausted")
    return False


def _portal_banner(page, label, filled, note):
    """Inject a status banner into the open browser page."""
    ok    = bool(filled)
    color = "#14532d" if ok else "#7c2d12"
    emoji = "✅" if ok else "⚠️"
    txt   = f"Filled: {', '.join(filled)}" if ok else "Fields not found — enter manually"
    note_esc = note.replace("'", "\\'")
    txt_esc  = txt.replace("'", "\\'")
    try:
        page.evaluate(f"""
            () => {{
                const old = document.getElementById('_cl_rpa_b');
                if (old) old.remove();
                const d  = document.createElement('div');
                d.id     = '_cl_rpa_b';
                d.style.cssText = [
                    'position:fixed','top:0','left:0','right:0',
                    'padding:0 14px',
                    'background:{color}',
                    'color:#fff',
                    'font:bold 13px/44px Arial,sans-serif',
                    'text-align:center',
                    'z-index:2147483647',
                    'box-shadow:0 3px 12px rgba(0,0,0,.5)'
                ].join(';');
                d.innerHTML = (
                    '{emoji} <b>ClientLedger RPA \u2014 {label}</b>'
                  + ' &nbsp;\u2502&nbsp; {txt_esc}'
                  + ' &nbsp;\u2502&nbsp; <span style="font-weight:normal">{note_esc}</span>'
                );
                document.body.insertBefore(d, document.body.firstChild);
                setTimeout(() => {{ if(d.parentNode) d.remove(); }}, 30000);
            }}
        """)
    except Exception as be:
        log.warning(f"  [RPA] Banner error: {be}")


def _itax_login_two_step(page, cfg, label, username, password):
    """
    Income Tax e-Portal 2-step login flow:

    STEP 1 — User ID page  (eportal.incometax.gov.in/iec/foservices/#/login)
      ├─ Fill User ID (PAN or TAN — same field, no tab switch needed)
      └─ Click "Continue" → navigates to /login/password

    STEP 2 — Password page  (#/login/password)
      ├─ Tick the "Secure Access Message" checkbox  (mandatory)
      ├─ Fill Password
      └─ Click "Continue" → logs in

    Returns list of filled field labels.
    """
    filled = []
    is_itax = True  # always True when this function is called

    # ── STEP 1: User ID ─────────────────────────────────────────────────────
    log.info("[RPA][IT] Step 1 — filling User ID")

    # Tab click: only if configured (itax_pan/itax_tan both use None now)
    if cfg["tab_labels"]:
        _portal_click_tab(page, cfg["tab_labels"])

    if username:
        if _portal_fill_field(page, cfg["uid_sels"], username, "User ID"):
            filled.append("User ID")
        time.sleep(0.1)

    # Click Continue to proceed to password page
    log.info("[RPA][IT] Clicking Continue (Step 1)...")
    continue_clicked = False
    for sel in [
        "button:has-text('Continue')",
        "button[type='submit']:has-text('Continue')",
        ".btn-primary:has-text('Continue')",
        "button.primary:has-text('Continue')",
        "input[type='submit'][value*='Continue' i]",
    ]:
        try:
            btn = page.locator(sel).first
            btn.wait_for(state="visible", timeout=500)  # fast-fail; first sel wins normally
            btn.click()
            log.info(f"  [RPA][IT] ✓ Continue (Step 1) clicked via '{sel}'")
            continue_clicked = True
            break
        except Exception:
            continue

    if not continue_clicked:
        # Fallback: find any visible button with text "Continue"
        try:
            page.evaluate("""
                () => {
                    const btns = Array.from(document.querySelectorAll('button'));
                    const b = btns.find(b => b.textContent.trim() === 'Continue'
                                         && b.offsetParent !== null);
                    if (b) b.click();
                }
            """)
            log.info("  [RPA][IT] Continue (Step 1) clicked via JS fallback")
            continue_clicked = True
        except Exception as e:
            log.warning(f"  [RPA][IT] Could not click Continue Step 1: {e}")

    if continue_clicked:
        # Wait for password page — watch for password field (DOM fires before URL updates)
        try:
            page.wait_for_selector(
                "input[type='password'], input[formcontrolname='password']",
                state="visible", timeout=8000)
            log.info("[RPA][IT] ✓ Password page ready")
        except Exception:
            # Fallback: give Angular a moment to finish routing
            time.sleep(0.5)

    # ── STEP 2: Password page ────────────────────────────────────────────────
    log.info("[RPA][IT] Step 2 — password page")

    # Password field is already confirmed visible by the wait above
    # Tick the "Secure Access Message" checkbox (mandatory — Continue stays
    # greyed out until this is checked)
    log.info("[RPA][IT] Ticking Secure Access Message checkbox...")
    checkbox_ticked = False
    for sel in [
        "input[type='checkbox']",
        "mat-checkbox input",
        "input[formcontrolname*='confirm' i]",
        "input[formcontrolname*='secure' i]",
        "input[formcontrolname*='check' i]",
        ".mat-checkbox-input",
        "[class*='checkbox'] input",
    ]:
        try:
            chk = page.locator(sel).first
            chk.wait_for(state="visible", timeout=300)  # fast-fail per selector
            if not chk.is_checked():
                chk.check()
                time.sleep(0.1)
            if chk.is_checked():
                log.info(f"  [RPA][IT] ✓ Checkbox ticked via '{sel}'")
                checkbox_ticked = True
                break
        except Exception:
            continue

    if not checkbox_ticked:
        # JS fallback — click the first visible checkbox
        try:
            page.evaluate("""
                () => {
                    const cb = document.querySelector('input[type=checkbox]');
                    if (cb && !cb.checked) {
                        cb.click();
                        cb.dispatchEvent(new Event('change', {bubbles:true}));
                    }
                }
            """)
            log.info("  [RPA][IT] Checkbox ticked via JS fallback")
            checkbox_ticked = True
        except Exception as e:
            log.warning(f"  [RPA][IT] Checkbox not found: {e}")

    # Fill Password
    if password:
        if _portal_fill_field(page, cfg["pwd_sels"], password, "Password"):
            filled.append("Password")
        time.sleep(0.1)

    # Click Continue (Step 2) to submit login
    log.info("[RPA][IT] Clicking Continue (Step 2)...")
    continue2_clicked = False
    for sel in [
        "button:has-text('Continue')",
        "button[type='submit']:has-text('Continue')",
        ".btn-primary:has-text('Continue')",
        "button.primary:has-text('Continue')",
        "input[type='submit'][value*='Continue' i]",
    ]:
        try:
            btn = page.locator(sel).first
            btn.wait_for(state="visible", timeout=500)  # fast-fail; first sel wins normally
            # Make sure it's enabled (not greyed out)
            is_disabled = btn.get_attribute("disabled") is not None
            if not is_disabled:
                btn.click()
                log.info(f"  [RPA][IT] ✓ Continue (Step 2) clicked via '{sel}'")
                continue2_clicked = True
                break
        except Exception:
            continue

    if not continue2_clicked:
        try:
            page.evaluate("""
                () => {
                    const btns = Array.from(document.querySelectorAll('button'));
                    const b = btns.find(b =>
                        b.textContent.trim() === 'Continue'
                        && b.offsetParent !== null
                        && !b.disabled);
                    if (b) b.click();
                }
            """)
            log.info("  [RPA][IT] Continue (Step 2) clicked via JS fallback")
            continue2_clicked = True
        except Exception as e:
            log.warning(f"  [RPA][IT] Could not click Continue Step 2: {e}")

    if checkbox_ticked:
        filled.append("Checkbox ✓")

    return filled


def _portal_login_thread(portal, username, password, tan, session_id):
    """
    Daemon thread — opens a visible browser directly via Playwright
    (launch_persistent_context, headless=False), fills credentials,
    shows banner, then exits WITHOUT closing the browser.

    The (playwright, context) objects are stored in _portal_sessions so
    Python does not garbage-collect them → browser stays alive.

    IT Portal (itax_pan / itax_tan) uses _itax_login_two_step() which
    handles the 2-page flow:
      Page 1: Enter User ID → click Continue
      Page 2: Tick checkbox + Enter Password → click Continue
    """
    cfg   = _PORTAL_CFG[portal]
    label = _PORTAL_LABELS[portal]
    url   = _PORTAL_URLS[portal]
    profile_dir = os.path.join(_tempfile.gettempdir(),
                               f"cl_rpa_{portal}_{session_id}")
    os.makedirs(profile_dir, exist_ok=True)

    log.info(f"[RPA] ══ {label} ══  session={session_id}")

    try:
        pw = sync_playwright().start()
    except Exception as e:
        log.error(f"[RPA] Playwright start failed: {e}"); return

    try:
        context = pw.chromium.launch_persistent_context(
            user_data_dir=profile_dir,
            **_get_browser_kwargs(no_viewport=True),
        )
    except Exception as e:
        log.error(f"[RPA] launch_persistent_context failed: {e}")
        try: pw.stop()
        except Exception: pass
        return

    with _portal_sessions_lock:
        _portal_sessions[session_id] = (pw, context)
    log.info(f"[RPA] Browser launched (session stored, will stay alive)")

    try:
        page = context.pages[0] if context.pages else context.new_page()
    except Exception as e:
        log.error(f"[RPA] Cannot get page: {e}"); return

    try:
        context.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
    except Exception: pass

    log.info(f"[RPA] goto {url}")
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)
    except Exception as nav_e:
        log.warning(f"[RPA] goto timeout ({nav_e}) — continuing")

    # For IT portal: wait for User ID field to appear (much faster than networkidle)
    # For other portals: fall back to a short domcontentloaded wait
    if portal in ("itax_pan", "itax_tan"):
        try:
            page.wait_for_selector(
                "input[formcontrolname='userId'], input[id='userId'], input[type='text']",
                state="visible", timeout=8000)
            log.info("[RPA] User ID field visible — starting fill")
        except Exception:
            log.info("[RPA] User ID selector timeout — proceeding anyway")
            time.sleep(0.3)
    else:
        try:
            page.wait_for_load_state("domcontentloaded", timeout=6000)
        except Exception:
            time.sleep(0.3)

    # ── Portal-specific login flow ──────────────────────────────────────────
    filled = []

    if portal in ("itax_pan", "itax_tan"):
        # IT Portal: 2-step flow (User ID → Continue → Checkbox + Password → Continue)
        filled = _itax_login_two_step(page, cfg, label, username, password)

    else:
        # GST / TRACES: single-page credential fill
        if cfg["tab_labels"]:
            _portal_click_tab(page, cfg["tab_labels"])

        try:
            inputs = page.evaluate("""
                () => Array.from(document.querySelectorAll('input'))
                           .filter(e => e.offsetParent !== null)
                           .map(e => ({id:e.id,name:e.name,type:e.type,
                                       ph:e.placeholder,
                                       fc:e.getAttribute('formcontrolname')}))
            """)
            log.info(f"[RPA] Visible inputs: {inputs}")
        except Exception: pass

        if cfg["tan_sels"] and tan:
            if _portal_fill_field(page, cfg["tan_sels"], tan, "TAN of Deductor"):
                filled.append("TAN of Deductor")
            time.sleep(0.3)

        if portal == "traces":
            # ── TRACES Angular Material: Playwright semantic locators ──────────
            # Angular Material FormControl ignores JS prototype setter.
            # get_by_placeholder() + .fill() fires the correct Angular InputEvent.
            log.info("[RPA] TRACES: Angular Material fill via get_by_placeholder()")
            time.sleep(0.5)

            # User ID = TAN
            if username:
                _uid_ok = False
                for ph in ["Enter TAN", "Enter User ID", "User ID", "TAN"]:
                    try:
                        loc = page.get_by_placeholder(ph)
                        loc.wait_for(state="visible", timeout=3000)
                        loc.click(); time.sleep(0.2)
                        loc.fill(username)
                        time.sleep(0.3)
                        log.info(f"[RPA] ✓ User ID filled via placeholder '{ph}'")
                        filled.append("User ID"); _uid_ok = True; break
                    except Exception: continue
                if not _uid_ok:
                    try:
                        loc = page.locator(
                            "input:not([type='password']):not([type='hidden'])"
                            ":not([type='radio']):not([type='checkbox'])").first
                        loc.wait_for(state="visible", timeout=3000)
                        loc.click(); time.sleep(0.2); loc.triple_click()
                        page.keyboard.type(username, delay=30)
                        filled.append("User ID"); log.info("[RPA] ✓ User ID via keyboard")
                    except Exception as _e:
                        log.warning(f"[RPA] User ID fill failed: {_e}")

            # Sub-user = No radio
            time.sleep(0.2)
            for sel in cfg.get("sub_user_no_sels", []):
                try:
                    r = page.locator(sel).first
                    r.wait_for(state="visible", timeout=2000)
                    r.click()
                    log.info("[RPA] ✓ Sub-user = No selected")
                    filled.append("Sub-user: No"); break
                except Exception: continue
            time.sleep(0.3)

            # Password
            if password:
                _pwd_ok = False
                for ph in ["Enter Password", "Password", "Enter your password"]:
                    try:
                        loc = page.get_by_placeholder(ph)
                        loc.wait_for(state="visible", timeout=3000)
                        loc.click(); time.sleep(0.2)
                        loc.fill(password)
                        time.sleep(0.3)
                        log.info(f"[RPA] ✓ Password filled via placeholder '{ph}'")
                        filled.append("Password"); _pwd_ok = True; break
                    except Exception: continue
                if not _pwd_ok:
                    try:
                        loc = page.locator("input[type='password']").first
                        loc.wait_for(state="visible", timeout=3000)
                        loc.click(); time.sleep(0.2); loc.triple_click()
                        page.keyboard.type(password, delay=30)
                        filled.append("Password"); log.info("[RPA] ✓ Password via keyboard")
                    except Exception as _e:
                        log.warning(f"[RPA] Password fill failed: {_e}")

        else:
            # GST / IT Portal: generic CSS selector fill
            if username:
                if _portal_fill_field(page, cfg["uid_sels"], username, "User ID"):
                    filled.append("User ID")
                time.sleep(0.3)

            if password:
                if _portal_fill_field(page, cfg["pwd_sels"], password, "Password"):
                    filled.append("Password")
                time.sleep(0.2)

    # ── Status banner ─────────────────────────────────────────────────────
    # Update banner message for IT portal
    banner_note = cfg["banner"]
    if portal in ("itax_pan", "itax_tan"):
        if "Password" in filled and "Checkbox ✓" in filled:
            banner_note = "All steps completed ✓  —  OTP / 2FA may be required next"
        elif "User ID" in filled:
            banner_note = "User ID entered, navigating to password page..."

    _portal_banner(page, label, filled, banner_note)

    log.info(f"[RPA] ══ Done: {label}  filled={filled}  "
             f"browser stays open (session_id={session_id}) ══")


@app.route("/portal_login", methods=["POST"])
def portal_login():
    """
    POST { portal, username, password, tan }
    Opens a visible browser and fills credentials automatically.
    Returns immediately — browser opens in background thread.
    """
    data     = request.get_json(force=True) or {}
    portal   = (data.get("portal")   or "").strip().lower()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    tan      = (data.get("tan")      or "").strip().upper()

    if portal not in _VALID_PORTALS:
        return jsonify({"error": f"Unknown portal '{portal}'. "
                        f"Valid: {sorted(_VALID_PORTALS)}"}), 400
    if not username:
        return jsonify({"error": "username is required"}), 400

    session_id = _uuid.uuid4().hex[:10]
    threading.Thread(
        target=_portal_login_thread,
        args=(portal, username, password, tan, session_id),
        daemon=True,
    ).start()

    return jsonify({
        "ok":         True,
        "portal":     portal,
        "label":      _PORTAL_LABELS.get(portal, portal),
        "session_id": session_id,
        "message":    (f"Opening {_PORTAL_LABELS.get(portal, portal)}. "
                       f"Credentials will be auto-filled. "
                       f"Solve captcha to complete login."),
    })




# ══════════════════════════════════════════════════════════════════════════════
# GSTIN DIRECTORY — Supplier Trade Name & Legal Name Repository
# Phase 1 : mine trdnm from all downloaded GSTR-2A + GSTR-2B JSON files
# Phase 2 : Playwright enrichment via GSTN Search Taxpayer portal (services.gst.gov.in)
# ══════════════════════════════════════════════════════════════════════════════

# ── State ─────────────────────────────────────────────────────────────────────
_gdir_state = {
    "status":         "idle",   # idle | running | waiting_captcha | waiting_otp | done | error
    "log":            [],
    "total":          0,
    "current":        0,
    "current_gstin":  "",
    "captcha_b64":    None,     # data:image/png;base64,… sent to UI
    "captcha_answer": None,     # set by /gstin_dir/enrich/submit_captcha
    "otp_answer":     None,     # set by /gstin_dir/enrich/submit_otp
    "results":        [],       # list of {gstin, result, legal_name, trade_name, status, …}
    "cancel":         False,
}
_gdir_lock = threading.Lock()


def _gdir_log(msg, level="info"):
    ts   = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    with _gdir_lock:
        _gdir_state["log"].append(line)
    # Route through the shared logging.FileHandler (opened once, kept
    # open) rather than a fresh open(LOG_FILE, "a") on every call —
    # see the matching comment in _comb_log for why the repeated-open
    # pattern is a real risk (confirmed to cause an indefinite hang in
    # that other function under real-world antivirus interference).
    (log.error if level == "error" else log.info)(f"[GDIR] {msg}")


def _gdir_set(**kw):
    with _gdir_lock:
        _gdir_state.update(kw)


# ══════════════════════════════════════════════════════════════════════════════
# Phase 1 — Scan all G2A + G2B JSON files for supplier GSTIN + trdnm
# ══════════════════════════════════════════════════════════════════════════════
@app.route("/gstin_dir/scan")
def gstin_dir_scan():
    """
    Walk G2A_DOWNLOAD_DIR, G2B_DOWNLOAD_DIR, and G1_DOWNLOAD_DIR, extract every
    unique GSTIN (suppliers from 2A/2B, buyers from G1) and trade name (trdnm).
    G1 buyer GSTINs are included so they appear in the Directory for enrichment.
    Returns a sorted list of dicts.
    """
    entries = {}   # ctin -> {gstin, trade_name, sources, fy_set, client_set}

    def _absorb(fpath, tag, owner_gstin, fy_label):
        try:
            with open(fpath, encoding="utf-8") as fh:
                data = json.load(fh)

            if tag == "G2A":
                # GSTR-2A offline JSON uses 'cdn'/'cdna' (NOT 'cdnr'/'cdnra')
                sections = [data.get("b2b", []), data.get("b2ba", []),
                            data.get("cdn", []), data.get("cdna", [])]
            else:                                   # G2B
                inner   = data.get("data", data)
                docdata = inner.get("docdata", {}) if isinstance(inner, dict) else {}
                sections = [docdata.get("b2b", []), docdata.get("cdnr", [])]

            for sec in sections:
                if not isinstance(sec, list):
                    continue
                for sup in sec:
                    if not isinstance(sup, dict):
                        continue
                    ctin  = (sup.get("ctin",  "") or "").strip().upper()
                    trdnm = (sup.get("trdnm", "") or "").strip()
                    if not ctin or len(ctin) != 15:
                        continue
                    if ctin not in entries:
                        entries[ctin] = {"gstin": ctin, "trade_name": "",
                                         "sources": set(), "fy_set": set(),
                                         "client_set": set()}
                    e = entries[ctin]
                    if trdnm and len(trdnm) > len(e["trade_name"]):
                        e["trade_name"] = trdnm
                    e["sources"].add(tag)
                    e["fy_set"].add(fy_label)
                    e["client_set"].add(owner_gstin)
        except Exception:
            pass

    for base_dir, tag in [(G2A_DOWNLOAD_DIR, "G2A"), (G2B_DOWNLOAD_DIR, "G2B")]:
        if not os.path.isdir(base_dir):
            continue
        for owner in sorted(os.listdir(base_dir)):
            od = os.path.join(base_dir, owner)
            if not os.path.isdir(od):
                continue
            for fy_dir in sorted(os.listdir(od)):
                fd = os.path.join(od, fy_dir)
                if not os.path.isdir(fd):
                    continue
                fy_lbl = fy_dir.replace("_", "-")
                for fname in sorted(os.listdir(fd)):
                    if fname.endswith(".json"):
                        _absorb(os.path.join(fd, fname), tag, owner, fy_lbl)

    # ── Also absorb G1 buyer GSTINs (ctin from b2b/b2ba sections) ────────────
    # G1 JSON has no 'trdnm' field — trade_name will be "" until enrichment runs.
    # These GSTINs are added to entries so they appear in the Directory table
    # and can be selected for enrichment via Search Taxpayer portal.
    if os.path.isdir(G1_DOWNLOAD_DIR):
        for owner in sorted(os.listdir(G1_DOWNLOAD_DIR)):
            od = os.path.join(G1_DOWNLOAD_DIR, owner)
            if not os.path.isdir(od):
                continue
            for fy_dir in sorted(os.listdir(od)):
                fd = os.path.join(od, fy_dir)
                if not os.path.isdir(fd):
                    continue
                fy_lbl = fy_dir.replace("_", "-")
                for fname in sorted(os.listdir(fd)):
                    if not fname.endswith(".json"):
                        continue
                    try:
                        with open(os.path.join(fd, fname), encoding="utf-8") as fh:
                            g1d = json.load(fh)
                        if isinstance(g1d, dict) and "data" in g1d:
                            g1d = g1d["data"]
                        for sec_key in ("b2b", "b2ba"):
                            for buyer in g1d.get(sec_key, []):
                                ctin = (buyer.get("ctin", "") or "").strip().upper()
                                if not ctin or len(ctin) != 15:
                                    continue
                                if ctin not in entries:
                                    entries[ctin] = {
                                        "gstin": ctin, "trade_name": "",
                                        "sources": set(), "fy_set": set(),
                                        "client_set": set()
                                    }
                                entries[ctin]["sources"].add("G1")
                                entries[ctin]["fy_set"].add(fy_lbl)
                                entries[ctin]["client_set"].add(owner)
                    except Exception:
                        pass

    out = []
    for gstin in sorted(entries):
        e = entries[gstin]
        out.append({
            "gstin":        gstin,
            "trade_name":   e["trade_name"],
            "sources":      sorted(e["sources"]),
            "fy_list":      sorted(e["fy_set"]),
            "client_count": len(e["client_set"]),
        })

    # ── Write per-return name files ──────────────────────────────────────────
    # 2a/: suppliers seen in G2A files (trdnm from G2B cross-scan since G2A has none)
    # 2b/: suppliers seen in G2B files (trdnm directly from G2B JSON)
    # 1/ : buyers seen in G1 b2b files (G1 has no trdnm, so trade_name is blank here;
    #      enrichment will populate it via portal search)
    def _build_per_return_files():
        g2a_entries = {}   # {gstin: trdnm} for G2A suppliers
        g2b_entries = {}   # {gstin: trdnm} for G2B suppliers
        g1_entries  = {}   # {gstin: ""}   for G1 buyers (no trdnm in G1 JSON)

        # G2B trdnm map (used to fill G2A entries too, since G2A has no trdnm)
        g2b_trdnm = {}
        if os.path.isdir(G2B_DOWNLOAD_DIR):
            for owner in os.listdir(G2B_DOWNLOAD_DIR):
                od = os.path.join(G2B_DOWNLOAD_DIR, owner)
                if not os.path.isdir(od): continue
                for fy_dir in os.listdir(od):
                    fd = os.path.join(od, fy_dir)
                    if not os.path.isdir(fd): continue
                    for fname in os.listdir(fd):
                        if not fname.endswith(".json"): continue
                        try:
                            with open(os.path.join(fd, fname), encoding="utf-8") as fh:
                                d = json.load(fh)
                            inner   = d.get("data", d)
                            docdata = inner.get("docdata", {}) if isinstance(inner, dict) else {}
                            for sec in ("b2b", "cdnr"):
                                for sup in docdata.get(sec, []):
                                    c = (sup.get("ctin","") or "").strip().upper()
                                    t = (sup.get("trdnm","") or "").strip()
                                    if c and t and c not in g2b_trdnm:
                                        g2b_trdnm[c] = t
                        except Exception:
                            pass

        for entry in out:
            gstin_key = entry["gstin"]
            srcs      = set(entry["sources"])
            trdnm     = entry["trade_name"] or g2b_trdnm.get(gstin_key, "")
            if "G2A" in srcs:
                g2a_entries[gstin_key] = g2b_trdnm.get(gstin_key, "")  # G2A has no trdnm
            if "G2B" in srcs:
                g2b_entries[gstin_key] = trdnm

        # G1 buyers — scan G1 JSON files for b2b ctin entries
        if os.path.isdir(G1_DOWNLOAD_DIR):
            for owner in os.listdir(G1_DOWNLOAD_DIR):
                od = os.path.join(G1_DOWNLOAD_DIR, owner)
                if not os.path.isdir(od): continue
                for fy_dir in os.listdir(od):
                    fd = os.path.join(od, fy_dir)
                    if not os.path.isdir(fd): continue
                    for fname in os.listdir(fd):
                        if not fname.endswith(".json"): continue
                        try:
                            with open(os.path.join(fd, fname), encoding="utf-8") as fh:
                                d = json.load(fh)
                            if isinstance(d, dict) and "data" in d:
                                d = d["data"]
                            for buyer in d.get("b2b", []) + d.get("b2ba", []):
                                c = (buyer.get("ctin","") or "").strip().upper()
                                if c and len(c) == 15:
                                    g1_entries[c] = g2b_trdnm.get(c, "")
                        except Exception:
                            pass

        if g2a_entries: _gnames_write_return_batch("2a", g2a_entries)
        if g2b_entries: _gnames_write_return_batch("2b", g2b_entries)
        if g1_entries:  _gnames_write_return_batch("1",  g1_entries)
        return len(g2a_entries), len(g2b_entries), len(g1_entries)

    try:
        n2a, n2b, n1 = _build_per_return_files()
        log.info(f"gstin_dir_scan: wrote per-return files "
                 f"(2a:{n2a} 2b:{n2b} 1:{n1} entries)")
    except Exception as _fe:
        log.warning(f"gstin_dir_scan: per-return file write failed: {_fe}")

    return jsonify({"ok": True, "count": len(out), "entries": out})


# ══════════════════════════════════════════════════════════════════════════════
# Phase 2 — Playwright enrichment: Login ONCE (G2B-style) → Search Taxpayer loop
# Flow mirrors g2b_do_browser_login:
#   Login captcha → OTP (if triggered) → dismiss popup →
#   nav to services.gst.gov.in/services/auth/searchtp →
#   loop: type GSTIN → click lotsearch → extract names → back to form
# No captcha per GSTIN — session persists across all lookups.
# ══════════════════════════════════════════════════════════════════════════════
def _gstin_dir_enrich_worker(gstins, username, password):
    """
    Thin wrapper around _gstin_dir_enrich_worker_impl that guarantees
    ANY exception is logged and reflected in _gdir_state, instead of
    silently killing the thread with nothing recorded anywhere — same
    reasoning as the matching wrapper on combined_download_worker.
    """
    try:
        _gstin_dir_enrich_worker_impl(gstins, username, password)
    except Exception as fatal:
        import traceback as _tb
        try:
            _gdir_log(f"✗ FATAL (uncaught): {fatal}", "error")
            _gdir_log(_tb.format_exc(), "error")
        except Exception:
            pass
        try:
            _gdir_set(status="error")
        except Exception:
            pass
        log.error(f"[GDIR] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _gstin_dir_enrich_worker_impl(gstins, username, password):
    """
    Login ONCE using G2B-style flow, then search each GSTIN on the
    authenticated Search Taxpayer page without any further CAPTCHAs.
    """
    import base64 as _b64

    # Diagnostic marker written to the persistent log FILE (not the
    # in-app activity panel, which gets wiped by the log=[] reset just
    # below — this line has to survive that reset to be useful). If we
    # ever see this line appear twice for what should be one enrichment
    # run, that proves two worker threads actually started; if it only
    # appears once, the duplication is happening somewhere else and the
    # route-level fix isn't the relevant piece.
    log.info(f"[GDIR-WORKER] ENTERED — thread={threading.current_thread().name} "
              f"gstins={len(gstins)}")

    _gdir_set(status="running", total=len(gstins), current=0,
              current_gstin="", results=[], log=[], cancel=False,
              captcha_b64=None, captcha_answer=None, otp_answer=None)

    # ── helpers ────────────────────────────────────────────────────────────────
    def _wait_field(field, timeout=300):
        deadline = time.time() + timeout
        while time.time() < deadline:
            if _gdir_state.get("cancel"):
                return None
            with _gdir_lock:
                val = _gdir_state.get(field)
                if val is not None:
                    _gdir_state[field] = None
                    return val
            time.sleep(0.35)
        return None

    def _cap_screenshot(page):
        """Capture login-page captcha image — mirrors g2b_do_browser_login."""
        for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                    "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                    ".captchaImage img", "img.captcha", "form img", "img"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    return "data:image/png;base64," + _b64.b64encode(png).decode()
            except Exception:
                continue
        # canvas fallback
        try:
            data_url = page.evaluate("""() => {
                for (const cv of document.querySelectorAll('canvas')) {
                    if (cv.width > 30 && cv.height > 10)
                        try { return cv.toDataURL('image/png'); } catch(e) {}
                }
                return null;
            }""")
            if data_url and data_url.startswith("data:image"):
                return data_url
        except Exception:
            pass
        # full-page screenshot last resort
        png = page.screenshot(full_page=False)
        return "data:image/png;base64," + _b64.b64encode(png).decode()

    def _fill_cap_field(page, answer):
        for sel in ["input[placeholder*='Characters' i]", "input#captcha",
                    "input[name='captcha']"]:
            try:
                page.locator(sel).first.fill(str(answer))
                return
            except Exception:
                continue

    def _submit_login(page):
        for sel in ["button[type='submit']", "input[type='submit']",
                    "button:has-text('LOGIN')"]:
            try:
                page.locator(sel).first.click()
                return
            except Exception:
                continue

    def _extract_names(page):
        """
        Extract Legal Name, Trade Name, Status from the authenticated searchtp result page.

        Actual portal DOM (from screenshot):
          <table>
            <tr>
              <td>Legal Name of Business</td>
              <td>Trade Name</td>
              <td>Additional Trade Name</td>
            </tr>
            <tr>
              <td><strong>GENERALI CENTRAL INSURANCE...</strong></td>
              <td><strong>GENERALI CENTRAL INSURANCE...</strong></td>
              <td>View</td>
            </tr>
          </table>
          Status is shown separately: "GSTIN / UIN Status" label → adjacent cell

        Strategy A: table-header column-index walk (primary — matches screenshot exactly)
        Strategy B: label→strong/b sibling within same cell
        Strategy C: regex on body innerText (fallback)
        """
        return page.evaluate(r"""() => {
            const c = s => (s || '').trim().replace(/\s+/g, ' ');
            const r = {lgnm: '', trdnm: '', sts: '', rgdt: '', ctb: '',
                       txptype: '', turnoverslab: ''};

            // ── Strategy A: table header row → next row same column ──────────
            // The portal renders Legal/Trade Name in a 2-row table:
            //   Row 1: "Legal Name of Business" | "Trade Name" | ...
            //   Row 2: <strong>ACTUAL NAME</strong> | ...
            const headerMap = {
                'legal name of business': 'lgnm',
                'legal name':             'lgnm',
                'trade name':             'trdnm',
                'trade name of business': 'trdnm',
            };
            document.querySelectorAll('table').forEach(tbl => {
                const rows = Array.from(tbl.querySelectorAll('tr'));
                rows.forEach((headerRow, ri) => {
                    const cells = Array.from(headerRow.querySelectorAll('td, th'));
                    cells.forEach((cell, ci) => {
                        const key = headerMap[c(cell.textContent).toLowerCase()];
                        if (!key || r[key]) return;
                        for (let ni = ri + 1; ni < Math.min(ri + 4, rows.length); ni++) {
                            const valCells = rows[ni].querySelectorAll('td, th');
                            if (ci < valCells.length) {
                                const v = c(valCells[ci].textContent);
                                if (v && v.length > 2 &&
                                    !v.toLowerCase().includes('legal name') &&
                                    !v.toLowerCase().includes('trade name')) {
                                    r[key] = v; break;
                                }
                            }
                        }
                    });
                });
            });

            // ── Strategy B: for names only — td/th siblings (safe, never div/span) ──
            if (!r.lgnm || !r.trdnm) {
                document.querySelectorAll('td, th').forEach(lbl => {
                    const t = c(lbl.textContent).toLowerCase();
                    // Only exact label matches to avoid partial matches in breadcrumbs etc.
                    if (!r.lgnm  && (t === 'legal name of business' || t === 'legal name')) {
                        const nxt = lbl.nextElementSibling;
                        if (nxt) { const v = c(nxt.textContent); if (v.length > 2) r.lgnm = v; }
                        const b = lbl.querySelector('strong, b');
                        if (!r.lgnm && b) r.lgnm = c(b.textContent);
                    }
                    if (!r.trdnm && (t === 'trade name' || t === 'trade name of business')) {
                        const nxt = lbl.nextElementSibling;
                        if (nxt) { const v = c(nxt.textContent); if (v.length > 2) r.trdnm = v; }
                        const b = lbl.querySelector('strong, b');
                        if (!r.trdnm && b) r.trdnm = c(b.textContent);
                    }
                });
            }

            // ── Strategy C: Angular scope (when available) ────────────────────
            if (!r.lgnm) {
                try {
                    const sc = angular && angular.element(document.body).scope();
                    const tp = sc && (sc.taxpayerDetails || sc.result || sc.tpData
                                   || sc.searchResult   || sc.tpDetails);
                    if (tp) {
                        if (!r.lgnm)  r.lgnm  = c(tp.lgnm    || tp.legalName    || '');
                        if (!r.trdnm) r.trdnm = c(tp.tradeNam || tp.tradeName   || '');
                    }
                } catch(e) {}
            }

            // ── Strategy D: regex on page innerText — ALWAYS runs for every field ──
            // The page starts with a CSS dump but the actual content appears as:
            //   "Legal Name of Business\nLE TRAVENUES TECHNOLOGY LIMITED\n"
            //   "GSTIN / UIN Status\nActive\n"
            //   "Constitution of Business\nPublic Limited Company\n"
            //   "Taxpayer Type\nRegular\n"
            //   "Annual Aggregate Turnover\nSlab: Rs. 0 to 40 lakhs (For FY 2024-2025)\n"
            // Regex anchors to the exact label text + newline so CSS junk never matches.
            // This runs unconditionally so it ALWAYS corrects garbage from earlier strategies.
            const full = (document.body || {}).innerText || '';

            // Names (regex fills if A/B/C missed)
            if (!r.lgnm) {
                const m = full.match(
                    /Legal Name(?:\s+of\s+Business)?\s*\n\s*([A-Z][A-Z0-9 &.,'\-()/]{2,120})/i);
                if (m) r.lgnm = c(m[1]);
            }
            if (!r.trdnm) {
                const m = full.match(
                    /\bTrade Name(?:\s+of\s+Business)?\s*\n\s*([A-Z][A-Z0-9 &.,'\-()/]{2,120})/i);
                if (m) r.trdnm = c(m[1]);
            }

            // ── sts: overwrite with regex result — ONLY known words accepted ──
            {
                const m = full.match(
                    /GSTIN\s*\/\s*UIN\s+Status\s*\n\s*(Active|Cancelled|Inactive|Suspended|Provisional)/i);
                r.sts = m ? c(m[1]) : '';
            }

            // ── rgdt: overwrite with regex result ────────────────────────────
            {
                const m = full.match(/Date of Registration\s*\n\s*(\d{2}\/\d{2}\/\d{4})/);
                r.rgdt = m ? c(m[1]) : (r.rgdt || '');
            }

            // ── ctb: overwrite with regex result ─────────────────────────────
            {
                const m = full.match(
                    /Constitution of Business\s*\n\s*([A-Za-z][A-Za-z0-9 /&\-()']{2,80})/i);
                r.ctb = m ? c(m[1]) : '';
            }

            // ── txptype: overwrite with regex result ──────────────────────────
            {
                const m = full.match(
                    /Taxpayer\s+Type\s*\n\s*([A-Za-z][A-Za-z0-9 \-/]{1,50})/i);
                r.txptype = m ? c(m[1]) : '';
            }

            // ── turnoverslab: overwrite with regex result ─────────────────────
            // Pattern: "Annual Aggregate Turnover\nSlab: Rs. 0 to 40 lakhs (For FY 2024-2025)"
            {
                const m = full.match(
                    /Annual\s+Aggregate\s+Turnover\s*\n\s*Slab\s*:\s*([^\n]{5,150})/i);
                r.turnoverslab = m ? c(m[1]) : '';
            }

            return r;
        }""")

    # ── Main worker ────────────────────────────────────────────────────────────
    # sync_playwright().start() launches Playwright's own Node.js driver
    # subprocess — if it throws (not just hangs), that exception was
    # previously uncaught here, which silently kills this thread with
    # nothing logged anywhere in a frozen/windowed build (see the
    # matching, more detailed comment on combined_download_worker).
    try:
        pw = sync_playwright().start()
    except Exception as e:
        _gdir_log(f"✗ Playwright driver failed to start: {e}", "error")
        _gdir_set(status="error"); return
    try:
        profile_dir = os.path.join(PATHS.profiles_dir, "gst_profile")
        os.makedirs(profile_dir, exist_ok=True)
        lock_file = os.path.join(profile_dir, "SingletonLock")
        if os.path.exists(lock_file):
            try: os.remove(lock_file)
            except Exception: pass

        _gdir_log("Launching browser…")
        browser = pw.chromium.launch(
            headless=False, slow_mo=80,
            args=["--disable-blink-features=AutomationControlled", "--no-sandbox"])
        ctx  = browser.new_context(viewport={"width": 1280, "height": 800})
        page = ctx.new_page()
        page.add_init_script("Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")

        # ══ STEP 1: Open login page ════════════════════════════════════════════
        _gdir_log("Opening GST login page…")
        page.goto(GST_LOGIN, wait_until="domcontentloaded", timeout=25000)
        try: page.wait_for_selector(
                "input#username, input[name='username'], input[placeholder*='username' i]",
                state="visible", timeout=8000)
        except Exception: pass

        # Fill username
        for sel in ["input#username", "input[name='username']",
                    "input[placeholder*='username' i]"]:
            try: page.locator(sel).first.fill(username); _gdir_log("  ✓ Username filled"); break
            except Exception: continue

        # Fill password
        for sel in ["input#user_pass", "input[name='user_pass']",
                    "input[type='password']", "input[placeholder*='password' i]"]:
            try: page.locator(sel).first.fill(password); _gdir_log("  ✓ Password filled"); break
            except Exception: continue
        time.sleep(0.3)

        # ══ STEP 2: Captcha (login page) — ONE TIME ════════════════════════════
        MAX_CAP = 3
        for cap_attempt in range(1, MAX_CAP + 1):
            cap_img = _cap_screenshot(page)
            _gdir_set(status="waiting_captcha", captcha_b64=cap_img, captcha_answer=None)
            _gdir_log(f"  ⏸ Waiting for login CAPTCHA (attempt {cap_attempt}/{MAX_CAP})…")

            answer = _wait_field("captcha_answer", timeout=300)
            _gdir_set(status="running", captcha_b64=None)
            if not answer:
                _gdir_log("  ✗ Captcha timeout — aborting", "error")
                _gdir_set(status="error"); browser.close(); pw.stop(); return

            _fill_cap_field(page, answer.strip())
            _gdir_log(f"  ✓ Captcha filled: {answer.strip()}")
            time.sleep(0.4)
            _submit_login(page)
            _gdir_log("  ✓ Login submitted")
            time.sleep(2)

            # ── OTP? ─────────────────────────────────────────────────────────
            otp_needed = False
            for _ in range(6):
                time.sleep(1)
                try:
                    otp_el = page.locator(
                        "input[placeholder*='OTP' i], input[id*='otp' i]").first
                    otp_el.wait_for(state="visible", timeout=1200)
                    otp_needed = True; break
                except Exception:
                    if check_login_success(page) is not False: break

            if otp_needed:
                _gdir_set(status="waiting_otp")
                _gdir_log("  ⏸ OTP required — waiting…")
                otp = _wait_field("otp_answer", timeout=180)
                if not otp:
                    _gdir_log("  ✗ OTP timeout — aborting")
                    _gdir_set(status="error"); browser.close(); pw.stop(); return
                _gdir_set(status="running")
                try:
                    otp_el = page.locator(
                        "input[placeholder*='OTP' i], input[id*='otp' i]").first
                    otp_el.click(); time.sleep(0.2); otp_el.fill(str(otp))
                    page.locator("button[type='submit'],input[type='submit']").first.click()
                    _gdir_log("  ✓ OTP submitted"); time.sleep(2)
                except Exception as e:
                    _gdir_log(f"  ✗ OTP error: {e}")

            # ── Check login result ────────────────────────────────────────────
            time.sleep(1)
            login_ok = check_login_success(page)
            if login_ok is not False:
                _gdir_log("  ✅ Logged in!")
                break
            if cap_attempt < MAX_CAP:
                _gdir_log(f"  ✗ Login failed — retrying captcha…")
                time.sleep(2)
                # re-fill credentials (page may have reset)
                for sel in ["input#username","input[name='username']"]:
                    try: page.locator(sel).first.fill(username); break
                    except Exception: continue
                for sel in ["input#user_pass","input[name='user_pass']","input[type='password']"]:
                    try: page.locator(sel).first.fill(password); break
                    except Exception: continue
            else:
                _gdir_log("  ✗ Login failed after all captcha attempts — aborting")
                _gdir_set(status="error"); browser.close(); pw.stop(); return

        # ══ STEP 3: Dismiss popup ══════════════════════════════════════════════
        try:
            g2a_dismiss_popup(page)
        except Exception as e:
            _gdir_log(f"  ⚠ Popup dismiss: {e}")

        # ══ STEP 4: Navigate to Search Taxpayer (authenticated URL) ═══════════
        _gdir_log("Navigating to Search Taxpayer…")
        page.goto("https://services.gst.gov.in/services/auth/searchtp",
                  wait_until="domcontentloaded", timeout=20000)
        time.sleep(1.5)

        if "login" in (page.url or "").lower():
            _gdir_log("  ✗ Session expired — redirected to login", "error")
            _gdir_set(status="error"); browser.close(); pw.stop(); return

        # Click "Search by GSTIN/UIN" tab
        page.evaluate("""() => {
            for (const a of document.querySelectorAll('a')) {
                const t = (a.textContent||'').trim().toLowerCase();
                if (t.includes('gstin') || t.includes('uin')) { a.click(); return; }
            }
        }""")
        time.sleep(1)
        try:
            page.wait_for_selector(
                "input[id*='gstin' i], input[name*='gstin' i], input[type='text']",
                state="visible", timeout=8000)
            _gdir_log("  ✓ Search Taxpayer page ready")
        except Exception:
            _gdir_log("  ⚠ GSTIN input not found — proceeding anyway")

        # ══ STEP 5: Loop — one GSTIN at a time, no further CAPTCHA ═══════════
        # The input field stays visible on the result page — just clear & retype.
        GSTIN_SELS = [
            "input[id*='gstin' i]", "input[name*='gstin' i]",
            "input[placeholder*='GSTIN' i]", "input[type='text']",
        ]

        for idx, gstin in enumerate(gstins):
            if _gdir_state.get("cancel"):
                _gdir_log("Cancelled by user."); break

            _gdir_set(current=idx, current_gstin=gstin)
            _gdir_log(f"━━ [{idx+1}/{len(gstins)}] {gstin}")

            try:
                # ── If first GSTIN, input is already on screen and ready.
                # ── For subsequent GSTINs the page shows the previous result
                #    but the input box is still at the top — just clear & retype.
                if idx > 0:
                    # Make sure we're still on the searchtp page
                    if "searchtp" not in (page.url or ""):
                        _gdir_log("  ↩ Re-navigating to searchtp…")
                        page.goto("https://services.gst.gov.in/services/auth/searchtp",
                                  wait_until="domcontentloaded", timeout=12000)
                        time.sleep(1.2)
                        page.evaluate("""() => {
                            for (const a of document.querySelectorAll('a')) {
                                const t = (a.textContent||'').trim().toLowerCase();
                                if (t.includes('gstin') || t.includes('uin'))
                                    { a.click(); return; }
                            }
                        }""")
                        time.sleep(0.8)

                # ── Fill GSTIN field (clears existing value via select_text+Delete) ──
                filled = False
                for sel in GSTIN_SELS:
                    try:
                        loc = page.locator(sel).first
                        loc.wait_for(state="visible", timeout=6000)
                        loc.scroll_into_view_if_needed()
                        loc.click(); time.sleep(0.15)
                        # Triple-click selects all, then type replaces
                        loc.click(click_count=3); time.sleep(0.1)
                        loc.fill("")             ; time.sleep(0.1)
                        loc.press_sequentially(gstin, delay=55)
                        time.sleep(0.3)
                        val = loc.input_value()
                        if val.strip():
                            filled = True; break
                    except Exception:
                        continue

                if not filled:
                    _gdir_log(f"  ✗ GSTIN input not found — skipping")
                    with _gdir_lock:
                        _gdir_state["results"].append({
                            "gstin": gstin, "result": "failed",
                            "legal_name": "", "trade_name": "", "status": "",
                        })
                    continue

                # ── Click SEARCH button ──────────────────────────────────────
                page.evaluate("""() => {
                    // id=lotsearch is the SEARCH button on the authenticated searchtp page
                    const b = document.getElementById('lotsearch');
                    if (b) { b.click(); return; }
                    for (const el of document.querySelectorAll('button, input[type=submit]')) {
                        const t = (el.textContent || el.value || '').trim().toUpperCase();
                        if (t === 'SEARCH') { el.click(); return; }
                    }
                }""")

                # ── Wait for result to render (look for "Legal Name" or result heading) ──
                result_appeared = False
                for _ in range(12):        # up to ~3.6 s
                    time.sleep(0.3)
                    try:
                        body_text = page.inner_text("body") or ""
                        if ("Legal Name" in body_text or
                                "Search Result based on" in body_text):
                            result_appeared = True
                            break
                    except Exception:
                        pass
                if not result_appeared:
                    time.sleep(1.5)        # graceful extra wait

                # ── Extract names ────────────────────────────────────────────
                rjs          = _extract_names(page)
                lgnm         = (rjs or {}).get("lgnm",         "").strip()
                trdnm        = (rjs or {}).get("trdnm",        "").strip()
                sts          = (rjs or {}).get("sts",          "").strip()
                rgdt         = (rjs or {}).get("rgdt",         "").strip()
                ctb          = (rjs or {}).get("ctb",          "").strip()
                txptype      = (rjs or {}).get("txptype",      "").strip()
                turnoverslab = (rjs or {}).get("turnoverslab", "").strip()

                if lgnm or trdnm:
                    _gdir_log(
                        f"  ✓  Legal: {lgnm or '—'}  Trade: {trdnm or '—'}  "
                        f"Status: {sts or '—'}  Type: {txptype or '—'}")
                    # ── Save to disk immediately — this is the primary store ──
                    _enrich_ts = datetime.now().isoformat()
                    _gnames_upsert(gstin, {
                        "trade_name": trdnm, "legal_name": lgnm,
                        "status": sts, "reg_date": rgdt,
                        "constitution": ctb, "taxpayer_type": txptype,
                        "turnover_slab": turnoverslab,
                        "enrich_ts": _enrich_ts,
                    })
                    _gdir_log(f"  💾 Saved → gstin_names/enriched/name_lookup.json")
                    with _gdir_lock:
                        _gdir_state["results"].append({
                            "gstin": gstin, "result": "ok",
                            "legal_name":    lgnm,    "trade_name":    trdnm,
                            "status":        sts,     "reg_date":      rgdt,
                            "constitution":  ctb,     "taxpayer_type": txptype,
                            "turnover_slab": turnoverslab,
                        })
                else:
                    _gdir_log(f"  ⚠ No data extracted for {gstin}")
                    with _gdir_lock:
                        _gdir_state["results"].append({
                            "gstin": gstin, "result": "failed",
                            "legal_name": "", "trade_name": "", "status": "",
                        })

            except Exception as exc:
                _gdir_log(f"  ✗ Error: {exc}")
                with _gdir_lock:
                    _gdir_state["results"].append({
                        "gstin": gstin, "result": "failed",
                        "legal_name": "", "trade_name": "", "status": "",
                    })

            time.sleep(0.2)   # brief breath before next GSTIN

        # Logout
        try:
            page.goto(f"{GST_PORTAL_URL}/services/logout", timeout=8000)
            _gdir_log("✓ Logged out")
        except Exception:
            pass
        browser.close()

    except Exception as exc:
        _gdir_log(f"Worker fatal: {exc}")
        _gdir_set(status="error"); pw.stop(); return
    finally:
        try: pw.stop()
        except Exception: pass

    done_count = len(_gdir_state["results"])
    ok_count   = sum(1 for r in _gdir_state["results"] if r.get("result") == "ok")
    _gdir_log(f"✅ Enrichment complete — {ok_count}/{done_count} GSTINs enriched")
    _gdir_set(status="done", current=len(gstins), captcha_b64=None)



# ── Flask routes ──────────────────────────────────────────────────────────────
@app.route("/gstin_dir/enrich/start", methods=["POST"])
def gstin_dir_enrich_start():
    body     = request.get_json(force=True) or {}
    gstins   = body.get("gstins",   [])
    username = (body.get("username", "") or "").strip()
    password = (body.get("password", "") or "").strip()
    if not gstins:
        return jsonify({"ok": False, "error": "No GSTINs provided"}), 400
    if not username or not password:
        return jsonify({"ok": False, "error": "Username and password are required"}), 400

    # Diagnostic marker (see matching comment in _gstin_dir_enrich_worker)
    # — written to the persistent log FILE for every single call this
    # route receives, whether accepted or rejected, so we can see exactly
    # how many requests actually arrived if this issue is reported again.
    log.info(f"[GDIR-ROUTE] /gstin_dir/enrich/start called — thread="
              f"{threading.current_thread().name}")

    # Claim the "running" state atomically, under the lock, BEFORE
    # spawning the worker thread — not after. Previously the check here
    # ("is a worker already running?") happened in this route, but the
    # answer only became true once the spawned thread got around to
    # calling _gdir_set(status="running", ...) itself, which can be
    # milliseconds (or longer, under load) after .start() returns. Two
    # near-simultaneous requests (e.g. a double-click on Start
    # Enrichment before the button visibly changes) could both see
    # status="idle" in that window and both spawn a worker — two
    # Chromium instances logging into the same GST account at once,
    # which is exactly what produced duplicated "Username filled" /
    # "Password filled" log lines and much slower, flakier logins.
    with _gdir_lock:
        st = _gdir_state.get("status", "idle")
        if st in ("running", "waiting_captcha", "waiting_otp"):
            log.info(f"[GDIR-ROUTE] REJECTED (409) — status was already '{st}'")
            return jsonify({"ok": False, "error": "Enrichment already running"}), 409
        _gdir_state.update(status="running", total=len(gstins), current=0,
                           current_gstin="", results=[], log=[], cancel=False,
                           captcha_b64=None, captcha_answer=None, otp_answer=None)
        log.info(f"[GDIR-ROUTE] ACCEPTED — claimed status='running', spawning worker")

    threading.Thread(target=_gstin_dir_enrich_worker,
                     args=(gstins, username, password), daemon=True).start()
    return jsonify({"ok": True, "total": len(gstins)})


@app.route("/gstin_dir/enrich/status")
def gstin_dir_enrich_status():
    with _gdir_lock:
        s = _gdir_state
        return jsonify({
            "status":        s.get("status", "idle"),
            "total":         s.get("total",  0),
            "current":       s.get("current", 0),
            "current_gstin": s.get("current_gstin", ""),
            "captcha_b64":   s.get("captcha_b64"),
            "results":       s.get("results", [])[-50:],
            "result_count":  len(s.get("results", [])),
            "log":           s.get("log", [])[-60:],
        })


@app.route("/gstin_dir/enrich/submit_captcha", methods=["POST"])
def gstin_dir_enrich_submit_captcha():
    body   = request.get_json(force=True) or {}
    answer = (body.get("captcha", "") or "").strip()
    if not answer:
        return jsonify({"ok": False, "error": "Empty captcha"}), 400
    with _gdir_lock:
        _gdir_state["captcha_answer"] = answer
    return jsonify({"ok": True})


@app.route("/gstin_dir/enrich/submit_otp", methods=["POST"])
def gstin_dir_enrich_submit_otp():
    body = request.get_json(force=True) or {}
    otp  = (body.get("otp", "") or "").strip()
    if not otp:
        return jsonify({"ok": False, "error": "Empty OTP"}), 400
    with _gdir_lock:
        _gdir_state["otp_answer"] = otp
    return jsonify({"ok": True})


@app.route("/gstin_dir/enrich/cancel", methods=["POST"])
def gstin_dir_enrich_cancel():
    with _gdir_lock:
        _gdir_state["cancel"] = True
    return jsonify({"ok": True})


@app.route("/gstin_dir/enrich/reset", methods=["POST"])
def gstin_dir_enrich_reset():
    with _gdir_lock:
        _gdir_state.update({
            "status": "idle", "log": [], "total": 0, "current": 0,
            "current_gstin": "", "captcha_b64": None, "captcha_answer": None,
            "results": [], "cancel": False,
        })
    return jsonify({"ok": True})


@app.route("/gstin_names", methods=["GET"])
def gstin_names_get():
    """Return the full enriched name_lookup.json for JS display."""
    store = _gnames_load()
    return jsonify({"ok": True, "count": len(store), "data": store})


@app.route("/gstin_names/save", methods=["POST"])
def gstin_names_save():
    """
    JS pushes enrichment results here as:
      { records: [{gstin, trade_name, legal_name, status, ...}] }
    Merges with existing — non-empty values always win.
    """
    body    = request.get_json(force=True) or {}
    records = body.get("records", []) or []
    if not records:
        return jsonify({"ok": False, "error": "No records"}), 400
    total = _gnames_save_batch(records)
    return jsonify({"ok": True, "total_in_store": total, "saved": len(records)})


@app.route("/gstin_names/status", methods=["GET"])
def gstin_names_status():
    """Count of names in each per-return and enriched file."""
    def _count(fpath):
        d = _gnames_load_file(fpath)
        return {
            "total":      len(d),
            "with_trade": sum(1 for v in d.values() if v.get("trade_name")),
            "with_legal": sum(1 for v in d.values() if v.get("legal_name")),
        }
    enriched = _count(_GNAMES_ENRICHED)
    return jsonify({
        "ok":      True,
        "enriched": enriched,
        "2a":       _count(_GNAMES_2A),
        "2b":       _count(_GNAMES_2B),
        "1":        _count(_GNAMES_1),
        # Convenience totals (enriched is primary, used for UI badge)
        "total":       enriched["total"],
        "with_trade":  enriched["with_trade"],
        "with_legal":  enriched["with_legal"],
        "file":        _GNAMES_ENRICHED,
    })


@app.route("/gstin_names/clear", methods=["POST"])
def gstin_names_clear():
    """Reset the enriched name store (delete the JSON file)."""
    with _gnames_lock:
        try:
            os.remove(_GNAMES_ENRICHED)
        except FileNotFoundError:
            pass
    return jsonify({"ok": True})



# ══════════════════════════════════════════════════════════════════════════════
#  COMBINED DOWNLOAD — Single login, all modules in one browser session
#  Modules: GSTR-1 | GSTR-2A | GSTR-2B | GSTR-3B | TDS/TCS
#  Login once → activate session once → download all selected reports
# ══════════════════════════════════════════════════════════════════════════════

_comb_lock  = threading.Lock()
_comb_state = {
    "status":         "idle",   # idle/login/running/done/error
    "log":            [],
    "progress":       0,
    "current_module": None,
    "modules_done":   [],
    "modules_failed": [],
    "error":          None,
    "captcha_image":  None,
    "captcha_answer": None,
    "otp_answer":     None,
    "gstin":          None,
    "fy":             None,
    "stop_requested": False,
}

def _comb_log(msg, level="info"):
    with _comb_lock:
        _comb_state["log"].append({"msg": msg, "level": level,
                                    "ts": time.strftime("%H:%M:%S")})
        if len(_comb_state["log"]) > 1500:
            _comb_state["log"] = _comb_state["log"][-1500:]
    # Route through the shared logging.FileHandler (opened once at
    # startup and kept open) instead of opening/closing LOG_FILE on
    # every single call. The previous version did open(LOG_FILE, "a")
    # fresh each time — for a feature that logs dozens of lines in
    # quick succession, that's a fresh open() on the same file over
    # and over, which is exactly the kind of repeated file-touch
    # pattern real-time antivirus scanning can lock up on. Confirmed:
    # a real run logged its first line successfully, then froze
    # completely on the very next _comb_log() call with zero code in
    # between them — the open() call itself is the only thing in that
    # gap capable of blocking indefinitely.
    (log.error if level == "error" else log.info)(f"[COMB] {msg}")

def _comb_set(d):
    with _comb_lock:
        _comb_state.update(d)


# ── Per-module run helpers ─────────────────────────────────────────────────────

def _comb_run_g1(page, gstin, fy):
    """Run GSTR-1 download on a shared authenticated page."""
    gstin_dir = os.path.join(G1_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    os.makedirs(gstin_dir, exist_ok=True)
    months = g1_months_for_fy(fy)
    if not months:
        _comb_log(f"  ⚠ GSTR-1: no months available for FY {fy}")
        return 0
    g1_set({"status": "running", "log": [], "error": None, "progress": 0,
            "total_months": len(months), "done_months": 0, "files": [],
            "gstin": gstin, "fy": fy, "current_month": None})
    g1_log(f"GSTR-1 | GSTIN: {gstin} | FY: {fy} | {len(months)} month(s)  [combined session]")
    # Trigger pass only (max_sweeps=1) — no 300s wait; harvest runs after other modules
    files_done, pending = g1_download_all_months(
        page, months, gstin, gstin_dir, fy,
        inter_sweep_wait=0, max_sweeps=1
    )
    g1_set({"status": "running", "current_month": None,
            "files": files_done, "done_months": len(files_done)})
    if pending:
        g1_log(f"  ⏩ Trigger pass: {len(files_done)} ready, {len(pending)} queued")
        _comb_log(f"  ⏩ GSTR-1: {len(files_done)} ready, {len(pending)} queued (portal generating)")
    else:
        g1_set({"status": "done", "progress": 100})
        _comb_log(f"  ✅ GSTR-1: all {len(files_done)} months downloaded in trigger pass")
    return files_done, pending



def _comb_run_g1_harvest(page, gstin, fy, pending_months):
    """Harvest pass: collect G1 months the portal had time to generate."""
    if not pending_months:
        return []
    gstin_dir = os.path.join(G1_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    _comb_log(f"  🔁 GSTR-1 harvest: {len(pending_months)} pending month(s)...")
    g1_log(f"\n🔁 GSTR-1 harvest — {len(pending_months)} month(s) to collect")
    harvest_files, still_pending = g1_download_all_months(
        page, pending_months, gstin, gstin_dir, fy, inter_sweep_wait=0
    )
    with g1_lock:
        existing = list(g1_state.get("files", []))
        all_files = existing + harvest_files
        g1_state["files"]       = all_files
        g1_state["done_months"] = len(all_files)
        g1_state["status"]      = "done"
        g1_state["progress"]    = 100
    if still_pending:
        _comb_log(f"  ⚠ GSTR-1: {len(still_pending)} month(s) still unavailable after harvest")
    else:
        _comb_log(f"  ✅ GSTR-1 harvest: {len(harvest_files)} more month(s) collected")
    return harvest_files
def _comb_run_g2a(page, gstin, fy):
    """Run GSTR-2A download on a shared authenticated page."""
    gstin_dir = os.path.join(G2A_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    os.makedirs(gstin_dir, exist_ok=True)
    months = g2a_months_for_fy(fy)
    if not months:
        _comb_log(f"  ⚠ GSTR-2A: no months available for FY {fy}")
        return 0
    fresh_results, months_needed = g2a_check_cached_months(gstin_dir, months, force=False)
    if not months_needed:
        _comb_log(f"  ✅ GSTR-2A: all {len(months)} months already cached")
        g2a_set({"status": "done", "progress": 100, "files": fresh_results,
                 "done_months": len(fresh_results), "gstin": gstin, "fy": fy})
        return len(fresh_results)
    g2a_set({"status": "running", "log": [], "error": None, "progress": 0,
             "total_months": len(months_needed), "done_months": 0,
             "files": fresh_results, "gstin": gstin, "fy": fy})
    g2a_log(f"GSTR-2A | GSTIN: {gstin} | FY: {fy} | {len(months_needed)} month(s)  [combined session]")
    new_files = g2a_download_all_months(page, months_needed, gstin, gstin_dir, fy)
    all_files  = fresh_results + new_files
    g2a_set({"status": "done", "progress": 100, "current_month": None,
             "files": all_files, "done_months": len(all_files)})
    g2a_log(f"✅ GSTR-2A complete — {len(new_files)} downloaded, {len(fresh_results)} cached")
    _comb_log(f"  ✅ GSTR-2A: {len(new_files)} downloaded, {len(fresh_results)} from cache")
    return len(all_files)


def _comb_run_g2b(page, gstin, fy):
    """Run GSTR-2B download on a shared authenticated page."""
    gstin_dir = os.path.join(G2B_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    os.makedirs(gstin_dir, exist_ok=True)
    months = g2b_months_for_fy(fy)
    if not months:
        _comb_log(f"  ⚠ GSTR-2B: no months available for FY {fy}")
        return 0
    g2b_set({"status": "running", "log": [], "error": None, "progress": 0,
             "total_months": len(months), "done_months": 0, "files": [],
             "gstin": gstin, "fy": fy})
    g2b_log(f"GSTR-2B | GSTIN: {gstin} | FY: {fy} | {len(months)} month(s)  [combined session]")
    files_done = g2b_download_all_months(page, months, gstin, gstin_dir, fy)
    g2b_set({"status": "done", "progress": 100, "current_month": None,
             "files": files_done, "done_months": len(files_done)})
    g2b_log(f"✅ GSTR-2B complete — {len(files_done)}/{len(months)} months")
    _comb_log(f"  ✅ GSTR-2B: {len(files_done)}/{len(months)} months downloaded")
    return len(files_done)


def _comb_run_g3b(page, context, gstin, fy):
    """
    Run GSTR-3B scrape on a shared authenticated page.

    Uses IDENTICAL per-month loop logic to the standalone g3b_worker:
    - Same g3b_navigate_to_period call signature: (page, context, fy, mon_num, already_qrmp=False)
    - Same 4-tuple return: (ok, reason, freq, nav_result)
    - Same nav_result tuple: (form_page, is_new_tab)
    - Same reason checks: tile_not_found / qrmp_pmt06 / view / prepare_online / resume
    - Same tab handling: close if new tab, navigate back to dashboard if same tab
    """
    fy_tag    = fy.replace("-", "_")
    gstin_dir = os.path.join(G3B_DOWNLOAD_DIR, gstin, fy_tag)
    os.makedirs(gstin_dir, exist_ok=True)
    fy_start     = int(fy.split("-")[0])
    months_in_fy = [4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3]
    now          = datetime.now()

    # ── Cache check (same as standalone) ────────────────────────────────────
    months_needed = []
    fresh_files   = []
    for m in months_in_fy:
        year  = fy_start if m >= 4 else fy_start + 1
        if datetime(year, m, 1) > now:
            continue
        fname = f"GSTR3B_{_G3B_MON[m]}_{fy_tag}.json"
        fpath = os.path.join(gstin_dir, fname)
        if os.path.exists(fpath):
            fresh_files.append(fpath)
        else:
            months_needed.append(m)

    if not months_needed:
        _comb_log(f"  ✅ GSTR-3B: all months already cached")
        g3b_set({"status": "done",
                 "files":       [os.path.basename(f) for f in fresh_files],
                 "fresh_files": [os.path.basename(f) for f in fresh_files]})
        return len(fresh_files)

    g3b_set({"status": "running", "log": [], "error": None,
             "files":       [os.path.basename(f) for f in fresh_files],
             "fresh_files": [os.path.basename(f) for f in fresh_files]})
    g3b_log(f"GSTR-3B | GSTIN: {gstin} | FY: {fy} | "
            f"{len(months_needed)} month(s)  [combined session]")
    g3b_log(f"  {len(fresh_files)} month(s) from cache, "
            f"{len(months_needed)} to scrape")

    new_files = []

    # ── Per-month scrape loop — identical to standalone g3b_worker ───────────
    for mon_num in months_needed:
        mon_name = _G3B_MON[mon_num]
        g3b_log(f"\n[{mon_name}] Processing...")
        g3b_set({"status": "running"})

        try:
            # Same call signature as standalone: returns (ok, reason, freq, nav_result)
            ok, reason, freq, nav_result = g3b_navigate_to_period(
                page, context, fy, mon_num, already_qrmp=False)

            mon_is_qrmp = (freq == "quarterly")
            if freq == "quarterly":
                g3b_log(f"  ℹ {mon_name}: quarterly frequency (QRMP)")
            elif freq == "monthly":
                g3b_log(f"  ℹ {mon_name}: monthly frequency")

            if not ok:
                if reason == "tile_not_found" and mon_num not in _G3B_QTR_END:
                    g3b_log(f"  ⏭ {mon_name}: no GSTR-3B tile — IFF month (QRMP), skipping")
                elif reason == "qrmp_pmt06":
                    g3b_log(f"  ⏭ {mon_name}: PMT-06/QRMP signal — IFF month, skipping")
                elif reason == "wrong_url":
                    g3b_log(f"  ✗ {mon_name}: portal navigated to wrong form — skip")
                else:
                    g3b_log(f"  ⚠ {mon_name}: {reason} — skipping")
                continue

            # Only scrape VIEW status (filed return with real data)
            if reason != "view":
                if reason == "prepare_online":
                    g3b_log(f"  ⏭ {mon_name} = PREPARE ONLINE — not yet filed, skipping")
                elif reason == "resume":
                    g3b_log(f"  ⏭ {mon_name} = RESUME (draft) — skipping")
                else:
                    g3b_log(f"  ⏭ {mon_name} = {reason} — skipping")
                continue

            # nav_result is (form_page, is_new_tab) — same as standalone
            form_page, is_new_tab = nav_result
            filing_type = "qrmp" if mon_is_qrmp else "monthly"
            fpath = g3b_scrape_period(
                form_page, fy, mon_num, gstin, gstin_dir, filing_type)
            if fpath:
                new_files.append(fpath)
                g3b_log(f"  ✅ {mon_name} scraped")

            # Tab handling — same as standalone
            if is_new_tab:
                try:
                    form_page.close()
                    g3b_log(f"    📄 Closed GSTR-3B tab for {mon_name}")
                except Exception:
                    pass
            else:
                # Same-tab: navigate back to dashboard for next iteration
                try:
                    page.evaluate(f"window.location.href = '{_G3B_DASHBOARD}'")
                    page.wait_for_load_state("domcontentloaded", timeout=8000)
                    time.sleep(1.0)
                except Exception:
                    pass

        except Exception as e:
            import traceback as _tb
            g3b_log(f"  ✗ {mon_name}: {e}")
            g3b_log(_tb.format_exc())

    all_files = fresh_files + new_files
    g3b_set({"status": "done",
             "files":       [os.path.basename(f) for f in all_files],
             "fresh_files": [os.path.basename(f) for f in fresh_files]})
    g3b_log(f"\n✅ GSTR-3B complete — {len(new_files)} scraped, "
            f"{len(fresh_files)} cached, {len(all_files)} total")
    _comb_log(f"  ✅ GSTR-3B: {len(new_files)} scraped, {len(fresh_files)} from cache")
    return len(all_files)


def _comb_run_tds(page, context, gstin, fy):
    """
    Run TDS/TCS download on a shared authenticated page.
    Identical logic to standalone tds_download_worker — same helpers,
    same two-sweep structure, same cache check, same file naming.
    No browser launch / login / session activation (shared session handles that).
    """
    OFFLINE_PATH = "gstr2xco/offlinedownload"

    fy_start = int(fy.split("-")[0])
    now      = datetime.now()

    # Build months list in same format as standalone worker: {display, period, num}
    _MN_FULL = {4:"April",5:"May",6:"June",7:"July",8:"August",9:"September",
                10:"October",11:"November",12:"December",1:"January",2:"February",3:"March"}
    months = []
    for m in [4,5,6,7,8,9,10,11,12,1,2,3]:
        yr = fy_start if m >= 4 else fy_start + 1
        if datetime(yr, m, 1) <= now:
            months.append({
                "display": f"{_MN_FULL[m]} {yr}",
                "period":  f"{str(m).zfill(2)}{yr}",
                "num":     m,
            })

    if not months:
        _comb_log(f"  ⚠ TDS: no months available for FY {fy}")
        return 0

    gstin_dir = os.path.join(TDS_DOWNLOAD_DIR, gstin, fy.replace("-", "_"))
    os.makedirs(gstin_dir, exist_ok=True)

    tds_set({"status": "running", "log": [], "error": None, "progress": 0,
             "done_months": 0, "files": [], "gstin": gstin, "fy": fy})
    tds_log(f"TDS/TCS | GSTIN: {gstin} | FY: {fy} | "
            f"{len(months)} month(s)  [combined session]")

    # ── Nested helpers — identical to standalone tds_download_worker ─────────

    def real_url(p):
        try:    return p.evaluate("location.href") or p.url
        except: return p.url

    def get_real_url():
        return real_url(page)

    def navigate_to_comptds(mon_name_str):
        nonlocal page
        if "comptds" in real_url(page).lower():
            _select_fy_and_month(mon_name_str)
            return
        tds_log(f"  -> Not on comptds — re-navigating via Services menu...")
        nav_ok = False
        for sel in ["ul.nav.navbar-nav li.dropdown > a:has-text('Services')",
                    "li.dropdown > a:has-text('Services')",
                    "nav a:has-text('Services')", "a:text-is('Services')"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=4000)
                loc.click(); tds_log("    ✓ Services clicked"); time.sleep(0.7); break
            except Exception: continue
        for sel in ["ul.nav-tabs a:text-is('Returns')",
                    ".nav-tabs li a:text-is('Returns')", "a:text-is('Returns')"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.hover(); tds_log("    ✓ Returns hovered"); time.sleep(0.8); break
            except Exception: continue
        for sel in ["a:text-is('TDS and TCS credit received')",
                    "a:has-text('TDS and TCS credit received')",
                    "a:has-text('TDS & TCS credit received')"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=4000)
                loc.click()
                tds_log("    ✓ 'TDS and TCS credit received' clicked")
                nav_ok = True; break
            except Exception: continue
        if not nav_ok:
            tds_log("    ⚠ Link not found — using tds_activate_session")
            ok2, page = tds_activate_session(page, context)
            if not ok2: return
        for _ in range(15):
            time.sleep(1)
            if "comptds" in real_url(page).lower(): break
            for p in context.pages:
                if "comptds" in real_url(p).lower():
                    page = p; break
        try: page.wait_for_load_state("domcontentloaded", timeout=8000)
        except Exception: pass
        time.sleep(1.0)
        _select_fy_and_month(mon_name_str)

    def _select_fy_and_month(mon_name_str):
        page.evaluate(f"""() => {{
            const sel = document.querySelectorAll('select')[0];
            if (!sel) return;
            for (const opt of sel.options) {{
                if (opt.text.trim() === '{fy}') {{
                    sel.value = opt.value;
                    sel.dispatchEvent(new Event('change', {{bubbles:true}}));
                    return;
                }}
            }}
        }}""")
        time.sleep(0.5)
        page.evaluate(f"""() => {{
            const sel = document.querySelectorAll('select')[1];
            if (!sel) return;
            for (const opt of sel.options) {{
                if (opt.text.trim() === '{mon_name_str}') {{
                    sel.value = opt.value;
                    sel.dispatchEvent(new Event('change', {{bubbles:true}}));
                    return;
                }}
            }}
        }}""")
        for s in ["button:has-text('SEARCH')", "button:has-text('Search')"]:
            try:
                page.locator(s).first.wait_for(state="visible", timeout=5000)
                page.locator(s).first.click(); break
            except Exception: continue
        time.sleep(2)

    def click_to_offline_download():
        result = page.evaluate("""() => {
            const btns = Array.from(document.querySelectorAll('button,a'));
            for (const b of btns) {
                if ((b.textContent||'').trim().toUpperCase() === 'DOWNLOAD') {
                    b.click(); return 'filed';
                }
            }
            for (const b of btns) {
                const t = (b.textContent||'').trim().toUpperCase();
                if (t === 'PREPARE OFFLINE' || t.includes('PREPARE OFFLINE')) {
                    b.click(); return 'not_filed';
                }
            }
            return null;
        }""")
        if not result: return None
        try: page.wait_for_url("*gstr2xco*", timeout=10000)
        except Exception: pass
        for _ in range(5):
            cur = get_real_url()
            if "gstr2xco" in cur or "offlineupload" in cur or "offlinedownload" in cur: break
            time.sleep(1)
        time.sleep(1)
        cur = get_real_url()
        tds_log(f"  -> After click, URL: {cur[30:80]}")
        if OFFLINE_PATH in cur: return result
        tds_log("  -> On Upload tab — clicking Download tab...")
        try:
            tab_loc = page.locator("a:has-text('Download'), li:has-text('Download')")
            for i in range(tab_loc.count()):
                if (tab_loc.nth(i).text_content() or "").strip().lower() == "download":
                    tab_loc.nth(i).click()
                    tds_log("    -> Download tab clicked"); break
            else:
                page.locator("text=Download").first.click()
        except Exception as e:
            tds_log(f"    ⚠ Tab click: {e}")
        try: page.wait_for_url(f"*{OFFLINE_PATH}*", timeout=8000)
        except Exception: pass
        time.sleep(1)
        return result

    def _save_download(dl_value, json_path, mon_display, period):
        import zipfile, io as _io
        try:
            raw = open(dl_value.path(), "rb").read()
            if raw[:2] == b"PK":
                with zipfile.ZipFile(_io.BytesIO(raw)) as zf:
                    names  = zf.namelist()
                    target = next((n for n in names if n.endswith(".json")), names[0])
                    data   = zf.read(target)
            else:
                data = raw
            with open(json_path, "wb") as jf: jf.write(data)
            size_kb = max(1, len(data) // 1024)
            tds_log(f"    ✅ {mon_display} — {size_kb} KB saved")
            return {"file": json_path, "period": period,
                    "month": mon_display, "size_kb": size_kb}
        except Exception as e:
            tds_log(f"    ⚠ Save error: {e}"); return None

    def click_json_and_intercept(mon_display, period):
        safe_m    = mon_display.replace(" ", "_").replace("/", "_")
        json_path = os.path.join(gstin_dir, f"TDSTCS_{safe_m}.json")
        clicked = False
        for s in ["button:has-text('DOWNLOAD TDS AND TCS CREDIT RECEIVED JSON FILE')",
                  "button:has-text('JSON FILE')", "button:has-text('JSON')"]:
            try:
                page.locator(s).first.wait_for(state="visible", timeout=5000)
                page.locator(s).first.click()
                tds_log("    -> JSON button clicked"); clicked = True; break
            except Exception: continue
        if not clicked:
            tds_log("    ⚠ JSON button not found"); return None
        def find_dl_link():
            return page.evaluate("""() => {
                for (const a of document.querySelectorAll('a')) {
                    const t = (a.textContent||'').trim().toLowerCase();
                    if (t.includes('click here to download') ||
                        (t.includes('click here') && t.includes('file')))
                        return true;
                }
                return false;
            }""")
        link_appeared = False
        for _ in range(5):
            time.sleep(1)
            if find_dl_link(): link_appeared = True; break
        if not link_appeared:
            tds_log("    -> Link not ready — marked for sweep 2"); return None
        tds_log("    -> Link appeared — downloading...")
        dl_result = None
        for attempt in range(2):
            try:
                with page.expect_download(timeout=30000) as dl_info:
                    page.evaluate("""() => {
                        for (const a of document.querySelectorAll('a')) {
                            const t = (a.textContent||'').trim().toLowerCase();
                            if (t.includes('click here to download') ||
                                (t.includes('click here') && t.includes('file'))) {
                                a.removeAttribute('target'); a.click(); return;
                            }
                        }
                    }""")
                dl_result = _save_download(dl_info.value, json_path, mon_display, period)
                break
            except Exception as e:
                tds_log(f"    ⚠ Attempt {attempt+1} failed: {e}")
                if attempt == 0:
                    for s in ["button:has-text('DOWNLOAD TDS AND TCS CREDIT RECEIVED JSON FILE')",
                              "button:has-text('JSON FILE')", "button:has-text('JSON')"]:
                        try: page.locator(s).first.click(); break
                        except Exception: continue
                    for _ in range(5):
                        time.sleep(1)
                        if find_dl_link(): break
        return dl_result

    # ══════════════════════════════════════════════════════════════════════════
    # SWEEP 1 — navigate to each month, download immediately if ready
    # ══════════════════════════════════════════════════════════════════════════
    files_done = []
    pending    = []

    for i, month in enumerate(months):
        mon_display = month["display"]
        period      = month["period"]
        mon_name    = mon_display.split()[0]   # "April" from "April 2025"

        tds_set({"current_month": mon_display,
                 "progress": int((i / len(months)) * 100)})
        tds_log(f"\n[Sweep 1 · {i+1}/{len(months)}] {mon_display}")

        safe_m = mon_display.replace(" ", "_").replace("/", "_")
        if os.path.isfile(os.path.join(gstin_dir, f"TDSTCS_{safe_m}.json")):
            tds_log("  ✓ Already downloaded — skipping"); continue

        navigate_to_comptds(mon_name)
        status = click_to_offline_download()
        if not status:
            tds_log(f"  ⚠ No DOWNLOAD/PREPARE OFFLINE button — "
                    f"no TDS/TCS records for {mon_display}"); continue

        tds_log(f"  -> Status: {status}")
        if OFFLINE_PATH not in get_real_url():
            tds_log(f"  ⚠ Not on offline page"); pending.append(month); continue

        result = click_json_and_intercept(mon_display, period)
        if result:
            files_done.append(result)
            tds_set({"files": files_done.copy(), "done_months": len(files_done)})
        else:
            pending.append(month)

    # ══════════════════════════════════════════════════════════════════════════
    # SWEEP 2 — revisit pending months (up to 5 min each)
    # ══════════════════════════════════════════════════════════════════════════
    if pending:
        tds_log("=" * 50)
        tds_log(f"SWEEP 2: {len(pending)} pending month(s)...")

    for i, month in enumerate(pending):
        mon_display = month["display"]
        period      = month["period"]
        mon_name    = mon_display.split()[0]

        safe_m = mon_display.replace(" ", "_").replace("/", "_")
        if os.path.isfile(os.path.join(gstin_dir, f"TDSTCS_{safe_m}.json")): continue

        tds_set({"current_month": mon_display})
        tds_log(f"\n[Sweep 2 · {i+1}/{len(pending)}] {mon_display}")

        navigate_to_comptds(mon_name)
        status = click_to_offline_download()
        if not status or OFFLINE_PATH not in get_real_url():
            tds_log(f"  ⚠ Still no offline page for {mon_display}"); continue

        deadline = time.time() + 300
        result   = None
        attempt  = 0
        while time.time() < deadline and not result:
            attempt += 1
            result = click_json_and_intercept(mon_display, period)
            if result: break
            if attempt < 10:
                tds_log(f"    [{attempt}] Retrying in 30s...")
                time.sleep(30)
                navigate_to_comptds(mon_name)
                click_to_offline_download()

        if result:
            files_done.append(result)
            tds_set({"files": files_done.copy(), "done_months": len(files_done),
                     "progress": int(((len(months)-len(pending)+i+1)/len(months))*100)})
        else:
            tds_log(f"  ✗ Could not download {mon_display} after 5 min", "error")

    tds_set({"status": "done", "progress": 100})
    tds_log(f"\n✅ TDS/TCS complete — {len(files_done)}/{len(months)} months downloaded")
    _comb_log(f"  ✅ TDS: {len(files_done)}/{len(months)} months downloaded")
    return len(files_done)


# ── Combined download orchestrator ────────────────────────────────────────────

def _comb_wait_field(field, timeout_sec=300):
    """Wait for a field to be set in _comb_state (captcha answer, OTP)."""
    deadline = time.time() + timeout_sec
    while time.time() < deadline:
        time.sleep(0.4)
        with _comb_lock:
            val = _comb_state.get(field)
            if val:
                _comb_state[field] = None
                return val
    return None


def comb_do_browser_login(page, username, password):
    """
    Login using _comb_state for captcha/OTP so the Download All tab
    can display and accept the captcha — NOT g1_state.
    Mirrors g1_do_browser_login exactly but writes to _comb_state.
    """
    _comb_log("Opening GST login page...")
    try:
        page.goto(GST_LOGIN, wait_until="domcontentloaded", timeout=25000)
        try:
            page.wait_for_selector(
                "input#username, input[name='username'], input[placeholder*='username' i]",
                state="visible", timeout=8000)
        except Exception:
            pass
    except Exception as e:
        _comb_log(f"  ✗ Could not open login page: {e}")
        _comb_set({"status": "error", "error": str(e)}); return False

    for sel in ["input#username", "input[name='username']",
                "input[placeholder*='username' i]"]:
        try:
            page.locator(sel).first.fill(username)
            _comb_log("  ✓ Username filled"); break
        except Exception:
            continue

    for sel in ["input#user_pass", "input[name='user_pass']",
                "input[type='password']", "input[placeholder*='password' i]"]:
        try:
            page.locator(sel).first.fill(password)
            _comb_log("  ✓ Password filled"); break
        except Exception:
            continue

    time.sleep(0.3)

    # Capture captcha
    cap_img = None
    try:
        import base64 as _b64
        for sel in ["#imgCaptcha", "img[id*='aptcha' i]", "img[src*='captcha' i]",
                    "img[src*='kaptcha' i]", "img[alt*='captcha' i]",
                    ".captchaImage img", "img.captcha", "form img", "img"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=2000)
                png = loc.screenshot()
                if png and len(png) > 500:
                    cap_img = _b64.b64encode(png).decode()  # raw base64 only
                    _comb_log(f"  ✓ Captcha captured ({sel})")
                    break
            except Exception:
                continue
        if not cap_img:
            png = page.screenshot(full_page=False)
            cap_img = _b64.b64encode(png).decode()  # raw base64 only
            _comb_log("  ✓ Captcha captured via screenshot fallback")
    except Exception as e:
        _comb_log(f"  ⚠ Captcha capture error: {e}")

    # Store in _comb_state so Download All tab shows the captcha panel
    _comb_set({"status": "waiting_captcha", "captcha_image": cap_img})
    _comb_log("  ⏸ Waiting for captcha... (enter in Download All tab)")

    answer = _comb_wait_field("captcha_answer", timeout_sec=300)
    if not answer:
        _comb_log("  ✗ Captcha timeout")
        _comb_set({"status": "error", "error": "Captcha timeout"}); return False

    _comb_set({"status": "running", "captcha_image": None})

    for sel in ["input[placeholder*='Characters' i]", "input#captcha",
                "input[name='captcha']"]:
        try:
            page.locator(sel).first.fill(str(answer))
            _comb_log(f"  ✓ Captcha filled: {repr(answer)}"); break
        except Exception:
            continue

    for sel in ["button[type='submit']", "input[type='submit']",
                "button:has-text('LOGIN')"]:
        try:
            page.locator(sel).first.click()
            _comb_log("  ✓ Login submitted"); break
        except Exception:
            continue
    time.sleep(1.5)

    # OTP handling
    otp_needed = False
    for _ in range(8):
        time.sleep(1)
        try:
            otp_el = page.locator(
                "input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.wait_for(state="visible", timeout=1500)
            otp_needed = True; break
        except Exception:
            if check_login_success(page) is not False:
                break

    if otp_needed:
        _comb_set({"status": "waiting_captcha"})   # reuse captcha panel for OTP
        _comb_log("  ⏸ OTP required — enter in Download All captcha box...")
        otp = _comb_wait_field("captcha_answer", timeout_sec=180)
        if not otp:
            _comb_log("  ✗ OTP timeout")
            _comb_set({"status": "error", "error": "OTP timeout"}); return False
        try:
            otp_el = page.locator(
                "input[placeholder*='OTP' i], input[id*='otp' i]").first
            otp_el.click(); time.sleep(0.2)
            otp_el.fill(str(otp))
            page.locator("button[type='submit'],input[type='submit']").first.click()
            _comb_log("  ✓ OTP submitted"); time.sleep(1)
        except Exception as e:
            _comb_log(f"  ✗ OTP error: {e}"); return False

    # Captcha retry loop (up to 3 attempts)
    for attempt in range(3):
        result = check_login_success(page)
        if result is not False:
            break
        remaining = 2 - attempt
        if remaining == 0:
            _comb_set({"status": "error",
                        "error": "Login failed after 3 captcha attempts"})
            _comb_log("  ✗ Login failed after 3 attempts")
            return False
        _comb_log(f"  ✗ Login failed — re-capturing captcha "
                   f"(attempt {attempt+2}/3)...")
        time.sleep(2)
        new_cap = None
        try:
            for sel in ["#imgCaptcha", "img[id*='aptcha' i]",
                        "img[src*='captcha' i]", "form img", "img"]:
                try:
                    loc = page.locator(sel).first
                    loc.wait_for(state="visible", timeout=2000)
                    png = loc.screenshot()
                    if png and len(png) > 500:
                        new_cap = _b64.b64encode(png).decode()
                        break
                except Exception:
                    continue
            if not new_cap:
                png = page.screenshot(full_page=False)
                new_cap = _b64.b64encode(png).decode()
        except Exception:
            pass
        _comb_set({"status": "running",  "captcha_image": None})
        time.sleep(0.1)
        _comb_set({"status": "waiting_captcha", "captcha_image": new_cap})
        _comb_log(f"  ⏸ Enter new captcha (attempt {attempt+2}/3)...")
        new_ans = _comb_wait_field("captcha_answer", timeout_sec=300)
        if not new_ans:
            _comb_set({"status": "error", "error": "Captcha retry timeout"})
            return False
        _comb_set({"status": "running", "captcha_image": None})
        for sel in ["input[placeholder*='Characters' i]", "input#captcha"]:
            try:
                loc = page.locator(sel).first
                loc.wait_for(state="visible", timeout=3000)
                loc.fill(""); loc.fill(str(new_ans))
                break
            except Exception:
                continue
        for sel in ["button[type='submit']", "button:has-text('LOGIN')"]:
            try:
                page.locator(sel).first.click(); break
            except Exception:
                continue
        time.sleep(2)

    _comb_log("  ✅ Logged in successfully")
    g2a_dismiss_popup(page)
    return True


def combined_download_worker(gstin, fy, username, password, modules):
    """
    Thin wrapper around _combined_download_worker_impl that guarantees
    ANY exception — including ones raised before browser launch even
    starts — gets logged and reflected in _comb_state, instead of
    silently killing the thread with nothing recorded anywhere.

    This matters specifically for frozen/windowed PyInstaller builds:
    Python's default behavior on an uncaught thread exception is to
    print a traceback to sys.stderr via threading.excepthook — but in
    a windowed build (console=False) there is no real stderr for that
    to go to, so the traceback vanishes completely. From the outside
    this looks exactly like "nothing is happening", indistinguishable
    from a genuine hang, with zero information to diagnose it by.
    """
    try:
        _combined_download_worker_impl(gstin, fy, username, password, modules)
    except Exception as fatal:
        import traceback as _tb
        try:
            _comb_log(f"✗ FATAL (uncaught): {fatal}", "error")
            _comb_log(_tb.format_exc(), "error")
        except Exception:
            pass
        try:
            _comb_set({"status": "error", "error": f"Uncaught: {fatal}"})
        except Exception:
            pass
        log.error(f"[COMB] Uncaught worker exception: {fatal}\n{_tb.format_exc()}")


def _combined_download_worker_impl(gstin, fy, username, password, modules):
    """
    Single-session downloader: ONE login, ONE session activation,
    then runs each selected module in sequence on the same browser page.

    Modules run in order: g1 → g2a → g2b → g3b → tds
    Each module updates its own state dict (g1_state, g2a_state, etc.)
    so the existing individual module UI panels also reflect progress.
    """
    _comb_log("🚀 Combined Download — single browser session")
    _comb_log(f"   GSTIN: {gstin}  FY: {fy}  Modules: {modules}")
    _comb_set({"gstin": gstin, "fy": fy, "modules_done": [], "modules_failed": []})


    # ── Browser launch ──────────────────────────────────────────────────────────
    profile_dir = os.path.join(PATHS.profiles_dir, "gst_profile_combined")
    os.makedirs(profile_dir, exist_ok=True)
    lock_file = os.path.join(profile_dir, "SingletonLock")
    if os.path.exists(lock_file):
        try: os.remove(lock_file); _comb_log("  ✓ Removed stale SingletonLock")
        except Exception: pass

    from playwright.sync_api import sync_playwright
    # This starts Playwright's own Node.js driver subprocess — a
    # SEPARATE thing from the Chromium browser binary. If this hangs
    # or fails silently, nothing after it (including the "Launching
    # browser..." log line) will ever run, which looks identical from
    # the outside to "nothing happening" with zero explanation. Logged
    # on both sides so a stuck run is now distinguishable from one that
    # never started at all.
    _comb_log("Starting Playwright driver...")
    try:
        pw = sync_playwright().start()
    except Exception as e:
        _comb_log(f"  ✗ Playwright driver failed to start: {e}", "error")
        _comb_set({"status": "error", "error": f"Playwright driver failed to start: {e}"})
        return
    _comb_log("  ✓ Playwright driver started")
    try:
        _comb_log("🌐 Launching browser...")
        try:
            context = pw.chromium.launch_persistent_context(
                user_data_dir=profile_dir,
                headless=False, slow_mo=40,
                args=["--disable-blink-features=AutomationControlled",
                      "--no-sandbox", "--start-maximized"],
                no_viewport=True, accept_downloads=True,
            )
        except Exception as e:
            _comb_log(f"  ✗ Browser launch error: {e}")
            _comb_set({"status": "error", "error": str(e)}); return

        page = context.pages[0] if context.pages else context.new_page()
        page.add_init_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})")
        g2a_install_interceptors(page)

        # ── Single login via comb_do_browser_login ─────────────────────────────
        # Uses _comb_state for captcha/OTP — shows in the Download All tab UI
        _comb_log("🔐 Logging in (single shared session)...")
        _comb_set({"status": "login"})
        if not comb_do_browser_login(page, username, password):
            _comb_set({"status": "error", "error": "Login failed"})
            context.close(); return

        # ── Single session activation ───────────────────────────────────────────
        _comb_log("🔄 Activating session on Returns portal (once for all modules)...")
        _comb_set({"status": "running", "current_module": "session"})
        ok, page = g2a_activate_session(page, context)
        if not ok:
            _comb_set({"status": "error",
                        "error": "Could not reach Returns portal"})
            context.close(); return

        g2a_install_interceptors(page)
        _comb_log(f"  ✓ Session active: {page.url[:70]}")
        _comb_log("━" * 55)

        # ── Run modules in order ────────────────────────────────────────────────
        MODULE_ORDER = ["g1", "g2a", "g2b", "g3b", "tds"]
        run_list     = [m for m in MODULE_ORDER if m in modules]
        total        = len(run_list)
        done_count   = 0

        # G1 runs in two phases — no 300s idle wait in combined session
        # Phase 1: G1 trigger pass (sweep 1 only, collect ready, log pending)
        # Phase 2: G2A → G2B → G3B → TDS (portal generates G1 files during this)
        # Phase 3: G1 harvest pass (collect remaining pending G1 months)
        g1_pending_months = []

        for mod in run_list:
            # Check stop flag before starting each module. _comb_log()
            # itself acquires _comb_lock — calling it while still
            # holding the lock here would self-deadlock this thread
            # forever (Python's threading.Lock is not reentrant), so
            # the log call happens AFTER the lock is released instead.
            with _comb_lock:
                stop_now = _comb_state.get("stop_requested", False)
            if stop_now:
                _comb_log("🛑 Stop confirmed — aborting remaining modules")
                break

            _comb_log(f"\n{'━'*55}")
            _comb_log(f"📥 [{done_count+1}/{total}] Starting {mod.upper()}...")
            _comb_set({"current_module": mod,
                        "progress": int(done_count / total * 100)})
            try:
                if mod == "g1":
                    # Trigger pass: GENERATE all months, collect ready, return pending
                    _, g1_pending_months = _comb_run_g1(page, gstin, fy)
                elif mod == "g2a": _comb_run_g2a(page, gstin, fy)
                elif mod == "g2b": _comb_run_g2b(page, gstin, fy)
                elif mod == "g3b": _comb_run_g3b(page, context, gstin, fy)
                elif mod == "tds": _comb_run_tds(page, context, gstin, fy)
                with _comb_lock:
                    _comb_state["modules_done"].append(mod)
            except Exception as e:
                import traceback as _tb
                _comb_log(f"  ✗ {mod.upper()} failed: {e}")
                _comb_log(_tb.format_exc())
                with _comb_lock:
                    _comb_state["modules_failed"].append(mod)
            done_count += 1

        # G1 Harvest: collect pending months (portal has had 20-40 min to generate)
        with _comb_lock:
            _stopped = _comb_state.get("stop_requested", False)
        if g1_pending_months and not _stopped:
            _comb_log(f"\n{'━'*55}")
            _comb_log(f"🔁 G1 HARVEST — {len(g1_pending_months)} pending month(s)")
            _comb_set({"current_module": "g1-harvest"})
            try:
                _comb_run_g1_harvest(page, gstin, fy, g1_pending_months)
            except Exception as e:
                import traceback as _tb
                _comb_log(f"  ✗ G1 harvest failed: {e}")
                _comb_log(_tb.format_exc())

        _comb_log(f"\n{'━'*55}")

        with _comb_lock:
            was_stopped = _comb_state.get("stop_requested", False)
        if was_stopped:
            _comb_log(f"🛑 Session stopped — {done_count}/{total} modules completed")
            _comb_set({"status": "stopped",
                        "progress": int(done_count / max(total, 1) * 100),
                        "current_module": None, "stop_requested": False})
        else:
            _comb_log(f"✅ Combined download complete — {done_count}/{total} modules")
            _comb_set({"status": "done", "progress": 100, "current_module": None})

        try: context.close(); _comb_log("🔒 Browser closed")
        except Exception as ce: _comb_log(f"  ⚠ Browser close: {ce}")

    except Exception as fatal:
        import traceback as _tb
        _comb_set({"status": "error", "error": str(fatal)})
        _comb_log(f"✗ Fatal error: {fatal}")
        _comb_log(_tb.format_exc())
    finally:
        try: pw.stop()
        except Exception: pass

    time.sleep(8)
    with _comb_lock:
        if _comb_state.get("status") == "done":
            _comb_state["status"] = "idle"
    _comb_log("🔁 Combined download reset to idle")


# ── Combined Download Flask Routes ────────────────────────────────────────────

@app.route("/combined/start", methods=["POST"])
def combined_start():
    data     = request.get_json(force=True) or {}
    gstin    = (data.get("gstin")    or "").strip().upper()
    fy       = (data.get("fy")       or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    modules  = data.get("modules", ["g1","g2a","g2b","g3b","tds"])
    if not all([gstin, fy, username, password]):
        return jsonify({"error": "gstin, fy, username and password required"}), 400
    if not modules:
        return jsonify({"error": "No modules selected"}), 400

    # Claim the "running" state atomically, in the SAME lock acquisition
    # as the "already running?" check — not in a separate one afterward.
    # The previous version checked status under the lock, released it,
    # then did request-body validation, THEN re-acquired the lock to
    # claim "running" — leaving a real window where two near-simultaneous
    # POSTs (e.g. a double-click on Download All before the button
    # visibly disables) could both pass the check and both spawn a
    # worker. Two Chromium instances logging into the same GST account
    # and downloading GSTR1/2A/2B/3B/TDS concurrently is exactly what
    # "Download All not working properly" looks like: interleaved,
    # corrupted progress state, doubled log lines, and files from two
    # sessions colliding in the same download folder.
    with _comb_lock:
        if _comb_state["status"] in ("login","running","waiting_captcha","downloading"):
            return jsonify({"error": "Already running — reset first"}), 409
        _comb_state.update({
            "status": "running", "log": [], "error": None,
            "progress": 0, "current_module": None,
            "modules_done": [], "modules_failed": [],
            "captcha_image": None, "captcha_answer": None,
            "gstin": gstin, "fy": fy,
        })

    threading.Thread(
        target=combined_download_worker,
        args=(gstin, fy, username, password, modules),
        daemon=True
    ).start()
    return jsonify({"ok": True, "modules": modules, "gstin": gstin, "fy": fy})


@app.route("/combined/state")
def combined_state_route():
    with _comb_lock:
        return jsonify({
            "status":         _comb_state["status"],
            "log":            _comb_state["log"][-60:],
            "progress":       _comb_state["progress"],
            "current_module": _comb_state["current_module"],
            "modules_done":   _comb_state["modules_done"],
            "modules_failed": _comb_state["modules_failed"],
            "error":          _comb_state["error"],
            "has_captcha":    _comb_state["captcha_image"] is not None,
            "stop_requested": _comb_state.get("stop_requested", False),
            "gstin":          _comb_state["gstin"],
            "fy":             _comb_state["fy"],
        })


@app.route("/combined/submit_captcha", methods=["POST"])
def combined_submit_captcha():
    data   = request.get_json(force=True) or {}
    answer = (data.get("captcha") or "").strip()
    if not answer:
        return jsonify({"error": "captcha required"}), 400
    with _comb_lock:
        _comb_state["captcha_answer"] = answer
        _comb_state["captcha_image"]  = None
    return jsonify({"ok": True})


@app.route("/combined/stop", methods=["POST"])
def combined_stop():
    """Signal the worker to stop after the current module finishes."""
    with _comb_lock:
        status = _comb_state.get("status", "idle")
        if status in ("running", "login", "waiting_captcha"):
            _comb_state["stop_requested"] = True
            _comb_state["status"] = "stopping"
            _comb_state["log"].append({
                "msg": "🛑 Stop requested — will finish current module then stop",
                "level": "warn", "ts": __import__("time").strftime("%H:%M:%S")
            })
    return jsonify({"ok": True, "status": _comb_state.get("status")})


@app.route("/combined/reset", methods=["POST"])
def combined_reset():
    _comb_set({
        "status": "idle", "log": [], "error": None, "progress": 0,
        "current_module": None, "modules_done": [], "modules_failed": [],
        "captcha_image": None, "captcha_answer": None,
        "gstin": None, "fy": None, "stop_requested": False,
    })
    return jsonify({"ok": True})


@app.route("/combined/captcha_image")
def combined_captcha_image():
    with _comb_lock:
        img = _comb_state.get("captcha_image")
    if not img:
        return jsonify({"error": "no captcha available"}), 404
    return jsonify({"image": img})


@app.route("/")
def serve_ui():
    """Serve ClientLedger-India.html — no-cache so browser always gets fresh copy."""
    candidates = [
        os.path.join(_APP_DIR, "templates", "ClientLedger-India.html"),
        os.path.join(_APP_DIR, "ClientLedger-India.html"),
    ]
    html_path = next((p for p in candidates if os.path.isfile(p)), None)
    if html_path:
        resp = send_file(html_path)
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
        resp.headers["Pragma"]        = "no-cache"
        resp.headers["Expires"]       = "0"
        return resp
    return "<h2>ClientLedger-India.html not found</h2>", 404

@app.route("/app")
def serve_ui_app():
    return serve_ui()


# ── Local data-folder info (read-only; shown in Settings) ──────────────────
@app.route("/config/data_dir")
def config_data_dir():
    return jsonify({
        "ok": True,
        "base": PATHS.base,
        "database": PATHS.db_file,
        "gstr1": PATHS.gstr1_dir,
        "gstr2a": PATHS.gstr2a_dir,
        "gstr2b": PATHS.gstr2b_dir,
        "gstr3b": PATHS.gstr3b_dir,
        "tdstcs": PATHS.tdstcs_dir,
    })


# ── SQLite-backed client store (replaces browser IndexedDB) ────────────────
@app.route("/db/clients", methods=["GET"])
def db_clients_list():
    return jsonify({"ok": True, "clients": dbstore.clients_get_all()})

@app.route("/db/clients/<int:rec_id>", methods=["GET"])
def db_clients_get(rec_id):
    rec = dbstore.clients_get(rec_id)
    if rec is None:
        return jsonify({"ok": False, "error": "not found"}), 404
    return jsonify({"ok": True, "client": rec})

@app.route("/db/clients", methods=["POST"])
def db_clients_add():
    rec = request.get_json(force=True, silent=True) or {}
    try:
        new_id = dbstore.clients_add(rec)
    except dbstore.DuplicateError as e:
        log.error(f"[CLIENT-ADD] Duplicate conflict: {e}")
        return jsonify({"ok": False, "error": f"A client with this PAN, Aadhaar or email already exists. ({e})"}), 409
    except Exception as e:
        import traceback as _tb
        log.error(f"[CLIENT-ADD] Failed to add client: {e}\n{_tb.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True, "id": new_id})

@app.route("/db/clients/<int:rec_id>", methods=["PUT"])
def db_clients_update(rec_id):
    rec = request.get_json(force=True, silent=True) or {}
    rec["id"] = rec_id
    try:
        dbstore.clients_update(rec)
    except dbstore.DuplicateError as e:
        # Previously this discarded the actual detail and always showed
        # the same generic message regardless of which field conflicted
        # or why — impossible to diagnose from the outside. Now includes
        # SQLite's own message (e.g. "UNIQUE constraint failed:
        # clients.pan") so it's clear exactly which field is the issue.
        log.error(f"[CLIENT-UPDATE] Duplicate conflict for client {rec_id}: {e}")
        return jsonify({"ok": False, "error": f"A client with this PAN, Aadhaar or email already exists. ({e})"}), 409
    except Exception as e:
        import traceback as _tb
        log.error(f"[CLIENT-UPDATE] Failed to update client {rec_id}: {e}\n{_tb.format_exc()}")
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True, "id": rec_id})

@app.route("/db/clients/<int:rec_id>", methods=["DELETE"])
def db_clients_delete(rec_id):
    dbstore.clients_delete(rec_id)
    return jsonify({"ok": True})


# ── SQLite-backed GSTIN directory store (replaces browser IndexedDB) ───────
@app.route("/db/gstin_directory", methods=["GET"])
def db_gdir_list():
    return jsonify({"ok": True, "records": dbstore.gdir_get_all()})

@app.route("/db/gstin_directory", methods=["POST"])
def db_gdir_put():
    rec = request.get_json(force=True, silent=True) or {}
    try:
        dbstore.gdir_put(rec)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True})

@app.route("/db/gstin_directory/batch", methods=["POST"])
def db_gdir_put_batch():
    body = request.get_json(force=True, silent=True) or {}
    records = body.get("records") or []
    try:
        dbstore.gdir_put_batch(records)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True, "count": len(records)})

@app.route("/db/gstin_directory/clear", methods=["POST"])
def db_gdir_clear():
    dbstore.gdir_clear()
    return jsonify({"ok": True})

# ── Start ─────────────────────────────────────────────────────
if __name__ == "__main__":
    import webbrowser
    print()
    print("  ✅ Open ClientLedger in your browser:")
    print("  http://localhost:8765")
    print()
    print("  Keep this window open while using ClientLedger.")
    print("=" * 56)
    # Auto-open browser after 1s
    threading.Timer(1.0, lambda: webbrowser.open("http://localhost:8765")).start()
    app.run(host="127.0.0.1", port=8765, debug=False, threaded=True)
