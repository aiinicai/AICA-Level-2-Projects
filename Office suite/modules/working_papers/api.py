from typing import List, Optional
import io
import re
from difflib import SequenceMatcher
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

try:
    from .database import get_db, init_db, DB_PATH
    from .models import FDRecord, AS26Entry, WPEntity
    from .calculator import FDCalculator
    from .template import generate_fd_template
    from .pdf_parser import parse_fd_pdf
    from .as26_parser import parse_26as_pdf, parse_26as_content
except ImportError:
    from database import get_db, init_db, DB_PATH
    from models import FDRecord, AS26Entry, WPEntity
    from calculator import FDCalculator
    from template import generate_fd_template
    from pdf_parser import parse_fd_pdf
    from as26_parser import parse_26as_pdf, parse_26as_content

# Ensure DB initialized
init_db()

router = APIRouter()

# --- Entity Pydantic Schemas ---
class CreateEntitySchema(BaseModel):
    name: str
    status: Optional[str] = "Active"

class UpdateEntitySchema(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None


# ==========================================
# 0. STANDALONE WORKING PAPER ENTITY CRUD
# ==========================================

@router.get("/entities")
def get_entities(db: Session = Depends(get_db)):
    """Fetch all standalone Working Paper entities."""
    entities = db.query(WPEntity).order_by(WPEntity.id.asc()).all()
    if not entities:
        e1 = WPEntity(id=1, name="Acme Enterprises Private Limited", status="Active")
        e2 = WPEntity(id=2, name="Tata Consultancy Services Limited", status="Active")
        db.add(e1)
        db.add(e2)
        db.commit()
        entities = db.query(WPEntity).order_by(WPEntity.id.asc()).all()

    return {
        "status": "success",
        "count": len(entities),
        "entities": [e.to_dict() for e in entities],
        "clients": [e.to_dict() for e in entities]
    }

@router.get("/entities/{entity_id}")
def get_entity(entity_id: int, db: Session = Depends(get_db)):
    """Fetch single Working Paper entity details."""
    entity = db.query(WPEntity).filter(WPEntity.id == entity_id).first()
    if not entity:
        if entity_id == 1:
            entity = WPEntity(id=1, name="Acme Enterprises Private Limited", status="Active")
            db.add(entity)
            db.commit()
            db.refresh(entity)
        elif entity_id == 2:
            entity = WPEntity(id=2, name="Tata Consultancy Services Limited", status="Active")
            db.add(entity)
            db.commit()
            db.refresh(entity)
        else:
            raise HTTPException(status_code=404, detail="Entity not found")

    return {
        "status": "success",
        "entity": entity.to_dict(),
        "client": entity.to_dict()
    }

@router.post("/entities")
def create_entity(payload: CreateEntitySchema, db: Session = Depends(get_db)):
    """Create a new Working Paper entity."""
    entity = WPEntity(name=payload.name.strip(), status=payload.status or "Active")
    db.add(entity)
    db.commit()
    db.refresh(entity)
    return {"status": "success", "entity": entity.to_dict(), "client": entity.to_dict()}

@router.put("/entities/{entity_id}")
def update_entity(entity_id: int, payload: UpdateEntitySchema, db: Session = Depends(get_db)):
    """Rename / edit a Working Paper entity."""
    entity = db.query(WPEntity).filter(WPEntity.id == entity_id).first()
    if not entity:
        raise HTTPException(status_code=404, detail="Entity not found")

    if payload.name:
        entity.name = payload.name.strip()
    if payload.status:
        entity.status = payload.status

    db.commit()
    db.refresh(entity)
    return {"status": "success", "entity": entity.to_dict(), "client": entity.to_dict()}

@router.delete("/entities/{entity_id}")
def delete_entity(entity_id: int, db: Session = Depends(get_db)):
    """Delete a Working Paper entity AND its associated FD & 26AS records without affecting other modules."""
    entity = db.query(WPEntity).filter(WPEntity.id == entity_id).first()
    if entity:
        db.delete(entity)

    # Purge associated Working Paper data for this entity ONLY
    db.query(FDRecord).filter(FDRecord.client_id == entity_id).delete()
    db.query(AS26Entry).filter(AS26Entry.client_id == entity_id).delete()

    db.commit()
    return {"status": "success", "deleted_id": entity_id}

# --- Helper for Bank Name Fuzzy Matching ---
def normalize_bank_name(name: str) -> str:
    if not name: return ""
    clean = name.upper()
    clean = re.sub(r"\b(LIMITED|LTD|CO\-OPERATIVE|BANK|INCORPORATED|INC|CORP|CORPORATION|PLC)\b", "", clean)
    clean = re.sub(r"[^\w\s]", "", clean)
    return " ".join(clean.split())

def fuzzy_bank_match(name1: str, name2: str) -> float:
    n1 = normalize_bank_name(name1)
    n2 = normalize_bank_name(name2)
    if not n1 or not n2: return 0.0
    if n1 in n2 or n2 in n1: return 0.9
    return SequenceMatcher(None, n1, n2).ratio()


# --- Pydantic Schemas ---
class CreateFDRecordSchema(BaseModel):
    client_id: int
    financial_year: str = "2024-25"
    bank_name: str
    fd_account_number: str
    principal_amount: float
    date_of_issue: str
    date_of_maturity: str
    interest_rate: float
    compounding_frequency: Optional[str] = "Quarterly"
    opening_accrued_interest: Optional[float] = 0.0
    tds_deducted: Optional[float] = 0.0
    status: Optional[str] = "Active"

class UpdateFDRecordSchema(BaseModel):
    bank_name: Optional[str] = None
    fd_account_number: Optional[str] = None
    principal_amount: Optional[float] = None
    date_of_issue: Optional[str] = None
    date_of_maturity: Optional[str] = None
    interest_rate: Optional[float] = None
    compounding_frequency: Optional[str] = None
    opening_accrued_interest: Optional[float] = None
    tds_deducted: Optional[float] = None
    status: Optional[str] = None

class RollForwardSchema(BaseModel):
    client_id: int
    target_fy: str


# ==========================================
# ROLL-FORWARD ENGINE FUNCTION
# ==========================================

def execute_roll_forward(client_id: int, target_fy: str, db: Session) -> int:
    """
    Rolls forward active FDs from prior financial year into target_fy.
    Deduplicates records and automatically purges orphaned roll-forward records if deleted in PY.
    """
    prior_fy = FDCalculator.get_prior_fy(target_fy)
    
    # Query prior FY records for client
    py_records = db.query(FDRecord).filter(
        FDRecord.client_id == client_id,
        FDRecord.financial_year == prior_fy
    ).all()

    valid_py_ids = {r.id for r in py_records}

    # Query target FY existing records
    target_records = db.query(FDRecord).filter(
        FDRecord.client_id == client_id,
        FDRecord.financial_year == target_fy
    ).all()

    # 1. PURGE ORPHANED ROLL-FORWARD RECORDS (If deleted in PY)
    purged_ids = set()
    for tr in target_records:
        if tr.is_roll_forward and tr.py_record_id is not None:
            if tr.py_record_id not in valid_py_ids:
                db.delete(tr)
                purged_ids.add(tr.id)

    if purged_ids:
        db.commit()
        target_records = db.query(FDRecord).filter(
            FDRecord.client_id == client_id,
            FDRecord.financial_year == target_fy
        ).all()

    if not py_records:
        return 0

    # 2. DEDUPLICATION MATCHING KEYS
    existing_py_ids = {r.py_record_id for r in target_records if r.py_record_id is not None}
    existing_keys = {(r.bank_name.lower().strip(), r.fd_account_number.lower().strip()) for r in target_records}

    calc = FDCalculator(financial_year=target_fy)
    rolled_count = 0

    for py_rec in py_records:
        if py_rec.id in existing_py_ids:
            continue
        
        key = (py_rec.bank_name.lower().strip(), py_rec.fd_account_number.lower().strip())
        if key in existing_keys:
            continue

        if py_rec.closing_principal > 0 or py_rec.status != "Matured":
            res = calc.process_fd(
                bank_name=py_rec.bank_name,
                fd_account_number=py_rec.fd_account_number,
                principal_amount=py_rec.principal_amount,
                date_of_issue=py_rec.date_of_issue,
                date_of_maturity=py_rec.date_of_maturity,
                interest_rate=py_rec.interest_rate,
                compounding_frequency=py_rec.compounding_frequency,
                opening_accrued_interest=py_rec.closing_accrued_interest,
                tds_deducted=0.0,
                status="Active",
                is_roll_forward=True,
                opening_principal_override=py_rec.closing_principal
            )

            new_rec = FDRecord(
                client_id=client_id,
                financial_year=target_fy,
                bank_name=res["bank_name"],
                fd_account_number=res["fd_account_number"],
                principal_amount=res["principal_amount"],
                date_of_issue=res["date_of_issue"],
                date_of_maturity=res["date_of_maturity"],
                interest_rate=res["interest_rate"],
                compounding_frequency=res["compounding_frequency"],
                opening_accrued_interest=res["opening_accrued_interest"],
                tds_deducted=0.0,
                status=res["status"],
                opening_principal=res["opening_principal"],
                created_principal=0.0,
                matured_principal=res["matured_principal"],
                settled_accrued_interest=res["settled_accrued_interest"],
                is_roll_forward=True,
                py_record_id=py_rec.id,
                original_maturity_days=res["original_maturity_days"],
                remaining_maturity_days=res["remaining_maturity_days"],
                interest_income=res["interest_income"],
                closing_accrued_interest=res["closing_accrued_interest"],
                closing_principal=res["closing_principal"],
                closing_total_balance=res["closing_total_balance"],
                classification_class=res["classification_class"],
                classification_label=res["classification_label"]
            )
            db.add(new_rec)
            existing_keys.add(key)
            rolled_count += 1

    if rolled_count > 0:
        db.commit()
    return rolled_count


def cascade_delete_fd_record(record_id: int, db: Session):
    """Recursively deletes downstream rolled-forward records across future financial years."""
    child_records = db.query(FDRecord).filter(FDRecord.py_record_id == record_id).all()
    for child in child_records:
        cascade_delete_fd_record(child.id, db)
        db.delete(child)


# ==========================================
# 1. IMPORTS & FORM 26AS UPLOADER
# ==========================================

@router.get("/template")
def download_template():
    """Download standard Excel import template for FDs with DataValidation dropdowns."""
    buf = generate_fd_template()
    filename = "FD_Import_Template.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.post("/upload_excel")
async def upload_excel(
    client_id: int = Form(...),
    financial_year: str = Form("2024-25"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Imports FD records from Excel spreadsheet and processes statutory classification."""
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    try:
        calc = FDCalculator(financial_year=financial_year)
        processed_count = 0
        new_records = []

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx < 3: continue
            if not row or not any(row): continue
            
            row_str = [str(c).strip() if c is not None else "" for c in row]
            if "Bank Name" in row_str[0] or "Principal Amount" in row_str: continue
            
            try:
                bank_name = str(row[0]).strip() if len(row) > 0 and row[0] else "Bank"
                fd_acc = str(row[1]).strip() if len(row) > 1 and row[1] else f"FD-{row_idx}"
                principal = float(row[2]) if len(row) > 2 and row[2] else 0.0
                issue_date = str(row[3]).strip() if len(row) > 3 and row[3] else "2024-01-01"
                maturity_date = str(row[4]).strip() if len(row) > 4 and row[4] else "2025-01-01"
                rate = float(row[5]) if len(row) > 5 and row[5] else 7.0
                freq = str(row[6]).strip() if len(row) > 6 and row[6] else "Quarterly"
                opening_accrued = float(row[7]) if len(row) > 7 and row[7] else 0.0
                tds = float(row[8]) if len(row) > 8 and row[8] else 0.0
                status = str(row[9]).strip() if len(row) > 9 and row[9] else "Active"

                res = calc.process_fd(
                    bank_name=bank_name,
                    fd_account_number=fd_acc,
                    principal_amount=principal,
                    date_of_issue=issue_date,
                    date_of_maturity=maturity_date,
                    interest_rate=rate,
                    compounding_frequency=freq,
                    opening_accrued_interest=opening_accrued,
                    tds_deducted=tds,
                    status=status
                )

                record = FDRecord(
                    client_id=client_id,
                    financial_year=financial_year,
                    bank_name=res["bank_name"],
                    fd_account_number=res["fd_account_number"],
                    principal_amount=res["principal_amount"],
                    date_of_issue=res["date_of_issue"],
                    date_of_maturity=res["date_of_maturity"],
                    interest_rate=res["interest_rate"],
                    compounding_frequency=res["compounding_frequency"],
                    opening_accrued_interest=res["opening_accrued_interest"],
                    tds_deducted=res["tds_deducted"],
                    status=res["status"],
                    opening_principal=res["opening_principal"],
                    created_principal=res["created_principal"],
                    matured_principal=res["matured_principal"],
                    settled_accrued_interest=res["settled_accrued_interest"],
                    original_maturity_days=res["original_maturity_days"],
                    remaining_maturity_days=res["remaining_maturity_days"],
                    interest_income=res["interest_income"],
                    closing_accrued_interest=res["closing_accrued_interest"],
                    closing_principal=res["closing_principal"],
                    closing_total_balance=res["closing_total_balance"],
                    classification_class=res["classification_class"],
                    classification_label=res["classification_label"]
                )
                db.add(record)
                new_records.append(record)
                processed_count += 1
            except Exception as err:
                print(f"Skipping row {row_idx} due to error: {err}")

        db.commit()
        return {
            "status": "success",
            "processed_count": processed_count,
            "records": [r.to_dict() for r in new_records]
        }
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error importing Excel schedule: {str(exc)}")

@router.post("/upload_pdf")
async def upload_pdf(
    client_id: int = Form(...),
    financial_year: str = Form("2024-25"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Imports FD advice/receipt PDF and processes statutory classification."""
    try:
        contents = await file.read()
        parsed_data = parse_fd_pdf(contents)

        calc = FDCalculator(financial_year=financial_year)
        res = calc.process_fd(
            bank_name=parsed_data["bank_name"],
            fd_account_number=parsed_data["fd_account_number"],
            principal_amount=parsed_data["principal_amount"],
            date_of_issue=parsed_data["date_of_issue"],
            date_of_maturity=parsed_data["date_of_maturity"],
            interest_rate=parsed_data["interest_rate"],
            compounding_frequency=parsed_data["compounding_frequency"],
            opening_accrued_interest=parsed_data["opening_accrued_interest"],
            tds_deducted=parsed_data["tds_deducted"],
            status=parsed_data["status"]
        )

        record = FDRecord(
            client_id=client_id,
            financial_year=financial_year,
            bank_name=res["bank_name"],
            fd_account_number=res["fd_account_number"],
            principal_amount=res["principal_amount"],
            date_of_issue=res["date_of_issue"],
            date_of_maturity=res["date_of_maturity"],
            interest_rate=res["interest_rate"],
            compounding_frequency=res["compounding_frequency"],
            opening_accrued_interest=res["opening_accrued_interest"],
            tds_deducted=res["tds_deducted"],
            status=res["status"],
            opening_principal=res["opening_principal"],
            created_principal=res["created_principal"],
            matured_principal=res["matured_principal"],
            settled_accrued_interest=res["settled_accrued_interest"],
            original_maturity_days=res["original_maturity_days"],
            remaining_maturity_days=res["remaining_maturity_days"],
            interest_income=res["interest_income"],
            closing_accrued_interest=res["closing_accrued_interest"],
            closing_principal=res["closing_principal"],
            closing_total_balance=res["closing_total_balance"],
            classification_class=res["classification_class"],
            classification_label=res["classification_label"]
        )
        db.add(record)
        db.commit()
        db.refresh(record)

        return {"status": "success", "record": record.to_dict()}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error parsing FD Receipt PDF: {str(exc)}")

@router.post("/upload_26as")
async def upload_26as(
    client_id: int = Form(...),
    financial_year: str = Form("2024-25"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Imports Form 26AS PDF or CSV file, parses Section 194A entries, and stores in database."""
    try:
        contents = await file.read()
        entries = parse_26as_content(contents, file.filename or "")

        if not entries:
            raise HTTPException(status_code=400, detail="No Section 194A entries found in Form 26AS file.")

        # Purge existing 26AS entries for client & FY
        db.query(AS26Entry).filter(
            AS26Entry.client_id == client_id,
            AS26Entry.financial_year == financial_year
        ).delete()

        new_entries = []
        for e in entries:
            obj = AS26Entry(
                client_id=client_id,
                financial_year=financial_year,
                deductor_name=e["deductor_name"],
                tan=e["tan"],
                section=e["section"],
                amount_paid=e["amount_paid"],
                tds_deducted=e["tds_deducted"]
            )
            db.add(obj)
            new_entries.append(obj)

        db.commit()
        return {
            "status": "success",
            "parsed_count": len(new_entries),
            "entries": [e.to_dict() for e in new_entries]
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Error parsing Form 26AS PDF: {str(exc)}")

@router.delete("/as26_all")
def delete_all_as26_entries(
    client_id: int = Query(...),
    financial_year: str = Query("2024-25"),
    db: Session = Depends(get_db)
):
    """Deletes all imported Form 26AS entries for a client and financial year."""
    deleted_count = db.query(AS26Entry).filter(
        AS26Entry.client_id == client_id,
        AS26Entry.financial_year == financial_year
    ).delete()

    db.commit()
    return {
        "status": "success",
        "client_id": client_id,
        "financial_year": financial_year,
        "deleted_count": deleted_count
    }


# ==========================================
# 2. CRUD, BULK DELETE & FUZZY RECONCILIATION
# ==========================================

@router.get("/records")
def get_records(
    client_id: int = Query(...),
    financial_year: str = Query("2024-25"),
    db: Session = Depends(get_db)
):
    """Fetch FD records for client_id and financial_year (with auto roll-forward check)."""
    execute_roll_forward(client_id, financial_year, db)

    records = db.query(FDRecord).filter(
        FDRecord.client_id == client_id,
        FDRecord.financial_year == financial_year
    ).order_by(FDRecord.date_of_maturity.asc()).all()

    return {
        "status": "success",
        "client_id": client_id,
        "financial_year": financial_year,
        "count": len(records),
        "records": [r.to_dict() for r in records]
    }

@router.get("/reconciliation_26as")
def get_reconciliation_26as(
    client_id: int = Query(...),
    financial_year: str = Query("2024-25"),
    db: Session = Depends(get_db)
):
    """
    Runs fuzzy string matching between FD Schedule banks and Form 26AS deductors u/s 194A.
    Returns bank-by-bank interest & TDS reconciliation.
    """
    fd_records = db.query(FDRecord).filter(
        FDRecord.client_id == client_id,
        FDRecord.financial_year == financial_year
    ).all()

    as26_entries = db.query(AS26Entry).filter(
        AS26Entry.client_id == client_id,
        AS26Entry.financial_year == financial_year
    ).all()

    # Group FD schedule data by Bank Name
    fd_bank_summary = {}
    for r in fd_records:
        bname = r.bank_name.strip()
        if bname not in fd_bank_summary:
            fd_bank_summary[bname] = {"interest": 0.0, "tds": 0.0, "count": 0}
        fd_bank_summary[bname]["interest"] += r.interest_income
        fd_bank_summary[bname]["tds"] += r.tds_deducted
        fd_bank_summary[bname]["count"] += 1

    # Fuzzy match 26AS entries to FD Banks
    reco_items = []
    used_26as_ids = set()

    for fd_bank, fd_data in fd_bank_summary.items():
        best_match = None
        best_score = 0.0

        for entry in as26_entries:
            if entry.id in used_26as_ids: continue
            score = fuzzy_bank_match(fd_bank, entry.deductor_name)
            if score > best_score:
                best_score = score
                best_match = entry

        if best_match and best_score >= 0.45:
            used_26as_ids.add(best_match.id)
            as26_amt = best_match.amount_paid
            as26_tds = best_match.tds_deducted
            tan_val = best_match.tan
            ded_name = best_match.deductor_name
        else:
            as26_amt = 0.0
            as26_tds = 0.0
            tan_val = "NOT FOUND IN 26AS"
            ded_name = "No 26AS Match"

        fd_interest = round(fd_data["interest"], 2)
        variance = round(fd_interest - as26_amt, 2)
        fd_tds = round(fd_data["tds"], 2)

        if best_match and abs(variance) <= 100:
            status = "Matched"
        elif best_match and variance > 100:
            status = "FD Interest Higher"
        elif best_match and variance < -100:
            status = "26AS Interest Higher"
        else:
            status = "Unmatched in 26AS"

        reco_items.append({
            "bank_name": fd_bank,
            "deductor_name_26as": ded_name,
            "tan": tan_val,
            "fd_interest": fd_interest,
            "as26_interest": round(as26_amt, 2),
            "variance": variance,
            "fd_tds": fd_tds,
            "as26_tds": round(as26_tds, 2),
            "status": status,
            "match_score": round(best_score, 2)
        })

    # Include remaining unmatched 26AS entries
    for entry in as26_entries:
        if entry.id not in used_26as_ids:
            reco_items.append({
                "bank_name": "Unmatched in FD Schedule",
                "deductor_name_26as": entry.deductor_name,
                "tan": entry.tan,
                "fd_interest": 0.0,
                "as26_interest": round(entry.amount_paid, 2),
                "variance": round(-entry.amount_paid, 2),
                "fd_tds": 0.0,
                "as26_tds": round(entry.tds_deducted, 2),
                "status": "Unmatched in FD Schedule",
                "match_score": 0.0
            })

    total_fd_interest = round(sum(r["fd_interest"] for r in reco_items), 2)
    total_26as_interest = round(sum(r["as26_interest"] for r in reco_items), 2)
    total_variance = round(total_fd_interest - total_26as_interest, 2)

    return {
        "status": "success",
        "financial_year": financial_year,
        "total_fd_interest": total_fd_interest,
        "total_26as_interest": total_26as_interest,
        "total_variance": total_variance,
        "items": reco_items,
        "as26_entries": [e.to_dict() for e in as26_entries]
    }

@router.post("/roll_forward")
def trigger_roll_forward(payload: RollForwardSchema, db: Session = Depends(get_db)):
    """Manually trigger Prior-Year Roll-Forward into target_fy."""
    count = execute_roll_forward(payload.client_id, payload.target_fy, db)
    return {
        "status": "success",
        "client_id": payload.client_id,
        "target_fy": payload.target_fy,
        "rolled_forward_count": count
    }

@router.post("/records")
def create_record(payload: CreateFDRecordSchema, db: Session = Depends(get_db)):
    """Manually add an FD record."""
    calc = FDCalculator(financial_year=payload.financial_year)
    res = calc.process_fd(
        bank_name=payload.bank_name,
        fd_account_number=payload.fd_account_number,
        principal_amount=payload.principal_amount,
        date_of_issue=payload.date_of_issue,
        date_of_maturity=payload.date_of_maturity,
        interest_rate=payload.interest_rate,
        compounding_frequency=payload.compounding_frequency or "Quarterly",
        opening_accrued_interest=payload.opening_accrued_interest or 0.0,
        tds_deducted=payload.tds_deducted or 0.0,
        status=payload.status or "Active"
    )

    record = FDRecord(
        client_id=payload.client_id,
        financial_year=payload.financial_year,
        bank_name=res["bank_name"],
        fd_account_number=res["fd_account_number"],
        principal_amount=res["principal_amount"],
        date_of_issue=res["date_of_issue"],
        date_of_maturity=res["date_of_maturity"],
        interest_rate=res["interest_rate"],
        compounding_frequency=res["compounding_frequency"],
        opening_accrued_interest=res["opening_accrued_interest"],
        tds_deducted=res["tds_deducted"],
        status=res["status"],
        opening_principal=res["opening_principal"],
        created_principal=res["created_principal"],
        matured_principal=res["matured_principal"],
        settled_accrued_interest=res["settled_accrued_interest"],
        original_maturity_days=res["original_maturity_days"],
        remaining_maturity_days=res["remaining_maturity_days"],
        interest_income=res["interest_income"],
        closing_accrued_interest=res["closing_accrued_interest"],
        closing_principal=res["closing_principal"],
        closing_total_balance=res["closing_total_balance"],
        classification_class=res["classification_class"],
        classification_label=res["classification_label"]
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return {"status": "success", "record": record.to_dict()}

@router.put("/records/{record_id}")
def update_record(record_id: int, payload: UpdateFDRecordSchema, db: Session = Depends(get_db)):
    """Edit an FD record and re-calculate statutory fields."""
    record = db.query(FDRecord).filter(FDRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="FD record not found")

    bank = payload.bank_name if payload.bank_name is not None else record.bank_name
    fd_acc = payload.fd_account_number if payload.fd_account_number is not None else record.fd_account_number
    principal = payload.principal_amount if payload.principal_amount is not None else record.principal_amount
    issue_date = payload.date_of_issue if payload.date_of_issue is not None else record.date_of_issue
    maturity_date = payload.date_of_maturity if payload.date_of_maturity is not None else record.date_of_maturity
    rate = payload.interest_rate if payload.interest_rate is not None else record.interest_rate
    freq = payload.compounding_frequency if payload.compounding_frequency is not None else record.compounding_frequency
    opening_accrued = payload.opening_accrued_interest if payload.opening_accrued_interest is not None else record.opening_accrued_interest
    tds = payload.tds_deducted if payload.tds_deducted is not None else record.tds_deducted
    status = payload.status if payload.status is not None else record.status

    calc = FDCalculator(financial_year=record.financial_year)
    res = calc.process_fd(
        bank_name=bank,
        fd_account_number=fd_acc,
        principal_amount=principal,
        date_of_issue=issue_date,
        date_of_maturity=maturity_date,
        interest_rate=rate,
        compounding_frequency=freq,
        opening_accrued_interest=opening_accrued,
        tds_deducted=tds,
        status=status,
        is_roll_forward=record.is_roll_forward,
        opening_principal_override=record.opening_principal if record.is_roll_forward else None
    )

    record.bank_name = res["bank_name"]
    record.fd_account_number = res["fd_account_number"]
    record.principal_amount = res["principal_amount"]
    record.date_of_issue = res["date_of_issue"]
    record.date_of_maturity = res["date_of_maturity"]
    record.interest_rate = res["interest_rate"]
    record.compounding_frequency = res["compounding_frequency"]
    record.opening_accrued_interest = res["opening_accrued_interest"]
    record.tds_deducted = res["tds_deducted"]
    record.status = res["status"]
    record.opening_principal = res["opening_principal"]
    record.created_principal = res["created_principal"]
    record.matured_principal = res["matured_principal"]
    record.settled_accrued_interest = res["settled_accrued_interest"]
    record.original_maturity_days = res["original_maturity_days"]
    record.remaining_maturity_days = res["remaining_maturity_days"]
    record.interest_income = res["interest_income"]
    record.closing_accrued_interest = res["closing_accrued_interest"]
    record.closing_principal = res["closing_principal"]
    record.closing_total_balance = res["closing_total_balance"]
    record.classification_class = res["classification_class"]
    record.classification_label = res["classification_label"]

    db.commit()
    db.refresh(record)
    return {"status": "success", "record": record.to_dict()}

@router.delete("/records/{record_id}")
def delete_record(record_id: int, db: Session = Depends(get_db)):
    """Delete an FD record and cascade-delete all downstream rolled-forward records."""
    record = db.query(FDRecord).filter(FDRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="FD record not found")

    cascade_delete_fd_record(record_id, db)
    db.delete(record)
    db.commit()
    return {"status": "success", "deleted_id": record_id}

@router.delete("/records_all")
def delete_all_records(
    client_id: int = Query(...),
    financial_year: str = Query("2024-25"),
    db: Session = Depends(get_db)
):
    """Deletes all FD records for a client and financial year, with cascade deletion of downstream roll-forwards."""
    records = db.query(FDRecord).filter(
        FDRecord.client_id == client_id,
        FDRecord.financial_year == financial_year
    ).all()

    deleted_count = len(records)
    for r in records:
        cascade_delete_fd_record(r.id, db)
        db.delete(r)

    db.commit()
    return {
        "status": "success",
        "client_id": client_id,
        "financial_year": financial_year,
        "deleted_count": deleted_count
    }


# ==========================================
# 3. STATUTORY SUMMARY & 4-SHEET EXCEL EXPORT
# ==========================================

@router.get("/summary")
def get_summary(
    client_id: int = Query(...),
    financial_year: str = Query("2024-25"),
    db: Session = Depends(get_db)
):
    """Calculates summary totals grouped by Class 1, Class 2, Class 3 statutory buckets including Interest Income and Accrued Interest."""
    execute_roll_forward(client_id, financial_year, db)

    records = db.query(FDRecord).filter(
        FDRecord.client_id == client_id,
        FDRecord.financial_year == financial_year
    ).all()

    class1 = [r for r in records if r.classification_class == "Class 1"]
    class2 = [r for r in records if r.classification_class == "Class 2"]
    class3 = [r for r in records if r.classification_class == "Class 3"]

    def calc_bucket(recs):
        return {
            "count": len(recs),
            "total_opening_principal": round(sum(r.opening_principal for r in recs), 2),
            "total_created_principal": round(sum(r.created_principal for r in recs), 2),
            "total_matured_principal": round(sum(r.matured_principal for r in recs), 2),
            "total_closing_principal": round(sum(r.closing_principal for r in recs), 2),
            "total_interest_income": round(sum(r.interest_income for r in recs), 2),
            "total_accrued_interest": round(sum(r.closing_accrued_interest for r in recs), 2),
            "total_balance": round(sum(r.closing_total_balance for r in recs), 2)
        }

    return {
        "status": "success",
        "financial_year": financial_year,
        "total_fds": len(records),
        "overall_closing_principal": round(sum(r.closing_principal for r in records), 2),
        "overall_accrued_interest": round(sum(r.closing_accrued_interest for r in records), 2),
        "overall_total_interest_income": round(sum(r.interest_income for r in records), 2),
        "overall_total_balance": round(sum(r.closing_total_balance for r in records), 2),
        "class_1_cash_equivalents": calc_bucket(class1),
        "class_2_other_current": calc_bucket(class2),
        "class_3_non_current": calc_bucket(class3)
    }

@router.get("/export_working_paper")
def export_working_paper(
    client_id: int = Query(...),
    financial_year: str = Query("2024-25"),
    db: Session = Depends(get_db)
):
    """
    Generates a 4-SHEET formula-linked Excel working paper workbook using openpyxl:
    - Sheet 1: Statutory Summary (Classification buckets, principal, interest, accrued, total balance)
    - Sheet 2: Detailed Schedule (FD movement schedule & cell formulas)
    - Sheet 3: 26AS Reconciliation (Bank-by-Bank reconciliation comparing FD Interest vs 26AS 194A Interest)
    - Sheet 4: 26AS Entries (Raw extracted Form 26AS Section 194A entries)
    """
    execute_roll_forward(client_id, financial_year, db)

    records = db.query(FDRecord).filter(
        FDRecord.client_id == client_id,
        FDRecord.financial_year == financial_year
    ).order_by(FDRecord.date_of_maturity.asc()).all()

    as26_entries = db.query(AS26Entry).filter(
        AS26Entry.client_id == client_id,
        AS26Entry.financial_year == financial_year
    ).all()

    calc = FDCalculator(financial_year=financial_year)
    rep_yr = calc.reporting_date.year
    rep_month = calc.reporting_date.month
    rep_day = calc.reporting_date.day

    wb = openpyxl.Workbook()

    # Styling Assets
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    rollforward_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    total_border = Border(
        top=Side(style='thin', color='1F4E78'),
        bottom=Side(style='double', color='1F4E78')
    )

    # ----------------------------------------------------
    # SHEET 1: MASTER STATUTORY SUMMARY SHEET
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Statutory Summary"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:F1")
    t1 = ws1["A1"]
    t1.value = f"FD WORKING PAPER - STATUTORY CLASSIFICATION SUMMARY ({financial_year})"
    t1.font = title_font
    t1.fill = header_fill
    t1.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:F2")
    t2 = ws1["A2"]
    t2.value = f"Reporting Date: March 31, {rep_yr} | Balance Sheet Classification & Revenue Mapping Schedule"
    t2.font = Font(name="Calibri", size=10, italic=True)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    ws1_headers = [
        "Statutory Classification Bucket",
        "Criteria / Maturity Rule",
        "Closing Principal (₹)",
        "Interest Income in FY (₹)",
        "Closing Accrued Interest (₹)",
        "Total Closing Balance (₹)"
    ]
    
    ws1.row_dimensions[4].height = 25
    for col_idx, h in enumerate(ws1_headers, 1):
        cell = ws1.cell(row=4, column=col_idx)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_data = [
        [
            "Class 1: Cash & Cash Equivalents",
            "Original Maturity ≤ 90 Days (3 Months)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 1*\", 'Detailed Schedule'!$M$4:$M$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 1*\", 'Detailed Schedule'!$O$4:$O$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 1*\", 'Detailed Schedule'!$R$4:$R$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 1*\", 'Detailed Schedule'!$S$4:$S$500)"
        ],
        [
            "Class 2: Other Current Bank Balances",
            "Original Maturity > 90 Days & Remaining ≤ 365 Days",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 2*\", 'Detailed Schedule'!$M$4:$M$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 2*\", 'Detailed Schedule'!$O$4:$O$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 2*\", 'Detailed Schedule'!$R$4:$R$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 2*\", 'Detailed Schedule'!$S$4:$S$500)"
        ],
        [
            "Class 3: Non-Current Assets",
            "Original Maturity > 90 Days & Remaining > 365 Days",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 3*\", 'Detailed Schedule'!$M$4:$M$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 3*\", 'Detailed Schedule'!$O$4:$O$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 3*\", 'Detailed Schedule'!$R$4:$R$500)",
            "=SUMIF('Detailed Schedule'!$U$4:$U$500, \"Class 3*\", 'Detailed Schedule'!$S$4:$S$500)"
        ]
    ]

    for row_idx, row_vals in enumerate(summary_data, 5):
        ws1.row_dimensions[row_idx].height = 22
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if col_idx >= 3:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    tot_row = 8
    ws1.row_dimensions[tot_row].height = 25
    ws1.cell(row=tot_row, column=1, value="GRAND TOTAL").font = bold_font
    ws1.cell(row=tot_row, column=2, value="=COUNTA('Detailed Schedule'!$A$4:$A$500) & \" Total FD Records\"").font = Font(name="Calibri", size=10, italic=True)
    
    for c_idx, formula_col in enumerate(['C', 'D', 'E', 'F'], 3):
        c = ws1.cell(row=tot_row, column=c_idx, value=f"=SUM({formula_col}5:{formula_col}7)")
        c.number_format = '₹#,##0.00'
        c.font = bold_font

    for c in range(1, 7):
        ws1.cell(row=tot_row, column=c).border = total_border

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 50
    ws1.column_dimensions['C'].width = 24
    ws1.column_dimensions['D'].width = 26
    ws1.column_dimensions['E'].width = 28
    ws1.column_dimensions['F'].width = 28

    # ----------------------------------------------------
    # SHEET 2: DETAILED COMPUTATION SCHEDULE
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Detailed Schedule")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:U1")
    t2_cell = ws2["A1"]
    t2_cell.value = f"DETAILED FIXED DEPOSIT MOVEMENT & STATUTORY SCHEDULE ({financial_year})"
    t2_cell.font = title_font
    t2_cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    t2_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    ws2_headers = [
        "FD ID", "Bank Name", "FD Account No", "Date of Issue", "Date of Maturity",
        "Interest Rate (%)", "Compounding Freq", "Original Days", "Remaining Days",
        "Opening Principal (₹)", "Created in FY (₹)", "Matured in FY (₹)", "Closing Principal (₹)",
        "Opening Accrued (₹)", "Interest Earned (₹)", "TDS Deducted (₹)", "Settled Accrued (₹)",
        "Closing Accrued (₹)", "Total Closing Balance (₹)", "Status", "Statutory Classification"
    ]

    ws2.row_dimensions[3].height = 25
    for col_idx, h in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=3, column=col_idx)
        cell.value = h
        cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, r in enumerate(records, 4):
        ws2.row_dimensions[row_idx].height = 20
        status_display = f"{r.status} (Roll-Forward from PY)" if r.is_roll_forward else r.status

        row_vals = [
            r.id, r.bank_name, r.fd_account_number, r.date_of_issue, r.date_of_maturity,
            r.interest_rate, r.compounding_frequency,
            f"=MAX(0, DATEVALUE(E{row_idx})-DATEVALUE(D{row_idx}))",
            f"=MAX(0, DATE({rep_yr},{rep_month},{rep_day})-DATEVALUE(E{row_idx}))",
            r.opening_principal, r.created_principal, r.matured_principal,
            f"=J{row_idx}+K{row_idx}-L{row_idx}",
            r.opening_accrued_interest, r.interest_income, r.tds_deducted, r.settled_accrued_interest,
            f"=N{row_idx}+O{row_idx}-P{row_idx}-Q{row_idx}",
            f"=M{row_idx}+R{row_idx}",
            status_display,
            f'=IF(H{row_idx}<=90, "Class 1: Cash & Cash Equivalents", IF(I{row_idx}<=365, "Class 2: Other Current Bank Balances", "Class 3: Non-Current Assets"))'
        ]

        for col_idx, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)

            if r.is_roll_forward:
                cell.fill = rollforward_fill

            if col_idx in [10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [6, 8, 9]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [4, 5]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    last_r = len(records) + 4
    ws2.row_dimensions[last_r].height = 25
    ws2.cell(row=last_r, column=1, value="TOTAL").font = bold_font
    
    for c_idx in [10, 11, 12, 13, 14, 15, 16, 17, 18, 19]:
        col_letter = get_column_letter(c_idx)
        c = ws2.cell(row=last_r, column=c_idx, value=f"=SUM({col_letter}4:{col_letter}{last_r-1})")
        c.number_format = '₹#,##0.00'
        c.font = bold_font

    for col_c in range(1, 22):
        ws2.cell(row=last_r, column=col_c).border = total_border

    for col in ws2.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]: continue
            if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # ----------------------------------------------------
    # SHEET 3: 26AS RECONCILIATION SHEET
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="26AS Reconciliation")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:H1")
    t3 = ws3["A1"]
    t3.value = f"FORM 26AS U/S 194A INTEREST INCOME RECONCILIATION ({financial_year})"
    t3.font = title_font
    t3.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    ws3_headers = [
        "Bank Name (FD Schedule)", "Deductor Name (26AS)", "TAN of Deductor",
        "FD Interest Income (₹)", "26AS Amount u/s 194A (₹)", "Variance (₹)",
        "FD Schedule TDS (₹)", "Reconciliation Status"
    ]

    ws3.row_dimensions[3].height = 25
    for col_idx, h in enumerate(ws3_headers, 1):
        cell = ws3.cell(row=3, column=col_idx)
        cell.value = h
        cell.fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Run reconciliation logic for Sheet 3
    reco_data = get_reconciliation_26as(client_id, financial_year, db)
    reco_items = reco_data.get("items", [])

    for r_idx, item in enumerate(reco_items, 4):
        ws3.row_dimensions[r_idx].height = 20
        row_vals = [
            item["bank_name"], item["deductor_name_26as"], item["tan"],
            item["fd_interest"], item["as26_interest"],
            f"=D{r_idx}-E{r_idx}", item["fd_tds"], item["status"]
        ]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws3.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if c_idx in [4, 5, 6, 7]:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    last_r3 = len(reco_items) + 4
    ws3.row_dimensions[last_r3].height = 25
    ws3.cell(row=last_r3, column=1, value="TOTAL RECONCILIATION").font = bold_font
    
    for c_idx in [4, 5, 6, 7]:
        col_letter = get_column_letter(c_idx)
        c = ws3.cell(row=last_r3, column=c_idx, value=f"=SUM({col_letter}4:{col_letter}{last_r3-1})")
        c.number_format = '₹#,##0.00'
        c.font = bold_font

    for col_c in range(1, 9):
        ws3.cell(row=last_r3, column=col_c).border = total_border

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 24
    ws3.column_dimensions['E'].width = 24
    ws3.column_dimensions['F'].width = 20
    ws3.column_dimensions['G'].width = 22
    ws3.column_dimensions['H'].width = 25

    # ----------------------------------------------------
    # SHEET 4: RAW 26AS ENTRIES SHEET
    # ----------------------------------------------------
    ws4 = wb.create_sheet(title="26AS Entries")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:E1")
    t4 = ws4["A1"]
    t4.value = f"EXTRACTED FORM 26AS SECTION 194A ENTRIES ({financial_year})"
    t4.font = title_font
    t4.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    ws4_headers = [
        "Deductor Name (Bank)", "TAN of Deductor", "Section", "Total Amount Paid / Credited (₹)", "Total Tax Deducted (₹)"
    ]

    ws4.row_dimensions[3].height = 25
    for col_idx, h in enumerate(ws4_headers, 1):
        cell = ws4.cell(row=3, column=col_idx)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, entry in enumerate(as26_entries, 4):
        ws4.row_dimensions[r_idx].height = 20
        row_vals = [
            entry.deductor_name, entry.tan, entry.section, entry.amount_paid, entry.tds_deducted
        ]
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws4.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if c_idx in [4, 5]:
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")

    last_r4 = len(as26_entries) + 4
    ws4.row_dimensions[last_r4].height = 25
    ws4.cell(row=last_r4, column=1, value="TOTAL 26AS ENTRIES").font = bold_font
    
    for c_idx in [4, 5]:
        col_letter = get_column_letter(c_idx)
        c = ws4.cell(row=last_r4, column=c_idx, value=f"=SUM({col_letter}4:{col_letter}{last_r4-1})")
        c.number_format = '₹#,##0.00'
        c.font = bold_font

    for col_c in range(1, 6):
        ws4.cell(row=last_r4, column=col_c).border = total_border

    ws4.column_dimensions['A'].width = 38
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 30
    ws4.column_dimensions['E'].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"FD_Working_Paper_{financial_year}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
