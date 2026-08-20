import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def generate_fd_template() -> io.BytesIO:
    """
    Generates a formatted Excel import template for Fixed Deposits (FDs) with DataValidation dropdowns.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FD Import Template"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "FIXED DEPOSIT WORKING PAPER - IMPORT TEMPLATE"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Instructions Row
    ws.merge_cells("A2:J2")
    inst_cell = ws["A2"]
    inst_cell.value = "Fill in FD details below. Use dropdown lists for Compounding Frequency and Status. Dates: YYYY-MM-DD or DD/MM/YYYY."
    inst_cell.font = Font(name="Calibri", size=9, italic=True, color="595959")
    inst_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Headers Definition
    headers = [
        "Bank Name",
        "FD Account Number",
        "Principal Amount",
        "Date of Issue",
        "Date of Maturity",
        "Interest Rate (%)",
        "Compounding Frequency",
        "Opening Accrued Interest",
        "TDS Deducted",
        "Status"
    ]

    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='medium', color='1F4E78')
    )

    ws.row_dimensions[3].height = 25
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header_title
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Data Validation Dropdown Lists
    dv_freq = DataValidation(type="list", formula1='"Quarterly,Monthly,Half-Yearly,Annual,Simple"', allow_blank=True)
    dv_freq.error = 'Your entry is not in the list. Please select from the dropdown.'
    dv_freq.errorTitle = 'Invalid Compounding Frequency'
    dv_freq.prompt = 'Select frequency from dropdown'
    dv_freq.promptTitle = 'Compounding Frequency'
    ws.add_data_validation(dv_freq)
    dv_freq.add("G4:G500")

    dv_status = DataValidation(type="list", formula1='"Active,Matured"', allow_blank=True)
    dv_status.error = 'Please select either Active or Matured.'
    dv_status.errorTitle = 'Invalid Status'
    dv_status.prompt = 'Select status from dropdown'
    dv_status.promptTitle = 'Status'
    ws.add_data_validation(dv_status)
    dv_status.add("J4:J500")

    # Sample Data Rows
    sample_rows = [
        ["State Bank of India", "FD-398249102", 500000.0, "2024-01-15", "2024-04-10", 7.25, "Quarterly", 2500.0, 250.0, "Active"],
        ["HDFC Bank", "FD-784019283", 1000000.0, "2023-06-01", "2025-05-31", 7.50, "Quarterly", 45000.0, 4500.0, "Active"],
        ["ICICI Bank Ltd", "FD-901823741", 2000000.0, "2022-10-01", "2027-09-30", 7.75, "Quarterly", 120000.0, 12000.0, "Active"],
        ["Axis Bank", "FD-112233445", 250000.0, "2024-02-01", "2024-04-30", 6.80, "Quarterly", 1000.0, 100.0, "Active"]
    ]

    row_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for row_idx, row_data in enumerate(sample_rows, 4):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.border = row_border
            cell.font = Font(name="Calibri", size=10)
            
            # Formatting
            if col_idx in [3, 8, 9]:  # Amounts
                cell.number_format = '₹#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx == 6:  # Interest Rate
                cell.number_format = '0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in [4, 5]:  # Dates
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]: continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
