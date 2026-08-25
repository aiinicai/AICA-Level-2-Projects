import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Realistic Base Chart of Accounts
# Format: (Account Code, Account Name, Category, Sub-Category, Account Type, Normal Balance, Base Debit, Base Credit)
ACCOUNTS_CHART = [
    # ── ASSETS ──────────────────────────────────────────────────
    ("1010", "Cash & Bank Balances", "Current Assets", "Cash & Bank", "Asset", "Debit", 28500.0, 0.0),
    ("1020", "Trade Accounts Receivable", "Current Assets", "Accounts Receivable", "Asset", "Debit", 34200.0, 0.0),
    ("1030", "Allowance for Doubtful Accounts", "Current Assets", "Accounts Receivable", "Asset", "Credit", 0.0, 1200.0),
    ("1040", "Raw Materials Inventory", "Current Assets", "Inventory", "Asset", "Debit", 14500.0, 0.0),
    ("1050", "Work-in-Progress Inventory", "Current Assets", "Inventory", "Asset", "Debit", 9800.0, 0.0),
    ("1060", "Finished Goods Inventory", "Current Assets", "Inventory", "Asset", "Debit", 21500.0, 0.0),
    ("1070", "Prepaid Expenses & Insurance", "Current Assets", "Prepaid Expenses", "Asset", "Debit", 4200.0, 0.0),
    ("1080", "Other Current Assets & Advances", "Current Assets", "Other Current Assets", "Asset", "Debit", 3100.0, 0.0),

    ("1510", "Land & Industrial Buildings", "Non-Current Assets", "Property Plant & Equipment", "Asset", "Debit", 75000.0, 0.0),
    ("1520", "Manufacturing Machinery & Plant", "Non-Current Assets", "Property Plant & Equipment", "Asset", "Debit", 62000.0, 0.0),
    ("1530", "Office & IT Infrastructure", "Non-Current Assets", "Property Plant & Equipment", "Asset", "Debit", 14500.0, 0.0),
    ("1540", "Accumulated Depreciation", "Non-Current Assets", "Accumulated Depreciation", "Asset", "Credit", 0.0, 28500.0),
    ("1610", "Software Licenses & Patents", "Non-Current Assets", "Intangible Assets", "Asset", "Debit", 12500.0, 0.0),
    ("1620", "Long-Term Strategic Investments", "Non-Current Assets", "Long-Term Investments", "Asset", "Debit", 15000.0, 0.0),

    # ── LIABILITIES ─────────────────────────────────────────────
    ("2010", "Trade Accounts Payable", "Current Liabilities", "Accounts Payable", "Liability", "Credit", 0.0, 24500.0),
    ("2020", "Short-Term Working Capital Loan", "Current Liabilities", "Short-Term Debt", "Liability", "Credit", 0.0, 14000.0),
    ("2030", "Accrued Salaries & Compensation", "Current Liabilities", "Accrued Expenses", "Liability", "Credit", 0.0, 6500.0),
    ("2040", "Statutory Taxes Payable (GST/TDS)", "Current Liabilities", "Accrued Expenses", "Liability", "Credit", 0.0, 4800.0),
    ("2050", "Deferred Customer Advances", "Current Liabilities", "Deferred Revenue", "Liability", "Credit", 0.0, 5200.0),

    ("2510", "Long-Term Bank Borrowings", "Non-Current Liabilities", "Long-Term Debt", "Liability", "Credit", 0.0, 45000.0),
    ("2520", "Finance Lease Liabilities", "Non-Current Liabilities", "Finance Lease Liabilities", "Liability", "Credit", 0.0, 8500.0),
    ("2530", "Deferred Tax Liabilities", "Non-Current Liabilities", "Deferred Tax Liability", "Liability", "Credit", 0.0, 6200.0),

    # ── EQUITY ──────────────────────────────────────────────────
    ("3010", "Paid-in Share Capital", "Equity", "Paid-in Capital", "Equity", "Credit", 0.0, 60000.0),
    ("3020", "Retained Earnings & Reserves", "Equity", "Retained Earnings", "Equity", "Credit", 0.0, 79900.0),

    # ── REVENUE ─────────────────────────────────────────────────
    ("4010", "Gross Sales & Services Revenue", "Revenue", "Gross Revenue", "Revenue", "Credit", 0.0, 125000.0),
    ("4020", "Sales Returns & Trade Discounts", "Revenue", "Sales Returns", "Revenue", "Debit", 2500.0, 0.0),

    # ── COST OF GOODS SOLD ──────────────────────────────────────
    ("5010", "Raw Material Consumed", "Cost of Goods Sold", "COGS Material", "Expense", "Debit", 42000.0, 0.0),
    ("5020", "Direct Manufacturing Labor & Wages", "Cost of Goods Sold", "COGS Labor", "Expense", "Debit", 18500.0, 0.0),
    ("5030", "Factory Power, Utilities & Overhead", "Cost of Goods Sold", "COGS Overhead", "Expense", "Debit", 7500.0, 0.0),
    ("5040", "Freight Inward & Logistics", "Cost of Goods Sold", "COGS Freight", "Expense", "Debit", 3200.0, 0.0),

    # ── OPERATING EXPENSES ──────────────────────────────────────
    ("6010", "Sales & Marketing Staff Salaries", "Operating Expenses", "Sales & Marketing", "Expense", "Debit", 7800.0, 0.0),
    ("6020", "Digital Marketing & Advertising", "Operating Expenses", "Sales & Marketing", "Expense", "Debit", 4500.0, 0.0),
    ("6030", "Executive & Staff Compensation", "Operating Expenses", "General & Admin", "Expense", "Debit", 9200.0, 0.0),
    ("6040", "Office Rent & Infrastructure Maintenance", "Operating Expenses", "General & Admin", "Expense", "Debit", 4200.0, 0.0),
    ("6050", "Research & Software Development", "Operating Expenses", "Research & Development", "Expense", "Debit", 3800.0, 0.0),
    ("6060", "Legal & Professional Consulting Fees", "Operating Expenses", "General & Admin", "Expense", "Debit", 2100.0, 0.0),
    ("6070", "Depreciation & Amortization Expense", "Operating Expenses", "Depreciation & Amortization", "Expense", "Debit", 3500.0, 0.0),

    # ── NON-OPERATING & TAX EXPENSES ────────────────────────────
    ("7010", "Interest & Finance Charges", "Non-Operating Expenses", "Interest Expense", "Expense", "Debit", 2200.0, 0.0),
    ("7020", "Provision for Income Tax", "Income Tax Expense", "Income Tax Expense", "Expense", "Debit", 3500.0, 0.0)
]

# Period-specific factors to generate realistic operational trends across quarters:
# Fluctuation in gross margin, operating leverage, scale efficiencies & working capital.
QUARTER_FACTORS = {
    "TB_Q1_FY2023": {"q": "Q1", "fy": "FY2023", "rev": 1.00, "cogs_mat": 1.00, "cogs_lab": 1.00, "cogs_ovh": 1.00, "sm": 1.00, "ga": 1.00, "rd": 1.00, "assets": 1.00, "liab": 1.00},
    "TB_Q2_FY2023": {"q": "Q2", "fy": "FY2023", "rev": 1.12, "cogs_mat": 1.03, "cogs_lab": 1.05, "cogs_ovh": 1.04, "sm": 1.05, "ga": 1.02, "rd": 1.04, "assets": 1.08, "liab": 1.04},
    "TB_Q3_FY2023": {"q": "Q3", "fy": "FY2023", "rev": 1.28, "cogs_mat": 1.10, "cogs_lab": 1.12, "cogs_ovh": 1.08, "sm": 1.10, "ga": 1.05, "rd": 1.08, "assets": 1.20, "liab": 1.10},
    "TB_Q4_FY2023": {"q": "Q4", "fy": "FY2023", "rev": 1.20, "cogs_mat": 1.14, "cogs_lab": 1.16, "cogs_ovh": 1.10, "sm": 1.22, "ga": 1.14, "rd": 1.10, "assets": 1.15, "liab": 1.12},
    
    "TB_Q1_FY2024": {"q": "Q1", "fy": "FY2024", "rev": 1.35, "cogs_mat": 1.30, "cogs_lab": 1.28, "cogs_ovh": 1.25, "sm": 1.28, "ga": 1.22, "rd": 1.25, "assets": 1.30, "liab": 1.22},
    "TB_Q2_FY2024": {"q": "Q2", "fy": "FY2024", "rev": 1.50, "cogs_mat": 1.36, "cogs_lab": 1.34, "cogs_ovh": 1.30, "sm": 1.32, "ga": 1.25, "rd": 1.28, "assets": 1.42, "liab": 1.28},
    "TB_Q3_FY2024": {"q": "Q3", "fy": "FY2024", "rev": 1.72, "cogs_mat": 1.48, "cogs_lab": 1.44, "cogs_ovh": 1.38, "sm": 1.38, "ga": 1.30, "rd": 1.35, "assets": 1.60, "liab": 1.35},
    "TB_Q4_FY2024": {"q": "Q4", "fy": "FY2024", "rev": 1.85, "cogs_mat": 1.62, "cogs_lab": 1.55, "cogs_ovh": 1.48, "sm": 1.48, "ga": 1.35, "rd": 1.42, "assets": 1.72, "liab": 1.42},
}

def get_account_multiplier(acct_code: str, subcat: str, cat: str, factors: dict) -> float:
    if acct_code in ["4010", "4020"]:
        return factors["rev"]
    elif acct_code == "5010":
        return factors["cogs_mat"]
    elif acct_code == "5020":
        return factors["cogs_lab"]
    elif acct_code in ["5030", "5040"]:
        return factors["cogs_ovh"]
    elif acct_code in ["6010", "6020"]:
        return factors["sm"]
    elif acct_code in ["6030", "6040", "6060", "6070", "7010", "7020"]:
        return factors["ga"]
    elif acct_code == "6050":
        return factors["rd"]
    elif cat in ["Current Assets", "Non-Current Assets"]:
        return factors["assets"]
    elif cat in ["Current Liabilities", "Non-Current Liabilities", "Equity"]:
        return factors["liab"]
    return factors["rev"]

def generate_trial_balance_excel(filename: str = "TrialBalance_COMP001_FY2023_FY2024.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )

    columns = [
        "Account Code", "Account Name", "Category", "Sub-Category",
        "Account Type", "Normal Balance", "Debit Amount", "Credit Amount",
        "Net Balance", "Quarter", "Fiscal Year"
    ]

    quarter_data_store = {}

    # 1. Generate 8 Quarter Sheets
    for sheet_name, factors in QUARTER_FACTORS.items():
        q, fy = factors["q"], factors["fy"]
        rows = []

        for acct in ACCOUNTS_CHART:
            code, name, cat, subcat, acct_type, norm, base_deb, base_cred = acct
            mult = get_account_multiplier(code, subcat, cat, factors)
            
            deb = round(base_deb * mult, 2)
            cred = round(base_cred * mult, 2)
            net = round(deb - cred, 2)

            rows.append({
                "code": code, "name": name, "cat": cat, "subcat": subcat,
                "acct_type": acct_type, "norm": norm, "deb": deb, "cred": cred,
                "net": net, "q": q, "fy": fy
            })

        # Balance Trial Balance by adjusting Retained Earnings (Account 3020)
        tot_deb = sum(r["deb"] for r in rows)
        other_cred = sum(r["cred"] for r in rows if r["code"] != "3020")
        
        re_idx = next(i for i, r in enumerate(rows) if r["code"] == "3020")
        rows[re_idx]["cred"] = round(tot_deb - other_cred, 2)
        rows[re_idx]["net"] = round(rows[re_idx]["deb"] - rows[re_idx]["cred"], 2)

        quarter_data_store[sheet_name] = rows

    # 2. Compute Annual Sheets as exact sum of 4 Quarters for P&L, closing Q4 for Balance Sheet
    for fy_name, q_keys in [("TB_Annual_FY2023", ["TB_Q1_FY2023", "TB_Q2_FY2023", "TB_Q3_FY2023", "TB_Q4_FY2023"]),
                            ("TB_Annual_FY2024", ["TB_Q1_FY2024", "TB_Q2_FY2024", "TB_Q3_FY2024", "TB_Q4_FY2024"])]:
        fy_year = "FY2023" if "2023" in fy_name else "FY2024"
        annual_rows = []

        for idx, acct in enumerate(ACCOUNTS_CHART):
            code, name, cat, subcat, acct_type, norm, _, _ = acct
            
            # For P&L accounts (Revenue & Expense), take exact sum of Q1 + Q2 + Q3 + Q4
            if acct_type in ["Revenue", "Expense"]:
                sum_deb = sum(quarter_data_store[q_k][idx]["deb"] for q_k in q_keys)
                sum_cred = sum(quarter_data_store[q_k][idx]["cred"] for q_k in q_keys)
            else:
                # For Balance Sheet accounts (Asset, Liability, Equity), take closing Q4 balance
                sum_deb = quarter_data_store[q_keys[-1]][idx]["deb"]
                sum_cred = quarter_data_store[q_keys[-1]][idx]["cred"]

            sum_deb = round(sum_deb, 2)
            sum_cred = round(sum_cred, 2)
            net = round(sum_deb - sum_cred, 2)

            annual_rows.append({
                "code": code, "name": name, "cat": cat, "subcat": subcat,
                "acct_type": acct_type, "norm": norm, "deb": sum_deb, "cred": sum_cred,
                "net": net, "q": "Annual", "fy": fy_year
            })

        # Balance Annual Sheet by adjusting Retained Earnings (Account 3020)
        tot_deb = sum(r["deb"] for r in annual_rows)
        other_cred = sum(r["cred"] for r in annual_rows if r["code"] != "3020")
        
        re_idx = next(i for i, r in enumerate(annual_rows) if r["code"] == "3020")
        annual_rows[re_idx]["cred"] = round(tot_deb - other_cred, 2)
        annual_rows[re_idx]["net"] = round(annual_rows[re_idx]["deb"] - annual_rows[re_idx]["cred"], 2)

        quarter_data_store[fy_name] = annual_rows

    # Order of sheets: Q1-Q4 FY23, Annual FY23, Q1-Q4 FY24, Annual FY24
    ordered_sheet_names = [
        "TB_Q1_FY2023", "TB_Q2_FY2023", "TB_Q3_FY2023", "TB_Q4_FY2023", "TB_Annual_FY2023",
        "TB_Q1_FY2024", "TB_Q2_FY2024", "TB_Q3_FY2024", "TB_Q4_FY2024", "TB_Annual_FY2024"
    ]

    for sheet_name in ordered_sheet_names:
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        ws.append(columns)
        for col_idx in range(1, 12):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, r_dict in enumerate(quarter_data_store[sheet_name], start=2):
            ws.append([
                r_dict["code"], r_dict["name"], r_dict["cat"], r_dict["subcat"],
                r_dict["acct_type"], r_dict["norm"], r_dict["deb"], r_dict["cred"],
                r_dict["net"], r_dict["q"], r_dict["fy"]
            ])
            for c_idx in range(1, 12):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.border = thin_border
                if c_idx in [7, 8, 9]:
                    cell.number_format = "#,##0.00"

    wb.save(filename)
    print(f"Successfully generated 10-sheet Dynamic Trend Trial Balance: {filename}")

if __name__ == "__main__":
    generate_trial_balance_excel()
