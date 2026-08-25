import io
import pandas as pd
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

KPI_FULL_NAMES = {
    "gross_profit_margin": "Gross Profit Margin",
    "net_profit_margin": "Net Profit Margin (PAT)",
    "operating_margin": "Operating Profit Margin (EBIT)",
    "ebitda_margin": "EBITDA Margin",
    "roa": "Return on Assets (ROA)",
    "roe": "Return on Equity (ROE)",
    "roce": "Return on Capital Employed (ROCE)",
    "cogs_ratio": "Cost of Goods Sold (COGS) Ratio",
    "cost_to_income_ratio": "Cost to Income Ratio",
    "current_ratio": "Current Ratio",
    "quick_ratio": "Quick (Acid Test) Ratio",
    "cash_ratio": "Cash Ratio",
    "nwc_ratio": "Net Working Capital (NWC) Ratio",
    "debt_to_equity": "Debt to Equity Ratio",
    "debt_to_assets": "Debt to Assets Ratio",
    "interest_coverage": "Interest Coverage Ratio (TIE)",
    "debt_service_coverage": "Debt Service Coverage Ratio (DSCR)",
    "net_debt_to_ebitda": "Net Debt to EBITDA Ratio",
    "asset_turnover": "Total Asset Turnover",
    "fixed_asset_turnover": "Fixed Asset Turnover",
    "inventory_turnover": "Inventory Turnover",
    "receivables_turnover": "Receivables Turnover",
    "payables_turnover": "Payables Turnover",
    "dso": "Days Sales Outstanding (DSO)",
    "dio": "Days Inventory Outstanding (DIO)",
    "dpo": "Days Payables Outstanding (DPO)",
    "ccc": "Cash Conversion Cycle (CCC)",
    "revenue_growth": "Revenue Growth Rate",
    "gross_profit_growth": "Gross Profit Growth Rate",
    "ebitda_growth": "EBITDA Growth Rate",
    "ebit_growth": "Operating Income (EBIT) Growth Rate",
    "net_income_growth": "Net Income Growth Rate",
    "eps": "Earnings Per Share (EPS)",
    "book_value_per_share": "Book Value Per Share (BVPS)",
    "revenue_per_employee": "Revenue Per Employee",
    "profit_per_employee": "Profit Per Employee"
}

def format_kpi_name(key: str) -> str:
    return KPI_FULL_NAMES.get(key, key.replace("_", " ").title())

def generate_excel_report(company_name: str, pnl_dict: Dict, bs_dict: Dict, kpi_dict: Dict) -> bytes:
    wb = Workbook()
    
    # ----------------------------------------------------
    # STYLES & COLOR PALETTE
    # ----------------------------------------------------
    title_font      = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill      = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Deep Navy
    
    header_font     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill     = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid") # Royal Blue
    
    cat_font        = Font(name="Calibri", size=11, bold=True, color="1E3A8A")
    cat_fill        = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid") # Light Blue Tint
    
    bold_font       = Font(name="Calibri", size=10, bold=True, color="0F172A")
    regular_font    = Font(name="Calibri", size=10, color="1E293B")
    
    # RAG Status Colors
    green_fill      = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    green_font      = Font(name="Calibri", size=10, bold=True, color="15803D")
    
    amber_fill      = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    amber_font      = Font(name="Calibri", size=10, bold=True, color="B45309")
    
    red_fill        = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_font        = Font(name="Calibri", size=10, bold=True, color="B91C1C")

    highlight_fill  = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # ----------------------------------------------------
    # TAB 1: KPI SCORECARD
    # ----------------------------------------------------
    ws_kpi = wb.active
    ws_kpi.title = "KPI Scorecard"
    ws_kpi.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws_kpi.merge_cells("A1:G1")
    title_cell = ws_kpi.cell(row=1, column=1, value=f" 📊 FINANCIAL KPI PERFORMANCE SCORECARD — {company_name.upper()}")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_kpi.row_dimensions[1].height = 40

    headers = ["Category", "KPI Name", "Current Value", "Unit", "QoQ Delta", "YoY Delta", "RAG Status"]
    ws_kpi.append(headers)
    ws_kpi.row_dimensions[2].height = 25
    for col_idx, h in enumerate(headers, 1):
        cell = ws_kpi.cell(row=2, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    curr_row = 3
    if kpi_dict:
        for cat, kpis in kpi_dict.items():
            # Category Row Header
            ws_kpi.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
            cat_cell = ws_kpi.cell(row=curr_row, column=1, value=f"  📌 CATEGORY: {cat.upper()}")
            cat_cell.font = cat_font
            cat_cell.fill = cat_fill
            cat_cell.alignment = Alignment(horizontal="left", vertical="center")
            ws_kpi.row_dimensions[curr_row].height = 22
            curr_row += 1

            for kpi_key, data in kpis.items():
                rag = data.get("rag_status", "GREEN").upper()
                if rag == "GREEN":
                    rag_text = "🟢 GREEN"
                    r_fill, r_font = green_fill, green_font
                elif rag == "AMBER":
                    rag_text = "🟡 AMBER"
                    r_fill, r_font = amber_fill, amber_font
                else:
                    rag_text = "🔴 RED"
                    r_fill, r_font = red_fill, red_font

                val = data.get("value", 0)
                unit = data.get("unit", "")
                val_display = val

                row_vals = [
                    cat.title(),
                    format_kpi_name(kpi_key),
                    val_display,
                    unit,
                    data.get("qoq_delta", 0),
                    data.get("yoy_delta", 0),
                    rag_text
                ]
                ws_kpi.append(row_vals)
                ws_kpi.row_dimensions[curr_row].height = 20

                for c_idx in range(1, 8):
                    c = ws_kpi.cell(row=curr_row, column=c_idx)
                    c.font = regular_font
                    c.border = thin_border
                    c.alignment = Alignment(vertical="center")
                    if c_idx in [3, 5, 6]:
                        c.alignment = Alignment(horizontal="right", vertical="center")
                        if unit == "₹":
                            c.number_format = '₹#,##0.00'
                        elif unit == "%":
                            c.number_format = '0.00"%"'
                        else:
                            c.number_format = '#,##0.00'
                    if c_idx == 7:
                        c.fill = r_fill
                        c.font = r_font
                        c.alignment = Alignment(horizontal="center", vertical="center")
                curr_row += 1

    for col_idx, col in enumerate(ws_kpi.columns, 1):
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col_idx)
        ws_kpi.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # ----------------------------------------------------
    # TAB 2: INCOME STATEMENT (P&L)
    # ----------------------------------------------------
    ws_pnl = wb.create_sheet(title="Income Statement")
    ws_pnl.views.sheetView[0].showGridLines = True
    
    ws_pnl.merge_cells("A1:B1")
    pnl_title = ws_pnl.cell(row=1, column=1, value=f" 🧾 INCOME STATEMENT (P&L) — {company_name.upper()}")
    pnl_title.font = title_font
    pnl_title.fill = title_fill
    pnl_title.alignment = Alignment(horizontal="left", vertical="center")
    ws_pnl.row_dimensions[1].height = 40

    ws_pnl.append(["Line Item Description", "Amount (₹)"])
    ws_pnl.row_dimensions[2].height = 25
    ws_pnl.cell(row=2, column=1).font = header_font; ws_pnl.cell(row=2, column=1).fill = header_fill; ws_pnl.cell(row=2, column=1).alignment = Alignment(vertical="center")
    ws_pnl.cell(row=2, column=2).font = header_font; ws_pnl.cell(row=2, column=2).fill = header_fill; ws_pnl.cell(row=2, column=2).alignment = Alignment(horizontal="right", vertical="center")

    if pnl_dict:
        items = [
            ("Gross Revenue", pnl_dict.get("gross_revenue", 0), True, False),
            ("Less: Sales Returns & Allowances", pnl_dict.get("sales_returns", 0), False, False),
            ("Net Revenue", pnl_dict.get("net_revenue", 0), True, True),
            ("Less: Cost of Goods Sold (COGS)", pnl_dict.get("cogs", {}).get("total", 0), False, False),
            ("Gross Profit", pnl_dict.get("gross_profit", 0), True, False),
            ("Less: Operating Expenses (OpEx)", pnl_dict.get("opex", {}).get("total", 0), False, False),
            ("EBITDA", pnl_dict.get("ebitda", 0), True, True),
            ("Less: Depreciation & Amortization", pnl_dict.get("depreciation_amortization", 0), False, False),
            ("EBIT (Operating Income)", pnl_dict.get("ebit", 0), True, False),
            ("Less: Interest Expense", pnl_dict.get("interest_expense", 0), False, False),
            ("Profit Before Tax (PBT / EBT)", pnl_dict.get("ebt", 0), True, True),
            ("Less: Income Tax Expense", pnl_dict.get("income_tax", 0), False, False),
            ("Profit After Tax (PAT / Net Income)", pnl_dict.get("net_income", 0), True, True)
        ]
        
        for r_idx, (name, val, is_bold, is_highlight) in enumerate(items, 3):
            ws_pnl.append([name, val])
            ws_pnl.row_dimensions[r_idx].height = 20
            
            c1 = ws_pnl.cell(row=r_idx, column=1)
            c2 = ws_pnl.cell(row=r_idx, column=2)
            
            c1.font = bold_font if is_bold else regular_font
            c2.font = bold_font if is_bold else regular_font
            
            c1.border = thin_border
            c2.border = thin_border
            
            c2.number_format = '₹#,##0.00'
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c1.alignment = Alignment(vertical="center")

            if is_highlight:
                c1.fill = highlight_fill
                c2.fill = highlight_fill

    ws_pnl.column_dimensions['A'].width = 40
    ws_pnl.column_dimensions['B'].width = 25

    # ----------------------------------------------------
    # TAB 3: BALANCE SHEET SUMMARY
    # ----------------------------------------------------
    ws_bs = wb.create_sheet(title="Balance Sheet")
    ws_bs.views.sheetView[0].showGridLines = True

    ws_bs.merge_cells("A1:B1")
    bs_title = ws_bs.cell(row=1, column=1, value=f" ⚖️ BALANCE SHEET SUMMARY — {company_name.upper()}")
    bs_title.font = title_font
    bs_title.fill = title_fill
    bs_title.alignment = Alignment(horizontal="left", vertical="center")
    ws_bs.row_dimensions[1].height = 40

    ws_bs.append(["Balance Sheet Line Item", "Amount (₹)"])
    ws_bs.row_dimensions[2].height = 25
    ws_bs.cell(row=2, column=1).font = header_font; ws_bs.cell(row=2, column=1).fill = header_fill; ws_bs.cell(row=2, column=1).alignment = Alignment(vertical="center")
    ws_bs.cell(row=2, column=2).font = header_font; ws_bs.cell(row=2, column=2).fill = header_fill; ws_bs.cell(row=2, column=2).alignment = Alignment(horizontal="right", vertical="center")

    if bs_dict:
        ca = bs_dict.get("assets", {}).get("current_assets", {})
        nca = bs_dict.get("assets", {}).get("non_current_assets", {})
        cl = bs_dict.get("liabilities", {}).get("current_liabilities", {})
        ncl = bs_dict.get("liabilities", {}).get("non_current_liabilities", {})
        eq = bs_dict.get("equity", {})

        bs_items = [
            ("Cash & Cash Equivalents", ca.get("cash_and_equivalents", 0), False, False),
            ("Accounts Receivable (net)", ca.get("accounts_receivable", 0), False, False),
            ("Inventories", ca.get("inventory", 0), False, False),
            ("Total Current Assets", ca.get("total", 0), True, False),
            ("Property, Plant & Equipment (net)", nca.get("property_plant_equipment_net", 0), False, False),
            ("TOTAL ASSETS", bs_dict.get("assets", {}).get("total_assets", 0), True, True),
            ("Total Current Liabilities", cl.get("total", 0), True, False),
            ("Total Non-Current Liabilities", ncl.get("total", 0), False, False),
            ("TOTAL LIABILITIES", bs_dict.get("liabilities", {}).get("total_liabilities", 0), True, False),
            ("Shareholders' Equity", eq.get("total_equity", 0), True, False),
            ("TOTAL LIABILITIES & EQUITY", bs_dict.get("total_liabilities_and_equity", 0), True, True)
        ]

        for r_idx, (name, val, is_bold, is_highlight) in enumerate(bs_items, 3):
            ws_bs.append([name, val])
            ws_bs.row_dimensions[r_idx].height = 20
            
            c1 = ws_bs.cell(row=r_idx, column=1)
            c2 = ws_bs.cell(row=r_idx, column=2)
            
            c1.font = bold_font if is_bold else regular_font
            c2.font = bold_font if is_bold else regular_font
            
            c1.border = thin_border
            c2.border = thin_border
            
            c2.number_format = '₹#,##0.00'
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c1.alignment = Alignment(vertical="center")

            if is_highlight:
                c1.fill = highlight_fill
                c2.fill = highlight_fill

    ws_bs.column_dimensions['A'].width = 40
    ws_bs.column_dimensions['B'].width = 25

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_pdf_report(company_name: str, period: str, year: str, pnl_dict: Dict, bs_dict: Dict, kpi_dict: Dict) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=4
    )

    sub_title_style = ParagraphStyle(
        name='SubTitleStyle',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor("#1E40AF"),
        spaceBefore=12,
        spaceAfter=6
    )

    meta_style = ParagraphStyle(
        name='MetaStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        textColor=colors.HexColor("#475569")
    )

    badge_green = ParagraphStyle('BG', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor("#15803D"), alignment=1)
    badge_amber = ParagraphStyle('BA', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor("#B45309"), alignment=1)
    badge_red   = ParagraphStyle('BR', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor("#B91C1C"), alignment=1)

    tbl_header  = ParagraphStyle('TH', fontName='Helvetica-Bold', fontSize=9, textColor=colors.whitesmoke)
    tbl_cell    = ParagraphStyle('TC', fontName='Helvetica', fontSize=8, textColor=colors.HexColor("#1E293B"))
    tbl_cell_b  = ParagraphStyle('TCB', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor("#0F172A"))
    tbl_cell_r  = ParagraphStyle('TCR', fontName='Helvetica', fontSize=8, textColor=colors.HexColor("#1E293B"), alignment=2)
    tbl_cell_rb = ParagraphStyle('TCRB', fontName='Helvetica-Bold', fontSize=8, textColor=colors.HexColor("#0F172A"), alignment=2)

    story = []
    
    # 1. Document Header Banner
    story.append(Paragraph(f"Financial KPI Analytics Report", title_style))
    story.append(Paragraph(f"<b>Company:</b> {company_name} &nbsp;|&nbsp; <b>Active Period:</b> {period} {year} &nbsp;|&nbsp; <b>Status:</b> Validated Multi-Period Analysis", meta_style))
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1E3A8A"), spaceAfter=12))

    # 2. Executive Metric Highlight Callouts (Cards)
    if pnl_dict:
        net_rev = f"INR {pnl_dict.get('net_revenue', 0):,.2f}"
        pbt     = f"INR {pnl_dict.get('ebt', 0):,.2f}"
        pat     = f"INR {pnl_dict.get('net_income', 0):,.2f}"
        gp_m    = f"{((pnl_dict.get('gross_profit', 0) / (pnl_dict.get('net_revenue', 1) or 1)) * 100):.1f}%"

        card_data = [
            [
                Paragraph("<b>NET REVENUE</b>", ParagraphStyle('C1', fontName='Helvetica', fontSize=8, textColor=colors.HexColor("#1E40AF"), alignment=1)),
                Paragraph("<b>PROFIT BEFORE TAX</b>", ParagraphStyle('C2', fontName='Helvetica', fontSize=8, textColor=colors.HexColor("#047857"), alignment=1)),
                Paragraph("<b>PROFIT AFTER TAX</b>", ParagraphStyle('C3', fontName='Helvetica', fontSize=8, textColor=colors.HexColor("#6D28D9"), alignment=1)),
                Paragraph("<b>GROSS MARGIN %</b>", ParagraphStyle('C4', fontName='Helvetica', fontSize=8, textColor=colors.HexColor("#B45309"), alignment=1))
            ],
            [
                Paragraph(f"<b>{net_rev}</b>", ParagraphStyle('V1', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#1E3A8A"), alignment=1)),
                Paragraph(f"<b>{pbt}</b>", ParagraphStyle('V2', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#065F46"), alignment=1)),
                Paragraph(f"<b>{pat}</b>", ParagraphStyle('V3', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#5B21B6"), alignment=1)),
                Paragraph(f"<b>{gp_m}</b>", ParagraphStyle('V4', fontName='Helvetica-Bold', fontSize=11, textColor=colors.HexColor("#92400E"), alignment=1))
            ]
        ]

        card_table = Table(card_data, colWidths=[130, 130, 130, 130])
        card_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor("#EFF6FF")),
            ('BACKGROUND', (1, 0), (1, -1), colors.HexColor("#ECFDF5")),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor("#F5F3FF")),
            ('BACKGROUND', (3, 0), (3, -1), colors.HexColor("#FFFBEB")),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BOX', (0, 0), (0, -1), 0.5, colors.HexColor("#BFDBFE")),
            ('BOX', (1, 0), (1, -1), 0.5, colors.HexColor("#A7F3D0")),
            ('BOX', (2, 0), (2, -1), 0.5, colors.HexColor("#DDD6FE")),
            ('BOX', (3, 0), (3, -1), 0.5, colors.HexColor("#FDE68A")),
        ]))
        story.append(card_table)
        story.append(Spacer(1, 12))

    # 3. Executive KPI RAG Scorecard Table
    story.append(Paragraph("Financial KPI Performance Scorecard", sub_title_style))
    kpi_rows = [[
        Paragraph("<b>Category</b>", tbl_header),
        Paragraph("<b>KPI Metric Name</b>", tbl_header),
        Paragraph("<b>Value</b>", ParagraphStyle('TH1', parent=tbl_header, alignment=2)),
        Paragraph("<b>Unit</b>", tbl_header),
        Paragraph("<b>RAG Status</b>", ParagraphStyle('TH2', parent=tbl_header, alignment=1))
    ]]

    if kpi_dict:
        for cat, kpis in kpi_dict.items():
            for kpi_name, data in kpis.items():
                rag = data.get("rag_status", "GREEN").upper()
                if rag == "GREEN":
                    status_p = Paragraph("GREEN", badge_green)
                elif rag == "AMBER":
                    status_p = Paragraph("AMBER", badge_amber)
                else:
                    status_p = Paragraph("RED", badge_red)

                val_raw = data.get("value", 0)
                unit_raw = data.get("unit", "")
                if unit_raw == "₹":
                    val_str = f"INR {float(val_raw):,.2f}"
                    unit_disp = "INR"
                else:
                    val_str = f"{float(val_raw):,.2f}"
                    unit_disp = unit_raw

                kpi_rows.append([
                    Paragraph(cat.title(), tbl_cell),
                    Paragraph(format_kpi_name(kpi_name), tbl_cell_b),
                    Paragraph(val_str, tbl_cell_r),
                    Paragraph(unit_disp, tbl_cell),
                    status_p
                ])

    kpi_table = Table(kpi_rows, colWidths=[90, 200, 80, 60, 90])
    ts = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]
    
    for r in range(1, len(kpi_rows)):
        if r % 2 == 0:
            ts.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor("#F8FAFC")))

    kpi_table.setStyle(TableStyle(ts))
    story.append(kpi_table)
    story.append(Spacer(1, 14))

    # 4. Income Statement Highlights Table
    story.append(Paragraph("Income Statement Summary", sub_title_style))
    pnl_rows = [[Paragraph("<b>Income Statement Line Item</b>", tbl_header), Paragraph("<b>Amount (INR)</b>", ParagraphStyle('TH2', parent=tbl_header, alignment=2))]]
    if pnl_dict:
        items = [
            ("Gross Revenue", pnl_dict.get("gross_revenue", 0), True),
            ("Less: Sales Returns", pnl_dict.get("sales_returns", 0), False),
            ("Net Revenue", pnl_dict.get("net_revenue", 0), True),
            ("Cost of Goods Sold (COGS)", pnl_dict.get("cogs", {}).get("total", 0), False),
            ("Gross Profit", pnl_dict.get("gross_profit", 0), True),
            ("Operating Expenses (OpEx)", pnl_dict.get("opex", {}).get("total", 0), False),
            ("EBITDA", pnl_dict.get("ebitda", 0), True),
            ("EBIT (Operating Income)", pnl_dict.get("ebit", 0), True),
            ("Profit Before Tax (PBT)", pnl_dict.get("ebt", 0), True),
            ("Profit After Tax (PAT)", pnl_dict.get("net_income", 0), True)
        ]
        for name, val, is_b in items:
            c_style = tbl_cell_b if is_b else tbl_cell
            r_style = tbl_cell_rb if is_b else tbl_cell_r
            pnl_rows.append([Paragraph(name, c_style), Paragraph(f"INR {val:,.2f}", r_style)])

    pnl_table = Table(pnl_rows, colWidths=[330, 190])
    pts = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]
    for r in range(1, len(pnl_rows)):
        if r % 2 == 0:
            pts.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor("#F8FAFC")))

    pnl_table.setStyle(TableStyle(pts))
    story.append(pnl_table)
    story.append(Spacer(1, 14))

    # 5. Balance Sheet Summary Table
    story.append(Paragraph("Balance Sheet Summary", sub_title_style))
    bs_rows = [[Paragraph("<b>Balance Sheet Line Item</b>", tbl_header), Paragraph("<b>Amount (INR)</b>", ParagraphStyle('TH2', parent=tbl_header, alignment=2))]]
    if bs_dict:
        ca = bs_dict.get("assets", {}).get("current_assets", {})
        nca = bs_dict.get("assets", {}).get("non_current_assets", {})
        cl = bs_dict.get("liabilities", {}).get("current_liabilities", {})
        ncl = bs_dict.get("liabilities", {}).get("non_current_liabilities", {})
        eq = bs_dict.get("equity", {})

        bs_items = [
            ("Cash & Cash Equivalents", ca.get("cash_and_equivalents", 0), False),
            ("Accounts Receivable (net)", ca.get("accounts_receivable", 0), False),
            ("Inventories", ca.get("inventory", 0), False),
            ("Total Current Assets", ca.get("total", 0), True),
            ("Property, Plant & Equipment (net)", nca.get("property_plant_equipment_net", 0), False),
            ("TOTAL ASSETS", bs_dict.get("assets", {}).get("total_assets", 0), True),
            ("Total Current Liabilities", cl.get("total", 0), True),
            ("Total Non-Current Liabilities", ncl.get("total", 0), False),
            ("TOTAL LIABILITIES", bs_dict.get("liabilities", {}).get("total_liabilities", 0), True),
            ("Shareholders' Equity", eq.get("total_equity", 0), True),
            ("TOTAL LIABILITIES & EQUITY", bs_dict.get("total_liabilities_and_equity", 0), True)
        ]
        for name, val, is_b in bs_items:
            c_style = tbl_cell_b if is_b else tbl_cell
            r_style = tbl_cell_rb if is_b else tbl_cell_r
            bs_rows.append([Paragraph(name, c_style), Paragraph(f"INR {val:,.2f}", r_style)])

    bs_table = Table(bs_rows, colWidths=[330, 190])
    bpts = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
    ]
    for r in range(1, len(bs_rows)):
        if r % 2 == 0:
            bpts.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor("#F8FAFC")))

    bs_table.setStyle(TableStyle(bpts))
    story.append(bs_table)

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()
