"""Excel Report Generator using openpyxl.
Produces a multi-tab client-ready workbook with formatted tables, KPIs, and conditional formatting.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from typing import Dict, Any, Optional

# Color Palette
NAVY_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
TEAL_HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
ACCENT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
RED_FLAG_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
GREEN_KPI_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E78")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
REGULAR_FONT = Font(name="Calibri", size=11)
RED_BOLD_FONT = Font(name="Calibri", size=11, bold=True, color="C00000")

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

def auto_fit_columns(ws, max_len_cap: int = 50):
    """Adjust column widths dynamically."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format and ('₹' in cell.number_format or '#' in cell.number_format):
                val_str = f"₹{val_str}"
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), max_len_cap)

def format_table_header(ws, row_idx: int, num_cols: int, fill=NAVY_HEADER_FILL):
    """Apply styling to a table header row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 25

def write_dataframe_to_sheet(
    ws,
    df: pd.DataFrame,
    start_row: int = 4,
    start_col: int = 1,
    header_fill=NAVY_HEADER_FILL,
    highlight_flags: bool = False
):
    """Write a DataFrame to worksheet with formatting."""
    if df is None or df.empty:
        ws.cell(row=start_row, column=start_col, value="No records found.").font = SUBTITLE_FONT
        return start_row + 2

    headers = df.columns.tolist()
    # Write Headers
    for c_idx, h in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + c_idx, value=str(h))
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[start_row].height = 24

    curr_row = start_row + 1
    for r_idx, row in df.iterrows():
        is_flagged = bool(row.get("is_flagged", False)) if highlight_flags else False
        row_fill = RED_FLAG_FILL if is_flagged else (ACCENT_ROW_FILL if curr_row % 2 == 0 else PatternFill(fill_type=None))

        for c_idx, h in enumerate(headers):
            val = row[h]
            cell = ws.cell(row=curr_row, column=start_col + c_idx)
            
            # Format numbers vs strings
            if isinstance(val, (int, float)):
                cell.value = val
                if "amount" in h.lower() or "receipt" in h.lower() or "payment" in h.lower() or "movement" in h.lower() or "balance" in h.lower() or "inr" in h.lower() or "cash" in h.lower():
                    cell.number_format = '₹#,##0.00'
                elif "%" in h.lower() or "share" in h.lower():
                    cell.number_format = '0.00"%"'
                else:
                    cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, list):
                cell.value = ", ".join(str(v) for v in val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.value = str(val) if pd.notna(val) else ""
                cell.alignment = Alignment(horizontal="left", vertical="center")

            cell.font = RED_BOLD_FONT if (is_flagged and c_idx == 0) else REGULAR_FONT
            if row_fill.fill_type:
                cell.fill = row_fill
            cell.border = THIN_BORDER

        curr_row += 1

    return curr_row

def export_excel_report(
    df: pd.DataFrame,
    output_path: str,
    client_name: str = "Client",
    analysis_data: Optional[Dict[str, Any]] = None
) -> str:
    """Generate multi-tab Excel report."""
    from analysis.summaries import (
        generate_executive_summary, generate_month_wise_summary,
        generate_nature_wise_summary, generate_party_wise_summary,
        generate_top_and_extrema_transactions, generate_cash_summary,
        generate_bank_charges_summary
    )
    from analysis.red_flags import detect_red_flags
    from analysis.presumptive_tax import analyze_presumptive_tax
    from analysis.reconciliation import validate_running_balances

    # Prepare Analysis
    exec_sum = generate_executive_summary(df)
    m_sum = generate_month_wise_summary(df)
    nat_cr, nat_dr = generate_nature_wise_summary(df)
    pty_cr, pty_dr = generate_party_wise_summary(df)
    top_ext = generate_top_and_extrema_transactions(df)
    cash_sum = generate_cash_summary(df)
    chg_sum = generate_bank_charges_summary(df)
    df_flagged, red_flag_sum = detect_red_flags(df, client_name=client_name)
    presump_sum = analyze_presumptive_tax(df)
    df_recon, recon_sum = validate_running_balances(df)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Executive Dashboard Sheet
    ws_dash = wb.create_sheet(title="Executive Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title
    ws_dash.merge_cells("A1:G1")
    ws_dash["A1"] = f"BANK STATEMENT ANALYSIS REPORT — {client_name.upper()}"
    ws_dash["A1"].font = TITLE_FONT
    
    ws_dash.merge_cells("A2:G2")
    ws_dash["A2"] = f"Prepared by Dharmjeet & Associates, Chartered Accountants | Period: {exec_sum['start_date']} to {exec_sum['end_date']}"
    ws_dash["A2"].font = SUBTITLE_FONT

    # KPI Summary Cards
    kpis = [
        ("Total Credits (Receipts)", exec_sum["total_credits"], "₹#,##0.00"),
        ("Total Debits (Payments)", exec_sum["total_debits"], "₹#,##0.00"),
        ("Net Cash Flow / Movement", exec_sum["net_movement"], "₹#,##0.00"),
        ("Total Transactions", exec_sum["total_transactions"], "#,##0"),
        ("Cash Deposits (Total)", cash_sum["total_cash_deposits"], "₹#,##0.00"),
        ("Cash Withdrawals (Total)", cash_sum["total_cash_withdrawals"], "₹#,##0.00"),
        ("Red Flag Anomalies Flagged", red_flag_sum["total_flagged_transactions"], "#,##0"),
        ("Reconciliation Status", recon_sum["status"], None)
    ]

    r = 4
    for idx, (label, val, fmt) in enumerate(kpis):
        col_start = 1 if idx % 2 == 0 else 4
        c_val = col_start + 1
        
        ws_dash.cell(row=r, column=col_start, value=label).font = BOLD_FONT
        val_cell = ws_dash.cell(row=r, column=c_val, value=val)
        val_cell.font = BOLD_FONT
        val_cell.fill = GREEN_KPI_FILL if "Total Credits" in label or "RECONCILED" in str(val) else (RED_FLAG_FILL if "Red Flag" in label and val > 0 else ACCENT_ROW_FILL)
        if fmt:
            val_cell.number_format = fmt
        
        if idx % 2 == 1:
            r += 2

    # Section 44AD / 44ADA Presumptive Check Box
    r += 2
    ws_dash.cell(row=r, column=1, value="PRESUMPTIVE TAXATION & AUDIT APPLICABILITY (FY 2024-25 / 2023-24)").font = BOLD_FONT
    r += 1
    p_headers = ["Gross Turnover / Receipts", "Digital Receipts %", "Sec 44AD Audit Required?", "Sec 44ADA Audit Required?", "Min 44AD Presumptive Profit"]
    p_vals = [presump_sum["total_turnover"], presump_sum["digital_percentage"], "YES (Audit Req.)" if presump_sum["audit_required_44ad"] else "NO (Eligible)", "YES (Audit Req.)" if presump_sum["audit_required_44ada"] else "NO (Eligible)", presump_sum["min_presumptive_income_44ad"]]
    
    for c_idx, (h, v) in enumerate(zip(p_headers, p_vals)):
        ws_dash.cell(row=r, column=c_idx + 1, value=h).font = HEADER_FONT
        ws_dash.cell(row=r, column=c_idx + 1).fill = TEAL_HEADER_FILL
        val_c = ws_dash.cell(row=r + 1, column=c_idx + 1, value=v)
        val_c.font = BOLD_FONT
        if isinstance(v, float) and "Profit" in h or "Gross" in h:
            val_c.number_format = '₹#,##0.00'
        elif "%" in h:
            val_c.number_format = '0.00"%"'
    
    auto_fit_columns(ws_dash)

    # 2. Month-wise Summary Sheet
    ws_month = wb.create_sheet(title="Month-wise Summary")
    ws_month.views.sheetView[0].showGridLines = True
    ws_month["A1"] = "MONTH-WISE & FY QUARTER BREAKDOWN"
    ws_month["A1"].font = TITLE_FONT
    write_dataframe_to_sheet(ws_month, m_sum, start_row=3, header_fill=NAVY_HEADER_FILL)
    auto_fit_columns(ws_month)

    # 3. Nature-wise Breakdown Sheet
    ws_nat = wb.create_sheet(title="Nature-wise Summary")
    ws_nat.views.sheetView[0].showGridLines = True
    ws_nat["A1"] = "RECEIPTS BY NATURE CATEGORY"
    ws_nat["A1"].font = TITLE_FONT
    end_cr = write_dataframe_to_sheet(ws_nat, nat_cr, start_row=3, header_fill=TEAL_HEADER_FILL)
    
    ws_nat.cell(row=end_cr + 1, column=1, value="PAYMENTS BY NATURE CATEGORY").font = TITLE_FONT
    write_dataframe_to_sheet(ws_nat, nat_dr, start_row=end_cr + 3, header_fill=NAVY_HEADER_FILL)
    auto_fit_columns(ws_nat)

    # 4. Party-wise Receipts Sheet
    ws_party_cr = wb.create_sheet(title="Party-wise Receipts")
    ws_party_cr.views.sheetView[0].showGridLines = True
    ws_party_cr["A1"] = "COUNTERPARTY RECEIPTS SUMMARY"
    ws_party_cr["A1"].font = TITLE_FONT
    write_dataframe_to_sheet(ws_party_cr, pty_cr, start_row=3, header_fill=TEAL_HEADER_FILL)
    auto_fit_columns(ws_party_cr)

    # 5. Party-wise Payments Sheet
    ws_party_dr = wb.create_sheet(title="Party-wise Payments")
    ws_party_dr.views.sheetView[0].showGridLines = True
    ws_party_dr["A1"] = "COUNTERPARTY PAYMENTS SUMMARY"
    ws_party_dr["A1"].font = TITLE_FONT
    write_dataframe_to_sheet(ws_party_dr, pty_dr, start_row=3, header_fill=NAVY_HEADER_FILL)
    auto_fit_columns(ws_party_dr)

    # 6. Top & Extrema Sheet
    ws_top = wb.create_sheet(title="Top 10 Transactions")
    ws_top.views.sheetView[0].showGridLines = True
    ws_top["A1"] = "TOP 10 LARGEST RECEIPTS"
    ws_top["A1"].font = TITLE_FONT
    end_top_cr = write_dataframe_to_sheet(ws_top, top_ext["top_receipts"], start_row=3, header_fill=TEAL_HEADER_FILL)
    
    ws_top.cell(row=end_top_cr + 1, column=1, value="TOP 10 LARGEST PAYMENTS").font = TITLE_FONT
    write_dataframe_to_sheet(ws_top, top_ext["top_payments"], start_row=end_top_cr + 3, header_fill=NAVY_HEADER_FILL)
    auto_fit_columns(ws_top)

    # 7. Cash Transactions Sheet
    ws_cash = wb.create_sheet(title="Cash Transactions")
    ws_cash.views.sheetView[0].showGridLines = True
    ws_cash["A1"] = "CASH DEPOSITS & WITHDRAWALS SUMMARY"
    ws_cash["A1"].font = TITLE_FONT
    
    ws_cash["A3"] = "Total Cash Deposits:"
    ws_cash["B3"] = cash_sum["total_cash_deposits"]
    ws_cash["B3"].number_format = '₹#,##0.00'
    ws_cash["A3"].font = BOLD_FONT
    ws_cash["B3"].font = BOLD_FONT

    ws_cash["D3"] = "Total Cash Withdrawals:"
    ws_cash["E3"] = cash_sum["total_cash_withdrawals"]
    ws_cash["E3"].number_format = '₹#,##0.00'
    ws_cash["D3"].font = BOLD_FONT
    ws_cash["E3"].font = BOLD_FONT
    
    write_dataframe_to_sheet(ws_cash, cash_sum["month_wise_cash"], start_row=5, header_fill=NAVY_HEADER_FILL)
    auto_fit_columns(ws_cash)

    # 8. Red Flags & Scrutiny Signals Sheet
    ws_flags = wb.create_sheet(title="Red Flags & Scrutiny")
    ws_flags.views.sheetView[0].showGridLines = True
    ws_flags["A1"] = "TAX SCRUTINY & ANOMALY SIGNALS"
    ws_flags["A1"].font = TITLE_FONT
    ws_flags["A2"] = f"Total Anomalies Flagged: {red_flag_sum['total_flagged_transactions']} | Total Flagged Volume: ₹{red_flag_sum['total_flagged_amount']:,.2f}"
    ws_flags["A2"].font = SUBTITLE_FONT

    flagged_df = df_flagged[df_flagged["is_flagged"]].copy() if "is_flagged" in df_flagged.columns else pd.DataFrame()
    if not flagged_df.empty:
        display_cols = ["transaction_date", "counterparty_name", "nature", "mode", "debit_amount", "credit_amount", "flag_reasons", "description"]
        flagged_df_display = flagged_df[[c for c in display_cols if c in flagged_df.columns]].copy()
        flagged_df_display = flagged_df_display.rename(columns={
            "transaction_date": "Date",
            "counterparty_name": "Counterparty",
            "nature": "Nature",
            "mode": "Mode",
            "debit_amount": "Debit (Dr)",
            "credit_amount": "Credit (Cr)",
            "flag_reasons": "Scrutiny Flag Reasons",
            "description": "Narration"
        })
        write_dataframe_to_sheet(ws_flags, flagged_df_display, start_row=4, header_fill=PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"), highlight_flags=True)
    else:
        ws_flags["A4"] = "No red flags or scrutiny anomalies detected under current thresholds."
        ws_flags["A4"].font = BOLD_FONT
    auto_fit_columns(ws_flags)

    # 9. Bank Charges Sheet
    ws_chg = wb.create_sheet(title="Bank Charges & Fees")
    ws_chg.views.sheetView[0].showGridLines = True
    ws_chg["A1"] = "BANK CHARGES & FEE REVERSALS"
    ws_chg["A1"].font = TITLE_FONT
    ws_chg["A3"] = "Total Bank Charges Incurred:"
    ws_chg["B3"] = chg_sum["total_bank_charges"]
    ws_chg["B3"].number_format = '₹#,##0.00'
    ws_chg["A3"].font = BOLD_FONT
    ws_chg["B3"].font = BOLD_FONT
    write_dataframe_to_sheet(ws_chg, chg_sum["charges_breakdown"], start_row=5, header_fill=NAVY_HEADER_FILL)
    auto_fit_columns(ws_chg)

    # 10. Raw Transaction Register Sheet
    ws_reg = wb.create_sheet(title="Transaction Register")
    ws_reg.views.sheetView[0].showGridLines = True
    ws_reg["A1"] = "COMPLETE NORMALIZED TRANSACTION REGISTER"
    ws_reg["A1"].font = TITLE_FONT
    
    reg_df = df_flagged.copy()
    write_dataframe_to_sheet(ws_reg, reg_df, start_row=3, header_fill=NAVY_HEADER_FILL, highlight_flags=True)
    # Enable AutoFilter
    ws_reg.auto_filter.ref = f"A3:{get_column_letter(len(reg_df.columns))}{len(reg_df) + 3}"
    ws_reg.freeze_panes = "A4"
    auto_fit_columns(ws_reg)

    # Save workbook
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    return output_path
