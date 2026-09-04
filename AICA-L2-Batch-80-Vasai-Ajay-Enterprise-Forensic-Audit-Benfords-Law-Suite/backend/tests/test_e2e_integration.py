"""
End-to-End Integration & Forensic Back-Testing Pipeline.
Tests the full system lifecycle: Consent -> Ingest -> DPDP Scrub -> Benford -> Forensics -> Chained Ledger -> PDF, Excel, Word Reports.
"""

import io
import json
import pytest
from fastapi.testclient import TestClient

from backend.app.main import app
from backend.app.engine.dpdp_compliance import generate_verhoeff

client = TestClient(app)


def test_full_forensic_pipeline_e2e():
    """Executes complete end-to-end audit workflow on synthetic enterprise ledger."""
    # 1. Step 1: Legal Disclaimer & DPDP Consent Gate
    consent_res = client.post("/api/consent/declare", json={
        "auditor_name": "Dr. Vikram Seth, CA",
        "organization_fiduciary": "National Forensic Audit Bureau",
        "audit_purpose": "Statutory Fraud & Benford Forensic Review",
        "disclaimer_acknowledged": True,
        "dpdp_mandate_acknowledged": True
    })
    assert consent_res.status_code == 200
    consent_data = consent_res.json()
    assert consent_data["success"] is True
    consent_token = consent_data["consent_token"]
    assert "DPDP-CONSENT-" in consent_token

    # 2. Step 2: Create Realistic Synthetic Enterprise Dataset with Injected PII & Fraud Anomalies
    sample_prefix = "21098765432"
    valid_aadhaar = sample_prefix + generate_verhoeff(sample_prefix)

    csv_rows = [
        "Txn_Date,Vendor_Name,Tax_PAN,GSTIN,Bank_Account,Invoice_ID,Amount,Narration"
    ]
    
    # Base Benford-like distribution
    for i in range(1, 200):
        amt = round((1.08 ** i) * 25, 2)
        csv_rows.append(f"2026-03-01,Apex Tech,AAACP1234F,27AAACC1234G1Z5,123456789012,INV-{1000+i},{amt},Consulting fees")

    # Injected Fraud Patterns:
    # A. Split Transactions / Smurfing under ₹50,000 PAN limit
    csv_rows.append("2026-03-10,Zenith Infra,AAAFZ9999K,27AAACC1234G1Z5,987654321098,INV-SPLIT1,49500.00,Material supplies")
    csv_rows.append("2026-03-10,Zenith Infra,AAAFZ9999K,27AAACC1234G1Z5,987654321098,INV-SPLIT2,49800.00,Material supplies")
    csv_rows.append("2026-03-11,Zenith Infra,AAAFZ9999K,27AAACC1234G1Z5,987654321098,INV-SPLIT3,49950.00,Material supplies")

    # B. Exact Duplicate Invoices
    csv_rows.append("2026-03-15,Global Trade,AAACC5555C,07ABCDE1234F1Z8,555566667777,INV-DUP-01,88500.00,Software license")
    csv_rows.append("2026-03-15,Global Trade,AAACC5555C,07ABCDE1234F1Z8,555566667777,INV-DUP-01,88500.00,Software license")

    # C. Vendor RSF Outlier (Single massive payment vs baseline)
    csv_rows.append("2026-03-20,HighRisk Vendor,AAACP7777P,27AAACC1234G1Z5,111122223333,INV-HR-1,50000.00,Equipment hire")
    csv_rows.append("2026-03-22,HighRisk Vendor,AAACP7777P,27AAACC1234G1Z5,111122223333,INV-HR-2,50000.00,Equipment hire")
    csv_rows.append("2026-03-25,HighRisk Vendor,AAACP7777P,27AAACC1234G1Z5,111122223333,INV-HR-3,750000.00,Emergency service") # RSF = 15.0!

    csv_data = "\n".join(csv_rows).encode('utf-8')

    # Step 3: Ingest Dataset via Multipart Upload
    files = {'file': ('audit_sample.csv', csv_data, 'text/csv')}
    ingest_res = client.post(
        "/api/ingest/upload",
        files=files,
        data={"consent_token": consent_token}
    )
    assert ingest_res.status_code == 200
    ingest_json = ingest_res.json()
    assert ingest_json["success"] is True
    assert ingest_json["row_count"] >= 200
    assert ingest_json["column_mapping"]["amount"] == "Amount"
    assert ingest_json["column_mapping"]["vendor"] == "Vendor_Name"

    # Step 4: Apply Indian DPDP Act 2023 Pseudonymization
    dpdp_res = client.post(
        "/api/dpdp/sanitize",
        data={"action_mode": "PSEUDONYMIZE", "consent_token": consent_token}
    )
    assert dpdp_res.status_code == 200
    dpdp_json = dpdp_res.json()
    assert dpdp_json["success"] is True
    assert dpdp_json["stats"]["pan"] > 0
    assert dpdp_json["stats"]["gstin"] > 0

    # Step 5: Execute Benford's Law Statistical Engine
    benford_res = client.post(
        "/api/benford/analyze",
        json={"amount_column": "Amount", "consent_token": consent_token}
    )
    assert benford_res.status_code == 200
    benford_json = benford_res.json()
    assert benford_json["success"] is True
    assert "first_two_digits" in benford_json
    assert benford_json["valid_rows"] >= 200

    # Step 6: Execute Advanced Forensic Anomaly Scanner
    forensic_res = client.post(
        "/api/forensics/analyze",
        json={"column_mapping": ingest_json["column_mapping"], "consent_token": consent_token}
    )
    assert forensic_res.status_code == 200
    forensic_json = forensic_res.json()
    assert forensic_json["success"] is True
    # Verify injected anomalies were caught
    assert forensic_json["duplicate_analysis"]["exact_duplicate_clusters"] >= 1
    assert forensic_json["split_transaction_analysis"]["total_split_anomalies"] >= 3
    assert forensic_json["rsf_analysis"]["outlier_vendor_count"] >= 1

    # Step 7: Verify Tamper-Evident SHA-256 Chained Audit Ledger
    verify_res = client.post("/api/audit/verify", json={})
    assert verify_res.status_code == 200
    verify_json = verify_res.json()
    assert verify_json["is_valid"] is True
    assert verify_json["total_blocks_verified"] >= 5

    # Step 8: Generate and Download Official PDF Audit Dossier
    pdf_res = client.get("/api/report/pdf")
    assert pdf_res.status_code == 200
    assert pdf_res.headers["content-type"] == "application/pdf"
    assert pdf_res.content.startswith(b"%PDF-")
    assert len(pdf_res.content) > 1000

    # Step 9: Generate and Download Detailed Multi-Tab Excel Workbook (.xlsx)
    excel_res = client.get("/api/report/excel")
    assert excel_res.status_code == 200
    assert "spreadsheetml" in excel_res.headers["content-type"]
    assert len(excel_res.content) > 1000

    # Validate Workbook Sheets & Non-empty Audit Trail
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(excel_res.content))
    expected_sheets = [
        "Executive Summary & DPDP", "Auditor Sampling Guide", "Master Sample Ledger",
        "Sampled - RSF Outliers", "Sampled - Duplicate Payments", "Sampled - Split Smurfing",
        "Benford F2D Digits Table", "Chained Audit Trail"
    ]
    for s in expected_sheets:
        assert s in wb.sheetnames, f"Missing expected sheet {s}"

    # Verify Chained Audit Trail is filled with all real audit blocks
    ws_audit = wb["Chained Audit Trail"]
    assert ws_audit.max_row >= 5, "Chained Audit Trail must contain populated audit blocks"
    assert ws_audit.cell(row=4, column=8).value == "VALID & CHAINED"

    # Step 10: Generate and Download Formatted Word Report (.docx)
    docx_res = client.get("/api/report/docx")
    assert docx_res.status_code == 200
    assert "wordprocessingml" in docx_res.headers["content-type"]
    assert len(docx_res.content) > 1000
