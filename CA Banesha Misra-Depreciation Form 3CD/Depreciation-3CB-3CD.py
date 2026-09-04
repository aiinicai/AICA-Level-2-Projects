import sys
import os
import json
import re
import traceback
import functools
from datetime import date, datetime
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
    QHBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QLabel,
    QMessageBox, QComboBox, QLineEdit, QFrame, QSizePolicy)
from PyQt6.QtGui import QColor, QFont, QKeySequence
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Editable depreciation rate table (Income-tax Act, WDV method) ----
DEP_RATES = {
    "Building (Residential)": 5,
    "Building (Other)": 10,
    "Furniture & Fittings": 10,
    "Plant & Machinery (General)": 15,
    "Motor Vehicles": 15,
    "Computers & Software": 40,
    "Intangible Assets": 25,
}
ASSET_OPTIONS = list(DEP_RATES.keys()) + ["Custom / Other"]

# Table columns - a single Addition Amount + Date of Addition is entered and the
# >=180 / <180 days classification is derived automatically for the selected FY.
# "Deduction Amount" covers both an asset removed from the block AND any sale
# consideration received for it - one figure, no duplication.
COLS = ["Asset/Block", "Rate %", "Opening WDV", "Addition Amount", "Date of Addition",
    "Deduction Amount (incl. Sale Consideration)", "Deduction Date", "Business Use %",
    "Period Used (Auto)", "Depreciation", "Closing WDV"]

READONLY_COLS = {8, 9, 10}   # Period, Depreciation, Closing WDV - computed
ASSET_COL, RATE_COL, ADD_DATE_COL, DED_DATE_COL = 0, 1, 4, 6

CURRENT_FY_START_YEAR = 2026  # today (per app context) falls in FY 2026-27
FY_OPTIONS = [f"{y}-{str(y + 1)[-2:]}" for y in range(CURRENT_FY_START_YEAR - 4, CURRENT_FY_START_YEAR + 2)]
DEFAULT_FY = f"{CURRENT_FY_START_YEAR}-{str(CURRENT_FY_START_YEAR + 1)[-2:]}"

# Only Asset/Block + Rate are pre-filled - everything else starts blank.
SAMPLE = [
    ["Plant & Machinery (General)"],
    ["Computers & Software"],
    ["Motor Vehicles"],
]

RED_TINT = QColor(255, 205, 205)
TRANSPARENT = QColor(255, 255, 255, 0)

# Saved records are stored per-user so both the script and any built .exe find
# the same file regardless of where they are run from.
DATA_FILE = os.path.join(os.path.expanduser("~"), "Depreciation3CD_SavedData.json")


def parse_date(s):
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def num(s):
    """Parse a user-typed amount into a float.

    Users commonly type amounts with thousand-separator commas (Indian
    grouping like '10,50,000.00', or Western grouping like '1,050,000.00'),
    or with a currency symbol/space. A plain float(s) throws on any of these
    and was silently swallowed, returning 0.0 - which made every amount with
    a comma in it (i.e. almost every real amount) compute as zero. Strip
    everything except digits, a leading minus sign, and the decimal point
    before parsing.
    """
    if s is None:
        return 0.0
    cleaned = re.sub(r"[^0-9.\-]", "", str(s).strip())
    if cleaned in ("", "-", ".", "-."):
        return 0.0
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def fy_bounds(fy_str):
    try:
        start_year = int(str(fy_str).strip()[:4])
    except (ValueError, IndexError):
        start_year = CURRENT_FY_START_YEAR
    return date(start_year, 4, 1), date(start_year + 1, 3, 31)


def compute_row(d, fy_start, fy_end):
    """d indices: 0 asset,1 rate,2 opening,3 addition,4 add_date,
    5 deduction(incl. sale),6 ded_date,7 biz%

    Returns: status, depreciation, closing, add_date_error, ded_date_error,
             effective_full_addition, effective_half_addition, effective_deduction
    Any addition/deduction whose date falls outside the selected Financial Year
    is EXCLUDED from this year's calculation and flagged as an error - it must
    not silently affect the depreciation figure.
    """
    rate = num(d[1]) / 100.0
    opening = num(d[2])
    addition = num(d[3])
    add_date = parse_date(d[4])
    deduction = num(d[5])
    ded_date = parse_date(d[6])
    biz_pct = num(d[7]) / 100.0 if d[7] else 1.0

    add_date_error = bool(addition and add_date and not (fy_start <= add_date <= fy_end))
    ded_date_error = bool(deduction and ded_date and not (fy_start <= ded_date <= fy_end))

    eff_addition = 0.0 if add_date_error else addition
    eff_deduction = 0.0 if ded_date_error else deduction

    full_addition, half_addition = 0.0, 0.0
    status_parts = []

    if add_date_error:
        status_parts.append("⚠ Addition date outside FY - ignored")
    elif eff_addition:
        if add_date:
            effective_start = max(add_date, fy_start)
            days_in_use = max((fy_end - effective_start).days + 1, 0)
            if days_in_use >= 180:
                full_addition = eff_addition
                status_parts.append(f"{days_in_use} days — Full Rate")
            else:
                half_addition = eff_addition
                status_parts.append(f"{days_in_use} days — Half Rate")
        else:
            full_addition = eff_addition
            status_parts.append("No date — treated as Full Rate")
    else:
        status_parts.append("No addition")

    if ded_date_error:
        status_parts.append("⚠ Deduction date outside FY - ignored")

    status = " | ".join(status_parts)

    base_full = opening - eff_deduction + full_addition
    dep_full = max(base_full, 0) * rate
    dep_half = max(half_addition, 0) * rate * 0.5
    depreciation = (dep_full + dep_half) * biz_pct
    closing = opening + eff_addition - eff_deduction - depreciation

    return (status, round(depreciation, 2), round(closing, 2),
            add_date_error, ded_date_error, full_addition, half_addition, eff_deduction)


# ---------- Simple JSON-based storage, keyed by Assessee Name + Financial Year ----------
def load_all_records():
    if not os.path.exists(DATA_FILE):
        return {}
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}


def save_all_records(data):
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        return True
    except OSError:
        return False


def record_key(name, fy):
    return f"{name.strip()}||{fy.strip()}"


def safe_slot(func):
    """Wrap a button/slot handler so any unexpected error is shown to the user
    in a message box instead of failing silently - this matters especially for
    a --windowed build where there is no console to print a traceback to.

    functools.wraps is essential here: several buttons connect directly to the
    decorated method (not via a lambda), and Qt's signal/slot machinery
    introspects the callable's signature to decide how many arguments to pass
    (e.g. the clicked(bool) signal's checked-state argument). Without wraps,
    the wrapper's generic (*args, **kwargs) signature would make Qt pass that
    extra bool through to methods that don't accept it (del_row, save_record,
    load_record, generate_excel), raising a TypeError on every click.
    """
    @functools.wraps(func)
    def wrapper(self, *args, **kwargs):
        try:
            return func(self, *args, **kwargs)
        except Exception as exc:  # noqa: BLE001 - deliberately broad, this is the top-level safety net
            QMessageBox.critical(self, "Unexpected Error",
                f"Something went wrong while running this action:\n\n{exc}\n\n"
                f"Details:\n{traceback.format_exc()}")
    return wrapper


class PasteableTable(QTableWidget):
    """Adds Excel-style Ctrl+C / Ctrl+V support so rows can be copied straight
    from a spreadsheet and pasted in, or copied out to paste elsewhere."""

    def __init__(self, rows, cols, host):
        super().__init__(rows, cols)
        self.host = host

    def keyPressEvent(self, event):
        if event.matches(QKeySequence.StandardKey.Paste):
            self.paste_from_clipboard()
        elif event.matches(QKeySequence.StandardKey.Copy):
            self.copy_to_clipboard()
        else:
            super().keyPressEvent(event)

    @safe_slot
    def paste_from_clipboard(self):
        text = QApplication.clipboard().text()
        if not text:
            return
        lines = text.replace("\r\n", "\n").replace("\r", "\n").strip("\n").split("\n")
        if not lines:
            return
        start_row = self.currentRow() if self.currentRow() >= 0 else 0
        start_col = self.currentColumn() if self.currentColumn() >= 0 else 0

        for i, line in enumerate(lines):
            cells = line.split("\t")
            target_row = start_row + i
            while target_row >= self.rowCount():
                self.host.add_row()
            for j, val in enumerate(cells):
                target_col = start_col + j
                if target_col >= self.columnCount() or target_col in READONLY_COLS:
                    continue
                val = val.strip()
                if target_col == ASSET_COL:
                    combo = self.cellWidget(target_row, ASSET_COL)
                    if combo is not None:
                        combo.setCurrentText(val)
                else:
                    item = self.item(target_row, target_col)
                    if item is None:
                        item = QTableWidgetItem()
                        self.setItem(target_row, target_col, item)
                    item.setText(val)
        self.host.calculate(silent=True)

    @safe_slot
    def copy_to_clipboard(self):
        ranges = self.selectedRanges()
        if not ranges:
            return
        sel = ranges[0]
        lines = []
        for r in range(sel.topRow(), sel.bottomRow() + 1):
            cells = []
            for c in range(sel.leftColumn(), sel.rightColumn() + 1):
                if c == ASSET_COL:
                    combo = self.cellWidget(r, ASSET_COL)
                    cells.append(combo.currentText() if combo else "")
                else:
                    item = self.item(r, c)
                    cells.append(item.text() if item else "")
            lines.append("\t".join(cells))
        QApplication.clipboard().setText("\n".join(lines))


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Depreciation - Form 3CB / 3CD")
        self.resize(1560, 660)
        self.setFont(QFont("Century Gothic", 10))
        self.setStyleSheet("""
            * { font-family: 'Century Gothic', 'Segoe UI Semibold', 'Segoe UI', Arial; }
            QMainWindow {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:1,
                    stop:0 #eef4ff, stop:0.5 #f5eeff, stop:1 #eefaf5);
            }
            QLabel#title {
                color: #241b4e; font-size: 24px; font-weight: 800;
                padding: 6px 4px 0px 4px; letter-spacing: 0.5px;
            }
            QLabel#accentbar {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #ff6b6b, stop:0.35 #ffb347, stop:0.65 #4d96ff, stop:1 #34d399);
                min-height: 4px; max-height: 4px; border-radius: 2px;
            }
            QLabel#subtitle { color: #5b5680; font-size: 12px; padding: 6px 4px 4px 4px; }
            QLabel#fieldlbl { color: #241b4e; font-weight: 700; font-size: 11px; letter-spacing: 0.3px; }
            QFrame#topbar {
                background: white; border: 1px solid #ded4f7; border-radius: 12px;
            }
            QLineEdit, QComboBox {
                background: #fbfaff; border: 1.5px solid #c8b9f7; border-radius: 7px;
                padding: 7px 9px; font-size: 12px; color: #241b4e; font-weight: 600;
            }
            QLineEdit:focus, QComboBox:focus { border: 1.5px solid #7c4dff; background: white; }
            QTableWidget {
                background: white; gridline-color: #e3e8f5; font-size: 12px;
                border: 1px solid #ded4f7; border-radius: 10px; alternate-background-color: #f6f4ff;
                selection-background-color: #d8e6ff; selection-color: #241b4e;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #3a2e78, stop:1 #241b4e);
                color: #ffffff; padding: 9px 6px; font-weight: 700; font-size: 11px;
                border: none; border-right: 1px solid #4a3d8f;
            }
            /* Base style gives every plain QPushButton (including dialog / message-box
               buttons like OK) a visible colourful background - previously they had
               white text with no background set, making them invisible. */
            QPushButton {
                color: white; padding: 8px 16px; border-radius: 8px; font-weight: 700; font-size: 12px;
                border: none; min-width: 90px;
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #4d96ff, stop:1 #2f6fe0);
            }
            QPushButton:hover {
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #3a82f0, stop:1 #1f5bcc);
            }
            QPushButton#addbtn, QPushButton#loadbtn, QPushButton#savebtn {
                min-width: 150px; padding: 11px 4px;
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #4d96ff, stop:1 #2f6fe0);
            }
            QPushButton#calc {
                min-width: 150px; padding: 11px 4px;
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #ffb347, stop:1 #ff8c1f);
            }
            QPushButton#calc:hover { background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #ff9f2e, stop:1 #f27a00); }
            QPushButton#gen {
                min-width: 220px; padding: 11px 4px;
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #34d399, stop:1 #0f9d6b);
            }
            QPushButton#gen:hover { background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #22c088, stop:1 #0c8259); }
            QPushButton#del {
                min-width: 150px; padding: 11px 4px;
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #ff6b6b, stop:1 #e0403f);
            }
            QPushButton#del:hover { background: qlineargradient(x1:0,y1:0,x2:0,y2:1, stop:0 #f0524f, stop:1 #c22e2d); }
            QMessageBox { background: white; }
            QMessageBox QLabel { color: #241b4e; font-size: 12px; }
        """)

        root = QWidget()
        self.setCentralWidget(root)
        v = QVBoxLayout(root)
        v.setContentsMargins(18, 14, 18, 14)
        v.setSpacing(8)

        v.addWidget(QLabel("Depreciation Working — Form 3CB / 3CD (Clause 18)", objectName="title"))
        accent = QLabel(objectName="accentbar")
        v.addWidget(accent)
        v.addWidget(QLabel("Enter the addition amount and date of purchase — the 180-day full/half rate "
                            "classification is calculated automatically for the selected Financial Year. "
                            "You can also copy cells from Excel and paste them directly into the table "
                            "(Ctrl+V), and copy rows out with Ctrl+C.", objectName="subtitle"))

        topbar = QFrame(objectName="topbar")
        tb = QHBoxLayout(topbar)
        tb.setContentsMargins(14, 12, 14, 12)
        tb.setSpacing(20)

        name_col = QVBoxLayout()
        name_col.setSpacing(4)
        name_col.addWidget(QLabel("ASSESSEE NAME", objectName="fieldlbl"))
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Enter assessee name")
        self.name_input.setMinimumWidth(260)
        name_col.addWidget(self.name_input)
        tb.addLayout(name_col)

        fy_col = QVBoxLayout()
        fy_col.setSpacing(4)
        fy_col.addWidget(QLabel("FINANCIAL YEAR", objectName="fieldlbl"))
        self.fy_combo = QComboBox()
        self.fy_combo.setEditable(True)
        self.fy_combo.addItems(FY_OPTIONS)
        self.fy_combo.setCurrentText(DEFAULT_FY)
        self.fy_combo.setMinimumWidth(120)
        self.fy_combo.currentTextChanged.connect(lambda _: self.calculate(silent=True))
        fy_col.addWidget(self.fy_combo)
        tb.addLayout(fy_col)

        rec_col = QVBoxLayout()
        rec_col.setSpacing(4)
        rec_col.addWidget(QLabel("SAVED RECORDS (Name + FY)", objectName="fieldlbl"))
        self.records_combo = QComboBox()
        self.records_combo.setMinimumWidth(280)
        rec_col.addWidget(self.records_combo)
        tb.addLayout(rec_col)
        tb.addStretch(1)
        v.addWidget(topbar)

        self.table = PasteableTable(0, len(COLS), self)
        self.table.setHorizontalHeaderLabels(COLS)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setDefaultSectionSize(30)
        v.addWidget(self.table, 1)

        h = QHBoxLayout()
        h.setSpacing(10)
        h.setContentsMargins(0, 4, 0, 0)
        # Every button uses an explicit no-arg lambda rather than connecting the
        # method directly. QPushButton.clicked emits a bool ("checked") argument;
        # connecting it straight to a method that doesn't expect that argument
        # is fragile across Qt/PyQt versions. A lambda with no parameters removes
        # any ambiguity - it always calls the handler with exactly the arguments
        # we intend, regardless of what the signal itself emits.
        b_add = QPushButton("➕  Add Row", objectName="addbtn")
        b_add.clicked.connect(lambda: self.add_row())
        b_del = QPushButton("🗑  Delete Row", objectName="del")
        b_del.clicked.connect(lambda: self.del_row())
        b_calc = QPushButton("🧮  Calculate", objectName="calc")
        b_calc.clicked.connect(lambda: self.calculate())
        b_save = QPushButton("💾  Save Record", objectName="savebtn")
        b_save.clicked.connect(lambda: self.save_record())
        b_load = QPushButton("📂  Load Record", objectName="loadbtn")
        b_load.clicked.connect(lambda: self.load_record())
        b_gen = QPushButton("📊  Generate Form 3CD Excel", objectName="gen")
        b_gen.clicked.connect(lambda: self.generate_excel())
        for b in (b_add, b_del, b_calc, b_save, b_load, b_gen):
            b.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Fixed)
            h.addWidget(b)
        h.addStretch(1)
        v.addLayout(h)

        for row in SAMPLE:
            self.add_row(row)
        self.refresh_records_combo()
        self.calculate(silent=True)

    # ---------- Row management ----------
    @safe_slot
    def add_row(self, data=None):
        r = self.table.rowCount()
        self.table.insertRow(r)

        combo = QComboBox()
        combo.setEditable(True)
        combo.addItems(ASSET_OPTIONS)
        preset_name = data[0] if data else ASSET_OPTIONS[0]
        combo.setCurrentText(preset_name if preset_name in ASSET_OPTIONS or not preset_name else preset_name)
        combo.currentTextChanged.connect(lambda text, row=r: self.on_asset_changed(row, text))
        self.table.setCellWidget(r, ASSET_COL, combo)

        # data[0] is the asset name (already consumed above by the combo box),
        # so data[c] maps directly to column c for every remaining column - NOT
        # data[c-1]. The previous off-by-one here silently shifted every field
        # one column left whenever a full saved/pasted row was loaded (e.g. Rate%
        # ended up holding the asset name, Deduction Date ended up holding the
        # deduction amount, etc.), which broke both the depreciation calculation
        # and the outside-FY date check.
        for c in range(1, len(COLS)):
            val = ""
            if data and c < len(data):
                val = data[c]
            item = QTableWidgetItem(str(val))
            if c in READONLY_COLS:
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                item.setBackground(QColor("#f1edfd"))
            self.table.setItem(r, c, item)

        if data is None or not (len(data) > 1 and data[1]):
            self.on_asset_changed(r, combo.currentText())

    def on_asset_changed(self, row, text):
        if text in DEP_RATES:
            rate_item = self.table.item(row, RATE_COL)
            if rate_item is None:
                rate_item = QTableWidgetItem()
                self.table.setItem(row, RATE_COL, rate_item)
            rate_item.setText(str(DEP_RATES[text]))

    @safe_slot
    def del_row(self):
        r = self.table.currentRow()
        if r >= 0:
            self.table.removeRow(r)

    def clear_all_rows(self):
        while self.table.rowCount() > 0:
            self.table.removeRow(0)

    def get_row_data(self, r):
        combo = self.table.cellWidget(r, ASSET_COL)
        asset_name = combo.currentText() if combo else ""
        rest = [self.table.item(r, c).text() if self.table.item(r, c) else "" for c in range(1, len(COLS))]
        return [asset_name] + rest

    # ---------- Calculation ----------
    @safe_slot
    def calculate(self, silent=False):
        fy_start, fy_end = fy_bounds(self.fy_combo.currentText())
        fy_label = self.fy_combo.currentText().strip() or DEFAULT_FY
        warnings = []
        for r in range(self.table.rowCount()):
            d = self.get_row_data(r)
            (status, dep, closing, add_err, ded_err,
             _full, _half, _ded) = compute_row(d, fy_start, fy_end)

            self.table.item(r, 8).setText(status)
            self.table.item(r, 9).setText(f"{dep:,.2f}")
            self.table.item(r, 10).setText(f"{closing:,.2f}")

            add_item = self.table.item(r, ADD_DATE_COL)
            if add_item:
                add_item.setBackground(RED_TINT if add_err else TRANSPARENT)
            ded_item = self.table.item(r, DED_DATE_COL)
            if ded_item:
                ded_item.setBackground(RED_TINT if ded_err else TRANSPARENT)

            asset_label = d[0] or f"Row {r + 1}"
            if add_err:
                warnings.append(f"• {asset_label} (row {r + 1}): Addition date is outside FY {fy_label}")
            if ded_err:
                warnings.append(f"• {asset_label} (row {r + 1}): Deduction date is outside FY {fy_label}")

        if not silent:
            if warnings:
                QMessageBox.warning(self, "Date Outside Financial Year",
                    "The following amounts were EXCLUDED from this year's depreciation "
                    "because their date falls outside FY " + fy_label + ":\n\n" +
                    "\n".join(warnings) +
                    "\n\nCorrect the date, or change the row's addition/deduction to the "
                    "correct year.")
            else:
                QMessageBox.information(self, "Done", "Depreciation calculated for all rows.")

    # ---------- Save / Load records (per Assessee Name + Financial Year) ----------
    def refresh_records_combo(self, select_key=None):
        data = load_all_records()
        self.records_combo.blockSignals(True)
        self.records_combo.clear()
        for key in sorted(data.keys()):
            name, _, fy = key.partition("||")
            self.records_combo.addItem(f"{name} — FY {fy}", key)
        if select_key is not None:
            idx = self.records_combo.findData(select_key)
            if idx >= 0:
                self.records_combo.setCurrentIndex(idx)
        self.records_combo.blockSignals(False)

    @safe_slot
    def save_record(self):
        name = self.name_input.text().strip()
        fy = self.fy_combo.currentText().strip() or DEFAULT_FY
        if not name:
            QMessageBox.warning(self, "Assessee Name Required",
                "Please enter an Assessee Name before saving, so the record can be found again later.")
            return
        self.calculate(silent=True)
        rows = [self.get_row_data(r) for r in range(self.table.rowCount())]
        data = load_all_records()
        key = record_key(name, fy)
        data[key] = rows
        if save_all_records(data):
            self.refresh_records_combo(select_key=key)
            QMessageBox.information(self, "Saved",
                f"Data saved for '{name}' — FY {fy}.\nUse Load Record next time to bring it back instantly.")
        else:
            QMessageBox.critical(self, "Save Failed", "Could not write the saved-data file on this computer.")

    @safe_slot
    def load_record(self):
        idx = self.records_combo.currentIndex()
        if idx < 0:
            QMessageBox.information(self, "No Record Selected",
                "There are no saved records yet, or none is selected. Save one first.")
            return
        key = self.records_combo.itemData(idx)
        data = load_all_records()
        rows = data.get(key)
        if rows is None:
            QMessageBox.warning(self, "Not Found", "That saved record could not be found.")
            self.refresh_records_combo()
            return
        name, _, fy = key.partition("||")
        confirm = QMessageBox.question(self, "Load Record",
            f"This will replace all rows currently in the table with the saved data for "
            f"'{name}' — FY {fy}. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if confirm != QMessageBox.StandardButton.Yes:
            return
        self.name_input.setText(name)
        self.fy_combo.setCurrentText(fy)
        self.clear_all_rows()
        for row in rows:
            self.add_row(row)
        if self.table.rowCount() == 0:
            self.add_row()
        self.calculate(silent=True)

    # ---------- Excel export ----------
    @safe_slot
    def generate_excel(self):
        self.calculate(silent=True)
        fy_start, fy_end = fy_bounds(self.fy_combo.currentText())
        fy_label = self.fy_combo.currentText().strip() or DEFAULT_FY
        assessee = self.name_input.text().strip() or "N/A"
        rows = [self.get_row_data(r) for r in range(self.table.rowCount())]

        wb = Workbook()
        header_fill = PatternFill("solid", fgColor="241B4E")
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        title_font = Font(bold=True, size=14, color="241B4E", name="Calibri")
        info_font = Font(bold=True, size=11, color="7C4DFF", name="Calibri")
        thin = Side(style="thin", color="D8CFF7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        total_fill = PatternFill("solid", fgColor="EDE7FB")

        def write_info_header(ws, last_col_letter):
            ws["A1"] = "FORM NO. 3CD — Clause 18: Particulars of Depreciation Allowable"
            ws["A1"].font = title_font
            ws.merge_cells(f"A1:{last_col_letter}1")
            ws["A2"] = f"Assessee: {assessee}      Financial Year: {fy_label}"
            ws["A2"].font = info_font
            ws.merge_cells(f"A2:{last_col_letter}2")

        # pre-compute every row once, shared by all sheets - avoids duplicated/
        # inconsistent logic that previously caused mismatched totals.
        computed = [compute_row(d, fy_start, fy_end) for d in rows]

        # ---- Sheet 1: Form 3CD Clause 18 ----
        ws1 = wb.active
        ws1.title = "Form 3CD Clause 18"
        write_info_header(ws1, "K")
        headers = ["Description of Block of Assets", "Rate of Depreciation (%)",
            "WDV / Actual Cost at beginning of year", "Additions", "Date of Addition",
            "Deductions (incl. Sale Consideration)", "Date of Deduction",
            "Depreciation Allowable", "WDV at end of year", "Business Use %", "180-Day Classification"]
        for c, h in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
        r0 = 5
        for d, res in zip(rows, computed):
            status, dep, closing, add_err, ded_err, full_add, half_add, eff_ded = res
            vals = [d[0], num(d[1]), num(d[2]), num(d[3]) if not add_err else 0, parse_date(d[4]),
                     eff_ded, parse_date(d[6]), dep, closing, num(d[7]), status]
            for c, v in enumerate(vals, 1):
                cell = ws1.cell(row=r0, column=c, value=v)
                cell.border = border
                if isinstance(v, date):
                    cell.number_format = "DD-MM-YYYY"
                if c in (2, 3, 4, 6, 8, 9, 10):
                    cell.number_format = "#,##0.00"
                if add_err and c in (4, 5):
                    cell.font = Font(color="C0392B", italic=True)
                if ded_err and c in (6, 7):
                    cell.font = Font(color="C0392B", italic=True)
            r0 += 1
        last_data_row = r0 - 1
        ws1.cell(row=r0, column=1, value="TOTAL").font = Font(bold=True)
        if last_data_row >= 5:
            for c in (3, 4, 6, 8, 9):
                col = get_column_letter(c)
                cell = ws1.cell(row=r0, column=c, value=f"=SUM({col}5:{col}{last_data_row})")
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.number_format = "#,##0.00"
        ws1.freeze_panes = "A5"
        ws1.auto_filter.ref = f"A4:K{last_data_row}"
        for i, w in enumerate([26, 10, 18, 14, 14, 20, 14, 16, 16, 12, 26], 1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        # ---- Sheet 2: Copy-Paste Data (PRIMARY) ----
        ws2 = wb.create_sheet("Copy-Paste Data")
        write_info_header(ws2, "J")
        cp_headers = ["Description of Asset/Block", "Rate (%)", "Opening WDV/Actual Cost",
            "Additions", "Date of Addition", "Deductions (incl. Sale Consideration)",
            "Date of Deduction", "Depreciation Allowable", "Closing WDV", "Business Use %"]
        for c, h in enumerate(cp_headers, 1):
            cell = ws2.cell(row=4, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = border
        rr = 5
        for d, res in zip(rows, computed):
            status, dep, closing, add_err, ded_err, full_add, half_add, eff_ded = res
            vals = [d[0], num(d[1]), num(d[2]), num(d[3]) if not add_err else 0, parse_date(d[4]),
                     eff_ded, parse_date(d[6]), dep, closing, num(d[7])]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=rr, column=c, value=v)
                cell.border = border
                if isinstance(v, date):
                    cell.number_format = "DD-MM-YYYY"
                if c in (2, 3, 4, 6, 8, 9):
                    cell.number_format = "#,##0.00"
            rr += 1
        last_cp = rr - 1
        ws2.cell(row=rr, column=1, value="TOTAL").font = Font(bold=True)
        if last_cp >= 5:
            for c in (3, 4, 6, 8, 9):
                col = get_column_letter(c)
                cell = ws2.cell(row=rr, column=c, value=f"=SUM({col}5:{col}{last_cp})")
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.number_format = "#,##0.00"
        ws2.freeze_panes = "A5"
        ws2.auto_filter.ref = f"A4:J{last_cp}"
        for i, w in enumerate([26, 10, 18, 14, 14, 20, 14, 16, 16, 12], 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # ---- Sheet 3: Block Summary ----
        ws3 = wb.create_sheet("Block Summary")
        write_info_header(ws3, "G")
        bs_headers = ["Block", "Rate %", "Opening WDV", "Additions", "Deductions", "Depreciation", "Closing WDV"]
        for c, h in enumerate(bs_headers, 1):
            cell = ws3.cell(row=4, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        blocks = {}
        for d, res in zip(rows, computed):
            status, dep, closing, add_err, ded_err, full_add, half_add, eff_ded = res
            key = d[0] or "(Unnamed Block)"
            b = blocks.setdefault(key, {"rate": d[1], "opening": 0, "add": 0, "ded": 0, "dep": 0, "closing": 0})
            b["opening"] += num(d[2])
            b["add"] += 0 if add_err else num(d[3])
            b["ded"] += eff_ded
            b["dep"] += dep
            b["closing"] += closing
        rr = 5
        for k, b in blocks.items():
            vals = [k, num(b["rate"]), b["opening"], b["add"], b["ded"], b["dep"], b["closing"]]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=rr, column=c, value=v)
                cell.border = border
                if c > 1:
                    cell.number_format = "#,##0.00"
            rr += 1
        ws3.freeze_panes = "A5"
        for i, w in enumerate([26, 10, 16, 14, 14, 16, 16], 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        # ---- Sheet 4: Depreciation Working ----
        ws4 = wb.create_sheet("Depreciation Working")
        write_info_header(ws4, "J")
        wk_headers = ["Block", "Rate %", "Opening WDV", "Full-Rate Additions", "Half-Rate Additions",
            "Deductions", "Dep on Full-Rate Portion", "Dep on Half-Rate Portion",
            "Total Depreciation", "Closing WDV"]
        for c, h in enumerate(wk_headers, 1):
            cell = ws4.cell(row=4, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        rr = 5
        for d, res in zip(rows, computed):
            status, dep, closing, add_err, ded_err, full_add, half_add, eff_ded = res
            rate = num(d[1]) / 100.0
            opening = num(d[2])
            biz_pct = num(d[7]) / 100.0 if d[7] else 1.0
            dep_full = max(opening - eff_ded + full_add, 0) * rate * biz_pct
            dep_half = max(half_add, 0) * rate * 0.5 * biz_pct
            vals = [d[0], num(d[1]), opening, full_add, half_add, eff_ded, dep_full, dep_half, dep, closing]
            for c, v in enumerate(vals, 1):
                cell = ws4.cell(row=rr, column=c, value=v)
                cell.border = border
                if c > 1:
                    cell.number_format = "#,##0.00"
            rr += 1
        ws4.freeze_panes = "A5"
        for i, w in enumerate([26, 10, 14, 16, 16, 14, 18, 18, 16, 16], 1):
            ws4.column_dimensions[get_column_letter(i)].width = w

        # ---- Sheet 5: Reconciliation ----
        ws5 = wb.create_sheet("Reconciliation")
        write_info_header(ws5, "F")
        # NOTE: headers must not start with '+', '-' or '=' - Excel/openpyxl treats
        # such leading characters as the start of a formula, which previously
        # produced #NAME? errors on this sheet.
        recon_headers = ["Block", "Opening WDV", "Additions (+)", "Deductions (-)",
            "Depreciation (-)", "Closing WDV (=)"]
        for c, h in enumerate(recon_headers, 1):
            cell = ws5.cell(row=4, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        rr = 5
        tot = [0.0] * 5
        for d, res in zip(rows, computed):
            status, dep, closing, add_err, ded_err, full_add, half_add, eff_ded = res
            opening = num(d[2])
            add = 0.0 if add_err else num(d[3])
            vals = [d[0], opening, add, eff_ded, dep, closing]
            for c, v in enumerate(vals, 1):
                cell = ws5.cell(row=rr, column=c, value=v)
                cell.border = border
                if c > 1:
                    cell.number_format = "#,##0.00"
            for i, v in enumerate([opening, add, eff_ded, dep, closing]):
                tot[i] += v
            rr += 1
        last_recon_row = rr - 1
        ws5.cell(row=rr, column=1, value="TOTAL").font = Font(bold=True)
        if last_recon_row >= 5:
            for i, v in enumerate(tot, 2):
                cell = ws5.cell(row=rr, column=i, value=v)
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.number_format = "#,##0.00"
        ws5.freeze_panes = "A5"
        for i, w in enumerate([26, 16, 14, 14, 16, 16], 1):
            ws5.column_dimensions[get_column_letter(i)].width = w

        fname = "Form_3CD_Depreciation_Report.xlsx"
        wb.save(fname)
        QMessageBox.information(self, "Excel Generated", f"Saved: {fname}")


def main():
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
