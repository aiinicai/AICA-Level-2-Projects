import io
import datetime as dt
from fastapi import APIRouter, Request, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..database import get_db
from .. import models, permissions
from ..auth import get_current_user, log_audit
from ..main import templates

router = APIRouter()

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _clean_filters(department_id, employee_id, start_date, end_date):
    """HTML <select>/<input> send an empty string for an unselected/blank
    field - treat that (and any other blank/whitespace value) as 'no filter'."""
    department_id = int(department_id) if department_id and department_id.strip() else None
    employee_id = int(employee_id) if employee_id and employee_id.strip() else None
    start_date = start_date if start_date and start_date.strip() else None
    end_date = end_date if end_date and end_date.strip() else None
    d_start = dt.date.fromisoformat(start_date) if start_date else dt.date.today().replace(day=1)
    d_end = dt.date.fromisoformat(end_date) if end_date else dt.date.today()
    return department_id, employee_id, d_start, d_end


def _gather_report(db: Session, user: models.User, department_id, employee_id, d_start, d_end):
    """Single source of truth for report data - used identically by the
    on-screen preview and the Excel export, so they can never disagree."""
    scope_users = permissions.export_scope_users(db, user)
    scope_ids = {u.id for u in scope_users}

    if employee_id:
        scope_ids &= {employee_id}
    elif department_id:
        scope_ids &= {u.id for u in scope_users if u.department_id == department_id}

    users = [u for u in scope_users if u.id in scope_ids]

    summary_rows, attendance_rows, task_rows = [], [], []
    for u in users:
        days_present = db.query(models.Attendance).filter(
            models.Attendance.user_id == u.id, models.Attendance.date >= d_start, models.Attendance.date <= d_end,
            models.Attendance.status.in_(["WFO", "WFH", "HALF_DAY"]),
        ).count()
        leaves_taken = db.query(models.LeaveRequest).filter(
            models.LeaveRequest.user_id == u.id, models.LeaveRequest.status == "APPROVED",
            models.LeaveRequest.start_date <= d_end, models.LeaveRequest.end_date >= d_start,
        ).count()
        pending_tasks = db.query(models.CalendarEvent).filter(
            models.CalendarEvent.owner_id == u.id, models.CalendarEvent.status.in_(["OPEN", "IN_PROGRESS"]),
            models.CalendarEvent.event_type.in_(["TASK", "PENDING_ACTION"]),
        ).count()
        summary_rows.append([u.employee_code, u.name, u.department.name if u.department else "", u.designation or "",
                              u.employment_status, days_present, leaves_taken, pending_tasks])

        for r in db.query(models.Attendance).filter(
            models.Attendance.user_id == u.id, models.Attendance.date >= d_start, models.Attendance.date <= d_end,
        ).order_by(models.Attendance.date).all():
            attendance_rows.append([u.employee_code, u.name, r.date.isoformat(), r.status])

        for e in db.query(models.CalendarEvent).filter(
            models.CalendarEvent.owner_id == u.id, models.CalendarEvent.start_date >= d_start,
            models.CalendarEvent.start_date <= d_end,
        ).order_by(models.CalendarEvent.start_date).all():
            task_rows.append([u.employee_code, u.name, e.event_type, e.title, e.start_date.isoformat(),
                               e.priority, e.status, e.raw_prompt_text or ""])

    confidential_rows = None
    permitted_all = all(permissions.can_view_confidential(db, user, u) for u in users) if users else False
    if permitted_all:
        confidential_rows = []
        for u in users:
            leaves = db.query(models.LeaveRequest).filter(models.LeaveRequest.user_id == u.id).order_by(models.LeaveRequest.start_date).all()
            base = [u.employee_code, u.name, u.resignation_date.isoformat() if u.resignation_date else "",
                    u.notice_period_days or "", u.last_working_day.isoformat() if u.last_working_day else "", u.employment_status]
            if not leaves:
                confidential_rows.append(base + ["", "", "", ""])
            for lv in leaves:
                confidential_rows.append(base + [lv.start_date.isoformat(), lv.end_date.isoformat(),
                                                  lv.leave_type.name if lv.leave_type else "", lv.status])

    return {
        "users": users, "d_start": d_start, "d_end": d_end,
        "summary_rows": summary_rows, "attendance_rows": attendance_rows,
        "task_rows": task_rows, "confidential_rows": confidential_rows,
    }


SUMMARY_HEADERS = ["Employee Code", "Name", "Department", "Designation", "Status", "Days Present", "Leaves Taken", "Pending Tasks"]
ATTENDANCE_HEADERS = ["Employee Code", "Name", "Date", "Status"]
TASK_HEADERS = ["Employee Code", "Name", "Type", "Title", "Date", "Priority", "Status", "Original Prompt"]
CONFIDENTIAL_HEADERS = ["Employee Code", "Name", "Resignation Date", "Notice Period (days)", "Last Working Day",
                         "Employment Status", "Planned Leave Start", "Planned Leave End", "Leave Type", "Leave Status"]


@router.get("/reports")
def reports_page(
    request: Request,
    department_id: str = Query(None), employee_id: str = Query(None),
    start_date: str = Query(None), end_date: str = Query(None),
    preview: str = Query(None),
    db: Session = Depends(get_db), user: models.User = Depends(get_current_user),
):
    scope_users = permissions.export_scope_users(db, user)
    if permissions.is_ceo(user):
        departments = db.query(models.Department).all()
    else:
        dept_ids = {u.department_id for u in scope_users if u.department_id}
        departments = db.query(models.Department).filter(models.Department.id.in_(dept_ids)).all()

    report = None
    if preview:
        dep_id, emp_id, d_start, d_end = _clean_filters(department_id, employee_id, start_date, end_date)
        report = _gather_report(db, user, dep_id, emp_id, d_start, d_end)

    return templates.TemplateResponse(request, "reports.html", {
        "request": request, "user": user, "scope_users": scope_users, "departments": departments,
        "report": report, "summary_headers": SUMMARY_HEADERS, "attendance_headers": ATTENDANCE_HEADERS,
        "task_headers": TASK_HEADERS, "confidential_headers": CONFIDENTIAL_HEADERS,
        "filters": {"department_id": department_id or "", "employee_id": employee_id or "",
                    "start_date": start_date or "", "end_date": end_date or ""},
    })


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    ws.freeze_panes = "A2"


@router.get("/reports/export")
def export_excel(
    request: Request,
    department_id: str = Query(None), employee_id: str = Query(None),
    start_date: str = Query(None), end_date: str = Query(None),
    db: Session = Depends(get_db), user: models.User = Depends(get_current_user),
):
    dep_id, emp_id, d_start, d_end = _clean_filters(department_id, employee_id, start_date, end_date)
    report = _gather_report(db, user, dep_id, emp_id, d_start, d_end)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(SUMMARY_HEADERS)
    for row in report["summary_rows"]:
        ws.append(row)
    _style_header(ws, len(SUMMARY_HEADERS))

    ws2 = wb.create_sheet("Attendance Detail")
    ws2.append(ATTENDANCE_HEADERS)
    for row in report["attendance_rows"]:
        ws2.append(row)
    _style_header(ws2, len(ATTENDANCE_HEADERS))

    ws3 = wb.create_sheet("Tasks & Calendar")
    ws3.append(TASK_HEADERS)
    for row in report["task_rows"]:
        ws3.append(row)
    _style_header(ws3, len(TASK_HEADERS))

    if report["confidential_rows"] is not None:
        ws4 = wb.create_sheet("Leave & Notice (Confidential)")
        ws4.append(CONFIDENTIAL_HEADERS)
        for row in report["confidential_rows"]:
            ws4.append(row)
        _style_header(ws4, len(CONFIDENTIAL_HEADERS))

    log_audit(db, user, "EXPORT_EXCEL", f"Exported report for {len(report['users'])} employee(s), {report['d_start']} to {report['d_end']}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"company_os_report_{dt.date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
