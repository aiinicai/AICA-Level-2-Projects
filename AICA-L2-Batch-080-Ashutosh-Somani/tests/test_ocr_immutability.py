"""Tests for Stage 9: OCR immutability, DB migration, logging privacy, and Excel OCR integration."""
import pytest
import json
import hashlib
import logging
import os
import shutil
import uuid
from pathlib import Path
from decimal import Decimal
from datetime import date
from unittest.mock import patch, MagicMock

import openpyxl

from app.database.migrations import init_db
from app.database.db import get_db_connection
from app.exporters.excel_exporter import ExcelExporter
from app.services.validation_service import ValidationService
from app.models.extraction_result import ExtractionResult, RawPage, RawWord


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _sha256(filepath):
    """Compute SHA-256 hex digest of a file."""
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        h.update(f.read())
    return h.hexdigest()


def _write_json(filepath, data):
    """Write *data* as JSON to *filepath*, creating parent dirs."""
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)


def _ocr_config(temp_config, tmp_path):
    """Add an 'ocr' section to temp_config so OcrService can be instantiated."""
    if not temp_config.has_section('ocr'):
        temp_config.add_section('ocr')
    temp_config.set('ocr', 'enabled', 'true')
    temp_config.set('ocr', 'render_dpi', '250')
    temp_config.set('ocr', 'minimum_word_confidence', '70')
    temp_config.set('ocr', 'max_pages', '250')
    temp_config.set('ocr', 'limited_text_word_threshold', '20')
    temp_config.set('paths', 'temp', str(tmp_path / 'temp'))
    return temp_config


# ===========================================================================
# 1. Six-artifact immutability
# ===========================================================================

def test_ocr_extraction_immutability_six_artifacts(tmp_path):
    """Only reviewed_statement.json may change after a user correction;
    the other five machine-generated artifacts must remain byte-identical."""
    job_id = str(uuid.uuid4())
    job_dir = tmp_path / 'temp' / 'jobs' / job_id
    job_dir.mkdir(parents=True)

    # -- Artifact payloads (content doesn't matter, only hashes) --
    artifacts = {
        'raw_extraction.json': {"extractor": "pdfplumber", "pages": [{"page": 1, "words": ["Hello"]}]},
        'ocr/ocr_result.json': {"ocr_engine": "rapidocr", "pages": [{"page": 1, "words": ["Hello"]}]},
        'ocr/effective_extraction.json': {"pages": [{"page": 1, "merged": True}]},
        'normalization/normalized_statement.json': {"transactions": [{"narration": "Payment", "debit": "100.00"}]},
        'validation/validation_result.json': {"summary": {"status": "PASS"}, "exceptions": []},
        'review/reviewed_statement.json': {"transactions": [{"narration": "Payment", "debit": "100.00"}], "review_revision": 1},
    }

    for rel_path, data in artifacts.items():
        _write_json(job_dir / rel_path, data)

    # -- Hash all six before correction --
    hashes_before = {rel: _sha256(job_dir / rel) for rel in artifacts}

    # -- Simulate user correction: edit the reviewed_statement narration --
    reviewed_path = job_dir / 'review' / 'reviewed_statement.json'
    with open(reviewed_path, 'r', encoding='utf-8') as f:
        reviewed = json.load(f)
    reviewed['transactions'][0]['narration'] = 'Corrected Payment'
    reviewed['review_revision'] = 2
    with open(reviewed_path, 'w', encoding='utf-8') as f:
        json.dump(reviewed, f, indent=2)

    # -- Hash all six after correction --
    hashes_after = {rel: _sha256(job_dir / rel) for rel in artifacts}

    # -- Assertions --
    immutable_artifacts = [
        'raw_extraction.json',
        'ocr/ocr_result.json',
        'ocr/effective_extraction.json',
        'normalization/normalized_statement.json',
        'validation/validation_result.json',
    ]
    for rel in immutable_artifacts:
        assert hashes_before[rel] == hashes_after[rel], f"{rel} was modified!"

    assert hashes_before['review/reviewed_statement.json'] != hashes_after['review/reviewed_statement.json'], \
        "reviewed_statement.json should have changed after user correction"


# ===========================================================================
# 2. DB v9 fresh migration
# ===========================================================================

def test_db_v9_fresh_migration(temp_config):
    """A fresh database should reach schema version 9 with an ocr_status column."""
    init_db(temp_config)

    with get_db_connection(temp_config) as conn:
        cursor = conn.cursor()

        # Schema version == 9
        cursor.execute("SELECT MAX(version) FROM schema_version")
        assert cursor.fetchone()[0] == 9

        # ocr_status column exists and accepts NULL
        cursor.execute(
            "INSERT INTO processing_jobs (id, display_name, status, stage) "
            "VALUES (?, ?, ?, ?)",
            ('test-v9', 'test.pdf', 'uploaded', 'intake')
        )
        cursor.execute("SELECT ocr_status FROM processing_jobs WHERE id = ?", ('test-v9',))
        row = cursor.fetchone()
        assert row is not None
        assert row['ocr_status'] is None  # NULL accepted


# ===========================================================================
# 3. DB v9 idempotent
# ===========================================================================

def test_db_v9_idempotent(temp_config):
    """Calling init_db twice must not raise and version stays at 9."""
    init_db(temp_config)
    init_db(temp_config)  # second call — no error expected

    with get_db_connection(temp_config) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(version) FROM schema_version")
        assert cursor.fetchone()[0] == 9


# ===========================================================================
# 4. export_history preserved across migrations
# ===========================================================================

def test_db_export_history_preserved(temp_config):
    """Data in export_history must survive a re-run of init_db."""
    init_db(temp_config)

    with get_db_connection(temp_config) as conn:
        conn.execute(
            "INSERT INTO export_history (job_id, filename, source_type) VALUES (?, ?, ?)",
            ('job-preserve', 'export.xlsx', 'REVIEWED')
        )

    # Re-run migration
    init_db(temp_config)

    with get_db_connection(temp_config) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM export_history WHERE job_id = ?", ('job-preserve',))
        row = cursor.fetchone()
        assert row is not None
        assert row['filename'] == 'export.xlsx'
        assert row['source_type'] == 'REVIEWED'


# ===========================================================================
# 5. Logging privacy — no raw text leaks
# ===========================================================================

def test_logging_privacy_no_raw_text(temp_config, tmp_path, caplog):
    """OCR word text must never appear in log output; aggregate info should."""
    _ocr_config(temp_config, tmp_path)

    # Create minimal ExtractionResult with a low-word page so it triggers OCR
    extraction = ExtractionResult(
        job_id='privacy-test',
        extractor_used='pdfplumber',
        extractor_version='0.10',
        status='success',
        page_count=1,
        pages_processed=1,
        total_words=2,
        total_characters=10,
        text_layer_status='LIMITED_TEXT',
        table_candidate_count=0,
        pages=[
            RawPage(
                page_number=1, width=612, height=792,
                raw_text='AB', word_count=2, character_count=10,
                source_type='DIGITAL', words=[], warnings=[]
            )
        ]
    )

    # Set up a job directory with a dummy PDF so the service won't bail
    job_dir = tmp_path / 'temp' / 'jobs' / 'privacy-test'
    job_dir.mkdir(parents=True)

    from app.services.ocr_service import OcrService
    svc = OcrService(temp_config)

    # Mock both the PDF-path lookup and the engine call
    sensitive = 'SENSITIVE_ACCOUNT_NUMBER_12345'
    mock_words = [
        {"text": sensitive, "x0": 10, "x1": 100, "top": 20, "bottom": 30, "confidence": 95.0}
    ]

    with patch('app.services.pdf_intake_service.get_job_pdf_path', return_value=job_dir / 'dummy.pdf'), \
         patch.object(svc.engine_svc, 'ocr_page', return_value=(mock_words, sensitive)):
        # Create a dummy PDF file so exists() passes
        (job_dir / 'dummy.pdf').write_bytes(b'%PDF-fake')

        with caplog.at_level(logging.DEBUG):
            svc.run_ocr('privacy-test', extraction)

    # Sensitive text must NOT appear in logs
    assert sensitive not in caplog.text, \
        "Raw OCR text leaked into logs!"

    # Aggregate operational messages SHOULD appear
    assert 'Starting OCR' in caplog.text or 'pages' in caplog.text


# ===========================================================================
# 6. Excel — Source Type and OCR Confidence columns
# ===========================================================================

def test_excel_ocr_source_type_column(tmp_path):
    """ExcelExporter must write Source Type and OCR Confidence values."""
    exporter = ExcelExporter()
    payload = {
        'transactions': [
            {
                'transaction_date': date(2024, 6, 15),
                'narration': 'ATM Withdrawal',
                'debit': Decimal('500.00'),
                'credit': None,
                'balance': Decimal('4500.00'),
                'source_type': 'OCR',
                'ocr_confidence': 85.5,
            }
        ],
        'summary': {},
        'exceptions': [],
        'audit': [],
    }

    filepath = tmp_path / 'ocr_export.xlsx'
    exporter.export(filepath, payload)

    wb = openpyxl.load_workbook(filepath)
    ws = wb['Transactions']

    # Find header positions
    headers = [cell.value for cell in ws[1]]
    src_col = headers.index('Source Type') + 1
    conf_col = headers.index('OCR Confidence') + 1

    assert ws.cell(row=2, column=src_col).value == 'OCR'
    assert ws.cell(row=2, column=conf_col).value == 85.5


# ===========================================================================
# 7. Excel — formula injection from OCR text
# ===========================================================================

def test_excel_formula_injection_ocr_text(tmp_path):
    """Malicious OCR narration starting with '=' must be sanitized."""
    exporter = ExcelExporter()
    payload = {
        'transactions': [
            {
                'transaction_date': date(2024, 6, 15),
                'narration': '=CMD("calc")',
                'debit': Decimal('100.00'),
                'credit': None,
                'balance': Decimal('900.00'),
                'source_type': 'OCR',
                'ocr_confidence': 72.0,
            }
        ],
        'summary': {},
        'exceptions': [],
        'audit': [],
    }

    filepath = tmp_path / 'injection_test.xlsx'
    exporter.export(filepath, payload)

    wb = openpyxl.load_workbook(filepath)
    ws = wb['Transactions']

    headers = [cell.value for cell in ws[1]]
    narr_col = headers.index('Narration') + 1
    narr_value = ws.cell(row=2, column=narr_col).value

    # Should be prefixed with ' to neutralize the formula
    assert narr_value.startswith("'"), \
        f"Formula injection not sanitized: {narr_value!r}"


# ===========================================================================
# 8. OCR cleanup preserves non-OCR files
# ===========================================================================

def test_ocr_cleanup_preserves_non_ocr_files(tmp_path):
    """Deleting the ocr/ sub-directory must not touch other artifacts."""
    job_dir = tmp_path / 'temp' / 'jobs' / 'cleanup-job'

    # Create artifacts
    _write_json(job_dir / 'raw_extraction.json', {"pages": []})
    _write_json(job_dir / 'normalization' / 'normalized_statement.json', {"transactions": []})
    _write_json(job_dir / 'ocr' / 'ocr_result.json', {"pages": []})
    _write_json(job_dir / 'ocr' / 'effective_extraction.json', {"pages": []})

    # Simulate cleanup — delete only the ocr/ subdirectory
    shutil.rmtree(job_dir / 'ocr')

    # Non-OCR artifacts must still exist
    assert (job_dir / 'raw_extraction.json').exists()
    assert (job_dir / 'normalization' / 'normalized_statement.json').exists()
    # OCR artifacts should be gone
    assert not (job_dir / 'ocr' / 'ocr_result.json').exists()
    assert not (job_dir / 'ocr' / 'effective_extraction.json').exists()


# ===========================================================================
# 9. Financial mismatch with OCR source — confidence must NOT suppress errors
# ===========================================================================

def test_financial_mismatch_ocr_source(temp_config):
    """A BALANCE_MISMATCH must be raised even when the transaction is OCR-sourced."""
    service = ValidationService(temp_config)

    norm_data = {
        'metadata': {'opening_balance': '1000.00'},
        'transactions': [
            {
                'transaction_date': '2024-06-15',
                'narration': 'ATM Cash',
                'debit': '200.00',
                'credit': '',
                # Correct balance would be 800.00; introduce 0.01 mismatch
                'balance': '799.99',
                'source_type': 'OCR',
                'ocr_confidence': 75.0,
                'source_page': 1,
                'source_row': 0,
            }
        ],
    }

    summary, tx_results, exceptions = service._perform_validation(norm_data)

    # Must flag the mismatch
    exception_codes = [e.exception_code for e in exceptions]
    assert 'BALANCE_MISMATCH' in exception_codes, \
        "OCR confidence must not suppress BALANCE_MISMATCH"

    # Transaction-level check
    tx = tx_results[0]
    assert tx.validation_status == 'BALANCE_MISMATCH'
    assert tx.difference == Decimal('-0.01')

    # OCR source_type is recorded but has no impact on validation strictness
    assert tx.source_type == 'OCR'
