import pytest
import json
import os
from pathlib import Path
from decimal import Decimal

from app.models.profile import BankProfile, ColumnDefinition, TableRegion
from app.services.profile_manager import ProfileManager
from app.services.profile_matcher import ProfileMatcher
from app.services.normalization_service import run_normalization, get_normalization_result
from app.services.review_service import ReviewService
from app.services.export_service import ExportService
from app.database.db import get_db_connection
from app.database.migrations import init_db
from openpyxl import load_workbook

def test_profile_persistence_and_list_after_restart(temp_config):
    init_db(temp_config)
    
    # 1. Create and save profile
    pm1 = ProfileManager(temp_config)
    prof = pm1.create_profile("Persistent Profile", "Bank of Baroda")
    prof.page_width = 800.0
    prof.page_height = 600.0
    prof.expected_header_signatures = ["Bank of Baroda"]
    prof.column_definitions = [
        ColumnDefinition(canonical_name="transaction_date", x0=10.0, x1=80.0),
        ColumnDefinition(canonical_name="narration", x0=160.0, x1=400.0),
        ColumnDefinition(canonical_name="debit", x0=510.0, x1=600.0),
        ColumnDefinition(canonical_name="credit", x0=620.0, x1=700.0),
        ColumnDefinition(canonical_name="balance", x0=730.0, x1=800.0)
    ]
    pm1.save_profile(prof)
    
    # 2. Simulate application restart (reinitialize ProfileManager)
    pm2 = ProfileManager(temp_config)
    profiles = pm2.list_profiles()
    
    assert len(profiles) >= 1
    found = next((p for p in profiles if p.profile_id == prof.profile_id), None)
    assert found is not None
    assert found.profile_name == "Persistent Profile"
    assert found.active is True
    assert found.revision_number == prof.revision_number
    
    # Verify SQLite index
    with get_db_connection(temp_config) as conn:
        row = conn.execute("SELECT profile_id, active FROM bank_profiles WHERE profile_id=?", (prof.profile_id,)).fetchone()
        assert row is not None
        assert row[0] == prof.profile_id
        assert row[1] == 1

def test_unknown_bank_auto_profile_match(temp_config):
    prof = BankProfile(
        profile_id="p_unknown_test",
        profile_name="BOB_Saving",
        bank_name="Bank of Baroda",
        page_width=800.0,
        page_height=600.0,
        expected_header_signatures=["Bank of Baroda"],
        column_definitions=[ColumnDefinition("transaction_date", 10.0, 80.0)]
    )
    
    matcher = ProfileMatcher([prof], temp_config)
    
    # Test layout match with Unknown Bank
    status, matched, score, details = matcher.match(
        bank_detected="Unknown Bank",
        page_width=800.0,
        page_height=600.0,
        extracted_text="Statement text containing Bank of Baroda header"
    )
    
    assert status == "AUTO_APPLIED"
    assert matched.profile_id == "p_unknown_test"
    assert score >= 80

def test_match_threshold_and_ambiguity_scenarios(temp_config):
    prof_a = BankProfile(
        profile_id="p_a",
        profile_name="Profile A",
        bank_name="Bank A",
        page_width=800.0,
        page_height=600.0,
        expected_header_signatures=["Header Signature A"]
    )
    prof_b = BankProfile(
        profile_id="p_b",
        profile_name="Profile B",
        bank_name="Bank B",
        page_width=800.0,
        page_height=600.0,
        expected_header_signatures=["Header Signature B"]
    )
    prof_inactive = BankProfile(
        profile_id="p_in",
        profile_name="Inactive",
        bank_name="Bank A",
        active=False
    )
    
    # 1. Inactive profile ignored
    matcher = ProfileMatcher([prof_a, prof_b, prof_inactive], temp_config)
    assert len(matcher.profiles) == 2
    
    # 2. Same bank label + wrong layout -> LOW_CONFIDENCE
    status, matched, score, _ = matcher.match(
        bank_detected="Bank A",
        page_width=100.0, # Wrong dimension
        page_height=100.0,
        extracted_text="No matching headers"
    )
    assert status == "SELECTION_REQUIRED"
    
    # 3. Similar dimensions + wrong headers -> LOW_CONFIDENCE
    status, matched, score, _ = matcher.match(
        bank_detected="Unknown Bank",
        page_width=800.0,
        page_height=600.0,
        extracted_text="Unrelated text"
    )
    assert status == "SELECTION_REQUIRED"
    
    # 4. Two profiles with similar score -> AMBIGUOUS
    status, matched, score, _ = matcher.match(
        bank_detected="Unknown Bank",
        page_width=800.0,
        page_height=600.0,
        extracted_text="Text with Header Signature A and Header Signature B"
    )
    assert status == "SELECTION_REQUIRED"
    assert matched is None

def test_stale_review_export_invalidation(temp_config, app):
    init_db(temp_config)
    
    # Setup job
    job_id = "test_stale_job"
    temp_dir = Path(temp_config.get('paths', 'temp'))
    job_dir = temp_dir / 'jobs' / job_id
    norm_dir = job_dir / 'normalization'
    norm_dir.mkdir(parents=True, exist_ok=True)
    
    # 1. Revision 1: generic normalization has 0 rows
    norm_data_rev1 = {
        "bank_detection": {"status": "unknown", "bank_name": "Unknown"},
        "metadata": {},
        "transactions": []
    }
    with open(norm_dir / 'normalized_statement.json', 'w') as f:
        json.dump(norm_data_rev1, f)
        
    # Create stale review artifact with 0 rows
    review_svc = ReviewService(temp_config)
    rev_stmt = review_svc.initialize_review(job_id)
    assert len(rev_stmt.transactions) == 0
    
    # Insert job into SQLite
    with get_db_connection(temp_config) as conn:
        conn.execute("INSERT OR REPLACE INTO processing_jobs (id, status, stage, display_name) VALUES (?, 'normalized', 'normalization', 'Test Job')", (job_id,))
        conn.commit()

    # 2. Revision 2: Profile normalization produces 3 transactions
    norm_data_rev2 = {
        "bank_detection": {"status": "unknown", "bank_name": "Unknown"},
        "metadata": {},
        "transactions": [
            {"transaction_date": "2026-01-01", "narration": "Tx 1", "debit": "50.00", "credit": None, "balance": "950.00"},
            {"transaction_date": "2026-01-02", "narration": "Tx 2", "debit": None, "credit": "200.00", "balance": "1150.00"},
            {"transaction_date": "2026-01-03", "narration": "Tx 3", "debit": "10.00", "credit": None, "balance": "1140.00"}
        ]
    }
    with open(norm_dir / 'normalized_statement.json', 'w') as f:
        json.dump(norm_data_rev2, f)
        
    # Also create machine validation artifact
    val_dir = job_dir / 'validation'
    val_dir.mkdir(parents=True, exist_ok=True)
    val_data = {
        "summary": {"validation_status": "PASS", "total_debits": "60.00", "total_credits": "200.00", "transaction_count": 3},
        "transactions": [],
        "exceptions": []
    }
    with open(val_dir / 'validation_result.json', 'w') as f:
        json.dump(val_data, f)

    # 3. Export to Excel
    with app.app_context():
        export_svc = ExportService(temp_config)
        wb_path = export_svc.export_excel(job_id)
        
    # 4. Verify Workbook content
    wb = load_workbook(wb_path)
    ws_tx = wb["Transactions"]
    # Check rows in Transactions sheet (header + 3 data rows = 4 rows)
    tx_rows = [row for row in ws_tx.iter_rows(values_only=True) if any(row)]
    assert len(tx_rows) >= 4
    
    ws_sum = wb["Summary"]
    summary_text = str(list(ws_sum.iter_rows(values_only=True)))
    assert "3" in summary_text # Transaction count 3

def test_review_after_profile_re_normalization(temp_config, app):
    init_db(temp_config)
    
    job_id = "test_review_rev2_job"
    temp_dir = Path(temp_config.get('paths', 'temp'))
    job_dir = temp_dir / 'jobs' / job_id
    norm_dir = job_dir / 'normalization'
    norm_dir.mkdir(parents=True, exist_ok=True)
    
    norm_data = {
        "bank_detection": {"status": "detected", "bank_name": "Test Bank"},
        "metadata": {},
        "transactions": [
            {"transaction_date": "2026-01-01", "narration": "Fresh Tx 1", "debit": "100.00", "credit": None, "balance": "900.00"}
        ]
    }
    with open(norm_dir / 'normalized_statement.json', 'w') as f:
        json.dump(norm_data, f)
        
    with get_db_connection(temp_config) as conn:
        conn.execute("INSERT OR REPLACE INTO processing_jobs (id, status, stage, display_name) VALUES (?, 'normalized', 'normalization', 'Test Job 2')", (job_id,))
        conn.commit()

    # Create fresh review after profile re-normalization
    review_svc = ReviewService(temp_config)
    rev_stmt = review_svc.initialize_review(job_id)
    assert len(rev_stmt.transactions) == 1
    assert rev_stmt.transactions[0].narration == "Fresh Tx 1"
    
    with app.app_context():
        export_svc = ExportService(temp_config)
        wb_path = export_svc.export_excel(job_id)
        
    wb = load_workbook(wb_path)
    ws_tx = wb["Transactions"]
    tx_rows = [row for row in ws_tx.iter_rows(values_only=True) if any(row)]
    assert len(tx_rows) >= 2 # Header + 1 transaction

def test_full_restart_auto_profile_and_export_e2e(client, app, temp_config):
    # Step 1: Upload PDF A & Create Profile
    pdf_path = Path('samples/synthetic_defect_test.pdf')
    with open(pdf_path, 'rb') as f:
        resp = client.post('/upload', data={'file': f})
    assert resp.status_code == 200
    job_id_a = resp.json['job_id']
    
    client.post(f'/jobs/{job_id_a}/extract')
    client.post(f'/jobs/{job_id_a}/normalize')
    
    # Get page dimensions from Job A extraction
    temp_dir = Path(temp_config.get('paths', 'temp'))
    raw_path = temp_dir / 'jobs' / job_id_a / 'extraction' / 'raw_extraction.json'
    with open(raw_path, 'r') as f:
        raw_data = json.load(f)
    p_w = raw_data['pages'][0]['width']
    p_h = raw_data['pages'][0]['height']

    # Create profile for synthetic PDF layout
    pm = ProfileManager(temp_config)
    prof = pm.create_profile("Synthetic Auto Profile", "Bank of Synthetic Baroda")
    prof.page_width = p_w
    prof.page_height = p_h
    prof.expected_header_signatures = ["Bank of Synthetic Baroda"]
    prof.column_definitions = [
        ColumnDefinition("date", 10.0, 80.0),
        ColumnDefinition("narration", 160.0, 440.0),
        ColumnDefinition("withdrawal", 510.0, 600.0),
        ColumnDefinition("deposit", 620.0, 700.0),
        ColumnDefinition("balance", 730.0, 800.0)
    ]
    prof.table_bbox = TableRegion(0.0, 0.0, p_w, p_h)
    pm.save_profile(prof)
    
    # Apply profile to Job A
    client.post(f'/jobs/{job_id_a}/normalize?profile_id={prof.profile_id}')
    
    # Step 2: Simulate Application Restart (re-read config/profiles)
    pm_restart = ProfileManager(temp_config)
    assert len(pm_restart.list_profiles()) >= 1
    
    # Step 3: Process PDF B with SAME layout
    with open(pdf_path, 'rb') as f:
        resp_b = client.post('/upload', data={'file': f})
    job_id_b = resp_b.json['job_id']
    
    client.post(f'/jobs/{job_id_b}/extract')
    
    # Trigger normalization on Job B -> Should AUTO-MATCH saved profile!
    resp_norm = client.post(f'/jobs/{job_id_b}/normalize')
    assert resp_norm.status_code == 302
    
    resp_sum = client.get(f'/jobs/{job_id_b}/normalization')
    html_b = resp_sum.data.decode()
    assert 'Rows Normalized:</strong> 3' in html_b
    assert 'Cannot resolve transaction table' not in html_b
    
    # Step 4: Export Excel for Job B
    resp_exp = client.post(f'/api/export/{job_id_b}')
    assert resp_exp.status_code == 200

def test_profile_privacy_no_confidential_data(temp_config):
    pm = ProfileManager(temp_config)
    prof = pm.create_profile("Privacy Test Profile", "Generic Bank")
    pm.save_profile(prof)
    
    prof_dict = prof.to_dict()
    forbidden_keys = [
        "customer_name", "account_number", "customer_id", "address",
        "narration", "amount", "balance", "password", "pdf_path"
    ]
    for key in forbidden_keys:
        assert key not in prof_dict
