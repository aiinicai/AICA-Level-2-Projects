import pytest
import os
import json
from pathlib import Path
import sqlite3
import datetime
import urllib.parse
from decimal import Decimal
import time

from app.database.migrations import init_db
from app.services.export_service import ExportService
from openpyxl import load_workbook
import openpyxl

def _create_mock_job(job_id, tmp_path, config):
    # Setup directories
    temp_dir = tmp_path / 'temp'
    jobs_dir = temp_dir / 'jobs'
    job_dir = jobs_dir / job_id
    
    norm_dir = job_dir / 'normalization'
    norm_dir.mkdir(parents=True, exist_ok=True)
    
    val_dir = job_dir / 'validation'
    val_dir.mkdir(parents=True, exist_ok=True)
    
    rev_dir = job_dir / 'review'
    rev_dir.mkdir(parents=True, exist_ok=True)
    
    # DB
    db_path = Path(config.get('paths', 'database'))
    # init_db is already called in the app fixture
    conn = sqlite3.connect(db_path)
    conn.execute("INSERT INTO processing_jobs (id, display_name, status, stage, source_filename, stored_filename, file_size, sha256, page_count, pdf_type, encrypted) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (job_id, f"{job_id}.pdf", "extracted", "normalization", f"{job_id}.pdf", job_id + ".pdf", 1000, job_id, 1, "digital", 0))
    conn.commit()
    conn.close()
    
    # Create normalized statement with 2000 transactions
    transactions = []
    for i in range(2000):
        transactions.append({
            "transaction_date": "2026-01-01",
            "value_date": "2026-01-01",
            "narration": f"=Row {i} =+-@",
            "debit": "100.00" if i % 2 == 0 else None,
            "credit": "100.00" if i % 2 != 0 else None,
            "balance": f"{100.00 * i:.2f}",
            "transaction_type": "DR" if i % 2 == 0 else "CR",
            "page_number": 1,
            "source_type": "digital"
        })
        
    norm_data = {
        "job_id": job_id,
        "metadata": {
            "bank_name": "Test Bank",
            "account_number": "12345678"
        },
        "transactions": transactions
    }
    with open(norm_dir / 'normalized_statement.json', 'w') as f:
        json.dump(norm_data, f)
        
    val_data = {
        "summary": {
            "validation_status": "PASS",
            "transaction_count": 2000,
            "validated_transaction_count": 2000,
            "balance_mismatch_count": 0,
            "transactions_not_verifiable": 0,
            "opening_balance": "0.00",
            "total_debits": "100000.00",
            "total_credits": "100000.00",
            "expected_closing_balance": "0.00",
            "statement_closing_balance": "0.00",
            "difference": "0.00"
        },
        "transactions": [{"transaction_index": i, "source_page": 1, "source_row": i, "validation_status": "VALID"} for i in range(2000)],
        "exceptions": []
    }
    with open(val_dir / 'validation_result.json', 'w') as f:
        json.dump(val_data, f)


def test_export_async_flow_and_2000_rows(client, app, temp_config, tmp_path):
    job_id = "test_export_async"
    _create_mock_job(job_id, tmp_path, temp_config)
    
    # 1. async export start response
    res = client.post(f"/api/export/{job_id}")
    assert res.status_code == 200
    data = res.json
    assert data['status'] == 'processing'
    assert 'progress_url' in data
    
    # Wait for background task to complete (should handle 2000 rows quickly enough)
    for _ in range(10):
        time.sleep(1)
        res_prog = client.get(f"/jobs/{job_id}/progress")
        prog_data = res_prog.json
        if prog_data.get('completed'):
            break
            
    assert prog_data.get('completed'), "Export worker failed or timed out"
    
    # 2. download url 
    result_data = prog_data.get('result_data', {})
    download_url = result_data.get('download_url')
    assert download_url
    
    # 3. Repeat export collision naming (duplicate export protection)
    # Call export again
    res2 = client.post(f"/api/export/{job_id}")
    assert res2.status_code == 200
    
    for _ in range(5):
        time.sleep(1)
        res_prog2 = client.get(f"/jobs/{job_id}/progress")
        if res_prog2.json.get('completed'):
            break
            
    assert res_prog2.json.get('completed')
    dl_url2 = res_prog2.json['result_data']['download_url']
    
    assert download_url != dl_url2, "Filename collision!"
    
    # 4. Verify 2000 row workbook
    output_dir = tmp_path / 'output'
    wb1_filename = download_url.split('/')[-1]
    wb_path = output_dir / wb1_filename
    
    wb = load_workbook(wb_path)
    assert 'Transactions' in wb.sheetnames
    assert 'Summary' in wb.sheetnames
    
    ws_tx = wb['Transactions']
    # +1 for header
    assert ws_tx.max_row == 2001
    
    # 5. Formula injection regression
    # Row 0 naration was '=+-@' -> expected to be prefixed with '
    assert ws_tx.cell(row=2, column=3).value.startswith("'")

def test_export_path_security(client, app, temp_config, tmp_path):
    # Try directory traversal on download
    res = client.get('/api/export/download/../config.ini')
    assert res.status_code == 404

def test_stale_review_export_fallback(client, app, temp_config, tmp_path):
    job_id = "test_export_stale"
    _create_mock_job(job_id, tmp_path, temp_config)
    
    # Create empty stale review
    rev_dir = tmp_path / 'temp' / 'jobs' / job_id / 'review'
    stale_data = {
        "job_id": job_id,
        "review_revision": 1,
        "transactions": [] # 0 rows, but normalized has 2000!
    }
    with open(rev_dir / 'reviewed_statement.json', 'w') as f:
        json.dump(stale_data, f)
        
    svc = ExportService(temp_config)
    payload = svc._get_export_payload(job_id)
    
    assert payload['summary']['export_source'] == "Machine Normalized"
    assert len(payload['transactions']) == 2000
