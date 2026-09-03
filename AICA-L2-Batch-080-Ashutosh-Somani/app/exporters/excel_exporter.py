import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import date

class ExcelExporter:
    """
    Renders structured bank statement data into a professional Excel Workbook.
    Contains no business logic or file/DB reads. Purely a rendering layer.
    """
    
    def __init__(self, amount_format="#,##0.00"):
        self.amount_format = amount_format
        self.dangerous_chars = ('=', '+', '-', '@')

    def _safe_decimal(self, value):
        """Converts Decimal to float safely or to string with warning if it exceeds Excel's 15-digit precision limit."""
        if not isinstance(value, Decimal):
            return value
        
        # Excel can only handle 15 significant digits.
        # Let's count digits before and after decimal point
        s = format(value, 'f')
        digits = len(s.replace('-', '').replace('.', '').lstrip('0'))
        if digits > 15:
            return f"PRECISION WARNING: {s}"
            
        # openpyxl natively handles numeric assignments via floats, 
        # but passing Decimal directly also works, openpyxl will cast it to float internally.
        # We will pass the Decimal directly, which avoids IEEE 754 drift on our end.
        return value

    def _sanitize(self, value):
        """Spreadsheet formula injection protection"""
        if value is None:
            return ""
        str_val = str(value)
        if str_val.startswith(self.dangerous_chars):
            return "'" + str_val
        return str_val

    def _apply_headers(self, ws, headers):
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_fit_columns(self, ws, max_width=60):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_len:
                        max_len = val_len
            
            adjusted_width = min(max_len + 2, max_width)
            ws.column_dimensions[col_letter].width = adjusted_width
            
            # Allow wrapping for wide columns
            if max_len > max_width:
                for cell in col:
                    if cell.row != 1:  # skip header
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def export(self, filepath, payload, progress_callback=None):
        """
        Export the given payload to the filepath.
        payload = {
            'transactions': [...],
            'summary': {...},
            'exceptions': [...],
            'audit': [...]
        }
        """
        wb = openpyxl.Workbook()
        
        # 1. Transactions
        if progress_callback: progress_callback(40, "Building Transactions sheet")
        ws_tx = wb.active
        ws_tx.title = "Transactions"
        self._build_transactions_sheet(ws_tx, payload.get('transactions', []))
        
        # 2. Summary
        if progress_callback: progress_callback(60, "Building Summary")
        ws_summary = wb.create_sheet("Summary")
        self._build_summary_sheet(ws_summary, payload.get('summary', {}))
        
        # 3. Exceptions
        if progress_callback: progress_callback(75, "Building Exceptions")
        ws_exceptions = wb.create_sheet("Exceptions")
        self._build_exceptions_sheet(ws_exceptions, payload.get('exceptions', []))
        
        # 4. Audit Trail
        if progress_callback: progress_callback(85, "Building Audit Trail")
        ws_audit = wb.create_sheet("Audit Trail")
        self._build_audit_sheet(ws_audit, payload.get('audit', []))
        
        if progress_callback: progress_callback(95, "Saving workbook")
        wb.save(filepath)

    def _build_transactions_sheet(self, ws, transactions):
        headers = [
            "Transaction Date", "Value Date", "Narration", "Reference No.", "Cheque No.",
            "Debit", "Credit", "Balance", "Transaction Type", "Validation Status", 
            "Review Status", "Source Page", "Source Type", "OCR Confidence", "Corrected?"
        ]
        self._apply_headers(ws, headers)
        
        for idx, tx in enumerate(transactions, start=2):
            row_data = [
                tx.get("transaction_date"),
                tx.get("value_date"),
                self._sanitize(tx.get("narration")),
                self._sanitize(tx.get("reference_number")),
                self._sanitize(tx.get("cheque_number")),
                self._safe_decimal(tx.get("debit")),
                self._safe_decimal(tx.get("credit")),
                self._safe_decimal(tx.get("balance")),
                tx.get("transaction_type"),
                tx.get("validation_status"),
                tx.get("review_status"),
                tx.get("source_page"),
                tx.get("source_type"),
                tx.get("ocr_confidence"),
                "Yes" if tx.get("user_corrected") else "No"
            ]
            
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(row=idx, column=c_idx, value=val)
                if isinstance(val, Decimal):
                    cell.number_format = self.amount_format
                if isinstance(val, date):
                    cell.number_format = "DD-MM-YYYY"
        
        self._auto_fit_columns(ws, max_width=60)
        ws.freeze_panes = "A2"
        
        # Apply table styling
        if transactions:
            tab = Table(displayName="TransactionsTable", ref=f"A1:O{len(transactions)+1}")
            style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)

    def _build_summary_sheet(self, ws, summary):
        bold_font = Font(bold=True)
        
        rows = [
            ("Export Source", summary.get("export_source")),
            ("OCR Used", summary.get("ocr_used", "No")),
            ("Review Revision", summary.get("review_revision")),
            ("Application Version", summary.get("app_version")),
            ("", ""),
            ("Bank Name", self._sanitize(summary.get("bank_name"))),
            ("Account Holder", self._sanitize(summary.get("account_holder"))),
            ("Account Number", self._sanitize(summary.get("account_number"))),
            ("IFSC", self._sanitize(summary.get("ifsc"))),
            ("Statement Period", self._sanitize(summary.get("statement_period"))),
            ("", ""),
            ("Opening Balance", self._safe_decimal(summary.get("opening_balance"))),
            ("Total Debits", self._safe_decimal(summary.get("total_debits"))),
            ("Total Credits", self._safe_decimal(summary.get("total_credits"))),
            ("Expected Closing Balance", self._safe_decimal(summary.get("expected_closing_balance"))),
            ("Statement Closing Balance", self._safe_decimal(summary.get("statement_closing_balance"))),
            ("Difference", self._safe_decimal(summary.get("difference"))),
            ("", ""),
            ("Transaction Count", summary.get("transaction_count")),
            ("Corrected Transaction Count", summary.get("corrected_transaction_count")),
            ("Exceptions Count", summary.get("exceptions_count")),
            ("Critical Exceptions", summary.get("critical_exceptions")),
            ("Error Exceptions", summary.get("error_exceptions")),
            ("Warning Exceptions", summary.get("warning_exceptions")),
            ("", ""),
            ("Validation Result", summary.get("validation_result")),
            ("Export Date", summary.get("export_date")),
            ("Source Filename", summary.get("source_filename")),
            ("Source SHA-256", summary.get("source_sha256"))
        ]
        
        for idx, (label, val) in enumerate(rows, start=1):
            cell_label = ws.cell(row=idx, column=1, value=label)
            cell_label.font = bold_font
            cell_val = ws.cell(row=idx, column=2, value=val)
            
            if isinstance(val, Decimal):
                cell_val.number_format = self.amount_format
                
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

    def _build_exceptions_sheet(self, ws, exceptions):
        headers = [
            "Exception Code", "Severity", "Transaction #", "Transaction ID",
            "Source Page", "Source Row", "Issue", "Financial Difference",
            "Review Required", "Corrected?"
        ]
        self._apply_headers(ws, headers)
        
        if not exceptions:
            ws.append(["No unresolved exceptions."])
        else:
            for idx, exc in enumerate(exceptions, start=2):
                row = [
                    exc.get("exception_code"),
                    exc.get("severity"),
                    exc.get("transaction_index"),
                    exc.get("transaction_id"),
                    exc.get("source_page"),
                    exc.get("source_row"),
                    self._sanitize(exc.get("message")),
                    exc.get("financial_difference"),
                    "Yes" if exc.get("review_required") else "No",
                    "Yes" if exc.get("user_corrected") else "No"
                ]
                ws.append(row)
                cell_diff = ws.cell(row=idx, column=8)
                if isinstance(cell_diff.value, Decimal):
                    cell_diff.number_format = self.amount_format
                    
            tab = Table(displayName="ExceptionsTable", ref=f"A1:J{len(exceptions)+1}")
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            tab.tableStyleInfo = style
            
        self._auto_fit_columns(ws, max_width=80)
        ws.freeze_panes = "A2"

    def _build_audit_sheet(self, ws, audit_events):
        headers = [
            "Timestamp", "Review Revision", "Action", "Transaction ID",
            "Affected Transaction IDs", "Field", "Before Value", "After Value",
            "Reason", "Source Page", "Source Row"
        ]
        self._apply_headers(ws, headers)
        
        if not audit_events:
            ws.append(["No user corrections recorded."])
        else:
            for ev in audit_events:
                affected = ev.get("affected_transaction_ids")
                aff_str = ", ".join(affected) if affected else ""
                row = [
                    ev.get("timestamp"),
                    ev.get("review_revision"),
                    ev.get("action"),
                    ev.get("transaction_id"),
                    aff_str,
                    ev.get("field_name"),
                    self._sanitize(ev.get("before_value")),
                    self._sanitize(ev.get("after_value")),
                    self._sanitize(ev.get("reason")),
                    ev.get("source_page"),
                    ev.get("source_row")
                ]
                ws.append(row)
                
            tab = Table(displayName="AuditTable", ref=f"A1:K{len(audit_events)+1}")
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            tab.tableStyleInfo = style

        self._auto_fit_columns(ws, max_width=60)
        ws.freeze_panes = "A2"
