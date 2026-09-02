import io
import os
import pandas as pd
from datetime import date, datetime
from typing import List, Dict, Any, Optional, Tuple
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter, landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics.charts.piecharts import Pie

from app.services.pdf_header import draw_client_header, pdf_document
from app.services.daybook_service import get_consolidated_daybook
from app.services.cash_service import get_previous_day_closing_balance
from app.services.card_qr_service import get_card_qr_reconciliations
from app.services.aggregator_service import get_aggregator_payout_matrix
from app.models.cash_rec import CashReconciliation
from app.models.card_qr_rec import CardQrReconciliation
from app.models.audit_log import AuditLog
from app.models.branch import Branch


_INR_FONT = "Helvetica"
_INR_FONT_BOLD = "Helvetica-Bold"


def _register_inr_fonts() -> Tuple[str, str]:
    """Helvetica has no ₹ glyph. Prefer a Windows Unicode font."""
    pairs = [
        (r"C:\Windows\Fonts\calibri.ttf", r"C:\Windows\Fonts\calibrib.ttf"),
        (r"C:\Windows\Fonts\segoeui.ttf", r"C:\Windows\Fonts\segoeuib.ttf"),
        (r"C:\Windows\Fonts\arial.ttf", r"C:\Windows\Fonts\arialbd.ttf"),
    ]
    for regular, bold in pairs:
        if not os.path.exists(regular):
            continue
        try:
            pdfmetrics.registerFont(TTFont("INR", regular))
            pdfmetrics.registerFont(TTFont("INR-Bold", bold if os.path.exists(bold) else regular))
            return "INR", "INR-Bold"
        except Exception:
            continue
    return "Helvetica", "Helvetica-Bold"


_INR_FONT, _INR_FONT_BOLD = _register_inr_fonts()


def _inr(val: float) -> str:
    return f"₹ {float(val or 0):,.2f}"


# NumberedCanvas for PDF Page Numbering Footer
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        width, height = self._pagesize
        draw_client_header(self, width, height, font=_INR_FONT, font_bold=_INR_FONT_BOLD)

        self.setStrokeColor(colors.HexColor("#D1FAE5"))
        self.setLineWidth(0.6)
        self.line(28, 30, width - 28, 30)

        self.setFont(_INR_FONT, 8)
        self.setFillColor(colors.HexColor("#64748B"))
        self.drawString(28, 16, "Harsh's RestoReconcile  ·  Property of Harsh Singhal & Associates")
        self.drawRightString(width - 28, 16, f"Page {self._pageNumber} of {page_count}")
        self.restoreState()


# Helper to format Excel Worksheets cleanly
def _format_excel_worksheet(ws, title: str, start_row: int = 4):
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    ws.cell(row=1, column=1, value="Harsh's RestoReconcile — " + title).font = title_font
    ws.cell(row=2, column=1, value=f"Generated on: {datetime.now().strftime('%b %d, %Y %I:%M %p')}").font = subtitle_font
    
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16

    # Style Header Row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if "Status" in str(cell.value) or "Date" in str(cell.value) else "left", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[start_row].height = 24

    # Format Data Rows
    for row in range(start_row + 1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            header_val = str(ws.cell(row=start_row, column=col).value or "")
            
            # Format numbers vs text
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and str(cell.value).startswith("=")):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif "Date" in header_val:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row < start_row:
                continue
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


# ==============================================================================
# DAYBOOK REPORT EXPORTS
# ==============================================================================
def generate_excel_daybook_report(
    db: Session,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    rows = list(reversed(get_consolidated_daybook(db, branch_id, start_date, end_date)))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Day Book"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    title_font = Font(name="Calibri", size=16, bold=True, color="166534")
    sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    head_font = Font(name="Calibri", size=10, bold=True, color="166534")
    num_font = Font(name="Calibri", size=10, color="0F172A")
    bold_font = Font(name="Calibri", size=10, bold=True, color="166534")
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    head_fill = PatternFill("solid", fgColor="EEF7F0")
    soft = PatternFill("solid", fgColor="F0FDF4")
    zebra = PatternFill("solid", fgColor="F8FAF9")

    period = _export_period_label(start_date, end_date)
    scope = "Consolidated" if not branch_id else "Branch"
    ws["A1"] = f"Day Book Report — {scope}"
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A2"] = f"Harsh's RestoReconcile  ·  {period or 'All dates'}  ·  Online = Card/QR + Aggregators + Other  ·  Total = Cash + Online  ·  Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = sub_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    headers = [
        "Date", "Branch", "Cash Balance", "Cash Sale", "Card / QR Code",
        "Zomato", "Swiggy", "Dineout", "Other Channels", "Online Payment",
        "Total Sales", "Status"
    ]
    start_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    for i, r in enumerate(rows):
        row_idx = start_row + 1 + i
        fill = zebra if i % 2 else PatternFill("solid", fgColor="FFFFFF")
        ws.cell(row=row_idx, column=1, value=r["date"].strftime("%d-%b-%y")).alignment = left
        ws.cell(row=row_idx, column=2, value=r["branch_name"]).alignment = left
        values = {
            3: float(r["cash_balance"] or 0),
            4: float(r["cash"] or 0),
            5: float(r["card_qr"] or 0),
            6: float(r["zomato"] or 0),
            7: float(r["swiggy"] or 0),
            8: float(r["dineout"] or 0),
            9: float(r["other_channels"] or 0),
        }
        for col, val in values.items():
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.number_format = _INR_XLS
            cell.alignment = right
            cell.font = num_font
        ws.cell(row=row_idx, column=10, value=f"=E{row_idx}+F{row_idx}+G{row_idx}+H{row_idx}+I{row_idx}")
        ws.cell(row=row_idx, column=11, value=f"=D{row_idx}+J{row_idx}")
        ws.cell(row=row_idx, column=12, value=r["status"])
        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).fill = fill
            if col in (10, 11):
                ws.cell(row=row_idx, column=col).number_format = _INR_XLS
                ws.cell(row=row_idx, column=col).alignment = right
                ws.cell(row=row_idx, column=col).font = bold_font
                ws.cell(row=row_idx, column=col).fill = soft

    total_row = start_row + 1 + len(rows)
    first, last = start_row + 1, start_row + len(rows)
    ws.cell(row=total_row, column=1, value="Total").font = bold_font
    ws.cell(row=total_row, column=2, value="All Branches" if not branch_id else "").font = bold_font
    if rows:
        for col in range(3, 10):
            letter = get_column_letter(col)
            cell = ws.cell(row=total_row, column=col, value=f"=SUM({letter}{first}:{letter}{last})")
            cell.number_format = _INR_XLS
            cell.font = bold_font
            cell.alignment = right
        ws.cell(row=total_row, column=10, value=f"=E{total_row}+F{total_row}+G{total_row}+H{total_row}+I{total_row}")
        ws.cell(row=total_row, column=11, value=f"=D{total_row}+J{total_row}")
        for col in (10, 11):
            ws.cell(row=total_row, column=col).number_format = _INR_XLS
            ws.cell(row=total_row, column=col).font = bold_font
            ws.cell(row=total_row, column=col).alignment = right
    for col in range(1, 13):
        ws.cell(row=total_row, column=col).fill = soft
        ws.cell(row=total_row, column=col).border = border

    note = total_row + 2
    ws.cell(row=note, column=1, value="Online Payment and Total Sales are Excel formulas. Totals are SUM / linked formulas. Amounts use ₹.").font = sub_font
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=8)
    widths = [12, 18, 14, 13, 14, 12, 12, 12, 15, 15, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_daybook_report(
    db: Session,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    from app.services.daybook_service import get_daybook_totals
    rows = get_consolidated_daybook(db, branch_id, start_date, end_date)
    totals = get_daybook_totals(rows)
    period = _export_period_label(start_date, end_date)
    scope = "Consolidated — All Branches" if not branch_id else (rows[0]["branch_name"] if rows else "Branch")

    buffer = io.BytesIO()
    page_w, _ = landscape(A4)
    avail = page_w - 44
    doc = pdf_document(buffer, landscape(A4), left=22, right=22, bottom=36)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("DbTitle", parent=styles["Heading1"], fontName=_INR_FONT_BOLD, fontSize=15, textColor=colors.HexColor("#14532D"), spaceAfter=2, leading=18)
    sub_style = ParagraphStyle("DbSub", parent=styles["Normal"], fontName=_INR_FONT, fontSize=8, textColor=colors.HexColor("#64748B"), spaceAfter=8)
    th = ParagraphStyle("DbTH", fontName=_INR_FONT_BOLD, fontSize=7, textColor=colors.HexColor("#166534"), alignment=1, leading=9)
    td = ParagraphStyle("DbTD", fontName=_INR_FONT, fontSize=7, textColor=colors.HexColor("#0F172A"), leading=9)
    tn = ParagraphStyle("DbTN", fontName=_INR_FONT, fontSize=7, textColor=colors.HexColor("#0F172A"), alignment=2, leading=9)
    tb = ParagraphStyle("DbTB", fontName=_INR_FONT_BOLD, fontSize=7, textColor=colors.HexColor("#166534"), alignment=2, leading=9)

    pie = _mix_pie_drawing("Sales Mix", [
        ("Cash", totals["cash"], colors.HexColor("#166534")),
        ("Card / QR", totals["card_qr"], colors.HexColor("#22C55E")),
        ("Zomato", totals["zomato"], colors.HexColor("#E11D48")),
        ("Swiggy", totals["swiggy"], colors.HexColor("#F97316")),
        ("Dineout", totals["dineout"], colors.HexColor("#CA8A04")),
        ("Other", totals["other_channels"], colors.HexColor("#64748B")),
    ], width=avail * 0.36, height=168)
    kpis = _kpi_row([
        ("Total Sales", totals["total_sales"]),
        ("Cash", totals["cash"]),
        ("Card / QR", totals["card_qr"]),
        ("Online / Aggregators", totals["online_payment"]),
    ], avail * 0.62)
    dash = Table([[kpis, pie]], colWidths=[avail * 0.62, avail * 0.38])
    dash.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (0, 0), 10),
    ]))

    headers = ["Date", "Branch", "Cash", "Card/QR", "Zomato", "Swiggy", "Dineout", "Other", "Online", "Total"]
    table_data = [[Paragraph(h, th) for h in headers]]
    for r in rows:
        table_data.append([
            Paragraph(r["date"].strftime("%d-%b-%y"), td),
            Paragraph(r["branch_name"], td),
            Paragraph(_inr(r["cash"]), tn),
            Paragraph(_inr(r["card_qr"]), tn),
            Paragraph(_inr(r["zomato"]), tn),
            Paragraph(_inr(r["swiggy"]), tn),
            Paragraph(_inr(r["dineout"]), tn),
            Paragraph(_inr(r["other_channels"]), tn),
            Paragraph(_inr(r["online_payment"]), tn),
            Paragraph(_inr(r["total_sales"]), tb),
        ])
    table_data.append([
        Paragraph("<b>Total</b>", td), Paragraph("Consolidated" if not branch_id else "", td),
        Paragraph(f"<b>{_inr(totals['cash'])}</b>", tb),
        Paragraph(f"<b>{_inr(totals['card_qr'])}</b>", tb),
        Paragraph(f"<b>{_inr(totals['zomato'])}</b>", tb),
        Paragraph(f"<b>{_inr(totals['swiggy'])}</b>", tb),
        Paragraph(f"<b>{_inr(totals['dineout'])}</b>", tb),
        Paragraph(f"<b>{_inr(totals['other_channels'])}</b>", tb),
        Paragraph(f"<b>{_inr(totals['online_payment'])}</b>", tb),
        Paragraph(f"<b>{_inr(totals['total_sales'])}</b>", tb),
    ])
    widths = [62, 88, 70, 70, 68, 68, 68, 64, 72, 78]
    t = Table(table_data, colWidths=widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF7F0")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F0FDF4")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D5DDD7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8FAF9")))
    t.setStyle(TableStyle(style))

    elements = [
        Paragraph(f"Day Book Report — {scope}", title_style),
        Paragraph(f"{period or 'All dates'}  ·  Branch-wise lines with consolidated totals  ·  Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}", sub_style),
        dash,
        Spacer(1, 8),
        t,
    ]
    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


# ==============================================================================
# CASH RECONCILIATION REPORT EXPORTS
# ==============================================================================
def generate_excel_cash_reconciliation_report(
    db: Session,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    from app.services.cash_service import sync_cash_recs_from_daybook
    sync_cash_recs_from_daybook(db, branch_id, start_date, end_date)
    query = db.query(CashReconciliation)
    if branch_id:
        query = query.filter(CashReconciliation.branch_id == branch_id)
    if start_date:
        query = query.filter(CashReconciliation.rec_date >= start_date)
    if end_date:
        query = query.filter(CashReconciliation.rec_date <= end_date)
    recs = query.order_by(CashReconciliation.branch_id.asc(), CashReconciliation.rec_date.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cash Rec"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    title_font = Font(name="Calibri", size=16, bold=True, color="166534")
    sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    head_font = Font(name="Calibri", size=9, bold=True, color="166534")
    num_font = Font(name="Calibri", size=10, color="0F172A")
    bold_font = Font(name="Calibri", size=10, bold=True, color="166534")
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    head_fill = PatternFill("solid", fgColor="EEF7F0")
    soft = PatternFill("solid", fgColor="F0FDF4")
    zebra = PatternFill("solid", fgColor="F8FAF9")

    period = _export_period_label(start_date, end_date)
    ws["A1"] = "Cash Reconciliation — " + ("Consolidated" if not branch_id else "Branch")
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws["A2"] = f"Harsh's RestoReconcile  ·  {period or 'All dates'}  ·  Expected, Difference, Opening carry-forward, and Totals are Excel formulas  ·  Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = sub_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    headers = [
        "Date", "Branch", "Opening Balance", "Cash Sale", "Expenses (Inv Recv)",
        "Expenses (Inv Not Recv)", "Salary Adv (1-5)", "Salary Adv (6-15)",
        "Salary Adv (16-31)", "Transfer Base Kitchen", "Service Charge", "Other Adj",
        "Expected Closing", "Actual Closing", "Difference", "Status"
    ]
    start_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    prev_branch = None
    for i, r in enumerate(recs):
        row_idx = start_row + 1 + i
        fill = zebra if i % 2 else PatternFill("solid", fgColor="FFFFFF")
        ws.cell(row=row_idx, column=1, value=r.rec_date.strftime("%d-%b-%y")).alignment = left
        ws.cell(row=row_idx, column=2, value=r.branch.name if r.branch else "N/A").alignment = left
        same_branch = prev_branch == r.branch_id
        if same_branch:
            ws.cell(row=row_idx, column=3, value=f"=N{row_idx - 1}")
        else:
            ws.cell(row=row_idx, column=3, value=float(r.opening_balance or 0))
        inputs = {
            4: float(r.cash_sale or 0),
            5: float(r.site_expenses_inv_rec or 0),
            6: float(r.site_expenses_inv_not_rec or 0),
            7: float(r.advance_salary_1_5 or 0),
            8: float(r.advance_salary_6_15 or 0),
            9: float(r.advance_salary_16_31 or 0),
            10: float(r.transfer_base_kitchen or 0),
            11: float(r.service_charge or 0),
            12: float(r.other_adjustments or 0),
            14: float(r.actual_closing_balance or 0),
        }
        for col, val in inputs.items():
            ws.cell(row=row_idx, column=col, value=val)
        ws.cell(row=row_idx, column=13, value=f"=C{row_idx}+D{row_idx}-E{row_idx}-F{row_idx}-G{row_idx}-H{row_idx}-I{row_idx}-J{row_idx}+K{row_idx}+L{row_idx}")
        ws.cell(row=row_idx, column=15, value=f"=N{row_idx}-M{row_idx}")
        ws.cell(row=row_idx, column=16, value=r.status)
        for col in range(1, 17):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            cell.fill = fill
            cell.font = num_font
            if col in range(3, 16):
                cell.number_format = _INR_XLS
                cell.alignment = right
            if col in (13, 15):
                cell.font = bold_font
                cell.fill = soft
        prev_branch = r.branch_id

    total_row = start_row + 1 + len(recs)
    first, last = start_row + 1, start_row + len(recs)
    ws.cell(row=total_row, column=1, value="Total").font = bold_font
    ws.cell(row=total_row, column=2, value="Consolidated" if not branch_id else "").font = bold_font
    if recs:
        for col in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14):
            letter = get_column_letter(col)
            ws.cell(row=total_row, column=col, value=f"=SUM({letter}{first}:{letter}{last})")
        ws.cell(row=total_row, column=13, value=f"=C{total_row}+D{total_row}-E{total_row}-F{total_row}-G{total_row}-H{total_row}-I{total_row}-J{total_row}+K{total_row}+L{total_row}")
        ws.cell(row=total_row, column=15, value=f"=N{total_row}-M{total_row}")
    for col in range(1, 17):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = soft
        cell.border = border
        cell.font = bold_font
        if col in range(3, 16):
            cell.number_format = _INR_XLS
            cell.alignment = right

    note = total_row + 2
    ws.cell(row=note, column=1, value="Opening for the next day = previous Actual Closing (same branch). Expected and Difference are formulas. Totals are SUM / linked formulas.").font = sub_font
    ws.merge_cells(start_row=note, start_column=1, end_row=note, end_column=10)
    for i, w in enumerate([12, 16, 14, 12, 14, 16, 13, 13, 14, 16, 13, 12, 15, 14, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_cash_reconciliation_report(
    db: Session,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    from app.services.cash_service import sync_cash_recs_from_daybook
    sync_cash_recs_from_daybook(db, branch_id, start_date, end_date)
    query = db.query(CashReconciliation)
    if branch_id:
        query = query.filter(CashReconciliation.branch_id == branch_id)
    if start_date:
        query = query.filter(CashReconciliation.rec_date >= start_date)
    if end_date:
        query = query.filter(CashReconciliation.rec_date <= end_date)
    recs = query.order_by(CashReconciliation.rec_date.asc(), CashReconciliation.branch_id.asc()).all()

    tot_sale = tot_exp = tot_sal = tot_xfer = tot_expected = tot_actual = tot_diff = 0.0
    for r in recs:
        tot_sale += float(r.cash_sale or 0)
        tot_exp += float((r.site_expenses_inv_rec or 0) + (r.site_expenses_inv_not_rec or 0))
        tot_sal += float((r.advance_salary_1_5 or 0) + (r.advance_salary_6_15 or 0) + (r.advance_salary_16_31 or 0))
        tot_xfer += float(r.transfer_base_kitchen or 0)
        tot_expected += float(r.expected_closing_balance or 0)
        tot_actual += float(r.actual_closing_balance or 0)
        tot_diff += float(r.difference or 0)

    period = _export_period_label(start_date, end_date)
    scope = "Consolidated — All Branches" if not branch_id else (recs[0].branch.name if recs and recs[0].branch else "Branch")
    buffer = io.BytesIO()
    page_w, _ = landscape(A4)
    avail = page_w - 44
    doc = pdf_document(buffer, landscape(A4), left=22, right=22, bottom=36)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CrTitle", parent=styles["Heading1"], fontName=_INR_FONT_BOLD, fontSize=15, textColor=colors.HexColor("#14532D"), spaceAfter=2, leading=18)
    sub_style = ParagraphStyle("CrSub", parent=styles["Normal"], fontName=_INR_FONT, fontSize=8, textColor=colors.HexColor("#64748B"), spaceAfter=8)
    th = ParagraphStyle("CrTH", fontName=_INR_FONT_BOLD, fontSize=7, textColor=colors.HexColor("#166534"), alignment=1, leading=9)
    td = ParagraphStyle("CrTD", fontName=_INR_FONT, fontSize=7, textColor=colors.HexColor("#0F172A"), leading=9)
    tn = ParagraphStyle("CrTN", fontName=_INR_FONT, fontSize=7, textColor=colors.HexColor("#0F172A"), alignment=2, leading=9)
    tb = ParagraphStyle("CrTB", fontName=_INR_FONT_BOLD, fontSize=7, textColor=colors.HexColor("#166534"), alignment=2, leading=9)

    pie = _mix_pie_drawing("Cash Mix", [
        ("Cash Sale", tot_sale, colors.HexColor("#166534")),
        ("Site Expenses", tot_exp, colors.HexColor("#F97316")),
        ("Salary Advance", tot_sal, colors.HexColor("#CA8A04")),
        ("Kitchen Transfer", tot_xfer, colors.HexColor("#64748B")),
    ], width=avail * 0.36, height=168)
    kpis = _kpi_row([
        ("Cash Sale", tot_sale),
        ("Expected Closing", tot_expected),
        ("Actual Closing", tot_actual),
        ("Difference", tot_diff),
    ], avail * 0.62)
    dash = Table([[kpis, pie]], colWidths=[avail * 0.62, avail * 0.38])
    dash.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (0, 0), 10),
    ]))

    headers = ["Date", "Branch", "Opening", "Cash Sale", "Expenses", "Salary", "Expected", "Actual", "Difference", "Status"]
    table_data = [[Paragraph(h, th) for h in headers]]
    for r in recs:
        exp = float((r.site_expenses_inv_rec or 0) + (r.site_expenses_inv_not_rec or 0))
        sal = float((r.advance_salary_1_5 or 0) + (r.advance_salary_6_15 or 0) + (r.advance_salary_16_31 or 0))
        diff = float(r.difference or 0)
        table_data.append([
            Paragraph(r.rec_date.strftime("%d-%b-%y"), td),
            Paragraph(r.branch.name if r.branch else "N/A", td),
            Paragraph(_inr(r.opening_balance), tn),
            Paragraph(_inr(r.cash_sale), tn),
            Paragraph(_inr(exp), tn),
            Paragraph(_inr(sal), tn),
            Paragraph(_inr(r.expected_closing_balance), tn),
            Paragraph(_inr(r.actual_closing_balance), tn),
            Paragraph(_inr(diff), tb),
            Paragraph(r.status or "", td),
        ])
    table_data.append([
        Paragraph("<b>Total</b>", td), Paragraph("Consolidated" if not branch_id else "", td),
        Paragraph("", tn),
        Paragraph(f"<b>{_inr(tot_sale)}</b>", tb),
        Paragraph(f"<b>{_inr(tot_exp)}</b>", tb),
        Paragraph(f"<b>{_inr(tot_sal)}</b>", tb),
        Paragraph(f"<b>{_inr(tot_expected)}</b>", tb),
        Paragraph(f"<b>{_inr(tot_actual)}</b>", tb),
        Paragraph(f"<b>{_inr(tot_diff)}</b>", tb),
        Paragraph("", td),
    ])
    t = Table(table_data, colWidths=[62, 95, 72, 72, 70, 68, 78, 78, 78, 70], repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF7F0")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F0FDF4")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#D5DDD7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            style.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#F8FAF9")))
    t.setStyle(TableStyle(style))

    elements = [
        Paragraph(f"Cash Reconciliation — {scope}", title_style),
        Paragraph(f"{period or 'All dates'}  ·  Branch-wise lines with consolidated totals  ·  Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}", sub_style),
        dash,
        Spacer(1, 8),
        t,
    ]
    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


# ==============================================================================
# CARD / QR RECONCILIATION REPORT EXPORTS
# ==============================================================================
def generate_excel_card_qr_report(
    db: Session,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None
) -> bytes:
    from app.services.card_qr_service import get_card_qr_settlement_matrix
    matrix = get_card_qr_settlement_matrix(db, branch_id, start_date, end_date, status)
    branches = matrix.get("branches") or []
    dates = matrix.get("dates") or []
    cells = matrix.get("cells") or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Final Settlement"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    title_font = Font(name="Calibri", size=16, bold=True, color="166534")
    sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    head_font = Font(name="Calibri", size=10, bold=True, color="166534")
    subhead_font = Font(name="Calibri", size=9, bold=True, color="4B6354")
    num_font = Font(name="Calibri", size=10, color="0F172A")
    bold_font = Font(name="Calibri", size=10, bold=True, color="166534")
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    branch_fill = PatternFill("solid", fgColor="EEF7F0")
    head_fill = PatternFill("solid", fgColor="F6FAF7")
    diff_fill = PatternFill("solid", fgColor="F0FDF4")
    soft = PatternFill("solid", fgColor="F0FDF4")
    zebra = PatternFill("solid", fgColor="F8FAF9")

    period = _export_period_label(start_date, end_date)
    ws["A1"] = "Final Settlement Report — Card / QR"
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(4, 1 + 3 * max(1, len(branches))))
    ws["A2"] = f"Harsh's RestoReconcile  ·  {period or 'All dates'}  ·  Difference = bank / payment-gateway charges  ·  Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = sub_font
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16

    ws["A4"] = "Date"
    ws["A4"].fill = head_fill
    ws["A4"].font = head_font
    ws["A4"].alignment = center
    ws["A4"].border = border
    ws.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)
    ws["A5"].fill = head_fill
    ws["A5"].border = border

    col = 2
    spans = []
    for b in branches:
        start_col = col
        header = ws.cell(row=4, column=col, value=str(b.get("name") or "Branch").upper())
        header.fill = branch_fill
        header.font = head_font
        header.alignment = center
        for offset, label in enumerate(("Card / QR Code", "Received Amount", "Difference")):
            cell = ws.cell(row=5, column=col + offset, value=label)
            cell.fill = diff_fill if offset == 2 else head_fill
            cell.font = subhead_font
            cell.alignment = center
            cell.border = border
            ws.cell(row=4, column=col + offset).fill = branch_fill
            ws.cell(row=4, column=col + offset).border = border
            ws.cell(row=4, column=col + offset).font = head_font
        if col + 2 > col:
            ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 2)
        spans.append({"id": str(b.get("id")), "start": start_col})
        col += 3

    last_col = max(2, col - 1)
    data_start = 6
    for i, iso in enumerate(dates):
        excel_row = data_start + i
        ws.cell(row=excel_row, column=1, value=datetime.strptime(iso, "%Y-%m-%d").strftime("%d-%b-%y")).alignment = left
        ws.cell(row=excel_row, column=1).font = bold_font
        ws.cell(row=excel_row, column=1).border = border
        row_fill = zebra if i % 2 else PatternFill("solid", fgColor="FFFFFF")
        ws.cell(row=excel_row, column=1).fill = row_fill
        day_cells = cells.get(iso) or {}
        for span in spans:
            cell = day_cells.get(span["id"]) or {}
            sales_col = span["start"]
            recv_col = span["start"] + 1
            diff_col = span["start"] + 2
            sc = ws.cell(row=excel_row, column=sales_col, value=float(cell.get("sales") or 0))
            rc = ws.cell(row=excel_row, column=recv_col, value=float(cell.get("received") or 0))
            dc = ws.cell(row=excel_row, column=diff_col, value=f"={get_column_letter(sales_col)}{excel_row}-{get_column_letter(recv_col)}{excel_row}")
            for c in (sc, rc, dc):
                c.number_format = _INR_XLS
                c.alignment = right
                c.border = border
                c.font = num_font
                c.fill = row_fill
            dc.fill = diff_fill
            dc.font = bold_font

    total_row = data_start + len(dates)
    ws.cell(row=total_row, column=1, value="Total").font = bold_font
    ws.cell(row=total_row, column=1).fill = soft
    ws.cell(row=total_row, column=1).border = border
    if dates:
        first = data_start
        last = data_start + len(dates) - 1
        for span in spans:
            sales_letter = get_column_letter(span["start"])
            recv_letter = get_column_letter(span["start"] + 1)
            for offset in range(3):
                cidx = span["start"] + offset
                letter = get_column_letter(cidx)
                if offset == 2:
                    formula = f"={sales_letter}{total_row}-{recv_letter}{total_row}"
                else:
                    formula = f"=SUM({letter}{first}:{letter}{last})"
                cell = ws.cell(row=total_row, column=cidx, value=formula)
                cell.number_format = _INR_XLS
                cell.font = bold_font
                cell.alignment = right
                cell.border = border
                cell.fill = diff_fill if offset == 2 else soft
    else:
        for span in spans:
            for offset in range(3):
                cell = ws.cell(row=total_row, column=span["start"] + offset, value=0)
                cell.number_format = _INR_XLS
                cell.font = bold_font
                cell.alignment = right
                cell.border = border
                cell.fill = soft

    note_row = total_row + 2
    ws.cell(
        row=note_row,
        column=1,
        value="Difference = Card/QR − Received Amount. Branch Totals are SUM formulas. Total Difference = Total Card/QR − Total Received."
    ).font = sub_font
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=min(7, last_col))

    ws.column_dimensions["A"].width = 14
    for cidx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(cidx)].width = 16

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_card_qr_report(
    db: Session,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None
) -> bytes:
    from app.services.card_qr_service import get_card_qr_settlement_matrix
    matrix = get_card_qr_settlement_matrix(db, branch_id, start_date, end_date, status)
    branches = matrix.get("branches") or []
    dates = matrix.get("dates") or []
    cells = matrix.get("cells") or {}
    totals = matrix.get("totals") or {}

    buffer = io.BytesIO()
    page_w, _ = landscape(A4)
    left_m, right_m = 22, 22
    avail = page_w - left_m - right_m
    doc = pdf_document(buffer, landscape(A4), left=left_m, right=right_m, bottom=36)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'SettleTitle', parent=styles['Heading1'], fontName=_INR_FONT_BOLD,
        fontSize=15, textColor=colors.HexColor('#14532D'), spaceAfter=2, leading=18
    )
    subtitle_style = ParagraphStyle(
        'SettleSub', parent=styles['Normal'], fontName=_INR_FONT,
        fontSize=8, textColor=colors.HexColor('#64748B'), spaceAfter=8, leading=11
    )
    th_style = ParagraphStyle('SettleTH', fontName=_INR_FONT_BOLD, fontSize=7, textColor=colors.HexColor('#166534'), alignment=1, leading=9)
    td_style = ParagraphStyle('SettleTD', fontName=_INR_FONT, fontSize=7.5, textColor=colors.HexColor('#0F172A'), leading=10)
    td_num = ParagraphStyle('SettleNum', fontName=_INR_FONT, fontSize=7.5, textColor=colors.HexColor('#0F172A'), alignment=2, leading=10)
    td_num_b = ParagraphStyle('SettleNumB', fontName=_INR_FONT_BOLD, fontSize=7.5, textColor=colors.HexColor('#0F172A'), alignment=2, leading=10)

    period = _export_period_label(start_date, end_date)
    elements = [
        Paragraph("Final Settlement Report — Card / QR", title_style),
        Paragraph(
            f"{period or 'All dates'}  ·  Day Book Card/QR vs bank receipts  ·  Difference = bank / PG charges  ·  Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
            subtitle_style
        ),
    ]

    if not branches:
        branches = [{"id": 0, "name": "All Branches"}]

    header_top = [Paragraph("Date", th_style)]
    header_sub = [Paragraph("", th_style)]
    spans = []
    for b in branches:
        header_top.extend([Paragraph(str(b.get("name") or "Branch"), th_style), "", ""])
        header_sub.extend([
            Paragraph("Card / QR Code", th_style),
            Paragraph("Received Amount", th_style),
            Paragraph("Difference", th_style),
        ])
        spans.append(str(b.get("id")))

    table_data = [header_top, header_sub]
    for iso in dates:
        try:
            label = datetime.strptime(iso, "%Y-%m-%d").strftime("%d-%b-%y")
        except ValueError:
            label = iso
        row = [Paragraph(label, td_style)]
        day = cells.get(iso) or {}
        for bid in spans:
            cell = day.get(bid) or {}
            row.append(Paragraph(_inr(cell.get("sales") or 0), td_num))
            row.append(Paragraph(_inr(cell.get("received") or 0), td_num))
            row.append(Paragraph(_inr(cell.get("difference") or 0), td_num_b))
        table_data.append(row)

    total_row = [Paragraph("<b>Total</b>", td_style)]
    for bid in spans:
        t = totals.get(str(bid)) or {}
        total_row.append(Paragraph(f"<b>{_inr(t.get('sales') or 0)}</b>", td_num_b))
        total_row.append(Paragraph(f"<b>{_inr(t.get('received') or 0)}</b>", td_num_b))
        total_row.append(Paragraph(f"<b>{_inr(t.get('difference') or 0)}</b>", td_num_b))
    table_data.append(total_row)

    date_w = 70
    other_w = max(72, (avail - date_w) / max(1, 3 * len(branches)))
    col_widths = [date_w] + [other_w] * (3 * len(branches))
    t = Table(table_data, colWidths=col_widths, repeatRows=2)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EEF7F0')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F6FAF7')),
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.HexColor('#166534')),
        ('SPAN', (0, 0), (0, 1)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 3.5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D5DDD7')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F0FDF4')),
    ]
    col = 1
    for _ in branches:
        style_cmds.append(('SPAN', (col, 0), (col + 2, 0)))
        style_cmds.append(('BACKGROUND', (col + 2, 2), (col + 2, -2), colors.HexColor('#F5FAF6')))
        col += 3
    for i in range(2, len(table_data) - 1):
        if i % 2 == 0:
            style_cmds.append(('BACKGROUND', (0, i), (0, i), colors.HexColor('#F8FAF9')))
    t.setStyle(TableStyle(style_cmds))
    elements.append(t)
    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


# ==============================================================================
# AGGREGATOR REPORT EXPORTS
# ==============================================================================
_NAVY = "166534"
_WHITE = "FFFFFF"
_MUTED = "64748B"
_INK = "0F172A"
_GRAY = "F8FAFC"
_INR_XLS = '"₹"#,##0.00'


def _thin_border():
    return Border(
        left=Side(style="thin", color="D4D4D4"),
        right=Side(style="thin", color="D4D4D4"),
        top=Side(style="thin", color="D4D4D4"),
        bottom=Side(style="thin", color="D4D4D4"),
    )


def _export_payout_group(matrix: Dict[str, Any], branch_id: Optional[int]) -> Dict[str, Any]:
    """One table only: combined All-Branches totals, or the selected branch."""
    if branch_id:
        groups = matrix.get("branch_groups") or []
        if groups:
            return groups[0]
    return {
        "branch_name": "All Branches",
        "branch_code": "ALL",
        "cycle_columns": matrix.get("cycle_columns") or [],
        "rows": matrix.get("rows") or [],
    }


def _export_period_label(start_date: Optional[date], end_date: Optional[date]) -> str:
    if start_date and end_date:
        if (
            start_date.month == 4
            and start_date.day == 1
            and end_date.month == 3
            and end_date.day == 31
            and end_date.year == start_date.year + 1
        ):
            return f"FY {start_date.year}-{str(end_date.year)[-2:]}"
        if start_date.month == end_date.month and start_date.year == end_date.year:
            return start_date.strftime("%B %Y")
        return f"{start_date.strftime('%d %b %Y')} – {end_date.strftime('%d %b %Y')}"
    if start_date:
        return start_date.strftime("%B %Y")
    return ""


def _export_aggregator_name(matrix: Dict[str, Any]) -> str:
    names = sorted({
        c.get("aggregator")
        for c in (matrix.get("columns") or [])
        if c.get("aggregator")
    })
    if len(names) == 1:
        return names[0]
    if names:
        return " / ".join(names)
    return ""


def _row_total(rows: List[Dict[str, Any]], code: str) -> float:
    for r in rows:
        if r.get("code") == code:
            return float(r.get("total") or 0)
    return 0.0


def _deduction_pie_drawing(rows: List[Dict[str, Any]], width: float = 300, height: float = 168) -> Drawing:
    slices = [
        ("COMMISSION", "Platform Charges", colors.HexColor("#166534")),
        ("PROMOTION", "Business Promotion", colors.HexColor("#22C55E")),
        ("TCS", "TCS", colors.HexColor("#0F766E")),
        ("TDS", "TDS Receivable", colors.HexColor("#CA8A04")),
        ("MISC", "Miscellaneous", colors.HexColor("#64748B")),
    ]
    data, labels, palette = [], [], []
    for code, label, color in slices:
        val = abs(_row_total(rows, code))
        if val > 0.009:
            data.append(val)
            labels.append(label)
            palette.append(color)

    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.HexColor("#F0FDF4"), strokeColor=colors.HexColor("#BBF7D0"), strokeWidth=0.6, rx=8, ry=8))
    d.add(String(14, height - 18, "Deduction Mix", fontName=_INR_FONT_BOLD, fontSize=10, fillColor=colors.HexColor("#14532D")))

    if not data:
        d.add(String(14, height / 2 - 4, "No deduction amounts in this period", fontName=_INR_FONT, fontSize=8, fillColor=colors.HexColor("#64748B")))
        return d

    pie = Pie()
    pie.x = 10
    pie.y = 12
    pie.width = 132
    pie.height = 132
    pie.data = data
    pie.labels = None
    pie.simpleLabels = 0
    pie.slices.strokeWidth = 1.2
    pie.slices.strokeColor = colors.white
    pie.slices.popout = 0
    if data:
        pie.slices[0].popout = 5
    for i, color in enumerate(palette):
        pie.slices[i].fillColor = color
    d.add(pie)

    y = height - 42
    total = sum(data) or 1.0
    for label, val, color in zip(labels, data, palette):
        d.add(Rect(158, y - 1, 8, 8, fillColor=color, strokeColor=None, rx=1, ry=1))
        d.add(String(172, y, label, fontName=_INR_FONT, fontSize=7.5, fillColor=colors.HexColor("#334155")))
        d.add(String(172, y - 11, f"{_inr(val)}   ({val / total * 100:.0f}%)", fontName=_INR_FONT_BOLD, fontSize=7, fillColor=colors.HexColor("#0F172A")))
        y -= 26
    return d


def _mix_pie_drawing(title: str, slices: List[tuple], width: float = 300, height: float = 168) -> Drawing:
    data, labels, palette = [], [], []
    for label, val, color in slices:
        amount = abs(float(val or 0))
        if amount > 0.009:
            data.append(amount)
            labels.append(label)
            palette.append(color)
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, fillColor=colors.HexColor("#F0FDF4"), strokeColor=colors.HexColor("#BBF7D0"), strokeWidth=0.6, rx=8, ry=8))
    d.add(String(14, height - 18, title, fontName=_INR_FONT_BOLD, fontSize=10, fillColor=colors.HexColor("#14532D")))
    if not data:
        d.add(String(14, height / 2 - 4, "No amounts in this period", fontName=_INR_FONT, fontSize=8, fillColor=colors.HexColor("#64748B")))
        return d
    pie = Pie()
    pie.x = 10
    pie.y = 12
    pie.width = 132
    pie.height = 132
    pie.data = data
    pie.labels = None
    pie.simpleLabels = 0
    pie.slices.strokeWidth = 1.2
    pie.slices.strokeColor = colors.white
    if data:
        pie.slices[0].popout = 5
    for i, color in enumerate(palette):
        pie.slices[i].fillColor = color
    d.add(pie)
    y = height - 42
    total = sum(data) or 1.0
    for label, val, color in zip(labels, data, palette):
        d.add(Rect(158, y - 1, 8, 8, fillColor=color, strokeColor=None, rx=1, ry=1))
        d.add(String(172, y, label, fontName=_INR_FONT, fontSize=7.5, fillColor=colors.HexColor("#334155")))
        d.add(String(172, y - 11, f"{_inr(val)}   ({val / total * 100:.0f}%)", fontName=_INR_FONT_BOLD, fontSize=7, fillColor=colors.HexColor("#0F172A")))
        y -= 26
    return d


def _kpi_row(cards: List[tuple], avail: float):
    fills = ["#ECFDF5", "#F0FDF4", "#FFFBEB", "#F8FAFC"]
    boxes = ["#86EFAC", "#BBF7D0", "#FDE68A", "#E2E8F0"]
    lbl = ParagraphStyle("KpiL", fontName=_INR_FONT, fontSize=7, textColor=colors.HexColor("#64748B"), alignment=1, leading=9)
    val = ParagraphStyle("KpiV", fontName=_INR_FONT_BOLD, fontSize=11, textColor=colors.HexColor("#14532D"), alignment=1, leading=14)
    cells = [[Paragraph(str(title).upper(), lbl), Paragraph(_inr(amount), val)] for title, amount in cards]
    width = avail / max(1, len(cards))
    table = Table([cells], colWidths=[width] * len(cards))
    cmds = [
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]
    for i, _ in enumerate(cards):
        cmds.append(("BACKGROUND", (i, 0), (i, 0), colors.HexColor(fills[i % 4])))
        cmds.append(("BOX", (i, 0), (i, 0), 0.6, colors.HexColor(boxes[i % 4])))
    table.setStyle(TableStyle(cmds))
    return table


def generate_excel_aggregator_report(
    db: Session,
    aggregator_id: Optional[int] = None,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    matrix = get_aggregator_payout_matrix(db, aggregator_id, branch_id, start_date, end_date)
    group = _export_payout_group(matrix, branch_id)
    groups = [group]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payout Breakup"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C6"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_title_rows = "1:5"

    title_font = Font(name="Calibri", size=18, bold=True, color=_INK)
    sub_font = Font(name="Calibri", size=10, italic=True, color=_MUTED)
    head_font = Font(name="Calibri", size=10, bold=True, color=_WHITE)
    label_font = Font(name="Calibri", size=10, bold=True, color=_INK)
    num_font = Font(name="Calibri", size=10, color=_INK)
    bold_num = Font(name="Calibri", size=10, bold=True, color=_INK)
    gst_font = bold_num
    unresolved_font = bold_num
    border = _thin_border()
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    period_label = _export_period_label(start_date, end_date)
    scope_name = group.get("branch_name") or "All Branches"
    if not branch_id:
        scope_name = "All Branches Combined"
    agg_name = _export_aggregator_name(matrix)
    title_bits = [p for p in (agg_name, "Payout Breakup", scope_name) if p]
    ws["A1"] = "  ·  ".join(title_bits)
    ws["A1"].font = title_font
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    generated = datetime.now().strftime("%d %b %Y, %I:%M %p")
    period_bit = f"{period_label}  ·  " if period_label else ""
    ws["A2"] = f"Harsh's RestoReconcile  ·  {period_bit}Combined totals  ·  Generated {generated}"
    if branch_id:
        ws["A2"] = f"Harsh's RestoReconcile  ·  {period_bit}Generated {generated}"
    ws["A2"].font = sub_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 16

    # Column map: 1 S.No, 2 Particular, then each branch's cycles + total
    row_codes = [r["code"] for r in _payout_row_defs()]
    data_start = 6
    row_excel = {code: data_start + i for i, code in enumerate(row_codes)}

    ws["A5"] = "S.No."
    ws["B5"] = "Particular"
    for cell in (ws["A5"], ws["B5"]):
        cell.fill = PatternFill("solid", fgColor=_NAVY)
        cell.font = head_font
        cell.alignment = center
        cell.border = border
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
    ws["A4"] = "Particulars"
    ws["A4"].fill = PatternFill("solid", fgColor=_NAVY)
    ws["A4"].font = head_font
    ws["A4"].alignment = center
    ws["A4"].border = border
    ws["B4"].fill = PatternFill("solid", fgColor=_NAVY)
    ws["B4"].border = border

    col = 3
    branch_spans = []
    for gi, group in enumerate(groups):
        cycles = [c for c in group["cycle_columns"] if c["key"] != "total"]
        if not cycles:
            cycles = [{"key": "empty", "label": "-"}]
        start_col = col
        for cycle in cycles:
            cell = ws.cell(row=5, column=col, value=cycle["label"])
            cell.fill = PatternFill("solid", fgColor=_NAVY)
            cell.font = head_font
            cell.alignment = center
            cell.border = border
            col += 1
        total_col = col
        tcell = ws.cell(row=5, column=total_col, value="Total")
        tcell.fill = PatternFill("solid", fgColor=_GRAY)
        tcell.font = Font(name="Calibri", size=10, bold=True, color=_INK)
        tcell.alignment = center
        tcell.border = border
        end_col = total_col
        header = ws.cell(row=4, column=start_col, value=str(group.get("branch_name") or "Branch").upper())
        header.fill = PatternFill("solid", fgColor=_NAVY)
        header.font = head_font
        header.alignment = center
        header.border = border
        if end_col > start_col:
            ws.merge_cells(start_row=4, start_column=start_col, end_row=4, end_column=end_col)
            for c in range(start_col, end_col + 1):
                ws.cell(row=4, column=c).fill = PatternFill("solid", fgColor=_NAVY)
                ws.cell(row=4, column=c).border = border
                ws.cell(row=4, column=c).font = head_font
        branch_spans.append({
            "group": group,
            "cycles": cycles,
            "start": start_col,
            "total": total_col,
            "end": end_col,
            "body": _WHITE,
        })
        col = end_col + 1

    last_col = max(2, col - 1)
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22

    input_codes = {"TOTAL_SALE", "PAYOUT", "COMMISSION", "PROMOTION", "TCS", "TDS", "MISC", "GST_9_5", "PACKING_CHARGES"}
    labels = {r["code"]: r["particular"] for r in _payout_row_defs()}
    sno = {r["code"]: r["s_no"] for r in _payout_row_defs()}

    for code, excel_row in row_excel.items():
        ws.cell(row=excel_row, column=1, value=sno[code]).alignment = center
        ws.cell(row=excel_row, column=2, value=labels[code]).alignment = left
        ws.row_dimensions[excel_row].height = 20

        is_unresolved = code == "DIFFERENCE_ADJUSTMENT"
        is_theme = code in ("DIFFERENCE", "TOTAL_DEDUCTIONS", "DIFFERENCE_ADJUSTMENT")

        for span in branch_spans:
            group = span["group"]
            row_map = {r["code"]: r for r in group["rows"]}
            src = row_map.get(code, {})
            first_letter = get_column_letter(span["start"])
            last_cycle_letter = get_column_letter(span["total"] - 1) if span["total"] > span["start"] else first_letter
            total_letter = get_column_letter(span["total"])

            for i, cycle in enumerate(span["cycles"]):
                cidx = span["start"] + i
                letter = get_column_letter(cidx)
                cell = ws.cell(row=excel_row, column=cidx)
                if code == "DIFFERENCE":
                    sale_r = row_excel["TOTAL_SALE"]
                    pay_r = row_excel["PAYOUT"]
                    cell.value = f"={letter}{sale_r}-{letter}{pay_r}"
                elif code == "TOTAL_DEDUCTIONS":
                    cell.value = (
                        f"={letter}{row_excel['COMMISSION']}"
                        f"+{letter}{row_excel['PROMOTION']}"
                        f"+{letter}{row_excel['TCS']}"
                        f"+{letter}{row_excel['TDS']}"
                        f"+{letter}{row_excel['MISC']}"
                    )
                elif code == "DIFFERENCE_ADJUSTMENT":
                    cell.value = f"={letter}{row_excel['DIFFERENCE']}-{letter}{row_excel['TOTAL_DEDUCTIONS']}"
                else:
                    cell.value = float(src.get(cycle["key"], 0.0) or 0.0)
                cell.number_format = _INR_XLS
                cell.alignment = right
                cell.border = border

            tcell = ws.cell(row=excel_row, column=span["total"])
            if code in input_codes and span["total"] > span["start"]:
                tcell.value = f"=SUM({first_letter}{excel_row}:{last_cycle_letter}{excel_row})"
            elif code == "DIFFERENCE":
                tcell.value = f"={total_letter}{row_excel['TOTAL_SALE']}-{total_letter}{row_excel['PAYOUT']}"
            elif code == "TOTAL_DEDUCTIONS":
                tcell.value = (
                    f"={total_letter}{row_excel['COMMISSION']}"
                    f"+{total_letter}{row_excel['PROMOTION']}"
                    f"+{total_letter}{row_excel['TCS']}"
                    f"+{total_letter}{row_excel['TDS']}"
                    f"+{total_letter}{row_excel['MISC']}"
                )
            elif code == "DIFFERENCE_ADJUSTMENT":
                tcell.value = f"={total_letter}{row_excel['DIFFERENCE']}-{total_letter}{row_excel['TOTAL_DEDUCTIONS']}"
            else:
                tcell.value = float(src.get("total", 0.0) or 0.0)
            tcell.number_format = _INR_XLS
            tcell.alignment = right
            tcell.border = border
            tcell.fill = PatternFill("solid", fgColor="F0FDF4" if is_theme else _GRAY)
            tcell.font = bold_num if is_unresolved else num_font

            body_fill = PatternFill("solid", fgColor=_WHITE)
            for cidx in range(span["start"], span["total"]):
                cell = ws.cell(row=excel_row, column=cidx)
                cell.fill = body_fill
                cell.font = bold_num if is_unresolved else num_font
                if is_theme:
                    cell.fill = PatternFill("solid", fgColor="F0FDF4")

        for cidx in (1, 2):
            lab = ws.cell(row=excel_row, column=cidx)
            lab.border = border
            lab.font = bold_num if is_unresolved else num_font
            lab.fill = PatternFill("solid", fgColor="F0FDF4" if is_theme else _WHITE)

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 48
    for cidx in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(cidx)].width = 14

    note_row = data_start + len(row_codes) + 1
    if not branch_id:
        note = "All branches are combined into one total. Difference, Total Deductions, Unresolved Difference, and Totals are Excel formulas."
    else:
        note = "Difference, Total Deductions, Unresolved Difference, and Totals are Excel formulas linked to the imported figures."
    ws.cell(row=note_row, column=1, value=note).font = sub_font
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=min(6, last_col))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _payout_row_defs() -> List[Dict[str, Any]]:
    from app.services.aggregator_service import _payout_row_templates
    return _payout_row_templates()


def generate_pdf_aggregator_report(
    db: Session,
    aggregator_id: Optional[int] = None,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    matrix = get_aggregator_payout_matrix(db, aggregator_id, branch_id, start_date, end_date)
    group = _export_payout_group(matrix, branch_id)
    cycle_cols = group.get("cycle_columns") or []
    rows = group.get("rows") or []

    scope = group.get("branch_name") or "All Branches"
    if not branch_id:
        scope = "All Branches Combined"
    period = _export_period_label(start_date, end_date)
    agg_name = _export_aggregator_name(matrix)
    heading = f"{agg_name} Payout Breakup" if agg_name else "Payout Breakup"

    buffer = io.BytesIO()
    page_w, page_h = landscape(A4)
    left_m, right_m = 22, 22
    avail = page_w - left_m - right_m
    doc = pdf_document(buffer, landscape(A4), left=left_m, right=right_m, bottom=36)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'PayoutTitle', parent=styles['Heading1'], fontName=_INR_FONT_BOLD,
        fontSize=16, textColor=colors.HexColor('#14532D'), spaceAfter=1, leading=18
    )
    subtitle_style = ParagraphStyle(
        'PayoutSub', parent=styles['Normal'], fontName=_INR_FONT,
        fontSize=8.5, textColor=colors.HexColor('#64748B'), spaceAfter=6, leading=11
    )
    kpi_lbl = ParagraphStyle(
        'KpiLbl', fontName=_INR_FONT, fontSize=7.5,
        textColor=colors.HexColor('#64748B'), alignment=1, leading=10
    )
    kpi_val = ParagraphStyle(
        'KpiVal', fontName=_INR_FONT_BOLD, fontSize=12,
        textColor=colors.HexColor('#14532D'), alignment=1, leading=15
    )
    th_style = ParagraphStyle(
        'PayoutTH', fontName=_INR_FONT_BOLD, fontSize=7,
        textColor=colors.white, alignment=1, leading=9
    )
    td_style = ParagraphStyle(
        'PayoutTD', fontName=_INR_FONT, fontSize=7,
        textColor=colors.HexColor('#1E293B'), leading=9
    )
    td_bold = ParagraphStyle(
        'PayoutTDBold', fontName=_INR_FONT_BOLD, fontSize=7,
        textColor=colors.HexColor('#0F172A'), leading=9
    )
    td_num = ParagraphStyle(
        'PayoutTDNum', fontName=_INR_FONT, fontSize=7,
        textColor=colors.HexColor('#0F172A'), alignment=2, leading=9
    )
    td_num_bold = ParagraphStyle(
        'PayoutTDNumB', fontName=_INR_FONT_BOLD, fontSize=7,
        textColor=colors.HexColor('#0F172A'), alignment=2, leading=9
    )
    td_center = ParagraphStyle(
        'PayoutTDCtr', fontName=_INR_FONT, fontSize=7,
        textColor=colors.HexColor('#64748B'), alignment=1, leading=9
    )

    sale = _row_total(rows, "TOTAL_SALE")
    payout = _row_total(rows, "PAYOUT")
    diff = _row_total(rows, "DIFFERENCE")
    deductions = _row_total(rows, "TOTAL_DEDUCTIONS")
    unresolved = _row_total(rows, "DIFFERENCE_ADJUSTMENT")

    generated = datetime.now().strftime("%d %b %Y, %I:%M %p")
    meta = "  ·  ".join(p for p in (
        scope,
        period,
        "Combined totals" if not branch_id else "",
        f"Generated {generated}",
    ) if p)

    kpi_cards = Table(
        [[
            [Paragraph("TOTAL SALE", kpi_lbl), Paragraph(_inr(sale), kpi_val)],
            [Paragraph("PAYOUT RECEIVED", kpi_lbl), Paragraph(_inr(payout), kpi_val)],
            [Paragraph("DIFFERENCE", kpi_lbl), Paragraph(_inr(diff), kpi_val)],
            [Paragraph("TOTAL DEDUCTIONS", kpi_lbl), Paragraph(_inr(deductions), kpi_val)],
        ]],
        colWidths=[avail * 0.16] * 4
    )
    kpi_cards.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#ECFDF5')),
        ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#F0FDF4')),
        ('BACKGROUND', (2, 0), (2, 0), colors.HexColor('#FFFBEB')),
        ('BACKGROUND', (3, 0), (3, 0), colors.HexColor('#F8FAFC')),
        ('BOX', (0, 0), (0, 0), 0.7, colors.HexColor('#86EFAC')),
        ('BOX', (1, 0), (1, 0), 0.7, colors.HexColor('#BBF7D0')),
        ('BOX', (2, 0), (2, 0), 0.7, colors.HexColor('#FDE68A')),
        ('BOX', (3, 0), (3, 0), 0.7, colors.HexColor('#E2E8F0')),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('RIGHTPADDING', (0, 0), (2, 0), 8),
    ]))

    pie = _deduction_pie_drawing(rows, width=avail * 0.36, height=168)
    dash = Table(
        [[kpi_cards, pie]],
        colWidths=[avail * 0.62, avail * 0.38]
    )
    dash.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (0, 0), 10),
        ('RIGHTPADDING', (1, 0), (1, 0), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))

    headers = ["S.No.", "Particulars"] + [c["label"] for c in cycle_cols]
    table_data = [[Paragraph(h, th_style) for h in headers]]
    theme_rows = set()
    highlight_rows = set()
    for idx, r in enumerate(rows, start=1):
        is_highlight = bool(r.get("highlight")) or r.get("code") == "DIFFERENCE_ADJUSTMENT"
        is_theme = bool(r.get("theme_row"))
        label_style = td_bold if is_highlight else td_style
        r_cells = [
            Paragraph(str(r.get("s_no") or idx), td_center),
            Paragraph(r.get("particular") or "", label_style),
        ]
        for c in cycle_cols:
            val = float(r.get(c["key"], 0.0) or 0.0)
            use_bold = c["key"] == "total" or is_highlight
            r_cells.append(Paragraph(_inr(val), td_num_bold if use_bold else td_num))
        table_data.append(r_cells)
        if is_theme:
            theme_rows.add(idx)
        if is_highlight:
            highlight_rows.add(idx)

    num_cols = len(headers)
    sno_w = 32
    part_w = 210
    other_w = max(58, (avail - sno_w - part_w) / max(1, num_cols - 2))
    payout_table = Table(
        table_data,
        colWidths=[sno_w, part_w] + [other_w] * (num_cols - 2)
    )
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#166534')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3.5),
        ('TOPPADDING', (0, 0), (-1, -1), 3.5),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#D1FAE5')),
        ('LINEBELOW', (0, 0), (-1, 0), 0, colors.HexColor('#166534')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ]
    for ridx in theme_rows:
        t_style.append(('BACKGROUND', (0, ridx), (-1, ridx), colors.HexColor('#ECFDF5')))
    for ridx in highlight_rows:
        t_style.append(('BACKGROUND', (0, ridx), (-1, ridx), colors.HexColor('#FEF3C7')))
    if cycle_cols:
        t_style.append(('BACKGROUND', (-1, 1), (-1, -1), colors.HexColor('#F8FAFC')))
        for ridx in theme_rows:
            t_style.append(('BACKGROUND', (-1, ridx), (-1, ridx), colors.HexColor('#DCFCE7')))
        for ridx in highlight_rows:
            t_style.append(('BACKGROUND', (-1, ridx), (-1, ridx), colors.HexColor('#FDE68A')))
    payout_table.setStyle(TableStyle(t_style))

    note_style = ParagraphStyle(
        'PayoutNote', fontName=_INR_FONT, fontSize=7,
        textColor=colors.HexColor('#64748B'), spaceBefore=4
    )
    if not branch_id:
        note = f"Figures above total every branch into one report. Unresolved difference is {_inr(unresolved)}."
    else:
        note = f"Single-branch report for {scope}. Unresolved difference is {_inr(unresolved)}."

    elements = [
        Paragraph(heading, title_style),
        Paragraph(meta, subtitle_style),
        dash,
        Spacer(1, 8),
        payout_table,
        Paragraph(note, note_style),
    ]
    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


# ==============================================================================
# AUDIT LOG REPORT EXPORTS
# ==============================================================================
def generate_excel_audit_report(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    query = db.query(AuditLog)
    if start_date:
        query = query.filter(AuditLog.timestamp >= start_date)
    if end_date:
        query = query.filter(AuditLog.timestamp <= end_date)

    logs = query.order_by(AuditLog.timestamp.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    headers = ["Timestamp", "User", "Action", "Entity", "Entity ID", "IP Address"]
    start_row = 4
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_num, value=header)

    for row_idx, r in enumerate(logs, start=start_row + 1):
        ws.cell(row=row_idx, column=1, value=r.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_idx, column=2, value=r.username or "System")
        ws.cell(row=row_idx, column=3, value=r.action)
        ws.cell(row=row_idx, column=4, value=r.entity_name)
        ws.cell(row=row_idx, column=5, value=r.entity_id or "")
        ws.cell(row=row_idx, column=6, value=r.ip_address or "")

    _format_excel_worksheet(ws, "System Audit Trail Log Report", start_row=start_row)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_pdf_audit_report(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> bytes:
    query = db.query(AuditLog)
    if start_date:
        query = query.filter(AuditLog.timestamp >= start_date)
    if end_date:
        query = query.filter(AuditLog.timestamp <= end_date)

    logs = query.order_by(AuditLog.timestamp.desc()).all()
    buffer = io.BytesIO()

    doc = pdf_document(buffer, letter, left=36, right=36, bottom=54)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('DocTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#0F172A'), spaceAfter=4)
    subtitle_style = ParagraphStyle('DocSubTitle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, textColor=colors.HexColor('#64748B'), spaceAfter=14)
    th_style = ParagraphStyle('TH', fontName='Helvetica-Bold', fontSize=8, textColor=colors.white, alignment=1)
    td_style = ParagraphStyle('TD', fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#1E293B'))

    elements = [
        Paragraph("System Audit Trail Report", title_style),
        Paragraph(f"Generated on {datetime.now().strftime('%d %b %Y, %I:%M %p')} | Security & Action Logs", subtitle_style),
        Spacer(1, 8)
    ]

    headers = ["Timestamp", "User", "Action", "Entity", "Entity ID", "IP Address"]
    table_data = [[Paragraph(h, th_style) for h in headers]]

    for r in logs:
        table_data.append([
            Paragraph(r.timestamp.strftime("%Y-%m-%d %H:%M:%S"), td_style),
            Paragraph(r.username or "System", td_style),
            Paragraph(r.action, td_style),
            Paragraph(r.entity_name, td_style),
            Paragraph(r.entity_id or "-", td_style),
            Paragraph(r.ip_address or "-", td_style)
        ])

    col_widths = [110, 100, 75, 100, 75, 80]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            t_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC')))

    t.setStyle(TableStyle(t_style))
    elements.append(t)

    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


def generate_pdf_attendance_report(
    db: Session,
    branch_id: Optional[int],
    year: int,
    month: int,
) -> bytes:
    from app.services.attendance_service import get_attendance_matrix

    if branch_id:
        branches = db.query(Branch).filter(Branch.id == branch_id).all()
    else:
        branches = db.query(Branch).filter(Branch.is_active == True).order_by(Branch.name).all()
    if not branches:
        branches = db.query(Branch).order_by(Branch.name).all()

    month_name = date(year, month, 1).strftime("%B %Y")
    buffer = io.BytesIO()
    doc = pdf_document(buffer, landscape(A4), left=18, right=18, bottom=36)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AttTitle", parent=styles["Heading1"], fontName=_INR_FONT_BOLD,
        fontSize=16, textColor=colors.HexColor("#166534"), spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "AttSub", parent=styles["Normal"], fontName=_INR_FONT,
        fontSize=9, textColor=colors.HexColor("#64748B"), spaceAfter=10,
    )
    th_style = ParagraphStyle("AttTH", fontName=_INR_FONT_BOLD, fontSize=6, textColor=colors.white, alignment=1)
    td_style = ParagraphStyle("AttTD", fontName=_INR_FONT, fontSize=6.5, textColor=colors.HexColor("#1E293B"))
    mark_style = ParagraphStyle("AttMark", fontName=_INR_FONT_BOLD, fontSize=6.5, alignment=1)

    elements = [
        Paragraph("Attendance Reconciliation Report", title_style),
        Paragraph(
            f"{month_name} · Generated {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
            subtitle_style,
        ),
    ]

    mark_colors = {
        "P": colors.HexColor("#166534"),
        "A": colors.HexColor("#B91C1C"),
        "O": colors.HexColor("#92400E"),
        "WO": colors.HexColor("#92400E"),
        "H": colors.HexColor("#92400E"),
        "L": colors.HexColor("#1D4ED8"),
    }
    mark_fills = {
        "P": colors.HexColor("#DCFCE7"),
        "A": colors.HexColor("#FEE2E2"),
        "O": colors.HexColor("#FEF3C7"),
        "WO": colors.HexColor("#FEF3C7"),
        "H": colors.HexColor("#FEF3C7"),
        "L": colors.HexColor("#DBEAFE"),
    }

    for branch in branches:
        matrix = get_attendance_matrix(db, branch.id, year, month)
        days = matrix["days"]
        elements.append(Paragraph(f"{branch.name} ({branch.code})", subtitle_style))
        headers = ["Name", "Rank", "Team"] + [str(d) for d in range(1, days + 1)] + ["P", "A", "WO", "L", "Gross", "Cash Adv", "Bank Adv", "Net"]
        table_data = [[Paragraph(h, th_style) for h in headers]]
        cell_marks: List[List[Optional[str]]] = []
        for emp in matrix["employees"]:
            row = [
                Paragraph(emp["name"] or "", td_style),
                Paragraph(emp.get("rank") or "", td_style),
                Paragraph(emp.get("team") or "", td_style),
            ]
            marks_row: List[Optional[str]] = [None, None, None]
            for d in range(1, days + 1):
                mark = (emp.get("marks") or {}).get(str(d)) or ""
                row.append(Paragraph(mark or "", mark_style))
                marks_row.append(mark or None)
            row.extend([
                Paragraph(str(emp.get("present") or 0), td_style),
                Paragraph(str(emp.get("absent") or 0), td_style),
                Paragraph(str(emp.get("weekly_off") or emp.get("off") or 0), td_style),
                Paragraph(str(emp.get("leave") or 0), td_style),
                Paragraph(f"{float(emp.get('gross_salary') or 0):,.0f}", td_style),
                Paragraph(f"{float(emp.get('cash_advance') or emp.get('advance') or 0):,.0f}", td_style),
                Paragraph(f"{float(emp.get('bank_advance') or 0):,.0f}", td_style),
                Paragraph(f"{float(emp.get('net_salary') or 0):,.0f}", td_style),
            ])
            marks_row.extend([None, None, None, None, None, None, None, None])
            table_data.append(row)
            cell_marks.append(marks_row)
        if len(table_data) == 1:
            table_data.append([Paragraph("No attendance recorded.", td_style)] + [""] * (len(headers) - 1))
            cell_marks.append([None] * len(headers))

        name_w, rank_w, team_w, tot_w, money_w = 64, 38, 34, 15, 32
        day_w = max(10, (780 - name_w - rank_w - team_w - tot_w * 4 - money_w * 4) / float(days))
        col_widths = [name_w, rank_w, team_w] + [day_w] * days + [tot_w, tot_w, tot_w, tot_w, money_w, money_w, money_w, money_w]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#166534")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (3, 1), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#D7E4DB")),
        ]
        for r_idx, marks_row in enumerate(cell_marks, start=1):
            if r_idx % 2 == 0:
                style_cmds.append(("BACKGROUND", (0, r_idx), (2, r_idx), colors.HexColor("#F8FAF9")))
            for c_idx, mark in enumerate(marks_row):
                if mark and mark in mark_fills:
                    style_cmds.append(("BACKGROUND", (c_idx, r_idx), (c_idx, r_idx), mark_fills[mark]))
                    style_cmds.append(("TEXTCOLOR", (c_idx, r_idx), (c_idx, r_idx), mark_colors[mark]))
        table.setStyle(TableStyle(style_cmds))
        elements.append(table)
        summary = matrix.get("summary") or {}
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(
            f"Staff {summary.get('staff', 0)} · Present {summary.get('present', 0)} · "
            f"Absent {summary.get('absent', 0)} · Weekly off {summary.get('weekly_off', 0)} · "
            f"Leave {summary.get('leave', 0)} (max {summary.get('leave_allowed', 2)})",
            subtitle_style,
        ))
        elements.append(Spacer(1, 10))

    doc.build(elements, canvasmaker=NumberedCanvas)
    return buffer.getvalue()


# ==============================================================================
# ANALYTICS DASHBOARD DATA GENERATOR
# ==============================================================================
def get_analytics_summary_data(
    db: Session,
    branch_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> Dict[str, Any]:
    daybook_res = get_consolidated_daybook(db, branch_id, start_date, end_date)
    
    # 1. Channel Distribution (Pie/Doughnut)
    totals = {
        "Cash": 0.0,
        "Card / QR": 0.0,
        "Zomato": 0.0,
        "Swiggy": 0.0,
        "Dineout": 0.0,
        "Other Channels": 0.0
    }

    # 2. Daily Sales Trend (Line Graph)
    trend_map: Dict[str, float] = {}

    # 3. Branch Performance (Bar Chart)
    branch_map: Dict[str, float] = {}

    # 4. Status Counts
    status_counts = {"RECONCILED": 0, "DIFFERENCE": 0, "PENDING": 0}

    for r in daybook_res:
        c_cash = float(r["cash"] or 0)
        c_card = float(r["card_qr"] or 0)
        c_zom = float(r["zomato"] or 0)
        c_swig = float(r["swiggy"] or 0)
        c_dine = float(r["dineout"] or 0)
        c_oth = float(r["other_channels"] or 0)
        tot_s = float(r["total_sales"] or 0)

        totals["Cash"] += c_cash
        totals["Card / QR"] += c_card
        totals["Zomato"] += c_zom
        totals["Swiggy"] += c_swig
        totals["Dineout"] += c_dine
        totals["Other Channels"] += c_oth

        d_str = r["date"].strftime("%Y-%m-%d")
        trend_map[d_str] = trend_map.get(d_str, 0.0) + tot_s

        b_name = r["branch_name"]
        branch_map[b_name] = branch_map.get(b_name, 0.0) + tot_s

        st = r["status"]
        if st in status_counts:
            status_counts[st] += 1
        else:
            status_counts["PENDING"] += 1

    sorted_dates = sorted(trend_map.keys())
    
    return {
        "channel_distribution": {
            "labels": list(totals.keys()),
            "data": list(totals.values())
        },
        "sales_trend": {
            "labels": sorted_dates,
            "data": [trend_map[d] for d in sorted_dates]
        },
        "branch_performance": {
            "labels": list(branch_map.keys()),
            "data": list(branch_map.values())
        },
        "reconciliation_status": {
            "labels": ["Reconciled", "Difference", "Pending"],
            "data": [status_counts["RECONCILED"], status_counts["DIFFERENCE"], status_counts["PENDING"]]
        }
    }
