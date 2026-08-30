"""
BOI Account Opening Audit & Document Scrutiny System
Excel Report Generator using OpenPyXL with Professional Banking Styles
"""

import io
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from checklists import get_checklist_for_type
from data_manager import calculate_account_metrics, calculate_global_summary

# BOI Brand Colors
NAVY_HEADER = "0B2545"        # Dark Navy Blue
ORANGE_ACCENT = "E65100"      # BOI Orange
LIGHT_BG_NAVY = "EBF3FA"      # Soft Navy highlight
LIGHT_GRAY_FILL = "F8F9FA"    # Zebra striping
ZEBRA_FILL = "F0F4F8"

# Status Colors
GREEN_FILL = "E8F5E9"         # Light Green
GREEN_TEXT = "1B5E20"         # Dark Green
ORANGE_FILL = "FFF3E0"        # Light Orange
ORANGE_TEXT = "E65100"        # Dark Orange
RED_FILL = "FFEBEE"           # Light Red
RED_TEXT = "B71C1C"           # Dark Red
BLUE_FILL = "E3F2FD"          # Light Blue
BLUE_TEXT = "0D47A1"          # Dark Blue


def create_excel_styles():
    """Defines reusable OpenPyXL styles for the audit report."""
    thin_border = Border(
        left=Side(style='thin', color='D0D5DD'),
        right=Side(style='thin', color='D0D5DD'),
        top=Side(style='thin', color='D0D5DD'),
        bottom=Side(style='thin', color='D0D5DD')
    )
    
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="0B2545")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="555555")
    
    section_font = Font(name="Calibri", size=12, bold=True, color=NAVY_HEADER)
    section_fill = PatternFill(start_color=LIGHT_BG_NAVY, end_color=LIGHT_BG_NAVY, fill_type="solid")
    
    data_font = Font(name="Calibri", size=10, color="1F2937")
    bold_data_font = Font(name="Calibri", size=10, bold=True, color="1F2937")
    
    return {
        "thin_border": thin_border,
        "header_font": header_font,
        "header_fill": header_fill,
        "title_font": title_font,
        "subtitle_font": subtitle_font,
        "section_font": section_font,
        "section_fill": section_fill,
        "data_font": data_font,
        "bold_data_font": bold_data_font
    }


def auto_fit_columns(worksheet):
    """Automatically adjusts column widths for optimal readability."""
    for col in worksheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if '\n' in val_str:
                lines = val_str.split('\n')
                max_len = max(max_len, max(len(l) for l in lines))
            else:
                max_len = max(max_len, len(val_str))
        # Add safety padding
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 55)


def apply_status_style(cell, status_text: str):
    """Applies colored badge fills depending on status text."""
    st = str(status_text).strip()
    if st in ["Passed", "Approved", "Re-check Completed", "Low"]:
        cell.fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type="solid")
        cell.font = Font(name="Calibri", size=10, bold=True, color=GREEN_TEXT)
    elif st in ["Ready for Approval", "Rectified", "Medium"]:
        cell.fill = PatternFill(start_color=BLUE_FILL, end_color=BLUE_FILL, fill_type="solid")
        cell.font = Font(name="Calibri", size=10, bold=True, color=BLUE_TEXT)
    elif st in ["Rectification Pending", "Under Scrutiny"]:
        cell.fill = PatternFill(start_color=ORANGE_FILL, end_color=ORANGE_FILL, fill_type="solid")
        cell.font = Font(name="Calibri", size=10, bold=True, color=ORANGE_TEXT)
    elif st in ["Discrepancy Found", "High", "Critical", "APPROVAL BLOCKED"]:
        cell.fill = PatternFill(start_color=RED_FILL, end_color=RED_FILL, fill_type="solid")
        cell.font = Font(name="Calibri", size=10, bold=True, color=RED_TEXT)


def generate_excel_report_bytes(accounts: List[Dict[str, Any]]) -> io.BytesIO:
    """
    Generates a professional 4-sheet Bank of India Audit Report Workbook (.xlsx)
    and returns it as an in-memory BytesIO stream.
    """
    wb = openpyxl.Workbook()
    styles = create_excel_styles()
    
    # -------------------------------------------------------------
    # SHEET 1: Executive Audit Summary
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    # Report Header
    ws1["A1"] = "BANK OF INDIA - CONCURRENT AUDIT & DOCUMENT SCRUTINY SYSTEM"
    ws1["A1"].font = styles["title_font"]
    ws1["A2"] = f"Training & Demo Audit Report (100% Synthetic Data) | Generated On: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}"
    ws1["A2"].font = styles["subtitle_font"]
    
    global_metrics = calculate_global_summary(accounts)
    
    # KPI Section
    ws1.append([])
    ws1.append(["AUDIT EXECUTIVE SCORECARD", "", "", ""])
    curr_row = 4
    ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
    ws1.cell(row=curr_row, column=1).font = styles["section_font"]
    ws1.cell(row=curr_row, column=1).fill = styles["section_fill"]
    
    kpis = [
        ("Total Accounts Audited", global_metrics["total_accounts"], "Saving Accounts (8 Checks)", global_metrics["saving_count"]),
        ("Approved Accounts", global_metrics["approved_count"], "Current Accounts (11 Checks)", global_metrics["current_count"]),
        ("Ready for Final Approval", global_metrics["ready_for_approval_count"], "Total Discrepancies Logged", global_metrics["total_discrepancies"]),
        ("Pending Discrepancies", global_metrics["pending_discrepancies"], "Rectified / Re-checked Items", global_metrics["rectified_count"]),
    ]
    
    for row_data in kpis:
        ws1.append(list(row_data))
        curr_row += 1
        for col_idx in [1, 3]:
            c = ws1.cell(row=curr_row, column=col_idx)
            c.font = styles["bold_data_font"]
            c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            c.border = styles["thin_border"]
        for col_idx in [2, 4]:
            c = ws1.cell(row=curr_row, column=col_idx)
            c.font = styles["data_font"]
            c.alignment = Alignment(horizontal="center")
            c.border = styles["thin_border"]
            
    # Discrepancy Breakdown by Category Table
    ws1.append([])
    curr_row += 2
    ws1.append(["DISCREPANCY BREAKDOWN BY CATEGORY", "", "", ""])
    ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=4)
    ws1.cell(row=curr_row, column=1).font = styles["section_font"]
    ws1.cell(row=curr_row, column=1).fill = styles["section_fill"]
    
    ws1.append(["Checklist Category / Area", "Deficiency Count", "Audit Impact", "Standard Action Required"])
    curr_row += 1
    for c_idx in range(1, 5):
        cell = ws1.cell(row=curr_row, column=c_idx)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center" if c_idx in [2, 3] else "left")
        cell.border = styles["thin_border"]
        
    for cat, count in global_metrics["category_discrepancies"].items():
        impact = "High / Approval Blocking" if count > 1 else "Medium"
        action = "Obtain missing documents from branch / verify OSV stamp"
        ws1.append([cat, count, impact, action])
        curr_row += 1
        for c_idx in range(1, 5):
            cell = ws1.cell(row=curr_row, column=c_idx)
            cell.font = styles["data_font"]
            cell.border = styles["thin_border"]
            if c_idx == 2:
                cell.alignment = Alignment(horizontal="center")
                apply_status_style(cell, "Discrepancy Found")
    
    auto_fit_columns(ws1)

    # -------------------------------------------------------------
    # SHEET 2: Master Account Register
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Master Account Register")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2["A1"] = "BANK OF INDIA - AUDITED ACCOUNTS MASTER REGISTER"
    ws2["A1"].font = styles["title_font"]
    ws2["A2"] = "Summary of all synthetic accounts, compliance rates, and scrutiny approval states."
    ws2["A2"].font = styles["subtitle_font"]
    ws2.append([])
    
    headers_ws2 = [
        "Account Ref ID", "Dummy Account No", "Synthetic Customer / Entity Name",
        "Account Type", "Branch Name", "Submission Date", "Risk Category",
        "Total Checks", "Passed Checks", "Pending Issues", "Compliance %",
        "Audit Status", "Approved?", "Approval Date & Time", "Auditor PF"
    ]
    
    ws2.append(headers_ws2)
    header_row_ws2 = 4
    for col_idx in range(1, len(headers_ws2) + 1):
        cell = ws2.cell(row=header_row_ws2, column=col_idx)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["thin_border"]
        
    row_idx = header_row_ws2 + 1
    for acc in accounts:
        metrics = calculate_account_metrics(acc)
        row_vals = [
            acc.get("account_id", ""),
            acc.get("dummy_account_no", ""),
            acc.get("customer_name", ""),
            acc.get("account_type", ""),
            acc.get("branch_name", ""),
            acc.get("submission_date", ""),
            acc.get("risk_category", "Low"),
            metrics["total_checks"],
            metrics["passed_checks"],
            metrics["pending_issues_count"],
            f"{metrics['compliance_pct']}%",
            metrics["overall_status"],
            "YES" if metrics["is_approved"] else "NO",
            acc.get("approval_date", "Pending"),
            acc.get("auditor_pf", "")
        ]
        ws2.append(row_vals)
        
        # Apply styling & colors
        for col_idx in range(1, len(row_vals) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.font = styles["data_font"]
            cell.border = styles["thin_border"]
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
            
            if col_idx in [8, 9, 10, 11, 13, 14, 15]:
                cell.alignment = Alignment(horizontal="center")
            if col_idx == 7:  # Risk Category
                apply_status_style(cell, acc.get("risk_category", "Low"))
            elif col_idx == 12:  # Audit Status
                apply_status_style(cell, metrics["overall_status"])
            elif col_idx == 13:  # Approved?
                apply_status_style(cell, "Approved" if metrics["is_approved"] else "Discrepancy Found")
                
        row_idx += 1
        
    auto_fit_columns(ws2)

    # -------------------------------------------------------------
    # SHEET 3: Checklist & Discrepancy Log
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="Checklist & Discrepancy Log")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3["A1"] = "BANK OF INDIA - GRANULAR CHECKLIST & SCRUTINY OBSERVATIONS LOG"
    ws3["A1"].font = styles["title_font"]
    ws3["A2"] = "Item-by-item verification records, scrutiny remarks, and rectification action details."
    ws3["A2"].font = styles["subtitle_font"]
    ws3.append([])
    
    headers_ws3 = [
        "Account Ref ID", "Customer Name", "Account Type", "Check Code",
        "Mandatory Check Title", "Checklist Category", "Severity",
        "Verification Status", "Auditor Observation Remarks", "Rectification Action Remarks"
    ]
    
    ws3.append(headers_ws3)
    header_row_ws3 = 4
    for col_idx in range(1, len(headers_ws3) + 1):
        cell = ws3.cell(row=header_row_ws3, column=col_idx)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["thin_border"]
        
    row_idx = header_row_ws3 + 1
    for acc in accounts:
        acc_type = acc.get("account_type", "Saving Account")
        checklist_items = get_checklist_for_type(acc_type)
        acc_checks = acc.get("checks", {})
        
        for item in checklist_items:
            cid = item["id"]
            check_data = acc_checks.get(cid, {"status": "Discrepancy Found", "remarks": "Pending verification", "rectification_notes": ""})
            status = check_data.get("status", "Discrepancy Found")
            
            row_vals = [
                acc.get("account_id", ""),
                acc.get("customer_name", ""),
                acc_type,
                cid,
                item["title"],
                item["category"],
                item["severity"],
                status,
                check_data.get("remarks", ""),
                check_data.get("rectification_notes", "")
            ]
            ws3.append(row_vals)
            
            for col_idx in range(1, len(row_vals) + 1):
                cell = ws3.cell(row=row_idx, column=col_idx)
                cell.font = styles["data_font"]
                cell.border = styles["thin_border"]
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
                if col_idx in [4, 7, 8]:
                    cell.alignment = Alignment(horizontal="center")
                if col_idx == 8:  # Verification Status
                    apply_status_style(cell, status)
                elif col_idx == 7:  # Severity
                    apply_status_style(cell, item["severity"])
                    
            row_idx += 1
            
    auto_fit_columns(ws3)

    # -------------------------------------------------------------
    # SHEET 4: High-Risk & Blocked Accounts
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="Blocked & High Risk Accounts")
    ws4.views.sheetView[0].showGridLines = True
    
    ws4["A1"] = "BANK OF INDIA - BLOCKED APPROVALS & HIGH RISK ACCOUNTS"
    ws4["A1"].font = styles["title_font"]
    ws4["A2"] = "Accounts requiring branch escalation or immediate document remediation before approval."
    ws4["A2"].font = styles["subtitle_font"]
    ws4.append([])
    
    headers_ws4 = [
        "Account Ref ID", "Customer / Entity Name", "Account Type", "Branch",
        "Risk Category", "Pending Discrepancies Count", "Approval State",
        "Outstanding Deficiencies Summary", "Auditor Action Mandate"
    ]
    
    ws4.append(headers_ws4)
    header_row_ws4 = 4
    for col_idx in range(1, len(headers_ws4) + 1):
        cell = ws4.cell(row=header_row_ws4, column=col_idx)
        cell.font = styles["header_font"]
        cell.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")  # Dark Red Header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = styles["thin_border"]
        
    row_idx = header_row_ws4 + 1
    for acc in accounts:
        metrics = calculate_account_metrics(acc)
        # Include if not approved or risk is High
        if not metrics["is_approved"] or acc.get("risk_category") == "High":
            discrepancy_titles = [f"[{d['id']}] {d['title']}" for d in metrics["discrepant_items"]]
            summary_str = "; ".join(discrepancy_titles) if discrepancy_titles else "High AML Risk Review Completed"
            mandate = "APPROVAL BLOCKED - Do not open in Finacle" if not metrics["is_approved"] else "Enhanced Due Diligence (EDD) Required"
            
            row_vals = [
                acc.get("account_id", ""),
                acc.get("customer_name", ""),
                acc.get("account_type", ""),
                acc.get("branch_name", ""),
                acc.get("risk_category", "Low"),
                metrics["pending_issues_count"],
                "APPROVAL BLOCKED" if not metrics["is_ready_for_approval"] else ("READY" if not metrics["is_approved"] else "APPROVED"),
                summary_str,
                mandate
            ]
            ws4.append(row_vals)
            
            for col_idx in range(1, len(row_vals) + 1):
                cell = ws4.cell(row=row_idx, column=col_idx)
                cell.font = styles["data_font"]
                cell.border = styles["thin_border"]
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
                if col_idx in [5, 6, 7]:
                    cell.alignment = Alignment(horizontal="center")
                if col_idx == 5:
                    apply_status_style(cell, acc.get("risk_category", "Low"))
                elif col_idx == 7:
                    apply_status_style(cell, cell.value)
                    
            row_idx += 1
            
    auto_fit_columns(ws4)

    # Save to buffer
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream
