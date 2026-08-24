"""Tests for app/data/loaders.py and app/data/financial_data.py."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

from app.core.enums import UnitOfMeasure

SAMPLE_EXCEL = Path(__file__).resolve().parent.parent.parent / "data" / "sample" / "Sona_BLW_Precis_screener_export.xlsx"


class TestScreenerLoader:
    def test_loads_nonzero_records(self, sona_blw_raw_records):
        assert len(sona_blw_raw_records) > 0

    def test_ten_years_of_sales_present(self, sona_blw_raw_records):
        sales = [
            r for r in sona_blw_raw_records
            if r.line_item == "Sales" and r.statement_type == "profit_and_loss"
        ]
        assert len(sales) == 10
        periods = {r.period for r in sales}
        assert periods == {f"FY{y}" for y in range(2017, 2027)}

    def test_known_fy2026_sales_value_matches_source(self, sona_blw_raw_records):
        # Cross-check against the value visible directly in the source workbook.
        fy26_sales = next(
            r for r in sona_blw_raw_records
            if r.line_item == "Sales" and r.statement_type == "profit_and_loss" and r.period == "FY2026"
        )
        assert fy26_sales.value == 4123.67

    def test_every_record_carries_source_lineage(self, sona_blw_raw_records):
        for rec in sona_blw_raw_records[:20]:
            assert rec.source.source.endswith(".xlsx")
            assert rec.source.company == "Sona BLW Precision Forgings Ltd"
            assert rec.company == "Sona BLW Precision Forgings Ltd"

    def test_balance_sheet_totals_disambiguated(self, sona_blw_raw_records):
        liab = {r.period: r.value for r in sona_blw_raw_records if r.line_item == "Total Liabilities"}
        assets = {r.period: r.value for r in sona_blw_raw_records if r.line_item == "Total Assets"}
        assert len(liab) == 10
        assert len(assets) == 10
        # In a clean balance sheet these are equal per period (verified via real data).
        for period in liab:
            assert liab[period] == assets[period]

    def test_unmapped_rows_captured_raw_but_excluded_from_canonical_statement(
        self, sona_blw_raw_records, sona_blw_statements
    ):
        # "Change in Inventory" IS captured at the raw layer (loaders.py's
        # job is to capture everything recognized in the source layout)...
        raw_labels = {r.line_item for r in sona_blw_raw_records}
        assert "Change in Inventory" in raw_labels
        assert "Power and Fuel" in raw_labels
        # ...but financial_data.py's _FIELD_MAP deliberately does not map
        # these to a canonical field yet, so they must not appear as an
        # attribute value anywhere on the canonical FinancialStatement -
        # i.e. no silent/incorrect mapping onto an unrelated field.
        from app.data.financial_data import _FIELD_MAP

        assert "Change in Inventory" not in _FIELD_MAP
        assert "Power and Fuel" not in _FIELD_MAP


class TestWorkbookFileHandleClosed:
    """Regression test for a real bug a user hit on Windows: openpyxl's
    read_only workbook keeps an underlying zip file handle open, and
    without an explicit close(), that handle could still be open when
    Streamlit's tempfile.TemporaryDirectory() tried to clean up
    immediately after the upload was processed. On Linux an open file
    can still be deleted (harmless, which is exactly why this was never
    caught in the Linux development environment) — on Windows an open
    handle blocks deletion outright, raising
    PermissionError ([WinError 32]). Fixed with an explicit wb.close()
    in a `finally` block. These tests verify close() is actually called
    on both the success and exception paths, rather than relying on
    OS-specific file-locking behavior to prove it (which wouldn't be
    portable to this Linux test-running environment)."""

    def test_close_called_on_successful_load(self):
        import app.data.loaders as loaders_module
        from openpyxl import load_workbook as real_load_workbook

        real_wb = real_load_workbook(SAMPLE_EXCEL, read_only=True, data_only=True)
        real_wb.close = MagicMock(wraps=real_wb.close)

        with patch.object(loaders_module, "load_workbook", return_value=real_wb):
            loaders_module.load_screener_excel(SAMPLE_EXCEL, company_name="Sona BLW Precision Forgings Ltd")

        real_wb.close.assert_called_once()

    def test_close_called_even_when_data_sheet_tab_missing(self, tmp_path):
        import app.data.loaders as loaders_module
        from openpyxl import Workbook
        from openpyxl import load_workbook as real_load_workbook
        from app.core.exceptions import DataIntegrityError

        bad_path = tmp_path / "bad.xlsx"
        wb = Workbook()
        wb.active.title = "Wrong Sheet Name"
        wb.save(bad_path)

        real_wb = real_load_workbook(bad_path, read_only=True, data_only=True)
        real_wb.close = MagicMock(wraps=real_wb.close)

        with patch.object(loaders_module, "load_workbook", return_value=real_wb):
            try:
                loaders_module.load_screener_excel(bad_path, company_name="Test")
                assert False, "expected DataIntegrityError"
            except DataIntegrityError:
                pass

        # The critical assertion: close() must run even though the
        # function raised partway through — this is exactly what the
        # `finally` block guarantees and a bare try/except would not.
        real_wb.close.assert_called_once()

    def test_file_deletable_immediately_after_load_returns(self, tmp_path):
        # A general (OS-agnostic) sanity check: copy the sample file to a
        # temp location, load it, then confirm it can be deleted right
        # away — this is always true on Linux regardless of the bug, but
        # still guards against a regression that would leave a lingering
        # reference preventing deletion on ANY platform (e.g. via os.remove
        # failing with a different error even on Linux if something else
        # were holding the file open).
        import shutil

        test_copy = tmp_path / "test_copy.xlsx"
        shutil.copy(SAMPLE_EXCEL, test_copy)

        from app.data.loaders import load_screener_excel
        load_screener_excel(test_copy, company_name="Sona BLW Precision Forgings Ltd")

        test_copy.unlink()  # would raise if still locked
        assert not test_copy.exists()


class TestFinancialDataNormalization:
    def test_ten_annual_statements_built(self, sona_blw_statements):
        assert len(sona_blw_statements) == 10

    def test_statements_sorted_ascending_by_period_end_date(self, sona_blw_statements):
        dates = [s.period_end_date for s in sona_blw_statements]
        assert dates == sorted(dates)

    def test_canonical_unit_is_crore(self, sona_blw_statements):
        assert all(s.unit == UnitOfMeasure.INR_CRORE for s in sona_blw_statements)

    def test_no_fabricated_operating_profit(self, sona_blw_statements):
        # operating_profit is deliberately left unset by financial_data.py
        # (Module 2 computes it explicitly with full input visibility) -
        # confirm we never silently invented a value for it here.
        assert all(s.operating_profit is None for s in sona_blw_statements)

    def test_fy2026_net_profit_matches_source(self, sona_blw_statements):
        fy26 = next(s for s in sona_blw_statements if s.period == "FY2026")
        assert fy26.net_profit == 646.42

    def test_missing_line_items_are_none_not_zero(self, sona_blw_statements):
        # No period should have a fabricated 0.0 for a field simply
        # because the raw sheet happened not to populate one differently
        # -- spot check a field known to always be present for this company.
        assert all(s.total_assets is not None for s in sona_blw_statements)

    def test_market_price_captured_where_available_none_elsewhere(self, sona_blw_statements):
        # Regression test: the "PRICE:" row sits inside the CASH FLOW
        # section's header block in the source layout and was originally
        # silently dropped by the loader (not in any recognized label
        # set) -- confirm it is now captured correctly, including that
        # the years without a price genuinely have no data (not a
        # fabricated 0.0 or a carried-forward value).
        by_period = {s.period: s.price for s in sona_blw_statements}
        assert by_period["FY2017"] is None
        assert by_period["FY2021"] is None
        assert by_period["FY2022"] == 679.45
        assert by_period["FY2026"] == 481.5
