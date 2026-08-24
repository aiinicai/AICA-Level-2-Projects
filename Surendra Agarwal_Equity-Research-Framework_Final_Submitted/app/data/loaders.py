"""Data loaders — Module 1.

Reads structured financial data from CSV/Excel/JSON into
`FinancialStatementRaw` records, each carrying full lineage metadata.

This module deliberately does NO normalization (unit conversion, ratio
derivation, etc.) — that happens in financial_data.py. Loaders only
parse a source file's layout into a consistent raw representation.

Currently implemented: the Screener.in "Data Sheet" export layout (the
format the person's own workbook uses). Additional loaders (generic CSV,
manual-entry JSON) are stubbed for Milestone 1 and extended as needed.
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from app.core.enums import Currency, DataSourceType, UnitOfMeasure
from app.core.exceptions import DataIntegrityError
from app.core.models import FinancialStatementRaw, SourceMetadata

logger = logging.getLogger(__name__)


# Row labels (as they appear in the Screener "Data Sheet" tab) mapped to
# the statement section they belong to. Kept as an explicit table rather
# than positional row indices so a re-export with a slightly different
# row order doesn't silently misparse — we look up by label, not offset.
_PNL_LABELS = {
    "Sales",
    "Raw Material Cost",
    "Change in Inventory",
    "Power and Fuel",
    "Other Mfr. Exp",
    "Employee Cost",
    "Selling and admin",
    "Other Expenses",
    "Other Income",
    "Depreciation",
    "Interest",
    "Profit before tax",
    "Tax",
    "Net profit",
    "Dividend Amount",
}

_BALANCE_SHEET_LABELS = {
    "Equity Share Capital",
    "Reserves",
    "Borrowings",
    "Other Liabilities",
    "Total",  # appears twice (liabilities total, assets total) — disambiguated by position
    "Net Block",
    "Capital Work in Progress",
    "Investments",
    "Other Assets",
    "Receivables",
    "Inventory",
    "Cash & Bank",
    "No. of Equity Shares",
    "Face value",
}

_CASH_FLOW_LABELS = {
    "Cash from Operating Activity",
    "Cash from Investing Activity",
    "Cash from Financing Activity",
    "Net Cash Flow",
}

# The PRICE row sits after the CASH FLOW section in the source layout,
# reusing that section's own "Report Date" header, but is conceptually a
# distinct market-data row, not a cash-flow line item — handled as a
# special case in the parse loop below rather than folded into
# _CASH_FLOW_LABELS (which would mislabel its statement_type).
_PRICE_LABEL = "PRICE:"


def _fy_label(period_end: date) -> str:
    """Convert a period-end date to an Indian fiscal-year label.

    Indian FY runs April-March. A period ending 2026-03-31 is FY2026
    (i.e. the year the FY *ends* in), consistent with how the source
    Screener export and Indian corporate reporting label fiscal years.
    """
    if period_end.month <= 3:
        return f"FY{period_end.year}"
    return f"FY{period_end.year + 1}"


def load_screener_excel(
    file_path: Path,
    company_name: str,
    *,
    assumed_unit: UnitOfMeasure = UnitOfMeasure.INR_CRORE,
) -> list[FinancialStatementRaw]:
    """Load a Screener.in-format Excel export's "Data Sheet" tab.

    Args:
        file_path: Path to the .xlsx file.
        company_name: Canonical company name to stamp on every record
            (the sheet's own "COMPANY NAME" cell is cross-checked
            against this, not blindly trusted, and a mismatch is logged).
        assumed_unit: Screener.in exports for Indian mid/large-cap
            companies are conventionally in INR crore, but the workbook
            itself does not state this explicitly anywhere in the sheet.
            This is therefore an explicit, labeled ASSUMPTION (Principle
            4: no silent assumptions) — every resulting SourceMetadata
            record carries this unit so it is visible and overridable
            downstream rather than silently baked in.

    Returns:
        List of FinancialStatementRaw, one per (line_item, period).

    Raises:
        DataIntegrityError: if the "Data Sheet" tab is missing or the
            file does not match the expected Screener layout closely
            enough to parse safely.
    """
    if not file_path.exists():
        raise DataIntegrityError(f"File not found: {file_path}")

    logger.info("Loading Screener Excel export: %s", file_path)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if "Data Sheet" not in wb.sheetnames:
            raise DataIntegrityError(
                f"Expected a 'Data Sheet' tab in {file_path.name}; found sheets: "
                f"{wb.sheetnames}. This loader only supports the Screener.in export "
                "layout — use a generic loader for other formats."
            )

        ws = wb["Data Sheet"]
        rows = list(ws.iter_rows(values_only=True))

        sheet_company = rows[0][1] if len(rows) > 0 and len(rows[0]) > 1 else None
        if sheet_company and company_name.strip().lower() not in str(sheet_company).strip().lower():
            logger.warning(
                "Company name mismatch: caller passed %r but sheet says %r. "
                "Proceeding, but verify this is the intended file.",
                company_name,
                sheet_company,
            )

        records: list[FinancialStatementRaw] = []
        statement_type: str | None = None
        header_periods: list[date | None] = []
        seen_totals_in_section = 0  # disambiguates the two "Total" rows in Balance Sheet

        for row in rows:
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip()

            # Section headers
            if label == "PROFIT & LOSS":
                statement_type = "profit_and_loss"
                seen_totals_in_section = 0
                continue
            if label == "Quarters":
                statement_type = "quarterly"
                seen_totals_in_section = 0
                continue
            if label == "BALANCE SHEET":
                statement_type = "balance_sheet"
                seen_totals_in_section = 0
                continue
            if label == "CASH FLOW:":
                statement_type = "cash_flow"
                seen_totals_in_section = 0
                continue

            if label == "Report Date":
                header_periods = []
                for cell in row[1:]:
                    if isinstance(cell, datetime):
                        header_periods.append(cell.date())
                    elif isinstance(cell, date):
                        header_periods.append(cell)
                    else:
                        header_periods.append(None)
                continue

            if statement_type is None:
                continue  # not yet inside a recognized section (META block, etc.)

            applicable_labels = {
                "profit_and_loss": _PNL_LABELS,
                "quarterly": _PNL_LABELS,  # quarterly reuses P&L-style labels
                "balance_sheet": _BALANCE_SHEET_LABELS,
                "cash_flow": _CASH_FLOW_LABELS,
            }.get(statement_type, set())

            if label == "Total":
                seen_totals_in_section += 1
                resolved_label = (
                    "Total Liabilities" if seen_totals_in_section == 1 else "Total Assets"
                )
                row_statement_type = statement_type
            elif label == _PRICE_LABEL:
                # Market price row: appears within the CASH FLOW section's
                # header block but is not itself a cash-flow line item -
                # tagged with its own statement_type so downstream code
                # (financial_data.py) doesn't try to unit-convert it as a
                # monetary crore-scale figure.
                resolved_label = "Price"
                row_statement_type = "market_price"
            elif label in applicable_labels:
                resolved_label = label
                row_statement_type = statement_type
            else:
                continue  # unrecognized row (ratios block, TRENDS block, etc.) — skip, don't guess

            if not header_periods:
                logger.warning("Row %r encountered before a 'Report Date' header; skipping.", label)
                continue

            for period_end, value in zip(header_periods, row[1:]):
                if period_end is None:
                    continue
                fy = _fy_label(period_end)
                period_label = (
                    fy if row_statement_type != "quarterly" else f"{fy}-{period_end.strftime('%b')}"
                )
                records.append(
                    FinancialStatementRaw(
                        company=company_name,
                        line_item=resolved_label,
                        statement_type=row_statement_type,
                        period=period_label,
                        period_end_date=period_end,
                        value=float(value) if isinstance(value, (int, float)) else None,
                        source=SourceMetadata(
                            company=company_name,
                            source=file_path.name,
                            source_type=DataSourceType.EXCEL_UPLOAD,
                            source_date=period_end,
                            reporting_period=period_label,
                            currency=Currency.INR,
                            unit=assumed_unit,
                            confidence="medium",  # unit is an assumption; see docstring
                        ),
                    )
                )

        logger.info("Parsed %d raw financial line items from %s", len(records), file_path.name)
        return records
    finally:
        # CRITICAL for Windows: openpyxl's read_only mode keeps an
        # underlying zip file handle open, and relying on garbage
        # collection to release it is unreliable — on Linux an open
        # file can still be deleted (harmless), but on Windows an open
        # handle BLOCKS deletion outright. This caused a real
        # PermissionError ([WinError 32]) when Streamlit's
        # tempfile.TemporaryDirectory() tried to clean up immediately
        # after this function returned, on a real Windows machine — a
        # class of bug that could never surface in a Linux dev
        # environment. Explicit close() in `finally` guarantees the
        # handle is released before this function returns, regardless
        # of whether it succeeded or raised.
        wb.close()


# --------------------------------------------------------------------------
# NSE Shareholding Pattern CSV — Promoter Holding history
# --------------------------------------------------------------------------
#
# This is a genuinely distinct data source from the Screener export above:
# NSE's own "Shareholding Pattern" corporate-filing export (available from
# NSE's website per-symbol), containing a quarterly time series of
# Promoter & Promoter Group holding percentage. It does NOT contain
# promoter pledge data — that lives in a separate NSE filing (pledge/
# encumbrance disclosures) not covered by this loader.


def load_nse_shareholding_pattern_csv(csv_path: Path) -> list[dict]:
    """Load NSE's Shareholding Pattern CSV export (columns include
    COMPANY, 'PROMOTER & PROMOTER GROUP (A)', 'PUBLIC (B)', 'AS ON DATE',
    STATUS, among others).

    Returns a list of dicts, one per filing, sorted ascending by date:
        {"as_on_date": date, "promoter_pct": float (0-1 scale),
         "public_pct": float, "status": str}

    Promoter/public percentages are converted from the source's 0-100
    scale to this project's 0-1 convention (matching
    FinancialStatement.promoter_holding_pct's documented scale).
    """
    import pandas as pd

    if not csv_path.exists():
        raise DataIntegrityError(f"File not found: {csv_path}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    df.columns = [c.strip() for c in df.columns]

    required = {"PROMOTER & PROMOTER GROUP (A)", "PUBLIC (B)", "AS ON DATE"}
    missing = required - set(df.columns)
    if missing:
        raise DataIntegrityError(
            f"{csv_path.name} is missing expected column(s): {sorted(missing)}. "
            f"Found columns: {list(df.columns)}. This loader expects the standard "
            "NSE Shareholding Pattern CSV export format."
        )

    records: list[dict] = []
    for _, row in df.iterrows():
        try:
            as_on_date = datetime.strptime(str(row["AS ON DATE"]).strip(), "%d-%b-%Y").date()
        except ValueError:
            logger.warning("Skipping row with unparseable AS ON DATE: %r", row["AS ON DATE"])
            continue
        try:
            promoter_pct = float(row["PROMOTER & PROMOTER GROUP (A)"]) / 100.0
            public_pct = float(row["PUBLIC (B)"]) / 100.0
        except (ValueError, TypeError):
            logger.warning("Skipping row with unparseable holding percentage on %s", as_on_date)
            continue
        records.append({
            "as_on_date": as_on_date, "promoter_pct": promoter_pct,
            "public_pct": public_pct, "status": str(row.get("STATUS", "")).strip(),
        })

    records.sort(key=lambda r: r["as_on_date"])
    logger.info("Parsed %d shareholding-pattern record(s) from %s", len(records), csv_path.name)
    return records


def apply_shareholding_history_to_statements(
    statements: list, shareholding_records: list[dict],
) -> tuple[list, int]:
    """Apply promoter-holding history to a FinancialStatement series by
    exact date match: a shareholding record's as_on_date is applied to
    any statement whose period_end_date equals it exactly.

    This is deliberately an EXACT match only, not a nearest-date
    approximation — the quarterly shareholding filings frequently do NOT
    land on a company's fiscal year-end, and silently applying the
    nearest available quarter's figure to an annual statement would
    misrepresent a point-in-time filing as if it were that period's
    actual year-end holding. Periods with no exact match keep
    promoter_holding_pct=None (still unavailable), same as if this
    function were never called.

    Pledge is never set here — this data source doesn't contain it.

    Returns (updated_statements, match_count) so callers can report how
    many periods were actually matched.
    """
    by_date = {r["as_on_date"]: r for r in shareholding_records}
    updated = []
    match_count = 0
    for stmt in statements:
        if stmt.period_end_date is not None and stmt.period_end_date in by_date:
            record = by_date[stmt.period_end_date]
            updated.append(stmt.model_copy(update={"promoter_holding_pct": record["promoter_pct"]}))
            match_count += 1
        else:
            updated.append(stmt)
    return updated, match_count
