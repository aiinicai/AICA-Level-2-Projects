"""
reports/excel_report.py
Generates the Excel (.xlsx) property analysis report using openpyxl.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import DISCLAIMER

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, size=12, color="1F4E78")


def _write_section(ws, row, title):
    ws.cell(row=row, column=1, value=title).font = SECTION_FONT
    return row + 1


def _write_kv(ws, row, pairs):
    for k, v in pairs:
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v)
        row += 1
    return row + 1


def generate_excel_report(path, property_input, city_name, locality_name, valuation_output):
    result = valuation_output["result"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30

    ws.merge_cells("A1:B1")
    ws["A1"] = "India Property Valuation Report"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E78")

    row = 3
    row = _write_section(ws, row, "1. Property Details")
    row = _write_kv(ws, row, [
        ("Location", f"{locality_name}, {city_name}"),
        ("Property Type", f"{property_input.bhk} BHK {property_input.property_type}"),
        ("Carpet Area (sqft)", property_input.carpet_area),
        ("Built-up Area (sqft)", property_input.builtup_area),
        ("Furnishing", property_input.furnishing),
        ("Age (years)", property_input.age_years),
        ("Asking Price (INR)", property_input.asking_price),
        ("Expected Monthly Rent (INR)", property_input.expected_rent),
    ])

    row = _write_section(ws, row, "2. Estimated Market Rent")
    row = _write_kv(ws, row, [
        ("Low (P25)", result.market_rent_low),
        ("Median", result.market_rent_median),
        ("High (P75)", result.market_rent_high),
        ("Comparable rental observations", valuation_output["rent_stats"]["count"]),
    ])

    row = _write_section(ws, row, "3. Estimated Fair Value")
    row = _write_kv(ws, row, [
        ("Comparable Market Value (INR)", result.comparable_value),
        ("Rental Capitalization Value (INR)", result.rental_cap_value),
        ("Adjusted Value (INR)", result.adjusted_value),
        ("Fair Value Low (INR)", result.fair_value_low),
        ("Fair Value High (INR)", result.fair_value_high),
    ])

    row = _write_section(ws, row, "4. Yield & Ratios")
    row = _write_kv(ws, row, [
        ("Gross Rental Yield (%)", result.gross_yield),
        ("Net Rental Yield (%)", result.net_yield),
        ("Price-to-Rent Ratio (years)", result.price_to_rent),
    ])

    row = _write_section(ws, row, "5. Verdict")
    row = _write_kv(ws, row, [
        ("Valuation Verdict", result.verdict),
        ("Premium/Discount vs Fair Value (%)", result.premium_pct),
        ("Investment Score", f"{result.investment_score} ({result.investment_score_label})"),
        ("Confidence (%)", result.confidence_pct),
    ])

    row = _write_section(ws, row, "6. Methodology")
    ws.cell(row=row, column=1, value=result.methodology_notes)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    row += 2

    row = _write_section(ws, row, "7. Disclaimer")
    ws.cell(row=row, column=1, value=DISCLAIMER)
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    # --- Comparable listings raw data sheet ---
    ws2 = wb.create_sheet("Comparable Data")
    headers = ["Source", "Kind", "BHK", "Area(sqft)", "Price", "Price/sqft", "Rent", "Rent/sqft", "Collected Date", "Outlier?"]
    for i, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    return wb.save(path) or path
