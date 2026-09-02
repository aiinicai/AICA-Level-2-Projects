"""
================================================================================
 TOP 200 INDIAN COMPANIES - STOCK ANALYSIS TOOL
================================================================================
What this script does:
  1. Fetches the Top N (default 200) Indian companies by Market Cap from
     companiesmarketcap.com. That site shows only 100 rows per page, so the
     script automatically pages through it (page=1, page=2, ...) -- this is
     the same as clicking "Next 100" on the website.
  2. For each company, fetches Market Cap, CMP, 52-Week High, 52-Week Low,
     ROE and ROCE -- all from screener.in's public company page in a SINGLE
     page fetch per company (screener.in shows all of these together in its
     "top ratios" box).
  3. Fetches daily closing-price history directly from moneycontrol.com's own
     price-chart API (the same data that powers the charts on their website)
     and computes 50-Day Moving Average (50 DMA) and 200-Day Moving Average
     (200 DMA) from it. This same price history is also used as a fallback
     for CMP / 52-Week High / 52-Week Low if screener.in doesn't have data
     for a particular company.
  4. Calculates Deviation % from 52-Week High for each stock
  5. Flags/highlights companies where:
        50 DMA < 200 DMA   AND   CMP < 50 DMA
     (a classic bearish / "watch out" technical signal)
  6. Saves everything to a formatted Excel file, with flagged rows highlighted

  NO YAHOO FINANCE IS USED ANYWHERE IN THIS SCRIPT. Every data point comes
  from either screener.in or moneycontrol.com, as requested.

IMPORTANT NOTE ON DATA SOURCES (please read):
  - Company list (names/rank)    : companiesmarketcap.com
  - Market Cap (Rs Cr), CMP,
    52 Week High, 52 Week Low,
    ROE, ROCE                    : screener.in "top ratios" box (one page
      fetch per company). screener.in shows these directly as e.g.:
        Market Cap Rs 17,72,762 Cr.   Current Price Rs 1,310
        High / Low Rs 1,612 / 1,250   ROCE 7.78 %   ROE 7.71 %
      The "High / Low" field on screener.in IS the 52-week high/low range.
  - 50 DMA / 200 DMA             : moneycontrol.com's own price-history API
      (priceapi.moneycontrol.com/techCharts/indianMarket/stock/history) --
      this is the same JSON data feed moneycontrol's own website charts use
      internally, fetched directly with the NSE symbol (e.g. "RELIANCE").
      It returns daily closing prices, from which 50/200-day simple moving
      averages are calculated here in the script.
  - Fallback only: if screener.in has no data for a company, CMP / 52W High
    / 52W Low are derived from that SAME moneycontrol.com price history
    instead (never from Yahoo Finance or any other source). The 'CMP Source'
    and '52W Source' columns in the output tell you exactly which of the two
    sites was used for each row.

  HONESTY NOTE: moneycontrol.com does not publish or document this
  price-history endpoint publicly -- it's the internal API their own website
  charts call, identified by inspecting network requests. It has worked
  reliably in community tooling, but because it isn't a documented public
  API, moneycontrol could change its format at any time. If 50 DMA / 200 DMA
  start showing "N/A" for most/all rows in a future run, that's the signal
  this endpoint's format has changed and the script needs a small update
  (the `fetch_moneycontrol_price_history()` function is the only place that
  would need touching).

HOW TO RUN (from IDLE):
  1. Install the required libraries first -- open Command Prompt (cmd) and run:
     pip install requests beautifulsoup4 pandas openpyxl
  2. Open this file in IDLE and press F5 (Run Module)
  3. Wait for it to finish. With TOP_N = 200 and polite delays between
     requests, a full run takes roughly 15-20 minutes -- this is expected,
     not a bug. Progress is printed to the console as it works through
     each company.
  4. An Excel file "Top200_Indian_Companies_Analysis_<date>.xlsx" will be
     created in the same folder as this script

DISCLAIMER: This is for informational/educational purposes only, not
investment advice. Web page structures / internal APIs change over time --
if a source site updates its layout or API, some fields may show "N/A"
until the script is updated.
================================================================================
"""

import re
import json
import time
import datetime
import traceback

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------------
# CONFIGURATION -- feel free to tweak these
# ------------------------------------------------------------------------
TOP_N = 200                    # how many companies to analyze
CMC_URL = "https://companiesmarketcap.com/india/largest-companies-in-india-by-market-cap/"
SCREENER_CONSOLIDATED = "https://www.screener.in/company/{code}/consolidated/"
SCREENER_STANDALONE = "https://www.screener.in/company/{code}/"
MC_PRICE_API = "https://priceapi.moneycontrol.com/techCharts/indianMarket/stock/history"
MC_SEARCH_URL = "https://www.moneycontrol.com/mccode/common/autosuggestion_solr.php/"
PRICE_HISTORY_DAYS_BACK = 450   # calendar days of history to request (~ >250 trading days)
OUTPUT_FILE = f"Top{TOP_N}_Indian_Companies_Analysis_{datetime.date.today().isoformat()}.xlsx"

REQUEST_TIMEOUT = 15           # seconds to wait for a web response
SLEEP_BETWEEN_REQUESTS = 1.5   # seconds -- politeness delay between companies
SLEEP_BETWEEN_PAGES = 1.5      # seconds -- politeness delay between listing pages
SLEEP_MICRO = 0.5              # seconds -- tiny delay between sub-requests for one company
MAX_PAGES_SAFETY = 10          # safety guard against an infinite paging loop

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.moneycontrol.com/",
}

# companiesmarketcap.com sometimes lists the US-listed ADR ticker instead of
# the NSE symbol for large dual-listed Indian companies. This map corrects
# those to their proper NSE code (used on screener.in / moneycontrol.com).
# Add more entries here if the script reports "Not Found" for a company you
# know is listed on NSE.
ADR_TO_NSE = {
    "HDB": "HDFCBANK",
    "IBN": "ICICIBANK",
    "WIT": "WIPRO",
    "RDY": "DRREDDY",
    "TTM": "TATAMOTORS",
    "SIFY": "SIFY",
    "INFY": "INFY",
}


# ------------------------------------------------------------------------
# Small shared helper
# ------------------------------------------------------------------------
def _to_float(text):
    """Strip currency symbols/commas/units and convert to a rounded float."""
    if text is None:
        return None
    num = re.sub(r"[^\d.\-]", "", str(text))
    try:
        return round(float(num), 2)
    except ValueError:
        return None


# ------------------------------------------------------------------------
# STEP 1: Get Top N Indian companies by market cap (auto-pages through
#         "Next 100" style pagination on companiesmarketcap.com)
# ------------------------------------------------------------------------
def get_top_companies(n=TOP_N):
    print(f"[1/3] Fetching Top {n} Indian companies by market cap from "
          f"companiesmarketcap.com (auto-paging through the listing, i.e. "
          f"clicking 'Next 100' as needed) ...")

    session = requests.Session()
    companies = []
    seen_symbols = set()
    page = 1

    while len(companies) < n and page <= MAX_PAGES_SAFETY:
        url = CMC_URL if page == 1 else f"{CMC_URL}?page={page}"
        try:
            resp = session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
        except requests.RequestException as exc:
            print(f"  WARNING: Could not fetch page {page} ({exc}). Stopping pagination.")
            break

        soup = BeautifulSoup(resp.text, "html.parser")
        found_this_page = 0

        for a in soup.find_all("a", href=True):
            if "/marketcap/" not in a["href"]:
                continue
            text = a.get_text(" ", strip=True)
            if not text or len(text.split()) < 2:
                continue

            tr = a.find_parent("tr")
            if tr is None:
                continue
            tds = [td.get_text(" ", strip=True) for td in tr.find_all("td")]

            rank = next((int(t) for t in tds if t.isdigit()), None)
            if rank is None:
                continue

            words = text.split()
            symbol_raw = words[-1]
            name = " ".join(words[:-1])

            if symbol_raw in seen_symbols:
                continue
            seen_symbols.add(symbol_raw)

            companies.append({
                "Rank": rank,
                "Company": name,
                "Symbol": symbol_raw,
            })
            found_this_page += 1

        print(f"  -> Page {page}: found {found_this_page} companies "
              f"(running total: {len(companies)})")

        if found_this_page == 0:
            break  # reached the end of the listing, or page layout changed
        page += 1
        time.sleep(SLEEP_BETWEEN_PAGES)

    companies = sorted(companies, key=lambda x: x["Rank"])[:n]
    print(f"  -> Total companies collected: {len(companies)}")
    return companies


def symbol_to_nse_code(symbol):
    """Convert a companiesmarketcap.com symbol into an NSE trading code."""
    if symbol.endswith(".NS") or symbol.endswith(".BO"):
        return symbol.split(".")[0]
    return ADR_TO_NSE.get(symbol, symbol)


# ------------------------------------------------------------------------
# STEP 2: Get Market Cap / CMP / 52W High-Low / ROE / ROCE, all from
#         screener.in's "top ratios" box (one page fetch per company)
# ------------------------------------------------------------------------
def fetch_screener_ratios(nse_code, session):
    """Return a dict of {ratio_name: value_text} scraped from screener.in."""
    urls = [
        SCREENER_CONSOLIDATED.format(code=nse_code),
        SCREENER_STANDALONE.format(code=nse_code),
    ]
    for url in urls:
        try:
            resp = session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            if resp.status_code != 200:
                continue
            soup = BeautifulSoup(resp.text, "html.parser")

            ratios_box = soup.find(id="top-ratios")
            if ratios_box is None:
                ratios_box = soup  # fallback: search whole page

            data = {}
            for li in ratios_box.find_all("li"):
                name_tag = li.find(class_="name")
                value_tag = li.find(class_="value") or li.find(class_="number")
                if name_tag and value_tag:
                    data[name_tag.get_text(strip=True)] = value_tag.get_text(" ", strip=True)

            if data:
                return data
        except requests.RequestException:
            continue
    return {}


def extract_ratio(data, keywords):
    """Find a value in the ratios dict whose key contains all keywords,
    and return it as a float (or None if not found / not numeric)."""
    for k, v in data.items():
        kl = k.lower()
        if all(kw in kl for kw in keywords):
            return _to_float(v)
    return None


FINANCIAL_KEYWORDS = (
    "bank", "finance", "financial", "insurance", "housing finance",
    "nbfc", "life insurance", "card", "asset management", "amc",
)


def is_probably_financial_company(company_name):
    name = company_name.lower()
    return any(kw in name for kw in FINANCIAL_KEYWORDS)


def fetch_ratio_table_fallback(nse_code, session, ratio_keywords):
    """screener.in's top-ratios box sometimes omits ROE/ROCE (common for
    finance/insurance companies, where ROCE isn't a meaningful metric, or
    when the label text differs slightly). This checks the full 'Ratios'
    table further down the company page (rows like 'ROE %', 'ROCE %') and
    returns the most recent (rightmost / TTM) numeric value found."""
    for url in (SCREENER_CONSOLIDATED.format(code=nse_code),
                SCREENER_STANDALONE.format(code=nse_code)):
        try:
            resp = session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            if resp.status_code != 200:
                continue
            soup = BeautifulSoup(resp.text, "html.parser")
            for table in soup.find_all("table"):
                first_row = table.find("tr")
                if not first_row:
                    continue
                header_text = first_row.get_text(" ", strip=True).lower()
                if not all(kw in header_text for kw in ratio_keywords):
                    continue
                cells = first_row.find_all(["td", "th"])
                values = [_to_float(c.get_text(strip=True)) for c in cells[1:]]
                values = [v for v in values if v is not None]
                if values:
                    return values[-1]  # most recent period, i.e. rightmost column
        except requests.RequestException:
            continue
    return None


def extract_52w_high_low(data):
    """screener.in shows the 52-week range as a single 'High / Low' ratio,
    e.g. 'High / Low' -> '1,612 / 1,250'. Split it into (high, low)."""
    for k, v in data.items():
        kl = k.lower()
        if "high" in kl and "low" in kl:
            parts = str(v).split("/")
            if len(parts) == 2:
                return _to_float(parts[0]), _to_float(parts[1])
    return None, None


# ------------------------------------------------------------------------
# STEP 3: Get daily closing-price history from moneycontrol.com's own
#         price-chart API -- used for 50/200 DMA, and as a fallback for
#         CMP / 52W High-Low if screener.in has no data for a company.
# ------------------------------------------------------------------------
def fetch_moneycontrol_price_history(nse_code, session, days_back=PRICE_HISTORY_DAYS_BACK):
    """Fetch daily closing prices for an NSE-listed stock directly from
    moneycontrol.com's own price-charting API (the same feed that powers
    the price charts on moneycontrol.com stock pages). Returns a pandas
    Series of closing prices indexed by date, or None if unavailable."""
    end = datetime.datetime.now()
    start = end - datetime.timedelta(days=days_back)
    params = {
        "symbol": nse_code,
        "resolution": "1D",
        "from": int(start.timestamp()),
        "to": int(end.timestamp()),
        "countback": str(days_back),
        "currencyCode": "INR",
    }
    for resolution in ("1D", "D"):
        params["resolution"] = resolution
        try:
            resp = session.get(MC_PRICE_API, params=params, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            if resp.status_code != 200:
                continue
            data = resp.json()
            if data.get("s") != "ok":
                continue
            closes = data.get("c")
            timestamps = data.get("t")
            if not closes or not timestamps:
                continue
            series = pd.Series(closes, index=pd.to_datetime(timestamps, unit="s")).dropna()
            if not series.empty:
                return series
        except (requests.RequestException, ValueError):
            continue
    return None


def clean_company_name_for_search(name):
    """Strip parenthetical suffixes etc. so moneycontrol's search matches
    better, e.g. 'Bajaj Housing Finance Limited (BHFL)' -> 'Bajaj Housing
    Finance Limited'."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", name).strip()


def moneycontrol_find_page_url(company_name, session):
    """Use moneycontrol.com's own search-suggestion endpoint to find the
    stock quote page URL for a given company name."""
    params = {
        "classic": "true",
        "query": company_name,
        "type": "1",
        "format": "json",
        "callback": "suggest1",
    }
    try:
        resp = session.get(MC_SEARCH_URL, params=params, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        text = resp.text.strip()
        start = text.find("(")
        end = text.rfind(")")
        if start == -1 or end == -1 or end <= start:
            return None
        data = json.loads(text[start + 1:end])
        if not data:
            return None

        best = None
        for item in data:
            name = str(item.get("stock_name", ""))
            if name.lower() == company_name.lower():
                best = item
                break
        if best is None:
            best = data[0]

        url = None
        for key in ("link_src", "url", "mc_url", "URL"):
            if best.get(key):
                url = best[key]
                break
        if url is None:
            for k, v in best.items():
                if isinstance(v, str) and v.startswith(("/india/stockpricequote", "http")):
                    url = v
                    break
        if not url:
            return None
        if not url.startswith("http"):
            url = "https://www.moneycontrol.com" + url
        return url
    except Exception:
        return None


def fetch_moneycontrol_market_cap(company_name, session):
    """Scrape the Market Cap figure off a company's moneycontrol.com stock
    quote page. Used ONLY for companies whose CMP was also sourced from
    moneycontrol.com (i.e. screener.in had no usable data for them), so
    Market Cap stays consistent with where the rest of that row came from."""
    page_url = moneycontrol_find_page_url(company_name, session)
    if not page_url:
        return None
    time.sleep(SLEEP_MICRO)
    try:
        resp = session.get(page_url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            return None
        soup = BeautifulSoup(resp.text, "html.parser")
        page_text = soup.get_text(" ", strip=True)
        m = re.search(r"Market\s*Cap[^\d]{0,15}([\d,]+\.?\d*)", page_text, re.I)
        if m:
            return _to_float(m.group(1))
    except requests.RequestException:
        return None
    return None


def moving_averages_from_closes(closes):
    dma50 = round(float(closes.tail(50).mean()), 2) if len(closes) >= 50 else None
    dma200 = round(float(closes.tail(200).mean()), 2) if len(closes) >= 200 else None
    return dma50, dma200


def high_low_from_closes(closes):
    """52-week high/low derived from the last ~365 calendar days of closes."""
    cutoff = closes.index.max() - pd.Timedelta(days=365)
    last_year = closes[closes.index >= cutoff]
    if last_year.empty:
        return None, None
    return round(float(last_year.max()), 2), round(float(last_year.min()), 2)


# ------------------------------------------------------------------------
# STEP 4: Build the Excel report
# ------------------------------------------------------------------------
def write_excel(df, highlight_flags, filename):
    export_df = df.fillna("N/A")
    export_df.to_excel(filename, index=False, sheet_name=f"Top{TOP_N} Analysis")

    wb = load_workbook(filename)
    ws = wb[f"Top{TOP_N} Analysis"]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    highlight_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    highlight_font = Font(color="9C0006", bold=True)
    thin_side = Side(style="thin", color="D9D9D9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(2, ws.max_row + 1):
        is_flagged = highlight_flags[row_idx - 2]
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if is_flagged:
                cell.fill = highlight_fill
                cell.font = highlight_font

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    ws.freeze_panes = "A2"

    # Add a short notes sheet
    notes = wb.create_sheet("Notes")
    notes_text = [
        "HIGHLIGHTED (red) rows = 50 Day Moving Average < 200 Day Moving Average",
        "                          AND Current Market Price < 50 Day Moving Average",
        "",
        "Deviation % from 52W High = (52W High - CMP) / 52W High x 100",
        "  (a positive number means CMP is that much % below its 52-week high)",
        "",
        "Deviation % from 52W Low = (CMP - 52W Low) / 52W Low x 100",
        "  (a positive number means CMP is that much % above its 52-week low)",
        "",
        "ROE/ROCE: if screener.in's top-ratios box doesn't have a value, the",
        "script also checks screener.in's full Ratios table further down the",
        "page before giving up. For finance/insurance companies (banks, NBFCs,",
        "insurers, housing finance, card companies) ROCE is not a meaningful",
        "metric and screener.in does not publish it -- these show 'Not",
        "applicable (financial company)' rather than a blank/failed lookup.",
        "",
        "Data sources (NO Yahoo Finance is used anywhere in this report):",
        " - Company list (names/rank)                    : companiesmarketcap.com",
        " - Market Cap (Rs Cr) / CMP / 52W High / 52W Low / ROE / ROCE :",
        "     screener.in 'top ratios' box (one page fetch per company).",
        "     screener.in's 'High / Low' ratio IS the 52-week high/low range.",
        " - 50 DMA / 200 DMA                              : moneycontrol.com's",
        "     own daily price-history feed (priceapi.moneycontrol.com), the",
        "     same data used by moneycontrol's own website charts.",
        "",
        " Fallback: if screener.in has no data for a company, CMP / 52W High /",
        " 52W Low are instead derived from that SAME moneycontrol.com price",
        " history. The 'CMP Source' / '52W Source' columns show which of the",
        " two sites (screener.in or moneycontrol.com) was actually used.",
        "",
        " Market Cap follows CMP's source: whenever CMP for a company came",
        " from moneycontrol.com (i.e. screener.in had no usable data for it),",
        " Market Cap for that same company is ALSO pulled from moneycontrol.com",
        " (never mixed with screener.in's figure for that row). See the",
        " 'Market Cap Source' column.",
        "",
        f"Companies analyzed: Top {TOP_N} by market cap",
        f"Report generated on: {datetime.datetime.now().strftime('%d-%b-%Y %H:%M:%S')}",
        "",
        "Disclaimer: For informational purposes only. Not investment advice.",
    ]
    for i, line in enumerate(notes_text, start=1):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions["A"].width = 95

    wb.save(filename)


# ------------------------------------------------------------------------
# MAIN
# ------------------------------------------------------------------------
def main():
    companies = get_top_companies(TOP_N)
    if not companies:
        print("\nCould not fetch the company list. Please check your internet "
              "connection, or companiesmarketcap.com may have changed its "
              "page layout.")
        return

    session = requests.Session()
    results = []
    highlight_flags = []

    print(f"\n[2-3/3] Fetching Market Cap / CMP / 52W High-Low / ROE / ROCE "
          f"(screener.in) and daily price history for 50DMA/200DMA "
          f"(moneycontrol.com) for each of {len(companies)} companies ...")

    for i, c in enumerate(companies, 1):
        nse_code = symbol_to_nse_code(c["Symbol"])
        print(f"  [{i}/{len(companies)}] {c['Company']} ({nse_code}) ...", end=" ")

        try:
            ratios = fetch_screener_ratios(nse_code, session)
            roe = extract_ratio(ratios, ["roe"])
            roce = extract_ratio(ratios, ["roce"])
            if roe is None:
                roe = fetch_ratio_table_fallback(nse_code, session, ["roe"])
            if roce is None:
                roce = fetch_ratio_table_fallback(nse_code, session, ["roce"])
            roce_display = roce
            if roce is None and is_probably_financial_company(c["Company"]):
                roce_display = "Not applicable (financial company)"
            mcap_screener = extract_ratio(ratios, ["market", "cap"])
            cmp_screener = extract_ratio(ratios, ["current", "price"])
            high_screener, low_screener = extract_52w_high_low(ratios)
            time.sleep(SLEEP_MICRO)

            closes = fetch_moneycontrol_price_history(nse_code, session)
            dma50 = dma200 = None
            cmp_mc = high_mc = low_mc = None
            if closes is not None:
                dma50, dma200 = moving_averages_from_closes(closes)
                cmp_mc = round(float(closes.iloc[-1]), 2)
                high_mc, low_mc = high_low_from_closes(closes)

            cmp_final = cmp_screener if cmp_screener is not None else cmp_mc
            cmp_source = "screener.in" if cmp_screener is not None else (
                "moneycontrol.com (fallback)" if cmp_mc is not None else "N/A")

            # If CMP for this company came from moneycontrol.com (i.e.
            # screener.in had no usable data for it), pull Market Cap from
            # moneycontrol.com too, instead of screener.in's value.
            if cmp_source.startswith("moneycontrol.com"):
                mcap_mc = fetch_moneycontrol_market_cap(
                    clean_company_name_for_search(c["Company"]), session)
                mcap_final = mcap_mc
                mcap_source = "moneycontrol.com" if mcap_mc is not None else "N/A"
            else:
                mcap_final = mcap_screener
                mcap_source = "screener.in" if mcap_screener is not None else "N/A"

            if high_screener is not None and low_screener is not None:
                week52_high, week52_low, source_52w = high_screener, low_screener, "screener.in"
            elif high_mc is not None and low_mc is not None:
                week52_high, week52_low, source_52w = high_mc, low_mc, "moneycontrol.com (fallback)"
            else:
                week52_high, week52_low, source_52w = None, None, "N/A"

            deviation_from_high_pct = None
            if week52_high not in (None, 0) and cmp_final is not None:
                deviation_from_high_pct = round((week52_high - cmp_final) / week52_high * 100, 2)

            deviation_from_low_pct = None
            if week52_low not in (None, 0) and cmp_final is not None:
                deviation_from_low_pct = round((cmp_final - week52_low) / week52_low * 100, 2)

            flagged = False
            if dma50 is not None and dma200 is not None and cmp_final is not None:
                if dma50 < dma200 and cmp_final < dma50:
                    flagged = True

            results.append({
                "Rank": c["Rank"],
                "Company": c["Company"],
                "NSE Symbol": nse_code,
                "Market Cap (Rs Cr)": mcap_final,
                "Market Cap Source": mcap_source,
                "CMP": cmp_final,
                "CMP Source": cmp_source,
                "52W High": week52_high,
                "52W Low": week52_low,
                "52W Source": source_52w,
                "Deviation from 52W High (%)": deviation_from_high_pct,
                "Deviation from 52W Low (%)": deviation_from_low_pct,
                "ROE (%)": roe,
                "ROCE (%)": roce_display,
                "50 DMA": dma50,
                "200 DMA": dma200,
                "Signal": "BEARISH WATCH" if flagged else "Normal",
            })
            highlight_flags.append(flagged)
            print("done")

        except Exception as exc:
            print(f"skipped ({exc})")
            results.append({
                "Rank": c["Rank"], "Company": c["Company"], "NSE Symbol": nse_code,
                "Market Cap (Rs Cr)": None, "Market Cap Source": "N/A",
                "CMP": None, "CMP Source": "N/A", "52W High": None, "52W Low": None,
                "52W Source": "N/A", "Deviation from 52W High (%)": None,
                "Deviation from 52W Low (%)": None,
                "ROE (%)": None, "ROCE (%)": None, "50 DMA": None, "200 DMA": None,
                "Signal": "N/A",
            })
            highlight_flags.append(False)

        time.sleep(SLEEP_BETWEEN_REQUESTS)

    df = pd.DataFrame(results)

    print(f"\nWriting Excel report -> {OUTPUT_FILE}")
    write_excel(df, highlight_flags, OUTPUT_FILE)

    flagged_count = sum(highlight_flags)
    print(f"\nDone! {flagged_count} of {len(companies)} companies flagged "
          f"(50 DMA < 200 DMA and CMP < 50 DMA).")
    print(f"Report saved as: {OUTPUT_FILE}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nStopped by user.")
    except Exception:
        print("\nAn unexpected error occurred:")
        traceback.print_exc()
