import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from config import APP_NAME, APP_FOOTER


def export_workbook(sheets, filepath, financial_year=""):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, df in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        ws["A1"] = APP_NAME
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = "Asset Depreciation and Deferred Tax Calculator"
        ws["A3"] = f"Financial Year: {financial_year}"
        ws["A4"] = f"Report Date: {datetime.now().strftime('%d-%b-%Y')}"
        ws["A5"] = APP_FOOTER
        start_row = 7
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
        for r_idx, row in enumerate(df.itertuples(index=False), start=start_row + 1):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(filepath)
    return filepath