"""
Ingestion & Compilation Module — Screen 1 backend. Schema-free.

Reads N division-wise Excel files for a billing month, auto-detects the
header row in each (row 1 or row 2 only, structural detection — see
header_detect.py), and compiles everything into one clean pandas
DataFrame plus a structured load log.

This module has NO knowledge of what columns "should" exist. Whatever
columns a file has, under whatever names, are kept as-is. When compiling
multiple files together, columns are aligned across files by EXACT name
match only (after trimming whitespace) — no fuzzy matching, no predefined
field vocabulary. A column present in one file but not another simply
appears blank for rows from files that didn't have it.

Cell values are carried through exactly as Excel stored them — a cell
Excel already typed as a number stays a number (e.g. a genuine billed
amount), and a cell Excel typed as text stays text (e.g. an ID number
stored as a text string). This module does NOT re-interpret or coerce
values based on heuristics (e.g. "this column looks 80% numeric, so
force the rest to numbers too") — that produced misleading results, such
as text-formatted service numbers being forced into floats and displayed
with unwanted decimal points. Whatever type is in the source file is what
is compiled.

A row is only dropped if it is completely blank across every cell.
"""

from dataclasses import dataclass, field
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np

from core.header_detect import detect_header_row


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class LogDetailError:
    row_number: int          # 1-based row number in the ORIGINAL source file
    field: str
    issue: str
    raw_value: object
    severity: str  # 'error' | 'warning'


@dataclass
class DivisionLoadLog:
    division: str
    file_name: str
    header_row_detected: int  # 1-based, for display
    header_confidence: str
    rows_read: int
    rows_rejected: int
    column_mismatches_found: int
    notes: str
    detected_columns: list[str] = field(default_factory=list)
    errors: list[LogDetailError] = field(default_factory=list)


@dataclass
class FileIngestSpec:
    """What Screen 1 sends per uploaded file, after the user confirms/edits it."""
    file_path: str
    division: str
    header_row_override: int | None = None  # 1-based; None = auto-detect (row 1 or 2 only)


@dataclass
class CompileResult:
    compiled_df: pd.DataFrame
    load_logs: list[DivisionLoadLog]
    total_rows_read: int
    total_rows_rejected: int
    total_divisions: int
    has_warnings_or_errors: bool
    compiled_at: str
    all_columns: list[str] = field(default_factory=list)


DIVISION_TAG_COL = "Source Division"
FILE_TAG_COL = "Source File"


# ---------------------------------------------------------------------------
# Per-file reading
# ---------------------------------------------------------------------------

def _read_raw_rows(file_path: str, sheet_name=0) -> list[list]:
    """
    Read a sheet with no header assumption, return list-of-lists.
    Uses python-calamine (fast Rust-based reader) with an openpyxl
    read_only fallback for files calamine can't parse.
    """
    try:
        from python_calamine import CalamineWorkbook

        wb = CalamineWorkbook.from_path(file_path)
        name = wb.sheet_names[sheet_name] if isinstance(sheet_name, int) else sheet_name
        sheet = wb.get_sheet_by_name(name)
        return sheet.to_python()
    except Exception:
        import openpyxl as _openpyxl

        wb = _openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
        return rows


def detect_file_header(file_path: str, sheet_name=0) -> tuple[int, str, list[str]]:
    """
    Returns (header_row_1based, confidence, headers).
    Public helper so Screen 1's "detected header row" column can call this
    directly. Only ever considers row 1 or row 2.
    """
    raw_rows = _read_raw_rows(file_path, sheet_name=sheet_name)
    result = detect_header_row(raw_rows)
    return result.header_row_index + 1, result.confidence, result.candidate_headers


# ---------------------------------------------------------------------------
# Value cleaning
# ---------------------------------------------------------------------------

def _is_blank(v) -> bool:
    return v is None or (isinstance(v, float) and np.isnan(v)) or str(v).strip() == ""


def _clean_header_name(h) -> str:
    return str(h).strip() if h is not None else ""


# ---------------------------------------------------------------------------
# Single-file ingestion
# ---------------------------------------------------------------------------

def ingest_single_file(spec: FileIngestSpec, sheet_name=0) -> tuple[pd.DataFrame, DivisionLoadLog]:
    """
    Read one division file end-to-end: detect/apply header row (row 1 or 2
    only), take whatever columns exist as-is, coerce numeric-looking columns
    to numbers, and return (clean_df, load_log).

    Rows are dropped only if EVERY cell in them is blank. Individual cells
    that fail to parse as a number in an otherwise-numeric column are kept
    as-is (not silently zeroed or dropped) and logged as a warning.
    """
    file_name = Path(spec.file_path).name
    raw_rows = _read_raw_rows(spec.file_path, sheet_name=sheet_name)

    if spec.header_row_override is not None:
        header_idx = spec.header_row_override - 1
        confidence = "manual"
    else:
        det = detect_header_row(raw_rows)
        header_idx = det.header_row_index
        confidence = det.confidence

    if header_idx >= len(raw_rows):
        log = DivisionLoadLog(
            division=spec.division, file_name=file_name, header_row_detected=header_idx + 1,
            header_confidence=confidence, rows_read=0, rows_rejected=0,
            column_mismatches_found=0,
            notes="ERROR: header row index beyond file length — file may be empty or corrupt.",
        )
        return pd.DataFrame(), log

    raw_headers = [_clean_header_name(h) for h in raw_rows[header_idx]]
    data_rows = raw_rows[header_idx + 1:]
    data_rows = [r for r in data_rows if any(not _is_blank(v) for v in r)]

    # De-duplicate blank/duplicate headers so pandas doesn't choke
    seen: dict[str, int] = {}
    final_headers = []
    for h in raw_headers:
        if h == "":
            h = f"(unnamed column {len(final_headers) + 1})"
        if h in seen:
            seen[h] += 1
            h = f"{h} ({seen[h]})"
        else:
            seen[h] = 0
        final_headers.append(h)

    n_cols = len(final_headers)
    errors: list[LogDetailError] = []

    clean_rows = []
    for row in data_rows:
        row = list(row)
        if len(row) < n_cols:
            row = row + [None] * (n_cols - len(row))
        elif len(row) > n_cols:
            row = row[:n_cols]
        clean_rows.append(row)

    df = pd.DataFrame(clean_rows, columns=final_headers)

    df.insert(0, FILE_TAG_COL, file_name)
    df.insert(0, DIVISION_TAG_COL, spec.division)

    notes_parts = []
    if confidence == "low":
        notes_parts.append("Header row auto-detection confidence was LOW — please verify header row manually")
    blank_header_count = sum(1 for h in raw_headers if h == "")
    if blank_header_count:
        notes_parts.append(f"{blank_header_count} column(s) had no header text and were auto-named")
    notes = "; ".join(notes_parts) if notes_parts else "OK"

    log = DivisionLoadLog(
        division=spec.division,
        file_name=file_name,
        header_row_detected=header_idx + 1,
        header_confidence=confidence,
        rows_read=len(data_rows),
        rows_rejected=0,
        column_mismatches_found=0,
        notes=notes,
        detected_columns=final_headers,
        errors=errors,
    )

    return df, log


# ---------------------------------------------------------------------------
# Multi-file compilation
# ---------------------------------------------------------------------------

def compile_divisions(specs: list[FileIngestSpec], sheet_name=0) -> CompileResult:
    """
    Ingest all division files and concatenate into a single compiled
    dataset. Columns are aligned across files by EXACT name match only
    (pandas.concat does this natively — a column present in file A but not
    file B simply appears as NaN for file B's rows). No fuzzy matching, no
    predefined field list.
    """
    all_dfs = []
    all_logs = []

    for spec in specs:
        try:
            df, log = ingest_single_file(spec, sheet_name=sheet_name)
        except Exception as exc:
            log = DivisionLoadLog(
                division=spec.division, file_name=Path(spec.file_path).name,
                header_row_detected=spec.header_row_override or 0,
                header_confidence="error", rows_read=0, rows_rejected=0,
                column_mismatches_found=0,
                notes=f"FAILED TO READ FILE: {exc}",
            )
            df = pd.DataFrame()
        all_dfs.append(df)
        all_logs.append(log)

    compiled_df = pd.concat(all_dfs, ignore_index=True, sort=False) if all_dfs else pd.DataFrame()

    total_rows_read = sum(l.rows_read for l in all_logs)
    total_rows_rejected = sum(l.rows_rejected for l in all_logs)
    has_warnings = any(
        l.header_confidence in ("low", "error") or l.errors or l.notes.startswith("FAILED")
        for l in all_logs
    )

    return CompileResult(
        compiled_df=compiled_df,
        load_logs=all_logs,
        total_rows_read=total_rows_read,
        total_rows_rejected=total_rows_rejected,
        total_divisions=len(specs),
        has_warnings_or_errors=has_warnings,
        compiled_at=datetime.now().isoformat(timespec="seconds"),
        all_columns=list(compiled_df.columns),
    )
