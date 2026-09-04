"""Generate 4 realistic, messily-formatted sample division Excel files for testing."""

import random
from pathlib import Path

import openpyxl
from openpyxl import Workbook

random.seed(42)

OUT_DIR = Path(__file__).parent.parent / "sample_data"
OUT_DIR.mkdir(exist_ok=True)

TARIFFS = ["Domestic (LT-1)", "Commercial (LT-2)", "Industrial Small Power (LT-3)",
           "Industrial Heavy HT (HT-1)", "Agricultural Pump Sets (AG-1)"]
METER_STATUSES = ["Normal", "Normal", "Normal", "Defective", "Burnt", "Door Lock"]
CONN_STATUSES = ["Active", "Active", "Active", "Active", "Disconnected"]


def gen_rows(n, division_name):
    rows = []
    for i in range(n):
        billed = round(random.uniform(500, 45000), 2)
        collected = round(billed * random.uniform(0.4, 1.0), 2)
        rows.append({
            "consumer_no": f"{division_name[:3].upper()}{100000+i}",
            "name": f"Consumer {i+1} - {division_name}",
            "tariff": random.choice(TARIFFS),
            "load_kw": round(random.uniform(1, 150), 2),
            "units": round(random.uniform(50, 5000), 1),
            "energy_amt": round(billed * 0.7, 2),
            "fc": round(billed * 0.15, 2),
            "ed": round(billed * 0.1, 2),
            "surcharge": round(billed * 0.05, 2),
            "bill_amount": billed,
            "collected_amount": collected,
            "arrears": round(random.uniform(0, 150000), 2),
            "meter_status": random.choice(METER_STATUSES),
            "status": random.choice(CONN_STATUSES),
        })
    return rows


# --- File 1: SAP-ISU style, clean header on row 1, 200 rows ---
def make_file_1():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = ["CA_NUMBER", "PARTNER_NAME", "DIV_CODE", "RATE_CATEGORY", "SANCT_KW",
               "BILLED_KWH", "ENERGY_AMT", "DEMAND_AMT", "ED_AMT", "DPS_SURCHARGE",
               "NET_BILL_AMT", "REALIZATION_AMT", "ACCUMULATED_ARREARS", "METER_CONDITION", "CONN_STATE"]
    ws.append(headers)
    for r in gen_rows(200, "Div01"):
        ws.append([r["consumer_no"], r["name"], "Division 01 - Metro North City", r["tariff"], r["load_kw"],
                   r["units"], r["energy_amt"], r["fc"], r["ed"], r["surcharge"], r["bill_amount"],
                   r["collected_amount"], r["arrears"], r["meter_status"], r["status"]])
    wb.save(OUT_DIR / "Division01_MetroNorth_Aug2026.xlsx")


# --- File 2: title banner rows + blank row before header (header on row 4) ---
def make_file_2():
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["TATA POWER ODISHA DISTRIBUTION LTD - DIVISION 02 BILLING REPORT"])
    ws.append(["Report Generated: 05-Sep-2026 10:32 AM"])
    ws.append([])
    headers = ["Consumer_ID", "Consumer_Name", "Division_Name", "Tariff_Type", "Conn_Load_KW",
               "Units_Billed", "EC_Charges", "FC_Charges", "Duty_State", "DPS_Amt",
               "Total_Demand", "Collection_Amt", "Arrear_Balance", "Meter_Status", "Status"]
    ws.append(headers)
    for r in gen_rows(180, "Div02"):
        ws.append([r["consumer_no"], r["name"], "Division 02 - Metro South Central", r["tariff"], r["load_kw"],
                   r["units"], r["energy_amt"], r["fc"], r["ed"], r["surcharge"], r["bill_amount"],
                   r["collected_amount"], r["arrears"], r["meter_status"], r["status"]])
    wb.save(OUT_DIR / "Division02_export_raw.xlsx")


# --- File 3: spot billing format, header on row 1, different column set/order ---
def make_file_3():
    wb = Workbook()
    ws = wb.active
    headers = ["SC_NO", "NAME", "DIV", "TARIFF", "KW", "KWH", "ENERGY_CHG", "FIXED_CHG",
               "ED", "SURCHARGE", "BILL_TOT", "PAID_AMT", "ARREARS", "STATUS_MTR", "SUPPLY_FLAG"]
    ws.append(headers)
    for r in gen_rows(220, "Div03"):
        ws.append([r["consumer_no"], r["name"], "Division 03 - Industrial Corridor East", r["tariff"], r["load_kw"],
                   r["units"], r["energy_amt"], r["fc"], r["ed"], r["surcharge"], r["bill_amount"],
                   r["collected_amount"], r["arrears"], r["meter_status"], r["status"]])
    wb.save(OUT_DIR / "Div03_spotbilling.xlsx")


# --- File 4: messy — two blank rows, merged-style title, header row 3, one required column MISSING (no arrears) ---
def make_file_4():
    wb = Workbook()
    ws = wb.active
    ws.append(["", "", "Division 04 Monthly Statement", "", ""])
    ws.append([])
    headers = ["Consumer No", "Consumer Name", "Division", "Category", "Units (kWh)",
               "Billed Amount", "Amount Collected", "Meter Status"]
    ws.append(headers)
    for r in gen_rows(150, "Div04"):
        ws.append([r["consumer_no"], r["name"], "Division 04 - Industrial Park - SEZ", r["tariff"],
                   r["units"], r["bill_amount"], r["collected_amount"], r["meter_status"]])
    # inject a couple of dirty rows: blank consumer no, text in a numeric field
    ws.append([None, "Bad Row Consumer", "Division 04 - Industrial Park - SEZ", "Domestic (LT-1)",
               100, "NOT_A_NUMBER", 500, "Normal"])
    wb.save(OUT_DIR / "DIV-04-billing-Aug26-FINAL-v2.xlsx")


make_file_1()
make_file_2()
make_file_3()
make_file_4()
print("Sample files created:")
for f in sorted(OUT_DIR.glob("*.xlsx")):
    print(" -", f.name)
