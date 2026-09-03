"""
The workpaper file.

Produces a multi-sheet Excel workbook that an engagement team can put on
the audit file as it stands: the analytical review, the Schedule III ratio
note, the SA 240 journal entry selections, the sample, and the CARO 2020
checklist ready for completion.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .pipeline import EngagementResult

INK = "1F2A24"
ACCENT = "1E5A4A"
HEADER_FILL = PatternFill("solid", fgColor="E3EDE7")
TITLE_FILL = PatternFill("solid", fgColor=ACCENT)
FLAG_FILL = PatternFill("solid", fgColor="FCEBE7")
OK_FILL = PatternFill("solid", fgColor="EAF3EE")
THIN = Side(style="thin", color="C9D2CC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RUPEE = '#,##0.00'


def _write_frame(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet: str,
    title: str,
    subtitle: str = "",
    money_columns: tuple[str, ...] = (),
) -> None:
    start = 3 if subtitle else 2
    df.to_excel(writer, sheet_name=sheet, index=False, startrow=start)
    ws = writer.sheets[sheet]

    ws.cell(row=1, column=1, value=title).font = Font(
        name="Calibri", size=13, bold=True, color="FFFFFF"
    )
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(df.columns), 4))
    ws.row_dimensions[1].height = 22
    ws.cell(row=1, column=1).alignment = Alignment(vertical="center", indent=1)

    if subtitle:
        c = ws.cell(row=2, column=1, value=subtitle)
        c.font = Font(name="Calibri", size=9, italic=True, color="55635E")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(df.columns), 4))

    header_row = start + 1
    for col_idx, name in enumerate(df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(name="Calibri", size=10, bold=True, color=INK)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        letter = get_column_letter(col_idx)
        longest = max(
            [len(str(name))] + [len(str(v)) for v in df[name].head(200).tolist()] or [10]
        )
        ws.column_dimensions[letter].width = min(max(12, longest + 3), 62)
        if name in money_columns:
            for r in range(header_row + 1, header_row + 1 + len(df)):
                ws.cell(row=r, column=col_idx).number_format = RUPEE

    for r in range(header_row + 1, header_row + 1 + len(df)):
        for c in range(1, len(df.columns) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def build_workbook(result: EngagementResult, path: str | Path) -> Path:
    path = Path(path)
    fy = result.inputs.financial_year
    client = result.inputs.client_name
    h = result.headlines()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # ---- 1. Cover and summary ------------------------------------
        summary = pd.DataFrame(
            [
                {"Item": "Client", "Value": client, "Note": ""},
                {"Item": "Financial year", "Value": fy, "Note": ""},
                {"Item": "Year end", "Value": result.inputs.year_end.strftime("%d %B %Y"), "Note": ""},
                {"Item": "Trial balance tallies", "Value": "Yes" if h["trial_balance_tallies"] else "No",
                 "Note": f"Difference Rs {h['trial_balance_difference']:,.2f}"},
                {"Item": "Schedule III mapping coverage", "Value": f"{h['mapping_coverage']:.1%}",
                 "Note": f"{h['ledgers_for_review']} ledger(s) require auditor review"},
                {"Item": "Balance sheet tallies", "Value": "Yes" if h["balance_sheet_tallies"] else "No",
                 "Note": h["balance_sheet_reconciliation"]},
                {"Item": "Revenue from operations", "Value": h["revenue"], "Note": ""},
                {"Item": "Profit before tax", "Value": h["profit_before_tax"], "Note": ""},
                {"Item": "Profit after tax", "Value": h["profit_after_tax"], "Note": ""},
                {"Item": "Overall materiality", "Value": h["overall_materiality"],
                 "Note": f"{result.materiality.percentage:.2%} of {result.materiality.benchmark_label.lower()}"},
                {"Item": "Performance materiality", "Value": h["performance_materiality"],
                 "Note": f"{result.materiality.performance_pct:.0%} of overall materiality"},
                {"Item": "Ratios requiring explanation", "Value": h["ratios_requiring_explanation"],
                 "Note": "Schedule III: movement beyond 25 per cent"},
                {"Item": "Journal entries flagged", "Value": h["je_entries_flagged"],
                 "Note": f"out of {h['je_total_entries']} entries tested under SA 240"},
                {"Item": "Benford first-digit test", "Value": h["benford_conclusion"], "Note": ""},
                {"Item": "CARO 2020 applicable", "Value": "Yes" if h["caro_applies"] else "No",
                 "Note": "; ".join(result.caro.applicability.reasons) if result.caro else ""},
                {"Item": "", "Value": "", "Note": ""},
                {"Item": "DISCLAIMER", "Value": result.disclaimer, "Note": ""},
            ]
        )
        _write_frame(
            writer, summary, "1. Summary",
            f"AuditLens - analytical review: {client}",
            f"Financial year {fy}. Generated from the trial balance and general ledger supplied by the engagement team.",
            money_columns=("Value",),
        )

        # ---- 2. Schedule III mapping ---------------------------------
        mapping_df = result.mapped.df.rename(
            columns={
                "account_code": "Account code",
                "account_name": "Ledger name",
                "debit": "Debit (Rs)",
                "credit": "Credit (Rs)",
                "head": "Schedule III head",
                "basis": "Mapped on",
                "confidence": "Confidence",
                "needs_review": "Review required",
            }
        )
        mapping_df["Review required"] = mapping_df["Review required"].map(
            {True: "Yes", False: "No"}
        )
        _write_frame(
            writer, mapping_df, "2. Schedule III mapping",
            "Trial balance mapped to Schedule III, Division I",
            "Every ledger mapped on a keyword, and every unmapped ledger, requires the auditor's confirmation.",
            money_columns=("Debit (Rs)", "Credit (Rs)"),
        )

        # ---- 3. Face of the statements -------------------------------
        bs_rows = [
            {
                "Particulars": ("    " * line.level) + line.label,
                "Current year (Rs)": line.current,
                "Previous year (Rs)": line.prior,
            }
            for line in result.statements.balance_sheet
        ]
        _write_frame(
            writer, pd.DataFrame(bs_rows), "3. Balance sheet",
            "Balance Sheet - Schedule III, Division I",
            result.statements.reconciliation,
            money_columns=("Current year (Rs)", "Previous year (Rs)"),
        )

        pl_rows = [
            {
                "Particulars": ("    " * line.level) + line.label,
                "Current year (Rs)": line.current,
                "Previous year (Rs)": line.prior,
            }
            for line in result.statements.profit_and_loss
        ]
        pl_rows += [
            {"Particulars": "", "Current year (Rs)": None, "Previous year (Rs)": None},
            {"Particulars": "Profit before tax", "Current year (Rs)": result.figures.profit_before_tax,
             "Previous year (Rs)": result.prior_figures.profit_before_tax if result.prior_figures else None},
            {"Particulars": "Profit for the year", "Current year (Rs)": result.figures.profit_after_tax,
             "Previous year (Rs)": result.prior_figures.profit_after_tax if result.prior_figures else None},
        ]
        _write_frame(
            writer, pd.DataFrame(pl_rows), "4. Profit and loss",
            "Statement of Profit and Loss - Schedule III, Division I",
            money_columns=("Current year (Rs)", "Previous year (Rs)"),
        )

        # ---- 5. Ratio note -------------------------------------------
        ratios_df = pd.DataFrame(result.ratios.as_rows())
        _write_frame(
            writer, ratios_df, "5. Ratios",
            "The eleven ratios required by Schedule III",
            "Inserted by MCA notification G.S.R. 207(E) dated 24 March 2021. An explanation is required in the notes wherever a ratio moves by more than 25 per cent.",
        )
        # Shade the ratios that Schedule III requires the notes to explain.
        ws = writer.sheets["5. Ratios"]
        first_data_row = 5
        for offset, needs in enumerate(ratios_df["Explanation required"]):
            if needs == "Yes":
                for c in range(1, len(ratios_df.columns) + 1):
                    ws.cell(row=first_data_row + offset, column=c).fill = FLAG_FILL

        # ---- 6. Materiality ------------------------------------------
        mat_df = pd.DataFrame(result.materiality.as_rows())
        _write_frame(
            writer, mat_df, "6. Materiality",
            "Materiality for the engagement - SA 320",
            f"Benchmark rationale: {result.materiality.rationale}",
            money_columns=("Amount (Rs)",),
        )

        # ---- 7. Journal entry testing --------------------------------
        if result.je_analysis:
            test_rows = [
                {
                    "Test": t.name,
                    "Reference": t.reference,
                    "What it selects": t.description,
                    "Population (entries)": t.population,
                    "Entries flagged": t.flagged,
                    "Flag rate": f"{t.rate:.2%}",
                }
                for t in result.je_analysis.tests
            ]
            b = result.je_analysis.benford
            test_rows.append(
                {
                    "Test": "Benford first-digit distribution",
                    "Reference": "SA 240 para A43",
                    "What it selects": f"Mean absolute deviation {b.mad:.5f} across {b.total} entries. {b.conclusion}",
                    "Population (entries)": b.total,
                    "Entries flagged": "n/a",
                    "Flag rate": "n/a",
                }
            )
            _write_frame(
                writer, pd.DataFrame(test_rows), "7. JE tests",
                "Journal entry testing - SA 240",
                "A flag is a selection for examination, not a finding.",
            )

            flags = result.je_analysis.flags_frame()
            if not flags.empty:
                _write_frame(
                    writer, flags, "8. JE selections",
                    "Journal entries selected for examination",
                    f"{len(result.je_analysis.flagged_entries)} distinct entries selected by one or more tests.",
                    money_columns=("Amount (Rs)",),
                )

        # ---- 9. Sample -----------------------------------------------
        if result.sample:
            plan = result.sample
            header = pd.DataFrame(
                [
                    {"Parameter": "Population size (items)", "Value": plan.population_size},
                    {"Parameter": "Population value (Rs)", "Value": plan.population_value},
                    {"Parameter": "Tolerable misstatement (Rs)", "Value": plan.tolerable_misstatement},
                    {"Parameter": "Confidence factor", "Value": plan.confidence_factor},
                    {"Parameter": "Sampling interval (Rs)", "Value": plan.sampling_interval},
                    {"Parameter": "Random start (Rs)", "Value": plan.random_start},
                    {"Parameter": "Seed (for re-performance)", "Value": plan.seed},
                    {"Parameter": "Sample size", "Value": plan.sample_size},
                    {"Parameter": "Value coverage", "Value": f"{plan.coverage:.1%}"},
                ]
                + [{"Parameter": "Warning", "Value": w} for w in plan.warnings]
            )
            _write_frame(
                writer, header, "9. Sample plan",
                "Monetary unit sample - SA 530",
                "The seed and random start are recorded so the selection can be re-performed by a reviewer.",
                money_columns=("Value",),
            )
            items = plan.as_frame()
            if not items.empty:
                _write_frame(
                    writer, items.head(500), "10. Sample items",
                    "Items selected for testing",
                    f"Showing the first {min(500, len(items))} of {len(items)} selected items.",
                    money_columns=("Amount (Rs)",),
                )

        # ---- 11. CARO ------------------------------------------------
        if result.caro:
            caro_df = pd.DataFrame(result.caro.as_rows())
            _write_frame(
                writer, caro_df, "11. CARO 2020",
                "CARO 2020 - paragraph 3 clause checklist",
                "Applicability: "
                + ("applies. " if result.caro.applicability.applies else "does not apply. ")
                + " ".join(result.caro.applicability.reasons),
            )

    return path
