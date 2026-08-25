import os
import re
import shutil
import datetime
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import customtkinter as ctk
import pandas as pd
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable, Image
from reportlab.lib.styles import getSampleStyleSheet

# Appearance and Theme Setup
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

DB_FILE = "veh_mis_database.xlsx"
INVOICE_DIR = "invoices"
RECEIPT_DIR = "receipts"
PHOTO_DIR = "driver_photos"
LEDGER_DIR = "ledgers"
REPORT_DIR = "reports_export"
DRIVER_CARD_DIR = "driver_cards"

os.makedirs(INVOICE_DIR, exist_ok=True)
os.makedirs(RECEIPT_DIR, exist_ok=True)
os.makedirs(PHOTO_DIR, exist_ok=True)
os.makedirs(LEDGER_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)
os.makedirs(DRIVER_CARD_DIR, exist_ok=True)

# GSTN Structure Validator: 11AAAAA1111A1AO
GSTIN_STRUCTURE_REGEX = r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}[A-Z]{1}[0-9A-Z]{1}$"

def is_valid_gstn(gstn_str):
    """Validates exact 15-character GSTN pattern (11AAAAA1111A1AO)."""
    if not gstn_str:
        return False
    clean_gstn = str(gstn_str).strip().upper()
    return bool(re.match(GSTIN_STRUCTURE_REGEX, clean_gstn))

def format_date_to_ddmmyyyy(date_val):
    """Converts any date string/object to DD/MM/YYYY format."""
    if not date_val:
        return datetime.date.today().strftime("%d/%m/%Y")
    s_val = str(date_val).strip()[:10]
    try:
        parts = s_val.split("/")
        if "/" in s_val and len(parts) == 3:
            return s_val
        dt = datetime.datetime.strptime(s_val, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except:
        try:
            dt = datetime.datetime.strptime(s_val, "%d/%m/%Y")
            return dt.strftime("%d/%m/%Y")
        except:
            return s_val

def parse_date_to_yyyymmdd(date_str):
    """Parses DD/MM/YYYY or YYYY-MM-DD back to standard YYYY-MM-DD for database."""
    if not date_str:
        return str(datetime.date.today())
    s_val = str(date_str).strip()
    try:
        if "/" in s_val:
            parts = s_val.split("/")
            if len(parts) == 3:
                return f"{parts[2]}-{parts[1].zfill(2)}-{parts[0].zfill(2)}"
        dt = datetime.datetime.strptime(s_val[:10], "%Y-%m-%d")
        return dt.strftime("%Y-%m-%d")
    except:
        return str(datetime.date.today())

# Database Initialization
def init_database():
    if not os.path.exists(DB_FILE):
        wb = Workbook()
        
        # 1. Company Profile
        ws_comp = wb.active
        ws_comp.title = "Company_Profile"
        ws_comp.append(["Org_Name", "Org_Status", "Reg_No", "GSTN", "PAN", "TAN", "Bank_Name", "Account_No", "IFSC", "Address", "Email", "Mobile", "Website"])
        ws_comp.append(["Travel Mate Rentals Pvt. Ltd.", "Private Limited Co.", "REG-12345", "09AAAAA0000A1Z5", "AAAAA0000A", "TAN12345", "State Bank of India", "123456789012", "SBIN0001234", "Civil Lines, New Delhi - 110001", "info@travelmate.com", "+91 9876543210", "www.travelmate.com"])
        
        # 2. Vehicle Master
        ws_veh = wb.create_sheet(title="Vehicle_Master")
        ws_veh.append(["Brand", "Type", "Reg_No", "Reg_Expiry", "Insurance_Expiry", "Pollution_Expiry", "State_Permit_Expiry", "All_India_Permit_Expiry", "Fitness_Expiry"])
        ws_veh.append(["Toyota", "Innova Crysta", "DL01AB1234", "2032-01-01", "2027-01-01", "2027-01-01", "2027-01-01", "2027-01-01", "2032-01-01"])
        ws_veh.append(["Maruti", "Dzire Sedan", "DL01CD5678", "2032-01-01", "2027-01-01", "2027-01-01", "2027-01-01", "2027-01-01", "2032-01-01"])
        
        # 3. Driver Master
        ws_drv = wb.create_sheet(title="Driver_Master")
        ws_drv.append(["Driver_ID", "Name", "ID_No", "DL_No", "DL_Expiry", "Photo_Path"])
        ws_drv.append(["DRV-101", "Ramesh Kumar", "ID-9988", "DL-99887766", "2028-01-01", ""])
        ws_drv.append(["DRV-102", "Suresh Singh", "ID-7766", "DL-11223344", "2028-01-01", ""])
        
        # 4. Inquiries
        ws_inq = wb.create_sheet(title="Inquiries")
        ws_inq.append(["Inquiry_No", "Date", "Customer_Name", "PAN_GSTN", "Address", "City", "PIN", "State", "Phone", "Whatsapp", "Email",
                       "Date_From", "Date_To", "From_Station", "To_Station", "Vehicle_Type", "Driver_Expected",
                       "Est_KM", "Est_Amount", "Rate_Per_Extra_KM", "Night_Charges", "Toll_Charges", "Parking_Charges", "Other_Charges", "Terms", "Deposit_Amount", "Status"])
        
        # 5. Bookings
        ws_bk = wb.create_sheet(title="Bookings")
        ws_bk.append(["Booking_ID", "Inquiry_No", "Customer_Name", "Vehicle_Assigned", "Driver_Assigned", "Date_From", "Date_To", "Advance_Amount", "Journey_Status", "Passengers_Summary"])
        
        # 6. Booking Passengers List
        ws_pax = wb.create_sheet(title="Booking_Passengers")
        ws_pax.append(["Booking_ID", "Passenger_Name", "Age", "Gender", "Contact_No"])
        
        # 7. Invoices
        ws_inv = wb.create_sheet(title="Invoices")
        ws_inv.append(["Invoice_No", "Date", "Inquiry_No", "Booking_ID", "Customer_Name", "GSTN", "Actual_KM", "Base_Amount", "Extra_KM_Charges", "Night_Charges", "Toll_Charges", "Parking_Charges", "Other_Charges", "Discount", "Taxable_Amount", "IGST", "CGST", "SGST", "Grand_Total", "Advance_Adjusted", "Net_Payable", "ITC_Claimable"])
        
        # 8. Receipts Master
        ws_rec = wb.create_sheet(title="Receipts")
        ws_rec.append(["Receipt_No", "Date", "Inquiry_No", "Customer_Name", "Payment_Type", "Payment_Mode", "Ref_No", "Amount", "Remarks"])
        
        # 9. Customer Ledger
        ws_led = wb.create_sheet(title="Customer_Ledger")
        ws_led.append(["Date", "Inquiry_No", "Customer_Name", "Particulars", "Voucher_Type", "Debit", "Credit", "Balance"])
        
        wb.save(DB_FILE)
    else:
        wb = load_workbook(DB_FILE)
        existing = wb.sheetnames
        if "Receipts" not in existing:
            ws_rec = wb.create_sheet(title="Receipts")
            ws_rec.append(["Receipt_No", "Date", "Inquiry_No", "Customer_Name", "Payment_Type", "Payment_Mode", "Ref_No", "Amount", "Remarks"])
            wb.save(DB_FILE)
            
        ws_veh_chk = wb["Vehicle_Master"]
        headers_veh = [cell.value for cell in ws_veh_chk[1]]
        if "Fitness_Expiry" not in headers_veh:
            ws_veh_chk.cell(row=1, column=len(headers_veh)+1, value="Fitness_Expiry")
            for r in range(2, ws_veh_chk.max_row+1):
                ws_veh_chk.cell(row=r, column=len(headers_veh)+1, value="2032-01-01")
                
        ws_drv_chk = wb["Driver_Master"]
        headers_drv = [cell.value for cell in ws_drv_chk[1]]
        if "DL_Expiry" not in headers_drv:
            ws_drv_chk.insert_cols(5)
            ws_drv_chk.cell(row=1, column=5, value="DL_Expiry")
            for r in range(2, ws_drv_chk.max_row+1):
                ws_drv_chk.cell(row=r, column=5, value="2028-01-01")
        wb.save(DB_FILE)

init_database()

# Reusable Interactive Date Picker Dialog (DD/MM/YYYY)
class DatePickerPopup(ctk.CTkToplevel):
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.title("Select Date (DD/MM/YYYY)")
        self.geometry("340x360")
        self.grab_set()
        
        self.callback = callback
        self.now = datetime.datetime.now()
        self.sel_year = self.now.year
        self.sel_month = self.now.month
        self.sel_day = self.now.day
        
        top_f = ctk.CTkFrame(self, fg_color="transparent")
        top_f.pack(fill="x", padx=15, pady=15)
        
        self.lbl_month_year = ctk.CTkLabel(top_f, text="", font=("Helvetica", 16, "bold"), text_color="#1a365d")
        self.lbl_month_year.pack(side="left")
        
        ctk.CTkButton(top_f, text=">", width=40, height=30, command=self.next_month).pack(side="right", padx=2)
        ctk.CTkButton(top_f, text="<", width=40, height=30, command=self.prev_month).pack(side="right", padx=2)
        
        self.cal_frame = ctk.CTkFrame(self)
        self.cal_frame.pack(fill="both", expand=True, padx=15, pady=5)
        
        self.render_calendar()

    def prev_month(self):
        self.sel_month -= 1
        if self.sel_month < 1:
            self.sel_month = 12
            self.sel_year -= 1
        self.render_calendar()

    def next_month(self):
        self.sel_month += 1
        if self.sel_month > 12:
            self.sel_month = 1
            self.sel_year += 1
        self.render_calendar()

    def render_calendar(self):
        for widget in self.cal_frame.winfo_children():
            widget.destroy()
            
        months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
        self.lbl_month_year.configure(text=f"{months[self.sel_month-1]} {self.sel_year}")
        
        days_header = ["Mo", "Tu", "We", "Th", "Fr", "Sa", "Su"]
        for col_idx, d_name in enumerate(days_header):
            lbl = ctk.CTkLabel(self.cal_frame, text=d_name, font=("Helvetica", 12, "bold"), text_color="#718096")
            lbl.grid(row=0, column=col_idx, padx=6, pady=4)
            
        import calendar
        cal = calendar.monthcalendar(self.sel_year, self.sel_month)
        for row_idx, week in enumerate(cal):
            for col_idx, day in enumerate(week):
                if day != 0:
                    btn = ctk.CTkButton(
                        self.cal_frame, text=str(day), width=38, height=34,
                        font=("Helvetica", 12),
                        fg_color="#e2e8f0" if (day != self.sel_day or self.sel_month != self.now.month or self.sel_year != self.now.year) else "#1a365d",
                        text_color="#1a202c" if (day != self.sel_day or self.sel_month != self.now.month or self.sel_year != self.now.year) else "#ffffff",
                        command=lambda d=day: self.select_day(d)
                    )
                    btn.grid(row=row_idx+1, column=col_idx, padx=2, pady=2)

    def select_day(self, day):
        formatted_date = f"{str(day).zfill(2)}/{str(self.sel_month).zfill(2)}/{self.sel_year}"
        self.callback(formatted_date)
        self.destroy()

# Main Application Class
class CabMISApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Travel Mate - Vehicle Rental & Travel MIS System")
        self.geometry("1420x920")
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        
        self.font_title = ctk.CTkFont(size=23, weight="bold")
        self.font_header = ctk.CTkFont(size=18, weight="bold")
        self.font_sub = ctk.CTkFont(size=15, weight="bold")
        self.font_body = ctk.CTkFont(size=14, weight="normal")
        self.font_label = ctk.CTkFont(size=13, weight="bold")
        self.font_btn = ctk.CTkFont(size=13, weight="bold")
        
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", font=("Helvetica", 13, "bold"), background="#e2e8f0", foreground="#1a365d")
        style.configure("Treeview", font=("Helvetica", 13), rowheight=34)
        
        self.sidebar_frame = ctk.CTkFrame(self, width=260, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, sticky="nsew")
        
        self.logo_label = ctk.CTkLabel(self.sidebar_frame, text="🚗 TRAVEL MATE", font=self.font_title, text_color="#1a73e8")
        self.logo_label.pack(pady=(25, 2), padx=20)
        
        self.sub_logo_label = ctk.CTkLabel(self.sidebar_frame, text="Vehicle MIS & Accounts Pro", font=ctk.CTkFont(size=13), text_color="#718096")
        self.sub_logo_label.pack(pady=(0, 25), padx=20)
        
        self.btn_dashboard = ctk.CTkButton(self.sidebar_frame, text="📈 Executive MIS Dashboard", font=self.font_btn, anchor="w", height=44, fg_color="#2b6cb0", hover_color="#2c5282", command=lambda: self.show_frame("dashboard"))
        self.btn_dashboard.pack(fill="x", padx=15, pady=6)

        self.btn_inquiry = ctk.CTkButton(self.sidebar_frame, text="📋 Inquiry Management", font=self.font_btn, anchor="w", height=44, command=lambda: self.show_frame("inquiry"))
        self.btn_inquiry.pack(fill="x", padx=15, pady=6)
        
        self.btn_booking = ctk.CTkButton(self.sidebar_frame, text="📑 Booking Management", font=self.font_btn, anchor="w", height=44, command=lambda: self.show_frame("booking"))
        self.btn_booking.pack(fill="x", padx=15, pady=6)
        
        self.btn_accounts = ctk.CTkButton(self.sidebar_frame, text="💳 Accounts & Billing", font=self.font_btn, anchor="w", height=44, command=lambda: self.show_frame("accounts"))
        self.btn_accounts.pack(fill="x", padx=15, pady=6)
        
        self.btn_reports = ctk.CTkButton(self.sidebar_frame, text="📊 Reports & MIS", font=self.font_btn, anchor="w", height=44, command=lambda: self.show_frame("reports"))
        self.btn_reports.pack(fill="x", padx=15, pady=6)
        
        self.btn_masters = ctk.CTkButton(self.sidebar_frame, text="⚙️ Master Setup", font=self.font_btn, anchor="w", height=44, command=lambda: self.show_frame("masters"))
        self.btn_masters.pack(fill="x", padx=15, pady=6)
        
        self.main_work_area = ctk.CTkFrame(self, fg_color="transparent")
        self.main_work_area.grid(row=0, column=1, sticky="nsew", padx=15, pady=15)
        self.main_work_area.grid_rowconfigure(1, weight=1)
        self.main_work_area.grid_columnconfigure(0, weight=1)
        
        self.branding_header = ctk.CTkFrame(self.main_work_area, corner_radius=10, fg_color="#1a365d", height=78)
        self.branding_header.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        self.branding_header.grid_columnconfigure(0, weight=1)
        self.branding_header.grid_columnconfigure(1, weight=1)
        
        left_brand_box = ctk.CTkFrame(self.branding_header, fg_color="transparent")
        left_brand_box.grid(row=0, column=0, sticky="w", padx=20, pady=8)
        
        self.lbl_brand_name = ctk.CTkLabel(left_brand_box, text="TRAVEL MATE RENTALS", font=self.font_title, text_color="#ffffff")
        self.lbl_brand_name.pack(anchor="w")
        
        self.lbl_brand_details = ctk.CTkLabel(left_brand_box, text="Loading enterprise details...", font=ctk.CTkFont(size=12), text_color="#cbd5e0")
        self.lbl_brand_details.pack(anchor="w", pady=(2, 0))
        
        right_brand_box = ctk.CTkFrame(self.branding_header, fg_color="transparent")
        right_brand_box.grid(row=0, column=1, sticky="e", padx=20, pady=8)
        
        self.lbl_developer = ctk.CTkLabel(right_brand_box, text="Developed by CA Kapil Agarwal\n(AICA L2 B77)", font=ctk.CTkFont(size=12, weight="bold"), text_color="#63b3ed", justify="right")
        self.lbl_developer.pack(anchor="e")
        
        self.content_frame = ctk.CTkFrame(self.main_work_area, corner_radius=10)
        self.content_frame.grid(row=1, column=0, sticky="nsew")
        self.content_frame.grid_rowconfigure(0, weight=1)
        self.content_frame.grid_columnconfigure(0, weight=1)
        
        self.frames = {}
        self.create_dashboard_view()
        self.create_inquiry_view()
        self.create_booking_view()
        self.create_accounts_view()
        self.create_reports_view()
        self.create_masters_view()
        
        self.sync_all_confirmed_inquiries_to_bookings()
        self.update_top_branding_header()
        self.show_frame("dashboard")
        
        self.blinking_state = False
        self.run_blinking_animation()

    def run_blinking_animation(self):
        try:
            if hasattr(self, 'alert_tree') and self.alert_tree.winfo_exists():
                self.blinking_state = not self.blinking_state
                current_color = "#991b1b" if self.blinking_state else "#dc2626"
                self.alert_tree.tag_configure("RED", foreground=current_color, font=('Helvetica', 13, 'bold'))
        except:
            pass
        self.after(600, self.run_blinking_animation)

    def calculate_total_advance_before_invoice(self, inq_id, df_inq=None, df_rec=None):
        total_advance = 0.0
        str_inq = str(inq_id).strip()
        
        if df_inq is None or df_rec is None:
            if os.path.exists(DB_FILE):
                excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
                df_inq = excel_dict.get("Inquiries", pd.DataFrame())
                df_rec = excel_dict.get("Receipts", pd.DataFrame())
            else:
                return 0.0

        if df_inq is not None and not df_inq.empty and "Inquiry_No" in df_inq.columns:
            match_inq = df_inq[df_inq["Inquiry_No"].astype(str) == str_inq]
            if not match_inq.empty:
                total_advance += float(match_inq.iloc[0].get("Deposit_Amount", 0) or 0)

        if df_rec is not None and not df_rec.empty and "Inquiry_No" in df_rec.columns:
            match_rec = df_rec[df_rec["Inquiry_No"].astype(str) == str_inq]
            for _, r in match_rec.iterrows():
                ref_no = str(r.get("Ref_No", "")).strip()
                p_type = str(r.get("Payment_Type", "")).strip().lower()
                amt = float(r.get("Amount", 0) or 0)
                if ref_no != "DEP-INIT":
                    if "refund" in p_type or "return" in p_type:
                        total_advance -= amt
                    else:
                        total_advance += amt

        return total_advance

    def sync_all_confirmed_inquiries_to_bookings(self):
        if not os.path.exists(DB_FILE):
            return
        try:
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_inq = excel_dict.get("Inquiries", pd.DataFrame())
            df_bk = excel_dict.get("Bookings", pd.DataFrame())
            df_rec = excel_dict.get("Receipts", pd.DataFrame())
            
            if df_inq.empty or "Status" not in df_inq.columns:
                return
                
            confirmed_mask = df_inq["Status"].astype(str).str.strip().str.lower() == "confirmed"
            confirmed_inqs = df_inq[confirmed_mask]
            
            confirmed_bk_ids = {f"BK-{str(inq_id).replace('INQ-', '')}": row for inq_id, row in confirmed_inqs.set_index("Inquiry_No").iterrows()}
            
            if not df_bk.empty and "Booking_ID" in df_bk.columns:
                df_bk = df_bk[df_bk["Booking_ID"].astype(str).isin(confirmed_bk_ids.keys())].copy()
            else:
                df_bk = pd.DataFrame()

            existing_bk_ids = set(df_bk["Booking_ID"].astype(str).tolist()) if not df_bk.empty else set()
            
            new_rows = []
            for bk_id, row in confirmed_bk_ids.items():
                inq_id = str(row.name) if hasattr(row, 'name') else str(row.get("Inquiry_No", ""))
                total_adv = self.calculate_total_advance_before_invoice(inq_id, df_inq, df_rec)
                
                if bk_id not in existing_bk_ids:
                    new_rows.append({
                        "Booking_ID": bk_id,
                        "Inquiry_No": inq_id,
                        "Customer_Name": row.get("Customer_Name", ""),
                        "Vehicle_Assigned": row.get("Vehicle_Type", ""),
                        "Driver_Assigned": row.get("Driver_Expected", ""),
                        "Date_From": format_date_to_ddmmyyyy(row.get("Date_From", "")),
                        "Date_To": format_date_to_ddmmyyyy(row.get("Date_To", "")),
                        "Advance_Amount": total_adv,
                        "Journey_Status": "Not Started",
                        "Passengers_Summary": "Adult: 0, Minor: 0, Male: 0, Female: 0"
                    })
                else:
                    idx = df_bk.index[df_bk["Booking_ID"].astype(str) == str(bk_id)].tolist()
                    if idx:
                        df_bk.at[idx[0], "Advance_Amount"] = total_adv
            
            if new_rows:
                df_bk = pd.concat([df_bk, pd.DataFrame(new_rows)], ignore_index=True)
                
            excel_dict["Bookings"] = df_bk
            with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                for s_name, s_df in excel_dict.items():
                    s_df.to_excel(writer, sheet_name=s_name, index=False)
        except Exception as e:
            print("Auto-sync error:", e)

    def update_top_branding_header(self):
        if os.path.exists(DB_FILE):
            df_comp = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
            if not df_comp.empty:
                c = df_comp.iloc[0]
                name = str(c.get("Org_Name", "TRAVEL MATE RENTALS PVT. LTD."))
                addr = str(c.get("Address", ""))
                phone = str(c.get("Mobile", ""))
                gstn = str(c.get("GSTN", ""))
                self.lbl_brand_name.configure(text=f"🏢 {name.upper()}")
                self.lbl_brand_details.configure(text=f"📍 {addr} | 📞 Phone: {phone} | 🧾 GSTIN: {gstn}")

    def show_frame(self, name):
        frame = self.frames.get(name)
        if frame:
            frame.tkraise()
            if hasattr(frame, 'refresh_data'):
                frame.refresh_data()
        self.update_top_branding_header()

    # =========================================================================
    # 0. EXECUTIVE MIS DASHBOARD
    # =========================================================================
    def create_dashboard_view(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid(row=0, column=0, sticky="nsew")
        self.frames["dashboard"] = frame
        
        scroll = ctk.CTkScrollableFrame(frame, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        kpi_row = ctk.CTkFrame(scroll, fg_color="transparent")
        kpi_row.pack(fill="x", pady=5)
        kpi_row.grid_columnconfigure((0, 1, 2, 3), weight=1)
        
        self.card_rev = self.create_kpi_card(kpi_row, "Total Billed Revenue", "₹0.00", "#2b6cb0", 0)
        self.card_rec = self.create_kpi_card(kpi_row, "Total Collections", "₹0.00", "#28a745", 1)
        self.card_due = self.create_kpi_card(kpi_row, "Outstanding Dues", "₹0.00", "#dc3545", 2)
        self.card_bk = self.create_kpi_card(kpi_row, "Confirmed Bookings", "0", "#6f42c1", 3)
        
        mid_row = ctk.CTkFrame(scroll, fg_color="transparent")
        mid_row.pack(fill="x", pady=10)
        mid_row.grid_columnconfigure((0, 1), weight=1)
        
        v_frame = ctk.CTkFrame(mid_row, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        v_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6), pady=5)
        ctk.CTkLabel(v_frame, text="🚗 All Fleet / Vehicle-Wise Performance & KM Covered", font=self.font_sub, text_color="#1a365d").pack(anchor="w", padx=12, pady=(10, 5))
        
        self.v_tree = ttk.Treeview(v_frame, columns=("Vehicle", "Trips", "KM", "Revenue"), show="headings", height=6)
        self.v_tree.heading("Vehicle", text="Vehicle Model / Reg")
        self.v_tree.heading("Trips", text="Trips")
        self.v_tree.heading("KM", text="Total KM")
        self.v_tree.heading("Revenue", text="Revenue (₹)")
        self.v_tree.column("Vehicle", width=140, anchor="w")
        self.v_tree.column("Trips", width=55, anchor="center")
        self.v_tree.column("KM", width=85, anchor="center")
        self.v_tree.column("Revenue", width=100, anchor="e")
        self.v_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        d_frame = ctk.CTkFrame(mid_row, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        d_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0), pady=5)
        ctk.CTkLabel(d_frame, text="👨‍✈️ All Drivers Performance & KM Driven", font=self.font_sub, text_color="#1a365d").pack(anchor="w", padx=12, pady=(10, 5))
        
        self.d_tree = ttk.Treeview(d_frame, columns=("Driver", "Trips", "KM", "Assigned"), show="headings", height=6)
        self.d_tree.heading("Driver", text="Driver Name")
        self.d_tree.heading("Trips", text="Completed")
        self.d_tree.heading("KM", text="Total KM")
        self.d_tree.heading("Assigned", text="Assigned")
        self.d_tree.column("Driver", width=130, anchor="w")
        self.d_tree.column("Trips", width=75, anchor="center")
        self.d_tree.column("KM", width=85, anchor="center")
        self.d_tree.column("Assigned", width=90, anchor="center")
        self.d_tree.pack(fill="both", expand=True, padx=10, pady=10)
        
        alert_frame = ctk.CTkFrame(scroll, fg_color="#fff5f5", border_width=1.5, border_color="#feb2b2", corner_radius=8)
        alert_frame.pack(fill="x", pady=10)
        
        ctk.CTkLabel(alert_frame, text="🚨 Compliance & Expiry Alert Radar (30-Day Threshold for Vehicles & Drivers)", font=self.font_sub, text_color="#c53030").pack(anchor="w", padx=15, pady=(10, 5))
        
        self.alert_tree = ttk.Treeview(alert_frame, columns=("Asset", "Type", "Expiry_Date", "Status"), show="headings", height=5)
        self.alert_tree.heading("Asset", text="Asset / Vehicle / Driver")
        self.alert_tree.heading("Type", text="Compliance Document")
        self.alert_tree.heading("Expiry_Date", text="Expiry Date")
        self.alert_tree.heading("Status", text="Urgency Status")
        self.alert_tree.column("Asset", width=180, anchor="w")
        self.alert_tree.column("Type", width=180, anchor="w")
        self.alert_tree.column("Expiry_Date", width=120, anchor="center")
        self.alert_tree.column("Status", width=160, anchor="center")
        
        self.alert_tree.tag_configure("RED", foreground="#dc2626", font=('Helvetica', 13, 'bold'))
        self.alert_tree.tag_configure("YELLOW", foreground="#d97706", font=('Helvetica', 13, 'bold'))
        self.alert_tree.tag_configure("BLUE", foreground="#2563eb", font=('Helvetica', 13, 'bold'))
        
        self.alert_tree.pack(fill="both", expand=True, padx=12, pady=10)
        
        frame.refresh_data = self.load_dashboard_data

    def create_kpi_card(self, parent, title, val_str, color_hex, col_idx):
        card = ctk.CTkFrame(parent, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8, height=95)
        card.grid(row=0, column=col_idx, sticky="nsew", padx=6, pady=5)
        card.grid_propagate(False)
        
        ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=13, weight="bold"), text_color="#718096").pack(anchor="w", padx=12, pady=(10, 2))
        lbl_val = ctk.CTkLabel(card, text=val_str, font=ctk.CTkFont(size=20, weight="bold"), text_color=color_hex)
        lbl_val.pack(anchor="w", padx=12, pady=(0, 10))
        return lbl_val

    def load_dashboard_data(self):
        if not os.path.exists(DB_FILE):
            return
            
        try:
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_inv = excel_dict.get("Invoices", pd.DataFrame())
            df_rec = excel_dict.get("Receipts", pd.DataFrame())
            df_led = excel_dict.get("Customer_Ledger", pd.DataFrame())
            df_bk = excel_dict.get("Bookings", pd.DataFrame())
            df_veh = excel_dict.get("Vehicle_Master", pd.DataFrame())
            df_drv = excel_dict.get("Driver_Master", pd.DataFrame())
            
            total_rev = float(df_inv["Grand_Total"].sum()) if not df_inv.empty and "Grand_Total" in df_inv.columns else 0.0
            total_col = float(df_rec["Amount"].sum()) if not df_rec.empty and "Amount" in df_rec.columns else 0.0
            
            outstanding = 0.0
            if not df_led.empty and "Debit" in df_led.columns and "Credit" in df_led.columns:
                outstanding = float(df_led["Debit"].sum() - df_led["Credit"].sum())
                
            conf_bks = len(df_bk) if not df_bk.empty else 0
            
            self.card_rev.configure(text=f"₹{total_rev:,.2f}")
            self.card_rec.configure(text=f"₹{total_col:,.2f}")
            self.card_due.configure(text=f"₹{max(0.0, outstanding):,.2f}")
            self.card_bk.configure(text=str(conf_bks))
            
            for item in self.v_tree.get_children():
                self.v_tree.delete(item)
                
            fleet_summary = []
            if not df_veh.empty:
                for _, v_row in df_veh.iterrows():
                    v_reg = str(v_row.get("Reg_No", ""))
                    v_brand = str(v_row.get("Brand", ""))
                    v_type = str(v_row.get("Type", ""))
                    v_label = f"{v_brand} {v_type} ({v_reg})" if v_brand else v_reg
                    
                    trips_count = 0
                    total_km = 0.0
                    rev_sum = 0.0
                    
                    if not df_bk.empty and "Vehicle_Assigned" in df_bk.columns:
                        v_bks = df_bk[df_bk["Vehicle_Assigned"].astype(str).str.strip().str.lower() == v_reg.lower()]
                        trips_count = len(v_bks)
                        if not df_inv.empty:
                            v_invs = pd.merge(v_bks, df_inv, on="Inquiry_No", how="inner")
                            if not v_invs.empty:
                                total_km = float(v_invs["Actual_KM"].sum() if "Actual_KM" in v_invs.columns else 0.0)
                                rev_sum = float(v_invs["Grand_Total"].sum() if "Grand_Total" in v_invs.columns else 0.0)
                                
                    fleet_summary.append((v_label, trips_count, f"{total_km:,.1f} KM", f"₹{rev_sum:,.2f}"))
                    
            for r in fleet_summary:
                self.v_tree.insert("", "end", values=r)
            
            for item in self.d_tree.get_children():
                self.d_tree.delete(item)
                
            drv_summary = []
            if not df_drv.empty:
                for _, d_row in df_drv.iterrows():
                    d_name = str(d_row.get("Name", ""))
                    d_bk_match = df_bk[df_bk["Driver_Assigned"].astype(str).str.strip().str.lower() == d_name.lower()] if not df_bk.empty else pd.DataFrame()
                    total_assigned = len(d_bk_match)
                    completed_trips = 0
                    total_km = 0.0
                    
                    if not d_bk_match.empty:
                        completed_trips = len(d_bk_match[d_bk_match["Journey_Status"].astype(str).str.lower() == "completed"])
                        if not df_inv.empty:
                            merged_d_inv = pd.merge(d_bk_match, df_inv, on="Inquiry_No", how="inner")
                            if "Actual_KM" in merged_d_inv.columns:
                                total_km = float(merged_d_inv["Actual_KM"].sum())
                                
                    drv_summary.append((d_name, completed_trips, f"{total_km:,.1f} KM", total_assigned))
                    
            for r in drv_summary:
                self.d_tree.insert("", "end", values=r)
                    
            for item in self.alert_tree.get_children():
                self.alert_tree.delete(item)
                
            today = datetime.date.today()
            alert_items = []
            
            if not df_veh.empty:
                for _, r in df_veh.iterrows():
                    reg = str(r.get("Reg_No", "Vehicle"))
                    for col_name in ["Reg_Expiry", "Insurance_Expiry", "Pollution_Expiry", "State_Permit_Expiry", "All_India_Permit_Expiry", "Fitness_Expiry"]:
                        if col_name in r:
                            exp_val = str(r.get(col_name, ""))[:10]
                            if exp_val and exp_val.lower() != "nan":
                                try:
                                    exp_dt = datetime.datetime.strptime(exp_val, "%Y-%m-%d").date()
                                    days_left = (exp_dt - today).days
                                    if days_left <= 30:
                                        if days_left < 0:
                                            status, tag = f"🔴 EXPIRED ({abs(days_left)}d ago)", "RED"
                                        elif days_left <= 7:
                                            status, tag = f"🔴 Due in {days_left} days", "RED"
                                        elif days_left <= 15:
                                            status, tag = f"🟠 Due in {days_left} days", "YELLOW"
                                        else:
                                            status, tag = f"🔵 Due in {days_left} days", "BLUE"
                                            
                                        alert_items.append((days_left, (reg, col_name.replace("_", " ").title(), format_date_to_ddmmyyyy(exp_dt), status), tag))
                                except:
                                    pass

            if not df_drv.empty:
                for _, r in df_drv.iterrows():
                    drv_name = str(r.get("Name", "Driver"))
                    dl_val = str(r.get("DL_Expiry", ""))[:10]
                    if dl_val and dl_val.lower() != "nan":
                        try:
                            exp_dt = datetime.datetime.strptime(dl_val, "%Y-%m-%d").date()
                            days_left = (exp_dt - today).days
                            if days_left <= 30:
                                if days_left < 0:
                                    status, tag = f"🔴 EXPIRED ({abs(days_left)}d ago)", "RED"
                                elif days_left <= 7:
                                    status, tag = f"🔴 Due in {days_left} days", "RED"
                                elif days_left <= 15:
                                    status, tag = f"🟠 Due in {days_left} days", "YELLOW"
                                else:
                                    status, tag = f"🔵 Due in {days_left} days", "BLUE"
                                    
                                alert_items.append((days_left, (drv_name, "Dl Expiry", format_date_to_ddmmyyyy(exp_dt), status), tag))
                        except:
                            pass
                                
            alert_items.sort(key=lambda x: x[0])
            for _, row_vals, tag_val in alert_items:
                self.alert_tree.insert("", "end", values=row_vals, tags=(tag_val,))
                
        except Exception as e:
            print("Dashboard data loading error:", e)

    # =========================================================================
    # 1. INQUIRY MANAGEMENT
    # =========================================================================
    def create_inquiry_view(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid(row=0, column=0, sticky="nsew")
        self.frames["inquiry"] = frame
        
        top_bar = ctk.CTkFrame(frame)
        top_bar.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkButton(top_bar, text="+ ADD", font=self.font_btn, width=80, height=38, fg_color="#007bff", command=lambda: self.open_inquiry_form_popup("ADD")).pack(side="left", padx=3)
        ctk.CTkButton(top_bar, text="✏️ MODIFY", font=self.font_btn, width=90, height=38, fg_color="#ffc107", text_color="black", command=lambda: self.open_inquiry_form_popup("MODIFY")).pack(side="left", padx=3)
        ctk.CTkButton(top_bar, text="📄 ESTIMATE PDF", font=self.font_btn, width=140, height=38, fg_color="#17a2b8", command=self.generate_estimate_pdf).pack(side="left", padx=3)
        
        status_bar = ctk.CTkFrame(top_bar, fg_color="transparent")
        status_bar.pack(side="left", padx=8)
        
        ctk.CTkLabel(status_bar, text="Set Status:", font=self.font_label).pack(side="left", padx=(0, 4))
        ctk.CTkButton(status_bar, text="🟢 Confirm", font=self.font_btn, width=85, height=38, fg_color="#28a745", hover_color="#218838", command=lambda: self.set_inquiry_quick_status("Confirmed")).pack(side="left", padx=2)
        ctk.CTkButton(status_bar, text="🟡 Pending", font=self.font_btn, width=85, height=38, fg_color="#e0a800", hover_color="#c69500", text_color="black", command=lambda: self.set_inquiry_quick_status("Pending")).pack(side="left", padx=2)
        ctk.CTkButton(status_bar, text="🔴 Cancel", font=self.font_btn, width=80, height=38, fg_color="#dc3545", hover_color="#c82333", command=lambda: self.set_inquiry_quick_status("Cancelled")).pack(side="left", padx=2)
        
        ctk.CTkButton(top_bar, text="🚪 QUIT", font=self.font_btn, width=70, height=38, fg_color="#6c757d", command=self.quit).pack(side="left", padx=3)
        
        self.inq_search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(top_bar, font=self.font_body, placeholder_text="🔍 Search Inquiry...", textvariable=self.inq_search_var, width=190, height=38)
        search_entry.pack(side="right", padx=10)
        self.inq_search_var.trace_add("write", lambda *args: self.load_inquiries())
        
        table_frame = ctk.CTkFrame(frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        columns = ("Inq_No", "Date", "Customer", "Phone", "Route", "Vehicle", "Est_Amount", "Deposit", "Status")
        self.inq_tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        for col in columns:
            self.inq_tree.heading(col, text=col)
            self.inq_tree.column(col, width=120, anchor="center")
            
        self.inq_tree.tag_configure("Confirmed", foreground="#16a34a", font=('Helvetica', 13, 'bold'))
        self.inq_tree.tag_configure("Pending", foreground="#d97706", font=('Helvetica', 13, 'bold'))
        self.inq_tree.tag_configure("Cancelled", foreground="#dc2626", font=('Helvetica', 13, 'bold'))
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.inq_tree.yview)
        self.inq_tree.configure(yscrollcommand=scrollbar.set)
        self.inq_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        frame.refresh_data = self.load_inquiries

    def load_inquiries(self):
        for item in self.inq_tree.get_children():
            self.inq_tree.delete(item)
        if os.path.exists(DB_FILE):
            try:
                df = pd.read_excel(DB_FILE, sheet_name="Inquiries")
                search_query = self.inq_search_var.get().lower()
                for _, row in df.iterrows():
                    row_str = " ".join(str(val) for val in row.values).lower()
                    if search_query in row_str:
                        st = str(row.get("Status", "Pending")).strip()
                        tag = st if st in ["Confirmed", "Pending", "Cancelled"] else "Pending"
                        self.inq_tree.insert("", "end", values=(
                            row.get("Inquiry_No", ""),
                            format_date_to_ddmmyyyy(row.get("Date", "")),
                            row.get("Customer_Name", ""),
                            row.get("Phone", ""),
                            f"{row.get('From_Station','')} -> {row.get('To_Station','')}",
                            row.get("Vehicle_Type", ""),
                            f"₹{float(row.get('Est_Amount', 0) or 0):,.2f}",
                            f"₹{float(row.get('Deposit_Amount', 0) or 0):,.2f}",
                            st
                        ), tags=(tag,))
            except Exception as e:
                print("Error loading inquiries:", e)

    def set_inquiry_quick_status(self, new_status):
        selected = self.inq_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an Inquiry from the table first!")
            return
        
        vals = self.inq_tree.item(selected[0])['values']
        inq_id = str(vals[0])
        bk_id = f"BK-{inq_id.replace('INQ-', '')}"
        
        excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
        df_inq = excel_dict.get("Inquiries", pd.DataFrame())
        df_bk = excel_dict.get("Bookings", pd.DataFrame())
        df_rec = excel_dict.get("Receipts", pd.DataFrame())
        
        idx = df_inq.index[df_inq["Inquiry_No"].astype(str) == inq_id].tolist()
        if idx:
            row_idx = idx[0]
            
            if not df_bk.empty and "Booking_ID" in df_bk.columns:
                b_match = df_bk[df_bk["Booking_ID"].astype(str) == str(bk_id)]
                if not b_match.empty and str(b_match.iloc[0].get("Journey_Status", "")).strip().lower() == "completed":
                    if new_status in ["Pending", "Cancelled"]:
                        messagebox.showerror(
                            "Fraud Prevention Warning",
                            f"Error: Journey for Inquiry {inq_id} is already marked as 'Completed'.\n\n"
                            "Completed journeys cannot be set to Pending or Cancelled."
                        )
                        return

            df_inq.at[row_idx, "Status"] = new_status
            
            if new_status == "Confirmed":
                total_adv = self.calculate_total_advance_before_invoice(inq_id, df_inq, df_rec)
                already_exists = False
                if not df_bk.empty and "Booking_ID" in df_bk.columns:
                    already_exists = not df_bk[df_bk["Booking_ID"].astype(str) == str(bk_id)].empty
                
                if not already_exists:
                    new_bk_row = {
                        "Booking_ID": bk_id,
                        "Inquiry_No": inq_id,
                        "Customer_Name": df_inq.at[row_idx, "Customer_Name"],
                        "Vehicle_Assigned": df_inq.at[row_idx, "Vehicle_Type"],
                        "Driver_Assigned": df_inq.at[row_idx, "Driver_Expected"],
                        "Date_From": format_date_to_ddmmyyyy(df_inq.at[row_idx, "Date_From"]),
                        "Date_To": format_date_to_ddmmyyyy(df_inq.at[row_idx, "Date_To"]),
                        "Advance_Amount": total_adv,
                        "Journey_Status": "Not Started",
                        "Passengers_Summary": "Adult: 0, Minor: 0, Male: 0, Female: 0"
                    }
                    df_bk = pd.concat([df_bk, pd.DataFrame([new_bk_row])], ignore_index=True)
            else:
                if not df_bk.empty and "Booking_ID" in df_bk.columns:
                    df_bk = df_bk[df_bk["Booking_ID"].astype(str) != str(bk_id)]

            excel_dict["Bookings"] = df_bk
            excel_dict["Inquiries"] = df_inq
            
            with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                for s_name, s_df in excel_dict.items():
                    s_df.to_excel(writer, sheet_name=s_name, index=False)
                    
            messagebox.showinfo("Status Updated", f"Inquiry {inq_id} status updated to '{new_status}'.")
            self.load_inquiries()
            self.load_bookings()

    def open_inquiry_form_popup(self, mode="ADD"):
        existing_data = None
        target_inq_id = None
        
        if mode == "MODIFY":
            selected = self.inq_tree.selection()
            if not selected:
                messagebox.showwarning("Warning", "Please select an Inquiry to modify.")
                return
            target_inq_id = str(self.inq_tree.item(selected[0])['values'][0])
            df = pd.read_excel(DB_FILE, sheet_name="Inquiries")
            matching = df[df["Inquiry_No"].astype(str) == target_inq_id]
            if not matching.empty:
                existing_data = matching.iloc[0]

        popup = ctk.CTkToplevel(self)
        popup.title(f"{mode} Inquiry Details - Travel Mate")
        popup.geometry("900x840")
        popup.grab_set()
        
        scroll = ctk.CTkScrollableFrame(popup)
        scroll.pack(fill="both", expand=True, padx=15, pady=15)
        
        inq_id = target_inq_id if mode == "MODIFY" else f"INQ-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"
        bk_id = f"BK-{inq_id.replace('INQ-', '')}"
        
        hdr_frame = ctk.CTkFrame(scroll, fg_color="#1a365d", corner_radius=8)
        hdr_frame.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(hdr_frame, text=f"📋 INQUIRY SPECIFICATION SHEET: {inq_id}", font=self.font_header, text_color="#ffffff").pack(side="left", padx=15, pady=10)
        
        sec1 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        sec1.pack(fill="x", pady=6, padx=2)
        ctk.CTkLabel(sec1, text="👤 Customer Profile & Contact Information", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        grid1 = ctk.CTkFrame(sec1, fg_color="transparent")
        grid1.pack(fill="x", padx=10, pady=(0, 10))
        grid1.grid_columnconfigure((0, 1), weight=1)
        
        default_inq_date = format_date_to_ddmmyyyy(existing_data.get("Date", datetime.date.today()) if existing_data is not None else datetime.date.today())
        e_date = self.create_date_field(grid1, "Date of Inquiry (DD/MM/YYYY):", row=0, col=0, default=default_inq_date)
        e_name = self.create_grid_field(grid1, "Customer Full Name:", row=0, col=1, default=str(existing_data.get("Customer_Name", "")) if existing_data is not None else "")
        e_phone = self.create_grid_field(grid1, "Phone Number:", row=1, col=0, default=str(existing_data.get("Phone", "")) if existing_data is not None else "")
        e_whatsapp = self.create_grid_field(grid1, "WhatsApp Number:", row=1, col=1, default=str(existing_data.get("Whatsapp", "")) if existing_data is not None else "")
        e_email = self.create_grid_field(grid1, "Email ID:", row=2, col=0, default=str(existing_data.get("Email", "")) if existing_data is not None else "")
        e_gstn = self.create_grid_field(grid1, "PAN / GSTIN (Format: 11AAAAA1111A1AO):", row=2, col=1, default=str(existing_data.get("PAN_GSTN", "")) if existing_data is not None else "")
        e_address = self.create_grid_field(grid1, "Official Address:", row=3, col=0, default=str(existing_data.get("Address", "")) if existing_data is not None else "")
        e_city = self.create_grid_field(grid1, "City / State / PIN:", row=3, col=1, default=str(existing_data.get("City", "")) if existing_data is not None else "")

        sec2 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        sec2.pack(fill="x", pady=6, padx=2)
        ctk.CTkLabel(sec2, text="🗺️ Itinerary, Dates & Vehicle Allocation", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        grid2 = ctk.CTkFrame(sec2, fg_color="transparent")
        grid2.pack(fill="x", padx=10, pady=(0, 10))
        grid2.grid_columnconfigure((0, 1), weight=1)
        
        default_from_date = format_date_to_ddmmyyyy(existing_data.get("Date_From", datetime.date.today()) if existing_data is not None else datetime.date.today())
        default_to_date = format_date_to_ddmmyyyy(existing_data.get("Date_To", datetime.date.today()) if existing_data is not None else datetime.date.today())
        
        e_from_date = self.create_date_field(grid2, "Journey Date From (DD/MM/YYYY):", row=0, col=0, default=default_from_date)
        e_to_date = self.create_date_field(grid2, "Journey Date To (DD/MM/YYYY):", row=0, col=1, default=default_to_date)
        e_from_station = self.create_grid_field(grid2, "From Station (Pickup Point):", row=1, col=0, default=str(existing_data.get("From_Station", "")) if existing_data is not None else "")
        e_to_station = self.create_grid_field(grid2, "To Station (Drop Destination):", row=1, col=1, default=str(existing_data.get("To_Station", "")) if existing_data is not None else "")
        
        veh_options = self.get_master_options("Vehicle_Master", "Reg_No") or ["Toyota Innova (DL01AB1234)", "Maruti Dzire (DL01CD5678)"]
        drv_options = self.get_master_options("Driver_Master", "Name") or ["Ramesh Kumar", "Suresh Singh"]
        
        f_v = ctk.CTkFrame(grid2, fg_color="transparent")
        f_v.grid(row=2, column=0, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(f_v, text="Vehicle Registration / Type:", font=self.font_label).pack(anchor="w")
        cb_veh = ctk.CTkComboBox(f_v, font=self.font_body, values=veh_options, height=38)
        if existing_data is not None and str(existing_data.get("Vehicle_Type", "")) in veh_options:
            cb_veh.set(str(existing_data.get("Vehicle_Type", "")))
        cb_veh.pack(fill="x")
        
        f_d = ctk.CTkFrame(grid2, fg_color="transparent")
        f_d.grid(row=2, column=1, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(f_d, text="Driver Expected:", font=self.font_label).pack(anchor="w")
        cb_drv = ctk.CTkComboBox(f_d, font=self.font_body, values=drv_options, height=38)
        if existing_data is not None and str(existing_data.get("Driver_Expected", "")) in drv_options:
            cb_drv.set(str(existing_data.get("Driver_Expected", "")))
        cb_drv.pack(fill="x")

        sec3 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        sec3.pack(fill="x", pady=6, padx=2)
        ctk.CTkLabel(sec3, text="💵 Fare Structure, Allowances & Advances", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        grid3 = ctk.CTkFrame(sec3, fg_color="transparent")
        grid3.pack(fill="x", padx=10, pady=(0, 10))
        grid3.grid_columnconfigure((0, 1, 2), weight=1)
        
        e_est_km = self.create_grid_field(grid3, "Estimated KM:", row=0, col=0, default=str(existing_data.get("Est_KM", 100)) if existing_data is not None else "100")
        e_est_amt = self.create_grid_field(grid3, "Estimated Base Amount (₹):", row=0, col=1, default=str(existing_data.get("Est_Amount", 2500)) if existing_data is not None else "2500")
        e_rate_extra_km = self.create_grid_field(grid3, "Rate/Extra KM (₹):", row=0, col=2, default=str(existing_data.get("Rate_Per_Extra_KM", 12)) if existing_data is not None else "12")
        
        e_night = self.create_grid_field(grid3, "Night Charges (₹):", row=1, col=0, default=str(existing_data.get("Night_Charges", 300)) if existing_data is not None else "300")
        e_toll = self.create_grid_field(grid3, "Toll Charges (₹):", row=1, col=1, default=str(existing_data.get("Toll_Charges", 0)) if existing_data is not None else "0")
        e_parking = self.create_grid_field(grid3, "Parking Charges (₹):", row=1, col=2, default=str(existing_data.get("Parking_Charges", 0)) if existing_data is not None else "0")
        
        e_deposit = self.create_grid_field(grid3, "Deposit Received (₹):", row=2, col=0, default=str(existing_data.get("Deposit_Amount", 0)) if existing_data is not None else "0")
        e_other = self.create_grid_field(grid3, "Other Charges (₹):", row=2, col=1, default=str(existing_data.get("Other_Charges", 0)) if existing_data is not None else "0")
        
        f_st = ctk.CTkFrame(grid3, fg_color="transparent")
        f_st.grid(row=2, column=2, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(f_st, text="Inquiry Status:", font=self.font_label).pack(anchor="w")
        cb_status = ctk.CTkComboBox(f_st, font=self.font_body, values=["Pending", "Confirmed", "Cancelled"], height=38)
        status_val = str(existing_data.get("Status", "Pending")) if existing_data is not None else "Pending"
        cb_status.set(status_val)
        cb_status.pack(fill="x")

        e_terms = self.create_grid_field(grid3, "Terms & Special Instructions:", row=3, col=0, default=str(existing_data.get("Terms", "Standard terms apply")) if existing_data is not None else "Standard terms apply")
        grid3.grid_slaves(row=3, column=0)[0].grid(columnspan=3)

        def save_record():
            if not e_name.get().strip():
                messagebox.showerror("Error", "Customer Name is required!")
                return
            
            input_gstn = e_gstn.get().strip().upper()
            if input_gstn and not is_valid_gstn(input_gstn):
                messagebox.showwarning(
                    "Invalid GSTN Structure",
                    f"The entered GSTIN '{input_gstn}' is invalid!\n\n"
                    "Required Structure: 11AAAAA1111A1AO\n"
                    "Example: 09AAAAA0000A1Z5"
                )
                return
            
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_inq = excel_dict.get("Inquiries", pd.DataFrame())
            df_bk = excel_dict.get("Bookings", pd.DataFrame())
            df_led = excel_dict.get("Customer_Ledger", pd.DataFrame())
            df_rec = excel_dict.get("Receipts", pd.DataFrame())
            
            if mode == "MODIFY" and not df_bk.empty and "Booking_ID" in df_bk.columns:
                b_match = df_bk[df_bk["Booking_ID"].astype(str) == str(bk_id)]
                if not b_match.empty and str(b_match.iloc[0].get("Journey_Status", "")).strip().lower() == "completed":
                    if cb_status.get() in ["Pending", "Cancelled"]:
                        messagebox.showerror(
                            "Fraud Prevention Warning",
                            f"Error: Journey for Inquiry {inq_id} is already marked as 'Completed'.\n\n"
                            "Completed journeys cannot be set to Pending or Cancelled."
                        )
                        return

            row_dict = {
                "Inquiry_No": inq_id,
                "Date": parse_date_to_yyyymmdd(e_date.get()),
                "Customer_Name": e_name.get(),
                "PAN_GSTN": input_gstn,
                "Address": e_address.get(),
                "City": e_city.get(),
                "PIN": "",
                "State": "",
                "Phone": e_phone.get(),
                "Whatsapp": e_whatsapp.get(),
                "Email": e_email.get(),
                "Date_From": parse_date_to_yyyymmdd(e_from_date.get()),
                "Date_To": parse_date_to_yyyymmdd(e_to_date.get()),
                "From_Station": e_from_station.get(),
                "To_Station": e_to_station.get(),
                "Vehicle_Type": cb_veh.get(),
                "Driver_Expected": cb_drv.get(),
                "Est_KM": float(e_est_km.get() or 0),
                "Est_Amount": float(e_est_amt.get() or 0),
                "Rate_Per_Extra_KM": float(e_rate_extra_km.get() or 0),
                "Night_Charges": float(e_night.get() or 0),
                "Toll_Charges": float(e_toll.get() or 0),
                "Parking_Charges": float(e_parking.get() or 0),
                "Other_Charges": float(e_other.get() or 0),
                "Terms": e_terms.get(),
                "Deposit_Amount": float(e_deposit.get() or 0),
                "Status": cb_status.get()
            }
            
            if mode == "MODIFY" and not df_inq.empty and "Inquiry_No" in df_inq.columns:
                df_inq = df_inq[df_inq["Inquiry_No"].astype(str) != str(inq_id)]
                
            df_inq = pd.concat([df_inq, pd.DataFrame([row_dict])], ignore_index=True)
            excel_dict["Inquiries"] = df_inq
            
            if cb_status.get() == "Confirmed":
                total_adv = self.calculate_total_advance_before_invoice(inq_id, df_inq, df_rec)
                if not df_bk.empty and "Booking_ID" in df_bk.columns:
                    df_bk = df_bk[df_bk["Booking_ID"].astype(str) != str(bk_id)]
                
                new_bk_row = {
                    "Booking_ID": bk_id,
                    "Inquiry_No": inq_id,
                    "Customer_Name": e_name.get(),
                    "Vehicle_Assigned": cb_veh.get(),
                    "Driver_Assigned": cb_drv.get(),
                    "Date_From": format_date_to_ddmmyyyy(e_from_date.get()),
                    "Date_To": format_date_to_ddmmyyyy(e_to_date.get()),
                    "Advance_Amount": total_adv,
                    "Journey_Status": "Not Started",
                    "Passengers_Summary": "Adult: 0, Minor: 0, Male: 0, Female: 0"
                }
                df_bk = pd.concat([df_bk, pd.DataFrame([new_bk_row])], ignore_index=True)
            else:
                if not df_bk.empty and "Booking_ID" in df_bk.columns:
                    df_bk = df_bk[df_bk["Booking_ID"].astype(str) != str(bk_id)]
                    
            excel_dict["Bookings"] = df_bk
            
            dep_val = float(e_deposit.get() or 0)
            if dep_val > 0 and mode == "ADD":
                new_led = {
                    "Date": parse_date_to_yyyymmdd(e_date.get()),
                    "Inquiry_No": inq_id,
                    "Customer_Name": e_name.get(),
                    "Particulars": "Inquiry Advance Deposit Received",
                    "Voucher_Type": "Receipt",
                    "Debit": 0,
                    "Credit": dep_val,
                    "Balance": -dep_val
                }
                df_led = pd.concat([df_led, pd.DataFrame([new_led])], ignore_index=True)
                excel_dict["Customer_Ledger"] = df_led
                
                rec_id = f"REC-{datetime.datetime.now().strftime('%Y%m%d%H%M')}"
                new_rec = {
                    "Receipt_No": rec_id,
                    "Date": parse_date_to_yyyymmdd(e_date.get()),
                    "Inquiry_No": inq_id,
                    "Customer_Name": e_name.get(),
                    "Payment_Type": "Advance / Inquiry Deposit",
                    "Payment_Mode": "Cash/Bank",
                    "Ref_No": "DEP-INIT",
                    "Amount": dep_val,
                    "Remarks": "Advance collected at inquiry"
                }
                df_rec = pd.concat([df_rec, pd.DataFrame([new_rec])], ignore_index=True)
                excel_dict["Receipts"] = df_rec
                
            with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                for s_name, s_df in excel_dict.items():
                    s_df.to_excel(writer, sheet_name=s_name, index=False)
                    
            messagebox.showinfo("Success", f"Inquiry {inq_id} saved successfully!")
            popup.destroy()
            self.load_inquiries()
            self.load_bookings()

        ctk.CTkButton(scroll, text="💾 SAVE / UPDATE INQUIRY SPECIFICATION", font=self.font_btn, fg_color="#28a745", height=44, command=save_record).pack(fill="x", pady=15)

    def create_grid_field(self, parent, label_text, row, col, default=""):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=col, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(frame, text=label_text, font=self.font_label).pack(anchor="w")
        entry = ctk.CTkEntry(frame, font=self.font_body, height=38)
        entry.insert(0, default)
        entry.pack(fill="x")
        return entry

    def create_date_field(self, parent, label_text, row, col=0, default=""):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        if label_text:
            if hasattr(frame, 'grid') and row is not None and col is not None:
                frame.grid(row=row, column=col, sticky="ew", padx=6, pady=4)
            else:
                frame.pack(fill="x", padx=2, pady=4)
            ctk.CTkLabel(frame, text=label_text, font=self.font_label).pack(anchor="w")
        else:
            frame.pack(fill="x", padx=2, pady=4)
            
        sub_f = ctk.CTkFrame(frame, fg_color="transparent")
        sub_f.pack(fill="x", expand=True)
        
        entry = ctk.CTkEntry(sub_f, font=self.font_body, height=38)
        entry.insert(0, default)
        entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        
        def open_picker():
            DatePickerPopup(self, lambda d_str: (entry.delete(0, 'end'), entry.insert(0, d_str)))
            
        ctk.CTkButton(sub_f, text="📅", width=42, height=38, fg_color="#17a2b8", command=open_picker).pack(side="right")
        return entry

    def generate_estimate_pdf(self):
        selected = self.inq_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select an Inquiry from the list.")
            return
            
        inq_id = str(self.inq_tree.item(selected[0])['values'][0])
        df = pd.read_excel(DB_FILE, sheet_name="Inquiries")
        df_comp = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
        
        row = df[df["Inquiry_No"].astype(str) == inq_id].iloc[0]
        comp = df_comp.iloc[0] if not df_comp.empty else {}
        
        pdf_path = os.path.join(INVOICE_DIR, f"Estimate_{inq_id}.pdf")
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []
        
        elements.append(Paragraph(f"<b><font size=16 color='#1a365d'>{comp.get('Org_Name', 'TRAVEL MATE')}</font></b>", styles['Normal']))
        elements.append(Paragraph(f"<font size=9 color='#4a5568'>{comp.get('Address', '')} | Phone: {comp.get('Mobile', '')} | GSTIN: {comp.get('GSTN', '')}</font>", styles['Normal']))
        elements.append(Spacer(1, 8))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1a365d'), spaceBefore=2, spaceAfter=10))
        
        elements.append(Paragraph("<b><font size=13 color='#2d3748'>VEHICLE BOOKING ESTIMATE / QUOTATION</font></b>", styles['Normal']))
        elements.append(Spacer(1, 8))
        
        data = [
            ["Estimate No:", str(row['Inquiry_No']), "Date:", format_date_to_ddmmyyyy(row['Date'])],
            ["Customer Name:", str(row['Customer_Name']), "Contact Phone:", str(row['Phone'])],
            ["Journey Route:", f"{row['From_Station']} to {row['To_Station']}", "Vehicle Type:", str(row['Vehicle_Type'])],
            ["Estimated KM:", f"{row['Est_KM']} KM", "Estimated Amount:", f"INR {row['Est_Amount']:,.2f}"],
            ["Extra KM Rate:", f"INR {row['Rate_Per_Extra_KM']}/KM", "Night Charges:", f"INR {row['Night_Charges']:,.2f}"],
            ["Deposit Received:", f"INR {row['Deposit_Amount']:,.2f}", "Current Status:", str(row['Status'])]
        ]
        
        t = Table(data, colWidths=[130, 150, 110, 145])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f7fafc')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e0')),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 9.5),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        
        elements.append(t)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"<b>Terms & Conditions:</b> {row.get('Terms', 'Standard terms apply.')}", styles['Normal']))
        elements.append(Spacer(1, 35))
        elements.append(Paragraph("Authorized Signatory: ____________________________", styles['Normal']))
        
        doc.build(elements)
        messagebox.showinfo("PDF Generated", f"Estimate PDF generated at:\n{pdf_path}")
        if os.name == 'nt':
            os.startfile(pdf_path)

    # =========================================================================
    # 2. BOOKING MANAGEMENT
    # =========================================================================
    def create_booking_view(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid(row=0, column=0, sticky="nsew")
        self.frames["booking"] = frame
        
        top_bar = ctk.CTkFrame(frame)
        top_bar.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkButton(top_bar, text="👥 MANAGE PASSENGERS & IMPORT", font=self.font_btn, width=240, height=38, fg_color="#007bff", command=self.manage_passengers_popup).pack(side="left", padx=4)
        ctk.CTkButton(top_bar, text="✅ MARK COMPLETED", font=self.font_btn, width=160, height=38, fg_color="#28a745", command=self.mark_journey_completed).pack(side="left", padx=4)
        ctk.CTkButton(top_bar, text="🚪 QUIT", font=self.font_btn, width=70, height=38, fg_color="#6c757d", command=self.quit).pack(side="left", padx=4)
        
        self.bk_search_var = tk.StringVar()
        search_entry = ctk.CTkEntry(top_bar, font=self.font_body, placeholder_text="🔍 Search Booking...", textvariable=self.bk_search_var, width=190, height=38)
        search_entry.pack(side="right", padx=10)
        self.bk_search_var.trace_add("write", lambda *args: self.load_bookings())
        
        table_frame = ctk.CTkFrame(frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        cols = ("Booking_ID", "Inq_No", "Customer", "Vehicle", "Driver", "Date_From", "Advance", "Journey_Status", "Pax_Summary")
        self.bk_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            self.bk_tree.heading(col, text=col)
            self.bk_tree.column(col, width=120, anchor="center")
            
        self.bk_tree.pack(side="left", fill="both", expand=True)
        frame.refresh_data = self.load_bookings

    def load_bookings(self):
        self.sync_all_confirmed_inquiries_to_bookings()
        
        for item in self.bk_tree.get_children():
            self.bk_tree.delete(item)
        if os.path.exists(DB_FILE):
            try:
                df = pd.read_excel(DB_FILE, sheet_name="Bookings")
                query = self.bk_search_var.get().lower()
                for _, r in df.iterrows():
                    row_str = " ".join(str(val) for val in r.values).lower()
                    if query in row_str:
                        self.bk_tree.insert("", "end", values=(
                            r.get("Booking_ID",""), r.get("Inquiry_No",""), r.get("Customer_Name",""),
                            r.get("Vehicle_Assigned",""), r.get("Driver_Assigned",""), format_date_to_ddmmyyyy(r.get("Date_From","")),
                            f"₹{float(r.get('Advance_Amount', 0) or 0):,.2f}", r.get("Journey_Status","Not Started"),
                            r.get("Passengers_Summary","")
                        ))
            except Exception as e:
                print("Error loading bookings:", e)

    def manage_passengers_popup(self):
        selected = self.bk_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a Booking from the list.")
            return
        bk_id = str(self.bk_tree.item(selected[0])['values'][0])
        
        popup = ctk.CTkToplevel(self)
        popup.title(f"Passenger Management - {bk_id}")
        popup.geometry("860x660")
        popup.grab_set()
        
        top_pax = ctk.CTkFrame(popup)
        top_pax.pack(fill="x", padx=15, pady=10)
        
        ctk.CTkLabel(top_pax, text=f"Booking ID: {bk_id}", font=self.font_header).pack(side="left", padx=5)
        
        lbl_summary = ctk.CTkLabel(popup, text="Summary: Loading...", font=self.font_sub, text_color="#17a2b8")
        lbl_summary.pack(anchor="w", padx=20, pady=2)
        
        table_pax_frame = ctk.CTkFrame(popup)
        table_pax_frame.pack(fill="both", expand=True, padx=15, pady=5)
        
        pax_cols = ("Name", "Age", "Gender", "Contact")
        pax_tree = ttk.Treeview(table_pax_frame, columns=pax_cols, show="headings", selectmode="browse")
        for c in pax_cols:
            pax_tree.heading(c, text=c)
            pax_tree.column(c, width=130, anchor="center")
            
        pax_tree.pack(side="left", fill="both", expand=True)
        
        def load_current_pax():
            for item in pax_tree.get_children():
                pax_tree.delete(item)
            df_pax = pd.read_excel(DB_FILE, sheet_name="Booking_Passengers")
            match = df_pax[df_pax["Booking_ID"].astype(str) == str(bk_id)]
            
            adults, minors, male, female = 0, 0, 0, 0
            for _, r in match.iterrows():
                pax_tree.insert("", "end", values=(r.get("Passenger_Name",""), r.get("Age",""), r.get("Gender",""), r.get("Contact_No","")))
                try:
                    age = int(r.get("Age", 0))
                    if age >= 18:
                        adults += 1
                    else:
                        minors += 1
                except:
                    adults += 1
                    
                g = str(r.get("Gender","")).strip().lower()
                if "m" in g and "f" not in g:
                    male += 1
                elif "f" in g:
                    female += 1
                    
            summary_txt = f"Summary: Total: {len(match)} | Adults: {adults}, Minors: {minors} | Male: {male}, Female: {female}"
            lbl_summary.configure(text=summary_txt)
            
            wb = load_workbook(DB_FILE)
            ws_bk = wb["Bookings"]
            for row in ws_bk.iter_rows(min_row=2):
                if str(row[0].value) == str(bk_id):
                    row[9].value = f"Adult: {adults}, Minor: {minors}, M: {male}, F: {female}"
                    break
            wb.save(DB_FILE)
            self.load_bookings()

        load_current_pax()

        add_box = ctk.CTkFrame(popup)
        add_box.pack(fill="x", padx=15, pady=10)
        
        e_pname = ctk.CTkEntry(add_box, font=self.font_body, placeholder_text="Passenger Name", width=170, height=38)
        e_pname.pack(side="left", padx=4)
        e_page = ctk.CTkEntry(add_box, font=self.font_body, placeholder_text="Age", width=70, height=38)
        e_page.pack(side="left", padx=4)
        cb_gender = ctk.CTkComboBox(add_box, font=self.font_body, values=["Male", "Female", "Other"], width=110, height=38)
        cb_gender.set("Male")
        cb_gender.pack(side="left", padx=4)
        e_pcontact = ctk.CTkEntry(add_box, font=self.font_body, placeholder_text="Contact No", width=130, height=38)
        e_pcontact.pack(side="left", padx=4)
        
        def add_single_pax():
            if not e_pname.get().strip():
                messagebox.showerror("Error", "Passenger Name is required.")
                return
            wb = load_workbook(DB_FILE)
            ws_pax = wb["Booking_Passengers"]
            ws_pax.append([bk_id, e_pname.get(), e_page.get(), cb_gender.get(), e_pcontact.get()])
            wb.save(DB_FILE)
            e_pname.delete(0, 'end')
            e_page.delete(0, 'end')
            e_pcontact.delete(0, 'end')
            load_current_pax()
            
        ctk.CTkButton(add_box, text="+ Add Pax", font=self.font_btn, width=95, height=38, fg_color="#28a745", command=add_single_pax).pack(side="left", padx=6)
        
        def download_excel_template():
            save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel file", "*.xlsx")], initialfile="Passenger_List_Template.xlsx")
            if not save_path:
                return
            try:
                sample_data = {
                    "Passenger_Name": ["Rahul Sharma", "Pooja Sharma", "Aarav Sharma"],
                    "Age": [34, 30, 8],
                    "Gender": ["Male", "Female", "Male"],
                    "Contact_No": ["9876543210", "9876543211", "9876543212"]
                }
                template_df = pd.DataFrame(sample_data)
                template_df.to_excel(save_path, index=False)
                messagebox.showinfo("Success", f"Template saved successfully:\n{save_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to save template: {str(e)}")

        def import_from_excel():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
            if not file_path:
                return
            try:
                imp_df = pd.read_excel(file_path)
                wb = load_workbook(DB_FILE)
                ws_pax = wb["Booking_Passengers"]
                
                cols = {str(c).strip().lower(): c for c in imp_df.columns}
                name_col = cols.get('passenger_name') or cols.get('name') or cols.get('passenger name') or list(imp_df.columns)[0]
                age_col = cols.get('age') or (list(imp_df.columns)[1] if len(imp_df.columns)>1 else None)
                gen_col = cols.get('gender') or (list(imp_df.columns)[2] if len(imp_df.columns)>2 else None)
                con_col = cols.get('contact_no') or cols.get('contact') or cols.get('phone') or (list(imp_df.columns)[3] if len(imp_df.columns)>3 else None)
                
                count = 0
                for _, row in imp_df.iterrows():
                    p_name = str(row[name_col]) if name_col and pd.notna(row[name_col]) else ""
                    if p_name:
                        p_age = str(row[age_col]) if age_col and pd.notna(row[age_col]) else ""
                        p_gen = str(row[gen_col]) if gen_col and pd.notna(row[gen_col]) else "Male"
                        p_con = str(row[con_col]) if con_col and pd.notna(row[con_col]) else ""
                        ws_pax.append([bk_id, p_name, p_age, p_gen, p_con])
                        count += 1
                        
                wb.save(DB_FILE)
                messagebox.showinfo("Success", f"{count} Passengers imported successfully from Excel!")
                load_current_pax()
            except Exception as e:
                messagebox.showerror("Error", f"Excel import failed: {str(e)}")

        btn_bar = ctk.CTkFrame(popup, fg_color="transparent")
        btn_bar.pack(fill="x", padx=15, pady=5)
        
        ctk.CTkButton(btn_bar, text="📥 Download Template (.xlsx)", font=self.font_btn, fg_color="#6f42c1", hover_color="#59359a", height=38, command=download_excel_template).pack(side="left", padx=4)
        ctk.CTkButton(btn_bar, text="📤 Upload / Import Excel List", font=self.font_btn, fg_color="#17a2b8", hover_color="#138496", height=38, command=import_from_excel).pack(side="left", padx=4)
        
        def delete_selected_pax():
            sel = pax_tree.selection()
            if not sel:
                messagebox.showwarning("Warning", "Please select a passenger first.")
                return
            pname = pax_tree.item(sel[0])['values'][0]
            
            wb = load_workbook(DB_FILE)
            ws_pax = wb["Booking_Passengers"]
            for r in range(2, ws_pax.max_row + 1):
                if str(ws_pax.cell(row=r, column=1).value) == str(bk_id) and str(ws_pax.cell(row=r, column=2).value) == str(pname):
                    ws_pax.delete_rows(r)
                    break
            wb.save(DB_FILE)
            load_current_pax()
            
        ctk.CTkButton(btn_bar, text="🗑️ Delete Selected Pax", font=self.font_btn, fg_color="#dc3545", height=38, command=delete_selected_pax).pack(side="left", padx=4)
        ctk.CTkButton(btn_bar, text="Done / Close", font=self.font_btn, fg_color="#6c757d", height=38, command=popup.destroy).pack(side="right", padx=4)

    def mark_journey_completed(self):
        selected = self.bk_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a Booking.")
            return
        bk_id = str(self.bk_tree.item(selected[0])['values'][0])
        wb = load_workbook(DB_FILE)
        ws_bk = wb["Bookings"]
        for row in ws_bk.iter_rows(min_row=2):
            if str(row[0].value) == str(bk_id):
                row[8].value = "Completed"
                break
        wb.save(DB_FILE)
        messagebox.showinfo("Success", "Journey Status updated to 'Completed'. You can now generate the Tax Invoice in Accounts.")
        self.load_bookings()

    # =========================================================================
    # 3. ACCOUNTS & FINANCIAL MANAGEMENT
    # =========================================================================
    def create_accounts_view(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid(row=0, column=0, sticky="nsew")
        self.frames["accounts"] = frame
        
        self.acc_tabview = ctk.CTkTabview(frame)
        self.acc_tabview.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.tab_billing = self.acc_tabview.add("🧾 Invoicing & Duty Clearance")
        self.tab_receipts = self.acc_tabview.add("💵 Money Receipts (Advance / Final / Refund)")
        self.tab_ledger = self.acc_tabview.add("📒 Customer Ledger / Statement")
        
        self.setup_billing_tab()
        self.setup_receipts_tab()
        self.setup_ledger_tab()
        
        frame.refresh_data = self.refresh_accounts_view

    def refresh_accounts_view(self):
        self.load_billing_data()
        self.load_receipts_data()
        self.load_ledger_data()

    def setup_billing_tab(self):
        top_bar = ctk.CTkFrame(self.tab_billing)
        top_bar.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkButton(top_bar, text="🖨️ CREATE / EDIT TAX INVOICE", font=self.font_btn, width=230, height=38, fg_color="#dc3545", command=self.open_create_invoice_popup).pack(side="left", padx=4)
        ctk.CTkButton(top_bar, text="👁️ VIEW / REPRINT INVOICE", font=self.font_btn, width=190, height=38, fg_color="#17a2b8", command=self.reprint_selected_invoice).pack(side="left", padx=4)
        ctk.CTkButton(top_bar, text="🚪 QUIT", font=self.font_btn, width=70, height=38, fg_color="#6c757d", command=self.quit).pack(side="right", padx=4)
        
        table_frame = ctk.CTkFrame(self.tab_billing)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        cols = ("Booking_ID", "Inquiry_No", "Customer", "Vehicle", "Advance_Paid", "Invoice_Status", "Invoice_No")
        self.acc_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            self.acc_tree.heading(col, text=col)
            self.acc_tree.column(col, width=130, anchor="center")
            
        self.acc_tree.pack(side="left", fill="both", expand=True)
        self.load_billing_data()

    def load_billing_data(self):
        for item in self.acc_tree.get_children():
            self.acc_tree.delete(item)
        if os.path.exists(DB_FILE):
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_bk = excel_dict.get("Bookings", pd.DataFrame())
            df_inv = excel_dict.get("Invoices", pd.DataFrame())
            df_inq = excel_dict.get("Inquiries", pd.DataFrame())
            df_rec = excel_dict.get("Receipts", pd.DataFrame())
            
            inv_map = {}
            if not df_inv.empty and "Inquiry_No" in df_inv.columns:
                for _, r in df_inv.iterrows():
                    inv_map[str(r["Inquiry_No"])] = str(r["Invoice_No"])
                    
            completed = df_bk[df_bk["Journey_Status"] == "Completed"]
            for _, r in completed.iterrows():
                inq_id = str(r.get("Inquiry_No",""))
                has_inv = inq_id in inv_map
                inv_status = "Generated" if has_inv else "Pending"
                inv_no_val = inv_map.get(inq_id, "N/A")
                
                total_adv = self.calculate_total_advance_before_invoice(inq_id, df_inq, df_rec)
                
                self.acc_tree.insert("", "end", values=(
                    r.get("Booking_ID",""), inq_id, r.get("Customer_Name",""),
                    r.get("Vehicle_Assigned",""), f"₹{total_adv:,.2f}",
                    inv_status, inv_no_val
                ))

    def reprint_selected_invoice(self):
        selected = self.acc_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a row from the list.")
            return
            
        vals = self.acc_tree.item(selected[0])['values']
        inq_id = str(vals[1]) if len(vals) > 1 else ""
        
        df_inv = pd.read_excel(DB_FILE, sheet_name="Invoices")
        match_inv = df_inv[df_inv["Inquiry_No"].astype(str) == str(inq_id)]
        
        if match_inv.empty:
            messagebox.showwarning("Warning", "Invoice has not been generated yet for this trip.")
            return
            
        inv_no = str(match_inv.iloc[0]["Invoice_No"])
        pdf_path = os.path.join(INVOICE_DIR, f"{inv_no}.pdf")
        if os.path.exists(pdf_path):
            if os.name == 'nt':
                os.startfile(pdf_path)
        else:
            messagebox.showinfo("Notice", f"PDF file not found ({pdf_path}). Please regenerate the invoice.")

    def open_create_invoice_popup(self):
        selected = self.acc_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a completed trip to invoice.")
            return
        
        vals = self.acc_tree.item(selected[0])['values']
        bk_id = str(vals[0]) if len(vals) > 0 else ""
        inq_id = str(vals[1]) if len(vals) > 1 else ""
        cust_name = str(vals[2]) if len(vals) > 2 else ""
        
        excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
        df_inv = excel_dict.get("Invoices", pd.DataFrame())
        df_inq = excel_dict.get("Inquiries", pd.DataFrame())
        df_rec = excel_dict.get("Receipts", pd.DataFrame())
        
        match_inv = df_inv[df_inv["Inquiry_No"].astype(str) == str(inq_id)] if not df_inv.empty else pd.DataFrame()
        
        is_editing = not match_inv.empty
        inv_row = match_inv.iloc[0] if is_editing else None
        existing_inv_no = str(inv_row["Invoice_No"]) if is_editing else ""
        
        inq_match = df_inq[df_inq["Inquiry_No"].astype(str) == str(inq_id)]
        if inq_match.empty:
            messagebox.showerror("Error", f"Record for Inquiry {inq_id} not found!")
            return
            
        inq_row = inq_match.iloc[0]
        
        est_km = float(inq_row.get("Est_KM", 0) or 0)
        base_fare = float(inq_row.get("Est_Amount", 0) or 0)
        rate_extra_km = float(inq_row.get("Rate_Per_Extra_KM", 12) or 12)
        night_charges_est = float(inq_row.get("Night_Charges", 0) or 0)
        
        advance_paid = self.calculate_total_advance_before_invoice(inq_id, df_inq, df_rec)
        default_cust_gstn = str(inv_row.get("GSTN", inq_row.get("PAN_GSTN", ""))) if inv_row is not None else str(inq_row.get("PAN_GSTN", ""))
        
        popup = ctk.CTkToplevel(self)
        popup.title(f"{'Edit/Update' if is_editing else 'Create'} Tax Invoice - Travel Mate")
        popup.geometry("880x840")
        popup.grab_set()
        
        scroll = ctk.CTkScrollableFrame(popup)
        scroll.pack(fill="both", expand=True, padx=15, pady=15)
        
        inv_no = existing_inv_no if is_editing else f"INV-{datetime.datetime.now().strftime('%Y%m%d%H%M')}"
        
        hdr_frame = ctk.CTkFrame(scroll, fg_color="#1a365d", corner_radius=8)
        hdr_frame.pack(fill="x", pady=(0, 10))
        title_text = f"🧾 TAX INVOICE SPECIFICATION: {inv_no}" if not is_editing else f"✏️ UPDATE TAX INVOICE: {inv_no}"
        ctk.CTkLabel(hdr_frame, text=title_text, font=self.font_header, text_color="#ffffff").pack(side="left", padx=15, pady=10)
        
        ic1 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        ic1.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(ic1, text="👤 Client Billed Details & Booking Reference", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        ig1 = ctk.CTkFrame(ic1, fg_color="transparent")
        ig1.pack(fill="x", padx=10, pady=(0, 10))
        ig1.grid_columnconfigure((0, 1), weight=1)
        
        self.create_grid_label_val(ig1, "Customer Name:", cust_name, row=0, col=0)
        self.create_grid_label_val(ig1, "Booking Ref / Inquiry Ref:", f"{bk_id} / {inq_id}", row=0, col=1)
        
        e_cust_gstn = self.create_grid_field(ig1, "Customer GSTIN / Tax ID (Format: 11AAAAA1111A1AO):", row=1, col=0, default=default_cust_gstn)
        default_inv_date = format_date_to_ddmmyyyy(inv_row.get("Date", datetime.date.today()) if inv_row is not None else datetime.date.today())
        e_inv_date = self.create_date_field(ig1, "Invoice Date (DD/MM/YYYY):", row=1, col=1, default=default_inv_date)

        ic2 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        ic2.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(ic2, text="🚗 Distance, Mileage & Base Rate Metrics", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        ig2 = ctk.CTkFrame(ic2, fg_color="transparent")
        ig2.pack(fill="x", padx=10, pady=(0, 10))
        ig2.grid_columnconfigure((0, 1, 2), weight=1)
        
        def_act_km = str(inv_row.get("Actual_KM", int(est_km))) if inv_row is not None else str(int(est_km))
        def_base = str(inv_row.get("Base_Amount", base_fare)) if inv_row is not None else str(base_fare)
        
        e_km = self.create_grid_field(ig2, "Actual KM Covered:", row=0, col=0, default=def_act_km)
        e_base_fare = self.create_grid_field(ig2, "Base Estimated Fare (₹):", row=0, col=1, default=def_base)
        e_extra_rate = self.create_grid_field(ig2, "Extra KM Rate (₹/KM):", row=0, col=2, default=str(rate_extra_km))
        
        lbl_calc_extra = ctk.CTkLabel(ic2, text="Computation: Extra KM: 0 KM | Extra Charges: ₹0.00", font=self.font_label, text_color="#28a745")
        lbl_calc_extra.pack(anchor="w", padx=15, pady=(0, 8))
        
        def update_extra_calc(*args):
            try:
                act_km = float(e_km.get() or 0)
                diff = max(0.0, act_km - est_km)
                extra_cost = diff * float(e_extra_rate.get() or 0)
                lbl_calc_extra.configure(text=f"Computation: Est KM: {est_km:.1f} KM | Extra KM: {diff:.1f} KM | Extra Charges: ₹{extra_cost:,.2f}")
            except:
                pass
                
        e_km.bind("<KeyRelease>", update_extra_calc)
        e_extra_rate.bind("<KeyRelease>", update_extra_calc)
        update_extra_calc()

        ic3 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        ic3.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(ic3, text="💵 Incidental Charges, Night Allowances & Discounts", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        ig3 = ctk.CTkFrame(ic3, fg_color="transparent")
        ig3.pack(fill="x", padx=10, pady=(0, 10))
        ig3.grid_columnconfigure((0, 1, 2), weight=1)
        
        def_night = str(inv_row.get("Night_Charges", night_charges_est)) if inv_row is not None else str(night_charges_est)
        def_toll = str(inv_row.get("Toll_Charges", 0)) if inv_row is not None else "0"
        def_parking = str(inv_row.get("Parking_Charges", 0)) if inv_row is not None else "0"
        def_other = str(inv_row.get("Other_Charges", 0)) if inv_row is not None else "0"
        def_disc = str(inv_row.get("Discount", 0)) if inv_row is not None else "0"
        
        e_night = self.create_grid_field(ig3, "Night Charges (₹):", row=0, col=0, default=def_night)
        e_toll = self.create_grid_field(ig3, "Toll Charges (₹):", row=0, col=1, default=def_toll)
        e_parking = self.create_grid_field(ig3, "Parking Charges (₹):", row=0, col=2, default=def_parking)
        
        e_other = self.create_grid_field(ig3, "Other Charges (₹):", row=1, col=0, default=def_other)
        e_disc = self.create_grid_field(ig3, "Less Discount (₹):", row=1, col=1, default=def_disc)
        self.create_grid_label_val(ig3, "Total Advance Paid (Adjusted):", f"₹{advance_paid:,.2f}", row=1, col=2)

        ic4 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        ic4.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(ic4, text="🏛️ GST Tax Engine & Compliance", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        ig4 = ctk.CTkFrame(ic4, fg_color="transparent")
        ig4.pack(fill="x", padx=10, pady=(0, 10))
        ig4.grid_columnconfigure((0, 1), weight=1)
        
        f_gst = ctk.CTkFrame(ig4, fg_color="transparent")
        f_gst.grid(row=0, column=0, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(f_gst, text="GST Applicable Type:", font=self.font_label).pack(anchor="w")
        gst_type = ctk.CTkComboBox(f_gst, font=self.font_body, values=["CGST + SGST (Intra-State)", "IGST (Inter-State)"], height=38)
        gst_type.pack(fill="x")
        
        f_itc = ctk.CTkFrame(ig4, fg_color="transparent")
        f_itc.grid(row=0, column=1, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(f_itc, text="ITC Option (Input Tax Credit):", font=self.font_label).pack(anchor="w")
        itc_opt = ctk.CTkComboBox(f_itc, font=self.font_body, values=["ITC Not Claimable (5% Total)", "ITC Claimable (18% Total)"], height=38)
        itc_opt.pack(fill="x")
        
        def save_and_generate_invoice():
            try:
                final_cust_gstn = str(e_cust_gstn.get()).strip().upper()
                is_18 = "18%" in itc_opt.get()
                
                if is_18:
                    if not final_cust_gstn:
                        messagebox.showwarning(
                            "GSTN Mandatory for ITC Claim",
                            "Customer GSTIN is required to claim Input Tax Credit (ITC).\n\n"
                            "Either provide a valid GSTIN/Tax ID or switch the ITC Option to 'ITC Not Claimable (5% Total)'."
                        )
                        return
                    elif not is_valid_gstn(final_cust_gstn):
                        messagebox.showwarning(
                            "Invalid GSTN Structure",
                            f"The entered GSTIN '{final_cust_gstn}' does not match the 15-character structure (11AAAAA1111A1AO)!\n\n"
                            "• 11 = 2 State Digits\n"
                            "• AAAAA = 5 PAN Letters\n"
                            "• 1111 = 4 PAN Digits\n"
                            "• A = 1 PAN Letter\n"
                            "• 1 = 1 Entity Code (Alphanumeric)\n"
                            "• A = 1 Alphabet (e.g. Z)\n"
                            "• O = 1 Check Character (Alphanumeric)\n\n"
                            "Example: 09AAAAA0000A1Z5"
                        )
                        return
                elif final_cust_gstn and not is_valid_gstn(final_cust_gstn):
                    messagebox.showwarning(
                        "Invalid GSTN Format",
                        f"The entered GSTIN '{final_cust_gstn}' does not match the 15-character GSTN structure (11AAAAA1111A1AO).\n\n"
                        "Please correct the GSTIN or leave it blank for Unregistered Persons (B2C)."
                    )
                    return
                
                act_km = float(e_km.get() or 0)
                base = float(e_base_fare.get() or 0)
                r_rate = float(e_extra_rate.get() or 0)
                extra_km = max(0.0, act_km - est_km)
                extra_charges = extra_km * r_rate
                
                night_c = float(e_night.get() or 0)
                toll_c = float(e_toll.get() or 0)
                park_c = float(e_parking.get() or 0)
                oth_c = float(e_other.get() or 0)
                disc = float(e_disc.get() or 0)
                
                taxable_subtotal = (base + extra_charges + night_c + toll_c + park_c + oth_c) - disc
                
                is_igst = "IGST" in gst_type.get()
                rate = 0.18 if is_18 else 0.05
                
                total_tax = round(taxable_subtotal * rate, 2)
                igst = total_tax if is_igst else 0.0
                cgst = round(total_tax / 2.0, 2) if not is_igst else 0.0
                sgst = round(total_tax / 2.0, 2) if not is_igst else 0.0
                
                grand_total = round(taxable_subtotal + total_tax, 2)
                net_payable = round(grand_total - advance_paid, 2)
                
                custom_inv_date = parse_date_to_yyyymmdd(e_inv_date.get())
                
                excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
                df_invoices = excel_dict.get("Invoices", pd.DataFrame())
                df_ledger = excel_dict.get("Customer_Ledger", pd.DataFrame())
                
                new_inv_data = {
                    "Invoice_No": inv_no,
                    "Date": custom_inv_date,
                    "Inquiry_No": inq_id,
                    "Booking_ID": bk_id,
                    "Customer_Name": cust_name,
                    "GSTN": final_cust_gstn,
                    "Actual_KM": act_km,
                    "Base_Amount": base,
                    "Extra_KM_Charges": extra_charges,
                    "Night_Charges": night_c,
                    "Toll_Charges": toll_c,
                    "Parking_Charges": park_c,
                    "Other_Charges": oth_c,
                    "Discount": disc,
                    "Taxable_Amount": taxable_subtotal,
                    "IGST": igst,
                    "CGST": cgst,
                    "SGST": sgst,
                    "Grand_Total": grand_total,
                    "Advance_Adjusted": advance_paid,
                    "Net_Payable": net_payable,
                    "ITC_Claimable": itc_opt.get()
                }
                
                if not df_invoices.empty and "Invoice_No" in df_invoices.columns:
                    df_invoices = df_invoices[df_invoices["Invoice_No"].astype(str) != str(inv_no)]
                
                df_invoices = pd.concat([df_invoices, pd.DataFrame([new_inv_data])], ignore_index=True)
                
                if not df_ledger.empty and "Particulars" in df_ledger.columns:
                    df_ledger = df_ledger[~df_ledger["Particulars"].astype(str).str.contains(inv_no)]
                
                new_ledger_entry = {
                    "Date": custom_inv_date,
                    "Inquiry_No": inq_id,
                    "Customer_Name": cust_name,
                    "Particulars": f"Tax Invoice Generated {inv_no}",
                    "Voucher_Type": "Invoice",
                    "Debit": grand_total,
                    "Credit": 0,
                    "Balance": grand_total
                }
                df_ledger = pd.concat([df_ledger, pd.DataFrame([new_ledger_entry])], ignore_index=True)
                
                excel_dict["Invoices"] = df_invoices
                excel_dict["Customer_Ledger"] = df_ledger
                
                with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                    for s_name, s_df in excel_dict.items():
                        s_df.to_excel(writer, sheet_name=s_name, index=False)
                
                self.build_professional_invoice_pdf(
                    inv_no=inv_no,
                    inv_date=format_date_to_ddmmyyyy(custom_inv_date),
                    inq_id=inq_id,
                    bk_id=bk_id,
                    cust_name=cust_name,
                    cust_gstn=final_cust_gstn,
                    inq_row=inq_row,
                    act_km=act_km,
                    est_km=est_km,
                    base=base,
                    extra_km=extra_km,
                    r_rate=r_rate,
                    extra_charges=extra_charges,
                    night_c=night_c,
                    toll_c=toll_c,
                    park_c=park_c,
                    oth_c=oth_c,
                    disc=disc,
                    taxable_subtotal=taxable_subtotal,
                    gst_type_text=gst_type.get(),
                    itc_opt_text=itc_opt.get(),
                    igst=igst,
                    cgst=cgst,
                    sgst=sgst,
                    grand_total=grand_total,
                    advance_paid=advance_paid,
                    net_payable=net_payable
                )
                
                messagebox.showinfo("Success", f"Invoice {inv_no} saved successfully and posted to Ledger!")
                popup.destroy()
                self.refresh_accounts_view()
            except Exception as ex:
                messagebox.showerror("Error", f"Invoice save error: {str(ex)}")

        btn_txt = "💾 UPDATE & RE-PRINT TAX INVOICE" if is_editing else "💾 SAVE & PRINT FINAL TAX INVOICE"
        ctk.CTkButton(scroll, text=btn_txt, font=self.font_btn, fg_color="#28a745", height=44, command=save_and_generate_invoice).pack(fill="x", pady=15)

    def create_grid_label_val(self, parent, label_text, val_text, row, col):
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.grid(row=row, column=col, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(frame, text=label_text, font=self.font_label).pack(anchor="w")
        lbl_box = ctk.CTkLabel(frame, text=val_text, font=self.font_body, anchor="w", fg_color="#edf2f7", height=38, corner_radius=6)
        lbl_box.pack(fill="x")

    def build_professional_invoice_pdf(self, inv_no, inv_date, inq_id, bk_id, cust_name, cust_gstn, inq_row,
                                       act_km, est_km, base, extra_km, r_rate, extra_charges,
                                       night_c, toll_c, park_c, oth_c, disc, taxable_subtotal,
                                       gst_type_text, itc_opt_text, igst, cgst, sgst, grand_total,
                                       advance_paid, net_payable):
        df_comp = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
        comp = df_comp.iloc[0] if not df_comp.empty else {}
        
        pdf_path = os.path.join(INVOICE_DIR, f"{inv_no}.pdf")
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []
        
        comp_name = str(comp.get('Org_Name', 'TRAVEL MATE RENTALS PVT. LTD.'))
        comp_details = f"{comp.get('Address', '')}<br/>Phone: {comp.get('Mobile', '')} | Email: {comp.get('Email', '')}<br/><b>GSTIN: {comp.get('GSTN', '')}</b> | PAN: {comp.get('PAN', '')}"
        
        header_data = [
            [Paragraph(f"<b><font size=16 color='#1a365d'>{comp_name}</font></b><br/><font size=8.5 color='#4a5568'>{comp_details}</font>", styles['Normal']),
             Paragraph(f"<b><font size=18 color='#2b6cb0'>TAX INVOICE</font></b><br/><font size=9><b>Invoice No:</b> {inv_no}<br/><b>Date:</b> {inv_date}</font>", styles['Normal'])]
        ]
        h_table = Table(header_data, colWidths=[330, 205])
        h_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(h_table)
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1a365d'), spaceBefore=2, spaceAfter=8))
        
        bill_data = [
            [Paragraph("<b>BILLED TO (CUSTOMER):</b>", styles['Normal']), Paragraph("<b>TRIP & BOOKING DETAILS:</b>", styles['Normal'])],
            [Paragraph(f"<b>Name:</b> {cust_name}<br/><b>GSTIN / Tax ID:</b> {cust_gstn or 'Unregistered / URP'}<br/><b>Contact:</b> {inq_row.get('Phone', '')}<br/><b>Address:</b> {inq_row.get('Address', '')}", styles['Normal']),
             Paragraph(f"<b>Inquiry Ref:</b> {inq_id}<br/><b>Booking Ref:</b> {bk_id}<br/><b>Route:</b> {inq_row.get('From_Station','')} to {inq_row.get('To_Station','')}<br/><b>Cab / Driver:</b> {inq_row.get('Vehicle_Type','')} / {inq_row.get('Driver_Expected','')}", styles['Normal'])]
        ]
        b_table = Table(bill_data, colWidths=[265, 270])
        b_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f7fafc')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('PADDING', (0,0), (-1,-1), 5),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(b_table)
        elements.append(Spacer(1, 8))
        
        items = [
            ["Description / Particulars", "Computation Details", "Amount (INR)"],
            ["Base Package Fare", f"Estimated limit: {est_km} KM", f"{base:,.2f}"],
            ["Extra Distance Charges", f"Actual: {act_km} KM (Extra: {extra_km:.1f} KM @ INR {r_rate}/KM)", f"{extra_charges:,.2f}"],
            ["Driver Night Charges", "Night service allowance", f"{night_c:,.2f}"],
            ["Toll & Parking Charges", "Actual receipts attached", f"{(toll_c + park_c):,.2f}"],
            ["Other Incidental Charges", "Special requests / permits", f"{oth_c:,.2f}"],
            ["Discount Allowed", "Promotional concession", f"-{disc:,.2f}"]
        ]
        
        i_table = Table(items, colWidths=[200, 225, 110])
        i_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8.5),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e0')),
            ('ALIGN', (2,0), (2,-1), 'RIGHT'),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(i_table)
        elements.append(Spacer(1, 6))
        
        summary_rows = [
            ["Taxable Value (Subtotal):", f"INR {taxable_subtotal:,.2f}"],
            [f"GST Breakdown ({itc_opt_text}):", f"{gst_type_text}"],
            ["CGST Amount:", f"INR {cgst:,.2f}"],
            ["SGST Amount:", f"INR {sgst:,.2f}"],
            ["IGST Amount:", f"INR {igst:,.2f}"],
            ["Grand Total Invoice Value:", f"INR {grand_total:,.2f}"],
            ["Less Advance Received:", f"INR {advance_paid:,.2f}"],
            ["Net Balance Amount Payable:", f"INR {net_payable:,.2f}"]
        ]
        
        s_table = Table(summary_rows, colWidths=[385, 150])
        s_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 8.5),
            ('PADDING', (0,0), (-1,-1), 3),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('BACKGROUND', (0,5), (-1,5), colors.HexColor('#edf2f7')),
            ('FONTNAME', (0,5), (-1,5), 'Helvetica-Bold'),
            ('BACKGROUND', (0,7), (-1,7), colors.HexColor('#feebc8')),
            ('FONTNAME', (0,7), (-1,7), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0,7), (-1,7), colors.HexColor('#7b341e')),
        ]))
        elements.append(s_table)
        elements.append(Spacer(1, 10))
        
        bank_box = [
            [Paragraph(f"<b>BANK PAYMENT DETAILS:</b><br/>Bank: {comp.get('Bank_Name','')}<br/>A/C No: {comp.get('Account_No','')}<br/>IFSC: {comp.get('IFSC','')}", styles['Normal']),
             Paragraph(f"<br/><br/><b>For {comp_name}</b><br/><br/>Authorized Signatory", styles['Normal'])]
        ]
        f_table = Table(bank_box, colWidths=[320, 215])
        f_table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e0')),
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f7fafc')),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
            ('PADDING', (0,0), (-1,-1), 5),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(f_table)
        
        doc.build(elements)

    # --- Sub Tab 2: Money Receipts & Refunds Management (Reframed & Fully Working) ---
    def setup_receipts_tab(self):
        top_rec = ctk.CTkFrame(self.tab_receipts)
        top_rec.pack(fill="x", padx=10, pady=10)
        
        ctk.CTkButton(top_rec, text="+ ISSUE MONEY RECEIPT / REFUND", font=self.font_btn, width=250, height=38, fg_color="#28a745", command=self.open_issue_receipt_popup).pack(side="left", padx=4)
        ctk.CTkButton(top_rec, text="🖨️ PRINT SELECTED VOUCHER", font=self.font_btn, width=200, height=38, fg_color="#17a2b8", command=self.reprint_selected_receipt).pack(side="left", padx=4)
        ctk.CTkButton(top_rec, text="🚪 QUIT", font=self.font_btn, width=70, height=38, fg_color="#6c757d", command=self.quit).pack(side="right", padx=4)
        
        table_frame = ctk.CTkFrame(self.tab_receipts)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        cols = ("Receipt_No", "Date", "Inquiry_No", "Customer_Name", "Payment_Type", "Payment_Mode", "Ref_No", "Amount_INR", "Remarks")
        self.rec_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            self.rec_tree.heading(col, text=col)
            self.rec_tree.column(col, width=125, anchor="center")
            
        self.rec_tree.pack(side="left", fill="both", expand=True)
        self.load_receipts_data()

    def load_receipts_data(self):
        for item in self.rec_tree.get_children():
            self.rec_tree.delete(item)
        if os.path.exists(DB_FILE):
            try:
                df_rec = pd.read_excel(DB_FILE, sheet_name="Receipts")
                if not df_rec.empty:
                    for _, r in df_rec.iterrows():
                        amt = float(r.get("Amount", 0) or 0)
                        self.rec_tree.insert("", "end", values=(
                            r.get("Receipt_No",""),
                            format_date_to_ddmmyyyy(r.get("Date","")),
                            r.get("Inquiry_No",""),
                            r.get("Customer_Name",""),
                            r.get("Payment_Type",""),
                            r.get("Payment_Mode",""),
                            r.get("Ref_No",""),
                            f"₹{amt:,.2f}",
                            r.get("Remarks","")
                        ))
            except Exception as e:
                print("Error loading receipts:", e)

    def open_issue_receipt_popup(self):
        popup = ctk.CTkToplevel(self)
        popup.title("Issue Official Money Receipt or Refund - Travel Mate")
        popup.geometry("680x750")
        popup.grab_set()
        
        # Dedicated container frame with proper padding and no geometry conflict
        container = ctk.CTkFrame(popup, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=20, pady=20)
        
        rec_no = f"REC-{datetime.datetime.now().strftime('%Y%m%d%H%M')}"
        ctk.CTkLabel(container, text=f"Voucher No: {rec_no}", font=self.font_header, text_color="#1a365d").pack(anchor="w", pady=(0, 10))
        
        # Date Field
        e_rec_date = self.create_date_field(container, "Voucher Date (DD/MM/YYYY):", row=None, col=None, default=format_date_to_ddmmyyyy(datetime.date.today()))
        
        # Inquiry Selection with Combined Name
        inq_options = ["INQ-DIRECT | Direct Cash / Walk-in"]
        if os.path.exists(DB_FILE):
            try:
                df_inq_master = pd.read_excel(DB_FILE, sheet_name="Inquiries")
                for _, r in df_inq_master.iterrows():
                    inq_no = str(r.get("Inquiry_No", ""))
                    c_name = str(r.get("Customer_Name", ""))
                    if inq_no:
                        inq_options.append(f"{inq_no} | {c_name}")
            except:
                pass

        ctk.CTkLabel(container, text="Select Linked Inquiry & Customer:", font=self.font_label).pack(anchor="w", pady=(8, 2))
        cb_inq = ctk.CTkComboBox(container, font=self.font_body, values=inq_options, height=38)
        cb_inq.pack(fill="x", pady=(0, 8))
        if inq_options:
            cb_inq.set(inq_options[0])

        e_cust = self.create_form_field(container, "Customer Name:")
        
        def auto_fill_cust(*args):
            sel_val = cb_inq.get()
            if " | " in sel_val:
                parts = sel_val.split(" | ")
                inq_code = parts[0].strip()
                if inq_code != "INQ-DIRECT" and os.path.exists(DB_FILE):
                    try:
                        df_inq = pd.read_excel(DB_FILE, sheet_name="Inquiries")
                        m = df_inq[df_inq["Inquiry_No"].astype(str) == inq_code]
                        if not m.empty:
                            e_cust.delete(0, 'end')
                            e_cust.insert(0, str(m.iloc[0].get("Customer_Name", "")))
                    except:
                        pass
                else:
                    e_cust.delete(0, 'end')
                    e_cust.insert(0, parts[1].strip() if len(parts) > 1 else "")
        
        cb_inq.configure(command=auto_fill_cust)
        auto_fill_cust()
        
        ctk.CTkLabel(container, text="Payment Type / Stage:", font=self.font_label).pack(anchor="w", pady=(8, 2))
        cb_type = ctk.CTkComboBox(container, font=self.font_body, values=["Advance / Booking Deposit", "Running / Midway Payment", "Final Bill Settlement", "Security Deposit", "Refund / Security Deposit Return"], height=38)
        cb_type.pack(fill="x", pady=(0, 8))
        cb_type.set("Advance / Booking Deposit")
        
        ctk.CTkLabel(container, text="Payment Mode:", font=self.font_label).pack(anchor="w", pady=(8, 2))
        cb_mode = ctk.CTkComboBox(container, font=self.font_body, values=["Cash", "UPI / GPay / PhonePe", "Bank Transfer (NEFT/RTGS)", "Cheque", "Credit Card / POS"], height=38)
        cb_mode.pack(fill="x", pady=(0, 8))
        cb_mode.set("Cash")
        
        e_ref = self.create_form_field(container, "Transaction Ref No / Cheque No / UTR:", default="CASH / UPI-REF")
        e_amt = self.create_form_field(container, "Amount (₹):", default="5000")
        e_rem = self.create_form_field(container, "Remarks / Particulars:", default="Payment received with thanks")
        
        def save_and_print_receipt():
            try:
                amt_val = float(e_amt.get() or 0)
                if amt_val <= 0:
                    messagebox.showerror("Error", "Please enter a valid Amount.")
                    return
                if not e_cust.get().strip():
                    messagebox.showerror("Error", "Customer Name is required.")
                    return
                    
                custom_date = parse_date_to_yyyymmdd(e_rec_date.get())
                p_type = cb_type.get()
                is_refund = "refund" in p_type.lower() or "return" in p_type.lower()
                
                selected_inq_combo = cb_inq.get()
                linked_inq_no = selected_inq_combo.split(" | ")[0].strip() if " | " in selected_inq_combo else selected_inq_combo
                
                excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
                df_rec = excel_dict.get("Receipts", pd.DataFrame())
                df_ledger = excel_dict.get("Customer_Ledger", pd.DataFrame())
                df_bk = excel_dict.get("Bookings", pd.DataFrame())
                df_inq = excel_dict.get("Inquiries", pd.DataFrame())
                
                new_rec = {
                    "Receipt_No": rec_no,
                    "Date": custom_date,
                    "Inquiry_No": linked_inq_no,
                    "Customer_Name": e_cust.get(),
                    "Payment_Type": p_type,
                    "Payment_Mode": cb_mode.get(),
                    "Ref_No": e_ref.get(),
                    "Amount": amt_val,
                    "Remarks": e_rem.get()
                }
                df_rec = pd.concat([df_rec, pd.DataFrame([new_rec])], ignore_index=True)
                
                if is_refund:
                    new_ledger = {
                        "Date": custom_date,
                        "Inquiry_No": linked_inq_no,
                        "Customer_Name": e_cust.get(),
                        "Particulars": f"Refund Voucher {rec_no} ({p_type} via {cb_mode.get()})",
                        "Voucher_Type": "Refund",
                        "Debit": amt_val,
                        "Credit": 0,
                        "Balance": amt_val
                    }
                else:
                    new_ledger = {
                        "Date": custom_date,
                        "Inquiry_No": linked_inq_no,
                        "Customer_Name": e_cust.get(),
                        "Particulars": f"Receipt {rec_no} ({p_type} via {cb_mode.get()})",
                        "Voucher_Type": "Receipt",
                        "Debit": 0,
                        "Credit": amt_val,
                        "Balance": -amt_val
                    }
                    
                df_ledger = pd.concat([df_ledger, pd.DataFrame([new_ledger])], ignore_index=True)
                
                bk_key = f"BK-{linked_inq_no.replace('INQ-', '')}"
                if not df_bk.empty and "Booking_ID" in df_bk.columns:
                    b_idx = df_bk.index[df_bk["Booking_ID"].astype(str) == bk_key].tolist()
                    if b_idx:
                        df_bk.at[b_idx[0], "Advance_Amount"] = self.calculate_total_advance_before_invoice(linked_inq_no, df_inq, df_rec)
                
                excel_dict["Receipts"] = df_rec
                excel_dict["Customer_Ledger"] = df_ledger
                excel_dict["Bookings"] = df_bk
                
                with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                    for s_name, s_df in excel_dict.items():
                        s_df.to_excel(writer, sheet_name=s_name, index=False)
                        
                self.build_money_receipt_pdf(
                    rec_no=rec_no,
                    rec_date=format_date_to_ddmmyyyy(custom_date),
                    inq_id=linked_inq_no,
                    cust_name=e_cust.get(),
                    pay_type=p_type,
                    pay_mode=cb_mode.get(),
                    ref_no=e_ref.get(),
                    amount=amt_val,
                    remarks=e_rem.get()
                )
                
                msg = f"Refund Voucher {rec_no} processed and debited from ledger!" if is_refund else f"Money Receipt {rec_no} generated and credited to ledger!"
                messagebox.showinfo("Success", msg)
                popup.destroy()
                self.refresh_accounts_view()
                self.load_bookings()
            except Exception as ex:
                messagebox.showerror("Error", f"Voucher save error: {str(ex)}")

        ctk.CTkButton(container, text="💾 SAVE & GENERATE VOUCHER", font=self.font_btn, fg_color="#28a745", height=44, command=save_and_print_receipt).pack(fill="x", pady=15)

    def reprint_selected_receipt(self):
        selected = self.rec_tree.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a voucher from the table.")
            return
        rec_no = str(self.rec_tree.item(selected[0])['values'][0])
        pdf_path = os.path.join(RECEIPT_DIR, f"{rec_no}.pdf")
        if os.path.exists(pdf_path):
            if os.name == 'nt':
                os.startfile(pdf_path)
        else:
            messagebox.showinfo("Notice", f"Voucher PDF file not found ({pdf_path}).")

    def build_money_receipt_pdf(self, rec_no, rec_date, inq_id, cust_name, pay_type, pay_mode, ref_no, amount, remarks):
        df_comp = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
        comp = df_comp.iloc[0] if not df_comp.empty else {}
        
        pdf_path = os.path.join(RECEIPT_DIR, f"{rec_no}.pdf")
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []
        
        is_refund = "refund" in pay_type.lower() or "return" in pay_type.lower()
        title_str = "REFUND VOUCHER" if is_refund else "PAYMENT RECEIPT"
        theme_color = colors.HexColor('#c53030') if is_refund else colors.HexColor('#28a745')
        
        comp_name = str(comp.get('Org_Name', 'TRAVEL MATE RENTALS PVT. LTD.'))
        comp_details = f"{comp.get('Address', '')}<br/>Phone: {comp.get('Mobile', '')} | Email: {comp.get('Email', '')}<br/><b>GSTIN: {comp.get('GSTN', '')}</b>"
        
        header_data = [
            [Paragraph(f"<b><font size=16 color='#1a365d'>{comp_name}</font></b><br/><font size=8.5 color='#4a5568'>{comp_details}</font>", styles['Normal']),
             Paragraph(f"<b><font size=18 color='{theme_color.hexval()}'>{title_str}</font></b><br/><font size=9><b>Voucher No:</b> {rec_no}<br/><b>Date:</b> {rec_date}</font>", styles['Normal'])]
        ]
        h_table = Table(header_data, colWidths=[330, 205])
        h_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elements.append(h_table)
        elements.append(HRFlowable(width="100%", thickness=1.5, color=theme_color, spaceBefore=2, spaceAfter=10))
        
        rec_data = [
            ["Customer Name:", str(cust_name), "Inquiry / Booking Ref:", str(inq_id)],
            ["Voucher Type / Purpose:", str(pay_type), "Payment Mode:", str(pay_mode)],
            ["Transaction Ref / UTR:", str(ref_no), "Voucher Amount:", f"INR {amount:,.2f}"],
            ["Remarks / Note:", str(remarks), "Status:", "Processed & Posted"]
        ]
        
        r_table = Table(rec_data, colWidths=[140, 160, 100, 135])
        r_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f7fafc')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e0')),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('PADDING', (0,0), (-1,-1), 6),
            ('BACKGROUND', (2,2), (3,2), colors.HexColor('#fed7d7') if is_refund else colors.HexColor('#d4edda')),
            ('FONTNAME', (2,2), (3,2), 'Helvetica-Bold'),
        ]))
        elements.append(r_table)
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("<b>Note:</b> This is an official computer-generated voucher and has been duly updated in the customer ledger.", styles['Normal']))
        elements.append(Spacer(1, 35))
        elements.append(Paragraph("Authorized Signatory / Cashier: ____________________________", styles['Normal']))
        
        doc.build(elements)
        if os.name == 'nt':
            os.startfile(pdf_path)

    # --- Sub Tab 3: Customer Ledger Management ---
    def setup_ledger_tab(self):
        top_led = ctk.CTkFrame(self.tab_ledger)
        top_led.pack(fill="x", padx=10, pady=8)
        
        ctk.CTkLabel(top_led, text="Filter Account (Inquiry | Customer):", font=self.font_label).pack(side="left", padx=4)
        self.cb_ledger_account = ctk.CTkComboBox(top_led, font=self.font_body, values=["All Accounts"], width=320, height=38, command=lambda e: self.load_ledger_data())
        self.cb_ledger_account.pack(side="left", padx=4)
        
        ctk.CTkButton(top_led, text="📥 EXPORT LEDGER PDF", font=self.font_btn, height=38, fg_color="#17a2b8", command=self.export_ledger_pdf).pack(side="left", padx=8)
        
        table_frame = ctk.CTkFrame(self.tab_ledger)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        cols = ("Date", "Inquiry_No", "Customer_Name", "Particulars", "Voucher_Type", "Debit_INR", "Credit_INR", "Balance_INR")
        self.led_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            self.led_tree.heading(col, text=col)
            self.led_tree.column(col, width=125, anchor="center")
            
        self.led_tree.pack(side="left", fill="both", expand=True)
        self.load_ledger_data()

    def load_ledger_data(self):
        for item in self.led_tree.get_children():
            self.led_tree.delete(item)
            
        if os.path.exists(DB_FILE):
            df_led = pd.read_excel(DB_FILE, sheet_name="Customer_Ledger")
            
            if not df_led.empty and "Inquiry_No" in df_led.columns and "Customer_Name" in df_led.columns:
                df_led["Account_Key"] = df_led["Inquiry_No"].astype(str) + " | " + df_led["Customer_Name"].astype(str)
                acc_list = ["All Accounts"] + sorted(list(df_led["Account_Key"].dropna().unique()))
                self.cb_ledger_account.configure(values=acc_list)
            else:
                df_led["Account_Key"] = ""
                
            selected_acc = self.cb_ledger_account.get()
            if selected_acc and selected_acc != "All Accounts":
                df_led = df_led[df_led["Account_Key"] == selected_acc]
            
            if not df_led.empty and "Date" in df_led.columns:
                df_led['__temp_dt'] = pd.to_datetime(df_led['Date'], errors='coerce')
                df_led = df_led.sort_values(by='__temp_dt', na_position='first').drop(columns=['__temp_dt'])
                
            running_balance = 0.0
            for _, r in df_led.iterrows():
                dr = float(r.get("Debit", 0) or 0)
                cr = float(r.get("Credit", 0) or 0)
                running_balance += (dr - cr)
                
                self.led_tree.insert("", "end", values=(
                    format_date_to_ddmmyyyy(r.get("Date", "")),
                    r.get("Inquiry_No", ""),
                    r.get("Customer_Name", ""),
                    r.get("Particulars", ""),
                    r.get("Voucher_Type", ""),
                    f"{dr:,.2f}",
                    f"{cr:,.2f}",
                    f"{running_balance:,.2f}"
                ))

    def export_ledger_pdf(self):
        selected_acc = self.cb_ledger_account.get()
        if not selected_acc or selected_acc == "All Accounts":
            messagebox.showwarning("Warning", "Please select a specific Account (Inquiry | Customer) to export the Statement PDF.")
            return
            
        df_led = pd.read_excel(DB_FILE, sheet_name="Customer_Ledger")
        df_comp = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
        comp = df_comp.iloc[0] if not df_comp.empty else {}
        
        df_led["Account_Key"] = df_led["Inquiry_No"].astype(str) + " | " + df_led["Customer_Name"].astype(str)
        match = df_led[df_led["Account_Key"] == selected_acc]
        
        if match.empty:
            messagebox.showinfo("Notice", "No ledger transactions found for this account.")
            return
            
        inq_id, cust_name = selected_acc.split(" | ") if " | " in selected_acc else ("", selected_acc)
        clean_file_acc = selected_acc.replace(" | ", "_").replace(" ", "_").replace("/", "_")
        pdf_path = os.path.join(LEDGER_DIR, f"Ledger_{clean_file_acc}.pdf")
        
        doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []
        
        elements.append(Paragraph(f"<b><font size=16 color='#1a365d'>{comp.get('Org_Name', 'TRAVEL MATE')}</font></b>", styles['Normal']))
        elements.append(Paragraph(f"<b>STATEMENT OF ACCOUNT / CUSTOMER LEDGER</b>", styles['Heading2']))
        elements.append(Paragraph(f"<b>Inquiry Ref:</b> {inq_id} | <b>Customer:</b> {cust_name} | <b>Statement Date:</b> {datetime.date.today().strftime('%d/%m/%Y')}", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        table_data = [["Date", "Ref No", "Particulars", "Type", "Debit (₹)", "Credit (₹)", "Balance (₹)"]]
        running_bal = 0.0
        for _, r in match.iterrows():
            dr = float(r.get("Debit", 0) or 0)
            cr = float(r.get("Credit", 0) or 0)
            running_bal += (dr - cr)
            table_data.append([
                format_date_to_ddmmyyyy(r.get("Date", "")),
                str(r.get("Inquiry_No", "")),
                str(r.get("Particulars", ""))[:25],
                str(r.get("Voucher_Type", "")),
                f"{dr:,.2f}",
                f"{cr:,.2f}",
                f"{running_bal:,.2f}"
            ])
            
        t = Table(table_data, colWidths=[65, 80, 160, 60, 60, 60, 65])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e0')),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (4,1), (-1,-1), 'RIGHT'),
            ('PADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(f"<b>Closing Outstanding Balance: INR {running_bal:,.2f}</b>", styles['Normal']))
        
        doc.build(elements)
        messagebox.showinfo("Success", f"Customer Statement PDF generated:\n{pdf_path}")
        if os.name == 'nt':
            os.startfile(pdf_path)

    # =========================================================================
    # 4. REPORTS & EXPORTS
    # =========================================================================
    def create_reports_view(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid(row=0, column=0, sticky="nsew")
        self.frames["reports"] = frame
        
        filter_bar = ctk.CTkFrame(frame)
        filter_bar.pack(fill="x", padx=10, pady=8)
        
        ctk.CTkLabel(filter_bar, text="Report Category:", font=self.font_label).pack(side="left", padx=4)
        
        self.report_categories = [
            "Inquiry: Pending Inquiries (Date-Wise)",
            "Inquiry: Cancelled Inquiries (Date-Wise)",
            "Inquiry: Confirmed Converted to Booking (Date-Wise)",
            "Booking: Confirmed Bookings (Date-Wise)",
            "Booking: Journey Completed & Invoice NOT Generated",
            "Booking: Journey Completed & Invoice Generated",
            "Booking: Journey Not Started",
            "Booking: Journey In-Progress (Started Not Completed)",
            "Accounts: Customers with Outstanding Dues",
            "Accounts: Confirmed Bookings Ledger (Date-Wise)",
            "Accounts: Tax Invoices Issued (Date-Wise)"
        ]
        
        self.cb_rep_type = ctk.CTkComboBox(filter_bar, font=self.font_body, values=self.report_categories, width=340, height=38, command=lambda e: self.generate_custom_report())
        self.cb_rep_type.pack(side="left", padx=4)
        
        ctk.CTkLabel(filter_bar, text="From:", font=self.font_label).pack(side="left", padx=(8, 2))
        self.e_rep_from = self.create_date_field(filter_bar, "", row=None, col=None, default=format_date_to_ddmmyyyy(datetime.date.today().replace(day=1)))
        self.e_rep_from.pack(side="left", padx=2)
        
        ctk.CTkLabel(filter_bar, text="To:", font=self.font_label).pack(side="left", padx=(6, 2))
        self.e_rep_to = self.create_date_field(filter_bar, "", row=None, col=None, default=format_date_to_ddmmyyyy(datetime.date.today()))
        self.e_rep_to.pack(side="left", padx=2)
        
        ctk.CTkButton(filter_bar, text="🔍 RUN REPORT", font=self.font_btn, width=120, height=38, fg_color="#007bff", command=self.generate_custom_report).pack(side="left", padx=6)
        ctk.CTkButton(filter_bar, text="📥 EXCEL", font=self.font_btn, width=85, height=38, fg_color="#28a745", command=self.export_report_excel).pack(side="left", padx=3)
        ctk.CTkButton(filter_bar, text="📄 PDF", font=self.font_btn, width=75, height=38, fg_color="#17a2b8", command=self.export_report_pdf).pack(side="left", padx=3)
        ctk.CTkButton(filter_bar, text="🚪 QUIT", font=self.font_btn, width=65, height=38, fg_color="#6c757d", command=self.quit).pack(side="right", padx=6)
        
        self.card_frame = ctk.CTkFrame(frame, fg_color="#edf2f7")
        self.card_frame.pack(fill="x", padx=10, pady=4)
        
        self.lbl_rep_title = ctk.CTkLabel(self.card_frame, text="Active Report: None", font=self.font_sub, text_color="#1a365d")
        self.lbl_rep_title.pack(side="left", padx=10, pady=6)
        
        self.lbl_rep_count = ctk.CTkLabel(self.card_frame, text="Total Records: 0", font=self.font_sub, text_color="#2b6cb0")
        self.lbl_rep_count.pack(side="right", padx=15, pady=6)
        
        table_frame = ctk.CTkFrame(frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.rep_tree = ttk.Treeview(table_frame, show="headings", selectmode="browse")
        scrollbar_y = ttk.Scrollbar(table_frame, orient="vertical", command=self.rep_tree.yview)
        scrollbar_x = ttk.Scrollbar(table_frame, orient="horizontal", command=self.rep_tree.xview)
        
        self.rep_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.rep_tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        
        self.current_report_df = pd.DataFrame()
        frame.refresh_data = self.generate_custom_report

    def filter_by_date(self, df, date_col):
        if df.empty or date_col not in df.columns:
            return df
        try:
            d_from = parse_date_to_yyyymmdd(self.e_rep_from.get())
            d_to = parse_date_to_yyyymmdd(self.e_rep_to.get())
            
            df_temp = df.copy()
            df_temp['__d'] = pd.to_datetime(df_temp[date_col], errors='coerce').dt.date
            
            if d_from:
                from_dt = datetime.datetime.strptime(d_from, "%Y-%m-%d").date()
                df_temp = df_temp[df_temp['__d'] >= from_dt]
            if d_to:
                to_dt = datetime.datetime.strptime(d_to, "%Y-%m-%d").date()
                df_temp = df_temp[df_temp['__d'] <= to_dt]
                
            return df_temp.drop(columns=['__d'], errors='ignore')
        except Exception:
            return df

    def generate_custom_report(self):
        rep_type = self.cb_rep_type.get()
        if not os.path.exists(DB_FILE):
            return
            
        excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
        df_inq = excel_dict.get("Inquiries", pd.DataFrame())
        df_bk = excel_dict.get("Bookings", pd.DataFrame())
        df_inv = excel_dict.get("Invoices", pd.DataFrame())
        df_led = excel_dict.get("Customer_Ledger", pd.DataFrame())
        
        result_df = pd.DataFrame()
        
        if "Inquiry: Pending" in rep_type:
            sub = df_inq[df_inq["Status"].astype(str).str.lower() == "pending"]
            result_df = self.filter_by_date(sub, "Date")
        elif "Inquiry: Cancelled" in rep_type:
            sub = df_inq[df_inq["Status"].astype(str).str.lower() == "cancelled"]
            result_df = self.filter_by_date(sub, "Date")
        elif "Inquiry: Confirmed" in rep_type:
            sub = df_inq[df_inq["Status"].astype(str).str.lower() == "confirmed"]
            result_df = self.filter_by_date(sub, "Date")
        elif "Booking: Confirmed Bookings" in rep_type:
            result_df = self.filter_by_date(df_bk, "Date_From")
        elif "Booking: Journey Completed & Invoice NOT Generated" in rep_type:
            inv_inqs = df_inv["Inquiry_No"].astype(str).unique() if not df_inv.empty and "Inquiry_No" in df_inv.columns else []
            sub = df_bk[(df_bk["Journey_Status"].astype(str) == "Completed") & (~df_bk["Inquiry_No"].astype(str).isin(inv_inqs))]
            result_df = self.filter_by_date(sub, "Date_From")
        elif "Booking: Journey Completed & Invoice Generated" in rep_type:
            inv_inqs = df_inv["Inquiry_No"].astype(str).unique() if not df_inv.empty and "Inquiry_No" in df_inv.columns else []
            sub = df_bk[(df_bk["Journey_Status"].astype(str) == "Completed") & (df_bk["Inquiry_No"].astype(str).isin(inv_inqs))]
            result_df = self.filter_by_date(sub, "Date_From")
        elif "Booking: Journey Not Started" in rep_type:
            sub = df_bk[df_bk["Journey_Status"].astype(str) == "Not Started"]
            result_df = self.filter_by_date(sub, "Date_From")
        elif "Booking: Journey In-Progress" in rep_type:
            sub = df_bk[df_bk["Journey_Status"].astype(str) == "In-Progress"]
            result_df = self.filter_by_date(sub, "Date_From")
        elif "Accounts: Customers with Outstanding Dues" in rep_type:
            if not df_led.empty and "Customer_Name" in df_led.columns:
                grouped = df_led.groupby("Customer_Name").apply(
                    lambda g: pd.Series({
                        "Total_Billed": g["Debit"].sum(),
                        "Total_Received": g["Credit"].sum(),
                        "Outstanding_Balance": g["Debit"].sum() - g["Credit"].sum()
                    })
                ).reset_index()
                result_df = grouped[grouped["Outstanding_Balance"] > 0]
        elif "Accounts: Confirmed Bookings Ledger" in rep_type:
            result_df = self.filter_by_date(df_led, "Date")
        elif "Accounts: Tax Invoices Issued" in rep_type:
            result_df = self.filter_by_date(df_inv, "Date")
            
        self.current_report_df = result_df
        self.render_report_to_table(result_df, rep_type)

    def render_report_to_table(self, df, rep_name):
        self.rep_tree.delete(*self.rep_tree.get_children())
        
        if df.empty:
            self.rep_tree["columns"] = ("Status",)
            self.rep_tree.heading("Status", text="Status")
            self.rep_tree.insert("", "end", values=("No records found for selected criteria",))
            self.lbl_rep_title.configure(text=f"Report: {rep_name}")
            self.lbl_rep_count.configure(text="Total Records: 0")
            return
            
        display_cols = list(df.columns)
        self.rep_tree["columns"] = display_cols
        
        for col in display_cols:
            self.rep_tree.heading(col, text=col)
            self.rep_tree.column(col, width=135, anchor="center")
            
        for _, row in df.iterrows():
            formatted_vals = []
            for col in display_cols:
                val = row[col]
                if isinstance(val, (float, int)) and ("Amount" in col or "Balance" in col or "Total" in col or "Debit" in col or "Credit" in col):
                    formatted_vals.append(f"₹{val:,.2f}")
                elif "Date" in col and pd.notna(val):
                    formatted_vals.append(format_date_to_ddmmyyyy(val))
                else:
                    formatted_vals.append(str(val) if pd.notna(val) else "")
            self.rep_tree.insert("", "end", values=formatted_vals)
            
        self.lbl_rep_title.configure(text=f"Report: {rep_name}")
        self.lbl_rep_count.configure(text=f"Total Records: {len(df)}")

    def export_report_excel(self):
        if self.current_report_df.empty:
            messagebox.showwarning("Warning", "No data available to export.")
            return
            
        clean_name = self.cb_rep_type.get().split(":")[1].strip().replace(" ", "_").replace("/", "_")
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=f"Report_{clean_name}.xlsx")
        if not file_path:
            return
            
        try:
            self.current_report_df.to_excel(file_path, index=False)
            messagebox.showinfo("Export Successful", f"Excel report saved successfully:\n{file_path}")
            if os.name == 'nt':
                os.startfile(file_path)
        except Exception as e:
            messagebox.showerror("Error", f"Excel export error: {str(e)}")

    def export_report_pdf(self):
        if self.current_report_df.empty:
            messagebox.showwarning("Warning", "No data available to export.")
            return
            
        clean_name = self.cb_rep_type.get().split(":")[1].strip().replace(" ", "_").replace("/", "_")
        pdf_path = os.path.join(REPORT_DIR, f"Report_{clean_name}_{datetime.datetime.now().strftime('%Y%m%d%H%M')}.pdf")
        
        try:
            df_comp = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
            comp = df_comp.iloc[0] if not df_comp.empty else {}
            
            doc = SimpleDocTemplate(pdf_path, pagesize=A4, rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
            styles = getSampleStyleSheet()
            elements = []
            
            elements.append(Paragraph(f"<b><font size=15 color='#1a365d'>{comp.get('Org_Name', 'TRAVEL MATE')}</font></b>", styles['Normal']))
            elements.append(Paragraph(f"<b><font size=12 color='#2b6cb0'>MIS REPORT: {self.cb_rep_type.get()}</font></b>", styles['Normal']))
            elements.append(Paragraph(f"<font size=8.5 color='#718096'>Date Range: {self.e_rep_from.get()} to {self.e_rep_to.get()} | Generated: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}</font>", styles['Normal']))
            elements.append(Spacer(1, 8))
            elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1a365d'), spaceBefore=2, spaceAfter=8))
            
            df_export = self.current_report_df.copy()
            if len(df_export.columns) > 6:
                df_export = df_export.iloc[:, :6]
                
            headers = [list(df_export.columns)]
            rows = []
            for _, r in df_export.iterrows():
                row_items = []
                for c in df_export.columns:
                    val = r[c]
                    if "Date" in c and pd.notna(val):
                        row_items.append(format_date_to_ddmmyyyy(val))
                    else:
                        row_items.append(str(val)[:20] if pd.notna(val) else "")
                rows.append(row_items)
                
            pdf_table_data = headers + rows
            t = Table(pdf_table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1a365d')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e0')),
                ('PADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 15))
            elements.append(Paragraph(f"<b>Total Count: {len(self.current_report_df)} Records</b>", styles['Normal']))
            
            doc.build(elements)
            messagebox.showinfo("Export Successful", f"PDF Report generated successfully:\n{pdf_path}")
            if os.name == 'nt':
                os.startfile(pdf_path)
        except Exception as e:
            messagebox.showerror("Error", f"PDF export error: {str(e)}")

    # =========================================================================
    # 5. MASTER DATA SETUP
    # =========================================================================
    def create_masters_view(self):
        frame = ctk.CTkFrame(self.content_frame, fg_color="transparent")
        frame.grid(row=0, column=0, sticky="nsew")
        self.frames["masters"] = frame
        
        top_bar = ctk.CTkFrame(frame)
        top_bar.pack(fill="x", padx=10, pady=8)
        
        ctk.CTkLabel(top_bar, text="⚙️ Master Database Setup", font=self.font_header).pack(side="left", padx=10)
        ctk.CTkButton(top_bar, text="🚪 QUIT", font=self.font_btn, width=70, height=38, fg_color="#6c757d", command=self.quit).pack(side="right", padx=10)
        
        self.master_tabview = ctk.CTkTabview(frame)
        self.master_tabview.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.tab_comp = self.master_tabview.add("🏢 Company Profile")
        self.tab_veh = self.master_tabview.add("🚗 Vehicle Master")
        self.tab_drv = self.master_tabview.add("👨‍✈️ Driver Master")
        
        self.setup_company_tab()
        self.setup_vehicle_tab()
        self.setup_driver_tab()
        
        frame.refresh_data = self.refresh_all_masters

    def refresh_all_masters(self):
        self.load_company_data()
        self.load_vehicle_data()
        self.load_driver_data()

    def setup_company_tab(self):
        scroll = ctk.CTkScrollableFrame(self.tab_comp)
        scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        self.comp_entries = {}
        
        c1 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        c1.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(c1, text="🏢 Legal & Registration Credentials", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        g1 = ctk.CTkFrame(c1, fg_color="transparent")
        g1.pack(fill="x", padx=10, pady=(0, 10))
        g1.grid_columnconfigure((0, 1), weight=1)
        
        self.comp_entries["Org_Name"] = self.create_grid_field(g1, "Name of Organisation / Company:", row=0, col=0)
        
        f_st = ctk.CTkFrame(g1, fg_color="transparent")
        f_st.grid(row=0, column=1, sticky="ew", padx=6, pady=4)
        ctk.CTkLabel(f_st, text="Organisation Legal Status:", font=self.font_label).pack(anchor="w")
        self.cb_comp_status = ctk.CTkComboBox(f_st, font=self.font_body, values=["Proprietorship", "Partnership", "HUF", "AOP", "Private Limited Co.", "Limited Co.", "LLP"], height=38)
        self.cb_comp_status.pack(fill="x")
        
        self.comp_entries["Reg_No"] = self.create_grid_field(g1, "Registration No (CIN / LLPIN / Reg):", row=1, col=0)
        self.comp_entries["GSTN"] = self.create_grid_field(g1, "Company GSTIN (Format: 11AAAAA1111A1AO):", row=1, col=1)
        self.comp_entries["PAN"] = self.create_grid_field(g1, "Company PAN:", row=2, col=0)
        self.comp_entries["TAN"] = self.create_grid_field(g1, "Company TAN:", row=2, col=1)

        c2 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        c2.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(c2, text="🏦 Banking & Financial Accounts", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        g2 = ctk.CTkFrame(c2, fg_color="transparent")
        g2.pack(fill="x", padx=10, pady=(0, 10))
        g2.grid_columnconfigure((0, 1, 2), weight=1)
        
        self.comp_entries["Bank_Name"] = self.create_grid_field(g2, "Bank Name:", row=0, col=0)
        self.comp_entries["Account_No"] = self.create_grid_field(g2, "Bank Account Number:", row=0, col=1)
        self.comp_entries["IFSC"] = self.create_grid_field(g2, "Bank IFSC Code:", row=0, col=2)

        c3 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        c3.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(c3, text="📍 Official Address & Communication Channels", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        g3 = ctk.CTkFrame(c3, fg_color="transparent")
        g3.pack(fill="x", padx=10, pady=(0, 10))
        g3.grid_columnconfigure((0, 1), weight=1)
        
        self.comp_entries["Address"] = self.create_grid_field(g3, "Complete Registered Address:", row=0, col=0)
        self.comp_entries["Website"] = self.create_grid_field(g3, "Official Website URL:", row=0, col=1)
        self.comp_entries["Email"] = self.create_grid_field(g3, "Support / Official Email ID:", row=1, col=0)
        self.comp_entries["Mobile"] = self.create_grid_field(g3, "Mobile / Helpline Phone:", row=1, col=1)
        
        def save_company_profile():
            comp_gstn = self.comp_entries["GSTN"].get().strip().upper()
            if comp_gstn and not is_valid_gstn(comp_gstn):
                messagebox.showwarning(
                    "Invalid Company GSTIN",
                    f"Company GSTIN '{comp_gstn}' does not match format 11AAAAA1111A1AO.\n\n"
                    "Example: 09AAAAA0000A1Z5"
                )
                return
                
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            comp_data = {
                "Org_Name": [self.comp_entries["Org_Name"].get()],
                "Org_Status": [self.cb_comp_status.get()],
                "Reg_No": [self.comp_entries["Reg_No"].get()],
                "GSTN": [comp_gstn],
                "PAN": [self.comp_entries["PAN"].get()],
                "TAN": [self.comp_entries["TAN"].get()],
                "Bank_Name": [self.comp_entries["Bank_Name"].get()],
                "Account_No": [self.comp_entries["Account_No"].get()],
                "IFSC": [self.comp_entries["IFSC"].get()],
                "Address": [self.comp_entries["Address"].get()],
                "Email": [self.comp_entries["Email"].get()],
                "Mobile": [self.comp_entries["Mobile"].get()],
                "Website": [self.comp_entries["Website"].get()]
            }
            excel_dict["Company_Profile"] = pd.DataFrame(comp_data)
            with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                for s_name, s_df in excel_dict.items():
                    s_df.to_excel(writer, sheet_name=s_name, index=False)
            messagebox.showinfo("Saved", "Company Profile successfully saved & updated!")
            self.update_top_branding_header()
            
        ctk.CTkButton(scroll, text="💾 SAVE / UPDATE COMPANY PROFILE", font=self.font_btn, fg_color="#28a745", height=44, command=save_company_profile).pack(fill="x", pady=15)
        self.load_company_data()

    def load_company_data(self):
        if os.path.exists(DB_FILE):
            df = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
            if not df.empty:
                c = df.iloc[0]
                self.cb_comp_status.set(str(c.get("Org_Status", "Private Limited Co.")))
                for key, entry in self.comp_entries.items():
                    entry.delete(0, 'end')
                    val = str(c.get(key, ""))
                    if val.lower() != "nan":
                        entry.insert(0, val)

    def setup_vehicle_tab(self):
        top_v = ctk.CTkFrame(self.tab_veh)
        top_v.pack(fill="x", padx=10, pady=8)
        
        ctk.CTkButton(top_v, text="+ ADD VEHICLE", font=self.font_btn, height=38, fg_color="#007bff", command=lambda: self.open_vehicle_form_popup("ADD")).pack(side="left", padx=4)
        ctk.CTkButton(top_v, text="✏️ MODIFY", font=self.font_btn, height=38, fg_color="#ffc107", text_color="black", command=lambda: self.open_vehicle_form_popup("MODIFY")).pack(side="left", padx=4)
        ctk.CTkButton(top_v, text="🗑️ DELETE", font=self.font_btn, height=38, fg_color="#dc3545", command=self.delete_vehicle).pack(side="left", padx=4)
        
        self.veh_search_var = tk.StringVar()
        s_entry = ctk.CTkEntry(top_v, font=self.font_body, placeholder_text="🔍 Search Vehicle...", textvariable=self.veh_search_var, width=190, height=38)
        s_entry.pack(side="right", padx=10)
        self.veh_search_var.trace_add("write", lambda *args: self.load_vehicle_data())
        
        table_frame = ctk.CTkFrame(self.tab_veh)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        cols = ("Brand", "Type", "Reg_No", "Reg_Expiry", "Insurance_Expiry", "Pollution_Expiry", "State_Permit", "All_India_Permit", "Fitness_Expiry")
        self.veh_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            self.veh_tree.heading(col, text=col)
            self.veh_tree.column(col, width=110, anchor="center")
            
        self.veh_tree.pack(side="left", fill="both", expand=True)
        self.load_vehicle_data()

    def load_vehicle_data(self):
        for item in self.veh_tree.get_children():
            self.veh_tree.delete(item)
        if os.path.exists(DB_FILE):
            df = pd.read_excel(DB_FILE, sheet_name="Vehicle_Master")
            query = self.veh_search_var.get().lower()
            for _, r in df.iterrows():
                row_str = " ".join(str(val) for val in r.values).lower()
                if query in row_str:
                    self.veh_tree.insert("", "end", values=(
                        r.get("Brand",""), r.get("Type",""), r.get("Reg_No",""),
                        format_date_to_ddmmyyyy(r.get("Reg_Expiry","")), format_date_to_ddmmyyyy(r.get("Insurance_Expiry","")),
                        format_date_to_ddmmyyyy(r.get("Pollution_Expiry","")), format_date_to_ddmmyyyy(r.get("State_Permit_Expiry","")),
                        format_date_to_ddmmyyyy(r.get("All_India_Permit_Expiry","")), format_date_to_ddmmyyyy(r.get("Fitness_Expiry",""))
                    ))

    def open_vehicle_form_popup(self, mode="ADD"):
        existing_data = None
        target_reg = None
        
        if mode == "MODIFY":
            sel = self.veh_tree.selection()
            if not sel:
                messagebox.showwarning("Warning", "Please select a Vehicle from the table.")
                return
            target_reg = str(self.veh_tree.item(sel[0])['values'][2])
            df = pd.read_excel(DB_FILE, sheet_name="Vehicle_Master")
            match = df[df["Reg_No"].astype(str) == target_reg]
            if not match.empty:
                existing_data = match.iloc[0]

        popup = ctk.CTkToplevel(self)
        popup.title(f"{mode} Vehicle Master - Travel Mate")
        popup.geometry("800x780")
        popup.grab_set()
        
        scroll = ctk.CTkScrollableFrame(popup)
        scroll.pack(fill="both", expand=True, padx=15, pady=15)
        
        hdr_frame = ctk.CTkFrame(scroll, fg_color="#1a365d", corner_radius=8)
        hdr_frame.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(hdr_frame, text=f"🚗 VEHICLE MASTER SPECIFICATION", font=self.font_header, text_color="#ffffff").pack(side="left", padx=15, pady=10)
        
        vc1 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        vc1.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(vc1, text="🚘 Vehicle Information & Registration", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        vg1 = ctk.CTkFrame(vc1, fg_color="transparent")
        vg1.pack(fill="x", padx=10, pady=(0, 10))
        vg1.grid_columnconfigure((0, 1), weight=1)
        
        b = self.create_grid_field(vg1, "Vehicle Brand (e.g. Toyota, Maruti):", row=0, col=0, default=str(existing_data.get("Brand","")) if existing_data is not None else "")
        t = self.create_grid_field(vg1, "Vehicle Model / Type (e.g. Innova Crysta, Sedan):", row=0, col=1, default=str(existing_data.get("Type","")) if existing_data is not None else "")
        r = self.create_grid_field(vg1, "Registration Plate Number (e.g. DL01AB1234):", row=1, col=0, default=str(existing_data.get("Reg_No","")) if existing_data is not None else "")
        reg_exp = self.create_date_field(vg1, "Registration Expiry Date (DD/MM/YYYY):", row=1, col=1, default=format_date_to_ddmmyyyy(existing_data.get("Reg_Expiry","2032-01-01") if existing_data is not None else "2032-01-01"))

        vc2 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        vc2.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(vc2, text="📅 Compliance, Fitness & Permit Expiry Validations", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        vg2 = ctk.CTkFrame(vc2, fg_color="transparent")
        vg2.pack(fill="x", padx=10, pady=(0, 10))
        vg2.grid_columnconfigure((0, 1), weight=1)
        
        ins_exp = self.create_date_field(vg2, "Insurance Expiry Date (DD/MM/YYYY):", row=0, col=0, default=format_date_to_ddmmyyyy(existing_data.get("Insurance_Expiry","2027-01-01") if existing_data is not None else "2027-01-01"))
        pol_exp = self.create_date_field(vg2, "Pollution (PUC) Expiry Date (DD/MM/YYYY):", row=0, col=1, default=format_date_to_ddmmyyyy(existing_data.get("Pollution_Expiry","2027-01-01") if existing_data is not None else "2027-01-01"))
        st_perm = self.create_date_field(vg2, "State Permit Expiry Date (DD/MM/YYYY):", row=1, col=0, default=format_date_to_ddmmyyyy(existing_data.get("State_Permit_Expiry","2027-01-01") if existing_data is not None else "2027-01-01"))
        ai_perm = self.create_date_field(vg2, "All India Permit Expiry Date (DD/MM/YYYY):", row=1, col=1, default=format_date_to_ddmmyyyy(existing_data.get("All_India_Permit_Expiry","2027-01-01") if existing_data is not None else "2027-01-01"))
        fit_exp = self.create_date_field(vg2, "Fitness Certificate Expiry (DD/MM/YYYY):", row=2, col=0, default=format_date_to_ddmmyyyy(existing_data.get("Fitness_Expiry","2027-01-01") if existing_data is not None else "2027-01-01"))

        def save_veh():
            if not r.get().strip():
                messagebox.showerror("Error", "Registration Number is required.")
                return
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_v = excel_dict.get("Vehicle_Master", pd.DataFrame())
            
            row_data = {
                "Brand": b.get(),
                "Type": t.get(),
                "Reg_No": r.get(),
                "Reg_Expiry": parse_date_to_yyyymmdd(reg_exp.get()),
                "Insurance_Expiry": parse_date_to_yyyymmdd(ins_exp.get()),
                "Pollution_Expiry": parse_date_to_yyyymmdd(pol_exp.get()),
                "State_Permit_Expiry": parse_date_to_yyyymmdd(st_perm.get()),
                "All_India_Permit_Expiry": parse_date_to_yyyymmdd(ai_perm.get()),
                "Fitness_Expiry": parse_date_to_yyyymmdd(fit_exp.get())
            }
            
            if mode == "MODIFY" and not df_v.empty and "Reg_No" in df_v.columns:
                df_v = df_v[df_v["Reg_No"].astype(str) != str(target_reg)]
                
            df_v = pd.concat([df_v, pd.DataFrame([row_data])], ignore_index=True)
            excel_dict["Vehicle_Master"] = df_v
            
            with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                for s_name, s_df in excel_dict.items():
                    s_df.to_excel(writer, sheet_name=s_name, index=False)
                    
            messagebox.showinfo("Saved", "Vehicle details successfully saved!")
            popup.destroy()
            self.load_vehicle_data()
            
        ctk.CTkButton(scroll, text="💾 SAVE / UPDATE VEHICLE SPECIFICATION", font=self.font_btn, fg_color="#28a745", height=44, command=save_veh).pack(fill="x", pady=15)

    def delete_vehicle(self):
        sel = self.veh_tree.selection()
        if not sel:
            messagebox.showwarning("Warning", "Please select a Vehicle from the table.")
            return
        reg_no = str(self.veh_tree.item(sel[0])['values'][2])
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete vehicle {reg_no}?"):
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_v = excel_dict.get("Vehicle_Master", pd.DataFrame())
            if not df_v.empty and "Reg_No" in df_v.columns:
                df_v = df_v[df_v["Reg_No"].astype(str) != str(reg_no)]
                excel_dict["Vehicle_Master"] = df_v
                with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                    for s_name, s_df in excel_dict.items():
                        s_df.to_excel(writer, sheet_name=s_name, index=False)
            self.load_vehicle_data()

    def setup_driver_tab(self):
        top_d = ctk.CTkFrame(self.tab_drv)
        top_d.pack(fill="x", padx=10, pady=8)
        
        ctk.CTkButton(top_d, text="+ ADD DRIVER", font=self.font_btn, height=38, fg_color="#007bff", command=lambda: self.open_driver_form_popup("ADD")).pack(side="left", padx=4)
        ctk.CTkButton(top_d, text="✏️ MODIFY", font=self.font_btn, height=38, fg_color="#ffc107", text_color="black", command=lambda: self.open_driver_form_popup("MODIFY")).pack(side="left", padx=4)
        ctk.CTkButton(top_d, text="🗑️ DELETE", font=self.font_btn, height=38, fg_color="#dc3545", command=self.delete_driver).pack(side="left", padx=4)
        ctk.CTkButton(top_d, text="🖨️ GENERATE DRIVER ID CARD", font=self.font_btn, height=38, fg_color="#6f42c1", hover_color="#59359a", command=self.generate_driver_card_pdf).pack(side="left", padx=4)
        
        self.drv_search_var = tk.StringVar()
        s_entry = ctk.CTkEntry(top_d, font=self.font_body, placeholder_text="🔍 Search Driver...", textvariable=self.drv_search_var, width=190, height=38)
        s_entry.pack(side="right", padx=10)
        self.drv_search_var.trace_add("write", lambda *args: self.load_driver_data())
        
        table_frame = ctk.CTkFrame(self.tab_drv)
        table_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        cols = ("Driver_ID", "Name", "ID_No", "DL_No", "DL_Expiry", "Photo_Path")
        self.drv_tree = ttk.Treeview(table_frame, columns=cols, show="headings", selectmode="browse")
        for col in cols:
            self.drv_tree.heading(col, text=col)
            self.drv_tree.column(col, width=115, anchor="center")
            
        self.drv_tree.pack(side="left", fill="both", expand=True)
        self.load_driver_data()

    def generate_driver_card_pdf(self):
        sel = self.drv_tree.selection()
        if not sel:
            messagebox.showwarning("Warning", "Please select a Driver from the table to generate their ID card.")
            return
            
        vals = self.drv_tree.item(sel[0])['values']
        drv_id = str(vals[0])
        name = str(vals[1])
        id_no = str(vals[2])
        dl_no = str(vals[3])
        dl_exp = str(vals[4]) if len(vals) > 4 else ""
        photo_path = str(vals[5]) if len(vals) > 5 else ""
        
        df_comp = pd.read_excel(DB_FILE, sheet_name="Company_Profile")
        comp = df_comp.iloc[0] if not df_comp.empty else {}
        comp_name = str(comp.get('Org_Name', 'TRAVEL MATE RENTALS PVT. LTD.'))
        comp_addr = str(comp.get('Address', ''))
        comp_phone = str(comp.get('Mobile', ''))
        
        card_pdf_path = os.path.join(DRIVER_CARD_DIR, f"Driver_Card_{drv_id}.pdf")
        doc = SimpleDocTemplate(card_pdf_path, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        styles = getSampleStyleSheet()
        elements = []
        
        header_text = [
            [Paragraph(f"<b><font size=16 color='#1a365d'>{comp_name}</font></b><br/><font size=8.5 color='#4a5568'>{comp_addr} | Phone: {comp_phone}</font>", styles['Normal'])]
        ]
        h_table = Table(header_text, colWidths=[525])
        h_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        
        elements.append(h_table)
        elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#1a365d'), spaceBefore=2, spaceAfter=12))
        
        elements.append(Paragraph("<b><font size=13 color='#2b6cb0'>OFFICIAL DRIVER IDENTITY & CREDENTIAL CARD</font></b>", styles['Normal']))
        elements.append(Spacer(1, 10))
        
        img_flowable = Paragraph("<b>No Photo Available</b>", styles['Normal'])
        if photo_path and os.path.exists(photo_path):
            try:
                img_flowable = Image(photo_path, width=105, height=125)
            except Exception:
                pass
                
        card_details = [
            [img_flowable, Paragraph(f"<b>Driver ID:</b> {drv_id}<br/><br/><b>Full Name:</b> {name}<br/><br/><b>ID Proof No:</b> {id_no}<br/><br/><b>Driving License (DL):</b> {dl_no} (Exp: {dl_exp})", styles['Normal'])]
        ]
        
        c_table = Table(card_details, colWidths=[130, 395])
        c_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f7fafc')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e0')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('PADDING', (0,0), (-1,-1), 10),
        ]))
        
        elements.append(c_table)
        elements.append(Spacer(1, 15))
        elements.append(Paragraph("<b>Terms & Conditions:</b><br/>1. This card must be carried by the driver at all times during duty.<br/>2. If found, please return to the company address mentioned above.", styles['Normal']))
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("Authorized Signatory / Director: ____________________________", styles['Normal']))
        
        doc.build(elements)
        messagebox.showinfo("Success", f"Driver ID Card PDF generated successfully:\n{card_pdf_path}")
        if os.name == 'nt':
            os.startfile(pdf_path)

    def load_driver_data(self):
        for item in self.drv_tree.get_children():
            self.drv_tree.delete(item)
        if os.path.exists(DB_FILE):
            df = pd.read_excel(DB_FILE, sheet_name="Driver_Master")
            query = self.drv_search_var.get().lower()
            for _, r in df.iterrows():
                row_str = " ".join(str(val) for val in r.values).lower()
                if query in row_str:
                    self.drv_tree.insert("", "end", values=(
                        r.get("Driver_ID",""), r.get("Name",""), r.get("ID_No",""),
                        r.get("DL_No",""), format_date_to_ddmmyyyy(r.get("DL_Expiry","")), r.get("Photo_Path","")
                    ))

    def open_driver_form_popup(self, mode="ADD"):
        existing_data = None
        target_did = None
        
        if mode == "MODIFY":
            sel = self.drv_tree.selection()
            if not sel:
                messagebox.showwarning("Warning", "Please select a Driver from the table.")
                return
            target_did = str(self.drv_tree.item(sel[0])['values'][0])
            df = pd.read_excel(DB_FILE, sheet_name="Driver_Master")
            match = df[df["Driver_ID"].astype(str) == target_did]
            if not match.empty:
                existing_data = match.iloc[0]

        popup = ctk.CTkToplevel(self)
        popup.title(f"{mode} Driver Master - Travel Mate")
        popup.geometry("780x680")
        popup.grab_set()
        
        scroll = ctk.CTkScrollableFrame(popup)
        scroll.pack(fill="both", expand=True, padx=15, pady=15)
        
        drv_id = target_did if mode == "MODIFY" else f"DRV-{datetime.datetime.now().strftime('%M%S')}"
        
        hdr_frame = ctk.CTkFrame(scroll, fg_color="#1a365d", corner_radius=8)
        hdr_frame.pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(hdr_frame, text=f"👨‍✈️ DRIVER CREDENTIAL SPECIFICATION: {drv_id}", font=self.font_header, text_color="#ffffff").pack(side="left", padx=15, pady=10)
        
        dc1 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        dc1.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(dc1, text="🪪 Identity, License & Verification", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        dg1 = ctk.CTkFrame(dc1, fg_color="transparent")
        dg1.pack(fill="x", padx=10, pady=(0, 10))
        dg1.grid_columnconfigure((0, 1), weight=1)
        
        n = self.create_grid_field(dg1, "Driver Full Name:", row=0, col=0, default=str(existing_data.get("Name","")) if existing_data is not None else "")
        id_no = self.create_grid_field(dg1, "Driver Identity / ID Card No:", row=0, col=1, default=str(existing_data.get("ID_No","")) if existing_data is not None else "")
        dl = self.create_grid_field(dg1, "Driving License (DL) Number:", row=1, col=0, default=str(existing_data.get("DL_No","")) if existing_data is not None else "")
        dl_exp = self.create_date_field(dg1, "Driving License Expiry (DD/MM/YYYY):", row=1, col=1, default=format_date_to_ddmmyyyy(existing_data.get("DL_Expiry","2028-01-01") if existing_data is not None else "2028-01-01"))
        
        photo_path_var = tk.StringVar(value=str(existing_data.get("Photo_Path","")) if existing_data is not None else "")
        
        dc2 = ctk.CTkFrame(scroll, fg_color="#f8fafc", border_width=1, border_color="#cbd5e0", corner_radius=8)
        dc2.pack(fill="x", pady=6, padx=4)
        ctk.CTkLabel(dc2, text="📷 Official Photograph Archive", font=self.font_sub, text_color="#2b6cb0").pack(anchor="w", padx=12, pady=(8, 4))
        
        f_p = ctk.CTkFrame(dc2, fg_color="transparent")
        f_p.pack(fill="x", padx=10, pady=10)
        
        lbl_photo_status = ctk.CTkLabel(f_p, text=f"File: {os.path.basename(photo_path_var.get()) if photo_path_var.get() else 'No Photo Uploaded'}", font=self.font_body, text_color="#4a5568")
        lbl_photo_status.pack(side="left", padx=10)
        
        def choose_photo():
            f = filedialog.askopenfilename(filetypes=[("Image files", "*.jpg;*.jpeg;*.png")])
            if f:
                dest = os.path.join(PHOTO_DIR, f"{drv_id}_{os.path.basename(f)}")
                shutil.copy(f, dest)
                photo_path_var.set(dest)
                lbl_photo_status.configure(text=f"Uploaded: {os.path.basename(dest)}")
                messagebox.showinfo("Uploaded", "Driver photo saved successfully!")
                
        ctk.CTkButton(f_p, text="📷 Browse / Upload Photo", font=self.font_btn, fg_color="#17a2b8", height=38, command=choose_photo).pack(side="right", padx=10)
        
        def save_drv():
            if not n.get().strip():
                messagebox.showerror("Error", "Driver Name is required.")
                return
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_d = excel_dict.get("Driver_Master", pd.DataFrame())
            
            row_data = {
                "Driver_ID": drv_id,
                "Name": n.get(),
                "ID_No": id_no.get(),
                "DL_No": dl.get(),
                "DL_Expiry": parse_date_to_yyyymmdd(dl_exp.get()),
                "Photo_Path": photo_path_var.get()
            }
            
            if mode == "MODIFY" and not df_d.empty and "Driver_ID" in df_d.columns:
                df_d = df_d[df_d["Driver_ID"].astype(str) != str(target_did)]
                
            df_d = pd.concat([df_d, pd.DataFrame([row_data])], ignore_index=True)
            excel_dict["Driver_Master"] = df_d
            
            with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                for s_name, s_df in excel_dict.items():
                    s_df.to_excel(writer, sheet_name=s_name, index=False)
                    
            messagebox.showinfo("Saved", "Driver details saved successfully!")
            popup.destroy()
            self.load_driver_data()
            
        ctk.CTkButton(scroll, text="💾 SAVE / UPDATE DRIVER RECORD", font=self.font_btn, fg_color="#28a745", height=44, command=save_drv).pack(fill="x", pady=15)

    def delete_driver(self):
        sel = self.drv_tree.selection()
        if not sel:
            messagebox.showwarning("Warning", "Please select a Driver from the table.")
            return
        drv_id = str(self.drv_tree.item(sel[0])['values'][0])
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete Driver {drv_id}?"):
            excel_dict = pd.read_excel(DB_FILE, sheet_name=None)
            df_d = excel_dict.get("Driver_Master", pd.DataFrame())
            if not df_d.empty and "Driver_ID" in df_d.columns:
                df_d = df_d[df_d["Driver_ID"].astype(str) != str(drv_id)]
                excel_dict["Driver_Master"] = df_d
                with pd.ExcelWriter(DB_FILE, engine='openpyxl', mode='w') as writer:
                    for s_name, s_df in excel_dict.items():
                        s_df.to_excel(writer, sheet_name=s_name, index=False)
            self.load_driver_data()

    def create_form_field(self, parent, label_text, default=""):
        ctk.CTkLabel(parent, text=label_text, font=self.font_label).pack(anchor="w", pady=(8, 2))
        entry = ctk.CTkEntry(parent, font=self.font_body, height=38)
        entry.insert(0, default)
        entry.pack(fill="x", pady=(0, 8))
        return entry

    def get_master_options(self, sheet, col_name):
        if os.path.exists(DB_FILE):
            try:
                df = pd.read_excel(DB_FILE, sheet_name=sheet)
                if col_name in df.columns:
                    return [str(val) for val in df[col_name].dropna().tolist() if str(val).strip() != ""]
            except Exception:
                pass
        return []

if __name__ == "__main__":
    app = CabMISApp()
    app.mainloop()