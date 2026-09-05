"""
MSME Section 43B(h) Compliance Ledger - Desktop Application (v2)
S. D. Nikam & Company

Adds: direct import of Tally "bill-wise" (Purchase Register / Payables) and
"payment-wise" (Bank Book / Cash Book) Excel exports, automatic matching of
payments to bills (by bill reference where available, falling back to
oldest-bill-first allocation per vendor), and a one-time-per-vendor MSME
category tagger that is remembered between runs.

Extra dependency: openpyxl (for reading .xlsx). Everything else is
still standard library.

    pip install openpyxl

If Tally exports .xls (old format) instead of .xlsx, open it once in
Excel and "Save As" .xlsx before importing here.
"""

import csv
import io
import json
import os
import re
import sqlite3
import sys
from collections import defaultdict
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                     Spacer, PageBreak)
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_OK = True

    # Base14 PDF fonts (Helvetica etc.) have no \u20b9 (Rupee) glyph, so amounts
    # would render as a missing-glyph box. DejaVu Sans does include it and
    # ships with most Windows/Linux systems (also bundled with LibreOffice,
    # Chrome, and many printers) \u2014 register it if found, else fall back to
    # plain 'Rs.' formatting so the PDF still renders correctly either way.
    PDF_FONT, PDF_FONT_BOLD = "Helvetica", "Helvetica-Bold"
    PDF_USES_RUPEE_SYMBOL = False
    _DEJAVU_CANDIDATES = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:\\Windows\\Fonts\\DejaVuSans.ttf",
        "C:\\Windows\\Fonts\\Arial.ttf",  # Arial doesn't have \u20b9, kept only as last resort below
    ]
    for _candidate in _DEJAVU_CANDIDATES[:2]:  # only the DejaVu ones actually have the glyph
        if os.path.exists(_candidate):
            try:
                bold_candidate = _candidate.replace("DejaVuSans.ttf", "DejaVuSans-Bold.ttf")
                pdfmetrics.registerFont(TTFont("DejaVuSans", _candidate))
                if os.path.exists(bold_candidate):
                    pdfmetrics.registerFont(TTFont("DejaVuSans-Bold", bold_candidate))
                    PDF_FONT_BOLD = "DejaVuSans-Bold"
                else:
                    PDF_FONT_BOLD = "DejaVuSans"
                PDF_FONT = "DejaVuSans"
                PDF_USES_RUPEE_SYMBOL = True
                break
            except Exception:
                pass
except ImportError:
    REPORTLAB_OK = False
    PDF_FONT, PDF_FONT_BOLD = "Helvetica", "Helvetica-Bold"
    PDF_USES_RUPEE_SYMBOL = False


def fmt_inr_pdf(n):
    """Same as fmt_inr but falls back to 'Rs.' when the PDF font in use has
    no Rupee glyph, to avoid a missing-character box in the exported PDF."""
    try:
        n = float(n)
    except (TypeError, ValueError):
        n = 0
    s = f"{n:,.0f}"
    return f"\u20b9{s}" if PDF_USES_RUPEE_SYMBOL else f"Rs. {s}"

APP_TITLE = "MSME 43B(h) Compliance Ledger"

# ---- modern web-style palette ----
BG_PAGE = "#EEF1F8"
CARD_BG = "#FFFFFF"
INK = "#1E293B"
MUTED = "#64748B"
BORDER = "#E2E8F0"

PRIMARY = "#4F46E5"        # indigo
PRIMARY_DARK = "#3730A3"
TEAL = "#0D9488"           # compute / positive
GOLD = "#F59E0B"           # import actions
CORAL = "#EF4444"          # disallow / danger
GREEN = "#10B981"          # compliant

ROW_OK_BG = "#ECFDF5"
ROW_DISALLOW_BG = "#FEF2F2"
ROW_NA_BG = "#F8FAFC"

MAROON = CORAL  # kept for backward-compat with existing code paths


def shade(hex_color, factor):
    """Lighten (factor>1) or darken (factor<1) a hex color for hover states."""
    hex_color = hex_color.lstrip("#")
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = max(0, min(255, int(r * factor)))
    g = max(0, min(255, int(g * factor)))
    b = max(0, min(255, int(b * factor)))
    return f"#{r:02x}{g:02x}{b:02x}"


def lerp_color(c1, c2, t):
    c1, c2 = c1.lstrip("#"), c2.lstrip("#")
    r1, g1, b1 = int(c1[0:2], 16), int(c1[2:4], 16), int(c1[4:6], 16)
    r2, g2, b2 = int(c2[0:2], 16), int(c2[2:4], 16), int(c2[4:6], 16)
    r = int(r1 + (r2 - r1) * t)
    g = int(g1 + (g2 - g1) * t)
    b = int(b1 + (b2 - b1) * t)
    return f"#{r:02x}{g:02x}{b:02x}"


def rounded_rect(canvas, x1, y1, x2, y2, radius=14, **kwargs):
    points = [
        x1 + radius, y1, x2 - radius, y1, x2, y1, x2, y1 + radius,
        x2, y2 - radius, x2, y2, x2 - radius, y2, x1 + radius, y2,
        x1, y2, x1, y2 - radius, x1, y1 + radius, x1, y1,
    ]
    return canvas.create_polygon(points, smooth=True, **kwargs)


class HoverButton(tk.Button):
    """A flat, colorful button that brightens slightly on hover."""

    def __init__(self, master, bg, fg="#FFFFFF", padx=16, pady=9, **kwargs):
        super().__init__(master, bg=bg, fg=fg, activebackground=shade(bg, 0.88),
                          activeforeground=fg, relief="flat", bd=0,
                          font=("Segoe UI", 10, "bold"), padx=padx, pady=pady,
                          cursor="hand2", **kwargs)
        self._base_bg = bg
        self._hover_bg = shade(bg, 1.12)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    def _on_enter(self, _e):
        self.config(bg=self._hover_bg)

    def _on_leave(self, _e):
        self.config(bg=self._base_bg)

SAMPLE_CSV = """Vendor Name,Category,Invoice Date,Invoice Amount,Agreement Days,Payment Date
Shree Enterprises,Micro,05/06/2025,185000,,20/06/2025
Konkan Fabricators,Small,12/06/2025,420000,45,10/08/2025
Bright Packaging Co,Micro,01/07/2025,96000,,25/07/2025
Om Traders,Small,15/07/2025,275000,30,01/09/2025
Vasant Industries,Micro,20/08/2025,54000,,
"Naik & Sons, Traders",Small,25/08/2025,318000,,05/09/2025
Sagar Enterprises,Medium,10/09/2025,150000,,15/12/2025
Mahalaxmi Steel Corp,Micro,02/09/2025,67000,,08/09/2025
Konkan Plastics,Small,18/09/2025,225000,45,25/10/2025
Ganesh Hardware Mart,Micro,22/09/2025,38000,,
"""

CATEGORIES = ["Micro", "Small", "Medium", "NotMSME"]


# ---------------------------------------------------------------- helpers --

def app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


CATEGORY_FILE = os.path.join(app_dir(), "vendor_categories.json")  # legacy, migrated on first run
DB_FILE = os.path.join(app_dir(), "msme_clients.db")


def db_connect():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = db_connect()
    conn.execute("""CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        trade_name TEXT,
        pan TEXT,
        gstin TEXT,
        created_at TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS bills (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
        fy_end TEXT NOT NULL,
        vendor TEXT, category TEXT, invoice_date TEXT, invoice_amount REAL,
        agreement_days TEXT, payment_date TEXT, amount_paid REAL,
        installments_json TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS vendor_categories (
        client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
        vendor_key TEXT NOT NULL,
        category TEXT,
        PRIMARY KEY (client_id, vendor_key)
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS report_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
        fy_end TEXT NOT NULL,
        report_type TEXT NOT NULL,
        file_name TEXT,
        file_path TEXT,
        generated_at TEXT
    )""")
    conn.commit()

    # ---- lightweight migration from the single-FY-per-client schema (v11) ----
    cols = [r[1] for r in conn.execute("PRAGMA table_info(clients)").fetchall()]
    if "fy_end" in cols:
        # v11 clients table had a single fy_end column; carry it forward as
        # each client's one existing financial year, then drop the column.
        old_rows = conn.execute("SELECT id, fy_end FROM clients").fetchall()
        bill_cols = [r[1] for r in conn.execute("PRAGMA table_info(bills)").fetchall()]
        if "fy_end" not in bill_cols:
            conn.execute("ALTER TABLE bills ADD COLUMN fy_end TEXT")
        for cid, old_fy in old_rows:
            conn.execute("UPDATE bills SET fy_end = ? WHERE client_id = ? AND (fy_end IS NULL OR fy_end = '')",
                         (old_fy or "31/03/2026", cid))
        # Rebuild clients table without fy_end (SQLite can't drop columns pre-3.35 reliably across all builds)
        conn.execute("""CREATE TABLE clients_new (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL, trade_name TEXT, pan TEXT, gstin TEXT, created_at TEXT
        )""")
        conn.execute("INSERT INTO clients_new (id, name, created_at) SELECT id, name, created_at FROM clients")
        conn.execute("DROP TABLE clients")
        conn.execute("ALTER TABLE clients_new RENAME TO clients")
        conn.commit()
    conn.close()


def list_clients():
    conn = db_connect()
    rows = conn.execute(
        "SELECT id, name, trade_name, pan, gstin FROM clients ORDER BY name COLLATE NOCASE").fetchall()
    conn.close()
    return rows  # [(id, name, trade_name, pan, gstin), ...]


def search_clients(query):
    conn = db_connect()
    q = f"%{query.strip()}%"
    rows = conn.execute(
        """SELECT id, name, trade_name, pan, gstin FROM clients
           WHERE name LIKE ? OR trade_name LIKE ? OR pan LIKE ? OR gstin LIKE ?
           ORDER BY name COLLATE NOCASE""", (q, q, q, q)).fetchall()
    conn.close()
    return rows


def create_client(name, trade_name="", pan="", gstin=""):
    conn = db_connect()
    conn.execute("INSERT INTO clients (name, trade_name, pan, gstin, created_at) VALUES (?, ?, ?, ?, ?)",
                 (name, trade_name, pan, gstin, datetime.now().isoformat()))
    conn.commit()
    cid = conn.execute("SELECT id FROM clients WHERE name = ?", (name,)).fetchone()[0]
    conn.close()
    return cid


def update_client_details(client_id, name, trade_name, pan, gstin):
    conn = db_connect()
    conn.execute("UPDATE clients SET name = ?, trade_name = ?, pan = ?, gstin = ? WHERE id = ?",
                 (name, trade_name, pan, gstin, client_id))
    conn.commit()
    conn.close()


def get_client(client_id):
    conn = db_connect()
    row = conn.execute("SELECT id, name, trade_name, pan, gstin FROM clients WHERE id = ?",
                        (client_id,)).fetchone()
    conn.close()
    return row


def delete_client_db(client_id):
    conn = db_connect()
    conn.execute("DELETE FROM bills WHERE client_id = ?", (client_id,))
    conn.execute("DELETE FROM vendor_categories WHERE client_id = ?", (client_id,))
    conn.execute("DELETE FROM report_history WHERE client_id = ?", (client_id,))
    conn.execute("DELETE FROM clients WHERE id = ?", (client_id,))
    conn.commit()
    conn.close()


def list_client_fys(client_id):
    """Every financial year that already has saved data for this client,
    most recent first (by FY-end date, descending)."""
    conn = db_connect()
    rows = conn.execute(
        "SELECT DISTINCT fy_end FROM bills WHERE client_id = ?", (client_id,)).fetchall()
    conn.close()
    fys = [r[0] for r in rows if r[0]]
    fys.sort(key=lambda s: parse_date(s) or datetime.min, reverse=True)
    return fys


def save_client_bills(client_id, fy_end, raw_rows):
    conn = db_connect()
    conn.execute("DELETE FROM bills WHERE client_id = ? AND fy_end = ?", (client_id, fy_end))
    for r in raw_rows:
        installments = r.get("payment_installments") or []
        installments_ser = json.dumps([[fmt_date(d) if d else None, amt] for d, amt in installments])
        conn.execute(
            """INSERT INTO bills (client_id, fy_end, vendor, category, invoice_date, invoice_amount,
                                   agreement_days, payment_date, amount_paid, installments_json)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (client_id, fy_end, r.get("vendor name", ""), r.get("category", ""), r.get("invoice date", ""),
             clean_amount(r.get("invoice amount", 0)), r.get("agreement days", ""),
             r.get("payment date", ""), clean_amount(r.get("amount paid", 0)), installments_ser))
    conn.commit()
    conn.close()


def load_client_bills(client_id, fy_end):
    conn = db_connect()
    rows = conn.execute(
        """SELECT vendor, category, invoice_date, invoice_amount, agreement_days,
                  payment_date, amount_paid, installments_json
           FROM bills WHERE client_id = ? AND fy_end = ?""", (client_id, fy_end)).fetchall()
    conn.close()
    raw_rows = []
    for vendor, category, invoice_date, invoice_amount, agreement_days, payment_date, amount_paid, inst_json in rows:
        installments = []
        if inst_json:
            for d_str, amt in json.loads(inst_json):
                installments.append((parse_date(d_str) if d_str else None, amt))
        raw_rows.append({
            "vendor name": vendor, "category": category, "invoice date": invoice_date,
            "invoice amount": invoice_amount, "agreement days": agreement_days,
            "payment date": payment_date, "amount paid": amount_paid,
            "payment_installments": installments,
        })
    return raw_rows


def save_client_categories(client_id, categories):
    # Vendor MSME category doesn't change per financial year (Udyam status is
    # a vendor attribute, not a yearly one), so this stays client-level.
    conn = db_connect()
    for vkey, cat in categories.items():
        conn.execute(
            """INSERT INTO vendor_categories (client_id, vendor_key, category) VALUES (?, ?, ?)
               ON CONFLICT(client_id, vendor_key) DO UPDATE SET category = excluded.category""",
            (client_id, vkey, cat))
    conn.commit()
    conn.close()


def load_client_categories(client_id):
    conn = db_connect()
    rows = conn.execute(
        "SELECT vendor_key, category FROM vendor_categories WHERE client_id = ?", (client_id,)).fetchall()
    conn.close()
    return dict(rows)


def log_report_export(client_id, fy_end, report_type, file_name, file_path):
    conn = db_connect()
    conn.execute(
        """INSERT INTO report_history (client_id, fy_end, report_type, file_name, file_path, generated_at)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (client_id, fy_end, report_type, file_name, file_path, datetime.now().isoformat()))
    conn.commit()
    conn.close()


def load_report_history(client_id, fy_end=None):
    conn = db_connect()
    if fy_end:
        rows = conn.execute(
            """SELECT generated_at, report_type, fy_end, file_name, file_path FROM report_history
               WHERE client_id = ? AND fy_end = ? ORDER BY generated_at DESC""", (client_id, fy_end)).fetchall()
    else:
        rows = conn.execute(
            """SELECT generated_at, report_type, fy_end, file_name, file_path FROM report_history
               WHERE client_id = ? ORDER BY generated_at DESC""", (client_id,)).fetchall()
    conn.close()
    return rows


def build_export_filename(client_name, report_type, fy_end, ext, target_dir=None):
    """Professional, collision-safe file name, e.g.
    ABC_PRIVATE_LIMITED_MSME_COMPLIANCE_FY2025-26.pdf \u2014 auto-versions
    (\u2026_v2, \u2026_v3\u2026) instead of silently overwriting an existing file."""
    fy_dt = parse_date(fy_end)
    fy_label = f"FY{fy_dt.year - 1}-{str(fy_dt.year)[-2:]}" if fy_dt else "FY"
    safe_client = re.sub(r"[^A-Za-z0-9]+", "_", client_name).strip("_").upper()
    safe_report = re.sub(r"[^A-Za-z0-9]+", "_", report_type).strip("_").upper()
    base = f"{safe_client}_{safe_report}_{fy_label}"
    name = f"{base}.{ext}"
    if target_dir:
        n = 2
        while os.path.exists(os.path.join(target_dir, name)):
            name = f"{base}_v{n}.{ext}"
            n += 1
    return name


def migrate_legacy_categories_if_any(client_id):
    """One-time import of the old single-file vendor_categories.json (from
    pre-multi-client versions of this app) into the first client created."""
    if not os.path.exists(CATEGORY_FILE):
        return
    try:
        with open(CATEGORY_FILE, "r", encoding="utf-8") as f:
            legacy = json.load(f)
        if legacy:
            save_client_categories(client_id, legacy)
        os.rename(CATEGORY_FILE, CATEGORY_FILE + ".migrated")
    except Exception:
        pass




def norm_vendor(name):
    return re.sub(r"[^A-Z0-9]", "", (name or "").upper())


def norm_ref(v):
    if v is None:
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(v).upper())


def norm_header(h):
    return re.sub(r"[^a-z0-9]", "", str(h or "").lower())


def parse_date(s):
    if isinstance(s, datetime):
        return s
    s = (s or "").strip() if isinstance(s, str) else s
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d-%b-%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(str(s), fmt)
        except ValueError:
            continue
    return None


def fmt_date(d):
    return d.strftime("%d/%m/%Y") if d else ""


def fmt_inr(n):
    try:
        n = float(n)
    except (TypeError, ValueError):
        n = 0
    return f"\u20b9{n:,.0f}"


def clean_amount(s):
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    s = str(s).replace(",", "").replace("\u20b9", "").strip()
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# ------------------------------------------------------- excel column map --

VOUCHER_ALIASES = {
    "vendor": ["vendor", "vendorname", "party", "particulars", "name", "ledgername", "ledger"],
    "date": ["billdate", "invoicedate", "paymentdate", "date", "voucherdate"],
    # Only genuine bill-reference fields go here. Deliberately EXCLUDES
    # "Vch No." / "Voucher No." \u2014 those are each report's own internal
    # running counter (Purchase vouchers numbered 1,2,3... independently of
    # Payment vouchers also numbered 1,2,3...), so matching on them produces
    # coincidental false matches rather than real bill-to-payment links.
    "ref": ["billref", "billreference", "againstref", "againstbill", "newref",
            "billno", "invoiceno"],
    "vchno": ["vchno", "voucherno", "vouchernumber"],
    "vchtype": ["vouchertype", "vchtype", "type"],
    "debit": ["debit", "debitamount", "dr"],
    "credit": ["credit", "creditamount", "cr"],
    "amount": ["amount", "billamount", "invoiceamount", "paidamount", "value", "grossamount"],
}

PURCHASE_TYPE_KEYWORDS = ["purchase"]
PAYMENT_TYPE_KEYWORDS = ["payment"]
# Anything that SETTLES / reduces a creditor's outstanding balance, even
# though it isn't a "Payment" voucher: Journal vouchers are routinely used
# for TDS adjustments, debit/credit-note knock-offs, and direct inter-party
# settlements, and a Payment Register alone will never show these.
SETTLEMENT_TYPE_KEYWORDS = ["payment", "journal", "debit note"]


def classify_voucher_type(vch_type):
    t = (str(vch_type) or "").strip().lower()
    if any(k in t for k in PURCHASE_TYPE_KEYWORDS):
        return "bill"
    if any(k in t for k in SETTLEMENT_TYPE_KEYWORDS):
        return "payment"
    return "ignore"


def detect_columns(headers, aliases):
    norm_headers = [norm_header(h) for h in headers]
    found = {}
    for field, alias_list in aliases.items():
        idx = None
        for i, nh in enumerate(norm_headers):
            if nh in alias_list:
                idx = i
                break
        if idx is None:
            for i, nh in enumerate(norm_headers):
                if any(a in nh for a in alias_list):
                    idx = i
                    break
        found[field] = idx
    return found


def read_excel_rows(path):
    """Returns (headers, list-of-row-tuples) from the first sheet with data.
    Tally exports typically carry 2-3 title/period rows above the real header,
    so we take the first row that looks like an actual column header (several
    short, non-numeric, non-date cells) rather than just '>=2 non-empty cells',
    which a title row can also satisfy."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(all_rows):
        non_empty = [c for c in row if c not in (None, "")]
        if len(non_empty) >= 3 and all(not isinstance(c, datetime) for c in non_empty):
            header_idx = i
            break
    if header_idx is None:
        for i, row in enumerate(all_rows):
            non_empty = [c for c in row if c not in (None, "")]
            if len(non_empty) >= 2:
                header_idx = i
                break
    if header_idx is None:
        return [], []
    headers = list(all_rows[header_idx])
    data_rows = all_rows[header_idx + 1:]
    return headers, data_rows


def extract_vouchers(headers, data_rows):
    """Tally 'register' exports (Purchase Register, Payment Register, Day Book)
    write one HEADER row per voucher (Date + Particulars + Vch Type + Vch No.,
    with the net amount under either Debit or Credit) followed by several
    breakdown lines (tax ledgers, bank ledger, etc.) that share no date. We
    only want the header rows \u2014 identified by the Date cell being populated
    \u2014 not the breakdown lines beneath them."""
    cols = detect_columns(headers, VOUCHER_ALIASES)
    if cols["vendor"] is None:
        return [], cols
    vouchers = []
    for row in data_rows:
        date_val = row[cols["date"]] if cols["date"] is not None else None
        inv_date = parse_date(date_val)
        if not inv_date:
            continue  # breakdown/continuation line, not a voucher header
        vendor = row[cols["vendor"]]
        if not vendor or str(vendor).strip().lower() in ("total", "grand total", ""):
            continue

        debit_v = clean_amount(row[cols["debit"]]) if cols["debit"] is not None else 0.0
        credit_v = clean_amount(row[cols["credit"]]) if cols["credit"] is not None else 0.0
        if debit_v and credit_v:
            amount = max(debit_v, credit_v)
        elif debit_v or credit_v:
            amount = debit_v or credit_v
        elif cols["amount"] is not None:
            amount = clean_amount(row[cols["amount"]])
        else:
            amount = 0.0
        if amount == 0:
            continue

        ref_val = row[cols["ref"]] if cols["ref"] is not None else None
        vchno_val = row[cols["vchno"]] if cols["vchno"] is not None else None
        vchtype_val = row[cols["vchtype"]] if cols["vchtype"] is not None else ""
        vouchers.append({
            "vendor": str(vendor).strip(),
            "date": inv_date,
            "amount": amount,
            "ref": norm_ref(ref_val),          # used for matching (only if genuine)
            "vchno": str(vchno_val).strip() if vchno_val not in (None, "") else "",  # display only
            "vchtype": str(vchtype_val or "").strip(),
        })
    return vouchers, cols


def load_bills_excel(path):
    headers, data_rows = read_excel_rows(path)
    vouchers, cols = extract_vouchers(headers, data_rows)
    if cols["vendor"] is None:
        raise ValueError("Couldn't find a Vendor/Particulars column in this file.")
    # If a Voucher Type column exists, keep only Purchase-type rows (drops any
    # Debit Notes / adjustment vouchers mixed into the same export).
    if cols["vchtype"] is not None:
        filtered = [v for v in vouchers if classify_voucher_type(v["vchtype"]) == "bill"]
        if filtered:
            vouchers = filtered
    return vouchers


def load_payments_excel(path):
    headers, data_rows = read_excel_rows(path)
    vouchers, cols = extract_vouchers(headers, data_rows)
    if cols["vendor"] is None:
        raise ValueError("Couldn't find a Vendor/Particulars column in this file.")
    if cols["vchtype"] is not None:
        filtered = [v for v in vouchers if classify_voucher_type(v["vchtype"]) == "payment"]
        if filtered:
            vouchers = filtered
    return vouchers


def is_grouped_ledger_format(all_rows, scan_limit=200):
    """Detects Tally's per-party 'Group Vouchers' / ledger-drill export, where
    each vendor gets its own 'Ledger: <name>' header row followed by its own
    repeated column-header row, rather than one flat table for everyone."""
    for row in all_rows[:scan_limit]:
        if row and str(row[0] or "").strip().lower() == "ledger:":
            return True
    return False


# Journal/Debit Note entries that INCREASE a creditor's balance are ambiguous
# by voucher type alone \u2014 real-world books route many different things
# through a creditor's ledger this way: genuine purchases of goods, genuine
# service bills (professional fees, AMC), but also things that can never be
# a vendor supply at all (salary, drawings, statutory dues, rounding). Rather
# than guess which trade categories count (too many ways to phrase "material
# purchase" \u2014 tiles, hardware, ceramics, spares...), we exclude only the
# narrow, unambiguous set that structurally can't be a purchase from a
# supplier, and let genuine-looking vendor items through for the accountant
# to classify (or dismiss as NotMSME) in the category-tagging step.
NON_VENDOR_JOURNAL_KEYWORDS = [
    "salary", "wages", "drawing", "interest", "dividend", "tds", "gst payable", "gst input",
    "provident fund", "esic", "bank charg", "rounded diff", "round off", "rounding",
    "depreciation", "provision for", "capital introduced", "loan from", "loan to",
    "advance tax", "income tax", "penalty", "fine paid", "donation", "profit and loss",
    "rent", "commission", "reimbursement", "remuneration", "incentive", "bonus",
    "conveyance", "electricity charg", "telephone charg", "mobile bill", "water charg",
    "staff welfare", "festival", "entertainment exp", "director's remuneration",
    "licence fee", "license fee",
]


def classify_grouped_row(vchtype, particulars, debit_v, credit_v):
    """Direction for a row in the per-party grouped ledger. Genuine Purchase
    and Payment vouchers are unambiguous regardless of column. A Journal or
    Debit Note is ambiguous \u2014 it's only a genuine settlement/adjustment
    against a real bill if it REDUCES the balance (Debit side). If it
    INCREASES the balance (Credit side), it's treated as a candidate bill
    UNLESS the narration matches an unambiguous non-vendor item (salary,
    drawings, statutory dues, rounding, etc.) \u2014 see NON_VENDOR_JOURNAL_KEYWORDS."""
    t = (vchtype or "").strip().lower()
    p = (particulars or "").strip().lower()
    if "purchase" in t:
        return "bill", (credit_v or debit_v)
    if "payment" in t:
        return "payment", (debit_v or credit_v)
    if "journal" in t or "debit note" in t:
        if debit_v > credit_v:
            return "payment", debit_v  # reduces balance -> genuine settlement/adjustment
        if any(k in p for k in NON_VENDOR_JOURNAL_KEYWORDS):
            return "ignore", 0.0  # structurally can't be a vendor supply
        return "bill", credit_v  # candidate purchase -- accountant reviews via category tagging
    return "ignore", 0.0


def load_grouped_ledger_rows(all_rows):
    """Parses the per-party grouped ledger shape:
        ('Ledger:', '<Vendor Name>', ...)          <- sets current vendor
        ('Date','Particulars',None,'Vch Type',...) <- repeated header, skip
        (date, 'Dr'/'Cr', <contra ledger>, VchType, VchNo, Debit, Credit)
        (<total>, None, ..., None, <total>)        <- per-vendor subtotal, skip
    The vendor name only appears on the 'Ledger:' row, not on each voucher
    line, so we carry it forward until the next 'Ledger:' marker. Direction
    is decided per classify_grouped_row \u2014 voucher type first, Dr/Cr side
    and narration only to disambiguate Journal/Debit Note."""
    vouchers = []
    ignored_count = 0
    current_vendor = None
    for row in all_rows:
        if not row or all(c in (None, "") for c in row):
            continue
        c0 = row[0]
        if isinstance(c0, str) and c0.strip().lower() == "ledger:":
            current_vendor = str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "") else None
            continue
        if isinstance(c0, str) and c0.strip().lower() == "date":
            continue  # repeated per-vendor column header
        inv_date = parse_date(c0)
        if not inv_date or not current_vendor:
            continue  # subtotal row, blank row, or voucher line before any Ledger: marker

        particulars = row[2] if len(row) > 2 else None
        vchtype = row[3] if len(row) > 3 else None
        vchno = row[4] if len(row) > 4 else None
        debit_v = clean_amount(row[5]) if len(row) > 5 else 0.0
        credit_v = clean_amount(row[6]) if len(row) > 6 else 0.0
        if not debit_v and not credit_v:
            continue

        kind, amount = classify_grouped_row(vchtype, particulars, debit_v, credit_v)
        if kind == "ignore" or not amount:
            ignored_count += 1
            continue

        vouchers.append({
            "vendor": current_vendor,
            "date": inv_date,
            "amount": amount,
            "ref": "",  # this report doesn't carry a genuine bill reference either
            "vchno": str(vchno).strip() if vchno not in (None, "") else "",
            "vchtype": str(vchtype or "").strip(),
            "kind": kind,
        })
    return vouchers, ignored_count


def load_multi_voucher_excel(path):
    """Reads a single export containing multiple voucher types \u2014 a
    whole-company Day Book, a flat 'Purchase/Payment Register'-style table,
    OR a per-party grouped Sundry Creditors ledger drill (auto-detected) \u2014
    and splits it into bills (Purchase vouchers) and settlements
    (Payment/Journal/Debit Note vouchers) using the Voucher Type column. Any
    other voucher type (Sales, Receipt, Contra, etc.) is counted but
    skipped."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    if is_grouped_ledger_format(all_rows):
        vouchers, ignored_count = load_grouped_ledger_rows(all_rows)
        if not vouchers:
            raise ValueError(
                "Found 'Ledger:' section markers (a per-party grouped export) but couldn't read any "
                "voucher rows underneath them. Check the export wasn't altered after being saved from Tally."
            )
        bills = [v for v in vouchers if v["kind"] == "bill"]
        payments = [v for v in vouchers if v["kind"] == "payment"]
        return bills, payments, ignored_count

    headers, data_rows = read_excel_rows(path)
    vouchers, cols = extract_vouchers(headers, data_rows)
    if cols["vchtype"] is None:
        raise ValueError(
            "Couldn't find a 'Voucher Type' column in this file, and it doesn't look like a "
            "per-party grouped ledger export either. This import needs a report with the "
            "Voucher Type column visible \u2014 check F12 configuration in Tally, or use the "
            "two-file import instead."
        )

    bills, payments, ignored_count = [], [], 0
    for v in vouchers:
        kind = classify_voucher_type(v["vchtype"])
        if kind == "bill":
            bills.append(v)
        elif kind == "payment":
            payments.append(v)
        else:
            ignored_count += 1
    return bills, payments, ignored_count


def merge_bills_and_payments(bills, payments, saved_categories):
    """Matches payments to bills: first by explicit bill reference (per vendor),
    then falls back to oldest-bill-first (FIFO) allocation per vendor for
    anything left unmatched. Returns raw_rows in the same shape compute() expects.
    Every individual installment is tracked (not just a summed total), so
    part-paid bills and same-year-but-late payments can be assessed correctly."""

    for b in bills:
        b["_installments"] = []  # list of (date, amount) actually applied to this bill
        b["_matched"] = False

    used_payment_idx = set()

    # Phase 1: reference-based matching
    ref_index = defaultdict(list)
    for i, p in enumerate(payments):
        if p["ref"]:
            ref_index[(norm_vendor(p["vendor"]), p["ref"])].append(i)

    for b in bills:
        key = (norm_vendor(b["vendor"]), b["ref"])
        if b["ref"] and key in ref_index:
            matched_idxs = [i for i in ref_index[key] if i not in used_payment_idx]
            if matched_idxs:
                for i in matched_idxs:
                    used_payment_idx.add(i)
                    b["_installments"].append((payments[i]["date"], payments[i]["amount"]))
                b["_matched"] = True

    # Phase 2: FIFO fallback per vendor for anything not ref-matched
    vendor_bills = defaultdict(list)
    vendor_payments = defaultdict(list)
    for b in bills:
        if not b["_matched"]:
            vendor_bills[norm_vendor(b["vendor"])].append(b)
    for i, p in enumerate(payments):
        if i not in used_payment_idx:
            vendor_payments[norm_vendor(p["vendor"])].append(dict(p, _remaining=p["amount"]))

    for vkey, blist in vendor_bills.items():
        blist.sort(key=lambda x: x["date"] or datetime.min)
        plist = vendor_payments.get(vkey, [])
        plist.sort(key=lambda x: x["date"] or datetime.min)
        pidx = 0
        for b in blist:
            need = b["amount"]
            while need > 0.01 and pidx < len(plist):
                avail = plist[pidx]["_remaining"]
                if avail <= 0.01:
                    pidx += 1
                    continue
                take = min(avail, need)
                need -= take
                plist[pidx]["_remaining"] -= take
                b["_installments"].append((plist[pidx]["date"], take))
                if plist[pidx]["_remaining"] <= 0.01:
                    pidx += 1

    raw_rows = []
    for b in bills:
        total_paid = sum(amt for _, amt in b["_installments"])
        latest_date = max((d for d, _ in b["_installments"] if d), default=None)
        raw_rows.append({
            "vendor name": b["vendor"],
            "category": saved_categories.get(norm_vendor(b["vendor"]), ""),
            "invoice date": fmt_date(b["date"]) if b["date"] else "",
            "invoice amount": b["amount"],
            "agreement days": "",
            "payment date": fmt_date(latest_date) if latest_date else "",
            "amount paid": total_paid,
            "payment_installments": b["_installments"],  # [(date, amount), ...] for precise FY-aware calc
        })
    return raw_rows


# --------------------------------------------------------------- the app --

class MSMEApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE + " \u2013 S. D. Nikam & Company")
        self.geometry("1340x820")
        self.minsize(1180, 680)
        self.configure(bg=INK)
        self.computed_rows = []
        self.raw_rows = []
        self.saved_categories = {}
        self.current_client_id = None
        self.current_client_name = None
        init_db()
        self._build_style()
        self._build_ui()
        self._init_clients()

    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=28,
                         background="#FFFFFF", fieldbackground="#FFFFFF", foreground=INK,
                         borderwidth=0)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"), foreground="#FFFFFF",
                         background=PRIMARY_DARK, relief="flat")
        style.map("Treeview.Heading", background=[("active", PRIMARY_DARK)])
        style.map("Treeview", background=[("selected", "#C7D2FE")], foreground=[("selected", INK)])

    def _gradient_header(self, parent):
        canvas = tk.Canvas(parent, height=118, highlightthickness=0, bd=0)
        canvas.pack(fill="x")

        def draw(event=None):
            canvas.delete("all")
            w = canvas.winfo_width() or 1200
            h = 118
            steps = 60
            for i in range(steps):
                t = i / steps
                color = lerp_color(PRIMARY, TEAL, t)
                x0 = int(w * i / steps)
                x1 = int(w * (i + 1) / steps) + 1
                canvas.create_rectangle(x0, 0, x1, h, fill=color, outline=color)
            canvas.create_oval(w - 150, -60, w + 40, 130, fill=shade(TEAL, 1.15), outline="", stipple="")
            canvas.create_text(34, 34, anchor="w", text="\U0001F4D2  WORKING PAPER \u00b7 TAX AUDIT ANNEXURE",
                                fill="#E0E7FF", font=("Segoe UI", 10, "bold"))
            canvas.create_text(34, 66, anchor="w", text=APP_TITLE,
                                fill="#FFFFFF", font=("Segoe UI", 22, "bold"))
            canvas.create_text(34, 96, anchor="w",
                                text="Section 43B(h), Income-tax Act 1961  \u00b7  Section 15, MSMED Act 2006  \u00b7  S. D. Nikam & Company",
                                fill="#E0E7FF", font=("Segoe UI", 9))

        canvas.bind("<Configure>", draw)
        return canvas

    def _build_ui(self):
        self.configure(bg=BG_PAGE)
        sheet = tk.Frame(self, bg=BG_PAGE)
        sheet.pack(fill="both", expand=True)

        self._gradient_header(sheet)

        body = tk.Frame(sheet, bg=BG_PAGE)
        body.pack(fill="both", expand=True, padx=22, pady=18)

        # ---- client bar ----
        client_card = tk.Frame(body, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
        client_card.pack(fill="x", pady=(0, 12))
        client_row = tk.Frame(client_card, bg=CARD_BG)
        client_row.pack(fill="x", padx=18, pady=(12, 6))

        tk.Label(client_row, text="\U0001F50D", bg=CARD_BG, fg=MUTED,
                  font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        self.client_search_var = tk.StringVar()
        search_entry = tk.Entry(client_row, textvariable=self.client_search_var, width=12,
                                 font=("Segoe UI", 10), relief="solid", bd=1)
        search_entry.pack(side="left", padx=(0, 8))
        search_entry.bind("<KeyRelease>", self._on_client_search)

        tk.Label(client_row, text="\U0001F4C1 Client", bg=CARD_BG, fg=MUTED,
                  font=("Segoe UI", 9, "bold")).pack(side="left", padx=(0, 6))
        self.client_var = tk.StringVar()
        self.client_combo = ttk.Combobox(client_row, textvariable=self.client_var, width=22,
                                          state="readonly", font=("Segoe UI", 10))
        self.client_combo.pack(side="left", padx=(0, 8))
        self.client_combo.bind("<<ComboboxSelected>>", self._on_client_selected)

        tk.Label(client_row, text="\U0001F4C5 FY End", bg=CARD_BG, fg=MUTED,
                  font=("Segoe UI", 9, "bold")).pack(side="left", padx=(4, 6))
        self.fy_var = tk.StringVar()
        self.fy_combo = ttk.Combobox(client_row, textvariable=self.fy_var, width=10,
                                      font=("Segoe UI", 10))
        self.fy_combo.pack(side="left", padx=(0, 3))
        self.fy_combo.bind("<<ComboboxSelected>>", self._on_fy_selected)
        HoverButton(client_row, text="Go", bg=TEAL, command=self._on_fy_go, padx=10).pack(side="left", padx=(0, 8))

        HoverButton(client_row, text="+ New", bg=PRIMARY, command=self.new_client, padx=10).pack(side="left", padx=2)
        HoverButton(client_row, text="\u270E Edit", bg=shade(MUTED, 1.0),
                    command=self.edit_client_details, padx=10).pack(side="left", padx=2)
        HoverButton(client_row, text="\U0001F5D1 Delete", bg=shade(CORAL, 0.9),
                    command=self.delete_current_client, padx=10).pack(side="left", padx=2)

        self.client_status_label = tk.Label(client_card, text="", bg=CARD_BG, fg=MUTED,
                                             font=("Segoe UI", 8, "italic"), anchor="w")
        self.client_status_label.pack(fill="x", padx=18, pady=(0, 10))

        # ---- tabbed workspace ----
        style = ttk.Style(self)
        style.configure("TNotebook.Tab", font=("Segoe UI", 10, "bold"), padding=(16, 8))
        notebook = ttk.Notebook(body)
        notebook.pack(fill="both", expand=True)

        tab_dashboard = tk.Frame(notebook, bg=BG_PAGE)
        tab_working = tk.Frame(notebook, bg=BG_PAGE)
        tab_clause26 = tk.Frame(notebook, bg=BG_PAGE)
        tab_history = tk.Frame(notebook, bg=BG_PAGE)
        notebook.add(tab_dashboard, text="\U0001F4CA  Dashboard")
        notebook.add(tab_working, text="\U0001F4D2  Ledger Working")
        notebook.add(tab_clause26, text="\u00A7  Clause 26 Report")
        notebook.add(tab_history, text="\U0001F553  Report History")
        self.notebook = notebook
        notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

        work_body = tk.Frame(tab_working, bg=BG_PAGE)
        work_body.pack(fill="both", expand=True, padx=0, pady=(12, 0))
        self._build_dashboard_tab(tab_dashboard)
        self._build_clause26_tab(tab_clause26)
        self._build_history_tab(tab_history)

        # ---- toolbar card ----
        toolbar_card = tk.Frame(work_body, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
        toolbar_card.pack(fill="x", pady=(0, 16))

        ctrl = tk.Frame(toolbar_card, bg=CARD_BG)
        ctrl.pack(fill="x", padx=18, pady=(16, 8))

        tk.Label(ctrl, text="Financial Year End", bg=CARD_BG, fg=MUTED,
                  font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", padx=(0, 8))
        self.fy_end_var = tk.StringVar(value="31/03/2026")
        tk.Label(ctrl, textvariable=self.fy_end_var, width=12, font=("Consolas", 10, "bold"),
                  bg=CARD_BG, fg=INK, anchor="w").grid(row=0, column=1, sticky="w")
        tk.Label(ctrl, text="(set via the Client/FY bar above)", bg=CARD_BG, fg=MUTED,
                  font=("Segoe UI", 7, "italic")).grid(row=1, column=0, columnspan=2, sticky="w")

        tk.Frame(ctrl, bg=BORDER, width=1, height=26).grid(row=0, column=2, padx=16)

        HoverButton(ctrl, text="\U0001F4D2 Creditors Ledger", bg=GOLD, padx=11,
                    command=self.import_creditors_ledger).grid(row=0, column=3, padx=(0, 6))
        HoverButton(ctrl, text="\U0001F4E5 Day Book", bg=shade(GOLD, 0.85), padx=11,
                    command=self.import_daybook).grid(row=0, column=4, padx=6)
        HoverButton(ctrl, text="\U0001F4C2 Bills + Payments", bg=PRIMARY, padx=11,
                    command=self.import_tally).grid(row=0, column=5, padx=6)

        ctrl2 = tk.Frame(toolbar_card, bg=CARD_BG)
        ctrl2.pack(fill="x", padx=18, pady=(0, 16))
        HoverButton(ctrl2, text="\U0001F4C4 Load CSV", bg=shade(MUTED, 1.0), padx=11,
                    command=self.load_csv_file).grid(row=0, column=0, padx=(0, 6))
        HoverButton(ctrl2, text="\U0001F9EA Sample Data", bg=shade(MUTED, 1.0), padx=11,
                    command=self.load_sample).grid(row=0, column=1, padx=6)
        HoverButton(ctrl2, text="\U0001F3F7\uFE0F Categories", bg=shade(MUTED, 1.0), padx=11,
                    command=self.manage_categories).grid(row=0, column=2, padx=6)
        HoverButton(ctrl2, text="\u2705 Compute", bg=TEAL, padx=11,
                    command=self.compute).grid(row=0, column=3, padx=6)
        HoverButton(ctrl2, text="\U0001F4E4 CSV", bg=PRIMARY_DARK, padx=11,
                    command=self.export_csv).grid(row=0, column=4, padx=6)
        HoverButton(ctrl2, text="\U0001F4D7 Excel", bg=GREEN, padx=11,
                    command=self.export_excel).grid(row=0, column=5, padx=6)
        HoverButton(ctrl2, text="\U0001F4C4 PDF", bg=shade(CORAL, 0.85), padx=11,
                    command=self.export_pdf).grid(row=0, column=6, padx=6)

        self.file_label = tk.Label(toolbar_card, text="No file loaded (using built-in sample data)",
                                    bg=CARD_BG, fg=MUTED, font=("Segoe UI", 8, "italic"), anchor="w")
        self.file_label.pack(fill="x", padx=18, pady=(0, 14))

        # ---- summary cards ----
        self.summary_frame = tk.Frame(work_body, bg=BG_PAGE)
        self.summary_frame.pack(fill="x", pady=(0, 16))
        self.summary_labels = {}
        self._build_summary_cards()

        # ---- table card ----
        table_card = tk.Frame(work_body, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
        table_card.pack(fill="both", expand=True, pady=(0, 16))

        table_frame = tk.Frame(table_card, bg=CARD_BG)
        table_frame.pack(fill="both", expand=True, padx=14, pady=14)

        cols = ("vendor", "category", "invdate", "amount", "paid", "duedate", "paydate", "status", "disallowed", "note")
        headers = ["Vendor", "Category", "Invoice Date", "Amount", "Paid", "Due Date", "Payment Date",
                   "Status", "Disallowed", "Note"]
        widths = [150, 90, 95, 100, 100, 130, 100, 100, 100, 260]

        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=14)
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.column("amount", anchor="e")
        self.tree.column("paid", anchor="e")
        self.tree.column("disallowed", anchor="e")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        self.tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self.tree.tag_configure("disallow", foreground=CORAL, background=ROW_DISALLOW_BG)
        self.tree.tag_configure("ok", foreground=TEAL, background=ROW_OK_BG)
        self.tree.tag_configure("na", foreground=MUTED, background=ROW_NA_BG)

        # ---- clause note ----
        clause_card = tk.Frame(work_body, bg="#EEF2FF", highlightbackground="#C7D2FE", highlightthickness=1)
        clause_card.pack(fill="x", pady=(0, 12))
        self.clause_var = tk.StringVar(value="")
        tk.Label(clause_card, textvariable=self.clause_var, bg="#EEF2FF", fg=INK,
                  font=("Segoe UI", 9), wraplength=1120, justify="left").pack(
            fill="x", padx=16, pady=12)

        footer = tk.Label(work_body, bg=BG_PAGE, fg=MUTED, font=("Segoe UI", 8), justify="left",
                           wraplength=1160,
                           text=("Due date: 15 days from acceptance if no written agreement; if a written "
                                 "agreement specifies a term, that term applies subject to a ceiling of 45 days "
                                 "(Sec 15, MSMED Act, 2006). Amounts unpaid beyond this limit are disallowed under "
                                 "Section 43B(h) in the year of accrual and become allowable only in the year of "
                                 "actual payment. Independently verify each vendor's Udyam registration before "
                                 "relying on this output for filing. Payment matching against bill references is "
                                 "automated best-effort \u2014 spot-check a few vendors after import."))
        footer.pack(fill="x")

    def _stat_card(self, parent, icon, label, color):
        wrap = tk.Frame(parent, bg=BG_PAGE)
        canvas = tk.Canvas(wrap, height=92, highlightthickness=0, bd=0, bg=BG_PAGE)
        canvas.pack(fill="both", expand=True)
        canvas._current_value = "\u2014"

        def draw(event=None):
            canvas.delete("all")
            w = canvas.winfo_width() or 260
            h = 92
            rounded_rect(canvas, 2, 2, w - 2, h - 2, radius=14, fill=CARD_BG, outline=BORDER)
            rounded_rect(canvas, 2, 2, 8, h - 2, radius=0, fill=color, outline=color)
            canvas.create_text(24, 22, anchor="w", text=icon, font=("Segoe UI", 16))
            canvas.create_text(w - 18, 40, anchor="e", text=canvas._current_value,
                                font=("Consolas", 20, "bold"), fill=color, tags="value")
            canvas.create_text(24, 70, anchor="w", text=label.upper(), font=("Segoe UI", 8, "bold"),
                                fill=MUTED, width=w - 40)

        canvas.bind("<Configure>", draw)
        canvas._redraw = draw
        return wrap, canvas

    def _build_summary_cards(self):
        for widget in self.summary_frame.winfo_children():
            widget.destroy()
        cards = [("msme", "\U0001F3E2", "MSME Vendors", PRIMARY),
                 ("purchase", "\U0001F4B0", "Total MSME Purchases", PRIMARY_DARK),
                 ("disallowed", "\u26A0\uFE0F", "Disallowed u/s 43B(h)", CORAL),
                 ("compliant", "\u2705", "Compliant Value", GREEN)]
        self.summary_canvases = {}
        for i, (key, icon, label, color) in enumerate(cards):
            wrap, canvas = self._stat_card(self.summary_frame, icon, label, color)
            wrap.grid(row=0, column=i, sticky="nsew", padx=(0 if i == 0 else 8, 0))
            self.summary_frame.grid_columnconfigure(i, weight=1)
            self.summary_canvases[key] = canvas
            self.summary_labels[key] = canvas  # keep attribute name for compatibility

    def _set_stat(self, key, text):
        canvas = self.summary_canvases[key]
        canvas._current_value = text
        canvas._redraw()

    # -------------------------------------------------------------- data --

    # ------------------------------------------------------------ clients --

    def _init_clients(self):
        clients = list_clients()
        if not clients:
            cid = create_client("Demo Client")
            migrate_legacy_categories_if_any(cid)
            self.current_client_id = cid
            self.current_client_name = "Demo Client"
            self._refresh_client_combo(select_id=cid)
            self.fy_end_var.set("31/03/2026")
            self.fy_var.set("31/03/2026")
            self._refresh_fy_combo()
            self.client_status_label.config(text="Working on: Demo Client")
            self.load_sample()  # friendly first-run content, gets persisted on compute
        else:
            self._refresh_client_combo(select_id=clients[0][0])
            self._load_client(clients[0][0])

    def _refresh_client_combo(self, select_id=None):
        clients = list_clients()
        self._client_lookup = {name: cid for cid, name, *_ in clients}
        self.client_combo["values"] = [name for _, name, *_ in clients]
        if select_id is not None:
            for cid, name, *_ in clients:
                if cid == select_id:
                    self.client_var.set(name)
                    break

    def _on_client_search(self, _event=None):
        q = self.client_search_var.get().strip()
        results = search_clients(q) if q else list_clients()
        self._client_lookup = {name: cid for cid, name, *_ in results}
        self.client_combo["values"] = [name for _, name, *_ in results]
        if results and self.client_var.get() not in self._client_lookup:
            self.client_var.set(results[0][1])

    def _on_client_selected(self, _event=None):
        name = self.client_var.get()
        cid = self._client_lookup.get(name)
        if cid is not None:
            self._load_client(cid)

    def _refresh_fy_combo(self):
        """Populates the FY dropdown with every year already saved for this
        client, plus the currently-typed year if it isn't saved yet."""
        fys = list_client_fys(self.current_client_id) if self.current_client_id else []
        current = self.fy_var.get()
        values = list(fys)
        if current and current not in values:
            values.insert(0, current)
        self.fy_combo["values"] = values

    def _on_fy_selected(self, _event=None):
        self._switch_fy(self.fy_var.get())

    def _on_fy_go(self):
        """Handles typing a brand-new FY-end date not yet in the dropdown
        (e.g. moving a client on to the next financial year)."""
        fy = self.fy_var.get().strip()
        if not fy or not parse_date(fy):
            messagebox.showwarning("Invalid date", "Enter the FY-end date as DD/MM/YYYY, e.g. 31/03/2027.")
            return
        self._switch_fy(fy)

    def _switch_fy(self, fy_end):
        """Reset the working screen for a (possibly new) financial year under
        the current client \u2014 never touches other years' saved data."""
        if self.current_client_id is None:
            return
        self.fy_end_var.set(fy_end)
        self.raw_rows = load_client_bills(self.current_client_id, fy_end)
        self._refresh_fy_combo()
        if self.raw_rows:
            self.file_label.config(text=f"Loaded saved data for FY ending {fy_end} ({len(self.raw_rows)} bills)")
            self.compute()
        else:
            self.file_label.config(text=f"No data yet for FY ending {fy_end} \u2014 use an import button below")
            self._clear_working_screen()
        self._refresh_dashboard()
        self._refresh_history()

    def _clear_working_screen(self):
        """Wipes the on-screen working area \u2014 used whenever we switch to a
        client/FY that has no saved data yet, so nothing from the previous
        selection lingers on screen."""
        self.tree.delete(*self.tree.get_children())
        self.computed_rows = []
        for key in ("msme", "purchase", "disallowed", "compliant"):
            self._set_stat(key, "\u2014")
        self.clause_var.set("")

    def _load_client(self, client_id):
        row = get_client(client_id)
        if not row:
            return
        _id, name, trade_name, pan, gstin = row
        self.current_client_id = client_id
        self.current_client_name = name
        self.saved_categories = load_client_categories(client_id)

        fys = list_client_fys(client_id)
        fy_end = fys[0] if fys else "31/03/2026"
        self.fy_var.set(fy_end)
        self.fy_end_var.set(fy_end)
        self._refresh_fy_combo()

        self.raw_rows = load_client_bills(client_id, fy_end)
        if self.raw_rows:
            self.file_label.config(text=f"Loaded saved data for {name}, FY ending {fy_end} ({len(self.raw_rows)} bills)")
            self.compute()
        else:
            self.file_label.config(text=f"No data yet for {name} \u2014 use an import button below")
            self._clear_working_screen()
        self.client_status_label.config(
            text=f"Working on: {name}" + (f"  ({trade_name})" if trade_name else "")
                 + (f"  \u00b7 PAN {pan}" if pan else "") + (f"  \u00b7 GSTIN {gstin}" if gstin else ""))
        self._refresh_dashboard()
        self._refresh_history()

    def new_client(self):
        self._open_client_form(mode="new")

    def edit_client_details(self):
        if self.current_client_id is None:
            return
        self._open_client_form(mode="edit")

    def _open_client_form(self, mode):
        existing = None
        if mode == "edit":
            existing = get_client(self.current_client_id)

        win = tk.Toplevel(self)
        win.title("New Client" if mode == "new" else "Edit Client Details")
        win.configure(bg=CARD_BG)
        win.geometry("420x320")

        title_bar = tk.Frame(win, bg=PRIMARY, height=50)
        title_bar.pack(fill="x")
        tk.Label(title_bar, text=("\U0001F4C1  New Client" if mode == "new" else "\u270E  Edit Client Details"),
                 bg=PRIMARY, fg="#FFFFFF", font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=16, pady=12)

        form = tk.Frame(win, bg=CARD_BG)
        form.pack(fill="both", expand=True, padx=20, pady=16)

        fields = [
            ("Client Name*", "name", existing[1] if existing else ""),
            ("Trade Name", "trade_name", existing[2] if existing else ""),
            ("PAN", "pan", existing[3] if existing else ""),
            ("GSTIN", "gstin", existing[4] if existing else ""),
        ]
        vars_ = {}
        for i, (label, key, val) in enumerate(fields):
            tk.Label(form, text=label, bg=CARD_BG, fg=MUTED, font=("Segoe UI", 9, "bold")).grid(
                row=i, column=0, sticky="w", pady=8)
            var = tk.StringVar(value=val or "")
            tk.Entry(form, textvariable=var, width=28, font=("Segoe UI", 10),
                      relief="solid", bd=1).grid(row=i, column=1, pady=8, padx=(10, 0))
            vars_[key] = var

        def save():
            name = vars_["name"].get().strip()
            if not name:
                messagebox.showwarning("Client name required", "Please enter a client name.")
                return
            trade_name, pan, gstin = (vars_["trade_name"].get().strip(),
                                       vars_["pan"].get().strip().upper(),
                                       vars_["gstin"].get().strip().upper())
            if mode == "new":
                existing_names = {n for _, n, *_ in list_clients()}
                if name in existing_names:
                    messagebox.showwarning("Already exists", f"A client named '{name}' already exists.")
                    return
                cid = create_client(name, trade_name, pan, gstin)
                win.destroy()
                self._refresh_client_combo(select_id=cid)
                self._load_client(cid)
            else:
                update_client_details(self.current_client_id, name, trade_name, pan, gstin)
                win.destroy()
                self._refresh_client_combo(select_id=self.current_client_id)
                self._load_client(self.current_client_id)

        btn_frame = tk.Frame(win, bg=CARD_BG)
        btn_frame.pack(fill="x", padx=20, pady=(0, 16), side="bottom")
        HoverButton(btn_frame, text="\u2705  Save", bg=TEAL, command=save).pack(side="right")

    def delete_current_client(self):
        if self.current_client_id is None:
            return
        name = self.current_client_name
        if not messagebox.askyesno("Delete client",
                                    f"Delete ALL saved data for '{name}' \u2014 every financial year, "
                                    f"every report \u2014 permanently? This cannot be undone."):
            return
        delete_client_db(self.current_client_id)
        remaining = list_clients()
        if remaining:
            self._refresh_client_combo(select_id=remaining[0][0])
            self._load_client(remaining[0][0])
        else:
            cid = create_client("Demo Client")
            self.current_client_id = cid
            self.current_client_name = "Demo Client"
            self._refresh_client_combo(select_id=cid)
            self.client_status_label.config(text="Working on: Demo Client")
            self.load_sample()

    def _persist_current_client(self):
        """Saves the current in-memory raw_rows + categories to disk for the
        active client and financial year, so they can be retrieved next time
        without re-import. Never touches other clients or other FYs."""
        if self.current_client_id is None:
            return
        fy_end = self.fy_end_var.get()
        save_client_bills(self.current_client_id, fy_end, self.raw_rows)
        save_client_categories(self.current_client_id, self.saved_categories)
        self._refresh_fy_combo()
        self._refresh_dashboard()
        self._refresh_history()

    # ------------------------------------------------------- dashboard tab --

    def _build_dashboard_tab(self, parent):
        wrap = tk.Frame(parent, bg=BG_PAGE)
        wrap.pack(fill="both", expand=True, padx=4, pady=12)

        info_card = tk.Frame(wrap, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
        info_card.pack(fill="x", pady=(0, 14))
        self.dash_info_label = tk.Label(info_card, text="", bg=CARD_BG, fg=INK, justify="left",
                                         font=("Segoe UI", 10), anchor="w")
        self.dash_info_label.pack(fill="x", padx=18, pady=14)

        tk.Label(wrap, text="Current financial year at a glance", bg=BG_PAGE, fg=MUTED,
                  font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 6))
        self.dash_stats_frame = tk.Frame(wrap, bg=BG_PAGE)
        self.dash_stats_frame.pack(fill="x", pady=(0, 16))

        tk.Label(wrap, text="Recently generated reports (all years)", bg=BG_PAGE, fg=MUTED,
                  font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 6))
        recent_card = tk.Frame(wrap, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
        recent_card.pack(fill="both", expand=True)
        cols = ("date", "report", "fy", "file")
        self.dash_recent_tree = ttk.Treeview(recent_card, columns=cols, show="headings", height=8)
        for c, h, w in zip(cols, ["Generated", "Report", "FY End", "File"], [140, 150, 90, 320]):
            self.dash_recent_tree.heading(c, text=h)
            self.dash_recent_tree.column(c, width=w, anchor="w")
        self.dash_recent_tree.pack(fill="both", expand=True, padx=10, pady=10)

    def _refresh_dashboard(self):
        if self.current_client_id is None:
            return
        row = get_client(self.current_client_id)
        if not row:
            return
        _id, name, trade_name, pan, gstin = row
        lines = [f"Client Name:  {name}"]
        if trade_name:
            lines.append(f"Trade Name:  {trade_name}")
        if pan:
            lines.append(f"PAN:  {pan}")
        if gstin:
            lines.append(f"GSTIN:  {gstin}")
        lines.append(f"Financial Year (current):  {self.fy_end_var.get()}")
        self.dash_info_label.config(text="\n".join(lines))

        for w in self.dash_stats_frame.winfo_children():
            w.destroy()
        if self.computed_rows:
            total_purchase, total_disallowed, compliant_val, msme_count = self._summary_totals()
            stats = [("MSME Vendors", str(msme_count), PRIMARY), ("Total Purchases", fmt_inr(total_purchase), PRIMARY_DARK),
                     ("Disallowed", fmt_inr(total_disallowed), CORAL), ("Compliant", fmt_inr(compliant_val), GREEN)]
        else:
            stats = [("MSME Vendors", "\u2014", MUTED), ("Total Purchases", "\u2014", MUTED),
                     ("Disallowed", "\u2014", MUTED), ("Compliant", "\u2014", MUTED)]
        for i, (label, val, color) in enumerate(stats):
            cell = tk.Frame(self.dash_stats_frame, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
            cell.grid(row=0, column=i, sticky="nsew", padx=(0 if i == 0 else 8, 0))
            self.dash_stats_frame.grid_columnconfigure(i, weight=1)
            tk.Label(cell, text=val, bg=CARD_BG, fg=color, font=("Consolas", 15, "bold")).pack(
                anchor="w", padx=12, pady=(10, 0))
            tk.Label(cell, text=label.upper(), bg=CARD_BG, fg=MUTED, font=("Segoe UI", 8)).pack(
                anchor="w", padx=12, pady=(0, 10))

        self.dash_recent_tree.delete(*self.dash_recent_tree.get_children())
        for generated_at, report_type, fy_end, file_name, _path in load_report_history(self.current_client_id)[:15]:
            ts = generated_at.split("T")[0] if "T" in generated_at else generated_at
            self.dash_recent_tree.insert("", "end", values=(ts, report_type, fy_end, file_name))

    # -------------------------------------------------- report history tab --

    # ------------------------------------------------------- clause 26 tab --

    def _build_clause26_tab(self, parent):
        wrap = tk.Frame(parent, bg=BG_PAGE)
        wrap.pack(fill="both", expand=True, padx=4, pady=12)

        header = tk.Frame(wrap, bg=BG_PAGE)
        header.pack(fill="x", pady=(0, 10))
        tk.Label(header, text="Form 3CD \u2014 Clause 26 Schedule", bg=BG_PAGE, fg=INK,
                  font=("Segoe UI", 12, "bold")).pack(side="left")
        HoverButton(header, text="\U0001F4D7 Export Clause 26 (Excel)", bg=GREEN, padx=11,
                    command=self.export_clause26_excel).pack(side="right", padx=(6, 0))
        HoverButton(header, text="\U0001F4C4 Export Clause 26 (PDF)", bg=shade(CORAL, 0.85), padx=11,
                    command=self.export_clause26_pdf).pack(side="right")

        tk.Label(wrap, text="Every Sundry Creditor in the imported ledger, whether or not Section 43B(h) "
                             "applies to them \u2014 nothing is filtered out of this list.",
                  bg=BG_PAGE, fg=MUTED, font=("Segoe UI", 9, "italic"), wraplength=1100, justify="left").pack(
            anchor="w", pady=(0, 10))

        table_card = tk.Frame(wrap, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
        table_card.pack(fill="both", expand=True)
        cols = ("vendor", "category", "bills", "amount", "paid", "disallowed", "status")
        headers_ = ["Vendor", "Category", "No. of Bills", "Total Purchases", "Total Paid",
                    "Disallowed u/s 43B(h)", "Status"]
        widths = [220, 90, 90, 130, 130, 150, 110]
        self.clause26_tree = ttk.Treeview(table_card, columns=cols, show="headings", height=16)
        for c, h, w in zip(cols, headers_, widths):
            self.clause26_tree.heading(c, text=h)
            self.clause26_tree.column(c, width=w, anchor="w")
        for c in ("bills", "amount", "paid", "disallowed"):
            self.clause26_tree.column(c, anchor="e")
        vsb = ttk.Scrollbar(table_card, orient="vertical", command=self.clause26_tree.yview)
        self.clause26_tree.configure(yscrollcommand=vsb.set)
        self.clause26_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        vsb.pack(side="right", fill="y", pady=10)
        self.clause26_tree.tag_configure("disallow", foreground=CORAL, background=ROW_DISALLOW_BG)
        self.clause26_tree.tag_configure("ok", foreground=TEAL, background=ROW_OK_BG)
        self.clause26_tree.tag_configure("na", foreground=MUTED, background=ROW_NA_BG)

    def _clause26_vendor_summary(self):
        """One row per creditor (not per bill) \u2014 the format a Clause 26
        schedule is normally prepared in. Includes every creditor from the
        imported ledger regardless of MSME status; non-MSME/uncategorised
        vendors show as N/A rather than being dropped."""
        by_vendor = defaultdict(list)
        for r in self.computed_rows:
            by_vendor[r["vendor"]].append(r)

        summary = []
        for vendor in sorted(by_vendor):
            rows = by_vendor[vendor]
            category = rows[0]["category"] or "Not set"
            total_amount = sum(r["amount"] for r in rows)
            total_paid = sum(r["amount_paid"] for r in rows)
            total_disallowed = sum(r["disallowed"] for r in rows)
            is_msme = rows[0]["is_msme"]
            if not is_msme:
                status = "na"
            elif total_disallowed > 0:
                status = "disallow"
            else:
                status = "ok"
            summary.append({
                "vendor": vendor, "category": category, "num_bills": len(rows),
                "amount": total_amount, "paid": total_paid, "disallowed": total_disallowed,
                "status": status,
            })
        return summary

    def _refresh_clause26_tab(self):
        self.clause26_tree.delete(*self.clause26_tree.get_children())
        if not self.computed_rows:
            return
        for v in self._clause26_vendor_summary():
            status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[v["status"]]
            self.clause26_tree.insert("", "end", values=(
                v["vendor"], v["category"], v["num_bills"], fmt_inr(v["amount"]), fmt_inr(v["paid"]),
                fmt_inr(v["disallowed"]) if v["disallowed"] else "\u2014", status_text
            ), tags=(v["status"],))

    def export_clause26_excel(self):
        if not self.computed_rows:
            messagebox.showinfo("Nothing to export", "Compute compliance first.")
            return
        if not OPENPYXL_OK:
            messagebox.showerror("Missing dependency",
                                  "This needs the 'openpyxl' package.\n\npip install openpyxl")
            return
        default_name = build_export_filename(self.current_client_name or "Client", "Clause 26",
                                              self.fy_end_var.get(), "xlsx")
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name,
                                             filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        try:
            wb = openpyxl.Workbook()

            ws1 = wb.active
            ws1.title = "Clause 26 Summary"
            self._write_clause26_summary_sheet(ws1)

            ws2 = wb.create_sheet("Bill-wise Detail")
            self._write_clause26_detail_sheet(ws2)

            wb.save(path)
            if self.current_client_id is not None:
                log_report_export(self.current_client_id, self.fy_end_var.get(), "Clause 26 (Excel)",
                                   os.path.basename(path), path)
                self._refresh_dashboard()
                self._refresh_history()
            messagebox.showinfo("Exported", f"Clause 26 report saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error saving file", str(e))

    def _write_clause26_summary_sheet(self, ws):
        navy = "1E293B"
        red_fill = PatternFill("solid", fgColor="FEF2F2")
        green_fill = PatternFill("solid", fgColor="ECFDF5")
        header_fill = PatternFill("solid", fgColor=navy)
        title_font = Font(name="Calibri", size=16, bold=True, color=navy)
        sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin = Side(style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["Vendor", "Category", "No. of Bills", "Total Purchases", "Total Paid",
                   "Disallowed u/s 43B(h)", "Status"]
        ncols = len(headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1,
                value=f"Form 3CD \u2014 Clause 26 Schedule \u2014 {self.current_client_name or ''}").font = title_font
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        ws.cell(row=2, column=1,
                value=f"Financial Year ending {self.fy_end_var.get()}  \u00b7  "
                      f"Report Date: {datetime.now().strftime('%d/%m/%Y')}  \u00b7  "
                      f"Every Sundry Creditor in the imported ledger is listed below").font = sub_font
        header_row = 4
        detail_line = self._client_detail_line()
        if detail_line:
            ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
            ws.cell(row=3, column=1, value=detail_line).font = sub_font
            header_row = 5

        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=j, value=h)
            cell.font, cell.fill, cell.border = header_font, header_fill, border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[header_row].height = 26

        r_idx = header_row + 1
        summary = self._clause26_vendor_summary()
        for v in summary:
            status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[v["status"]]
            values = [v["vendor"], v["category"], v["num_bills"], v["amount"], v["paid"],
                      v["disallowed"], status_text]
            for j, val in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=j, value=val)
                cell.border = border
                if j in (4, 5, 6):
                    cell.number_format = "#,##0"
                if v["status"] == "disallow":
                    cell.fill = red_fill
                elif v["status"] == "ok":
                    cell.fill = green_fill
            r_idx += 1

        r_idx += 1
        total_purchase, total_disallowed, compliant_val, msme_count = self._summary_totals()
        for label, value in [("Total Creditors Listed", len(summary)), ("MSME Vendors", msme_count),
                              ("Total MSME Purchases", total_purchase),
                              ("Disallowed u/s 43B(h)", total_disallowed),
                              ("Compliant Value", compliant_val)]:
            ws.cell(row=r_idx, column=1, value=label).font = bold_font
            vcell = ws.cell(row=r_idx, column=2, value=value)
            vcell.font = bold_font
            if isinstance(value, (int, float)) and "Vendors" not in label and "Listed" not in label:
                vcell.number_format = "#,##0"
            r_idx += 1

        r_idx += 1
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx + 3, end_column=ncols)
        note_cell = ws.cell(row=r_idx, column=1, value=self.clause_var.get())
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
        note_cell.font = Font(name="Calibri", size=10, italic=True)

        widths = [28, 12, 12, 16, 14, 18, 12]
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = f"{header_row}:{header_row}"

    def _write_clause26_detail_sheet(self, ws):
        navy = "1E293B"
        red_fill = PatternFill("solid", fgColor="FEF2F2")
        green_fill = PatternFill("solid", fgColor="ECFDF5")
        header_fill = PatternFill("solid", fgColor=navy)
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        thin = Side(style="thin", color="E2E8F0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["Vendor", "Category", "Invoice Date", "Invoice Amount", "Amount Paid",
                   "Due Date", "Payment Date", "Status", "Disallowed", "Note"]
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.font, cell.fill, cell.border = header_font, header_fill, border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 26

        r_idx = 2
        for r in self.computed_rows:
            status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[r["status"]]
            values = [r["vendor"], r["category"] or "\u2014", fmt_date(r["inv_date"]), r["amount"],
                      r["amount_paid"], fmt_date(r["due_date"]) if r["due_date"] else "",
                      fmt_date(r["pay_date"]) if r["pay_date"] else "", status_text,
                      r["disallowed"], r["note"]]
            for j, v in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=j, value=v)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(j == 10))
                if j in (4, 5, 9):
                    cell.number_format = "#,##0"
                if r["status"] == "disallow":
                    cell.fill = red_fill
                elif r["status"] == "ok":
                    cell.fill = green_fill
            r_idx += 1

        widths = [26, 10, 13, 14, 13, 13, 13, 12, 14, 46]
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = ws.cell(row=2, column=1)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = "1:1"

    def export_clause26_pdf(self):
        if not self.computed_rows:
            messagebox.showinfo("Nothing to export", "Compute compliance first.")
            return
        if not REPORTLAB_OK:
            messagebox.showerror("Missing dependency",
                                  "This needs the 'reportlab' package.\n\npip install reportlab")
            return
        default_name = build_export_filename(self.current_client_name or "Client", "Clause 26",
                                              self.fy_end_var.get(), "pdf")
        path = filedialog.asksaveasfilename(defaultextension=".pdf", initialfile=default_name,
                                             filetypes=[("PDF files", "*.pdf")])
        if not path:
            return
        try:
            doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                     leftMargin=14 * mm, rightMargin=14 * mm,
                                     topMargin=12 * mm, bottomMargin=12 * mm)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("TitleX", parent=styles["Title"], fontName=PDF_FONT_BOLD, fontSize=16,
                                          textColor=colors.HexColor("#1E293B"), spaceAfter=2)
            sub_style = ParagraphStyle("SubX", parent=styles["Normal"], fontName=PDF_FONT, fontSize=9,
                                        textColor=colors.HexColor("#64748B"), spaceAfter=10)
            note_style = ParagraphStyle("NoteX", parent=styles["Normal"], fontName=PDF_FONT, fontSize=8, leading=10)
            cell_style = ParagraphStyle("CellX", parent=styles["Normal"], fontName=PDF_FONT, fontSize=7.5, leading=9)

            elements = [
                Paragraph(f"Form 3CD \u2014 Clause 26 Schedule \u2014 {self.current_client_name or ''}", title_style),
                Paragraph(f"Financial Year ending {self.fy_end_var.get()} &nbsp;&middot;&nbsp; "
                          f"Report Date: {datetime.now().strftime('%d/%m/%Y')} &middot; "
                          f"Every Sundry Creditor in the imported ledger is listed below", sub_style),
            ]
            detail_line = self._client_detail_line()
            if detail_line:
                elements.append(Paragraph(detail_line.replace("\u00b7", "&middot;"), sub_style))

            total_purchase, total_disallowed, compliant_val, msme_count = self._summary_totals()
            summary = self._clause26_vendor_summary()
            summary_data = [
                ["Creditors Listed", "MSME Vendors", "Total Purchases", "Disallowed u/s 43B(h)", "Compliant"],
                [str(len(summary)), str(msme_count), fmt_inr_pdf(total_purchase),
                 fmt_inr_pdf(total_disallowed), fmt_inr_pdf(compliant_val)],
            ]
            summary_table = Table(summary_data, colWidths=[52 * mm] * 5)
            summary_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, 1), PDF_FONT_BOLD),
                ("FONTNAME", (1, 1), (-1, 1), PDF_FONT),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("TEXTCOLOR", (3, 1), (3, 1), colors.HexColor("#EF4444")),
                ("TEXTCOLOR", (4, 1), (4, 1), colors.HexColor("#10B981")),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 10 * mm))

            headers = ["Vendor", "Category", "Bills", "Total Purchases", "Total Paid",
                       "Disallowed u/s 43B(h)", "Status"]
            data = [headers]
            row_status = []
            for v in summary:
                status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[v["status"]]
                data.append([
                    Paragraph(v["vendor"], cell_style), v["category"], str(v["num_bills"]),
                    fmt_inr_pdf(v["amount"]), fmt_inr_pdf(v["paid"]),
                    fmt_inr_pdf(v["disallowed"]) if v["disallowed"] else "\u2014", status_text,
                ])
                row_status.append(v["status"])

            col_widths = [60 * mm, 24 * mm, 18 * mm, 34 * mm, 34 * mm, 40 * mm, 30 * mm]
            table = Table(data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("FONTNAME", (0, 1), (-1, -1), PDF_FONT),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
            for i, status in enumerate(row_status, start=1):
                if status == "disallow":
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FEF2F2")))
                elif status == "ok":
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#ECFDF5")))
            table.setStyle(TableStyle(style_cmds))
            elements.append(table)

            elements.append(Spacer(1, 6 * mm))
            elements.append(Paragraph(self.clause_var.get(), note_style))

            doc.build(elements)
            if self.current_client_id is not None:
                log_report_export(self.current_client_id, self.fy_end_var.get(), "Clause 26 (PDF)",
                                   os.path.basename(path), path)
                self._refresh_dashboard()
                self._refresh_history()
            messagebox.showinfo("Exported", f"Clause 26 report saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error saving file", str(e))

    def _build_history_tab(self, parent):
        wrap = tk.Frame(parent, bg=BG_PAGE)
        wrap.pack(fill="both", expand=True, padx=4, pady=12)
        tk.Label(wrap, text="Every export generated for this client, across all financial years",
                  bg=BG_PAGE, fg=MUTED, font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 8))
        hist_card = tk.Frame(wrap, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
        hist_card.pack(fill="both", expand=True)
        cols = ("date", "report", "fy", "file", "path")
        self.history_tree = ttk.Treeview(hist_card, columns=cols, show="headings", height=20)
        for c, h, w in zip(cols, ["Generated", "Report", "FY End", "File Name", "Location"],
                            [150, 160, 90, 260, 320]):
            self.history_tree.heading(c, text=h)
            self.history_tree.column(c, width=w, anchor="w")
        vsb = ttk.Scrollbar(hist_card, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=vsb.set)
        self.history_tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        vsb.pack(side="right", fill="y", pady=10)

    def _refresh_history(self):
        if self.current_client_id is None:
            return
        self.history_tree.delete(*self.history_tree.get_children())
        for generated_at, report_type, fy_end, file_name, file_path in load_report_history(self.current_client_id):
            ts = generated_at.replace("T", "  ")[:19] if generated_at else ""
            self.history_tree.insert("", "end", values=(ts, report_type, fy_end, file_name, file_path))

    def _on_tab_changed(self, _event=None):
        try:
            tab_text = self.notebook.tab(self.notebook.select(), "text")
        except tk.TclError:
            return
        if "Dashboard" in tab_text:
            self._refresh_dashboard()
        elif "Clause 26" in tab_text:
            self._refresh_clause26_tab()
        elif "History" in tab_text:
            self._refresh_history()

    def load_sample(self):
        rows = list(csv.DictReader(io.StringIO(SAMPLE_CSV)))
        self.raw_rows = [{k.strip().lower(): v for k, v in r.items()} for r in rows]
        self.file_label.config(text="Using built-in sample data (10 vendors)")
        self.compute()
        self._persist_current_client()

    def load_csv_file(self):
        path = filedialog.askopenfilename(title="Select vendor ledger CSV",
                                           filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if not path:
            return
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.DictReader(f))
            self.raw_rows = [{k.strip().lower(): v for k, v in r.items()} for r in rows]
            self.file_label.config(text=f"Loaded: {os.path.basename(path)} ({len(self.raw_rows)} rows)")
            self.compute()
            self._persist_current_client()
        except Exception as e:
            messagebox.showerror("Error reading file", str(e))

    def import_tally(self):
        if not OPENPYXL_OK:
            messagebox.showerror(
                "Missing dependency",
                "Reading Excel files needs the 'openpyxl' package.\n\n"
                "Close this app and run:\n    pip install openpyxl\n\nthen restart."
            )
            return

        bills_path = filedialog.askopenfilename(
            title="Step 1 of 2 \u2014 Select Tally BILLS export (Purchase Register / Payables)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not bills_path:
            return
        payments_path = filedialog.askopenfilename(
            title="Step 2 of 2 \u2014 Select Tally PAYMENTS export (Bank Book / Cash Book)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not payments_path:
            return

        try:
            bills = load_bills_excel(bills_path)
            payments = load_payments_excel(payments_path)
        except Exception as e:
            messagebox.showerror("Error reading Excel file(s)", str(e))
            return

        if not bills:
            messagebox.showwarning(
                "No bills detected",
                "Couldn't find recognisable Vendor/Amount columns in the bills file.\n"
                "Check the export includes bill-wise details (F12 > Bill-wise Details) and try again.")
            return

        self.raw_rows = merge_bills_and_payments(bills, payments, self.saved_categories)
        self.file_label.config(
            text=f"Imported: {os.path.basename(bills_path)} + {os.path.basename(payments_path)} "
                 f"({len(bills)} bills, {len(payments)} payments matched)")

        # prompt for any vendor categories not yet on file
        unknown_vendors = sorted({r["vendor name"] for r in self.raw_rows
                                   if not self.saved_categories.get(norm_vendor(r["vendor name"]))})
        if unknown_vendors:
            self.manage_categories(prefill_vendors=unknown_vendors)
        else:
            self.compute()
        self._persist_current_client()

    def _import_multi_voucher_file(self, path, source_label):
        try:
            bills, payments, ignored_count = load_multi_voucher_excel(path)
        except Exception as e:
            messagebox.showerror(f"Error reading {source_label}", str(e))
            return

        if not bills:
            messagebox.showwarning(
                "No purchase vouchers detected",
                "Couldn't find any rows classified as Purchase vouchers.\n"
                "Check the export includes a Voucher Type column with 'Purchase' entries, "
                "and Bill-wise Details is switched on (F12) so bill references carry through.")
            return

        self.raw_rows = merge_bills_and_payments(bills, payments, self.saved_categories)
        self.file_label.config(
            text=f"Imported from {source_label}: {os.path.basename(path)} "
                 f"({len(bills)} purchase vouchers, {len(payments)} settlement vouchers, "
                 f"{ignored_count} other vouchers skipped)")

        unknown_vendors = sorted({r["vendor name"] for r in self.raw_rows
                                   if not self.saved_categories.get(norm_vendor(r["vendor name"]))})
        if unknown_vendors:
            self.manage_categories(prefill_vendors=unknown_vendors)
        else:
            self.compute()
        self._persist_current_client()

    def import_creditors_ledger(self):
        if not OPENPYXL_OK:
            messagebox.showerror(
                "Missing dependency",
                "Reading Excel files needs the 'openpyxl' package.\n\n"
                "Close this app and run:\n    pip install openpyxl\n\nthen restart."
            )
            return
        path = filedialog.askopenfilename(
            title="Select the Creditors Ledger export (all parties, bulk) \u2014 Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        self._import_multi_voucher_file(path, "Creditors Ledger")

    def import_daybook(self):
        if not OPENPYXL_OK:
            messagebox.showerror(
                "Missing dependency",
                "Reading Excel files needs the 'openpyxl' package.\n\n"
                "Close this app and run:\n    pip install openpyxl\n\nthen restart."
            )
            return
        path = filedialog.askopenfilename(
            title="Select the full Tally Day Book / All Vouchers export (Excel)",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not path:
            return
        self._import_multi_voucher_file(path, "Day Book")

    def manage_categories(self, prefill_vendors=None):
        vendors = prefill_vendors or sorted({r["vendor name"] for r in self.raw_rows})
        if not vendors:
            messagebox.showinfo("No vendors", "Load some data first.")
            return

        win = tk.Toplevel(self)
        win.title("Assign MSME category per vendor")
        win.configure(bg=CARD_BG)
        win.geometry("540x560")

        title_bar = tk.Frame(win, bg=PRIMARY, height=54)
        title_bar.pack(fill="x")
        tk.Label(title_bar, text="\U0001F3F7\uFE0F  Assign MSME category per vendor", bg=PRIMARY, fg="#FFFFFF",
                 font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=16, pady=14)

        tk.Label(win, text="Tag each vendor's Udyam category once \u2014 remembered for next time.",
                 bg=CARD_BG, fg=MUTED, font=("Segoe UI", 9), wraplength=490, justify="left").pack(
            anchor="w", padx=16, pady=(14, 8))

        canvas = tk.Canvas(win, bg=CARD_BG, highlightthickness=0)
        scroll_frame = tk.Frame(canvas, bg=CARD_BG)
        vsb = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        canvas.pack(side="left", fill="both", expand=True, padx=(16, 0))
        vsb.pack(side="right", fill="y")
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        combo_vars = {}
        for i, v in enumerate(vendors):
            row_bg = ROW_NA_BG if i % 2 == 0 else CARD_BG
            row = tk.Frame(scroll_frame, bg=row_bg)
            row.grid(row=i, column=0, sticky="ew", pady=1)
            scroll_frame.grid_columnconfigure(0, weight=1)
            tk.Label(row, text=v, bg=row_bg, fg=INK, font=("Segoe UI", 9),
                      width=34, anchor="w").grid(row=0, column=0, sticky="w", padx=(6, 8), pady=6)
            var = tk.StringVar(value=self.saved_categories.get(norm_vendor(v), "Small"))
            combo = ttk.Combobox(row, textvariable=var, values=CATEGORIES,
                                  width=12, state="readonly")
            combo.grid(row=0, column=1, pady=6, padx=(0, 6))
            combo_vars[v] = var

        def save_and_close():
            for v, var in combo_vars.items():
                self.saved_categories[norm_vendor(v)] = var.get()
                for r in self.raw_rows:
                    if r["vendor name"] == v:
                        r["category"] = var.get()
            win.destroy()
            self.compute()
            self._persist_current_client()

        btn_frame = tk.Frame(win, bg=CARD_BG)
        btn_frame.pack(fill="x", padx=16, pady=12, side="bottom")
        HoverButton(btn_frame, text="\u2705  Save & Compute", bg=TEAL,
                    command=save_and_close).pack(side="right")

    # ----------------------------------------------------------- compute --

    def compute(self):
        if not self.raw_rows:
            messagebox.showinfo("No data", "Load a CSV file, sample data, or import from Tally first.")
            return

        fy_end = parse_date(self.fy_end_var.get())
        rows = []
        for raw in self.raw_rows:
            r = {k.strip().lower(): (v if v not in (None, "") else "") for k, v in raw.items()}
            category = str(r.get("category", "")).strip()
            is_msme = category.lower() in ("micro", "small")
            inv_date = parse_date(r.get("invoice date", ""))
            amount = clean_amount(r.get("invoice amount", 0))

            agreement_days_raw = str(r.get("agreement days", "")).strip()
            try:
                agreement_days = int(float(agreement_days_raw)) if agreement_days_raw else None
            except ValueError:
                agreement_days = None

            due_days = 15
            if agreement_days and agreement_days > 0:
                due_days = min(agreement_days, 45)

            due_date = (inv_date + timedelta(days=due_days)) if inv_date else None

            # Installment-aware payment reconstruction. Imported rows carry the
            # full list of matched installments; manually entered / CSV rows
            # only ever have one lump "payment date", so treat that as a
            # single installment for the full amount.
            installments = raw.get("payment_installments")
            if not installments:
                single_date = parse_date(r.get("payment date", ""))
                installments = [(single_date, amount)] if single_date else []

            # Only installments actually paid on or before the FY-end count
            # towards this year's compliance; anything paid after FY-end (or
            # not yet paid at all) leaves that slice outstanding for this year.
            paid_by_fyend = 0.0
            latest_qualifying_date = None
            any_late_installment = False
            for d, amt in installments:
                if d and (not fy_end or d <= fy_end):
                    paid_by_fyend += amt
                    if latest_qualifying_date is None or d > latest_qualifying_date:
                        latest_qualifying_date = d
                    if due_date and d > due_date:
                        any_late_installment = True

            outstanding = max(0.0, amount - paid_by_fyend)
            tolerance = 1.0

            status, disallowed, note = "na", 0.0, ""
            if not category:
                note = "MSME category not set \u2014 use 'Manage MSME Categories'"
            elif not is_msme:
                note = f"Not covered \u2014 {category}"
            elif not inv_date:
                note = "Invalid/missing invoice date"
            elif outstanding <= tolerance:
                # Fully covered by payments made within this FY. Per Sec 43B,
                # paying late but still within the same year as accrual
                # creates no timing mismatch \u2014 nothing is disallowed.
                status = "ok"
                if any_late_installment:
                    note = (f"Fully paid within FY (last on {fmt_date(latest_qualifying_date)}) \u2014 "
                             f"some installment(s) exceeded the {due_days}d limit, but since settled "
                             f"within the same FY, no disallowance arises.")
                else:
                    days_taken = (latest_qualifying_date - inv_date).days if latest_qualifying_date else None
                    if days_taken is not None and days_taken < 0:
                        note = f"Paid in advance of invoice date (limit {due_days}d)"
                    elif len(installments) > 1:
                        note = f"Fully paid in {len(installments)} installments, last on {fmt_date(latest_qualifying_date)} (limit {due_days}d)"
                    else:
                        note = f"Paid in {days_taken}d (limit {due_days}d)" if days_taken is not None else "Paid within limit"
            elif paid_by_fyend > tolerance:
                # Partially paid: only the shortfall is disallowed, not the
                # whole original bill.
                status = "disallow"
                disallowed = outstanding
                note = (f"Partly paid: {fmt_inr(paid_by_fyend)} of {fmt_inr(amount)}"
                        + (f" (last installment {fmt_date(latest_qualifying_date)})" if latest_qualifying_date else "")
                        + f" \u2014 balance {fmt_inr(outstanding)} unpaid as on {fmt_date(fy_end)}, "
                        f"disallowed to that extent. Allowable when the balance is actually paid.")
            else:
                status = "disallow"
                disallowed = amount
                note = (f"Unpaid as on {fmt_date(fy_end)} \u2014 allowable in year actually paid."
                        if fy_end else "Unpaid \u2014 allowable in year actually paid.")

            rows.append(dict(vendor=r.get("vendor name", "\u2014"), category=category,
                              is_msme=is_msme, inv_date=inv_date, pay_date=latest_qualifying_date,
                              amount=amount, amount_paid=paid_by_fyend,
                              due_days=due_days, due_date=due_date, status=status,
                              disallowed=disallowed, note=note))

        self.computed_rows = rows
        self._render_table(rows)
        self._render_summary(rows)
        self._refresh_clause26_tab()

    def _render_table(self, rows):
        self.tree.delete(*self.tree.get_children())
        for r in rows:
            status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[r["status"]]
            self.tree.insert("", "end", values=(
                r["vendor"], r["category"] or "\u2014", fmt_date(r["inv_date"]), fmt_inr(r["amount"]),
                fmt_inr(r["amount_paid"]) if r["amount_paid"] else "\u2014",
                f'{fmt_date(r["due_date"])} ({r["due_days"]}d)' if r["due_date"] else "\u2014",
                fmt_date(r["pay_date"]), status_text,
                fmt_inr(r["disallowed"]) if r["disallowed"] else "\u2014", r["note"]
            ), tags=(r["status"],))

    def _render_summary(self, rows):
        msme_rows = [r for r in rows if r["is_msme"]]
        total_purchase = sum(r["amount"] for r in msme_rows)
        total_disallowed = sum(r["disallowed"] for r in msme_rows)
        compliant_val = total_purchase - total_disallowed

        self._set_stat("msme", str(len(msme_rows)))
        self._set_stat("purchase", fmt_inr(total_purchase))
        self._set_stat("disallowed", fmt_inr(total_disallowed))
        self._set_stat("compliant", fmt_inr(compliant_val))

        flagged = [r for r in msme_rows if r["status"] == "disallow"]
        if total_disallowed > 0:
            self.clause_var.set(
                f"Form 3CD \u2014 Clause 26 draft note: Amounts of {fmt_inr(total_disallowed)}, payable to "
                f"{len(flagged)} micro/small enterprise(s) beyond the time limit specified under Section 15 of "
                f"the MSMED Act, 2006, are disallowable under Section 43B(h) of the Income-tax Act, 1961 for the "
                f"year, to be allowed as deduction in the year of actual payment.")
        else:
            self.clause_var.set(
                "Form 3CD \u2014 Clause 26 draft note: No disallowance arises under Section 43B(h) \u2014 all "
                "traced MSME payments were made within the time limit specified under Section 15 of the MSMED "
                "Act, 2006.")

    def _client_detail_line(self):
        """PAN/GSTIN/Trade Name line for report headers, e.g.
        'Trade Name: ABC Traders  \u00b7  PAN: ABCDE1234F  \u00b7  GSTIN: 27ABCDE1234F1Z5'.
        Omits any field the client record doesn't have."""
        if self.current_client_id is None:
            return ""
        row = get_client(self.current_client_id)
        if not row:
            return ""
        _id, _name, trade_name, pan, gstin = row
        parts = []
        if trade_name:
            parts.append(f"Trade Name: {trade_name}")
        if pan:
            parts.append(f"PAN: {pan}")
        if gstin:
            parts.append(f"GSTIN: {gstin}")
        return "  \u00b7  ".join(parts)

    def export_csv(self):
        if not self.computed_rows:
            messagebox.showinfo("Nothing to export", "Compute compliance first.")
            return
        default_name = build_export_filename(
            self.current_client_name or "Client", "MSME Compliance", self.fy_end_var.get(), "csv")
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                             initialfile=default_name,
                                             filetypes=[("CSV files", "*.csv")])
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["Vendor", "Category", "Invoice Date", "Amount", "Amount Paid", "Due Date",
                                  "Payment Date", "Status", "Disallowed Amount", "Note"])
                for r in self.computed_rows:
                    status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[r["status"]]
                    writer.writerow([
                        r["vendor"], r["category"], fmt_date(r["inv_date"]), r["amount"], r["amount_paid"],
                        fmt_date(r["due_date"]) if r["due_date"] else "", fmt_date(r["pay_date"]) if r["pay_date"] else "",
                        status_text, r["disallowed"], r["note"]
                    ])
            if self.current_client_id is not None:
                log_report_export(self.current_client_id, self.fy_end_var.get(), "MSME Compliance (CSV)",
                                   os.path.basename(path), path)
                self._refresh_dashboard()
                self._refresh_history()
            messagebox.showinfo("Exported", f"Schedule saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error saving file", str(e))

    def _summary_totals(self):
        msme_rows = [r for r in self.computed_rows if r["is_msme"]]
        total_purchase = sum(r["amount"] for r in msme_rows)
        total_disallowed = sum(r["disallowed"] for r in msme_rows)
        return total_purchase, total_disallowed, total_purchase - total_disallowed, len(msme_rows)

    def export_excel(self):
        if not self.computed_rows:
            messagebox.showinfo("Nothing to export", "Compute compliance first.")
            return
        if not OPENPYXL_OK:
            messagebox.showerror("Missing dependency",
                                  "This needs the 'openpyxl' package.\n\npip install openpyxl")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=build_export_filename(self.current_client_name or "Client", "MSME Compliance",
                                               self.fy_end_var.get(), "xlsx"),
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "43B(h) Schedule"

            navy = "1E293B"
            gold = "F59E0B"
            red_fill = PatternFill("solid", fgColor="FEF2F2")
            green_fill = PatternFill("solid", fgColor="ECFDF5")
            header_fill = PatternFill("solid", fgColor=navy)
            title_font = Font(name="Calibri", size=16, bold=True, color=navy)
            sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
            header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            bold_font = Font(name="Calibri", size=11, bold=True)
            thin = Side(style="thin", color="E2E8F0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            headers = ["Vendor", "Category", "Invoice Date", "Invoice Amount", "Amount Paid",
                       "Due Date", "Payment Date", "Status", "Disallowed u/s 43B(h)", "Note"]
            ncols = len(headers)

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
            c = ws.cell(row=1, column=1, value=f"MSME 43B(h) Compliance Schedule \u2014 {self.current_client_name or ''}")
            c.font = title_font
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
            c = ws.cell(row=2, column=1,
                        value=f"Financial Year ending {self.fy_end_var.get()}  \u00b7  "
                              f"Report Date: {datetime.now().strftime('%d/%m/%Y')}  \u00b7  "
                              f"Section 43B(h), Income-tax Act 1961  \u00b7  Section 15, MSMED Act 2006")
            c.font = sub_font
            detail_line = self._client_detail_line()
            header_row = 4
            if detail_line:
                ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
                c = ws.cell(row=3, column=1, value=detail_line)
                c.font = sub_font
                header_row = 5
            for j, h in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=j, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[header_row].height = 26

            r_idx = header_row + 1
            for r in self.computed_rows:
                status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[r["status"]]
                values = [
                    r["vendor"], r["category"] or "\u2014", fmt_date(r["inv_date"]), r["amount"],
                    r["amount_paid"], fmt_date(r["due_date"]) if r["due_date"] else "",
                    fmt_date(r["pay_date"]) if r["pay_date"] else "", status_text,
                    r["disallowed"], r["note"],
                ]
                for j, v in enumerate(values, start=1):
                    cell = ws.cell(row=r_idx, column=j, value=v)
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=(j == 10))
                    if j in (4, 5, 9):
                        cell.number_format = "#,##0"
                    if r["status"] == "disallow":
                        cell.fill = red_fill
                    elif r["status"] == "ok":
                        cell.fill = green_fill
                r_idx += 1

            total_purchase, total_disallowed, compliant_val, msme_count = self._summary_totals()
            r_idx += 1
            for label, value in [
                ("MSME Vendors", msme_count),
                ("Total MSME Purchases", total_purchase),
                ("Disallowed u/s 43B(h)", total_disallowed),
                ("Compliant Value", compliant_val),
            ]:
                ws.cell(row=r_idx, column=1, value=label).font = bold_font
                vcell = ws.cell(row=r_idx, column=2, value=value)
                vcell.font = bold_font
                if isinstance(value, (int, float)) and label != "MSME Vendors":
                    vcell.number_format = "#,##0"
                r_idx += 1

            r_idx += 1
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx + 3, end_column=ncols)
            note_cell = ws.cell(row=r_idx, column=1, value=self.clause_var.get())
            note_cell.alignment = Alignment(wrap_text=True, vertical="top")
            note_cell.font = Font(name="Calibri", size=10, italic=True)

            widths = [26, 10, 13, 14, 13, 13, 13, 12, 16, 46]
            for j, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(j)].width = w
            ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

            # Print/page setup so it also looks right if printed or exported
            # to PDF straight from Excel, not just when viewed on-screen.
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.print_title_rows = f"{header_row}:{header_row}"

            wb.save(path)
            if self.current_client_id is not None:
                log_report_export(self.current_client_id, self.fy_end_var.get(), "MSME Compliance (Excel)",
                                   os.path.basename(path), path)
                self._refresh_dashboard()
                self._refresh_history()
            messagebox.showinfo("Exported", f"Excel report saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error saving file", str(e))

    def export_pdf(self):
        if not self.computed_rows:
            messagebox.showinfo("Nothing to export", "Compute compliance first.")
            return
        if not REPORTLAB_OK:
            messagebox.showerror("Missing dependency",
                                  "This needs the 'reportlab' package.\n\npip install reportlab")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=build_export_filename(self.current_client_name or "Client", "MSME Compliance",
                                               self.fy_end_var.get(), "pdf"),
            filetypes=[("PDF files", "*.pdf")])
        if not path:
            return

        try:
            doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                                     leftMargin=14 * mm, rightMargin=14 * mm,
                                     topMargin=12 * mm, bottomMargin=12 * mm)
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle("TitleX", parent=styles["Title"], fontName=PDF_FONT_BOLD, fontSize=16,
                                          textColor=colors.HexColor("#1E293B"), spaceAfter=2)
            sub_style = ParagraphStyle("SubX", parent=styles["Normal"], fontName=PDF_FONT, fontSize=9,
                                        textColor=colors.HexColor("#64748B"), spaceAfter=10)
            note_style = ParagraphStyle("NoteX", parent=styles["Normal"], fontName=PDF_FONT, fontSize=8, leading=10)
            cell_style = ParagraphStyle("CellX", parent=styles["Normal"], fontName=PDF_FONT, fontSize=7.5, leading=9)

            elements = [
                Paragraph(f"MSME 43B(h) Compliance Schedule \u2014 {self.current_client_name or ''}", title_style),
                Paragraph(f"Financial Year ending {self.fy_end_var.get()} &nbsp;&middot;&nbsp; "
                          f"Report Date: {datetime.now().strftime('%d/%m/%Y')} &middot; "
                          f"Section 43B(h), Income-tax Act 1961 &middot; Section 15, MSMED Act 2006", sub_style),
            ]
            detail_line = self._client_detail_line()
            if detail_line:
                elements.append(Paragraph(detail_line.replace("\u00b7", "&middot;"), sub_style))

            total_purchase, total_disallowed, compliant_val, msme_count = self._summary_totals()
            summary_data = [
                ["MSME Vendors", "Total MSME Purchases", "Disallowed u/s 43B(h)", "Compliant Value"],
                [str(msme_count), fmt_inr_pdf(total_purchase), fmt_inr_pdf(total_disallowed), fmt_inr_pdf(compliant_val)],
            ]
            summary_table = Table(summary_data, colWidths=[65 * mm] * 4)
            summary_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("FONTNAME", (1, 1), (-1, 1), PDF_FONT),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, 1), PDF_FONT_BOLD),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor("#EF4444")),
                ("TEXTCOLOR", (3, 1), (3, 1), colors.HexColor("#10B981")),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 10 * mm))

            headers = ["Vendor", "Category", "Inv. Date", "Amount", "Paid", "Due Date",
                       "Payment Date", "Status", "Disallowed", "Note"]
            data = [headers]
            row_status = []
            for r in self.computed_rows:
                status_text = {"ok": "Within Limit", "disallow": "Disallowed", "na": "N/A"}[r["status"]]
                data.append([
                    Paragraph(r["vendor"], cell_style), r["category"] or "\u2014",
                    fmt_date(r["inv_date"]), fmt_inr_pdf(r["amount"]), fmt_inr_pdf(r["amount_paid"]),
                    fmt_date(r["due_date"]) if r["due_date"] else "\u2014",
                    fmt_date(r["pay_date"]) if r["pay_date"] else "\u2014",
                    status_text, fmt_inr_pdf(r["disallowed"]) if r["disallowed"] else "\u2014",
                    Paragraph(r["note"], cell_style),
                ])
                row_status.append(r["status"])

            col_widths = [40 * mm, 16 * mm, 18 * mm, 20 * mm, 20 * mm, 18 * mm, 20 * mm, 18 * mm, 20 * mm, 62 * mm]
            table = Table(data, colWidths=col_widths, repeatRows=1)
            style_cmds = [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), PDF_FONT_BOLD),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("FONTNAME", (0, 1), (-1, -1), PDF_FONT),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
            for i, status in enumerate(row_status, start=1):
                if status == "disallow":
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#FEF2F2")))
                elif status == "ok":
                    style_cmds.append(("BACKGROUND", (0, i), (-1, i), colors.HexColor("#ECFDF5")))
            table.setStyle(TableStyle(style_cmds))
            elements.append(table)

            elements.append(Spacer(1, 6 * mm))
            elements.append(Paragraph(self.clause_var.get(), note_style))

            doc.build(elements)
            if self.current_client_id is not None:
                log_report_export(self.current_client_id, self.fy_end_var.get(), "MSME Compliance (PDF)",
                                   os.path.basename(path), path)
                self._refresh_dashboard()
                self._refresh_history()
            messagebox.showinfo("Exported", f"PDF report saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Error saving file", str(e))


if __name__ == "__main__":
    app = MSMEApp()
    app.mainloop()
