"""
GST 2B → Munim Template Populator
Extracts matched/NCP invoices from GSTR-2B vs Books reconciliation
and populates the Munim b2b template sheet.

Built for: IJR & Co., Panchkula
Usage: Copy-paste into Claude Code or run directly with Python.
"""

import sys
import subprocess
import os
import shutil
import re
from datetime import datetime, date

# ─── Auto-install ───
def ensure_packages():
    for mod, pkg in {'pandas': 'pandas', 'openpyxl': 'openpyxl'}.items():
        try:
            __import__(mod)
        except ImportError:
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', '--quiet', pkg])

ensure_packages()

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, numbers
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


# ═══════════════════════════════════════════════════════════════
#  GST State Code Mapping
# ═══════════════════════════════════════════════════════════════

STATE_CODE_MAP = {
    'jammu and kashmir': '01-Jammu & Kashmir',
    'jammu & kashmir': '01-Jammu & Kashmir',
    'himachal pradesh': '02-Himachal Pradesh',
    'punjab': '03-Punjab',
    'chandigarh': '04-Chandigarh',
    'uttarakhand': '05-Uttarakhand',
    'haryana': '06-Haryana',
    'delhi': '07-Delhi',
    'rajasthan': '08-Rajasthan',
    'uttar pradesh': '09-Uttar Pradesh',
    'bihar': '10-Bihar',
    'sikkim': '11-Sikkim',
    'arunachal pradesh': '12-Arunachal Pradesh',
    'nagaland': '13-Nagaland',
    'manipur': '14-Manipur',
    'mizoram': '15-Mizoram',
    'tripura': '16-Tripura',
    'meghalaya': '17-Meghalaya',
    'assam': '18-Assam',
    'west bengal': '19-West Bengal',
    'jharkhand': '20-Jharkhand',
    'odisha': '21-Odisha',
    'chhattisgarh': '22-Chhattisgarh',
    'madhya pradesh': '23-Madhya Pradesh',
    'gujarat': '24-Gujarat',
    'dadra and nagar haveli and daman and diu': '26-Dadra & Nagar Haveli and Daman & Diu',
    'dadra & nagar haveli and daman & diu': '26-Dadra & Nagar Haveli and Daman & Diu',
    'daman and diu': '25-Daman & Diu',
    'daman & diu': '25-Daman & Diu',
    'maharashtra': '27-Maharashtra',
    'andhra pradesh': '37-Andhra Pradesh',
    'karnataka': '29-Karnataka',
    'goa': '30-Goa',
    'lakshadweep': '31-Lakshadweep',
    'kerala': '32-Kerala',
    'tamil nadu': '33-Tamil Nadu',
    'puducherry': '34-Puducherry',
    'andaman and nicobar islands': '35-Andaman & Nicobar Islands',
    'andaman & nicobar islands': '35-Andaman & Nicobar Islands',
    'telangana': '36-Telangana',
    'ladakh': '38-Ladakh',
    'other territory': '97-Other Territory',
}

# Also map from GSTIN first 2 digits
GSTIN_STATE_MAP = {
    '01': '01-Jammu & Kashmir', '02': '02-Himachal Pradesh',
    '03': '03-Punjab', '04': '04-Chandigarh',
    '05': '05-Uttarakhand', '06': '06-Haryana',
    '07': '07-Delhi', '08': '08-Rajasthan',
    '09': '09-Uttar Pradesh', '10': '10-Bihar',
    '11': '11-Sikkim', '12': '12-Arunachal Pradesh',
    '13': '13-Nagaland', '14': '14-Manipur',
    '15': '15-Mizoram', '16': '16-Tripura',
    '17': '17-Meghalaya', '18': '18-Assam',
    '19': '19-West Bengal', '20': '20-Jharkhand',
    '21': '21-Odisha', '22': '22-Chhattisgarh',
    '23': '23-Madhya Pradesh', '24': '24-Gujarat',
    '25': '25-Daman & Diu',
    '26': '26-Dadra & Nagar Haveli and Daman & Diu',
    '27': '27-Maharashtra', '28': '28-Andhra Pradesh',
    '29': '29-Karnataka', '30': '30-Goa',
    '31': '31-Lakshadweep', '32': '32-Kerala',
    '33': '33-Tamil Nadu', '34': '34-Puducherry',
    '35': '35-Andaman & Nicobar Islands',
    '36': '36-Telangana', '37': '37-Andhra Pradesh',
    '38': '38-Ladakh', '97': '97-Other Territory',
}




# ═══════════════════════════════════════════════════════════════
#  Utility Functions
# ═══════════════════════════════════════════════════════════════

def safe_float(val):
    """Convert to float, handle NaN/blanks/commas."""
    if pd.isna(val) or val == '' or val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def derive_rate(taxable, igst, cgst, sgst):
    """Simple rate: 18 if any tax exists, 0 if no tax."""
    total_tax = abs(igst) + abs(cgst) + abs(sgst)
    return 18 if total_tax > 0 else 0


def map_place_of_supply(pos_value, gstin=None):
    """Convert Place of Supply name to Munim code format."""
    if pd.notna(pos_value) and str(pos_value).strip():
        key = str(pos_value).strip().lower()
        if key in STATE_CODE_MAP:
            return STATE_CODE_MAP[key]
        # Check if already in code format like "01-Jammu & Kashmir"
        if re.match(r'^\d{2}-', str(pos_value).strip()):
            return str(pos_value).strip()

    # Fallback: derive from GSTIN first 2 digits
    if gstin and len(str(gstin).strip()) >= 2:
        code = str(gstin).strip()[:2]
        if code in GSTIN_STATE_MAP:
            return GSTIN_STATE_MAP[code]

    return str(pos_value) if pd.notna(pos_value) else ''


def parse_excel_date(dt_val):
    """Parse date from the recon file and return as datetime object."""
    if pd.isna(dt_val) or dt_val == '' or dt_val is None:
        return None
    if isinstance(dt_val, (datetime, date)):
        if isinstance(dt_val, date) and not isinstance(dt_val, datetime):
            return datetime(dt_val.year, dt_val.month, dt_val.day)
        return dt_val
    if isinstance(dt_val, pd.Timestamp):
        return dt_val.to_pydatetime()
    s = str(dt_val).strip()
    # Handle Excel serial date numbers
    if s.replace('.', '').isdigit():
        try:
            serial = int(float(s))
            # Excel serial date: 1 = Jan 1 1900
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return base + timedelta(days=serial)
        except (ValueError, OverflowError):
            pass
    # Try common string formats
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y',
                '%d-%b-%Y', '%d.%m.%Y', '%d/%m/%y', '%d-%m-%y']:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ═══════════════════════════════════════════════════════════════
#  Main Application
# ═══════════════════════════════════════════════════════════════

class ReconToMunimApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("GST 2B → Munim Template — IJR & Co.")
        self.root.geometry("750x520")
        self.root.resizable(False, False)

        style = ttk.Style()
        style.configure('Title.TLabel', font=('Segoe UI', 14, 'bold'))
        style.configure('Sub.TLabel', font=('Segoe UI', 9, 'italic'), foreground='#555')
        style.configure('Run.TButton', font=('Segoe UI', 11, 'bold'))
        style.configure('Status.TLabel', font=('Segoe UI', 9))

        self.recon_path = tk.StringVar()
        self.munim_path = tk.StringVar()
        self.selected_month = tk.StringVar(value='April')
        self.output_dir = tk.StringVar(value=os.path.join(os.path.expanduser('~'), 'Desktop'))
        self.status_var = tk.StringVar(value='Ready — Select files and month')

        self._build_ui()

    def _build_ui(self):
        main = ttk.Frame(self.root, padding=15)
        main.pack(fill='both', expand=True)

        ttk.Label(main, text="GST 2B → Munim Template Populator", style='Title.TLabel').pack(pady=(0, 2))
        ttk.Label(main, text="Extract matched 2B data → Invoices to b2b, Credit Notes to cdnr",
                  style='Sub.TLabel').pack(pady=(0, 12))

        # ── Recon File ──
        f1 = ttk.LabelFrame(main, text=" 2B vs Books ITC Reconciliation File ", padding=10)
        f1.pack(fill='x', pady=4)
        row1 = ttk.Frame(f1)
        row1.pack(fill='x')
        ttk.Entry(row1, textvariable=self.recon_path, width=62, state='readonly').pack(
            side='left', fill='x', expand=True)
        ttk.Button(row1, text="Browse", command=self._browse_recon).pack(side='left', padx=(8, 0))

        # ── Munim Template ──
        f2 = ttk.LabelFrame(main, text=" Munim Template File ", padding=10)
        f2.pack(fill='x', pady=4)
        row2 = ttk.Frame(f2)
        row2.pack(fill='x')
        ttk.Entry(row2, textvariable=self.munim_path, width=62, state='readonly').pack(
            side='left', fill='x', expand=True)
        ttk.Button(row2, text="Browse", command=self._browse_munim).pack(side='left', padx=(8, 0))

        # ── Month Selection ──
        f3 = ttk.LabelFrame(main, text=" Select Month for Extraction ", padding=10)
        f3.pack(fill='x', pady=4)
        month_row = ttk.Frame(f3)
        month_row.pack(fill='x')
        ttk.Label(month_row, text="Return Period Month:").pack(side='left')
        months = ['April', 'May', 'June', 'July', 'August', 'September',
                  'October', 'November', 'December', 'January', 'February', 'March']
        self.month_combo = ttk.Combobox(month_row, textvariable=self.selected_month,
                                         values=months, width=15, state='readonly')
        self.month_combo.pack(side='left', padx=(8, 0))

        ttk.Label(f3, text="Filters: Data From=2B | Remark='Matched in [month]' or 'Nil in 2B' | Invoice→b2b, CN→cdnr",
                  foreground='#666', font=('Segoe UI', 8)).pack(anchor='w', pady=(5, 0))

        # ── Output ──
        f4 = ttk.LabelFrame(main, text=" Output Folder ", padding=10)
        f4.pack(fill='x', pady=4)
        row4 = ttk.Frame(f4)
        row4.pack(fill='x')
        ttk.Label(row4, text="Save to:").pack(side='left')
        ttk.Entry(row4, textvariable=self.output_dir, width=54).pack(
            side='left', padx=(5, 0), fill='x', expand=True)
        ttk.Button(row4, text="Browse", command=self._browse_output).pack(side='left', padx=(8, 0))

        # ── Run ──
        ttk.Button(main, text="▶  Extract & Populate Munim", style='Run.TButton',
                   command=self._run).pack(pady=14, ipadx=25, ipady=6)

        # ── Status ──
        status_frame = ttk.Frame(main, relief='sunken', borderwidth=1)
        status_frame.pack(fill='x')
        ttk.Label(status_frame, textvariable=self.status_var, style='Status.TLabel',
                  padding=4).pack(side='left', fill='x')

    # ── Browse Handlers ──

    def _browse_recon(self):
        path = filedialog.askopenfilename(
            title="Select 2B vs Books Reconciliation File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.recon_path.set(path)
            self.status_var.set(f"Recon: {os.path.basename(path)}")

    def _browse_munim(self):
        path = filedialog.askopenfilename(
            title="Select Munim Template File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.munim_path.set(path)
            self.status_var.set(f"Template: {os.path.basename(path)}")

    def _browse_output(self):
        path = filedialog.askdirectory(title="Select Output Folder")
        if path:
            self.output_dir.set(path)

    # ── Main Processing ──

    def _run(self):
        if not self.recon_path.get():
            messagebox.showerror("Missing", "Please select the Reconciliation file.")
            return
        if not self.munim_path.get():
            messagebox.showerror("Missing", "Please select the Munim Template file.")
            return
        if not os.path.isdir(self.output_dir.get()):
            messagebox.showerror("Invalid", "Output directory does not exist.")
            return

        month = self.selected_month.get()
        if not month:
            messagebox.showerror("Missing", "Please select a month.")
            return

        try:
            # ── Step 1: Read reconciliation file ──
            self.status_var.set("Step 1/5 — Reading reconciliation file ...")
            self.root.update_idletasks()

            df = pd.read_excel(self.recon_path.get(), header=5, dtype=str)

            # Verify expected columns exist
            required = ['Data From', 'Month', 'GSTIN-UIN', 'Document Type',
                        'Document number', 'Document Date', 'Document Value',
                        'Taxable', 'Integrated Tax', 'Central Tax', 'State Tax',
                        'Cess', 'Remarks1 for GSTR 3B']
            missing = [c for c in required if c not in df.columns]
            if missing:
                messagebox.showerror("Column Error",
                    f"These expected columns are missing from the recon file:\n{', '.join(missing)}\n\n"
                    f"Available columns:\n{', '.join(df.columns[:20])}")
                self.status_var.set("Error — column mismatch")
                return

            # ── Step 2: Filter data ──
            self.status_var.set(f"Step 2/5 — Filtering: 2B + Matched in {month} / Nil in 2B ...")
            self.root.update_idletasks()

            # Filter 1: Data From == '2B'
            mask_2b = df['Data From'].str.strip() == '2B'

            # Filter 2: Remarks-based month filter
            # "Matched in [Month]" → match by the REMARK month (not the Month column)
            #   so "Matched in July" is picked even if Month column says April/May/June
            # "Nil in 2B" → match by the Month column (remark has no month)
            remarks_col = df['Remarks1 for GSTR 3B'].fillna('').str.strip()
            remarks_lower = remarks_col.str.lower()

            mask_matched_in_month = remarks_lower == f'matched in {month.lower()}'
            mask_nil_in_2b = remarks_lower == 'nil in 2b'
            mask_nil_month = df['Month'].str.strip().str.lower() == month.lower()

            # Combine: (2B AND Matched in [selected month])
            #        OR (2B AND Nil in 2B AND Month column = selected month)
            filtered = df[
                mask_2b & (mask_matched_in_month | (mask_nil_in_2b & mask_nil_month))
            ].copy()

            if len(filtered) == 0:
                total_2b = mask_2b.sum()
                total_matched_any = (mask_2b & remarks_lower.str.startswith('matched in')).sum()
                messagebox.showinfo("No Data",
                    f"No records found for:\n"
                    f"  Data From: 2B\n"
                    f"  Remark: 'Matched in {month}' or 'Nil in 2B' (for {month})\n\n"
                    f"Total 2B rows: {total_2b}\n"
                    f"2B with any 'Matched in ...' remark: {total_matched_any}")
                self.status_var.set("No matching records found.")
                return

            # Split by Document Type
            doc_type_col = filtered['Document Type'].fillna('').str.strip().str.lower()
            invoices_df = filtered[doc_type_col == 'invoice'].copy()
            credit_notes_df = filtered[doc_type_col == 'credit note'].copy()

            matched_count = (mask_2b & mask_matched_in_month).sum()
            nil_count = (mask_2b & mask_nil_in_2b & mask_nil_month).sum()

            # ── Step 3: Map Invoices to b2b format ──
            self.status_var.set(f"Step 3/5 — Mapping {len(invoices_df)} invoices to b2b ...")
            self.root.update_idletasks()

            b2b_rows = []
            for idx, row in invoices_df.iterrows():
                b2b_rows.append(self._map_b2b_row(row))

            # ── Step 4: Map Credit Notes to cdnr format ──
            self.status_var.set(f"Step 4/5 — Mapping {len(credit_notes_df)} credit notes to cdnr ...")
            self.root.update_idletasks()

            cdnr_rows = []
            for idx, row in credit_notes_df.iterrows():
                cdnr_rows.append(self._map_cdnr_row(row))

            # ── Step 5: Write to Munim template copy ──
            self.status_var.set("Step 5/5 — Writing to Munim template ...")
            self.root.update_idletasks()

            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"Munim_{month}_{timestamp}.xlsx"
            output_path = os.path.join(self.output_dir.get(), output_filename)

            # Copy template
            shutil.copy2(self.munim_path.get(), output_path)

            wb = load_workbook(output_path)
            data_font = Font(name='Arial', size=10)
            date_fmt = 'DD-MMM-YYYY'

            # ── Write b2b sheet ──
            ws_b2b = wb['b2b']
            # Clear existing data (row 5 onwards)
            for r in range(5, ws_b2b.max_row + 1):
                for c in range(1, 23):
                    ws_b2b.cell(row=r, column=c).value = None

            # b2b columns (1-indexed):
            # A=Return Period, B=GSTIN, C=Invoice Number, D=Invoice date,
            # E=Invoice Value, F=Place Of Supply, G=Reverse Charge,
            # H=Invoice Type, I=Rate, J=Taxable Value, K=Integrated Tax,
            # L=Central Tax, M=State/UT Tax, N=Cess
            for i, rec in enumerate(b2b_rows):
                r = 5 + i
                ws_b2b.cell(row=r, column=1).value = None  # Return Period
                ws_b2b.cell(row=r, column=2).value = rec['gstin']
                ws_b2b.cell(row=r, column=3).value = rec['doc_no']
                cell_d = ws_b2b.cell(row=r, column=4)
                cell_d.value = rec['doc_date']
                cell_d.number_format = date_fmt
                ws_b2b.cell(row=r, column=5).value = rec['doc_value']
                ws_b2b.cell(row=r, column=6).value = rec['pos']
                ws_b2b.cell(row=r, column=7).value = 'N'
                ws_b2b.cell(row=r, column=8).value = 'Regular'
                ws_b2b.cell(row=r, column=9).value = rec['rate']
                ws_b2b.cell(row=r, column=10).value = rec['taxable']
                ws_b2b.cell(row=r, column=11).value = rec['igst']
                ws_b2b.cell(row=r, column=12).value = rec['cgst']
                ws_b2b.cell(row=r, column=13).value = rec['sgst']
                ws_b2b.cell(row=r, column=14).value = rec['cess']
                for c in range(1, 15):
                    ws_b2b.cell(row=r, column=c).font = data_font

            # ── Write cdnr sheet ──
            ws_cdnr = wb['cdnr']
            # Clear existing data (row 5 onwards)
            for r in range(5, ws_cdnr.max_row + 1):
                for c in range(1, 23):
                    ws_cdnr.cell(row=r, column=c).value = None

            # cdnr columns (1-indexed):
            # A=Return Period, B=GSTIN of Supplier,
            # C=Note/Refund Voucher Number, D=Note/Refund Voucher date,
            # E=Invoice/Advance Payment Voucher Number, F=Invoice/Advance Payment Voucher date,
            # G=Pre GST, H=Document Type, I=Reason For Issuing document,
            # J=Supply Type, K=Note/Refund Voucher Value, L=Rate,
            # M=Taxable Value, N=Integrated Tax, O=Central Tax,
            # P=State/UT Tax, Q=Cess
            for i, rec in enumerate(cdnr_rows):
                r = 5 + i
                ws_cdnr.cell(row=r, column=1).value = None  # Return Period
                ws_cdnr.cell(row=r, column=2).value = rec['gstin']
                ws_cdnr.cell(row=r, column=3).value = rec['doc_no']
                cell_d = ws_cdnr.cell(row=r, column=4)
                cell_d.value = rec['doc_date']
                cell_d.number_format = date_fmt
                ws_cdnr.cell(row=r, column=5).value = None  # Original Invoice No (not in 2B recon)
                ws_cdnr.cell(row=r, column=6).value = None  # Original Invoice Date
                ws_cdnr.cell(row=r, column=7).value = 'N'   # Pre GST
                ws_cdnr.cell(row=r, column=8).value = 'C'   # Document Type = Credit Note
                ws_cdnr.cell(row=r, column=9).value = None   # Reason (not in 2B recon)
                ws_cdnr.cell(row=r, column=10).value = rec.get('supply_type', 'Regular')
                ws_cdnr.cell(row=r, column=11).value = rec['doc_value']
                ws_cdnr.cell(row=r, column=12).value = rec['rate']
                ws_cdnr.cell(row=r, column=13).value = rec['taxable']
                ws_cdnr.cell(row=r, column=14).value = rec['igst']
                ws_cdnr.cell(row=r, column=15).value = rec['cgst']
                ws_cdnr.cell(row=r, column=16).value = rec['sgst']
                ws_cdnr.cell(row=r, column=17).value = rec['cess']
                for c in range(1, 18):
                    ws_cdnr.cell(row=r, column=c).font = data_font

            wb.save(output_path)
            wb.close()

            self.status_var.set(f"Done! b2b={len(b2b_rows)}, cdnr={len(cdnr_rows)}")

            # ── Summary ──
            msg = (
                f"Munim Template Populated Successfully!\n"
                f"{'─' * 45}\n"
                f"Month selected: {month}\n"
                f"Total 2B records extracted: {len(filtered)}\n"
                f"  → 'Matched in {month}': {matched_count}\n"
                f"  → 'Nil in 2B' ({month}): {nil_count}\n"
                f"{'─' * 45}\n"
                f"b2b sheet  (Invoices)     : {len(b2b_rows)}\n"
                f"cdnr sheet (Credit Notes) : {len(cdnr_rows)}\n"
                f"{'─' * 45}\n"
                f"Output: {output_filename}\n"
                f"\nOpen the output file?"
            )

            if messagebox.askyesno("Success", msg):
                os.startfile(output_path)

        except Exception as e:
            self.status_var.set(f"Error: {e}")
            messagebox.showerror("Error", f"An error occurred:\n\n{type(e).__name__}: {e}")

    # ── Row Mapping Helpers ──

    def _map_common(self, row):
        """Extract common fields from a recon row."""
        gstin = str(row['GSTIN-UIN']).strip() if pd.notna(row['GSTIN-UIN']) else ''
        doc_no = str(row['Document number']).strip() if pd.notna(row['Document number']) else ''
        if doc_no.endswith('.0'):
            doc_no = doc_no[:-2]

        doc_date = parse_excel_date(row['Document Date'])
        doc_value = safe_float(row['Document Value'])
        taxable = safe_float(row['Taxable'])
        igst = safe_float(row['Integrated Tax'])
        cgst = safe_float(row['Central Tax'])
        sgst = safe_float(row['State Tax'])
        cess = safe_float(row['Cess'])
        rate = derive_rate(taxable, igst, cgst, sgst)
        pos = map_place_of_supply(row.get('Place of Supply', None), gstin)

        return {
            'gstin': gstin, 'doc_no': doc_no, 'doc_date': doc_date,
            'doc_value': round(abs(doc_value), 2),
            'pos': pos, 'rate': rate,
            'taxable': round(abs(taxable), 2),
            'igst': round(abs(igst), 2),
            'cgst': round(abs(cgst), 2),
            'sgst': round(abs(sgst), 2),
            'cess': round(abs(cess), 2),
        }

    def _map_b2b_row(self, row):
        """Map a recon Invoice row to b2b format."""
        return self._map_common(row)

    def _map_cdnr_row(self, row):
        """Map a recon Credit Note row to cdnr format."""
        rec = self._map_common(row)
        # Add supply type from recon (defaults to Regular)
        supply_type = str(row.get('Note Supply type', 'Regular')).strip()
        if not supply_type or supply_type == 'nan':
            supply_type = 'Regular'
        rec['supply_type'] = supply_type
        return rec

    def run(self):
        self.root.mainloop()


# ═══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    app = ReconToMunimApp()
    app.run()
