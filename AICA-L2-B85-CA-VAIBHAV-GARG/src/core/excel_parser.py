"""Excel parsing module for Schedule III financial statement workbooks."""
from dataclasses import dataclass, field
import hashlib
import os
import re
from typing import Dict, List, Optional, Tuple, Any
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ParsedLineItem:
    sheet: str
    row_no: int
    raw_label: str
    normalised_label: str
    amount_reporting: Optional[float]
    amount_comparative: Optional[float]
    section: Optional[str] = None


@dataclass
class SheetMetadata:
    sheet_type: str
    sheet_name: str
    header_row: int
    reporting_year_col: int
    comparative_year_col: Optional[int]
    reporting_period: str
    comparative_period: Optional[str]
    reporting_year: int
    comparative_year: Optional[int]


@dataclass
class WorkbookParseResult:
    file_path: str
    file_name: str
    file_hash: str
    units: str
    reporting_year: int
    comparative_year: Optional[int]
    reporting_period_label: str
    comparative_period_label: Optional[str]
    sheet_metadata: Dict[str, SheetMetadata]
    line_items: List[ParsedLineItem] = field(default_factory=list)


class ExcelParsingError(Exception):
    pass


SHEET_SYNONYMS = {
    "BS": {"bs", "balancesheet", "bal sheet", "b/s", "balance sheet"},
    "PL": {"pl", "p&l", "p l", "profit and loss", "statement of profit and loss", "pandl", "profit & loss", "statement of p&l"},
    "CF": {"cf", "cash flow", "cashflow", "cash flow statement", "cashflows"},
}


def compute_file_sha256(file_path: str) -> str:
    sha256_hash = hashlib.sha256()
    with open(file_path, "rb") as f:
        for byte_block in iter(lambda: f.read(65536), b""):
            sha256_hash.update(byte_block)
    return sha256_hash.hexdigest()


def normalize_whitespace(text: str) -> str:
    if not text:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


def strip_leading_enumerators(text: str) -> str:
    """Strip leading enumerators like '(a)', '(b)', '(i)', 'I.', 'IV', '1.', 'A.', and bullets."""
    cleaned = normalize_whitespace(text)
    # Strip leading dashes / bullets
    cleaned = re.sub(r"^[\-–—•\*\.\s]+", "", cleaned)
    # 1. Parenthesized e.g. (a), (b), (i), (ii), (iv), (1)
    cleaned = re.sub(r"^\([a-zA-Z0-9ivxIVX]{1,4}\)\s*", "", cleaned)
    # 2. Roman numerals e.g. "I.", "II.", "III.", "IV.", "IV ", "V.", "VI.", "VII."
    cleaned = re.sub(r"^(?:I|II|III|IV|V|VI|VII|VIII|IX|X|XI|XII)[\.\s:\-]+\s*", "", cleaned, flags=re.IGNORECASE)
    # 3. Single letter or number with dot or closing parenthesis e.g. "A.", "B.", "1.", "2.", "a)", "1)"
    cleaned = re.sub(r"^[a-zA-Z0-9][\.\)]\s*", "", cleaned)
    # Strip any remaining leading dashes
    cleaned = re.sub(r"^[\-–—•\*\.\s]+", "", cleaned)
    return cleaned.strip()


def parse_numeric_cell_value(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    
    val_str = normalize_whitespace(str(val))
    if not val_str or val_str in ("-", "—", "–", "nil", "Nil", "NIL", "N/A", "n/a", "NA"):
        return 0.0
    
    is_negative = False
    if val_str.startswith("(") and val_str.endswith(")"):
        is_negative = True
        val_str = val_str[1:-1].strip()
    elif val_str.startswith("-"):
        is_negative = True
        val_str = val_str[1:].strip()
    
    val_str = re.sub(r"[,₹$Rs\.\s]", lambda m: "." if m.group(0) == "." else "", val_str)
    
    try:
        num = float(val_str)
        return -num if is_negative else num
    except ValueError:
        return None


def extract_year_from_header_text(header_text: str) -> Optional[Tuple[int, str]]:
    if not header_text:
        return None
    text = normalize_whitespace(header_text)
    date_match = re.search(r"31[\./\-]03[\./\-](20\d{2})", text)
    if date_match:
        return int(date_match.group(1)), text
    month_match = re.search(r"(?:March\s+31(?:st)?,?|31(?:st)?\s+March,?)\s+(20\d{2})", text, re.IGNORECASE)
    if month_match:
        return int(month_match.group(1)), text
    fy_match = re.search(r"(?:FY\s*)?(20\d{2})\s*[-–/]\s*(\d{2,4})", text, re.IGNORECASE)
    if fy_match:
        start_year = int(fy_match.group(1))
        end_part = fy_match.group(2)
        end_year = int(str(start_year)[:2] + end_part) if len(end_part) == 2 else int(end_part)
        return end_year, text
    year_match = re.search(r"\b(20\d{2})\b", text)
    if year_match:
        return int(year_match.group(1)), text
    return None


def detect_units(worksheet: Worksheet) -> str:
    unit_patterns = [
        (r"(?:Rs\.?\s*in\s*lacs?|in\s*lacs?|Rs\.?\s*in\s*lakhs?|in\s*lakhs?)", "Lacs"),
        (r"(?:Rs\.?\s*in\s*crores?|in\s*crores?)", "Crores"),
        (r"(?:in\s*'000|in\s*thousands?|Rs\.?\s*in\s*thousands?)", "'000"),
        (r"(?:Rs\.?\s*in\s*millions?|in\s*millions?)", "Millions"),
        (r"(?:in\s*rupees?|in\s*Rs\.?)", "Rupees"),
    ]
    for r in range(1, min(10, worksheet.max_row + 1)):
        for c in range(1, min(15, worksheet.max_column + 1)):
            val = worksheet.cell(r, c).value
            if val:
                val_str = str(val)
                for pattern, unit_name in unit_patterns:
                    if re.search(pattern, val_str, re.IGNORECASE):
                        return unit_name
    return "Lacs"


def identify_sheets(workbook: openpyxl.Workbook) -> Dict[str, str]:
    sheet_map = {}
    for actual_name in workbook.sheetnames:
        cleaned = normalize_whitespace(actual_name).lower()
        for sheet_type, synonyms in SHEET_SYNONYMS.items():
            if cleaned in synonyms and sheet_type not in sheet_map:
                sheet_map[sheet_type] = actual_name
                break
    missing = [st for st in ("BS", "PL", "CF") if st not in sheet_map]
    if missing:
        raise ExcelParsingError(f"Required financial statement sheet(s) missing: {', '.join(missing)}.")
    return sheet_map


def find_header_row_and_columns(
    sheet: Worksheet,
    sheet_type: str
) -> Tuple[int, int, Optional[int], str, Optional[str], int, Optional[int]]:
    header_row = -1
    for r in range(1, min(16, sheet.max_row + 1)):
        for c in range(1, min(6, sheet.max_column + 1)):
            val = sheet.cell(r, c).value
            if val and "particular" in normalize_whitespace(str(val)).lower():
                header_row = r
                break
        if header_row != -1:
            break
            
    if header_row == -1:
        raise ExcelParsingError(f"Could not find 'Particulars' header row in sheet '{sheet.title}'.")
        
    candidates = []
    for c in range(1, sheet.max_column + 1):
        cell_val = sheet.cell(header_row, c).value
        if not cell_val:
            continue
        header_str = normalize_whitespace(str(cell_val))
        if header_str.upper() in ("A", "B", "C", "D", "E", "NOTE", "NOTE NO", "NOTE NO."):
            continue
        res = extract_year_from_header_text(header_str)
        if res:
            year, period_label = res
            candidates.append((c, year, period_label))
            
    if not candidates:
        raise ExcelParsingError(f"No period/year headers found on row {header_row} in sheet '{sheet.title}'.")
        
    candidates.sort(key=lambda item: item[1], reverse=True)
    rep_col, rep_year, rep_label = candidates[0]
    comp_col, comp_year, comp_label = (None, None, None)
    if len(candidates) > 1:
        comp_col, comp_year, comp_label = candidates[1]
        
    return header_row, rep_col, comp_col, rep_label, comp_label, rep_year, comp_year


def parse_workbook(file_path: str) -> WorkbookParseResult:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
        
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xls":
        raise ExcelParsingError("Legacy .xls format is not supported. Please convert to .xlsx or .xlsm.")
    if ext not in (".xlsx", ".xlsm"):
        raise ExcelParsingError(f"Unsupported file format: {ext}. Only .xlsx and .xlsm files are supported.")
        
    file_hash = compute_file_sha256(file_path)
    file_name = os.path.basename(file_path)
    
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        if "password" in str(e).lower() or "encrypted" in str(e).lower():
            raise ExcelParsingError("The Excel file is password-protected and cannot be opened.")
        raise ExcelParsingError(f"Failed to load workbook: {str(e)}")
        
    sheet_map = identify_sheets(wb_data)
    units = "Lacs"
    sheet_metadata: Dict[str, SheetMetadata] = {}
    all_line_items: List[ParsedLineItem] = []
    
    primary_rep_year = None
    primary_comp_year = None
    primary_rep_label = ""
    primary_comp_label = None
    
    for sheet_type in ("BS", "PL", "CF"):
        sheet_name = sheet_map[sheet_type]
        ws = wb_data[sheet_name]
        
        if sheet_type == "BS":
            units = detect_units(ws)
            
        header_row, rep_col, comp_col, rep_label, comp_label, rep_year, comp_year = find_header_row_and_columns(
            ws, sheet_type
        )
        
        if primary_rep_year is None:
            primary_rep_year = rep_year
            primary_comp_year = comp_year
            primary_rep_label = rep_label
            primary_comp_label = comp_label
            
        sheet_metadata[sheet_type] = SheetMetadata(
            sheet_type=sheet_type,
            sheet_name=sheet_name,
            header_row=header_row,
            reporting_year_col=rep_col,
            comparative_year_col=comp_col,
            reporting_period=rep_label,
            comparative_period=comp_label,
            reporting_year=rep_year,
            comparative_year=comp_year,
        )
        
        current_cf_section = None
        for r in range(header_row + 1, ws.max_row + 1):
            col_a_val = ws.cell(r, 1).value
            col_b_val = ws.cell(r, 2).value
            
            str_a = normalize_whitespace(str(col_a_val)) if col_a_val is not None else ""
            str_b = normalize_whitespace(str(col_b_val)) if col_b_val is not None else ""
            
            if sheet_type == "CF":
                if str_a.upper() in ("A.", "A", "B.", "B", "C.", "C"):
                    current_cf_section = str_a.upper().replace(".", "")
            
            if str_a and str_b:
                raw_label = f"{str_a} {str_b}"
            elif str_a:
                raw_label = str_a
            elif str_b:
                raw_label = str_b
            else:
                raw_label = ""
                
            normalised_label = strip_leading_enumerators(raw_label)
            
            rep_val = ws.cell(r, rep_col).value
            comp_val = ws.cell(r, comp_col).value if comp_col else None
            
            amt_rep = parse_numeric_cell_value(rep_val)
            amt_comp = parse_numeric_cell_value(comp_val) if comp_col else None
            
            if not raw_label and amt_rep is None and amt_comp is None:
                continue
                
            all_line_items.append(
                ParsedLineItem(
                    sheet=sheet_type,
                    row_no=r,
                    raw_label=raw_label,
                    normalised_label=normalised_label,
                    amount_reporting=amt_rep,
                    amount_comparative=amt_comp,
                    section=current_cf_section,
                )
            )
            
    return WorkbookParseResult(
        file_path=file_path,
        file_name=file_name,
        file_hash=file_hash,
        units=units,
        reporting_year=primary_rep_year or 2026,
        comparative_year=primary_comp_year,
        reporting_period_label=primary_rep_label,
        comparative_period_label=primary_comp_label,
        sheet_metadata=sheet_metadata,
        line_items=all_line_items,
    )
