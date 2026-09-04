"""Excel exporter generating live-formula ratio analysis workbook (§10)."""
from datetime import datetime
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.core.calculator import SingleRatioResult, CalculationResultSet
from src.core.assumptions import AssumptionItem
from src.core.integrity import IntegrityCheckResult


def export_ratios_to_excel(
    file_path: str,
    client_name: str,
    fy_end_date: str,
    units: str,
    result_set: CalculationResultSet,
    assumptions: Dict[str, AssumptionItem],
    integrity_results: List[IntegrityCheckResult],
    note_number: str = ""
) -> None:
    """Export Schedule III analytical ratios to styled Excel file with live variance formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule III Ratios"
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    navy_header_fill = PatternFill(start_color="0B4F8C", end_color="0B4F8C", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    title_font = Font(name="Segoe UI", size=14, bold=True, color="073763")
    subtitle_font = Font(name="Segoe UI", size=11, bold=False, color="1A2330")
    body_font = Font(name="Segoe UI", size=9.5, color="1A2330")
    bold_body_font = Font(name="Segoe UI", size=9.5, bold=True, color="1A2330")
    green_font = Font(name="Segoe UI", size=9.5, bold=True, color="1E8E5A")
    red_font = Font(name="Segoe UI", size=9.5, bold=True, color="C0392B")
    muted_font = Font(name="Segoe UI", size=8.5, italic=True, color="5B6B7F")
    
    thin_border_side = Side(style="thin", color="D6DEE7")
    cell_border = Border(top=thin_border_side, bottom=thin_border_side, left=thin_border_side, right=thin_border_side)
    
    # 1. Title rows
    ws.merge_cells("A1:H1")
    ws["A1"] = client_name
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Notes forming part of the Financial Statements for the year ended {fy_end_date}"
    ws["A2"].font = subtitle_font
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    note_txt = f"Note {note_number}: Analytical Ratios" if note_number else "Note: Analytical Ratios"
    ws["A4"] = note_txt
    ws["A4"].font = Font(name="Segoe UI", size=11, bold=True, color="0B4F8C")
    
    ws["H4"] = f"(Rs. in {units})"
    ws["H4"].font = muted_font
    ws["H4"].alignment = Alignment(horizontal="right")
    
    # 2. Table Headers
    headers = [
        "S. No.", "Ratio", "Numerator", "Denominator",
        f"Current Year ({result_set.cy_label})",
        f"Previous Year ({result_set.py_label})",
        "% Variance",
        "Reason for variance (where 25% or more)"
    ]
    
    start_row = 6
    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=h_text)
        cell.fill = navy_header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="right" if col_idx in (5, 6, 7) else ("center" if col_idx == 1 else "left"),
            vertical="center",
            wrap_text=True
        )
        cell.border = cell_border
        
    ws.row_dimensions[start_row].height = 28
    
    # 3. Data Rows with Live Formulas
    for r_idx, r_data in enumerate(result_set.schedule_iii_ratios):
        curr_row = start_row + 1 + r_idx
        ws.row_dimensions[curr_row].height = 24
        fill = alt_row_fill if r_idx % 2 == 1 else white_fill
        
        # Col A: S. No.
        c_a = ws.cell(row=curr_row, column=1, value=r_data.id)
        c_a.alignment = Alignment(horizontal="center", vertical="center")
        
        # Col B: Name
        c_b = ws.cell(row=curr_row, column=2, value=r_data.name)
        c_b.alignment = Alignment(horizontal="left", vertical="center")
        
        # Col C: Num Desc
        c_c = ws.cell(row=curr_row, column=3, value=r_data.numerator_desc)
        c_c.alignment = Alignment(horizontal="left", vertical="center")
        
        # Col D: Den Desc
        c_d = ws.cell(row=curr_row, column=4, value=r_data.denominator_desc)
        c_d.alignment = Alignment(horizontal="left", vertical="center")
        
        # Col E: CY Value
        c_e = ws.cell(row=curr_row, column=5)
        if r_data.value_cy is not None:
            if r_data.is_percentage:
                c_e.value = r_data.value_cy / 100.0
                c_e.number_format = "0.00%"
            else:
                c_e.value = r_data.value_cy
                c_e.number_format = "0.00"
        else:
            c_e.value = "Not meaningful"
        c_e.alignment = Alignment(horizontal="right", vertical="center")
        
        # Col F: PY Value
        c_f = ws.cell(row=curr_row, column=6)
        if r_data.value_py is not None:
            if r_data.is_percentage:
                c_f.value = r_data.value_py / 100.0
                c_f.number_format = "0.00%"
            else:
                c_f.value = r_data.value_py
                c_f.number_format = "0.00"
        else:
            c_f.value = "Not meaningful"
        c_f.alignment = Alignment(horizontal="right", vertical="center")
        
        # Col G: % Variance (Live Excel Formula)
        c_g = ws.cell(row=curr_row, column=7)
        if r_data.value_cy is not None and r_data.value_py is not None and r_data.value_py != 0:
            c_g.value = f"=(E{curr_row}-F{curr_row})/ABS(F{curr_row})"
            c_g.number_format = "+0.00%;-0.00%;0.00%"
            if r_data.is_flagged:
                c_g.font = green_font if (r_data.variance_pct or 0) > 0 else red_font
        else:
            c_g.value = "NM"
            c_g.alignment = Alignment(horizontal="right", vertical="center")
            
        # Col H: Reason
        reason_txt = r_data.reason_final if (r_data.is_flagged or r_data.status == "Not meaningful") else "—"
        c_h = ws.cell(row=curr_row, column=8, value=reason_txt)
        c_h.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        for c in (c_a, c_b, c_c, c_d, c_e, c_f, c_g, c_h):
            c.fill = fill
            c.border = cell_border
            if not c.font or c.font.color.rgb == "000000":
                c.font = body_font

    # Column Widths
    col_widths = {1: 8, 2: 26, 3: 28, 4: 28, 5: 16, 6: 16, 7: 16, 8: 45}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    # 4. Assumptions & Basis of Preparation Tab
    ws_assump = wb.create_sheet(title="Assumptions & Disclosures")
    ws_assump.views.sheetView[0].showGridLines = True
    
    ws_assump["A1"] = "Assumptions and Basis of Preparation"
    ws_assump["A1"].font = title_font
    
    r_pos = 3
    for k, item in assumptions.items():
        ws_assump.cell(row=r_pos, column=1, value=item.name).font = bold_body_font
        ws_assump.cell(row=r_pos, column=2, value=item.disclosure_text).font = body_font
        r_pos += 1
        
    ws_assump.column_dimensions["A"].width = 35
    ws_assump.column_dimensions["B"].width = 80
    
    wb.save(file_path)
