"""Finished, reproducible audit deliverables for The 45-Day Clock.

The public entry point is :func:`export_all`.  It writes only to the folder
the caller supplies and returns the four exact paths created there.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Mapping, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .classify import (
    EVIDENCE_RANK,
    GATE_ACTIVITY,
    GATE_CLASS,
    GATE_REGISTRATION,
    GATE_TIMING,
    SRC_DECLARATION,
    UdyamRecord,
)
from .engine import (
    ACC_POLICY_TEXT,
    ALLOWED_LATE_INTEREST_ONLY,
    ComputationRun,
    PaymentLine,
    PurchaseLine,
    action_list,
    exclusion_register,
)
from .rules import BANK_RATE_TABLE, fy_bounds, resolve_credit_period


INK = "12233A"
TEAL = "0F766E"
PALE_TEAL = "E8F5F3"
BACKGROUND = "F7F8FA"
WHITE = "FFFFFF"
MUTED = "667085"
LINE = "D9DEE7"
RED = "B91C1C"
PALE_RED = "FEF2F2"
AMBER = "B45309"
PALE_AMBER = "FFFBEB"
GREEN = "15803D"
PALE_GREEN = "F0FDF4"
INDIAN_CURRENCY = '"Rs "#,##,##0.00;[Red]-"Rs "#,##,##0.00'
INDIAN_NUMBER = "#,##,##0"
DATE_FORMAT = "dd-mmm-yyyy"
THIN_LINE = Side(style="thin", color=LINE)

INDIA_CODE_SECTION_2 = (
    "MSMED Act, 2006, s.2(n) (definition of supplier): "
    "https://www.indiacode.nic.in/show-data?actid=AC_CEN_46_77_00002_200627_1517807324919&orderno=2&sectionId=9884&sectionno=2"
)
INDIA_CODE_SECTION_8 = (
    "MSMED Act, 2006, s.8(1) (filing of memorandum): "
    "https://www.indiacode.nic.in/show-data?actid=AC_CEN_46_77_00002_200627_1517807324919&orderno=8"
)
TRADER_CITATION = (
    "Ministry of MSME OMs dated 02.07.2021 and 01.09.2021; Ministry of MSME "
    "Annual Report 2021-22, 'Inclusion of Retail and Wholesale Trade' "
    "(benefits restricted to Priority Sector Lending): "
    "https://msme.gov.in/sites/default/files/MSMEENGLISHANNUALREPORT2021-22.pdf"
)

GATE_LABELS = {
    GATE_CLASS: "Gate 1 - Enterprise class",
    GATE_ACTIVITY: "Gate 2 - Trader activity",
    GATE_REGISTRATION: "Gate 3 - Udyam registration",
    GATE_TIMING: "Gate 4 - Registration timing",
}
GATE_CITATIONS = {
    GATE_CLASS: INDIA_CODE_SECTION_2 + "; MSMED Act, 2006, s.7(1).",
    GATE_ACTIVITY: TRADER_CITATION,
    GATE_REGISTRATION: INDIA_CODE_SECTION_2 + "; " + INDIA_CODE_SECTION_8,
    GATE_TIMING: INDIA_CODE_SECTION_2 + "; " + INDIA_CODE_SECTION_8,
}


@dataclass(frozen=True)
class ExportContext:
    run: ComputationRun
    purchases: list[PurchaseLine]
    payments: list[PaymentLine]
    udyam: Mapping[str, UdyamRecord]
    entity_pan: str = "Not provided"
    preparer: str = ""
    reviewer: str = ""
    firm_name: str = "FIRM PROFILE NOT CONFIGURED"
    firm_frn: str = ""
    firm_address: str = ""
    firm_email: str = ""
    firm_phone: str = ""
    document_status: str = "DRAFT"
    vendor_contacts: Mapping[str, str] = field(default_factory=dict)
    tax_rate: Decimal = Decimal("0.2517")
    tax_rate_basis: str = (
        "Illustrative effective rate; confirm against the client's applicable tax regime"
    )
    vendor_metadata: Mapping[str, Mapping[str, str]] = field(default_factory=dict)
    evidence_documents: tuple[Mapping[str, object], ...] = ()


@dataclass(frozen=True)
class ExportBundle:
    clause_22: Path
    action_list: Path
    exclusion_register: Path
    working_paper: Path

    def paths(self) -> tuple[Path, Path, Path, Path]:
        return (
            self.clause_22,
            self.action_list,
            self.exclusion_register,
            self.working_paper,
        )


@dataclass(frozen=True)
class AuditPackBundle:
    root: Path
    files: tuple[Path, ...]

    def paths(self) -> tuple[Path, ...]:
        return self.files


def _purchase_map(context: ExportContext) -> dict[str, PurchaseLine]:
    return {line.invoice_id: line for line in context.purchases}


def _payments_map(context: ExportContext) -> dict[str, list[PaymentLine]]:
    grouped: dict[str, list[PaymentLine]] = {}
    for payment in context.payments:
        grouped.setdefault(payment.invoice_id, []).append(payment)
    return grouped


def _unpaid_at_year_end(
    purchase: PurchaseLine,
    payments: list[PaymentLine],
    financial_year: str,
) -> Decimal:
    _, year_end = fy_bounds(financial_year)
    paid = sum(
        (payment.amount for payment in payments if payment.payment_date <= year_end),
        Decimal("0.00"),
    )
    return max(Decimal("0.00"), purchase.amount - paid).quantize(Decimal("0.01"))


def _bank_rate_for_year_end(financial_year: str) -> tuple[Decimal, date]:
    _, year_end = fy_bounds(financial_year)
    applicable = [(effective, rate) for effective, rate in BANK_RATE_TABLE if effective <= year_end]
    if not applicable:
        raise ValueError(f"No RBI Bank Rate recorded for {year_end}")
    effective, rate = applicable[-1]
    return rate, effective


def _weak_vendor_count(context: ExportContext) -> int:
    vendor_ids = {line.vendor_id for line in context.purchases}
    declaration_rank = EVIDENCE_RANK[SRC_DECLARATION]
    return sum(
        EVIDENCE_RANK.get(context.udyam.get(vendor_id, UdyamRecord(vendor_id)).source, 0)
        < declaration_rank
        for vendor_id in vendor_ids
    )


def _set_document_properties(workbook: Workbook, context: ExportContext, title: str) -> None:
    workbook.properties.title = title
    workbook.properties.subject = (
        f"{context.run.entity_name} - FY {context.run.fy} - Run {context.run.run_hash()}"
    )
    workbook.properties.creator = "The 45-Day Clock"
    workbook.properties.description = (
        f"Reproducible output using rule pack {context.run.rule_pack_version}. "
        f"Run hash {context.run.run_hash()}."
    )


def _title_block(sheet, title: str, subtitle: str, last_column: int, run_hash: str) -> None:
    last = get_column_letter(last_column)
    sheet.merge_cells(f"A1:{last}1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=INK)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells(f"A2:{last}2")
    sheet["A2"] = f"{subtitle}  |  Run hash: {run_hash}"
    sheet["A2"].font = Font(name="Aptos", size=9, color=MUTED)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 22


def _header_row(sheet, row_number: int, headers: list[str]) -> None:
    for column, header in enumerate(headers, 1):
        cell = sheet.cell(row_number, column, header)
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=TEAL))
    sheet.row_dimensions[row_number].height = 30


def _body_style(
    sheet,
    start_row: int,
    end_row: int,
    last_column: int,
    *,
    currency_columns: tuple[int, ...] = (),
    date_columns: tuple[int, ...] = (),
    wrap_columns: tuple[int, ...] = (),
) -> None:
    if end_row < start_row:
        return
    for row in range(start_row, end_row + 1):
        if row % 2 == 0:
            for column in range(1, last_column + 1):
                sheet.cell(row, column).fill = PatternFill("solid", fgColor="FAFBFC")
        for column in range(1, last_column + 1):
            cell = sheet.cell(row, column)
            cell.font = Font(name="Aptos", size=9, color=INK)
            cell.border = Border(bottom=THIN_LINE)
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=column in wrap_columns,
                horizontal="right" if column in currency_columns else "left",
            )
        for column in currency_columns:
            sheet.cell(row, column).number_format = INDIAN_CURRENCY
        for column in date_columns:
            sheet.cell(row, column).number_format = DATE_FORMAT


def _finish_sheet(
    sheet,
    *,
    widths: Mapping[int, float],
    header_row: int,
    last_row: int,
    last_column: int,
    run_hash: str,
    landscape: bool = True,
) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = f"A{header_row + 1}"
    if last_row >= header_row:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(last_column)}{last_row}"
    sheet.print_title_rows = f"1:{header_row}"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.orientation = "landscape" if landscape else "portrait"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5
    sheet.oddHeader.center.text = f"&B{sheet.title}"
    sheet.oddFooter.left.text = f"Run hash: {run_hash}"
    sheet.oddFooter.right.text = "Page &P of &N"
    sheet.sheet_properties.outlinePr.summaryBelow = True


def build_clause_22_workbook(path: str | Path, context: ExportContext) -> Path:
    output = Path(path)
    workbook = Workbook()
    _set_document_properties(workbook, context, "Clause 22 Workbook")
    summary = workbook.active
    summary.title = "Summary"
    run = context.run
    run_hash = run.run_hash()

    _title_block(
        summary,
        "CLAUSE 22 - MSME PAYABLE EXPOSURE",
        f"{run.entity_name} | FY {run.fy}",
        8,
        run_hash,
    )
    labels = [
        ("Entity", run.entity_name),
        ("PAN", context.entity_pan),
        ("Financial year", run.fy),
        ("Governing provision", f"{run.statute['act']}, Section {run.statute['section']}"),
        ("Rule pack version", run.rule_pack_version),
        ("Run hash", run_hash),
    ]
    row = 4
    for label, value in labels:
        summary.cell(row, 1, label)
        summary.cell(row, 2, value)
        summary.cell(row, 1).font = Font(name="Aptos", size=9, bold=True, color=MUTED)
        summary.cell(row, 2).font = Font(name="Aptos", size=10, bold=label == "Run hash", color=INK)
        if label == "Run hash":
            summary.cell(row, 2).fill = PatternFill("solid", fgColor=PALE_TEAL)
        row += 1

    firm_lines = [
        ("CA firm", context.firm_name),
        ("FRN", context.firm_frn or "Not provided"),
        ("Document status", context.document_status),
    ]
    for offset, (label, value) in enumerate(firm_lines, 8):
        summary.cell(offset, 4, label)
        summary.merge_cells(start_row=offset, start_column=5, end_row=offset, end_column=7)
        summary.cell(offset, 5, value)
        summary.cell(offset, 4).font = Font(name="Aptos", size=9, bold=True, color=MUTED)
        summary.cell(offset, 5).font = Font(name="Aptos", size=9, color=INK)

    summary.merge_cells("D4:E4")
    summary["D4"] = "DISALLOWANCE"
    summary.merge_cells("D5:E6")
    summary["D5"] = run.disallowance_total
    summary["D5"].number_format = INDIAN_CURRENCY
    summary.merge_cells("F4:G4")
    summary["F4"] = "MSMED s.16 INTEREST"
    summary.merge_cells("F5:G6")
    summary["F5"] = run.interest_total
    summary["F5"].number_format = INDIAN_CURRENCY
    for label_cell, value_cell, fill, color in (
        ("D4", "D5", PALE_RED, RED),
        ("F4", "F5", PALE_AMBER, AMBER),
    ):
        summary[label_cell].fill = PatternFill("solid", fgColor=fill)
        summary[label_cell].font = Font(name="Aptos", size=9, bold=True, color=color)
        summary[label_cell].alignment = Alignment(horizontal="center", vertical="center")
        summary[value_cell].fill = PatternFill("solid", fgColor=fill)
        summary[value_cell].font = Font(name="Aptos Display", size=18, bold=True, color=color)
        summary[value_cell].alignment = Alignment(horizontal="center", vertical="center")

    summary.merge_cells("A12:H12")
    summary["A12"] = "CONTROL TOTALS"
    summary["A12"].fill = PatternFill("solid", fgColor=INK)
    summary["A12"].font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    source_hash = run.control_totals.get("source_file_sha256", "")
    control_rows = [
        ("Source file", run.control_totals.get("source_file_name", "Not recorded"), "General"),
        ("Source SHA-256", source_hash or "Not recorded", "General"),
        ("Imported at", run.control_totals.get("imported_at", "Not recorded"), "General"),
        ("Source purchase lines", run.control_totals.get("source_purchase_lines", 0), INDIAN_NUMBER),
        ("Source purchase value", run.control_totals.get("source_purchase_value", Decimal("0")), INDIAN_CURRENCY),
        ("Lines in selected FY", run.control_totals.get("ledger_lines_in_year", 0), INDIAN_NUMBER),
        ("Value in selected FY", run.control_totals.get("ledger_value", Decimal("0")), INDIAN_CURRENCY),
        ("Out-of-period lines", run.control_totals.get("out_of_period_lines", 0), INDIAN_NUMBER),
        ("Out-of-period value", run.control_totals.get("out_of_period_value", Decimal("0")), INDIAN_CURRENCY),
        ("Value accounted for in FY", run.control_totals.get("value_accounted_for", Decimal("0")), INDIAN_CURRENCY),
        ("Totals tie", "YES" if run.control_totals.get("ties") else "NO - DO NOT SIGN", "General"),
    ]
    for offset, (label, value, number_format) in enumerate(control_rows, 13):
        summary.cell(offset, 1, label)
        summary.merge_cells(start_row=offset, start_column=2, end_row=offset, end_column=3)
        summary.cell(offset, 2, value)
        summary.cell(offset, 2).number_format = number_format
        summary.cell(offset, 1).font = Font(name="Aptos", size=9, bold=True, color=MUTED)
        summary.cell(offset, 2).font = Font(name="Aptos", size=10, bold=True, color=INK)
    tie_row = 12 + len(control_rows)
    tie_cell = summary.cell(tie_row, 2)
    tie_cell.fill = PatternFill(
        "solid", fgColor=PALE_GREEN if run.control_totals.get("ties") else PALE_RED
    )
    tie_cell.font = Font(
        name="Aptos", size=10, bold=True,
        color=GREEN if run.control_totals.get("ties") else RED,
    )

    signoff_row = tie_row + 3
    summary.merge_cells(start_row=signoff_row, start_column=1, end_row=signoff_row, end_column=8)
    summary.cell(signoff_row, 1, "SIGN-OFF")
    summary.cell(signoff_row, 1).fill = PatternFill("solid", fgColor=INK)
    summary.cell(signoff_row, 1).font = Font(name="Aptos", size=9, bold=True, color=WHITE)
    summary.cell(signoff_row + 2, 1, "Prepared by")
    summary.cell(signoff_row + 2, 2, context.preparer)
    summary.cell(signoff_row + 2, 4, "Date")
    summary.cell(signoff_row + 2, 5, "")
    summary.cell(signoff_row + 4, 1, "Reviewed by")
    summary.cell(signoff_row + 4, 2, context.reviewer)
    summary.cell(signoff_row + 4, 4, "Date")
    summary.cell(signoff_row + 4, 5, "")
    for cell_ref in (
        f"B{signoff_row + 2}", f"E{signoff_row + 2}",
        f"B{signoff_row + 4}", f"E{signoff_row + 4}",
    ):
        summary[cell_ref].border = Border(bottom=Side(style="thin", color=INK))
    disclaimer_row = signoff_row + 7
    summary.cell(disclaimer_row, 1, (
        "Strictly private and confidential. This workbook is a computation aid, not "
        "professional advice. Validate source data, evidence and legal conclusions before sign-off."
    ))
    summary.cell(disclaimer_row, 1).font = Font(name="Aptos", size=8, italic=True, color=MUTED)
    summary.merge_cells(start_row=disclaimer_row, start_column=1, end_row=disclaimer_row, end_column=8)
    _finish_sheet(
        summary,
        widths={1: 24, 2: 26, 3: 4, 4: 18, 5: 18, 6: 18, 7: 18, 8: 4},
        header_row=2,
        last_row=disclaimer_row,
        last_column=8,
        run_hash=run_hash,
        landscape=False,
    )
    summary.freeze_panes = "A4"
    summary.auto_filter.ref = None
    summary.print_area = f"A1:H{disclaimer_row}"

    purchase_by_invoice = _purchase_map(context)
    payments_by_invoice = _payments_map(context)

    invoice_detail = workbook.create_sheet("Invoice detail")
    invoice_headers = [
        "Invoice no.", "Vendor", "Udyam no.", "Invoice date", "Acceptance date",
        "Credit days", "Credit basis", "Due date", "Appointed day", "Invoice amount",
        "Unpaid at year end", "Disallowance", "Interest",
    ]
    _title_block(
        invoice_detail,
        "DISALLOWED INVOICE DETAIL",
        f"{run.entity_name} | Section {run.statute['section']}",
        len(invoice_headers),
        run_hash,
    )
    _header_row(invoice_detail, 4, invoice_headers)
    row = 5
    for finding in (item for item in run.findings if item.disallowance > 0):
        purchase = purchase_by_invoice[finding.invoice_id]
        credit = resolve_credit_period(purchase.agreement_days)
        record = context.udyam.get(finding.vendor_id)
        values = [
            finding.invoice_id,
            finding.vendor_name,
            record.udyam_no if record else "",
            finding.invoice_date,
            finding.acceptance_date,
            credit.days,
            credit.basis,
            finding.due_date,
            finding.appointed_day,
            finding.amount,
            _unpaid_at_year_end(
                purchase, payments_by_invoice.get(finding.invoice_id, []), run.fy
            ),
            finding.disallowance,
            finding.interest,
        ]
        for column, value in enumerate(values, 1):
            invoice_detail.cell(row, column, value)
        row += 1
    detail_last = row - 1
    _body_style(
        invoice_detail, 5, detail_last, len(invoice_headers),
        currency_columns=(10, 11, 12, 13),
        date_columns=(4, 5, 8, 9),
        wrap_columns=(2, 7),
    )
    _finish_sheet(
        invoice_detail,
        widths={1: 15, 2: 28, 3: 23, 4: 14, 5: 15, 6: 11, 7: 44,
                8: 14, 9: 14, 10: 17, 11: 18, 12: 17, 13: 17},
        header_row=4,
        last_row=detail_last,
        last_column=len(invoice_headers),
        run_hash=run_hash,
    )

    exclusions = workbook.create_sheet("Exclusions")
    exclusion_headers = [
        "Vendor", "Gate failed", "Invoices", "Turnover reviewed",
        "Correctly not disallowed", "Evidence strength", "Full legal reason", "Citation",
    ]
    _title_block(
        exclusions,
        "EXCLUSIONS - CORRECTLY NOT DISALLOWED",
        f"{run.entity_name} | Four coverage gates",
        len(exclusion_headers),
        run_hash,
    )
    _header_row(exclusions, 4, exclusion_headers)
    row = 5
    for item in exclusion_register(run):
        values = [
            item["vendor"], GATE_LABELS.get(item["gate"], item["gate"]), item["invoices"],
            item["turnover"], item["wrongly_disallowable"], item["evidence_strength"],
            item["reason"], GATE_CITATIONS.get(item["gate"], ""),
        ]
        for column, value in enumerate(values, 1):
            exclusions.cell(row, column, value)
        row += 1
    exclusions_last = row - 1
    _body_style(
        exclusions, 5, exclusions_last, len(exclusion_headers),
        currency_columns=(4, 5), wrap_columns=(1, 2, 7, 8),
    )
    for current in range(5, exclusions_last + 1):
        exclusions.row_dimensions[current].height = 45
    _finish_sheet(
        exclusions,
        widths={1: 28, 2: 27, 3: 10, 4: 19, 5: 22, 6: 15, 7: 70, 8: 70},
        header_row=4,
        last_row=exclusions_last,
        last_column=len(exclusion_headers),
        run_hash=run_hash,
    )

    interest_sheet = workbook.create_sheet("Interest only")
    interest_headers = [
        "Invoice no.", "Vendor", "Udyam no.", "Invoice date", "Acceptance date",
        "Credit days", "Credit basis", "Due date", "Payment date(s)",
        "Invoice amount", "Interest", "Conclusion",
    ]
    _title_block(
        interest_sheet,
        "PAID LATE WITHIN THE YEAR - INTEREST ONLY",
        "Principal deductible; MSMED s.16 interest is not deductible under s.23",
        len(interest_headers),
        run_hash,
    )
    _header_row(interest_sheet, 4, interest_headers)
    row = 5
    for finding in (
        item for item in run.findings if item.status == ALLOWED_LATE_INTEREST_ONLY
    ):
        purchase = purchase_by_invoice[finding.invoice_id]
        credit = resolve_credit_period(purchase.agreement_days)
        record = context.udyam.get(finding.vendor_id)
        payment_dates = ", ".join(
            payment.payment_date.strftime("%d-%b-%Y")
            for payment in payments_by_invoice.get(finding.invoice_id, [])
        )
        values = [
            finding.invoice_id, finding.vendor_name, record.udyam_no if record else "",
            finding.invoice_date, finding.acceptance_date, credit.days, credit.basis,
            finding.due_date, payment_dates, finding.amount, finding.interest, finding.reason,
        ]
        for column, value in enumerate(values, 1):
            interest_sheet.cell(row, column, value)
        row += 1
    interest_last = row - 1
    _body_style(
        interest_sheet, 5, interest_last, len(interest_headers),
        currency_columns=(10, 11), date_columns=(4, 5, 8),
        wrap_columns=(2, 7, 9, 12),
    )
    _finish_sheet(
        interest_sheet,
        widths={1: 15, 2: 28, 3: 23, 4: 14, 5: 15, 6: 11, 7: 42,
                8: 14, 9: 24, 10: 17, 11: 17, 12: 60},
        header_row=4,
        last_row=interest_last,
        last_column=len(interest_headers),
        run_hash=run_hash,
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def build_action_list_workbook(path: str | Path, context: ExportContext) -> Path:
    output = Path(path)
    workbook = Workbook()
    _set_document_properties(workbook, context, "31 March Action List")
    sheet = workbook.active
    sheet.title = "31 March Action List"
    run = context.run
    run_hash = run.run_hash()
    headers = [
        "Vendor", "Contact", "No. of invoices", "Pay by date", "Amount to pay",
        "Disallowance avoided", "Interest avoided", "Running cumulative total",
        "Estimated tax saved",
    ]
    _title_block(
        sheet,
        "31 MARCH ACTION LIST",
        f"{run.entity_name} | Work down this list before year end",
        len(headers),
        run_hash,
    )
    sheet.merge_cells("A3:I3")
    sheet["A3"] = (
        f"Priority is ranked by estimated tax saved at {context.tax_rate * 100:.2f}%. "
        f"Basis recorded by the firm: {context.tax_rate_basis}. "
        "Because one tax rate applies throughout, this is the same descending order as "
        "disallowance avoided. Update contact details before circulation."
    )
    sheet["A3"].font = Font(name="Aptos", size=9, italic=True, color=MUTED)
    sheet["A3"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[3].height = 32
    _header_row(sheet, 5, headers)
    rows = action_list(run, top_n=len(context.purchases))
    rows.sort(
        key=lambda item: item["disallowance_saved"] * context.tax_rate,
        reverse=True,
    )
    vendor_id_by_name = {}
    for finding in run.findings:
        vendor_id_by_name.setdefault(finding.vendor_name, finding.vendor_id)
    row = 6
    for item in rows:
        vendor_id = vendor_id_by_name.get(item["vendor"], "")
        values = [
            item["vendor"], context.vendor_contacts.get(vendor_id, ""), item["invoices"],
            item["earliest_due"], item["pay_now"], item["disallowance_saved"],
            item["interest_exposure"], None, None,
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row, column, value)
        sheet.cell(row, 8, f"=SUM($E$6:E{row})")
        sheet.cell(row, 9, f"=F{row}*{context.tax_rate}")
        row += 1
    last_row = row - 1
    _body_style(
        sheet, 6, last_row, len(headers),
        currency_columns=(5, 6, 7, 8, 9), date_columns=(4,), wrap_columns=(1, 2),
    )
    if last_row >= 6:
        sheet.conditional_formatting.add(
            f"B6:B{last_row}",
            FormulaRule(formula=["LEN(B6)=0"], fill=PatternFill("solid", fgColor=PALE_AMBER)),
        )
    _finish_sheet(
        sheet,
        widths={1: 30, 2: 27, 3: 14, 4: 15, 5: 18, 6: 21, 7: 18, 8: 24, 9: 20},
        header_row=5,
        last_row=last_row,
        last_column=len(headers),
        run_hash=run_hash,
    )
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def build_exclusion_register_workbook(path: str | Path, context: ExportContext) -> Path:
    output = Path(path)
    workbook = Workbook()
    _set_document_properties(workbook, context, "Exclusion Register")
    sheet = workbook.active
    sheet.title = "Exclusion Register"
    run = context.run
    run_hash = run.run_hash()
    _title_block(
        sheet,
        "EXCLUSION REGISTER",
        f"{run.entity_name} | Correctly not disallowed: {run.excluded_total}",
        7,
        run_hash,
    )
    sheet.merge_cells("A3:G3")
    sheet["A3"] = (
        "A failed coverage gate means the delayed-payment disallowance does not apply. "
        "The counterfactual amount is what a tool without these gates would have added back."
    )
    sheet["A3"].font = Font(name="Aptos", size=9, italic=True, color=MUTED)
    sheet["A3"].alignment = Alignment(wrap_text=True)
    grouped: dict[str, list[dict]] = {}
    for item in exclusion_register(run):
        grouped.setdefault(item["gate"], []).append(item)
    row = 5
    for gate in (GATE_CLASS, GATE_ACTIVITY, GATE_REGISTRATION, GATE_TIMING):
        entries = grouped.get(gate, [])
        if not entries:
            continue
        group_start = row
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        group_total = sum(
            (entry["wrongly_disallowable"] for entry in entries), Decimal("0.00")
        )
        sheet.cell(row, 1, f"{GATE_LABELS[gate]}  |  Group total: {group_total}")
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor=INK)
        sheet.cell(row, 1).font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        sheet.cell(row, 1).number_format = INDIAN_CURRENCY
        sheet.row_dimensions[row].height = 25
        row += 1
        headers = [
            "Vendor", "Invoices", "Turnover reviewed", "Correct exclusion",
            "Evidence strength", "Legal reason", "Citation",
        ]
        _header_row(sheet, row, headers)
        header_row = row
        row += 1
        detail_start = row
        for entry in entries:
            values = [
                entry["vendor"], entry["invoices"], entry["turnover"],
                entry["wrongly_disallowable"], entry["evidence_strength"],
                entry["reason"], GATE_CITATIONS[gate],
            ]
            for column, value in enumerate(values, 1):
                sheet.cell(row, column, value)
            row += 1
        _body_style(
            sheet, detail_start, row - 1, 7,
            currency_columns=(3, 4), wrap_columns=(1, 6, 7),
        )
        for current in range(detail_start, row):
            sheet.row_dimensions[current].height = 58
        sheet.row_dimensions.group(detail_start, row - 1, outline_level=1, hidden=False)
        row += 1
    last_row = row - 1
    for column, width in {1: 30, 2: 10, 3: 19, 4: 19, 5: 15, 6: 72, 7: 72}.items():
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.print_title_rows = "1:3"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_margins.left = 0.25
    sheet.page_margins.right = 0.25
    sheet.oddFooter.left.text = f"Run hash: {run_hash}"
    sheet.oddFooter.right.text = "Page &P of &N"
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def build_working_paper_pdf(path: str | Path, context: ExportContext) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from xml.sax.saxutils import escape

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    run = context.run
    run_hash = run.run_hash()
    bank_rate, bank_effective = _bank_rate_for_year_end(run.fy)
    weak_count = _weak_vendor_count(context)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="WPTitle", parent=styles["Title"], fontName="Helvetica-Bold",
        fontSize=18, leading=21, textColor=colors.HexColor(f"#{INK}"),
        alignment=TA_LEFT, spaceAfter=3 * mm,
    ))
    styles.add(ParagraphStyle(
        name="WPSection", parent=styles["Heading2"], fontName="Helvetica-Bold",
        fontSize=10, leading=12, textColor=colors.white,
        backColor=colors.HexColor(f"#{INK}"), leftIndent=4 * mm,
        borderPadding=(3 * mm, 3 * mm, 2.5 * mm, 3 * mm),
        spaceBefore=4 * mm, spaceAfter=3 * mm,
    ))
    styles.add(ParagraphStyle(
        name="WPBody", parent=styles["BodyText"], fontName="Helvetica",
        fontSize=8.7, leading=12, textColor=colors.HexColor(f"#{INK}"),
        spaceAfter=2 * mm,
    ))
    styles.add(ParagraphStyle(
        name="WPSmall", parent=styles["BodyText"], fontName="Helvetica",
        fontSize=7.5, leading=10, textColor=colors.HexColor(f"#{MUTED}"),
    ))
    styles.add(ParagraphStyle(
        name="WPKPI", parent=styles["BodyText"], fontName="Helvetica-Bold",
        fontSize=15, leading=18, alignment=TA_CENTER,
    ))

    def money_text(value: Decimal) -> str:
        amount = Decimal(value).quantize(Decimal("0.01"))
        whole, fraction = f"{abs(amount):.2f}".split(".")
        if len(whole) > 3:
            head, tail = whole[:-3], whole[-3:]
            groups = []
            while len(head) > 2:
                groups.insert(0, head[-2:])
                head = head[:-2]
            if head:
                groups.insert(0, head)
            whole = ",".join(groups + [tail])
        return f"{'-' if amount < 0 else ''}Rs {whole}{'' if fraction == '00' else '.' + fraction}"

    def page_frame(canvas, document):
        canvas.saveState()
        width, height = A4
        canvas.setStrokeColor(colors.HexColor(f"#{LINE}"))
        canvas.line(18 * mm, 16 * mm, width - 18 * mm, 16 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor(f"#{MUTED}"))
        canvas.drawString(
            18 * mm, 10 * mm,
            f"STRICTLY PRIVATE & CONFIDENTIAL | {context.document_status} | Run hash: {run_hash}",
        )
        canvas.drawRightString(
            width - 18 * mm, 10 * mm, f"Page {document.page} | The 45-Day Clock"
        )
        canvas.restoreState()

    document = SimpleDocTemplate(
        str(output), pagesize=A4,
        rightMargin=18 * mm, leftMargin=18 * mm,
        topMargin=15 * mm, bottomMargin=20 * mm,
        title=f"Working Paper - {run.entity_name} - {run.fy}",
        author="The 45-Day Clock",
        subject=f"Run hash {run_hash}",
    )
    story = []
    letterhead = Table(
        [[
            Paragraph(
                f"<b>{escape(context.firm_name)}</b><br/>"
                f"<font size='7'>Chartered Accountants"
                f"{' | FRN: ' + escape(context.firm_frn) if context.firm_frn else ''}</font><br/>"
                f"<font size='6'>{escape(context.firm_address)}</font>",
                styles["WPBody"],
            ),
            Paragraph(
                f"AUDIT WORKING PAPER<br/><font size='7'>MSME payable exposure | "
                f"{escape(context.document_status)}</font>",
                      ParagraphStyle("Right", parent=styles["WPBody"], alignment=TA_RIGHT)),
        ]],
        colWidths=[95 * mm, 65 * mm],
    )
    letterhead.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor(f"#{INK}")),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{BACKGROUND}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5 * mm),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5 * mm),
        ("TOPPADDING", (0, 0), (-1, -1), 4 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4 * mm),
    ]))
    story.extend([letterhead, Spacer(1, 6 * mm)])
    story.append(Paragraph("MSME PAYABLE EXPOSURE", styles["WPTitle"]))
    metadata = [
        ["Client", run.entity_name, "Financial year", run.fy],
        ["PAN", context.entity_pan, "Governing provision", f"{run.statute['act']}, s.{run.statute['section']}"],
        ["Rule pack", run.rule_pack_version, "Run hash", run_hash],
        ["Run date", run.run_at.strftime("%d-%b-%Y %H:%M"), "Prepared for", "Tax audit working papers"],
    ]
    metadata_table = Table(metadata, colWidths=[25 * mm, 55 * mm, 35 * mm, 45 * mm])
    metadata_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{PALE_TEAL}")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor(f"#{PALE_TEAL}")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.6),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor(f"#{INK}")),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{LINE}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.3 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.3 * mm),
    ]))
    story.extend([metadata_table, Paragraph("CONCLUSION", styles["WPSection"])])
    kpi_data = [
        ["DISALLOWANCE", "MSMED s.16 INTEREST", "CORRECTLY NOT DISALLOWED"],
        [money_text(run.disallowance_total), money_text(run.interest_total), money_text(run.excluded_total)],
        [f"Section {run.statute['section']}", "NOT DEDUCTIBLE - s.23", "Four coverage gates"],
    ]
    kpi_table = Table(kpi_data, colWidths=[53.3 * mm] * 3)
    kpi_table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor(f"#{LINE}")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{LINE}")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{PALE_RED}")),
        ("BACKGROUND", (1, 0), (1, -1), colors.HexColor(f"#{PALE_AMBER}")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor(f"#{PALE_GREEN}")),
        ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor(f"#{RED}")),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor(f"#{AMBER}")),
        ("TEXTCOLOR", (2, 1), (2, 1), colors.HexColor(f"#{GREEN}")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("FONTSIZE", (0, 2), (-1, 2), 6.8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.5 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5 * mm),
    ]))
    story.append(kpi_table)
    story.append(Paragraph("BASIS AND ASSUMPTION", styles["WPSection"]))
    story.append(Paragraph(
        f"<b>Acceptance-date policy:</b> {ACC_POLICY_TEXT[run.acceptance_policy]}",
        styles["WPBody"],
    ))
    story.append(Paragraph(
        f"<b>Interest basis:</b> RBI Bank Rate {bank_rate}% effective "
        f"{bank_effective.strftime('%d-%b-%Y')}; MSMED Act s.16 rate is three times "
        "the Bank Rate, compounded with monthly rests. Rate changes are applied by "
        "dated segment in the invoice arithmetic.",
        styles["WPBody"],
    ))
    story.append(Paragraph(
        f"<b>Action List tax-impact rate:</b> {context.tax_rate * 100:.2f}%. "
        f"{escape(context.tax_rate_basis)}. This management estimate ranks cash-saving "
        "priorities only and does not alter the statutory disallowance or interest computation.",
        styles["WPBody"],
    ))
    story.append(Paragraph("CONTROL TOTALS AND EVIDENCE", styles["WPSection"]))
    control_data = [
        ["Source file", run.control_totals.get("source_file_name", "Not recorded") or "Not recorded"],
        ["Source SHA-256", run.control_totals.get("source_file_sha256", "Not recorded") or "Not recorded"],
        ["Source purchase lines", f"{run.control_totals.get('source_purchase_lines', 0):,}"],
        ["Source purchase value", money_text(run.control_totals.get("source_purchase_value", Decimal("0")))],
        ["Lines in selected FY", f"{run.control_totals.get('ledger_lines_in_year', 0):,}"],
        ["Value in selected FY", money_text(run.control_totals.get("ledger_value", Decimal("0")))],
        ["Out-of-period lines", f"{run.control_totals.get('out_of_period_lines', 0):,}"],
        ["Out-of-period value", money_text(run.control_totals.get("out_of_period_value", Decimal("0")))],
        ["Value accounted for in FY", money_text(run.control_totals.get("value_accounted_for", Decimal("0")))],
        ["Control totals tie", "YES" if run.control_totals.get("ties") else "NO - DO NOT SIGN"],
        ["Vendors on weak evidence", str(weak_count)],
    ]
    control_table = Table(control_data, colWidths=[75 * mm, 85 * mm])
    control_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{BACKGROUND}")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{LINE}")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 2.2 * mm),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2 * mm),
    ]))
    story.append(control_table)
    if weak_count:
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(
            f"<b>Attention:</b> {weak_count} vendor(s) rely on weak evidence. These "
            "classifications require resolution before the working paper is signed.",
            ParagraphStyle(
                "Warning", parent=styles["WPBody"], textColor=colors.HexColor(f"#{AMBER}"),
                backColor=colors.HexColor(f"#{PALE_AMBER}"), borderPadding=3 * mm,
            ),
        ))
    story.append(Paragraph("PROFESSIONAL USE AND SIGN-OFF", styles["WPSection"]))
    story.append(Paragraph(
        "Strictly private and confidential. This document is a computation aid, not "
        "professional advice. The engagement "
        "team remains responsible for validating source data, vendor evidence, legal "
        "classification, the acceptance-date policy and the conclusions recorded in "
        "the tax audit file. The document must not be signed or circulated as final "
        "while control totals do not tie or evidence exceptions remain unresolved.",
        styles["WPBody"],
    ))
    signature = Table([
        ["Prepared by", context.preparer or "", "Reviewed by", context.reviewer or ""],
        ["Signature", "", "Signature", ""],
        ["Date", "", "Date", ""],
    ], colWidths=[24 * mm, 56 * mm, 24 * mm, 56 * mm], rowHeights=[10 * mm] * 3)
    signature.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor(f"#{LINE}")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor(f"#{BACKGROUND}")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor(f"#{BACKGROUND}")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(signature)
    document.build(story, onFirstPage=page_frame, onLaterPages=page_frame)
    return output


def export_all(output_folder: str | Path, context: ExportContext) -> ExportBundle:
    """Create all four deliverables in the exact folder chosen by the user."""
    folder = Path(output_folder).expanduser().resolve()
    folder.mkdir(parents=True, exist_ok=True)
    safe_entity = "".join(
        character if character.isalnum() else "-" for character in context.run.entity_name
    ).strip("-")
    while "--" in safe_entity:
        safe_entity = safe_entity.replace("--", "-")
    prefix = f"{safe_entity}-{context.run.fy}"
    bundle = ExportBundle(
        clause_22=folder / f"{prefix}-Clause-22-Workbook.xlsx",
        action_list=folder / f"{prefix}-31-March-Action-List.xlsx",
        exclusion_register=folder / f"{prefix}-Exclusion-Register.xlsx",
        working_paper=folder / f"{prefix}-Working-Paper.pdf",
    )
    build_clause_22_workbook(bundle.clause_22, context)
    build_action_list_workbook(bundle.action_list, context)
    build_exclusion_register_workbook(bundle.exclusion_register, context)
    build_working_paper_pdf(bundle.working_paper, context)
    return bundle


def _audit_pack_workbook(path: Path, context: ExportContext) -> Path:
    workbook = Workbook()
    purchases = workbook.active
    purchases.title = "Reconciled Ledger"
    _set_document_properties(workbook, context, "Reconciled purchase and payment ledger")
    _title_block(
        purchases,
        "RECONCILED LEDGER",
        f"{context.run.entity_name} | FY {context.run.fy} | Run hash {context.run.run_hash()}",
        14,
        context.run.run_hash(),
    )
    headers = [
        "Invoice number", "Invoice date", "Vendor", "Vendor ID / PAN / GSTIN",
        "Udyam number", "Enterprise class", "Evidence source", "GRN date",
        "Agreement credit days", "Invoice amount", "Payments recorded",
        "Outstanding at year end", "Treatment", "Reason",
    ]
    _header_row(purchases, 4, headers)
    payments = _payments_map(context)
    findings = {item.invoice_id: item for item in context.run.findings}
    _, year_end = fy_bounds(context.run.fy)
    for row_number, line in enumerate(context.purchases, 5):
        record = context.udyam.get(line.vendor_id, UdyamRecord(line.vendor_id))
        finding = findings.get(line.invoice_id)
        paid = sum(
            (item.amount for item in payments.get(line.invoice_id, []) if item.payment_date <= year_end),
            Decimal("0.00"),
        )
        outstanding = max(Decimal("0.00"), line.amount - paid)
        values = [
            line.invoice_id, line.invoice_date, line.vendor_name_as_written, line.vendor_id,
            record.udyam_no or "", record.enterprise_class or "", record.source,
            line.grn_date, line.agreement_days, line.amount, paid, outstanding,
            (finding.status if finding else "Not calculated"), (finding.reason if finding else ""),
        ]
        for column, value in enumerate(values, 1):
            purchases.cell(row_number, column, value)
        _body_style(purchases, row_number, row_number, len(headers), currency_columns=(10, 11, 12), date_columns=(2, 8))
    _finish_sheet(
        purchases, widths=dict(enumerate([18, 14, 28, 23, 25, 18, 21, 14, 20, 18, 18, 20, 24, 58], 1)),
        header_row=4, last_row=max(4, purchases.max_row), last_column=len(headers),
        run_hash=context.run.run_hash(),
    )

    payment_sheet = workbook.create_sheet("Payments")
    _title_block(
        payment_sheet, "PAYMENTS RECORDED",
        f"{context.run.entity_name} | FY {context.run.fy} | Run hash {context.run.run_hash()}",
        3, context.run.run_hash(),
    )
    _header_row(payment_sheet, 4, ["Invoice number", "Payment date", "Amount"])
    for row_number, line in enumerate(context.payments, 5):
        payment_sheet.append([line.invoice_id, line.payment_date, line.amount])
        _body_style(payment_sheet, row_number, row_number, 3, currency_columns=(3,), date_columns=(2,))
    _finish_sheet(
        payment_sheet, widths={1: 22, 2: 16, 3: 20}, header_row=4,
        last_row=max(4, payment_sheet.max_row), last_column=3, run_hash=context.run.run_hash(),
    )
    workbook.save(path)
    return path


def _vendor_master_workbook(path: Path, context: ExportContext) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vendor Master"
    _set_document_properties(workbook, context, "Vendor master and MSME evidence status")
    _title_block(
        sheet, "VENDOR MASTER",
        f"{context.run.entity_name} | FY {context.run.fy} | Run hash {context.run.run_hash()}",
        13, context.run.run_hash(),
    )
    headers = [
        "Vendor", "Vendor ID", "PAN", "GSTIN", "Contact", "Udyam number",
        "Enterprise class", "NIC code", "Major activity", "Registration date",
        "Evidence source", "Confirmed by", "Evidence hash",
    ]
    _header_row(sheet, 4, headers)
    names: dict[str, str] = {}
    for line in context.purchases:
        names.setdefault(line.vendor_id, line.vendor_name_as_written)
    for row_number, vendor_id in enumerate(sorted(names, key=lambda item: names[item].casefold()), 5):
        record = context.udyam.get(vendor_id, UdyamRecord(vendor_id))
        metadata = context.vendor_metadata.get(vendor_id, {})
        values = [
            names[vendor_id], vendor_id, metadata.get("pan", ""), metadata.get("gstin", ""),
            metadata.get("contact", ""), record.udyam_no or "", record.enterprise_class or "",
            record.nic_code or "", record.activity_label or "", record.registration_date,
            record.source, record.confirmed_by or "", record.evidence_file_hash or "",
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
        _body_style(sheet, row_number, row_number, len(headers), date_columns=(10,))
    _finish_sheet(
        sheet, widths=dict(enumerate([30, 23, 16, 20, 22, 25, 18, 12, 25, 16, 22, 20, 42], 1)),
        header_row=4, last_row=max(4, sheet.max_row), last_column=len(headers),
        run_hash=context.run.run_hash(),
    )
    workbook.save(path)
    return path


def _evidence_register_workbook(path: Path, context: ExportContext) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence Register"
    _set_document_properties(workbook, context, "Udyam evidence register")
    _title_block(
        sheet, "UDYAM EVIDENCE REGISTER",
        f"{context.run.entity_name} | FY {context.run.fy} | Run hash {context.run.run_hash()}",
        8, context.run.run_hash(),
    )
    headers = ["Vendor ID", "Original filename", "Stored filename", "Media type", "SHA-256", "Added by", "Added at", "Bytes"]
    _header_row(sheet, 4, headers)
    for row_number, item in enumerate(context.evidence_documents, 5):
        values = [
            item.get("vendor_id", ""), item.get("filename", ""), item.get("stored_filename", ""),
            item.get("media_type", ""), item.get("sha256", ""), item.get("added_by", ""),
            item.get("added_at", ""), item.get("bytes", 0),
        ]
        for column, value in enumerate(values, 1):
            sheet.cell(row_number, column, value)
        _body_style(sheet, row_number, row_number, len(headers))
    _finish_sheet(
        sheet, widths=dict(enumerate([23, 30, 30, 24, 70, 20, 22, 12], 1)),
        header_row=4, last_row=max(4, sheet.max_row), last_column=len(headers),
        run_hash=context.run.run_hash(),
    )
    workbook.save(path)
    return path


def export_complete_audit_pack(output_folder: str | Path, context: ExportContext) -> AuditPackBundle:
    """Create a structured, non-overwriting audit pack in the chosen folder."""
    selected = Path(output_folder).expanduser().resolve()
    selected.mkdir(parents=True, exist_ok=True)
    safe_entity = "".join(
        character if character.isalnum() else "-" for character in context.run.entity_name
    ).strip("-") or "Client"
    while "--" in safe_entity:
        safe_entity = safe_entity.replace("--", "-")
    base = selected / f"{safe_entity}-{context.run.fy}-Audit-Pack"
    root = base
    sequence = 2
    while root.exists():
        root = Path(f"{base}-{sequence}")
        sequence += 1
    input_folder = root / "01_Input"
    evidence_folder = root / "02_Evidence"
    calculations_folder = root / "03_Calculations"
    results_folder = root / "04_Results"
    working_folder = root / "05_Working_Papers"
    evidence_documents_folder = evidence_folder / "Documents"
    for folder in (input_folder, evidence_folder, calculations_folder, results_folder, working_folder, evidence_documents_folder):
        folder.mkdir(parents=True, exist_ok=True)

    prefix = f"{safe_entity}-{context.run.fy}"
    created: list[Path] = []
    created.append(_audit_pack_workbook(input_folder / f"{prefix}-Reconciled-Ledger.xlsx", context))
    created.append(_vendor_master_workbook(evidence_folder / f"{prefix}-Vendor-Master.xlsx", context))

    used_names: set[str] = set()
    for item in context.evidence_documents:
        content = item.get("content")
        if not isinstance(content, (bytes, bytearray)):
            continue
        original = Path(str(item.get("filename") or "evidence.bin")).name
        candidate = original
        counter = 2
        while candidate.casefold() in used_names:
            source = Path(original)
            candidate = f"{source.stem}-{counter}{source.suffix}"
            counter += 1
        used_names.add(candidate.casefold())
        target = evidence_documents_folder / candidate
        target.write_bytes(bytes(content))
        created.append(target)
        if isinstance(item, dict):
            item["stored_filename"] = str(Path("Documents") / candidate)
    created.append(_evidence_register_workbook(evidence_folder / f"{prefix}-Udyam-Evidence-Register.xlsx", context))
    created.append(build_clause_22_workbook(calculations_folder / f"{prefix}-Clause-22-Workbook.xlsx", context))
    created.append(build_action_list_workbook(results_folder / f"{prefix}-31-March-Action-List.xlsx", context))
    created.append(build_exclusion_register_workbook(results_folder / f"{prefix}-Exclusion-Register.xlsx", context))
    created.append(build_working_paper_pdf(working_folder / f"{prefix}-Working-Paper.pdf", context))
    return AuditPackBundle(root=root, files=tuple(created))
