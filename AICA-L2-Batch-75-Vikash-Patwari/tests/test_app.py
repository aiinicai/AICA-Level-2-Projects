"""Desktop API integration checks. Runs with pytest OR standalone."""

import os
import sys
import tempfile
from io import BytesIO

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from fastapi.testclient import TestClient  # noqa: E402
from docx import Document  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app import DesktopState, create_application  # noqa: E402
from clock45.store import Store  # noqa: E402


TOKEN = "test-session-token"
results = []


def check(name, got, want):
    ok = got == want
    results.append((ok, name, got, want))
    assert ok, f"{name}: got {got!r}, want {want!r}"


def client():
    application = create_application(DesktopState(), TOKEN)
    return TestClient(application), {"X-Session-Token": TOKEN}


def test_01_security_and_utf8_shell():
    browser, headers = client()
    check("01 API rejects no token", browser.get("/api/home").status_code, 403)
    response = browser.get("/")
    check("01 HTML UTF-8", "charset=utf-8" in response.headers["content-type"].lower(), True)
    check("01 token embedded", TOKEN in response.text, True)
    check("01 all seven screens", sum(f'id="screen-{name}"' in response.text for name in
          ["home", "load", "vendors", "assumptions", "results", "exclusions", "export"]), 7)
    check("01 client-year setup", 'id="analysis-modal"' in response.text, True)
    check("01 firm profile setup", 'id="firm-modal"' in response.text, True)
    check("01 vendor editor", 'id="vendor-modal"' in response.text, True)


def test_02_demo_workflow_and_drilldown():
    browser, headers = client()
    loaded = browser.post("/api/demo/load", headers=headers)
    check("02 demo load", loaded.status_code, 200)
    check("02 controls tie", loaded.json()["control_totals"]["ties"], True)
    check(
        "02 demo source fingerprint",
        len(loaded.json()["control_totals"]["source_file_sha256"]),
        64,
    )
    vendors = browser.get("/api/vendors", headers=headers).json()
    check("02 demo has vendors", len(vendors["vendors"]) > 100, True)
    check("02 no assumed blocker", vendors["assumed_count"], 0)
    vendor_id = vendors["vendors"][0]["vendor_id"]
    declaration = browser.get(f"/api/vendors/{vendor_id}/declaration", headers=headers)
    check("02 declaration is DOCX", declaration.content[:2], b"PK")
    declaration_text = "\n".join(
        paragraph.text for paragraph in Document(BytesIO(declaration.content)).paragraphs
    )
    check("02 declaration is finished", "MSME / UDYAM STATUS DECLARATION" in declaration_text, True)
    assumptions = browser.get("/api/assumptions", headers=headers).json()
    check("02 three policies", len(assumptions["policies"]), 3)
    run = browser.post(
        "/api/assumptions/confirm", headers=headers,
        json={"policy": "INVOICE_DATE", "plus_days": 0},
    )
    check("02 run succeeds", run.status_code, 200)
    output = browser.get("/api/results", headers=headers).json()
    check("02 result section", output["section"], "43B(h)")
    check("02 positive disallowance", float(output["disallowance"]) > 0, True)
    findings = browser.get("/api/findings?kind=disallowance", headers=headers).json()["findings"]
    check("02 drilldown rows", len(findings) > 0, True)
    invoice_id = findings[0]["invoice_id"]
    detail = browser.get(f"/api/findings/{invoice_id}", headers=headers).json()
    check("02 full arithmetic", all(key in detail for key in
          ["acceptance_date", "credit_basis", "due_date", "appointed_day", "payments", "interest_segments"]), True)


def test_03_exclusions_and_exports():
    browser, headers = client()
    browser.post("/api/demo/load", headers=headers)
    browser.post("/api/assumptions/confirm", headers=headers,
                 json={"policy": "GRN_DATE", "plus_days": 0})
    exclusions = browser.get("/api/exclusions", headers=headers).json()
    check("03 exclusion groups", len(exclusions["groups"]) >= 3, True)
    check("03 legal reason present", all(group["reason"] for group in exclusions["groups"]), True)
    preview = browser.get("/api/exports/preview", headers=headers).json()
    check("03 four previews", len(preview["files"]), 4)
    clause = browser.get("/api/exports/clause22", headers=headers)
    check("03 Excel export", clause.content[:2], b"PK")
    working = browser.get("/api/exports/working-paper", headers=headers)
    check("03 PDF export", working.content[:4], b"%PDF")


def test_04_close_and_reopen_saved_analysis():
    with tempfile.TemporaryDirectory() as folder:
        first_store = Store(folder)
        first_state = DesktopState(first_store)
        first_app = create_application(first_state, TOKEN)
        first_browser = TestClient(first_app)
        headers = {"X-Session-Token": TOKEN}
        first_browser.post("/api/demo/load", headers=headers)
        completed = first_browser.post(
            "/api/assumptions/confirm", headers=headers,
            json={"policy": "INVOICE_DATE", "plus_days": 0},
        )
        check("04 first run completes", completed.status_code, 200)
        run_hash = completed.json()["run_hash"]
        first_store.close()

        second_store = Store(folder)
        second_state = DesktopState(second_store)
        second_app = create_application(second_state, TOKEN)
        second_browser = TestClient(second_app)
        home = second_browser.get("/api/home", headers=headers).json()
        check("04 run survives restart", home["runs"][0]["run_hash"], run_hash)
        check("04 workflow resumes at results", home["active_stage"], "results")
        result = second_browser.get("/api/results", headers=headers)
        check("04 saved results available", result.status_code, 200)
        check("04 saved hash reproduces", second_state.run.run_hash(), run_hash)
        second_store.close()


def test_05_scope_block_firm_profile_and_template():
    with tempfile.TemporaryDirectory() as folder:
        store = Store(folder)
        state = DesktopState(store)
        application = create_application(state, TOKEN)
        browser = TestClient(application)
        headers = {"X-Session-Token": TOKEN}
        started = browser.post(
            "/api/analysis/start", headers=headers,
            json={"entity_name": "Synthetic Client Pvt Ltd", "entity_pan": "ABCDE1234F", "fy": "2025-26"},
        )
        check("05 analysis setup", started.status_code, 200)
        csv_data = (
            "Invoice Number,Invoice Date,Vendor Name,Vendor PAN or GSTIN,Amount\n"
            "OUT-001,31-05-2026,Synthetic Vendor,AAAAA1111A,501141.28\n"
        )
        inspected = browser.post(
            "/api/import/file?record_type=purchase", headers=headers,
            files={"file": ("out-of-period.csv", csv_data, "text/csv")},
        )
        check("05 inspect out of period", inspected.status_code, 200)
        mapped = browser.post(
            "/api/import/map", headers=headers,
            json={"record_type": "purchase", "mapping": {
                "invoice_number": "Invoice Number", "invoice_date": "Invoice Date",
                "vendor_name": "Vendor Name", "vendor_pan_or_gstin": "Vendor PAN or GSTIN",
                "amount": "Amount",
            }},
        )
        controls = mapped.json()["control_totals"]
        check("05 import itself ties", controls["ties"], True)
        check("05 no in-year lines", controls["in_period_lines"], 0)
        check("05 out-of-period disclosed", controls["out_of_period_value"], "501141.28")
        check("05 continuation blocked", controls["can_continue"], False)
        attempted = browser.post(
            "/api/assumptions/confirm", headers=headers,
            json={"policy": "INVOICE_DATE", "plus_days": 0},
        )
        check("05 run blocked", attempted.status_code, 409)
        check("05 reason names FY", "No purchase invoice falls within FY 2025-26" in attempted.json()["detail"], True)

        profile = browser.post(
            "/api/firm-profile", headers=headers,
            json={"firm_name": "ABC & Co.", "frn": "012345N", "address": "Synthetic address",
                  "email": "", "phone": "", "preparer": "CA Preparer",
                  "reviewer": "CA Reviewer", "document_status": "DRAFT", "tax_rate_pct": "25.17"},
        )
        check("05 firm profile saved", profile.json()["firm_name"], "ABC & Co.")
        check("05 firm profile persisted", browser.get("/api/firm-profile", headers=headers).json()["frn"], "012345N")
        template = browser.get("/api/import/template", headers=headers)
        check("05 template workbook", template.content[:2], b"PK")
        workbook = load_workbook(BytesIO(template.content), read_only=True)
        headers_in_template = [cell.value for cell in next(workbook["Purchase Invoices"].iter_rows())]
        check("05 template has Udyam and GRN", "Udyam Registration Number" in headers_in_template and "GRN Date" in headers_in_template, True)
        store.close()


def test_06_vendor_classification_editor_and_audit_log():
    with tempfile.TemporaryDirectory() as folder:
        store = Store(folder)
        state = DesktopState(store)
        browser = TestClient(create_application(state, TOKEN))
        headers = {"X-Session-Token": TOKEN}
        browser.post(
            "/api/analysis/start", headers=headers,
            json={"entity_name": "Vendor Editor Test", "entity_pan": "ABCDE1234F", "fy": "2025-26"},
        )
        csv_data = (
            "Invoice Number,Invoice Date,Vendor Name,Vendor PAN or GSTIN,Amount\n"
            "INV-001,01-04-2025,Alpha Engineering,AAAAA1111A,100000\n"
        )
        browser.post(
            "/api/import/file?record_type=purchase", headers=headers,
            files={"file": ("vendor-edit.csv", csv_data, "text/csv")},
        )
        browser.post(
            "/api/import/map", headers=headers,
            json={"record_type": "purchase", "mapping": {
                "invoice_number": "Invoice Number", "invoice_date": "Invoice Date",
                "vendor_name": "Vendor Name", "vendor_pan_or_gstin": "Vendor PAN or GSTIN",
                "amount": "Amount",
            }},
        )
        invalid = browser.put(
            "/api/vendors/AAAAA1111A", headers=headers,
            json={"udyam_no": "WRONG", "enterprise_class": "SMALL", "nic_code": "25999",
                  "activity_label": "Manufacturing", "registration_date": "2024-01-01",
                  "source": "VENDOR_DECLARATION", "confirmed_by": "CA Reviewer"},
        )
        check("06 malformed Udyam rejected", invalid.status_code, 422)
        updated = browser.put(
            "/api/vendors/AAAAA1111A", headers=headers,
            json={"udyam_no": "UDYAM-MH-26-0123456", "enterprise_class": "SMALL",
                  "nic_code": "25999", "activity_label": "Manufacturing",
                  "registration_date": "2024-01-01", "source": "VENDOR_DECLARATION",
                  "confirmed_by": "CA Reviewer"},
        )
        check("06 vendor update accepted", updated.status_code, 200)
        check("06 Udyam saved", updated.json()["udyam_no"], "UDYAM-MH-26-0123456")
        check("06 vendor now covered", updated.json()["coverage"], "Covered")
        check("06 sign-off blocker cleared", browser.get("/api/vendors", headers=headers).json()["assumed_count"], 0)
        audit = store.vendor_audit_log(state.client_id, "AAAAA1111A")
        check("06 classification audit recorded", len(audit) >= 1, True)
        check("06 reviewer attributed", audit[-1]["changed_by"], "Vendor classification confirmed by CA Reviewer")
        store.close()


def test_07_payment_completeness_must_be_confirmed():
    state = DesktopState()
    browser = TestClient(create_application(state, TOKEN))
    headers = {"X-Session-Token": TOKEN}
    browser.post("/api/demo/load", headers=headers)
    state.payment_information_confirmed = False
    blocked = browser.post(
        "/api/assumptions/confirm", headers=headers,
        json={"policy": "INVOICE_DATE", "plus_days": 0, "payments_confirmed": False},
    )
    check("07 unconfirmed payments blocked", blocked.status_code, 409)
    check(
        "07 payment reason actionable",
        "payment information is complete" in blocked.json()["detail"],
        True,
    )
    confirmed = browser.post(
        "/api/assumptions/confirm", headers=headers,
        json={"policy": "INVOICE_DATE", "plus_days": 0, "payments_confirmed": True},
    )
    check("07 explicit payment confirmation runs", confirmed.status_code, 200)


if __name__ == "__main__":
    functions = [value for key, value in sorted(globals().items()) if key.startswith("test_")]
    failed = 0
    for function in functions:
        try:
            function()
        except AssertionError as exc:
            failed += 1
            print(f"  FAIL  {function.__name__}: {exc}")
        except Exception as exc:
            failed += 1
            print(f"  ERROR {function.__name__}: {type(exc).__name__}: {exc}")
    for ok, name, got, want in results:
        print(f"  {'PASS' if ok else 'FAIL'}  {name:<34} {got}")
    print(
        f"\n{len(results)} assertions across {len(functions)} cases · "
        f"{'ALL PASSED' if failed == 0 else f'{failed} FAILED'}"
    )
    sys.exit(1 if failed else 0)
