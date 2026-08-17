"""Finished export deliverables. Runs with pytest OR standalone."""

import os
import sys
import tempfile
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook  # noqa: E402
from pypdf import PdfReader  # noqa: E402

from clock45.demo_data import build_demo_dataset  # noqa: E402
from clock45.engine import ACC_INVOICE_DATE, run_assessment  # noqa: E402
from clock45.export import ExportContext, export_all, export_complete_audit_pack  # noqa: E402


results = []


def check(name, got, want):
    ok = got == want
    results.append((ok, name, got, want))
    assert ok, f"{name}: got {got!r}, want {want!r}"


def build_exports(folder: Path):
    demo = build_demo_dataset()
    run = run_assessment(
        entity_name=demo["entity_name"],
        fy=demo["fy"],
        operator="CA Test Preparer",
        purchases=demo["purchases"],
        payments=demo["payments"],
        udyam=demo["udyam"],
        acceptance_policy=ACC_INVOICE_DATE,
    )
    context = ExportContext(
        run=run,
        purchases=demo["purchases"],
        payments=demo["payments"],
        udyam=demo["udyam"],
        entity_pan="AAECS1234F",
        preparer="CA Test Preparer",
        reviewer="CA Test Reviewer",
        firm_name="ABC & Co.",
        firm_frn="012345N",
        firm_address="Test address",
        document_status="DRAFT",
    )
    return context, export_all(folder, context)


def test_01_all_files_and_clause_workbook():
    with tempfile.TemporaryDirectory() as temporary:
        context, bundle = build_exports(Path(temporary))
        check("01 four outputs", len(bundle.paths()), 4)
        check("01 all files exist", all(path.exists() and path.stat().st_size > 1000 for path in bundle.paths()), True)
        check("01 chosen folder only", {path.parent for path in bundle.paths()}, {Path(temporary).resolve()})
        workbook = load_workbook(bundle.clause_22, data_only=False)
        check("01 required sheets", workbook.sheetnames,
              ["Summary", "Invoice detail", "Exclusions", "Interest only"])
        check("01 PAN", workbook["Summary"]["B5"].value, "AAECS1234F")
        check("01 run hash summary", workbook["Summary"]["B9"].value, context.run.run_hash())
        summary = workbook["Summary"]
        tie_row = next(
            row for row in range(1, summary.max_row + 1)
            if summary.cell(row, 1).value == "Totals tie"
        )
        check("01 control ties", summary.cell(tie_row, 2).value, "YES")
        check("01 firm identity", "ABC & Co." in [cell.value for cell in summary["E"]], True)
        check("01 source controls", any(
            summary.cell(row, 1).value == "Out-of-period lines"
            for row in range(1, summary.max_row + 1)
        ), True)
        disallowed_count = sum(finding.disallowance > 0 for finding in context.run.findings)
        check("01 disallowed detail", workbook["Invoice detail"].max_row - 4, disallowed_count)
        check("01 detail frozen", workbook["Invoice detail"].freeze_panes, "A5")
        check("01 detail filter", workbook["Invoice detail"].auto_filter.ref.startswith("A4:"), True)
        check("01 currency format", "#,##,##0" in workbook["Invoice detail"]["J5"].number_format, True)
        check("01 hash in every footer", all(context.run.run_hash() in sheet.oddFooter.left.text for sheet in workbook), True)


def test_02_action_and_exclusion_workbooks():
    with tempfile.TemporaryDirectory() as temporary:
        context, bundle = build_exports(Path(temporary))
        action = load_workbook(bundle.action_list, data_only=False)["31 March Action List"]
        values = [action.cell(row, 6).value for row in range(6, action.max_row + 1)]
        check("02 descending priority", values, sorted(values, reverse=True))
        check("02 cumulative formula", action["H6"].value, "=SUM($E$6:E6)")
        check("02 action hash", context.run.run_hash() in action["A2"].value, True)
        register = load_workbook(bundle.exclusion_register, data_only=False)["Exclusion Register"]
        all_text = " | ".join(str(cell.value or "") for row in register.iter_rows() for cell in row)
        check("02 grouped gates", "Gate 1 - Enterprise class" in all_text and "Gate 2 - Trader activity" in all_text, True)
        check("02 citations", "indiacode.nic.in" in all_text and "msme.gov.in" in all_text, True)
        check("02 exclusion hash", context.run.run_hash() in register["A2"].value, True)


def test_03_working_paper_content():
    with tempfile.TemporaryDirectory() as temporary:
        context, bundle = build_exports(Path(temporary))
        reader = PdfReader(bundle.working_paper)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        check("03 PDF pages", len(reader.pages) >= 1, True)
        check("03 client and year", context.run.entity_name in text and context.run.fy in text, True)
        check("03 assumption in full", "Date of acceptance has been taken as the invoice date" in text, True)
        check("03 RBI effective date", "05-Dec-2025" in text, True)
        check("03 rule pack and hash", context.run.rule_pack_version in text and context.run.run_hash() in text, True)
        check("03 professional caveat", "computation aid, not professional advice" in text, True)
        check("03 signature blocks", "Prepared by" in text and "Reviewed by" in text, True)
        check("03 firm and status", "ABC & Co." in text and "DRAFT" in text, True)
        check("03 source scope", "Out-of-period lines" in text, True)


def test_04_complete_audit_pack_is_structured_and_non_overwriting():
    with tempfile.TemporaryDirectory() as temporary:
        context, _ = build_exports(Path(temporary))
        first = export_complete_audit_pack(temporary, context)
        second = export_complete_audit_pack(temporary, context)
        check("04 pack roots differ", first.root != second.root, True)
        check("04 seven pack files", len(first.paths()), 7)
        check("04 every pack file valid", all(path.is_file() and path.stat().st_size > 500 for path in first.paths()), True)
        relative = {path.relative_to(first.root).parts[0] for path in first.paths()}
        check(
            "04 structured folders",
            relative,
            {"01_Input", "02_Evidence", "03_Calculations", "04_Results", "05_Working_Papers"},
        )
        ledger = load_workbook(next(path for path in first.paths() if "Reconciled-Ledger" in path.name))
        check("04 input hash carried", context.run.run_hash() in ledger["Reconciled Ledger"]["A2"].value, True)
        action_text = "\n".join(
            str(cell.value or "")
            for row in load_workbook(
                next(path for path in first.paths() if "Action-List" in path.name), data_only=False
            )["31 March Action List"].iter_rows()
            for cell in row
        )
        check("04 tax rate disclosed", "25.17%" in action_text, True)


if __name__ == "__main__":
    functions = [value for key, value in sorted(globals().items()) if key.startswith("test_")]
    failed = 0
    for function in functions:
        try:
            function()
        except AssertionError as exc:
            failed += 1
            print(f"  FAIL  {function.__name__}: {exc}")
        except Exception as exc:
            failed += 1
            print(f"  ERROR {function.__name__}: {type(exc).__name__}: {exc}")
    for ok, name, got, want in results:
        print(f"  {'PASS' if ok else 'FAIL'}  {name:<34} {got}")
    print(
        f"\n{len(results)} assertions across {len(functions)} cases · "
        f"{'ALL PASSED' if failed == 0 else f'{failed} FAILED'}"
    )
    sys.exit(1 if failed else 0)
