from http.server import ThreadingHTTPServer, BaseHTTPRequestHandler
from pathlib import Path
from urllib.parse import urlparse, parse_qs
import json, socket, webbrowser, io, traceback, re
from datetime import datetime, date
from openpyxl import Workbook, load_workbook

# These globals are deliberately overridable by the client wrapper.
BASE = Path(__file__).resolve().parent
DASHBOARD = BASE / "dashboard"
DATA = DASHBOARD / "data"
OUTPUT = BASE / "Tally_Output"
ASSETS = DASHBOARD / "assets"
HTML = DASHBOARD / "Tally_Financial_Intelligence_Dashboard.html"

for p in (DASHBOARD, DATA, OUTPUT, ASSETS):
    p.mkdir(parents=True, exist_ok=True)


def free_port(start=8765):
    for p in range(start, start + 100):
        with socket.socket() as s:
            s.settimeout(0.15)
            if s.connect_ex(("127.0.0.1", p)) != 0:
                return p
    raise RuntimeError("No free local port available")


def _json_value(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        try:
            if v != v:
                return None
        except Exception:
            pass
        return v
    return str(v)


def read_pnl_workbook(path: Path):
    """Read the formatted Profit & Loss workbook produced by the extractor.

    The P&L workbook is a presentation workbook, not a flat table: title and
    control rows occupy rows 1-6, the actual headers are on row 7, and income
    and expenses are presented side-by-side. Convert that layout into the
    normalized records expected by the HTML dashboard.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        records = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            vals = list(row) + [None] * max(0, 8 - len(row))
            # Stop at the presentation totals / result rows. Main/Sub detail
            # rows before these are the source for the dashboard P&L.
            left_name = str(vals[0] or '').strip()
            right_name = str(vals[4] or '').strip()
            stop_names = {'TOTAL INCOME', 'TOTAL EXPENSES', 'NET PROFIT / (LOSS)', 'NET PROFIT / LOSS'}
            if left_name.upper() in stop_names or right_name.upper() in stop_names:
                break

            if left_name:
                level = str(vals[1] or 'Sub').strip()
                debit = float(vals[2] or 0)
                credit = float(vals[3] or 0)
                amount = credit - debit
                if abs(amount) > 1e-9:
                    records.append({
                        'Section': 'Income',
                        'Particulars': left_name,
                        'Amount': amount,
                        'Debit': debit,
                        'Credit': credit,
                        'Level': level,
                    })

            if right_name:
                level = str(vals[5] or 'Sub').strip()
                debit = float(vals[6] or 0)
                credit = float(vals[7] or 0)
                # Expenses are represented as negative signed P&L amounts.
                amount = credit - debit
                if abs(amount) > 1e-9:
                    records.append({
                        'Section': 'Expenses',
                        'Particulars': right_name,
                        'Amount': amount,
                        'Debit': debit,
                        'Credit': credit,
                        'Level': level,
                    })
        return records
    finally:
        wb.close()


def read_first_sheet(path: Path):
    """Read an exported Tally workbook without depending on pandas engines."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows))
        except StopIteration:
            return []
        headers = [str(h).strip() if h is not None else "" for h in headers]
        # Remove completely blank trailing headers.
        while headers and headers[-1] == "":
            headers.pop()
        records = []
        for vals in rows:
            vals = list(vals)[:len(headers)]
            if len(vals) < len(headers):
                vals += [None] * (len(headers) - len(vals))
            if not any(v not in (None, "") for v in vals):
                continue
            records.append({headers[i]: _json_value(vals[i]) for i in range(len(headers))})
        return records
    finally:
        wb.close()


def _find_latest(pattern):
    """Find latest workbook, including if the extractor created a subfolder."""
    if not OUTPUT.exists():
        return None
    matches = list(OUTPUT.glob(pattern)) + list(OUTPUT.rglob(pattern))
    matches = [p for p in matches if p.is_file()]
    return max(matches, key=lambda p: p.stat().st_mtime) if matches else None


def _norm_header(x):
    return re.sub(r'[^a-z0-9]+', '', str(x or '').strip().lower())


def _find_header(headers, aliases):
    normed = {_norm_header(h): h for h in headers}
    for alias in aliases:
        a = _norm_header(alias)
        if a in normed:
            return normed[a]
    for h in headers:
        nh = _norm_header(h)
        for alias in aliases:
            if _norm_header(alias) in nh or nh in _norm_header(alias):
                return h
    return ''


def read_table_auto(path: Path):
    """Read a Tally Excel sheet whose header may not be row 1."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        raw = list(ws.iter_rows(values_only=True))
        if not raw:
            return [], []
        best_idx, best_score = 0, -1
        for i, vals in enumerate(raw[:25]):
            headers = [str(v).strip() if v is not None else '' for v in vals]
            score = 0
            joined = ' '.join(_norm_header(h) for h in headers)
            for aliases in [
                ['date','voucher date'], ['voucher type','type'],
                ['ledger/party name','ledger name','ledger','particulars'],
                ['debit','dr'], ['credit','cr']
            ]:
                if any(_norm_header(a) in joined for a in aliases):
                    score += 1
            if score > best_score:
                best_idx, best_score = i, score
        headers = [str(v).strip() if v is not None else '' for v in raw[best_idx]]
        while headers and not headers[-1]:
            headers.pop()
        records=[]
        for vals in raw[best_idx+1:]:
            vals=list(vals)[:len(headers)]
            if len(vals)<len(headers): vals += [None]*(len(headers)-len(vals))
            if not any(v not in (None,'') for v in vals): continue
            records.append({headers[i]: _json_value(vals[i]) for i in range(len(headers))})
        return headers, records
    finally:
        wb.close()


def ensure_data_from_selected_file(path: Path):
    """Build dashboard data from a user-selected Tally workbook.

    Voucher-wise/Day Book files are used directly. Trial Balance/ledger files
    are normalized into the dashboard's common ledger/debit/credit structure.
    The source workbook is never changed.
    """
    path = Path(path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"Selected Excel file not found: {path}")
    DATA.mkdir(parents=True, exist_ok=True)
    headers, raw = read_table_auto(path)
    if not raw:
        raise RuntimeError(f"Selected workbook contains no data rows: {path.name}")

    date_k = _find_header(headers, ['Date','Voucher Date','Transaction Date'])
    vtype_k = _find_header(headers, ['Voucher Type','VoucherType','Type'])
    vno_k = _find_header(headers, ['Voucher Number','Voucher No','Voucher No.','VoucherNumber'])
    ledger_k = _find_header(headers, ['Ledger/Party Name','Ledger / Party Name','Ledger/Party','Ledger Name','Party Name','Particulars'])
    debit_k = _find_header(headers, ['Debit','Dr','Debit Amount','Debit Balance'])
    credit_k = _find_header(headers, ['Credit','Cr','Credit Amount','Credit Balance'])
    particulars_k = _find_header(headers, ['Particulars','Particular','Narration','Description'])
    narration_k = _find_header(headers, ['Narration','Remarks','Description'])

    transaction_like = bool(date_k and vtype_k and ledger_k and debit_k and credit_k)
    source_type = 'Voucher Wise / Day Book' if transaction_like else 'Trial Balance / Ledger'
    normalized=[]
    if transaction_like:
        normalized = raw
    else:
        if not (ledger_k and debit_k and credit_k):
            raise RuntimeError(
                "The selected Excel file is not recognized. It should contain "
                "either Date + Voucher Type + Ledger + Debit + Credit, or Ledger + Debit + Credit columns."
            )
        for i,r in enumerate(raw, start=1):
            normalized.append({
                'Date': _json_value(r.get(date_k)) if date_k else '',
                'Voucher Type': str(r.get(vtype_k) or 'Trial Balance') if vtype_k else 'Trial Balance',
                'Voucher Number': str(r.get(vno_k) or i) if vno_k else str(i),
                'Ledger/Party Name': r.get(ledger_k),
                'Particulars': r.get(particulars_k) if particulars_k else r.get(ledger_k),
                'Debit': r.get(debit_k),
                'Credit': r.get(credit_k),
                'Narration': r.get(narration_k) if narration_k else f'Selected file: {path.name}',
            })

    (DATA / 'latest_daybook.json').write_text(json.dumps(normalized, ensure_ascii=False), encoding='utf-8')

    # If a native P&L workbook exists in Tally_Output, preserve the proper
    # Tally P&L even in selected-file mode. Otherwise leave it empty rather
    # than inventing a P&L from a Trial Balance.
    pnl_file = _find_latest('Profit_Loss_*.xlsx')
    pnl_rows=[]
    if pnl_file:
        pnl_rows = read_pnl_workbook(pnl_file)
        (DATA / 'latest_pnl.json').write_text(json.dumps(pnl_rows, ensure_ascii=False), encoding='utf-8')
    else:
        (DATA / 'latest_pnl.json').write_text('[]', encoding='utf-8')

    diagnostics={
        'checked_at': datetime.now().isoformat(timespec='seconds'),
        'source_mode': 'selected_file',
        'source_file': str(path),
        'source_type': source_type,
        'daybook_rows': len(normalized),
        'pnl_source': str(pnl_file) if pnl_file else '',
        'pnl_rows': len(pnl_rows),
        'ok': bool(normalized),
        'note': 'Selected workbook is read-only. Native Tally P&L is used when available.'
    }
    (DATA/'meta.json').write_text(json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding='utf-8')
    (DATA/'dashboard_data_status.txt').write_text('Selected-file dashboard bridge OK\n'+json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding='utf-8')
    return diagnostics


def source_files():
    return _find_latest("Day_Book_*.xlsx"), _find_latest("Profit_Loss_*.xlsx")


def ensure_data_from_excel():
    """Always rebuild the dashboard bridge from the latest Tally Excel files."""
    DATA.mkdir(parents=True, exist_ok=True)
    day_file, pnl_file = source_files()
    diagnostics = {
        "checked_at": datetime.now().isoformat(timespec="seconds"),
        "output_dir": str(OUTPUT),
        "output_exists": OUTPUT.exists(),
        "daybook_source": str(day_file) if day_file else "",
        "pnl_source": str(pnl_file) if pnl_file else "",
    }

    errors = []
    if day_file:
        try:
            day_rows = read_first_sheet(day_file)
            (DATA / "latest_daybook.json").write_text(
                json.dumps(day_rows, ensure_ascii=False), encoding="utf-8"
            )
            diagnostics["daybook_rows"] = len(day_rows)
        except Exception as exc:
            diagnostics["daybook_error"] = repr(exc)
            errors.append(f"Day Book: {exc}")
    else:
        diagnostics["daybook_rows"] = 0
        errors.append("No Day_Book_*.xlsx found in Tally_Output")

    if pnl_file:
        try:
            pnl_rows = read_pnl_workbook(pnl_file)
            (DATA / "latest_pnl.json").write_text(
                json.dumps(pnl_rows, ensure_ascii=False), encoding="utf-8"
            )
            diagnostics["pnl_rows"] = len(pnl_rows)
        except Exception as exc:
            diagnostics["pnl_error"] = repr(exc)
            errors.append(f"P&L: {exc}")
    else:
        diagnostics["pnl_rows"] = 0
        errors.append("No Profit_Loss_*.xlsx found in Tally_Output")

    diagnostics["ok"] = not errors and diagnostics.get("daybook_rows", 0) > 0
    diagnostics["errors"] = errors
    (DATA / "meta.json").write_text(
        json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (DATA / "dashboard_data_status.txt").write_text(
        ("Dashboard bridge OK\n" if diagnostics["ok"] else "Dashboard bridge FAILED\n")
        + json.dumps(diagnostics, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if errors:
        raise RuntimeError("; ".join(errors))
    return diagnostics


def load(name):
    p = DATA / name
    if not p.exists():
        ensure_data_from_excel()
    if not p.exists():
        return []
    return json.loads(p.read_text(encoding="utf-8"))


def amt(v):
    try:
        if v is None:
            return 0.0
        s = str(v).replace(",", "").replace("₹", "").replace("Rs.", "").strip()
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        return float(s or 0)
    except Exception:
        return 0.0


def cashbank(x):
    x = str(x or "").lower().strip()
    if not x:
        return False
    if any(t in x for t in ["bank charges", "bank interest", "interest", "loan", "borrowing", "finance cost", "processing fee"]):
        return False
    return any(t in x for t in ["cash", "bank", "od account", "overdraft", "cc account", "current account", "saving account", "savings account"])


def filtered(rows, q):
    qs = parse_qs(q)
    frm = qs.get("from", [""])[0]
    to = qs.get("to", [""])[0]
    vt = qs.get("vtype", [""])[0]
    text = qs.get("q", [""])[0].lower()
    out = []
    for r in rows:
        d = str(r.get("Date", ""))[:10]
        if frm and d < frm:
            continue
        if to and d > to:
            continue
        if vt and str(r.get("Voucher Type", "")) != vt:
            continue
        if text and text not in " ".join(str(r.get(k, "")) for k in ["Ledger/Party Name", "Particulars", "Narration"]).lower():
            continue
        out.append(r)
    return out


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        pass

    def send(self, b, ctype, status=200, filename=None):
        self.send_response(status)
        self.send_header("Content-Type", ctype)
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate")
        self.send_header("Pragma", "no-cache")
        self.send_header("Content-Length", str(len(b)))
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(b)

    def json_response(self, obj, status=200):
        self.send(json.dumps(obj, ensure_ascii=False).encode("utf-8"), "application/json; charset=utf-8", status)

    def do_GET(self):
        p = urlparse(self.path)
        try:
            # Health/status endpoint: useful for troubleshooting and prevents
            # the HTML from showing a generic "data unavailable" message.
            if p.path == "/api/status.json":
                try:
                    d = ensure_data_from_excel()
                    self.json_response(d)
                except Exception as exc:
                    self.json_response({"ok": False, "error": str(exc), "output_dir": str(OUTPUT)}, 500)
                return

            if p.path == "/":
                if not HTML.exists():
                    self.send(b"Dashboard HTML not found", "text/plain; charset=utf-8", 500)
                    return
                self.send(HTML.read_bytes(), "text/html; charset=utf-8")
                return

            if p.path == "/api/daybook.json":
                try:
                    rows = load("latest_daybook.json")
                    if not rows:
                        raise RuntimeError("Day Book workbook was found but contains no data rows.")
                    self.json_response(rows)
                except Exception as exc:
                    self.json_response({"error": str(exc), "output_dir": str(OUTPUT)}, 500)
                return

            if p.path == "/api/pnl.json":
                try:
                    rows = load("latest_pnl.json")
                    self.json_response(rows)
                except Exception as exc:
                    self.json_response({"error": str(exc), "output_dir": str(OUTPUT)}, 500)
                return

            if p.path == "/api/meta.json":
                f = DATA / "meta.json"
                if not f.exists():
                    try:
                        ensure_data_from_excel()
                    except Exception:
                        pass
                self.send(f.read_bytes() if f.exists() else b"{}", "application/json; charset=utf-8")
                return

            if p.path == "/api/export.xlsx":
                rows = filtered(load("latest_daybook.json"), p.query)
                pnl = load("latest_pnl.json")
                wb = Workbook()
                ws = wb.active
                ws.title = "Executive KPI"
                ws.append(["TALLY FINANCIAL INTELLIGENCE - MIS"])
                ws.append(["Transactions", len(rows)])
                ws.append(["Total Debit", sum(amt(r.get("Debit")) for r in rows)])
                ws.append(["Total Credit", sum(amt(r.get("Credit")) for r in rows)])
                ws.append(["Cash/Bank Inflow", sum(amt(r.get("Credit")) for r in rows if cashbank(r.get("Ledger/Party Name")))])
                ws.append(["Cash/Bank Outflow", sum(amt(r.get("Debit")) for r in rows if cashbank(r.get("Ledger/Party Name")))])
                w = wb.create_sheet("Filtered Day Book")
                if rows:
                    hs = list(rows[0]); w.append(hs)
                    for r in rows:
                        w.append([r.get(h) for h in hs])
                w = wb.create_sheet("Tally P&L")
                if pnl:
                    hs = list(pnl[0]); w.append(hs)
                    for r in pnl:
                        w.append([r.get(h) for h in hs])
                buf = io.BytesIO(); wb.save(buf)
                self.send(buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="Tally_Financial_Intelligence_MIS.xlsx")
                return

            if p.path.startswith("/assets/"):
                rel = p.path[len("/assets/"):]
                f = (ASSETS / rel).resolve()
                if ASSETS.resolve() not in f.parents or not f.exists():
                    self.send(b"Asset not found", "text/plain", 404)
                    return
                ctype = "application/javascript; charset=utf-8" if f.suffix == ".js" else "application/octet-stream"
                self.send(f.read_bytes(), ctype)
                return

            self.send(b"404 - Not Found", "text/plain; charset=utf-8", 404)
        except Exception as exc:
            self.json_response({"error": str(exc), "trace": traceback.format_exc()}, 500)
