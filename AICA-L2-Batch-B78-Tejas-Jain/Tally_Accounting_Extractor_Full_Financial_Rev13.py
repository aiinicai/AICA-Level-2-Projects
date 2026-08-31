# -*- coding: utf-8 -*-
"""
TallyPrime Accounting Extractor - Streamlit Edition
====================================================

Single-file application.

Run by double-clicking this .py file:
    - The file launches Streamlit automatically.
    - Your browser opens the attractive GUI.
    - Enter From Date, To Date, DSN, host/port and optional company.
    - Click "Extract Accounting Data".

Or run manually:
    streamlit run Tally_Accounting_Extractor.py

Requirements:
    py -m pip install pandas openpyxl pyodbc streamlit

Tally connection:
    - TallyPrime must be running.
    - HTTP server must be enabled, normally port 9000.
    - ODBC server may be enabled; the configured DSN is tested.
    - At least one company must be loaded.

Important design:
    Tally XML/HTTP is used for accounting reports and voucher hierarchy.
    pyodbc is used to verify the configured Tally ODBC DSN and to inspect
    exposed schema information. This avoids inventing Tally SQL table names.

Reports created directly in Tally_Output:
    Day_Book_YYYYMMDD_YYYYMMDD.xlsx
    Voucher_Wise_YYYYMMDD_YYYYMMDD.xlsx
    Ledger_Wise_YYYYMMDD_YYYYMMDD.xlsx
    Trial_Balance_YYYYMMDD_YYYYMMDD.xlsx
    Ledger_Summary_YYYYMMDD_YYYYMMDD.xlsx
    Profit_Loss_YYYYMMDD_YYYYMMDD.xlsx
    Balance_Sheet_YYYYMMDD_YYYYMMDD.xlsx
    Parameters_YYYYMMDD_YYYYMMDD.xlsx

Profit & Loss and Balance Sheet are fetched from TallyPrime's built-in
financial reports over XML/HTTP rather than reconstructed from vouchers.
The Profit & Loss A/c ledger is explicitly excluded from Trial Balance.

Ledger Wise is deliberately GROUP-WISE, not one worksheet per ledger.
This keeps large workbooks substantially smaller and less likely to become
corrupted.

Trial Balance:
    Opening Balance, Debit, Credit and Closing Balance are taken directly
    from the verified Tally ODBC Ledger query:

        SELECT $Name, $Parent, $_PrimaryGroup,
               $OpeningBalance, $DebitTotals,
               $CreditTotals, $_ClosingBalance
        FROM Ledger

    XML voucher lines remain the source for detailed Day Book, Voucher Wise
    and Ledger Wise transactions. The ODBC Ledger values are NOT replaced
    by calculations from XML. A control section exposes reconciliation
    information without altering accounting values.
"""

from __future__ import annotations

import hashlib
import html
import logging
import os
import re
import subprocess
import sys
import time
import traceback
import urllib.error
import urllib.request
import webbrowser
import xml.etree.ElementTree as ET

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import pyodbc

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


# ================================================================
# PATHS / CONFIG
# ================================================================

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "Tally_Output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

LOG_FILE = OUTPUT_DIR / "tally_extractor.log"
STREAMLIT_LOG_FILE = OUTPUT_DIR / "streamlit_startup.log"

DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 9000
DEFAULT_DSN = "TallyODBC64_9000"

HTTP_TIMEOUT = 180
ODBC_TIMEOUT = 15
MAX_ROWS_PER_SHEET = 450_000
EPSILON = 0.005


logging.basicConfig(
    filename=str(LOG_FILE),
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
LOG = logging.getLogger("tally_extractor")


# ================================================================
# GENERAL HELPERS
# ================================================================

def clean(value: Any) -> str:
    """
    Normalize Tally text for both pandas and Excel.

    Tally can return control characters in XML/ODBC text fields, for example
    the XML entity &#4; in GST classification text. openpyxl correctly rejects
    those characters with IllegalCharacterError. Remove all characters that
    Excel worksheet cells cannot contain.
    """
    if value is None:
        return ""

    text = str(value)

    # Remove NUL and other C0 control characters that Excel/openpyxl rejects.
    # Keep tab/newline/carriage-return temporarily so whitespace normalization
    # can turn them into spaces.
    text = "".join(
        ch
        for ch in text
        if (
            ch in "\t\n\r"
            or ord(ch) >= 0x20
        )
    )

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text.strip()


def xml_escape(value: Any) -> str:
    return html.escape(str(value), quote=True)


def excel_safe_value(value: Any) -> Any:
    """
    Final defensive sanitization immediately before writing to openpyxl.
    This protects against control characters that entered a DataFrame through
    a source other than the normal clean() path.
    """
    if isinstance(value, str):
        return clean(value)
    return value


def parse_amount(value: Any) -> float:
    s = clean(value)
    if not s:
        return 0.0
    s = (
        s.replace(",", "")
        .replace("₹", "")
        .replace("$", "")
        .replace("Rs.", "")
        .strip()
    )
    neg = s.startswith("(") and s.endswith(")")
    if neg:
        s = s[1:-1]
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return 0.0
    n = float(m.group(0))
    return -abs(n) if neg else n


def parse_tally_date(value: Any) -> Optional[date]:
    s = clean(value)
    if not s:
        return None

    for fmt in (
        "%Y%m%d",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    return None


def tally_xml_date(d: date) -> str:
    return d.strftime("%d-%b-%Y")


def local_name(tag: str) -> str:
    if "}" in tag:
        tag = tag.rsplit("}", 1)[1]
    if ":" in tag:
        tag = tag.rsplit(":", 1)[1]
    return tag


def unique_join(values: Iterable[str]) -> str:
    seen = set()
    result = []
    for value in values:
        value = clean(value)
        if not value:
            continue
        key = value.casefold()
        if key not in seen:
            seen.add(key)
            result.append(value)
    return " ; ".join(result)


def bool_tally(value: Any) -> bool:
    return clean(value).casefold() in {
        "yes", "y", "true", "1"
    }


def sanitize_xml(data: bytes) -> bytes:
    text = data.decode("utf-8", errors="replace")

    def replace_ref(match):
        token = match.group(0)
        try:
            if token.lower().startswith("&#x"):
                number = int(token[3:-1], 16)
            else:
                number = int(token[2:-1])
        except Exception:
            return ""

        valid = (
            number in (9, 10, 13)
            or 0x20 <= number <= 0xD7FF
            or 0xE000 <= number <= 0xFFFD
            or 0x10000 <= number <= 0x10FFFF
        )
        return token if valid else ""

    text = re.sub(
        r"&#(?:x[0-9A-Fa-f]+|[0-9]+);",
        replace_ref,
        text,
    )

    text = "".join(
        ch for ch in text
        if ch in "\t\n\r" or ord(ch) >= 0x20
    )

    return text.encode("utf-8")


def first_direct_text(element: ET.Element, name: str) -> str:
    wanted = name.upper()
    for child in element:
        if local_name(child.tag).upper() == wanted:
            return clean(child.text)
    return ""


def first_deep_text(element: ET.Element, name: str) -> str:
    wanted = name.upper()
    for child in element.iter():
        if local_name(child.tag).upper() == wanted:
            return clean(child.text)
    return ""


def first_text(element: ET.Element, *names: str) -> str:
    for name in names:
        value = first_direct_text(element, name)
        if value:
            return value

    for name in names:
        value = first_deep_text(element, name)
        if value:
            return value

    return ""


# ================================================================
# TALLY HTTP CLIENT
# ================================================================

class TallyClient:
    def __init__(
        self,
        host: str,
        port: int,
        timeout: int = HTTP_TIMEOUT,
        log_callback=None,
    ):
        self.host = host.strip()
        self.port = int(port)
        self.timeout = timeout
        self.url = f"http://{self.host}:{self.port}"
        self.log_callback = log_callback

    def log(self, message: str):
        LOG.info(message)
        if self.log_callback:
            self.log_callback(message)

    def post(
        self,
        xml_request: str,
        request_name: str,
        retries: int = 3,
    ) -> ET.Element:

        last_error = None

        for attempt in range(1, retries + 1):

            self.log(
                f"{request_name}: HTTP attempt {attempt}/{retries}"
            )

            data = xml_request.encode("utf-8")

            request = urllib.request.Request(
                self.url,
                data=data,
                method="POST",
                headers={
                    "Content-Type": "text/xml; charset=utf-8",
                    "Accept": "text/xml",
                    "User-Agent": "TallyPrime-Accounting-Extractor/1.0",
                    "Connection": "close",
                },
            )

            opener = urllib.request.build_opener(
                urllib.request.ProxyHandler({})
            )

            try:
                with opener.open(
                    request,
                    timeout=self.timeout,
                ) as response:
                    raw = response.read()

                # Keep the latest response for diagnosis.
                try:
                    (OUTPUT_DIR / "last_tally_response.xml").write_bytes(raw)
                except Exception:
                    pass

                try:
                    root = ET.fromstring(
                        sanitize_xml(raw)
                    )
                except ET.ParseError as exc:
                    raise RuntimeError(
                        f"{request_name}: Tally returned invalid XML: {exc}"
                    )

                status = first_deep_text(
                    root,
                    "STATUS",
                )

                if status == "0":
                    errors = []

                    for element in root.iter():
                        tag = local_name(
                            element.tag
                        ).upper()

                        if tag in {
                            "LINEERROR",
                            "ERROR",
                        }:
                            value = clean(element.text)
                            if value:
                                errors.append(value)

                    detail = "\n".join(errors)

                    raise RuntimeError(
                        f"{request_name}: Tally rejected the request."
                        + (f"\n{detail}" if detail else "")
                    )

                return root

            except (
                urllib.error.URLError,
                urllib.error.HTTPError,
                TimeoutError,
                ConnectionError,
                OSError,
            ) as exc:

                last_error = exc

                self.log(
                    f"{request_name}: connection error: {exc}"
                )

                if attempt < retries:
                    time.sleep(2 * attempt)
                    continue

                break

            except RuntimeError:
                raise

            except Exception as exc:
                last_error = exc
                self.log(
                    f"{request_name}: unexpected error: {exc}"
                )
                if attempt < retries:
                    time.sleep(2 * attempt)
                    continue
                break

        raise RuntimeError(
            f"Could not get {request_name} from Tally at "
            f"{self.url}.\n\n"
            f"Last error: {last_error}\n\n"
            "Check that TallyPrime is running, a company is loaded, "
            "HTTP Server is enabled and the configured port is correct."
        )


# ================================================================
# ODBC
# ================================================================

def test_odbc(
    dsn: str,
    log_callback=None,
) -> Tuple[bool, str, List[str]]:

    def emit(message):
        LOG.info(message)
        if log_callback:
            log_callback(message)

    emit(f"Testing ODBC DSN: {dsn}")

    connection = None
    cursor = None

    try:
        connection = pyodbc.connect(
            f"DSN={dsn};",
            timeout=ODBC_TIMEOUT,
            autocommit=True,
        )

        cursor = connection.cursor()

        tables = []

        try:
            for row in cursor.tables(
                tableType="TABLE"
            ):
                if row.table_name:
                    tables.append(
                        str(row.table_name)
                    )
        except Exception as exc:
            emit(
                f"ODBC connected, but table discovery failed: {exc}"
            )

        emit(
            f"ODBC connection successful; "
            f"{len(tables)} tables exposed."
        )

        return True, "ODBC connection successful.", sorted(
            set(tables),
            key=str.casefold,
        )

    except Exception as exc:
        emit(
            f"ODBC connection failed: {exc}"
        )
        return False, str(exc), []

    finally:
        try:
            if cursor:
                cursor.close()
        except Exception:
            pass

        try:
            if connection:
                connection.close()
        except Exception:
            pass


# ================================================================
# ODBC LEDGER BALANCES - AUTHORITATIVE TRIAL BALANCE SOURCE
# ================================================================

ODBC_LEDGER_SQL = """
SELECT
    $Name,
    $Parent,
    $_PrimaryGroup,
    $OpeningBalance,
    $DebitTotals,
    $CreditTotals,
    $_ClosingBalance
FROM Ledger
""".strip()


def _normalise_odbc_column(name: Any) -> str:
    """Normalise a Tally ODBC column label for matching."""
    return re.sub(
        r"[^a-z0-9]+",
        "",
        clean(name).lower(),
    )


def fetch_odbc_ledger_balances(
    dsn: str,
    log_callback=None,
) -> Dict[str, Dict[str, Any]]:
    """
    Read the exact Ledger query verified by the user in TallyPrime.

    This is deliberately the authoritative source for:
      - Ledger Name
      - Group
      - Parent Group / Primary Group
      - Opening Balance
      - Debit
      - Credit
      - Closing Balance

    Tally's current ODBC documentation confirms that the Ledger collection
    is exposed through SQL and that ledger balances can be extracted through
    ODBC. The query below is the exact query configured for this application.
    """
    def emit(message: str):
        LOG.info(message)
        if log_callback:
            log_callback(message)

    connection = None
    cursor = None

    try:
        emit("Connecting to Tally ODBC for authoritative Ledger balances...")
        connection = pyodbc.connect(
            f"DSN={dsn};",
            timeout=ODBC_TIMEOUT,
            autocommit=True,
        )
        cursor = connection.cursor()

        emit("Executing verified Tally Ledger balance query:")
        emit(ODBC_LEDGER_SQL.replace("\n", " "))

        cursor.execute(ODBC_LEDGER_SQL)

        description = [
            col[0]
            for col in cursor.description
        ]

        rows = cursor.fetchall()

        emit(
            f"ODBC Ledger query returned {len(rows):,} rows."
        )

        # Map columns by their actual returned names. Do not assume that
        # pyodbc preserves a particular case.
        index_by_name = {
            _normalise_odbc_column(name): i
            for i, name in enumerate(description)
        }

        required = {
            "name": _normalise_odbc_column("$Name"),
            "parent": _normalise_odbc_column("$Parent"),
            "primary_group": _normalise_odbc_column("$_PrimaryGroup"),
            "opening": _normalise_odbc_column("$OpeningBalance"),
            "debit": _normalise_odbc_column("$DebitTotals"),
            "credit": _normalise_odbc_column("$CreditTotals"),
            "closing": _normalise_odbc_column("$_ClosingBalance"),
        }

        missing = [
            key
            for key, normalised in required.items()
            if normalised not in index_by_name
        ]

        if missing:
            raise RuntimeError(
                "Tally ODBC Ledger query succeeded, but the driver did not "
                "return the expected columns: "
                + ", ".join(missing)
                + "\nReturned columns: "
                + ", ".join(map(str, description))
            )

        result: Dict[str, Dict[str, Any]] = {}

        for raw_row in rows:

            def value(field: str) -> Any:
                return raw_row[
                    index_by_name[
                        required[field]
                    ]
                ]

            name = clean(
                value("name")
            )

            if not name:
                continue

            key = name.casefold()

            result[key] = {
                "name": name,
                "group": clean(
                    value("parent")
                ),
                "parent_group": clean(
                    value("primary_group")
                ),
                "opening": parse_amount(
                    value("opening")
                ),
                "debit": parse_amount(
                    value("debit")
                ),
                "credit": parse_amount(
                    value("credit")
                ),
                "closing": parse_amount(
                    value("closing")
                ),
            }

        if not result:
            raise RuntimeError(
                "The Tally ODBC Ledger query returned no usable ledger rows."
            )

        emit(
            f"Usable ODBC Ledger balances: {len(result):,}"
        )

        # Verify that all text values destined for Excel are free of illegal
        # control characters. This is diagnostic only; values are not changed
        # numerically.
        bad_text = 0
        for item in result.values():
            for field in ("name", "group", "parent_group"):
                value_text = str(item.get(field, ""))
                if any(
                    ord(ch) < 0x20 and ch not in "\t\n\r"
                    for ch in value_text
                ):
                    bad_text += 1
        if bad_text:
            emit(
                f"WARNING: {bad_text:,} ODBC text fields contained "
                "Excel-illegal control characters; they were sanitized."
            )

        return result

    except pyodbc.Error as exc:
        raise RuntimeError(
            "Tally ODBC Ledger query failed.\n\n"
            f"DSN: {dsn}\n"
            f"SQL: {ODBC_LEDGER_SQL}\n"
            f"ODBC error: {exc}"
        ) from exc

    finally:
        try:
            if cursor:
                cursor.close()
        except Exception:
            pass

        try:
            if connection:
                connection.close()
        except Exception:
            pass



def get_odbc_voucher_date_range(
    dsn: str,
    log_callback=None,
) -> Tuple[Optional[date], Optional[date]]:
    """
    Diagnostic only.

    Tally's ODBC Voucher collection is used to determine the earliest and
    latest voucher dates visible to the same DSN/company context. This does
    NOT replace XML transaction extraction; it tells the user whether the
    selected period actually exists in the ODBC-visible company data.
    """
    def emit(message: str):
        LOG.info(message)
        if log_callback:
            log_callback(message)

    connection = None
    cursor = None

    try:
        connection = pyodbc.connect(
            f"DSN={dsn};",
            timeout=ODBC_TIMEOUT,
            autocommit=True,
        )
        cursor = connection.cursor()

        earliest = None
        latest = None

        queries = [
            (
                "earliest",
                """
                SELECT TOP 1 $Date
                FROM Voucher
                WHERE NOT $$IsEmpty:$Date
                ORDER BY $Date ASC
                """,
            ),
            (
                "latest",
                """
                SELECT TOP 1 $Date
                FROM Voucher
                WHERE NOT $$IsEmpty:$Date
                ORDER BY $Date DESC
                """,
            ),
        ]

        for label, sql in queries:
            try:
                cursor.execute(sql)
                row = cursor.fetchone()

                if row and row[0] not in (None, ""):
                    parsed = parse_tally_date(
                        str(row[0])
                    )

                    if label == "earliest":
                        earliest = parsed
                    else:
                        latest = parsed

            except Exception as exc:
                emit(
                    f"ODBC Voucher {label}-date diagnostic failed: {exc}"
                )

        if earliest or latest:
            emit(
                "ODBC-visible Voucher date range: "
                f"{earliest.isoformat() if earliest else 'unknown'} "
                "to "
                f"{latest.isoformat() if latest else 'unknown'}"
            )

        return earliest, latest

    except Exception as exc:
        emit(
            f"Could not inspect ODBC Voucher date range: {exc}"
        )
        return None, None

    finally:
        try:
            if cursor:
                cursor.close()
        except Exception:
            pass

        try:
            if connection:
                connection.close()
        except Exception:
            pass


# ================================================================
# TALLY XML REQUESTS
# ================================================================

def company_xml(company: Optional[str]) -> str:
    if not company:
        return ""
    return (
        "<SVCURRENTCOMPANY TYPE=\"String\">"
        f"{xml_escape(company)}"
        "</SVCURRENTCOMPANY>"
    )


def build_daybook_report_request(
    from_date: date,
    to_date: date,
    company: Optional[str],
) -> str:
    """
    Official Tally Day Book report request.

    Tally's current XML documentation shows DayBook with TYPE=Data and
    SVFROMDATE/SVTODATE as the supported way to request transactions for
    a specific period. This is used as a fallback if the custom Voucher
    collection returns an empty collection.
    """
    return f"""
<ENVELOPE>
<HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Data</TYPE>
    <ID>DayBook</ID>
</HEADER>
<BODY>
<DESC>
<STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <SVFROMDATE TYPE="Date">{tally_xml_date(from_date)}</SVFROMDATE>
    <SVTODATE TYPE="Date">{tally_xml_date(to_date)}</SVTODATE>
    <EXPLODEFLAG>Yes</EXPLODEFLAG>
    {company_xml(company)}
</STATICVARIABLES>
</DESC>
</BODY>
</ENVELOPE>
""".strip()


def build_voucher_request(
    from_date: date,
    to_date: date,
    company: Optional[str],
) -> str:
    """
    Explicit Voucher collection with server-side date filtering.
    Local filtering is also applied after parsing.

    This avoids relying only on a Day Book report response when a Tally
    installation has report-period behaviour that differs from expectation.
    """

    return f"""
<ENVELOPE>
<HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>EXPORT</TALLYREQUEST>
    <TYPE>COLLECTION</TYPE>
    <ID>PY_TALLY_VOUCHERS</ID>
</HEADER>
<BODY>
<DESC>
<STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <SVFROMDATE TYPE="Date">{tally_xml_date(from_date)}</SVFROMDATE>
    <SVTODATE TYPE="Date">{tally_xml_date(to_date)}</SVTODATE>
    <SVCURRENTDATE TYPE="Date">{tally_xml_date(to_date)}</SVCURRENTDATE>
    {company_xml(company)}
</STATICVARIABLES>

<TDL>
<TDLMESSAGE>

<COLLECTION NAME="PY_TALLY_VOUCHERS"
    ISMODIFY="No"
    ISFIXED="No"
    ISINITIALIZE="Yes"
    ISOPTION="No"
    ISINTERNAL="No">

    <TYPE>Voucher</TYPE>

    <FILTER>PY_TALLY_DATE_FILTER</FILTER>

    <FETCH>
        DATE,
        VOUCHERTYPENAME,
        VOUCHERNUMBER,
        PARTYLEDGERNAME,
        PARTYNAME,
        NARRATION,
        REFERENCE,
        MASTERID,
        GUID,
        ALLLEDGERENTRIES.*
    </FETCH>

</COLLECTION>

<SYSTEM TYPE="Formulae"
    NAME="PY_TALLY_DATE_FILTER"
    ISMODIFY="No"
    ISFIXED="No"
    ISINTERNAL="No">
    $Date &gt;= ##SVFromDate
    AND
    $Date &lt;= ##SVToDate
</SYSTEM>

</TDLMESSAGE>
</TDL>
</DESC>
</BODY>
</ENVELOPE>
""".strip()


def build_ledger_master_request(
    company: Optional[str],
) -> str:

    return f"""
<ENVELOPE>
<HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>EXPORT</TALLYREQUEST>
    <TYPE>COLLECTION</TYPE>
    <ID>PY_LEDGER_MASTER</ID>
</HEADER>
<BODY>
<DESC>
<STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    {company_xml(company)}
</STATICVARIABLES>

<TDL>
<TDLMESSAGE>

<COLLECTION NAME="PY_LEDGER_MASTER"
    ISMODIFY="No"
    ISFIXED="No"
    ISINITIALIZE="Yes"
    ISOPTION="No"
    ISINTERNAL="No">

    <TYPE>Ledger</TYPE>

    <FETCH>
        NAME,
        PARENT
    </FETCH>

</COLLECTION>

</TDLMESSAGE>
</TDL>
</DESC>
</BODY>
</ENVELOPE>
""".strip()


def build_group_request(
    company: Optional[str],
) -> str:

    return f"""
<ENVELOPE>
<HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>EXPORT</TALLYREQUEST>
    <TYPE>COLLECTION</TYPE>
    <ID>PY_GROUP_MASTER</ID>
</HEADER>
<BODY>
<DESC>
<STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    {company_xml(company)}
</STATICVARIABLES>

<TDL>
<TDLMESSAGE>

<COLLECTION NAME="PY_GROUP_MASTER"
    ISMODIFY="No"
    ISFIXED="No"
    ISINITIALIZE="Yes"
    ISOPTION="No"
    ISINTERNAL="No">

    <TYPE>Group</TYPE>

    <FETCH>
        NAME,
        PARENT
    </FETCH>

</COLLECTION>

</TDLMESSAGE>
</TDL>
</DESC>
</BODY>
</ENVELOPE>
""".strip()


def build_trial_balance_request(
    as_of_date: date,
    company: Optional[str],
) -> str:
    """
    Tally's documented Trial Balance XML report.

    EXPLODEFLAG=Yes asks Tally for detailed ledger-level information.
    """

    return f"""
<ENVELOPE>
<HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>EXPORT</TALLYREQUEST>
    <TYPE>DATA</TYPE>
    <ID>Trial Balance</ID>
</HEADER>
<BODY>
<DESC>
<STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <EXPLODEFLAG>Yes</EXPLODEFLAG>
    <SVCURRENTDATE TYPE="Date">{tally_xml_date(as_of_date)}</SVCURRENTDATE>
    <SVFROMDATE TYPE="Date">{tally_xml_date(as_of_date)}</SVFROMDATE>
    <SVTODATE TYPE="Date">{tally_xml_date(as_of_date)}</SVTODATE>
    {company_xml(company)}
</STATICVARIABLES>
</DESC>
</BODY>
</ENVELOPE>
""".strip()


# ================================================================
# XML PARSING - VOUCHERS
# ================================================================

def find_voucher_nodes(
    root: ET.Element,
) -> List[ET.Element]:

    result = []

    for element in root.iter():
        if local_name(element.tag).upper() == "VOUCHER":
            result.append(element)

    return result


def extract_ledger_lines_from_voucher(
    voucher: ET.Element,
) -> List[Dict[str, Any]]:

    nodes = [
        child
        for child in voucher
        if local_name(child.tag).upper()
        in {
            "ALLLEDGERENTRIES.LIST",
            "LEDGERENTRIES.LIST",
        }
    ]

    if not nodes:
        nodes = [
            child
            for child in voucher.iter()
            if local_name(child.tag).upper()
            in {
                "ALLLEDGERENTRIES.LIST",
                "LEDGERENTRIES.LIST",
            }
        ]

    vdate = parse_tally_date(
        first_text(voucher, "DATE")
    )

    vtype = first_text(
        voucher,
        "VOUCHERTYPENAME",
    )

    vno = first_text(
        voucher,
        "VOUCHERNUMBER",
    )

    master_id = (
        first_text(
            voucher,
            "MASTERID",
            "GUID",
            "VOUCHERKEY",
        )
        or
        f"{vdate}|{vtype}|{vno}"
    )

    narration = first_text(
        voucher,
        "NARRATION",
    )

    reference = first_text(
        voucher,
        "REFERENCE",
    )

    party = (
        first_text(
            voucher,
            "PARTYLEDGERNAME",
        )
        or
        first_text(
            voucher,
            "PARTYNAME",
        )
    )

    result = []

    for line_index, line in enumerate(
        nodes,
        start=1,
    ):

        ledger_name = first_text(
            line,
            "LEDGERNAME",
        )

        if not ledger_name:
            continue

        amount = parse_amount(
            first_text(
                line,
                "AMOUNT",
            )
        )

        deemed = first_text(
            line,
            "ISDEEMEDPOSITIVE",
        )

        if deemed:
            if bool_tally(deemed):
                debit = abs(amount)
                credit = 0.0
            else:
                debit = 0.0
                credit = abs(amount)
        else:
            if amount < 0:
                debit = abs(amount)
                credit = 0.0
            else:
                debit = 0.0
                credit = abs(amount)

        bills = []

        for bill in line.iter():
            if (
                local_name(
                    bill.tag
                ).upper()
                != "BILLALLOCATIONS.LIST"
            ):
                continue

            bill_name = first_text(
                bill,
                "NAME",
            )

            bill_type = first_text(
                bill,
                "BILLTYPE",
            )

            bill_amount = first_text(
                bill,
                "AMOUNT",
            )

            bits = []

            if bill_name:
                bits.append(
                    f"Bill={bill_name}"
                )

            if bill_type:
                bits.append(
                    f"Type={bill_type}"
                )

            if bill_amount:
                bits.append(
                    f"Amount={bill_amount}"
                )

            if bits:
                bills.append(
                    " | ".join(bits)
                )

        costs = []

        for cc in line.iter():
            if (
                local_name(
                    cc.tag
                ).upper()
                not in {
                    "COSTCENTREALLOCATIONS.LIST",
                    "CATEGORYALLOCATIONS.LIST",
                }
            ):
                continue

            cc_name = first_text(
                cc,
                "NAME",
            )

            cc_amount = first_text(
                cc,
                "AMOUNT",
            )

            if cc_name:
                value = cc_name

                if cc_amount:
                    value += (
                        f" ({cc_amount})"
                    )

                costs.append(value)

        tax_values = []

        for element in line.iter():

            tag = local_name(
                element.tag
            ).upper()

            if not any(
                keyword in tag
                for keyword in (
                    "GST",
                    "CGST",
                    "SGST",
                    "IGST",
                    "CESS",
                    "TAX",
                    "HSN",
                    "SAC",
                )
            ):
                continue

            value = clean(
                element.text
            )

            if value:
                tax_values.append(
                    f"{tag}={value}"
                )

        result.append(
            {
                "Voucher ID":
                    clean(master_id),

                "Line No":
                    line_index,

                "Date":
                    vdate,

                "Voucher Type":
                    vtype,

                "Voucher Number":
                    vno,

                "Ledger/Party Name":
                    clean(ledger_name),

                "Party Name":
                    party,

                "Debit":
                    debit,

                "Credit":
                    credit,

                "Narration":
                    clean(
                        first_text(
                            line,
                            "NARRATION",
                        )
                        or narration
                    ),

                "Reference/Bill Number":
                    reference,

                "Bill Details":
                    unique_join(bills),

                "GST/Tax Information":
                    unique_join(tax_values),

                "Cost Centre":
                    unique_join(costs),
            }
        )

    return result


def extract_vouchers_and_ledgers(
    root: ET.Element,
    from_date: date,
    to_date: date,
) -> Tuple[
    List[Dict[str, Any]],
    List[Dict[str, Any]],
]:

    vouchers = []
    ledger_rows = []

    seen_vouchers = set()

    for voucher in find_voucher_nodes(root):

        vdate = parse_tally_date(
            first_text(
                voucher,
                "DATE",
            )
        )

        if not vdate:
            continue

        if not (
            from_date
            <= vdate
            <= to_date
        ):
            continue

        vtype = first_text(
            voucher,
            "VOUCHERTYPENAME",
        )

        vno = first_text(
            voucher,
            "VOUCHERNUMBER",
        )

        master_id = (
            first_text(
                voucher,
                "MASTERID",
                "GUID",
                "VOUCHERKEY",
            )
            or
            f"{vdate.isoformat()}|{vtype}|{vno}"
        )

        master_id = clean(
            master_id
        )

        # A voucher can occur more than once in an exploded response.
        # We process the first full voucher only.
        if master_id in seen_vouchers:
            continue

        seen_vouchers.add(
            master_id
        )

        party = (
            first_text(
                voucher,
                "PARTYLEDGERNAME",
            )
            or
            first_text(
                voucher,
                "PARTYNAME",
            )
        )

        vouchers.append(
            {
                "Voucher ID":
                    master_id,

                "Voucher Date":
                    vdate,

                "Voucher Type":
                    vtype,

                "Voucher Number":
                    vno,

                "Party Name":
                    party,

                "Narration":
                    first_text(
                        voucher,
                        "NARRATION",
                    ),

                "Reference/Bill Number":
                    first_text(
                        voucher,
                        "REFERENCE",
                    ),
            }
        )

        # IMPORTANT:
        # Do not deduplicate two identical-looking ledger lines.
        # A voucher can legitimately contain the same ledger twice.
        for line in extract_ledger_lines_from_voucher(
            voucher
        ):
            line["Voucher ID"] = master_id
            ledger_rows.append(line)

    return vouchers, ledger_rows


# ================================================================
# MASTER DATA
# ================================================================

def fetch_ledger_master(
    client: TallyClient,
    company: Optional[str],
) -> Dict[str, Dict[str, Any]]:

    root = client.post(
        build_ledger_master_request(
            company
        ),
        "Ledger master",
    )

    result = {}

    for element in root.iter():

        if (
            local_name(
                element.tag
            ).upper()
            != "LEDGER"
        ):
            continue

        name = first_text(
            element,
            "NAME",
        )

        if not name:
            name = clean(
                element.attrib.get(
                    "NAME",
                    "",
                )
            )

        if not name:
            continue

        parent = first_text(
            element,
            "PARENT",
        )

        key = name.casefold()

        result[key] = {
            "name": name,
            "group": parent,
        }

    return result


def fetch_groups(
    client: TallyClient,
    company: Optional[str],
) -> Dict[str, Dict[str, str]]:

    root = client.post(
        build_group_request(
            company
        ),
        "Group master",
    )

    result = {}

    for element in root.iter():

        if (
            local_name(
                element.tag
            ).upper()
            != "GROUP"
        ):
            continue

        name = first_text(
            element,
            "NAME",
        )

        if not name:
            name = clean(
                element.attrib.get(
                    "NAME",
                    "",
                )
            )

        if not name:
            continue

        result[
            name.casefold()
        ] = {
            "name": name,
            "parent": first_text(
                element,
                "PARENT",
            ),
        }

    return result


def parent_group(
    group_name: str,
    groups: Dict[str, Dict[str, str]],
) -> str:

    group_name = clean(
        group_name
    )

    if not group_name:
        return ""

    info = groups.get(
        group_name.casefold()
    )

    if not info:
        return ""

    parent = clean(
        info.get(
            "parent",
            "",
        )
    )

    return parent


# ================================================================
# TALLY TRIAL BALANCE
# ================================================================

def parse_trial_balance_report(
    root: ET.Element,
    known_ledgers: Dict[str, Dict[str, Any]],
) -> Dict[str, float]:
    """
    Parse Tally's Trial Balance XML.

    Tally's documented response contains alternating DSPACCNAME and
    DSPACCINFO elements. DSPCLDRAMTA is the debit amount (typically
    represented as a negative value) and DSPCLCRAMTA is the credit amount.

    We only keep names that are actually present in the Ledger master, so
    report group headings are not accidentally treated as ledgers.
    """

    result: Dict[str, float] = {}

    # The normal response has DSPACCNAME followed by DSPACCINFO as siblings.
    # Parse at the parent level first; this is much safer than keeping one
    # global "pending" name while walking the entire document.
    for parent in root.iter():

        children = list(parent)

        if not children:
            continue

        for idx, child in enumerate(children):

            if (
                local_name(child.tag).upper()
                != "DSPACCNAME"
            ):
                continue

            if idx + 1 >= len(children):
                continue

            info = children[idx + 1]

            if (
                local_name(info.tag).upper()
                != "DSPACCINFO"
            ):
                continue

            name = clean(
                first_text(
                    child,
                    "DSPDISPNAME",
                )
            )

            if not name:
                continue

            key = name.casefold()

            if key not in known_ledgers:
                continue

            debit_value = abs(
                parse_amount(
                    first_text(
                        info,
                        "DSPCLDRAMTA",
                    )
                )
            )

            credit_value = abs(
                parse_amount(
                    first_text(
                        info,
                        "DSPCLCRAMTA",
                    )
                )
            )

            # Signed convention used throughout this application:
            # debit balance = negative
            # credit balance = positive
            result[key] = (
                credit_value
                - debit_value
            )

    # Fallback for releases where DSPACCINFO is nested differently.
    if not result:

        elements = list(root.iter())

        for index, element in enumerate(elements):

            if (
                local_name(element.tag).upper()
                != "DSPACCNAME"
            ):
                continue

            name = clean(
                first_text(
                    element,
                    "DSPDISPNAME",
                )
            )

            if not name:
                continue

            key = name.casefold()

            if key not in known_ledgers:
                continue

            # Search only forward until the next account name, avoiding
            # accidental pairing with a later ledger.
            info = None

            for candidate in elements[
                index + 1:
                index + 8
            ]:

                candidate_tag = (
                    local_name(
                        candidate.tag
                    ).upper()
                )

                if candidate_tag == "DSPACCNAME":
                    break

                if candidate_tag == "DSPACCINFO":
                    info = candidate
                    break

            if info is None:
                continue

            debit_value = abs(
                parse_amount(
                    first_text(
                        info,
                        "DSPCLDRAMTA",
                    )
                )
            )

            credit_value = abs(
                parse_amount(
                    first_text(
                        info,
                        "DSPCLCRAMTA",
                    )
                )
            )

            result[key] = (
                credit_value
                - debit_value
            )

    return result


def fetch_trial_balance_as_of(
    client: TallyClient,
    as_of_date: date,
    company: Optional[str],
    known_ledgers: Dict[str, Dict[str, Any]],
) -> Dict[str, float]:

    root = client.post(
        build_trial_balance_request(
            as_of_date,
            company,
        ),
        f"Trial Balance {as_of_date:%Y-%m-%d}",
    )

    result = parse_trial_balance_report(
        root,
        known_ledgers,
    )

    LOG.info(
        "Trial Balance %s: %d ledger balances parsed.",
        as_of_date,
        len(result),
    )

    return result


# ================================================================
# DATAFRAMES
# ================================================================

# ================================================================
# FINANCIAL REPORTS: PROFIT & LOSS / BALANCE SHEET
# ================================================================

def build_financial_report_request(
    report_name: str,
    from_date: date,
    to_date: date,
    company: Optional[str],
) -> str:
    """
    Request a native TallyPrime financial report over XML/HTTP.

    TallyPrime documents report export with:
        TALLYREQUEST=Export
        TYPE=Data
        ID=<report name>

    Profit & Loss is period-sensitive, so both SVFROMDATE and SVTODATE
    are supplied. Balance Sheet is an as-on-date statement; SVTODATE is
    the authoritative date, while SVFROMDATE is also supplied for
    releases that expose comparative/period context.
    """
    return f"""
<ENVELOPE>
<HEADER>
    <VERSION>1</VERSION>
    <TALLYREQUEST>Export</TALLYREQUEST>
    <TYPE>Data</TYPE>
    <ID>{xml_escape(report_name)}</ID>
</HEADER>
<BODY>
<DESC>
<STATICVARIABLES>
    <SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
    <SVFROMDATE TYPE="Date">{tally_xml_date(from_date)}</SVFROMDATE>
    <SVTODATE TYPE="Date">{tally_xml_date(to_date)}</SVTODATE>
    <EXPLODEFLAG>Yes</EXPLODEFLAG>
    {company_xml(company)}
</STATICVARIABLES>
</DESC>
</BODY>
</ENVELOPE>
""".strip()


def _amount_from_first_tags(
    element: ET.Element,
    tag_names: Iterable[str],
) -> float:
    for tag_name in tag_names:
        value = first_text(
            element,
            tag_name,
        )
        if value not in (None, ""):
            return parse_amount(value)
    return 0.0


def _parse_financial_pairs(
    root: ET.Element,
    section: str = "",
) -> List[Dict[str, Any]]:
    """
    Parse Tally's native Profit & Loss / Balance Sheet XML.

    Important: Tally has more than one XML layout for financial reports.

    Trial Balance commonly uses:
        DSPACCNAME + DSPACCINFO

    Balance Sheet commonly uses:
        BSNAME + BSAMT
        BSAMT/BSMAINAMT for main/group amounts
        BSAMT/BSSUBAMT for sub-ledger amounts

    Profit & Loss commonly uses:
        DSPACCNAME + PLAMT
        PLAMT/BSMAINAMT and/or PLAMT/PLSUBAMT

    The user's actual Balance Sheet response uses BSNAME + BSAMT, so the
    parser must explicitly support that layout. The previous implementation
    only looked for DSPACCINFO and therefore returned zero rows even though
    Tally had returned the complete Balance Sheet.
    """
    rows: List[Dict[str, Any]] = []

    def amount_from_container(
        amount_node: ET.Element,
        main_tags: Tuple[str, ...],
        sub_tags: Tuple[str, ...],
    ) -> Tuple[float, str]:
        """
        Read the Tally hierarchy from the XML tag itself.

        BSMAINAMT means the report row is a Main/group total.
        BSSUBAMT / PLSUBAMT means the row is a Sub/detail row.

        This is deliberately based on tag presence, not on whether the
        numeric value happens to be zero. A zero-valued Main group must still
        remain Main.
        """
        for tag in main_tags:
            value = first_direct_text(
                amount_node,
                tag,
            )
            if value not in ("", None):
                return parse_amount(value), "Main"

        for tag in sub_tags:
            value = first_direct_text(
                amount_node,
                tag,
            )
            if value not in ("", None):
                return parse_amount(value), "Sub"

        return 0.0, ""

    def add_row(
        row_section: str,
        name_node: ET.Element,
        amount_node: ET.Element,
        amount_type: str,
    ) -> None:
        if local_name(name_node.tag).upper() == "BSNAME":
            name = clean(
                first_deep_text(
                    name_node,
                    "DSPDISPNAME",
                )
            )
        else:
            name = clean(
                first_deep_text(
                    name_node,
                    "DSPDISPNAME",
                )
            )

        if not name:
            return

        if amount_type == "BS":
            amount, level = amount_from_container(
                amount_node,
                ("BSMAINAMT", "BSCLOSINGAMT", "BSCLOSINGBAL"),
                ("BSSUBAMT",),
            )
        elif amount_type == "PL":
            amount, level = amount_from_container(
                amount_node,
                ("BSMAINAMT",),
                ("PLSUBAMT",),
            )
        else:
            # DSPACCINFO / Trial-Balance-style structure.
            debit = abs(
                _amount_from_first_tags(
                    amount_node,
                    (
                        "DSPCLDRAMTA",
                        "DSPDRAMTA",
                        "DEBIT",
                    ),
                )
            )
            credit = abs(
                _amount_from_first_tags(
                    amount_node,
                    (
                        "DSPCLCRAMTA",
                        "DSPCRAMTA",
                        "CREDIT",
                    ),
                )
            )
            amount = credit - debit
            level = ""

            rows.append(
                {
                    "Section": clean(row_section),
                    "Particulars": name,
                    "Amount": amount,
                    "Debit": debit,
                    "Credit": credit,
                }
            )
            return

        rows.append(
            {
                "Section": clean(row_section),
                "Particulars": name,
                "Amount": amount,
                "Debit": 0.0,
                "Credit": 0.0,
                "Level": level,
            }
        )

    # ------------------------------------------------------------
    # 1. Explicit Balance Sheet sections, if the Tally release returns them.
    # ------------------------------------------------------------

    explicit_sections = False

    for node in root.iter():
        tag = local_name(node.tag).upper()

        if tag not in {"BSSOURCES", "BSAPP"}:
            continue

        explicit_sections = True

        row_section = (
            "Sources of Funds"
            if tag == "BSSOURCES"
            else "Application of Funds"
        )

        # Each detail normally contains BSNAME followed by BSAMT.
        for detail in node.iter():
            children = list(detail)

            for idx, child in enumerate(children):
                child_tag = local_name(child.tag).upper()

                if child_tag not in {"BSNAME", "DSPACCNAME"}:
                    continue

                for candidate in children[idx + 1:idx + 4]:
                    candidate_tag = local_name(
                        candidate.tag
                    ).upper()

                    if candidate_tag == "BSAMT":
                        add_row(
                            row_section,
                            child,
                            candidate,
                            "BS",
                        )
                        break

                    if candidate_tag in {
                        "BSNAME",
                        "DSPACCNAME",
                    }:
                        break

    # ------------------------------------------------------------
    # 2. Flat Balance Sheet layout:
    #
    # BSNAME -> DSPACCNAME -> BSAMT
    #
    # This is the exact structure returned by the user's TallyPrime.
    # ------------------------------------------------------------

    if not explicit_sections:
        children = list(root)

        # In the flat XML layout returned by this TallyPrime environment,
        # source-of-funds rows are followed by application-of-funds rows.
        # Once the first primary asset group is encountered, subsequent rows
        # belong to Application of Funds until the report ends.
        application_started = False

        application_groups = {
            "fixed assets",
            "current assets",
            "loans & advances (asset)",
            "investments",
            "misc. expenses (asset)",
            "branch / divisions",
        }

        for idx, child in enumerate(children):

            if local_name(child.tag).upper() != "BSNAME":
                continue

            for candidate in children[idx + 1:idx + 4]:

                candidate_tag = local_name(
                    candidate.tag
                ).upper()

                if candidate_tag == "BSAMT":

                    name = clean(
                        first_deep_text(
                            child,
                            "DSPDISPNAME",
                        )
                    )

                    if (
                        name.casefold()
                        in application_groups
                    ):
                        application_started = True

                    row_section = (
                        "Application of Funds"
                        if application_started
                        else "Sources of Funds"
                    )

                    add_row(
                        row_section,
                        child,
                        candidate,
                        "BS",
                    )
                    break

                if candidate_tag == "BSNAME":
                    break

    # ------------------------------------------------------------
    # 3. Profit & Loss layout:
    #
    # DSPACCNAME -> PLAMT
    #
    # Some releases return DSPACCNAME and PLAMT as adjacent top-level
    # elements; others wrap them. Search both direct siblings and nested
    # children.
    # ------------------------------------------------------------

    def scan_pl_children(parent: ET.Element) -> None:
        children = list(parent)

        for idx, child in enumerate(children):

            tag = local_name(child.tag).upper()

            if tag != "DSPACCNAME":
                continue

            for candidate in children[idx + 1:idx + 4]:

                candidate_tag = local_name(
                    candidate.tag
                ).upper()

                if candidate_tag == "PLAMT":
                    add_row(
                        section or "Profit & Loss",
                        child,
                        candidate,
                        "PL",
                    )
                    break

                if candidate_tag in {
                    "DSPACCNAME",
                    "BSNAME",
                }:
                    break

        # Nested report parts can have their own DSPACCNAME/PLAMT pairs.
        for child in children:
            if list(child):
                scan_pl_children(child)

    scan_pl_children(root)

    # ------------------------------------------------------------
    # 4. Generic Trial-Balance-style fallback.
    # ------------------------------------------------------------

    if not rows:

        for parent in root.iter():

            children = list(parent)

            for idx, child in enumerate(children):

                if local_name(
                    child.tag
                ).upper() != "DSPACCNAME":
                    continue

                for candidate in children[idx + 1:idx + 5]:

                    if local_name(
                        candidate.tag
                    ).upper() == "DSPACCINFO":

                        add_row(
                            section or "Report",
                            child,
                            candidate,
                            "DSP",
                        )
                        break

                    if local_name(
                        candidate.tag
                    ).upper() in {
                        "DSPACCNAME",
                        "BSNAME",
                    }:
                        break

    return rows


def parse_financial_report(
    root: ET.Element,
    report_name: str,
) -> pd.DataFrame:
    report_upper = report_name.upper()

    rows = _parse_financial_pairs(
        root,
        "Profit & Loss" if "PROFIT" in report_upper else "Balance Sheet",
    )

    columns = [
        "Section",
        "Particulars",
        "Amount",
        "Debit",
        "Credit",
        "Level",
    ]

    if not rows:
        return pd.DataFrame(
            columns=columns
        )

    df = pd.DataFrame(rows)

    for col in columns:
        if col not in df.columns:
            df[col] = ""

    df = df[columns]

    df = df.drop_duplicates(
        subset=[
            "Section",
            "Particulars",
            "Amount",
            "Debit",
            "Credit",
            "Level",
        ],
        keep="first",
    )

    for col in [
        "Amount",
        "Debit",
        "Credit",
    ]:
        df[col] = pd.to_numeric(
            df[col],
            errors="coerce",
        ).fillna(0.0)

    df["Section"] = df["Section"].map(clean)
    df["Particulars"] = df["Particulars"].map(clean)
    df["Level"] = df["Level"].map(clean)

    return df.reset_index(drop=True)


def fetch_financial_report(
    client: "TallyClient",
    report_name: str,
    from_date: date,
    to_date: date,
    company: Optional[str],
) -> pd.DataFrame:
    """
    Fetch a native Tally financial report.

    TallyPrime's user-facing name for the P&L report is "Profit & Loss A/c",
    but the internal report identifier is not guaranteed to be identical
    across Tally releases/configurations. In this environment the literal
    "Profit & Loss A/c" was rejected by Tally.

    We therefore try the documented/common internal identifiers in order and
    accept the first successful XML report. Balance Sheet is normally
    "Balance Sheet" and is kept as-is.

    No accounting values are invented: if every supported identifier is
    rejected, the error contains each Tally response.
    """

    candidates = [report_name]

    normalized = re.sub(
        r"\s+",
        " ",
        report_name.strip(),
    ).casefold()

    if normalized in {
        "profit & loss a/c",
        "profit and loss a/c",
        "profit & loss account",
        "profit and loss account",
    }:
        # Tally's current report UI calls it Profit & Loss A/c, while the
        # internal report identifier commonly used by integrations is
        # "Profit and Loss A/c". Try this first for compatibility.
        candidates = [
            "Profit and Loss A/c",
            "Profit & Loss A/c",
            "Profit and Loss Account",
            "Profit & Loss Account",
            "Profit and Loss",
            "Profit & Loss",
        ]

    errors = []

    for candidate in candidates:

        try:
            LOG.info(
                "Requesting Tally financial report identifier: %s",
                candidate,
            )

            root = client.post(
                build_financial_report_request(
                    candidate,
                    from_date,
                    to_date,
                    company,
                ),
                f"{candidate} report",
            )

            df = parse_financial_report(
                root,
                candidate,
            )

            LOG.info(
                "%s report rows parsed: %s",
                candidate,
                len(df),
            )

            # Do not silently accept a structurally valid but empty report.
            # If the report is empty, try the next compatible identifier.
            if df.empty:
                errors.append(
                    f"{candidate}: Tally accepted the request but "
                    "the response contained no financial-report rows."
                )
                LOG.warning(
                    "%s returned a successful response but 0 report rows.",
                    candidate,
                )
                continue

            return df

        except Exception as exc:

            error_text = str(exc)

            errors.append(
                f"{candidate}: {error_text}"
            )

            LOG.warning(
                "Tally financial report identifier '%s' failed: %s",
                candidate,
                error_text,
            )

    raise RuntimeError(
        f"Could not fetch Tally financial report '{report_name}'. "
        "Tried the following identifiers:\n"
        + "\n".join(
            f" - {item}"
            for item in errors
        )
    )


def _save_financial_report(
    df: pd.DataFrame,
    report_title: str,
    filename_prefix: str,
    from_date: date,
    to_date: date,
) -> Path:
    path = OUTPUT_DIR / (
        f"{filename_prefix}_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    wb.remove(
        wb.active
    )

    ws = wb.create_sheet(
        "Report"
    )
    ws.sheet_view.showGridLines = False

    ws["A1"] = report_title.upper()
    ws["A1"].font = Font(
        bold=True,
        size=16,
    )
    ws["A1"].fill = TITLE_FILL

    ws["A2"] = "Period"
    ws["B2"] = (
        f"{from_date:%d-%m-%Y} to "
        f"{to_date:%d-%m-%Y}"
    )

    start_row = 4

    if df.empty:
        ws.cell(
            start_row,
            1,
            "No records returned by Tally for this report.",
        )
        ws.cell(
            start_row,
            1,
        ).font = Font(
            bold=True
        )

    else:
        columns = list(
            df.columns
        )

        for c, header in enumerate(
            columns,
            start=1,
        ):
            cell = ws.cell(
                start_row,
                c,
                header,
            )
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for r, values in enumerate(
            df.itertuples(
                index=False,
                name=None,
            ),
            start=start_row + 1,
        ):
            for c, value in enumerate(
                values,
                start=1,
            ):
                cell = ws.cell(
                    r,
                    c,
                    excel_safe_value(value),
                )

                if columns[c - 1] in {
                    "Amount",
                    "Debit",
                    "Credit",
                }:
                    cell.number_format = (
                        '#,##0.00;[Red]-#,##0.00'
                    )

        ws.auto_filter.ref = (
            f"A{start_row}:"
            f"{get_column_letter(len(columns))}"
            f"{start_row + len(df)}"
        )

        ws.freeze_panes = (
            f"A{start_row + 1}"
        )

        # Total amount/debit/credit where meaningful.
        total_row = (
            start_row
            + len(df)
            + 1
        )

        ws.cell(
            total_row,
            1,
            "TOTAL",
        ).font = Font(
            bold=True
        )

        for c, header in enumerate(
            columns,
            start=1,
        ):
            if header not in {
                "Amount",
                "Debit",
                "Credit",
            }:
                continue

            letter = get_column_letter(
                c
            )

            cell = ws.cell(
                total_row,
                c,
                (
                    f"=SUM("
                    f"{letter}{start_row + 1}:"
                    f"{letter}{total_row - 1}"
                    f")"
                ),
            )

            cell.number_format = (
                '#,##0.00;[Red]-#,##0.00'
            )

            cell.fill = TOTAL_FILL
            cell.font = Font(
                bold=True
            )

    widths = {
        "A": 28,
        "B": 48,
        "C": 22,
        "D": 22,
        "E": 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[
            col
        ].width = width

    wb.save(
        path
    )

    return path


def _pnl_category(particulars: str, amount: float) -> str:
    """
    Presentation classification for P&L.

    Tally's native P&L response in this environment uses signed amounts:
      positive  -> income / credit
      negative  -> expense / debit

    The sign is therefore authoritative. Text is used only for zero-valued
    rows where the report itself gives no monetary direction.
    """
    if amount > EPSILON:
        return "Income"

    if amount < -EPSILON:
        return "Expenses"

    name = clean(particulars).casefold()

    expense_terms = (
        "expense",
        "expenses",
        "purchase",
        "depreciation",
        "tax expense",
        "bank charges",
        "audit fee",
        "salary",
        "rent",
        "interest exp",
        "interest on",
        "insurance",
        "electric",
        "courier",
        "legal",
        "repair",
        "advertisement",
    )

    if any(term in name for term in expense_terms):
        return "Expenses"

    return "Income"


def save_profit_loss(
    df: pd.DataFrame,
    from_date: date,
    to_date: date,
) -> Path:
    """
    Save P&L in a concise, readable financial-statement layout.

    Layout:
      - Left side: INCOME
      - Right side: EXPENSES
      - Main/Sub hierarchy is clearly shown.
      - Debit and Credit balances are separate.
      - Income is shown as a credit balance when positive.
      - Expenses are shown as a debit balance when negative.
      - Net Profit/Loss is calculated from the two Tally report totals.
      - Tally signed values are never altered.

    This is a brief management-style P&L presentation. It is not represented
    as a statutory Schedule III filing.
    """
    path = OUTPUT_DIR / (
        f"Profit_Loss_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Profit & Loss"
    ws.sheet_view.showGridLines = False

    # ------------------------------------------------------------
    # Title
    # ------------------------------------------------------------
    ws.merge_cells("A1:H1")
    ws["A1"] = "PROFIT & LOSS A/C"
    ws["A1"].font = Font(
        bold=True,
        size=16,
    )
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    ws.row_dimensions[1].height = 25

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Period: {from_date:%d-%m-%Y} to "
        f"{to_date:%d-%m-%Y}"
    )
    ws["A2"].font = Font(
        italic=True,
        size=10,
    )

    if df is None or df.empty:
        ws.merge_cells("A4:H4")
        ws["A4"] = (
            "No Profit & Loss records were returned by Tally."
        )
        ws["A4"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 34
        wb.save(path)
        return path

    data = df.copy()

    for col in [
        "Particulars",
        "Level",
        "Amount",
    ]:
        if col not in data.columns:
            data[col] = "" if col != "Amount" else 0.0

    data["Amount"] = pd.to_numeric(
        data["Amount"],
        errors="coerce",
    ).fillna(0.0)

    data["Particulars"] = data["Particulars"].map(clean)
    data["Level"] = data["Level"].map(clean)

    data = data[
        data["Particulars"].ne("")
    ].copy()

    # Remove exact duplicate report rows without changing source values.
    data = data.drop_duplicates(
        subset=[
            "Particulars",
            "Amount",
            "Level",
        ],
        keep="first",
    )

    data["Category"] = data.apply(
        lambda r: _pnl_category(
            r["Particulars"],
            float(r["Amount"]),
        ),
        axis=1,
    )

    income = data[
        data["Category"].eq("Income")
    ].copy()

    expenses = data[
        data["Category"].eq("Expenses")
    ].copy()

    # ------------------------------------------------------------
    # IMPORTANT ACCOUNTING RULE:
    # Tally's Main rows are already group totals. Sub rows are the
    # components of those totals. Therefore totals and the current-period
    # profit/loss MUST use Main rows only, otherwise every group is counted
    # twice (once at Main level and again through its Sub ledgers).
    # ------------------------------------------------------------
    main_mask = (
        data["Level"]
        .str.casefold()
        .eq("main")
    )

    main_data = data[
        main_mask
    ].copy()

    main_income = main_data[
        main_data["Amount"] > EPSILON
    ]

    main_expenses = main_data[
        main_data["Amount"] < -EPSILON
    ]

    income_total = float(
        main_income["Amount"].sum()
    )

    expense_total = float(
        -main_expenses["Amount"].sum()
    )

    # Accounting result from Main P&L group totals only.
    net_profit = (
        income_total
        - expense_total
    )

    ws["A4"] = "Total Income"
    ws["B4"] = income_total
    ws["D4"] = "Total Expenses"
    ws["E4"] = expense_total
    ws["G4"] = "Net Profit / (Loss)"
    ws["H4"] = net_profit

    ws["A5"] = "Calculation"
    ws["B5"] = (
        "Main group totals only "
        "(Sub rows are details and are excluded from totals)"
    )
    ws["A5"].font = Font(
        italic=True,
        size=9,
    )
    ws["B5"].font = Font(
        italic=True,
        size=9,
    )

    for ref in ["B4", "E4", "H4"]:
        ws[ref].number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )

    for ref in ["A4", "D4", "G4"]:
        ws[ref].font = Font(
            bold=True,
        )

    for ref in ["B4", "E4", "H4"]:
        ws[ref].font = Font(
            bold=True,
        )
        ws[ref].fill = TOTAL_FILL

    # ------------------------------------------------------------
    # Two-sided statement
    # ------------------------------------------------------------
    header_row = 7

    left_headers = [
        "INCOME",
        "Level",
        "Debit Balance",
        "Credit Balance",
    ]

    right_headers = [
        "EXPENSES",
        "Level",
        "Debit Balance",
        "Credit Balance",
    ]

    for c, value in enumerate(
        left_headers,
        start=1,
    ):
        cell = ws.cell(
            header_row,
            c,
            value,
        )
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for c, value in enumerate(
        right_headers,
        start=5,
    ):
        cell = ws.cell(
            header_row,
            c,
            value,
        )
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.freeze_panes = "A8"

    def ordered_rows(part: pd.DataFrame) -> List[Tuple[str, Any]]:
        result: List[Tuple[str, Any]] = []

        if part.empty:
            return result

        # Main rows first, then sub rows, while preserving Tally order within
        # each level.
        main = part[
            part["Level"].str.casefold().eq("main")
        ]
        sub = part[
            ~part["Level"].str.casefold().eq("main")
        ]

        for _, row in main.iterrows():
            result.append(("data", row))

        for _, row in sub.iterrows():
            result.append(("data", row))

        return result

    left_rows = ordered_rows(income)
    right_rows = ordered_rows(expenses)

    max_len = max(
        len(left_rows),
        len(right_rows),
    )

    current_row = header_row + 1

    def write_pnl_side_row(
        row_number: int,
        start_col: int,
        item: Tuple[str, Any],
    ) -> None:
        kind, value = item

        if kind != "data":
            return

        row = value

        particulars = clean(
            row["Particulars"]
        )

        level = clean(
            row["Level"]
        ) or "Sub"

        amount = float(
            row["Amount"]
        )

        # Income is normally positive/credit in Tally's P&L response.
        # Expenses are normally negative/debit.
        if amount < -EPSILON:
            debit = abs(amount)
            credit = 0.0
        elif amount > EPSILON:
            debit = 0.0
            credit = amount
        else:
            debit = 0.0
            credit = 0.0

        name_cell = ws.cell(
            row_number,
            start_col,
            particulars,
        )
        level_cell = ws.cell(
            row_number,
            start_col + 1,
            level,
        )
        debit_cell = ws.cell(
            row_number,
            start_col + 2,
            debit,
        )
        credit_cell = ws.cell(
            row_number,
            start_col + 3,
            credit,
        )

        for cell in [
            name_cell,
            level_cell,
            debit_cell,
            credit_cell,
        ]:
            cell.border = THIN_BORDER

        is_main = (
            level.casefold() == "main"
        )

        name_cell.font = Font(
            bold=is_main,
        )
        name_cell.alignment = Alignment(
            horizontal="left",
            indent=0 if is_main else 2,
        )

        level_cell.font = Font(
            bold=is_main,
        )
        level_cell.alignment = Alignment(
            horizontal="center",
        )

        for cell in [
            debit_cell,
            credit_cell,
        ]:
            cell.number_format = (
                '#,##0.00;[Red]-#,##0.00'
            )

    for i in range(max_len):

        if i < len(left_rows):
            write_pnl_side_row(
                current_row,
                1,
                left_rows[i],
            )

        if i < len(right_rows):
            write_pnl_side_row(
                current_row,
                5,
                right_rows[i],
            )

        current_row += 1

    # ------------------------------------------------------------
    # Side totals
    # ------------------------------------------------------------
    total_row = current_row + 1

    ws.cell(
        total_row,
        1,
        "TOTAL INCOME",
    )
    ws.cell(
        total_row,
        5,
        "TOTAL EXPENSES",
    )

    for c in [1, 5]:
        ws.cell(
            total_row,
            c,
        ).fill = TOTAL_FILL
        ws.cell(
            total_row,
            c,
        ).font = Font(
            bold=True,
        )

    # Income is credit; expenses are debit.
    ws.cell(
        total_row,
        3,
        0.0,
    )
    ws.cell(
        total_row,
        4,
        income_total,
    )
    ws.cell(
        total_row,
        7,
        expense_total,
    )
    ws.cell(
        total_row,
        8,
        0.0,
    )

    for c in [3, 4, 7, 8]:
        ws.cell(
            total_row,
            c,
        ).number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )
        ws.cell(
            total_row,
            c,
        ).fill = TOTAL_FILL
        ws.cell(
            total_row,
            c,
        ).font = Font(
            bold=True,
        )

    # ------------------------------------------------------------
    # Net Profit / Loss line
    # ------------------------------------------------------------
    result_row = total_row + 2

    ws.merge_cells(
        start_row=result_row,
        start_column=1,
        end_row=result_row,
        end_column=2,
    )
    ws.cell(
        result_row,
        1,
        "NET PROFIT / (LOSS)",
    )

    if net_profit >= -EPSILON:
        ws.cell(
            result_row,
            3,
            0.0,
        )
        ws.cell(
            result_row,
            4,
            max(net_profit, 0.0),
        )
    else:
        ws.cell(
            result_row,
            3,
            abs(net_profit),
        )
        ws.cell(
            result_row,
            4,
            0.0,
        )

    ws.merge_cells(
        start_row=result_row,
        start_column=5,
        end_row=result_row,
        end_column=8,
    )
    ws.cell(
        result_row,
        5,
        (
            "Profit"
            if net_profit >= -EPSILON
            else "Loss"
        ),
    )

    for c in range(1, 9):
        ws.cell(
            result_row,
            c,
        ).fill = SECTION_FILL
        ws.cell(
            result_row,
            c,
        ).font = Font(
            bold=True,
        )

    for c in [3, 4]:
        ws.cell(
            result_row,
            c,
        ).number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )

    # ------------------------------------------------------------
    # Reconciliation note
    # ------------------------------------------------------------
    reconcile_row = result_row + 1

    ws.merge_cells(
        start_row=reconcile_row,
        start_column=1,
        end_row=reconcile_row,
        end_column=8,
    )

    ws.cell(
        reconcile_row,
        1,
        (
            "Control: Total Income - Total Expenses = "
            "Net Profit / (Loss). P&L totals use Main rows only; "
            "Sub rows are displayed for detail and are excluded from "
            "the calculation."
        ),
    )

    ws.cell(
        reconcile_row,
        1,
    ).font = Font(
        italic=True,
        size=9,
    )

    # ------------------------------------------------------------
    # Formatting
    # ------------------------------------------------------------
    widths = {
        "A": 36,
        "B": 12,
        "C": 20,
        "D": 20,
        "E": 36,
        "F": 12,
        "G": 20,
        "H": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[
            col
        ].width = width

    ws.auto_filter.ref = (
        f"A{header_row}:H"
        f"{max(header_row, current_row - 1)}"
    )

    # Add a note so users understand the presentation.
    note_row = result_row + 2
    ws.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=8,
    )
    ws.cell(
        note_row,
        1,
        "Note: Tally signed balances are presented as "
        "Debit/Credit. Main and Sub levels are retained from "
        "the Tally report. This is a brief management presentation, "
        "not a statutory Schedule III filing.",
    )
    ws.cell(
        note_row,
        1,
    ).font = Font(
        italic=True,
        size=9,
    )
    ws.cell(
        note_row,
        1,
    ).alignment = Alignment(
        wrap_text=True,
    )

    wb.save(path)

    LOG.info(
        "Saved P&L workbook: %s "
        "(Income/Expense totals and Net Profit/Loss calculated from Main rows only)",
        path,
    )

    return path


def _schedule_iii_bucket(
    particulars: str,
    section: str,
) -> str:
    """
    Map a Tally MAIN group to a brief Schedule III-style bucket.

    This function is intentionally conservative. It classifies the group
    heading, and the Balance Sheet writer propagates that classification to
    its Sub rows. Individual ledger names are not guessed into statutory
    categories.

    This is a presentation classification, not a claim of statutory
    Schedule III compliance.
    """
    n = clean(particulars).casefold()

    if section == "Sources of Funds":

        if n in {
            "capital account",
            "reserves & surplus",
            "reserves and surplus",
            "profit & loss a/c",
            "profit and loss a/c",
            "profit & loss account",
            "profit and loss account",
        }:
            return "I. Equity"

        if n == "loans (liability)":
            return "II. Non-current Liabilities"

        if n in {
            "current liabilities",
            "duties & taxes",
            "duties and taxes",
            "provisions",
            "sundry creditors",
            "bank od a/c",
            "credit card",
            "deferred tax asset/liability",
            "suspense a/c",
        }:
            return "III. Current Liabilities"

        # If Tally itself gives a main group not known to this presentation,
        # keep it on the liabilities side rather than guessing an asset type.
        return "III. Current Liabilities"

    # Application of Funds / Assets
    if n in {
        "fixed assets",
        "fixed asset",
        "property, plant and equipment",
        "property plant and equipment",
        "investments",
        "long term investments",
        "long-term investments",
        "loans & advances (asset)",
        "loans and advances (asset)",
        "deferred tax asset",
        "misc. expenses (asset)",
        "misc expenses (asset)",
        "miscellaneous expenses (asset)",
    }:
        return "I. Non-current Assets"

    if n in {
        "current assets",
        "sundry debtors",
        "bank accounts",
        "cash-in-hand",
        "cash in hand",
        "stock-in-hand",
        "stock in hand",
        "inventory",
        "deposits (asset)",
        "branch / divisions",
        "branch/divisions",
    }:
        return "II. Current Assets"

    # For an unknown Tally main asset group, keep it as current only as a
    # presentation fallback; it is not claimed to be statutory classification.
    return "II. Current Assets"


def save_balance_sheet(
    df: pd.DataFrame,
    from_date: date,
    to_date: date,
) -> Path:
    """
    Save Balance Sheet in a concise, readable Schedule III-style layout.

    Design:
      - Two clearly separated sides: Equity & Liabilities / Assets.
      - Main groups are bold with a shaded row.
      - Sub-groups/ledgers are indented below their main group.
      - Debit and Credit balances are displayed in separate columns.
      - Tally's signed amount is not changed; negative values are shown as
        Debit Balance and positive values as Credit Balance.
      - The report is a brief Schedule III-style presentation, not a statutory
        Schedule III filing.
    """
    path = OUTPUT_DIR / (
        f"Balance_Sheet_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "BALANCE SHEET"
    ws["A1"].font = Font(
        bold=True,
        size=16,
    )
    ws["A1"].fill = TITLE_FILL
    ws["A1"].alignment = Alignment(
        horizontal="left",
        vertical="center",
    )
    ws.row_dimensions[1].height = 25

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"As at {to_date:%d-%m-%Y} "
        f"(Period: {from_date:%d-%m-%Y} to {to_date:%d-%m-%Y})"
    )
    ws["A2"].font = Font(
        italic=True,
        size=10,
    )

    if df is None or df.empty:
        ws.merge_cells("A4:H4")
        ws["A4"] = (
            "No Balance Sheet records were returned by Tally."
        )
        ws["A4"].font = Font(
            bold=True,
        )
        ws.column_dimensions["A"].width = 34
        wb.save(path)
        return path

    data = df.copy()

    for col in [
        "Section",
        "Particulars",
        "Level",
    ]:
        if col not in data.columns:
            data[col] = ""

    if "Amount" not in data.columns:
        data["Amount"] = 0.0

    data["Amount"] = pd.to_numeric(
        data["Amount"],
        errors="coerce",
    ).fillna(0.0)

    data["Particulars"] = data["Particulars"].map(clean)
    data["Section"] = data["Section"].map(clean)
    data["Level"] = data["Level"].map(clean)

    # Remove blank particulars and duplicate report rows.
    data = data[
        data["Particulars"].ne("")
    ].copy()

    data = data.drop_duplicates(
        subset=[
            "Section",
            "Particulars",
            "Amount",
            "Level",
        ],
        keep="first",
    )

    # ------------------------------------------------------------
    # Build the actual Tally Main Group for every row.
    #
    # The XML is hierarchical but flattened by the parser. A Main row is
    # followed by its Sub rows until the next Main row. Propagating the
    # current Main group preserves that hierarchy and prevents a ledger name
    # such as "COMPUTER" or "AC" from being independently classified.
    # ------------------------------------------------------------
    current_main_group = ""
    main_groups = []

    for _, row in data.iterrows():

        level = clean(
            row["Level"]
        ).casefold()

        if level == "main":
            current_main_group = clean(
                row["Particulars"]
            )

        main_groups.append(
            current_main_group
        )

    data["Main Group"] = main_groups

    # Preserve Tally's hierarchy and classify by the Main group only.
    data["Schedule Bucket"] = data.apply(
        lambda r: _schedule_iii_bucket(
            r["Main Group"] or r["Particulars"],
            r["Section"],
        ),
        axis=1,
    )

    # ------------------------------------------------------------
    # Summary / control area
    # ------------------------------------------------------------

    credit_total = float(
        data.loc[
            data["Amount"] > EPSILON,
            "Amount",
        ].sum()
    )

    debit_total = float(
        -data.loc[
            data["Amount"] < -EPSILON,
            "Amount",
        ].sum()
    )

    ws["A4"] = "Presentation"
    ws["A4"].font = Font(
        bold=True,
    )
    ws["B4"] = (
        "Brief Schedule III-style presentation; "
        "Tally hierarchy and signed balances are preserved."
    )

    ws["A5"] = "Total Debit Balance"
    ws["B5"] = debit_total
    ws["A6"] = "Total Credit Balance"
    ws["B6"] = credit_total
    ws["A7"] = "Difference"
    ws["B7"] = credit_total - debit_total

    for cell_ref in ["B5", "B6", "B7"]:
        ws[cell_ref].number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )

    ws["A5"].font = Font(bold=True)
    ws["A6"].font = Font(bold=True)
    ws["A7"].font = Font(bold=True)

    if abs(credit_total - debit_total) <= EPSILON:
        ws["C7"] = "BALANCED"
        ws["C7"].font = Font(bold=True)
    else:
        ws["C7"] = "CHECK"
        ws["C7"].font = Font(bold=True)

    # ------------------------------------------------------------
    # Two-column balance-sheet presentation.
    #
    # A:D = Equity & Liabilities
    # E:H = Assets
    # ------------------------------------------------------------

    header_row = 9

    left_headers = [
        "Equity & Liabilities",
        "Level",
        "Debit Balance",
        "Credit Balance",
    ]

    right_headers = [
        "Assets",
        "Level",
        "Debit Balance",
        "Credit Balance",
    ]

    for c, value in enumerate(
        left_headers,
        start=1,
    ):
        cell = ws.cell(
            header_row,
            c,
            value,
        )
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for c, value in enumerate(
        right_headers,
        start=5,
    ):
        cell = ws.cell(
            header_row,
            c,
            value,
        )
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    ws.freeze_panes = "A10"

    # ------------------------------------------------------------
    # Build rows grouped by Schedule III bucket.
    # Main rows are shown separately from their sub rows.
    # ------------------------------------------------------------

    left = data[
        data["Section"].eq("Sources of Funds")
    ].copy()

    right = data[
        data["Section"].eq("Application of Funds")
    ].copy()

    bucket_order = [
        "I. Equity",
        "II. Non-current Liabilities",
        "III. Current Liabilities",
        "I. Non-current Assets",
        "II. Current Assets",
    ]

    def prepare_side(
        side: pd.DataFrame,
        allowed_buckets: List[str],
    ) -> List[Tuple[str, Any]]:

        result: List[Tuple[str, Any]] = []

        if side.empty:
            return result

        # Keep Tally's report order. Insert a Schedule III-style bucket only
        # when the main-group classification changes.
        last_bucket = None

        for _, row in side.iterrows():

            bucket = clean(
                row["Schedule Bucket"]
            )

            if bucket not in allowed_buckets:
                continue

            if bucket != last_bucket:
                result.append(
                    ("bucket", bucket)
                )
                last_bucket = bucket

            result.append(
                ("data", row)
            )

        return result

    left_rows = prepare_side(
        left,
        [
            "I. Equity",
            "II. Non-current Liabilities",
            "III. Current Liabilities",
        ],
    )

    right_rows = prepare_side(
        right,
        [
            "I. Non-current Assets",
            "II. Current Assets",
        ],
    )

    max_len = max(
        len(left_rows),
        len(right_rows),
    )

    current_row = header_row + 1

    def write_side_row(
        row_number: int,
        start_col: int,
        item: Tuple[str, Any],
    ) -> None:

        kind, value = item

        if kind == "bucket":
            # Broad Schedule III-style heading.
            ws.cell(
                row_number,
                start_col,
                value,
            )
            ws.merge_cells(
                start_row=row_number,
                start_column=start_col,
                end_row=row_number,
                end_column=start_col + 3,
            )

            cell = ws.cell(
                row_number,
                start_col,
            )
            cell.fill = SECTION_FILL
            cell.font = Font(
                bold=True,
            )
            cell.alignment = Alignment(
                horizontal="left",
            )
            return

        row = value

        particulars = clean(
            row["Particulars"]
        )

        level = clean(
            row["Level"]
        ) or "Sub"

        amount = float(
            row["Amount"]
        )

        debit = (
            abs(amount)
            if amount < -EPSILON
            else 0.0
        )

        credit = (
            amount
            if amount > EPSILON
            else 0.0
        )

        name_cell = ws.cell(
            row_number,
            start_col,
            particulars,
        )

        level_cell = ws.cell(
            row_number,
            start_col + 1,
            level,
        )

        debit_cell = ws.cell(
            row_number,
            start_col + 2,
            debit,
        )

        credit_cell = ws.cell(
            row_number,
            start_col + 3,
            credit,
        )

        for cell in [
            name_cell,
            level_cell,
            debit_cell,
            credit_cell,
        ]:
            cell.border = THIN_BORDER

        is_main = (
            level.casefold() == "main"
        )

        name_cell.font = Font(
            bold=is_main,
        )

        name_cell.alignment = Alignment(
            horizontal="left",
            indent=0 if is_main else 2,
        )

        level_cell.font = Font(
            bold=is_main,
        )
        level_cell.alignment = Alignment(
            horizontal="center",
        )

        debit_cell.number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )
        credit_cell.number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )

    for i in range(max_len):
        left_item = (
            left_rows[i]
            if i < len(left_rows)
            else None
        )

        right_item = (
            right_rows[i]
            if i < len(right_rows)
            else None
        )

        if left_item is not None:
            write_side_row(
                current_row,
                1,
                left_item,
            )

        if right_item is not None:
            write_side_row(
                current_row,
                5,
                right_item,
            )

        current_row += 1

    # ------------------------------------------------------------
    # Closing totals
    # ------------------------------------------------------------

    total_row = current_row + 1

    ws.cell(
        total_row,
        1,
        "TOTAL EQUITY & LIABILITIES",
    )
    ws.cell(
        total_row,
        5,
        "TOTAL ASSETS",
    )

    for c in [1, 5]:
        ws.cell(
            total_row,
            c,
        ).fill = TOTAL_FILL
        ws.cell(
            total_row,
            c,
        ).font = Font(
            bold=True,
        )

    # ------------------------------------------------------------
    # IMPORTANT ACCOUNTING RULE:
    # Main rows are Tally's group totals. Sub rows are their details.
    # Therefore Balance Sheet totals must use Main rows only.
    # Summing Main + Sub rows would double-count the same balances.
    # ------------------------------------------------------------
    left_main = left[
        left["Level"]
        .str.casefold()
        .eq("main")
    ]

    right_main = right[
        right["Level"]
        .str.casefold()
        .eq("main")
    ]

    left_debit = float(
        -left_main.loc[
            left_main["Amount"] < -EPSILON,
            "Amount",
        ].sum()
    )
    left_credit = float(
        left_main.loc[
            left_main["Amount"] > EPSILON,
            "Amount",
        ].sum()
    )

    right_debit = float(
        -right_main.loc[
            right_main["Amount"] < -EPSILON,
            "Amount",
        ].sum()
    )
    right_credit = float(
        right_main.loc[
            right_main["Amount"] > EPSILON,
            "Amount",
        ].sum()
    )

    ws.cell(
        total_row,
        3,
        left_debit,
    )
    ws.cell(
        total_row,
        4,
        left_credit,
    )
    ws.cell(
        total_row,
        7,
        right_debit,
    )
    ws.cell(
        total_row,
        8,
        right_credit,
    )

    for c in [3, 4, 7, 8]:
        cell = ws.cell(
            total_row,
            c,
        )
        cell.fill = TOTAL_FILL
        cell.font = Font(
            bold=True,
        )
        cell.number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )

    # ------------------------------------------------------------
    # Reconciliation control
    # ------------------------------------------------------------
    control_row = total_row + 2

    equity_liability_net = (
        left_credit
        - left_debit
    )

    assets_net = (
        right_debit
        - right_credit
    )

    balance_difference = (
        equity_liability_net
        - assets_net
    )

    ws.cell(
        control_row,
        1,
        "Statement Difference",
    )
    ws.cell(
        control_row,
        2,
        balance_difference,
    )
    ws.cell(
        control_row,
        3,
        (
            "BALANCED"
            if abs(balance_difference) <= EPSILON
            else "CHECK"
        ),
    )

    ws.cell(
        control_row,
        1,
    ).font = Font(
        bold=True,
    )

    ws.cell(
        control_row,
        2,
    ).number_format = (
        '#,##0.00;[Red]-#,##0.00'
    )

    ws.cell(
        control_row,
        3,
    ).font = Font(
        bold=True,
    )

    # ------------------------------------------------------------
    # Reconciliation note
    # ------------------------------------------------------------

    note_row = control_row + 2

    ws.merge_cells(
        start_row=note_row,
        start_column=1,
        end_row=note_row,
        end_column=8,
    )

    ws.cell(
        note_row,
        1,
        (
            "Note: Tally's signed Balance Sheet amounts are presented as "
            "separate Debit/Credit balances. Main/Sub classification follows "
            "the XML hierarchy. Balance Sheet totals use Main group totals "
            "only; Sub rows are details and are not added again. "
            "Schedule III-style classification is a brief management "
            "presentation and is not a statutory filing."
        ),
    )

    ws.cell(
        note_row,
        1,
    ).font = Font(
        italic=True,
        size=9,
    )

    ws.cell(
        note_row,
        1,
    ).alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    # Widths
    widths = {
        "A": 38,
        "B": 12,
        "C": 18,
        "D": 18,
        "E": 38,
        "F": 12,
        "G": 18,
        "H": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[
            col
        ].width = width

    ws.auto_filter.ref = (
        f"A{header_row}:H{current_row - 1}"
    )

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"

    wb.save(path)

    LOG.info(
        "Saved formatted Schedule III-style Balance Sheet: %s "
        "(totals calculated from Main rows only)",
        path,
    )

    return path



def build_day_book(
    ledger_rows: List[Dict[str, Any]],
) -> pd.DataFrame:

    columns = [
        "Date",
        "Voucher Type",
        "Voucher Number",
        "Ledger/Party Name",
        "Particulars",
        "Debit",
        "Credit",
        "Narration",
        "Reference/Bill Number",
        "Bill Details",
        "GST/Tax Information",
        "Cost Centre",
        "Voucher ID",
        "Line No",
    ]

    if not ledger_rows:
        return pd.DataFrame(
            columns=columns
        )

    df = pd.DataFrame(
        ledger_rows
    )

    voucher_ledgers = defaultdict(list)

    for row in ledger_rows:
        voucher_ledgers[
            row["Voucher ID"]
        ].append(
            row["Ledger/Party Name"]
        )

    particulars = []

    for _, row in df.iterrows():

        current = clean(
            row["Ledger/Party Name"]
        )

        others = [
            value
            for value in dict.fromkeys(
                voucher_ledgers[
                    row["Voucher ID"]
                ]
            )
            if clean(value) != current
        ]

        particulars.append(
            unique_join(others)
        )

    df["Particulars"] = particulars

    df = df[
        columns
    ]

    return df.sort_values(
        [
            "Date",
            "Voucher Type",
            "Voucher Number",
            "Line No",
        ],
        kind="stable",
    ).reset_index(
        drop=True
    )


def build_voucher_wise(
    vouchers: List[Dict[str, Any]],
    ledger_rows: List[Dict[str, Any]],
) -> pd.DataFrame:

    columns = [
        "Voucher Date",
        "Voucher Type",
        "Voucher Number",
        "Party Name",
        "Ledger Entries",
        "Debit",
        "Credit",
        "Narration",
        "Reference/Bill Number",
        "Bill Details",
        "GST/Tax Information",
        "Cost Centre",
        "Ledger Entry Count",
        "Voucher ID",
    ]

    by_voucher = {}

    for voucher in vouchers:
        by_voucher[
            voucher["Voucher ID"]
        ] = {
            **voucher,
            "entries": [],
            "debit": 0.0,
            "credit": 0.0,
            "bills": [],
            "tax": [],
            "cost": [],
        }

    for row in ledger_rows:

        item = by_voucher.setdefault(
            row["Voucher ID"],
            {
                "Voucher ID":
                    row["Voucher ID"],

                "Voucher Date":
                    row["Date"],

                "Voucher Type":
                    row["Voucher Type"],

                "Voucher Number":
                    row["Voucher Number"],

                "Party Name":
                    row["Party Name"],

                "Narration":
                    row["Narration"],

                "Reference/Bill Number":
                    row["Reference/Bill Number"],

                "entries": [],
                "debit": 0.0,
                "credit": 0.0,
                "bills": [],
                "tax": [],
                "cost": [],
            },
        )

        item["entries"].append(
            (
                f"{row['Ledger/Party Name']} "
                f"(Dr {row['Debit']:.2f}, "
                f"Cr {row['Credit']:.2f})"
            )
        )

        item["debit"] += float(
            row["Debit"]
        )

        item["credit"] += float(
            row["Credit"]
        )

        if row.get(
            "Bill Details"
        ):
            item["bills"].append(
                row["Bill Details"]
            )

        if row.get(
            "GST/Tax Information"
        ):
            item["tax"].append(
                row[
                    "GST/Tax Information"
                ]
            )

        if row.get(
            "Cost Centre"
        ):
            item["cost"].append(
                row["Cost Centre"]
            )

    output = []

    for item in by_voucher.values():

        output.append(
            {
                "Voucher Date":
                    item["Voucher Date"],

                "Voucher Type":
                    item["Voucher Type"],

                "Voucher Number":
                    item["Voucher Number"],

                "Party Name":
                    item["Party Name"],

                "Ledger Entries":
                    unique_join(
                        item["entries"]
                    ),

                "Debit":
                    item["debit"],

                "Credit":
                    item["credit"],

                "Narration":
                    item["Narration"],

                "Reference/Bill Number":
                    item[
                        "Reference/Bill Number"
                    ],

                "Bill Details":
                    unique_join(
                        item["bills"]
                    ),

                "GST/Tax Information":
                    unique_join(
                        item["tax"]
                    ),

                "Cost Centre":
                    unique_join(
                        item["cost"]
                    ),

                "Ledger Entry Count":
                    len(
                        item["entries"]
                    ),

                "Voucher ID":
                    item["Voucher ID"],
            }
        )

    df = pd.DataFrame(
        output,
        columns=columns,
    )

    if not df.empty:
        df = df.sort_values(
            [
                "Voucher Date",
                "Voucher Type",
                "Voucher Number",
            ],
            kind="stable",
        )

    return df.reset_index(
        drop=True
    )


def build_ledger_wise(
    ledger_rows: List[Dict[str, Any]],
    opening_balances: Dict[str, float],
    closing_balances: Dict[str, float],
    ledger_master: Dict[str, Dict[str, Any]],
    groups: Dict[str, Dict[str, str]],
) -> pd.DataFrame:

    columns = [
        "Ledger Name",
        "Group",
        "Parent Group",
        "Voucher Date",
        "Voucher Type",
        "Voucher Number",
        "Particulars",
        "Debit",
        "Credit",
        "Opening Balance",
        "Running Balance",
        "Closing Balance",
        "Narration",
        "Reference/Bill Number",
        "Bill Details",
        "GST/Tax Information",
        "Cost Centre",
        "Voucher ID",
        "Line No",
    ]

    if ledger_rows:
        tx = pd.DataFrame(
            ledger_rows
        )
    else:
        tx = pd.DataFrame(
            columns=[
                "Date",
                "Voucher Type",
                "Voucher Number",
                "Ledger/Party Name",
                "Debit",
                "Credit",
                "Narration",
                "Reference/Bill Number",
                "Bill Details",
                "GST/Tax Information",
                "Cost Centre",
                "Voucher ID",
                "Line No",
            ]
        )

    totals = {}

    if not tx.empty:

        for ledger_name, group in tx.groupby(
            "Ledger/Party Name",
            dropna=False,
        ):

            key = clean(
                ledger_name
            ).casefold()

            totals[key] = {
                "name":
                    clean(ledger_name),

                "debit":
                    float(
                        group["Debit"].sum()
                    ),

                "credit":
                    float(
                        group["Credit"].sum()
                    ),
            }

    # Active ledgers:
    # include if opening OR movement OR closing is non-zero.
    all_names = {}

    for key, master in ledger_master.items():
        all_names[
            key
        ] = master["name"]

    for key, value in totals.items():
        all_names.setdefault(
            key,
            value["name"],
        )

    active = set()

    for key in all_names:

        opening = float(
            opening_balances.get(
                key,
                0.0,
            )
        )

        debit = float(
            totals.get(
                key,
                {},
            ).get(
                "debit",
                0.0,
            )
        )

        credit = float(
            totals.get(
                key,
                {},
            ).get(
                "credit",
                0.0,
            )
        )

        closing = float(
            closing_balances.get(
                key,
                opening + credit - debit,
            )
        )

        if (
            abs(opening) > EPSILON
            or
            abs(debit) > EPSILON
            or
            abs(credit) > EPSILON
            or
            abs(closing) > EPSILON
        ):
            active.add(key)

    if not tx.empty:

        tx["_key"] = (
            tx["Ledger/Party Name"]
            .astype(str)
            .str.strip()
            .str.casefold()
        )

        tx = tx[
            tx["_key"].isin(
                active
            )
        ].copy()

        tx.drop(
            columns=["_key"],
            inplace=True,
        )

        tx.sort_values(
            [
                "Ledger/Party Name",
                "Date",
                "Voucher Type",
                "Voucher Number",
                "Line No",
            ],
            kind="stable",
            inplace=True,
        )

    voucher_ledgers = defaultdict(list)

    for row in ledger_rows:
        voucher_ledgers[
            row["Voucher ID"]
        ].append(
            row["Ledger/Party Name"]
        )

    running = {}
    output = []

    for _, row in tx.iterrows():

        name = clean(
            row["Ledger/Party Name"]
        )

        key = name.casefold()

        opening = float(
            opening_balances.get(
                key,
                0.0,
            )
        )

        running.setdefault(
            key,
            opening,
        )

        debit = float(
            row["Debit"]
        )

        credit = float(
            row["Credit"]
        )

        # Signed convention:
        # debit movement increases debit side => reduces signed
        # credit-side balance; credit movement increases signed value.
        running[key] += (
            credit
            -
            debit
        )

        closing = float(
            closing_balances.get(
                key,
                running[key],
            )
        )

        group = clean(
            ledger_master.get(
                key,
                {},
            ).get(
                "group",
                "",
            )
        )

        parent = parent_group(
            group,
            groups,
        )

        others = [
            x
            for x in dict.fromkeys(
                voucher_ledgers[
                    row["Voucher ID"]
                ]
            )
            if clean(x) != name
        ]

        output.append(
            {
                "Ledger Name":
                    name,

                "Group":
                    group,

                "Parent Group":
                    parent,

                "Voucher Date":
                    row["Date"],

                "Voucher Type":
                    row["Voucher Type"],

                "Voucher Number":
                    row["Voucher Number"],

                "Particulars":
                    unique_join(others),

                "Debit":
                    debit,

                "Credit":
                    credit,

                "Opening Balance":
                    opening,

                "Running Balance":
                    running[key],

                "Closing Balance":
                    closing,

                "Narration":
                    row["Narration"],

                "Reference/Bill Number":
                    row["Reference/Bill Number"],

                "Bill Details":
                    row.get(
                        "Bill Details",
                        "",
                    ),

                "GST/Tax Information":
                    row.get(
                        "GST/Tax Information",
                        "",
                    ),

                "Cost Centre":
                    row.get(
                        "Cost Centre",
                        "",
                    ),

                "Voucher ID":
                    row["Voucher ID"],

                "Line No":
                    row.get(
                        "Line No",
                        "",
                    ),
            }
        )

    # Balance-only ledgers are retained.
    transaction_keys = set(
        tx[
            "Ledger/Party Name"
        ]
        .astype(str)
        .str.strip()
        .str.casefold()
    ) if not tx.empty else set()

    for key in sorted(active):

        if key in transaction_keys:
            continue

        name = all_names[key]

        opening = float(
            opening_balances.get(
                key,
                0.0,
            )
        )

        closing = float(
            closing_balances.get(
                key,
                opening,
            )
        )

        group = clean(
            ledger_master.get(
                key,
                {},
            ).get(
                "group",
                "",
            )
        )

        output.append(
            {
                "Ledger Name":
                    name,

                "Group":
                    group,

                "Parent Group":
                    parent_group(
                        group,
                        groups,
                    ),

                "Voucher Date":
                    None,

                "Voucher Type":
                    "",

                "Voucher Number":
                    "",

                "Particulars":
                    "Opening/Closing Balance",

                "Debit":
                    0.0,

                "Credit":
                    0.0,

                "Opening Balance":
                    opening,

                "Running Balance":
                    closing,

                "Closing Balance":
                    closing,

                "Narration":
                    "",

                "Reference/Bill Number":
                    "",

                "Bill Details":
                    "",

                "GST/Tax Information":
                    "",

                "Cost Centre":
                    "",

                "Voucher ID":
                    "",

                "Line No":
                    "",
            }
        )

    result = pd.DataFrame(
        output,
        columns=columns,
    )

    if not result.empty:
        result.sort_values(
            [
                "Group",
                "Ledger Name",
                "Voucher Date",
                "Voucher Type",
                "Voucher Number",
                "Line No",
            ],
            kind="stable",
            inplace=True,
        )

    return result.reset_index(
        drop=True
    )


def build_trial_balance(
    odbc_ledgers: Dict[str, Dict[str, Any]],
    ledger_master: Dict[str, Dict[str, Any]],
    groups: Dict[str, Dict[str, str]],
) -> pd.DataFrame:
    """
    Build Trial Balance directly from the verified Tally ODBC Ledger query.

    IMPORTANT:
    The previous implementation reconstructed Debit/Credit from XML voucher
    lines and Opening/Closing from XML Trial Balance snapshots. That can
    disagree with Tally because the XML report contains hierarchy/group
    rows and report-specific balance presentation.

    Tally ODBC is now authoritative because the user has independently
    verified this exact query in TallyPrime:

        SELECT $Name, $Parent, $_PrimaryGroup,
               $OpeningBalance, $DebitTotals,
               $CreditTotals, $_ClosingBalance
        FROM Ledger

    We therefore do NOT recalculate these four accounting values.
    """
    columns = [
        "Ledger Name",
        "Group",
        "Parent Group",
        "Opening Balance",
        "Debit",
        "Credit",
        "Closing Balance",
    ]

    rows = []

    for key, item in sorted(
        odbc_ledgers.items(),
        key=lambda kv: kv[1].get("name", "").casefold(),
    ):
        name = clean(
            item.get("name", "")
        )

        if not name:
            continue

        # Tally's predefined Profit & Loss A/c is not a Trial Balance
        # ledger for the selected period. It represents the carried-forward
        # previous-year result and is presented in the Balance Sheet.
        if name.casefold() in {
            "profit & loss a/c",
            "profit and loss a/c",
            "profit & loss account",
            "profit and loss account",
        }:
            continue

        opening = float(
            item.get("opening", 0.0) or 0.0
        )
        debit = abs(
            float(
                item.get("debit", 0.0) or 0.0
            )
        )
        credit = abs(
            float(
                item.get("credit", 0.0) or 0.0
            )
        )
        closing = float(
            item.get("closing", 0.0) or 0.0
        )

        # User requirement:
        # exclude ledgers having no opening, no transaction movement,
        # and no closing balance.
        if (
            abs(opening) <= EPSILON
            and abs(debit) <= EPSILON
            and abs(credit) <= EPSILON
            and abs(closing) <= EPSILON
        ):
            continue

        group = clean(
            item.get("group", "")
        )

        # $_PrimaryGroup is authoritative. Fall back to the XML group
        # hierarchy only when ODBC did not return it.
        primary_group = clean(
            item.get("parent_group", "")
        )

        if not primary_group:
            primary_group = parent_group(
                group,
                groups,
            )

        rows.append(
            {
                "Ledger Name": name,
                "Group": group,
                "Parent Group": primary_group,
                "Opening Balance": opening,
                "Debit": debit,
                "Credit": credit,
                "Closing Balance": closing,
            }
        )

    if not rows:
        return pd.DataFrame(
            columns=columns
        )

    return (
        pd.DataFrame(
            rows,
            columns=columns,
        )
        .sort_values(
            [
                "Group",
                "Parent Group",
                "Ledger Name",
            ],
            kind="stable",
        )
        .reset_index(drop=True)
    )

def build_ledger_summary(
    trial_balance: pd.DataFrame,
) -> pd.DataFrame:

    columns = [
        "Ledger Name",
        "Group",
        "Parent Group",
        "Opening Balance",
        "Debit",
        "Credit",
        "Net Movement",
        "Closing Balance",
    ]

    if trial_balance.empty:
        return pd.DataFrame(
            columns=columns
        )

    df = trial_balance.copy()

    df["Net Movement"] = (
        df["Credit"]
        -
        df["Debit"]
    )

    return df[
        columns
    ].sort_values(
        "Ledger Name",
        kind="stable",
    ).reset_index(
        drop=True
    )


# ================================================================
# CONTROL / VALIDATION
# ================================================================

def voucher_balance_control(
    ledger_rows: List[Dict[str, Any]],
) -> pd.DataFrame:

    if not ledger_rows:
        return pd.DataFrame(
            columns=[
                "Voucher ID",
                "Debit",
                "Credit",
                "Difference",
            ]
        )

    df = pd.DataFrame(
        ledger_rows
    )

    result = (
        df.groupby(
            "Voucher ID",
            as_index=False,
        )
        .agg(
            Debit=("Debit", "sum"),
            Credit=("Credit", "sum"),
        )
    )

    result["Difference"] = (
        result["Debit"]
        -
        result["Credit"]
    )

    return result


def trial_balance_controls(
    trial_balance: pd.DataFrame,
) -> Dict[str, Any]:

    if trial_balance.empty:
        return {
            "Opening Balance Total": 0.0,
            "Debit Total": 0.0,
            "Credit Total": 0.0,
            "Closing Balance Total": 0.0,
            "Opening Difference": 0.0,
            "Debit/Credit Difference": 0.0,
            "Closing Difference": 0.0,
            "Opening Balance Status": "BALANCED",
            "Debit/Credit Status": "BALANCED",
            "Closing Balance Status": "BALANCED",
            "Accounting Equation Difference": 0.0,
            "Accounting Equation Status": "BALANCED",
        }

    opening = float(
        trial_balance["Opening Balance"].sum()
    )
    debit = float(
        trial_balance["Debit"].sum()
    )
    credit = float(
        trial_balance["Credit"].sum()
    )
    closing = float(
        trial_balance["Closing Balance"].sum()
    )

    # For signed ledger balances:
    #   Closing = Opening + Credit - Debit
    #
    # Therefore:
    #   Opening + Credit - Debit - Closing = 0
    equation_difference = (
        opening
        + credit
        - debit
        - closing
    )

    opening_difference = opening
    debit_credit_difference = debit - credit
    closing_difference = closing

    return {
        "Opening Balance Total": opening,
        "Debit Total": debit,
        "Credit Total": credit,
        "Closing Balance Total": closing,
        "Opening Difference": opening_difference,
        "Debit/Credit Difference": debit_credit_difference,
        "Closing Difference": closing_difference,
        "Accounting Equation Difference":
            equation_difference,

        "Opening Balance Status":
            (
                "BALANCED"
                if abs(opening_difference) <= EPSILON
                else "CHECK TALLY OPENING BALANCES"
            ),

        "Debit/Credit Status":
            (
                "BALANCED"
                if abs(debit_credit_difference) <= EPSILON
                else "CHECK TALLY DEBIT/CREDIT TOTALS"
            ),

        "Closing Balance Status":
            (
                "BALANCED"
                if abs(closing_difference) <= EPSILON
                else "CHECK TALLY CLOSING BALANCES"
            ),

        "Accounting Equation Status":
            (
                "BALANCED"
                if abs(equation_difference) <= EPSILON
                else "CHECK PERIOD / BALANCE SOURCE"
            ),
    }


# ================================================================
# EXCEL
# ================================================================

HEADER_FILL = PatternFill(
    "solid",
    fgColor="1F4E78",
)

HEADER_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TITLE_FILL = PatternFill(
    "solid",
    fgColor="D9EAF7",
)

# Ledger-wise section heading fill.
SECTION_FILL = PatternFill(
    "solid",
    fgColor="EAF2F8",
)

TOTAL_FILL = PatternFill(
    "solid",
    fgColor="E2F0D9",
)

THIN_BORDER = Border(
    bottom=Side(
        style="thin",
        color="D9E2F3",
    )
)


def safe_sheet_name(
    name: str,
    used: set,
) -> str:

    name = clean(name)

    if not name:
        name = "Unnamed"

    name = re.sub(
        r'[:\\/?*\[\]]',
        "_",
        name,
    )

    name = name.strip("'") or "Unnamed"

    name = name[:31]

    original = name
    counter = 1

    while name in used:

        suffix = f"_{counter}"

        name = (
            original[
                :31 - len(suffix)
            ]
            + suffix
        )

        counter += 1

    used.add(name)

    return name


def unique_table_name(
    workbook: Workbook,
    seed: str,
) -> str:

    digest = hashlib.sha1(
        seed.encode(
            "utf-8"
        )
    ).hexdigest()[:12]

    base = re.sub(
        r"[^A-Za-z0-9_]",
        "_",
        seed,
    )

    if not base:
        base = "Data"

    if base[0].isdigit():
        base = "T_" + base

    base = (
        "T_"
        + base[:18]
        + "_"
        + digest
    )

    existing = set()

    for ws in workbook.worksheets:
        for table in ws.tables.values():
            existing.add(
                table.name
            )

    candidate = base
    counter = 1

    while candidate in existing:
        candidate = (
            f"{base[:25]}_{counter}"
        )
        counter += 1

    return candidate


def write_df_sheet(
    workbook: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    add_total: bool = True,
    freeze: str = "A2",
):
    """
    Writes a normal worksheet with filter/freeze/formatting.

    Deliberately does NOT create an Excel Table for large sheets.
    This substantially reduces workbook overhead and avoids duplicate
    table-name/corruption problems.
    """

    ws = workbook.create_sheet(
        sheet_name
    )

    ws.sheet_view.showGridLines = False

    if df is None or df.empty:

        ws["A1"] = (
            f"No records found in {sheet_name}"
        )

        ws["A1"].font = Font(
            bold=True,
            size=14,
        )

        return ws

    df = df.copy()

    date_columns = {
        "Date",
        "Voucher Date",
    }

    number_columns = {
        "Debit",
        "Credit",
        "Opening Balance",
        "Running Balance",
        "Closing Balance",
        "Net Movement",
    }

    headers = list(
        df.columns
    )

    # Header
    for col_no, header in enumerate(
        headers,
        start=1,
    ):

        cell = ws.cell(
            1,
            col_no,
            header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # Data
    for row_no, values in enumerate(
        df.itertuples(
            index=False,
            name=None,
        ),
        start=2,
    ):

        for col_no, value in enumerate(
            values,
            start=1,
        ):

            header = headers[
                col_no - 1
            ]

            cell = ws.cell(
                row_no,
                col_no,
                value,
            )

            if header in date_columns:
                cell.number_format = (
                    "dd-mm-yyyy"
                )

            elif header in number_columns:
                cell.number_format = (
                    '#,##0.00;[Red]-#,##0.00'
                )

    # Totals
    if add_total:

        total_row = len(df) + 2

        ws.cell(
            total_row,
            1,
            "TOTAL",
        )

        ws.cell(
            total_row,
            1,
        ).font = Font(
            bold=True,
        )

        ws.cell(
            total_row,
            1,
        ).fill = TOTAL_FILL

        for col_no, header in enumerate(
            headers,
            start=1,
        ):

            if header not in {
                "Debit",
                "Credit",
                "Net Movement",
            }:
                continue

            letter = get_column_letter(
                col_no
            )

            cell = ws.cell(
                total_row,
                col_no,
                (
                    f"=SUM("
                    f"{letter}2:"
                    f"{letter}{total_row - 1}"
                    f")"
                ),
            )

            cell.number_format = (
                '#,##0.00;[Red]-#,##0.00'
            )

            cell.font = Font(
                bold=True,
            )

            cell.fill = TOTAL_FILL

    ws.freeze_panes = freeze

    # Filter
    ws.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(headers))}"
        f"{len(df) + 1}"
    )

    # Width
    for col_no, header in enumerate(
        headers,
        start=1,
    ):

        width = max(
            12,
            len(str(header)) + 2,
        )

        sample = df[
            header
        ].head(200)

        for value in sample:

            if pd.isna(value):
                continue

            width = max(
                width,
                len(str(value)) + 2,
            )

        ws.column_dimensions[
            get_column_letter(col_no)
        ].width = min(
            width,
            45,
        )

    ws.row_dimensions[1].height = 30

    return ws


def add_index_sheet(
    workbook: Workbook,
    entries: List[Tuple[str, int, str]],
    name_header: str,
    count_header: str,
):
    ws = workbook.create_sheet(
        "Index",
        0,
    )

    ws.sheet_view.showGridLines = False

    headers = [
        name_header,
        count_header,
        "Worksheet",
        "Open",
    ]

    for col_no, header in enumerate(
        headers,
        start=1,
    ):

        cell = ws.cell(
            1,
            col_no,
            header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center"
        )

    for row_no, (
        name,
        count,
        sheet,
    ) in enumerate(
        entries,
        start=2,
    ):

        ws.cell(
            row_no,
            1,
            name,
        )

        ws.cell(
            row_no,
            2,
            count,
        )

        ws.cell(
            row_no,
            3,
            sheet,
        )

        link = ws.cell(
            row_no,
            4,
            "Open",
        )

        link.hyperlink = (
            f"#'{sheet}'!A1"
        )

        link.style = "Hyperlink"

    ws.freeze_panes = "A2"

    if entries:
        ws.auto_filter.ref = (
            f"A1:D{len(entries)+1}"
        )

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 12

    return ws


def save_daybook(
    df: pd.DataFrame,
    from_date: date,
    to_date: date,
) -> Path:

    path = OUTPUT_DIR / (
        f"Day_Book_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    wb.remove(
        wb.active
    )

    write_df_sheet(
        wb,
        "Day Book",
        df,
    )

    wb.save(
        path
    )

    return path


def save_voucher_wise(
    df: pd.DataFrame,
    from_date: date,
    to_date: date,
) -> Path:

    path = OUTPUT_DIR / (
        f"Voucher_Wise_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    wb.remove(
        wb.active
    )

    used = {"Index"}
    entries = []

    if df.empty:

        add_index_sheet(
            wb,
            [],
            "Voucher Type",
            "Voucher Count",
        )

    else:

        for voucher_type, group in df.groupby(
            "Voucher Type",
            dropna=False,
            sort=True,
        ):

            voucher_type = (
                clean(voucher_type)
                or
                "Unknown Voucher Type"
            )

            sheet = safe_sheet_name(
                voucher_type,
                used,
            )

            write_df_sheet(
                wb,
                sheet,
                group.reset_index(
                    drop=True
                ),
            )

            entries.append(
                (
                    voucher_type,
                    len(group),
                    sheet,
                )
            )

        add_index_sheet(
            wb,
            entries,
            "Voucher Type",
            "Voucher Count",
        )

    wb.save(
        path
    )

    return path


def save_ledger_wise_grouped(
    df: pd.DataFrame,
    from_date: date,
    to_date: date,
) -> Path:
    """
    Create one worksheet per Ledger Group.

    IMPORTANT:
    Each Group worksheet is internally organised LEDGER-WISE:
      - Ledger name section heading
      - Ledger transactions
      - Ledger subtotal
      - Next ledger

    Thus, for example, the "Sundry Debtors" worksheet contains all
    debtor ledgers, but each debtor is presented as its own logical
    section. This keeps the workbook much smaller than creating one
    worksheet per ledger while preserving ledger-wise readability.
    """

    path = OUTPUT_DIR / (
        f"Ledger_Wise_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    wb.remove(
        wb.active
    )

    used = {"Index"}
    entries = []

    if df.empty:

        add_index_sheet(
            wb,
            [],
            "Ledger Group",
            "Ledger Count",
        )

    else:

        # --------------------------------------------------------
        # One worksheet per GROUP
        # --------------------------------------------------------

        for group_name, group_df in df.groupby(
            "Group",
            dropna=False,
            sort=True,
        ):

            group_name = (
                clean(group_name)
                or
                "Ungrouped"
            )

            # Sort first by ledger, then date/voucher.
            sort_columns = [
                col
                for col in [
                    "Ledger Name",
                    "Voucher Date",
                    "Date",
                    "Voucher Type",
                    "Voucher Number",
                ]
                if col in group_df.columns
            ]

            if sort_columns:

                group_df = (
                    group_df
                    .sort_values(
                        sort_columns,
                        kind="stable",
                    )
                    .reset_index(drop=True)
                )

            # ----------------------------------------------------
            # Split very large GROUPS into workbook sheets.
            #
            # We estimate rows conservatively because each ledger
            # gets a heading and subtotal.
            # ----------------------------------------------------

            ledger_names = (
                group_df[
                    "Ledger Name"
                ]
                .fillna("Unnamed Ledger")
                .astype(str)
                .map(clean)
                .replace("", "Unnamed Ledger")
                .drop_duplicates()
                .tolist()
            )

            group_chunks = []

            current_rows = []
            estimated_rows = 0

            for ledger_name in ledger_names:

                ledger_df = group_df[
                    group_df[
                        "Ledger Name"
                    ]
                    .fillna("Unnamed Ledger")
                    .astype(str)
                    .map(clean)
                    == ledger_name
                ].copy()

                # Section overhead:
                # title + header + subtotal + blank
                required_rows = (
                    len(ledger_df)
                    + 4
                )

                if (
                    current_rows
                    and
                    estimated_rows
                    + required_rows
                    > MAX_ROWS_PER_SHEET
                ):

                    group_chunks.append(
                        pd.concat(
                            current_rows,
                            ignore_index=True,
                        )
                    )

                    current_rows = []
                    estimated_rows = 0

                current_rows.append(
                    ledger_df
                )

                estimated_rows += (
                    required_rows
                )

            if current_rows:

                group_chunks.append(
                    pd.concat(
                        current_rows,
                        ignore_index=True,
                    )
                )

            if not group_chunks:

                group_chunks = [
                    group_df
                ]

            # ----------------------------------------------------
            # WRITE EACH GROUP SHEET
            # ----------------------------------------------------

            for part_no, chunk in enumerate(
                group_chunks,
                start=1,
            ):

                base = group_name

                if len(group_chunks) > 1:

                    base = (
                        f"{group_name} - Part {part_no}"
                    )

                sheet = safe_sheet_name(
                    base,
                    used,
                )

                ws = wb.create_sheet(
                    sheet
                )

                ws.sheet_view.showGridLines = False

                # ------------------------------------------------
                # Group title
                # ------------------------------------------------

                ws["A1"] = (
                    f"Ledger Group: {group_name}"
                )

                ws["A1"].font = Font(
                    bold=True,
                    size=15,
                )

                ws["A1"].fill = TITLE_FILL

                ws.merge_cells(
                    start_row=1,
                    start_column=1,
                    end_row=1,
                    end_column=max(
                        len(chunk.columns),
                        1,
                    ),
                )

                ws["A2"] = (
                    f"Period: "
                    f"{from_date:%d-%m-%Y}"
                    f" to "
                    f"{to_date:%d-%m-%Y}"
                )

                ws["A2"].font = Font(
                    italic=True,
                )

                # ------------------------------------------------
                # Data starts at row 4
                # ------------------------------------------------

                current_row = 4

                columns = list(
                    chunk.columns
                )

                ledger_count = 0

                # Keep totals for the group
                group_debit = 0.0
                group_credit = 0.0

                # ------------------------------------------------
                # Ledger-wise sections
                # ------------------------------------------------

                for ledger_name, ledger_df in chunk.groupby(
                    "Ledger Name",
                    dropna=False,
                    sort=True,
                ):

                    ledger_name = (
                        clean(ledger_name)
                        or
                        "Unnamed Ledger"
                    )

                    ledger_count += 1

                    # --------------------------------------------
                    # Ledger section heading
                    # --------------------------------------------

                    ws.cell(
                        current_row,
                        1,
                        f"Ledger: {ledger_name}",
                    )

                    ws.cell(
                        current_row,
                        1,
                    ).font = Font(
                        bold=True,
                        size=12,
                    )

                    ws.cell(
                        current_row,
                        1,
                    ).fill = SECTION_FILL

                    ws.merge_cells(
                        start_row=current_row,
                        start_column=1,
                        end_row=current_row,
                        end_column=max(
                            len(columns),
                            1,
                        ),
                    )

                    current_row += 1

                    # --------------------------------------------
                    # Ledger table header
                    # --------------------------------------------

                    header_row = current_row

                    for col_no, header in enumerate(
                        columns,
                        start=1,
                    ):

                        cell = ws.cell(
                            header_row,
                            col_no,
                            header,
                        )

                        cell.fill = HEADER_FILL
                        cell.font = HEADER_FONT

                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                            wrap_text=True,
                        )

                    current_row += 1

                    # --------------------------------------------
                    # Ledger transactions
                    # --------------------------------------------

                    data_start = current_row

                    for values in ledger_df.itertuples(
                        index=False,
                        name=None,
                    ):

                        for col_no, value in enumerate(
                            values,
                            start=1,
                        ):

                            cell = ws.cell(
                                current_row,
                                col_no,
                                value,
                            )

                            header = columns[
                                col_no - 1
                            ]

                            if header in {
                                "Voucher Date",
                                "Date",
                            }:

                                if pd.notna(value):

                                    cell.number_format = (
                                        "dd-mm-yyyy"
                                    )

                            elif header in {
                                "Debit",
                                "Credit",
                                "Opening Balance",
                                "Running Balance",
                                "Closing Balance",
                                "Net Movement",
                            }:

                                cell.number_format = (
                                    '#,##0.00;[Red]-#,##0.00'
                                )

                        current_row += 1

                    data_end = current_row - 1

                    # --------------------------------------------
                    # Ledger subtotal
                    # --------------------------------------------

                    subtotal_row = current_row

                    ws.cell(
                        subtotal_row,
                        1,
                        f"{ledger_name} - TOTAL",
                    )

                    ws.cell(
                        subtotal_row,
                        1,
                    ).font = Font(
                        bold=True,
                    )

                    for col_no, header in enumerate(
                        columns,
                        start=1,
                    ):

                        if header not in {
                            "Debit",
                            "Credit",
                            "Net Movement",
                        }:

                            continue

                        letter = get_column_letter(
                            col_no
                        )

                        if data_end >= data_start:

                            formula = (
                                f"=SUM("
                                f"{letter}{data_start}:"
                                f"{letter}{data_end}"
                                f")"
                            )

                        else:

                            formula = "=0"

                        cell = ws.cell(
                            subtotal_row,
                            col_no,
                            formula,
                        )

                        cell.number_format = (
                            '#,##0.00;[Red]-#,##0.00'
                        )

                        cell.fill = TOTAL_FILL

                        cell.font = Font(
                            bold=True,
                        )

                    # Calculate group totals from source values,
                    # not from Excel formulas.
                    if "Debit" in ledger_df.columns:

                        group_debit += float(
                            pd.to_numeric(
                                ledger_df[
                                    "Debit"
                                ],
                                errors="coerce",
                            )
                            .fillna(0)
                            .sum()
                        )

                    if "Credit" in ledger_df.columns:

                        group_credit += float(
                            pd.to_numeric(
                                ledger_df[
                                    "Credit"
                                ],
                                errors="coerce",
                            )
                            .fillna(0)
                            .sum()
                        )

                    current_row += 2

                # ------------------------------------------------
                # Group control totals
                # ------------------------------------------------

                control_row = current_row

                ws.cell(
                    control_row,
                    1,
                    f"{group_name} GROUP TOTAL",
                )

                ws.cell(
                    control_row,
                    1,
                ).font = Font(
                    bold=True,
                    size=12,
                )

                debit_col = (
                    columns.index("Debit") + 1
                    if "Debit" in columns
                    else None
                )

                credit_col = (
                    columns.index("Credit") + 1
                    if "Credit" in columns
                    else None
                )

                if debit_col:

                    ws.cell(
                        control_row,
                        debit_col,
                        group_debit,
                    )

                    ws.cell(
                        control_row,
                        debit_col,
                    ).number_format = (
                        '#,##0.00;[Red]-#,##0.00'
                    )

                if credit_col:

                    ws.cell(
                        control_row,
                        credit_col,
                        group_credit,
                    )

                    ws.cell(
                        control_row,
                        credit_col,
                    ).number_format = (
                        '#,##0.00;[Red]-#,##0.00'
                    )

                # ------------------------------------------------
                # Freeze
                # ------------------------------------------------

                ws.freeze_panes = "A4"

                # ------------------------------------------------
                # Auto filter cannot span multiple repeated
                # ledger sections correctly, so deliberately
                # don't apply one large filter to this layout.
                # Each section is already ledger-separated.
                # ------------------------------------------------

                # ------------------------------------------------
                # Column widths
                # ------------------------------------------------

                for col_no, header in enumerate(
                    columns,
                    start=1,
                ):

                    width = len(
                        str(header)
                    )

                    sample = chunk[
                        header
                    ].head(200)

                    for value in sample:

                        if pd.isna(value):
                            continue

                        width = max(
                            width,
                            len(
                                str(value)
                            )
                        )

                    ws.column_dimensions[
                        get_column_letter(
                            col_no
                        )
                    ].width = min(
                        max(
                            width + 2,
                            12,
                        ),
                        45,
                    )

                entries.append(
                    (
                        base,
                        len(chunk),
                        sheet,
                    )
                )

        # --------------------------------------------------------
        # INDEX
        # --------------------------------------------------------

        add_index_sheet(
            wb,
            entries,
            "Ledger Group / Sheet",
            "Record Count",
        )

    wb.save(
        path
    )

    return path

def save_trial_balance(
    df: pd.DataFrame,
    controls: Dict[str, float],
    from_date: date,
    to_date: date,
) -> Path:

    path = OUTPUT_DIR / (
        f"Trial_Balance_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.sheet_view.showGridLines = False

    # Control section first
    ws["A1"] = "TRIAL BALANCE CONTROL"
    ws["A1"].font = Font(
        bold=True,
        size=14,
    )
    ws["A1"].fill = TITLE_FILL

    ws["A2"] = "Selected Period"
    ws["B2"] = (
        f"{from_date:%d-%m-%Y} to "
        f"{to_date:%d-%m-%Y}"
    )

    control_rows = [
        (
            "Total Opening Balance",
            controls[
                "Opening Balance Total"
            ],
        ),
        (
            "Total Debit",
            controls[
                "Debit Total"
            ],
        ),
        (
            "Total Credit",
            controls[
                "Credit Total"
            ],
        ),
        (
            "Total Closing Balance",
            controls[
                "Closing Balance Total"
            ],
        ),
        (
            "Opening Difference",
            controls[
                "Opening Difference"
            ],
        ),
        (
            "Debit/Credit Difference",
            controls[
                "Debit/Credit Difference"
            ],
        ),
        (
            "Closing Difference",
            controls[
                "Closing Difference"
            ],
        ),
    ]

    row_no = 3

    for label, value in control_rows:

        ws.cell(
            row_no,
            1,
            label,
        )

        ws.cell(
            row_no,
            2,
            value,
        )

        ws.cell(
            row_no,
            2,
        ).number_format = (
            '#,##0.00;[Red]-#,##0.00'
        )

        row_no += 1

    data_start = 12

    columns = list(
        df.columns
    )

    for col_no, header in enumerate(
        columns,
        start=1,
    ):

        cell = ws.cell(
            data_start,
            col_no,
            header,
        )

        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center"
        )

    for r, values in enumerate(
        df.itertuples(
            index=False,
            name=None,
        ),
        start=data_start + 1,
    ):

        for c, value in enumerate(
            values,
            start=1,
        ):

            cell = ws.cell(
                r,
                c,
                excel_safe_value(value),
            )

            if columns[c - 1] in {
                "Opening Balance",
                "Debit",
                "Credit",
                "Closing Balance",
            }:
                cell.number_format = (
                    '#,##0.00;[Red]-#,##0.00'
                )

    if not df.empty:

        total_row = (
            data_start
            + len(df)
            + 1
        )

        ws.cell(
            total_row,
            1,
            "TOTAL",
        ).font = Font(
            bold=True
        )

        for c, header in enumerate(
            columns,
            start=1,
        ):

            if header not in {
                "Opening Balance",
                "Debit",
                "Credit",
                "Closing Balance",
            }:
                continue

            letter = get_column_letter(c)

            cell = ws.cell(
                total_row,
                c,
                (
                    f"=SUM("
                    f"{letter}{data_start+1}:"
                    f"{letter}{total_row-1}"
                    f")"
                ),
            )

            cell.number_format = (
                '#,##0.00;[Red]-#,##0.00'
            )

            cell.fill = TOTAL_FILL
            cell.font = Font(
                bold=True
            )

        ws.auto_filter.ref = (
            f"A{data_start}:"
            f"{get_column_letter(len(columns))}"
            f"{data_start + len(df)}"
        )

    ws.freeze_panes = (
        f"A{data_start+1}"
    )

    widths = {
        "A": 40,
        "B": 28,
        "C": 28,
        "D": 20,
        "E": 20,
        "F": 20,
        "G": 20,
    }

    for col, width in widths.items():
        ws.column_dimensions[
            col
        ].width = width

    try:
        wb.save(path)
    except Exception as exc:
        LOG.exception(
            "Could not save Trial Balance workbook: %s",
            path,
        )
        raise RuntimeError(
            f"Could not save Trial Balance workbook.\n"
            f"Path: {path.resolve()}\n"
            f"Error: {exc}"
        ) from exc

    return path


def save_ledger_summary(
    df: pd.DataFrame,
    from_date: date,
    to_date: date,
) -> Path:

    path = OUTPUT_DIR / (
        f"Ledger_Summary_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    wb.remove(
        wb.active
    )

    write_df_sheet(
        wb,
        "Ledger Summary",
        df,
    )

    wb.save(
        path
    )

    return path


def save_parameters(
    from_date: date,
    to_date: date,
    company: Optional[str],
    host: str,
    port: int,
    dsn: str,
    stats: Dict[str, Any],
) -> Path:

    path = OUTPUT_DIR / (
        f"Parameters_{from_date:%Y%m%d}_"
        f"{to_date:%Y%m%d}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Parameters"
    ws.sheet_view.showGridLines = False

    rows = [
        (
            "TALLY ACCOUNTING DATA EXTRACTION",
            "",
        ),
        (
            "From Date",
            from_date.isoformat(),
        ),
        (
            "To Date",
            to_date.isoformat(),
        ),
        (
            "Company",
            company
            or
            "(currently loaded company)",
        ),
        (
            "Tally Host",
            host,
        ),
        (
            "Tally Port",
            port,
        ),
        (
            "ODBC DSN",
            dsn,
        ),
        (
            "Generated At",
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
        ),
        (
            "",
            "",
        ),
        (
            "EXTRACTION STATISTICS",
            "",
        ),
    ]

    for key, value in stats.items():
        rows.append(
            (
                key,
                value,
            )
        )

    for r, row in enumerate(
        rows,
        start=1,
    ):

        for c, value in enumerate(
            row,
            start=1,
        ):

            cell = ws.cell(
                r,
                c,
                excel_safe_value(value),
            )

            if r in {
                1,
                10,
            }:
                cell.font = Font(
                    bold=True,
                    size=14,
                )
                cell.fill = TITLE_FILL

    ws.column_dimensions[
        "A"
    ].width = 40

    ws.column_dimensions[
        "B"
    ].width = 75

    wb.save(
        path
    )

    return path


# ================================================================
# EXTRACTION PIPELINE
# ================================================================

def run_extraction(
    from_date: date,
    to_date: date,
    host: str,
    port: int,
    dsn: str,
    company: Optional[str],
    progress=None,
    log_callback=None,
) -> Dict[str, Any]:

    if from_date > to_date:
        raise ValueError(
            "From Date cannot be greater than To Date."
        )

    def emit(message):
        LOG.info(message)
        if log_callback:
            log_callback(message)

    def set_progress(value, message):
        if progress:
            progress(value, message)
        emit(message)

    # ------------------------------------------------------------
    # Connection
    # ------------------------------------------------------------

    set_progress(
        0.05,
        "Connecting to Tally HTTP server...",
    )

    client = TallyClient(
        host,
        port,
        log_callback=log_callback,
    )

    # Simple Tally test.
    test_request = f"""
<ENVELOPE>
<HEADER>
<VERSION>1</VERSION>
<TALLYREQUEST>Export</TALLYREQUEST>
<TYPE>Collection</TYPE>
<ID>PY_CONNECTION_TEST</ID>
</HEADER>
<BODY>
<DESC>
<STATICVARIABLES>
<SVEXPORTFORMAT>$$SysName:XML</SVEXPORTFORMAT>
{company_xml(company)}
</STATICVARIABLES>
<TDL>
<TDLMESSAGE>
<COLLECTION NAME="PY_CONNECTION_TEST">
<TYPE>Company</TYPE>
<FETCH>NAME</FETCH>
</COLLECTION>
</TDLMESSAGE>
</TDL>
</DESC>
</BODY>
</ENVELOPE>
""".strip()

    client.post(
        test_request,
        "Tally connection test",
    )

    set_progress(
        0.12,
        "Tally HTTP connection successful.",
    )

    # ------------------------------------------------------------
    # ODBC
    # ------------------------------------------------------------

    odbc_ok, odbc_message, odbc_tables = test_odbc(
        dsn,
        log_callback=log_callback,
    )

    set_progress(
        0.18,
        (
            "ODBC test successful."
            if odbc_ok
            else
            "ODBC test failed; continuing with Tally XML extraction."
        ),
    )

    # ------------------------------------------------------------
    # Ledger / Group master
    # ------------------------------------------------------------

    set_progress(
        0.25,
        "Fetching ledger master and group hierarchy...",
    )

    ledger_master = fetch_ledger_master(
        client,
        company,
    )

    groups = fetch_groups(
        client,
        company,
    )

    emit(
        f"Ledger master records: {len(ledger_master):,}"
    )

    # ------------------------------------------------------------
    # AUTHORITATIVE ODBC LEDGER BALANCES
    # ------------------------------------------------------------
    #
    # These values are the source of truth for Trial Balance and
    # ledger opening/debit/credit/closing balances.
    # ------------------------------------------------------------

    set_progress(
        0.30,
        "Reading authoritative Ledger balances through ODBC...",
    )

    odbc_ledger_balances = fetch_odbc_ledger_balances(
        dsn,
        emit,
    )

    emit(
        "Authoritative Trial Balance source: "
        "Tally ODBC Ledger collection."
    )

    emit(
        f"Group records: {len(groups):,}"
    )

    # ------------------------------------------------------------
    # Transactions
    # ------------------------------------------------------------

    set_progress(
        0.35,
        "Fetching vouchers and ledger entries...",
    )

    emit(
        "Requested transaction period: "
        f"{from_date:%d-%m-%Y} to {to_date:%d-%m-%Y}"
    )

    # ------------------------------------------------------------
    # Transaction extraction
    # ------------------------------------------------------------
    #
    # First use the explicit Voucher collection because it allows us to
    # request ALLLEDGERENTRIES.* directly.
    #
    # If that collection is empty, fall back to Tally's documented Day Book
    # report request. The returned XML is still locally date-filtered by
    # extract_vouchers_and_ledgers(), so a voucher outside the requested
    # period can never silently enter the output.
    # ------------------------------------------------------------

    root = client.post(
        build_voucher_request(
            from_date,
            to_date,
            company,
        ),
        "Voucher transactions",
    )

    vouchers, ledger_rows = (
        extract_vouchers_and_ledgers(
            root,
            from_date,
            to_date,
        )
    )

    if not vouchers:

        emit(
            "Custom Voucher collection returned no vouchers. "
            "Trying Tally's official Day Book report request as fallback..."
        )

        daybook_root = client.post(
            build_daybook_report_request(
                from_date,
                to_date,
                company,
            ),
            "Day Book fallback",
        )

        daybook_vouchers, daybook_ledger_rows = (
            extract_vouchers_and_ledgers(
                daybook_root,
                from_date,
                to_date,
            )
        )

        if daybook_vouchers:
            vouchers = daybook_vouchers
            ledger_rows = daybook_ledger_rows

            emit(
                f"Day Book fallback succeeded: "
                f"{len(vouchers):,} vouchers and "
                f"{len(ledger_rows):,} ledger lines."
            )

        else:

            earliest, latest = (
                get_odbc_voucher_date_range(
                    dsn,
                    emit,
                )
            )

            selected_text = (
                f"{from_date:%d-%m-%Y} to "
                f"{to_date:%d-%m-%Y}"
            )

            available_text = "unknown"

            if earliest or latest:
                available_text = (
                    f"{earliest:%d-%m-%Y}"
                    if earliest
                    else "unknown"
                )
                available_text += " to "
                available_text += (
                    f"{latest:%d-%m-%Y}"
                    if latest
                    else "unknown"
                )

            raise RuntimeError(
                "No vouchers were returned by Tally for the selected "
                "period.\n\n"
                f"Selected period: {selected_text}\n"
                f"Company: {company or '(currently loaded company)'}\n"
                f"ODBC-visible voucher range: {available_text}\n\n"
                "The program tried both:\n"
                "1. Explicit Voucher collection with a date filter.\n"
                "2. Tally's official Day Book XML report for the same dates.\n\n"
                "The XML responses contained no voucher dated inside the "
                "selected period. If the ODBC-visible date range does not "
                "include the selected period, load the correct company or "
                "choose a period containing transactions. If the ODBC range "
                "does include the period, the saved Tally XML response "
                "should be inspected because the Tally report context is "
                "not matching the ODBC context.\n\n"
                f"Latest response: "
                f"{OUTPUT_DIR / 'last_tally_response.xml'}"
            )

    emit(
        f"Vouchers extracted: {len(vouchers):,}"
    )

    emit(
        f"Ledger lines extracted: {len(ledger_rows):,}"
    )

    if vouchers and not ledger_rows:
        emit(
            "WARNING: Tally returned vouchers but no ALLLEDGERENTRIES.LIST "
            "objects were present in the XML response. The voucher request "
            "now explicitly fetches ALLLEDGERENTRIES.*. Check "
            "last_tally_response.xml if this warning remains."
        )

    # ------------------------------------------------------------
    # Native Tally financial reports
    # ------------------------------------------------------------

    set_progress(
        0.55,
        "Fetching native Tally Profit & Loss report...",
    )

    profit_loss_df = fetch_financial_report(
        client,
        "Profit & Loss A/c",
        from_date,
        to_date,
        company,
    )

    emit(
        f"Profit & Loss rows extracted: "
        f"{len(profit_loss_df):,}"
    )

    set_progress(
        0.58,
        "Fetching native Tally Balance Sheet report...",
    )

    balance_sheet_df = fetch_financial_report(
        client,
        "Balance Sheet",
        from_date,
        to_date,
        company,
    )

    emit(
        f"Balance Sheet rows extracted: "
        f"{len(balance_sheet_df):,}"
    )

    if profit_loss_df.empty:
        LOG.error(
            "Profit & Loss parser returned 0 rows. "
            "The raw financial-report XML should be inspected."
        )

    if balance_sheet_df.empty:
        LOG.error(
            "Balance Sheet parser returned 0 rows. "
            "The raw financial-report XML should be inspected."
        )

    # ------------------------------------------------------------
    # Trial Balance
    # ------------------------------------------------------------
    #
    # Do NOT reconstruct Trial Balance from XML Trial Balance snapshots.
    # The verified ODBC Ledger query is authoritative.
    # ------------------------------------------------------------

    set_progress(
        0.62,
        "Using authoritative ODBC Ledger balances for Trial Balance...",
    )

    opening_balances = {
        key: item["opening"]
        for key, item in odbc_ledger_balances.items()
    }

    closing_balances = {
        key: item["closing"]
        for key, item in odbc_ledger_balances.items()
    }

    emit(
        f"Authoritative ODBC ledger balances: "
        f"{len(odbc_ledger_balances):,}"
    )

    # ------------------------------------------------------------
    # Build reports
    # ------------------------------------------------------------

    set_progress(
        0.70,
        "Building report datasets...",
    )

    daybook_df = build_day_book(
        ledger_rows
    )

    voucher_df = build_voucher_wise(
        vouchers,
        ledger_rows,
    )

    ledger_df = build_ledger_wise(
        ledger_rows,
        opening_balances,
        closing_balances,
        ledger_master,
        groups,
    )

    trial_balance_df = build_trial_balance(
        odbc_ledger_balances,
        ledger_master,
        groups,
    )

    ledger_summary_df = build_ledger_summary(
        trial_balance_df
    )

    voucher_controls = voucher_balance_control(
        ledger_rows
    )

    tb_controls = trial_balance_controls(
        trial_balance_df
    )

    # Diagnostic only: compare detailed XML voucher movement with the
    # authoritative ODBC totals. We do not overwrite ODBC values.
    xml_debit = float(
        sum(
            float(r.get("Debit", 0.0) or 0.0)
            for r in ledger_rows
        )
    )
    xml_credit = float(
        sum(
            float(r.get("Credit", 0.0) or 0.0)
            for r in ledger_rows
        )
    )

    odbc_debit = float(
        trial_balance_df["Debit"].sum()
        if not trial_balance_df.empty
        else 0.0
    )
    odbc_credit = float(
        trial_balance_df["Credit"].sum()
        if not trial_balance_df.empty
        else 0.0
    )

    xml_odbc_debit_difference = (
        xml_debit - odbc_debit
    )
    xml_odbc_credit_difference = (
        xml_credit - odbc_credit
    )

    emit(
        "ODBC vs XML transaction control: "
        f"XML Debit={xml_debit:.2f}, "
        f"ODBC Debit={odbc_debit:.2f}, "
        f"Difference={xml_odbc_debit_difference:.2f}; "
        f"XML Credit={xml_credit:.2f}, "
        f"ODBC Credit={odbc_credit:.2f}, "
        f"Difference={xml_odbc_credit_difference:.2f}"
    )

    unbalanced_vouchers = (
        voucher_controls[
            voucher_controls[
                "Difference"
            ].abs()
            > EPSILON
        ]
        if not voucher_controls.empty
        else pd.DataFrame()
    )

    emit(
        f"Unbalanced voucher lines: "
        f"{len(unbalanced_vouchers):,}"
    )

    emit(
        "Trial Balance controls: "
        f"Opening={tb_controls['Opening Balance Total']:.2f}, "
        f"Debit={tb_controls['Debit Total']:.2f}, "
        f"Credit={tb_controls['Credit Total']:.2f}, "
        f"Closing={tb_controls['Closing Balance Total']:.2f}"
    )

    # ------------------------------------------------------------
    # Save eight workbooks
    # ------------------------------------------------------------

    set_progress(
        0.76,
        "Creating Day Book workbook...",
    )

    files = {}

    files["Day Book"] = save_daybook(
        daybook_df,
        from_date,
        to_date,
    )

    set_progress(
        0.80,
        "Creating Voucher Wise workbook...",
    )

    files["Voucher Wise"] = save_voucher_wise(
        voucher_df,
        from_date,
        to_date,
    )

    set_progress(
        0.84,
        "Creating group-wise Ledger Wise workbook...",
    )

    files["Ledger Wise"] = (
        save_ledger_wise_grouped(
            ledger_df,
            from_date,
            to_date,
        )
    )

    set_progress(
        0.88,
        "Creating Trial Balance workbook...",
    )

    files["Trial Balance"] = save_trial_balance(
        trial_balance_df,
        tb_controls,
        from_date,
        to_date,
    )

    set_progress(
        0.91,
        "Creating Profit & Loss workbook...",
    )

    files["Profit & Loss"] = save_profit_loss(
        profit_loss_df,
        from_date,
        to_date,
    )

    set_progress(
        0.93,
        "Creating Balance Sheet workbook...",
    )

    files["Balance Sheet"] = save_balance_sheet(
        balance_sheet_df,
        from_date,
        to_date,
    )

    set_progress(
        0.95,
        "Creating Ledger Summary workbook...",
    )

    files["Ledger Summary"] = save_ledger_summary(
        ledger_summary_df,
        from_date,
        to_date,
    )

    set_progress(
        0.98,
        "Creating Parameters workbook...",
    )

    files["Parameters"] = save_parameters(
        from_date,
        to_date,
        company,
        host,
        port,
        dsn,
        {
            "Vouchers Extracted":
                len(vouchers),

            "Day Book Rows":
                len(daybook_df),

            "Voucher Wise Rows":
                len(voucher_df),

            "Ledger Wise Rows":
                len(ledger_df),

            "Active Ledgers":
                trial_balance_df[
                    "Ledger Name"
                ].nunique()
                if not trial_balance_df.empty
                else 0,

            "Trial Balance Ledgers":
                len(trial_balance_df),

            "Profit & Loss Rows":
                len(profit_loss_df),

            "Balance Sheet Rows":
                len(balance_sheet_df),

            "Ledger Summary Rows":
                len(ledger_summary_df),

            "Unbalanced Vouchers":
                len(unbalanced_vouchers),

            "Opening Difference":
                tb_controls[
                    "Opening Difference"
                ],

            "Debit/Credit Difference":
                tb_controls[
                    "Debit/Credit Difference"
                ],

            "Closing Difference":
                tb_controls[
                    "Closing Difference"
                ],

            "Accounting Equation Difference":
                tb_controls[
                    "Accounting Equation Difference"
                ],

            "Accounting Equation Status":
                tb_controls[
                    "Accounting Equation Status"
                ],

            "XML Debit Total":
                xml_debit,

            "ODBC Debit Total":
                odbc_debit,

            "XML vs ODBC Debit Difference":
                xml_odbc_debit_difference,

            "XML Credit Total":
                xml_credit,

            "ODBC Credit Total":
                odbc_credit,

            "XML vs ODBC Credit Difference":
                xml_odbc_credit_difference,

            "ODBC Available":
                "Yes" if odbc_ok else "No",

            "ODBC Tables Discovered":
                len(odbc_tables),

            "Output Directory":
                str(
                    OUTPUT_DIR.resolve()
                ),
        },
    )

    set_progress(
        1.0,
        "Extraction completed successfully.",
    )

    return {
        "files": files,
        "daybook": daybook_df,
        "voucher": voucher_df,
        "ledger": ledger_df,
        "trial_balance": trial_balance_df,
        "ledger_summary": ledger_summary_df,
        "profit_loss": profit_loss_df,
        "balance_sheet": balance_sheet_df,
        "voucher_controls":
            voucher_controls,
        "unbalanced_vouchers":
            unbalanced_vouchers,
        "tb_controls":
            tb_controls,
        "odbc_ok":
            odbc_ok,
        "odbc_message":
            odbc_message,
        "odbc_tables":
            odbc_tables,
        "opening_balances":
            opening_balances,
        "closing_balances":
            closing_balances,
    }



def streamlit_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize mixed pandas object columns before Streamlit/pyarrow receives
    them. Tally data can legitimately contain mixed int/string values in
    fields such as Line No, and pyarrow otherwise raises ArrowTypeError.
    """
    if df is None:
        return pd.DataFrame()

    out = df.copy()

    for column in out.columns:

        if pd.api.types.is_object_dtype(
            out[column]
        ):
            out[column] = out[column].map(
                lambda value:
                    "" if pd.isna(value)
                    else str(value)
            )

    return out



# ================================================================
# STREAMLIT GUI
# ================================================================

def run_streamlit_app():

    import streamlit as st

    try:
        LOG.info(
            "Streamlit GUI loaded successfully. "
            "Waiting for user to click Extract Accounting Data."
        )
    except Exception:
        pass

    st.set_page_config(
        page_title="TallyPrime Accounting Extractor",
        page_icon="📊",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    st.markdown(
        """
        <style>
        .main {
            background: #f5f7fb;
        }

        .block-container {
            padding-top: 1.5rem;
            padding-bottom: 2rem;
            max-width: 1450px;
        }

        .hero {
            padding: 1.5rem 2rem;
            border-radius: 18px;
            background: linear-gradient(
                135deg,
                #0b3954 0%,
                #087e8b 55%,
                #2a9d8f 100%
            );
            color: white;
            margin-bottom: 1.5rem;
            box-shadow: 0 10px 30px rgba(0,0,0,.10);
        }

        .hero h1 {
            margin: 0;
            font-size: 2.2rem;
        }

        .hero p {
            margin: .45rem 0 0 0;
            opacity: .92;
            font-size: 1.02rem;
        }

        .section-card {
            background: white;
            padding: 1.15rem 1.25rem;
            border-radius: 14px;
            border: 1px solid #e4e8ef;
            margin-bottom: 1rem;
            box-shadow: 0 3px 12px rgba(0,0,0,.04);
        }

        .status-card {
            padding: .8rem 1rem;
            border-radius: 12px;
            background: #eef7f2;
            border: 1px solid #cfe8d9;
            color: #1f5134;
            margin-bottom: 1rem;
        }

        .warning-card {
            padding: .8rem 1rem;
            border-radius: 12px;
            background: #fff7e6;
            border: 1px solid #f2d39a;
            color: #76521b;
            margin-bottom: 1rem;
        }

        .metric-label {
            font-size: .85rem;
            color: #667085;
        }

        div.stButton > button {
            border-radius: 10px;
            font-weight: 700;
            min-height: 2.8rem;
        }

        .small-note {
            color: #667085;
            font-size: .85rem;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.markdown(
        """
        <div class="hero">
            <h1>📊 TallyPrime Accounting Extractor</h1>
            <p>
                Professional accounting-data extraction from TallyPrime
                into eight structured Excel workbooks.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------------
    # Sidebar
    # ------------------------------------------------------------

    with st.sidebar:

        st.markdown(
            "## ⚙️ Extraction Settings"
        )

        from_date = st.date_input(
            "From Date",
            value=date(
                datetime.now().year - 1,
                4,
                1,
            ),
            format="DD-MM-YYYY",
        )

        to_date = st.date_input(
            "To Date",
            value=date.today(),
            format="DD-MM-YYYY",
        )

        st.divider()

        host = st.text_input(
            "Tally Host",
            value=DEFAULT_HOST,
            help="Usually 127.0.0.1 for Tally running on this PC.",
        )

        port = st.number_input(
            "Tally HTTP/ODBC Port",
            min_value=1,
            max_value=65535,
            value=DEFAULT_PORT,
            step=1,
        )

        dsn = st.text_input(
            "Tally ODBC DSN",
            value=DEFAULT_DSN,
        )

        company = st.text_input(
            "Company Name (optional)",
            value="",
            help="Leave blank to use the company currently loaded in TallyPrime.",
        ).strip() or None

        st.divider()

        st.markdown(
            "### 📁 Output"
        )

        st.code(
            str(
                OUTPUT_DIR.resolve()
            ),
            language="text",
        )

        st.caption(
            "All eight Excel workbooks are created directly in this folder. "
            "Ledger Wise is grouped by Ledger Group to keep the workbook smaller."
        )

        st.divider()

        run_button = st.button(
            "🚀 Extract Accounting Data",
            type="primary",
            width="stretch",
        )

    # ------------------------------------------------------------
    # Top information
    # ------------------------------------------------------------

    col1, col2, col3, col4 = st.columns(4)

    with col1:
        st.metric(
            "Selected Period",
            (
                f"{from_date:%d-%m-%Y} → "
                f"{to_date:%d-%m-%Y}"
            ),
        )

    with col2:
        st.metric(
            "Tally Server",
            f"{host}:{port}",
        )

    with col3:
        st.metric(
            "ODBC DSN",
            dsn,
        )

    with col4:
        st.metric(
            "Output",
            "8 Excel Workbooks",
        )

    if from_date > to_date:
        st.error(
            "From Date cannot be greater than To Date."
        )
        return

    # ------------------------------------------------------------
    # Information
    # ------------------------------------------------------------

    st.markdown(
        """
        <div class="section-card">
        <b>What will be generated?</b><br>
        Day Book • Voucher Wise • Group-wise Ledger Wise • Trial Balance •
        Ledger Summary • Profit & Loss • Balance Sheet • Parameters
        <br><br>
        <span class="small-note">
        The Trial Balance uses Tally Trial Balance snapshots for the opening
        date and closing date. The program does not artificially force an
        accounting difference to zero.
        </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------------
    # Run
    # ------------------------------------------------------------

    if run_button:

        progress_bar = st.progress(
            0
        )

        status_box = st.empty()

        logs_box = st.expander(
            "Live extraction log",
            expanded=True,
        )

        log_lines = []

        def progress(value, message):
            progress_bar.progress(
                min(
                    max(
                        float(value),
                        0.0,
                    ),
                    1.0,
                )
            )
            status_box.info(
                message
            )

        def log_callback(message):
            log_lines.append(
                message
            )

        try:

            result = run_extraction(
                from_date=from_date,
                to_date=to_date,
                host=host,
                port=int(port),
                dsn=dsn,
                company=company,
                progress=progress,
                log_callback=log_callback,
            )

            progress_bar.progress(
                1.0
            )

            status_box.success(
                "Extraction completed successfully."
            )

            with logs_box:
                st.code(
                    "\n".join(
                        log_lines[-100:]
                    ),
                    language="text",
                )

            # ----------------------------------------------------
            # Metrics
            # ----------------------------------------------------

            st.markdown(
                "## 📈 Extraction Summary"
            )

            c1, c2, c3, c4, c5 = st.columns(5)

            with c1:
                st.metric(
                    "Vouchers",
                    f"{len(result['voucher']):,}",
                )

            with c2:
                st.metric(
                    "Day Book Rows",
                    f"{len(result['daybook']):,}",
                )

            with c3:
                st.metric(
                    "Ledger Rows",
                    f"{len(result['ledger']):,}",
                )

            with c4:
                st.metric(
                    "Trial Balance Ledgers",
                    f"{len(result['trial_balance']):,}",
                )

            with c5:
                st.metric(
                    "Unbalanced Vouchers",
                    f"{len(result['unbalanced_vouchers']):,}",
                )

            f1, f2 = st.columns(2)

            with f1:
                st.metric(
                    "P&L Rows",
                    f"{len(result['profit_loss']):,}",
                )

            with f2:
                st.metric(
                    "Balance Sheet Rows",
                    f"{len(result['balance_sheet']):,}",
                )

            # ----------------------------------------------------
            # Trial Balance controls
            # ----------------------------------------------------

            st.markdown(
                "## ⚖️ Trial Balance Control"
            )

            controls = result[
                "tb_controls"
            ]

            cc1, cc2, cc3, cc4 = st.columns(4)

            with cc1:
                st.metric(
                    "Opening Balance Total",
                    f"{controls['Opening Balance Total']:,.2f}",
                )

            with cc2:
                st.metric(
                    "Debit Total",
                    f"{controls['Debit Total']:,.2f}",
                )

            with cc3:
                st.metric(
                    "Credit Total",
                    f"{controls['Credit Total']:,.2f}",
                )

            with cc4:
                st.metric(
                    "Closing Balance Total",
                    f"{controls['Closing Balance Total']:,.2f}",
                )

            if (
                abs(
                    controls[
                        "Debit/Credit Difference"
                    ]
                ) <= EPSILON
                and
                abs(
                    controls[
                        "Opening Difference"
                    ]
                ) <= EPSILON
                and
                abs(
                    controls[
                        "Closing Difference"
                    ]
                ) <= EPSILON
            ):

                st.success(
                    "Trial Balance control is balanced within the 0.01 tolerance."
                )

            else:

                st.warning(
                    "The extracted Trial Balance has a non-zero control difference. "
                    "The program has NOT artificially changed the accounting values. "
                    "See the Trial Balance Control section in the workbook."
                )

            # ----------------------------------------------------
            # Voucher control
            # ----------------------------------------------------

            if not result[
                "unbalanced_vouchers"
            ].empty:

                with st.expander(
                    "⚠️ Unbalanced vouchers",
                    expanded=False,
                ):

                    st.dataframe(
                        streamlit_safe_dataframe(
                            result[
                                "unbalanced_vouchers"
                            ]
                        ),
                        width="stretch",
                        hide_index=True,
                    )

            # ----------------------------------------------------
            # Files
            # ----------------------------------------------------

            st.markdown(
                "## 📁 Generated Workbooks"
            )

            for title, path in result[
                "files"
            ].items():

                col_a, col_b = st.columns(
                    [4, 1]
                )

                with col_a:

                    size_mb = (
                        path.stat().st_size
                        /
                        (
                            1024
                            *
                            1024
                        )
                    )

                    st.write(
                        f"**{title}**  \n"
                        f"`{path.name}` — {size_mb:.2f} MB"
                    )

                with col_b:

                    with open(
                        path,
                        "rb",
                    ) as file_handle:

                        st.download_button(
                            "⬇️ Download",
                            data=file_handle.read(),
                            file_name=path.name,
                            mime=(
                                "application/vnd.openxmlformats-"
                                "officedocument.spreadsheetml.sheet"
                            ),
                            key=(
                                "download_"
                                + re.sub(
                                    r"[^A-Za-z0-9]+",
                                    "_",
                                    title,
                                )
                            ),
                            width="stretch",
                        )

            # ----------------------------------------------------
            # Previews
            # ----------------------------------------------------

            st.markdown(
                "## 🔎 Data Preview"
            )

            tabs = st.tabs(
                [
                    "Day Book",
                    "Voucher Wise",
                    "Ledger Wise",
                    "Trial Balance",
                    "Ledger Summary",
                    "Profit & Loss",
                    "Balance Sheet",
                ]
            )

            preview_map = [
                result["daybook"],
                result["voucher"],
                result["ledger"],
                result["trial_balance"],
                result["ledger_summary"],
                result["profit_loss"],
                result["balance_sheet"],
            ]

            for tab, dataframe in zip(
                tabs,
                preview_map,
            ):

                with tab:
                    st.dataframe(
                        streamlit_safe_dataframe(
                            dataframe.head(200)
                        ),
                        width="stretch",
                        hide_index=True,
                    )

            st.markdown(
                "## 📝 Output Location"
            )

            st.code(
                str(
                    OUTPUT_DIR.resolve()
                ),
                language="text",
            )

            st.caption(
                "The detailed log is stored in tally_extractor.log."
            )

        except Exception as exc:

            LOG.exception(
                "Extraction failed"
            )

            progress_bar.progress(
                0
            )

            status_box.error(
                "Extraction failed."
            )

            st.error(
                str(exc)
            )

            with st.expander(
                "Technical error details",
                expanded=True,
            ):

                st.code(
                    traceback.format_exc(),
                    language="text",
                )

                st.info(
                    "Check Tally_Output/tally_extractor.log "
                    "and last_tally_response.xml for the raw Tally response."
                )


# ================================================================
# DOUBLE-CLICK LAUNCHER
# ================================================================

def running_inside_streamlit() -> bool:

    try:
        from streamlit.runtime.scriptrunner import (
            get_script_run_ctx,
        )

        return (
            get_script_run_ctx()
            is not None
        )

    except Exception:
        return False


def _find_free_local_port(start_port=8501, max_tries=20):
    """Find an available localhost TCP port without requiring extra packages."""
    import socket

    for port in range(int(start_port), int(start_port) + max_tries):
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        try:
            sock.settimeout(0.25)
            if sock.connect_ex(("127.0.0.1", port)) != 0:
                return port
        finally:
            sock.close()

    return None


def launch_streamlit_from_double_click():
    """
    Reliable double-click launcher.

    The previous launcher only logged that Streamlit was started. If the child
    process failed immediately, the user had no useful diagnostic information.
    This version captures Streamlit stdout/stderr in streamlit_startup.log,
    selects a free port, waits for the server, and then opens the browser.
    """

    script = str(
        Path(__file__).resolve()
    )

    # ------------------------------------------------------------
    # Check Streamlit installation
    # ------------------------------------------------------------

    try:
        import streamlit  # noqa: F401
    except ImportError:
        message = (
            "Streamlit is not installed.\n\n"
            "Open Command Prompt and run:\n\n"
            "py -m pip install pandas openpyxl pyodbc streamlit\n\n"
            "Then double-click this file again."
        )

        try:
            import tkinter as tk
            from tkinter import messagebox

            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "TallyPrime Extractor",
                message,
            )
            root.destroy()
        except Exception:
            print(message)

        return

    # ------------------------------------------------------------
    # Find available port
    # ------------------------------------------------------------

    port = _find_free_local_port(
        start_port=8501,
        max_tries=30,
    )

    if port is None:
        message = (
            "Could not find a free local port for Streamlit.\n\n"
            "Try closing other Streamlit applications and run again."
        )

        try:
            import tkinter as tk
            from tkinter import messagebox

            root = tk.Tk()
            root.withdraw()
            messagebox.showerror(
                "TallyPrime Extractor",
                message,
            )
            root.destroy()
        except Exception:
            print(message)

        return

    # ------------------------------------------------------------
    # Capture child process output
    # ------------------------------------------------------------

    startup_log = (
        OUTPUT_DIR /
        "streamlit_startup.log"
    )

    try:
        startup_handle = open(
            startup_log,
            "a",
            encoding="utf-8",
            buffering=1,
        )
    except Exception:
        startup_handle = None

    env = os.environ.copy()

    env[
        "TALLY_EXTRACTOR_STREAMLIT_CHILD"
    ] = "1"

    command = [
        sys.executable,
        "-m",
        "streamlit",
        "run",
        script,
        "--server.address=127.0.0.1",
        f"--server.port={port}",
        "--server.headless=true",
        "--browser.gatherUsageStats=false",
    ]

    LOG.info(
        "Launching Streamlit: %s",
        command,
    )

    if startup_handle:
        startup_handle.write(
            "\n"
            + "=" * 80
            + "\n"
            + f"Started: {datetime.now():%Y-%m-%d %H:%M:%S}\n"
            + f"Python: {sys.executable}\n"
            + f"Script: {script}\n"
            + f"Port: {port}\n"
            + f"Command: {command}\n"
            + "=" * 80
            + "\n"
        )
        startup_handle.flush()

    try:

        creationflags = 0

        if os.name == "nt":
            creationflags = (
                subprocess.CREATE_NEW_PROCESS_GROUP
                | subprocess.CREATE_NO_WINDOW
            )

        process = subprocess.Popen(
            command,
            env=env,
            cwd=str(
                Path(__file__).resolve().parent
            ),
            stdout=(
                startup_handle
                if startup_handle
                else subprocess.DEVNULL
            ),
            stderr=(
                startup_handle
                if startup_handle
                else subprocess.DEVNULL
            ),
            creationflags=creationflags,
        )

        # --------------------------------------------------------
        # Wait until Streamlit actually starts.
        # --------------------------------------------------------

        import socket

        ready = False

        for _ in range(30):

            if process.poll() is not None:

                if startup_handle:
                    startup_handle.write(
                        "\n"
                        f"Streamlit process exited with code "
                        f"{process.returncode}.\n"
                    )
                    startup_handle.flush()

                break

            sock = socket.socket(
                socket.AF_INET,
                socket.SOCK_STREAM,
            )

            try:

                sock.settimeout(0.5)

                if sock.connect_ex(
                    ("127.0.0.1", port)
                ) == 0:

                    ready = True
                    break

            finally:

                sock.close()

            time.sleep(0.5)

        if startup_handle:
            startup_handle.write(
                f"\nServer ready: {ready}\n"
                f"URL: http://127.0.0.1:{port}\n"
            )
            startup_handle.flush()

        if ready:

            url = (
                f"http://127.0.0.1:{port}"
            )

            LOG.info(
                "Streamlit server ready at %s",
                url,
            )

            webbrowser.open(
                url
            )

        else:

            LOG.error(
                "Streamlit did not become ready. "
                "See %s",
                startup_log,
            )

            message = (
                "Streamlit could not start.\n\n"
                f"Diagnostic log:\n{startup_log.resolve()}\n\n"
                "Open that file to see the exact error."
            )

            try:

                import tkinter as tk
                from tkinter import messagebox

                root = tk.Tk()
                root.withdraw()

                messagebox.showerror(
                    "TallyPrime Extractor",
                    message,
                )

                root.destroy()

            except Exception:

                print(message)

    except Exception as exc:

        LOG.exception(
            "Could not launch Streamlit"
        )

        if startup_handle:

            startup_handle.write(
                "\nFATAL LAUNCH ERROR:\n"
                + traceback.format_exc()
                + "\n"
            )

            startup_handle.flush()

        print(
            "Could not start Streamlit."
        )

        print(
            f"Diagnostic log: {startup_log.resolve()}"
        )

    finally:

        if startup_handle:
            startup_handle.close()


# ================================================================
# ENTRY POINT
# ================================================================

if __name__ == "__main__":

    if running_inside_streamlit():

        run_streamlit_app()

    elif (
        os.environ.get(
            "TALLY_EXTRACTOR_STREAMLIT_CHILD"
        )
        == "1"
    ):

        run_streamlit_app()

    else:

        launch_streamlit_from_double_click()
